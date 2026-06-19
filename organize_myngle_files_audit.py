"""
organize_myngle_files_audit.py

Safe file-organization audit for local Myngle data folders.

Workflow:
  1. Dry-run  → writes an Excel + CSV plan file (all rows start with approve=NO)
  2. User     → opens plan, sets approve=YES for chosen rows
  3. Apply    → reads approved plan, executes only YES rows

Usage:
  python organize_myngle_files_audit.py \
      --target-root "C:\\Users\\...\\Myngle" \
      --queues Italy50 Italy100 Italy200 \
      --dry-run

  python organize_myngle_files_audit.py \
      --target-root "C:\\Users\\...\\Myngle" \
      --apply-plan "C:\\...\\file_organization_plan_20260619_120000.xlsx" \
      [--allow-delete]
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import shutil
import sys
import unicodedata
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Dependency check
# ---------------------------------------------------------------------------
try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

EXPECTED_FOLDERS = {
    "00_raw",
    "01_cleaned_domains",
    "02_lead_prioritized",
    "03_opportunity_radar",
    "04_caller_briefs",
    "_archive",
}
OPTIONAL_FOLDERS = {
    "_logs",
    "02_lead_prioritized_url_patched",
    "03_opportunity_input_url_patched",
    "_url_patch_reports",
}

SCRIPT_EXTENSIONS = {".py", ".bat", ".ps1", ".sh"}

# Filename patterns
_RAW_RE = re.compile(
    r"^(?P<queue>[A-Za-z0-9]+)_(?P<batch>\d+)_R(?P<start>\d+)[_\-](?P<end>\d+)\.xlsx$",
    re.IGNORECASE,
)
_CLEANED_RE = re.compile(
    r"^(?P<queue>[A-Za-z0-9]+)_(?P<batch>\d+)_R(?P<start>\d+)[_\-](?P<end>\d+)_cleaned_",
    re.IGNORECASE,
)
_ENRICHED_RE = re.compile(
    r"^(?P<queue>[A-Za-z0-9]+)_(?P<batch>\d+)_R(?P<start>\d+)[_\-](?P<end>\d+)_enriched_",
    re.IGNORECASE,
)
_GENERIC_ENRICHED_RE = re.compile(
    r"^enrichedResults[_\-]?\d*", re.IGNORECASE
)
_GENERIC_RESULTS_RE = re.compile(
    r"^Results[_\-]?\d*", re.IGNORECASE
)
_RANGE_RE = re.compile(r"[Rr](\d+)[_\-](\d+)")
_TS_RE = re.compile(r"\d{8}[_\-]?\d{4,6}")
_DOUBLE_SUFFIX_RE = re.compile(
    r"(url_patched|url_patch).*?(url_patched|url_patch)", re.IGNORECASE
)

PLAN_FIELDS = [
    "approve",
    "queue",
    "current_path",
    "current_name",
    "item_type",
    "detected_category",
    "suggested_action",
    "suggested_new_path",
    "suggested_new_name",
    "reason",
    "confidence",
    "risk_level",
    "batch_id_detected",
    "range_detected",
    "related_file",
    "notes",
    "requires_manual_check",
]

LOG_FIELDS = [
    "timestamp",
    "approved_action",
    "old_path",
    "new_path",
    "status",
    "error",
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_open_wb(path: Path):
    """Return (workbook, error_str). workbook is None if unreadable."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        return wb, None
    except (zipfile.BadZipFile, KeyError, OSError, Exception) as exc:
        return None, str(exc)


def _parse_range(filename: str) -> tuple[int, int] | None:
    m = _RANGE_RE.search(filename)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None


def _safe_new_path(target: Path) -> Path:
    if not target.exists():
        return target
    stem, suffix = target.stem, target.suffix
    parent = target.parent
    n = 2
    while True:
        candidate = parent / f"{stem}_{n}{suffix}"
        if not candidate.exists():
            return candidate
        n += 1


def _ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _plan_row(**kwargs) -> dict:
    defaults = {f: "" for f in PLAN_FIELDS}
    defaults["approve"] = "NO"
    defaults["requires_manual_check"] = "NO"
    defaults["risk_level"] = "LOW"
    defaults["confidence"] = "HIGH"
    defaults.update(kwargs)
    return defaults


# ---------------------------------------------------------------------------
# Detection functions
# ---------------------------------------------------------------------------

def detect_lockfiles(folder: Path, queue: str, rows: list[dict]) -> None:
    for f in folder.rglob("~$*"):
        if not f.is_file():
            continue
        rows.append(_plan_row(
            queue=queue,
            current_path=str(f.parent),
            current_name=f.name,
            item_type="file",
            detected_category="LOCKFILE_TEMP",
            suggested_action="DELETE_LOCKFILE",
            suggested_new_path="",
            suggested_new_name="",
            reason="Excel lock/temp file left by an open workbook",
            confidence="HIGH",
            risk_level="LOW",
            notes="Only safe to delete when Excel is fully closed",
            requires_manual_check="YES",
        ))


def detect_corrupt_workbooks(folder: Path, queue: str, rows: list[dict]) -> int:
    count = 0
    for f in folder.rglob("*.xlsx"):
        if f.name.startswith("~$"):
            continue
        if "_url_patched" in str(f) or "_archive" in str(f).lower():
            continue
        wb, err = _safe_open_wb(f)
        if wb is None:
            archive_target = folder / "_archive" / "corrupt" / f.name
            rows.append(_plan_row(
                queue=queue,
                current_path=str(f.parent),
                current_name=f.name,
                item_type="file",
                detected_category="CORRUPT_WORKBOOK",
                suggested_action="MOVE_TO_ARCHIVE_CORRUPT",
                suggested_new_path=str(archive_target.parent),
                suggested_new_name=f.name,
                reason=f"Cannot open workbook: {err}",
                confidence="HIGH",
                risk_level="MEDIUM",
                notes="Verify file is not currently open in Excel",
                requires_manual_check="YES",
            ))
            count += 1
        else:
            wb.close()
    return count


def detect_double_suffix_folders(queue_dir: Path, queue: str, rows: list[dict]) -> None:
    for item in queue_dir.rglob("*"):
        if not item.is_dir():
            continue
        if _DOUBLE_SUFFIX_RE.search(item.name):
            file_count = sum(1 for _ in item.rglob("*") if _.is_file())
            risk = "HIGH" if file_count > 0 else "LOW"
            clean_name = _DOUBLE_SUFFIX_RE.sub(
                lambda m: m.group(1), item.name
            )
            archive_target = queue_dir / "_archive" / f"cleanup_{_ts()}" / item.name
            rows.append(_plan_row(
                queue=queue,
                current_path=str(item.parent),
                current_name=item.name,
                item_type="folder",
                detected_category="DOUBLE_SUFFIX_FOLDER",
                suggested_action="MOVE_TO_ARCHIVE",
                suggested_new_path=str(archive_target.parent),
                suggested_new_name=item.name,
                reason="Folder name contains duplicate url_patched suffix",
                confidence="HIGH",
                risk_level=risk,
                notes=f"Contains {file_count} file(s). Consider archiving or renaming.",
                requires_manual_check="YES" if file_count > 0 else "NO",
            ))


def detect_unexpected_folders(queue_dir: Path, queue: str, rows: list[dict]) -> None:
    known = EXPECTED_FOLDERS | OPTIONAL_FOLDERS
    for item in queue_dir.iterdir():
        if not item.is_dir():
            continue
        if item.name not in known and not _DOUBLE_SUFFIX_RE.search(item.name):
            file_count = sum(1 for _ in item.rglob("*") if _.is_file())
            rows.append(_plan_row(
                queue=queue,
                current_path=str(item.parent),
                current_name=item.name,
                item_type="folder",
                detected_category="UNEXPECTED_FOLDER",
                suggested_action="REVIEW_MANUALLY",
                suggested_new_path="",
                suggested_new_name="",
                reason="Folder not in expected queue structure",
                confidence="MEDIUM",
                risk_level="LOW",
                notes=f"Contains {file_count} file(s). Verify purpose.",
                requires_manual_check="YES",
            ))


def detect_generic_enriched(
    folder: Path, queue: str, queue_dir: Path, rows: list[dict]
) -> int:
    count = 0
    enriched_dir = queue_dir / "02_lead_prioritized"
    if not enriched_dir.exists():
        return 0
    for f in enriched_dir.glob("*.xlsx"):
        if f.name.startswith("~$"):
            continue
        if _GENERIC_ENRICHED_RE.match(f.name):
            count += 1
            rows.append(_plan_row(
                queue=queue,
                current_path=str(f.parent),
                current_name=f.name,
                item_type="file",
                detected_category="GENERIC_ENRICHED_FILENAME",
                suggested_action="RENAME_WITH_BATCH_PREFIX",
                suggested_new_path=str(f.parent),
                suggested_new_name=f"[MANUAL] {queue}_<batch>_R<start>_<end>_enriched_{_ts()}.xlsx",
                reason="Generic enriched filename not following queue naming convention",
                confidence="LOW",
                risk_level="MEDIUM",
                notes=(
                    "Open workbook to read first company_name row, "
                    "compare with cleaned/raw files to infer batch range"
                ),
                requires_manual_check="YES",
            ))
    return count


def detect_root_loose_files(
    target_root: Path, rows: list[dict], include_root: bool
) -> None:
    if not include_root:
        return
    archive_base = target_root / "_archive" / f"manual_working_files_{_ts()}"
    for f in target_root.iterdir():
        if not f.is_file():
            continue
        if f.suffix.lower() not in {".xlsx", ".xls", ".csv", ".txt"}:
            continue
        rows.append(_plan_row(
            queue="(root)",
            current_path=str(f.parent),
            current_name=f.name,
            item_type="file",
            detected_category="LOOSE_ROOT_FILE",
            suggested_action="MOVE_TO_ARCHIVE",
            suggested_new_path=str(archive_base),
            suggested_new_name=f.name,
            reason="Data file found in root folder, not inside a queue subfolder",
            confidence="MEDIUM",
            risk_level="LOW",
            notes="Verify this is not a register or active source file before archiving",
            requires_manual_check="YES",
        ))


def detect_scripts_in_data_folders(
    queue_dir: Path, queue: str, rows: list[dict]
) -> None:
    for f in queue_dir.rglob("*"):
        if not f.is_file():
            continue
        if f.suffix.lower() not in SCRIPT_EXTENSIONS:
            continue
        rows.append(_plan_row(
            queue=queue,
            current_path=str(f.parent),
            current_name=f.name,
            item_type="file",
            detected_category="SCRIPT_IN_DATA_FOLDER",
            suggested_action="MOVE_TO_REPO_ROOT_OR_ARCHIVE",
            suggested_new_path="<repo_root> or _archive\\scripts",
            suggested_new_name=f.name,
            reason="Script file found inside a data queue folder",
            confidence="HIGH",
            risk_level="LOW",
            notes="Move to repo root if reusable; otherwise archive",
            requires_manual_check="YES",
        ))


def detect_org_folders(queue_dir: Path, queue: str, rows: list[dict]) -> None:
    for item in queue_dir.rglob("org"):
        if not item.is_dir():
            continue
        file_count = sum(1 for _ in item.rglob("*") if _.is_file())
        archive_target = queue_dir / "_archive" / f"old_{item.parent.name}_org"
        rows.append(_plan_row(
            queue=queue,
            current_path=str(item.parent),
            current_name=item.name,
            item_type="folder",
            detected_category="OLD_ORG_FOLDER",
            suggested_action="MOVE_TO_ARCHIVE",
            suggested_new_path=str(archive_target),
            suggested_new_name="org",
            reason="Legacy 'org' subfolder inside a queue data folder",
            confidence="MEDIUM",
            risk_level="LOW",
            notes=f"Contains {file_count} file(s). Low urgency – review before archiving.",
            requires_manual_check="YES",
        ))


# ---------------------------------------------------------------------------
# Missing batch detection
# ---------------------------------------------------------------------------

def _collect_ranges(folder: Path) -> set[tuple[int, int]]:
    if not folder.exists():
        return set()
    ranges: set[tuple[int, int]] = set()
    for f in folder.glob("*.xlsx"):
        if f.name.startswith("~$"):
            continue
        r = _parse_range(f.name)
        if r:
            ranges.add(r)
    return ranges


def detect_missing_batches(queue_dir: Path, queue: str, rows: list[dict]) -> None:
    raw_ranges = _collect_ranges(queue_dir / "00_raw")
    cleaned_ranges = _collect_ranges(queue_dir / "01_cleaned_domains")
    enriched_ranges = _collect_ranges(queue_dir / "02_lead_prioritized")

    for r in sorted(raw_ranges):
        label = f"R{r[0]:04d}_{r[1]:04d}"
        if r not in cleaned_ranges:
            rows.append(_plan_row(
                queue=queue,
                current_path=str(queue_dir / "00_raw"),
                current_name=label,
                item_type="batch_range",
                detected_category="MISSING_CLEANED_BATCH",
                suggested_action="NO_ACTION_REPORT_ONLY",
                reason=f"Raw range {label} has no matching cleaned file",
                confidence="MEDIUM",
                risk_level="LOW",
                notes="Run cleaner for this batch",
                requires_manual_check="YES",
            ))
        if r not in enriched_ranges:
            rows.append(_plan_row(
                queue=queue,
                current_path=str(queue_dir / "01_cleaned_domains"),
                current_name=label,
                item_type="batch_range",
                detected_category="MISSING_ENRICHED_BATCH",
                suggested_action="NO_ACTION_REPORT_ONLY",
                reason=f"Raw range {label} has no matching enriched file",
                confidence="MEDIUM",
                risk_level="LOW",
                notes="Run enricher for this batch",
                requires_manual_check="YES",
            ))


# ---------------------------------------------------------------------------
# Report writer
# ---------------------------------------------------------------------------

def _write_xlsx_plan(rows: list[dict], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "File Organization Plan"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    approve_fill = PatternFill("solid", fgColor="FFF2CC")  # yellow for approve col

    ws.append(PLAN_FIELDS)
    for col_idx, header in enumerate(PLAN_FIELDS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([row.get(f, "") for f in PLAN_FIELDS])

    # Highlight approve column
    for row_idx in range(2, len(rows) + 2):
        ws.cell(row=row_idx, column=1).fill = approve_fill
        ws.cell(row=row_idx, column=1).font = Font(bold=True)

    # Column widths
    col_widths = {
        "approve": 10,
        "queue": 12,
        "current_path": 55,
        "current_name": 50,
        "item_type": 12,
        "detected_category": 30,
        "suggested_action": 30,
        "suggested_new_path": 55,
        "suggested_new_name": 45,
        "reason": 50,
        "confidence": 12,
        "risk_level": 12,
        "batch_id_detected": 15,
        "range_detected": 15,
        "related_file": 40,
        "notes": 55,
        "requires_manual_check": 22,
    }
    for col_idx, field in enumerate(PLAN_FIELDS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(field, 20)

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)


def _write_csv_plan(rows: list[dict], path: Path) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=PLAN_FIELDS, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in PLAN_FIELDS})


def _write_apply_log(log_rows: list[dict], path: Path) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=LOG_FIELDS, extrasaction="ignore")
        w.writeheader()
        for r in log_rows:
            w.writerow({k: r.get(k, "") for k in LOG_FIELDS})


# ---------------------------------------------------------------------------
# Apply mode
# ---------------------------------------------------------------------------

def _read_plan(plan_path: Path) -> list[dict]:
    suffix = plan_path.suffix.lower()
    rows: list[dict] = []
    if suffix == ".xlsx":
        wb, err = _safe_open_wb(plan_path)
        if wb is None:
            sys.exit(f"Cannot open plan file: {err}")
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))})
        wb.close()
    elif suffix == ".csv":
        with open(plan_path, newline="", encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f))
    else:
        sys.exit(f"Unsupported plan file format: {suffix}")
    return rows


SUPPORTED_ACTIONS = {
    "MOVE_TO_ARCHIVE",
    "MOVE_TO_ARCHIVE_CORRUPT",
    "MOVE_TO_REPO_ROOT_OR_ARCHIVE",
    "RENAME_WITH_BATCH_PREFIX",
    "DELETE_LOCKFILE",
}


def apply_plan(
    plan_path: Path,
    report_dir: Path,
    allow_delete: bool,
) -> None:
    rows = _read_plan(plan_path)
    log_rows: list[dict] = []
    approved = [r for r in rows if str(r.get("approve", "")).strip().upper() == "YES"]

    print(f"\n[APPLY] Plan rows total    : {len(rows)}")
    print(f"        Approved rows      : {len(approved)}")

    completed = skipped = errors = 0

    for r in approved:
        action = str(r.get("suggested_action", "")).strip()
        src = Path(str(r.get("current_path", "")).strip()) / str(r.get("current_name", "")).strip()
        dst_dir = str(r.get("suggested_new_path", "")).strip()
        dst_name = str(r.get("suggested_new_name", "")).strip()

        log: dict = {
            "timestamp": datetime.now().isoformat(),
            "approved_action": action,
            "old_path": str(src),
            "new_path": "",
            "status": "",
            "error": "",
        }

        # Guard: source must exist
        if not src.exists():
            log["status"] = "SKIPPED_SRC_MISSING"
            log["error"] = "Source path no longer exists"
            print(f"  [SKIP] {src.name} — source missing")
            skipped += 1
            log_rows.append(log)
            continue

        # Guard: no-delete safety
        if action == "DELETE_LOCKFILE":
            if not allow_delete:
                log["status"] = "SKIPPED_NO_DELETE"
                log["error"] = "--allow-delete not passed; skipping delete"
                print(f"  [SKIP] {src.name} — delete skipped (use --allow-delete)")
                skipped += 1
                log_rows.append(log)
                continue
            # Actually delete
            try:
                src.unlink()
                log["status"] = "DELETED"
                print(f"  [DEL ] {src.name}")
                completed += 1
            except OSError as exc:
                log["status"] = "ERROR"
                log["error"] = str(exc)
                print(f"  [ERR ] {src.name} — {exc}")
                errors += 1
            log_rows.append(log)
            continue

        # Guard: unsupported / review-only actions
        if action not in SUPPORTED_ACTIONS or action in (
            "NO_ACTION_REPORT_ONLY",
            "REVIEW_MANUALLY",
            "MOVE_TO_REPO_ROOT_OR_ARCHIVE",
        ):
            log["status"] = "SKIPPED_MANUAL_ACTION"
            log["error"] = f"Action '{action}' requires manual execution"
            print(f"  [SKIP] {src.name} — manual action: {action}")
            skipped += 1
            log_rows.append(log)
            continue

        # Move / rename
        if not dst_dir or not dst_name or dst_name.startswith("[MANUAL]"):
            log["status"] = "SKIPPED_NO_TARGET"
            log["error"] = "No valid target path/name in plan row"
            print(f"  [SKIP] {src.name} — no target path")
            skipped += 1
            log_rows.append(log)
            continue

        dst = _safe_new_path(Path(dst_dir) / dst_name)
        log["new_path"] = str(dst)
        try:
            dst.parent.mkdir(parents=True, exist_ok=True)
            shutil.move(str(src), str(dst))
            log["status"] = "MOVED"
            print(f"  [MOVE] {src.name} → {dst}")
            completed += 1
        except OSError as exc:
            log["status"] = "ERROR"
            log["error"] = str(exc)
            print(f"  [ERR ] {src.name} — {exc}")
            errors += 1

        log_rows.append(log)

    # Write apply log
    ts = _ts()
    log_path = report_dir / f"file_org_apply_log_{ts}.csv"
    report_dir.mkdir(parents=True, exist_ok=True)
    _write_apply_log(log_rows, log_path)

    print(f"\n[APPLY] Completed : {completed}")
    print(f"        Skipped   : {skipped}")
    print(f"        Errors    : {errors}")
    print(f"        Log       : {log_path}\n")


# ---------------------------------------------------------------------------
# Dry-run / audit
# ---------------------------------------------------------------------------

def run_audit(
    target_root: Path,
    queues: list[str],
    report_dir: Path,
    include_root_files: bool,
) -> list[dict]:
    all_rows: list[dict] = []

    # Root-level loose files
    detect_root_loose_files(target_root, all_rows, include_root_files)

    for queue in queues:
        queue_dir = target_root / queue
        if not queue_dir.exists():
            print(f"[WARN] Queue folder not found, skipping: {queue_dir}")
            continue

        print(f"\nScanning queue: {queue}")

        detect_unexpected_folders(queue_dir, queue, all_rows)
        detect_double_suffix_folders(queue_dir, queue, all_rows)
        detect_lockfiles(queue_dir, queue, all_rows)
        detect_scripts_in_data_folders(queue_dir, queue, all_rows)
        detect_org_folders(queue_dir, queue, all_rows)
        detect_generic_enriched(queue_dir / "02_lead_prioritized", queue, queue_dir, all_rows)
        detect_missing_batches(queue_dir, queue, all_rows)

        # Corrupt workbook scan (can be slow on large folders)
        print(f"  Scanning for corrupt workbooks in {queue}…")
        corrupt_count = detect_corrupt_workbooks(queue_dir, queue, all_rows)
        print(f"  Corrupt workbooks found: {corrupt_count}")

    return all_rows


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Safe file-organization audit for Myngle data folders.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("--target-root", required=True,
                   help="Root folder, e.g. C:\\Users\\...\\Myngle")
    p.add_argument("--queues", nargs="+", default=["Italy50", "Italy100", "Italy200"],
                   help="Queue folder names to scan")
    p.add_argument("--dry-run", action="store_true",
                   help="Scan and write plan file (no changes)")
    p.add_argument("--apply-plan",
                   help="Path to approved plan .xlsx or .csv to execute")
    p.add_argument("--report-dir", default=None,
                   help="Directory for plan and log files "
                        "(default: <target-root>\\_file_organization_reports)")
    p.add_argument("--interactive", action="store_true", default=False,
                   help="(reserved) Not yet implemented")
    p.add_argument("--max-depth", type=int, default=6,
                   help="Maximum folder scan depth (default: 6)")
    p.add_argument("--include-root-files", action="store_true", default=True,
                   help="Include loose files in target-root (default: true)")
    p.add_argument("--no-delete", action="store_true", default=True,
                   help="(informational) Deletes are skipped unless --allow-delete is passed")
    p.add_argument("--allow-delete", action="store_true", default=False,
                   help="Allow DELETE_LOCKFILE actions in apply mode")
    return p


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    target_root = Path(args.target_root)
    if not target_root.exists():
        sys.exit(f"Target root not found: {target_root}")

    report_dir = (
        Path(args.report_dir) if args.report_dir
        else target_root / "_file_organization_reports"
    )

    # -----------------------------------------------------------------------
    # Apply mode
    # -----------------------------------------------------------------------
    if args.apply_plan:
        plan_path = Path(args.apply_plan)
        if not plan_path.exists():
            sys.exit(f"Plan file not found: {plan_path}")
        apply_plan(plan_path, report_dir, allow_delete=args.allow_delete)
        return

    # -----------------------------------------------------------------------
    # Dry-run mode (default)
    # -----------------------------------------------------------------------
    if not args.dry_run:
        print("Neither --dry-run nor --apply-plan specified. Running dry-run.")

    print(f"\n[DRY-RUN] organize_myngle_files_audit.py")
    print(f"  Target root : {target_root}")
    print(f"  Queues      : {args.queues}")

    all_rows = run_audit(
        target_root=target_root,
        queues=args.queues,
        report_dir=report_dir,
        include_root_files=args.include_root_files,
    )

    # Write plan
    report_dir.mkdir(parents=True, exist_ok=True)
    ts = _ts()
    xlsx_path = report_dir / f"file_organization_plan_{ts}.xlsx"
    csv_path = report_dir / f"file_organization_plan_{ts}.csv"

    _write_xlsx_plan(all_rows, xlsx_path)
    _write_csv_plan(all_rows, csv_path)

    # Summary
    def _count(cat: str) -> int:
        return sum(1 for r in all_rows if r.get("detected_category") == cat)

    cats: dict[str, int] = {}
    for r in all_rows:
        c = r.get("detected_category", "")
        cats[c] = cats.get(c, 0) + 1

    print("\n" + "=" * 62)
    print("AUDIT SUMMARY")
    print("=" * 62)
    print(f"  Queues scanned               : {len(args.queues)}")
    print(f"  Total plan rows              : {len(all_rows)}")
    print(f"  Corrupt workbooks            : {_count('CORRUPT_WORKBOOK')}")
    print(f"  Generic enriched filenames   : {_count('GENERIC_ENRICHED_FILENAME')}")
    print(f"  Lock/temp files              : {_count('LOCKFILE_TEMP')}")
    print(f"  Double-suffix folders        : {_count('DOUBLE_SUFFIX_FOLDER')}")
    print(f"  Unexpected folders           : {_count('UNEXPECTED_FOLDER')}")
    print(f"  Scripts in data folders      : {_count('SCRIPT_IN_DATA_FOLDER')}")
    print(f"  Old org folders              : {_count('OLD_ORG_FOLDER')}")
    print(f"  Loose root files             : {_count('LOOSE_ROOT_FILE')}")
    print(f"  Missing cleaned batches      : {_count('MISSING_CLEANED_BATCH')}")
    print(f"  Missing enriched batches     : {_count('MISSING_ENRICHED_BATCH')}")
    print()
    print(f"  Plan (Excel) : {xlsx_path}")
    print(f"  Plan (CSV)   : {csv_path}")
    print("=" * 62)
    print("\nNext step:")
    print("  1. Open the Excel plan file.")
    print("  2. Set approve=YES for rows you want to execute.")
    print("  3. Run:")
    print(f"     python organize_myngle_files_audit.py "
          f"--target-root \"{target_root}\" "
          f"--apply-plan \"{xlsx_path}\"\n")


if __name__ == "__main__":
    main()
