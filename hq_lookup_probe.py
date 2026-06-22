"""
hq_lookup_probe.py

Standalone HQ-detection probe for a small batch of companies.

Given an Excel (or CSV) input with company names and domains, runs
neutral headquarters-focused Serper searches, extracts the HQ location
deterministically from snippets, and optionally uses a Claude model for
harder cases.

Outputs a reviewable Excel workbook with old enrichment columns alongside
new probe columns so the two approaches can be compared side-by-side.

Usage:
    python hq_lookup_probe.py \\
        --input  "path\\to\\input.xlsx" \\
        --output "path\\to\\hq_lookup_probe_results.xlsx" \\
        --serper-key  sk-... \\
        --limit 50

    # with model extraction:
    python hq_lookup_probe.py \\
        --input  "path\\to\\input.xlsx" \\
        --use-model \\
        --anthropic-key sk-ant-...

Environment variables (all optional):
    SERPER_API_KEY   or  SERPER_KEY
    ANTHROPIC_API_KEY
"""

from __future__ import annotations

import argparse
import io
import json
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Dependency check
# ---------------------------------------------------------------------------
try:
    import requests
except ImportError:
    sys.exit("requests is required:  pip install requests")

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_SERPER_URL = "https://google.serper.dev/search"
_REQUEST_TIMEOUT = 12  # seconds
_INTER_REQUEST_SLEEP = 0.5  # seconds between Serper calls

DEFAULT_COMPANY_COL = "company_name"
DEFAULT_DOMAIN_COL  = "domain"
DEFAULT_COUNTRY_COL = "input_country"
DEFAULT_INPUT_COUNTRY = "Italy"
DEFAULT_LIMIT = 50

# Old enrichment columns to carry through if present in input
OLD_ENRICHMENT_COLS = [
    "sig_foreign_hq_score",
    "sig_foreign_hq_evidence",
    "foreign_hq_sanitized",
    "foreign_hq_sanitizer_reason",
    "foreign_hq_original_score",
    "final_commercial_fit_score",
    "commercial_fit_score",
]

# Output probe columns (in order)
PROBE_COLS = [
    "hq_detected_city",
    "hq_detected_region",
    "hq_detected_country",
    "hq_confidence",
    "foreign_hq_simple",
    "input_country_used",
    "needs_manual_review",
    "hq_reason",
    "hq_evidence_url",
    "hq_evidence_quote",
    "hq_query_used",
    "serper_knowledge_graph_location",
    "serper_answer_box",
    "top_organic_title_1",
    "top_organic_snippet_1",
    "top_organic_url_1",
    "top_organic_title_2",
    "top_organic_snippet_2",
    "top_organic_url_2",
    "top_organic_title_3",
    "top_organic_snippet_3",
    "top_organic_url_3",
    "probe_error",
]

# ---------------------------------------------------------------------------
# Known city / country aliases for deterministic extraction
# ---------------------------------------------------------------------------

# Maps normalised lowercase alias -> (city_display, country)
_ITALY_CITIES: dict[str, tuple[str, str]] = {
    "milan":          ("Milan",          "Italy"),
    "milano":         ("Milano",         "Italy"),
    "rome":           ("Rome",           "Italy"),
    "roma":           ("Roma",           "Italy"),
    "turin":          ("Turin",          "Italy"),
    "torino":         ("Torino",         "Italy"),
    "naples":         ("Naples",         "Italy"),
    "napoli":         ("Napoli",         "Italy"),
    "bologna":        ("Bologna",        "Italy"),
    "bergamo":        ("Bergamo",        "Italy"),
    "brescia":        ("Brescia",        "Italy"),
    "verona":         ("Verona",         "Italy"),
    "padova":         ("Padova",         "Italy"),
    "padua":          ("Padua",          "Italy"),
    "vicenza":        ("Vicenza",        "Italy"),
    "treviso":        ("Treviso",        "Italy"),
    "modena":         ("Modena",         "Italy"),
    "parma":          ("Parma",          "Italy"),
    "firenze":        ("Firenze",        "Italy"),
    "florence":       ("Florence",       "Italy"),
    "genova":         ("Genova",         "Italy"),
    "genoa":          ("Genoa",          "Italy"),
    "venice":         ("Venice",         "Italy"),
    "venezia":        ("Venezia",        "Italy"),
    "trieste":        ("Trieste",        "Italy"),
    "bari":           ("Bari",           "Italy"),
    "catania":        ("Catania",        "Italy"),
    "palermo":        ("Palermo",        "Italy"),
    "perugia":        ("Perugia",        "Italy"),
    "ancona":         ("Ancona",         "Italy"),
    "trento":         ("Trento",         "Italy"),
    "bolzano":        ("Bolzano",        "Italy"),
    "reggio":         ("Reggio",         "Italy"),
    "cagliari":       ("Cagliari",       "Italy"),
    "lecce":          ("Lecce",          "Italy"),
    "pescara":        ("Pescara",        "Italy"),
    "pisa":           ("Pisa",           "Italy"),
    "siena":          ("Siena",          "Italy"),
    "udine":          ("Udine",          "Italy"),
    "novara":         ("Novara",         "Italy"),
    "monza":          ("Monza",          "Italy"),
    # Additional municipalities
    "bentivoglio":    ("Bentivoglio",    "Italy"),
    "interporto":     ("Interporto Bologna", "Italy"),
    "castel maggiore":("Castel Maggiore","Italy"),
    "calderara":      ("Calderara di Reno", "Italy"),
    "imola":          ("Imola",          "Italy"),
    "faenza":         ("Faenza",         "Italy"),
    "rimini":         ("Rimini",         "Italy"),
    "ravenna":        ("Ravenna",        "Italy"),
    "ferrara":        ("Ferrara",        "Italy"),
    "forlì":          ("Forlì",          "Italy"),
    "forli":          ("Forlì",          "Italy"),
    "piacenza":       ("Piacenza",       "Italy"),
    "reggio emilia":  ("Reggio Emilia",  "Italy"),
    "mantova":        ("Mantova",        "Italy"),
    "mantua":         ("Mantua",         "Italy"),
    "cremona":        ("Cremona",        "Italy"),
    "como":           ("Como",           "Italy"),
    "varese":         ("Varese",         "Italy"),
    "lecco":          ("Lecco",          "Italy"),
    "pavia":          ("Pavia",          "Italy"),
    "lodi":           ("Lodi",           "Italy"),
    "sesto san giovanni": ("Sesto San Giovanni", "Italy"),
    "cinisello balsamo":  ("Cinisello Balsamo",  "Italy"),
    "busto arsizio":  ("Busto Arsizio",  "Italy"),
    "gallarate":      ("Gallarate",      "Italy"),
    "saronno":        ("Saronno",        "Italy"),
    "rho":            ("Rho",            "Italy"),
    "segrate":        ("Segrate",        "Italy"),
    "assago":         ("Assago",         "Italy"),
    "cernusco":       ("Cernusco sul Naviglio", "Italy"),
    "agrate brianza": ("Agrate Brianza", "Italy"),
    "vimercate":      ("Vimercate",      "Italy"),
    "cassina de' pecchi": ("Cassina de' Pecchi", "Italy"),
    "sesto fiorentino": ("Sesto Fiorentino", "Italy"),
    "prato":          ("Prato",          "Italy"),
    "livorno":        ("Livorno",        "Italy"),
    "lucca":          ("Lucca",          "Italy"),
    "pistoia":        ("Pistoia",        "Italy"),
    "arezzo":         ("Arezzo",         "Italy"),
    "grosseto":       ("Grosseto",       "Italy"),
    "la spezia":      ("La Spezia",      "Italy"),
    "savona":         ("Savona",         "Italy"),
    "imperia":        ("Imperia",        "Italy"),
    "salerno":        ("Salerno",        "Italy"),
    "caserta":        ("Caserta",        "Italy"),
    "foggia":         ("Foggia",         "Italy"),
    "taranto":        ("Taranto",        "Italy"),
    "brindisi":       ("Brindisi",       "Italy"),
    "cosenza":        ("Cosenza",        "Italy"),
    "catanzaro":      ("Catanzaro",      "Italy"),
    "reggio calabria": ("Reggio Calabria", "Italy"),
    "messina":        ("Messina",        "Italy"),
    "agrigento":      ("Agrigento",      "Italy"),
    "ragusa":         ("Ragusa",         "Italy"),
    "siracusa":       ("Siracusa",       "Italy"),
    "trapani":        ("Trapani",        "Italy"),
    "sassari":        ("Sassari",        "Italy"),
    "nuoro":          ("Nuoro",          "Italy"),
    "oristano":       ("Oristano",       "Italy"),
    "macerata":       ("Macerata",       "Italy"),
    "pesaro":         ("Pesaro",         "Italy"),
    "ascoli piceno":  ("Ascoli Piceno",  "Italy"),
    "teramo":         ("Teramo",         "Italy"),
    "chieti":         ("Chieti",         "Italy"),
    "campobasso":     ("Campobasso",     "Italy"),
    "potenza":        ("Potenza",        "Italy"),
    "matera":         ("Matera",         "Italy"),
}

# International city aliases (city_lower -> (city_display, country))
_INTL_CITIES: dict[str, tuple[str, str]] = {
    # Germany
    "berlin": ("Berlin", "Germany"), "munich": ("Munich", "Germany"),
    "münchen": ("München", "Germany"), "hamburg": ("Hamburg", "Germany"),
    "frankfurt": ("Frankfurt", "Germany"), "cologne": ("Cologne", "Germany"),
    "köln": ("Köln", "Germany"), "düsseldorf": ("Düsseldorf", "Germany"),
    "stuttgart": ("Stuttgart", "Germany"), "dortmund": ("Dortmund", "Germany"),
    # France
    "paris": ("Paris", "France"), "lyon": ("Lyon", "France"),
    "marseille": ("Marseille", "France"), "toulouse": ("Toulouse", "France"),
    "bordeaux": ("Bordeaux", "France"), "strasbourg": ("Strasbourg", "France"),
    # Netherlands
    "amsterdam": ("Amsterdam", "Netherlands"), "rotterdam": ("Rotterdam", "Netherlands"),
    "the hague": ("The Hague", "Netherlands"), "den haag": ("Den Haag", "Netherlands"),
    "eindhoven": ("Eindhoven", "Netherlands"), "utrecht": ("Utrecht", "Netherlands"),
    # Spain
    "madrid": ("Madrid", "Spain"), "barcelona": ("Barcelona", "Spain"),
    "valencia": ("Valencia", "Spain"), "seville": ("Seville", "Spain"),
    "sevilla": ("Sevilla", "Spain"),
    # Switzerland
    "zurich": ("Zurich", "Switzerland"), "zürich": ("Zürich", "Switzerland"),
    "geneva": ("Geneva", "Switzerland"), "genève": ("Genève", "Switzerland"),
    "bern": ("Bern", "Switzerland"), "basel": ("Basel", "Switzerland"),
    # United Kingdom
    "london": ("London", "United Kingdom"), "manchester": ("Manchester", "United Kingdom"),
    "birmingham": ("Birmingham", "United Kingdom"), "edinburgh": ("Edinburgh", "United Kingdom"),
    # Austria
    "vienna": ("Vienna", "Austria"), "wien": ("Wien", "Austria"),
    "graz": ("Graz", "Austria"), "salzburg": ("Salzburg", "Austria"),
    # Belgium
    "brussels": ("Brussels", "Belgium"), "bruxelles": ("Bruxelles", "Belgium"),
    "antwerp": ("Antwerp", "Belgium"), "antwerpen": ("Antwerpen", "Belgium"),
    # US
    "new york": ("New York", "United States"), "san francisco": ("San Francisco", "United States"),
    "chicago": ("Chicago", "United States"), "boston": ("Boston", "United States"),
    "los angeles": ("Los Angeles", "United States"), "seattle": ("Seattle", "United States"),
    # Nordics
    "stockholm": ("Stockholm", "Sweden"), "oslo": ("Oslo", "Norway"),
    "copenhagen": ("Copenhagen", "Denmark"), "helsinki": ("Helsinki", "Finland"),
    # Others
    "tokyo": ("Tokyo", "Japan"), "beijing": ("Beijing", "China"),
    "shanghai": ("Shanghai", "China"), "singapore": ("Singapore", "Singapore"),
    "dublin": ("Dublin", "Ireland"), "warsaw": ("Warsaw", "Poland"),
    "lisbon": ("Lisbon", "Portugal"), "lisboa": ("Lisboa", "Portugal"),
    "luxembourg": ("Luxembourg", "Luxembourg"),
}

_COUNTRY_ALIASES: dict[str, str] = {
    "italy": "Italy", "italia": "Italy", "italian": "Italy",
    "germany": "Germany", "deutschland": "Germany", "german": "Germany",
    "france": "France", "french": "France",
    "spain": "Spain", "españa": "Spain", "spanish": "Spain",
    "netherlands": "Netherlands", "holland": "Netherlands", "dutch": "Netherlands",
    "belgium": "Belgium", "belgian": "Belgium",
    "switzerland": "Switzerland", "swiss": "Switzerland", "svizzera": "Switzerland",
    "austria": "Austria", "austrian": "Austria",
    "united kingdom": "United Kingdom", "uk": "United Kingdom",
    "great britain": "United Kingdom",
    "united states": "United States", "usa": "United States", "us": "United States",
    "america": "United States",
    "japan": "Japan", "japanese": "Japan",
    "china": "China", "chinese": "China",
    "sweden": "Sweden", "swedish": "Sweden",
    "denmark": "Denmark", "danish": "Denmark",
    "norway": "Norway", "norwegian": "Norway",
    "finland": "Finland", "finnish": "Finland",
    "portugal": "Portugal", "portuguese": "Portugal",
    "poland": "Poland", "polish": "Poland",
    "luxembourg": "Luxembourg",
    "ireland": "Ireland", "irish": "Ireland",
}

# Italian province codes → (provincial capital, "Italy")
_ITALY_PROVINCE_CODES: dict[str, tuple[str, str]] = {
    "AG": ("Agrigento", "Italy"), "AL": ("Alessandria", "Italy"),
    "AN": ("Ancona",    "Italy"), "AO": ("Aosta",       "Italy"),
    "AR": ("Arezzo",    "Italy"), "AP": ("Ascoli Piceno","Italy"),
    "AT": ("Asti",      "Italy"), "AV": ("Avellino",    "Italy"),
    "BA": ("Bari",      "Italy"), "BT": ("Barletta",    "Italy"),
    "BL": ("Belluno",   "Italy"), "BN": ("Benevento",   "Italy"),
    "BG": ("Bergamo",   "Italy"), "BI": ("Biella",      "Italy"),
    "BO": ("Bologna",   "Italy"), "BZ": ("Bolzano",     "Italy"),
    "BS": ("Brescia",   "Italy"), "BR": ("Brindisi",    "Italy"),
    "CA": ("Cagliari",  "Italy"), "CL": ("Caltanissetta","Italy"),
    "CB": ("Campobasso","Italy"), "CE": ("Caserta",     "Italy"),
    "CT": ("Catania",   "Italy"), "CZ": ("Catanzaro",   "Italy"),
    "CH": ("Chieti",    "Italy"), "CO": ("Como",        "Italy"),
    "CS": ("Cosenza",   "Italy"), "CR": ("Cremona",     "Italy"),
    "KR": ("Crotone",   "Italy"), "CN": ("Cuneo",       "Italy"),
    "EN": ("Enna",      "Italy"), "FM": ("Fermo",       "Italy"),
    "FE": ("Ferrara",   "Italy"), "FI": ("Firenze",     "Italy"),
    "FG": ("Foggia",    "Italy"), "FC": ("Forlì",       "Italy"),
    "FR": ("Frosinone", "Italy"), "GE": ("Genova",      "Italy"),
    "GO": ("Gorizia",   "Italy"), "GR": ("Grosseto",    "Italy"),
    "IM": ("Imperia",   "Italy"), "IS": ("Isernia",     "Italy"),
    "SP": ("La Spezia", "Italy"), "AQ": ("L'Aquila",    "Italy"),
    "LT": ("Latina",    "Italy"), "LE": ("Lecce",       "Italy"),
    "LC": ("Lecco",     "Italy"), "LI": ("Livorno",     "Italy"),
    "LO": ("Lodi",      "Italy"), "LU": ("Lucca",       "Italy"),
    "MC": ("Macerata",  "Italy"), "MN": ("Mantova",     "Italy"),
    "MS": ("Massa",     "Italy"), "MT": ("Matera",      "Italy"),
    "ME": ("Messina",   "Italy"), "MI": ("Milano",      "Italy"),
    "MO": ("Modena",    "Italy"), "MB": ("Monza",       "Italy"),
    "NA": ("Napoli",    "Italy"), "NO": ("Novara",      "Italy"),
    "NU": ("Nuoro",     "Italy"), "OR": ("Oristano",    "Italy"),
    "PD": ("Padova",    "Italy"), "PA": ("Palermo",     "Italy"),
    "PR": ("Parma",     "Italy"), "PV": ("Pavia",       "Italy"),
    "PG": ("Perugia",   "Italy"), "PU": ("Pesaro",      "Italy"),
    "PE": ("Pescara",   "Italy"), "PC": ("Piacenza",    "Italy"),
    "PI": ("Pisa",      "Italy"), "PT": ("Pistoia",     "Italy"),
    "PN": ("Pordenone", "Italy"), "PZ": ("Potenza",     "Italy"),
    "PO": ("Prato",     "Italy"), "RG": ("Ragusa",      "Italy"),
    "RA": ("Ravenna",   "Italy"), "RC": ("Reggio Calabria","Italy"),
    "RE": ("Reggio Emilia","Italy"),"RI": ("Rieti",     "Italy"),
    "RN": ("Rimini",    "Italy"), "RM": ("Roma",        "Italy"),
    "RO": ("Rovigo",    "Italy"), "SA": ("Salerno",     "Italy"),
    "SS": ("Sassari",   "Italy"), "SV": ("Savona",      "Italy"),
    "SI": ("Siena",     "Italy"), "SR": ("Siracusa",    "Italy"),
    "SO": ("Sondrio",   "Italy"), "TA": ("Taranto",     "Italy"),
    "TE": ("Teramo",    "Italy"), "TR": ("Terni",       "Italy"),
    "TO": ("Torino",    "Italy"), "TP": ("Trapani",     "Italy"),
    "TN": ("Trento",    "Italy"), "TV": ("Treviso",     "Italy"),
    "TS": ("Trieste",   "Italy"), "UD": ("Udine",       "Italy"),
    "VA": ("Varese",    "Italy"), "VE": ("Venezia",     "Italy"),
    "VB": ("Verbania",  "Italy"), "VC": ("Vercelli",    "Italy"),
    "VR": ("Verona",    "Italy"), "VV": ("Vibo Valentia","Italy"),
    "VI": ("Vicenza",   "Italy"), "VT": ("Viterbo",     "Italy"),
}

# Province code in parentheses or following a comma then Italy/Italia,
# e.g. "(BO)", ", BO)", ", BO Italy", ", BO Italia", ", BO\n"
_PROVINCE_CODE_RE = re.compile(
    r"(?:[\(,\s])([A-Z]{2})(?:\)|(?:\s*[-,]\s*(?:Ital(?:y|ia)))|\s*$)",
    re.MULTILINE,
)

# Postal code pattern (5-digit Italian CAP) immediately resolves to Italy
_ITALIAN_CAP_RE = re.compile(r"\b([1-9]\d{4})\b")

# Group/parent context signal — presence means the snippet is about the group,
# not necessarily the exact operating company.
_GROUP_CONTEXT_RE = re.compile(
    r"\b(?:"
    r"part\s+of|owned\s+by|subsidiary\s+of|member\s+of|division\s+of"
    r"|affiliated\s+(?:with|to)|belongs?\s+to"
    r"|parent\s+company|holding\s+company|holding\s+group"
    r"|gruppo\b|gruppo\s+\w+|appartenente\s+(?:a|al)"
    r"|fa\s+parte\s+(?:del|di)|filiale\s+di|controllata\s+da"
    r")\b",
    re.IGNORECASE,
)

# Regex patterns for deterministic HQ extraction.
# Each yields the city/country text in group 1.
_HQ_PATTERNS = [
    # English
    re.compile(
        r"headquarter(?:s|ed)\s+(?:in|at)\s+([A-Z][A-Za-zÀ-ÿ\s,]{2,40}?)(?:\.|,|\s*\b(?:with|and|since|in\s+\d))",
        re.IGNORECASE,
    ),
    re.compile(r"head\s+office\s+(?:in|at|is\s+in)\s+([A-Z][A-Za-zÀ-ÿ\s,]{2,40}?)(?:\.|,)", re.IGNORECASE),
    re.compile(r"based\s+in\s+([A-Z][A-Za-zÀ-ÿ\s,]{2,35}?)(?:\.|,|\s+(?:and|with|since))", re.IGNORECASE),
    re.compile(r"Headquarters?\s*:\s*([A-Za-zÀ-ÿ\s,]{2,50}?)(?:\n|$|\.)", re.IGNORECASE),
    re.compile(r"Head\s+[Oo]ffice\s*:\s*([A-Za-zÀ-ÿ\s,]{2,50}?)(?:\n|$|\.)", re.IGNORECASE),
    # English – additional
    re.compile(r"registered\s+office\s*(?:in|at|:)\s*([A-Za-zÀ-ÿ\s,]{2,50}?)(?:\n|$|\.|,)", re.IGNORECASE),
    re.compile(r"corporate\s+office\s*(?:in|at|:)\s*([A-Za-zÀ-ÿ\s,]{2,50}?)(?:\n|$|\.|,)", re.IGNORECASE),
    re.compile(r"principal\s+office\s*(?:in|at|:)\s*([A-Za-zÀ-ÿ\s,]{2,50}?)(?:\n|$|\.|,)", re.IGNORECASE),
    re.compile(r"registered\s+in\s+([A-Za-zÀ-ÿ\s,]{2,40}?)(?:\n|$|\.|,)", re.IGNORECASE),
    # Italian
    re.compile(r"sede\s+legale\s*[:\s]+(?:in\s+)?([A-Za-zÀ-ÿ\s,]{2,50}?)(?:\n|$|\.|,)", re.IGNORECASE),
    re.compile(r"sede\s+principale\s*[:\s]+(?:in\s+)?([A-Za-zÀ-ÿ\s,]{2,50}?)(?:\n|$|\.|,)", re.IGNORECASE),
    re.compile(r"sede\s+amministrativa\s+(?:in\s+)?([A-Za-zÀ-ÿ\s,]{2,40}?)(?:\n|$|\.|,)", re.IGNORECASE),
    re.compile(r"sede\s+operativa\s+(?:in\s+)?([A-Za-zÀ-ÿ\s,]{2,40}?)(?:\n|$|\.|,)", re.IGNORECASE),
    re.compile(r"con\s+sede\s+(?:a|in)\s+([A-Za-zÀ-ÿ\s,]{2,40}?)(?:\n|$|\.|,)", re.IGNORECASE),
    re.compile(r"ha\s+sede\s+(?:a|in)\s+([A-Za-zÀ-ÿ\s,]{2,40}?)(?:\n|$|\.|,)", re.IGNORECASE),
    re.compile(r"(?:la\s+)?(?:sua\s+)?sede\s+(?:è\s+)?(?:a|in)\s+([A-Za-zÀ-ÿ\s,]{2,40}?)(?:\n|$|\.|,)", re.IGNORECASE),
    re.compile(r"uffici\s+(?:a|in|di)\s+([A-Za-zÀ-ÿ\s,]{2,40}?)(?:\n|$|\.|,)", re.IGNORECASE),
    re.compile(r"ufficio\s+(?:a|in|di)\s+([A-Za-zÀ-ÿ\s,]{2,40}?)(?:\n|$|\.|,)", re.IGNORECASE),
    # c/o or address-style patterns
    re.compile(r"c/o\s+[A-Za-zÀ-ÿ\s]+?,\s*([A-Za-zÀ-ÿ\s,]{2,40}?)(?:\n|$|\.|,)", re.IGNORECASE),
    # "located in / located at"
    re.compile(r"located\s+(?:in|at)\s+([A-Za-zÀ-ÿ\s,]{2,40}?)(?:\n|$|\.|,)", re.IGNORECASE),
]


# ---------------------------------------------------------------------------
# Serper
# ---------------------------------------------------------------------------

def _serper_search(query: str, api_key: str) -> tuple[dict, str]:
    """Return (full_response_dict, error_str)."""
    headers = {"X-API-KEY": api_key, "Content-Type": "application/json"}
    payload = {"q": query, "num": 5}
    try:
        resp = requests.post(_SERPER_URL, headers=headers, json=payload, timeout=_REQUEST_TIMEOUT)
        resp.raise_for_status()
        return resp.json(), ""
    except requests.HTTPError as exc:
        return {}, f"HTTP {exc.response.status_code}: {exc.response.text[:120]}"
    except Exception as exc:
        return {}, str(exc)[:200]


def _build_queries(company_name: str, domain: str) -> list[str]:
    """Return ordered list of HQ-focused queries (most direct first)."""
    n = company_name.strip()
    d = (domain or "").strip().lstrip("https://").lstrip("http://").rstrip("/")
    queries = [
        f'"{n}" headquarters',
        f'"{n}" head office',
        f'"{n}" sede legale',
        f'"{n}" sede principale',
        f'"{n}" sede amministrativa',
    ]
    if d:
        queries += [
            f"site:{d} sede",
            f"site:{d} headquarters",
            f"site:{d} head office",
        ]
    return queries


# ---------------------------------------------------------------------------
# Deterministic extraction
# ---------------------------------------------------------------------------

def _resolve_city_country(text: str) -> tuple[str, str]:
    """
    Given a captured location string, try to identify a known city and country.
    Returns ("", "") if nothing recognised.
    """
    text_lc = text.lower().strip(" ,.")
    # Italy cities first (most likely for this use case)
    for alias, (city, country) in _ITALY_CITIES.items():
        if alias in text_lc:
            return city, country
    # Italian province codes in parentheses, e.g. (BO), , BO Italy
    for m in _PROVINCE_CODE_RE.finditer(text):
        code = m.group(1)
        if code in _ITALY_PROVINCE_CODES:
            city, country = _ITALY_PROVINCE_CODES[code]
            return city, country
    # Italian CAP (5-digit postal code) → Italy
    if _ITALIAN_CAP_RE.search(text):
        return "", "Italy"
    # Italian fiscal / VAT identifiers → Italy
    _ITALIAN_FISCAL_RE = re.compile(
        r"\b(p\.?\s*iva|partita\s+iva|codice\s+fiscale|c\.f\.)\b", re.IGNORECASE
    )
    if _ITALIAN_FISCAL_RE.search(text):
        return "", "Italy"
    # Italian address language → Italy
    _ITALIAN_ADDR_PHRASES = (
        "sede legale", "sede principale", "sede amministrativa", "sede operativa",
        "ufficio", "uffici", "registered office in italy", "registered in italy",
        "incorporata in italia", "societa italiana", "società italiana",
    )
    for phrase in _ITALIAN_ADDR_PHRASES:
        if phrase in text_lc:
            return "", "Italy"
    # International cities
    for alias, (city, country) in _INTL_CITIES.items():
        if alias in text_lc:
            return city, country
    # Country name/alias only
    for alias, country in _COUNTRY_ALIASES.items():
        if alias in text_lc:
            return "", country
    return "", ""


def resolve_country_from_location(location_text: str) -> tuple[str, str]:
    """
    Public helper: given free-form location/address/HQ text, return (country, city).

    Returns ("Italy", city) for any recognised Italian city, province abbreviation,
    region, postal code (CAP), or Italian address language.
    Returns (country, city) for other recognised international cities.
    Returns ("", "") when nothing is recognised.

    Examples:
        resolve_country_from_location("Milano")              -> ("Italy", "Milano")
        resolve_country_from_location("Milan")               -> ("Italy", "Milan")
        resolve_country_from_location("Bentivoglio (BO)")    -> ("Italy", "Bentivoglio")
        resolve_country_from_location("Prato, Tuscany")      -> ("Italy", "Prato")
        resolve_country_from_location("registered office in Milano, Italy")
                                                             -> ("Italy", "Milano")
    """
    city, country = _resolve_city_country(location_text)
    return country, city


def _has_group_context(text: str) -> bool:
    """Return True if the text signals parent/group context (not the exact company's own HQ)."""
    return bool(_GROUP_CONTEXT_RE.search(text))


def _italy_in_text(text: str) -> bool:
    """Return True if any Italian city, province code, CAP, or 'Italy/Italia' appears in text."""
    text_lc = text.lower()
    if "italy" in text_lc or "italia" in text_lc:
        return True
    for alias in _ITALY_CITIES:
        if alias in text_lc:
            return True
    if _PROVINCE_CODE_RE.search(text):
        code = _PROVINCE_CODE_RE.search(text).group(1)
        if code in _ITALY_PROVINCE_CODES:
            return True
    if _ITALIAN_CAP_RE.search(text):
        return True
    return False


def _extract_deterministic(
    organic: list[dict],
    kg_location: str,
    answer_box: str,
    input_country: str = "",
) -> dict[str, Any]:
    """
    Scan snippets, titles, knowledge graph, and answer box for HQ patterns.

    Group-context safety: if a snippet contains parent/group language AND the
    matched country is foreign, we do not immediately accept it.  We continue
    scanning for a clean Italy/input-country hit.  Only if no clean hit is
    found do we fall back to the group-context match (marked Low confidence).
    """
    texts: list[tuple[str, str]] = []
    if kg_location:
        texts.append((kg_location, ""))
    if answer_box:
        texts.append((answer_box, ""))
    for item in organic:
        combined = f"{item.get('title', '')} {item.get('snippet', '')}"
        texts.append((combined, item.get("link", "")))

    input_country_std = _COUNTRY_ALIASES.get(input_country.lower(), input_country)
    fallback_group_hit: dict[str, Any] = {}

    for text, url in texts:
        is_group = _has_group_context(text)
        for pat in _HQ_PATTERNS:
            m = pat.search(text)
            if not m:
                continue
            captured = m.group(1).strip(" ,.")
            city, country = _resolve_city_country(captured)
            if not city and not country:
                continue

            quote = text[:250].strip()
            candidate = {
                "hq_detected_city":    city,
                "hq_detected_country": country,
                "hq_confidence":       "High",
                "hq_reason":           f"Pattern match: '{m.group(0)[:80]}'",
                "hq_evidence_url":     url,
                "hq_evidence_quote":   quote,
            }

            # Always accept if the detected country matches the input country
            country_std = _COUNTRY_ALIASES.get(country.lower(), country)
            if country_std.lower() == input_country_std.lower():
                return candidate

            # For foreign country: reject if group context in same snippet
            if is_group:
                # Check whether Italy also appears in this snippet
                # (company may have Italian address mentioned alongside group HQ)
                if _italy_in_text(text) and input_country_std == "Italy":
                    # Something Italian is here — skip to next snippet to find cleaner hit
                    pass
                elif not fallback_group_hit:
                    # Keep as low-confidence fallback
                    candidate["hq_confidence"] = "Low"
                    candidate["hq_reason"] = (
                        "[group context detected] " + candidate["hq_reason"]
                    )
                    fallback_group_hit = candidate
                continue  # Don't accept group-context foreign hit immediately

            # Clean foreign hit (no group context)
            return candidate

    # Try a direct Italy scan of all snippets even without a pattern match
    # (handles "Bentivoglio, Bologna" in an address line without a pattern trigger)
    for text, url in texts:
        if _italy_in_text(text):
            # Find the first Italian city/province mentioned
            text_lc = text.lower()
            for alias, (city, country) in _ITALY_CITIES.items():
                if alias in text_lc:
                    quote = text[:250].strip()
                    return {
                        "hq_detected_city":    city,
                        "hq_detected_country": "Italy",
                        "hq_confidence":       "Medium",
                        "hq_reason":           f"Italian city '{city}' found in evidence (no explicit HQ pattern)",
                        "hq_evidence_url":     url,
                        "hq_evidence_quote":   quote,
                    }

    return fallback_group_hit


# ---------------------------------------------------------------------------
# Model extraction (optional)
# ---------------------------------------------------------------------------

_MODEL_EXTRACTION_PROMPT = """\
You are a headquarters-location extractor.

You will receive a JSON array of search result snippets for a specific company.
Your task is to identify the EXACT headquarters location of THIS specific company.

Rules:
- Extract headquarters of the EXACT company named, not of a parent group or subsidiary.
- Do NOT infer HQ from: international activity, branches, distributors, customers, \
training providers, competitor mentions, or group ownership.
- If the evidence says the company is PART OF a foreign group but also gives an Italian \
registered office or corporate address for the exact company itself, return Italy — \
group context is not the same as the company's own HQ.
- If the evidence is ONLY about a parent/holding group with no specific address for the \
exact company, set is_group_or_parent_only to true and confidence to "Low".
- entity_match: "Exact" if evidence explicitly names this company, "Likely" if very \
similar name, "Weak" if indirect, "No" if evidence is about a different entity entirely.
- If uncertain, set confidence to "Low" or "Unknown".

Return ONLY valid JSON, no other text:
{
  "hq_city": "",
  "hq_region": "",
  "hq_country": "",
  "confidence": "High|Medium|Low|Unknown",
  "reason": "",
  "evidence_url": "",
  "evidence_quote": "",
  "entity_match": "Exact|Likely|Weak|No",
  "is_group_or_parent_only": false
}
"""


def _model_extract(
    company_name: str,
    snippets: list[dict],
    anthropic_key: str,
    model: str = "claude-haiku-4-5-20251001",
) -> dict[str, Any]:
    """Call Claude to extract HQ from snippets. Returns parsed JSON dict or {}."""
    try:
        import anthropic as _anthropic
    except ImportError:
        return {"probe_error": "anthropic package not installed (pip install anthropic)"}

    evidence = json.dumps(snippets[:5], ensure_ascii=False)
    user_msg = f"Company: {company_name}\n\nSearch results:\n{evidence}"

    client = _anthropic.Anthropic(api_key=anthropic_key)
    try:
        resp = client.messages.create(
            model=model,
            max_tokens=400,
            messages=[{"role": "user", "content": user_msg}],
            system=_MODEL_EXTRACTION_PROMPT,
        )
        raw = resp.content[0].text.strip()
        # Strip any markdown fences
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        parsed = json.loads(raw)
        # Normalise new fields with safe defaults
        parsed.setdefault("entity_match", "")
        parsed.setdefault("is_group_or_parent_only", False)
        return parsed
    except json.JSONDecodeError as exc:
        return {"probe_error": f"Model JSON parse error: {exc}"}
    except Exception as exc:
        return {"probe_error": f"Model error: {exc}"}


# ---------------------------------------------------------------------------
# Classification
# ---------------------------------------------------------------------------

def _classify(
    hq_country: str,
    input_country: str,
    confidence: str,
    reason: str,
    organic: list[dict],
) -> tuple[str, bool]:
    """
    Returns (foreign_hq_simple, needs_manual_review).
    foreign_hq_simple: "True", "False", or "" (unknown)
    """
    if not hq_country or confidence in ("Unknown", ""):
        return "", True

    hq_norm    = hq_country.strip().lower()
    input_norm = input_country.strip().lower()

    # normalise via aliases
    hq_std    = _COUNTRY_ALIASES.get(hq_norm, hq_country)
    input_std = _COUNTRY_ALIASES.get(input_norm, input_country)

    foreign = hq_std.lower() != input_std.lower()
    foreign_str = "True" if foreign else "False"

    # needs_manual_review heuristics
    needs_review = confidence in ("Low", "Unknown")
    if not needs_review and organic:
        # Check if multiple countries appear in top snippets
        all_text = " ".join(
            f"{r.get('title','')} {r.get('snippet','')}" for r in organic[:3]
        ).lower()
        country_hits = sum(1 for alias in _COUNTRY_ALIASES if alias in all_text)
        if country_hits > 2:
            needs_review = True

    return foreign_str, needs_review


# ---------------------------------------------------------------------------
# Per-company probe
# ---------------------------------------------------------------------------

def probe_company(
    company_name: str,
    domain: str,
    input_country: str,
    serper_key: str,
    use_model: bool,
    anthropic_key: str,
    cache: dict,
) -> dict[str, Any]:
    """Run all queries for one company; return probe column dict."""
    result: dict[str, Any] = {col: "" for col in PROBE_COLS}

    if not company_name.strip():
        result["probe_error"] = "blank company name"
        return result

    queries = _build_queries(company_name, domain)

    best: dict[str, Any] = {}
    used_query = ""
    all_organic: list[dict] = []
    kg_location = ""
    answer_box_text = ""
    errors: list[str] = []

    for query in queries:
        cache_key = ("serper", query)
        if cache_key in cache:
            data, err = cache[cache_key]
        else:
            data, err = _serper_search(query, serper_key)
            cache[cache_key] = (data, err)
            time.sleep(_INTER_REQUEST_SLEEP)

        if err:
            errors.append(f"{query!r}: {err}")
            continue

        organic = data.get("organic", [])

        # Grab knowledge graph location once
        if not kg_location:
            kg = data.get("knowledgeGraph", {})
            kg_location = (
                kg.get("address", "")
                or kg.get("headquarters", "")
                or kg.get("location", "")
            )

        # Grab answer box once
        if not answer_box_text:
            ab = data.get("answerBox", {})
            answer_box_text = ab.get("answer", "") or ab.get("snippet", "")

        if not all_organic and organic:
            all_organic = organic

        extracted = _extract_deterministic(organic, kg_location, answer_box_text, input_country=input_country)
        if extracted and extracted.get("hq_detected_country"):
            best = extracted
            used_query = query
            # Keep the organic from the query that yielded the hit
            all_organic = organic
            break

    # Fall back to model extraction if no deterministic hit
    if not best.get("hq_detected_country") and use_model and anthropic_key:
        snippets = [
            {"title": r.get("title", ""), "snippet": r.get("snippet", ""), "url": r.get("link", "")}
            for r in all_organic[:5]
        ]
        model_result = _model_extract(company_name, snippets, anthropic_key)
        is_group_only = model_result.get("is_group_or_parent_only", False)
        entity_match  = model_result.get("entity_match", "")
        if model_result.get("hq_country") and not is_group_only and entity_match != "No":
            reason_prefix = f"[model entity={entity_match}] " if entity_match else "[model] "
            best = {
                "hq_detected_city":    model_result.get("hq_city", ""),
                "hq_detected_region":  model_result.get("hq_region", ""),
                "hq_detected_country": model_result.get("hq_country", ""),
                "hq_confidence":       model_result.get("confidence", ""),
                "hq_reason":           reason_prefix + model_result.get("reason", ""),
                "hq_evidence_url":     model_result.get("evidence_url", ""),
                "hq_evidence_quote":   model_result.get("evidence_quote", ""),
            }
            if not used_query and queries:
                used_query = queries[0]
        elif model_result.get("hq_country") and is_group_only:
            # Model says evidence is group-only: keep as low-confidence fallback
            best = {
                "hq_detected_city":    model_result.get("hq_city", ""),
                "hq_detected_region":  model_result.get("hq_region", ""),
                "hq_detected_country": model_result.get("hq_country", ""),
                "hq_confidence":       "Low",
                "hq_reason":           "[model – group/parent only] " + model_result.get("reason", ""),
                "hq_evidence_url":     model_result.get("evidence_url", ""),
                "hq_evidence_quote":   model_result.get("evidence_quote", ""),
            }
            if not used_query and queries:
                used_query = queries[0]
        if model_result.get("probe_error"):
            errors.append(model_result["probe_error"])

    # Populate result
    for k, v in best.items():
        if k in result:
            result[k] = v

    result["hq_query_used"] = used_query
    result["serper_knowledge_graph_location"] = kg_location
    result["serper_answer_box"] = answer_box_text

    for i, item in enumerate(all_organic[:3], start=1):
        result[f"top_organic_title_{i}"]   = item.get("title", "")
        result[f"top_organic_snippet_{i}"] = item.get("snippet", "")
        result[f"top_organic_url_{i}"]     = item.get("link", "")

    # Classification
    foreign_str, needs_review = _classify(
        result["hq_detected_country"],
        input_country,
        result["hq_confidence"],
        result["hq_reason"],
        all_organic,
    )
    result["foreign_hq_simple"]     = foreign_str
    result["input_country_used"]    = _COUNTRY_ALIASES.get(input_country.strip().lower(), input_country)
    result["needs_manual_review"]   = "Yes" if needs_review else "No"

    if errors:
        result["probe_error"] = "; ".join(errors)

    return result


# ---------------------------------------------------------------------------
# Input reading
# ---------------------------------------------------------------------------

def _read_input(
    input_path: Path,
    company_col: str,
    domain_col: str,
    country_col: str,
    limit: int,
) -> list[dict[str, Any]]:
    """Read Excel or CSV input. Returns list of row dicts."""
    suffix = input_path.suffix.lower()
    rows: list[dict[str, Any]] = []

    if suffix in (".xlsx", ".xls"):
        wb = load_workbook(input_path, read_only=True, data_only=True)
        ws = wb.active
        iter_rows = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else "" for h in next(iter_rows)]
        for raw in iter_rows:
            row = {headers[i]: raw[i] for i in range(len(headers))}
            if all(v is None or str(v).strip() == "" for v in raw):
                continue
            rows.append(row)
            if len(rows) >= limit:
                break
        wb.close()
    elif suffix == ".csv":
        import csv
        with open(input_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(dict(row))
                if len(rows) >= limit:
                    break
    else:
        sys.exit(f"Unsupported input format: {suffix}. Use .xlsx or .csv")

    return rows


# ---------------------------------------------------------------------------
# Public input helpers (used by Streamlit app and CLI alike)
# ---------------------------------------------------------------------------

def get_excel_sheet_names(fileobj: "io.IOBase | Path") -> list[str]:
    """Return sheet names from an Excel file or file-like object."""
    wb = load_workbook(fileobj, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def read_input_from_fileobj(
    fileobj: "io.IOBase",
    suffix: str,
    limit: int = DEFAULT_LIMIT,
    sheet_name: "str | None" = None,
) -> list[dict[str, Any]]:
    """
    Read Excel or CSV from a file-like object (e.g. Streamlit UploadedFile).
    Returns list of row dicts keyed by header name.
    All rows returned; caller selects company/domain/country by key.
    """
    rows: list[dict[str, Any]] = []
    suffix = suffix.lower()

    if suffix in (".xlsx", ".xls"):
        wb = load_workbook(fileobj, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        iter_rows = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else "" for h in next(iter_rows)]
        for raw in iter_rows:
            if all(v is None or str(v).strip() == "" for v in raw):
                continue
            rows.append({headers[i]: raw[i] for i in range(len(headers))})
            if len(rows) >= limit:
                break
        wb.close()
    elif suffix == ".csv":
        import csv
        text = fileobj.read()
        if isinstance(text, bytes):
            text = text.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append(dict(row))
            if len(rows) >= limit:
                break
    return rows


# ---------------------------------------------------------------------------
# Public batch runner (used by Streamlit app)
# ---------------------------------------------------------------------------

def run_probe_on_rows(
    rows: list[dict[str, Any]],
    company_col: str,
    domain_col: str,
    country_col: str,
    default_country: str,
    serper_key: str,
    use_model: bool,
    anthropic_key: str,
    model: str = "claude-haiku-4-5-20251001",
    cache: "dict | None" = None,
    progress_cb: "Any | None" = None,
) -> list[dict[str, Any]]:
    """
    Run the HQ probe on a list of row dicts.
    progress_cb: optional callable(current_idx, total) for progress reporting.
    Returns list of probe result dicts (one per input row).
    """
    if cache is None:
        cache = {}
    results: list[dict[str, Any]] = []
    total = len(rows)
    for i, row in enumerate(rows):
        company = str(row.get(company_col) or "").strip()
        domain  = str(row.get(domain_col)  or "").strip()
        country = str(row.get(country_col) or default_country).strip() or default_country
        probe = probe_company(
            company_name=company,
            domain=domain,
            input_country=country,
            serper_key=serper_key,
            use_model=use_model,
            anthropic_key=anthropic_key,
            cache=cache,
        )
        results.append(probe)
        if progress_cb:
            progress_cb(i + 1, total)
    return results


# ---------------------------------------------------------------------------
# Output writing
# ---------------------------------------------------------------------------

_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_PROBE_FILL   = PatternFill("solid", fgColor="E2EFDA")
_PROBE_FONT   = Font(bold=True)
_OLD_FILL     = PatternFill("solid", fgColor="D6E4F0")
_WARN_FILL    = PatternFill("solid", fgColor="FCE4D6")

_WRAP_COLS = {
    "hq_reason", "hq_evidence_quote",
    "top_organic_snippet_1", "top_organic_snippet_2", "top_organic_snippet_3",
    "sig_foreign_hq_evidence",
}

_COL_WIDTHS: dict[str, int] = {
    "source_row": 10, "company_name": 38, "domain": 28, "input_country": 14,
    "hq_detected_city": 18, "hq_detected_region": 18, "hq_detected_country": 18,
    "hq_confidence": 12, "foreign_hq_simple": 16, "needs_manual_review": 18,
    "hq_reason": 50, "hq_evidence_url": 45, "hq_evidence_quote": 60,
    "hq_query_used": 45, "serper_knowledge_graph_location": 35, "serper_answer_box": 45,
    "top_organic_title_1": 45, "top_organic_snippet_1": 60, "top_organic_url_1": 45,
    "top_organic_title_2": 45, "top_organic_snippet_2": 60, "top_organic_url_2": 45,
    "top_organic_title_3": 45, "top_organic_snippet_3": 60, "top_organic_url_3": 45,
    "probe_error": 40,
    "sig_foreign_hq_score": 20, "sig_foreign_hq_evidence": 55,
    "foreign_hq_sanitized": 18, "foreign_hq_sanitizer_reason": 35,
    "foreign_hq_original_score": 22, "final_commercial_fit_score": 24,
    "commercial_fit_score": 22,
}


def _build_workbook(
    input_rows: list[dict],
    probe_results: list[dict],
    present_old_cols: list[str],
    company_col: str,
    domain_col: str,
    country_col: str,
    qa_meta: dict,
    output_label: str = "",
) -> "Workbook":
    """Build and return the openpyxl Workbook. Caller decides how to save it."""
    wb = Workbook()

    # ── Sheet 1: results ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "HQ Probe Results"

    input_identity_cols = ["source_row", company_col, domain_col, country_col]
    seen: set[str] = set()
    all_cols: list[str] = []
    for c in input_identity_cols + present_old_cols + PROBE_COLS:
        if c not in seen:
            all_cols.append(c)
            seen.add(c)

    ws.append(all_cols)
    ws.row_dimensions[1].height = 18
    probe_col_set = set(PROBE_COLS)
    old_col_set   = set(present_old_cols)
    for col_idx, h in enumerate(all_cols, start=1):
        cell = ws.cell(row=1, column=col_idx)
        if h in probe_col_set:
            cell.fill = _PROBE_FILL
            cell.font = _PROBE_FONT
        elif h in old_col_set:
            cell.fill = _OLD_FILL
            cell.font = Font(bold=True)
        else:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=False)
        ws.column_dimensions[get_column_letter(col_idx)].width = _COL_WIDTHS.get(h, max(14, min(len(h) + 4, 35)))

    for row_idx, (in_row, probe) in enumerate(zip(input_rows, probe_results), start=2):
        values: list[Any] = []
        for h in all_cols:
            if h == "source_row":
                values.append(row_idx - 1)
            elif h in probe:
                values.append(probe[h])
            else:
                v = in_row.get(h)
                values.append(v if v is not None else "")
        ws.append(values)
        ws.row_dimensions[row_idx].height = 15
        if probe.get("needs_manual_review") == "Yes":
            for col_idx in range(1, len(all_cols) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = _WARN_FILL
        for col_idx, h in enumerate(all_cols, start=1):
            if h in _WRAP_COLS:
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions

    # ── Sheet 2: Run QA ─────────────────────────────────────────────────────
    ws_qa = wb.create_sheet("Run QA")

    detected_italy   = sum(
        1 for p in probe_results
        if _COUNTRY_ALIASES.get((p.get("hq_detected_country") or "").lower(),
                                 p.get("hq_detected_country", "")) == "Italy"
    )
    detected_foreign = sum(1 for p in probe_results if p.get("foreign_hq_simple") == "True")
    detected_unknown = sum(1 for p in probe_results if not p.get("hq_detected_country"))
    needs_review_cnt = sum(1 for p in probe_results if p.get("needs_manual_review") == "Yes")
    has_errors       = sum(1 for p in probe_results if p.get("probe_error"))

    qa_rows_data = [
        ("Run QA – hq_lookup_probe.py", ""),
        ("", ""),
        ("timestamp",             qa_meta.get("timestamp", "")),
        ("input_file",            qa_meta.get("input_file", "")),
        ("output_file",           output_label or qa_meta.get("output_file", "")),
        ("company_col",           company_col),
        ("domain_col",            domain_col),
        ("country_col",           country_col),
        ("rows_processed",        len(probe_results)),
        ("", ""),
        ("── detection summary ──", ""),
        ("detected_italy_hq",     detected_italy),
        ("detected_foreign_hq",   detected_foreign),
        ("detected_unknown",      detected_unknown),
        ("needs_manual_review",   needs_review_cnt),
        ("rows_with_errors",      has_errors),
        ("", ""),
        ("── settings ──", ""),
        ("model_extraction_used", qa_meta.get("use_model", False)),
        ("model_used",            qa_meta.get("model", "")),
        ("serper_available",      qa_meta.get("serper_available", False)),
        ("limit",                 qa_meta.get("limit", "")),
    ]

    ws_qa.column_dimensions["A"].width = 30
    ws_qa.column_dimensions["B"].width = 80
    for r_idx, (k, v) in enumerate(qa_rows_data, start=1):
        cell_a = ws_qa.cell(row=r_idx, column=1, value=k)
        ws_qa.cell(row=r_idx, column=2, value=v)
        if r_idx == 1:
            cell_a.font = Font(bold=True, size=13)
        elif k and not k.startswith(" "):
            cell_a.font = Font(bold=True)
        ws_qa.row_dimensions[r_idx].height = 15

    errors_present = [
        (i + 1, p.get("probe_error"))
        for i, p in enumerate(probe_results)
        if p.get("probe_error")
    ]
    if errors_present:
        r_idx = len(qa_rows_data) + 2
        ws_qa.cell(row=r_idx, column=1, value="── per-row errors ──").font = Font(bold=True)
        r_idx += 1
        for src_row, err in errors_present:
            ws_qa.cell(row=r_idx, column=1, value=f"row {src_row}")
            ws_qa.cell(row=r_idx, column=2, value=err)
            r_idx += 1

    return wb


def _write_output(
    output_path: Path,
    input_rows: list[dict],
    probe_results: list[dict],
    present_old_cols: list[str],
    company_col: str,
    domain_col: str,
    country_col: str,
    qa_meta: dict,
) -> None:
    wb = _build_workbook(
        input_rows=input_rows,
        probe_results=probe_results,
        present_old_cols=present_old_cols,
        company_col=company_col,
        domain_col=domain_col,
        country_col=country_col,
        qa_meta=qa_meta,
        output_label=str(output_path),
    )
    wb.save(output_path)


def build_excel_bytes(
    input_rows: list[dict],
    probe_results: list[dict],
    present_old_cols: list[str],
    company_col: str,
    domain_col: str,
    country_col: str,
    qa_meta: dict,
) -> bytes:
    """Build the output workbook in memory and return raw bytes (for Streamlit download)."""
    wb = _build_workbook(
        input_rows=input_rows,
        probe_results=probe_results,
        present_old_cols=present_old_cols,
        company_col=company_col,
        domain_col=domain_col,
        country_col=country_col,
        qa_meta=qa_meta,
        output_label="(in-memory)",
    )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Safe output path
# ---------------------------------------------------------------------------

def _safe_output_path(path: Path, overwrite: bool) -> Path:
    if not path.exists() or overwrite:
        return path
    stem, suffix, parent = path.stem, path.suffix, path.parent
    n = 2
    while True:
        candidate = parent / f"{stem}_{n}{suffix}"
        if not candidate.exists():
            return candidate
        n += 1


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description=(
            "HQ lookup probe — neutral headquarters detection for a batch of companies.\n"
            "Outputs a reviewable Excel workbook comparing old enrichment signals with\n"
            "new simple HQ detection."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("--input",  default=None,   help="Input Excel (.xlsx) or CSV file")
    p.add_argument("--output", default=None,   help="Output Excel path (default: <input>_hq_probe_results.xlsx)")
    p.add_argument("--company-col",  default=DEFAULT_COMPANY_COL, help=f"Column name for company name (default: {DEFAULT_COMPANY_COL})")
    p.add_argument("--domain-col",   default=DEFAULT_DOMAIN_COL,  help=f"Column name for domain (default: {DEFAULT_DOMAIN_COL})")
    p.add_argument("--country-col",  default=DEFAULT_COUNTRY_COL, help=f"Column name for input country (default: {DEFAULT_COUNTRY_COL}). If missing from input, uses --default-country.")
    p.add_argument("--default-country", default=DEFAULT_INPUT_COUNTRY, help=f"Fallback country when column is absent (default: {DEFAULT_INPUT_COUNTRY})")
    p.add_argument("--limit",  type=int, default=DEFAULT_LIMIT, help=f"Max rows to process (default: {DEFAULT_LIMIT})")
    p.add_argument("--serper-key", default=None, help="Serper API key. Also read from SERPER_API_KEY or SERPER_KEY env vars.")
    p.add_argument("--use-model", action="store_true", default=False, help="Use Claude model for cases not resolved by deterministic extraction.")
    p.add_argument("--anthropic-key", default=None, help="Anthropic API key. Also read from ANTHROPIC_API_KEY env var.")
    p.add_argument("--model", default="claude-haiku-4-5-20251001", help="Claude model to use when --use-model is set.")
    p.add_argument("--overwrite", action="store_true", default=False, help="Overwrite existing output file.")
    p.add_argument("--dry-run", action="store_true", default=False, help="Read input and show what would be searched; do not call Serper.")
    p.add_argument("--self-test", action="store_true", default=False, help="Run a tiny self-test with no network calls.")
    return p


def _self_test() -> None:
    """Minimal offline self-test of deterministic extraction and classification."""
    print("[self-test] Running offline extraction tests...")

    tests = [
        ("Acme srl is headquartered in Milan, Italy with 200 employees.", "Milan", "Italy"),
        ("Con sede a Roma, la società opera in tutta Europa.", "Roma", "Italy"),
        ("Sede legale: Torino. Uffici in Germania e Francia.", "Torino", "Italy"),
        ("Head office: Frankfurt, Germany.", "", "Germany"),
        ("Based in Amsterdam, the company serves clients worldwide.", "", "Netherlands"),
    ]

    for text, expected_city, expected_country in tests:
        result = _extract_deterministic([{"title": "", "snippet": text, "link": "http://test.com"}], "", "")
        city    = result.get("hq_detected_city", "")
        country = result.get("hq_detected_country", "")
        ok = (city == expected_city or not expected_city) and country == expected_country
        status = "PASS" if ok else "FAIL"
        print(f"  [{status}] text={text[:60]!r}  → city={city!r}, country={country!r}")

    # FARMACIE PRATESI: domain phoenixpharmaitalia.it is part of German PHOENIX group,
    # but the company itself has sede/registered office in Bentivoglio (Bologna), Italy.
    # Group context snippets must NOT cause a false foreign HQ detection.
    # resolve_country_from_location tests (public API)
    print("[self-test] resolve_country_from_location tests...")
    loc_tests = [
        ("Milano",                                    "Italy"),
        ("Milan",                                     "Italy"),
        ("Bentivoglio (BO)",                          "Italy"),
        ("Prato, Tuscany",                            "Italy"),
        ("registered office in Milano, Italy",        "Italy"),
        ("sede legale in Roma",                       "Italy"),
        ("Bologna",                                   "Italy"),
        ("Bergamo",                                   "Italy"),
        ("Brescia",                                   "Italy"),
        ("Verona",                                    "Italy"),
        ("Padova",                                    "Italy"),
        ("Vicenza",                                   "Italy"),
        ("Treviso",                                   "Italy"),
        ("Modena",                                    "Italy"),
        ("Parma",                                     "Italy"),
        ("Genova",                                    "Italy"),
        ("Genoa",                                     "Italy"),
        ("Turin",                                     "Italy"),
        ("Torino",                                    "Italy"),
        ("Naples",                                    "Italy"),
        ("Rome",                                      "Italy"),
        ("P.IVA 01234567890",                         "Italy"),
        ("40010 Bentivoglio",                         "Italy"),
        ("Berlin, Germany",                           "Germany"),
        ("",                                          ""),
    ]
    loc_failures = 0
    for loc_text, exp_country in loc_tests:
        got_country, _ = resolve_country_from_location(loc_text)
        ok = got_country == exp_country
        status = "PASS" if ok else "FAIL"
        print(f"  [{status}] {loc_text!r:50s} → {got_country!r}")
        if not ok:
            loc_failures += 1
    if loc_failures:
        print(f"  {loc_failures} location test(s) FAILED")
    else:
        print("  All location tests passed.")

    print("[self-test] FARMACIE PRATESI false-foreign regression...")
    fp_snippets = [
        {
            "title": "PHOENIX group – international pharmaceutical wholesaler",
            "snippet": (
                "PHOENIX group is one of the leading pharmaceutical companies in Europe, "
                "headquartered in Mannheim, Germany. It is part of the PHOENIX group holding."
            ),
            "link": "https://www.phoenixgroup.eu/en/",
        },
        {
            "title": "Farmacie Pratesi Pratofarma S.p.A. – chi siamo",
            "snippet": (
                "Farmacie Pratesi Pratofarma S.p.A. ha sede legale a Bentivoglio (BO), "
                "Via dell'Artigianato 1, 40010 Bentivoglio, Italia."
            ),
            "link": "https://www.phoenixpharmaitalia.it/chi-siamo",
        },
    ]
    fp_result = _extract_deterministic(fp_snippets, "", "", input_country="Italy")
    fp_country = fp_result.get("hq_detected_country", "")
    fp_city    = fp_result.get("hq_detected_city", "")
    fp_conf    = fp_result.get("hq_confidence", "")
    fp_ok      = fp_country == "Italy"
    fp_status  = "PASS" if fp_ok else "FAIL"
    print(
        f"  [{fp_status}] FARMACIE PRATESI → city={fp_city!r}, country={fp_country!r}, "
        f"confidence={fp_conf!r}"
    )
    if not fp_ok:
        print(f"         reason={fp_result.get('hq_reason', '')!r}")
    # Verify classification
    fp_foreign, _ = _classify(fp_country, "Italy", fp_conf, fp_result.get("hq_reason", ""), [])
    fp_cls_ok = fp_foreign == "False"
    fp_cls_status = "PASS" if fp_cls_ok else "FAIL"
    print(f"  [{fp_cls_status}] FARMACIE PRATESI foreign_hq_simple={fp_foreign!r} (expected 'False')")

    # Classification test
    foreign, review = _classify("Germany", "Italy", "High", "found it", [])
    assert foreign == "True", f"Expected True, got {foreign}"
    assert review is False, f"Expected False, got {review}"

    foreign, review = _classify("Italy", "Italy", "High", "found it", [])
    assert foreign == "False", f"Expected False, got {foreign}"

    foreign, review = _classify("", "Italy", "Unknown", "", [])
    assert review is True

    print("[self-test] Classification tests passed.")
    print("[self-test] Done.\n")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = _build_parser()
    args = parser.parse_args()

    if args.self_test:
        _self_test()
        sys.exit(0)

    # Resolve keys
    serper_key = (
        args.serper_key
        or os.environ.get("SERPER_API_KEY", "")
        or os.environ.get("SERPER_KEY", "")
    ).strip()

    anthropic_key = (
        args.anthropic_key
        or os.environ.get("ANTHROPIC_API_KEY", "")
    ).strip()

    if not args.dry_run and not serper_key:
        sys.exit(
            "ERROR: Serper API key required.\n"
            "  Pass --serper-key, or set SERPER_API_KEY / SERPER_KEY env var.\n"
            "  Use --dry-run to test without a key."
        )

    if args.use_model and not anthropic_key:
        print("[WARN] --use-model set but no Anthropic key found. Model extraction will be skipped.")
        args.use_model = False

    if not args.input:
        sys.exit("ERROR: --input is required (unless --self-test or --dry-run with --help).")
    input_path = Path(args.input)
    if not input_path.exists():
        sys.exit(f"Input file not found: {input_path}")

    if args.output:
        raw_output = Path(args.output)
    else:
        raw_output = input_path.parent / f"{input_path.stem}_hq_probe_results.xlsx"
    output_path = _safe_output_path(raw_output, args.overwrite)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    print(f"\n[HQ-PROBE] hq_lookup_probe.py")
    print(f"  Input          : {input_path}")
    print(f"  Output         : {output_path}")
    print(f"  Limit          : {args.limit}")
    print(f"  Use model      : {args.use_model}")
    print(f"  Dry run        : {args.dry_run}")
    print(f"  Serper key     : {'set' if serper_key else 'NOT SET'}")

    # Read input
    input_rows = _read_input(
        input_path,
        args.company_col,
        args.domain_col,
        args.country_col,
        args.limit,
    )
    print(f"  Rows to process: {len(input_rows)}")

    if not input_rows:
        sys.exit("No rows found in input file.")

    # Detect which old enrichment cols are present
    sample_keys = set(input_rows[0].keys()) if input_rows else set()
    present_old_cols = [c for c in OLD_ENRICHMENT_COLS if c in sample_keys]

    # Run probes
    cache: dict = {}
    probe_results: list[dict] = []

    for i, row in enumerate(input_rows, start=1):
        company = str(row.get(args.company_col) or "").strip()
        domain  = str(row.get(args.domain_col)  or "").strip()
        country = str(row.get(args.country_col) or args.default_country).strip() or args.default_country

        print(f"  [{i:>3}/{len(input_rows)}] {company[:45]:<45} ({domain})", end="", flush=True)

        if args.dry_run:
            queries = _build_queries(company, domain)
            print(f"  → DRY RUN, would search: {queries[0]!r}")
            probe_results.append({col: "" for col in PROBE_COLS})
            probe_results[-1]["hq_query_used"] = queries[0] if queries else ""
            probe_results[-1]["probe_error"]   = "dry_run"
            continue

        probe = probe_company(
            company_name=company,
            domain=domain,
            input_country=country,
            serper_key=serper_key,
            use_model=args.use_model,
            anthropic_key=anthropic_key,
            cache=cache,
        )
        probe_results.append(probe)

        country_out = probe.get("hq_detected_country") or "?"
        conf_out    = probe.get("hq_confidence") or "?"
        review_flag = " [REVIEW]" if probe.get("needs_manual_review") == "Yes" else ""
        err_flag    = f" [ERR: {probe.get('probe_error','')[:30]}]" if probe.get("probe_error") else ""
        print(f"  → {country_out} ({conf_out}){review_flag}{err_flag}")

    if args.dry_run:
        print("\n[DRY RUN] No output file written.")
        return

    # Write output
    output_path.parent.mkdir(parents=True, exist_ok=True)
    _write_output(
        output_path=output_path,
        input_rows=input_rows,
        probe_results=probe_results,
        present_old_cols=present_old_cols,
        company_col=args.company_col,
        domain_col=args.domain_col,
        country_col=args.country_col,
        qa_meta={
            "timestamp":       ts,
            "input_file":      str(input_path),
            "use_model":       args.use_model,
            "model":           args.model if args.use_model else "",
            "serper_available": bool(serper_key),
            "limit":           args.limit,
        },
    )

    detected_italy   = sum(1 for p in probe_results if _COUNTRY_ALIASES.get((p.get("hq_detected_country") or "").lower(), p.get("hq_detected_country", "")) == "Italy")
    detected_foreign = sum(1 for p in probe_results if p.get("foreign_hq_simple") == "True")
    detected_unknown = sum(1 for p in probe_results if not p.get("hq_detected_country"))
    needs_review_cnt = sum(1 for p in probe_results if p.get("needs_manual_review") == "Yes")

    print(f"\n{'='*60}")
    print("HQ PROBE SUMMARY")
    print(f"{'='*60}")
    print(f"  Rows processed       : {len(probe_results)}")
    print(f"  Detected Italy HQ    : {detected_italy}")
    print(f"  Detected foreign HQ  : {detected_foreign}")
    print(f"  Unknown / no hit     : {detected_unknown}")
    print(f"  Needs manual review  : {needs_review_cnt}")
    print(f"  Output               : {output_path}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
