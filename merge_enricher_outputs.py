"""
merge_enricher_outputs.py
=========================
Merge multiple Client Enricher batch-output Excel files into one combined workbook.

Tabular sheets are merged row-by-row (union of columns).
Non-tabular layout sheets (Company Profiles, Scoring Settings, Run Settings)
are handled separately — see --include-company-profiles.

Usage:
    python merge_enricher_outputs.py \\
        --input-dir "C:/Users/.../Italy200/02_lead_prioritized" \\
        --output "Italy200_merged_20260615_1400.xlsx" \\
        [--pattern "*.xlsx"] \\
        [--include-company-profiles]

Sheet order in output:
  1. Tabular sheets in priority order (Opportunity Input first)
  2. Company Profiles (only if --include-company-profiles)
  3. Merge QA (always last)
"""

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Sheet classification ──────────────────────────────────────────────────────

# Sheets that are merged row-by-row as normal tables (first row = header).
TABULAR_SHEETS: list[str] = [
    "Opportunity Input",
    "Lead Scores",
    "Enriched",
    "Advanced Evidence",
    "Input",
    "model_features",
    "qa_evidence",
]

# Sheets that are NOT normal tables — merged cells, card layout, or settings blocks.
NON_TABULAR_SHEETS: list[str] = [
    "Company Profiles",
    "Scoring Settings",
    "Run Settings",
]

# Priority order for tabular sheets in the output workbook.
TABULAR_PRIORITY: list[str] = [
    "Opportunity Input",
    "Lead Scores",
    "Enriched",
    "Advanced Evidence",
    "Input",
    "model_features",
    "qa_evidence",
]

# Evidence/JSON columns that must always be written as plain text.
TEXT_COLUMNS: set[str] = {
    "raw_google_evidence_json",
    "raw_google_evidence_json_01",
    "raw_google_evidence_json_02",
    "raw_google_evidence_json_03",
    "raw_google_evidence_combined",
    "raw_google_evidence_urls",
    "raw_google_evidence_truncated",
    "serper_query_summary",
    "serper_source_urls",
    "serper_result_titles",
    "serper_snippets",
    "raw_evidence_summary",
    "evidence_source_urls",
    "source_file",
    "source_batch",
}

# Long-text columns that should wrap in Excel.
WRAP_COLUMNS: set[str] = {
    "raw_google_evidence_combined",
    "raw_google_evidence_urls",
    "serper_snippets",
    "icp_evidence",
    "icp_why_relevant",
    "icp_buying_signals",
    "icp_likely_training_interest",
    "scoring_notes",
    "caller_angle",
    "business_output_reason",
}

# Duplicate-check columns in Opportunity Input.
DUP_COL_NUMBER = "company_number"
DUP_COL_DOMAIN = "canonical_company_domain"


# =============================================================================
# FORMATTING HELPERS
# =============================================================================

def _header_fill() -> PatternFill:
    return PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")


def _header_font() -> Font:
    return Font(bold=True, color="FFFFFF", size=10)


def _col_width(col_name: str) -> int:
    if col_name in WRAP_COLUMNS:
        return 40
    if "json" in col_name.lower():
        return 20
    if "snippet" in col_name.lower() or "evidence" in col_name.lower():
        return 36
    if "url" in col_name.lower():
        return 32
    return min(max(len(str(col_name)) + 2, 12), 48)


def _write_df_to_sheet(ws, df: pd.DataFrame) -> None:
    """Write a DataFrame to an openpyxl worksheet with standard formatting."""
    hfill = _header_fill()
    hfont = _header_font()
    cols = list(df.columns)

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=str(col))
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = _col_width(col)

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, col in enumerate(cols, 1):
            val = row[col]
            if isinstance(val, float) and val != val:
                val = ""
            if col in TEXT_COLUMNS:
                val = str(val) if val not in ("", None) else ""
            cell = ws.cell(row=ri, column=ci, value=val)
            if col in WRAP_COLUMNS:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")

    ws.freeze_panes = "A2"
    if cols:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    ws.row_dimensions[1].height = 18


def _write_qa_sheet(ws, qa: dict) -> None:
    hfill = _header_fill()
    hfont = _header_font()
    bold  = Font(bold=True)

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 80

    for ci, val in enumerate(["Metric", "Value"], 1):
        c = ws.cell(row=1, column=ci, value=val)
        c.fill = hfill
        c.font = hfont
    ws.freeze_panes = "A2"

    rows: list[tuple[str, str]] = [
        ("Merge timestamp",    qa["timestamp"]),
        ("Input folder",       qa["input_dir"]),
        ("Output file",        qa["output_file"]),
        ("Input files found",  str(qa["n_files"])),
        ("", ""),
        ("Input files", ""),
    ]
    for fname in qa.get("input_files", []):
        rows.append((f"  {fname}", ""))
    rows.append(("", ""))

    rows.append(("Sheets merged", ""))
    for sname, nrows in qa.get("rows_per_sheet", {}).items():
        rows.append((f"  {sname}", f"{nrows} rows"))
    rows.append(("", ""))

    rows.append(("Sheets skipped", ""))
    for s in qa.get("sheets_skipped", []):
        rows.append((f"  {s}", ""))
    rows.append(("", ""))

    rows.append(("Rows per input file", ""))
    for fname, n in qa.get("rows_per_file", {}).items():
        rows.append((f"  {fname}", f"{n} rows"))
    rows.append(("", ""))

    rows.append(("Warnings / notes", ""))
    warnings = qa.get("warnings", [])
    for w in warnings:
        rows.append(("  ⚠", w))
    if not warnings:
        rows.append(("  —", "No warnings"))

    for ri, (k, v) in enumerate(rows, 2):
        ck = ws.cell(row=ri, column=1, value=str(k))
        cv = ws.cell(row=ri, column=2, value=str(v))
        if k and not k.startswith(" ") and k not in ("", " "):
            ck.font = bold
        cv.alignment = Alignment(wrap_text=True, vertical="top")

    ws.auto_filter.ref = "A1:B1"


# =============================================================================
# FILE HELPERS
# =============================================================================

def _derive_batch_label(filename: str) -> str:
    stem = Path(filename).stem
    cleaned = re.sub(
        r"_(lead_prioritized|enrichedResults|enriched_results)_\d{8}_?\d{4,6}$",
        "",
        stem,
        flags=re.IGNORECASE,
    )
    return cleaned if cleaned != stem else stem


def collect_files(input_dir: Path, pattern: str) -> list[Path]:
    return sorted(
        p for p in input_dir.glob(pattern)
        if p.is_file() and not p.name.startswith("~$")
    )


def classify_sheet(sheet_name: str) -> str:
    """Return 'tabular', 'non_tabular', or 'unknown'."""
    if sheet_name in TABULAR_SHEETS:
        return "tabular"
    if sheet_name in NON_TABULAR_SHEETS:
        return "non_tabular"
    # Unknown sheets: treat as tabular by default (safe fallback).
    return "tabular"


def read_tabular_sheet(xlsx_path: Path, sheet_name: str) -> pd.DataFrame | None:
    """Read a tabular sheet as a DataFrame (all values as str)."""
    try:
        df = pd.read_excel(xlsx_path, sheet_name=sheet_name, dtype=str, engine="openpyxl")
        df = df.fillna("").replace("nan", "").replace("NaN", "")
        return df
    except Exception as e:
        print(f"  [warn] Could not read sheet '{sheet_name}' in {xlsx_path.name}: {e}")
        return None


def read_raw_sheet_values(xlsx_path: Path, sheet_name: str) -> list[list] | None:
    """Read a non-tabular sheet as a raw list-of-rows (values only)."""
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        return rows
    except Exception as e:
        print(f"  [warn] Could not read raw sheet '{sheet_name}' in {xlsx_path.name}: {e}")
        return None


# =============================================================================
# MERGE HELPERS
# =============================================================================

def merge_tabular_frames(
    frames: list[tuple[str, str, pd.DataFrame]],
) -> pd.DataFrame:
    """Merge (source_file, source_batch, df) list into one DataFrame (union columns)."""
    if not frames:
        return pd.DataFrame()

    all_cols: list[str] = []
    seen: set[str] = set()
    for _, _, df in frames:
        for c in df.columns:
            if c not in seen:
                all_cols.append(c)
                seen.add(c)

    parts = []
    for source_file, source_batch, df in frames:
        aligned = df.reindex(columns=all_cols, fill_value="")
        aligned.insert(0, "source_row",   range(1, len(aligned) + 1))
        aligned.insert(0, "source_batch", source_batch)
        aligned.insert(0, "source_file",  source_file)
        parts.append(aligned)

    return pd.concat(parts, ignore_index=True)


def add_duplicate_flags(df: pd.DataFrame, warnings: list[str]) -> pd.DataFrame:
    for col, flag_col in [
        (DUP_COL_NUMBER, "possible_duplicate_company_number"),
        (DUP_COL_DOMAIN, "possible_duplicate_domain"),
    ]:
        if col in df.columns:
            vals = df[col].astype(str).str.strip()
            counts = vals.map(vals.value_counts())
            df[flag_col] = (counts > 1) & (vals != "") & (vals.str.lower() != "nan")
            df[flag_col] = df[flag_col].map({True: "YES", False: ""})
        else:
            warnings.append(
                f"Duplicate check skipped: column '{col}' not found in Opportunity Input."
            )
    return df


def write_company_profiles_sheet(
    ws,
    raw_blocks: list[tuple[str, list[list]]],
) -> None:
    """
    Append Company Profiles blocks from each file below each other.
    Inserts a separator row with the source filename between batches.
    No merged cells — values only.
    """
    sep_font = Font(bold=True, color="FFFFFF")
    sep_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    ri = 1
    for source_file, rows in raw_blocks:
        # Separator row
        sep_cell = ws.cell(row=ri, column=1, value=f"── {source_file} ──")
        sep_cell.font = sep_font
        sep_cell.fill = sep_fill
        ri += 1
        # Data rows
        for row in rows:
            for ci, val in enumerate(row, 1):
                if val is not None:
                    ws.cell(row=ri, column=ci, value=val)
            ri += 1
        ri += 1  # blank separator line

    # Reasonable default column widths
    for ci in range(1, 6):
        ws.column_dimensions[get_column_letter(ci)].width = 32


# =============================================================================
# MAIN MERGE LOGIC
# =============================================================================

def run_merge(
    input_dir: Path,
    output_path: Path,
    pattern: str = "*.xlsx",
    include_company_profiles: bool = False,
) -> None:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[merge] Start:      {ts}")
    print(f"[merge] Input dir:  {input_dir}")
    print(f"[merge] Output:     {output_path}")
    print(f"[merge] Pattern:    {pattern}")
    print(f"[merge] Company Profiles: {'included' if include_company_profiles else 'skipped'}")

    files = collect_files(input_dir, pattern)
    if not files:
        print(f"[merge] ERROR: No files found matching '{pattern}' in {input_dir}")
        sys.exit(1)

    print(f"[merge] Found {len(files)} file(s):")
    for f in files:
        print(f"  {f.name}")

    # ── Discover sheets in all files ─────────────────────────────────────────
    all_sheet_names: list[str] = []
    seen_sheet_set: set[str] = set()
    for xlsx_path in files:
        try:
            xf = pd.ExcelFile(xlsx_path, engine="openpyxl")
            for s in xf.sheet_names:
                if s not in seen_sheet_set:
                    all_sheet_names.append(s)
                    seen_sheet_set.add(s)
        except Exception as e:
            print(f"  [warn] Could not open {xlsx_path.name} for sheet discovery: {e}")

    tabular_found    = [s for s in all_sheet_names if classify_sheet(s) == "tabular"]
    non_tabular_found = [s for s in all_sheet_names if classify_sheet(s) == "non_tabular"]
    print(f"[merge] Tabular sheets detected:     {tabular_found}")
    print(f"[merge] Non-tabular sheets detected: {non_tabular_found}")

    # ── Read all tabular sheets ───────────────────────────────────────────────
    # sheet_name -> list of (source_file, source_batch, df)
    tabular_data: dict[str, list[tuple[str, str, pd.DataFrame]]] = {}
    rows_per_file: dict[str, int] = {}
    warnings: list[str] = []

    for xlsx_path in files:
        label = xlsx_path.name
        batch = _derive_batch_label(xlsx_path.name)
        file_rows = 0

        try:
            xf = pd.ExcelFile(xlsx_path, engine="openpyxl")
            available = set(xf.sheet_names)
        except Exception as e:
            warnings.append(f"Could not open {label}: {e}")
            continue

        for sname in tabular_found:
            if sname not in available:
                continue
            df = read_tabular_sheet(xlsx_path, sname)
            if df is None:
                continue
            tabular_data.setdefault(sname, []).append((label, batch, df))
            file_rows += len(df)

        rows_per_file[label] = file_rows

    # ── Read Company Profiles (raw values) ───────────────────────────────────
    cp_blocks: list[tuple[str, list[list]]] = []
    if include_company_profiles and "Company Profiles" in seen_sheet_set:
        for xlsx_path in files:
            rows = read_raw_sheet_values(xlsx_path, "Company Profiles")
            if rows:
                cp_blocks.append((xlsx_path.name, rows))
    elif "Company Profiles" in seen_sheet_set:
        warnings.append(
            "Company Profiles skipped because it is a formatted card-layout sheet, "
            "not a tabular sheet. Use --include-company-profiles to append raw values."
        )

    # Note Scoring Settings / Run Settings
    for s in ("Scoring Settings", "Run Settings"):
        if s in seen_sheet_set:
            warnings.append(
                f"'{s}' is a settings/layout sheet and was not merged. "
                "Check the first input file for settings context."
            )

    # ── Merge tabular sheets ──────────────────────────────────────────────────
    merged: dict[str, pd.DataFrame] = {}
    rows_per_sheet: dict[str, int] = {}
    sheets_merged: list[str] = []
    sheets_skipped: list[str] = list(non_tabular_found)
    if include_company_profiles and "Company Profiles" in sheets_skipped:
        sheets_skipped.remove("Company Profiles")

    # Build output sheet order: priority first, then remaining tabular
    priority_present = [s for s in TABULAR_PRIORITY if s in seen_sheet_set]
    rest_tabular = [s for s in tabular_found if s not in set(TABULAR_PRIORITY)]
    ordered_tabular = priority_present + rest_tabular

    for sname in ordered_tabular:
        frames = tabular_data.get(sname, [])
        if not frames:
            warnings.append(f"Sheet '{sname}' found in file list but no rows read — skipped.")
            continue
        df_merged = merge_tabular_frames(frames)

        # Column-uniformity warning
        col_sets = [set(df.columns) for _, _, df in frames]
        if len(col_sets) > 1:
            union_cols = set.union(*col_sets)
            for i, cs in enumerate(col_sets):
                missing = union_cols - cs
                if missing:
                    fname = frames[i][0]
                    sample = ", ".join(sorted(missing)[:4])
                    more   = f" (+{len(missing)-4} more)" if len(missing) > 4 else ""
                    warnings.append(
                        f"Sheet '{sname}' in {fname} missing {len(missing)} col(s): {sample}{more}"
                    )

        if sname == "Opportunity Input" and not df_merged.empty:
            df_merged = add_duplicate_flags(df_merged, warnings)

        merged[sname] = df_merged
        rows_per_sheet[sname] = len(df_merged)
        sheets_merged.append(sname)
        print(f"[merge] '{sname}': {len(df_merged)} rows from {len(frames)} file(s)")

    if not merged:
        print("[merge] ERROR: no tabular data was merged — nothing to write.")
        sys.exit(1)

    # ── Write output workbook ─────────────────────────────────────────────────
    print(f"[merge] Writing: {output_path}")
    wb = Workbook()
    wb.remove(wb.active)

    for sname in ordered_tabular:
        if sname not in merged:
            continue
        ws = wb.create_sheet(sname)
        df = merged[sname]
        if df.empty:
            ws.cell(row=1, column=1, value="(No data)")
        else:
            _write_df_to_sheet(ws, df)

    if include_company_profiles and cp_blocks:
        ws_cp = wb.create_sheet("Company Profiles")
        write_company_profiles_sheet(ws_cp, cp_blocks)
        sheets_merged.append("Company Profiles")
        rows_per_sheet["Company Profiles"] = sum(len(rows) for _, rows in cp_blocks)
        print(f"[merge] 'Company Profiles': raw values from {len(cp_blocks)} file(s)")

    # Merge QA — always last
    ws_qa = wb.create_sheet("Merge QA")
    _write_qa_sheet(ws_qa, {
        "timestamp":      ts,
        "input_dir":      str(input_dir),
        "output_file":    str(output_path),
        "n_files":        len(files),
        "input_files":    [f.name for f in files],
        "rows_per_file":  rows_per_file,
        "rows_per_sheet": rows_per_sheet,
        "sheets_merged":  sheets_merged,
        "sheets_skipped": sheets_skipped,
        "warnings":       warnings,
    })

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"[merge] Saved: {output_path}")

    # ── Console QA ───────────────────────────────────────────────────────────
    print()
    print(f"[QA] Sheets merged:  {sheets_merged}")
    print(f"[QA] Sheets skipped: {sheets_skipped}")
    print(f"[QA] Rows per sheet: {rows_per_sheet}")
    if warnings:
        print(f"[QA] Warnings ({len(warnings)}):")
        for w in warnings:
            print(f"     ⚠ {w}")
    else:
        print("[QA] No warnings.")


# =============================================================================
# CLI
# =============================================================================

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Merge Client Enricher batch-output Excel files into one workbook.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python merge_enricher_outputs.py \\
      --input-dir "C:/Users/gmeijer4/Nextcloud/Myngle/Italy200/02_lead_prioritized" \\
      --output "Italy200_merged_20260615.xlsx"

  python merge_enricher_outputs.py \\
      --input-dir ./outputs \\
      --output merged.xlsx \\
      --pattern "Italy200_*_lead_prioritized_*.xlsx" \\
      --include-company-profiles
""",
    )
    parser.add_argument("--input-dir",  required=True,  help="Directory with enricher .xlsx files")
    parser.add_argument("--output",     required=True,  help="Path for the merged output .xlsx")
    parser.add_argument("--pattern",    default="*.xlsx", help="Glob pattern (default: *.xlsx)")
    parser.add_argument(
        "--include-company-profiles",
        action="store_true",
        default=False,
        help="Append raw Company Profiles values from each file (layout not preserved)",
    )
    args = parser.parse_args()

    input_dir   = Path(args.input_dir).resolve()
    output_path = Path(args.output).resolve()

    if not input_dir.exists() or not input_dir.is_dir():
        print(f"ERROR: --input-dir not found or not a directory: {input_dir}", file=sys.stderr)
        sys.exit(1)

    run_merge(
        input_dir,
        output_path,
        pattern=args.pattern,
        include_company_profiles=args.include_company_profiles,
    )


if __name__ == "__main__":
    main()
