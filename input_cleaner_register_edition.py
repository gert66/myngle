"""
input_cleaner_register_edition.py — Layer 0: mYngle Input Cleaner · Register Edition
======================================================================================
Cleans and enriches Italian Business Register exports before Lead Prioritizer.
Handles missing websites (common in register data), PEC email detection,
multi-website fields, and location-aware Serper search queries.

Website Discovery Upgrade v2:
- Multi-variant brand name extraction
- 8 Serper query strategies with configurable cap
- Aggressive email-domain usage with Serper confirmation
- Richer scoring (rank, title/snippet signals, .it TLD, email match, location)
- New diagnostic output columns
- Expanded blacklist

Entry point:  streamlit run input_cleaner_register_edition.py
"""

import argparse
import csv
import hashlib
import io
import json
import os
import re
import shutil
import sys
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

import pandas as pd
import requests
import streamlit as st

try:
    import anthropic as _anthropic_sdk
    _ANTHROPIC_AVAILABLE = True
except ImportError:
    _anthropic_sdk = None
    _ANTHROPIC_AVAILABLE = False

# =============================================================================
# PAGE CONFIG
# =============================================================================

st.set_page_config(
    page_title="Input Cleaner · Register Edition",
    page_icon="🧹",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# =============================================================================
# CONSTANTS
# =============================================================================

SERPER_URL      = "https://google.serper.dev/search"
_AUTOSAVE_DIR   = Path("autosave")

# Module-level Serper call counter (reset per run via _reset_serper_counter())
_serper_call_count: int = 0

def _reset_serper_counter() -> None:
    global _serper_call_count
    _serper_call_count = 0

def _get_serper_count() -> int:
    return _serper_call_count
_AUTOSAVE_EVERY = 10   # write checkpoint every N processed rows

# Generic / directory / social / database domains to skip (global + Italian-specific)
_GENERIC_DOMAINS: frozenset = frozenset({
    # Social networks
    "linkedin.com", "facebook.com", "twitter.com", "x.com", "instagram.com",
    "youtube.com", "xing.com",
    # Global business directories / data providers
    "bloomberg.com", "crunchbase.com", "zoominfo.com", "dnb.com",
    "glassdoor.com", "indeed.com", "angel.co", "pitchbook.com",
    "opencorporates.com", "companieshouse.gov.uk",
    "rocketreach.co", "signalhire.com", "apollo.io", "hunter.io",
    "trustpilot.com", "yelp.com", "reuters.com", "ft.com",
    "github.com", "amazon.com", "app.lusha.com", "wikipedia.org",
    "google.com", "bing.com", "yahoo.com",
    # Job boards
    "jobrapido.it", "monster.it", "infojobs.it", "jobbydoo.it",
    "lavoro.corriere.it", "subito.it", "kijiji.it",
    # Italian company registers / directories / data sources
    "registroimprese.it", "infocamere.it", "imprese.it",
    "ufficiocamerale.it", "companywall.it", "reportaziende.it",
    "companyreports.it", "atoka.io",
    "paginegialle.it", "paginebianche.it",
    "europages.it", "europages.com",
    "kompass.com", "kompass.it",
    "cerved.com", "cervedgroup.it",
    "aziende.it", "icecat.it", "businessit.it",
    "italianmade.com", "viesus.com",
    "madeintaly.com", "italyexport.com",
    "nixonpowerseo.it", "dnbItaly.com",
    # Italian business-profile / financial-data directory sites
    "fatturatoaziende.com", "fatturatoitalia.it",
    "registroaziende.it", "registroaziende.com",
    "informazione-aziende.it", "aziendit.com",
    "dati-aziende.it", "ufficio-camerale.it",
    "companiesitaly.com", "italianbusinessregister.it",
    "visura.pro", "abbrevia.it",
    # News aggregators, price comparison, marketplaces
    "corriere.it", "repubblica.it", "ilsole24ore.com", "sole24ore.com",
    "trovaprezzi.it", "idealo.it", "amazon.it",
    # Business-data / profile lookup services
    "visura.pro", "abbrevia.it",
    # Free hosted-site platforms — subdomains are never official company sites
    "altervista.org", "wordpress.com", "blogspot.com",
    "wixsite.com", "weebly.com", "sites.google.com",
})

# Subdomain / path prefix check — any domain that contains these base domains is generic too
_GENERIC_DOMAIN_BASES: tuple = (
    "linkedin.com", "facebook.com", "twitter.com", "x.com", "instagram.com",
    "wikipedia.org", "google.com", "bing.com", "youtube.com",
    "registroimprese.it", "infocamere.it", "atoka.io", "kompass.com", "kompass.it",
    "europages.com", "europages.it", "paginegialle.it", "paginebianche.it",
    "cerved.com", "cervedgroup.it", "dnb.com", "zoominfo.com",
    "bloomberg.com", "crunchbase.com", "glassdoor.com", "indeed.com",
    # Italian business-profile directory base domains (catches subdomains like m.fatturatoitalia.it)
    "fatturatoaziende.com", "fatturatoitalia.it",
    "registroaziende.it", "registroaziende.com",
    "informazione-aziende.it", "aziendit.com",
    "dati-aziende.it", "ufficio-camerale.it",
    "companiesitaly.com", "italianbusinessregister.it",
    # Free hosted-site platforms (catches e.g. rsuibmsegrate.altervista.org, x.abbrevia.it)
    "altervista.org", "blogspot.com", "wordpress.com",
    "wixsite.com", "weebly.com", "sites.google.com",
    "visura.pro", "abbrevia.it",
)

# PEC (Posta Elettronica Certificata) domains — never use as company website
_PEC_DOMAIN_PATTERNS: tuple = (
    "pec.it", "pec.com", "pec.eu", "legalmail.it", "legalmail.com",
    "postecert.it", "arubapec.it", "arubapec.eu", "pecactually.it",
    "cert.it", "pecimprese.it", "certificata.it", "pecaziende.it",
    "pecmail.it", "pec.tiscali.it", "pecimprese.com",
    "ordineavvocati", "ordinedottori", "caf", "patronato",
    "libero.it", "yahoo.it", "gmail.com", "hotmail.it",
    "alice.it", "tin.it", "virgilio.it", "live.com", "outlook.com",
    "tiscali.it", "hotmail.com", "icloud.com", "protonmail.com",
)

# Expected column names for an Italian Business Register export
_REG_COL_COMPANY   = "Company Name"
_REG_COL_WEBSITE   = "Website"
_REG_COL_EMAIL     = "Email address"
_REG_COL_CITY      = "City"
_REG_COL_PROVINCE  = "National statistical institute Province"
_REG_COL_POSTCODE  = "Postal Code"
_REG_COL_PHONE     = "Phone number"
_REG_COL_SERIAL    = "Serial Number"
_REG_COL_REGOFFICE = "Registered Office"

# Fallback detection candidates (normalised lowercase)
_NAME_CANDIDATES = (
    "company name", "company_name", "ragione sociale", "denominazione",
    "nome", "company", "name",
)
_WEBSITE_CANDIDATES = (
    "website", "sito web", "sito", "url", "web", "homepage",
    "website_url", "domain",
)
_EMAIL_CANDIDATES = (
    "email address", "email", "e-mail", "posta elettronica",
    "indirizzo email",
)
_CITY_CANDIDATES = (
    "city", "città", "comune", "citta",
)
_PROVINCE_CANDIDATES = (
    "national statistical institute province", "province", "provincia",
    "prov",
)
_POSTCODE_CANDIDATES = (
    "postal code", "cap", "postcode", "zip", "codice postale",
)
_PHONE_CANDIDATES = (
    "phone number", "telefono", "tel", "phone",
)

# Legal suffix patterns (Italian + common European)
_LEGAL_TOKENS = re.compile(
    r"\b(s\.?r\.?l\.?|s\.?p\.?a\.?|s\.?a\.?s?|snc|s\.?n\.?c\.?|"
    r"s\.?a\.?p\.?a?\.?|ltd|limited|b\.?v\.?|n\.?v\.?|gmbh|ag|"
    r"llc|inc|corp|plc|holding|holdings|group|co|company|pty|"
    r"se|pte|bhd|sarl|eurl|scs|cv|impresa|ditta|studio|"
    r"kg|k\.g\.|ohg|o\.h\.g\.|ug|kgaa|k\.g\.a\.a\.|gbr|g\.b\.r\.|ek|e\.k\.|partg)\b\.?",
    re.IGNORECASE,
)

# Tokens that are purely legal suffixes and must never be used as the brand name
_BRAND_LEGAL_REJECT_TOKENS: frozenset = frozenset({
    # Italian
    "srl", "spa", "sas", "snc", "srls", "sapa",
    # German
    "gmbh", "ag", "kg", "ohg", "ug", "se", "kgaa", "gbr", "ek", "partg",
    # International
    "ltd", "llc", "inc", "corp", "plc", "bv", "nv", "co",
})

# Generic single-word tokens that must never be used as a standalone brand for
# German companies. If the extracted brand reduces to one of these, the candidate
# requires the exact legal company name (or a very close variant) in title/snippet
# before it can be accepted — preventing false positives like "Global" → geglobalsales.com.
_DE_GENERIC_BRAND_TOKENS: frozenset = frozenset({
    "global", "holding", "holdings", "group", "international", "investments",
    "investment", "media", "entertainment", "management", "services", "service",
    "solutions", "solution", "technologies", "technology", "systems", "system",
    "consulting", "sales", "trading", "industries", "industrial", "capital",
    "partners", "enterprise", "enterprises", "logistics", "digital",
})

# Italian descriptor words that are NOT part of the brand name
_ITALIAN_DESCRIPTORS = re.compile(
    r"\b(societ[aà]|societa|aziend[ae]|azienda|impres[ae]|impresa|"
    r"industri[ae]|industria|industriale|commerciale|agricol[ae]|agricola|"
    r"gruppo|gruppi|holding|cooperativ[ae]|cooperativa|manifattur[ae]|"
    r"manifatturiero|costruzioni|costruttori|distribuzione|lavorazione|"
    r"produzione|prodotti|fratelli|f\.lli|flli|figli|eredi|successori|"
    r"succ\.?|consorzio|consorzi|associazione|fondazione|istituto)\b\.?",
    re.IGNORECASE,
)

# Matches "O, IN FORMA ABBREVIATA, <short name>" — Italian register long-form pattern
# Captures everything after the phrase as the preferred short name
_FORMA_ABBREVIATA_RE = re.compile(
    r"\bO?,?\s*IN\s+FORMA\s+ABBREVIATA[,\s]+(.+)",
    re.IGNORECASE,
)

# "O IN BREVE", "IN BREVE" → use text after phrase (the short name)
_IN_BREVE_RE = re.compile(
    r"\bO?\s*IN\s+BREVE[,\s]+(.+)",
    re.IGNORECASE,
)

# "DA INDICARE ANCHE COME", "INDICARE ANCHE COME", "ANCHE COME" → use text after
_ANCHE_COME_RE = re.compile(
    r"\b(?:DA\s+INDICARE\s+)?(?:INDICARE\s+)?ANCHE\s+COME[,\s]+(.+)",
    re.IGNORECASE,
)

# "IN SIGLA ...", "SIGLABILE ..." → strip from this phrase onward (main name is before)
_IN_SIGLA_RE = re.compile(
    r"\s*[,\-–]\s*(?:IN\s+SIGLA|SIGLABILE)\b.*$",
    re.IGNORECASE,
)

# Regex to extract SIGLABILE/IN SIGLA content for brand extraction
_SIGLABILE_EXTRACT_RE = re.compile(
    r"[\(\[]*\s*(?:IN\s+SIGLA|SIGLABILE)\s+(.+?)[\)\]]*$", re.I
)

# First-token brand rule: generic descriptors that indicate brand is first token
_GENERIC_DESCRIPTORS_RE = re.compile(
    r"^(TECNOLOGIE|IMPIANTI|SISTEMI|SERVIZI|SICUREZZA|ENERGIA|INNOVAZIONE|"
    r"SOLUZIONI|AUTOMAZIONE|COSTRUZIONI|PRODUZIONE|DISTRIBUZIONE|LOGISTICA|"
    r"TRASPORTI|CONSULENZA|INGEGNERIA|INFORMATICA|ELETTRONICA|MECCANICA|"
    r"CHIMICA|FARMACEUTICA|ALIMENTARE|INDUSTRIALE|COMMERCIALE|INTERNAZIONALE)$", re.I
)
_LEGAL_SUFFIX_TOKENS = {"spa", "srl", "sas", "snc", "srls", "sapa", "ss"}

# Extra Italian register legal phrases not caught by _LEGAL_TOKENS / _ITALIAN_DESCRIPTORS
_EXTRA_LEGAL_PHRASES_RE = re.compile(
    r"\bSOCIETA\'?\s+PER\s+AZIONI\b"
    r"|\bPER\s+AZIONI\b"
    r"|\bIN\s+FORMA\s+ABBREVIATA\b"
    r"|\bO,?\s*IN\s+FORMA\s+ABBREVIATA\b",
    re.IGNORECASE,
)

_NOISE_TOKENS: frozenset = frozenset({
    "the", "and", "for", "global", "international", "services", "solutions",
    "consulting", "management", "technology", "technologies", "systems",
    "software", "digital", "enterprise", "enterprises", "partners",
    "italia", "italy", "italian", "europe", "european",
    "snc", "srl", "spa", "sas", "del", "della", "degli", "dei",
    "di", "da", "in", "con", "su", "per", "tra", "fra",
    "group", "holding", "co", "ltd", "inc", "bv",
})

_TLDS: frozenset = frozenset({
    "com", "net", "org", "it", "eu", "nl", "de", "fr", "be", "uk", "co",
    "io", "biz", "info", "at", "ch", "es", "pl", "cz", "se", "no",
    "dk", "fi", "pt", "hu", "ro", "hr", "gr", "gov", "edu",
})

# Keywords that signal an official/home page in title/snippet
_OFFICIAL_SIGNALS = frozenset({
    "official", "sito ufficiale", "home page", "homepage",
    "benvenuti", "welcome", "chi siamo", "about us",
    "sito web ufficiale", "official website", "official site",
})

# Government domain patterns — always reject as company website
_GOVT_PATTERNS = re.compile(
    r"\.gov\.it$|\.gov\b|agenziaentrate|"
    r"(?:^|\.)comune\.|(?:^|\.)regione\.|(?:^|\.)provincia\.|"
    r"prefettura|questura|tribunale|ministero|"
    r"inps\.it$|inail\.it$|agenziademanio|"
    r"camera\.it$|senato\.it$|governo\.it$|quirinale\.it$|mef\.gov",
    re.IGNORECASE,
)

# Religious institution patterns — reject as company website
_RELIGIOUS_PATTERNS = re.compile(
    r"basilica|diocesi|parrocchia|chiesa(?:cattolica)?|santuario|"
    r"abbazia|convento|vescovado|cattedrale|arcidiocesi|"
    r"seminario|oratorio|vaticano|pontific|caritas|"
    r"cappella|pieve|fraternita|confraternita",
    re.IGNORECASE,
)

# Directory / hours / aggregator patterns not already in _GENERIC_DOMAINS
_DIRECTORY_EXTRA_PATTERNS = re.compile(
    r"oraridiapertura|aperturenegozi|tuttopmi|impresaitalia|"
    r"businessfinder|b2bnetwork|catalogoimprese|trovimprese|"
    r"ioimpresa|businessregister|italiabusiness|infobel|"
    r"fatturato|bilanci|dati-aziend|scheda-aziend|scheda-impres|"
    r"visura-aziend|report-aziend|company-profile|business-profile",
    re.IGNORECASE,
)

# Strong content signals that identify a page as a business-profile/directory entry.
# Checked against title+snippet when domain is not in _GENERIC_DOMAINS.
_DIRECTORY_PROFILE_TITLE_SIGNALS = re.compile(
    r"\bfatturato\b|\bbilancio\b|\butili\b|\bricavi\b|\bpartita\s+iva\b|"
    r"\bp\.?\s*iva\b|\bscheda\s+azienda\b|\bscheda\s+impresa\b|"
    r"\bdati\s+aziendali\b|\breport\s+azienda\b|\bvisura\b|"
    r"\bregistro\s+aziende\b|\bcompany\s+profile\b|\bbusiness\s+profile\b|"
    r"\bcodice\s+ateco\b|\bforma\s+giuridica\b|\bcapitale\s+sociale\b",
    re.IGNORECASE,
)

# Academic / university patterns
_ACADEMIC_PATTERNS = re.compile(
    r"\.edu$|\.ac\.[a-z]{2,}$|universit[aà]|polimi|polito|"
    r"unimi|unibo|unitn|luiss|bocconi|sapienza|unipd|unifi|politecnico",
    re.IGNORECASE,
)

# Brand similarity gate thresholds
_MIN_BRAND_SIM_TO_SCORE    = 0.15  # below this → score × 0.10 (near-rejection)
_WEAK_BRAND_SIM_MULTIPLIER = 0.35  # between _MIN and 0.25 → score × this
_WEAK_BRAND_SIM_THRESHOLD  = 0.25
_HIGH_CONF_BRAND_THRESHOLD = 0.60  # brand must reach this for High-confidence rule A

# Row colours (openpyxl ARGB hex)
_ACTION_COLORS = {
    "OK":                   "C6EFCE",
    "LIKELY_OK":            "E2EFDA",
    "REVIEW":               "FFEB9C",
    "SUGGEST_REPLACE":      "FCE4D6",
    "MISSING_DOMAIN_FIXED": "FCE4D6",
    "NO_CONFIDENT_MATCH":   "FFC7CE",
    "MISSING_DOMAIN":       "FFC7CE",
    "EMAIL_DERIVED":        "DDEEFF",
}

# Source labels for domain_source column
SRC_ORIGINAL              = "original_website"
SRC_EMAIL                 = "email_domain"
SRC_SERPER                = "serper_search"
SRC_SERPER_EMAIL          = "serper_confirmed_email_domain"
SRC_NONE                  = ""

# Claude Haiku review mode constants
_DEFAULT_HAIKU_MODEL  = "claude-haiku-4-5-20251001"
_HAIKU_MODE_PYTHON    = "Python only"
_HAIKU_MODE_UNCERTAIN = "Haiku for uncertain rows only"
_HAIKU_MODE_ALL       = "Haiku for all rows"

# Jina website verifier mode constants
_JINA_MODE_OFF       = "Off"
_JINA_MODE_UNCERTAIN = "For uncertain candidates only"
_JINA_MODE_ALL_DEBUG = "For all selected candidates in debug mode"
_JINA_MODES          = [_JINA_MODE_OFF, _JINA_MODE_UNCERTAIN, _JINA_MODE_ALL_DEBUG]

_JINA_READER_BASE = "https://r.jina.ai/"
_JINA_MAX_CHARS   = 3000   # chars to keep per fetched page
_JINA_SLUGS       = ["/", "/about", "/about-us", "/chi-siamo", "/contatti", "/contact"]

_JINA_RISKY_DOMAIN_RE = re.compile(
    r"\b(?:forum|fan|club|community|dealer|shop|store|wiki|museum|foundation)\b",
    re.IGNORECASE,
)

_RISKY_MARKERS = [
    "forum", "foro", "fan", "fans", "club", "community",
    "archive", "archivio", "wiki", "museum",
    "dealer", "reseller", "shop", "store",
    "directory", "profile",
    # Hosted-platform subdomains are never official company sites
    "altervista", "blogspot", "wordpress", "wixsite", "weebly",
]

# Hosted-platform base domains whose subdomains must always be rejected.
# These overlap with _GENERIC_DOMAIN_BASES but are kept separate so that
# is_hosted_platform() can provide a specific rejection reason.
_HOSTED_PLATFORM_BASES: tuple = (
    "altervista.org", "blogspot.com", "wordpress.com",
    "wixsite.com", "weebly.com", "sites.google.com",
)


def _domain_has_risky_marker(domain: str) -> tuple[bool, str]:
    """
    Return (is_risky, reason) for a domain name.
    Detects risky markers even when glued to a brand (e.g. ferrariforum.it).
    Use this instead of _JINA_RISKY_DOMAIN_RE when checking domain strings.
    """
    d = (domain or "").lower()
    base = d.split(".")[0]                      # e.g. "ferrariforum"
    compact = re.sub(r"[^a-z0-9]", "", base)   # e.g. "ferrariforum"
    parts = re.split(r"[-._]", d)              # e.g. ["ferrariforum", "it"]

    for m in _RISKY_MARKERS:
        if m in parts:                          # exact token match
            return True, f"contains risky marker: {m}"
        if m in compact:                        # glued pattern (ferrariforum)
            return True, f"contains risky marker: {m}"

    return False, ""


# Famous / generic single-word brands that need extra Jina verification
_JINA_FAMOUS_BRANDS: frozenset = frozenset({
    "ferrari", "lamborghini", "fiat", "alfa", "romeo", "lancia", "ducati",
    "barilla", "lavazza", "campari", "pirelli", "benetton", "prada", "gucci",
    "versace", "armani", "valentino", "bulgari", "ferrero",
    "delta", "omega", "sigma", "alpha", "beta", "gamma",
    "atlas", "titan", "apollo", "mercury", "saturn", "orion",
})

# Module-level Jina page cache: (domain, slug) -> (text, fetch_status)
_JINA_CACHE: dict[tuple[str, str], tuple[str, str]] = {}

# Firecrawl verifier constants
_FC_API_URL   = "https://api.firecrawl.dev/v1/scrape"
_FC_MAX_CHARS = 5000  # chars to keep per scraped page
_FC_CACHE: dict[tuple, tuple[str, str, dict]] = {}  # (url, country, languages) -> (text, status, meta)


def _make_fc_cache_key(url: str, fc_location: dict | None) -> tuple:
    """Build a hashable cache key from URL + location (list values are converted to tuple)."""
    loc = fc_location or {}
    return (url, loc.get("country", ""), tuple(loc.get("languages", []) or []))


_REDIRECT_CACHE: dict[str, dict] = {}   # domain -> redirect result dict


def _resolve_redirect(domain: str) -> dict:
    """
    Lightweight HEAD/GET redirect resolver.  Follows redirects and returns
    the final URL and domain.  Results are cached per domain.

    Returns a dict with keys:
      redirect_checked       bool
      redirect_status        str  "ok" | "timeout" | "error:<msg>"
      redirect_final_url     str  final URL after redirects (may equal original)
      redirect_final_domain  str  root domain of final URL (empty if same as input)
    """
    if domain in _REDIRECT_CACHE:
        return _REDIRECT_CACHE[domain]

    result = {
        "redirect_checked":       True,
        "redirect_status":        "ok",
        "redirect_final_url":     "",
        "redirect_final_domain":  "",
    }
    _headers = {"User-Agent": "Mozilla/5.0 (compatible; mYngle-Verifier/1.0)"}
    _origin  = domain.lower().lstrip("www.")

    def _try(scheme: str) -> str | None:
        try:
            r = requests.get(
                f"{scheme}://{domain}",
                allow_redirects=True,
                timeout=5,
                headers=_headers,
            )
            return r.url
        except requests.Timeout:
            result["redirect_status"] = "timeout"
        except Exception as exc:
            result["redirect_status"] = f"error:{str(exc)[:80]}"
        return None

    final_url = _try("https") or _try("http")

    if final_url:
        result["redirect_final_url"] = final_url
        _rd = re.sub(r"^https?://([^/]+).*$", r"\1", final_url).lower().lstrip("www.")
        if _rd and _rd != _origin:
            result["redirect_final_domain"] = _rd
    else:
        result["redirect_checked"] = True  # attempted but failed

    _REDIRECT_CACHE[domain] = result
    return result


# =============================================================================
# ORGANIZATION ELIGIBILITY PRE-FILTER
# =============================================================================

_ELI_KEEP    = "KEEP"
_ELI_MAYBE   = "MAYBE"
_ELI_EXCLUDE = "EXCLUDE"

_PF_SEND    = "send_to_website_discovery"
_PF_SKIP    = "skip_website_discovery"
_PF_LATER   = "optional_later_review"

_PF_MODE_COMMERCIAL = "Commercial only"
_PF_MODE_MAYBE      = "Commercial + Maybe"
_PF_MODE_ALL        = "All rows"
_PF_MODES           = [_PF_MODE_COMMERCIAL, _PF_MODE_MAYBE, _PF_MODE_ALL]

# Legal forms → KEEP (commercial companies)
_KEEP_FORMS_RE = re.compile(
    r"\b(s\.?\s*p\.?\s*a\.?|spa|societ[àa]\s+per\s+azioni"
    r"|s\.?\s*r\.?\s*l\.?\s*s?\.?|srl[s]?|societ[àa]\s+a\s+responsabilit[àa]\s+limitata"
    r"|s\.?\s*a\.?\s*p\.?\s*a\.?"
    r"|s\.?\s*n\.?\s*c\.?|snc|societ[àa]\s+in\s+nome\s+collettivo"
    r"|s\.?\s*a\.?\s*s\.?|sas|societ[àa]\s+in\s+accomandita\s+semplice)\b",
    re.I,
)

# Legal forms / keywords → MAYBE
_MAYBE_FORMS_RE = re.compile(
    r"\b(cooperativa|societ[àa]\s+cooperativa|coop\.?"
    r"|consorzio|societ[àa]\s+consortile"
    r"|societ[àa]\s+agricola|societ[àa]\s+semplice\s+agricola|ss\s+agricola"
    r"|azienda\s+speciale)\b",
    re.I,
)

# Keywords → EXCLUDE (non-commercial or public bodies)
_EXCLUDE_FORMS_RE = re.compile(
    r"\b(associazione|fondazione|ets|onlus|odv|aps"
    r"|comitato\b|caritas|parrocchia|diocesi|congregazione"
    r"|universit[àa]|politecnico"
    r"|scuola\b|istituto\s+statale|istituto\s+tecnico|istituto\s+comprensivo"
    r"|comune\s+di|comune\b|regione\b|provincia\s+di|ministero"
    r"|asl\b|ausl\b|ats\b|asp\b"
    r"|azienda\s+sanitaria|azienda\s+ospedaliera|ospedale|irccs"
    r"|ente\s+ecclesiastico)\b",
    re.I,
)

# Public/health override signals that downgrade even legal commercial forms
_PUBLIC_OVERRIDE_RE = re.compile(
    r"\b(sanitaservice|asl\s+\w+|ausl\s+\w+|ats\s+\w+|asp\s+\w+"
    r"|azienda\s+sanitaria|azienda\s+ospedaliera|ospedale\s+\w+"
    r"|irccs|ente\s+pubblico|ente\s+locale|partecipata\s+pubblica"
    r"|in\s+house\s+provid|societ[àa]\s+in\s+house)\b",
    re.I,
)


_DE_KEEP_FORMS_RE = re.compile(
    r"\b(GmbH\s*&\s*Co\.?\s*KG|GmbH\s*&\s*Co\.?\s*KGaA"
    r"|Gesellschaft\s+mit\s+beschr[äa]nkter\s+Haftung"
    r"|GmbH|AG|Aktiengesellschaft"
    r"|KG|OHG|KGaA|SE\b|UG\s*\(?haftungsbeschr[äa]nkt\)?"
    r"|UG\b|PartG|GbR|Einzelunternehmen|e\.?\s*K\.?)\b",
    re.I,
)

_DE_MAYBE_FORMS_RE = re.compile(
    r"\b(eG\b|Genossenschaft|Genossenschaftsbank"
    r"|GmbH\s*&\s*Co\.?\s*eG|Kommanditgesellschaft\s+auf\s+Aktien)\b",
    re.I,
)

_DE_EXCLUDE_FORMS_RE = re.compile(
    r"\b(e\.?\s*V\.?\b|Verein|Stiftung|gemeinn[üu]tzig"
    r"|Stadtwerke|Gemeinde\b|Landkreis\b|Kreisverwaltung"
    r"|Krankenhaus|Klinikum|Klinik\b|Universit[äa]t\b|Hochschule\b"
    r"|Schule\b|Gymnasium\b|Realschule\b|Grundschule\b|Berufsschule\b"
    r"|Kirche\b|Pfarr\w*|Diözese|Bistum|Caritas\b"
    r"|Bundesamt\b|Bundesanstalt\b|Ministerium\b|Amt\s+f[üu]r|Beh[öo]rde)\b",
    re.I,
)

_DE_PUBLIC_OVERRIDE_RE = re.compile(
    r"\b(Stadt\s+\w+|Stadtwerke\s+\w+|Landratsamt\b|Landkreis\s+\w+"
    r"|Klinikum\s+\w+|Universit[äa]tsklinikum|Beh[öo]rde\b"
    r"|GmbH\s+der\s+Stadt|mbH\s+der\s+Gemeinde)\b",
    re.I,
)


def _classify_organization_de(
    company_name: str,
    legal_form_hint: str = "",
) -> tuple[str, str, str, str]:
    """Classify German company for mYngle eligibility."""
    # Prefer legal_form_detected from register; fall back to company name
    probe = (legal_form_hint or company_name or "").strip()
    name_probe = (company_name or "").strip()

    has_keep    = bool(_DE_KEEP_FORMS_RE.search(probe) or _DE_KEEP_FORMS_RE.search(name_probe))
    has_maybe   = bool(_DE_MAYBE_FORMS_RE.search(probe) or _DE_MAYBE_FORMS_RE.search(name_probe))
    has_exclude = bool(_DE_EXCLUDE_FORMS_RE.search(probe) or _DE_EXCLUDE_FORMS_RE.search(name_probe))
    has_public  = bool(_DE_PUBLIC_OVERRIDE_RE.search(probe) or _DE_PUBLIC_OVERRIDE_RE.search(name_probe))

    nl = (probe + " " + name_probe).lower()

    if has_exclude:
        if any(t in nl for t in ("verein", "e.v", "e. v")):
            org_type = "nonprofit_association"
        elif any(t in nl for t in ("stiftung",)):
            org_type = "foundation"
        elif any(t in nl for t in ("gemeinde", "landkreis", "kreisverwaltung", "stadtwerke", "amt ", "ministerium", "bundesamt", "bundesanstalt")):
            org_type = "government_body"
        elif any(t in nl for t in ("klinik", "krankenhaus", "klinikum")):
            org_type = "public_health"
        elif any(t in nl for t in ("universit", "hochschule", "schule", "gymnasium", "realschule", "grundschule", "berufsschule")):
            org_type = "university_education"
        elif any(t in nl for t in ("kirche", "pfarr", "diöze", "bistum", "caritas")):
            org_type = "religious_body"
        else:
            org_type = "non_commercial"
        return org_type, _ELI_EXCLUDE, _PF_SKIP, f"Non-commercial entity (DE): {org_type}"

    if has_public and has_keep:
        return "public_health_or_public_owned", _ELI_MAYBE, _PF_LATER, "Commercial form but public-body name signals (DE)"

    if has_keep:
        return "commercial_company", _ELI_KEEP, _PF_SEND, "German commercial legal form"

    if has_maybe:
        return "cooperative_or_consortium", _ELI_MAYBE, _PF_LATER, "German cooperative/eG — lower priority"

    return "unknown", _ELI_MAYBE, _PF_LATER, "No recognized German legal form — unknown organization type"


def _normalize_legal_text(s: str) -> str:
    """Strip accents, apostrophes, and non-alphanumeric chars for regex matching."""
    s = (s or "").upper()
    for old, new in [("'", ""), ("’", ""), ("`", ""), ("´", ""),
                     ("À", "A"), ("È", "E"), ("É", "E"),
                     ("Ì", "I"), ("Ò", "O"), ("Ù", "U")]:
        s = s.replace(old, new)
    s = re.sub(r"[^A-Z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def classify_organization(
    company_name: str,
    legal_form_hint: str = "",
    country_code: str = "IT",
) -> tuple[str, str, str, str]:
    """
    Classify a company name for mYngle eligibility.
    Returns (organization_type, myngle_target_eligibility, pre_filter_decision, pre_filter_reason).
    """
    # ── German fast-path ─────────────────────────────────────────────────────
    if country_code == "DE":
        return _classify_organization_de(company_name, legal_form_hint)

    n = (company_name or "").strip()
    nl = n.lower()
    n_norm = _normalize_legal_text(n)

    has_keep    = bool(_KEEP_FORMS_RE.search(n) or _KEEP_FORMS_RE.search(n_norm))
    has_maybe   = bool(_MAYBE_FORMS_RE.search(n) or _MAYBE_FORMS_RE.search(n_norm))
    has_exclude = bool(_EXCLUDE_FORMS_RE.search(n) or _EXCLUDE_FORMS_RE.search(n_norm))
    has_public  = bool(_PUBLIC_OVERRIDE_RE.search(n) or _PUBLIC_OVERRIDE_RE.search(n_norm))

    # Determine organization type label
    nl_norm = n_norm.lower()
    if _EXCLUDE_FORMS_RE.search(n) or _EXCLUDE_FORMS_RE.search(n_norm):
        # Find the first matching exclude term
        _em = _EXCLUDE_FORMS_RE.search(n) or _EXCLUDE_FORMS_RE.search(n_norm)
        _et = _em.group(0).strip().lower() if _em else "non_commercial"
        if any(t in nl or t in nl_norm for t in ("asl", "ausl", "ats", "asp", "sanitaria", "ospedaliera", "irccs", "ospedale")):
            org_type = "public_health"
        elif any(t in nl or t in nl_norm for t in ("universit", "politecnico")):
            org_type = "university_education"
        elif any(t in nl or t in nl_norm for t in ("scuola", "istituto statale", "istituto tecnico", "istituto comprensivo")):
            org_type = "public_school"
        elif any(t in nl or t in nl_norm for t in ("associazione", "odv", "aps", "onlus", "ets")):
            org_type = "nonprofit_association"
        elif any(t in nl or t in nl_norm for t in ("fondazione",)):
            org_type = "foundation"
        elif any(t in nl or t in nl_norm for t in ("comune", "regione", "provincia", "ministero")):
            org_type = "government_body"
        elif any(t in nl or t in nl_norm for t in ("parrocchia", "diocesi", "congregazione", "caritas", "ente ecclesiastico")):
            org_type = "religious_body"
        else:
            org_type = "non_commercial"
    elif has_public and has_keep:
        org_type = "public_health_or_public_owned"
    elif has_maybe:
        org_type = "cooperative_or_consortium"
    elif has_keep:
        org_type = "commercial_company"
    else:
        org_type = "unknown"

    # Eligibility decision
    if has_exclude:
        return org_type, _ELI_EXCLUDE, _PF_SKIP, f"Non-commercial entity: {org_type}"

    if has_public and has_keep:
        # Commercial legal form but public-body name signals
        return org_type, _ELI_MAYBE, _PF_LATER, "Commercial form but public/healthcare name signals"

    if has_keep:
        return org_type, _ELI_KEEP, _PF_SEND, "Commercial legal form"

    if has_maybe:
        return org_type, _ELI_MAYBE, _PF_LATER, "Cooperative/consortium — lower priority"

    # No recognizable legal form
    return "unknown", _ELI_MAYBE, _PF_LATER, "No recognized legal form — unknown organization type"


# Unified website verifier provider options
_VP_OFF       = "Off"
_VP_JINA      = "Jina"
_VP_FIRECRAWL = "Firecrawl"
_VP_FC_JINA   = "Firecrawl first, Jina fallback"
_VP_OPTIONS   = [_VP_OFF, _VP_JINA, _VP_FIRECRAWL, _VP_FC_JINA]

# Verification trigger mode options
_VM_UNCERTAIN = "Uncertain candidates only"
_VM_ALL_DEBUG = "All selected candidates in debug mode"
_VM_OPTIONS   = [_VM_UNCERTAIN, _VM_ALL_DEBUG]

# Firecrawl speed mode constants
_FC_SPEED_FAST      = "Fast"       # homepage only, 1 candidate, 1 page, 6s timeout
_FC_SPEED_BALANCED  = "Balanced"   # homepage + about/contact if weak, 1 cand, 2 pages, 8s
_FC_SPEED_THOROUGH  = "Thorough"   # homepage + about + contact, configurable
_FC_SPEED_OPTIONS   = [_FC_SPEED_FAST, _FC_SPEED_BALANCED, _FC_SPEED_THOROUGH]

# Firecrawl location constants
_FC_LOC_AUTO    = "Auto from detected country"
_FC_LOC_ITALY   = "Italy"
_FC_LOC_GERMANY = "Germany"
_FC_LOC_USA     = "United States"
_FC_LOC_DEFAULT = "Default Firecrawl"
_FC_LOC_OPTIONS = [_FC_LOC_AUTO, _FC_LOC_ITALY, _FC_LOC_GERMANY, _FC_LOC_USA, _FC_LOC_DEFAULT]

_FC_LOC_PAYLOADS = {
    _FC_LOC_AUTO:    None,  # resolved at runtime from detected country
    _FC_LOC_ITALY:   {"country": "IT", "languages": ["it", "en"]},
    _FC_LOC_GERMANY: {"country": "DE", "languages": ["de", "en"]},
    _FC_LOC_USA:     {"country": "US", "languages": ["en"]},
    _FC_LOC_DEFAULT: {},
}

def _fc_loc_payload_for_country(country_code: str) -> dict:
    """Return the Firecrawl location payload for the given country code."""
    if country_code == "DE":
        return {"country": "DE", "languages": ["de", "en"]}
    return {"country": "IT", "languages": ["it", "en"]}

# =============================================================================
# COUNTRY CONFIG
# =============================================================================

from dataclasses import dataclass, field as _dc_field

@dataclass
class CountryConfig:
    """All country-specific settings for the input cleaner pipeline."""
    country_code:          str          # "IT" or "DE"
    country_name:          str          # "Italy" or "Germany"
    serper_gl:             str          # Serper geolocation code
    serper_hl:             str          # Serper interface language
    country_search_name:   str          # How to name the country in queries
    official_search_terms: list         # e.g. ["sito ufficiale", "official website"]
    preferred_tlds:        list         # e.g. [".it"] or [".de"]
    directory_blacklist_domains: set    # exact domains to reject
    directory_blacklist_bases:   set    # partial base-domain substrings to reject
    email_provider_blacklist:    set    # generic email providers (not company)
    legal_tokens:          list         # legal form tokens for this country
    descriptor_cleanup_patterns: list   # (regex_str, replacement) tuples
    firecrawl_location:    dict         # Firecrawl location payload
    # German-style columns (empty for IT)
    name_col_primary:      str  = ""    # e.g. "company_name_clean"
    name_col_fallback:     str  = ""    # e.g. "company_name_raw"
    city_col_primary:      str  = ""    # e.g. "city_or_registered_office"
    state_col:             str  = ""    # e.g. "federal_state"
    address_col:           str  = ""    # e.g. "registered_address"
    # Negative domain patterns (word-boundary safe)
    negative_domain_patterns: list = _dc_field(default_factory=list)


_DE_DIRECTORY_DOMAINS: set = {
    "handelsregister.de", "unternehmensregister.de", "bundesanzeiger.de",
    "northdata.de", "firmenwissen.de", "companyhouse.de", "implisense.com",
    "wer-zu-wem.de", "die-deutsche-wirtschaft.de", "gelbeseiten.de",
    "dasoertliche.de", "meinestadt.de", "cylex.de", "kompany.com",
    "opencorporates.com", "firmenabc.de", "europages.de", "europages.com",
    "kompass.com", "kompass.de", "indeed.com", "stepstone.de",
    "kununu.com", "xing.com", "linkedin.com", "facebook.com",
    "instagram.com",
}

_DE_DIRECTORY_BASES: set = {
    "handelsregister", "unternehmensregister", "bundesanzeiger",
    "northdata", "firmenwissen", "companyhouse", "implisense",
    "wer-zu-wem", "gelbeseiten", "dasoertliche", "meinestadt",
    "cylex", "firmenabc", "europages", "kompass", "kununu",
    "stepstone", "opencorporates",
}

_DE_EMAIL_GENERICS: set = {
    "gmail.com", "yahoo.com", "outlook.com", "hotmail.com",
    "web.de", "gmx.de", "gmx.net", "t-online.de", "protonmail.com",
    "icloud.com", "me.com", "live.com", "msn.com",
}

# German public/non-commercial domain patterns — checked with word-boundary regex
# so names like "Schulenburg", "Kammerer", "Neustadt", "Speicherstadt" are not rejected.
_DE_NEGATIVE_DOMAIN_WORDS: list = [
    "stadt", "gemeinde", "landkreis", "kreisverwaltung", "rathaus",
    "ihk", "handwerkskammer", "universitaet", "universitä",
    "hochschule", "schule", "kita", "kindergarten",
    "verein", "kirche", "pfarr", "behoerde", "behoerd",
]
# Compiled as whole-token pattern: domain split on [-_.] must contain exact token
_DE_NEGATIVE_DOMAIN_RE = re.compile(
    r"(?<![a-z])(?:"
    + "|".join(re.escape(w) for w in _DE_NEGATIVE_DOMAIN_WORDS)
    + r")(?![a-z])",
    re.IGNORECASE,
)

_DE_LEGAL_TOKENS: list = [
    "gmbh", "ag", "kg", "ohg", "gbr", "kgaa", "se", "ug",
    "gmbh & co. kg", "gmbh & co.kg", "gmbh&co.kg",
    "eingetragener kaufmann", "e.k.", "e. k.", "e.kfm.",
]

_DE_DESCRIPTOR_PATTERNS: list = [
    (r"\bUnternehmen\b", ""),
    (r"\bGesellschaft\b", ""),
    (r"\bBetrieb\b", ""),
    (r"\bWerkstatt\b", ""),
    (r"\bHandel\b", ""),
    (r"\bBau\b", ""),
]


IT_CONFIG = CountryConfig(
    country_code          = "IT",
    country_name          = "Italy",
    serper_gl             = "it",
    serper_hl             = "it",
    country_search_name   = "Italy",
    official_search_terms = ["sito ufficiale", "official website"],
    preferred_tlds        = [".it"],
    directory_blacklist_domains = set(),   # existing _GENERIC_DOMAINS already handles IT
    directory_blacklist_bases   = set(),
    email_provider_blacklist    = set(),   # existing _PEC_DOMAIN_PATTERNS handles IT
    legal_tokens          = [],            # existing Italy logic handles legal forms
    descriptor_cleanup_patterns = [],
    firecrawl_location    = {"country": "IT", "languages": ["it", "en"]},
    name_col_primary      = "",
    name_col_fallback     = "",
    city_col_primary      = "",
    state_col             = "",
    address_col           = "",
    negative_domain_patterns = [],
)

DE_CONFIG = CountryConfig(
    country_code          = "DE",
    country_name          = "Germany",
    serper_gl             = "de",
    serper_hl             = "de",
    country_search_name   = "Deutschland",
    official_search_terms = ["offizielle Website", "official website", "Impressum", "Kontakt"],
    preferred_tlds        = [".de", ".com", ".eu"],
    directory_blacklist_domains = _DE_DIRECTORY_DOMAINS,
    directory_blacklist_bases   = _DE_DIRECTORY_BASES,
    email_provider_blacklist    = _DE_EMAIL_GENERICS,
    legal_tokens          = _DE_LEGAL_TOKENS,
    descriptor_cleanup_patterns = _DE_DESCRIPTOR_PATTERNS,
    firecrawl_location    = {"country": "DE", "languages": ["de", "en"]},
    name_col_primary      = "company_name_clean",
    name_col_fallback     = "company_name_raw",
    city_col_primary      = "city_or_registered_office",
    state_col             = "federal_state",
    address_col           = "registered_address",
    negative_domain_patterns = _DE_NEGATIVE_DOMAIN_WORDS,
)

COUNTRY_CONFIGS: dict[str, CountryConfig] = {
    "IT": IT_CONFIG,
    "DE": DE_CONFIG,
}


def detect_country_from_path(input_path: str) -> str | None:
    """Infer country code from pipeline folder structure or filename."""
    p = input_path.replace("\\", "/").lower()
    # Folder-level detection (pipeline structure)
    if "/germany/" in p or "/germany\\" in input_path.lower():
        return "DE"
    if "/italy" in p:
        return "IT"
    # Filename-level detection (bare upload names like Germany_1_R0001_0500.xlsx)
    import os as _os
    fname = _os.path.basename(p)
    if fname.startswith("germany"):
        return "DE"
    if fname.startswith("italy"):
        return "IT"
    return None


def detect_country_from_columns(df: "pd.DataFrame") -> str | None:
    """Infer country code by checking for known country-specific column names."""
    cols_lc = {str(c).lower().strip() for c in df.columns}
    de_cols = {"company_name_clean", "company_name_raw", "city_or_registered_office",
               "federal_state", "company_number", "register_nummer"}
    it_cols = {"company name", "national statistical institute province",
               "email address", "postal code"}
    if cols_lc & de_cols:
        return "DE"
    if cols_lc & it_cols:
        return "IT"
    return None


def resolve_country(cli_country: str, input_path: str, df: "pd.DataFrame") -> str:
    """Return the resolved country code ("IT" or "DE") from CLI arg, path, or columns."""
    if cli_country and cli_country.upper() in COUNTRY_CONFIGS:
        return cli_country.upper()
    from_path = detect_country_from_path(input_path)
    if from_path:
        return from_path
    from_cols = detect_country_from_columns(df)
    if from_cols:
        return from_cols
    return "IT"   # backward-compatible default


def detect_columns_generic(df: "pd.DataFrame", config: "CountryConfig") -> dict:
    """
    Detect column roles for both Italian and German inputs.

    For Germany: tries config-specific primary/fallback columns first,
    then falls back to the standard detect_columns() for anything not found.
    For Italy: delegates entirely to existing detect_columns().
    """
    if config.country_code == "IT":
        return detect_columns(df)

    # Germany: map canonical roles from known DE column names
    cols_avail = set(df.columns)

    def _first(*candidates):
        for c in candidates:
            if c and c in cols_avail:
                return c
        return None

    col_map = {
        "company": _first(
            config.name_col_primary, config.name_col_fallback,
            "company_name", "name",
        ),
        "website": _first(
            "website", "domain", "url", "homepage",
            "canonical_company_url", "validated_domain",
        ),
        "email": _first(
            "email", "email address", "email_address",
            "kontakt_email", "contact_email",
        ),
        "city": _first(
            config.city_col_primary,
            "registered_office", "city", "ort",
        ),
        "province": _first(
            config.state_col,
            "state", "bundesland", "province",
        ),
        "postcode": _first(
            "postcode", "postal_code", "plz",
        ),
        "phone": _first(
            "phone", "phone_number", "telefon", "tel",
        ),
    }
    return col_map


def normalize_register_columns_for_cleaner(
    df: "pd.DataFrame",
    cfg: "CountryConfig",
) -> "tuple[pd.DataFrame, dict]":
    """
    Return a (copy_of_df, mapping_report) where canonical columns used by
    process_dataframe are guaranteed to exist.

    Canonical columns added (never overwrite if already present and non-blank):
        company_name, website, email, city, province, postcode, phone,
        country_code, country_name, legal_form, registered_address

    The mapping_report dict records which source column was chosen for each role.
    Original columns are always preserved.
    """
    out = df.copy()
    report: dict = {}

    cols_avail = set(df.columns)

    def _coalesce(*candidates) -> tuple[str, str]:
        """Return (source_col_name, value_from_first_non_blank_col)."""
        for c in candidates:
            if c and c in cols_avail:
                return c, None  # lazy — actual values resolved per-row
        return "", ""

    def _add_canonical(canonical: str, *candidates: str) -> None:
        """Add `canonical` column by picking first non-blank candidate column."""
        # If already present and at least one non-blank value → keep it
        if canonical in cols_avail:
            non_blank = out[canonical].astype(str).str.strip().ne("").any()
            if non_blank:
                report[canonical] = canonical  # identity
                return
        # Find first candidate that exists and has at least one non-blank value
        for c in candidates:
            if c and c in cols_avail:
                has_data = out[c].astype(str).str.strip().ne("").any()
                if has_data:
                    out[canonical] = out[c].astype(str).str.strip()
                    report[canonical] = c
                    return
        # All blank — still add the column (empty) from first existing candidate
        for c in candidates:
            if c and c in cols_avail:
                out[canonical] = out[c].astype(str).str.strip()
                report[canonical] = f"{c} (blank)"
                return
        # Column doesn't exist at all
        out[canonical] = ""
        report[canonical] = "(not found)"

    if cfg.country_code == "DE":
        _add_canonical("company_name",
            "company_name_clean", "company_name_raw", "Company Name", "company_name", "name")
        _add_canonical("website",
            "website", "Website", "domain", "url", "homepage",
            "canonical_company_url", "validated_domain")
        _add_canonical("email",
            "email", "Email", "email_address", "contact_email", "kontakt_email")
        _add_canonical("city",
            "city_or_registered_office", "registered_office", "city", "ort", "City")
        _add_canonical("province",
            "federal_state", "bundesland", "state", "province")
        _add_canonical("postcode",
            "postcode", "postal_code", "plz", "Postal Code")
        _add_canonical("phone",
            "phone", "phone_number", "telefon", "tel", "Phone number")
        _add_canonical("legal_form",
            "legal_form_detected", "legal_form", "rechtsform")
        _add_canonical("registered_address",
            "registered_address", "address", "anschrift")
        out["country_code"] = "DE"
        out["country_name"] = "Germany"
        report["country_code"] = "constant"
        report["country_name"] = "constant"
    else:  # IT
        _add_canonical("company_name",
            "Company Name", "company_name", "company_name_clean",
            "ragione sociale", "denominazione")
        _add_canonical("website",
            "Website", "website", "domain", "url")
        _add_canonical("email",
            "Email address", "email", "email_address")
        _add_canonical("city",
            "City", "city", "comune")
        _add_canonical("province",
            "National statistical institute Province", "Province", "province", "provincia")
        _add_canonical("postcode",
            "Postal Code", "postal_code", "postcode", "cap")
        _add_canonical("phone",
            "Phone number", "phone", "telefono")
        _add_canonical("legal_form",
            "legal_form_detected", "legal_form", "forma giuridica")
        _add_canonical("registered_address",
            "registered_address", "address", "indirizzo")
        out["country_code"] = "IT"
        out["country_name"] = "Italy"
        report["country_code"] = "constant"
        report["country_name"] = "constant"

    return out, report


def _cols_from_normalized(run_df: "pd.DataFrame", cfg: "CountryConfig") -> dict:
    """
    Build a cols dict that always points to canonical column names.
    Falls back to detect_columns_generic if a canonical column is entirely blank.
    The fallback itself is validated: blank fallback columns are never returned.
    """
    canonical_map = {
        "company":  "company_name",
        "website":  "website",
        "email":    "email",
        "city":     "city",
        "province": "province",
        "postcode": "postcode",
        "phone":    "phone",
    }
    cols: dict = {}
    fallback = detect_columns_generic(run_df, cfg)

    def _col_has_data(col_name: str) -> bool:
        if not col_name or col_name not in run_df.columns:
            return False
        return run_df[col_name].astype(str).str.strip().ne("").any()

    for role, canon in canonical_map.items():
        if _col_has_data(canon):
            cols[role] = canon
        else:
            fb = fallback.get(role)
            cols[role] = fb if _col_has_data(fb) else None
    return cols


def _is_de_negative_domain(domain: str) -> tuple[bool, str]:
    """
    Return (True, reason) when domain contains a German public/non-commercial token
    as a whole word (split on -, _, .), so names like Schulenburg are safe.
    """
    d = (domain or "").lower()
    parts = set(re.split(r"[-._/]", d))
    for word in _DE_NEGATIVE_DOMAIN_WORDS:
        if word in parts:
            return True, f"German public/non-commercial token in domain: {word}"
    return False, ""



# Per-speed-mode defaults: (max_cands, max_pages, timeout_secs)
_FC_SPEED_DEFAULTS  = {
    _FC_SPEED_FAST:     (1, 1, 6),
    _FC_SPEED_BALANCED: (1, 2, 8),
    _FC_SPEED_THOROUGH: (3, 3, 15),
}

# Negative source-type patterns that block replacement
_NEG_SOURCE_RES: dict[str, re.Pattern] = {
    "news_media":       re.compile(
        # Only flag pages that are clearly third-party editorial/journalism
        r"\b(testata\s+giornalistica|giornale\s+online|quotidiano|settimanale|mensile|"
        r"rivista\s+di\s+settore|redazione\s+di|notizie\s+di\s+cronaca|ultime\s+notizie|"
        r"breaking\s+news|newsroom|press\s+release\s+wire|comunicato\s+stampa\s+agenzia)\b",
        re.I),
    "directory":        re.compile(
        r"\b(fatturato|bilancio|visura|scheda\s+azienda|scheda\s+impresa|"
        r"company\s+profile|business\s+profile|registro\s+imprese|dati\s+aziendali)\b", re.I),
    "foundation":       re.compile(
        r"\b(fondazione(?!\s+di\s+(?:Pompe|Delta|Garbarino))|onlus|ente\s+non\s+profit)\b", re.I),
    "event_conference": re.compile(
        r"\b(conferenza|convegno|fiera\s+di|exhibition|congress|summit)\b", re.I),
    "association":      re.compile(
        r"\b(associazione\s+di\s+categoria|federazione\s+nazionale|confindustria|confcommercio)\b", re.I),
    "marketplace":      re.compile(
        r"\b(confronta\s+prezzi|trova\s+prezzi|acquista\s+online|e-commerce\s+shop)\b", re.I),
    "dealer_reseller":  re.compile(
        r"\b(concessionaria\s+ufficiale|rivenditore\s+autorizzato|franchising)\b", re.I),
    "job_board":        re.compile(
        r"\b(offerte\s+di\s+lavoro\s+candidatura|curriculum\s+vitae\s+candidati|job\s+listing)\b", re.I),
    "government":       re.compile(
        r"\b(comune\s+di|regione\s+[a-z]+|provincia\s+di|ministero|agenzia\s+delle\s+entrate)\b", re.I),
    "public_school":    re.compile(
        r"\b(istituto\s+tecnico\s+statale|istituto\s+statale|istituto\s+comprensivo|"
        r"liceo\s+statale|scuola\s+media\s+statale|scuola\s+primaria|"
        r"ministero\s+dell.istruzione|ministero\s+dell.istruzione\s+e\s+del\s+merito|"
        r"scuola\s+in\s+chiaro|ptof\b|piano\s+triennale\s+dell.offerta\s+formativa|"
        r"\bdocenti?\b.*\bstudenti?\b|\bstudenti?\b.*\bdocenti?\b|"
        r"ptof|piano\s+triennale|registro\s+elettronico|docenti|studenti|"
        r"scuola\s+in\s+chiaro|ministero\s+dell.istruzione|miur|ministero\s+istruzione)\b",
        re.I),
    "university":       re.compile(
        r"\b(universit[àa]\s+degli\s+studi|universit[àa]\s+di|dipartimento\s+di|"
        r"facolt[àa]\s+di|corso\s+di\s+laurea|rettore|prorettore|dottorato\s+di\s+ricerca)\b",
        re.I),
}

# Public school / education signal — applied to the domain TLD as well
_PUBLIC_SCHOOL_RE = re.compile(
    r"\b(istituto\s+tecnico\s+statale|istituto\s+statale|istituto\s+comprensivo|"
    r"liceo\s+statale|scuola\s+media\s+statale|scuola\s+primaria|"
    r"ministero\s+dell.istruzione|ministero\s+dell.istruzione\s+e\s+del\s+merito|"
    r"scuola\s+in\s+chiaro|ptof\b|piano\s+triennale\s+dell.offerta\s+formativa|"
    r"\bdocenti?\b|\bstudenti?\b|istruzione\s+e\s+merito)\b",
    re.I,
)

# Domain TLD/SLD patterns that signal non-commercial entities for Italian companies
_EDU_IT_RE = re.compile(r"\.edu\.it$", re.I)
_GOV_IT_RE = re.compile(r"\.(gov|governo|istruzione|giustizia|interno|esteri)\.it$", re.I)

# Hard negative source types — always block even for High-confidence Python domains
_HARD_NEG_SOURCES = frozenset({
    "directory", "government", "marketplace", "job_board",
    "public_school", "university",
    # v12 — URL/domain-identified hard negatives
    "business_directory", "marketplace_product_page", "pdf_list",
    "document_hosting", "third_party_client_reference",
})

_HAIKU_MODES          = [_HAIKU_MODE_PYTHON, _HAIKU_MODE_UNCERTAIN, _HAIKU_MODE_ALL]

_HAIKU_SYSTEM_PROMPT = (
    "You are a conservative B2B sales intelligence reviewer for Italian companies.\n"
    "Your task: given a company name, location, and candidate search results, decide whether "
    "the Python-selected domain is the correct official website.\n\n"
    "Classify the Python domain as one of:\n"
    "  accept          — confirmed official company or group/subsidiary site\n"
    "  replace         — a different domain already in the evidence is clearly better\n"
    "  reject          — none of the results represent the real official site\n"
    "  uncertain       — insufficient evidence to decide\n"
    "  needs_firecrawl — plausible candidate, but live page confirmation required\n\n"
    "The search results include:\n"
    "  [SCORED]   — passed Python brand-similarity filter; scored and ranked\n"
    "  [FILTERED] — removed by Python heuristics; you may override if clearly correct\n\n"
    "IMPORTANT DOMAIN RESTRICTION:\n"
    "You may ONLY suggest a domain that already appears in the evidence block.\n"
    "You must NOT invent, guess, or construct any domain not present in the evidence.\n\n"
    "For group/subsidiary pages, accept when ALL of the following hold:\n"
    "  - the candidate URL path contains /aziende/, /companies/, /company/, /business-sector/,\n"
    "    /business-sectors/, /settori/, /settori-di-attivita/, /subsidiary/, /societa/, /gruppo/\n"
    "  - the title or snippet contains the company name or a close variant\n"
    "  - the snippet contains at least one of: activity, contact, registered office, city,\n"
    "    province, address, or business sector\n\n"
    "Risk flags to include when relevant (use exact strings):\n"
    "  group_domain_not_company_domain, snippet_only_evidence, no_address_city_evidence,\n"
    "  name_domain_mismatch, possible_directory_profile, generic_or_ambiguous_name,\n"
    "  no_location_match\n\n"
    "Return ONLY a JSON object. No text before or after. No markdown fences.\n"
    "Use exactly this schema:\n"
    "{\n"
    "  \"decision\": \"accept|replace|reject|uncertain|needs_firecrawl\",\n"
    "  \"domain\": \"selected domain or empty string\",\n"
    "  \"candidate_url\": \"exact full URL from evidence or empty string\",\n"
    "  \"confidence\": \"High|Medium|Low|None\",\n"
    "  \"reason\": \"max 180 chars\",\n"
    "  \"risk_flags\": [\"...\"],\n"
    "  \"recommended_action\": \"skip_firecrawl|firecrawl_exact_url|firecrawl_root|manual_review\"\n"
    "}"
)

_HAIKU_USER_TEMPLATE = (
    "Company: {company_name}\n"
    "Location: {city}, {province} (Italy)\n"
    "Email domain: {email_domain}\n"
    "Original website in register: {original_website}\n"
    "Python-suggested domain: {python_domain} (confidence: {python_confidence})\n\n"
    "Search results (pay attention to full URLs and path segments):\n{results_block}\n\n"
    "Decide whether the Python-suggested domain is the correct official website for this "
    "Italian company. If a group/subsidiary page URL exists, evaluate it carefully. "
    "Reply with JSON only."
)

# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================


def _normalize_col_key(col: str) -> str:
    return re.sub(r"[\s_\-]+", " ", str(col).strip().lower())


def normalize_domain(raw: str) -> str:
    """Strip protocol, www, path, query. Return root domain lowercase."""
    if not raw or not isinstance(raw, str):
        return ""
    d = raw.strip().lower()
    d = re.sub(r"^https?://", "", d)
    d = re.sub(r"^www\.", "", d)
    d = d.split("/")[0].split("?")[0].split("#")[0].strip()
    if not d or " " in d or d in ("nan", "none", "n/a", "-", "—"):
        return ""
    if "." not in d:
        return ""
    return d


def split_multi_website(raw: str) -> list[str]:
    """
    Split a website field that may contain multiple URLs separated by
    commas, semicolons, spaces, or double spaces. Return normalised domains.
    """
    if not raw or not isinstance(raw, str):
        return []
    parts = re.split(r"[,;\s]+", raw.strip())
    domains = []
    for p in parts:
        d = normalize_domain(p)
        if d:
            domains.append(d)
    return domains


def best_website_domain(raw: str) -> str:
    """
    Parse a multi-website field and return the best single domain.
    Prefers non-generic, then .it TLD.
    """
    domains = split_multi_website(raw)
    if not domains:
        return ""
    non_generic = [d for d in domains if not is_generic(d)]
    if not non_generic:
        return ""
    it_domains = [d for d in non_generic if d.endswith(".it")]
    return it_domains[0] if it_domains else non_generic[0]


def strip_legal(name: str) -> str:
    cleaned = _LEGAL_TOKENS.sub(" ", name)
    return re.sub(r"\s+", " ", cleaned).strip(" .,/-")


def strip_descriptors(name: str) -> str:
    """Remove legal suffixes AND Italian descriptor words."""
    cleaned = _LEGAL_TOKENS.sub(" ", name)
    cleaned = _ITALIAN_DESCRIPTORS.sub(" ", cleaned)
    return re.sub(r"\s+", " ", cleaned).strip(" .,/-")


def _pre_clean_register_name(name: str) -> str:
    """
    Pre-process Italian Chamber of Commerce long-form names before variant extraction.

    Resolution priority (first match wins):
      1. "IN FORMA ABBREVIATA …" → use text after phrase
      2. "O IN BREVE …" / "IN BREVE …" → use text after phrase
      3. "ANCHE COME …" / "DA INDICARE ANCHE COME …" → use text after phrase
      4. "IN SIGLA …" / "SIGLABILE …" → strip phrase and everything after (keep main name)

    After resolution, strips residual legal boilerplate, apostrophes, and normalises
    whitespace. Trailing dots are preserved (e.g. "S.P.A.").
    """
    s = name.strip()

    # Priority 1: IN FORMA ABBREVIATA
    m = _FORMA_ABBREVIATA_RE.search(s)
    if m:
        short = m.group(1).strip().strip(",").strip()
        if short:
            s = short
    else:
        # Priority 2: O IN BREVE / IN BREVE
        m = _IN_BREVE_RE.search(s)
        if m:
            short = m.group(1).strip().strip(",").strip()
            if short:
                s = short
        else:
            # Priority 3: ANCHE COME / DA INDICARE ANCHE COME
            m = _ANCHE_COME_RE.search(s)
            if m:
                short = m.group(1).strip().strip(",").strip()
                if short:
                    s = short
            else:
                # Priority 4: IN SIGLA / SIGLABILE — keep only main name before phrase
                s = _IN_SIGLA_RE.sub("", s).strip()

    # Strip remaining extra legal phrases not covered by _LEGAL_TOKENS
    s = _EXTRA_LEGAL_PHRASES_RE.sub(" ", s)

    # Remove stray apostrophes (e.g. left by SOCIETA')
    s = s.replace("'", "")

    # Collapse whitespace; trim only leading/trailing spaces and commas — not dots
    # (dots are part of legal abbreviations like S.P.A.)
    s = re.sub(r"\s+", " ", s).strip(" ,")
    return s


_ACRONYM_RE = re.compile(r"^[A-Z](\.[A-Z])+\.?$")


def _is_acronym(token: str) -> bool:
    """True if token looks like a dotted acronym: I.M.E.S.A. or I.T."""
    return bool(_ACRONYM_RE.match(token.strip()))


def _acronym_nodot(token: str) -> str:
    """Return the dotless form of a dotted acronym: I.M.E.S.A. → IMESA."""
    return re.sub(r"\.", "", token).upper()


def extract_name_variants(name: str) -> dict:
    """
    Build multiple name variants for search query generation.

    Returns dict with keys:
      full            — cleaned name (pre-processed to remove register legal boilerplate)
      no_legal        — full with legal suffix removed
      no_desc         — full with legal suffix + Italian descriptors removed
      brand           — the core brand string used for domain matching and focused queries
      brand_nodot     — for dotted acronyms (I.M.E.S.A. → IMESA), else same as brand
      original        — raw input name before any cleaning
      is_acronym      — True if brand is a dotted acronym
    """
    original = name.strip()

    # ── Extract SIGLABILE brand before stripping ──────────────────────────────
    _siglabile_raw = ""
    _siglabile_m = _SIGLABILE_EXTRACT_RE.search(original)
    if _siglabile_m:
        _siglabile_raw = _siglabile_m.group(1).strip().rstrip(")].").strip()

    siglabile_variants: list[str] = []
    if _siglabile_raw:
        _sig_no_legal = strip_legal(_siglabile_raw)
        _sig_compact = re.sub(r"[^a-zA-Z0-9]", "", _sig_no_legal).lower()
        if _sig_no_legal:
            siglabile_variants.append(_sig_no_legal)
        if _sig_compact and _sig_compact != _sig_no_legal.lower():
            siglabile_variants.append(_sig_compact)

    # ── First-token brand rule ────────────────────────────────────────────────
    first_token_brand = ""
    _name_tokens = re.split(r"[\s\-_/&,]+", original)
    if len(_name_tokens) >= 2:
        _ft = _name_tokens[0].strip().rstrip(".'")
        _ft_clean = re.sub(r"[^A-Za-z0-9]", "", _ft)
        if (3 <= len(_ft_clean) <= 6
                and _ft_clean.upper() == _ft_clean
                and _ft_clean.lower() not in _LEGAL_SUFFIX_TOKENS
                and _GENERIC_DESCRIPTORS_RE.match(_name_tokens[1].strip())):
            first_token_brand = _ft_clean

    # Pre-clean: resolve IN FORMA ABBREVIATA and strip extra legal phrases
    full = _pre_clean_register_name(original)
    if not full:
        full = original  # safety fallback

    no_legal = strip_legal(full)
    no_desc  = strip_descriptors(full)

    # Dotted-acronym protection: strip_legal can corrupt "I.M.E.S.A." → "I.M.E"
    # because the sub-pattern s\.?a\.?s? matches .S.A. inside the acronym.
    # If the stripped result is much shorter AND the original full contains a dotted
    # acronym, use full tokens directly (legal suffix at end already removed by full split).
    _full_toks_raw = [
        t for t in re.split(r"[\s\-_/&,]+", full)
        if len(t) >= 2 and t.lower() not in _NOISE_TOKENS and not re.match(r"^\d+$", t)
    ]
    _has_acronym_in_full = any(_is_acronym(t) for t in _full_toks_raw)
    _stripping_corrupted = (
        _has_acronym_in_full
        and len(no_desc.strip()) < len(full.strip()) * 0.6
    )
    if _stripping_corrupted:
        # Use full tokens; exclude standalone legal-suffix tokens at end
        _legal_suffix_re = re.compile(r"^[Ss]\.?[Pp]?\.?[Aa]?\.?$|^[Ss]\.[Pp]\.[Aa]\.$", re.I)
        raw_toks = [t for t in _full_toks_raw if not _legal_suffix_re.match(t)]
        if not raw_toks:
            raw_toks = _full_toks_raw
        # Also fix no_desc so subsequent brand logic has a clean base
        no_desc = " ".join(raw_toks)
    else:
        # Normal path
        raw_toks = [
            t for t in re.split(r"[\s\-_/&,]+", no_desc)
            if len(t) >= 2 and t.lower() not in _NOISE_TOKENS
            and not re.match(r"^\d+$", t)
        ]

    if not raw_toks:
        # Strip trailing punctuation/operators from fallback (e.g. "Global &" → "Global")
        _fb = no_desc or no_legal or full
        brand = re.sub(r"[\s&,./\\-]+$", "", _fb).strip()
        brand = re.sub(r"^[\s&,./\\-]+", "", brand).strip()
        if not brand:
            brand = full
    elif len(raw_toks) == 1:
        brand = raw_toks[0]
    elif len(raw_toks) == 2:
        # Two-token brand names should be kept together: "POMPE GARBARINO", "SAN NICOLA"
        brand = " ".join(raw_toks)
    else:
        # Three or more tokens: keep last two significant long tokens together
        long_toks = [t for t in raw_toks if len(t) >= 4]
        if len(long_toks) >= 2:
            brand = " ".join(long_toks[-2:])
        elif long_toks:
            brand = long_toks[-1]
        else:
            brand = raw_toks[-1]

    # Acronym handling: if brand is a single dotted-acronym token, build nodot form
    brand_is_acronym = _is_acronym(brand)
    brand_nodot = _acronym_nodot(brand) if brand_is_acronym else brand

    # ── Descriptor transformation variants ────────────────────────────────────
    descriptor_variants: list[tuple[str, str]] = []
    _nolegal_upper = (no_legal or "").upper()

    if "RISERIA" in _nolegal_upper:
        _riso = re.sub(r"\bRISERIA\b", "RISO", no_legal, flags=re.I)
        descriptor_variants.append(("riso_variant", _riso.strip()))

    if "IMMOBILIARE" in _nolegal_upper:
        _re_variant = re.sub(r"\bIMMOBILIARE\b", "REAL ESTATE", no_legal, flags=re.I)
        descriptor_variants.append(("real_estate_variant", _re_variant.strip()))
        _compact = re.sub(r"\s+", "", _re_variant.lower())
        if len(_compact) <= 30:
            descriptor_variants.append(("real_estate_compact", _compact))

    if "TECNOLOGIE" in _nolegal_upper:
        _tech_variant = re.sub(r"\bTECNOLOGIE\b", "TECHNOLOGIES", no_legal, flags=re.I)
        descriptor_variants.append(("technologies_variant", _tech_variant.strip()))

    # Safety: if the resulting brand is purely a legal suffix (e.g. "KG", "AG"),
    # fall back to no_legal or full so we never search with a generic legal token.
    _brand_lower = re.sub(r"[^a-z0-9]", "", brand.lower())
    if _brand_lower in _BRAND_LEGAL_REJECT_TOKENS:
        # Use the cleaned name without descriptors, fall back to no_legal, then full
        brand = no_desc.strip() or no_legal.strip() or full
        # If that also reduces to a legal token, use full
        if re.sub(r"[^a-z0-9]", "", brand.lower()) in _BRAND_LEGAL_REJECT_TOKENS:
            brand = full
        brand_is_acronym = _is_acronym(brand)
        brand_nodot = _acronym_nodot(brand) if brand_is_acronym else brand

    # Brand-is-generic guard for German names: a single generic token like "Global"
    # must never be treated as a distinctive brand signal.
    _brand_stripped = re.sub(r"[^a-z]", "", brand.lower())
    brand_is_de_generic = (
        bool(_brand_stripped)
        and " " not in brand.strip()
        and _brand_stripped in _DE_GENERIC_BRAND_TOKENS
    )

    return {
        "full":               full,
        "no_legal":           no_legal,
        "no_desc":            no_desc,
        "brand":              brand,
        "brand_nodot":        brand_nodot,
        "original":           original,
        "is_acronym":         brand_is_acronym,
        "siglabile_brand":    siglabile_variants[0] if siglabile_variants else "",
        "siglabile_compact":  siglabile_variants[1] if len(siglabile_variants) > 1 else "",
        "first_token_brand":  first_token_brand,
        "descriptor_variants": descriptor_variants,
        "brand_is_de_generic": brand_is_de_generic,
    }


def company_tokens(name: str) -> set:
    clean = strip_legal(name)
    clean = re.sub(r"[^\w\s\-]", " ", clean)
    toks = {t.lower() for t in re.split(r"[\s\-_]+", clean) if len(t) >= 2}
    return toks - _NOISE_TOKENS


def domain_tokens(domain: str) -> set:
    if not domain:
        return set()
    parts = domain.split(".")
    while len(parts) > 1 and parts[-1].lower() in _TLDS:
        parts = parts[:-1]
    base = ".".join(parts)
    toks = {t for t in re.split(r"[-.]", base.lower()) if t and len(t) >= 2}
    return toks - _TLDS


def token_overlap(name: str, domain: str) -> float:
    """Overlap between company name tokens and domain tokens."""
    ctok = company_tokens(name)
    dtok = domain_tokens(domain)
    if not ctok or not dtok:
        return 0.0
    overlap: set = ctok & dtok
    for c in ctok:
        for d in dtok:
            if c in d or d in c:
                overlap.add(c)
    return len(overlap) / min(len(ctok), len(dtok))


def brand_overlap(brand: str, domain: str) -> float:
    """
    Direct brand-name / domain overlap.
    Returns 1.0 if brand (lowercased, stripped) appears literally in domain base.
    For dotted acronyms (I.M.E.S.A.) also tries the nodot form (IMESA).
    """
    if not brand or not domain:
        return 0.0
    b = re.sub(r"[^\w]", "", brand.lower())
    # Get domain base (strip TLD)
    parts = domain.split(".")
    while len(parts) > 1 and parts[-1].lower() in _TLDS:
        parts = parts[:-1]
    base = re.sub(r"[^\w]", "", ".".join(parts).lower())
    if not b or not base:
        return 0.0
    if b == base:
        return 1.0
    if b in base or base in b:
        return 0.8
    # Token-level check
    b_toks = set(re.split(r"[-.]", b)) - _TLDS
    base_toks = set(re.split(r"[-.]", base)) - _TLDS
    if b_toks and base_toks:
        hit = b_toks & base_toks
        return len(hit) / min(len(b_toks), len(base_toks))
    return 0.0


def brand_overlap_variants(name_variants: dict, domain: str) -> float:
    """Return the best brand_overlap across brand and brand_nodot variants."""
    bo = brand_overlap(name_variants.get("brand", ""), domain)
    if name_variants.get("is_acronym"):
        bo = max(bo, brand_overlap(name_variants.get("brand_nodot", ""), domain))
    return bo


def is_generic(domain: str, country_config: "CountryConfig | None" = None) -> bool:
    """Return True if domain is in the generic/directory blacklist (incl. subdomains)."""
    if not domain:
        return False
    dl = domain.lower()
    if dl in _GENERIC_DOMAINS:
        return True
    # Check subdomain containment for known bad base domains
    for base in _GENERIC_DOMAIN_BASES:
        if dl == base or dl.endswith("." + base):
            return True
    # Country-specific additional blacklists
    if country_config:
        if dl in country_config.directory_blacklist_domains:
            return True
        for base in country_config.directory_blacklist_bases:
            if dl == base or dl.endswith("." + base):
                return True
    return False


def is_discovery_blocked(domain: str) -> bool:
    """
    Return True if the domain must never become an official company domain
    (music stores, streaming, social media, marketplaces, encyclopedias).
    These may remain in Raw Search Evidence as rejected evidence but must not
    be validated, recommended, or sent to Firecrawl.
    """
    if not domain:
        return False
    dl = domain.lower()
    for blocked in _DISCOVERY_BLOCKED_DOMAINS:
        if dl == blocked or dl.endswith("." + blocked):
            return True
    return False


# URL shortener domains — must never become validated/recommended/final domain,
# must never be sent to Firecrawl, must never be used in size inference.
_URL_SHORTENER_DOMAINS: frozenset = frozenset({
    "t.co", "bit.ly", "bitly.com", "tinyurl.com", "ow.ly", "buff.ly",
    "shorturl.at", "rebrand.ly", "cutt.ly", "lnkd.in", "linktr.ee",
    "goo.gl", "is.gd", "s.id", "trib.al",
})


def is_url_shortener(domain: str) -> bool:
    """Return True if the domain is a known URL shortener service."""
    if not domain:
        return False
    dl = domain.lower()
    return any(dl == blocked or dl.endswith("." + blocked) for blocked in _URL_SHORTENER_DOMAINS)


def is_pec_or_personal_email(email_domain: str) -> bool:
    """Return True if the email domain is a PEC provider or personal mailbox."""
    if not email_domain:
        return True
    dl = email_domain.lower()
    return any(dl == pat or dl.endswith("." + pat) for pat in _PEC_DOMAIN_PATTERNS)


def extract_email_domain(email: str) -> str:
    """Extract the domain part of an email address."""
    if not email or "@" not in email:
        return ""
    parts = email.strip().split("@")
    if len(parts) < 2:
        return ""
    domain = parts[-1].strip().lower()
    return domain if "." in domain else ""


def location_in_text(text: str, city: str, province: str) -> bool:
    """Return True if city or province appears in the text snippet/title."""
    t = text.lower()
    if city and len(city) >= 3 and city.lower() in t:
        return True
    if province and len(province) >= 2 and province.lower() in t:
        return True
    return False


def has_official_signal(text: str) -> bool:
    """Return True if text contains official-page keywords."""
    tl = text.lower()
    return any(sig in tl for sig in _OFFICIAL_SIGNALS)


def classify_domain(domain: str, title: str = "", snippet: str = "") -> str | None:
    """
    Return a rejection-category string if the domain belongs to a known
    non-commercial category, or None if the domain looks acceptable.

    Categories: "government" | "religious" | "directory" | "academic" | None

    Checks domain string first; for religious also checks the page title
    because a domain like 'sannicola.it' is ambiguous without title context.
    Also rejects pages that look like business-profile/financial-data pages
    based on strong title/snippet signals (e.g. "fatturato", "visura").
    """
    dl = domain.lower()

    if _GOVT_PATTERNS.search(dl):
        return "government"

    # Religious: domain match OR title match (e.g. "Basilica di San Nicola" in title)
    if _RELIGIOUS_PATTERNS.search(dl) or _RELIGIOUS_PATTERNS.search(title.lower()):
        return "religious"

    if _DIRECTORY_EXTRA_PATTERNS.search(dl):
        return "directory"

    if _ACADEMIC_PATTERNS.search(dl):
        return "academic"

    # Content-based: title or snippet strongly signals a business-profile/directory page
    combined = (title + " " + snippet).lower()
    if _DIRECTORY_PROFILE_TITLE_SIGNALS.search(combined):
        return "directory"

    return None


def _conf_label(conf: float) -> str:
    if conf >= 0.70:
        return "High"
    if conf >= 0.40:
        return "Medium"
    return "Low"


def is_hosted_platform(domain: str) -> bool:
    """Return True if domain is a subdomain of a free hosted-site platform."""
    dl = (domain or "").lower()
    return any(dl == base or dl.endswith("." + base) for base in _HOSTED_PLATFORM_BASES)


# Weak/generic tokens that alone cannot justify accepting a candidate domain.
_WEAK_BRAND_TOKENS: frozenset = frozenset({
    "global", "group", "services", "solutions", "italia", "italy",
    "industry", "industries", "holding", "holdings", "international",
    "management", "consulting", "digital", "technology", "technologies",
    "systems", "system", "enterprise", "enterprises", "partners",
})


def evaluate_domain_candidate(
    company_name: str,
    candidate_domain: str,
    title: str = "",
    snippet: str = "",
    url: str = "",
    email_domain: str = "",
    city: str = "",
    province: str = "",
    variants: "dict | None" = None,
    country_config: "CountryConfig | None" = None,
) -> dict:
    """
    Structured quality gate for a single domain candidate.

    Returns::
        {
          "accepted":     bool,
          "score":        int,     # 0-100 scale
          "reason":       str,
          "source_type":  str,     # e.g. "social", "directory", "hosted_platform", "official"
          "needs_review": bool,
        }

    Acceptance thresholds (maps internal 0–3+ float to 0–100):
      score >= 75 → accepted=True,  needs_review=False
      50–74       → accepted=True,  needs_review=True  (suggest but flag)
      < 50        → accepted=False, needs_review=False (reject / leave blank)

    Hard rejections (score=0) override all thresholds.
    """
    cfg = country_config or IT_CONFIG

    def _reject(reason: str, source_type: str) -> dict:
        return {
            "accepted": False, "score": 0,
            "reason": reason, "source_type": source_type,
            "needs_review": False,
        }

    if not candidate_domain:
        return _reject("empty domain", "none")

    dl = candidate_domain.lower()

    # ── Hard rejections ────────────────────────────────────────────────────────
    if is_generic(candidate_domain, country_config=cfg):
        return _reject(f"blacklisted/generic domain: {candidate_domain}", "directory_or_social")

    if is_url_shortener(candidate_domain):
        return _reject(f"URL shortener: {candidate_domain}", "shortener")

    if is_discovery_blocked(candidate_domain):
        return _reject(f"media/streaming/social blocked: {candidate_domain}", "media_blocked")

    if is_hosted_platform(candidate_domain):
        return _reject(
            f"free hosted-site platform (not an official company website): {candidate_domain}",
            "hosted_platform",
        )

    cat = classify_domain(candidate_domain, title, snippet)
    if cat:
        return _reject(f"rejected category: {cat}", cat)

    risky, risky_reason = _domain_has_risky_marker(candidate_domain)
    if risky:
        return _reject(f"risky domain marker — {risky_reason}", "risky_marker")

    # ── Compute internal score and map to 0-100 ────────────────────────────────
    nv = variants or extract_name_variants(company_name)
    internal_score = _score_candidate(
        candidate_domain, 0, title, snippet,
        nv, email_domain, city, province,
        country_config=cfg,
    )

    # Internal score is 0–3+; map to 0–100 (cap at 3.0 for ceiling)
    raw_100 = min(100, int(round(internal_score / 3.0 * 100)))

    # Extra boost for email match (already in internal_score via +0.5, but be explicit)
    if email_domain and candidate_domain == email_domain:
        raw_100 = min(100, raw_100 + 15)

    # Penalty: domain has no meaningful brand overlap (only weak/generic tokens)
    brand = (nv.get("brand") or "").lower().strip()
    brand_tokens = set(re.split(r"[\W_]+", brand)) - _WEAK_BRAND_TOKENS - _TLDS
    if brand_tokens:
        bov = brand_overlap_variants(nv, candidate_domain)
        if bov < _MIN_BRAND_SIM_TO_SCORE:
            raw_100 = min(raw_100, 20)
    else:
        # Brand reduced entirely to weak/generic tokens — demand exact name in title/snippet
        combined = (title + " " + snippet).lower()
        full_name_lower = (nv.get("full") or "").lower()
        if full_name_lower and full_name_lower not in combined:
            raw_100 = min(raw_100, 30)

    score = raw_100
    needs_review = False
    accepted = False

    if score >= 75:
        accepted = True
        needs_review = False
        reason = f"strong match (score {score})"
        source_type = "official"
    elif score >= 50:
        accepted = True
        needs_review = True
        reason = f"plausible match (score {score}) — manual review recommended"
        source_type = "official_uncertain"
    else:
        accepted = False
        needs_review = False
        reason = f"weak or no brand match (score {score})"
        source_type = "low_confidence"

    return {
        "accepted": accepted,
        "score": score,
        "reason": reason,
        "source_type": source_type,
        "needs_review": needs_review,
    }


# =============================================================================
# COLUMN DETECTION
# =============================================================================


def detect_columns(df: pd.DataFrame) -> dict:
    """
    Detect register column names. Returns dict of role → actual_col_name.
    Tries exact match first, then normalised-key fallback.
    """
    cols_norm = {_normalize_col_key(c): c for c in df.columns}

    def _find(candidates, exact_first=None):
        if exact_first and exact_first in df.columns:
            return exact_first
        for cand in candidates:
            key = _normalize_col_key(cand)
            if key in cols_norm:
                return cols_norm[key]
        return None

    return {
        "company":   _find(_NAME_CANDIDATES,     _REG_COL_COMPANY),
        "website":   _find(_WEBSITE_CANDIDATES,  _REG_COL_WEBSITE),
        "email":     _find(_EMAIL_CANDIDATES,    _REG_COL_EMAIL),
        "city":      _find(_CITY_CANDIDATES,     _REG_COL_CITY),
        "province":  _find(_PROVINCE_CANDIDATES, _REG_COL_PROVINCE),
        "postcode":  _find(_POSTCODE_CANDIDATES, _REG_COL_POSTCODE),
        "phone":     _find(_PHONE_CANDIDATES,    _REG_COL_PHONE),
    }


# =============================================================================
# SERPER SEARCH
# =============================================================================


def _call_serper(
    query: str,
    serper_key: str,
    timeout: int = 12,
    gl: str = "it",
    hl: str = "it",
) -> tuple:
    global _serper_call_count
    _serper_call_count += 1
    try:
        resp = requests.post(
            SERPER_URL,
            headers={"X-API-KEY": serper_key, "Content-Type": "application/json"},
            json={"q": query, "gl": gl, "hl": hl, "num": 5},
            timeout=timeout,
        )
        resp.raise_for_status()
        data = resp.json()
        return data.get("organic", []), None
    except requests.Timeout:
        return [], "Serper timeout"
    except Exception as e:
        return [], str(e)


def _extract_domain(url: str) -> str:
    try:
        p = urlparse(url if url.startswith("http") else f"https://{url}")
        host = p.hostname or ""
        return re.sub(r"^www\.", "", host.lower())
    except Exception:
        return ""


_URL_IN_TEXT_RE = re.compile(
    r"(?:https?://|www\.)[a-zA-Z0-9\-._~:/?#\[\]@!$&'()*+,;=%]{4,200}", re.I
)


def _extract_urls_from_text(text: str) -> list:
    """Extract http/https/www URLs embedded in snippet/title text."""
    urls = []
    for m in _URL_IN_TEXT_RE.finditer(text or ""):
        u = m.group(0).rstrip(".,;)'\"")
        if not u.startswith("http"):
            u = "https://" + u
        urls.append(u)
    return urls


_GROUP_PATH_SEGMENTS = (
    "/aziende/",
    "/companies/",
    "/company/",
    "/business-sector/",
    "/business-sectors/",
    "/settori/",
    "/settori-di-attivita/",
    "/subsidiary/",
    "/societa/",
    "/gruppo/",
)


def _classify_candidate_type(url: str, domain: str, title: str, snippet: str) -> str:
    """Classify the candidate type based on URL path and content signals."""
    path = ""
    try:
        path = urlparse(url).path or ""
    except Exception:
        pass
    path_lower = path.lower()
    # Ensure trailing slash for segment matching
    path_check = path_lower if path_lower.endswith("/") else path_lower + "/"
    if any(seg in path_check for seg in _GROUP_PATH_SEGMENTS):
        return "group_site_subsidiary_page"
    path_parts = [p for p in path.strip("/").split("/") if p]
    if len(path_parts) >= 2:
        combined = (title + " " + snippet).lower()
        if any(sig in combined for sig in ["companies", "aziende", "business-sector", "gruppo", "group"]):
            return "group_site_subsidiary_page"
        return "deep_page"
    return "root_domain"


def _brand_is_ambiguous(brand: str) -> bool:
    """
    True if brand is likely too generic or short to confidently identify an Italian company
    without additional location context.

    Heuristics:
    - Single word ≤ 10 chars: ambiguous (SIMONETTI, RAINBOW, MINO, FERRARI)
    - Multi-word where NO token exceeds 7 chars: ambiguous (DELTA MOTORS, MINI BIKE)
      Brands with at least one distinctive long token are not ambiguous (POMPE GARBARINO).
    """
    if not brand:
        return True
    toks = brand.split()
    if len(toks) == 1 and len(brand) <= 10:
        return True
    if len(toks) >= 2 and max(len(t) for t in toks) <= 7:
        return True
    return False


def _build_search_queries(
    name_variants: dict,
    city: str,
    province: str,
    postcode: str,
    max_queries: int = 5,
    country_config: "CountryConfig | None" = None,
) -> list[str]:
    """
    Build up to max_queries Serper search queries.

    For ambiguous/generic brands, location queries are prioritised early to
    reduce false positives from foreign or unrelated companies.
    For dotted acronyms (I.M.E.S.A.), the nodot form (IMESA) is also used.
    country_config drives country-specific terminology and TLD preferences.
    """
    cfg = country_config or IT_CONFIG

    clean_name  = name_variants.get("no_desc") or name_variants.get("no_legal") or name_variants["full"]
    brand       = name_variants.get("brand") or clean_name
    brand_nodot = name_variants.get("brand_nodot") or brand
    is_acronym  = name_variants.get("is_acronym", False)
    ambiguous   = _brand_is_ambiguous(brand)

    use_brand_queries = (brand.lower() != clean_name.lower() and len(brand) >= 3)
    use_nodot_queries = is_acronym and brand_nodot.lower() != brand.lower()

    loc      = city or province
    country  = cfg.country_search_name
    tld      = cfg.preferred_tlds[0] if cfg.preferred_tlds else ".com"
    off_term = cfg.official_search_terms[0] if cfg.official_search_terms else "official website"

    queries: list[str] = []

    if cfg.country_code == "DE":
        # ── German query set ──────────────────────────────────────────────────
        if ambiguous and loc:
            queries.append(f'"{clean_name}" {loc} offizielle Website')
            queries.append(f'"{clean_name}" Unternehmen Deutschland')
            queries.append(f'"{clean_name}" Impressum')
            if city and province:
                queries.append(f'"{clean_name}" "{city}" "{province}" Deutschland')
            elif city:
                queries.append(f'"{clean_name}" "{city}" Deutschland')
            elif province:
                queries.append(f'"{clean_name}" "{province}" Deutschland')
            queries.append(f'site:.de "{clean_name}"')
        else:
            queries.append(f'"{clean_name}" offizielle Website')
            queries.append(f'"{clean_name}" Impressum')
            queries.append(f'"{clean_name}" Kontakt')
            if city:
                queries.append(f'"{clean_name}" "{city}" Deutschland')
            elif province:
                queries.append(f'"{clean_name}" "{province}" Deutschland')
            queries.append(f'"{clean_name}" Unternehmen Deutschland')
            queries.append(f'site:.de "{clean_name}"')
            if use_brand_queries:
                queries.append(f'"{brand}" Deutschland offizielle Website')
                queries.append(f'site:.de "{brand}"')
        if use_nodot_queries:
            queries.append(f'"{brand_nodot}" offizielle Website Deutschland')
            queries.append(f'site:.de "{brand_nodot}"')

    else:
        # ── Italian query set (original behavior) ─────────────────────────────
        if ambiguous and loc:
            queries.append(f'"{clean_name}" {loc} Italy sito ufficiale')
            queries.append(f'"{clean_name}" Italy official website')
            queries.append(f'"{clean_name}" sito ufficiale')
            if city and province:
                queries.append(f'"{clean_name}" "{city}" "{province}" Italy')
            elif city:
                queries.append(f'"{clean_name}" "{city}" Italy')
            elif province:
                queries.append(f'"{clean_name}" "{province}" Italy')
            queries.append(f'site:.it "{clean_name}"')
        else:
            queries.append(f'"{clean_name}" official website')
            queries.append(f'"{clean_name}" sito ufficiale')
            if city:
                queries.append(f'"{clean_name}" "{city}" Italy')
            elif province:
                queries.append(f'"{clean_name}" "{province}" Italy')
            queries.append(f'"{clean_name}" Italy')
            if province and city:
                queries.append(f'"{clean_name}" "{province}" Italy')
            queries.append(f'site:.it "{clean_name}"')
            if use_brand_queries:
                queries.append(f'"{brand}" Italy official website')
                queries.append(f'site:.it "{brand}"')
        if use_nodot_queries:
            queries.append(f'"{brand_nodot}" Italy sito ufficiale')
            queries.append(f'site:.it "{brand_nodot}"')

        # ── SIGLABILE brand queries (Italy-only, high priority) ───────────────
        siglabile_brand   = name_variants.get("siglabile_brand", "")
        siglabile_compact = name_variants.get("siglabile_compact", "")
        if siglabile_brand and siglabile_brand.lower() != clean_name.lower():
            queries.insert(0, f'"{siglabile_brand}" sito ufficiale')
            queries.insert(1, f'"{siglabile_brand}" official website')
            if siglabile_compact and siglabile_compact != siglabile_brand.lower():
                queries.append(f'"{siglabile_compact}"')

        # ── First-token brand queries (Italy-only) ────────────────────────────
        first_token_brand = name_variants.get("first_token_brand", "")
        if first_token_brand and first_token_brand.lower() != brand.lower():
            queries.insert(0, f'"{first_token_brand}" "{clean_name}" sito ufficiale')
            if loc:
                queries.append(f'"{first_token_brand}" "{loc}" Italy official website')
            queries.append(f'site:.it "{first_token_brand}"')

        # ── Descriptor variant queries (Italy-only) ───────────────────────────
        for _dv_label, _dv_val in name_variants.get("descriptor_variants", []):
            if _dv_val and _dv_val.lower() != clean_name.lower():
                queries.append(f'"{_dv_val}" official website')
                queries.append(f'"{_dv_val}" sito ufficiale')
                if loc:
                    queries.append(f'"{_dv_val}" {loc}')

    # Deduplicate while preserving order
    seen: set = set()
    unique: list[str] = []
    for q in queries:
        if q not in seen:
            seen.add(q)
            unique.append(q)

    return unique[:max_queries]


def _score_candidate(
    domain: str,
    rank: int,
    title: str,
    snippet: str,
    name_variants: dict,
    email_domain: str,
    city: str,
    province: str,
    country_config: "CountryConfig | None" = None,
) -> float:
    """
    Score a candidate domain on a 0–3+ scale.
    Higher is better.
    country_config controls TLD bonus and country-specific penalties.
    """
    cfg = country_config or IT_CONFIG
    score = 0.0

    # 1. Position weight (rank 0 = 1.0, rank 4 = 0.2)
    position_w = 1.0 / (rank + 1)

    # 2. Name overlap signals (use best across variants, incl. acronym nodot form)
    full_overlap  = token_overlap(name_variants["full"], domain)
    desc_overlap  = token_overlap(name_variants.get("no_desc", ""), domain)
    brand_ov      = brand_overlap_variants(name_variants, domain)

    best_name_overlap = max(full_overlap, desc_overlap, brand_ov)
    score += position_w * (0.5 + best_name_overlap * 1.5)

    # 3. Brand name directly in domain (strong signal)
    if brand_ov >= 0.8:
        score += 0.4

    if best_name_overlap < _MIN_BRAND_SIM_TO_SCORE:
        score *= 0.10
    elif best_name_overlap < _WEAK_BRAND_SIM_THRESHOLD:
        score *= _WEAK_BRAND_SIM_MULTIPLIER

    # 4. Title / snippet contains official-page keywords
    combined_text = (title + " " + snippet).lower()
    if has_official_signal(combined_text):
        score += 0.25

    # 5. Title or snippet contains company / brand name (any variant)
    brand_lower = (name_variants.get("brand") or "").lower()
    if brand_lower and brand_lower in combined_text:
        score += 0.2

    # 6. Location signal
    if location_in_text(combined_text, city, province):
        score += 0.3

    # 7. Email domain match (strong confirmation)
    if email_domain and domain == email_domain:
        score += 0.5

    # 8. TLD bonus — config-driven (primary TLD = full bonus; secondary = partial)
    if cfg.preferred_tlds:
        primary_tld = cfg.preferred_tlds[0]
        if domain.endswith(primary_tld):
            score += 0.15
        elif len(cfg.preferred_tlds) > 1:
            # Secondary TLDs get a smaller bonus (Germany: .com/.eu acceptable)
            for _tld in cfg.preferred_tlds[1:]:
                if domain.endswith(_tld):
                    score += 0.07
                    break

    # 9. Italy-specific hard penalties
    if cfg.country_code == "IT":
        if _EDU_IT_RE.search(domain):
            score -= 0.6
        elif re.search(r"\b(forum|foro|archive|archivio)\b", domain, re.I):
            score -= 0.3
        elif re.search(r"\b(associazione|fondazione|onlus|odv|aps)\b", domain, re.I):
            score -= 0.3

    # 9b. Hosted-platform penalty (any country) — free hosting is never an official site
    if is_hosted_platform(domain):
        score -= 2.0

    # 10. Germany-specific penalties (whole-token, word-boundary safe)
    if cfg.country_code == "DE":
        _neg, _neg_reason = _is_de_negative_domain(domain)
        if _neg:
            score -= 0.5
        elif re.search(r"\b(forum|archive)\b", domain, re.I):
            score -= 0.3
        # Germany directory blacklist penalty
        _domain_base = domain.split(".")[0] if domain else ""
        if _domain_base in cfg.directory_blacklist_bases:
            score -= 1.0

    if score < 0:
        score = 0.0

    return round(score, 4)


# ---------------------------------------------------------------------------
# Top Serper Override
#
# When the #1 Google/Serper result passes all official-site guardrails and
# shows a strong name match, we accept it directly — before the multi-
# candidate scoring chooses a winner.  This prevents the normal scoring from
# accidentally preferring a lower-ranked result when the answer is obvious.
#
# The override is deliberately conservative:
#   • All existing blacklists are applied first (generic, hosted, blocked).
#   • Deep profile/directory pages are rejected by URL-path check.
#   • For ambiguous short brands, location or official signal is required.
#   • If ANY guardrail fails, override is skipped and normal scoring runs.
# ---------------------------------------------------------------------------

# Minimum brand overlap to consider a top-hit at all for override.
_TOPSER_MIN_BRAND_OVERLAP = 0.35

# Path patterns that identify a deep profile / directory / employer page —
# these disqualify a URL from being accepted as the official company site.
_TOPSER_DEEP_PATH_RE = re.compile(
    r"/in/|/company/|/companies/|/aziende?/|/scheda[_-]|/profile[_-]|"
    r"/profil[eo]/|/employer/|/org/|/bilanci[_-]|/dati[_-]|/register[_-]|"
    r"/imprese?/|/jobs?/|/lavoro/|/offerte[_-]lavoro|/search\?|/results\?",
    re.IGNORECASE,
)


def _top_serper_override_candidate(
    first_result: dict,
    query: str,
    name_variants: dict,
    email_domain: str,
    city: str,
    province: str,
    country_config: "CountryConfig | None" = None,
) -> tuple[dict | None, str]:
    """
    Inspect the first organic Serper result and decide whether it is safe to
    accept directly as the official company domain, bypassing multi-query scoring.

    Returns:
        (override_evidence_dict, skip_reason)

        override_evidence_dict — filled when override is approved; None when rejected.
        skip_reason            — human-readable explanation (empty string on approval).

    Guardrail checklist (all must pass):
      G1  Domain extractable.
      G2  Domain not generic / blacklisted.
      G3  Domain not a free hosted platform (altervista, wixsite, …).
      G4  Domain not discovery-blocked (media, social, streaming, …).
      G5  Domain not classified as directory/government/religious/academic.
      G6  URL is not a deep profile/directory page.
      G7  Brand overlap ≥ _TOPSER_MIN_BRAND_OVERLAP  OR  email domain matches.
      G8  For ambiguous brands: must also have location match OR official signal
          OR preferred TLD (e.g. .it / .de)  OR email domain match.
    """
    cfg         = country_config or IT_CONFIG
    url         = first_result.get("link", "")
    title       = first_result.get("title", "")
    snippet     = first_result.get("snippet", "")
    domain      = _extract_domain(url)

    # G1
    if not domain:
        return None, "no_extractable_domain"

    # G2
    if is_generic(domain, country_config=cfg):
        return None, f"generic/blacklisted: {domain}"

    # G3
    if is_hosted_platform(domain):
        return None, f"hosted_platform: {domain}"

    # G4
    if is_discovery_blocked(domain):
        return None, f"discovery_blocked: {domain}"

    # G5
    cat = classify_domain(domain, title, snippet)
    if cat:
        return None, f"category_rejected: {cat}"

    # G6 — deep profile or directory URL
    if _TOPSER_DEEP_PATH_RE.search(url):
        return None, f"deep_profile_url: {url[:120]}"

    # G7 — brand overlap
    b_ov       = brand_overlap_variants(name_variants, domain)
    email_hit  = bool(email_domain and domain == email_domain)
    if b_ov < _TOPSER_MIN_BRAND_OVERLAP and not email_hit:
        return None, (
            f"brand_overlap_too_low ({b_ov:.2f} < {_TOPSER_MIN_BRAND_OVERLAP})"
        )

    # G8 — ambiguous brand needs corroboration
    brand     = name_variants.get("brand", "")
    ambiguous = _brand_is_ambiguous(brand)
    if ambiguous:
        combined  = (title + " " + snippet).lower()
        loc_ok    = location_in_text(combined, city, province)
        offic_ok  = has_official_signal(combined)
        pref_tlds = tuple(cfg.preferred_tlds) if cfg.preferred_tlds else (".it",)
        tld_ok    = any(domain.endswith(t) for t in pref_tlds)
        if not (loc_ok or offic_ok or tld_ok or email_hit):
            return None, (
                f"ambiguous_brand_no_corroboration "
                f"(brand={brand!r}, loc={loc_ok}, official={offic_ok}, "
                f"tld={tld_ok}, email={email_hit})"
            )

    # All guardrails passed — build confidence
    combined = (title + " " + snippet).lower()
    loc_ok   = location_in_text(combined, city, province)
    offic_ok = has_official_signal(combined)

    if b_ov >= 0.70 and (loc_ok or offic_ok or not ambiguous):
        confidence      = 0.82
        confidence_str  = "High"
    elif b_ov >= 0.50 or email_hit:
        confidence      = 0.68
        confidence_str  = "Medium"
    else:
        confidence      = 0.55
        confidence_str  = "Medium"

    ev = {
        "query":                query,
        "title":                title[:120],
        "url":                  url,
        "snippet":              snippet[:200],
        "domain":               domain,
        "score":                round(confidence, 3),
        "brand_overlap":        round(b_ov, 3),
        "full_overlap":         round(token_overlap(name_variants["full"], domain), 3),
        "location_match":       loc_ok,
        "email_match":          email_hit,
        "official_signal":      offic_ok,
        "used":                 True,
        "candidate_url":        url,
        "candidate_type":       _classify_candidate_type(url, domain, title, snippet),
        "candidate_source":     "top_serper_override",
        "evidence_source_url":  url,
        # Carry override metadata so downstream diagnostics can explain the decision
        "_override_confidence_str": confidence_str,
    }
    return ev, ""


def search_official_domain_register(
    company_name: str,
    city: str,
    province: str,
    postcode: str,
    email_domain: str,
    serper_key: str,
    max_queries: int = 5,
    country_config: "CountryConfig | None" = None,
) -> tuple[str, float, str, list, str, str, list, dict]:
    """
    Run up to max_queries Serper queries with multi-variant brand scoring.

    Returns:
      (suggested_domain, confidence, reason, evidence_rows,
       query_used, name_variant_used, top_3_domains, rejection_counts)

    rejection_counts: dict with keys directory/government/religious/academic/low_similarity
    """
    cfg = country_config or IT_CONFIG
    name_variants = extract_name_variants(company_name)
    queries = _build_search_queries(
        name_variants, city, province, postcode, max_queries,
        country_config=cfg,
    )

    candidates: dict[str, float] = {}    # domain → best score seen
    domain_variant: dict[str, str] = {}  # domain → which name variant matched best
    evidence: list[dict] = []
    query_used = queries[0] if queries else ""
    rejection_notes: list[str] = []
    rejection_counts: dict[str, int] = {
        "directory": 0, "government": 0, "religious": 0,
        "academic": 0, "low_similarity": 0,
    }

    for query in queries:
        results, err = _call_serper(
            query, serper_key, gl=cfg.serper_gl, hl=cfg.serper_hl,
        )
        if err:
            rejection_notes.append(f"Serper error: {err}")
            evidence.append({
                "query": query, "title": "", "url": "", "snippet": "",
                "domain": "", "used": False,
                "skip_reason": f"serper_error: {err[:120]}",
                "rejection_category": "serper_error",
                "score": "",
            })
            break
        if not results:
            evidence.append({
                "query": query, "title": "", "url": "", "snippet": "",
                "domain": "", "used": False,
                "skip_reason": "no_serper_results",
                "rejection_category": "serper_no_results",
                "score": "",
            })
            continue

        # ── Top Serper Override — first result of first query only ────────────
        # Before scoring all candidates normally, check whether the #1 Google
        # result already passes all official-site guardrails.  If it does, we
        # accept it directly, record it in evidence, and return early.  This
        # prevents the scoring algorithm from accidentally preferring a lower-
        # ranked hit when the answer is obvious (e.g. the company's own homepage
        # ranks #1 but scores slightly below a directory that happens to contain
        # the brand name).
        # Override only runs once: on the very first Serper query.
        if query == queries[0]:
            _ov_ev, _ov_skip = _top_serper_override_candidate(
                results[0], query, name_variants,
                email_domain, city, province,
                country_config=cfg,
            )
            if _ov_ev is not None:
                # Guardrails passed — use this domain directly.
                _ov_domain     = _ov_ev["domain"]
                _ov_conf_str   = _ov_ev.pop("_override_confidence_str", "High")
                _ov_conf       = _ov_ev["score"]
                _ov_reason     = (
                    f"Top Serper result accepted: passed official-site guardrails "
                    f"[brand_overlap={_ov_ev['brand_overlap']:.2f}, confidence={_ov_conf_str}]"
                )
                evidence.append(_ov_ev)
                # Also record all remaining results from this query as skipped
                # so Raw Search Evidence remains complete.
                for _rank2, _item2 in enumerate(results[1:], start=1):
                    _u2 = _item2.get("link", "")
                    _d2 = _extract_domain(_u2)
                    evidence.append({
                        "query": query, "title": _item2.get("title", "")[:80],
                        "url": _u2, "snippet": _item2.get("snippet", "")[:200],
                        "domain": _d2 or "", "used": False,
                        "skip_reason": "skipped_after_top_serper_override",
                        "score": "",
                    })
                return (
                    _ov_domain, _ov_conf, _ov_reason,
                    evidence, query, name_variants.get("brand", ""),
                    [_ov_domain],
                    rejection_counts,
                )
            else:
                # Guardrails failed — record why and continue with normal scoring.
                _ov_url  = results[0].get("link", "")
                _ov_dom  = _extract_domain(_ov_url) or ""
                evidence.append({
                    "query": query, "title": results[0].get("title", "")[:80],
                    "url": _ov_url, "snippet": results[0].get("snippet", "")[:200],
                    "domain": _ov_dom, "used": False,
                    "skip_reason": f"top_serper_override_rejected: {_ov_skip}",
                    "rejection_category": "top_serper_override_rejected",
                    "score": "",
                })
        # ── End Top Serper Override ───────────────────────────────────────────

        for rank, item in enumerate(results):
            url     = item.get("link", "")
            title   = item.get("title", "")
            snippet = item.get("snippet", "")
            domain  = _extract_domain(url)

            # Separate: URL gave no extractable domain vs domain is generic/blacklisted
            if not domain:
                evidence.append({
                    "query": query, "title": title[:80], "url": url,
                    "snippet": snippet[:200],
                    "domain": "", "used": False,
                    "skip_reason": "no_extractable_domain",
                    "rejection_category": "domain_extraction_failed",
                    "score": "",
                })
                continue

            if is_generic(domain, country_config=cfg):
                evidence.append({
                    "query": query, "title": title[:80], "url": url,
                    "snippet": snippet[:200],
                    "domain": domain, "used": False,
                    "skip_reason": "generic/blacklisted", "score": 0,
                })
                rejection_notes.append(f"{domain}: blacklisted")
                rejection_counts["directory"] += 1
                continue

            # Hard-block URL shorteners — they must never become a company domain.
            if is_url_shortener(domain):
                evidence.append({
                    "query": query, "title": title[:80], "url": url,
                    "snippet": snippet[:200],
                    "domain": domain, "used": False,
                    "skip_reason": "url_shortener_blocked", "score": 0,
                    "rejection_category": "rejected_shortener",
                })
                rejection_notes.append(f"{domain}: url shortener blocked")
                rejection_counts["rejected_shortener"] = rejection_counts.get("rejected_shortener", 0) + 1
                continue

            # Hard-block media, streaming, social, marketplace domains — they can
            # never be an official company domain regardless of snippet content.
            if is_discovery_blocked(domain):
                evidence.append({
                    "query": query, "title": title[:80], "url": url,
                    "snippet": snippet[:200],
                    "domain": domain, "used": False,
                    "skip_reason": "media/streaming/social domain blocked", "score": 0,
                    "rejection_category": "media_blocked",
                })
                rejection_notes.append(f"{domain}: media/streaming blocked")
                rejection_counts["media_blocked"] = rejection_counts.get("media_blocked", 0) + 1
                continue

            # Category check — reject government, religious, directory, academic
            cat = classify_domain(domain, title, snippet)
            if cat:
                evidence.append({
                    "query": query, "title": title[:80], "url": url,
                    "snippet": snippet[:200],
                    "domain": domain, "used": False,
                    "skip_reason": f"category:{cat}", "score": 0,
                    "rejection_category": cat,
                })
                rejection_counts[cat] = rejection_counts.get(cat, 0) + 1
                rejection_notes.append(f"{domain}: rejected ({cat})")
                continue

            score = _score_candidate(
                domain, rank, title, snippet,
                name_variants, email_domain, city, province,
                country_config=cfg,
            )

            # Very low score after brand gate — note it but don't include in candidates
            if score < 0.08:
                evidence.append({
                    "query": query, "title": title[:80], "url": url,
                    "snippet": snippet[:200],
                    "domain": domain, "used": False,
                    "skip_reason": f"low_similarity_score({score:.3f})", "score": score,
                })
                rejection_counts["low_similarity"] += 1
                rejection_notes.append(f"{domain}: low similarity ({score:.3f})")
                continue

            if domain not in candidates or score > candidates[domain]:
                candidates[domain] = score
                bov = brand_overlap_variants(name_variants, domain)
                dov = token_overlap(name_variants.get("no_desc", ""), domain)
                fov = token_overlap(name_variants["full"], domain)
                if bov >= dov and bov >= fov:
                    domain_variant[domain] = f"brand:{name_variants.get('brand','')}"
                elif dov >= fov:
                    domain_variant[domain] = f"no_desc:{name_variants.get('no_desc','')}"
                else:
                    domain_variant[domain] = f"full:{name_variants['full']}"
            else:
                candidates[domain] = max(candidates[domain], score)

            evidence.append({
                "query": query, "title": title[:120], "url": url,
                "snippet": snippet[:200],
                "domain": domain, "score": round(score, 3),
                "brand_overlap": round(brand_overlap_variants(name_variants, domain), 3),
                "full_overlap":  round(token_overlap(name_variants["full"], domain), 3),
                "location_match": location_in_text(title + " " + snippet, city, province),
                "email_match": (domain == email_domain),
                "official_signal": has_official_signal(title + " " + snippet),
                "used": True,
                "candidate_url": url,
                "candidate_type": _classify_candidate_type(url, domain, title, snippet),
                "candidate_source": "serper_result",
                "evidence_source_url": url,
            })

            # Check for URLs embedded in snippet (e.g. LinkedIn showing "Sito Web: http://...")
            _snippet_urls = _extract_urls_from_text(snippet)
            for _su in _snippet_urls:
                _su_domain = _extract_domain(_su)
                if (_su_domain and not is_generic(_su_domain)
                        and not is_url_shortener(_su_domain)
                        and not is_discovery_blocked(_su_domain)
                        and not classify_domain(_su_domain, title, snippet)):
                    _su_score = _score_candidate(
                        _su_domain, rank, title, snippet,
                        name_variants, email_domain, city, province,
                        country_config=cfg,
                    )
                    if _su_score >= 0.08:
                        if _su_domain not in candidates or _su_score > candidates[_su_domain]:
                            candidates[_su_domain] = _su_score
                            domain_variant[_su_domain] = f"snippet_url:{_su_domain}"
                        evidence.append({
                            "query": query, "title": title[:120], "url": _su,
                            "snippet": snippet[:200],
                            "domain": _su_domain, "score": round(_su_score, 3),
                            "brand_overlap": round(brand_overlap_variants(name_variants, _su_domain), 3),
                            "full_overlap": round(token_overlap(name_variants["full"], _su_domain), 3),
                            "location_match": location_in_text(title + " " + snippet, city, province),
                            "email_match": (_su_domain == email_domain),
                            "official_signal": has_official_signal(title + " " + snippet),
                            "used": True,
                            "candidate_source": "snippet_extracted_url",
                            "candidate_url": _su,
                            "candidate_type": "snippet_url",
                            "evidence_source_url": url,
                        })

        time.sleep(0.25)

    if not candidates:
        return (
            "", 0.0,
            "No candidate domain found. " + "; ".join(rejection_notes[:4]),
            evidence, query_used, "", [],
            rejection_counts,
        )

    # ── Generic DE brand guard ────────────────────────────────────────────────
    # If the brand is a single generic token (e.g. "Global"), require the full
    # legal company name or no_legal variant to appear in title+snippet evidence
    # before accepting any candidate domain.
    if name_variants.get("brand_is_de_generic"):
        _full_lower   = name_variants.get("full", "").lower()
        _nolegal_lower = name_variants.get("no_legal", "").lower()
        _brand_name   = name_variants.get("brand", "")
        _filtered_cands: dict[str, float] = {}
        for _dom, _sc in candidates.items():
            _dom_evs = [e for e in evidence if e.get("domain") == _dom and e.get("used")]
            _name_found = False
            for _ev in _dom_evs:
                _ct = (_ev.get("title", "") + " " + _ev.get("snippet", "")).lower()
                if ((_full_lower and _full_lower in _ct)
                        or (_nolegal_lower and len(_nolegal_lower) >= 5 and _nolegal_lower in _ct)):
                    _name_found = True
                    break
            if _name_found:
                _filtered_cands[_dom] = _sc
            else:
                # Demote: mark evidence rows as rejected
                for _ev in evidence:
                    if _ev.get("domain") == _dom and _ev.get("used"):
                        _ev["used"] = False
                        _ev["skip_reason"] = (
                            f"generic_de_brand_no_exact_name_match "
                            f"(brand='{_brand_name}' is generic; need full company name in evidence)"
                        )
                        _ev["rejection_category"] = "generic_brand_rejected"
                rejection_notes.append(
                    f"{_dom}: rejected (generic DE brand '{_brand_name}', "
                    f"full company name not found in evidence)"
                )
                rejection_counts["generic_brand_rejected"] = (
                    rejection_counts.get("generic_brand_rejected", 0) + 1
                )
        candidates = _filtered_cands
        if not candidates:
            return (
                "", 0.0,
                f"Generic DE brand '{_brand_name}': no candidate with exact company name in evidence. "
                + "; ".join(rejection_notes[:4]),
                evidence, query_used, "", [],
                rejection_counts,
            )

    # Sort by score — compare all candidates, pick best
    sorted_cands = sorted(candidates.items(), key=lambda x: x[1], reverse=True)
    best, best_score = sorted_cands[0]
    top3 = [d for d, _ in sorted_cands[:3]]

    b_ov         = brand_overlap_variants(name_variants, best)
    f_ov         = token_overlap(name_variants["full"], best)
    brand_lower  = (name_variants.get("brand") or "").lower()
    brand_nodot_lower = (name_variants.get("brand_nodot") or brand_lower).lower()
    best_variant = domain_variant.get(best, "full")

    # Top evidence entry for supplementary signals
    top_ev = next(
        (e for e in evidence if e.get("domain") == best and e.get("used")), {}
    )
    loc_match = top_ev.get("location_match", False)
    official  = top_ev.get("official_signal", False)
    combined_title_snip = (top_ev.get("title", "") + " " + top_ev.get("snippet", "")).lower()
    brand_in_title = bool(
        (brand_lower and brand_lower in combined_title_snip)
        or (brand_nodot_lower and brand_nodot_lower in combined_title_snip)
    )

    # ── High-confidence rules ────────────────────────────────────────────────
    # A: Brand clearly in domain (exact/near-exact match or strong substring)
    #    A_strong: brand constitutes majority of domain base (e.g. pompegarbarino.com)
    #    A_weak:   brand is substring of longer domain (e.g. goblinsimonetti.com, deltamotorsofconcord.com)
    #    A_weak requires corroborating location or .it TLD to qualify for High.
    domain_base_clean = re.sub(r"[^\w]", "", re.sub(r"\.[a-z]{2,}$", "", best.lower()))
    brand_clean_str   = re.sub(r"[^\w]", "", brand_lower)
    brand_fills_domain = (
        len(brand_clean_str) >= len(domain_base_clean) * 0.7
        if domain_base_clean else False
    )
    rule_A_strong = b_ov >= _HIGH_CONF_BRAND_THRESHOLD and brand_fills_domain
    rule_A_weak   = b_ov >= _HIGH_CONF_BRAND_THRESHOLD and not brand_fills_domain
    # A_weak only counts for High when location or .it is present
    rule_A = rule_A_strong or (rule_A_weak and (loc_match or best.endswith(".it")))

    # B: Domain matches email domain (external corroboration)
    rule_B = bool(email_domain and best == email_domain)

    # C: Brand in title alone — no longer sufficient for High by itself
    #    (title match can be fabricated in directory pages)
    rule_C = brand_in_title  # kept for rule_D counting and extras display

    # D: Multiple independent signals agree (requires score ≥ 1.0)
    rule_D = (
        best_score >= 1.0
        and sum([loc_match, official, rule_B, rule_A_strong or rule_A_weak, rule_C]) >= 2
    )

    is_high = rule_A or rule_B or rule_D

    # Score guardrail: score < 1.0 → High only if email match or exact brand-in-domain
    if best_score < 1.0 and not rule_B:
        is_high = is_high and rule_A_strong

    # Low-score guardrail: if score < 0.8 and no email match, cap at Low/Medium
    below_threshold = best_score < 0.8 and not rule_B

    # Assign confidence
    if rule_B and best_score >= 0.6:
        conf   = 0.88
        reason = f"Email domain '{best}' confirmed by Serper."
    elif is_high and best_score >= 1.2:
        conf   = 0.85
        reason = "Strong brand match in domain + search position."
    elif is_high and best_score >= 0.7:
        conf   = 0.78
        reason = "Brand confirmed in domain + reasonable search position."
    elif is_high:
        conf   = 0.72
        reason = "Brand confirmed in domain or email match."
    elif rule_C and best_score >= 0.8 and not below_threshold:
        conf   = 0.62
        reason = "Brand in search result title, score acceptable."
    elif best_score >= 0.60 and not below_threshold:
        conf   = 0.52
        reason = "Reasonable position + partial name match."
    elif best_score >= 0.35:
        conf   = 0.38
        reason = "Weak brand-domain relationship. Likely needs manual review."
    else:
        conf   = 0.20
        reason = "Very weak match — high false-positive risk."

    # If below_threshold, cap at Medium regardless
    if below_threshold and conf > 0.65:
        conf   = 0.55
        reason += " [capped: score < 0.8]"

    extras = []
    if rule_A_strong:
        extras.append(f"brand '{name_variants.get('brand','')}' fills domain")
    elif rule_A_weak:
        extras.append(f"brand '{name_variants.get('brand','')}' in domain (partial)")
    if rule_B:
        extras.append(f"matches email domain ({best})")
    if rule_C:
        extras.append("brand in search result title/snippet")
    if loc_match:
        extras.append("city/province in result")
    if official:
        extras.append("official-page keyword in title/snippet")
    if best.endswith(".it"):
        extras.append(".it domain")
    if extras:
        reason += " — " + "; ".join(extras) + "."

    return best, conf, reason, evidence, query_used, best_variant, top3, rejection_counts


# =============================================================================
# CORE VALIDATION (register-aware)
# =============================================================================


def validate_register_row(
    company_name: str,
    raw_website: str,
    raw_email: str,
    city: str,
    province: str,
    postcode: str,
    serper_key: str | None,
    max_queries: int = 5,
    country_config: "CountryConfig | None" = None,
) -> tuple[dict, list]:
    """
    Validate one register row. Returns result fields dict.

    Decision flow:
      1. Parse and clean website → normalized_input_website
      2. If website valid and non-generic → OK / LIKELY_OK
      3. If website missing/invalid → try email domain (aggressively)
      4. If email domain plausible → EMAIL_DERIVED, then try Serper to confirm
      5. If still missing → Serper search with multi-variant queries
      6. If Serper finds confident result → MISSING_DOMAIN_FIXED / SUGGEST_REPLACE
      7. Otherwise → MISSING_DOMAIN / NO_CONFIDENT_MATCH

    New diagnostic columns added v2:
      name_variant_used, candidate_domains_considered,
      best_candidate_score, top_3_candidate_domains,
      rejection_reason_if_missing, website_discovery_method
    """
    name     = str(company_name or "").strip()
    email    = str(raw_email or "").strip()
    city     = str(city or "").strip()
    province = str(province or "").strip()
    postcode = str(postcode or "").strip()

    _all_raw_evidence: list[dict] = []

    email_domain = extract_email_domain(email)
    email_is_pec = is_pec_or_personal_email(email_domain)

    norm_website = best_website_domain(raw_website)
    # Hard guard: URL shorteners must never become validated_domain
    if is_url_shortener(norm_website):
        norm_website = ""
    name_variants = extract_name_variants(name)

    result = {
        # Core output
        "cleaned_company_name":       name,
        "normalized_input_website":   norm_website,
        "email_domain":               email_domain,
        "validated_domain":           norm_website,
        "recommended_domain":         "",
        "domain_source":              SRC_ORIGINAL if norm_website else SRC_NONE,
        "domain_action":              "",
        "domain_confidence":          "",
        "domain_reason":              "",
        "manual_review_needed":       False,
        "search_query_used":          "",
        "serper_top_result_title":    "",
        "serper_top_result_url":      "",
        "serper_top_result_domain":   "",
        # v2 diagnostic columns
        "name_variant_used":            "",
        "candidate_domains_considered": "",
        "best_candidate_score":         "",
        "top_3_candidate_domains":      "",
        "rejection_reason_if_missing":  "",
        "website_discovery_method":     "",
        # v3 rejection counts (accumulated across Serper calls for this row)
        "rejected_directory":           0,
        "rejected_government":          0,
        "rejected_religious":           0,
        "rejected_academic":            0,
        "rejected_low_similarity":      0,
    }

    if not name:
        result.update(
            domain_action="NO_CONFIDENT_MATCH",
            domain_confidence="None",
            domain_reason="Company name is blank.",
            manual_review_needed=True,
            rejection_reason_if_missing="Company name is blank.",
            website_discovery_method="none",
        )
        return result, _all_raw_evidence

    # Helper: run Serper and fill result fields
    def _run_serper(existing_email_domain=""):
        sug, conf, reason, ev, query, variant, top3, rej = search_official_domain_register(
            name, city, province, postcode,
            existing_email_domain or email_domain,
            serper_key, max_queries,
            country_config=country_config,
        )
        _fill_serper_top(result, ev, query)
        result["name_variant_used"] = variant
        all_doms = [e.get("domain", "") for e in ev if e.get("domain")]
        result["candidate_domains_considered"] = ", ".join(dict.fromkeys(filter(None, all_doms)))
        result["top_3_candidate_domains"] = ", ".join(top3)
        if sug:
            result["best_candidate_score"] = str(round(conf, 3))
        # Accumulate rejection counts across multiple Serper calls for this row
        for cat, cnt in rej.items():
            key = f"rejected_{cat}"
            result[key] = result.get(key, 0) + cnt
        _all_raw_evidence.extend(ev)
        return sug, conf, reason, ev

    # ── Case 1: website present, non-generic ─────────────────────────────────
    if norm_website and not is_generic(norm_website):
        overlap = token_overlap(name, norm_website)
        b_ov    = brand_overlap(name_variants.get("brand", ""), norm_website)
        best_ov = max(overlap, b_ov)

        if best_ov >= 0.45:
            result.update(
                domain_action="OK",
                domain_confidence="High",
                domain_reason="Website domain matches company name / brand tokens closely.",
                manual_review_needed=False,
                website_discovery_method="original_website_accepted",
            )
            return result, _all_raw_evidence

        if best_ov >= 0.15:
            result.update(
                domain_action="LIKELY_OK",
                domain_confidence="Medium",
                domain_reason="Website present; partial name-domain overlap (group/abbreviation likely).",
                manual_review_needed=False,
                website_discovery_method="original_website_partial_match",
            )
            return result, _all_raw_evidence

        # Low overlap — search to confirm or find a better domain
        if serper_key:
            suggested, conf, reason, ev = _run_serper()
            if suggested and conf >= 0.40 and suggested != norm_website:
                result.update(
                    validated_domain=suggested,
                    recommended_domain=suggested,
                    domain_source=SRC_SERPER,
                    domain_action="SUGGEST_REPLACE",
                    domain_confidence=_conf_label(conf),
                    domain_reason=f"Low name-website overlap ({best_ov:.2f}). {reason}",
                    manual_review_needed=(conf < 0.70),
                    website_discovery_method="serper_replaced_low_overlap_website",
                )
                return result, _all_raw_evidence
            if suggested and suggested == norm_website:
                result.update(
                    domain_action="LIKELY_OK",
                    domain_confidence="Medium",
                    domain_reason=f"Search confirms website despite low token overlap ({best_ov:.2f}).",
                    manual_review_needed=False,
                    website_discovery_method="serper_confirmed_original_website",
                )
                return result, _all_raw_evidence

        result.update(
            domain_action="REVIEW",
            domain_confidence="Low",
            domain_reason=f"Website present but low name-domain overlap ({best_ov:.2f}). Manual check recommended.",
            manual_review_needed=True,
            website_discovery_method="original_website_low_confidence",
        )
        return result, _all_raw_evidence

    # ── Case 2: website is a generic/directory site ──────────────────────────
    if norm_website and is_generic(norm_website):
        result["rejection_reason_if_missing"] = f"Input website '{norm_website}' is a directory/blacklisted domain."
        if serper_key:
            suggested, conf, reason, ev = _run_serper()
            if suggested and conf >= 0.40:
                result.update(
                    validated_domain=suggested,
                    recommended_domain=suggested,
                    domain_source=SRC_SERPER,
                    domain_action="SUGGEST_REPLACE",
                    domain_confidence=_conf_label(conf),
                    domain_reason=f"Register website ({norm_website}) is a directory. {reason}",
                    manual_review_needed=(conf < 0.70),
                    website_discovery_method=(
                        "top_serper_override"
                        if "Top Serper result accepted" in reason
                        else "serper_found_after_blacklisted_website"
                    ),
                )
                return result, _all_raw_evidence
        result.update(
            validated_domain="",
            domain_action="REVIEW",
            domain_confidence="Low",
            domain_reason=f"Register website ({norm_website}) is a generic directory site.",
            manual_review_needed=True,
            website_discovery_method="none_website_blacklisted",
        )
        return result, _all_raw_evidence

    # ── Case 3: website missing — try email domain aggressively ──────────────
    # v2: Use email domain with much lower bar; Serper will confirm if needed.
    if email_domain and not email_is_pec and not is_generic(email_domain):
        email_name_overlap = token_overlap(name, email_domain)
        email_brand_overlap = brand_overlap(name_variants.get("brand", ""), email_domain)
        email_best_overlap = max(email_name_overlap, email_brand_overlap)

        # Any non-PEC, non-generic, non-personal email domain is a candidate
        # (v2: we no longer require 0.3 overlap — Serper will validate)
        email_plausible = (email_best_overlap >= 0.15) or (email_best_overlap >= 0.0 and serper_key)

        if email_plausible:
            # Try Serper to confirm or find better
            if serper_key:
                suggested, conf, reason, ev = _run_serper(email_domain)
                if suggested == email_domain:
                    result.update(
                        validated_domain=email_domain,
                        recommended_domain=email_domain,
                        domain_source=SRC_SERPER_EMAIL,
                        domain_action="MISSING_DOMAIN_FIXED",
                        domain_confidence=_conf_label(max(conf, 0.70)),
                        domain_reason=f"Website missing. Serper confirms email domain '{email_domain}': {reason}",
                        manual_review_needed=(conf < 0.70),
                        website_discovery_method="serper_confirmed_email_domain",
                    )
                    return result, _all_raw_evidence
                if suggested and conf >= 0.50:
                    # Serper found something better than the email domain
                    result.update(
                        validated_domain=suggested,
                        recommended_domain=suggested,
                        domain_source=SRC_SERPER,
                        domain_action="MISSING_DOMAIN_FIXED",
                        domain_confidence=_conf_label(conf),
                        domain_reason=f"Website missing. Email domain was proxy; Serper found better: {reason}",
                        manual_review_needed=(conf < 0.55),
                        website_discovery_method="serper_found_overrides_email_domain",
                    )
                    return result, _all_raw_evidence
                if suggested and conf >= 0.30:
                    # Weak Serper hit — fall back to email domain with Medium confidence
                    result.update(
                        validated_domain=email_domain,
                        recommended_domain=email_domain,
                        domain_source=SRC_EMAIL,
                        domain_action="EMAIL_DERIVED",
                        domain_confidence="Medium",
                        domain_reason=(
                            f"Website missing. Email domain '{email_domain}' used "
                            f"(overlap {email_best_overlap:.2f}); Serper inconclusive."
                        ),
                        manual_review_needed=True,
                        website_discovery_method="email_domain_serper_inconclusive",
                    )
                    return result, _all_raw_evidence

            # No Serper or Serper found nothing — use email domain if overlap reasonable
            if email_best_overlap >= 0.15:
                result.update(
                    validated_domain=email_domain,
                    recommended_domain=email_domain,
                    domain_source=SRC_EMAIL,
                    domain_action="EMAIL_DERIVED",
                    domain_confidence="Medium" if email_best_overlap >= 0.30 else "Low",
                    domain_reason=(
                        f"Website missing. Email domain '{email_domain}' used as proxy "
                        f"(overlap {email_best_overlap:.2f}). Verify manually."
                    ),
                    manual_review_needed=True,
                    website_discovery_method="email_domain_proxy",
                )
                return result, _all_raw_evidence

    # ── Case 4: website missing — Serper search (no email signal) ────────────
    if serper_key:
        suggested, conf, reason, ev = _run_serper()
        if suggested and conf >= 0.40:
            result.update(
                validated_domain=suggested,
                recommended_domain=suggested,
                domain_source=SRC_SERPER,
                domain_action="MISSING_DOMAIN_FIXED",
                domain_confidence=_conf_label(conf),
                domain_reason=f"Website missing in register. {reason}",
                manual_review_needed=(conf < 0.70),
                website_discovery_method=(
                    "top_serper_override"
                    if "Top Serper result accepted" in reason
                    else "serper_search"
                ),
            )
        elif suggested and conf >= 0.15:
            # Plausible candidate — output with manual review flag instead of blank
            result.update(
                validated_domain=suggested,
                recommended_domain=suggested,
                domain_source=SRC_SERPER,
                domain_action="PLAUSIBLE_NEEDS_REVIEW",
                domain_confidence="Low",
                domain_reason=f"Website missing. Plausible candidate found but confidence is low ({conf:.2f}). Manual review needed. {reason}",
                manual_review_needed=True,
                website_discovery_method="serper_search_plausible",
            )
        else:
            result.update(
                validated_domain="",
                domain_source=SRC_NONE,
                domain_action="MISSING_DOMAIN",
                domain_confidence="None",
                domain_reason="Website missing and no confident result found in search.",
                manual_review_needed=True,
                rejection_reason_if_missing="No sufficiently confident candidate in Serper results.",
                website_discovery_method="none_serper_failed",
            )
    else:
        # No Serper — email fallback with very low bar as last resort
        if email_domain and not email_is_pec and not is_generic(email_domain):
            result.update(
                validated_domain=email_domain,
                recommended_domain=email_domain,
                domain_source=SRC_EMAIL,
                domain_action="EMAIL_DERIVED",
                domain_confidence="Low",
                domain_reason=(
                    f"Website missing. No Serper key. Email domain '{email_domain}' "
                    "used as best guess — verify manually."
                ),
                manual_review_needed=True,
                website_discovery_method="email_domain_no_serper",
            )
        else:
            result.update(
                validated_domain="",
                domain_source=SRC_NONE,
                domain_action="MISSING_DOMAIN",
                domain_confidence="None",
                domain_reason="Website missing. No Serper key. No usable email domain.",
                manual_review_needed=True,
                rejection_reason_if_missing="No website, no Serper, no usable email domain.",
                website_discovery_method="none",
            )

    return result, _all_raw_evidence


def _fill_serper_top(result: dict, evidence: list, query: str) -> None:
    result["search_query_used"] = query
    used = [e for e in evidence if e.get("used")]
    top  = used or [e for e in evidence if e.get("domain")]
    if top:
        result["serper_top_result_title"]  = str(top[0].get("title", ""))[:120]
        result["serper_top_result_url"]    = top[0].get("url", "")
        result["serper_top_result_domain"] = top[0].get("domain", "")


# =============================================================================
# CLAUDE HAIKU REVIEW LAYER
# =============================================================================


def _build_haiku_results_block(raw_evidence: list[dict]) -> str:
    """
    Format ALL Serper evidence for the Haiku prompt, grouped by query.
    Shows up to 10 results per query: used results first, then filtered/rejected ones
    with their rejection reason so Haiku can override if appropriate.
    """
    if not raw_evidence:
        return "(no search results available)"

    # Group by query, preserving insertion order
    from collections import OrderedDict
    by_query: OrderedDict[str, list[dict]] = OrderedDict()
    for e in raw_evidence:
        q = e.get("query", "(unknown query)")
        by_query.setdefault(q, []).append(e)

    sections: list[str] = []
    for query, items in by_query.items():
        lines: list[str] = [f'Query: "{query}"']
        seen_domains: set[str] = set()
        count = 0
        for e in items:
            if count >= 10:
                break
            domain  = e.get("domain", "") or "(no domain)"
            title   = (e.get("title", "") or "")[:80]
            snippet = (e.get("snippet", "") or "")[:120]
            url     = e.get("url", "")
            used    = e.get("used", False)

            if domain in seen_domains:
                continue
            seen_domains.add(domain)
            count += 1

            cand_url  = e.get("candidate_url", "") or url
            cand_type = e.get("candidate_type", "") or _classify_candidate_type(url, domain, title, snippet)
            cand_src  = e.get("candidate_source", "serper_result")
            if used:
                score   = e.get("score", "?")
                b_ov    = e.get("brand_overlap", "?")
                f_ov    = e.get("full_overlap", "?")
                em      = "Yes" if e.get("email_match") else "No"
                loc     = "Yes" if e.get("location_match") else "No"
                off     = "Yes" if e.get("official_signal") else "No"
                lines.append(
                    f"  {count}. [SCORED] {domain}\n"
                    f"     Title:       {title}\n"
                    f"     Snippet:     {snippet}\n"
                    f"     URL:         {cand_url}\n"
                    f"     Type:        {cand_type} | Source: {cand_src}\n"
                    f"     Score: {score} | Brand overlap: {b_ov} | Full overlap: {f_ov} | "
                    f"Email match: {em} | Location: {loc} | Official signal: {off}"
                )
            else:
                reason = e.get("skip_reason", "filtered")
                score  = e.get("score", "")
                score_str = f" | Score: {score}" if score else ""
                lines.append(
                    f"  {count}. [FILTERED: {reason}] {domain}\n"
                    f"     Title:   {title}\n"
                    f"     Snippet: {snippet}\n"
                    f"     URL:     {cand_url}\n"
                    f"     Type:    {cand_type}{score_str}"
                )
        sections.append("\n".join(lines))

    return "\n\n".join(sections)


def _haiku_review_domain(
    company_name: str,
    city: str,
    province: str,
    email_domain: str,
    original_website: str,
    python_result: dict,
    raw_evidence: list[dict],
    api_key: str,
    model: str = _DEFAULT_HAIKU_MODEL,
    country_config: "CountryConfig | None" = None,
) -> dict:
    """
    Call Claude Haiku to validate the Python-suggested domain.
    Returns a dict with haiku_* fields.
    """
    out = {
        "haiku_used":               True,
        "haiku_decision":           "",
        "haiku_domain":             "",
        "haiku_confidence":         "",
        "haiku_reason":             "",
        "haiku_risk_flags":         "",
        "haiku_error":              "",
        "haiku_candidate_url":      "",
        "haiku_recommended_action": "",
    }

    # Do not call Haiku if there is no Serper evidence to reason about
    if not raw_evidence:
        out["haiku_used"]               = False
        out["haiku_decision"]           = "skipped_no_serper_evidence"
        out["haiku_error"]              = "no Serper results available for this row"
        out["haiku_candidate_url"]      = ""
        out["haiku_recommended_action"] = ""
        return out

    if not _ANTHROPIC_AVAILABLE or not api_key:
        out["haiku_used"]  = False
        out["haiku_error"] = "anthropic SDK not installed or API key missing"
        return out

    cfg = country_config or IT_CONFIG
    python_domain     = str(python_result.get("validated_domain", "") or "")
    python_confidence = str(python_result.get("domain_confidence", "") or "")
    results_block     = _build_haiku_results_block(raw_evidence)

    # Build country-aware location line and system/user prompt
    _country_name = cfg.country_name
    _state_label  = "federal state" if cfg.country_code == "DE" else "province"
    _loc_line = f"{city}, {province} ({_country_name})"
    _system_prompt = _HAIKU_SYSTEM_PROMPT.replace(
        "for Italian companies", f"for {_country_name} companies"
    )
    _user_template_country = (
        f"Company: {{company_name}}\n"
        f"Location: {{city}}, {{province}} ({_country_name})\n"
        f"Email domain: {{email_domain}}\n"
        f"Original website in register: {{original_website}}\n"
        f"Python-suggested domain: {{python_domain}} (confidence: {{python_confidence}})\n\n"
        f"Search results (pay attention to full URLs and path segments):\n{{results_block}}\n\n"
        f"Decide whether the Python-suggested domain is the correct official website for this "
        f"{_country_name} company. If a group/subsidiary page URL exists, evaluate it carefully. "
        f"Reply with JSON only."
    )

    user_msg = _user_template_country.format(
        company_name=company_name,
        city=city,
        province=province,
        email_domain=email_domain or "(none)",
        original_website=original_website or "(none)",
        python_domain=python_domain or "(none)",
        python_confidence=python_confidence or "(none)",
        results_block=results_block,
    )

    try:
        client = _anthropic_sdk.Anthropic(api_key=api_key)
        resp = client.messages.create(
            model=model,
            max_tokens=256,
            system=_system_prompt,
            messages=[{"role": "user", "content": user_msg}],
        )
        raw_text = resp.content[0].text.strip()
        # Strip optional markdown fences
        raw_text = re.sub(r"^```[a-z]*\n?", "", raw_text)
        raw_text = re.sub(r"\n?```$", "", raw_text)
        parsed = json.loads(raw_text)
        out["haiku_decision"]           = str(parsed.get("decision", "uncertain"))
        out["haiku_domain"]             = str(parsed.get("domain", ""))
        out["haiku_confidence"]         = str(parsed.get("confidence", ""))
        out["haiku_reason"]             = str(parsed.get("reason", ""))[:200]
        flags = parsed.get("risk_flags", [])
        out["haiku_risk_flags"]         = ", ".join(flags) if isinstance(flags, list) else str(flags)
        out["haiku_candidate_url"]      = str(parsed.get("candidate_url", ""))
        out["haiku_recommended_action"] = str(parsed.get("recommended_action", ""))
    except Exception as exc:
        out["haiku_used"]     = True
        out["haiku_error"]    = str(exc)[:200]
        out["haiku_decision"] = "uncertain"

    return out


def _serper_exact_page_identity_evidence(
    company_name: str,
    city: str,
    province: str,
    evidence_row: dict,
) -> dict:
    """
    Evaluate how strongly a single Serper evidence row identifies the exact company.
    Uses title + snippet + URL path.
    Returns a dict with strength rating and component flags.
    """
    title   = (evidence_row.get("title",   "") or "").lower()
    snippet = (evidence_row.get("snippet", "") or "").lower()
    url     = (evidence_row.get("url",     "") or "").lower()
    try:
        path = urlparse(url).path.lower()
    except Exception:
        path = ""

    combined = title + " " + snippet + " " + path

    # Tokenise company name
    _name_clean = re.sub(r"[^\w\s]", " ", company_name.lower())
    _tokens = [t for t in _name_clean.split() if len(t) >= 3
               and t not in ("spa", "srl", "srl", "snc", "sas", "spa", "soc", "per", "azioni",
                             "societa", "delle", "degli", "della", "dello", "dei", "gli", "the",
                             "and", "di", "da", "del", "dal", "con", "tra", "fra")]

    _has_name       = any(tok in combined for tok in _tokens) if _tokens else False
    _most_tokens    = (sum(1 for t in _tokens if t in combined) / len(_tokens) >= 0.6) if _tokens else False

    _city_prov = [v.lower() for v in [city, province] if v and len(v) >= 3]
    _has_loc   = any(v in combined for v in _city_prov) if _city_prov else False

    _contact_kw  = ("telefono", "fax", "tel.", "tel:", "email", "indirizzo", "sede legale",
                    "sede operativa", "p.iva", "partita iva", "cap ", "via ", "viale ", "corso ")
    _activity_kw = ("attivita", "settore", "servizi", "prodotti", "trasporti", "logistica",
                    "produzione", "commercio", "industria", "lavorazione", "costruzioni",
                    "impianti", "tecnologie", "forniture", "distribuzione")

    _has_contact  = any(kw in combined for kw in _contact_kw)
    _has_activity = any(kw in combined for kw in _activity_kw)

    # Determine strength
    if _most_tokens and (_has_loc or _has_contact):
        strength = "strong"
        reason   = "company name tokens + location/contact evidence"
    elif _most_tokens and _has_activity:
        strength = "strong"
        reason   = "company name tokens + activity description"
    elif _has_name and (_has_loc or _has_contact):
        strength = "medium"
        reason   = "partial name match + location or contact"
    elif _has_name and _has_activity:
        strength = "medium"
        reason   = "partial name match + activity signal"
    elif _has_name:
        strength = "weak"
        reason   = "partial name match only"
    else:
        strength = "none"
        reason   = "no company name tokens found in evidence"

    return {
        "has_company_name":              _has_name,
        "has_most_company_tokens":       _most_tokens,
        "has_location":                  _has_loc,
        "has_contact_or_registered_office": _has_contact,
        "has_activity_signal":           _has_activity,
        "serper_identity_strength":      strength,
        "serper_identity_reason":        reason,
    }


def _apply_haiku_decision(
    python_result: dict,
    haiku_result: dict,
    mode: str,
    raw_evidence: list[dict] | None = None,
) -> dict:
    """
    Merge Python result and Haiku result into final_* fields.
    Returns dict with final_selected_domain, final_decision_source, final_confidence.
    """
    python_domain = str(python_result.get("validated_domain", "") or "")
    python_conf   = str(python_result.get("domain_confidence", "") or "")

    if mode == _HAIKU_MODE_PYTHON:
        return {
            "final_selected_domain": python_domain,
            "final_decision_source": "python",
            "final_confidence":      python_conf,
        }

    if not haiku_result.get("haiku_used"):
        decision_h = haiku_result.get("haiku_decision", "")
        source = "haiku_skipped" if decision_h == "skipped_no_serper_evidence" else "python"
        return {
            "final_selected_domain": python_domain,
            "final_decision_source": source,
            "final_confidence":      python_conf,
        }

    decision     = haiku_result.get("haiku_decision", "uncertain")
    haiku_domain = str(haiku_result.get("haiku_domain", "") or "")
    haiku_conf   = haiku_result.get("haiku_confidence", "")

    # Build set of domains present in evidence for replace validation
    _evidence_domains: set[str] = set()
    if raw_evidence:
        for _e in raw_evidence:
            _d = _e.get("domain", "")
            if _d:
                _evidence_domains.add(_d.lower())

    if decision == "accept":
        return {
            "final_selected_domain": python_domain,
            "final_decision_source": "haiku_accept",
            "final_confidence":      haiku_conf or python_conf,
        }

    if decision == "replace" and haiku_domain:
        _hd_norm = haiku_domain.lower().lstrip("www.").split("/")[0]
        _in_evidence = (
            _hd_norm in _evidence_domains
            or haiku_domain.lower() in _evidence_domains
            or any(haiku_domain.lower() in d or d in haiku_domain.lower() for d in _evidence_domains)
        )
        if not _in_evidence:
            # Haiku suggested a domain not in evidence — ignore replace
            return {
                "final_selected_domain": python_domain,
                "final_decision_source": "haiku_invalid_replace_ignored",
                "final_confidence":      python_conf,
                "manual_review_needed":  True,
            }
        _review = haiku_conf != "High"
        return {
            "final_selected_domain": haiku_domain,
            "final_decision_source": "haiku_replace",
            "final_confidence":      haiku_conf or "Medium",
            "manual_review_needed":  _review,
        }

    if decision == "needs_firecrawl":
        return {
            "final_selected_domain": python_domain,
            "final_decision_source": "haiku_needs_firecrawl",
            "final_confidence":      python_conf,
        }

    if decision == "reject":
        _py_conf_low = python_conf.lower() in ("low", "none", "")
        if _py_conf_low:
            return {
                "final_selected_domain": "",
                "final_decision_source": "haiku_reject",
                "final_confidence":      "None",
                "manual_review_needed":  True,
            }
        # Python confidence is Medium or High — protect the domain
        return {
            "final_selected_domain": python_domain,
            "final_decision_source": "haiku_reject_protected",
            "final_confidence":      python_conf,
            "manual_review_needed":  True,
        }

    if decision == "uncertain":
        return {
            "final_selected_domain": python_domain,
            "final_decision_source": "haiku_uncertain",
            "final_confidence":      python_conf,
            "manual_review_needed":  True,
        }

    # skipped / error / unknown — keep python result
    source = "haiku_skipped" if decision == "skipped_no_serper_evidence" else "haiku_uncertain"
    return {
        "final_selected_domain": python_domain,
        "final_decision_source": source,
        "final_confidence":      python_conf,
    }


# =============================================================================
# JINA WEBSITE VERIFIER
# =============================================================================


def _jina_should_run(
    result: dict,
    name_variants: dict,
    raw_ev: list[dict],
    jina_mode: str,
    debug_mode: bool = False,
) -> bool:
    """Return True if Jina verification should be triggered for this row."""
    if jina_mode == _JINA_MODE_OFF:
        return False
    if jina_mode == _JINA_MODE_ALL_DEBUG and debug_mode:
        return True

    conf = str(result.get("final_confidence") or result.get("domain_confidence") or "").strip().lower()
    manual = str(result.get("manual_review_needed", "")).lower() in ("true", "1", "yes")
    final_dom = str(result.get("final_selected_domain") or result.get("validated_domain") or "")

    # Trigger 1: confidence is Medium, Low, None, or empty
    if conf in ("medium", "low", "none", ""):
        return True

    # Trigger 2: manual_review_needed
    if manual:
        return True

    # Trigger 3: top-2 candidate scores are close (delta < 0.25)
    scored: dict[str, float] = {}
    for e in raw_ev:
        if not e.get("used"):
            continue
        dom = e.get("domain", "")
        try:
            sc = float(e.get("score", 0))
        except (TypeError, ValueError):
            sc = 0.0
        if dom and sc > scored.get(dom, -1.0):
            scored[dom] = sc
    if len(scored) >= 2:
        top2 = sorted(scored.values(), reverse=True)[:2]
        if top2[0] - top2[1] < 0.25:
            return True

    # Trigger 4: selected domain came from a site:.it query
    query_used = str(result.get("search_query_used", "") or "")
    if query_used.lower().startswith("site:.it"):
        return True

    # Trigger 5: no location_match and no email_match for selected domain
    sel = final_dom.lower()
    if sel and result.get("domain_source") not in (SRC_ORIGINAL, SRC_EMAIL, SRC_SERPER_EMAIL):
        has_loc   = any(e.get("domain", "").lower() == sel and e.get("location_match") for e in raw_ev)
        has_email = any(e.get("domain", "").lower() == sel and e.get("email_match") for e in raw_ev)
        if not has_loc and not has_email:
            return True

    # Trigger 6: brand is single-word, short, generic, or famous
    brand = (name_variants.get("brand") or "").strip()
    brand_clean = re.sub(r"[^\w]", "", brand.lower())
    if brand_clean and (
        len(brand_clean) <= 5
        or brand_clean in _JINA_FAMOUS_BRANDS
        or _brand_is_ambiguous(brand)
    ):
        return True

    # Trigger 7: selected domain contains risky patterns
    if final_dom and _domain_has_risky_marker(final_dom)[0]:
        return True

    return False


def _jina_fetch_page(
    domain: str,
    slug: str,
    jina_api_key: str | None = None,
    timeout: int = 12,
) -> tuple[str, str]:
    """
    Fetch domain+slug via Jina Reader. Returns (text, fetch_status).
    Caches results by (domain, slug) for the lifetime of the Streamlit session.
    """
    if not slug.startswith("/"):
        slug = f"/{slug}"
    cache_key = (domain, slug)
    if cache_key in _JINA_CACHE:
        return _JINA_CACHE[cache_key]

    jina_url = f"{_JINA_READER_BASE}https://{domain}{slug}"
    headers = {"Accept": "text/plain", "X-Return-Format": "text"}
    if jina_api_key:
        headers["Authorization"] = f"Bearer {jina_api_key}"

    text = ""
    try:
        resp = requests.get(jina_url, headers=headers, timeout=timeout)
        if resp.status_code == 200:
            text  = resp.text[:_JINA_MAX_CHARS]
            status = "ok"
        elif resp.status_code == 404:
            status = "404"
        else:
            status = f"http_{resp.status_code}"
    except requests.Timeout:
        status = "timeout"
    except Exception as exc:
        status = f"error:{str(exc)[:60]}"

    _JINA_CACHE[cache_key] = (text, status)
    return text, status


def _jina_extract_evidence(text: str, domain: str) -> dict:
    """Extract identity signals from Jina-fetched page text."""
    t = text.lower()

    _pi_re    = re.compile(r"p\.?\s*iva[:\s]*([0-9]{11})", re.IGNORECASE)
    _phone_re = re.compile(
        r"(?:tel\.?|fax\.?|phone|telefono|cellulare)[:\s]*([\+0][\d\s\-\(\)\.]{7,20})",
        re.IGNORECASE,
    )
    _email_re = re.compile(r"[\w.\-]+@[\w.\-]+\.[a-z]{2,6}", re.IGNORECASE)

    pi_match    = _pi_re.search(text)
    phone_match = _phone_re.search(text)
    emails      = _email_re.findall(text)
    domain_email = next((e for e in emails if domain.split(".")[0] in e.lower()), "")

    directory_signal = bool(re.search(
        r"fatturato|bilancio|visura|scheda\s+azienda|registro\s+imprese|"
        r"company\s+profile|business\s+profile|similar\s+companies|competitors|"
        r"employees\s+count|revenue|founded\s+in\s+\d{4}\s+·",
        t,
    ))
    wrong_country_signal = bool(
        re.search(r"\b(?:united\s+states|usa|uk\s+company|british|deutschland|français)\b", t)
        and not re.search(r"\b(?:italia|italian|italy|italiano|italiana)\b", t)
    )
    official_signal = bool(re.search(
        r"sito\s+ufficiale|official\s+(?:website|site)|benvenuti|chi\s+siamo|"
        r"about\s+us|la\s+nostra\s+azienda|our\s+company|contattaci",
        t,
    ))
    is_italian = bool(re.search(
        r"\b(?:italia|italiano|italiana|italiani|azienda|prodotti|servizi|contatti|"
        r"via\s+[a-z]|piazza\s+[a-z]|corso\s+[a-z]|srl|spa|snc|sas)\b",
        t,
    ))

    return {
        "partita_iva_on_site":        pi_match.group(1)     if pi_match    else "",
        "phone_on_site":              phone_match.group(1).strip() if phone_match else "",
        "email_on_site":              domain_email or (emails[0] if emails else ""),
        "directory_or_profile_signal": directory_signal,
        "wrong_country_signal":       wrong_country_signal,
        "official_site_signal":       official_signal,
        "language_country_signal":    "it" if is_italian else "unknown",
    }


def _jina_score_pages(
    domain: str,
    company_name: str,
    city: str,
    province: str,
    email_domain: str,
    name_variants: dict,
    pages_data: list[dict],
) -> tuple[float, str, list[str]]:
    """
    Aggregate Jina page evidence into a score.
    Returns (score, hint_str, signal_list).
    """
    brand_lower  = re.sub(r"[^\w]", "", (name_variants.get("brand") or company_name).lower())
    brand_nodot  = re.sub(r"[^\w]", "", (name_variants.get("brand_nodot") or brand_lower).lower())

    any_ok = any(p["fetch_status"] == "ok" and p["text"] for p in pages_data)
    if not any_ok:
        return 0.0, "no_pages_fetched", []

    score   = 0.0
    signals: list[str] = []

    has_dir = has_wrong = has_official = has_italian = has_email = has_brand = has_city = has_prov = False

    for p in pages_data:
        ev = p.get("evidence", {})
        tl = p.get("text", "").lower()
        if ev.get("directory_or_profile_signal"):
            has_dir = True
        if ev.get("wrong_country_signal"):
            has_wrong = True
        if ev.get("official_site_signal"):
            has_official = True
        if ev.get("language_country_signal") == "it":
            has_italian = True
        if email_domain and email_domain in ev.get("email_on_site", "").lower():
            has_email = True
        if brand_lower in tl or brand_nodot in tl:
            has_brand = True
        if city and len(city) >= 3 and city.lower() in tl:
            has_city = True
        if province and len(province) >= 2 and province.lower() in tl:
            has_prov = True

    if has_dir:
        score -= 1.5
        signals.append("directory/profile page detected")
    if has_wrong:
        score -= 0.8
        signals.append("wrong country signal")
    if has_official:
        score += 0.5
        signals.append("official site signal")
    if has_italian:
        score += 0.3
        signals.append("Italian language detected")
    if has_email:
        score += 0.8
        signals.append(f"email domain {email_domain} found on site")
    if has_brand:
        score += 0.6
        signals.append(f"brand '{brand_lower}' found on site")
    if has_city:
        score += 0.4
        signals.append(f"city '{city}' found on site")
    if has_prov:
        score += 0.3
        signals.append(f"province '{province}' found on site")
    if domain.endswith(".it"):
        score += 0.2
        signals.append(".it domain")

    if score >= 1.5:
        hint = "strong_match"
    elif score >= 0.5:
        hint = "likely_match"
    elif score <= -1.0:
        hint = "directory_or_wrong"
    elif score < 0:
        hint = "weak_negative"
    else:
        hint = "insufficient_evidence"

    return round(score, 3), hint, signals


def _jina_verify_candidates(
    company_name: str,
    city: str,
    province: str,
    email_domain: str,
    name_variants: dict,
    candidates: list[str],
    current_domain: str,
    jina_api_key: str | None = None,
) -> tuple[dict, list[dict]]:
    """
    Run Jina verification on candidate domains.
    Returns (jina_result_dict, jina_debug_rows).
    """
    base_out = {
        "jina_verifier_used":       True,
        "jina_verified_domain":     "",
        "jina_verifier_confidence": "",
        "jina_verifier_decision":   "",
        "jina_verifier_reason":     "",
        "jina_evidence_legal_name": "",
        "jina_evidence_address":    "",
        "jina_evidence_city":       "",
        "jina_evidence_phone":      "",
        "jina_evidence_email":      "",
        "jina_evidence_partita_iva": "",
        "jina_pages_fetched":       0,
        "jina_fetch_status":        "",
    }
    debug_rows: list[dict] = []

    if not candidates:
        base_out.update(
            jina_verifier_used=False,
            jina_verifier_decision="no_candidates",
            jina_verifier_reason="No candidates to verify",
        )
        return base_out, debug_rows

    domain_scores: dict[str, tuple[float, str, list[str]]] = {}
    domain_best_ev: dict[str, dict] = {}
    total_fetched  = 0
    all_statuses: list[str] = []

    for domain in candidates[:5]:
        if not domain:
            continue
        pages_data: list[dict] = []
        for slug in _JINA_SLUGS:
            text, status = _jina_fetch_page(domain, slug, jina_api_key)
            all_statuses.append(status)
            ev: dict = {}
            if text:
                total_fetched += 1
                ev = _jina_extract_evidence(text, domain)
            pages_data.append({"slug": slug, "text": text, "fetch_status": status,
                                "chars_fetched": len(text), "evidence": ev})
            debug_rows.append({
                "company_name":       company_name,
                "candidate_domain":   domain,
                "candidate_url":      f"https://{domain}{slug}",
                "page_slug":          slug,
                "fetch_status":       status,
                "chars_fetched":      len(text),
                "extracted_legal_name": "",
                "extracted_address":  "",
                "extracted_phone":    ev.get("phone_on_site", ""),
                "extracted_email":    ev.get("email_on_site", ""),
                "extracted_partita_iva": ev.get("partita_iva_on_site", ""),
                "verifier_score":     "",  # filled after scoring
                "verifier_decision":  "",
                "verifier_reason":    "",
            })

        sc, hint, sigs = _jina_score_pages(
            domain, company_name, city, province, email_domain, name_variants, pages_data
        )
        domain_scores[domain] = (sc, hint, sigs)

        # Best evidence page for the output columns
        best_pg_ev = next((p["evidence"] for p in pages_data if p["fetch_status"] == "ok" and p["text"]), {})
        domain_best_ev[domain] = best_pg_ev

        # Back-fill score into debug rows for this domain
        reason_str = "; ".join(sigs[:3]) if sigs else hint
        for row in debug_rows:
            if row["candidate_domain"] == domain and row["verifier_score"] == "":
                row["verifier_score"]   = sc
                row["verifier_decision"] = hint
                row["verifier_reason"]  = reason_str

    base_out["jina_pages_fetched"] = total_fetched
    base_out["jina_fetch_status"]  = "; ".join(dict.fromkeys(all_statuses))[:200]

    if not domain_scores:
        base_out.update(
            jina_verifier_decision="fetch_failed",
            jina_verifier_reason="No pages could be fetched",
            jina_verifier_confidence="None",
        )
        return base_out, debug_rows

    sorted_doms = sorted(domain_scores.items(), key=lambda x: x[1][0], reverse=True)
    best_dom, (best_sc, best_hint, best_sigs) = sorted_doms[0]
    cur_sc = domain_scores.get(current_domain, (0.0, "not_fetched", []))[0]
    reason_str = "; ".join(best_sigs[:3]) if best_sigs else best_hint

    def _fill_ev(dom: str) -> dict:
        ev = domain_best_ev.get(dom, {})
        return {
            "jina_evidence_phone":       ev.get("phone_on_site", ""),
            "jina_evidence_email":       ev.get("email_on_site", ""),
            "jina_evidence_partita_iva": ev.get("partita_iva_on_site", ""),
        }

    if best_sc <= -1.0:
        # Best candidate is a directory/wrong site
        alt = [(d, s, h, sg) for d, (s, h, sg) in sorted_doms if d != best_dom and s > 0.3]
        if alt and best_dom == current_domain:
            alt_dom, alt_sc, _, alt_sg = alt[0]
            base_out.update(
                jina_verified_domain=alt_dom,
                jina_verifier_decision="replace",
                jina_verifier_confidence="Medium" if alt_sc >= 0.8 else "Low",
                jina_verifier_reason=f"Current domain is {best_hint}; {alt_dom} shows: {'; '.join(alt_sg[:2]) or 'better signals'}",
                **_fill_ev(alt_dom),
            )
        else:
            base_out.update(
                jina_verified_domain="",
                jina_verifier_decision="reject" if current_domain == best_dom else "uncertain",
                jina_verifier_confidence="None",
                jina_verifier_reason=f"All Jina candidates show negative signals: {reason_str}",
            )
    elif best_dom != current_domain and best_sc > cur_sc + 0.4:
        # A different domain has markedly stronger evidence
        base_out.update(
            jina_verified_domain=best_dom,
            jina_verifier_decision="replace",
            jina_verifier_confidence="High" if best_sc >= 1.5 else "Medium",
            jina_verifier_reason=f"{best_dom} has stronger evidence than {current_domain}: {reason_str}",
            **_fill_ev(best_dom),
        )
    elif best_sc >= 0.5:
        # Confirm Python selection (or best domain)
        confirmed = current_domain if current_domain in domain_scores else best_dom
        conf_label = "High" if best_sc >= 1.5 else ("Medium" if best_sc >= 0.8 else "Low")
        base_out.update(
            jina_verified_domain=confirmed,
            jina_verifier_decision="confirm",
            jina_verifier_confidence=conf_label,
            jina_verifier_reason=f"Jina confirms {confirmed}: {reason_str}",
            **_fill_ev(confirmed),
        )
    else:
        # Insufficient evidence
        base_out.update(
            jina_verified_domain=current_domain,
            jina_verifier_decision="uncertain",
            jina_verifier_confidence="Low",
            jina_verifier_reason="Jina could not gather sufficient identity evidence",
            **_fill_ev(current_domain),
        )

    return base_out, debug_rows


def _apply_jina_decision(result: dict, jina_result: dict) -> dict:
    """
    Apply Jina verifier decision on top of Python+Haiku final_* fields.
    Returns a dict of fields to update on result.
    """
    if not jina_result.get("jina_verifier_used"):
        return {}

    decision     = jina_result.get("jina_verifier_decision", "")
    verified_dom = str(jina_result.get("jina_verified_domain", "") or "")
    jina_conf    = jina_result.get("jina_verifier_confidence", "")

    if decision == "replace" and verified_dom:
        return {
            "final_selected_domain": verified_dom,
            "final_decision_source": "jina_replace",
            "final_confidence":      jina_conf or "Medium",
            "manual_review_needed":  False,
        }
    if decision == "reject":
        return {
            "final_selected_domain": "",
            "final_decision_source": "jina_reject",
            "final_confidence":      "None",
            "manual_review_needed":  True,
        }
    if decision == "confirm":
        updates: dict = {"final_decision_source": "jina_confirm"}
        # Upgrade confidence if Jina is more certain
        current_conf = (result.get("final_confidence") or "").lower()
        if jina_conf == "High" and current_conf in ("medium", "low", "none", ""):
            updates["final_confidence"] = "High"
        return updates
    # uncertain / fetch_failed / no_candidates
    return {"final_decision_source": f"jina_{decision or 'uncertain'}"}


# =============================================================================
# PROFESSIONAL SITE SIGNALS SCORING
# =============================================================================

_PRO_SITE_NAV_RE = re.compile(
    r"\b(chi\s+siamo|la\s+nostra\s+azienda|azienda|servizi|prodotti|settori|"
    r"business\s+sectors|companies|sustainability|sostenibilit[àa]|media|"
    r"contatti|contacts|lavora\s+con\s+noi|work\s+with\s+us|about\s+us)\b", re.I
)
_PRO_SITE_LEGAL_RE = re.compile(
    r"\b(privacy\s+policy|cookie\s+policy|terms\s+and\s+conditions|"
    r"p\.?\s*iva|c\.?\s*f\.|codice\s+fiscale|vat\s+number|"
    r"sede\s+legale|capitale\s+sociale|rea\s+\w+|pec\b|"
    r"registered\s+office)\b", re.I
)
_PRO_SITE_MULTI_RE = re.compile(
    r"\b(italiano|english|français|deutsch|español|"
    r"language|lingua)\b|/en/|/it/|/fr/|/de/", re.I
)
_PRO_SITE_MEDIA_RE = re.compile(
    r"\b(wp-content|elementor|webflow|cms|slider|carousel|hero|banner|"
    r"\.jpg|\.jpeg|\.png|\.webp|\.mp4|\.svg)\b", re.I
)


def _score_professional_site(text: str, domain: str) -> dict:
    """Score professional site signals from Firecrawl markdown text."""
    if not text:
        return {
            "professional_site_score": 0,
            "professional_site_level": "none",
            "professional_site_signals": "",
        }
    score = 0
    signals = []

    # HTTPS bonus
    if not (domain or "").startswith("http://"):
        score += 1
        signals.append("https")

    nav_matches = _PRO_SITE_NAV_RE.findall(text)
    if len(nav_matches) >= 3:
        score += 1
        signals.append(f"navigation({len(nav_matches)} labels)")

    legal_matches = _PRO_SITE_LEGAL_RE.findall(text)
    if legal_matches:
        score += 1
        signals.append(f"legal/footer({len(legal_matches)} terms)")

    if _PRO_SITE_MULTI_RE.search(text):
        score += 1
        signals.append("multilanguage")

    if _PRO_SITE_MEDIA_RE.search(text):
        score += 1
        signals.append("media/cms")

    if len(nav_matches) >= 5:
        score += 1
        signals.append("rich_navigation")

    if len(legal_matches) >= 3:
        score += 1
        signals.append("rich_footer")

    level = "none" if score == 0 else "weak" if score <= 2 else "medium" if score <= 4 else "strong"

    return {
        "professional_site_score": score,
        "professional_site_level": level,
        "professional_site_signals": "; ".join(signals),
    }


# =============================================================================
# FIRECRAWL WEBSITE VERIFIER
# =============================================================================


def _fc_load_keys() -> list[str]:
    """
    Load all Firecrawl API keys from st.secrets and environment variables.
    Sources tried in order:
      1. st.secrets["FIRECRAWL_API_KEYS"]  (list)
      2. st.secrets["FIRECRAWL_API_KEY"] / ["firecrawl_api_key"]
      3. st.secrets["FIRECRAWL_API_KEY_1"] … ["FIRECRAWL_API_KEY_5"]
      4. os.getenv("FIRECRAWL_API_KEYS") (comma-separated)
      5. os.getenv("FIRECRAWL_API_KEY") / os.getenv("firecrawl_api_key")
      6. os.getenv("FIRECRAWL_API_KEY_1") … os.getenv("FIRECRAWL_API_KEY_5")
    Never logs or displays key values.
    """
    raw: list[str] = []
    try:
        _keys_list = st.secrets.get("FIRECRAWL_API_KEYS")
        if _keys_list:
            if isinstance(_keys_list, (list, tuple)):
                raw.extend(str(k) for k in _keys_list)
            else:
                raw.append(str(_keys_list))
        for _name in ("FIRECRAWL_API_KEY", "firecrawl_api_key"):
            _v = st.secrets.get(_name)
            if _v:
                raw.append(str(_v))
        for _i in range(1, 6):
            _v = st.secrets.get(f"FIRECRAWL_API_KEY_{_i}")
            if _v:
                raw.append(str(_v))
    except Exception:
        pass
    _env_multi = os.getenv("FIRECRAWL_API_KEYS", "")
    if _env_multi:
        raw.extend(_env_multi.split(","))
    for _name in ("FIRECRAWL_API_KEY", "firecrawl_api_key"):
        _v = os.getenv(_name)
        if _v:
            raw.append(_v)
    for _i in range(1, 6):
        _v = os.getenv(f"FIRECRAWL_API_KEY_{_i}")
        if _v:
            raw.append(_v)
    # Deduplicate preserving order, strip whitespace, drop empty
    seen: set[str] = set()
    result: list[str] = []
    for k in raw:
        k = k.strip()
        if k and k not in seen:
            seen.add(k)
            result.append(k)
    return result


def _fc_load_key() -> str | None:
    """Load a single Firecrawl API key (backward-compatible wrapper)."""
    keys = _fc_load_keys()
    return keys[0] if keys else None


# HTTP status codes / response text patterns that indicate a key/account problem
# and should trigger failover to the next key.
_FC_KEY_FAILURE_CODES = {401, 402, 403, 429}
_FC_KEY_FAILURE_TERMS = (
    "quota", "credits", "rate limit", "payment", "billing",
    "unauthorized", "forbidden", "invalid api key", "invalid_api_key",
)

# Fail-fast thresholds — checked at runtime during batch processing.
_FC_FAIL_FAST_CONSECUTIVE_FAILURES: int   = 10
_FC_FAIL_FAST_MIN_ATTEMPTS:         int   = 30
_FC_FAIL_FAST_MIN_SUCCESS_RATE:     float = 0.20
_FC_HEALTH_LOG_EVERY:               int   = 25   # print status every N FC attempts (CLI)


class FcFailFastError(RuntimeError):
    """Raised when Firecrawl health exceeds a fail-fast threshold."""


def firecrawl_preflight_check(
    fc_keys: list[str],
    test_url: str = "https://www.myngle.com",
    timeout: int = 10,
    fc_location: dict | None = None,
) -> dict:
    """
    Test each Firecrawl API key independently (no failover).
    Returns a summary dict — never exposes actual key values.

    preflight_status values:
      "OK"       — all keys work
      "DEGRADED" — at least one key works but not all
      "FAILED"   — no keys work
    """
    key_statuses: dict[str, str] = {}   # "key1" -> "ok" / "http_402" / "timeout" / …
    keys_ok: int   = 0
    keys_fail: int = 0
    first_error: str = ""

    _payload: dict = {"url": test_url, "formats": ["markdown"]}
    if fc_location:
        _payload["location"] = fc_location

    for _ki, _key in enumerate(fc_keys, start=1):
        _label = f"key{_ki}"
        try:
            resp = requests.post(
                _FC_API_URL,
                headers={"Authorization": f"Bearer {_key}", "Content-Type": "application/json"},
                json=_payload,
                timeout=timeout,
            )
            if resp.status_code == 200:
                key_statuses[_label] = "ok"
                keys_ok += 1
            else:
                _raw = f"http_{resp.status_code}"
                if _fc_is_key_failure(resp.status_code, resp.text):
                    _raw = f"key_failure:{resp.status_code}"
                key_statuses[_label] = _raw
                keys_fail += 1
                if not first_error:
                    first_error = f"{_label}: {_raw}"
        except requests.Timeout:
            key_statuses[_label] = "timeout"
            keys_fail += 1
            if not first_error:
                first_error = f"{_label}: timeout"
        except Exception as exc:
            _err = f"error:{str(exc)[:80]}"
            key_statuses[_label] = _err
            keys_fail += 1
            if not first_error:
                first_error = f"{_label}: {_err}"

    if not fc_keys:
        preflight_status = "FAILED"
        preflight_error  = "No Firecrawl API keys provided."
    elif keys_ok == 0:
        preflight_status = "FAILED"
        preflight_error  = f"No working Firecrawl keys. First error: {first_error}"
    elif keys_fail > 0:
        preflight_status = "DEGRADED"
        preflight_error  = f"{keys_fail} of {len(fc_keys)} key(s) failed. First: {first_error}"
    else:
        preflight_status = "OK"
        preflight_error  = ""

    return {
        "keys_total":       len(fc_keys),
        "keys_ok":          keys_ok,
        "keys_failed":      keys_fail,
        "key_statuses":     key_statuses,   # {"key1": "ok", "key2": "key_failure:402"}
        "preflight_status": preflight_status,
        "preflight_error":  preflight_error,
    }


def _make_fc_health() -> dict:
    """Return a zeroed Firecrawl health-counter dict."""
    return {
        "requests_attempted":       0,
        "pages_successful":         0,
        "key_failovers":            0,
        "key_failure_events":       0,
        "quota_or_billing_errors":  0,
        "rate_limit_errors":        0,
        "timeouts":                 0,
        "exceptions":               0,
        "consecutive_failures":     0,
        "consecutive_failures_max": 0,
        "last_statuses":            [],
        "fail_fast_triggered":      False,
        "fail_fast_reason":         "",
    }


def _safe_extract_http_status(status_value) -> int:
    """Extract the first HTTP status code from a Firecrawl status string.

    Handles clean values ("http_500") and combined values ("http_500; ok",
    "http_429; http_200").  Returns 0 when no three-digit code is found.
    Never raises.
    """
    import re as _re
    if not status_value:
        return 0
    m = _re.search(r"http_(\d{3})", str(status_value).lower())
    return int(m.group(1)) if m else 0


def _fc_health_update(
    health: dict,
    fetch_status: str,
    key_statuses_str: str,
    failover_count: int,
    fc_was_used: bool,
) -> None:
    """
    Update runtime health counters from a single row's Firecrawl result.
    fc_was_used: True when Firecrawl was actually invoked for this row.
    """
    if not fc_was_used:
        return

    health["requests_attempted"] = health.get("requests_attempted", 0) + 1
    health["key_failovers"]      = health.get("key_failovers", 0) + max(failover_count, 0)
    if failover_count > 0:
        health["key_failure_events"] = health.get("key_failure_events", 0) + 1

    _st = fetch_status.lower()
    if _st == "ok":
        health["pages_successful"]    = health.get("pages_successful", 0) + 1
        health["consecutive_failures"] = 0
    elif _st == "timeout":
        health["timeouts"]             = health.get("timeouts", 0) + 1
        health["consecutive_failures"] = health.get("consecutive_failures", 0) + 1
    elif _st.startswith("error:"):
        health["exceptions"]           = health.get("exceptions", 0) + 1
        health["consecutive_failures"] = health.get("consecutive_failures", 0) + 1
    else:
        # HTTP error or other non-ok, non-timeout status
        if _fc_is_key_failure(
            _safe_extract_http_status(_st),
            _st,
        ):
            health["quota_or_billing_errors"] = health.get("quota_or_billing_errors", 0) + 1
        if "429" in _st or "rate" in _st:
            health["rate_limit_errors"] = health.get("rate_limit_errors", 0) + 1
        health["consecutive_failures"] = health.get("consecutive_failures", 0) + 1

    health["consecutive_failures_max"] = max(
        health.get("consecutive_failures_max", 0),
        health.get("consecutive_failures", 0),
    )
    _last: list[str] = health.get("last_statuses", [])
    _last.append(fetch_status or "unknown")
    health["last_statuses"] = _last[-10:]


def _fc_health_check_fail_fast(
    health: dict,
    enabled: bool = True,
) -> str | None:
    """
    Return a human-readable fail-fast reason if a threshold is exceeded.
    Returns None when processing should continue.
    """
    if not enabled:
        return None
    consec = health.get("consecutive_failures", 0)
    if consec >= _FC_FAIL_FAST_CONSECUTIVE_FAILURES:
        return (
            f"Firecrawl fail-fast: {consec} consecutive failures "
            f"(threshold {_FC_FAIL_FAST_CONSECUTIVE_FAILURES}). "
            "Check API keys and account status."
        )
    att  = health.get("requests_attempted", 0)
    succ = health.get("pages_successful", 0)
    if att >= _FC_FAIL_FAST_MIN_ATTEMPTS:
        rate = succ / att if att else 0.0
        if rate < _FC_FAIL_FAST_MIN_SUCCESS_RATE:
            return (
                f"Firecrawl fail-fast: success rate {rate:.1%} after {att} attempts "
                f"is below minimum {_FC_FAIL_FAST_MIN_SUCCESS_RATE:.0%}."
            )
    return None


def _fc_health_summary_line(health: dict) -> str:
    """Return a one-line Firecrawl health status string for logging."""
    att   = health.get("requests_attempted", 0)
    succ  = health.get("pages_successful", 0)
    rate  = round(succ / att * 100, 1) if att else 0.0
    fo    = health.get("key_failovers", 0)
    quota = health.get("quota_or_billing_errors", 0)
    to    = health.get("timeouts", 0)
    consec = health.get("consecutive_failures", 0)
    ff    = health.get("fail_fast_triggered", False)
    status = "FAILED" if ff else ("OK" if rate >= 50 or att == 0 else "DEGRADED")
    return (
        f"attempts={att}, ok={succ}, success_rate={rate}%, "
        f"failovers={fo}, quota_errors={quota}, timeouts={to}, "
        f"consecutive_failures={consec}, status={status}"
    )


def _fc_is_key_failure(status_code: int, resp_text: str) -> bool:
    """Return True if the response indicates a key/account/quota problem."""
    if status_code in _FC_KEY_FAILURE_CODES:
        return True
    _lower = resp_text.lower()
    return any(t in _lower for t in _FC_KEY_FAILURE_TERMS)


def _fc_scrape(
    url: str,
    fc_key: str | list[str],
    timeout: int = 15,
    fc_location: dict | None = None,
) -> tuple[str, str, dict]:
    """
    Scrape a single URL via Firecrawl v1 scrape endpoint.
    Returns (markdown_text, fetch_status, metadata_dict).
    metadata_dict includes:
      - redirect_url / redirect_domain / canonical_url / final_url
      - firecrawl_key_index_used   (1-based)
      - firecrawl_key_failover_count
      - firecrawl_key_statuses     (e.g. "key1:http_402; key2:ok")
    Accepts a single key string or a list of keys.
    On key/quota/auth failure tries each key in order; does NOT failover for
    timeout / 404 / page-content issues.
    Results cached by (URL, location) for the lifetime of the Streamlit session.
    """
    _cache_key = _make_fc_cache_key(url, fc_location)
    if _cache_key in _FC_CACHE:
        return _FC_CACHE[_cache_key]

    # Normalise to list
    if isinstance(fc_key, list):
        keys = [k for k in fc_key if k]
    else:
        keys = [fc_key] if fc_key else []
    if not keys:
        result = ("", "no_key", {"firecrawl_key_index_used": 0,
                                  "firecrawl_key_failover_count": 0,
                                  "firecrawl_key_statuses": "no_key"})
        _FC_CACHE[_cache_key] = result
        return result

    _payload: dict = {"url": url, "formats": ["markdown"]}
    if fc_location:
        _payload["location"] = fc_location

    text, status, meta = "", "not_attempted", {}
    key_statuses: list[str] = []
    key_index_used = 1
    failover_count = 0

    for _ki, _key in enumerate(keys, start=1):
        try:
            resp = requests.post(
                _FC_API_URL,
                headers={"Authorization": f"Bearer {_key}", "Content-Type": "application/json"},
                json=_payload,
                timeout=timeout,
            )
            if resp.status_code == 200:
                body = resp.json()
                page_data = body.get("data") or {}
                text  = (page_data.get("markdown") or "")[:_FC_MAX_CHARS]
                raw_meta = page_data.get("metadata") or {}
                meta = dict(raw_meta)

                redirect_url  = meta.get("sourceURL") or meta.get("url") or meta.get("ogUrl") or ""
                canonical_url = meta.get("canonicalUrl") or meta.get("canonical") or ""
                final_url     = canonical_url or redirect_url or ""
                meta["redirect_url"]    = redirect_url
                meta["canonical_url"]   = canonical_url
                meta["final_url"]       = final_url

                _redirect_domain = ""
                _origin_domain = re.sub(r"^https?://([^/]+).*$", r"\1", url).lower().lstrip("www.")
                if final_url:
                    _rd = re.sub(r"^https?://([^/]+).*$", r"\1", final_url).lower().lstrip("www.")
                    if _rd and _rd != _origin_domain:
                        _redirect_domain = _rd
                meta["redirect_domain"] = _redirect_domain

                status = "ok"
                key_statuses.append(f"key{_ki}:ok")
                key_index_used = _ki
                break  # success — stop trying keys

            elif resp.status_code == 404:
                status = "404"
                key_statuses.append(f"key{_ki}:404")
                break  # page problem, not a key problem

            else:
                _raw_status = f"http_{resp.status_code}"
                key_statuses.append(f"key{_ki}:{_raw_status}")
                if _fc_is_key_failure(resp.status_code, resp.text):
                    # Key/quota problem — try next key
                    status = _raw_status
                    key_index_used = _ki
                    if _ki < len(keys):
                        failover_count += 1
                    continue
                # Non-key HTTP error — stop trying
                status = _raw_status
                key_index_used = _ki
                break

        except requests.Timeout:
            status = "timeout"
            key_statuses.append(f"key{_ki}:timeout")
            key_index_used = _ki
            break  # timeout is a page/network problem, not a key problem
        except Exception as exc:
            status = f"error:{str(exc)[:80]}"
            key_statuses.append(f"key{_ki}:{status}")
            key_index_used = _ki
            break

    meta["firecrawl_key_index_used"]   = key_index_used
    meta["firecrawl_key_failover_count"] = failover_count
    meta["firecrawl_key_statuses"]     = "; ".join(key_statuses)

    # Only cache if the final status is not a key/account failure
    # (prevents blocking a future run from trying another key)
    _is_key_fail = status.startswith("http_") and any(
        str(c) in status for c in _FC_KEY_FAILURE_CODES
    )
    if not _is_key_fail:
        _FC_CACHE[_cache_key] = (text, status, meta)
    return text, status, meta


def _detect_neg_source(text: str, domain: str = "") -> str:
    """
    Return the first matching negative-source-type key, or empty string if none.

    Thresholds:
    - news_media: 3+ matches (company websites often have news sections)
    - public_school / university: 1+ match (very distinctive vocabulary)
    - others: 2+ matches

    Also checks domain TLD: .edu.it → public_school, .gov.it etc. → government.
    """
    if not text and not domain:
        return ""

    # Domain-level hard signals (checked before text)
    if domain:
        if _EDU_IT_RE.search(domain.lower()):
            return "public_school"
        if _GOV_IT_RE.search(domain.lower()):
            return "government"

    if not text:
        return ""
    tl = text.lower()

    _thresholds = {
        "news_media": 3,
        "public_school": 1,   # very specific vocabulary, 1 match is enough
        "university": 1,
    }
    for src_type, pat in _NEG_SOURCE_RES.items():
        hits = pat.findall(tl)
        threshold = _thresholds.get(src_type, 2)
        if len(hits) >= threshold:
            return src_type
    return ""


_BUSINESS_DIR_DOMAINS = frozenset({
    "creditsafe.com", "infobel.com", "paginegialle.it", "tuttitalia.it",
    "dnb.com", "kompass.com", "europages.it", "europages.com",
    "registroimprese.it", "atoka.io", "offertedilavoro.it",
})
_DOCUMENT_HOSTING_DOMAINS = frozenset({
    "yumpu.com", "issuu.com", "scribd.com", "slideshare.net",
    "calameo.com", "docplayer.it", "docplayer.net",
})
_PRODUCT_PATH_RE = re.compile(
    r"/(product|products|shop|catalog|catalogue|catalogo|cart|checkout|"
    r"negozio|articolo|article|item|sku)/",
    re.I,
)
_PDF_URL_RE = re.compile(r"\.pdf(\?|#|$)", re.I)


def _detect_neg_source_url(url: str, candidate_domain: str = "") -> str:
    """
    Detect hard negative source type purely from the URL.
    Returns a string key or empty string.
    """
    if not url:
        return ""
    try:
        _parsed = urlparse(url)
        _url_domain = _parsed.netloc.lower().lstrip("www.")
        _path = (_parsed.path or "").lower()
    except Exception:
        _url_domain = ""
        _path = url.lower()

    # PDF hosted elsewhere (not on the candidate's own domain)
    if _PDF_URL_RE.search(_path) or _path.endswith(".pdf"):
        if candidate_domain and _url_domain and _url_domain != candidate_domain.lstrip("www."):
            return "pdf_list"
        if not candidate_domain:
            return "pdf_list"
    # /wp-content/uploads/ PDF on any domain (typically blog/client reference)
    if "/wp-content/uploads/" in _path and _PDF_URL_RE.search(_path):
        return "pdf_list"

    # Known business directory domains
    if _url_domain in _BUSINESS_DIR_DOMAINS:
        return "business_directory"
    for _bd in _BUSINESS_DIR_DOMAINS:
        if _url_domain.endswith("." + _bd):
            return "business_directory"

    # Document hosting
    if _url_domain in _DOCUMENT_HOSTING_DOMAINS:
        return "document_hosting"

    # Product / marketplace path pattern
    if _PRODUCT_PATH_RE.search(_path):
        if not candidate_domain or _url_domain != candidate_domain.lstrip("www."):
            return "marketplace_product_page"

    # Known marketplace / e-commerce aggregators (also caught by text detector
    # but url check is faster and fires even with no page text)
    _MARKETPLACE_DOMAINS = frozenset({
        "gosupps.com", "amazon.it", "amazon.com", "ebay.it", "ebay.com",
        "aliexpress.com", "etsy.com",
    })
    if _url_domain in _MARKETPLACE_DOMAINS:
        return "marketplace_product_page"

    return ""


def _detect_neg_source(text: str, domain: str = "", url: str = "") -> str:
    """
    Return the first matching negative-source-type key, or empty string if none.
    Checks URL first (fast), then domain TLD, then page text patterns.

    Thresholds:
    - news_media: 3+ matches (company websites often have news sections)
    - public_school / university: 1+ match (very distinctive vocabulary)
    - others: 2+ matches

    Also checks domain TLD: .edu.it → public_school, .gov.it etc. → government.
    """
    # URL-based check (highest confidence, no text needed)
    if url:
        _url_neg = _detect_neg_source_url(url, domain)
        if _url_neg:
            return _url_neg

    if not text and not domain:
        return ""

    # Domain-level hard signals (checked before text)
    if domain:
        if _EDU_IT_RE.search(domain.lower()):
            return "public_school"
        if _GOV_IT_RE.search(domain.lower()):
            return "government"

    if not text:
        return ""
    tl = text.lower()

    _thresholds = {
        "news_media": 3,
        "public_school": 1,   # very specific vocabulary, 1 match is enough
        "university": 1,
    }
    for src_type, pat in _NEG_SOURCE_RES.items():
        hits = pat.findall(tl)
        threshold = _thresholds.get(src_type, 2)
        if len(hits) >= threshold:
            return src_type
    return ""


def _extract_fc_evidence(
    text: str,
    domain: str,
    company_name: str,
    city: str,
    province: str,
    input_email_domain: str,
    input_partita_iva: str,
    name_variants: dict,
    source_url: str = "",
) -> dict:
    """
    Extract identity evidence from Firecrawl-scraped page text.
    Returns evidence dict used to compute evidence_strength.
    """
    if not text:
        return {
            "legal_name_match": False, "partita_iva_match": False,
            "city_match": False, "province_match": False,
            "email_domain_match": False, "brand_strong": False,
            "brand_token": False, "official_signal": False,
            "group_italy_mention": False,
            "italian_language": False, "extracted_iva": "",
            "extracted_email": "", "extracted_phone": "",
            "negative_source_type": "",
        }

    tl = text.lower()
    brand_lower   = re.sub(r"[^\w\s]", "", (name_variants.get("brand") or company_name).lower())
    brand_nodot   = re.sub(r"[^\w\s]", "", (name_variants.get("brand_nodot") or brand_lower).lower())
    name_lower    = re.sub(r"[^\w\s]", "", company_name.lower())
    name_stripped = re.sub(r"[^\w\s]", "", strip_legal(company_name).lower())

    # Legal/company name match — require most tokens to appear in the page
    name_toks = [t for t in re.split(r"\s+", name_stripped) if len(t) >= 3]
    legal_name_match = False
    if name_toks:
        matched = sum(1 for t in name_toks if t in tl)
        legal_name_match = (matched >= max(2, len(name_toks) * 0.75))

    # Partita IVA match
    _pi_re = re.compile(r"\b(\d{11})\b")
    extracted_ivas = _pi_re.findall(text)
    partita_iva_match = bool(
        input_partita_iva and any(v == input_partita_iva.strip() for v in extracted_ivas)
    )
    extracted_iva = extracted_ivas[0] if extracted_ivas else ""

    # City / province match
    city_match     = bool(city     and len(city)     >= 3 and city.lower()     in tl)
    province_match = bool(province and len(province) >= 2 and province.lower() in tl)

    # Email domain match (the scraped page contains an email ending with @candidate_domain)
    _em_re = re.compile(r"[\w.\-]+@([\w.\-]+\.[a-z]{2,6})", re.I)
    page_email_domains = [m.lower() for m in _em_re.findall(text)]
    email_domain_match = bool(
        domain and any(ed == domain or ed.endswith("." + domain) for ed in page_email_domains)
    )
    extracted_email = next(
        (f for f in re.findall(r"[\w.\-]+@[\w.\-]+\.[a-z]{2,6}", text, re.I)
         if domain in f.lower()), ""
    ) or (re.findall(r"[\w.\-]+@[\w.\-]+\.[a-z]{2,6}", text, re.I) or [""])[0]

    # Brand token presence
    brand_token  = bool(brand_lower in tl or brand_nodot in tl)
    brand_strong = bool(
        brand_lower and (
            len(brand_lower) >= 5 and brand_lower in tl
            or brand_nodot and len(brand_nodot) >= 5 and brand_nodot in tl
        )
    )

    # Official-site signal
    official_signal = bool(re.search(
        r"\b(sito\s+ufficiale|official\s+(website|site)|benvenuti\s+sul\s+sito|"
        r"chi\s+siamo|about\s+us|la\s+nostra\s+azienda|our\s+company)\b", tl
    ))

    # Group corporate site with Italian operations mention
    group_italy_mention = bool(re.search(
        r"\b(italia|italy|italian|sede\s+italiana|operazioni\s+in\s+italia|"
        r"filiale\s+italiana|stabilimento|plant\s+in\s+italy|cementificio|cemento|cement"
        r"|subsidiary|sussidiar|gruppo\s+\w+|group\s+\w+|holding)\b", tl
    ))

    # Italian language signal
    italian_language = bool(re.search(
        r"\b(azienda|prodotti|servizi|contatti|via\s+[a-z]|piazza|corso\s+[a-z]|"
        r"srl|spa|snc|sas|p\.iva|partita\s+iva)\b", tl
    ))

    # Phone
    _ph_re = re.compile(r"(?:tel\.?|telefono|phone)[:\s]*([\+0][\d\s\-\(\).]{7,20})", re.I)
    ph_m = _ph_re.search(text)
    extracted_phone = ph_m.group(1).strip() if ph_m else ""

    # Negative source type (check URL first, then domain TLD, then text)
    negative_source_type = _detect_neg_source(text, domain, source_url)

    # Wrong entity type: the detected source type is non-commercial and hard
    wrong_entity_type_signal = negative_source_type in _HARD_NEG_SOURCES

    # Wrong location: a different Italian city appears on the page that contradicts the register
    wrong_location_signal = False
    if city and len(city) >= 3 and not city_match and not province_match:
        _it_city_re = re.compile(
            r"\b(milano|roma|torino|napoli|bologna|firenze|venezia|palermo|genova|"
            r"bari|catania|verona|messina|padova|trieste|brescia|taranto|prato|"
            r"modena|perugia|livorno|ravenna|cagliari|foggia|rimini|salerno|ferrara|"
            r"sassari|latina|monza|bergamo|forl[ìi]|trento|vicenza|terni|novara|"
            r"piacenza|ancona|udine|cesena|lecce|pesaro|alessandria|"
            r"casale\s+monferrato)\b", re.I
        )
        if _it_city_re.search(tl) or wrong_entity_type_signal:
            wrong_location_signal = True

    return {
        "legal_name_match":         legal_name_match,
        "partita_iva_match":        partita_iva_match,
        "city_match":               city_match,
        "province_match":           province_match,
        "email_domain_match":       email_domain_match,
        "brand_strong":             brand_strong,
        "brand_token":              brand_token,
        "official_signal":          official_signal,
        "group_italy_mention":      group_italy_mention,
        "italian_language":         italian_language,
        "extracted_iva":            extracted_iva,
        "extracted_email":          extracted_email,
        "extracted_phone":          extracted_phone,
        "negative_source_type":     negative_source_type,
        "wrong_entity_type_signal": wrong_entity_type_signal,
        "wrong_location_signal":    wrong_location_signal,
    }


def _compute_evidence_strength(
    pages_evidence: list[dict],
    negative_source_type: str,
) -> tuple[str, bool, str]:
    """
    Aggregate per-page evidence into an overall evidence_strength and replace_allowed flag.

    Returns (evidence_strength, replace_allowed, reason).

    evidence_strength levels:
      "none"   — nothing meaningful found
      "weak"   — soft signals only (Italian language, brand token, official keyword)
      "medium" — partial identity (city+brand, email domain match, address context)
      "strong" — hard identity (legal name match, VAT match, or city+brand+email together)

    replace_allowed is True ONLY when evidence_strength == "strong" AND no negative_source_type.
    False positives are worse than missing domains — be conservative.
    """
    if negative_source_type:
        return "none", False, f"Blocked by negative source type: {negative_source_type}"

    # Aggregate flags across all pages
    legal_name        = any(e.get("legal_name_match")     for e in pages_evidence)
    iva_match         = any(e.get("partita_iva_match")     for e in pages_evidence)
    city_match        = any(e.get("city_match")            for e in pages_evidence)
    prov_match        = any(e.get("province_match")        for e in pages_evidence)
    email_match       = any(e.get("email_domain_match")    for e in pages_evidence)
    brand_strong      = any(e.get("brand_strong")          for e in pages_evidence)
    brand_token       = any(e.get("brand_token")           for e in pages_evidence)
    official          = any(e.get("official_signal")       for e in pages_evidence)
    group_italy       = any(e.get("group_italy_mention")   for e in pages_evidence)
    italian           = any(e.get("italian_language")      for e in pages_evidence)

    # Group corporate site with Italian operations → strong when brand + location match
    if group_italy and brand_strong and (city_match or prov_match):
        return "strong", True, "Group site with Italian operations + brand + city/province match"

    # Hard evidence combinations → strong
    if legal_name and brand_strong:
        return "strong", True, "Legal name + brand confirmed on site"
    if iva_match:
        return "strong", True, "Partita IVA matched on site"
    if email_match and brand_strong and (city_match or prov_match or official):
        return "strong", True, "Email domain + brand + location/official signal"
    if city_match and prov_match and brand_strong and official:
        return "strong", True, "City + province + brand + official signal"
    if legal_name:
        return "strong", True, "Legal company name matched on site"

    # Partial identity → medium (confirms but does NOT allow replacement alone)
    if email_match and brand_strong:
        return "medium", False, "Email domain + brand (no location corroboration)"
    if (city_match or prov_match) and brand_strong:
        return "medium", False, "City/province + brand (no email or VAT)"
    if email_match and brand_token:
        return "medium", False, "Email domain found + brand token"

    # Soft signals only → weak
    if brand_token or official or italian:
        return "weak", False, "Soft signals only (brand token, Italian language, official keyword)"

    return "none", False, "No meaningful identity evidence found"


def _fc_verify_candidates(
    company_name: str,
    city: str,
    province: str,
    email_domain: str,
    input_partita_iva: str,
    name_variants: dict,
    candidates: list[str],
    current_domain: str,
    fc_key: str | list[str],
    max_pages: int = 3,
    page_timeout: int = 15,
    fc_speed_mode: str = _FC_SPEED_FAST,
    python_confidence: str = "",  # "High"/"Medium"/"Low"/"None" from Python scoring
    progress_update_fn=None,  # optional callback(candidate_i, total_cands, page_i, total_pages, domain)
    # live-counter dict — caller passes {} and reads back keys after the call
    live_counters: dict | None = None,
    fc_location: dict | None = None,
    candidate_hints: dict | None = None,  # domain -> best candidate_url from Serper evidence
) -> tuple[dict, list[dict]]:
    """
    Verify candidate domains via Firecrawl scrape.
    Returns (fc_result_dict, verif_debug_rows).

    max_pages counts ATTEMPTED requests (not only successes) — the per-candidate
    request budget stops after max_pages attempts regardless of outcome.

    Early-stop rules:
    - After homepage: if evidence_strength is "strong" → skip about/contact.
    - After homepage: if negative_source_type is set → skip remaining pages for this candidate.
    - In Balanced mode: skip about/contact when homepage already gives medium+ evidence.
    """
    fc_out = {
        "firecrawl_used":               True,
        "firecrawl_verified_domain":    "",
        "firecrawl_decision":           "uncertain",
        "firecrawl_confidence":         "Low",
        "firecrawl_reason":             "",
        "firecrawl_evidence_strength":  "none",
        "firecrawl_negative_source_type": "",
        "firecrawl_pages_fetched":      0,
        "firecrawl_fetch_status":       "",
        "firecrawl_error":              "",
        "firecrawl_evidence_url":       "",
    }
    debug_rows: list[dict] = []

    if live_counters is None:
        live_counters = {}
    live_counters.setdefault("fc_requests_attempted", 0)
    live_counters.setdefault("fc_pages_successful", 0)
    live_counters.setdefault("fc_timeouts", 0)
    live_counters.setdefault("fc_total_secs", 0.0)

    _fc_keys_norm = fc_key if isinstance(fc_key, list) else ([fc_key] if fc_key else [])
    if not candidates or not _fc_keys_norm:
        fc_out.update(
            firecrawl_used=bool(_fc_keys_norm),
            firecrawl_decision="no_candidates" if not candidates else "no_key",
            firecrawl_reason="No candidates to verify" if not candidates else "No Firecrawl key",
        )
        return fc_out, debug_rows

    # Page slots available per mode
    _ALL_SLOTS = [
        ("homepage",  [""]),
        ("about",     ["/chi-siamo", "/about", "/about-us"]),
        ("contact",   ["/contatti", "/contact", "/contacts"]),
    ]
    if fc_speed_mode == _FC_SPEED_FAST:
        _active_slots = _ALL_SLOTS[:1]   # homepage only
    elif fc_speed_mode == _FC_SPEED_BALANCED:
        _active_slots = _ALL_SLOTS[:2]   # homepage + about (contact dropped; balanced decides below)
    else:
        _active_slots = _ALL_SLOTS       # all three slots

    domain_results: dict[str, dict] = {}  # domain -> aggregated result
    domain_page_texts: dict[str, list] = {}  # domain -> list of scraped texts
    all_statuses: list[str] = []
    total_pages_fetched = 0
    fc_errors: list[str] = []
    _best_evidence_url: str = ""
    import time as _time_mod

    # ── Pre-flight redirect check ─────────────────────────────────────────────
    # For each candidate, run a cheap HTTP redirect check.  If the candidate
    # redirects to a different root domain, insert that as the primary candidate
    # (it becomes the first domain Firecrawl actually scrapes).
    _redirect_info: dict[str, dict] = {}   # original domain -> redirect result
    # Maps domain -> full URL to scrape first (redirect URL or Serper candidate_url)
    # Seeded from candidate_hints (Serper evidence URLs for group/subsidiary pages)
    _domain_hint_url: dict[str, str] = {}
    # Tracks slot type label: "candidate_url_hint" (from Serper) vs "redirect_hint"
    _domain_hint_type: dict[str, str] = {}
    if candidate_hints:
        for _ch_dom, _ch_url in (candidate_hints or {}).items():
            if _ch_dom and _ch_url:
                # Never send shorteners or discovery-blocked domains to Firecrawl
                if is_url_shortener(_ch_dom) or is_discovery_blocked(_ch_dom):
                    continue
                _ch_path = re.sub(r"^https?://[^/]+", "", _ch_url) or ""
                # Only use as a hint when there's a meaningful path (not just root)
                if _ch_path and _ch_path != "/" and len(_ch_path) > 1:
                    _domain_hint_url[_ch_dom] = _ch_url
                    _domain_hint_type[_ch_dom] = "candidate_url_hint"
    _expanded_candidates: list[str] = []
    _suspicious_re = re.compile(r"\b(forum|fan|club|community|archive|directory)\b", re.I)
    for _cdom in candidates:
        if not _cdom:
            continue
        # Skip shorteners and blocked domains entirely — never send to Firecrawl
        if is_url_shortener(_cdom) or is_discovery_blocked(_cdom):
            continue
        _rr = _resolve_redirect(_cdom)
        _redirect_info[_cdom] = _rr
        _redir_final  = _rr.get("redirect_final_domain", "")
        _redir_url    = _rr.get("redirect_final_url", "")
        if _redir_final:
            # Remember the exact redirect URL so Firecrawl scrapes it first
            if _redir_url and _redir_final not in _domain_hint_url:
                _domain_hint_url[_redir_final] = _redir_url
                _domain_hint_type[_redir_final] = "redirect_hint"
            if _redir_final not in _expanded_candidates:
                # Prioritise the redirect destination when original looks suspicious
                if _suspicious_re.search(_cdom):
                    _expanded_candidates.insert(0, _redir_final)
                else:
                    _expanded_candidates.append(_redir_final)
        if _cdom not in _expanded_candidates:
            _expanded_candidates.append(_cdom)
    # Deduplicate preserving order
    _seen_cands: set[str] = set()
    _deduped: list[str] = []
    for _c in _expanded_candidates:
        if _c not in _seen_cands:
            _deduped.append(_c)
            _seen_cands.add(_c)
    candidates = _deduped

    for ci, domain in enumerate(candidates):
        if not domain:
            continue

        pages_evidence: list[dict] = []
        neg_src_domain = ""
        requests_attempted = 0  # budget: counts every FC request attempt

        # Build per-domain slot list: prepend the known hint URL if we have one.
        # Hint may come from a redirect destination (redirect_hint) or directly from
        # the Serper candidate_url stored in evidence (candidate_url_hint).
        _hint_url  = _domain_hint_url.get(domain, "")
        _hint_type = _domain_hint_type.get(domain, "redirect_hint")
        if _hint_url:
            # Extract the path from the hint URL to use as a specific slug
            _hint_path = re.sub(r"^https?://[^/]+", "", _hint_url) or ""
            _dom_slots = [(_hint_type, [_hint_path, ""])] + list(_active_slots)
        else:
            _dom_slots = list(_active_slots)

        for slot_name, slug_candidates in _dom_slots:
            if requests_attempted >= max_pages:
                break

            # Early-stop after homepage/hint slots: strong evidence or confirmed negative source
            if slot_name not in ("homepage", "redirect_hint", "candidate_url_hint") and pages_evidence:
                _interim_str, _, _ = _compute_evidence_strength(pages_evidence, neg_src_domain)
                if _interim_str == "strong":
                    break  # no need to fetch more pages
                if neg_src_domain:
                    break  # negative source confirmed — stop fetching this candidate

                # Balanced: skip about/contact when homepage gives medium evidence
                if fc_speed_mode == _FC_SPEED_BALANCED and _interim_str == "medium":
                    break

            # Try slugs in order for this slot; stop at first success or hard failure
            for slug in slug_candidates:
                if requests_attempted >= max_pages:
                    break

                # For hint slots (redirect or candidate_url), use the full hint URL
                if slot_name in ("redirect_hint", "candidate_url_hint") and _hint_url and slug == _hint_path:
                    url = _hint_url
                else:
                    url = f"https://{domain}{slug}"
                if progress_update_fn:
                    progress_update_fn(ci, len(candidates), requests_attempted, max_pages, domain)

                _t0 = _time_mod.time()
                _exc_text = ""
                try:
                    text, status, meta = _fc_scrape(url, _fc_keys_norm, timeout=page_timeout, fc_location=fc_location)
                except Exception as _scrape_exc:
                    text, status, meta = "", "exception", {}
                    _exc_text = str(_scrape_exc)[:200]
                    fc_errors.append(f"{domain}{slug}:exception:{_exc_text}")
                    live_counters.setdefault("fc_exceptions", 0)
                    live_counters["fc_exceptions"] = live_counters.get("fc_exceptions", 0) + 1
                _elapsed = _time_mod.time() - _t0

                requests_attempted += 1
                live_counters["fc_requests_attempted"] += 1
                live_counters["fc_total_secs"] += _elapsed
                all_statuses.append(status)

                if status == "timeout":
                    live_counters["fc_timeouts"] += 1
                    fc_errors.append(f"{domain}{slug}:timeout")

                # Always emit a debug row for every attempt (even failures)
                _ev_partial: dict = {}
                if status == "ok" and text:
                    total_pages_fetched += 1
                    live_counters["fc_pages_successful"] += 1
                    _ev_partial = _extract_fc_evidence(
                        text, domain, company_name, city, province,
                        email_domain, input_partita_iva, name_variants,
                        source_url=url,
                    )
                    if _ev_partial.get("negative_source_type") and not neg_src_domain:
                        neg_src_domain = _ev_partial["negative_source_type"]
                    pages_evidence.append(_ev_partial)
                    domain_page_texts.setdefault(domain, []).append(text)

                    # Track best evidence URL for firecrawl_evidence_url
                    if _best_evidence_url == "":
                        _interim_str, _, _ = _compute_evidence_strength(pages_evidence, neg_src_domain)
                        if _interim_str in ("medium", "strong"):
                            _best_evidence_url = url

                _redirect_url    = meta.get("final_url", "") or meta.get("redirect_url", "")
                _redirect_domain = meta.get("redirect_domain", "")
                _canonical_domain_used = _redirect_domain or domain

                _url_neg_src = _detect_neg_source_url(url, domain)
                _neg_src_for_row = _ev_partial.get("negative_source_type", "") or _url_neg_src
                debug_rows.append({
                    "company_name":             company_name,
                    "candidate_domain":         domain,
                    "original_candidate_url":   url,
                    "page_url":                 _redirect_url or url,
                    "page_type":                slot_name,
                    "redirect_final_url":       _redirect_url,
                    "redirect_final_domain":    _redirect_domain,
                    "canonical_domain_used":    _canonical_domain_used,
                    "fetch_status":             status,
                    "chars_fetched":            len(text),
                    "elapsed_secs":             round(_elapsed, 2),
                    "source_type":              _neg_src_for_row or ("company" if status == "ok" else ""),
                    "wrong_entity_type_signal": _ev_partial.get("wrong_entity_type_signal", False),
                    "wrong_location_signal":    _ev_partial.get("wrong_location_signal", False),
                    "evidence_strength":        "",   # filled after domain scoring
                    "negative_source_type":     _neg_src_for_row,
                    "third_party_evidence_flag": bool(_url_neg_src or _neg_src_for_row),
                    "verifier_decision":        "",
                    "verifier_reason":          "",
                    "replace_allowed":          "",
                    "firecrawl_error":          _exc_text,
                    "firecrawl_key_index_used":     meta.get("firecrawl_key_index_used", ""),
                    "firecrawl_key_failover_count": meta.get("firecrawl_key_failover_count", ""),
                    "firecrawl_key_statuses":       meta.get("firecrawl_key_statuses", ""),
                })

                if status != "ok":
                    if status == "timeout":
                        break  # don't try more slugs on timeout
                    elif status.startswith("error:") or status == "exception":
                        break  # hard error — don't try more slugs
                    elif status == "404":
                        continue  # try next slug
                    else:
                        continue  # other HTTP error — try next slug
                else:
                    # Success: handle redirect evidence
                    ev = _ev_partial
                    if _redirect_domain and _redirect_domain != domain:
                        _redir_ev = _extract_fc_evidence(
                            text, _redirect_domain, company_name, city, province,
                            email_domain, input_partita_iva, name_variants,
                            source_url=_redirect_url or url,
                        )
                        ev["redirect_domain"]       = _redirect_domain
                        ev["redirect_ev"]           = _redir_ev
                        ev["redirect_final_url"]    = _redirect_url
                    break  # slot satisfied; move to next slot

        # Aggregate wrong entity / wrong location signals across pages
        _any_wrong_entity = any(e.get("wrong_entity_type_signal") for e in pages_evidence)
        _any_wrong_loc    = any(e.get("wrong_location_signal")    for e in pages_evidence)

        # Check for canonical redirect to a different domain that passes evidence
        _redirect_domain_winner = ""
        _redirect_domain_ev: dict = {}
        for _pe in pages_evidence:
            _rd = _pe.get("redirect_domain", "")
            _rev = _pe.get("redirect_ev")
            if _rd and _rev:
                _r_str, _r_repl, _r_rsn = _compute_evidence_strength([_rev], "")
                if _r_str in ("strong", "medium") and not _redirect_domain_winner:
                    _redirect_domain_winner = _rd
                    _redirect_domain_ev = {
                        "evidence_strength": _r_str,
                        "replace_allowed":   _r_repl,
                        "reason":            _r_rsn,
                        "redirect_url":      _pe.get("redirect_final_url", ""),
                    }

        # Score this domain (use the redirected domain evidence if it's stronger)
        strength, replace_ok, reason = _compute_evidence_strength(pages_evidence, neg_src_domain)

        # If original domain is a wrong entity type but redirects cleanly to a strong domain,
        # override the score to allow replacement with the redirect destination
        if _any_wrong_entity and _redirect_domain_winner and _redirect_domain_ev.get("replace_allowed"):
            strength   = _redirect_domain_ev["evidence_strength"]
            replace_ok = True
            reason     = (
                f"Original candidate is {neg_src_domain or 'wrong entity type'}; "
                f"redirects to {_redirect_domain_winner}: "
                f"{_redirect_domain_ev['reason']}"
            )
            neg_src_domain = ""  # clear the blocker so redirect can win

        # Score professional site quality from all scraped text
        _all_page_texts = " ".join(domain_page_texts.get(domain, []))
        _pro_site_info = _score_professional_site(_all_page_texts, domain)

        domain_results[domain] = {
            "pages_evidence":           pages_evidence,
            "evidence_strength":        strength,
            "replace_allowed":          replace_ok,
            "reason":                   reason,
            "negative_source_type":     neg_src_domain,
            "wrong_entity_type_signal": _any_wrong_entity,
            "wrong_location_signal":    _any_wrong_loc,
            "redirect_domain_winner":   _redirect_domain_winner,
            "redirect_domain_ev":       _redirect_domain_ev,
            "professional_site_score":  _pro_site_info["professional_site_score"],
            "professional_site_level":  _pro_site_info["professional_site_level"],
            "professional_site_signals": _pro_site_info["professional_site_signals"],
        }

        # Back-fill debug rows for this domain
        for row in debug_rows:
            if row["candidate_domain"] == domain and row["evidence_strength"] == "":
                row["evidence_strength"]        = strength
                row["wrong_entity_type_signal"] = _any_wrong_entity
                row["wrong_location_signal"]    = _any_wrong_loc
                row["verifier_decision"]        = "replace" if replace_ok and domain != current_domain else "confirm" if replace_ok else "uncertain"
                row["verifier_reason"]          = reason
                row["replace_allowed"]          = replace_ok

    fc_out["firecrawl_pages_fetched"] = total_pages_fetched
    fc_out["firecrawl_fetch_status"]  = "; ".join(dict.fromkeys(all_statuses))[:200]
    fc_out["firecrawl_evidence_url"]  = _best_evidence_url
    # Aggregate key-failover count across all pages fetched for this company
    fc_out["firecrawl_key_failover_count"] = sum(
        int(r.get("firecrawl_key_failover_count") or 0) for r in debug_rows
    )

    # Store redirect info for the original current_domain candidate
    _cur_rr = _redirect_info.get(current_domain, {})
    if _cur_rr.get("redirect_final_domain"):
        fc_out.setdefault("redirect_final_url",    _cur_rr.get("redirect_final_url", ""))
        fc_out.setdefault("redirect_final_domain", _cur_rr.get("redirect_final_domain", ""))
        fc_out.setdefault("original_candidate_domain", current_domain)
    if fc_errors:
        fc_out["firecrawl_error"] = "; ".join(fc_errors[:5])

    if not domain_results:
        fc_out.update(firecrawl_decision="fetch_failed", firecrawl_confidence="None",
                      firecrawl_reason="No pages fetched for any candidate")
        return fc_out, debug_rows

    # Decision: find the best candidate
    # Priority: strong evidence > medium > weak > none
    # Among same strength: prefer current_domain (avoid false replacements)
    _STRENGTH_ORDER = {"strong": 3, "medium": 2, "weak": 1, "none": 0}

    current_res   = domain_results.get(current_domain, {})
    current_str   = _STRENGTH_ORDER.get(current_res.get("evidence_strength", "none"), 0)
    current_neg   = current_res.get("negative_source_type", "")

    # Find best alternative (replace_allowed=True and stronger than current)
    best_alt: str | None = None
    best_alt_str = 0
    for dom, dr in domain_results.items():
        if dom == current_domain:
            continue
        if dr.get("replace_allowed") and _STRENGTH_ORDER.get(dr["evidence_strength"], 0) > best_alt_str:
            best_alt     = dom
            best_alt_str = _STRENGTH_ORDER[dr["evidence_strength"]]

    # Determine final decision
    cur_dr = domain_results.get(current_domain, {})
    _is_high_conf_python = python_confidence.strip().lower() == "high"

    # Check if any candidate domain redirected to a winner domain
    _global_redirect_winner = ""
    _global_redirect_ev: dict = {}
    for _dom, _dr in domain_results.items():
        _rw = _dr.get("redirect_domain_winner", "")
        _rev = _dr.get("redirect_domain_ev", {})
        if _rw and _rev.get("replace_allowed") and not _global_redirect_winner:
            _global_redirect_winner = _rw
            _global_redirect_ev    = _rev
            _global_redirect_from  = _dom

    cur_neg = cur_dr.get("negative_source_type", "")
    _cur_neg_is_hard = cur_neg in _HARD_NEG_SOURCES
    _cur_wrong_entity = cur_dr.get("wrong_entity_type_signal", False)
    _cur_wrong_loc    = cur_dr.get("wrong_location_signal", False)

    # If the current domain redirects to a provably better domain, use that as replacement
    _cur_redirect_winner = cur_dr.get("redirect_domain_winner", "")
    _cur_redirect_ev     = cur_dr.get("redirect_domain_ev", {})

    # Also check global (any candidate redirected to a winner)
    if not _cur_redirect_winner and _global_redirect_winner:
        _cur_redirect_winner = _global_redirect_winner
        _cur_redirect_ev     = _global_redirect_ev

    # ── Redirect canonical domain wins ───────────────────────────────────────
    # If the original candidate is a wrong entity type (e.g. fan forum, school) but
    # cleanly redirects to a strong official domain, replace with the redirect destination.
    if _cur_redirect_winner and _cur_redirect_ev.get("replace_allowed"):
        _redir_url = _cur_redirect_ev.get("redirect_url", "")
        fc_out.update(
            firecrawl_verified_domain=_cur_redirect_winner,
            firecrawl_decision="replace",
            firecrawl_confidence="High",
            firecrawl_reason=(
                f"Candidate {current_domain} redirected to {_cur_redirect_winner}: "
                f"{_cur_redirect_ev.get('reason', 'strong identity evidence on redirect destination')}"
            ),
            firecrawl_evidence_strength=_cur_redirect_ev.get("evidence_strength", "strong"),
            firecrawl_negative_source_type="",
            firecrawl_redirect_final_url=_redir_url,
            firecrawl_redirect_final_domain=_cur_redirect_winner,
        )
        return fc_out, debug_rows

    if cur_neg and not cur_dr.get("replace_allowed"):
        if _is_high_conf_python and not _cur_neg_is_hard:
            # Soft negative source (news_media, foundation, event, association, dealer) on a
            # High-confidence Python domain — do NOT reject or blank the domain.
            # Flag it as uncertain and let the human review if needed.
            fc_out.update(
                firecrawl_verified_domain=current_domain,
                firecrawl_decision="uncertain",
                firecrawl_confidence="Low",
                firecrawl_reason=(
                    f"Possible {cur_neg} signal on page — but Python confidence is High; "
                    "keeping domain; manual review recommended"
                ),
                firecrawl_evidence_strength=cur_dr.get("evidence_strength", "weak"),
                firecrawl_negative_source_type=cur_neg,
            )
        elif best_alt:
            # Current domain is a hard negative source (or non-High confidence) — replace
            alt_dr = domain_results[best_alt]
            fc_out.update(
                firecrawl_verified_domain=best_alt,
                firecrawl_decision="replace",
                firecrawl_confidence="High" if best_alt_str >= 3 else "Medium",
                firecrawl_reason=f"Current domain is {cur_neg}; {best_alt}: {alt_dr['reason']}",
                firecrawl_evidence_strength=alt_dr["evidence_strength"],
                firecrawl_negative_source_type="",
            )
        else:
            if _is_high_conf_python:
                # No strong alternative and High Python confidence — flag, don't blank
                fc_out.update(
                    firecrawl_verified_domain=current_domain,
                    firecrawl_decision="uncertain",
                    firecrawl_confidence="Low",
                    firecrawl_reason=(
                        f"Current domain shows {cur_neg} signals; no verified alternative found; "
                        "Python confidence High — keeping domain"
                    ),
                    firecrawl_evidence_strength="weak",
                    firecrawl_negative_source_type=cur_neg,
                )
            else:
                fc_out.update(
                    firecrawl_verified_domain="",
                    firecrawl_decision="reject",
                    firecrawl_confidence="None",
                    firecrawl_reason=f"Current domain is {cur_neg}; no strong alternative found",
                    firecrawl_evidence_strength="none",
                    firecrawl_negative_source_type=cur_neg,
                )

    elif best_alt and best_alt_str > current_str and best_alt_str >= 3:
        # A different domain has strong evidence and current domain is weaker.
        # For High-confidence Python rows only replace if the alternative is very strong (IVA match
        # or legal name match) — otherwise just flag uncertain.
        alt_dr = domain_results[best_alt]
        if _is_high_conf_python and alt_dr.get("evidence_strength") != "strong":
            fc_out.update(
                firecrawl_verified_domain=current_domain,
                firecrawl_decision="uncertain",
                firecrawl_confidence="Low",
                firecrawl_reason=(
                    f"Alternative {best_alt} has stronger evidence but Python confidence is High — "
                    "not replacing without hard identity proof"
                ),
                firecrawl_evidence_strength=cur_dr.get("evidence_strength", "none"),
                firecrawl_negative_source_type=cur_neg,
            )
        else:
            fc_out.update(
                firecrawl_verified_domain=best_alt,
                firecrawl_decision="replace",
                firecrawl_confidence="High",
                firecrawl_reason=f"Stronger evidence on {best_alt}: {alt_dr['reason']}",
                firecrawl_evidence_strength=alt_dr["evidence_strength"],
                firecrawl_negative_source_type="",
            )

    elif cur_dr.get("replace_allowed") or cur_dr.get("evidence_strength", "none") in ("strong", "medium"):
        # Current domain confirmed (or soft-confirmed)
        strength = cur_dr.get("evidence_strength", "none")
        conf = {"strong": "High", "medium": "Medium", "weak": "Low", "none": "Low"}.get(strength, "Low")
        fc_out.update(
            firecrawl_verified_domain=current_domain,
            firecrawl_decision="confirm",
            firecrawl_confidence=conf,
            firecrawl_reason=cur_dr.get("reason", "Firecrawl confirms Python selection"),
            firecrawl_evidence_strength=strength,
            firecrawl_negative_source_type=cur_neg,
        )

    elif cur_dr.get("evidence_strength", "none") == "weak":
        fc_out.update(
            firecrawl_verified_domain=current_domain,
            firecrawl_decision="uncertain",
            firecrawl_confidence="Low",
            firecrawl_reason="Only weak/soft evidence found — keeping Python selection but flagging",
            firecrawl_evidence_strength="weak",
            firecrawl_negative_source_type=cur_neg,
        )

    else:
        fc_out.update(
            firecrawl_verified_domain=current_domain,
            firecrawl_decision="uncertain",
            firecrawl_confidence="Low",
            firecrawl_reason="Insufficient identity evidence across all candidates",
            firecrawl_evidence_strength="none",
        )

    # Add professional site scoring for the winning domain
    _winning_dom = fc_out.get("firecrawl_verified_domain") or current_domain
    _win_dr = domain_results.get(_winning_dom, {})
    fc_out["professional_site_score"]   = _win_dr.get("professional_site_score", 0)
    fc_out["professional_site_level"]   = _win_dr.get("professional_site_level", "none")
    fc_out["professional_site_signals"] = _win_dr.get("professional_site_signals", "")

    return fc_out, debug_rows


# ---------------------------------------------------------------------------
# Unified verifier layer
# ---------------------------------------------------------------------------


def _should_verify(
    result: dict,
    name_variants: dict,
    raw_ev: list[dict],
    verifier_mode: str,
    debug_mode: bool = False,
) -> tuple[bool, str]:
    """
    Decide whether to run the website verifier for this row.
    Returns (should_run, trigger_reason).

    Pre-skip rule: High-confidence rows with strong brand-domain overlap and
    clear score separation are skipped even if minor soft triggers are present.
    This prevents wasting Firecrawl credits on obvious correct matches.
    """
    if verifier_mode == _VM_ALL_DEBUG and debug_mode:
        return True, "all_debug_mode"

    conf      = str(result.get("final_confidence") or result.get("domain_confidence") or "").strip().lower()
    manual    = str(result.get("manual_review_needed", "")).lower() in ("true", "1", "yes")
    final_dom = str(result.get("final_selected_domain") or result.get("validated_domain") or "")

    # ── Collect top-2 score delta ─────────────────────────────────────────────
    scored: dict[str, float] = {}
    for e in raw_ev:
        if not e.get("used"):
            continue
        dom = e.get("domain", "")
        try:
            sc = float(e.get("score", 0))
        except (TypeError, ValueError):
            sc = 0.0
        if dom and sc > scored.get(dom, -1.0):
            scored[dom] = sc
    sorted_scores = sorted(scored.values(), reverse=True)
    score_delta = (sorted_scores[0] - sorted_scores[1]) if len(sorted_scores) >= 2 else 1.0

    # ── Brand-domain overlap for selected domain ──────────────────────────────
    bov = brand_overlap_variants(name_variants, final_dom) if final_dom else 0.0

    brand       = (name_variants.get("brand") or "").strip()
    brand_clean = re.sub(r"[^\w]", "", brand.lower())

    # ── Pre-skip: High confidence + strong brand-domain match ────────────────
    # If all these hold, verification adds no value and wastes credits.
    _is_high_conf   = conf == "high"
    _domain_ok      = bool(final_dom) and not is_generic(final_dom)
    _strong_overlap = bov >= 0.60
    _clear_winner   = score_delta >= 0.25
    _not_famous     = brand_clean not in _JINA_FAMOUS_BRANDS
    _not_risky      = not _domain_has_risky_marker(final_dom)[0] if final_dom else True

    if _is_high_conf and _domain_ok and _strong_overlap and _clear_winner and _not_famous and _not_risky:
        return False, f"skipped: high confidence + strong brand-domain overlap ({round(bov, 2)}) + clear score winner"

    # ── Positive triggers ─────────────────────────────────────────────────────
    reasons: list[str] = []

    if conf in ("medium", "low", "none", ""):
        reasons.append(f"confidence={conf or 'empty'}")
    if manual:
        reasons.append("manual_review_needed")
    if len(sorted_scores) >= 2 and score_delta < 0.25:
        reasons.append("close_scores")

    # site:.it query used
    if str(result.get("search_query_used", "") or "").lower().startswith("site:.it"):
        reasons.append("site_it_query")

    # No location or email evidence — only trigger for non-High confidence rows
    # or when brand-domain overlap is weak (< 0.40).
    if not _is_high_conf or bov < 0.40:
        sel = final_dom.lower()
        if sel and result.get("domain_source") not in (SRC_ORIGINAL, SRC_EMAIL, SRC_SERPER_EMAIL):
            has_loc   = any(e.get("domain", "").lower() == sel and e.get("location_match") for e in raw_ev)
            has_email = any(e.get("domain", "").lower() == sel and e.get("email_match")    for e in raw_ev)
            if not has_loc and not has_email:
                reasons.append("no_location_or_email_evidence")

    # Ambiguous/famous brand — only trigger when confidence is not already High
    if not _is_high_conf:
        if brand_clean and (len(brand_clean) <= 5 or brand_clean in _JINA_FAMOUS_BRANDS or _brand_is_ambiguous(brand)):
            reasons.append("ambiguous_brand")
    else:
        # For High-confidence rows, only flag famous brands as a trigger
        if brand_clean in _JINA_FAMOUS_BRANDS:
            reasons.append("famous_brand")

    # Risky domain pattern — always a trigger
    if final_dom and _domain_has_risky_marker(final_dom)[0]:
        reasons.append("risky_domain_pattern")

    should_run = bool(reasons)
    return should_run, "; ".join(reasons) if reasons else ""


def _root_domain(url_or_domain: str) -> str:
    """Extract bare root domain from a URL or domain string (strips scheme, path, www.)."""
    s = (url_or_domain or "").strip().lower()
    s = re.sub(r"^https?://", "", s)
    s = re.sub(r"/.*$", "", s)   # strip path
    s = re.sub(r"^www\.", "", s)
    return s


def _apply_final_safety_guard(result: dict) -> dict:
    """
    Final safety pass applied after Python + Haiku + verifier decisions are merged.
    Detects and corrects unsafe combinations where the final result would be
    High-confidence / no-review despite weak or absent verification evidence.

    Returns a dict of fields to update (empty if nothing needs correcting).
    False positives are worse than missing domains.
    """
    updates: dict = {}

    conf       = str(result.get("final_confidence", "") or "").strip().lower()
    manual     = str(result.get("manual_review_needed", "")).lower() in ("true", "1", "yes")
    dec_src    = str(result.get("final_decision_source", "") or "")
    haiku_dec  = str(result.get("haiku_decision", "") or "")

    verif_used   = str(result.get("verifier_used", "")).lower() in ("true", "1")
    verif_dec    = str(result.get("verifier_decision", "") or "")
    verif_ev_str = str(result.get("verifier_evidence_strength", "") or "").lower()

    fc_used   = str(result.get("firecrawl_used", "")).lower() in ("true", "1")
    fc_dec    = str(result.get("firecrawl_decision", "") or "")
    fc_ev_str = str(result.get("firecrawl_evidence_strength", "") or "").lower()
    fc_status = str(result.get("firecrawl_fetch_status", "") or "")

    wrong_loc    = str(result.get("wrong_location_signal", "")).lower() in ("true", "1")
    wrong_entity = str(result.get("wrong_entity_type_signal", "")).lower() in ("true", "1")
    had_timeout  = "timeout" in fc_status

    _verif_inconclusive = verif_dec in ("uncertain", "fetch_failed", "no_candidates", "")
    _verif_weak_ev      = verif_ev_str in ("none", "weak", "")
    _fc_uncertain       = fc_dec in ("uncertain", "")
    _fc_weak_ev         = fc_ev_str in ("none", "weak", "")
    _verif_confirmed    = verif_dec == "confirm" and verif_ev_str in ("medium", "strong")
    _verif_replaced     = verif_dec == "replace" and result.get("verifier_replace_allowed")

    # Pre-save snapshot of current values (populated only if we change something)
    def _snapshot():
        return {
            "pre_safety_final_confidence":       result.get("final_confidence", ""),
            "pre_safety_manual_review_needed":   result.get("manual_review_needed", ""),
            "pre_safety_final_decision_source":  result.get("final_decision_source", ""),
        }

    # ── Rule 0: blank domain — ALWAYS the first check ────────────────────────
    final_dom = str(result.get("final_selected_domain", "") or "").strip()
    if not final_dom:
        if conf in ("high", "medium") or not manual:
            updates = _snapshot()
            updates["final_confidence"]      = "None"
            updates["manual_review_needed"]  = True
            updates["final_decision_source"] = "blank_domain_safety_guard"
            updates["safety_guard_applied"]  = True
            updates["safety_guard_reason"]   = "final_selected_domain_blank"
            return updates

    # ── Rule 6 (evidence URL check) — before verifier-confirm early return ───
    _ev_url = str(result.get("verifier_evidence_url", "") or
                  result.get("firecrawl_evidence_url", "") or "")
    if _ev_url and verif_dec == "confirm" and verif_ev_str in ("medium", "strong"):
        _ev_neg = _detect_neg_source_url(_ev_url, str(result.get("final_selected_domain", "") or ""))
        if _ev_neg in _HARD_NEG_SOURCES:
            updates = _snapshot()
            updates["safety_guard_applied"] = True
            if haiku_dec == "reject":
                updates["final_selected_domain"] = ""
                updates["final_confidence"]      = "None"
                updates["manual_review_needed"]  = True
                updates["final_decision_source"] = "haiku_reject_third_party_verifier_guard"
                updates["safety_guard_reason"]   = "haiku_reject_and_verifier_source_not_official"
            else:
                updates["final_confidence"]      = "Low"
                updates["manual_review_needed"]  = True
                updates["final_decision_source"] = "third_party_evidence_safety_guard"
                updates["safety_guard_reason"]   = f"verifier_evidence_url_hard_negative:{_ev_neg}"
            return updates

    # Rule 5 / 6: verifier confirmed or replaced with good evidence — do not downgrade
    if _verif_confirmed or (_verif_replaced and not wrong_loc and not wrong_entity):
        # Only allow no-review for strong confirmation
        if _verif_confirmed and verif_ev_str == "strong":
            return {}
        # Medium confirmation: keep manual_review as-is, but ensure not High without review
        if _verif_confirmed and verif_ev_str == "medium" and conf == "high" and not manual:
            updates = _snapshot()
            updates["manual_review_needed"] = True
            updates["safety_guard_applied"] = True
            updates["safety_guard_reason"] = "verifier_medium_evidence_requires_manual_review"
            updates["final_decision_source"] = dec_src or "verifier_confirm_medium_guard"
        return updates

    # Rule 3: wrong location or entity type signal — always flag
    if wrong_loc or wrong_entity:
        if conf == "high" or not manual:
            updates = _snapshot()
            updates["manual_review_needed"] = True
            if conf == "high":
                updates["final_confidence"] = "Low"
            updates["safety_guard_applied"] = True
            updates["safety_guard_reason"] = "wrong_location_or_entity_signal"
            updates["final_decision_source"] = "wrong_location_or_entity_safety_guard"
        return updates

    # Rule 4: Haiku rejected but verifier did not confirm/replace
    if haiku_dec == "reject" and not _verif_confirmed and not _verif_replaced:
        py_conf = str(result.get("domain_confidence", "") or "").lower()
        _py_weak = py_conf in ("low", "none", "")
        if conf in ("high", "medium") or not manual:
            updates = _snapshot()
            updates["manual_review_needed"] = True
            updates["safety_guard_applied"] = True
            updates["safety_guard_reason"] = "haiku_reject_without_verifier_confirmation"
            updates["final_decision_source"] = "haiku_reject_safety_guard"
            if _py_weak:
                updates["final_selected_domain"] = ""
                updates["final_confidence"] = "None"
            else:
                updates["final_confidence"] = "Low"
        return updates

    # Rule 1: verifier ran and returned inconclusive with weak evidence
    if verif_used and _verif_inconclusive and _verif_weak_ev:
        if conf == "high" or not manual:
            updates = _snapshot()
            updates["manual_review_needed"] = True
            updates["safety_guard_applied"] = True
            if conf == "high":
                if had_timeout or wrong_loc or wrong_entity:
                    updates["final_confidence"] = "Low"
                else:
                    updates["final_confidence"] = "Medium"
            updates["safety_guard_reason"] = "verifier_uncertain_with_weak_evidence"
            updates["final_decision_source"] = "verifier_uncertain_safety_guard"
        return updates

    # Rule 2: Firecrawl ran and returned uncertain with weak evidence
    if fc_used and _fc_uncertain and _fc_weak_ev:
        if conf == "high" or not manual:
            updates = _snapshot()
            updates["manual_review_needed"] = True
            updates["safety_guard_applied"] = True
            if conf == "high":
                if had_timeout or wrong_loc or wrong_entity:
                    updates["final_confidence"] = "Low"
                else:
                    updates["final_confidence"] = "Medium"
            updates["safety_guard_reason"] = "firecrawl_uncertain_with_weak_evidence"
            updates["final_decision_source"] = "verifier_uncertain_safety_guard"
        return updates

    return updates


def _apply_verifier_decision(result: dict, verif_res: dict) -> dict:
    """
    Apply the unified verifier result on top of final_* fields.
    Replacement is only allowed when verifier_replace_allowed is True.

    Protection rule: for High-confidence Python rows a verifier "reject" decision
    is downgraded to "uncertain" — the domain is kept and manual_review_needed is
    set True, but final_selected_domain is never blanked.
    Returns a dict of fields to merge into result.
    """
    if not verif_res.get("verifier_used"):
        return {}

    decision    = verif_res.get("verifier_decision", "")
    sel_domain  = str(verif_res.get("verifier_selected_domain", "") or "")
    replace_ok  = verif_res.get("verifier_replace_allowed", False)
    confidence  = verif_res.get("verifier_confidence", "")
    cur_conf_py = (result.get("final_confidence") or "").strip().lower()
    _py_high    = cur_conf_py == "high"

    # Shared context
    _fc_status   = str(verif_res.get("firecrawl_fetch_status", "") or "")
    _had_timeout = "timeout" in _fc_status
    _verify_rsn  = str(result.get("verification_reason", "") or "")
    _final_dom   = str(result.get("final_selected_domain") or result.get("validated_domain") or "")
    _redir_dom   = str(verif_res.get("redirect_final_domain", "") or result.get("redirect_final_domain", "") or "")
    _redir_url   = str(verif_res.get("redirect_final_url", "")   or result.get("redirect_final_url", "")   or "")
    _brand_clean = re.sub(r"[^\w]", "", (result.get("cleaned_company_name") or "").lower().split()[0] if result.get("cleaned_company_name") else "")
    _suspicious_orig, _risky_reason = _domain_has_risky_marker(_final_dom or "")
    _is_famous   = _brand_clean in _JINA_FAMOUS_BRANDS
    _is_high_risk = (
        "famous_brand"    in _verify_rsn
        or "close_scores" in _verify_rsn
        or "risky_domain" in _verify_rsn
        or _suspicious_orig
        or _is_famous
    )

    # ── Fix 1: Canonical redirect → use redirect domain even on timeout ───────
    # When original candidate is suspicious AND a clean redirect domain exists
    # with strong brand overlap, use the redirect domain as final even if
    # Firecrawl could not confirm (timeout).  Mark manual_review_needed = True.
    if (
        _redir_dom
        and _suspicious_orig
        and decision not in ("confirm", "replace")
        and not is_generic(_redir_dom)
        and not _EDU_IT_RE.search(_redir_dom)
        and not _GOV_IT_RE.search(_redir_dom)
    ):
        # Brand-domain overlap check on redirect domain
        _name_variants_check = result.get("_name_variants", {})
        _redir_brand_ov = brand_overlap_variants(_name_variants_check, _redir_dom) if _name_variants_check else 0.0
        # Fallback: check if brand token appears in redirect domain
        if not _name_variants_check and _brand_clean and len(_brand_clean) >= 4:
            _redir_brand_ov = 1.0 if _brand_clean in _redir_dom.replace("-", "").replace(".", "") else 0.0
        if _redir_brand_ov >= 0.5 or (_brand_clean and len(_brand_clean) >= 4 and _brand_clean in _redir_dom):
            _cv_status = "firecrawl_timeout" if _had_timeout else "unverified"
            return {
                "final_selected_domain":          _root_domain(_redir_dom),
                "final_decision_source":          "redirect_canonical_unverified",
                "final_confidence":               "Medium",
                "manual_review_needed":           True,
                "verifier_reason":                (
                    f"Suspicious candidate ({_final_dom}) redirected to clean canonical domain "
                    f"({_redir_dom}), but Firecrawl did not fully confirm "
                    f"({'timeout' if _had_timeout else 'uncertain'}). Manual review required."
                ),
                "canonical_domain_verification_status":  _cv_status,
                "redirect_checked":                      True,
                "redirect_final_domain":                 _redir_dom,
                "redirect_final_url":                    _redir_url,
                "original_candidate_domain":             _final_dom,
                "original_domain_risky_marker":          True,
                "risky_marker_reason":                   _risky_reason,
                "redirect_canonical_override_applied":   True,
            }

    # ── Fix 2: High-risk timeout that is NOT a suspicious-redirect case ───────
    # Downgrade confidence and force manual review but keep the current domain.
    if _is_high_risk and _had_timeout and decision not in ("confirm", "replace"):
        _timeout_note = (
            "Firecrawl timeout on high-risk row; selected domain may be a redirect/forum. "
            "Manual review required."
        )
        _conf_downgrade = "Low" if _is_famous else "Medium"
        return {
            "final_decision_source": "verifier_timeout_high_risk",
            "final_confidence":      _conf_downgrade,
            "manual_review_needed":  True,
            "verifier_reason":       _timeout_note,
            "firecrawl_reason":      _timeout_note,
            "canonical_domain_verification_status": "firecrawl_timeout",
        }

    if decision == "replace" and replace_ok and sel_domain:
        return {
            "final_selected_domain": _root_domain(sel_domain) or sel_domain,
            "final_decision_source": "verifier_replace",
            "final_confidence":      confidence or "Medium",
            "manual_review_needed":  False,
            "verifier_evidence_url": verif_res.get("verifier_evidence_url", ""),
            "canonical_domain_verification_status": "verified",
        }
    if decision == "reject":
        if _py_high:
            return {
                "final_decision_source": "verifier_flag_high_conf",
                "manual_review_needed":  True,
                "canonical_domain_verification_status": "rejected_high_conf_protected",
            }
        return {
            "final_selected_domain": "",
            "final_decision_source": "verifier_reject",
            "final_confidence":      "None",
            "manual_review_needed":  True,
            "canonical_domain_verification_status": "rejected",
        }
    if decision == "confirm":
        # If the evidence came from a deep subsidiary page, mark as group_subsidiary_page
        _ev_url   = str(verif_res.get("firecrawl_evidence_url", "") or verif_res.get("verifier_evidence_url", "") or "")
        _ev_path  = re.sub(r"^https?://[^/]+", "", _ev_url) if _ev_url else ""
        _ev_deep  = bool(_ev_path and _ev_path.strip("/") and len(_ev_path.strip("/").split("/")) >= 2)
        _dec_src  = "group_subsidiary_page" if _ev_deep else "verifier_confirm"
        updates: dict = {
            "final_decision_source": _dec_src,
            "canonical_domain_verification_status": "confirmed",
            "verifier_evidence_url": _ev_url,
        }
        if confidence == "High" and cur_conf_py in ("medium", "low", "none", ""):
            updates["final_confidence"] = "High"
        return updates
    # uncertain / fetch_failed / no_candidates
    _updates: dict = {
        "final_decision_source": f"verifier_{decision or 'uncertain'}",
        "canonical_domain_verification_status": f"verifier_{decision or 'uncertain'}",
    }
    if _had_timeout:
        _updates["verifier_reason"] = (
            str(verif_res.get("verifier_reason", "") or "")
            or "Firecrawl timed out; could not verify selected domain."
        )
        # Fix 3: downgrade confidence on timeout even for non-famous brands
        if _is_high_risk and cur_conf_py == "high":
            _updates["final_confidence"] = "Medium"
            _updates["manual_review_needed"] = True

        # ── Part 8: Haiku timeout fallback ────────────────────────────────────
        # When Firecrawl timed out on an exact candidate_url with a meaningful path,
        # and Serper evidence is medium or strong, accept the candidate domain.
        _cand_url_hint = str(result.get("candidate_url", "") or "")
        _cand_type_r   = str(result.get("candidate_type", "") or "")
        _haiku_dec_r   = str(result.get("haiku_decision", "") or "")
        _haiku_cand_u  = str(result.get("haiku_candidate_url", "") or "")
        try:
            _hint_path = urlparse(_cand_url_hint).path if _cand_url_hint else ""
        except Exception:
            _hint_path = ""
        _hint_meaningful = bool(_hint_path and len(_hint_path.strip("/").split("/")) >= 2)
        _eligible_type = _cand_type_r in ("group_site_subsidiary_page", "deep_page")
        if (
            _hint_meaningful
            and _eligible_type
            and decision in ("uncertain", "fetch_failed", "")
            and _final_dom
        ):
            # Compute Serper identity strength using all evidence for this domain
            _company_nm  = str(result.get("cleaned_company_name", "") or "")
            _city_r      = str(result.get("city", "") or "")
            _prov_r      = str(result.get("province", "") or "")
            # Find the best evidence row for the candidate URL
            _best_hint_ev: dict = {}
            for _chk_e in (result.get("_raw_ev_ref") or []):
                if _chk_e.get("candidate_url", "") == _cand_url_hint or _chk_e.get("url", "") == _cand_url_hint:
                    _best_hint_ev = _chk_e
                    break
            # Fallback: build a synthetic evidence row from result fields
            if not _best_hint_ev:
                _best_hint_ev = {
                    "title":   result.get("serper_top_result_title", ""),
                    "snippet": "",
                    "url":     _cand_url_hint,
                }
            _id_ev = _serper_exact_page_identity_evidence(
                _company_nm, _city_r, _prov_r, _best_hint_ev
            )
            _id_strength = _id_ev.get("serper_identity_strength", "none")

            if _id_strength in ("medium", "strong"):
                # Haiku confirmed this candidate if decision was accept/needs_firecrawl
                _haiku_confirmed = _haiku_dec_r in ("accept", "needs_firecrawl")
                _timeout_src = (
                    "haiku_timeout_fallback"
                    if _haiku_confirmed
                    else "serper_exact_page_timeout_fallback"
                )
                _ev_url_fb = _haiku_cand_u or _cand_url_hint
                _updates.update({
                    "final_selected_domain":          _root_domain(_final_dom) or _final_dom,
                    "final_decision_source":          _timeout_src,
                    "final_confidence":               "Medium",
                    "manual_review_needed":           _id_strength != "strong",
                    "verifier_evidence_url":          _ev_url_fb,
                    "firecrawl_evidence_url":         _ev_url_fb,
                    "canonical_domain_verification_status": "firecrawl_timeout_serper_or_haiku_confirmed",
                })

    return _updates


def _run_website_verifier(
    company_name: str,
    city: str,
    province: str,
    email_domain: str,
    input_partita_iva: str,
    name_variants: dict,
    candidates: list[str],
    current_domain: str,
    verifier_provider: str,
    jina_api_key: str | None,
    fc_key: str | list[str] | None,
    max_cands: int = 3,
    max_pages: int = 3,
    page_timeout: int = 15,
    fc_speed_mode: str = _FC_SPEED_FAST,
    python_confidence: str = "",
    live_counters: dict | None = None,
    progress_update_fn=None,
    fc_location: dict | None = None,
    candidate_hints: dict | None = None,  # domain -> best candidate_url from Serper
) -> tuple[dict, list[dict]]:
    """
    Route verification to Firecrawl, Jina, or both based on verifier_provider.
    Returns (unified_verifier_result_dict, verif_debug_rows).

    The returned dict contains both provider-specific columns (firecrawl_*, jina_verifier_*)
    and unified verifier_* columns.
    """
    cands = [c for c in candidates[:max_cands] if c]
    if not cands:
        cands = [current_domain] if current_domain else []

    _verif_defaults = {
        "verifier_provider_used":       verifier_provider,
        "verifier_used":                False,
        "verifier_decision":            "skipped",
        "verifier_confidence":          "",
        "verifier_selected_domain":     current_domain,
        "verifier_replace_allowed":     False,
        "verifier_reason":              "",
        "verifier_evidence_strength":   "none",
        "verifier_negative_source_type": "",
        "verifier_pages_fetched":       0,
        "verifier_fetch_status":        "",
        "verifier_error":               "",
        # Provider-specific defaults
        "firecrawl_used":               False,
        "firecrawl_verified_domain":    "",
        "firecrawl_decision":           "",
        "firecrawl_confidence":         "",
        "firecrawl_reason":             "",
        "firecrawl_evidence_strength":  "",
        "firecrawl_negative_source_type": "",
        "firecrawl_pages_fetched":      0,
        "firecrawl_fetch_status":       "",
        "firecrawl_error":              "",
        "firecrawl_evidence_url":       "",
        "verifier_evidence_url":        "",
    }

    all_debug: list[dict] = []
    fc_res: dict = {}
    jina_res: dict = {}
    _provider_ran = False  # tracks whether any provider actually attempted requests

    # ── Firecrawl ────────────────────────────────────────────────────────────
    _fc_key_norm = (
        fc_key if isinstance(fc_key, list)
        else ([fc_key] if fc_key else [])
    )
    if verifier_provider in (_VP_FIRECRAWL, _VP_FC_JINA):
        if not _fc_key_norm:
            # Key missing — do not count as verified
            _verif_defaults.update(
                firecrawl_used=False,
                firecrawl_decision="skipped_no_firecrawl_key",
                verifier_decision="skipped_no_firecrawl_key",
                verifier_error="Firecrawl selected but no API key was provided.",
                firecrawl_error="No Firecrawl API key.",
            )
        else:
            try:
                fc_res, fc_debug = _fc_verify_candidates(
                    company_name, city, province, email_domain, input_partita_iva,
                    name_variants, cands, current_domain, _fc_key_norm,
                    max_pages=max_pages, page_timeout=page_timeout,
                    fc_speed_mode=fc_speed_mode,
                    python_confidence=python_confidence,
                    live_counters=live_counters,
                    progress_update_fn=progress_update_fn,
                    fc_location=fc_location,
                    candidate_hints=candidate_hints,
                )
                _verif_defaults.update(fc_res)
                all_debug.extend(fc_debug)
                _provider_ran = True
            except Exception as exc:
                _err = str(exc)[:200]
                _verif_defaults["firecrawl_error"] = _err
                _verif_defaults["verifier_error"] = _err
                _provider_ran = True  # attempted but errored

    # ── Jina (standalone or fallback) ────────────────────────────────────────
    _jina_needed = (
        verifier_provider == _VP_JINA
        or (verifier_provider == _VP_FC_JINA and
            _verif_defaults.get("firecrawl_decision") in ("uncertain", "fetch_failed", ""))
    )
    if _jina_needed and jina_api_key:
        try:
            jina_res, jina_debug = _jina_verify_candidates(
                company_name, city, province, email_domain,
                name_variants, cands, current_domain, jina_api_key,
            )
            _verif_defaults.update({
                k: v for k, v in jina_res.items()
                if k.startswith("jina_")
            })
            all_debug.extend(jina_debug)
            _provider_ran = True
        except Exception as exc:
            _verif_defaults["jina_fetch_status"] = f"jina_exception:{str(exc)[:120]}"
            _provider_ran = True

    # ── Build unified verifier_* fields from the winning provider ────────────
    # verifier_used=True only when a provider actually ran (not just "selected")
    _verif_defaults["verifier_used"] = _provider_ran

    # Prefer Firecrawl result if available and actionable
    if fc_res.get("firecrawl_decision") in ("confirm", "replace", "reject"):
        src = "firecrawl"
        _verif_defaults.update(
            verifier_decision=fc_res["firecrawl_decision"],
            verifier_confidence=fc_res.get("firecrawl_confidence", ""),
            verifier_selected_domain=fc_res.get("firecrawl_verified_domain", current_domain) or current_domain,
            verifier_replace_allowed=(fc_res["firecrawl_decision"] == "replace"),
            verifier_reason=fc_res.get("firecrawl_reason", ""),
            verifier_evidence_strength=fc_res.get("firecrawl_evidence_strength", "none"),
            verifier_negative_source_type=fc_res.get("firecrawl_negative_source_type", ""),
            verifier_pages_fetched=fc_res.get("firecrawl_pages_fetched", 0),
            verifier_fetch_status=fc_res.get("firecrawl_fetch_status", ""),
            verifier_error=fc_res.get("firecrawl_error", ""),
            verifier_evidence_url=fc_res.get("firecrawl_evidence_url", ""),
        )
    elif jina_res.get("jina_verifier_decision") in ("confirm", "replace", "reject"):
        src = "jina"
        jd = jina_res.get("jina_verifier_decision", "")
        jv = jina_res.get("jina_verified_domain", current_domain) or current_domain
        jr = (jd == "replace") and bool(jv)
        _verif_defaults.update(
            verifier_decision=jd,
            verifier_confidence=jina_res.get("jina_verifier_confidence", ""),
            verifier_selected_domain=jv,
            verifier_replace_allowed=jr,
            verifier_reason=jina_res.get("jina_verifier_reason", ""),
            verifier_evidence_strength="weak",
            verifier_pages_fetched=jina_res.get("jina_pages_fetched", 0),
            verifier_fetch_status=jina_res.get("jina_fetch_status", ""),
        )
    else:
        _verif_defaults.update(
            verifier_decision="uncertain",
            verifier_confidence="Low",
            verifier_selected_domain=current_domain,
            verifier_replace_allowed=False,
            verifier_reason="No provider returned actionable evidence",
        )

    return _verif_defaults, all_debug


# =============================================================================
# AUTOSAVE / RESUME
# =============================================================================

_MODE_SAFE: dict[str, str] = {
    _HAIKU_MODE_PYTHON:    "pythononly",
    _HAIKU_MODE_UNCERTAIN: "haikuuncertain",
    _HAIKU_MODE_ALL:       "haikuall",
}


def _make_run_label(
    mode: str, batch_n: int, max_queries: int, debug: bool,
    ts: str | None = None,
) -> str:
    """Return a human-readable run label: YYYYMMDD_HHMM_mode_Nrows_Qq_debug."""
    if ts is None:
        ts = pd.Timestamp.now().strftime("%Y%m%d_%H%M")
    mode_safe = _MODE_SAFE.get(mode, "pythononly")
    debug_str = "debug" if debug else "nodebug"
    return f"{ts}_{mode_safe}_{batch_n}rows_{max_queries}q_{debug_str}"


def _make_filename(run_label: str, run_id: str) -> str:
    """Return a readable Excel output filename."""
    return f"register_cleaned_{run_label}_{run_id[:8]}.xlsx"


def _file_hash(data: bytes) -> str:
    """Return a short, stable identifier for a file's byte content."""
    return hashlib.sha1(data).hexdigest()[:16]


def parse_pipeline_filename(input_path: str) -> dict:
    """
    Parse batch filename convention: {cohort}_{batch_number}_{row_range}.xlsx
    e.g. Italy100_1_R0001_0500.xlsx  → cohort=Italy100, batch_number=1, row_range=R0001_0500

    Accepts any file; returns valid=False when the name doesn't match the pattern.
    The pattern is: one or more non-underscore parts forming cohort, then a numeric
    batch-number segment, then an optional row-range segment starting with R.
    """
    stem = Path(input_path).stem  # drop extension
    parts = stem.split("_")
    result: dict = {
        "cohort": stem,
        "batch_number": None,
        "row_range": None,
        "batch_stem": stem,
        "valid": False,
    }
    # Scan left-to-right for the first purely-numeric segment (= batch number).
    # Everything before it = cohort; everything after = row_range.
    # Italy100_1_R0001_0500 → cohort=Italy100, batch=1, row_range=R0001_0500
    batch_idx = None
    for i in range(len(parts)):
        if re.fullmatch(r"\d+", parts[i]):
            batch_idx = i
            break
    if batch_idx is None or batch_idx == 0:
        return result
    result["cohort"]        = "_".join(parts[:batch_idx])
    result["batch_number"]  = int(parts[batch_idx])
    row_parts               = parts[batch_idx + 1:]
    result["row_range"]     = "_".join(row_parts) if row_parts else None
    result["batch_stem"]    = stem
    result["valid"]         = True
    return result


def resolve_pipeline_output_paths(
    input_path: str,
    project_root: str | None = None,
    ts: str | None = None,
) -> dict:
    """
    Resolve standard pipeline output paths from an input file path.

    Expected layout (project_root auto-detected as grandparent of 00_raw/ when not given):
      {project_root}/
        {cohort}/
          00_raw/          ← input file lives here
          01_cleaned_domains/
          _logs/
          _archive/

    Returns a dict with keys:
      cohort, batch_stem, batch_number, row_range, valid_name,
      project_root, cohort_dir,
      cleaned_dir, logs_dir, archive_dir,
      output_xlsx, partial_prefix, run_log_csv
    """
    if ts is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M")

    input_p   = Path(input_path).resolve()
    meta      = parse_pipeline_filename(str(input_p))
    batch_stem = meta["batch_stem"]
    cohort     = meta["cohort"]

    # Auto-detect project root: if input lives in a folder named "00_raw",
    # its parent is the cohort folder and grandparent is project_root.
    if project_root is not None:
        _pr = Path(project_root).resolve()
        cohort_dir = _pr / cohort
    elif input_p.parent.name.lower() == "00_raw":
        cohort_dir = input_p.parent.parent
        _pr        = cohort_dir.parent
    else:
        # Fallback: cohort folder = input file's parent
        cohort_dir = input_p.parent
        _pr        = cohort_dir.parent

    cleaned_dir = cohort_dir / "01_cleaned_domains"
    logs_dir    = cohort_dir / "_logs"
    archive_dir = cohort_dir / "_archive"

    output_xlsx    = cleaned_dir / f"{batch_stem}_cleaned_{ts}.xlsx"
    partial_prefix = cleaned_dir / f"{batch_stem}_cleaned_PARTIAL_"
    run_log_csv    = logs_dir    / f"{cohort}_cleaner_runlog.csv"

    return {
        "cohort":        cohort,
        "batch_stem":    batch_stem,
        "batch_number":  meta["batch_number"],
        "row_range":     meta["row_range"],
        "valid_name":    meta["valid"],
        "project_root":  str(_pr),
        "cohort_dir":    str(cohort_dir),
        "cleaned_dir":   str(cleaned_dir),
        "logs_dir":      str(logs_dir),
        "archive_dir":   str(archive_dir),
        "output_xlsx":   str(output_xlsx),
        "partial_prefix": str(partial_prefix),
        "run_log_csv":   str(run_log_csv),
        "ts":            ts,
    }


_RUN_LOG_FIELDS = [
    "timestamp", "cohort", "batch_stem", "batch_number", "row_range",
    "input_path", "output_xlsx", "rows_processed", "rows_input",
    "coverage_pct", "manual_review_count", "no_match_count",
    "haiku_mode", "verifier_provider", "serper_queries",
    "firecrawl_keys_loaded", "run_id", "run_label", "status", "notes",
]


def _append_run_log_csv(csv_path: Path, log_row: dict) -> None:
    """Append one row to the pipeline run-log CSV; creates file with headers on first write."""
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    write_header = not csv_path.exists()
    with open(csv_path, "a", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=_RUN_LOG_FIELDS, extrasaction="ignore")
        if write_header:
            writer.writeheader()
        writer.writerow(log_row)


def _write_pipeline_output(
    output_path: Path,
    excel_bytes: bytes,
    partial: bool = False,
) -> None:
    """Write excel_bytes to output_path, creating parent directories as needed."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(excel_bytes)


def _cp_dir(run_id: str, run_label: str = "") -> Path:
    folder = f"{run_label}_{run_id}" if run_label else run_id
    return _AUTOSAVE_DIR / folder


def _save_checkpoint(
    run_id: str,
    all_results: list[dict],
    all_evidence: list[dict],
    row_idx: int,        # number of rows completed so far
    total_rows: int,
    input_df: pd.DataFrame,
    cols: dict,
    settings: dict,
    run_label: str = "",
) -> None:
    """Persist current progress to autosave/{run_label}_{run_id}/ (or autosave/{run_id}/ if no label)."""
    d = _cp_dir(run_id, run_label)
    d.mkdir(parents=True, exist_ok=True)

    meta = {
        "run_id":     run_id,
        "run_label":  run_label,
        "folder_name": d.name,
        "row_idx":    row_idx,
        "total_rows": total_rows,
        "timestamp":  pd.Timestamp.now().isoformat(timespec="seconds"),
        "cols":       {k: v for k, v in cols.items() if v},
        "settings":   settings,
        "complete":   row_idx >= total_rows,
    }
    (d / "meta.json").write_text(json.dumps(meta, indent=2, default=str), encoding="utf-8")

    if all_results:
        pd.DataFrame(all_results).to_csv(d / "results.csv", index=False)

    if all_evidence:
        (d / "evidence.json").write_text(
            json.dumps(all_evidence, ensure_ascii=False, default=str), encoding="utf-8"
        )

    # Input snapshot — written once; never overwritten (needed for crash-resume)
    input_path = d / "input.csv"
    if not input_path.exists() and input_df is not None:
        input_df.to_csv(input_path, index=False)


def _load_checkpoint_from_dir(d: Path) -> dict | None:
    """Load checkpoint data from a specific directory."""
    try:
        meta = json.loads((d / "meta.json").read_text(encoding="utf-8"))
        results: list[dict] = []
        results_path = d / "results.csv"
        if results_path.exists():
            results = pd.read_csv(results_path, dtype=str).fillna("").to_dict("records")
        evidence: list[dict] = []
        ev_path = d / "evidence.json"
        if ev_path.exists():
            evidence = json.loads(ev_path.read_text(encoding="utf-8"))
        input_df: pd.DataFrame | None = None
        input_path = d / "input.csv"
        if input_path.exists():
            input_df = pd.read_csv(input_path, dtype=str).fillna("")
        return {"meta": meta, "results": results, "evidence": evidence, "input_df": input_df}
    except Exception:
        return None


def _load_checkpoint(run_id: str) -> dict | None:
    """
    Load checkpoint by run_id.  Tries legacy hash-only folder first, then scans all
    checkpoint directories for a meta.json whose run_id matches.
    """
    legacy = _AUTOSAVE_DIR / run_id
    if legacy.is_dir() and (legacy / "meta.json").exists():
        return _load_checkpoint_from_dir(legacy)
    for cp_meta in _list_checkpoints():
        if cp_meta.get("run_id") == run_id:
            folder = _AUTOSAVE_DIR / cp_meta["_folder"]
            if folder.is_dir():
                return _load_checkpoint_from_dir(folder)
    return None


def _list_checkpoints() -> list[dict]:
    """Return all checkpoint meta dicts sorted newest-first.  Adds '_folder' key."""
    if not _AUTOSAVE_DIR.exists():
        return []
    out = []
    for d in _AUTOSAVE_DIR.iterdir():
        if not d.is_dir():
            continue
        mp = d / "meta.json"
        if not mp.exists():
            continue
        try:
            meta = json.loads(mp.read_text(encoding="utf-8"))
            meta["_folder"] = d.name
            out.append(meta)
        except Exception:
            continue
    return sorted(out, key=lambda m: m.get("timestamp", ""), reverse=True)


def _delete_checkpoint(run_id: str) -> None:
    for cp_meta in _list_checkpoints():
        if cp_meta.get("run_id") == run_id:
            shutil.rmtree(_AUTOSAVE_DIR / cp_meta["_folder"], ignore_errors=True)
            return
    shutil.rmtree(_AUTOSAVE_DIR / run_id, ignore_errors=True)


def _checkpoint_excel_bytes(run_id: str, cols: dict) -> bytes | None:
    """Build and return an Excel file from a checkpoint's saved data."""
    cp = _load_checkpoint(run_id)
    if not cp or not cp["results"] or cp["input_df"] is None:
        return None
    try:
        input_df   = cp["input_df"]
        results    = cp["results"]
        evidence   = cp["evidence"]
        n_done     = len(results)

        result_df  = pd.DataFrame(results)
        partial_in = input_df.iloc[:n_done].copy().reset_index(drop=True)
        result_df  = result_df.reset_index(drop=True)

        enriched   = pd.concat([partial_in, result_df], axis=1)
        enriched   = enriched.loc[:, ~enriched.columns.duplicated()]

        return build_excel(enriched, input_df, evidence, cols,
                           debug_rows=None, debug_mode=False, run_meta=None)
    except Exception:
        return None


# =============================================================================
# DATAFRAME PROCESSOR
# =============================================================================

# Output columns added by this tool (v2 includes diagnostic columns)
_OUTPUT_COLS = [
    "cleaned_company_name",
    "normalized_input_website",
    "email_domain",
    "validated_domain",
    "recommended_domain",
    "domain_source",
    "domain_action",
    "domain_confidence",
    "domain_reason",
    "manual_review_needed",
    "search_query_used",
    "serper_top_result_title",
    "serper_top_result_url",
    "serper_top_result_domain",
    # v2 diagnostic
    "name_variant_used",
    "candidate_domains_considered",
    "best_candidate_score",
    "top_3_candidate_domains",
    "rejection_reason_if_missing",
    "website_discovery_method",
    # v3 rejection counts
    "rejected_directory",
    "rejected_government",
    "rejected_religious",
    "rejected_academic",
    "rejected_low_similarity",
    # v4 Haiku review columns
    "haiku_used",
    "haiku_decision",
    "haiku_domain",
    "haiku_confidence",
    "haiku_reason",
    "haiku_risk_flags",
    "haiku_error",
    "final_selected_domain",
    "final_decision_source",
    "final_confidence",
    # v5 Jina verifier columns
    "jina_verifier_used",
    "jina_verified_domain",
    "jina_verifier_confidence",
    "jina_verifier_decision",
    "jina_verifier_reason",
    "jina_evidence_legal_name",
    "jina_evidence_address",
    "jina_evidence_city",
    "jina_evidence_phone",
    "jina_evidence_email",
    "jina_evidence_partita_iva",
    "jina_pages_fetched",
    "jina_fetch_status",
    # v6 unified website verifier columns
    "verification_needed",
    "verification_reason",
    "verifier_provider_used",
    "verifier_used",
    "verifier_decision",
    "verifier_confidence",
    "verifier_selected_domain",
    "verifier_replace_allowed",
    "verifier_reason",
    "verifier_evidence_strength",
    "verifier_negative_source_type",
    "verifier_pages_fetched",
    "verifier_fetch_status",
    "verifier_error",
    # v6 Firecrawl-specific columns
    "firecrawl_used",
    "firecrawl_verified_domain",
    "firecrawl_decision",
    "firecrawl_confidence",
    "firecrawl_reason",
    "firecrawl_evidence_strength",
    "firecrawl_negative_source_type",
    "firecrawl_pages_fetched",
    "firecrawl_fetch_status",
    "firecrawl_error",
    # v7 redirect / wrong entity columns
    "original_candidate_domain",
    "redirect_final_url",
    "redirect_final_domain",
    "canonical_domain_used",
    "wrong_entity_type_signal",
    "wrong_location_signal",
    "firecrawl_redirect_final_url",
    "firecrawl_redirect_final_domain",
    # v8 evidence URL fields
    "verifier_evidence_url",
    "firecrawl_evidence_url",
    # v9 redirect resolution status
    "redirect_checked",
    "redirect_resolution_status",
    "canonical_domain_verification_status",
    "original_domain_risky_marker",
    "risky_marker_reason",
    "redirect_canonical_override_applied",
    # v8 organization eligibility pre-filter columns
    "organization_type",
    "myngle_target_eligibility",
    "pre_filter_decision",
    "pre_filter_reason",
    # v10 professional site scoring
    "professional_site_score",
    "professional_site_level",
    "professional_site_signals",
    # v10 candidate metadata
    "candidate_url",
    "candidate_type",
    "candidate_path",
    "candidate_source",
    # v11 haiku extended fields
    "haiku_candidate_url",
    "haiku_recommended_action",
    # v12 final safety guard fields
    "safety_guard_applied",
    "safety_guard_reason",
    "pre_safety_final_confidence",
    "pre_safety_manual_review_needed",
    "pre_safety_final_decision_source",
    # size inference fields
    "size_inference_enabled",
    "employee_size_band",
    "employee_size_confidence",
    "employee_count_estimate",
    "employee_count_min",
    "employee_count_max",
    "employee_size_evidence_text",
    "employee_size_evidence_url",
    "employee_size_source",
    "hrm_likelihood_score",
    "hrm_signals",
    "career_page_found",
    "size_manual_review_needed",
    "size_inference_stop_reason",
    "firecrawl_used_for_domain_verification",
    "firecrawl_used_for_size_inference",
    "firecrawl_pages_used_for_size",
    "firecrawl_pages_used_total",
    "firecrawl_purpose",
]

# ── Size inference constants ──────────────────────────────────────────────────
SIZE_100_PLUS_CONFIRMED = "SIZE_100_PLUS_CONFIRMED"
SIZE_100_PLUS_LIKELY    = "SIZE_100_PLUS_LIKELY"
SIZE_50_99_LIKELY       = "SIZE_50_99_LIKELY"
SIZE_BELOW_50_LIKELY    = "SIZE_BELOW_50_LIKELY"
SIZE_UNKNOWN            = "SIZE_UNKNOWN"

SIZE_INFERENCE_FIELDS: list[str] = [
    "size_inference_enabled",
    "employee_size_band",
    "employee_size_confidence",
    "employee_count_estimate",
    "employee_count_min",
    "employee_count_max",
    "employee_size_evidence_text",
    "employee_size_evidence_url",
    "employee_size_source",
    "hrm_likelihood_score",
    "hrm_signals",
    "career_page_found",
    "size_manual_review_needed",
    "size_inference_stop_reason",
    # FC tracking by purpose
    "firecrawl_used_for_domain_verification",
    "firecrawl_used_for_size_inference",
    "firecrawl_pages_used_for_size",
    "firecrawl_pages_used_total",
    "firecrawl_purpose",
]


def _size_inference_empty() -> dict:
    """Return a blank size-inference result dict."""
    return {
        "size_inference_enabled": False,
        "employee_size_band": SIZE_UNKNOWN,
        "employee_size_confidence": "none",
        "employee_count_estimate": "",
        "employee_count_min": "",
        "employee_count_max": "",
        "employee_size_evidence_text": "",
        "employee_size_evidence_url": "",
        "employee_size_source": "",
        "hrm_likelihood_score": 0,
        "hrm_signals": "",
        "career_page_found": False,
        "size_manual_review_needed": False,
        "size_inference_stop_reason": "",
        "firecrawl_used_for_domain_verification": False,
        "firecrawl_used_for_size_inference": False,
        "firecrawl_pages_used_for_size": 0,
        "firecrawl_pages_used_total": 0,
        "firecrawl_purpose": "",
    }


def _extract_employee_size_regex(text: str, country_code: str = "IT") -> dict:
    """
    Extract employee count from page text using language-specific regex patterns.
    Returns dict with keys: count_min, count_max, estimate, band, confidence, evidence_text.
    """
    import re as _re

    text = (text or "")[:20_000]  # cap to avoid slow regex on huge pages

    # German patterns
    _DE_PATTERNS = [
        # "über 500 Mitarbeiter", "mehr als 200 Beschäftigte"
        (r"(?:über|mehr als|ca\.?|rund|etwa)\s+(\d[\d\.,]*)\s*(?:Mitarbeiter|Beschäftigte|Angestellte|Mitarbeitende)", "de_ueber"),
        # "500 Mitarbeiter", "1.200 Beschäftigte"
        (r"(\d[\d\.,]*)\s*(?:Mitarbeiter(?:innen)?|Beschäftigte|Angestellte|Mitarbeitende)", "de_plain"),
        # "Team von 50", "Team mit 120 Personen"
        (r"Team\s+(?:von|mit)\s+(\d[\d\.,]*)\s*(?:Personen|Mitarbeiter|Menschen)?", "de_team"),
        # "150 employees" (English on DE sites)
        (r"(\d[\d\.,]*)\s+employees?", "de_en_empl"),
        # "Headcount: 300" — require colon to avoid matching "Headcount 100 songs"
        (r"[Hh]eadcount\s*:\s*(\d[\d\.,]*)", "de_headcount"),
    ]
    # Italian patterns
    _IT_PATTERNS = [
        # "oltre 500 dipendenti", "più di 200 collaboratori"
        (r"(?:oltre|più di|circa|ca\.?|quasi)\s+(\d[\d\.,]*)\s*(?:dipendenti|collaboratori|lavoratori|persone)", "it_oltre"),
        # "500 dipendenti", "1.200 collaboratori"
        (r"(\d[\d\.,]*)\s*(?:dipendenti|collaboratori|lavoratori|risorse umane)", "it_plain"),
        # "team di 50 persone"
        (r"team\s+di\s+(\d[\d\.,]*)\s*(?:persone|dipendenti)?", "it_team"),
        # "150 employees" (English on IT sites)
        (r"(\d[\d\.,]*)\s+employees?", "it_en_empl"),
    ]

    patterns = _DE_PATTERNS if country_code == "DE" else _IT_PATTERNS

    def _parse_num(s: str) -> int | None:
        s = s.replace(".", "").replace(",", "")
        try:
            return int(s)
        except ValueError:
            return None

    best_count: int | None = None
    best_evidence = ""

    for pattern, label in patterns:
        for m in _re.finditer(pattern, text, _re.IGNORECASE):
            n = _parse_num(m.group(1))
            if n is None or n < 2 or n > 500_000:
                continue
            span_start = max(0, m.start() - 40)
            span_end   = min(len(text), m.end() + 40)
            snippet    = text[span_start:span_end].replace("\n", " ").strip()
            if best_count is None or n > best_count:
                best_count    = n
                best_evidence = snippet

    if best_count is None:
        return {"count_min": "", "count_max": "", "estimate": "", "band": SIZE_UNKNOWN,
                "confidence": "none", "evidence_text": ""}

    # Classify band
    if best_count >= 200:
        band = SIZE_100_PLUS_CONFIRMED
        conf = "high"
        cmin, cmax = best_count, ""
    elif best_count >= 100:
        band = SIZE_100_PLUS_CONFIRMED
        conf = "medium"
        cmin, cmax = best_count, ""
    elif best_count >= 50:
        band = SIZE_50_99_LIKELY
        conf = "medium"
        cmin, cmax = 50, 99
    else:
        band = SIZE_BELOW_50_LIKELY
        conf = "low"
        cmin, cmax = "", best_count

    return {
        "count_min": cmin,
        "count_max": cmax,
        "estimate":  best_count,
        "band":      band,
        "confidence": conf,
        "evidence_text": best_evidence,
    }


def _classify_hrm_signals(
    text_corpus: list[str],
    country_code: str = "IT",
    official_domain: str = "",
    evidence_urls: list[str] | None = None,
) -> tuple[int, str, bool]:
    """
    Score HRM likelihood 0-10 from page text corpus.
    Returns (score, signals_csv, career_page_found).
    """
    combined = " ".join((t or "") for t in text_corpus)[:40_000].lower()

    # Signal definitions: (keyword_list, points, label)
    _DE_SIGNALS = [
        (["karriere", "karriereseite", "jobs", "stellenangebote"], 2, "career_page"),
        (["ausbildung", "ausbildungsplatz", "azubi"], 1, "ausbildung"),
        (["personalentwicklung", "weiterbildung", "l&d", "learning & development"], 1, "personalentwicklung"),
        (["hr-abteilung", "human resources", "personalwesen"], 1, "hr_dept"),
        (["onboarding", "einarbeitung"], 1, "onboarding"),
        (["benefits", "mitarbeitervorteile", "corporate benefits"], 1, "benefits"),
        (["mehrere standorte", "niederlassungen", "international"], 1, "multi_location"),
        (["academy", "akademie", "lernpfad"], 1, "academy"),
        (["sprachen", "fremdsprachen", "sprachkenntnisse"], 1, "languages"),
    ]
    _IT_SIGNALS = [
        (["lavora con noi", "posizioni aperte", "carriere", "careers"], 2, "career_page"),
        (["formazione", "sviluppo professionale", "l&d"], 1, "formazione"),
        (["risorse umane", "hr", "human resources"], 1, "hr_dept"),
        (["onboarding", "inserimento"], 1, "onboarding"),
        (["benefit", "welfare aziendale", "fringe benefit"], 1, "benefits"),
        (["più sedi", "sedi", "internazionale"], 1, "multi_location"),
        (["academy", "accademia", "percorso formativo"], 1, "academy"),
        (["lingue", "lingua straniera", "corsi di lingua"], 1, "languages"),
        (["stage", "tirocinio", "apprendistato"], 1, "apprenticeship"),
    ]

    signals_def = _DE_SIGNALS if country_code == "DE" else _IT_SIGNALS
    score = 0
    found_signals: list[str] = []
    career_found = False

    # career_page_found requires BOTH: keyword match AND URL from the official domain
    _evidence_urls_low = [u.lower() for u in (evidence_urls or [])]
    _career_path_signals = ("karriere", "/career", "/jobs", "stellenangebote", "ausbildung",
                            "lavora-con-noi", "posizioni-aperte", "/carriere", "/careers")
    _official_domain_low = (official_domain or "").lower()
    _career_from_official = False
    if _official_domain_low and _evidence_urls_low:
        for _eu in _evidence_urls_low:
            # URL must be from the official domain
            _eu_domain = _extract_domain(_eu)
            _on_official = _eu_domain and (
                _eu_domain == _official_domain_low
                or _eu_domain.endswith("." + _official_domain_low)
                or _official_domain_low.endswith("." + _eu_domain)
            )
            if _on_official:
                # And URL path or domain must signal career/jobs page
                if any(sig in _eu for sig in _career_path_signals):
                    _career_from_official = True
                    break
                # Or the homepage itself (if on official domain and career keywords in text)
                _career_from_official = True  # on official domain = at least trusted
                break

    for keywords, pts, label in signals_def:
        if any(kw in combined for kw in keywords):
            score += pts
            found_signals.append(label)
            if label == "career_page":
                career_found = _career_from_official

    score = min(score, 10)
    return score, ",".join(found_signals), career_found


# Hard-blocked source domains for size inference — media, app stores, music, lyrics sites
_SIZE_INFERENCE_BLOCKED_DOMAINS: frozenset = frozenset({
    "music.apple.com", "apple.com", "spotify.com", "youtube.com", "soundcloud.com",
    "deezer.com", "tidal.com", "bandcamp.com", "genius.com", "azlyrics.com",
    "lyrics.com", "musixmatch.com", "amazon.com", "amazon.de", "amazon.it",
    "ebay.com", "ebay.de", "ebay.it", "facebook.com", "instagram.com",
    "twitter.com", "x.com", "tiktok.com", "linkedin.com", "xing.com",
    "reddit.com", "wikipedia.org", "wikidata.org",
    "play.google.com", "apps.apple.com", "podcasts.apple.com",
})

# Hard-blocked domains for official-domain discovery — music, media, streaming,
# app stores, marketplaces. These must never become python_validated_domain or
# be sent to Firecrawl. Separate from _SIZE_INFERENCE_BLOCKED_DOMAINS to allow
# fine-tuned control of each gate.
_DISCOVERY_BLOCKED_DOMAINS: frozenset = frozenset({
    "music.apple.com", "apps.apple.com", "podcasts.apple.com", "apple.com",
    "spotify.com", "youtube.com", "soundcloud.com", "deezer.com",
    "tidal.com", "bandcamp.com", "iheart.com", "iheartradio.com",
    "last.fm", "tunein.com", "pandora.com", "audiomack.com", "napster.com",
    "amazon.com", "amazon.de", "amazon.it", "amazon.co.uk", "amazon.fr",
    "amazon.es", "amazon.com.au",
    "ebay.com", "ebay.de", "ebay.it",
    "play.google.com", "store.google.com",
    "facebook.com", "instagram.com", "twitter.com", "x.com", "tiktok.com",
    "linkedin.com", "xing.com", "reddit.com",
    "wikipedia.org", "wikidata.org", "wikimedia.org",
    "genius.com", "azlyrics.com", "lyrics.com", "musixmatch.com",
})


def _is_allowed_size_source(
    evidence_url: str,
    company_name: str,
    validated_domain: str,
) -> bool:
    """
    Return True only if the evidence URL is a safe source for employee-size inference.
    Rules:
    - Allow if URL is from the validated company domain.
    - Hard-reject known unrelated domains (music, media, stores, social).
    - For any other domain: require company name words in the URL hostname.
    """
    if not evidence_url:
        return False
    ev_domain = _extract_domain(evidence_url.lower())
    if not ev_domain:
        return False
    # Allow if it is the validated company domain
    if validated_domain and (
        ev_domain == validated_domain
        or ev_domain.endswith("." + validated_domain)
        or validated_domain.endswith("." + ev_domain)
    ):
        return True
    # Hard-reject URL shorteners
    if is_url_shortener(ev_domain):
        return False
    # Hard-reject blocked domains and their subdomains
    for blocked in _SIZE_INFERENCE_BLOCKED_DOMAINS:
        if ev_domain == blocked or ev_domain.endswith("." + blocked):
            return False
    # For other domains: require company-name signal in the domain
    _name_words = [w.lower() for w in (company_name or "").split() if len(w) > 3
                   and w.lower() not in {"gmbh", "co.", "co", "kg", "und", "the", "and",
                                         "srl", "spa", "snc", "sas", "etal", "italia"}]
    if _name_words and any(w in ev_domain for w in _name_words):
        return True
    # Unknown domain with no name signal — reject
    return False


def _build_size_serper_queries(domain: str, country_config: "CountryConfig") -> list[str]:
    """Build up to 2 targeted Serper queries for size inference."""
    cc = country_config.country_code
    queries: list[str] = []
    if cc == "DE":
        queries.append(f"site:{domain} Mitarbeiter OR Beschäftigte OR Mitarbeitende")
        queries.append(f"site:{domain} Karriere OR Stellenangebote OR Jobs")
    else:  # IT default
        queries.append(f"site:{domain} dipendenti OR collaboratori OR team")
        queries.append(f"site:{domain} lavora con noi OR posizioni aperte OR careers")
    return queries[:2]


def _infer_company_size(
    domain: str,
    company_name: str,
    serper_key: str | None,
    fc_keys,
    country_config: "CountryConfig",
    fc_location: dict | None,
    max_size_serper_queries: int = 2,
    max_size_fc_pages: int = 2,
    fc_pages_already_used: int = 0,
    max_total_fc_pages: int = 3,
    existing_texts: list[str] | None = None,
    existing_evidence_url: str = "",
    fc_speed_mode: str = "fast",
    page_timeout: int = 15,
    live_counters: dict | None = None,  # updated in-place with SI FC calls
) -> dict:
    """
    Orchestrate size + HRM inference for one company.
    Returns a dict compatible with SIZE_INFERENCE_FIELDS.
    """
    result = _size_inference_empty()
    result["size_inference_enabled"] = True
    result["firecrawl_used_for_domain_verification"] = fc_pages_already_used > 0

    if not domain:
        result["size_inference_stop_reason"] = "no_domain"
        return result

    # text_corpus: safe Serper snippets + safe existing texts (never raw Apple Music / blocked domains)
    # _fc_texts: only FC-fetched page content from the official company domain (always trusted)
    text_corpus: list[str] = []
    _fc_texts: list[str] = []
    evidence_url = ""
    fc_pages_for_size = 0
    stop_reason = ""

    # Step 1: mine existing Serper/FC evidence already in the row — only if URL is safe
    if (existing_evidence_url
            and _is_allowed_size_source(existing_evidence_url, company_name, domain)):
        for _t in (existing_texts or []):
            if str(_t).strip():
                text_corpus.append(str(_t))
        evidence_url = existing_evidence_url
    # (if existing_evidence_url is blocked / unsafe → discard both URL and texts entirely)

    size_info = _extract_employee_size_regex(" ".join(text_corpus), country_config.country_code)
    if size_info["band"] == SIZE_100_PLUS_CONFIRMED and size_info["confidence"] == "high":
        stop_reason = "exact_count_from_existing_evidence"

    # Step 2: Serper snippets — only from validated domain or company-identified pages
    if not stop_reason and serper_key:
        queries = _build_size_serper_queries(domain, country_config)
        gl = country_config.serper_gl
        hl = country_config.serper_hl
        serper_used = 0
        for q in queries[:max_size_serper_queries]:
            if serper_used >= max_size_serper_queries:
                break
            try:
                _hits, _err = _call_serper(q, serper_key, gl=gl, hl=hl)
                serper_used += 1
                if _hits:
                    for hit in _hits[:3]:
                        _link = hit.get("link", "") or ""
                        # Source-domain safety check — hard-block media/store/social results
                        if not _is_allowed_size_source(_link, company_name, domain):
                            continue
                        snippet = (hit.get("snippet") or "") + " " + (hit.get("title") or "")
                        if snippet.strip():
                            text_corpus.append(snippet)
                            if not evidence_url:
                                evidence_url = _link
            except Exception:
                pass

        # After Serper: require HIGH confidence AND safe evidence URL
        size_info = _extract_employee_size_regex(" ".join(text_corpus), country_config.country_code)
        if (size_info["band"] == SIZE_100_PLUS_CONFIRMED
                and size_info["confidence"] == "high"
                and _is_allowed_size_source(evidence_url, company_name, domain)):
            stop_reason = "confirmed_from_serper_snippets"

    # Step 3: FC pages — only from the validated company domain (domain is already validated)
    fc_budget_remaining = max_total_fc_pages - fc_pages_already_used
    size_fc_budget = min(max_size_fc_pages, fc_budget_remaining)
    _fc_evidence_url = ""  # track FC-specific evidence URL separately

    if not stop_reason and size_fc_budget > 0 and fc_keys:
        cc = country_config.country_code
        career_path = "/karriere" if cc == "DE" else "/lavora-con-noi"
        # Only fetch from the validated company domain — never from third-party URLs
        fc_urls_to_try = [
            f"https://{domain}",
            f"https://{domain}{career_path}",
        ]

        fc_key_list = fc_keys if isinstance(fc_keys, list) else [fc_keys]
        for fc_url in fc_urls_to_try[:size_fc_budget]:
            if fc_pages_for_size >= size_fc_budget:
                break
            try:
                _md, _status, _meta = _fc_scrape(
                    fc_url,
                    fc_key_list,
                    timeout=page_timeout,
                    fc_location=fc_location,
                )
                fc_pages_for_size += 1
                if live_counters is not None:
                    # si_* sub-counters track size-inference FC separately from DV
                    live_counters["si_requests_attempted"] = live_counters.get("si_requests_attempted", 0) + 1
                    live_counters["requests_attempted"] = live_counters.get("requests_attempted", 0) + 1
                page_text = _md or ""
                if page_text:
                    _fc_texts.append(page_text)   # FC-only list (always from official domain)
                    text_corpus.append(page_text)
                    if not _fc_evidence_url:
                        _fc_evidence_url = fc_url  # FC URLs are always from validated domain
                    if live_counters is not None:
                        live_counters["si_pages_successful"] = live_counters.get("si_pages_successful", 0) + 1
                        live_counters["pages_successful"] = live_counters.get("pages_successful", 0) + 1
            except Exception:
                fc_pages_for_size += 1  # count attempt
                if live_counters is not None:
                    live_counters["si_requests_attempted"] = live_counters.get("si_requests_attempted", 0) + 1
                    live_counters["requests_attempted"] = live_counters.get("requests_attempted", 0) + 1

        result["firecrawl_used_for_size_inference"] = fc_pages_for_size > 0
        if _fc_evidence_url:
            evidence_url = _fc_evidence_url
        # Size regex on FC texts only — they come exclusively from the official domain
        if _fc_texts:
            size_info = _extract_employee_size_regex(" ".join(_fc_texts), country_config.country_code)
        else:
            size_info = _extract_employee_size_regex(" ".join(text_corpus), country_config.country_code)
        if size_info["band"] in (SIZE_100_PLUS_CONFIRMED, SIZE_100_PLUS_LIKELY) and size_info["confidence"] != "none":
            stop_reason = "confirmed_from_firecrawl"

    if not stop_reason:
        stop_reason = "max_pages_reached" if (fc_pages_for_size >= size_fc_budget and size_fc_budget > 0) else "no_evidence_found"

    # Final source-safety check: if evidence_url is not from official domain, reset to SIZE_UNKNOWN
    if not _is_allowed_size_source(evidence_url, company_name, domain):
        size_info = {"band": SIZE_UNKNOWN, "confidence": "none", "estimate": "",
                     "count_min": "", "count_max": "", "evidence_text": ""}
        evidence_url = ""
        stop_reason = "unsafe_or_unrelated_source"

    # HRM signals — use ONLY FC-fetched texts from the official domain to avoid false positives
    # from Serper snippets (which may include music bands, app stores, etc.)
    _hrm_texts = _fc_texts  # never use Serper snippets or unvalidated existing_texts for HRM
    hrm_score, hrm_signals, career_found = _classify_hrm_signals(
        _hrm_texts,
        country_config.country_code,
        official_domain=domain,
        evidence_urls=[evidence_url] if evidence_url else [],
    )

    band = size_info["band"]
    confidence = size_info["confidence"]

    # Upgrade band based on HRM signals (size unknown but strong HRM signal → likely 100+)
    if band == SIZE_UNKNOWN and hrm_score >= 5:
        band = SIZE_100_PLUS_LIKELY
        confidence = "low"

    result.update({
        "employee_size_band":       band,
        "employee_size_confidence": confidence,
        "employee_count_estimate":  size_info.get("estimate", ""),
        "employee_count_min":       size_info.get("count_min", ""),
        "employee_count_max":       size_info.get("count_max", ""),
        "employee_size_evidence_text": size_info.get("evidence_text", "")[:300],
        "employee_size_evidence_url":  evidence_url,
        "employee_size_source":     "serper+firecrawl" if result["firecrawl_used_for_size_inference"] else "serper",
        "hrm_likelihood_score":     hrm_score,
        "hrm_signals":              hrm_signals,
        "career_page_found":        career_found,
        "size_manual_review_needed": band == SIZE_UNKNOWN,
        "size_inference_stop_reason": stop_reason,
        "firecrawl_pages_used_for_size": fc_pages_for_size,
        "firecrawl_pages_used_total": fc_pages_already_used + fc_pages_for_size,
        "firecrawl_purpose": "domain_verification+size_inference" if (fc_pages_already_used > 0 and fc_pages_for_size > 0)
                             else "size_inference" if fc_pages_for_size > 0
                             else "domain_verification" if fc_pages_already_used > 0
                             else "",
    })
    return result


def process_dataframe(
    df: pd.DataFrame,
    cols: dict,
    serper_key: str | None,
    max_queries: int = 5,
    progress_cb=None,
    live_counters_out: dict | None = None,  # caller-supplied dict updated in-place with FC stats
    # Autosave / resume parameters
    run_id: str | None = None,
    resume_from: int = 0,
    prior_results: list[dict] | None = None,
    prior_evidence: list[dict] | None = None,
    settings: dict | None = None,
    run_label: str = "",
    # Claude Haiku review layer
    haiku_mode: str = _HAIKU_MODE_PYTHON,
    haiku_api_key: str | None = None,
    haiku_model: str = _DEFAULT_HAIKU_MODEL,
    haiku_max_rows: int = 0,   # 0 = no limit
    # Jina website verifier (legacy — now routed through unified verifier)
    jina_mode: str = _JINA_MODE_UNCERTAIN,
    jina_api_key: str | None = None,
    # Unified website verifier
    verifier_provider: str = _VP_OFF,
    verifier_mode: str = _VM_UNCERTAIN,
    fc_key: str | list[str] | None = None,
    max_cands_per_company: int = 3,
    max_pages_per_cand: int = 3,
    page_timeout: int = 15,
    fc_speed_mode: str = _FC_SPEED_FAST,
    fc_location: dict | None = None,
    # Organization eligibility pre-filter
    eligibility_filter_mode: str = _PF_MODE_COMMERCIAL,
    # Firecrawl health / fail-fast
    fc_fail_fast: bool = True,
    # Debug
    debug_mode: bool = False,
    # Country config
    country_config: "CountryConfig | None" = None,
    # Optional size inference
    infer_size: bool = False,
) -> tuple[pd.DataFrame, list[dict], list[dict], list[dict]]:
    """
    Process rows resume_from..len(df)-1, prepending prior_results for already-done rows.
    Saves a checkpoint to disk every _AUTOSAVE_EVERY rows and on completion.

    Returns (enriched_df_for_all_rows, all_evidence_rows, all_debug_rows, all_jina_debug_rows).
    """
    _reset_serper_counter()
    new_results:   list[dict] = []
    new_evidence:  list[dict] = []
    new_debug:     list[dict] = []
    new_jina_debug: list[dict] = []
    n = len(df)
    process_dataframe._jina_debug = new_jina_debug  # type: ignore[attr-defined]
    _live_fc_counters: dict = (
        live_counters_out
        if live_counters_out is not None
        else {"fc_requests_attempted": 0, "fc_pages_successful": 0,
              "fc_timeouts": 0, "fc_total_secs": 0.0}
    )
    _live_fc_counters.setdefault("fc_requests_attempted", 0)
    _live_fc_counters.setdefault("fc_pages_successful", 0)
    _live_fc_counters.setdefault("fc_timeouts", 0)
    _live_fc_counters.setdefault("fc_total_secs", 0.0)

    company_col  = cols.get("company") or ""
    website_col  = cols.get("website") or ""
    email_col    = cols.get("email") or ""
    city_col     = cols.get("city") or ""
    province_col = cols.get("province") or ""
    postcode_col = cols.get("postcode") or ""

    rows_list = list(df.iterrows())

    for local_i, (_, row) in enumerate(rows_list[resume_from:]):
        global_i = resume_from + local_i

        def _sv(col, _row=row):
            return str(_row.get(col, "") or "").strip() if col else ""

        name     = _sv(company_col)
        website  = _sv(website_col)
        email    = _sv(email_col)
        city     = _sv(city_col)
        province = _sv(province_col)
        postcode = _sv(postcode_col)

        # ── Organization eligibility pre-filter ──────────────────────────────
        _legal_form_col = (country_config.name_col_primary if country_config else "") or ""
        _legal_form_detected = str(row.get("legal_form_detected", "") or "").strip()
        org_type, eligibility, pf_decision, pf_reason = classify_organization(
            name,
            legal_form_hint=_legal_form_detected,
            country_code=country_config.country_code if country_config else "IT",
        )
        _pf_skip = (
            (eligibility == _ELI_EXCLUDE and eligibility_filter_mode in (_PF_MODE_COMMERCIAL, _PF_MODE_MAYBE))
            or (eligibility == _ELI_MAYBE and eligibility_filter_mode == _PF_MODE_COMMERCIAL)
        )
        if _pf_skip:
            res = {
                "organization_type": org_type,
                "myngle_target_eligibility": eligibility,
                "pre_filter_decision": _PF_SKIP,
                "pre_filter_reason": pf_reason,
                "cleaned_company_name": name,
                "normalized_input_website": website,
                "email_domain": "",
                "final_selected_domain": "",
                "final_decision_source": "pre_filter_skip",
                "final_confidence": "",
                "manual_review_needed": False,
                "domain_confidence": "",
                "domain_action": "skip",
                "domain_reason": pf_reason,
            }
            new_results.append(res)
            new_evidence.append({})
            if progress_cb:
                progress_cb(global_i + 1, n)
            continue

        res, raw_ev = validate_register_row(
            name, website, email, city, province, postcode, serper_key, max_queries,
            country_config=country_config,
        )
        res["organization_type"] = org_type
        res["myngle_target_eligibility"] = eligibility
        res["pre_filter_decision"] = pf_decision
        res["pre_filter_reason"] = pf_reason

        # Default safety-guard fields (overwritten by _apply_final_safety_guard if triggered)
        res.update({
            "safety_guard_applied":           False,
            "safety_guard_reason":            "",
            "pre_safety_final_confidence":    "",
            "pre_safety_manual_review_needed": "",
            "pre_safety_final_decision_source": "",
        })

        # Default Haiku fields (Python-only values)
        res.update({
            "haiku_used": False, "haiku_decision": "", "haiku_domain": "",
            "haiku_confidence": "", "haiku_reason": "", "haiku_risk_flags": "",
            "haiku_error": "", "haiku_candidate_url": "", "haiku_recommended_action": "",
        })

        # Determine if Haiku should run for this row
        _haiku_rows_done = global_i - resume_from + 1
        _haiku_limit_ok  = (haiku_max_rows <= 0 or _haiku_rows_done <= haiku_max_rows)
        _is_uncertain    = (
            str(res.get("manual_review_needed", "")).lower() in ("true", "1", "yes")
            or str(res.get("domain_confidence", "")).lower() in ("low", "medium", "none", "")
        )
        _run_haiku = (
            haiku_mode != _HAIKU_MODE_PYTHON
            and haiku_api_key
            and _haiku_limit_ok
            and (haiku_mode == _HAIKU_MODE_ALL or _is_uncertain)
        )

        if _run_haiku:
            email_domain_h = str(res.get("email_domain", "") or "")
            orig_website_h = str(res.get("normalized_input_website", "") or "")
            haiku_res = _haiku_review_domain(
                name, city, province, email_domain_h, orig_website_h,
                res, raw_ev, haiku_api_key, haiku_model,
                country_config=country_config,
            )
            res.update(haiku_res)
        else:
            haiku_res = {"haiku_used": False}

        final_fields = _apply_haiku_decision(res, haiku_res, haiku_mode, raw_evidence=raw_ev)
        res.update(final_fields)

        # ── Populate candidate metadata from best evidence row ───────────────
        _final_sel_dom = str(res.get("final_selected_domain") or res.get("validated_domain") or "")
        _best_ev_row: dict = {}
        _best_ev_score: float = -1.0
        for _ev in raw_ev:
            _ev_dom = _ev.get("domain", "")
            if _ev_dom and _ev_dom == _final_sel_dom:
                try:
                    _ev_sc = float(_ev.get("score", -1))
                except (TypeError, ValueError):
                    _ev_sc = -1.0
                if _ev_sc > _best_ev_score:
                    _best_ev_score = _ev_sc
                    _best_ev_row = _ev
        if not _best_ev_row:
            # Fallback: any evidence row for this domain
            _best_ev_row = next(
                (_ev for _ev in raw_ev if _ev.get("domain", "") == _final_sel_dom), {}
            )
        _cand_url_out  = _best_ev_row.get("candidate_url", "") or ""
        _cand_type_out = _best_ev_row.get("candidate_type", "") or ""
        if _cand_url_out and not _cand_type_out:
            _t = _best_ev_row.get("title", "")
            _s = _best_ev_row.get("snippet", "")
            _cand_type_out = _classify_candidate_type(_cand_url_out, _final_sel_dom, _t, _s)
        try:
            _cand_path_out = urlparse(_cand_url_out).path if _cand_url_out else ""
        except Exception:
            _cand_path_out = ""
        _cand_src_out = _best_ev_row.get("candidate_source", "serper_result") if _best_ev_row else ""
        res["candidate_url"]    = _cand_url_out
        res["candidate_type"]   = _cand_type_out
        res["candidate_path"]   = _cand_path_out
        res["candidate_source"] = _cand_src_out

        # ── Haiku-controlled Firecrawl forcing (Part 6) ───────────────────────
        # Stored on res so _should_verify can read it; also used directly below
        _haiku_rec_action = str(res.get("haiku_recommended_action", "") or "")
        _haiku_dec_final  = str(res.get("haiku_decision", "") or "")
        _haiku_force_fc   = (
            _haiku_rec_action in ("firecrawl_exact_url", "firecrawl_root")
            or _haiku_dec_final == "needs_firecrawl"
        )
        res["_haiku_force_firecrawl"] = _haiku_force_fc

        # ── Default Jina verifier fields ─────────────────────────────────────
        _jina_defaults = {
            "jina_verifier_used": False, "jina_verified_domain": "",
            "jina_verifier_confidence": "", "jina_verifier_decision": "",
            "jina_verifier_reason": "", "jina_evidence_legal_name": "",
            "jina_evidence_address": "", "jina_evidence_city": "",
            "jina_evidence_phone": "", "jina_evidence_email": "",
            "jina_evidence_partita_iva": "", "jina_pages_fetched": 0,
            "jina_fetch_status": "",
        }
        res.update(_jina_defaults)

        # ── Default unified verifier fields ──────────────────────────────────
        _verifier_defaults = {
            "verification_needed": False, "verification_reason": "",
            "verifier_provider_used": verifier_provider,
            "verifier_used": False, "verifier_decision": "",
            "verifier_confidence": "", "verifier_selected_domain": "",
            "verifier_replace_allowed": False, "verifier_reason": "",
            "verifier_evidence_strength": "none",
            "verifier_negative_source_type": "",
            "verifier_pages_fetched": 0, "verifier_fetch_status": "",
            "verifier_error": "",
            "firecrawl_used": False, "firecrawl_verified_domain": "",
            "firecrawl_decision": "", "firecrawl_confidence": "",
            "firecrawl_reason": "", "firecrawl_evidence_strength": "",
            "firecrawl_negative_source_type": "",
            "firecrawl_pages_fetched": 0, "firecrawl_fetch_status": "",
            "firecrawl_error": "",
            "firecrawl_evidence_url": "",
            "verifier_evidence_url": "",
            "redirect_checked": False,
            "redirect_resolution_status": "",
            "canonical_domain_verification_status": "",
            "original_domain_risky_marker": False,
            "risky_marker_reason": "",
            "redirect_canonical_override_applied": False,
            # v7 redirect / wrong entity
            "original_candidate_domain": "",
            "redirect_final_url": "",
            "redirect_final_domain": "",
            "canonical_domain_used": "",
            "wrong_entity_type_signal": False,
            "wrong_location_signal": False,
            "firecrawl_redirect_final_url": "",
            "firecrawl_redirect_final_domain": "",
        }
        res.update(_verifier_defaults)

        # ── Unified website verifier ──────────────────────────────────────────
        _verifier_debug_rows: list[dict] = []
        _verifier_active = verifier_provider != _VP_OFF
        if _verifier_active:
            _name_variants_v = extract_name_variants(name)
            _cur_dom = str(res.get("final_selected_domain") or res.get("validated_domain") or "")
            _should_run_v, _verify_reason = _should_verify(
                res, _name_variants_v, raw_ev, verifier_mode, debug_mode
            )
            # Haiku may force Firecrawl even when _should_verify would skip
            if res.pop("_haiku_force_firecrawl", False) and not _should_run_v:
                _should_run_v   = True
                _verify_reason  = ("haiku_requested_firecrawl_exact_url; " + _verify_reason).strip("; ")
            res["verification_needed"] = _should_run_v
            res["verification_reason"] = _verify_reason
            if _should_run_v:
                # Collect top scored candidates + best candidate_url per domain
                _cand_scores: dict[str, float] = {}
                _cand_urls:   dict[str, str]   = {}  # domain -> best candidate_url
                for _e in raw_ev:
                    if not _e.get("used"):
                        continue
                    _d = _e.get("domain", "")
                    try:
                        _s = float(_e.get("score", 0))
                    except (TypeError, ValueError):
                        _s = 0.0
                    if _d and _s > _cand_scores.get(_d, -1.0):
                        _cand_scores[_d] = _s
                        _curl = _e.get("candidate_url", "")
                        if _curl:
                            _cand_urls[_d] = _curl
                _sorted_cands = sorted(_cand_scores.items(), key=lambda x: x[1], reverse=True)
                _v_cands = [d for d, _ in _sorted_cands[:max_cands_per_company]]
                if _cur_dom and _cur_dom not in _v_cands:
                    _v_cands = [_cur_dom] + _v_cands[: max_cands_per_company - 1]

                try:
                    _verif_res, _verifier_debug_rows = _run_website_verifier(
                        company_name=name,
                        city=city,
                        province=province,
                        email_domain=str(res.get("email_domain", "") or ""),
                        input_partita_iva=str(res.get("partita_iva", "") or ""),
                        name_variants=_name_variants_v,
                        candidates=_v_cands,
                        current_domain=_cur_dom,
                        verifier_provider=verifier_provider,
                        jina_api_key=jina_api_key,
                        fc_key=fc_key,
                        max_cands=max_cands_per_company,
                        max_pages=max_pages_per_cand,
                        page_timeout=page_timeout,
                        fc_speed_mode=fc_speed_mode,
                        python_confidence=str(res.get("final_confidence") or res.get("domain_confidence") or ""),
                        live_counters=_live_fc_counters,
                        fc_location=fc_location,
                        candidate_hints=_cand_urls,
                    )
                    res.update(_verif_res)
                    res["_name_variants"] = _name_variants_v  # temp: used in _apply_verifier_decision
                    _verif_upd = _apply_verifier_decision(res, _verif_res)
                    res.pop("_name_variants", None)
                    if _verif_upd:
                        res.update(_verif_upd)
                except Exception as _vex:
                    res["verifier_error"] = f"verifier_exception:{str(_vex)[:120]}"

        # ── Final safety guard ────────────────────────────────────────────────
        _safety_upd = _apply_final_safety_guard(res)
        if _safety_upd:
            res.update(_safety_upd)

        # ── Optional size inference ───────────────────────────────────────────
        if infer_size and country_config is not None:
            _fc_pages_used_dv = int(res.get("firecrawl_pages_fetched") or 0)
            _existing_texts: list[str] = []
            _ev_url = str(res.get("firecrawl_evidence_url") or res.get("verifier_evidence_url") or "")
            # Reuse any FC-fetched content stored in res
            for _fk in ("firecrawl_reason", "verifier_reason"):
                _ft = str(res.get(_fk) or "")
                if _ft:
                    _existing_texts.append(_ft)
            _size_result = _infer_company_size(
                domain=str(res.get("validated_domain") or res.get("final_selected_domain") or ""),
                company_name=name,
                serper_key=serper_key,
                fc_keys=fc_key,
                country_config=country_config,
                fc_location=fc_location,
                max_size_serper_queries=2,
                max_size_fc_pages=2,
                fc_pages_already_used=_fc_pages_used_dv,
                max_total_fc_pages=3,
                existing_texts=_existing_texts,
                existing_evidence_url=_ev_url,
                fc_speed_mode=fc_speed_mode,
                page_timeout=page_timeout,
                live_counters=_live_fc_counters,
            )
            res.update(_size_result)
        elif not infer_size:
            res["size_inference_enabled"] = False

        # Clean NaN strings from output
        for _k, _v in list(res.items()):
            if isinstance(_v, float) and (_v != _v):  # NaN check
                res[_k] = ""
            elif isinstance(_v, str) and _v.lower() in ("nan", "none", "null"):
                res[_k] = ""
        # Ensure professional_site_level is always set when score exists
        _pro_score = res.get("professional_site_score")
        if _pro_score is not None and _pro_score != "" and not res.get("professional_site_level"):
            try:
                _s = int(_pro_score)
                res["professional_site_level"] = (
                    "strong" if _s >= 5 else "medium" if _s >= 3 else "weak" if _s >= 1 else "none"
                )
            except (ValueError, TypeError):
                res["professional_site_level"] = "none"

        new_results.append(res)

        # ── Firecrawl runtime health update + fail-fast check ─────────────────
        if verifier_provider in (_VP_FIRECRAWL, _VP_FC_JINA):
            _fc_was_used = str(res.get("firecrawl_used", "") or "").lower() in ("true", "1")
            _fc_fetch_st = str(
                res.get("firecrawl_fetch_status")
                or res.get("verifier_fetch_status")
                or ""
            )
            _fc_fo_count = int(res.get("firecrawl_key_failover_count") or 0)
            _fc_key_sts  = str(res.get("firecrawl_key_statuses") or "")
            _fc_health_update(
                _live_fc_counters, _fc_fetch_st, _fc_key_sts,
                _fc_fo_count, _fc_was_used,
            )
            _ff = _fc_health_check_fail_fast(_live_fc_counters, enabled=fc_fail_fast)
            if _ff:
                _live_fc_counters["fail_fast_triggered"] = True
                _live_fc_counters["fail_fast_reason"]    = _ff
                raise FcFailFastError(_ff)
            _att = _live_fc_counters.get("requests_attempted", 0)
            if _att > 0 and _att % _FC_HEALTH_LOG_EVERY == 0:
                print(
                    f"[FC HEALTH] {_fc_health_summary_line(_live_fc_counters)}",
                    flush=True,
                )

        # Accumulate verifier debug rows; back-fill Haiku + candidate metadata
        if _verifier_debug_rows:
            _haiku_meta = {
                "haiku_decision":          res.get("haiku_decision", ""),
                "haiku_confidence":        res.get("haiku_confidence", ""),
                "haiku_candidate_url":     res.get("haiku_candidate_url", ""),
                "haiku_recommended_action": res.get("haiku_recommended_action", ""),
                "candidate_url":           res.get("candidate_url", ""),
                "candidate_type":          res.get("candidate_type", ""),
                "candidate_source":        res.get("candidate_source", ""),
                "candidate_path":          res.get("candidate_path", ""),
            }
            for _vdr in _verifier_debug_rows:
                for _k, _v in _haiku_meta.items():
                    if not _vdr.get(_k):
                        _vdr[_k] = _v
                # Derive candidate_path from original_candidate_url if still missing
                if not _vdr.get("candidate_path"):
                    _oc_url = _vdr.get("original_candidate_url", "")
                    if _oc_url:
                        try:
                            _vdr["candidate_path"] = urlparse(_oc_url).path
                        except Exception:
                            pass
        new_jina_debug.extend(_verifier_debug_rows)

        query = res.get("search_query_used", "")
        if query:
            new_evidence.append({
                "company_name":      name,
                "city":              city,
                "province":          province,
                "search_query_used": query,
                "serper_top_title":  res.get("serper_top_result_title", ""),
                "serper_top_url":    res.get("serper_top_result_url", ""),
                "serper_top_domain": res.get("serper_top_result_domain", ""),
                "validated_domain":  res.get("validated_domain", ""),
                "domain_source":     res.get("domain_source", ""),
                "domain_action":     res.get("domain_action", ""),
                "domain_confidence": res.get("domain_confidence", ""),
                "name_variant_used": res.get("name_variant_used", ""),
                "top_3_candidates":  res.get("top_3_candidate_domains", ""),
                "rejection_reason":  res.get("rejection_reason_if_missing", ""),
                "discovery_method":  res.get("website_discovery_method", ""),
            })

        # Candidate Discovery Debug rows — one row per Serper result (or sentinel)
        if debug_mode:
            _final_py  = res.get("validated_domain", "")
            _final_sel = res.get("final_selected_domain", "")
            _haiku_dec = res.get("haiku_decision", "")
            if raw_ev:
                # Find best-score evidence row for the selected domain (for selected_best_candidate)
                _best_sel_score: float = -1.0
                _best_sel_idx: int = -1
                _row_start = len(new_debug)
                query_counters: dict[str, int] = {}
                for e in raw_ev:
                    q = e.get("query", "")
                    query_counters[q] = query_counters.get(q, 0) + 1
                    is_sel = e.get("domain", "") == _final_sel and bool(_final_sel)
                    try:
                        _sc = float(e.get("score", -1))
                    except (TypeError, ValueError):
                        _sc = -1.0
                    if is_sel and _sc > _best_sel_score:
                        _best_sel_score = _sc
                        _best_sel_idx = len(new_debug)
                    _e_url  = e.get("url", "")
                    _e_dom  = e.get("domain", "")
                    _e_curl = e.get("candidate_url", "") or _e_url
                    _e_ctyp = e.get("candidate_type", "") or _classify_candidate_type(
                        _e_curl, _e_dom, e.get("title", ""), e.get("snippet", ""))
                    try:
                        _e_cpath = urlparse(_e_curl).path if _e_curl else ""
                    except Exception:
                        _e_cpath = ""
                    new_debug.append({
                        "company_name":          name,
                        "row_number":            global_i + 1,
                        "search_query":          q,
                        "result_rank":           query_counters[q] if e.get("url") else "",
                        "title":                 e.get("title", ""),
                        "snippet":               e.get("snippet", ""),
                        "url":                   _e_url,
                        "extracted_domain":      _e_dom,
                        "candidate_url":         _e_curl,
                        "candidate_type":        _e_ctyp,
                        "candidate_source":      e.get("candidate_source", "serper_result"),
                        "candidate_path":        _e_cpath,
                        "score":                 e.get("score", ""),
                        "used":                  e.get("used", ""),
                        "skip_reason":           e.get("skip_reason", ""),
                        "rejection_category":    e.get("rejection_category", ""),
                        "brand_overlap":         e.get("brand_overlap", ""),
                        "full_overlap":          e.get("full_overlap", ""),
                        "location_match":        e.get("location_match", ""),
                        "email_match":           e.get("email_match", ""),
                        "official_signal":       e.get("official_signal", ""),
                        "final_python_domain":   _final_py,
                        "haiku_mode":            haiku_mode,
                        "haiku_decision":        _haiku_dec,
                        "haiku_candidate_url":   res.get("haiku_candidate_url", ""),
                        "haiku_recommended_action": res.get("haiku_recommended_action", ""),
                        "final_selected_domain": _final_sel,
                        "selected_candidate":    is_sel,
                        "selected_best_candidate": False,  # back-filled below
                    })
                # Back-fill the single best evidence row for the selected domain
                if _best_sel_idx >= 0:
                    new_debug[_best_sel_idx]["selected_best_candidate"] = True
            else:
                # No Serper evidence at all (no key, pre-Serper failure, or website
                # already accepted without search). Always emit one sentinel row so
                # every input company appears in the debug sheet.
                new_debug.append({
                    "company_name":          name,
                    "row_number":            global_i + 1,
                    "search_query":          res.get("search_query_used", ""),
                    "result_rank":           "",
                    "title":                 "",
                    "snippet":               "",
                    "url":                   "",
                    "extracted_domain":      "",
                    "score":                 "",
                    "used":                  False,
                    "skip_reason":           "no_candidates_generated",
                    "rejection_category":    "candidate_generation_failed",
                    "brand_overlap":         "",
                    "full_overlap":          "",
                    "location_match":        "",
                    "email_match":           "",
                    "official_signal":       "",
                    "final_python_domain":   _final_py,
                    "haiku_mode":            haiku_mode,
                    "haiku_decision":        _haiku_dec,
                    "final_selected_domain":   _final_sel,
                    "selected_candidate":      False,
                    "selected_best_candidate": False,
                })

        rows_done = global_i + 1
        # Checkpoint every N rows and on the final row
        if run_id and (rows_done % _AUTOSAVE_EVERY == 0 or rows_done == n):
            all_r = list(prior_results or []) + new_results
            all_e = list(prior_evidence or []) + new_evidence
            _save_checkpoint(
                run_id, all_r, all_e, rows_done, n,
                df, cols, settings or {}, run_label=run_label,
            )

        if progress_cb:
            progress_cb(rows_done, n)

    # Merge prior completed rows with newly processed rows
    all_results  = list(prior_results or []) + new_results
    all_evidence = list(prior_evidence or []) + new_evidence
    all_debug    = new_debug       # debug rows only cover the newly processed rows
    all_jina_debug = new_jina_debug

    result_df = pd.DataFrame(all_results, index=df.index)
    enriched  = pd.concat([df.copy(), result_df], axis=1)
    enriched  = enriched.loc[:, ~enriched.columns.duplicated()]
    process_dataframe._last_serper_count = _get_serper_count()
    return enriched, all_evidence, all_debug, all_jina_debug


# =============================================================================
# EXCEL BUILDER
# =============================================================================


def _header_style():
    from openpyxl.styles import PatternFill, Font
    fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=10)
    return fill, font


def _action_fill(action: str):
    from openpyxl.styles import PatternFill
    hex_color = _ACTION_COLORS.get(str(action), "FFFFFF")
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _write_sheet(ws, df: pd.DataFrame) -> None:
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    hdr_fill, hdr_font = _header_style()
    cols = list(df.columns)

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = min(
            max(len(str(col)) + 2, 12), 52
        )

    action_idx   = (cols.index("domain_action") + 1)   if "domain_action"   in cols else None
    val_dom_idx  = (cols.index("validated_domain") + 1) if "validated_domain" in cols else None
    norm_web_idx = (cols.index("normalized_input_website") + 1) if "normalized_input_website" in cols else None

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        action = str(row.get("domain_action", "") or "")
        fill   = _action_fill(action)
        for ci, col in enumerate(cols, 1):
            val = row[col]
            if isinstance(val, float) and val != val:
                val = ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=False, vertical="top")

        if val_dom_idx and norm_web_idx:
            v = str(row.get("validated_domain", "") or "")
            n = str(row.get("normalized_input_website", "") or "")
            if v and v != n:
                from openpyxl.styles import Font
                ws.cell(row=ri, column=val_dom_idx).font = Font(bold=True, color="C00000")

    ws.freeze_panes = "A2"
    if len(df) > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    ws.row_dimensions[1].height = 18


def _build_best_guess_df(
    enriched_df: pd.DataFrame,
    cols: dict,
) -> pd.DataFrame:
    """
    Best Guess Input — contains all fields useful for Lead Prioritizer:
    company_name, website_url (best guess), email, city, province, phone,
    plus key diagnostic columns.

    Column naming:
    - python_validated_domain   = what Python scoring selected (may be a forum/redirect)
    - python_recommended_domain = what Python suggested as replacement
    - final_selected_domain     = the actual business output domain (may differ from Python)
    - website_url               = same as final_selected_domain (or Python fallback)
    - firecrawl_original_checked_domain = which domain Firecrawl actually scraped
    """
    company_col  = cols.get("company") or ""
    email_col    = cols.get("email") or ""
    city_col     = cols.get("city") or ""
    province_col = cols.get("province") or ""
    phone_col    = cols.get("phone") or ""

    # Priority fallback chains for Best Guess columns — handle both IT and DE output
    _company_fallbacks = [
        company_col, "company_name", "cleaned_company_name",
        "company_name_clean", "company_name_raw",
    ]
    _city_fallbacks    = [city_col, "city", "city_or_registered_office"]
    _province_fallbacks = [province_col, "province", "federal_state"]

    rows = []
    for _, r in enriched_df.iterrows():
        def _sv(col):
            return str(r.get(col, "") or "").strip() if col else ""

        def _sv_first(*cols_chain):
            for c in cols_chain:
                if c:
                    v = str(r.get(c, "") or "").strip()
                    if v:
                        return v
            return ""

        action       = str(r.get("domain_action", "") or "")
        norm         = str(r.get("normalized_input_website", "") or "").strip()
        recom        = str(r.get("recommended_domain", "") or "").strip()
        final        = str(r.get("final_selected_domain", "") or "").strip()
        orig_cand    = str(r.get("original_candidate_domain", "") or "").strip()
        redir_dom    = str(r.get("redirect_final_domain", "") or "").strip()
        redir_url    = str(r.get("redirect_final_url", "") or "").strip()
        py_validated = str(r.get("validated_domain", "") or "").strip()
        dec_src      = str(r.get("final_decision_source", "") or "")

        # Business output URL: final_selected_domain wins, then Python fallback
        if final:
            url = final
        elif action in ("OK", "LIKELY_OK"):
            url = norm
        elif action in ("SUGGEST_REPLACE", "MISSING_DOMAIN_FIXED", "EMAIL_DERIVED"):
            url = recom or norm
        elif action == "REVIEW":
            url = norm or recom
        else:
            url = norm

        # Was the final domain overridden from what Python selected?
        _override = dec_src in (
            "redirect_canonical_unverified", "verifier_replace",
            "verifier_flag_high_conf", "verifier_reject",
        )
        _final_is_output = bool(url)

        # Human-readable explanation of how the final domain was chosen
        if dec_src == "redirect_canonical_unverified":
            _reason = (
                f"Python first selected {orig_cand or py_validated or 'unknown'}, "
                f"but redirect resolver found {redir_dom}. "
                f"Final output uses {final}; manual review kept because Firecrawl evidence was weak."
            )
        elif dec_src == "verifier_replace":
            _reason = (
                f"Firecrawl verified and replaced Python domain "
                f"({orig_cand or py_validated}) with {final}."
            )
        elif dec_src == "verifier_confirm":
            _reason = f"Firecrawl confirmed Python-selected domain ({final})."
        elif dec_src == "verifier_reject":
            _reason = (
                f"Firecrawl rejected Python-selected domain ({orig_cand or py_validated}). "
                "No confident replacement found; manual review required."
            )
        elif dec_src == "verifier_timeout_high_risk":
            _reason = (
                f"Firecrawl timed out on high-risk row. "
                f"Domain {final or py_validated} kept but requires manual review."
            )
        elif dec_src == "pre_filter_skip":
            _reason = f"Row skipped by eligibility pre-filter: {str(r.get('pre_filter_reason', '') or '')}."
        elif dec_src == "plausible_serper_result":
            _reason = (
                f"Plausible candidate ({final or py_validated}) found in search but with low confidence. "
                "Manual review required before use."
            )
        elif dec_src == "group_subsidiary_page":
            _reason = (
                f"Domain ({final or py_validated}) matched via group/subsidiary page in search results. "
                "Verify this is the correct company entity."
            )
        elif final and final == py_validated:
            _reason = f"Python scoring selected {final} (no verifier override)."
        elif final and final != py_validated and py_validated:
            _reason = f"Python selected {py_validated}; haiku/verifier overrode to {final}."
        else:
            _reason = ""

        _cname = _sv_first(*_company_fallbacks)
        _curl  = url
        _cdom  = re.sub(r"^https?://(www\.)?", "", url).split("/")[0].strip() if url else ""

        rows.append({
            # ── Business-facing (lead prioritiser input) ──────────────────────
            "company_name":                      _cname,
            "website_url":                       _curl,
            "email":                             _sv(email_col),
            "city":                              _sv_first(*_city_fallbacks),
            "province":                          _sv_first(*_province_fallbacks),
            "phone":                             _sv(phone_col),
            # ── Downstream-required canonical aliases ─────────────────────────
            # buyer_contact_finder / opportunity_radar detect these by name;
            # canonical_company_url is in _DOMAIN_CANDIDATES so it wins over website_url.
            "canonical_company_name":            _cname,
            "canonical_company_url":             _curl,
            "canonical_company_domain":          _cdom,
            "source_contact_count":              "",
            "input_type":                        "enriched_export",
            # ── Remaining business columns ────────────────────────────────────
            "final_selected_domain":             final,
            "final_domain_is_business_output":   _final_is_output,
            "final_decision_source":             dec_src,
            "final_confidence":                  str(r.get("final_confidence", "") or ""),
            "manual_review_needed":              r.get("manual_review_needed", False),
            "business_output_reason":            _reason,
            # ── Redirect / canonical resolution ───────────────────────────────
            "original_candidate_domain":         orig_cand,
            "redirect_final_domain":             redir_dom,
            "redirect_final_url":                redir_url,
            "redirect_canonical_override_applied": bool(r.get("redirect_canonical_override_applied", False)),
            "canonical_domain_verification_status": str(r.get("canonical_domain_verification_status", "") or ""),
            # ── Python scoring (for traceability, renamed for clarity) ────────
            "python_validated_domain":           py_validated,
            "python_recommended_domain":         recom,
            "domain_action":                     action,
            "domain_confidence":                 str(r.get("domain_confidence", "") or ""),
            "domain_source":                     str(r.get("domain_source", "") or ""),
            "website_discovery_method":          str(r.get("website_discovery_method", "") or ""),
            # ── Verifier summary ──────────────────────────────────────────────
            "verification_needed":               r.get("verification_needed", False),
            "verification_reason":               str(r.get("verification_reason", "") or ""),
            "verifier_used":                     r.get("verifier_used", False),
            "verifier_decision":                 str(r.get("verifier_decision", "") or ""),
            "verifier_reason":                   str(r.get("verifier_reason", "") or ""),
            "verifier_evidence_strength":        str(r.get("verifier_evidence_strength", "") or ""),
            "verifier_evidence_url":             str(r.get("verifier_evidence_url", "") or ""),
            "verifier_negative_source_type":     str(r.get("verifier_negative_source_type", "") or ""),
            # ── Firecrawl detail ──────────────────────────────────────────────
            "firecrawl_original_checked_domain": str(r.get("firecrawl_verified_domain", "") or ""),
            "firecrawl_decision":                str(r.get("firecrawl_decision", "") or ""),
            "firecrawl_reason":                  str(r.get("firecrawl_reason", "") or ""),
            "firecrawl_pages_fetched":           r.get("firecrawl_pages_fetched", 0),
            "firecrawl_fetch_status":            str(r.get("firecrawl_fetch_status", "") or ""),
            "firecrawl_evidence_url":            str(r.get("firecrawl_evidence_url", "") or ""),
            # ── Wrong entity / location signals ───────────────────────────────
            "wrong_entity_type_signal":          r.get("wrong_entity_type_signal", False),
            "wrong_location_signal":             r.get("wrong_location_signal", False),
            "original_domain_risky_marker":      bool(r.get("original_domain_risky_marker", False)),
            "risky_marker_reason":               str(r.get("risky_marker_reason", "") or ""),
            # ── Eligibility pre-filter ────────────────────────────────────────
            "organization_type":                 str(r.get("organization_type", "") or ""),
            "myngle_target_eligibility":         str(r.get("myngle_target_eligibility", "") or ""),
            # ── Professional site scoring ─────────────────────────────────────
            "professional_site_score":           r.get("professional_site_score", ""),
            "professional_site_level":           str(r.get("professional_site_level", "") or ""),
            "professional_site_signals":         str(r.get("professional_site_signals", "") or ""),
        })
    bg = pd.DataFrame(rows)
    # Scrub any NaN values that leak from rows without FC / size-inference data
    for _c in bg.select_dtypes(include="object").columns:
        bg[_c] = bg[_c].fillna("").replace("nan", "")
    return bg


def _write_best_guess_sheet(ws, bg_df: pd.DataFrame, enriched_df: pd.DataFrame) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    hdr_fill, hdr_font = _header_style()
    cols = list(bg_df.columns)

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = min(
            max(len(str(col)) + 2, 16), 48
        )

    changed_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    blank_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ok_fill      = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    email_fill   = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")

    orig_norms = enriched_df.get(
        "normalized_input_website",
        pd.Series("", index=enriched_df.index)
    )

    url_col_idx = (cols.index("website_url") + 1) if "website_url" in cols else 2

    for ri, ((_, bg_row), orig_norm) in enumerate(
        zip(bg_df.iterrows(), orig_norms), 2
    ):
        action = str(bg_row.get("domain_action", "") or "")
        url    = str(bg_row.get("website_url", "") or "").strip()
        orig   = str(orig_norm or "").strip()

        for ci, col in enumerate(cols, 1):
            val = bg_row[col]
            if isinstance(val, float) and val != val:
                val = ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(vertical="top")

        url_cell = ws.cell(row=ri, column=url_col_idx)
        src = str(bg_row.get("domain_source", "") or "")
        if not url:
            url_cell.fill = blank_fill
        elif src in (SRC_EMAIL, SRC_SERPER_EMAIL) or action == "EMAIL_DERIVED":
            url_cell.fill = email_fill
            url_cell.font = Font(italic=True, color="1F497D")
        elif url != orig:
            url_cell.fill = changed_fill
            url_cell.font = Font(bold=True, color="C00000")
        else:
            url_cell.fill = ok_fill

    ws.freeze_panes = "A2"
    if len(bg_df) > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    ws.row_dimensions[1].height = 18


def _write_run_summary_sheet(ws, run_meta: dict) -> None:
    """Write a two-column (Field / Value) run summary sheet."""
    from openpyxl.styles import Alignment, Font, PatternFill
    hdr_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    key_font  = Font(bold=True, size=10)
    for ci, header in enumerate(["Field", "Value"], 1):
        cell = ws.cell(row=1, column=ci, value=header)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="left")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 52
    for ri, (field, value) in enumerate(run_meta.items(), 2):
        a = ws.cell(row=ri, column=1, value=str(field))
        a.font = key_font
        a.alignment = Alignment(vertical="top")
        b = ws.cell(row=ri, column=2, value=str(value) if value is not None else "")
        b.alignment = Alignment(vertical="top", wrap_text=False)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 18


def _build_rank_diagnostics(
    debug_rows: list[dict],
    enriched_df: pd.DataFrame,
    cols: dict,
) -> pd.DataFrame:
    """Build one row per company with rank/score diagnostics for the selected domain."""
    if not debug_rows:
        return pd.DataFrame()

    company_col = cols.get("company", "")
    rows = []
    # Group debug rows by company
    from itertools import groupby as _groupby
    sorted_debug = sorted(debug_rows, key=lambda r: r.get("company_name", ""))
    for company_name, group_iter in _groupby(sorted_debug, key=lambda r: r.get("company_name", "")):
        group = list(group_iter)
        final_sel = group[0].get("final_selected_domain", "")

        # Find enriched row for this company to get serper_top and confidence
        if company_col and company_col in enriched_df.columns:
            match_mask = enriched_df[company_col].astype(str).str.strip() == company_name
            matched = enriched_df[match_mask]
        else:
            matched = pd.DataFrame()
        serper_top_domain = (
            matched["serper_top_result_domain"].iloc[0]
            if not matched.empty and "serper_top_result_domain" in matched.columns
            else ""
        )
        final_confidence = (
            matched["final_confidence"].iloc[0]
            if not matched.empty and "final_confidence" in matched.columns
            else matched["domain_confidence"].iloc[0]
            if not matched.empty and "domain_confidence" in matched.columns
            else ""
        )

        # Scored candidates only (have a numeric score)
        scored = []
        for r in group:
            try:
                s = float(r.get("score", ""))
                scored.append((r, s))
            except (TypeError, ValueError):
                pass

        # Top Serper domain = first scored candidate in rank 1 of first query
        if not serper_top_domain:
            q_order: dict[str, int] = {}
            for r in group:
                q = r.get("search_query", "")
                if q not in q_order:
                    q_order[q] = len(q_order)
            rank1_q0 = [
                r for r in group
                if q_order.get(r.get("search_query", ""), 99) == 0
                and r.get("result_rank") == 1
                and r.get("extracted_domain")
            ]
            serper_top_domain = rank1_q0[0].get("extracted_domain", "") if rank1_q0 else ""

        # Selected domain stats
        sel_entries = [(r, s) for r, s in scored if r.get("extracted_domain") == final_sel]
        if sel_entries:
            best_sel_r, best_sel_s = max(sel_entries, key=lambda x: x[1])
            sel_first_seen_query  = best_sel_r.get("search_query", "")
            sel_first_seen_rank   = best_sel_r.get("result_rank", "")
            sel_best_score        = round(best_sel_s, 3)
            sel_best_title        = best_sel_r.get("title", "")
            sel_best_url          = best_sel_r.get("url", "")
            # Find first occurrence (lowest score rank within first query seen)
            first_q_entries = [
                (r, s) for r, s in sel_entries
                if r.get("search_query") == sel_first_seen_query
            ]
            if first_q_entries:
                first_r = min(first_q_entries, key=lambda x: x[0].get("result_rank") or 99)
                sel_first_seen_rank = first_r[0].get("result_rank", "")
        else:
            sel_first_seen_query = sel_first_seen_rank = ""
            sel_best_score = sel_best_title = sel_best_url = ""

        # Top serper domain stats
        top_entries = [(r, s) for r, s in scored if r.get("extracted_domain") == serper_top_domain]
        top_serper_score = round(max(s for _, s in top_entries), 3) if top_entries else ""

        # Reason label
        if not final_sel:
            reason = "no_domain_selected"
        elif final_sel == serper_top_domain:
            reason = "top_result_selected"
        elif not serper_top_domain:
            reason = "no_serper_top_available"
        else:
            reason = "lower_rank_selected"

        # Query index of selected domain
        q_order2: dict[str, int] = {}
        for r in group:
            q = r.get("search_query", "")
            if q not in q_order2:
                q_order2[q] = len(q_order2)
        sel_query_idx = q_order2.get(sel_first_seen_query, "") if sel_first_seen_query else ""

        rows.append({
            "company_name":                  company_name,
            "final_selected_domain":         final_sel,
            "final_confidence":              final_confidence,
            "selected_domain_first_seen_query": sel_first_seen_query,
            "selected_domain_query_index":   sel_query_idx,
            "selected_domain_first_seen_rank": sel_first_seen_rank,
            "selected_domain_best_score":    sel_best_score,
            "selected_domain_best_title":    sel_best_title,
            "selected_domain_best_url":      sel_best_url,
            "top_serper_domain":             serper_top_domain,
            "top_serper_domain_score":       top_serper_score,
            "selected_vs_top_reason":        reason,
        })
    return pd.DataFrame(rows)


def _build_validation_summary(
    enriched_df: pd.DataFrame,
    debug_rows: list[dict],
    cols: dict,
) -> dict:
    """Compact validation metrics beyond Run Summary basics."""
    n = len(enriched_df)

    final_col = "final_selected_domain" if "final_selected_domain" in enriched_df.columns \
        else "validated_domain"
    finals = enriched_df.get(final_col, pd.Series(dtype=str)).astype(str).str.strip()
    unique_domains = int(finals.replace("", pd.NA).dropna().nunique())

    confs = enriched_df.get("final_confidence",
                enriched_df.get("domain_confidence", pd.Series(dtype=str))).astype(str)
    n_high   = int(confs.str.lower().eq("high").sum())
    n_medium = int(confs.str.lower().eq("medium").sum())
    n_low    = int(confs.str.lower().isin(["low", "none"]).sum())

    review = int(
        enriched_df.get("manual_review_needed", pd.Series(dtype=str))
        .astype(str).str.lower().isin(["true", "1", "yes"]).sum()
    )

    # Serper top vs final divergence (from enriched_df)
    top_domain_col = "serper_top_result_domain"
    if top_domain_col in enriched_df.columns:
        top_doms  = enriched_df[top_domain_col].astype(str).str.strip()
        diverged  = int(
            ((top_doms != "") & (finals != "") & (top_doms != finals)).sum()
        )
    else:
        diverged = "n/a"

    # Rank-based diagnostics from debug_rows
    below_rank1 = 0
    later_query  = 0
    if debug_rows:
        rd = _build_rank_diagnostics(debug_rows, enriched_df, cols)
        if not rd.empty:
            has_sel = rd["final_selected_domain"].astype(str).str.strip() != ""
            # selected domain appeared below rank 1
            below_rank1 = int(
                (has_sel & (rd["selected_domain_first_seen_rank"].apply(
                    lambda v: (int(v) > 1) if str(v).isdigit() else False
                ))).sum()
            )
            # selected domain found in query index > 0
            later_query = int(
                (has_sel & (rd["selected_domain_query_index"].apply(
                    lambda v: (int(v) > 0) if str(v).isdigit() else False
                ))).sum()
            )

    return {
        "processed_rows":                       n,
        "unique_final_domains":                 unique_domains,
        "confidence_High":                      n_high,
        "confidence_Medium":                    n_medium,
        "confidence_Low_or_None":               n_low,
        "manual_review_needed":                 review,
        "top_serper_differs_from_selected":     diverged,
        "selected_domain_below_rank_1":         below_rank1,
        "selected_domain_from_later_query":     later_query,
    }


def _xl_write_fc_audit(ws, fc_audit: dict) -> None:
    """Write Firecrawl Audit sheet — key-value table, no actual API key values."""
    import openpyxl.styles as _oxl_styles
    _bold = _oxl_styles.Font(bold=True)
    ws.cell(row=1, column=1, value="Field").font       = _bold
    ws.cell(row=1, column=2, value="Value").font       = _bold
    ws.cell(row=1, column=3, value="Description").font = _bold

    _rows = [
        ("firecrawl_enabled",            fc_audit.get("firecrawl_enabled", ""),
         "Whether Firecrawl verifier was selected for this run"),
        ("firecrawl_preflight_status",   fc_audit.get("firecrawl_preflight_status", ""),
         "OK / DEGRADED / FAILED / SKIPPED"),
        ("firecrawl_keys_total",         fc_audit.get("firecrawl_keys_total", ""),
         "Number of Firecrawl API keys available"),
        ("firecrawl_keys_ok",            fc_audit.get("firecrawl_keys_ok", ""),
         "Keys that passed the preflight test"),
        ("firecrawl_keys_failed",        fc_audit.get("firecrawl_keys_failed", ""),
         "Keys that failed the preflight test (key labels only — key1, key2, …)"),
        ("firecrawl_key_statuses_preflight", fc_audit.get("firecrawl_key_statuses_preflight", ""),
         "Per-key preflight result: {key1: ok, key2: failed, …} — no actual key values"),
        ("firecrawl_requests_attempted", fc_audit.get("firecrawl_requests_attempted", ""),
         "Total Firecrawl rows attempted during processing"),
        ("firecrawl_pages_successful",   fc_audit.get("firecrawl_pages_successful", ""),
         "Rows where Firecrawl returned OK"),
        ("firecrawl_success_rate",       fc_audit.get("firecrawl_success_rate", ""),
         "pages_successful / requests_attempted"),
        ("firecrawl_key_failovers",      fc_audit.get("firecrawl_key_failovers", ""),
         "Total failover-to-next-key events"),
        ("firecrawl_key_failure_events", fc_audit.get("firecrawl_key_failure_events", ""),
         "Rows that triggered at least one key failover"),
        ("firecrawl_quota_or_billing_errors", fc_audit.get("firecrawl_quota_or_billing_errors", ""),
         "Quota / billing / auth errors across all keys"),
        ("firecrawl_rate_limit_errors",  fc_audit.get("firecrawl_rate_limit_errors", ""),
         "HTTP 429 / rate-limit errors"),
        ("firecrawl_timeouts",           fc_audit.get("firecrawl_timeouts", ""),
         "Request timeouts"),
        ("firecrawl_exceptions",         fc_audit.get("firecrawl_exceptions", ""),
         "Unexpected exceptions during Firecrawl calls"),
        ("firecrawl_consecutive_failures_max", fc_audit.get("firecrawl_consecutive_failures_max", ""),
         "Longest streak of consecutive failures in this run"),
        ("batch_firecrawl_status",       fc_audit.get("batch_firecrawl_status", ""),
         "Overall batch outcome: ok / degraded / fail_fast / not_used"),
        ("batch_firecrawl_notes",        fc_audit.get("batch_firecrawl_notes", ""),
         "Human-readable notes or fail-fast reason"),
    ]

    for ri, (field, value, desc) in enumerate(_rows, 2):
        ws.cell(row=ri, column=1, value=field)
        ws.cell(row=ri, column=2, value=str(value) if value != "" else "")
        ws.cell(row=ri, column=3, value=desc)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 60
    ws.freeze_panes = "A2"


def build_excel(
    enriched_df: pd.DataFrame,
    original_df: pd.DataFrame,
    evidence_rows: list[dict],
    cols: dict,
    debug_rows: list[dict] | None = None,
    debug_mode: bool = False,
    run_meta: dict | None = None,
    jina_debug_rows: list[dict] | None = None,
    fc_audit: dict | None = None,
) -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()

    # Pre-compute eligibility masks so Commercial Input can go in position 2
    _pf_col     = "myngle_target_eligibility"
    _keep_mask  = enriched_df.get(_pf_col, pd.Series("", index=enriched_df.index)).astype(str) == _ELI_KEEP
    _maybe_mask = enriched_df.get(_pf_col, pd.Series("", index=enriched_df.index)).astype(str) == _ELI_MAYBE
    _excl_mask  = enriched_df.get(_pf_col, pd.Series("", index=enriched_df.index)).astype(str) == _ELI_EXCLUDE

    _pf_key_cols = [
        "organization_type", "myngle_target_eligibility",
        "pre_filter_decision", "pre_filter_reason",
        "cleaned_company_name", "normalized_input_website",
        "email_domain", "domain_action", "final_selected_domain",
        "final_confidence", "final_decision_source",
    ]

    def _make_pf_sheet(ws, df_subset):
        if df_subset.empty:
            ws.cell(row=1, column=1, value="No rows in this category.")
            return
        out_cols = [c for c in _pf_key_cols if c in df_subset.columns]
        _write_sheet(ws, df_subset[out_cols])

    # Sheet 1: Best Guess Input
    ws0 = wb.active
    ws0.title = "Best Guess Input"
    bg_df = _build_best_guess_df(enriched_df, cols)
    _write_best_guess_sheet(ws0, bg_df, enriched_df)

    # Sheet 2: Commercial Input (eligibility keep — needed early for Lead Prioritizer)
    ws_comm = wb.create_sheet("Commercial Input")
    _make_pf_sheet(ws_comm, enriched_df[_keep_mask])

    # Sheet 3: Cleaned Register Input
    ws1 = wb.create_sheet("Cleaned Register Input")
    _write_sheet(ws1, enriched_df)

    # Sheet 3: Review Needed
    ws2 = wb.create_sheet("Review Needed")
    review_mask = (
        enriched_df.get("manual_review_needed", pd.Series(False, index=enriched_df.index))
        .astype(str).str.lower().isin(["true", "1", "yes"])
    )
    review_df = enriched_df[review_mask] if review_mask.any() else enriched_df.iloc[:0]
    _write_sheet(ws2, review_df.copy())

    # Sheet 4: Original Input
    ws3 = wb.create_sheet("Original Input")
    _write_sheet(ws3, original_df)

    # Sheet 5: Raw Search Evidence
    ws4 = wb.create_sheet("Raw Search Evidence")
    ev_df = pd.DataFrame(evidence_rows) if evidence_rows else pd.DataFrame(columns=[
        "company_name", "city", "province", "search_query_used",
        "serper_top_title", "serper_top_url", "serper_top_domain",
        "validated_domain", "domain_source", "domain_action", "domain_confidence",
        "name_variant_used", "top_3_candidates", "rejection_reason", "discovery_method",
    ])
    _write_sheet(ws4, ev_df)

    # Sheet 6: Python vs Haiku Comparison (only shown when Haiku was run)
    haiku_cols = [
        cols.get("company"),
        "validated_domain", "domain_confidence", "domain_action",
        "haiku_used", "haiku_decision", "haiku_domain", "haiku_confidence",
        "haiku_reason", "haiku_risk_flags", "haiku_error",
        "final_selected_domain", "final_decision_source", "final_confidence",
    ]
    avail_haiku_cols = [c for c in haiku_cols if c and c in enriched_df.columns]
    haiku_ran = (
        "haiku_used" in enriched_df.columns
        and enriched_df["haiku_used"].astype(str).str.lower().isin(["true", "1"]).any()
    )
    ws5 = wb.create_sheet("Python vs Haiku Comparison")
    if haiku_ran and avail_haiku_cols:
        _write_sheet(ws5, enriched_df[avail_haiku_cols].copy())
    else:
        ws5.cell(row=1, column=1, value="Haiku review was not used in this run.")

    # Sheet 7: Candidate Rank Diagnostics (always present when debug_rows available)
    ws_rank = wb.create_sheet("Candidate Rank Diagnostics")
    if debug_rows:
        rank_df = _build_rank_diagnostics(debug_rows, enriched_df, cols)
        if not rank_df.empty:
            _write_sheet(ws_rank, rank_df)
        else:
            ws_rank.cell(row=1, column=1, value="No rank diagnostics available.")
    else:
        ws_rank.cell(row=1, column=1, value="No debug rows collected (debug mode was off).")

    # Sheet 8: Candidate Discovery Debug (only when debug_mode is active)
    if debug_mode:
        ws6 = wb.create_sheet("Candidate Discovery Debug")
        _debug_cols = [
            "company_name", "row_number", "search_query", "result_rank",
            "title", "snippet", "url", "extracted_domain",
            "candidate_url", "candidate_type", "candidate_source", "candidate_path",
            "score", "used", "selected_candidate", "selected_best_candidate",
            "skip_reason", "rejection_category",
            "brand_overlap", "full_overlap", "location_match", "email_match",
            "official_signal", "final_python_domain",
            "haiku_mode", "haiku_decision", "haiku_candidate_url", "haiku_recommended_action",
            "final_selected_domain",
        ]
        if debug_rows:
            debug_df = pd.DataFrame(debug_rows)
            for c in _debug_cols:
                if c not in debug_df.columns:
                    debug_df[c] = ""
            _write_sheet(ws6, debug_df[_debug_cols])
        else:
            ws6.cell(row=1, column=1,
                     value="Debug mode was enabled but no Serper results were collected "
                           "(no Serper key, or no rows required search).")

    # Manual Review Queue sheet
    _mrq_cols = [
        cols.get("company") or "company_name",
        cols.get("city") or "city",
        cols.get("province") or "province",
        cols.get("website") or "website_url",
        "final_selected_domain", "final_confidence", "final_decision_source",
        "domain_action", "domain_confidence",
        "final_domain_is_business_output", "business_output_reason",
        "manual_review_needed",
        "verification_needed", "verification_reason",
        "verifier_decision", "verifier_reason",
        "verifier_evidence_strength", "verifier_negative_source_type",
        "original_candidate_domain", "redirect_final_domain", "redirect_final_url",
        "redirect_canonical_override_applied",
        "wrong_entity_type_signal", "wrong_location_signal",
        "python_validated_domain", "python_recommended_domain",
        "firecrawl_original_checked_domain",
        "firecrawl_decision", "firecrawl_reason", "firecrawl_fetch_status",
        "verifier_evidence_url", "firecrawl_evidence_url",
        "organization_type", "myngle_target_eligibility",
        "top_3_candidate_domains", "serper_top_result_title", "serper_top_result_url",
    ]
    ws_mrq = wb.create_sheet("Manual Review Queue")
    _mrq_mask = (
        enriched_df.get("manual_review_needed", pd.Series(False, index=enriched_df.index))
        .astype(str).str.lower().isin(["true", "1", "yes"])
    )
    _mrq_df = enriched_df[_mrq_mask].copy() if _mrq_mask.any() else enriched_df.iloc[:0].copy()
    if not _mrq_df.empty:
        _avail_mrq = [c for c in _mrq_cols if c and c in _mrq_df.columns]
        _write_sheet(ws_mrq, _mrq_df[_avail_mrq])
    else:
        ws_mrq.cell(row=1, column=1, value="No rows require manual review.")

    # Website Verification Debug sheet (always shown when any verifier was used)
    jina_ran = (
        "jina_verifier_used" in enriched_df.columns
        and enriched_df["jina_verifier_used"].astype(str).str.lower().isin(["true", "1"]).any()
    )
    _verif_ran = (
        "verifier_used" in enriched_df.columns
        and enriched_df["verifier_used"].astype(str).str.lower().isin(["true", "1"]).any()
    )
    if jina_ran or _verif_ran or jina_debug_rows:
        ws_jina = wb.create_sheet("Website Verification Debug")
        _verif_debug_cols = [
            "company_name", "candidate_domain",
            "candidate_url", "candidate_type", "candidate_source", "candidate_path",
            "original_candidate_url", "page_url",
            "page_type", "redirect_final_url", "redirect_final_domain",
            "canonical_domain_used", "fetch_status", "chars_fetched", "elapsed_secs",
            "source_type", "wrong_entity_type_signal", "wrong_location_signal",
            "evidence_strength", "negative_source_type",
            "verifier_decision", "verifier_reason", "replace_allowed",
            "haiku_decision", "haiku_confidence",
            "firecrawl_key_index_used", "firecrawl_key_failover_count", "firecrawl_key_statuses",
            "third_party_evidence_flag",
        ]
        if jina_debug_rows:
            jd_df = pd.DataFrame(jina_debug_rows)
            for c in _verif_debug_cols:
                if c not in jd_df.columns:
                    jd_df[c] = ""
            _write_sheet(ws_jina, jd_df[_verif_debug_cols])
        else:
            ws_jina.cell(row=1, column=1, value="Verifier ran but no debug rows were collected.")

    # Validation Diagnostics sheet — always present
    ws_val = wb.create_sheet("Validation Diagnostics")
    val_summary = _build_validation_summary(enriched_df, debug_rows or [], cols)
    import openpyxl.styles as _oxl_styles
    _key_font_val = _oxl_styles.Font(bold=True)
    ws_val.cell(row=1, column=1, value="Metric").font = _key_font_val
    ws_val.cell(row=1, column=2, value="Value").font  = _key_font_val
    for ri, (k, v) in enumerate(val_summary.items(), 2):
        ws_val.cell(row=ri, column=1, value=str(k))
        ws_val.cell(row=ri, column=2, value=str(v) if v is not None else "")
    ws_val.column_dimensions["A"].width = 42
    ws_val.column_dimensions["B"].width = 14
    ws_val.freeze_panes = "A2"

    # ── Organization eligibility pre-filter sheets ───────────────────────────
    # Commercial Input already written at sheet 2 above; write Maybe / Excluded here.
    ws_maybe_sheet = wb.create_sheet("Maybe Review")
    _make_pf_sheet(ws_maybe_sheet, enriched_df[_maybe_mask])

    ws_excl = wb.create_sheet("Excluded Organizations")
    _make_pf_sheet(ws_excl, enriched_df[_excl_mask])

    # Pre-filter Summary
    ws_pfs = wb.create_sheet("Pre-filter Summary")
    _pf_rows = []
    if _pf_col in enriched_df.columns and "organization_type" in enriched_df.columns:
        for (ot, eli, dec), grp in enriched_df.groupby(
            ["organization_type", "myngle_target_eligibility", "pre_filter_decision"],
            dropna=False,
        ):
            reasons = enriched_df.loc[grp.index, "pre_filter_reason"].value_counts()
            top_reason = reasons.index[0] if len(reasons) else ""
            _pf_rows.append({
                "organization_type": ot,
                "eligibility": eli,
                "pre_filter_decision": dec,
                "count": len(grp),
                "most_common_reason": top_reason,
            })
    if _pf_rows:
        _write_sheet(ws_pfs, pd.DataFrame(_pf_rows))
    else:
        ws_pfs.cell(row=1, column=1, value="Pre-filter columns not present in output.")

    # Firecrawl Audit sheet (always present when fc_audit provided)
    if fc_audit is not None:
        ws_fc_audit = wb.create_sheet("Firecrawl Audit")
        _xl_write_fc_audit(ws_fc_audit, fc_audit)

    # API Usage Summary — always present when run_meta available
    if run_meta:
        ws_usage = wb.create_sheet("API Usage Summary")
        _xl_write_api_usage_summary(ws_usage, run_meta)

    # Run Summary — always last
    ws_summary = wb.create_sheet("Run Summary")
    if run_meta:
        _write_run_summary_sheet(ws_summary, run_meta)
    else:
        ws_summary.cell(row=1, column=1, value="Run metadata not available.")

    # ── QA print ──────────────────────────────────────────────────────────────
    _qa_sheets = wb.sheetnames
    _qa_first_cols = list(bg_df.columns[:10]) if not bg_df.empty else []
    _required_downstream = [
        "canonical_company_name", "canonical_company_url",
        "canonical_company_domain", "source_contact_count", "input_type",
    ]
    _bg_col_set = set(bg_df.columns)
    _present    = [c for c in _required_downstream if c in _bg_col_set]
    _missing    = [c for c in _required_downstream if c not in _bg_col_set]
    print("[QA] Sheet names:          ", _qa_sheets)
    print("[QA] Best Guess row count: ", len(bg_df))
    print("[QA] First 10 columns:     ", _qa_first_cols)
    print("[QA] Required cols present:", _present)
    print("[QA] Required cols missing:", _missing)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================================
# SUMMARY METRICS
# =============================================================================


def _summary_metrics(df: pd.DataFrame, cols: dict) -> None:
    actions  = df.get("domain_action",  pd.Series(dtype=str)).astype(str)
    sources  = df.get("domain_source",  pd.Series(dtype=str)).astype(str)
    confs    = df.get("domain_confidence", pd.Series(dtype=str)).astype(str)

    total     = len(df)
    has_web   = int(df.get(cols.get("website") or "_", pd.Series("")).astype(str)
                    .str.strip().replace("", pd.NA).notna().sum()) if cols.get("website") else 0
    has_email = int(df.get(cols.get("email") or "_", pd.Series("")).astype(str)
                    .str.strip().replace("", pd.NA).notna().sum()) if cols.get("email") else 0
    has_phone = int(df.get(cols.get("phone") or "_", pd.Series("")).astype(str)
                    .str.strip().replace("", pd.NA).notna().sum()) if cols.get("phone") else 0

    accepted        = int(actions.isin(["OK", "LIKELY_OK"]).sum())
    from_email      = int(sources.isin([SRC_EMAIL]).sum())
    serper_conf_email = int(sources.isin([SRC_SERPER_EMAIL]).sum())
    from_serper     = int(sources.isin([SRC_SERPER]).sum())
    no_match        = int(actions.isin(["MISSING_DOMAIN", "NO_CONFIDENT_MATCH"]).sum())

    # website found = any non-empty validated_domain
    has_domain_after = int(
        df.get("validated_domain", pd.Series(dtype=str)).astype(str)
        .str.strip().replace("", pd.NA).notna().sum()
    )
    review = int(
        df.get("manual_review_needed", pd.Series(dtype=str))
        .astype(str).str.lower().isin(["true", "1", "yes"]).sum()
    )
    coverage_pct = round(has_domain_after / total * 100) if total else 0

    def card(col, label, val, color, hint=""):
        col.markdown(
            f"<div style='border-left:4px solid {color};padding:8px 12px;"
            f"background:#f7f9fc;border-radius:4px;margin-bottom:6px'>"
            f"<div style='font-size:1.45em;font-weight:700;color:{color}'>{val}</div>"
            f"<div style='font-size:0.78em;color:#555'>{label}</div>"
            + (f"<div style='font-size:0.72em;color:#888'>{hint}</div>" if hint else "")
            + "</div>",
            unsafe_allow_html=True,
        )

    # Row 1 — input data
    row1 = st.columns(4)
    card(row1[0], "Total companies",       total,      "#0B4A92")
    card(row1[1], "With original website", has_web,    "#2E7D32",
         f"{round(has_web/total*100) if total else 0}% of rows")
    card(row1[2], "With email",            has_email,  "#1565C0")
    card(row1[3], "With phone",            has_phone,  "#37474F")

    st.markdown("")
    # Row 2 — discovery outcome
    row2 = st.columns(4)
    card(row2[0], "Website found after cleaning", has_domain_after, "#2E7D32",
         f"{coverage_pct}% coverage")
    card(row2[1], "Original website accepted",    accepted,          "#43A047",
         "OK + LIKELY_OK")
    card(row2[2], "Email domain used",            from_email + serper_conf_email, "#1565C0",
         f"{from_email} proxy · {serper_conf_email} Serper-confirmed")
    card(row2[3], "Found by Serper search",       from_serper,       "#E65100")

    st.markdown("")
    # Row 3 — review / gaps
    row3 = st.columns(4)
    card(row3[0], "Need manual review",    review,    "#B71C1C")
    card(row3[1], "No confident match",    no_match,  "#C62828",
         "MISSING_DOMAIN or NO_CONFIDENT_MATCH")
    card(row3[2], "Serper-confirmed email domain", serper_conf_email, "#6A1B9A",
         "strongest email signal")
    card(row3[3], "High confidence rows",
         int(confs.str.lower().eq("high").sum()),
         "#2E7D32",
         "manual_review_needed = False")

    # Row 3b — Haiku stats (only shown when Haiku was used)
    haiku_used_col = df.get("haiku_used", pd.Series(dtype=str)).astype(str).str.lower()
    n_haiku = int(haiku_used_col.isin(["true", "1"]).sum())
    if n_haiku > 0:
        st.markdown("")
        row3b = st.columns(4)
        decisions = df.get("haiku_decision", pd.Series(dtype=str)).astype(str)
        sources   = df.get("final_decision_source", pd.Series(dtype=str)).astype(str)
        card(row3b[0], "Rows reviewed by Haiku",      n_haiku,                             "#7B1FA2")
        card(row3b[1], "Haiku: accepted Python",       int(decisions.eq("accept").sum()),  "#2E7D32",
             "agreed with Python scoring")
        card(row3b[2], "Haiku: replaced domain",       int(decisions.eq("replace").sum()), "#E65100",
             "found better domain than Python")
        card(row3b[3], "Haiku: rejected / uncertain",
             int(decisions.isin(["reject", "uncertain"]).sum()),                            "#B71C1C",
             "no confident match")

    # Row 3c — Jina stats (only shown when Jina was used)
    jina_used_col = df.get("jina_verifier_used", pd.Series(dtype=str)).astype(str).str.lower()
    n_jina = int(jina_used_col.isin(["true", "1"]).sum())
    if n_jina > 0:
        st.markdown("")
        row3c = st.columns(4)
        jina_dec = df.get("jina_verifier_decision", pd.Series(dtype=str)).astype(str)
        card(row3c[0], "Rows verified by Jina",        n_jina,                                 "#00796B")
        card(row3c[1], "Jina: confirmed Python",        int(jina_dec.eq("confirm").sum()),      "#2E7D32",
             "Jina evidence matches Python selection")
        card(row3c[2], "Jina: replaced domain",         int(jina_dec.eq("replace").sum()),      "#E65100",
             "Jina found stronger alternative")
        card(row3c[3], "Jina: rejected / uncertain",
             int(jina_dec.isin(["reject", "uncertain", "fetch_failed"]).sum()),                  "#B71C1C",
             "no confident Jina evidence")

    # Row — FC / Serper usage (shown whenever there is any FC usage)
    _fc_pages_dv   = int(pd.to_numeric(df.get("firecrawl_pages_fetched",
        pd.Series(dtype=int)), errors="coerce").fillna(0).sum())
    _fc_pages_si   = int(pd.to_numeric(df.get("firecrawl_pages_used_for_size",
        pd.Series(dtype=int)), errors="coerce").fillna(0).sum())
    _fc_total      = _fc_pages_dv
    _fc_credits    = _fc_total + _fc_pages_si  # total estimated credits
    _fc_rows       = int(df.get("firecrawl_used", pd.Series(dtype=str)).astype(str)
                         .str.lower().isin(["true","1"]).sum())
    _avg_fc_row    = round(_fc_credits / _fc_rows, 2) if _fc_rows else 0.0
    _avg_fc_proc   = round(_fc_credits / total, 2)    if total    else 0.0
    _si_enabled    = "firecrawl_pages_used_for_size" in df.columns and _fc_pages_si > 0

    st.markdown("")
    _row_usage = st.columns(4)
    card(_row_usage[0], "Est. Firecrawl credits", _fc_credits, "#0277BD",
         "1 credit per fetched page; exact billing not returned by API")
    card(_row_usage[1], "FC pages · domain verify", _fc_pages_dv, "#01579B",
         "pages used for website verification")
    if _si_enabled:
        card(_row_usage[2], "FC pages · size inference", _fc_pages_si, "#006064",
             "new FC pages for employee-size inference")
        card(_row_usage[3], f"FC avg / row ({_fc_rows} FC rows)", f"{_avg_fc_row:.2f}", "#0277BD",
             f"avg {_avg_fc_proc:.2f} pages per all processed rows")
    else:
        card(_row_usage[2], f"FC avg / row ({_fc_rows} FC rows)", f"{_avg_fc_row:.2f}", "#0277BD",
             f"avg {_avg_fc_proc:.2f} pages per all processed rows")
        card(_row_usage[3], "Size inference", "disabled", "#78909C")

    st.markdown("")
    # Row 4 — false-positive rejections (sum across all rows)
    def _sum_col(col_name):
        col = df.get(col_name, pd.Series(0, index=df.index))
        return int(pd.to_numeric(col, errors="coerce").fillna(0).sum())

    row4 = st.columns(5)
    card(row4[0], "Rejected: directory",      _sum_col("rejected_directory"),      "#5D4037",
         "oraridiapertura, pagine gialle, etc.")
    card(row4[1], "Rejected: government",     _sum_col("rejected_government"),     "#37474F",
         ".gov.it, agenziaentrate, comune, etc.")
    card(row4[2], "Rejected: religious",      _sum_col("rejected_religious"),      "#6A1B9A",
         "basilica, diocesi, parrocchia, etc.")
    card(row4[3], "Rejected: academic",       _sum_col("rejected_academic"),       "#1565C0",
         ".edu, università, politecnico, etc.")
    card(row4[4], "Rejected: low similarity", _sum_col("rejected_low_similarity"), "#E65100",
         "domain unrelated to company name")


# =============================================================================
# STREAMLIT UI
# =============================================================================


def _build_fc_usage_fields(
    enriched_df: pd.DataFrame,
    firecrawl_keys_loaded: int,
    fc_health: dict,
) -> dict:
    """
    Build Firecrawl credit/usage fields for the Run Summary sheet.
    Uses runtime health counters when available; falls back to per-row df columns.
    Never exposes actual API key values.
    """
    # Pages fetched — from df column (ground truth, same as before)
    pages_total = int(
        pd.to_numeric(
            enriched_df.get("firecrawl_pages_fetched", pd.Series(dtype=int)), errors="coerce"
        ).fillna(0).sum()
    )
    # Firecrawl API does not return exact credit cost in its v1 scrape response.
    # We estimate: 1 credit per page fetched (Firecrawl's standard pricing unit).
    estimated_credits = pages_total
    credit_method = (
        "Estimated as 1 credit per fetched page; "
        "exact Firecrawl billing not returned by API"
    )

    # Runtime health counters (populated when _make_fc_health() was used)
    att      = fc_health.get("requests_attempted", 0)
    succ     = fc_health.get("pages_successful",   0)
    timeouts = fc_health.get("timeouts", 0)
    excepts  = fc_health.get("exceptions", 0)
    failovers= fc_health.get("key_failovers", 0)
    failed   = att - succ if att > 0 else 0

    # Firecrawl decisions (from enriched df)
    fc_dec = enriched_df.get("firecrawl_decision", pd.Series(dtype=str)).astype(str)
    fc_used = enriched_df.get("firecrawl_used", pd.Series(dtype=str)).astype(str).str.lower().isin(["true","1"])

    # Per-key usage (from firecrawl_key_statuses column — "key1:ok; key2:http_402" etc.)
    _per_key: dict[str, dict] = {}
    if "firecrawl_key_statuses" in enriched_df.columns:
        for _ks_str in enriched_df["firecrawl_key_statuses"].astype(str):
            for _part in _ks_str.split(";"):
                _part = _part.strip()
                if ":" not in _part:
                    continue
                _klabel, _kst = _part.split(":", 1)
                _klabel = _klabel.strip()
                _kst    = _kst.strip()
                if _klabel not in _per_key:
                    _per_key[_klabel] = {"requests": 0, "successes": 0, "failures": 0}
                _per_key[_klabel]["requests"] += 1
                if _kst == "ok":
                    _per_key[_klabel]["successes"] += 1
                else:
                    _per_key[_klabel]["failures"] += 1

    # Purpose-split from per-row df columns (size_inference vs domain_verification)
    # domain_verification pages = firecrawl_pages_fetched (set by the verifier path)
    dv_pages  = int(pd.to_numeric(enriched_df.get("firecrawl_pages_fetched",
        pd.Series(dtype=int)), errors="coerce").fillna(0).sum())
    # size_inference pages = new FC fetches made during size inference
    si_new    = int(pd.to_numeric(enriched_df.get("firecrawl_pages_used_for_size",
        pd.Series(dtype=int)), errors="coerce").fillna(0).sum())
    si_reused = int(pd.to_numeric(enriched_df.get("firecrawl_size_inference_reused_pages",
        pd.Series(dtype=int)), errors="coerce").fillna(0).sum())
    # Total new pages = DV + SI new (reused pages cost 0 credits)
    total_new_pages    = dv_pages + si_new
    total_credits      = total_new_pages  # 1 credit per new fetched page
    # rows where size inference ran
    si_rows_col = enriched_df.get("firecrawl_used_for_size_inference", pd.Series(dtype=str))
    si_rows = int(si_rows_col.astype(str).str.lower().isin(["true","1"]).sum())
    fc_rows = int(fc_used.sum())
    _proc = len(enriched_df)
    _avg_pages_processed = round(total_credits / _proc, 2) if _proc else 0.0
    _avg_pages_fc_row    = round(total_credits / fc_rows, 2) if fc_rows else 0.0
    _avg_pages_dv_row    = round(dv_pages      / fc_rows, 2) if fc_rows else 0.0
    _avg_pages_si_row    = round(si_new        / si_rows, 2) if si_rows else 0.0

    # Total requests from health (includes both DV and SI since _infer_company_size now updates it)
    total_att  = att if att > 0 else total_new_pages
    total_succ = succ if att > 0 else total_new_pages

    # DV requests ≈ DV pages (each page is one Firecrawl API call)
    dv_att  = dv_pages   # pages fetched = API calls for domain verification
    dv_succ = dv_pages   # all fetched pages count as successful
    # SI requests from health si_* sub-counters (or fall back to column)
    si_att  = fc_health.get("si_requests_attempted", si_new)
    si_succ = fc_health.get("si_pages_successful",   si_new)

    fields: dict = {
        # ── Canonical purpose-split fields (the authoritative numbers) ─────────
        "firecrawl_domain_verification_requests_attempted":  dv_att,
        "firecrawl_domain_verification_pages_successful":    dv_succ,
        "firecrawl_domain_verification_estimated_credits":   dv_succ,
        "firecrawl_size_inference_requests_attempted":       si_att,
        "firecrawl_size_inference_pages_successful_new":     si_succ,
        "firecrawl_size_inference_pages_reused":             si_reused,
        "firecrawl_size_inference_estimated_credits":        si_succ,
        # ── Canonical totals ───────────────────────────────────────────────────
        "firecrawl_total_requests_attempted":                dv_att + si_att,
        "firecrawl_total_pages_successful_new":              total_new_pages,
        "firecrawl_total_estimated_credits":                 total_credits,
        "firecrawl_credit_estimation_method":                credit_method,
        # ── Row-level counts (rows, not pages) ────────────────────────────────
        "firecrawl_rows_attempted":                          fc_rows,
        "firecrawl_rows_with_successful_page":               int(fc_dec.isin(["confirm","replace","replace_domain"]).sum()),
        "firecrawl_rows_confirmed":                          int(fc_dec.eq("confirm").sum()),
        "firecrawl_rows_uncertain":                          int(fc_dec.eq("uncertain").sum()),
        "firecrawl_rows_failed":                             int(fc_dec.isin(["failed", "error", "no_fetch"]).sum()),
        "firecrawl_size_inference_rows":                     si_rows,
        # ── Runtime health (for debugging) ────────────────────────────────────
        "firecrawl_runtime_requests_attempted":              total_att,
        "firecrawl_runtime_pages_successful":                total_succ,
        "firecrawl_runtime_requests_failed":                 failed if att > 0 else 0,
        "firecrawl_timeouts":                                timeouts,
        "firecrawl_exceptions":                              excepts,
        "firecrawl_key_failovers_total":                     failovers,
        "firecrawl_keys_loaded":                             firecrawl_keys_loaded,
        # ── Averages ───────────────────────────────────────────────────────────
        "firecrawl_avg_pages_per_processed_row":             _avg_pages_processed,
        "firecrawl_avg_pages_per_firecrawl_used_row":        _avg_pages_fc_row,
        "firecrawl_avg_pages_per_domain_verification_row":   _avg_pages_dv_row,
        "firecrawl_avg_pages_per_size_inference_row":        _avg_pages_si_row,
        # ── Legacy aliases (backward compat) ──────────────────────────────────
        "firecrawl_domain_verification_pages":               dv_pages,
        "firecrawl_size_inference_pages_new":                si_new,
        "firecrawl_size_inference_reused_pages":             si_reused,
        "firecrawl_estimated_credits_used":                  total_credits,
        "firecrawl_pages_fetched_total":                     dv_pages,
    }
    # Per-key usage (key labels only — no actual key values)
    for _klabel in sorted(_per_key.keys()):
        _kd = _per_key[_klabel]
        fields[f"firecrawl_{_klabel}_requests"]  = _kd["requests"]
        fields[f"firecrawl_{_klabel}_successes"] = _kd["successes"]
        fields[f"firecrawl_{_klabel}_failures"]  = _kd["failures"]

    return fields


def _xl_write_api_usage_summary(ws, run_meta: dict) -> None:
    """Write the API Usage Summary sheet — provider breakdown, credits, key usage."""
    from openpyxl.styles import Alignment, Font, PatternFill

    hdr_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    sec_font = Font(bold=True, size=10, color="1F497D")
    val_font = Font(size=10)

    def _hdr(row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="left")

    def _sec(row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.font = sec_font

    def _row(ws_row, label, value):
        a = ws.cell(row=ws_row, column=1, value=label)
        a.font = Font(bold=True, size=10)
        b = ws.cell(row=ws_row, column=2, value=str(value) if value is not None else "")
        b.font = val_font

    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 60
    ws.freeze_panes = "A2"

    ri = 1
    _hdr(ri, 1, "Metric")
    _hdr(ri, 2, "Value")
    ri += 1

    # ── Firecrawl section ─────────────────────────────────────────────────────
    _sec(ri, 1, "── Firecrawl ──")
    ri += 1
    _fc_fields = [
        ("firecrawl_total_estimated_credits",                 "Total estimated credits (DV + SI new pages)"),
        ("firecrawl_credit_estimation_method",                "Credit estimation method"),
        ("firecrawl_total_pages_successful_new",              "Total new pages fetched (DV + SI)"),
        ("firecrawl_domain_verification_pages",               "Domain verification pages"),
        ("firecrawl_domain_verification_estimated_credits",   "Domain verification estimated credits"),
        ("firecrawl_size_inference_pages_new",                "Size inference pages (new fetches)"),
        ("firecrawl_size_inference_reused_pages",             "Size inference pages (reused, 0 credits)"),
        ("firecrawl_size_inference_estimated_credits",        "Size inference estimated credits"),
        ("firecrawl_size_inference_rows",                     "Rows with size inference FC"),
        ("firecrawl_avg_pages_per_processed_row",             "Avg pages / processed row"),
        ("firecrawl_avg_pages_per_firecrawl_used_row",        "Avg pages / FC-used row"),
        ("firecrawl_avg_pages_per_domain_verification_row",   "Avg pages / domain-verification row"),
        ("firecrawl_avg_pages_per_size_inference_row",        "Avg pages / size-inference row"),
        ("firecrawl_requests_attempted",                      "Requests attempted (runtime)"),
        ("firecrawl_pages_successful",                        "Pages successful (runtime)"),
        ("firecrawl_requests_failed",                         "Requests failed"),
        ("firecrawl_timeouts",                                "Timeouts"),
        ("firecrawl_exceptions",                              "Exceptions"),
        ("firecrawl_key_failovers_total",                     "Key failovers (total)"),
        ("firecrawl_keys_loaded",                             "Keys loaded"),
        ("firecrawl_rows_attempted",                          "Rows attempted"),
        ("firecrawl_rows_with_successful_page",               "Rows with successful page"),
        ("firecrawl_rows_confirmed",                          "Rows confirmed (domain verification)"),
        ("firecrawl_rows_uncertain",                          "Rows uncertain"),
        ("firecrawl_rows_failed",                             "Rows failed"),
    ]
    for _key, _label in _fc_fields:
        val = run_meta.get(_key, "")
        if val == "" and _key.startswith("firecrawl_avg"):
            val = "0.00"
        _row(ri, _label, val)
        ri += 1

    # Per-key usage (key labels only — no actual key values)
    _per_key_rows = [(k, v) for k, v in run_meta.items() if k.startswith("firecrawl_key") and
                     any(k.endswith(s) for s in ("_requests", "_successes", "_failures"))]
    if _per_key_rows:
        ri += 1
        _sec(ri, 1, "── Firecrawl per-key usage (labels only, no actual key values) ──")
        ri += 1
        for _key, _val in sorted(_per_key_rows):
            _label = _key.replace("firecrawl_", "").replace("_", " ").title()
            _row(ri, _label, _val)
            ri += 1

    # ── Serper section ────────────────────────────────────────────────────────
    ri += 1
    _sec(ri, 1, "── Serper ──")
    ri += 1
    _serper_fields = [
        ("serper_queries_total",                  "Queries total (actual API calls)"),
        ("serper_avg_queries_per_processed_row",  "Avg queries / processed row"),
        ("serper_avg_queries_per_serper_row",     "Avg queries / row where Serper ran"),
        ("max_serper_queries",                    "Max queries per company (setting)"),
        ("serper_key_present",                    "Serper key present"),
    ]
    for _key, _label in _serper_fields:
        val = run_meta.get(_key, "")
        if val == "" and "avg" in _key:
            val = "0.00"
        _row(ri, _label, val)
        ri += 1

    # ── Claude Haiku section ──────────────────────────────────────────────────
    ri += 1
    _sec(ri, 1, "── Claude Haiku ──")
    ri += 1
    _haiku_fields = [
        ("haiku_calls_total",               "Calls total"),
        ("haiku_avg_calls_per_processed_row", "Avg calls / processed row"),
        ("rows_reviewed_by_haiku",          "Rows reviewed by Haiku"),
        ("haiku_accepted_python",           "Haiku accepted Python decision"),
        ("haiku_replaced_python",           "Haiku replaced Python decision"),
        ("haiku_rejected",                  "Haiku rejected domain"),
        ("haiku_uncertain",                 "Haiku uncertain"),
        ("haiku_mode",                      "Haiku mode"),
        ("haiku_model",                     "Haiku model"),
        ("anthropic_key_present",           "Anthropic key present"),
    ]
    for _key, _label in _haiku_fields:
        val = run_meta.get(_key, "")
        if val == "" and "avg" in _key:
            val = "0.00"
        _row(ri, _label, val)
        ri += 1

    ws.row_dimensions[1].height = 18


def _fc_usage_console_summary(run_meta: dict) -> str:
    """Return a multi-line usage summary for CLI output (Serper + FC + Haiku)."""
    _proc = run_meta.get("processed_rows", 0) or 1
    _fc_pages = run_meta.get("firecrawl_pages_fetched_total", 0)
    _fc_cred  = run_meta.get("firecrawl_estimated_credits_used", 0)
    _fc_dv    = run_meta.get("firecrawl_domain_verification_pages", 0)
    _fc_si    = run_meta.get("firecrawl_size_inference_pages_new", 0)
    _si_rows  = run_meta.get("firecrawl_size_inference_rows", 0)
    _fc_rows  = run_meta.get("firecrawl_rows_attempted", 0)
    lines = [
        "-- Usage summary -----------------------------------------------",
        f"  Rows processed:             {run_meta.get('processed_rows', 0)}",
        f"  Serper queries total:       {run_meta.get('serper_queries_total', 'n/a')}",
        f"  Serper avg / processed row: {run_meta.get('serper_avg_queries_per_processed_row', 'n/a')}",
        f"  Haiku calls total:          {run_meta.get('haiku_calls_total', 0)}",
        f"  Haiku avg / processed row:  {run_meta.get('haiku_avg_calls_per_processed_row', 0.0):.3f}",
        "  -------------------------------------------------------------",
        f"  FC estimated credits total: {_fc_cred}",
        f"  FC pages total:             {_fc_pages}",
        f"  FC avg pages / processed:   {run_meta.get('firecrawl_avg_pages_per_processed_row', 0.0):.2f}",
        f"  FC avg pages / FC row:      {run_meta.get('firecrawl_avg_pages_per_firecrawl_used_row', 0.0):.2f}  ({_fc_rows} FC rows)",
        f"  FC domain verification:     {_fc_dv} pages",
        f"  FC size inference (new):    {_fc_si} pages  ({_si_rows} rows)",
        f"  FC size inference (reused): {run_meta.get('firecrawl_size_inference_reused_pages', 0)} pages",
        f"  FC requests attempted:      {run_meta.get('firecrawl_requests_attempted', 0)}",
        f"  FC timeouts:                {run_meta.get('firecrawl_timeouts', 0)}",
        f"  FC key failovers:           {run_meta.get('firecrawl_key_failovers_total', 0)}",
        f"  FC keys loaded:             {run_meta.get('firecrawl_keys_loaded', 0)}",
        f"  Note: Estimated credits = fetched pages. Exact billing not returned by FC API.",
        "----------------------------------------------------------------",
    ]
    return "\n".join(lines)


def _build_run_meta(
    enriched_df: pd.DataFrame,
    input_filename: str,
    run_id: str,
    run_label: str,
    total_rows_input: int,
    batch_n: int,
    max_queries: int,
    haiku_mode: str,
    haiku_model: str,
    haiku_max_rows: int,
    debug_mode: bool,
    serper_key_present: bool,
    anthropic_key_present: bool,
    jina_mode: str = _JINA_MODE_UNCERTAIN,
    verifier_provider: str = _VP_OFF,
    verifier_mode: str = _VM_UNCERTAIN,
    firecrawl_keys_loaded: int = 0,
    fc_health: dict | None = None,
    serper_queries_total: int = 0,
    detected_country: str = "",
    country_detection_source: str = "",
    firecrawl_location_mode: str = "",
    firecrawl_location_used: str = "",
    firecrawl_location_country: str = "",
    firecrawl_location_languages: str = "",
) -> dict:
    """Build the ordered dict that populates the Run Summary Excel sheet."""
    actions  = enriched_df.get("domain_action",  pd.Series(dtype=str)).astype(str)
    has_final = int(
        enriched_df.get("final_selected_domain", enriched_df.get("validated_domain",
            pd.Series(dtype=str))).astype(str).str.strip().replace("", pd.NA).notna().sum()
    )
    coverage = round(has_final / batch_n * 100) if batch_n else 0
    decisions = enriched_df.get("haiku_decision", pd.Series(dtype=str)).astype(str)
    n_haiku   = int(
        enriched_df.get("haiku_used", pd.Series(dtype=str)).astype(str)
        .str.lower().isin(["true", "1"]).sum()
    )
    n_jina = int(
        enriched_df.get("jina_verifier_used", pd.Series(dtype=str)).astype(str)
        .str.lower().isin(["true", "1"]).sum()
    )
    jina_decisions = enriched_df.get("jina_verifier_decision", pd.Series(dtype=str)).astype(str)
    review = int(
        enriched_df.get("manual_review_needed", pd.Series(dtype=str))
        .astype(str).str.lower().isin(["true", "1", "yes"]).sum()
    )
    no_match = int(actions.isin(["MISSING_DOMAIN", "NO_CONFIDENT_MATCH"]).sum())
    return {
        "timestamp":                 pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
        "input_filename":            input_filename,
        "input_file_hash":           run_id,
        "run_label":                 run_label,
        "total_rows_in_input":       total_rows_input,
        "processed_rows":            batch_n,
        "max_serper_queries":        max_queries,
        "haiku_mode":                haiku_mode,
        "haiku_model":               haiku_model,
        "haiku_max_rows":            haiku_max_rows if haiku_max_rows > 0 else "all",
        "jina_mode":                 jina_mode,
        "candidate_debug_mode":      "on" if debug_mode else "off",
        "serper_key_present":        "yes" if serper_key_present else "no",
        "anthropic_key_present":     "yes" if anthropic_key_present else "no",
        "rows_with_final_domain":    has_final,
        "final_website_coverage_%":  f"{coverage}%",
        "rows_reviewed_by_haiku":    n_haiku,
        "haiku_accepted_python":     int(decisions.eq("accept").sum()),
        "haiku_replaced_python":     int(decisions.eq("replace").sum()),
        "haiku_rejected":            int(decisions.eq("reject").sum()),
        "haiku_uncertain":           int(decisions.eq("uncertain").sum()),
        "rows_verified_by_jina":     n_jina,
        "jina_confirmed":            int(jina_decisions.eq("confirm").sum()),
        "jina_replaced":             int(jina_decisions.eq("replace").sum()),
        "jina_rejected":             int(jina_decisions.eq("reject").sum()),
        "jina_uncertain":            int(jina_decisions.isin(["uncertain", "insufficient_evidence"]).sum()),
        "website_verifier_provider": verifier_provider,
        "verification_mode":         verifier_mode,
        "rows_verified_by_website_verifier": int(
            enriched_df.get("verifier_used", pd.Series(dtype=str)).astype(str)
            .str.lower().isin(["true", "1"]).sum()
        ),
        "firecrawl_confirmed":       int(
            enriched_df.get("firecrawl_decision", pd.Series(dtype=str)).astype(str).eq("confirm").sum()
        ),
        "firecrawl_replaced":        int(
            enriched_df.get("firecrawl_decision", pd.Series(dtype=str)).astype(str)
            .str.startswith("replace").sum()
        ),
        "firecrawl_rejected":        int(
            enriched_df.get("firecrawl_decision", pd.Series(dtype=str)).astype(str)
            .str.startswith("reject").sum()
        ),
        "firecrawl_uncertain":       int(
            enriched_df.get("firecrawl_decision", pd.Series(dtype=str)).astype(str).eq("uncertain").sum()
        ),
        "firecrawl_pages_fetched_total": int(
            pd.to_numeric(
                enriched_df.get("firecrawl_pages_fetched", pd.Series(dtype=int)), errors="coerce"
            ).fillna(0).sum()
        ),
        "no_confident_match_count":  no_match,
        "manual_review_count":       review,
        # v12 Firecrawl multi-key counters
        "firecrawl_keys_loaded":     firecrawl_keys_loaded,
        "firecrawl_key_failovers":   int(
            pd.to_numeric(
                enriched_df.get("firecrawl_key_failover_count", pd.Series(dtype=int)), errors="coerce"
            ).fillna(0).sum()
        ),
        # v11 Haiku extended counters
        "haiku_needs_firecrawl":     int(decisions.eq("needs_firecrawl").sum()),
        "firecrawl_forced_by_haiku": int(
            enriched_df.get("verification_reason", pd.Series(dtype=str)).astype(str)
            .str.contains("haiku_requested_firecrawl", na=False).sum()
        ),
        "timeout_fallback_applied":  int(
            enriched_df.get("final_decision_source", pd.Series(dtype=str)).astype(str)
            .str.contains("timeout_fallback", na=False).sum()
        ),
        # ── Serper usage ──────────────────────────────────────────────────────
        "serper_queries_total":                serper_queries_total,
        "serper_avg_queries_per_processed_row": round(serper_queries_total / batch_n, 2) if batch_n else 0.0,
        "serper_avg_queries_per_serper_row":   round(
            serper_queries_total
            / max(int(enriched_df.get("search_query_used", pd.Series(dtype=str)).astype(str)
                      .str.strip().replace("", pd.NA).notna().sum()), 1), 2
        ) if serper_queries_total else 0.0,
        # ── Haiku usage ───────────────────────────────────────────────────────
        "haiku_calls_total":                   n_haiku,
        "haiku_avg_calls_per_processed_row":   round(n_haiku / batch_n, 3) if batch_n else 0.0,
        # ── Size inference summary ────────────────────────────────────────────
        "size_inference_rows":                 int(
            enriched_df.get("size_inference_enabled", pd.Series(dtype=str))
            .astype(str).str.lower().isin(["true","1"]).sum()
        ),
        # ── Detected country and Firecrawl location ───────────────────────────
        "detected_country":             detected_country,
        "country_detection_source":     country_detection_source,
        "firecrawl_location_mode":      firecrawl_location_mode,
        "firecrawl_location_used":      firecrawl_location_used,
        "firecrawl_location_country":   firecrawl_location_country,
        "firecrawl_location_languages": firecrawl_location_languages,
        # ── Firecrawl credit / usage counters (from runtime health) ──────────
        **(_build_fc_usage_fields(enriched_df, firecrawl_keys_loaded, fc_health or {})),
    }


def _show_run_info_block(
    run_label: str,
    run_id: str,
    batch_n,
    max_queries: int,
    haiku_mode: str,
    debug_mode: bool,
    filename: str,
) -> None:
    """Show a compact post-run settings summary in the main area."""
    mode_safe = _MODE_SAFE.get(haiku_mode, "pythononly")
    debug_str = "on" if debug_mode else "off"
    st.markdown(
        f"<div style='background:#f0f4f8;border-radius:6px;padding:10px 16px;"
        f"font-size:0.82em;margin:8px 0;line-height:1.7'>"
        f"<b>Run summary</b> · "
        f"Mode: <code>{mode_safe}</code> · "
        f"Rows: <code>{batch_n}</code> · "
        f"Queries: <code>{max_queries}</code> · "
        f"Debug: <code>{debug_str}</code><br>"
        f"File: <code>{filename}</code> · "
        f"Autosave: <code>{_AUTOSAVE_DIR}/{run_label}_{run_id[:8] if run_label else run_id}</code>"
        f"</div>",
        unsafe_allow_html=True,
    )


def _load_secrets_key() -> str | None:
    try:
        return st.secrets.get("SERPER_API_KEY") or st.secrets.get("serper_api_key")
    except Exception:
        return None


def _parse_bytes(raw: bytes, filename: str) -> pd.DataFrame | None:
    """Parse CSV or Excel bytes into a DataFrame."""
    try:
        if filename.lower().endswith(".csv"):
            return pd.read_csv(io.BytesIO(raw), dtype=str).fillna("")
        else:
            return pd.read_excel(io.BytesIO(raw), dtype=str).fillna("")
    except Exception as e:
        st.error(f"Could not read file: {e}")
        return None


def _show_results(enriched_df: pd.DataFrame, stored_cols: dict) -> None:
    """Render the results table and review expander."""
    show_cols = [c for c in [
        stored_cols.get("company"),
        stored_cols.get("website"),
        "normalized_input_website",
        "validated_domain",
        "domain_source",
        "domain_action",
        "domain_confidence",
        "website_discovery_method",
        "manual_review_needed",
    ] if c and c in enriched_df.columns]
    st.dataframe(enriched_df[show_cols], use_container_width=True, height=360)

    review_mask = (
        enriched_df.get("manual_review_needed", pd.Series(False))
        .astype(str).str.lower().isin(["true", "1", "yes"])
    )
    if review_mask.any():
        with st.expander(
            f"🔴 Rows needing manual review ({int(review_mask.sum())})", expanded=False
        ):
            st.dataframe(enriched_df[review_mask][show_cols], use_container_width=True)


def _download_section(
    enriched_df, original_df, evidence_rows, stored_cols,
    filename="register_cleaned_output.xlsx",
    debug_rows: list[dict] | None = None,
    debug_mode: bool = False,
    run_meta: dict | None = None,
    jina_debug_rows: list[dict] | None = None,
    fc_audit: dict | None = None,
) -> None:
    """Render the output-sheet legend + download button."""
    excel_bytes = build_excel(
        enriched_df, original_df, evidence_rows, stored_cols,
        debug_rows=debug_rows, debug_mode=debug_mode, run_meta=run_meta,
        jina_debug_rows=jina_debug_rows, fc_audit=fc_audit,
    )
    jina_ran = (
        "jina_verifier_used" in enriched_df.columns
        and enriched_df["jina_verifier_used"].astype(str).str.lower().isin(["true", "1"]).any()
    )
    sheet_list = (
        "1. **Best Guess Input** — company, website, email, city, province, phone "
        "+ discovery method (ready for Lead Prioritizer)  \n"
        "2. **Cleaned Register Input** — all original columns + validation + diagnostic columns  \n"
        "3. **Review Needed** — rows requiring manual check  \n"
        "4. **Original Input** — unchanged source data  \n"
        "5. **Raw Search Evidence** — Serper queries, results, name variants, rejection reasons  \n"
        "6. **Python vs Haiku Comparison** — side-by-side comparison (populated when Haiku mode is active)  \n"
        "7. **Candidate Rank Diagnostics** — per-company: selected domain rank, score vs. top Serper result  \n"
    )
    sheet_n = 8
    if debug_mode:
        sheet_list += (
            f"  \n{sheet_n}. **Candidate Discovery Debug** — every Serper result per company with full scoring detail"
        )
        sheet_n += 1
    if jina_ran or jina_debug_rows:
        sheet_list += (
            f"  \n{sheet_n}. **Website Verification Debug** — one row per company/candidate/page fetched by verifier"
        )
        sheet_n += 1
    sheet_list += f"  \n{sheet_n}. **Validation Diagnostics** — coverage, confidence, divergence counts  \n"
    sheet_n += 1
    if fc_audit is not None:
        sheet_list += (
            f"  \n{sheet_n}. **Firecrawl Audit** — preflight result + runtime health counters (no API key values)  \n"
        )
        sheet_n += 1
    sheet_list += f"  \n{sheet_n}. **API Usage Summary** — Firecrawl credits, Serper queries, Haiku calls  \n"
    sheet_list += f"  \n{sheet_n + 1}. **Run Summary** — run settings and outcome metrics"
    st.markdown("**Output sheets:**  \n" + sheet_list)
    st.download_button(
        "⬇ Download cleaned register Excel",
        data=excel_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )


def _smoke_test_firecrawl_health() -> None:
    """
    Self-contained smoke test for Firecrawl health helpers.
    Simulates 6 scenarios and asserts expected outcomes.
    Call from the REPL or CI: python -c "from input_cleaner_register_edition import _smoke_test_firecrawl_health; _smoke_test_firecrawl_health()"
    """
    def _apply(health, statuses, failovers=None):
        """Feed a sequence of fetch_status strings into the health counter."""
        if failovers is None:
            failovers = [0] * len(statuses)
        for st, fo in zip(statuses, failovers):
            _fc_health_update(health, st, "", fo, fc_was_used=True)

    # Case 1: all OK — no fail-fast
    h1 = _make_fc_health()
    _apply(h1, ["ok"] * 50)
    assert h1["pages_successful"] == 50
    assert h1["consecutive_failures"] == 0
    assert _fc_health_check_fail_fast(h1) is None, "Case 1 fail"

    # Case 2: 10 consecutive non-ok → fail-fast triggered
    h2 = _make_fc_health()
    _apply(h2, ["ok"] * 5 + ["timeout"] * 10)
    reason2 = _fc_health_check_fail_fast(h2)
    assert reason2 is not None and "consecutive" in reason2.lower(), f"Case 2 fail: {reason2}"

    # Case 3: 30+ attempts with <20% success rate → fail-fast (interleave ok+fail to avoid consecutive threshold)
    h3 = _make_fc_health()
    # Pattern: 1 ok then 8 fails, repeat — keeps consecutive < 10 but success rate ~11%
    _pattern3 = (["ok"] + ["http_500"] * 8) * 4  # 36 attempts, 4 ok = 11%
    _apply(h3, _pattern3)
    reason3 = _fc_health_check_fail_fast(h3)
    assert reason3 is not None and "success rate" in reason3.lower(), f"Case 3 fail: {reason3}"

    # Case 4: fail-fast disabled → no trigger
    h4 = _make_fc_health()
    _apply(h4, ["timeout"] * 10)
    assert _fc_health_check_fail_fast(h4, enabled=False) is None, "Case 4 fail"

    # Case 5: failovers counted correctly
    h5 = _make_fc_health()
    _apply(h5, ["ok", "ok", "ok"], failovers=[2, 0, 1])
    assert h5["key_failovers"] == 3, f"Case 5 fail: key_failovers={h5['key_failovers']}"
    assert h5["key_failure_events"] == 2, f"Case 5 fail: key_failure_events={h5['key_failure_events']}"

    # Case 6: quota errors counted
    h6 = _make_fc_health()
    _apply(h6, ["http_402", "http_429", "ok"])
    assert h6["quota_or_billing_errors"] >= 1, f"Case 6 fail: quota_or_billing_errors={h6['quota_or_billing_errors']}"
    assert h6["rate_limit_errors"] >= 1, f"Case 6 fail: rate_limit_errors={h6['rate_limit_errors']}"

    # Case 7: fc_was_used=False → counters unchanged
    h7 = _make_fc_health()
    _fc_health_update(h7, "ok", "", 0, fc_was_used=False)
    assert h7["requests_attempted"] == 0, "Case 7 fail"

    # Case 8: _safe_extract_http_status handles combined / malformed status strings
    assert _safe_extract_http_status("http_500")          == 500,  "Case 8a fail"
    assert _safe_extract_http_status("http_500; ok")      == 500,  "Case 8b fail"
    assert _safe_extract_http_status("http_429; http_200") == 429, "Case 8c fail"
    assert _safe_extract_http_status("ok")                == 0,    "Case 8d fail"
    assert _safe_extract_http_status(None)                == 0,    "Case 8e fail"
    # Verify that a combined status no longer crashes _fc_health_update
    h8 = _make_fc_health()
    _fc_health_update(h8, "http_500; ok", "", 0, fc_was_used=True)
    assert h8["requests_attempted"] == 1, "Case 8f fail"

    print("[SMOKE TEST] _smoke_test_firecrawl_health: all 8 cases passed.", flush=True)


def _smoke_test_country_config() -> None:
    """
    Smoke test for country detection, column mapping, Serper params, and path routing.
    Run with:
        python -c "from input_cleaner_register_edition import _smoke_test_country_config; _smoke_test_country_config()"
    """
    import pandas as _pd

    # ── 1. Column detection: Italy-style dataframe ────────────────────────────
    it_df = _pd.DataFrame(columns=[
        "Company Name", "Website", "Email address",
        "City", "National statistical institute Province", "Postal Code", "Phone number",
    ])
    it_cols = detect_columns_generic(it_df, IT_CONFIG)
    assert it_cols["company"] == "Company Name",  f"IT company col: {it_cols['company']}"
    assert it_cols["email"]   == "Email address", f"IT email col: {it_cols['email']}"
    assert it_cols["province"] == "National statistical institute Province", \
        f"IT province col: {it_cols['province']}"

    # ── 2. Column detection: Germany-style dataframe ──────────────────────────
    de_df = _pd.DataFrame(columns=[
        "company_name_clean", "company_name_raw", "city_or_registered_office",
        "federal_state", "registered_address", "company_number", "register_nummer",
    ])
    de_cols = detect_columns_generic(de_df, DE_CONFIG)
    assert de_cols["company"] == "company_name_clean",       f"DE company col: {de_cols['company']}"
    assert de_cols["city"]    == "city_or_registered_office", f"DE city col: {de_cols['city']}"
    assert de_cols["province"] == "federal_state",            f"DE province col: {de_cols['province']}"

    # ── 3. Country auto-detection from file path ──────────────────────────────
    assert detect_country_from_path(r"C:\Users\gertm\Nextcloud\Myngle\Germany\00_raw\Germany_1_R0001_0500.xlsx") == "DE"
    assert detect_country_from_path("/Users/gertm/Nextcloud/Myngle/Italy100/00_raw/Italy100_1_R0001_0500.xlsx") == "IT"
    assert detect_country_from_path("/tmp/unknown_file.xlsx") is None

    # ── 4. Country auto-detection from columns ────────────────────────────────
    assert detect_country_from_columns(de_df) == "DE"
    assert detect_country_from_columns(it_df) == "IT"

    # ── 5. Serper payload uses gl/hl from country config ─────────────────────
    assert IT_CONFIG.serper_gl == "it" and IT_CONFIG.serper_hl == "it", "IT Serper locale"
    assert DE_CONFIG.serper_gl == "de" and DE_CONFIG.serper_hl == "de", "DE Serper locale"

    # ── 6. Dry-run paths resolve correctly ────────────────────────────────────
    it_paths = resolve_pipeline_output_paths(
        r"C:\Users\gertm\Nextcloud\Myngle\Italy100\00_raw\Italy100_1_R0001_0500.xlsx",
        project_root=r"C:\Users\gertm\Nextcloud\Myngle",
        ts="20260613_1200",
    )
    assert "01_cleaned_domains" in it_paths["output_xlsx"].replace("\\", "/"), \
        f"IT output_xlsx: {it_paths['output_xlsx']}"
    assert "Italy100" in it_paths["cohort"], f"IT cohort: {it_paths['cohort']}"

    de_paths = resolve_pipeline_output_paths(
        r"C:\Users\gertm\Nextcloud\Myngle\Germany\00_raw\Germany_1_R0001_0500.xlsx",
        project_root=r"C:\Users\gertm\Nextcloud\Myngle",
        ts="20260613_1200",
    )
    assert "01_cleaned_domains" in de_paths["output_xlsx"].replace("\\", "/"), \
        f"DE output_xlsx: {de_paths['output_xlsx']}"
    assert "Germany" in de_paths["cohort"], f"DE cohort: {de_paths['cohort']}"

    # ── 7. German negative domain checker (word-boundary safe) ────────────────
    assert _is_de_negative_domain("gemeinde-musterhausen.de")[0],     "gemeinde should be rejected"
    assert _is_de_negative_domain("rathaus.de")[0],                    "rathaus should be rejected"
    assert not _is_de_negative_domain("schulenburg.de")[0],            "schulenburg should be safe"
    assert not _is_de_negative_domain("neustadt-gmbh.de")[0],          "neustadt-gmbh should be safe (not a city token)"
    assert not _is_de_negative_domain("kammerer.de")[0],               "kammerer should be safe"

    # ── 8. DE blacklist in is_generic ─────────────────────────────────────────
    assert is_generic("handelsregister.de", country_config=DE_CONFIG), "handelsregister.de should be generic"
    assert is_generic("northdata.de",       country_config=DE_CONFIG), "northdata.de should be generic"
    assert not is_generic("acme-gmbh.de",   country_config=DE_CONFIG), "acme-gmbh.de should not be generic"

    # ── 9. Search queries are country-aware ───────────────────────────────────
    _it_qs = _build_search_queries(
        {"full": "ACME SRL", "brand": "ACME", "no_desc": "ACME SRL", "is_acronym": False},
        city="Milano", province="MI", postcode="", max_queries=3, country_config=IT_CONFIG,
    )
    assert any("sito ufficiale" in q or "official website" in q for q in _it_qs), \
        f"IT queries missing Italian terms: {_it_qs}"

    _de_qs = _build_search_queries(
        {"full": "Mustermann GmbH", "brand": "Mustermann", "no_desc": "Mustermann GmbH", "is_acronym": False},
        city="Berlin", province="Berlin", postcode="", max_queries=3, country_config=DE_CONFIG,
    )
    assert any("Deutschland" in q or "offizielle" in q or "Impressum" in q for q in _de_qs), \
        f"DE queries missing German terms: {_de_qs}"

    # ── 10. Filename-only country detection (Streamlit upload scenario) ───────
    # detect_country_from_path works on bare filenames too
    assert detect_country_from_path("Germany_1_R0001_0500.xlsx") == "DE", \
        "bare German filename should resolve to DE"
    assert detect_country_from_path("Italy100_1_R0001_0500.xlsx") == "IT", \
        "bare Italian filename should resolve to IT"
    assert detect_country_from_path("companies_export.xlsx") is None, \
        "generic filename should return None"

    # resolve_country with bare filename + empty df
    _empty = _pd.DataFrame()
    assert resolve_country("auto", "Germany_1_R0001_0500.xlsx", _empty) == "DE"
    assert resolve_country("auto", "Italy100_1_R0001_0500.xlsx", _empty) == "IT"
    assert resolve_country("DE",   "anything.xlsx",              _empty) == "DE"
    assert resolve_country("IT",   "Germany_1_R0001_0500.xlsx",  _empty) == "IT"

    # ── 11. Full German register column set ───────────────────────────────────
    _de_full = _pd.DataFrame(columns=[
        "source", "source_row_id", "company_number", "current_status",
        "jurisdiction_code", "company_name_raw", "company_name_clean",
        "legal_form_detected", "registered_address", "city_or_registered_office",
        "federal_state", "native_company_number", "registrar",
        "register_art", "register_nummer", "retrieved_at",
        "pre_score", "pre_label", "positive_reasons", "exclude_reasons",
        "low_priority_reasons",
    ])
    _de_full_cols = detect_columns_generic(_de_full, DE_CONFIG)
    assert _de_full_cols["company"]  == "company_name_clean",        f"DE full: company={_de_full_cols['company']}"
    assert _de_full_cols["city"]     == "city_or_registered_office", f"DE full: city={_de_full_cols['city']}"
    assert _de_full_cols["province"] == "federal_state",             f"DE full: province={_de_full_cols['province']}"

    # ── 12. normalize_register_columns_for_cleaner: Germany ──────────────────
    _de_norm_in = _pd.DataFrame([{
        "company_name_clean": "Mustermann GmbH",
        "company_name_raw":   "Mustermann GmbH i.G.",
        "city_or_registered_office": "Berlin",
        "federal_state": "Berlin",
        "registered_address": "Unter den Linden 1",
        "legal_form_detected": "GmbH",
    }])
    _de_norm_out, _de_norm_rep = normalize_register_columns_for_cleaner(_de_norm_in, DE_CONFIG)
    assert _de_norm_out["company_name"].iloc[0] == "Mustermann GmbH", \
        f"DE norm company_name: {_de_norm_out['company_name'].iloc[0]}"
    assert _de_norm_out["city"].iloc[0] == "Berlin", \
        f"DE norm city: {_de_norm_out['city'].iloc[0]}"
    assert _de_norm_out["province"].iloc[0] == "Berlin", \
        f"DE norm province: {_de_norm_out['province'].iloc[0]}"
    assert _de_norm_out["country_code"].iloc[0] == "DE", \
        f"DE norm country_code: {_de_norm_out['country_code'].iloc[0]}"
    # Original columns must still be present
    assert "company_name_clean" in _de_norm_out.columns, "DE norm: original column missing"

    # _cols_from_normalized must point to canonical columns
    _de_canon_cols = _cols_from_normalized(_de_norm_out, DE_CONFIG)
    assert _de_canon_cols["company"] == "company_name", \
        f"DE canonical company col: {_de_canon_cols['company']}"
    assert _de_canon_cols["city"] == "city", \
        f"DE canonical city col: {_de_canon_cols['city']}"

    # ── 13. normalize_register_columns_for_cleaner: Italy ────────────────────
    _it_norm_in = _pd.DataFrame([{
        "Company Name": "ACME SRL",
        "Website": "acme.it",
        "Email address": "info@acme.it",
        "City": "Milano",
        "National statistical institute Province": "MI",
        "Postal Code": "20100",
    }])
    _it_norm_out, _it_norm_rep = normalize_register_columns_for_cleaner(_it_norm_in, IT_CONFIG)
    assert _it_norm_out["company_name"].iloc[0] == "ACME SRL", \
        f"IT norm company_name: {_it_norm_out['company_name'].iloc[0]}"
    assert _it_norm_out["city"].iloc[0] == "Milano", \
        f"IT norm city: {_it_norm_out['city'].iloc[0]}"
    assert _it_norm_out["province"].iloc[0] == "MI", \
        f"IT norm province: {_it_norm_out['province'].iloc[0]}"
    assert _it_norm_out["country_code"].iloc[0] == "IT", \
        f"IT norm country_code: {_it_norm_out['country_code'].iloc[0]}"

    # ── 14. detect_country_from_path: bare filenames (spot-check) ────────────
    assert detect_country_from_path("Germany_1_R0001_0500.xlsx") == "DE"
    assert detect_country_from_path("Italy100_1_R0001_0500.xlsx") == "IT"

    # ── 15. Generic German brand: brand_is_de_generic flag ───────────────────
    _gh_variants = extract_name_variants("Global Holding GmbH & Co. KG")
    assert _gh_variants.get("brand_is_de_generic"), (
        f"'Global Holding GmbH & Co. KG' must set brand_is_de_generic=True; "
        f"brand='{_gh_variants.get('brand')}'"
    )
    # Verify that a real distinctive brand does NOT set the flag
    _real_variants = extract_name_variants("Mustermann GmbH")
    assert not _real_variants.get("brand_is_de_generic"), (
        f"'Mustermann GmbH' must NOT set brand_is_de_generic; brand='{_real_variants.get('brand')}'"
    )
    # Single-token company with non-generic brand
    _acme_variants = extract_name_variants("Acme GmbH")
    assert not _acme_variants.get("brand_is_de_generic"), (
        f"'Acme GmbH' must NOT set brand_is_de_generic; brand='{_acme_variants.get('brand')}'"
    )

    # ── 16. URL shorteners: is_url_shortener detects all required domains ─────
    for _short in ["t.co", "bit.ly", "bitly.com", "tinyurl.com", "lnkd.in", "linktr.ee", "goo.gl"]:
        assert is_url_shortener(_short), f"'{_short}' must be detected as url shortener"
    assert not is_url_shortener("acme-gmbh.de"), "acme-gmbh.de must NOT be a shortener"
    assert not is_url_shortener("linkedin.com"),  "linkedin.com must NOT be a shortener"

    # t.co must be blocked from validated_domain via norm_website guard
    _tco_norm = best_website_domain("https://t.co/somelink")
    _tco_blocked = is_url_shortener(_tco_norm) if _tco_norm else True
    assert _tco_blocked, f"t.co must be blocked as url shortener; norm='{_tco_norm}'"

    # ── 17. Firecrawl country default: DE_CONFIG.firecrawl_location ──────────
    assert DE_CONFIG.firecrawl_location.get("country") == "DE", \
        f"DE_CONFIG.firecrawl_location must have country=DE: {DE_CONFIG.firecrawl_location}"
    assert "de" in DE_CONFIG.firecrawl_location.get("languages", []), \
        f"DE_CONFIG.firecrawl_location must include 'de': {DE_CONFIG.firecrawl_location}"
    assert IT_CONFIG.firecrawl_location.get("country") == "IT", \
        f"IT_CONFIG.firecrawl_location must have country=IT: {IT_CONFIG.firecrawl_location}"

    # ── 18. CLI country resolution drives Firecrawl location ─────────────────
    # resolve_country("auto", "Germany_1_R0001_0500.xlsx", empty_df) → DE
    _de_country = resolve_country("auto", "Germany_1_R0001_0500.xlsx", _pd.DataFrame())
    _de_cfg = COUNTRY_CONFIGS.get(_de_country, IT_CONFIG)
    assert _de_cfg.firecrawl_location.get("country") == "DE", (
        f"German filename auto-detected as {_de_country} must use DE FC location, "
        f"got {_de_cfg.firecrawl_location}"
    )

    print("[SMOKE TEST] _smoke_test_country_config: all 18 cases passed.", flush=True)


def _smoke_test_german_pipeline() -> None:
    """
    Regression tests for the German batch pipeline.
    Covers: normalize columns, _cols_from_normalized, _build_best_guess_df fallbacks.
    """
    import pandas as _pd

    # ── T1. normalize_register_columns_for_cleaner produces canonical DE columns ─
    _de_raw = _pd.DataFrame([{
        "company_name_clean": "Mustermann GmbH",
        "homepage_url": "mustermann.de",
        "email": "info@mustermann.de",
        "city_or_registered_office": "Berlin",
        "federal_state": "Berlin",
        "postcode": "10115",
    }])
    _de_norm, _ = normalize_register_columns_for_cleaner(_de_raw, DE_CONFIG)
    assert "company_name" in _de_norm.columns, "T1: company_name canonical col missing after norm"
    assert _de_norm["company_name"].iloc[0] == "Mustermann GmbH", \
        f"T1: company_name value wrong: {_de_norm['company_name'].iloc[0]}"
    assert "city" in _de_norm.columns, "T1: city canonical col missing after norm"
    assert _de_norm["city"].iloc[0] == "Berlin", \
        f"T1: city value wrong: {_de_norm['city'].iloc[0]}"

    # ── T2. _cols_from_normalized returns populated canonical columns ──────────
    _de_cols = _cols_from_normalized(_de_norm, DE_CONFIG)
    assert _de_cols.get("company") == "company_name", \
        f"T2: company col should be 'company_name', got: {_de_cols.get('company')}"
    assert _de_cols.get("city") == "city", \
        f"T2: city col should be 'city', got: {_de_cols.get('city')}"

    # ── T3. _cols_from_normalized does NOT return a blank canonical column ─────
    _de_blank = _pd.DataFrame([{
        "company_name": "",   # canonical exists but is blank
        "company_name_clean": "Blank Test GmbH",
        "city": "Hamburg",
    }])
    _de_blank_cols = _cols_from_normalized(_de_blank, DE_CONFIG)
    # company_name col is blank → should fall back to None or fallback with data
    assert _de_blank_cols.get("company") != "company_name" or \
        _de_blank["company_name"].astype(str).str.strip().ne("").any(), \
        "T3: _cols_from_normalized must not return blank canonical col as company col"

    # ── T4. _build_best_guess_df produces non-blank company_name via fallback ──
    _de_multi = _pd.DataFrame([{
        "company_name": "",
        "company_name_clean": "Fallback GmbH",
        "website": "fallback.de",
        "city": "München",
        "federal_state": "Bavaria",
    }])
    _de_multi_norm, _ = normalize_register_columns_for_cleaner(_de_multi, DE_CONFIG)
    _de_multi_cols = _cols_from_normalized(_de_multi_norm, DE_CONFIG)
    _bg_df = _build_best_guess_df(_de_multi_norm, _de_multi_cols)
    assert len(_bg_df) > 0, "T4: _build_best_guess_df returned empty DataFrame"
    _bg_name = str(_bg_df.iloc[0].get("company_name", "")).strip()
    assert _bg_name, f"T4: company_name is blank in Best Guess output; cols={_de_multi_cols}"

    # ── T5. detect_columns_generic handles German cleaned output file ─────────
    _de_cleaned = _pd.DataFrame([{
        "company_name": "Mustermann GmbH",
        "website": "mustermann.de",
        "city": "Berlin",
        "organization_type": "GmbH",
        "federal_state": "Berlin",
    }])
    _de_gc = detect_columns_generic(_de_cleaned, DE_CONFIG)
    assert _de_gc.get("company") == "company_name", \
        f"T5: detect_columns_generic should find 'company_name', got: {_de_gc.get('company')}"
    assert _de_gc.get("website") == "website", \
        f"T5: detect_columns_generic should find 'website', got: {_de_gc.get('website')}"

    # ── T6. brand_is_de_generic: single-word generic brand detected ───────────
    for _generic_name in ["Global GmbH", "Solutions GmbH & Co. KG", "Digital GmbH"]:
        _v = extract_name_variants(_generic_name)
        assert _v.get("brand_is_de_generic"), \
            f"T6: '{_generic_name}' must set brand_is_de_generic (brand='{_v.get('brand')}')"

    # ── T7. url_shortener guard: t.co must not survive as validated domain ─────
    _tco_domain = best_website_domain("https://t.co/xyz123")
    assert is_url_shortener(_tco_domain) if _tco_domain else True, \
        f"T7: t.co must be flagged as url shortener; got domain='{_tco_domain}'"

    # ── T8. resolve_country + COUNTRY_CONFIGS round-trip for DE ──────────────
    _country = resolve_country("auto", "Germany_2_R0001_0500.xlsx", _pd.DataFrame())
    assert _country == "DE", f"T8: resolve_country for German file must return 'DE', got '{_country}'"
    _cfg = COUNTRY_CONFIGS.get(_country)
    assert _cfg is not None, "T8: COUNTRY_CONFIGS must contain 'DE' key"
    assert _cfg.country_code == "DE", f"T8: DE config country_code must be 'DE': {_cfg.country_code}"
    assert _cfg.firecrawl_location.get("country") == "DE", \
        f"T8: DE firecrawl_location must have country=DE: {_cfg.firecrawl_location}"

    print("[SMOKE TEST] _smoke_test_german_pipeline: all 8 cases passed.", flush=True)


def _smoke_test_size_inference() -> None:
    """
    Regression tests for size inference safety.
    - Apple Music / Headcount false positive must be blocked.
    - German commercial legal forms must be classified correctly.
    - Firecrawl total credits = DV pages + SI new pages.
    """
    _bad_snippets = [
        "Apple Music / Headcount / 100 Miles",
        "Headcount 100 songs on Apple Music",
        "Listen to Headcount on music.apple.com",
    ]
    _bad_url = "https://music.apple.com/us/artist/headcount/40703962"
    _domain  = "global-holding.de"
    _company = "Global Holding GmbH & Co. KG"

    # 1. Regex must not extract employee count from Apple Music snippets
    _combined = " ".join(_bad_snippets)
    _re_result = _extract_employee_size_regex(_combined, "DE")
    assert _re_result["band"] == SIZE_UNKNOWN, \
        f"Apple Music snippets must produce SIZE_UNKNOWN, got: {_re_result}"

    # 2. Source-domain safety: music.apple.com must be rejected
    assert not _is_allowed_size_source(_bad_url, _company, _domain), \
        f"music.apple.com must be rejected by _is_allowed_size_source"

    # 3. career_page_found must be False for Apple Music URL
    _, _, career_found = _classify_hrm_signals(
        _bad_snippets,
        country_code="DE",
        official_domain=_domain,
        evidence_urls=[_bad_url],
    )
    assert not career_found, \
        f"career_page_found must be False for Apple Music URL"

    # 4. German commercial legal forms: GmbH & Co. KG → commercial_company / KEEP
    from input_cleaner_register_edition import classify_organization
    _org, _eli, _pf, _reason = classify_organization(
        "Global Holding GmbH & Co. KG",
        legal_form_hint="GmbH & Co. KG",
        country_code="DE",
    )
    assert _org == "commercial_company", f"GmbH & Co. KG must be commercial_company, got {_org}"
    assert _eli == _ELI_KEEP, f"GmbH & Co. KG must be KEEP, got {_eli}"

    _org2, _eli2, _, _ = classify_organization("Mustermann AG", legal_form_hint="AG", country_code="DE")
    assert _org2 == "commercial_company" and _eli2 == _ELI_KEEP, f"AG must be commercial_company/KEEP"

    # 5. Firecrawl total credits = DV + SI new pages
    import pandas as _pd
    _test_df = _pd.DataFrame({
        "firecrawl_pages_fetched":         [2],
        "firecrawl_pages_used_for_size":   [4],
        "firecrawl_size_inference_reused_pages": [0],
        "firecrawl_used":                  ["True"],
        "firecrawl_used_for_size_inference": ["True"],
        "firecrawl_decision":              ["confirm"],
    })
    _fc_fields = _build_fc_usage_fields(_test_df, 1, {})
    _total = _fc_fields["firecrawl_total_estimated_credits"]
    assert _total == 6, f"Total estimated credits must be 6 (2 DV + 4 SI), got {_total}"

    print("[SMOKE TEST] _smoke_test_size_inference: all 5 cases passed.", flush=True)


def _add_alias_diagnostic_cols(df: "pd.DataFrame") -> "pd.DataFrame":
    """
    Ensure friendly diagnostic column aliases exist in the output dataframe.
    Maps from existing pipeline columns where possible; leaves blank otherwise.
    These columns are NEVER overwritten if already present.
    """
    aliases = {
        # Req col name          : source col (first that exists wins) or ""
        "needs_domain_review":  ["manual_review_needed"],
        "domain_score":         ["domain_confidence", "best_candidate_score"],
        "candidate_1_domain":   [],   # derived below from top_3_candidate_domains
        "candidate_1_score":    ["best_candidate_score"],
        "candidate_1_reason":   ["domain_reason"],
        "candidate_2_domain":   [],   # derived below
        "candidate_2_score":    [],
        "candidate_2_reason":   [],
    }

    # Derive candidate_1/2_domain from top_3_candidate_domains (comma-separated)
    if "top_3_candidate_domains" in df.columns and "candidate_1_domain" not in df.columns:
        _parts = df["top_3_candidate_domains"].astype(str).str.split(r",\s*", expand=False)
        df["candidate_1_domain"] = _parts.apply(
            lambda xs: xs[0].strip() if isinstance(xs, list) and len(xs) >= 1 else ""
        )
        df["candidate_2_domain"] = _parts.apply(
            lambda xs: xs[1].strip() if isinstance(xs, list) and len(xs) >= 2 else ""
        )
    else:
        for col in ("candidate_1_domain", "candidate_2_domain"):
            if col not in df.columns:
                df[col] = ""

    for new_col, sources in aliases.items():
        if new_col in df.columns:
            continue
        if new_col in ("candidate_1_domain", "candidate_2_domain"):
            continue  # already handled above
        filled = False
        for src in sources:
            if src in df.columns:
                df[new_col] = df[src]
                filled = True
                break
        if not filled:
            df[new_col] = ""

    return df


def _no_overwrite_path(p: Path) -> Path:
    """
    Return p unchanged if it does not exist; otherwise return p with _2, _3, … suffix
    inserted before the extension until a free name is found.
    """
    if not p.exists():
        return p
    stem   = p.stem
    suffix = p.suffix
    parent = p.parent
    n = 2
    while True:
        candidate = parent / f"{stem}_{n}{suffix}"
        if not candidate.exists():
            return candidate
        n += 1


def _self_test_domain_quality() -> None:
    """
    Quick self-test for domain rejection logic.
    Verifies that known bad domains are rejected and known good domains are kept.
    """
    import sys as _sys

    cases = [
        # (domain, expect_rejected, label)
        ("it.linkedin.com",              True,  "it.linkedin.com subdomain"),
        ("it.wikipedia.org",             True,  "it.wikipedia.org subdomain"),
        ("it.kompass.com",               True,  "it.kompass.com subdomain"),
        ("fatturatoitalia.it",           True,  "fatturatoitalia.it exact"),
        ("visura.pro",                   True,  "visura.pro exact"),
        ("x.abbrevia.it",               True,  "x.abbrevia.it subdomain"),
        ("rsuibmsegrate.altervista.org", True,  "rsuibmsegrate.altervista.org subdomain"),
        ("ibm.com",                      False, "ibm.com (legit)"),
        ("zf.com",                       False, "zf.com (legit)"),
        ("q8.it",                        False, "q8.it (legit)"),
        ("solutions30.com",              False, "solutions30.com (legit)"),
    ]

    failures: list[str] = []
    for domain, expect_rejected, label in cases:
        rejected = is_generic(domain)
        if rejected != expect_rejected:
            failures.append(
                f"  FAIL [{label}] is_generic({domain!r}) = {rejected}; "
                f"expected {expect_rejected}"
            )

    if failures:
        print("[SELF-TEST domain-quality] FAILED:", flush=True)
        for f in failures:
            print(f, flush=True)
        _sys.exit(1)
    else:
        print(
            f"[SELF-TEST domain-quality] All {len(cases)} cases passed.",
            flush=True,
        )


def cli_batch_run() -> None:
    """
    Non-Streamlit batch entry point.

    Usage:
        python input_cleaner_register_edition.py \\
            --input  Italy100/00_raw/Italy100_1_R0001_0500.xlsx \\
            --project-root C:\\Users\\gertm\\Nextcloud\\Myngle \\
            --serper-key sk-xxx \\
            --max-rows 0

    Options:
        --input           Path to the input .xlsx or .csv file (or a folder of .xlsx files).
        --project-root    Project root folder (auto-detected from 00_raw/ when omitted).
        --serper-key      Serper API key (falls back to SERPER_API_KEY env var / secrets).
        --max-rows        Process only the first N rows (0 = all). Default 0.
        --max-queries     Serper queries per company (3/5/8). Default 5.
        --haiku-mode      Haiku review mode. Default 'uncertain'.
        --verifier        Website verifier: off/firecrawl/jina/fc_jina. Default 'off'.
        --debug           Enable debug output sheet.
        --dry-run-paths   Print resolved output paths and exit without processing.
    """
    parser = argparse.ArgumentParser(
        description="Input Cleaner · Register Edition — CLI batch mode",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--input",          default=None,
                        help="Input .xlsx or .csv file, or a folder of .xlsx files")
    parser.add_argument("--self-test-domain-quality", action="store_true",
                        help="Run domain-quality self-test and exit")
    parser.add_argument("--project-root",   default=None,   help="Pipeline project root folder")
    parser.add_argument("--serper-key",     default=None,   help="Serper API key")
    parser.add_argument("--anthropic-key",  default=None,   help="Anthropic API key for Haiku")
    parser.add_argument("--firecrawl-key",  default=None,   help="Firecrawl API key")
    parser.add_argument("--max-rows",       type=int, default=0,  help="Rows to process (0=all)")
    parser.add_argument("--max-queries",    type=int, default=5,  choices=[3, 5, 8])
    parser.add_argument("--haiku-mode",     default=_HAIKU_MODE_UNCERTAIN,
                        choices=_HAIKU_MODES, help="Haiku review mode")
    parser.add_argument("--verifier",       default=_VP_OFF,
                        choices=_VP_OPTIONS, help="Website verifier provider")
    parser.add_argument("--debug",          action="store_true", help="Enable debug output sheet")
    parser.add_argument("--dry-run-paths",  action="store_true",
                        help="Print resolved pipeline paths and exit (no processing)")
    parser.add_argument("--skip-firecrawl-preflight", action="store_true",
                        help="Skip the per-key Firecrawl preflight health check before processing")
    parser.add_argument("--no-firecrawl-fail-fast", action="store_true",
                        help="Disable the runtime Firecrawl fail-fast safety check")
    parser.add_argument("--country", default="auto", choices=["auto", "IT", "DE"],
                        help="Country pipeline: auto (default), IT (Italy), DE (Germany)")
    parser.add_argument("--infer-size", action="store_true",
                        help="Infer company size and HR signals using Serper + Firecrawl evidence")
    # ── Firecrawl speed / budget controls ────────────────────────────────────
    parser.add_argument("--fc-speed",
                        default=_FC_SPEED_FAST, choices=_FC_SPEED_OPTIONS,
                        help=(
                            "Firecrawl speed mode (default: Fast). "
                            "Fast=1 cand/1 page/6s, Balanced=1/2/8s, Thorough=3/3/15s."
                        ))
    parser.add_argument("--fc-max-cands", type=int, default=None,
                        help="Override max candidates per company (default: from --fc-speed).")
    parser.add_argument("--fc-max-pages", type=int, default=None,
                        help="Override max pages per candidate (default: from --fc-speed).")
    parser.add_argument("--fc-page-timeout", type=int, default=None,
                        help="Override page timeout in seconds (default: from --fc-speed).")
    parser.add_argument("--fc-budget", type=int, default=0,
                        help=(
                            "Hard Firecrawl page budget. Processing stops cleanly when "
                            "successful pages reach this limit (0 = no limit)."
                        ))
    args = parser.parse_args()

    # ── Self-test mode ────────────────────────────────────────────────────────
    if getattr(args, "self_test_domain_quality", False):
        _run_domain_quality_self_test()
        sys.exit(0)

    # ── Resolve input: file or folder ─────────────────────────────────────────
    if not args.input:
        print("ERROR: --input is required (file or folder of .xlsx files).", file=sys.stderr)
        sys.exit(1)

    _input_raw = Path(args.input).resolve()
    if not _input_raw.exists():
        print(f"ERROR: input path not found: {_input_raw}", file=sys.stderr)
        sys.exit(1)

    if _input_raw.is_dir():
        _input_files = sorted(_input_raw.glob("*.xlsx"))
        if not _input_files:
            print(f"ERROR: no .xlsx files found in folder: {_input_raw}", file=sys.stderr)
            sys.exit(1)
        print(
            f"[cleaner] Folder mode: {len(_input_files)} file(s) in {_input_raw}",
            flush=True,
        )
        for _f in _input_files:
            print(f"  {_f.name}", flush=True)
        print(flush=True)
        import subprocess as _sp
        _base_argv = [a for a in sys.argv if not a.startswith("--input")]
        for _input_file in _input_files:
            print(f"\n[cleaner] ═══ Processing: {_input_file.name} ═══", flush=True)
            _sub_argv = [sys.executable] + _base_argv + ["--input", str(_input_file)]
            _proc = _sp.run(_sub_argv)
            if _proc.returncode != 0:
                print(
                    f"[cleaner] WARNING: {_input_file.name} exited with code "
                    f"{_proc.returncode} — continuing with next file.",
                    flush=True,
                )
        return

    input_path = _input_raw

    ts       = datetime.now().strftime("%Y%m%d_%H%M")
    pl_paths = resolve_pipeline_output_paths(str(input_path), args.project_root, ts=ts)

    if args.dry_run_paths:
        import json as _json
        # Determine country from path for dry-run (no df loaded yet)
        _dry_country = (
            args.country.upper()
            if args.country.upper() in COUNTRY_CONFIGS
            else (detect_country_from_path(str(input_path)) or "IT")
        )
        pl_paths["country"] = _dry_country
        print(_json.dumps(pl_paths, indent=2))
        sys.exit(0)

    # ── API keys: CLI arg → env var → secrets file ────────────────────────────
    serper_key = (
        args.serper_key
        or os.environ.get("SERPER_API_KEY", "")
        or _load_secrets_key()
        or ""
    )
    anthropic_key = (
        args.anthropic_key
        or os.environ.get("ANTHROPIC_API_KEY", "")
        or ""
    ) or None
    fc_keys_cli: list[str] = []
    if args.firecrawl_key:
        fc_keys_cli = [args.firecrawl_key]
    else:
        fc_keys_cli = _fc_load_keys()
    fc_key_arg: str | list[str] | None = fc_keys_cli if fc_keys_cli else None

    # ── Firecrawl preflight ───────────────────────────────────────────────────
    _fc_preflight: dict = {}
    _fc_fail_fast = not args.no_firecrawl_fail_fast
    if args.verifier in (_VP_FIRECRAWL, _VP_FC_JINA) and fc_keys_cli:
        if args.skip_firecrawl_preflight:
            print("[FC PREFLIGHT] Skipped (--skip-firecrawl-preflight).", flush=True)
            _fc_preflight = {"preflight_status": "SKIPPED"}
        else:
            print(f"[FC PREFLIGHT] Testing {len(fc_keys_cli)} key(s)...", flush=True)
            _fc_preflight = firecrawl_preflight_check(fc_keys_cli)
            _pst = _fc_preflight.get("preflight_status", "UNKNOWN")
            _pok  = _fc_preflight.get("keys_ok", 0)
            _ptot = _fc_preflight.get("keys_total", 0)
            print(
                f"[FC PREFLIGHT] Status: {_pst} · "
                f"{_pok}/{_ptot} keys ok · "
                f"statuses: {_fc_preflight.get('key_statuses', {})}",
                flush=True,
            )
            if _pst == "FAILED":
                print(
                    "[FC PREFLIGHT] ERROR: All Firecrawl keys failed preflight. "
                    "Check API keys and account status. Aborting.",
                    file=sys.stderr,
                )
                sys.exit(2)
            if _pst == "DEGRADED":
                print(
                    "[FC PREFLIGHT] WARNING: Some Firecrawl keys failed — "
                    "will use available working keys.",
                    flush=True,
                )

    # ── Load input ────────────────────────────────────────────────────────────
    raw_bytes = input_path.read_bytes()
    file_hash = _file_hash(raw_bytes)
    df = _parse_bytes(raw_bytes, input_path.name)
    if df is None:
        print(f"ERROR: could not parse input file: {input_path}", file=sys.stderr)
        sys.exit(1)

    batch_n = len(df) if args.max_rows <= 0 else min(args.max_rows, len(df))
    run_df  = df.head(batch_n).copy()

    # ── Country resolution ───────────────────────────────────────────────────
    country_code = resolve_country(args.country, str(input_path), run_df)
    cfg          = COUNTRY_CONFIGS.get(country_code, IT_CONFIG)

    # ── Normalize to canonical columns ───────────────────────────────────────
    run_df, _norm_report = normalize_register_columns_for_cleaner(run_df, cfg)
    cols = _cols_from_normalized(run_df, cfg)

    # Firecrawl location follows resolved country (never silently default to Italy)
    _fc_loc_payload = cfg.firecrawl_location or {"country": "IT", "languages": ["it", "en"]}

    run_label    = _make_run_label(args.haiku_mode, batch_n, args.max_queries, args.debug, ts=ts)
    run_filename = _make_filename(run_label, file_hash)

    import time as _time_mod
    _cli_started_at = _time_mod.time()
    _fc_loc_country_str = _fc_loc_payload.get("country", "auto") if _fc_loc_payload else "auto"
    _fc_loc_langs_str   = ", ".join(_fc_loc_payload.get("languages", [])) if _fc_loc_payload else ""
    _infer_size_str     = "enabled" if getattr(args, "infer_size", False) else "disabled"

    print(f"[cleaner] Input:            {input_path}", flush=True)
    print(f"[cleaner] Country:          {cfg.country_name} ({cfg.country_code})", flush=True)
    print(f"[cleaner] Rows:             {batch_n} / {len(df)}", flush=True)
    print(f"[cleaner] Output dir:       {pl_paths['output_xlsx']}", flush=True)
    print(f"[cleaner] Verifier:         {args.verifier}", flush=True)
    print(f"[cleaner] Firecrawl loc:    country={_fc_loc_country_str} languages=[{_fc_loc_langs_str}]", flush=True)
    print(f"[cleaner] Haiku mode:       {args.haiku_mode}", flush=True)
    print(f"[cleaner] Size inference:   {_infer_size_str}", flush=True)
    print(f"[cleaner] Started:          {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}", flush=True)

    # ── Column-map debug output ───────────────────────────────────────────────
    _col_summary = ", ".join(f"{role}={col}" for role, col in cols.items() if col)
    print(f"[cleaner] Column map:       {_col_summary}", flush=True)
    _name_col = cols.get("company") or ""
    _first_names: list[str] = []
    if _name_col and _name_col in run_df.columns:
        _first_names = (
            run_df[_name_col].astype(str).str.strip()
            .replace("", None).dropna()
            .head(5).tolist()
        )
    print(f"[cleaner] First names:      {_first_names}", flush=True)

    # ── Hard guardrail: abort if company column is missing or all blank ──────
    if not _name_col:
        print(
            f"\nERROR: Could not find a company-name column.\n"
            f"  Resolved country: {cfg.country_name} ({cfg.country_code})\n"
            f"  Available columns: {list(run_df.columns)}\n"
            f"  Normalization report: {_norm_report}\n"
            f"  Hint: pass --country DE or --country IT to force the country.",
            file=sys.stderr,
        )
        sys.exit(3)

    _sample_rows = min(5, len(run_df))
    _sample_names = run_df[_name_col].astype(str).str.strip().head(_sample_rows).tolist()
    if all(n == "" for n in _sample_names):
        print(
            f"\nERROR: company-name column '{_name_col}' is blank in the first "
            f"{_sample_rows} rows.\n"
            f"  Resolved country: {cfg.country_name} ({cfg.country_code})\n"
            f"  Available columns: {list(run_df.columns)}\n"
            f"  Normalization report: {_norm_report}\n"
            f"  First 5 raw values from likely source columns:\n"
            + "\n".join(
                f"    [{c}]: {run_df[c].head(5).tolist()}"
                for c in ["company_name_clean", "company_name_raw",
                           "Company Name", "company_name"]
                if c in run_df.columns
            ),
            file=sys.stderr,
        )
        sys.exit(3)

    # ── Resolve Firecrawl speed defaults ─────────────────────────────────────
    _cli_fc_speed = args.fc_speed
    _spd = _FC_SPEED_DEFAULTS.get(_cli_fc_speed, _FC_SPEED_DEFAULTS[_FC_SPEED_FAST])
    _cli_fc_max_cands   = args.fc_max_cands   if args.fc_max_cands   is not None else _spd[0]
    _cli_fc_max_pages   = args.fc_max_pages   if args.fc_max_pages   is not None else _spd[1]
    _cli_fc_page_timeout= args.fc_page_timeout if args.fc_page_timeout is not None else _spd[2]
    _cli_fc_budget      = max(0, args.fc_budget)

    if args.infer_size:
        print(
            "[cleaner] WARNING: --infer-size is enabled. "
            "Size inference may add up to 2 Firecrawl pages per company.",
            flush=True,
        )

    if args.verifier in (_VP_FIRECRAWL, _VP_FC_JINA):
        print(
            f"[cleaner] Firecrawl mode: {_cli_fc_speed} · "
            f"max_cands={_cli_fc_max_cands} · max_pages={_cli_fc_max_pages} · "
            f"timeout={_cli_fc_page_timeout}s · "
            f"infer_size={'on' if args.infer_size else 'off'}"
            + (f" · budget={_cli_fc_budget} pages" if _cli_fc_budget else " · budget=unlimited"),
            flush=True,
        )

    _cli_fc_health: dict = _make_fc_health()
    # Current-name box: updated before each row's progress callback via index lookup
    _name_col_for_cb = cols.get("company") or ""

    # Budget-exceeded sentinel — raised from progress callback to stop processing cleanly
    class _FcBudgetExceeded(Exception):
        pass

    def _cli_progress(i, total):
        # Hard budget guard: stop when successful pages reach the limit
        if _cli_fc_budget > 0:
            _pages_so_far = _cli_fc_health.get("pages_successful", 0)
            if _pages_so_far >= _cli_fc_budget:
                raise _FcBudgetExceeded(
                    f"Firecrawl page budget reached ({_pages_so_far}/{_cli_fc_budget})"
                )
        # Look up the company name for the row just completed (0-indexed row = i-1)
        _cur = ""
        if _name_col_for_cb and _name_col_for_cb in run_df.columns and i > 0:
            try:
                _cur = str(run_df[_name_col_for_cb].iloc[i - 1] or "").strip()
            except (IndexError, KeyError):
                _cur = ""
        _counters = {
            **_cli_fc_health,
            "serper_queries": getattr(process_dataframe, "_last_serper_count", 0),
        }
        _print_cli_progress(
            "[cleaner]", i, total, _cli_started_at,
            current_name=_cur,
            counters=_counters,
        )

    _budget_hit = False
    try:
        enriched_df, evidence_rows, debug_rows, jina_debug_rows = process_dataframe(
            run_df, cols, serper_key or None, args.max_queries,
            progress_cb=_cli_progress,
            live_counters_out=_cli_fc_health,
            run_id=file_hash,
            resume_from=0,
            prior_results=[],
            prior_evidence=[],
            run_label=run_label,
            haiku_mode=args.haiku_mode,
            haiku_api_key=anthropic_key,
            haiku_model=_DEFAULT_HAIKU_MODEL,
            haiku_max_rows=0,
            jina_mode=_JINA_MODE_OFF,
            verifier_provider=args.verifier,
            verifier_mode=_VM_UNCERTAIN,
            fc_key=fc_key_arg,
            fc_location=_fc_loc_payload,
            eligibility_filter_mode=_PF_MODE_MAYBE,
            debug_mode=args.debug,
            fc_fail_fast=_fc_fail_fast,
            country_config=cfg,
            infer_size=args.infer_size,
            max_cands_per_company=_cli_fc_max_cands,
            max_pages_per_cand=_cli_fc_max_pages,
            page_timeout=_cli_fc_page_timeout,
            fc_speed_mode=_cli_fc_speed,
        )
    except _FcBudgetExceeded as _bexc:
        _budget_hit = True
        print(f"\n[cleaner] WARNING: {_bexc}. Saving partial results.", flush=True)
        # process_dataframe returns partial results via _cli_fc_health and autosave;
        # reconstruct what we have from autosave or use an empty fallback.
        # Attempt to load the latest autosave checkpoint for this run.
        try:
            from pathlib import Path as _Path2
            _as_dir = _AUTOSAVE_DIR / file_hash
            _cp_files = sorted(_as_dir.glob("checkpoint_*.xlsx")) if _as_dir.exists() else []
            if _cp_files:
                import openpyxl as _opx
                _cp_wb = _opx.load_workbook(_cp_files[-1], read_only=True, data_only=True)
                _cp_ws = _cp_wb.active
                _cp_hdr = [c.value for c in next(_cp_ws.iter_rows(min_row=1, max_row=1))]
                _cp_rows = [[c.value for c in r] for r in _cp_ws.iter_rows(min_row=2)]
                enriched_df  = pd.DataFrame(_cp_rows, columns=_cp_hdr)
                evidence_rows = []
                debug_rows    = []
                jina_debug_rows = []
                print(f"[cleaner] Loaded {len(enriched_df)} rows from checkpoint: {_cp_files[-1].name}", flush=True)
            else:
                enriched_df     = pd.DataFrame()
                evidence_rows   = []
                debug_rows      = []
                jina_debug_rows = []
                print("[cleaner] No checkpoint found - output will be empty.", flush=True)
        except Exception as _cp_exc:
            print(f"[cleaner] Checkpoint load failed: {_cp_exc}. Output may be empty.", flush=True)
            enriched_df     = pd.DataFrame()
            evidence_rows   = []
            debug_rows      = []
            jina_debug_rows = []
    print()  # newline after progress

    _fc_loc_pl = _fc_loc_payload or {}
    run_meta = _build_run_meta(
        enriched_df=enriched_df,
        input_filename=input_path.name,
        run_id=file_hash, run_label=run_label,
        total_rows_input=len(df), batch_n=batch_n,
        max_queries=args.max_queries, haiku_mode=args.haiku_mode,
        haiku_model=_DEFAULT_HAIKU_MODEL, haiku_max_rows=0,
        debug_mode=args.debug,
        serper_key_present=bool(serper_key),
        anthropic_key_present=bool(anthropic_key),
        jina_mode=_JINA_MODE_OFF,
        verifier_provider=args.verifier,
        verifier_mode=_VM_UNCERTAIN,
        firecrawl_keys_loaded=len(fc_keys_cli),
        fc_health=_cli_fc_health,
        serper_queries_total=getattr(process_dataframe, "_last_serper_count", 0),
        detected_country=country_code,
        country_detection_source="cli_arg" if args.country.upper() in COUNTRY_CONFIGS else "auto",
        firecrawl_location_mode="cli_auto",
        firecrawl_location_used=cfg.country_name,
        firecrawl_location_country=_fc_loc_pl.get("country", ""),
        firecrawl_location_languages=", ".join(_fc_loc_pl.get("languages", [])),
    )

    # Build fc_audit dict for the Firecrawl Audit sheet
    _fc_enabled = args.verifier in (_VP_FIRECRAWL, _VP_FC_JINA)
    _fc_att  = _cli_fc_health.get("requests_attempted", 0)
    _fc_succ = _cli_fc_health.get("pages_successful", 0)
    _fc_audit_dict: dict | None = None
    if _fc_enabled:
        _pf_key_sts = _fc_preflight.get("key_statuses", {})
        _fc_audit_dict = {
            "firecrawl_enabled":                True,
            "firecrawl_preflight_status":       _fc_preflight.get("preflight_status", "SKIPPED"),
            "firecrawl_keys_total":             _fc_preflight.get("keys_total", len(fc_keys_cli)),
            "firecrawl_keys_ok":                _fc_preflight.get("keys_ok", ""),
            "firecrawl_keys_failed":            _fc_preflight.get("keys_failed", ""),
            "firecrawl_key_statuses_preflight": str(_pf_key_sts),
            "firecrawl_requests_attempted":     _fc_att,
            "firecrawl_pages_successful":       _fc_succ,
            "firecrawl_success_rate":           f"{_fc_succ / _fc_att:.1%}" if _fc_att else "n/a",
            "firecrawl_key_failovers":          _cli_fc_health.get("key_failovers", 0),
            "firecrawl_key_failure_events":     _cli_fc_health.get("key_failure_events", 0),
            "firecrawl_quota_or_billing_errors": _cli_fc_health.get("quota_or_billing_errors", 0),
            "firecrawl_rate_limit_errors":      _cli_fc_health.get("rate_limit_errors", 0),
            "firecrawl_timeouts":               _cli_fc_health.get("timeouts", 0),
            "firecrawl_exceptions":             _cli_fc_health.get("exceptions", 0),
            "firecrawl_consecutive_failures_max": _cli_fc_health.get("consecutive_failures_max", 0),
            "batch_firecrawl_status":           (
                "fail_fast" if _cli_fc_health.get("fail_fast_triggered")
                else "ok" if _fc_att > 0
                else "not_used"
            ),
            "batch_firecrawl_notes":            _cli_fc_health.get("fail_fast_reason", ""),
        }

    enriched_df = _add_alias_diagnostic_cols(enriched_df)

    excel_bytes = build_excel(
        enriched_df, run_df, evidence_rows, cols,
        debug_rows=debug_rows, debug_mode=args.debug,
        run_meta=run_meta, jina_debug_rows=jina_debug_rows,
        fc_audit=_fc_audit_dict,
    )

    out_path = _no_overwrite_path(Path(pl_paths["output_xlsx"]))
    _write_pipeline_output(out_path, excel_bytes)
    print(f"[cleaner] Saved:      {out_path}")

    # Run log
    _pl_log_row = {
        "timestamp":             ts,
        "cohort":                pl_paths["cohort"],
        "batch_stem":            pl_paths["batch_stem"],
        "batch_number":          pl_paths["batch_number"],
        "row_range":             pl_paths["row_range"],
        "input_path":            str(input_path),
        "output_xlsx":           pl_paths["output_xlsx"],
        "rows_processed":        batch_n,
        "rows_input":            len(df),
        "coverage_pct":          run_meta.get("coverage_pct", ""),
        "manual_review_count":   run_meta.get("manual_review_count", ""),
        "no_match_count":        run_meta.get("no_confident_match_count", ""),
        "haiku_mode":            args.haiku_mode,
        "verifier_provider":     args.verifier,
        "serper_queries":        args.max_queries,
        "firecrawl_keys_loaded": len(fc_keys_cli),
        "run_id":                file_hash,
        "run_label":             run_label,
        "status":                "complete",
        "notes":                 "",
    }
    _append_run_log_csv(Path(pl_paths["run_log_csv"]), _pl_log_row)
    import time as _time_mod2
    _elapsed_total = _time_mod2.time() - _cli_started_at

    print(f"[cleaner] Run log:    {pl_paths['run_log_csv']}", flush=True)
    _summary_text = f"\n{_fc_usage_console_summary(run_meta)}"
    try:
        print(_summary_text, flush=True)
    except UnicodeEncodeError:
        print(_summary_text.encode("ascii", errors="replace").decode("ascii"), flush=True)

    # ── End-of-run summary ────────────────────────────────────────────────────
    _rows_with_domain = int(
        enriched_df.get("final_selected_domain", pd.Series(dtype=str))
        .astype(str).str.strip().replace("", pd.NA).notna().sum()
    )
    _rows_failed = _cli_fc_health.get("exceptions", 0)
    _proc_n      = max(batch_n, 1)
    print(f"\n[cleaner] Summary:", flush=True)
    print(f"  rows_processed:                    {batch_n}", flush=True)
    print(f"  rows_with_final_domain:            {_rows_with_domain}", flush=True)
    print(f"  rows_failed:                       {_rows_failed}", flush=True)
    print(f"  serper_queries_total:              {getattr(process_dataframe, '_last_serper_count', 0)}", flush=True)
    print(f"  haiku_calls_total:                 {run_meta.get('rows_reviewed_by_haiku', 0)}", flush=True)
    _tot_att  = run_meta.get("firecrawl_total_requests_attempted", 0)
    _tot_succ = run_meta.get("firecrawl_total_pages_successful_new", 0)
    _tot_cred = run_meta.get("firecrawl_total_estimated_credits", 0)
    _fc_dv    = run_meta.get("firecrawl_domain_verification_pages", 0)
    _fc_si    = run_meta.get("firecrawl_size_inference_pages_new", 0)
    _avg_cred = round(_tot_cred / _proc_n, 2) if _tot_cred else 0.0
    print(f"  firecrawl_speed_mode:              {_cli_fc_speed} / max_cands={_cli_fc_max_cands} / max_pages={_cli_fc_max_pages} / timeout={_cli_fc_page_timeout}s", flush=True)
    if _cli_fc_budget:
        print(f"  firecrawl_budget:                  {_cli_fc_budget} pages ({'LIMIT HIT - partial output' if _budget_hit else 'not reached'})", flush=True)
    print(f"  firecrawl_total_requests_attempted:{_tot_att}", flush=True)
    print(f"  firecrawl_total_pages_successful:  {_tot_succ}", flush=True)
    print(f"  firecrawl_domain_verification_pages:{_fc_dv}", flush=True)
    print(f"  firecrawl_size_inference_pages:    {_fc_si}", flush=True)
    print(f"  firecrawl_total_estimated_credits: {_tot_cred}", flush=True)
    print(f"  firecrawl_avg_credits_per_row:     {_avg_cred}", flush=True)
    print(f"  elapsed:                           {_format_duration(_elapsed_total)}", flush=True)
    print(f"  output_file:                       {out_path}", flush=True)

    print(f"\n[cleaner] Done.", flush=True)
    if args.infer_size and "employee_size_band" in enriched_df.columns:
        _band_counts = enriched_df["employee_size_band"].value_counts().to_dict()
        _hrm_vals = pd.to_numeric(enriched_df.get("hrm_likelihood_score", pd.Series(dtype=float)), errors="coerce")
        _hrm_avg = _hrm_vals.mean()
        _fc_size_total = pd.to_numeric(enriched_df.get("firecrawl_pages_used_for_size", pd.Series(dtype=float)), errors="coerce").sum()
        print(f"\n[size inference] Employee size band distribution: {_band_counts}")
        print(f"[size inference] HRM likelihood avg: {_hrm_avg:.1f}")
        print(f"[size inference] FC pages used for size inference: {int(_fc_size_total)}")


# ── Streamlit autosave / download helpers ─────────────────────────────────────

def _prepared_download_key(suffix: str) -> str:
    """Stable session_state key for a lazily-prepared checkpoint download."""
    return f"_cp_dl_{suffix}"


def _clear_prepared_checkpoint_downloads() -> None:
    """Remove all prepared checkpoint download bytes from session_state."""
    to_remove = [k for k in st.session_state if k.startswith("_cp_dl_")]
    for k in to_remove:
        del st.session_state[k]


def _archive_autosave_dir() -> None:
    """Rename autosave/ to autosave_archive_YYYYMMDD_HHMMSS; recreate empty autosave/."""
    from datetime import datetime as _dt
    src = _AUTOSAVE_DIR
    if src.exists():
        dst = src.parent / f"autosave_archive_{_dt.now().strftime('%Y%m%d_%H%M%S')}"
        src.rename(dst)
    _AUTOSAVE_DIR.mkdir(parents=True, exist_ok=True)


def _format_duration(seconds) -> str:
    """Format a duration in seconds to MM:SS or HH:MM:SS. Safe against None/NaN."""
    try:
        s = float(seconds)
        if s != s or s < 0:  # NaN or negative
            return "??:??"
        s = int(s)
        h, rem = divmod(s, 3600)
        m, sec = divmod(rem, 60)
        if h > 0:
            return f"{h:02d}:{m:02d}:{sec:02d}"
        return f"{m:02d}:{sec:02d}"
    except (TypeError, ValueError):
        return "??:??"


def _print_cli_progress(
    prefix: str,
    i: int,
    total: int,
    started_at: float,
    current_name: str = "",
    counters: dict | None = None,
    every: int = 10,
) -> None:
    """
    Print a CLI progress line if this row warrants one.
    Prints: first row, last row, every row when total<=5, every 5 when total<=100,
    every `every` rows otherwise.
    """
    import time as _time
    if total <= 5 or i == 1 or i == total:
        do_print = True
    elif total <= 100:
        do_print = (i % 5 == 0)
    else:
        do_print = (i % every == 0)
    if not do_print:
        return

    elapsed = _time.time() - started_at
    pct = i / total * 100 if total else 0.0
    if i > 0 and elapsed > 0:
        eta = elapsed / i * (total - i)
    else:
        eta = 0.0

    name_part = (current_name[:57] + "...") if len(current_name) > 60 else current_name
    c = counters or {}
    serper_n  = c.get("serper_queries", 0)
    fc_att    = c.get("requests_attempted", 0)
    fc_succ   = c.get("pages_successful", 0)
    haiku_n   = c.get("haiku_calls", 0)
    errors_n  = c.get("errors", 0)
    skipped_n = c.get("skipped", 0)

    line = (
        f"{prefix} {i}/{total} ({pct:.1f}%) | "
        f"elapsed {_format_duration(elapsed)} | ETA {_format_duration(eta)} | "
        f"current: {name_part} | "
        f"Serper: {serper_n} | FC: {fc_att}req/{fc_succ}ok | "
        f"Haiku: {haiku_n} | errors: {errors_n}"
    )
    if skipped_n:
        line += f" | skipped: {skipped_n}"
    print(line, flush=True)


def main():
    st.title("🌍 Input Cleaner · Register Edition")
    st.caption(
        "Layer 0 · mYngle Sales Intelligence · "
        "Cleans company register exports (Italy / Germany) before Lead Prioritizer enrichment  \n"
        "Website Discovery v2 — multi-variant brand extraction, 8 query strategies, "
        "category rejection, brand-similarity gate"
    )

    # ── Sidebar: API key + settings ───────────────────────────────────────────
    st.sidebar.header("Settings")

    # Country selector (top of sidebar — affects column detection and FC location default)
    _COUNTRY_OPTIONS = ["Auto-detect", "Italy", "Germany"]
    _sidebar_country = st.sidebar.selectbox(
        "Country / pipeline",
        options=_COUNTRY_OPTIONS,
        index=0,
        key="reg_country_mode",
        help=(
            "Auto-detect: determined from the uploaded filename and column names.  \n"
            "Italy: force Italian pipeline (ATECO, PEC, .it TLD).  \n"
            "Germany: force German pipeline (Handelsregister, .de TLD, federal_state)."
        ),
    )
    _SIDEBAR_COUNTRY_MAP = {"Auto-detect": "auto", "Italy": "IT", "Germany": "DE"}
    _sidebar_country_code = _SIDEBAR_COUNTRY_MAP[_sidebar_country]
    st.sidebar.markdown("---")

    serper_key = _load_secrets_key()
    if serper_key:
        st.sidebar.success("✓ Serper API key loaded from secrets.")
    else:
        st.sidebar.warning(
            "No Serper API key found in `.streamlit/secrets.toml`.\n\n"
            "Serper domain search will be skipped. "
            "Only website normalisation and email-domain fallback will run."
        )
        manual_key = st.sidebar.text_input(
            "Paste Serper API key (optional)", type="password", key="reg_serper"
        )
        if manual_key.strip():
            serper_key = manual_key.strip()

    # Anthropic API key for Haiku
    anthropic_key = None
    try:
        anthropic_key = st.secrets.get("ANTHROPIC_API_KEY") or st.secrets.get("anthropic_api_key")
    except Exception:
        pass

    st.sidebar.markdown("---")
    max_queries = st.sidebar.selectbox(
        "Max Serper queries per company",
        options=[3, 5, 8],
        index=1,
        help=(
            "3 = fast/cheap · 5 = default, good balance · 8 = maximum discovery.\n\n"
            "Each query costs 1 Serper credit. For 200 companies: "
            "3 queries ≈ up to 600 credits, 5 ≈ up to 1000, 8 ≈ up to 1600."
        ),
    )
    st.sidebar.caption(
        f"Each missing website tries up to {max_queries} search strategies."
    )
    st.sidebar.markdown("---")
    st.sidebar.subheader("Claude Haiku Review (Experiment)")
    if not _ANTHROPIC_AVAILABLE:
        st.sidebar.warning(
            "Install `anthropic` package to enable Haiku review:  \n"
            "`pip install anthropic`"
        )
    haiku_mode = st.sidebar.selectbox(
        "Haiku review mode",
        options=_HAIKU_MODES,
        index=_HAIKU_MODES.index(_HAIKU_MODE_UNCERTAIN),
        help=(
            "Python only: use existing scoring only (no Haiku calls, no cost).  \n"
            "Uncertain rows: call Haiku only for rows where Python is not confident.  \n"
            "All rows: call Haiku for every row (most accurate, higher cost)."
        ),
    )
    if haiku_mode != _HAIKU_MODE_PYTHON:
        if anthropic_key:
            st.sidebar.success("✓ Anthropic API key loaded from secrets.")
        else:
            manual_anthropic = st.sidebar.text_input(
                "Paste Anthropic API key", type="password", key="reg_anthropic"
            )
            if manual_anthropic.strip():
                anthropic_key = manual_anthropic.strip()
            if not anthropic_key:
                st.sidebar.warning("Haiku review requires an Anthropic API key.")

        haiku_model = st.sidebar.text_input(
            "Haiku model ID",
            value=_DEFAULT_HAIKU_MODEL,
            key="reg_haiku_model",
        )
        haiku_max_rows = st.sidebar.number_input(
            "Max rows to send to Haiku (0 = all)",
            min_value=0, max_value=5000, value=0, step=10,
            key="reg_haiku_max_rows",
            help=(
                "0 = all rows selected by the current Haiku mode.  \n"
                "For 'Haiku for uncertain rows only', this means all uncertain rows, "
                "not all input rows.  \n"
                "Set a positive number to cap Haiku calls during testing."
            ),
        )
        _est_calls = haiku_max_rows if haiku_max_rows > 0 else "all eligible"
        st.sidebar.caption(
            f"Estimated Haiku calls: up to **{_est_calls}** rows.  \n"
            "Haiku input/output tokens ≈ 800/100 per row."
        )
    else:
        haiku_model    = _DEFAULT_HAIKU_MODEL
        haiku_max_rows = 0

    st.sidebar.markdown("---")
    st.sidebar.subheader("Organization Eligibility Filter")
    eligibility_filter_mode = st.sidebar.selectbox(
        "Eligibility filter mode",
        options=_PF_MODES,
        index=_PF_MODES.index(_PF_MODE_MAYBE),
        key="reg_eligibility_filter_mode",
        help=(
            "Commercial only: process SPA, SRL, SNC, SAS, etc. Skip associations, schools, "
            "public bodies, cooperatives.  \n"
            "Commercial + Maybe: also process cooperatives, consortia, agricultural companies.  \n"
            "All rows: no pre-filtering — process every row regardless of legal form."
        ),
    )

    st.sidebar.markdown("---")
    st.sidebar.subheader("Website Verifier")
    verifier_provider = st.sidebar.selectbox(
        "Website verifier provider",
        options=_VP_OPTIONS,
        index=_VP_OPTIONS.index(_VP_FIRECRAWL),
        key="reg_verifier_provider",
        help=(
            "Off: skip website verification entirely.  \n"
            "Jina: lightweight free verifier (no API key required).  \n"
            "Firecrawl: high-fidelity scraper (API key required).  \n"
            "Firecrawl first, Jina fallback: use Firecrawl; fall back to Jina when uncertain."
        ),
    )
    verifier_mode = st.sidebar.selectbox(
        "Verification mode",
        options=_VM_OPTIONS,
        index=0,
        key="reg_verifier_mode",
        help=(
            "Uncertain candidates only: verify when Python scoring is ambiguous.  \n"
            "All selected candidates in debug mode: verify every row when debug mode is ON."
        ),
    )
    verifier_speed_mode = st.sidebar.selectbox(
        "Firecrawl speed mode",
        options=_FC_SPEED_OPTIONS,
        index=_FC_SPEED_OPTIONS.index(_FC_SPEED_FAST),
        key="reg_verifier_speed_mode",
        help=(
            "Fast: homepage only — 1 candidate, 1 page, 6 s timeout.  \n"
            "Balanced: homepage + about/contact only when homepage is weak — 1 candidate, 2 pages, 8 s.  \n"
            "Thorough: homepage + about + contact — configurable below."
        ),
    ) if verifier_provider in (_VP_FIRECRAWL, _VP_FC_JINA) else _FC_SPEED_FAST

    # Default max_cands / max_pages / timeout from speed mode; allow override in Thorough
    _spd_defaults = _FC_SPEED_DEFAULTS.get(verifier_speed_mode, _FC_SPEED_DEFAULTS[_FC_SPEED_FAST])
    if verifier_speed_mode == _FC_SPEED_THOROUGH and verifier_provider in (_VP_FIRECRAWL, _VP_FC_JINA):
        verifier_max_candidates = int(st.sidebar.number_input(
            "Max candidates per company", min_value=1, max_value=10, value=_spd_defaults[0], step=1,
            key="reg_verifier_max_cands",
        ))
        verifier_max_pages = int(st.sidebar.number_input(
            "Max pages per candidate (request budget)", min_value=1, max_value=6, value=_spd_defaults[1], step=1,
            key="reg_verifier_max_pages",
        ))
        verifier_page_timeout = int(st.sidebar.number_input(
            "Page timeout (seconds)", min_value=5, max_value=60, value=_spd_defaults[2], step=5,
            key="reg_verifier_page_timeout",
        ))
    else:
        verifier_max_candidates = _spd_defaults[0]
        verifier_max_pages      = _spd_defaults[1]
        verifier_page_timeout   = _spd_defaults[2]
        if verifier_provider in (_VP_FIRECRAWL, _VP_FC_JINA):
            st.sidebar.caption(
                f"Speed defaults: {verifier_max_candidates} cand · "
                f"{verifier_max_pages} page(s)/cand · {verifier_page_timeout}s timeout"
            )

    # Firecrawl location mode — default is "Auto from detected country"
    fc_location_label = st.sidebar.selectbox(
        "Firecrawl location",
        options=_FC_LOC_OPTIONS,
        index=0,  # Auto from detected country is always index 0
        key="reg_fc_location",
        help=(
            "Auto from detected country: uses Germany location for DE files, Italy for IT files.  \n"
            "Italy: sends scrape requests as if from Italy (country=IT, languages=[it,en]).  \n"
            "Germany: sends scrape requests as if from Germany (country=DE, languages=[de,en]).  \n"
            "United States: US infrastructure.  \n"
            "Default Firecrawl: no location override."
        ),
    ) if verifier_provider in (_VP_FIRECRAWL, _VP_FC_JINA) else _FC_LOC_AUTO
    # fc_location_payload is resolved after country detection (below, when file is uploaded)
    # For sidebar-only state (no file yet), default to auto
    _fc_location_mode = fc_location_label  # store the mode label for later resolution

    # ── API keys ──────────────────────────────────────────────────────────────
    # Jina API key (optional — Jina Reader works without a key at lower rate limits)
    jina_api_key: str | None = None
    try:
        jina_api_key = st.secrets.get("JINA_API_KEY") or st.secrets.get("jina_api_key")
    except Exception:
        pass
    if verifier_provider in (_VP_JINA, _VP_FC_JINA):
        if jina_api_key:
            st.sidebar.success("✓ Jina API key loaded from secrets.")
        else:
            _manual_jina = st.sidebar.text_input(
                "Jina API key (optional — improves rate limits)",
                type="password",
                key="reg_jina_key",
            )
            if _manual_jina.strip():
                jina_api_key = _manual_jina.strip()

    # Firecrawl API keys: secrets → env → manual input (never displayed)
    _fc_keys: list[str] = _fc_load_keys()
    if verifier_provider in (_VP_FIRECRAWL, _VP_FC_JINA):
        if _fc_keys:
            st.sidebar.success(f"✓ Firecrawl API keys loaded: {len(_fc_keys)}")
        else:
            _fc_key_input = st.sidebar.text_input(
                "Firecrawl API key",
                type="password",
                key="reg_firecrawl_key",
            )
            if _fc_key_input.strip():
                _fc_keys = [_fc_key_input.strip()]
        if st.sidebar.button("Preflight-check Firecrawl keys", key="reg_fc_preflight"):
            if not _fc_keys:
                st.sidebar.warning("No Firecrawl API key provided.")
            else:
                with st.sidebar:
                    with st.spinner("Running preflight check…"):
                        _pf = firecrawl_preflight_check(_fc_keys)
                _pf_st = _pf.get("preflight_status", "UNKNOWN")
                _pf_ok = _pf.get("keys_ok", 0)
                _pf_tot = _pf.get("keys_total", 0)
                _pf_msg = f"Preflight: **{_pf_st}** · {_pf_ok}/{_pf_tot} keys OK"
                if _pf_st == "OK":
                    st.sidebar.success(_pf_msg)
                elif _pf_st == "DEGRADED":
                    st.sidebar.warning(_pf_msg)
                else:
                    st.sidebar.error(_pf_msg + "  \nAll keys failed — check account status.")
                st.session_state["_fc_preflight_result"] = _pf

    if st.sidebar.button("Test Firecrawl connection", key="reg_firecrawl_test"):
            if not _fc_keys:
                st.sidebar.warning("No Firecrawl API key provided.")
            else:
                _test_url = "https://www.firecrawl.dev"
                _any_ok = False
                _first_ok_idx = -1
                for _tki, _tkey in enumerate(_fc_keys, start=1):
                    _fc_start = time.time()
                    try:
                        _fc_resp = requests.post(
                            "https://api.firecrawl.dev/v1/scrape",
                            headers={
                                "Authorization": f"Bearer {_tkey}",
                                "Content-Type": "application/json",
                            },
                            json={"url": _test_url, "formats": ["markdown"]},
                            timeout=20,
                        )
                        _fc_elapsed = round(time.time() - _fc_start, 2)
                        if _fc_resp.status_code == 200:
                            _fc_data = _fc_resp.json()
                            _fc_chars = len(
                                ((_fc_data.get("data") or {}).get("markdown", "") or str(_fc_data))
                            )
                            st.sidebar.success(
                                f"Key {_tki}: ✓ OK · {_fc_chars:,} chars · {_fc_elapsed}s"
                            )
                            _any_ok = True
                            if _first_ok_idx < 0:
                                _first_ok_idx = _tki
                        else:
                            _is_quota = _fc_is_key_failure(_fc_resp.status_code, _fc_resp.text)
                            _label = "quota/auth" if _is_quota else f"HTTP {_fc_resp.status_code}"
                            st.sidebar.error(
                                f"Key {_tki}: ✗ {_label} · {_fc_elapsed}s"
                            )
                    except requests.Timeout:
                        _fc_elapsed = round(time.time() - _fc_start, 2)
                        st.sidebar.error(f"Key {_tki}: ✗ Timeout after {_fc_elapsed}s")
                    except Exception as _fc_exc:
                        _fc_elapsed = round(time.time() - _fc_start, 2)
                        st.sidebar.error(
                            f"Key {_tki}: ✗ Error · {_fc_elapsed}s  \n{str(_fc_exc)[:120]}"
                        )
                if _any_ok and _first_ok_idx > 1:
                    st.sidebar.info(
                        f"✓ Firecrawl failover available: key {_first_ok_idx} works"
                    )
    # Expose as a single key for callers that still use the old scalar API
    _fc_key: str | list[str] | None = _fc_keys if _fc_keys else None

    # legacy jina_mode — kept for backward compat with checkpoint resume logic
    jina_mode = _JINA_MODE_OFF

    st.sidebar.markdown("---")
    debug_mode = st.sidebar.checkbox(
        "Candidate Discovery Debug Mode",
        value=False,
        key="reg_debug_mode",
        help=(
            "When enabled, exports an extra Excel sheet with every Serper result for "
            "every company: title, snippet, url, score, rejection reason, brand overlap, etc. "
            "Useful for diagnosing why certain companies fail. Increases file size."
        ),
    )
    if debug_mode:
        st.sidebar.caption(
            "🔍 Debug sheet will include all raw Serper candidates + rejection reasons."
        )

    st.sidebar.markdown("---")
    infer_size = st.sidebar.checkbox(
        "Infer company size and HR signals",
        value=False,
        help=(
            "Uses Serper snippets and up to 2 Firecrawl pages per company to estimate "
            "employee count and HRM signal strength. Uses extra credits."
        ),
    )

    st.sidebar.markdown("---")
    st.sidebar.subheader("Pipeline output (optional)")
    st.sidebar.caption(
        "When enabled, the cleaned Excel is automatically written to the standard "
        "`01_cleaned_domains/` folder next to the source `00_raw/` folder, "
        "and a run-log entry is appended to `_logs/{cohort}_cleaner_runlog.csv`."
    )
    pipeline_output_enabled = st.sidebar.checkbox(
        "Save to pipeline folder automatically",
        value=False,
        key="reg_pipeline_output",
        help=(
            "Requires the input file to live inside a …/{cohort}/00_raw/ folder.  \n"
            "Output: …/{cohort}/01_cleaned_domains/{batch_stem}_cleaned_YYYYMMDD_HHMM.xlsx  \n"
            "Log:    …/{cohort}/_logs/{cohort}_cleaner_runlog.csv"
        ),
    )
    pipeline_project_root = ""
    if pipeline_output_enabled:
        pipeline_project_root = st.sidebar.text_input(
            "Project root (leave blank to auto-detect from file path)",
            value="",
            key="reg_pipeline_project_root",
            help=(
                "e.g. C:\\Users\\gertm\\Nextcloud\\Myngle  \n"
                "Leave blank: auto-detected as grandparent of 00_raw/."
            ),
        ).strip()

    st.sidebar.markdown("---")
    st.sidebar.caption(
        f"Autosave every **{_AUTOSAVE_EVERY} rows** → `{_AUTOSAVE_DIR}/`  \n"
        "If the app crashes or your browser refreshes, reopen the app and use "
        "**Resume previous run** to continue without reprocessing completed rows."
    )

    # ── Previous-run panel (lazy — only rendered when checkbox is checked) ────
    checkpoints = _list_checkpoints()
    if checkpoints:
        n_cp = len(checkpoints)
        st.sidebar.markdown(f"📂 Previous runs available: **{n_cp}**")
        if st.sidebar.button("🗄 Archive all previous runs", key="reg_archive_autosave",
                             help="Renames autosave/ to autosave_archive_YYYYMMDD_HHMMSS. Nothing is deleted."):
            _archive_autosave_dir()
            _clear_prepared_checkpoint_downloads()
            st.rerun()
        show_prev = st.sidebar.checkbox("Show previous runs", value=False, key="reg_show_prev_runs")
    else:
        show_prev = False

    if show_prev and checkpoints:
        st.markdown("### 📂 Previous runs")
        for idx, cp_meta in enumerate(checkpoints[:8]):
            run_id_cp  = cp_meta.get("run_id", "?")
            folder_cp  = cp_meta.get("_folder", "") or run_id_cp
            key_suffix = f"{idx}_{folder_cp}"
            row_idx    = cp_meta.get("row_idx", 0)
            total      = cp_meta.get("total_rows", "?")
            ts         = str(cp_meta.get("timestamp", "?"))[:19]
            complete   = cp_meta.get("complete", False)
            pct        = f"{row_idx / total * 100:.0f}%" if isinstance(total, int) and total else "?"
            run_label_cp = cp_meta.get("run_label", "") or cp_meta.get("_folder", run_id_cp[:8])
            label_str  = (
                f"{'✅ Complete' if complete else '⏸ Partial'} · "
                f"**{row_idx}/{total}** rows ({pct}) · saved {ts}  \n"
                f"`{run_label_cp}`"
            )

            c1, c2, c3, c4 = st.columns([6, 2, 2, 2])
            c1.markdown(label_str)

            if c2.button("Resume", key=f"resume_{key_suffix}", use_container_width=True):
                cp_data = _load_checkpoint(run_id_cp)
                if cp_data and cp_data.get("input_df") is not None:
                    st.session_state["reg_resume_data"] = cp_data
                    st.session_state["reg_run_id"]      = run_id_cp
                    st.rerun()
                else:
                    st.error("Could not load checkpoint (input snapshot missing). Re-upload the file.")

            # Lazy download: only build Excel when user clicks "Prepare download"
            saved_cols = cp_meta.get("cols", {})
            dl_key = _prepared_download_key(key_suffix)
            prepared = st.session_state.get(dl_key)
            if prepared:
                c3.download_button(
                    "⬇ Download",
                    data=prepared["bytes"],
                    file_name=prepared["filename"],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_{key_suffix}",
                    use_container_width=True,
                )
            else:
                if c3.button("Prepare download", key=f"prep_{key_suffix}", use_container_width=True):
                    xl = _checkpoint_excel_bytes(run_id_cp, saved_cols)
                    if xl:
                        st.session_state[dl_key] = {
                            "bytes": xl,
                            "filename": f"partial_{run_id_cp[:8]}.xlsx",
                        }
                        st.rerun()
                    else:
                        st.error("Could not build Excel from checkpoint.")

            if c4.button("Delete", key=f"del_{key_suffix}", use_container_width=True):
                _delete_checkpoint(run_id_cp)
                st.session_state.pop(dl_key, None)
                st.rerun()

    # ── Handle resume-from-checkpoint (no upload needed if snapshot present) ─
    resume_data = st.session_state.get("reg_resume_data")
    if resume_data is not None:
        cp_meta    = resume_data["meta"]
        run_id     = cp_meta["run_id"]
        resume_from = cp_meta["row_idx"]
        total_rows  = cp_meta["total_rows"]
        prior_results  = resume_data["results"]
        prior_evidence = resume_data["evidence"]
        resume_input_df = resume_data["input_df"]
        saved_cols = cp_meta.get("cols", {})

        st.info(
            f"⏸ **Resuming run `{run_id}`** — "
            f"{resume_from} of {total_rows} rows already completed.  \n"
            "Re-upload your file below to continue from where processing stopped, "
            "or click **Start fresh** to reprocess from the beginning."
        )

        uploaded = st.file_uploader(
            "Re-upload the same file to continue (or upload a new file for a fresh run)",
            type=["csv", "xlsx"],
            key="reg_upload_resume",
        )

        col_a, col_b = st.columns(2)
        if col_b.button("✖ Start fresh instead", use_container_width=True):
            st.session_state.pop("reg_resume_data", None)
            st.session_state.pop("reg_run_id", None)
            st.rerun()

        if uploaded is None:
            # Lazy download: only build Excel when user clicks the prepare button
            _resume_dl_key = _prepared_download_key(f"resume_{run_id[:8]}")
            _resume_prepared = st.session_state.get(_resume_dl_key)
            if _resume_prepared:
                st.download_button(
                    f"⬇ Download partial results ({resume_from} rows so far)",
                    data=_resume_prepared["bytes"],
                    file_name=_resume_prepared["filename"],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            else:
                if st.button("Prepare partial download", key=f"prep_resume_{run_id[:8]}",
                             use_container_width=True):
                    _xl = _checkpoint_excel_bytes(run_id, saved_cols)
                    if _xl:
                        st.session_state[_resume_dl_key] = {
                            "bytes": _xl,
                            "filename": f"partial_{run_id[:8]}.xlsx",
                        }
                        st.rerun()
                    else:
                        st.error("Could not build partial Excel from checkpoint.")
            return

        raw_bytes  = uploaded.read()
        file_hash  = _file_hash(raw_bytes)
        upload_run_id = file_hash   # run_id based on file content

        if upload_run_id != run_id:
            # Different file uploaded — treat as fresh run
            st.warning(
                "The uploaded file does not match the previous run's input.  \n"
                "Starting a fresh run with this file."
            )
            st.session_state.pop("reg_resume_data", None)
            prior_results  = []
            prior_evidence = []
            resume_from    = 0
            run_id = upload_run_id
        else:
            st.success(
                f"✅ File matches previous run. Will continue from row **{resume_from + 1}**."
            )

        df = _parse_bytes(raw_bytes, uploaded.name)
        if df is None:
            return

        # Resolve country for resume path using saved cols meta or sidebar setting
        if _sidebar_country_code in ("IT", "DE"):
            _resume_country = _sidebar_country_code
        else:
            _resume_country = resolve_country("auto", uploaded.name, df)
        _resume_cfg = COUNTRY_CONFIGS.get(_resume_country, IT_CONFIG)
        df, _ = normalize_register_columns_for_cleaner(df, _resume_cfg)
        cols = _cols_from_normalized(df, _resume_cfg)
        # Apply saved column mapping (overrides detection when columns are preserved)
        for role, col_name in saved_cols.items():
            if col_name and col_name in df.columns:
                cols[role] = col_name

        run_df  = df.head(total_rows).copy()
        # Recompute a run label for this resume session (new timestamp + current settings)
        _resume_label = _make_run_label(haiku_mode, total_rows, int(max_queries), debug_mode)
        _resume_filename = _make_filename(_resume_label, run_id)
        settings_dict = {
            "max_queries": int(max_queries), "batch_n": total_rows,
            "haiku_mode": haiku_mode, "debug_mode": debug_mode,
            "jina_mode": jina_mode,
            "run_label": _resume_label,
        }

        if col_a.button(
            f"▶ Continue from row {resume_from + 1}", type="primary", use_container_width=True
        ):
            progress_bar = st.progress(resume_from / total_rows if total_rows else 0.0)
            status_text  = st.empty()

            def progress_cb(i, total):
                progress_bar.progress(i / total)
                status_text.caption(f"Processing {i} / {total}…")

            enriched_df, evidence_rows, debug_rows, jina_debug_rows = process_dataframe(
                run_df, cols, serper_key, int(max_queries),
                progress_cb=progress_cb,
                run_id=run_id,
                resume_from=resume_from,
                prior_results=prior_results,
                prior_evidence=prior_evidence,
                settings=settings_dict,
                run_label=_resume_label,
                haiku_mode=haiku_mode,
                haiku_api_key=anthropic_key,
                haiku_model=haiku_model,
                haiku_max_rows=int(haiku_max_rows),
                jina_mode=jina_mode,
                jina_api_key=jina_api_key,
                verifier_provider=verifier_provider,
                verifier_mode=verifier_mode,
                fc_key=_fc_key,
                max_cands_per_company=verifier_max_candidates,
                max_pages_per_cand=verifier_max_pages,
                page_timeout=verifier_page_timeout,
                fc_speed_mode=verifier_speed_mode,
                fc_location=fc_location_payload,
                eligibility_filter_mode=eligibility_filter_mode,
                debug_mode=debug_mode,
                infer_size=infer_size,
                country_config=_resume_cfg,
            )

            progress_bar.progress(1.0)
            status_text.caption(f"Done — {total_rows} companies processed.")

            # Build run_meta for summary sheet
            _run_meta_resume = _build_run_meta(
                enriched_df=enriched_df,
                input_filename="(resumed run)",
                run_id=run_id, run_label=_resume_label,
                total_rows_input=total_rows, batch_n=total_rows,
                max_queries=int(max_queries), haiku_mode=haiku_mode,
                haiku_model=haiku_model, haiku_max_rows=int(haiku_max_rows),
                debug_mode=debug_mode,
                serper_key_present=bool(serper_key),
                anthropic_key_present=bool(anthropic_key),
                jina_mode=jina_mode,
                verifier_provider=verifier_provider,
                verifier_mode=verifier_mode,
                firecrawl_keys_loaded=len(_fc_keys),
                fc_health={},
                serper_queries_total=getattr(process_dataframe, "_last_serper_count", 0),
            )

            st.session_state["reg_enriched"]    = enriched_df
            st.session_state["reg_evidence"]    = evidence_rows
            st.session_state["reg_debug"]       = debug_rows
            st.session_state["reg_jina_debug"]  = jina_debug_rows
            st.session_state["reg_original"]   = run_df
            st.session_state["reg_cols"]       = cols
            st.session_state["reg_run_id"]     = run_id
            st.session_state["reg_run_label"]  = _resume_label
            st.session_state["reg_filename"]   = _resume_filename
            st.session_state["reg_run_meta"]   = _run_meta_resume
            st.session_state["reg_debug_mode_value"] = debug_mode
            st.session_state.pop("reg_resume_data", None)

            # Mark checkpoint complete
            _save_checkpoint(
                run_id, list(prior_results or []) + evidence_rows,
                evidence_rows, total_rows, total_rows, run_df, cols, settings_dict,
                run_label=_resume_label,
            )

        # Show partial download while waiting for user to click Continue (lazy)
        else:
            _cont_dl_key = _prepared_download_key(f"continue_{run_id[:8]}")
            _cont_prepared = st.session_state.get(_cont_dl_key)
            if _cont_prepared:
                st.download_button(
                    f"⬇ Download partial results ({resume_from} rows so far)",
                    data=_cont_prepared["bytes"],
                    file_name=_cont_prepared["filename"],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            else:
                if st.button("Prepare partial download", key=f"prep_continue_{run_id[:8]}",
                             use_container_width=True):
                    _xl = _checkpoint_excel_bytes(run_id, cols)
                    if _xl:
                        st.session_state[_cont_dl_key] = {
                            "bytes": _xl,
                            "filename": f"partial_{run_id[:8]}.xlsx",
                        }
                        st.rerun()
                    else:
                        st.error("Could not build partial Excel from checkpoint.")

        # Fall through to results section if enriched_df is available
        enriched_df = st.session_state.get("reg_enriched")
        if enriched_df is not None:
            st.markdown("---")
            st.markdown("### Results")
            stored_cols  = st.session_state.get("reg_cols", cols)
            _stored_label = st.session_state.get("reg_run_label", _resume_label)
            _stored_fn    = st.session_state.get("reg_filename", _resume_filename)
            _stored_dm    = st.session_state.get("reg_debug_mode_value", debug_mode)
            _summary_metrics(enriched_df, stored_cols)
            st.markdown("")
            _show_results(enriched_df, stored_cols)
            _show_run_info_block(
                run_label=_stored_label, run_id=run_id,
                batch_n=total_rows, max_queries=int(max_queries),
                haiku_mode=haiku_mode, debug_mode=_stored_dm,
                filename=_stored_fn,
            )
            st.markdown("---")
            _download_section(
                enriched_df,
                st.session_state.get("reg_original", run_df),
                st.session_state.get("reg_evidence", []),
                stored_cols,
                filename=_stored_fn,
                debug_rows=st.session_state.get("reg_debug", []),
                debug_mode=_stored_dm,
                run_meta=st.session_state.get("reg_run_meta"),
                jina_debug_rows=st.session_state.get("reg_jina_debug", []),
                fc_audit=st.session_state.get("_fc_audit"),
            )
        return   # ← resume path ends here

    # =========================================================================
    # NORMAL (fresh) UPLOAD PATH
    # =========================================================================

    uploaded = st.file_uploader(
        "Upload company register export (CSV or Excel .xlsx)",
        type=["csv", "xlsx"],
        key="reg_upload",
        help="Upload an Italian or German company-register export. The app will auto-detect the country and map the relevant columns.",
    )

    if uploaded is None:
        _hint_country = _sidebar_country_code if _sidebar_country_code != "auto" else "IT/DE"
        if _sidebar_country_code == "DE":
            _col_hint = "**company_name_clean**, **city_or_registered_office**, **federal_state**, **registered_address**"
        elif _sidebar_country_code == "IT":
            _col_hint = "**Company Name**, **Website**, **Email address**, **City**, **Province**, **Postal Code**, **Phone number**"
        else:
            _col_hint = "**Company Name / company_name_clean**, **Website**, **Email / city_or_registered_office**, …"
        st.info(
            f"Upload a CSV or Excel export from an Italian or German company register.  \n"
            f"Expected columns ({_hint_country}): {_col_hint}."
        )
        return

    raw_bytes = uploaded.read()
    file_hash = _file_hash(raw_bytes)
    run_id    = file_hash   # one run_id per unique file

    df = _parse_bytes(raw_bytes, uploaded.name)
    if df is None:
        return

    # ── Resolve country from sidebar setting, filename, and columns ───────────
    if _sidebar_country_code in ("IT", "DE"):
        _resolved_country = _sidebar_country_code
        _country_source = "manual"
    else:
        _resolved_country = resolve_country("auto", uploaded.name, df)
        _country_source = "auto"
    _country_cfg = COUNTRY_CONFIGS.get(_resolved_country, IT_CONFIG)

    # Resolve FC location payload from the mode label + detected country
    if _fc_location_mode == _FC_LOC_AUTO:
        fc_location_payload = _fc_loc_payload_for_country(_resolved_country)
        _fc_location_used_label = f"Auto → {_country_cfg.country_name}"
    else:
        fc_location_payload = _FC_LOC_PAYLOADS.get(_fc_location_mode) or _fc_loc_payload_for_country(_resolved_country)
        _fc_location_used_label = _fc_location_mode

    _country_flag = {"IT": "🇮🇹", "DE": "🇩🇪"}.get(_resolved_country, "🌍")
    st.title(f"{_country_flag} Input Cleaner · Register Edition")

    _country_label = _country_cfg.country_name
    if _country_source == "manual":
        st.info(
            f"Country selected manually: **{_country_label} ({_resolved_country})**  \n"
            f"Firecrawl location: **{_fc_location_used_label}**"
        )
    else:
        st.info(
            f"Detected country: **{_country_label} ({_resolved_country})**  \n"
            f"Firecrawl location: **{_fc_location_used_label}**"
        )

    st.success(f"✅ Loaded **{len(df)} companies**, {len(df.columns)} columns from `{uploaded.name}`")

    # Offer resume if a checkpoint exists for this exact file
    existing_cp = _load_checkpoint(run_id)
    if existing_cp and not existing_cp["meta"].get("complete", False):
        row_idx   = existing_cp["meta"].get("row_idx", 0)
        total_rows_saved = existing_cp["meta"].get("total_rows", len(df))
        ts = str(existing_cp["meta"].get("timestamp", ""))[:19]
        st.warning(
            f"⏸ **Unfinished run found** for this file — "
            f"**{row_idx}/{total_rows_saved}** rows completed (saved {ts}).  \n"
            "Click **Resume** to continue, or **Start fresh** to reprocess from the beginning."
        )
        rc1, rc2 = st.columns(2)
        if rc1.button("⏩ Resume previous run", type="primary", use_container_width=True):
            st.session_state["reg_resume_data"] = existing_cp
            st.session_state["reg_run_id"]      = run_id
            st.rerun()
        if rc2.button("🔄 Start fresh (discard saved progress)", use_container_width=True):
            _delete_checkpoint(run_id)
            st.rerun()
        return

    # ── Normalize to canonical columns ────────────────────────────────────────
    df, _st_norm_report = normalize_register_columns_for_cleaner(df, _country_cfg)
    cols = _cols_from_normalized(df, _country_cfg)

    with st.expander("Column mapping (auto-detected)", expanded=False):
        col_options = ["(none)"] + list(df.columns)

        def _sel(label, role, default):
            cur = cols.get(role)
            idx = col_options.index(cur) if cur and cur in col_options else 0
            chosen = st.selectbox(label, col_options, index=idx, key=f"col_{role}")
            return None if chosen == "(none)" else chosen

        cols["company"]  = _sel("Company name column",  "company",  _REG_COL_COMPANY)
        cols["website"]  = _sel("Website column",        "website",  _REG_COL_WEBSITE)
        cols["email"]    = _sel("Email column",          "email",    _REG_COL_EMAIL)
        cols["city"]     = _sel("City column",           "city",     _REG_COL_CITY)
        cols["province"] = _sel("Province column",       "province", _REG_COL_PROVINCE)
        cols["postcode"] = _sel("Postal code column",    "postcode", _REG_COL_POSTCODE)
        cols["phone"]    = _sel("Phone column",          "phone",    _REG_COL_PHONE)

    if not cols.get("company"):
        st.error("Company name column could not be detected. Please select it above.")
        return

    # ── Detection summary ─────────────────────────────────────────────────────
    detected_info = {k: v for k, v in cols.items() if v}
    missing_info  = {k for k, v in cols.items() if not v}
    st.caption(
        "Detected: " + " · ".join(f"**{k}** → `{v}`" for k, v in detected_info.items())
        + (f"  \n⚠ Not found: {', '.join(missing_info)}" if missing_info else "")
    )

    # ── Input preview ─────────────────────────────────────────────────────────
    preview_cols = [v for v in [
        cols.get("company"), cols.get("website"), cols.get("email"),
        cols.get("city"), cols.get("province"), cols.get("phone"),
    ] if v and v in df.columns]
    st.dataframe(df[preview_cols].head(8), use_container_width=True)

    # ── Batch size option ─────────────────────────────────────────────────────
    max_rows = len(df)
    with st.expander("Batch options", expanded=False):
        run_all = st.checkbox("Process all rows", value=True, key="reg_run_all")
        if not run_all:
            batch_n = st.number_input(
                "Process top N rows (demo / credit-saving mode)",
                min_value=1, max_value=max_rows, value=min(10, max_rows), step=5,
                key="reg_batch_n",
            )
        else:
            batch_n = max_rows

    # ── Verifier Preview ──────────────────────────────────────────────────────
    if verifier_provider not in (_VP_OFF,):
        _preview_n = int(batch_n)
        _spd_def = _FC_SPEED_DEFAULTS.get(verifier_speed_mode, _FC_SPEED_DEFAULTS[_FC_SPEED_FAST])
        _preview_max_cands = verifier_max_candidates
        _preview_max_pages = verifier_max_pages

        # Quick heuristic: count how many rows in the batch are likely to trigger verification.
        # We apply only the fast pre-skip rule (high-conf + strong brand-domain) on already-known
        # company names; since we haven't run Python scoring yet, we use the input domain column.
        _company_col_p = cols.get("company") or ""
        _website_col_p = cols.get("website") or ""
        _preview_df = df.head(_preview_n)
        _estimated_verified = 0
        _preview_rows: list[dict] = []
        for _, _prow in _preview_df.iterrows():
            _pname = str(_prow.get(_company_col_p, "") or "").strip()
            _pdom  = str(_prow.get(_website_col_p, "") or "").strip()
            # Conservative estimate: assume verification runs unless domain exactly matches brand
            _nv = extract_name_variants(_pname) if _pname else {}
            _bov = brand_overlap_variants(_nv, _pdom) if _pdom and _nv else 0.0
            _will_verify = not (_bov >= 0.60 and _pdom and not is_generic(_pdom))
            if _will_verify:
                _estimated_verified += 1
            if len(_preview_rows) < 20 and _will_verify:
                _preview_rows.append({
                    "company": _pname,
                    "input_domain": _pdom,
                    "brand_domain_overlap": round(_bov, 2),
                    "estimated_reason": "will verify" if _will_verify else "pre-skip (strong match)",
                })

        _est_requests = _estimated_verified * _preview_max_cands * _preview_max_pages
        with st.expander(
            f"Verifier preview — est. {_estimated_verified}/{_preview_n} rows → "
            f"~{_est_requests} max FC requests",
            expanded=_estimated_verified > 0,
        ):
            st.caption(
                f"Provider: **{verifier_provider}** · Speed: **{verifier_speed_mode}** · "
                f"Max cands: {_preview_max_cands} · Max pages/cand: {_preview_max_pages} · "
                f"Timeout: {verifier_page_timeout}s"
            )
            st.markdown(
                f"- Total rows in batch: **{_preview_n}**  \n"
                f"- Estimated rows using Firecrawl: **{_estimated_verified}**  \n"
                f"- Estimated max Firecrawl requests: **{_est_requests}**  \n"
                f"- *(Actual count will be lower due to early-stop and high-confidence pre-skips)*"
            )
            if _preview_rows:
                st.markdown("**First ≤20 companies estimated to be verified:**")
                st.dataframe(
                    pd.DataFrame(_preview_rows),
                    use_container_width=True,
                    hide_index=True,
                )

    # ── Run ───────────────────────────────────────────────────────────────────
    if st.button("🧹 Clean and validate register data", type="primary", use_container_width=True):
        run_df = df.head(int(batch_n)).copy()
        n      = len(run_df)
        # Compute label + filename at run start
        _run_label    = _make_run_label(haiku_mode, n, int(max_queries), debug_mode)
        _run_filename = _make_filename(_run_label, run_id)
        settings_dict = {
            "max_queries": int(max_queries), "batch_n": n,
            "haiku_mode": haiku_mode, "debug_mode": debug_mode,
            "jina_mode": jina_mode,
            "run_label": _run_label,
        }

        progress_bar = st.progress(0.0)
        status_text  = st.empty()
        _fc_live: dict = {}  # live counters written by process_dataframe via _live_fc_counters

        _run_t0 = time.time()

        def progress_cb(i, total):
            progress_bar.progress(i / total)
            _elapsed = time.time() - _run_t0
            _elapsed_str = f"{int(_elapsed // 60)}m{int(_elapsed % 60)}s"
            _remain_str = ""
            if i > 0:
                _eta = _elapsed / i * (total - i)
                _remain_str = f" · ETA {int(_eta // 60)}m{int(_eta % 60)}s"
            if verifier_provider != _VP_OFF and _fc_live:
                _att  = _fc_live.get("fc_requests_attempted", 0)
                _succ = _fc_live.get("fc_pages_successful", 0)
                _to   = _fc_live.get("fc_timeouts", 0)
                _secs = _fc_live.get("fc_total_secs", 0.0)
                _avg  = round(_secs / _att, 1) if _att else 0.0
                _avg_proc = round(_succ / i, 2) if i else 0.0
                _avg_fc_row = round(_succ / max(_att, 1), 2) if _att else 0.0
                status_text.caption(
                    f"Row {i}/{total} · {_elapsed_str} elapsed{_remain_str}  \n"
                    f"FC: {_att} req / {_succ} pages / {_to} timeouts · "
                    f"avg {_avg}s/req · {_avg_proc:.2f} pages/processed row · "
                    f"~{_succ} estimated credits"
                )
            else:
                status_text.caption(f"Processing {i} / {total} · {_elapsed_str} elapsed{_remain_str}")

        enriched_df, evidence_rows, debug_rows, jina_debug_rows = process_dataframe(
            run_df, cols, serper_key, int(max_queries),
            progress_cb=progress_cb,
            live_counters_out=_fc_live,
            run_id=run_id,
            resume_from=0,
            prior_results=[],
            prior_evidence=[],
            settings=settings_dict,
            run_label=_run_label,
            haiku_mode=haiku_mode,
            haiku_api_key=anthropic_key,
            haiku_model=haiku_model,
            haiku_max_rows=int(haiku_max_rows),
            jina_mode=jina_mode,
            jina_api_key=jina_api_key,
            verifier_provider=verifier_provider,
            verifier_mode=verifier_mode,
            fc_key=_fc_key,
            max_cands_per_company=verifier_max_candidates,
            max_pages_per_cand=verifier_max_pages,
            page_timeout=verifier_page_timeout,
            fc_speed_mode=verifier_speed_mode,
            fc_location=fc_location_payload,
            eligibility_filter_mode=eligibility_filter_mode,
            debug_mode=debug_mode,
            infer_size=infer_size,
            country_config=_country_cfg,
        )

        progress_bar.progress(1.0)
        status_text.caption(f"✅ Done — {n} companies processed.")

        # Build run_meta for summary sheet
        _run_meta = _build_run_meta(
            enriched_df=enriched_df,
            input_filename=uploaded.name,
            run_id=run_id, run_label=_run_label,
            total_rows_input=len(df), batch_n=n,
            max_queries=int(max_queries), haiku_mode=haiku_mode,
            haiku_model=haiku_model, haiku_max_rows=int(haiku_max_rows),
            debug_mode=debug_mode,
            serper_key_present=bool(serper_key),
            anthropic_key_present=bool(anthropic_key),
            jina_mode=jina_mode,
            verifier_provider=verifier_provider,
            verifier_mode=verifier_mode,
            firecrawl_keys_loaded=len(_fc_keys),
            fc_health=_fc_live,
            serper_queries_total=getattr(process_dataframe, "_last_serper_count", 0),
        )

        # Build Firecrawl Audit dict
        _fc_enabled_st = verifier_provider in (_VP_FIRECRAWL, _VP_FC_JINA)
        _fc_pf_result  = st.session_state.get("_fc_preflight_result", {})
        _fc_att_st     = _fc_live.get("requests_attempted", 0)
        _fc_succ_st    = _fc_live.get("pages_successful", 0)
        _fc_audit_st: dict | None = None
        if _fc_enabled_st:
            _pf_ks = _fc_pf_result.get("key_statuses", {})
            _fc_audit_st = {
                "firecrawl_enabled":                True,
                "firecrawl_preflight_status":       _fc_pf_result.get("preflight_status", "SKIPPED"),
                "firecrawl_keys_total":             _fc_pf_result.get("keys_total", len(_fc_keys)),
                "firecrawl_keys_ok":                _fc_pf_result.get("keys_ok", ""),
                "firecrawl_keys_failed":            _fc_pf_result.get("keys_failed", ""),
                "firecrawl_key_statuses_preflight": str(_pf_ks),
                "firecrawl_requests_attempted":     _fc_att_st,
                "firecrawl_pages_successful":       _fc_succ_st,
                "firecrawl_success_rate":           f"{_fc_succ_st / _fc_att_st:.1%}" if _fc_att_st else "n/a",
                "firecrawl_key_failovers":          _fc_live.get("key_failovers", 0),
                "firecrawl_key_failure_events":     _fc_live.get("key_failure_events", 0),
                "firecrawl_quota_or_billing_errors": _fc_live.get("quota_or_billing_errors", 0),
                "firecrawl_rate_limit_errors":      _fc_live.get("rate_limit_errors", 0),
                "firecrawl_timeouts":               _fc_live.get("timeouts", 0),
                "firecrawl_exceptions":             _fc_live.get("exceptions", 0),
                "firecrawl_consecutive_failures_max": _fc_live.get("consecutive_failures_max", 0),
                "batch_firecrawl_status":           (
                    "fail_fast" if _fc_live.get("fail_fast_triggered")
                    else "ok" if _fc_att_st > 0
                    else "not_used"
                ),
                "batch_firecrawl_notes":            _fc_live.get("fail_fast_reason", ""),
            }
        st.session_state["_fc_audit"] = _fc_audit_st

        # Mark complete in checkpoint
        _save_checkpoint(run_id, [], evidence_rows, n, n, run_df, cols, settings_dict,
                         run_label=_run_label)

        # ── Pipeline folder auto-save ─────────────────────────────────────────
        if pipeline_output_enabled:
            try:
                _pl_ts    = datetime.now().strftime("%Y%m%d_%H%M")
                _pl_paths = resolve_pipeline_output_paths(
                    getattr(uploaded, "name", ""),
                    project_root=pipeline_project_root or None,
                    ts=_pl_ts,
                )
                _pl_excel = build_excel(
                    enriched_df, run_df, evidence_rows, cols,
                    debug_rows=debug_rows, debug_mode=debug_mode,
                    run_meta=_run_meta, jina_debug_rows=jina_debug_rows,
                )
                _pl_out_path = Path(_pl_paths["output_xlsx"])
                _write_pipeline_output(_pl_out_path, _pl_excel)
                _pl_rm = _run_meta or {}
                _pl_log_row = {
                    "timestamp":             _pl_ts,
                    "cohort":                _pl_paths["cohort"],
                    "batch_stem":            _pl_paths["batch_stem"],
                    "batch_number":          _pl_paths["batch_number"],
                    "row_range":             _pl_paths["row_range"],
                    "input_path":            getattr(uploaded, "name", ""),
                    "output_xlsx":           _pl_paths["output_xlsx"],
                    "rows_processed":        n,
                    "rows_input":            len(df),
                    "coverage_pct":          _pl_rm.get("coverage_pct", ""),
                    "manual_review_count":   _pl_rm.get("manual_review_count", ""),
                    "no_match_count":        _pl_rm.get("no_confident_match_count", ""),
                    "haiku_mode":            haiku_mode,
                    "verifier_provider":     verifier_provider,
                    "serper_queries":        int(max_queries),
                    "firecrawl_keys_loaded": len(_fc_keys),
                    "run_id":                run_id,
                    "run_label":             _run_label,
                    "status":                "complete",
                    "notes":                 "",
                }
                _append_run_log_csv(Path(_pl_paths["run_log_csv"]), _pl_log_row)
                st.success(
                    f"✅ Pipeline output saved → `{_pl_paths['output_xlsx']}`  \n"
                    f"Run log → `{_pl_paths['run_log_csv']}`"
                )
            except Exception as _pl_exc:
                st.warning(f"⚠ Pipeline auto-save failed: {_pl_exc}")

        st.session_state["reg_enriched"]         = enriched_df
        st.session_state["reg_evidence"]         = evidence_rows
        st.session_state["reg_debug"]            = debug_rows
        st.session_state["reg_jina_debug"]       = jina_debug_rows
        st.session_state["reg_original"]         = run_df
        st.session_state["reg_cols"]             = cols
        st.session_state["reg_run_id"]           = run_id
        st.session_state["reg_run_label"]        = _run_label
        st.session_state["reg_filename"]         = _run_filename
        st.session_state["reg_run_meta"]         = _run_meta
        st.session_state["reg_debug_mode_value"] = debug_mode
        st.session_state["_fc_live_counters"]    = _fc_live

    # ── Results ───────────────────────────────────────────────────────────────
    enriched_df = st.session_state.get("reg_enriched")
    if enriched_df is None:
        return

    st.markdown("---")
    st.markdown("### Results")
    stored_cols   = st.session_state.get("reg_cols", cols)
    active_run_id = st.session_state.get("reg_run_id", run_id)
    _out_label    = st.session_state.get("reg_run_label", "")
    _out_filename = st.session_state.get("reg_filename",
                        f"register_cleaned_{active_run_id[:8]}.xlsx")
    _out_dm       = st.session_state.get("reg_debug_mode_value", debug_mode)

    _summary_metrics(enriched_df, stored_cols)
    st.markdown("")
    _show_results(enriched_df, stored_cols)

    # ── Firecrawl diagnostics (shown when Firecrawl was selected) ────────────
    _active_vp = st.session_state.get("reg_verifier_provider", _VP_OFF)
    if _active_vp in (_VP_FIRECRAWL, _VP_FC_JINA):
        _fc_pages_total = int(pd.to_numeric(
            enriched_df.get("firecrawl_pages_fetched", pd.Series(dtype=int)),
            errors="coerce").fillna(0).sum())
        _fc_rows_used = int(
            enriched_df.get("firecrawl_used", pd.Series(dtype=str)).astype(str)
            .str.lower().isin(["true", "1"]).sum()
        )
        _rows_planned = int(
            enriched_df.get("verification_needed", pd.Series(dtype=str)).astype(str)
            .str.lower().isin(["true", "1"]).sum()
        )
        _verifier_rows = int(
            enriched_df.get("verifier_used", pd.Series(dtype=str)).astype(str)
            .str.lower().isin(["true", "1"]).sum()
        )

        # ── Sanity warning (Fix 6) ──────────────────────────────────────────
        if _rows_planned > 0 and _fc_pages_total == 0:
            st.error(
                f"⚠️ Firecrawl was selected and {_rows_planned} rows were planned for verification, "
                "but **no Firecrawl pages were fetched**. "
                "This usually means the Firecrawl cache key raised a TypeError (unhashable list) "
                "or the API key is missing/invalid. Check _fc_scrape, the API key, and the location payload."
            )

        _fc_confirmed = int(
            enriched_df.get("firecrawl_decision", pd.Series(dtype=str))
            .astype(str).eq("confirm").sum()
        )
        _fc_replaced = int(
            enriched_df.get("firecrawl_decision", pd.Series(dtype=str))
            .astype(str).str.startswith("replace").sum()
        )
        _fc_rejected = int(
            enriched_df.get("firecrawl_decision", pd.Series(dtype=str))
            .astype(str).str.startswith("reject").sum()
        )
        _fc_uncertain = int(
            enriched_df.get("firecrawl_decision", pd.Series(dtype=str))
            .astype(str).eq("uncertain").sum()
        )
        # Collect errors for diagnostics (Fix 7)
        _fc_errors_series = (
            enriched_df.get("firecrawl_error", pd.Series(dtype=str))
            .astype(str).replace("", pd.NA).dropna()
        )
        _live = st.session_state.get("_fc_live_counters", {})

        with st.expander(
            f"Firecrawl diagnostics — {_fc_pages_total} pages fetched / {_rows_planned} planned",
            expanded=(_fc_pages_total == 0 and _rows_planned > 0),
        ):
            st.markdown(
                f"**Planned rows (verification_needed):** {_rows_planned}  \n"
                f"**Rows with verifier_used=True:** {_verifier_rows}  \n"
                f"**Rows with firecrawl_used=True:** {_fc_rows_used}  \n"
                f"**Firecrawl requests attempted:** {_live.get('fc_requests_attempted', '—')}  \n"
                f"**Pages fetched (successful):** {_fc_pages_total}  \n"
                f"**Timeouts:** {_live.get('fc_timeouts', '—')}  \n"
                f"**Exceptions:** {_live.get('fc_exceptions', '—')}  \n"
                f"**Avg pages/row verified:** "
                f"{round(_fc_pages_total / _fc_rows_used, 1) if _fc_rows_used else 0}  \n"
                f"**Decisions:** confirm={_fc_confirmed} · replace={_fc_replaced} · "
                f"reject={_fc_rejected} · uncertain={_fc_uncertain}"
            )
            if not _fc_errors_series.empty:
                st.markdown("**First 10 firecrawl_error values:**")
                st.code("\n".join(_fc_errors_series.head(10).tolist()))

    _show_run_info_block(
        run_label=_out_label, run_id=active_run_id,
        batch_n=int(st.session_state.get("reg_run_meta", {}).get("processed_rows", "?")),
        max_queries=int(max_queries), haiku_mode=haiku_mode,
        debug_mode=_out_dm, filename=_out_filename,
    )

    # ── Download ──────────────────────────────────────────────────────────────
    st.markdown("---")
    _download_section(
        enriched_df,
        st.session_state.get("reg_original", df),
        st.session_state.get("reg_evidence", []),
        stored_cols,
        filename=_out_filename,
        debug_rows=st.session_state.get("reg_debug", []),
        debug_mode=_out_dm,
        run_meta=st.session_state.get("reg_run_meta"),
        jina_debug_rows=st.session_state.get("reg_jina_debug", []),
        fc_audit=st.session_state.get("_fc_audit"),
    )


def _run_domain_quality_self_test() -> None:
    """
    Lightweight self-test for domain quality logic.
    No Serper calls, no API keys required.
    Run via: python input_cleaner_register_edition.py --self-test-domain-quality
    """
    import traceback

    CASES_REJECT = [
        # (domain, description)
        ("it.linkedin.com",           "LinkedIn country subdomain"),
        ("it.wikipedia.org",          "Wikipedia country subdomain"),
        ("it.kompass.com",            "Kompass country subdomain"),
        ("fatturatoitalia.it",        "Italian financial-data directory"),
        ("visura.pro",                "Business data lookup service"),
        ("x.abbrevia.it",            "Abbrevia subdomain profile service"),
        ("rsuibmsegrate.altervista.org", "Altervista hosted blog"),
        ("mybrand.blogspot.com",      "Blogspot hosted blog"),
        ("myfirm.wixsite.com",        "Wix hosted site"),
        ("myfirm.weebly.com",         "Weebly hosted site"),
        ("myfirm.wordpress.com",      "WordPress hosted blog"),
    ]

    CASES_ACCEPT = [
        # (domain, description) — should NOT be hard-rejected by is_generic/is_hosted_platform/classify_domain
        ("ibm.com",           "IBM global domain"),
        ("zf.com",            "ZF global domain"),
        ("q8.it",             "Q8 Italy"),
        ("solutions30.com",   "Solutions30 corporate"),
        ("pirelli.com",       "Pirelli corporate"),
    ]

    passed = 0
    failed = 0

    print("=" * 60)
    print("Domain Quality Self-Test")
    print("=" * 60)

    print("\n── Should-REJECT cases ──")
    for domain, desc in CASES_REJECT:
        try:
            generic  = is_generic(domain)
            blocked  = is_discovery_blocked(domain)
            platform = is_hosted_platform(domain)
            cat      = classify_domain(domain)
            rejected = generic or blocked or platform or bool(cat)
            status = "PASS" if rejected else "FAIL"
            reason = (
                "generic" if generic else
                "media_blocked" if blocked else
                "hosted_platform" if platform else
                cat if cat else "not_rejected"
            )
            print(f"  [{status}] {domain!s:<40} ({desc}) → {reason}")
            if rejected:
                passed += 1
            else:
                failed += 1
        except Exception:
            print(f"  [ERROR] {domain}: {traceback.format_exc(limit=1)}")
            failed += 1

    print("\n── Should-ACCEPT cases (must NOT be hard-rejected) ──")
    for domain, desc in CASES_ACCEPT:
        try:
            generic  = is_generic(domain)
            blocked  = is_discovery_blocked(domain)
            platform = is_hosted_platform(domain)
            cat      = classify_domain(domain)
            hard_rejected = generic or blocked or platform or bool(cat)
            status = "PASS" if not hard_rejected else "FAIL"
            reason = (
                "generic" if generic else
                "media_blocked" if blocked else
                "hosted_platform" if platform else
                cat if cat else "ok"
            )
            print(f"  [{status}] {domain!s:<40} ({desc}) → {reason}")
            if not hard_rejected:
                passed += 1
            else:
                failed += 1
        except Exception:
            print(f"  [ERROR] {domain}: {traceback.format_exc(limit=1)}")
            failed += 1

    print("\n── evaluate_domain_candidate spot-checks ──")
    spot_checks = [
        ("it.linkedin.com",  "Rossi SRL", False, "LinkedIn rejected"),
        ("ibm.com",          "IBM Italia SPA", None, "IBM not hard-rejected"),
    ]
    for dom, cname, expected_accepted, label in spot_checks:
        try:
            res = evaluate_domain_candidate(cname, dom)
            ok = (res["accepted"] == expected_accepted) if expected_accepted is not None else (res["score"] > 0 or not res["accepted"])
            # For IBM: just verify it is NOT hard-rejected with score 0 due to blacklist
            if expected_accepted is None:
                ok = res["source_type"] not in ("directory_or_social", "hosted_platform", "media_blocked", "shortener")
            status = "PASS" if ok else "FAIL"
            print(f"  [{status}] evaluate({dom!r}, {cname!r}) → accepted={res['accepted']}, score={res['score']}, {res['reason'][:60]}")
            if ok:
                passed += 1
            else:
                failed += 1
        except Exception:
            print(f"  [ERROR] {label}: {traceback.format_exc(limit=1)}")
            failed += 1

    print(f"\n{'='*60}")
    print(f"Results: {passed} passed, {failed} failed")
    print("=" * 60)

    if failed:
        sys.exit(1)


if __name__ == "__main__":
    # CLI batch mode when --input (file or folder) or --self-test-domain-quality is passed.
    if "--input" in sys.argv or "--self-test-domain-quality" in sys.argv:
        cli_batch_run()
    else:
        main()
