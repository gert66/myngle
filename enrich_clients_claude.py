"""
Claude + Jina AI Two-Step Company Enrichment
=============================================
Upload a file with company names and URLs.
Each row gets TWO enrichment passes:

  Step 1 — Basic firmographics (Jina AI Reader → Claude extraction)
  Step 2 — Mingle ICP signals  (Claude with web_search tool)

Architecture
------------
- One row is processed per Streamlit rerun so the Stop button works at any point.
- All mutable run state lives in st.session_state (keys prefixed with _).
- Anthropic API key is read ONLY from st.secrets['ANTHROPIC_API_KEY'].
- Debug mode: toggled via sidebar checkbox.
"""

import base64
import io
import json
import os
import re
import sys
import time
import unicodedata
import zipfile
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from urllib.parse import quote

import anthropic
import pandas as pd
import warnings as _warnings
from pandas.errors import PerformanceWarning as _PerformanceWarning
_warnings.simplefilter("ignore", _PerformanceWarning)
import requests
st = None          # lazy — populated by get_streamlit() in UI mode only
components = None  # lazy — populated by get_streamlit() in UI mode only


def get_streamlit():
    global st, components
    if st is None:
        import streamlit as _st
        import streamlit.components.v1 as _components
        st = _st
        components = _components
    return st, components
# ─────────────────────────────────────────────────────────────────────────────
# CLI / Streamlit mode detection
# ─────────────────────────────────────────────────────────────────────────────

_CLI_FLAGS = {
    "--input", "--dry-run-paths", "--output-dir", "--project-root",
    "--max-rows", "--debug", "--anthropic-key", "--serper-key",
    "--self-test-competitor-override", "--self-test-output", "--no-eta",
}


def cli_args_present() -> bool:
    """Return True when any known CLI flag is present in sys.argv."""
    for arg in sys.argv[1:]:
        if arg in _CLI_FLAGS:
            return True
        if any(arg.startswith(flag + "=") for flag in _CLI_FLAGS):
            return True
    return False


def running_under_streamlit() -> bool:
    """Return True when executed via `streamlit run` (does not import streamlit)."""
    if "streamlit" not in sys.modules:
        return False
    try:
        import streamlit.runtime as _sr
        return _sr.exists()
    except Exception:
        return False


def is_cli_mode() -> bool:
    """Return True when running in CLI mode (not via Streamlit)."""
    return cli_args_present() or not running_under_streamlit()


from bs4 import BeautifulSoup

try:
    from human_scraper import scrape_with_human_behaviour
    _PLAYWRIGHT_AVAILABLE = True
except ImportError:
    _PLAYWRIGHT_AVAILABLE = False

try:
    from commercial_fit_scoring import score_dataframe as _score_dataframe, SCORE_OUTPUT_COLS as _SCORE_OUTPUT_COLS
    _SCORING_AVAILABLE = True
except ImportError:
    _score_dataframe = None  # type: ignore[assignment]
    _SCORE_OUTPUT_COLS = []
    _SCORING_AVAILABLE = False

# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

JINA_READER_URL  = "https://r.jina.ai/"
JINA_SEARCH_URL  = "https://s.jina.ai/"
CACHE_DIR        = Path("claude_json_cache")
DEBUG_LOG_DIR    = Path("debug_logs")
SEARCH_OUTPUT_DIR = DEBUG_LOG_DIR / "search_outputs"
AUTOSAVE_PATH         = "/tmp/enrichment_autosave.csv"
CHECKPOINT_EVERY      = 50   # write checkpoint_NNN.xlsx every N companies
_AUTO_DL_EVERY        = 100  # auto browser-download every N companies
_DEFAULT_DOWNLOAD_DIR = os.path.expanduser("~/Downloads")
_XL_AUTOSAVE_DIR     = "enrichment_outputs"   # folder next to the app
_XL_AUTOSAVE_DEFAULT = "Enriched results.xlsx"
_XL_AUTOSAVE_EVERY   = 5                       # default: save every 5 companies
_PER_COMPANY_AUTOSAVE_DEFAULT_DIR = os.path.expanduser(
    "~/Downloads/company_enrichment_runs"
)
MODEL_STEP1      = "claude-haiku-4-5-20251001"
MODEL_STEP2      = "claude-haiku-4-5-20251001"
MODEL_ID         = MODEL_STEP1   # legacy alias used in a few places
# WEB_SEARCH_TOOL is retained only for audit/reference — it must NEVER be passed
# to any Anthropic API call. Serper is the only permitted Step 2 search provider.
WEB_SEARCH_TOOL  = {"type": "web_search_20250305", "name": "web_search"}

# Set to True (or via SHOW_ADVANCED_SETTINGS env var / Streamlit secret) to show
# the full sidebar, column picker, preview, debug sections, and technical logs.
# Cloud deployments set SHOW_ADVANCED_SETTINGS = false in Streamlit secrets to
# hide the sidebar and show only the minimal user flow.
SHOW_ADVANCED_SETTINGS: bool = True

SERPER_SEARCH_URL    = "https://google.serper.dev/search"
STEP2_PROVIDER_CLAUDE = "Claude Web Search"
STEP2_PROVIDER_SERPER = "Serper Google Search"

AVAILABLE_MODELS = {
    "Haiku 4.5 (fast, cheap)":                   "claude-haiku-4-5-20251001",
    "Sonnet 4.5 (better reasoning, higher cost)": "claude-sonnet-4-5-20250929",
}
# Per-row cost estimates for the pre-run info banner
_COST_EST = {
    ("claude-haiku-4-5-20251001",  "claude-haiku-4-5-20251001"):  0.01,
    ("claude-haiku-4-5-20251001",  "claude-sonnet-4-5-20250929"): 0.05,
    ("claude-sonnet-4-5-20250929", "claude-haiku-4-5-20251001"):  0.04,
    ("claude-sonnet-4-5-20250929", "claude-sonnet-4-5-20250929"): 0.08,
}

# Pricing per million tokens (claude-haiku-4-5)
_COST_INPUT_PER_M  = 0.80
_COST_OUTPUT_PER_M = 4.00

# Step 1 extraction prompt
_STEP1_PROMPT = (
    "Extract company information from this webpage content. "
    "Return ONLY a raw JSON object with these exact fields: "
    "company_name, domain, description, founded_year, employee_range, revenue_range, "
    "main_industry, sub_industry, company_type, country, city, continent, "
    "linkedin_url, specialties (as comma-separated string), "
    "technologies (as comma-separated string), "
    "total_funding_amount, total_funding_rounds, last_round_type, last_round_amount, "
    "last_round_date, ipo_status. "
    "Use empty string for any field not found."
)

# ── Step 2 prompt — static cacheable prefix + dynamic per-company suffix ──────
#
# The static prefix is sent with cache_control so Anthropic caches it after the
# first company.  Subsequent companies pay ~10 % of normal input cost for it.
# Minimum cacheable size: 1 024 tokens (Sonnet) / 2 048 tokens (Haiku).
# The web_search tool definition (~300 tokens) counts toward the threshold too.

STEP2_STATIC_PREFIX = """\
You are analyzing companies to identify whether they may have a potential interest in \
language training, business English, communication training, leadership training, \
negotiation training, team-building, onboarding, or broader employee development.

Your task is to evaluate each company based on the 10 buying signals below. \
Prioritize the signals in this order:

1. International footprint - The company has offices, teams, subsidiaries, clients, \
production sites, or operations in multiple countries or regions.

2. Foreign headquarters, parent company, or group structure - The company operates in \
one country but has its headquarters, parent company, group ownership, regional HQ, or \
reporting lines in another country.

3. Competitor signal — check all three categories below and populate the exact matching \
output field. Each provider belongs to EXACTLY ONE category; do not move it to another.

  Category 1 — Direct corporate language training competitors (STRONG signal). \
These providers must ONLY go into competitor_signal and direct_language_competitor_signal. \
Do NOT put them into online_language_learning_signal or broader_lnd_platform_signal. \
Providers: goFLUENT, Learnlight, Speexx, Voxy, Learnship, Berlitz, \
EF Corporate Solutions, Babbel for Business, Rosetta Stone Enterprise, Preply Business, \
Talaera, Busuu for Business, Lingoda for Business, Fluentify, Twenix, Cambly.

  Category 2 — Online language learning brands (MEDIUM signal). \
These must ONLY go into online_language_learning_signal, and only when found in a \
corporate, HR, L&D, employee benefit, or company-wide training context. \
Do NOT put them into competitor_signal, direct_language_competitor_signal, \
or broader_lnd_platform_signal. \
Providers: Duolingo, Babbel, Busuu, Rosetta Stone, Preply, Memrise, Mondly, ELSA Speak, \
FluentU, italki, Lingoda, Open English, Mango Languages, Pimsleur, Drops, HelloTalk, \
Tandem.

  Category 3 — Broader corporate learning / L&D platforms (L&D maturity signal). \
These must ONLY go into broader_lnd_platform_signal. \
Do NOT put them into competitor_signal, direct_language_competitor_signal, \
or online_language_learning_signal. \
Providers: OpenSesame, Coursera for Business, Udemy Business, LinkedIn Learning, \
Skillsoft, Docebo, Degreed, Cornerstone, 360Learning, Moodle Workplace, Absorb LMS, \
TalentLMS, LearnUpon, Pluralsight.

  mYngle — mYngle is the company running this analysis and is NOT a competitor. \
Do NOT place mYngle in any competitor or provider signal field under any circumstances. \
If mYngle is mentioned in the search results, note it only in the evidence field as a \
reference or client signal.

4. Merger, acquisition, integration, or new group ownership.

5. Explicit learning and development focus.

6. International customer base or client-facing international work.

7. Multicultural or multilingual workforce.

8. Employer branding and employee satisfaction.

9. Rapid growth, hiring, or expansion.

10. Leadership, management, sales, or negotiation-heavy roles.

Return ONLY a raw JSON object with exactly these fields and no others:
{"lead_score": "High or Medium or Low", \
"buying_signals": "comma-separated list of signal names actually supported by evidence", \
"competitor_signal": "comma-separated Category 1 provider names found, otherwise empty string — Category 3 platforms such as LinkedIn Learning or Skillsoft must never appear here", \
"direct_language_competitor_signal": "comma-separated Category 1 provider names found, otherwise empty string", \
"online_language_learning_signal": "comma-separated Category 2 provider names found in corporate/HR/L&D context only, otherwise empty string — Category 3 platforms must never appear here", \
"broader_lnd_platform_signal": "comma-separated Category 3 provider names found, otherwise empty string", \
"evidence": "brief description of what was found and source types", \
"likely_training_interest": \
"comma-separated list from: Language training / Business English, \
Intercultural communication, Leadership training, Negotiation training, \
Sales or client communication, Team collaboration / team-building, \
Onboarding / employee development, Broader professional training", \
"why_relevant": "brief explanation", \
"potential_buyer_function": \
"most likely buyer such as HR, Learning and Development, Talent Development, \
People and Culture, Leadership Development, Sales Enablement, Customer Success, \
Operations, Procurement"}

Do not invent evidence. Do not wrap in markdown.\
"""


# ── Model-signal extraction prompt ────────────────────────────────────────────
#
# This prompt is sent AFTER Step 1 + Step 2 enrichment to extract structured
# numeric signals for logistic-regression readiness.  It uses only evidence
# already gathered — no additional web searches are performed.

MODEL_SIGNAL_PROMPT_TEMPLATE = """\
You are extracting structured model signals for a company from available enrichment context.
Your ONLY task is to return a valid JSON object with exactly the fields listed below.

Rules:
- Use ONLY evidence from the provided enrichment context (website, search snippets, firmographic data).
- Do NOT invent facts, infer company size, or estimate employee count.
- Do NOT generate or fill any employee_size_score field.
- If evidence is weak or ambiguous, score 1 or 0.
- Use score 3 ONLY for explicit or very strong evidence.
- Keep each evidence field to one short sentence. Use empty string if score is 0.
- Return ONLY raw JSON — no markdown, no backticks, no explanation.
- All score fields must be integers.
- All binary fields must be 0 or 1.

Scoring rubric (0–3):
  0 = no evidence found
  1 = weak or indirect evidence
  2 = clear evidence
  3 = strong or explicit evidence

International footprint vs foreign HQ — CRITICAL distinction:
  sig_intl_footprint_score: Score 1–3 when the company has international operations,
    subsidiaries, offices, production sites, clients, export activity, multilingual teams,
    or cross-border business. This applies even when the company is headquartered in the
    input country. A domestic company with international subsidiaries scores high here.
  sig_foreign_hq_score: Score 1–3 ONLY when the company's headquarters, parent company,
    group ownership, regional HQ, or reporting line is OUTSIDE the input country.
    Examples that justify a positive score: "part of a German group", "owned by a US parent",
    "subsidiary of a French group", "reports to the Swedish parent company".
    Do NOT score foreign HQ based on words like: global, international, multinational,
    worldwide, export, subsidiaries, countries, offices abroad, plants in other countries.
    If the company is headquartered in the input country and has international subsidiaries
    or global operations, set sig_foreign_hq_score = 0 and score sig_intl_footprint instead.
    If evidence says "headquartered in [input country city]", set sig_foreign_hq_score = 0.

Provider category rules:
  Category 1 direct language-training competitors (goes into has_language_competitor):
    goFLUENT, Learnlight, Speexx, Voxy, Learnship, Berlitz, EF Corporate Solutions,
    Babbel for Business, Rosetta Stone Enterprise, Preply Business, Talaera,
    Busuu for Business, Lingoda for Business, Fluentify, Twenix, Cambly.
  Category 2 online language-learning brands (goes into has_online_learning_signal ONLY
    when found in a corporate/HR/L&D/employee-benefit/company-wide training context):
    Duolingo, Babbel, Busuu, Rosetta Stone, Preply, Memrise, Mondly, ELSA Speak,
    FluentU, italki, Lingoda, Open English, Mango Languages, Pimsleur, Drops,
    HelloTalk, Tandem.
  Category 3 broader L&D platforms (goes into has_lnd_platform_signal):
    OpenSesame, Coursera for Business, Udemy Business, LinkedIn Learning, Skillsoft,
    Docebo, Degreed, Cornerstone, 360Learning, Moodle Workplace, Absorb LMS,
    TalentLMS, LearnUpon, Pluralsight.
  has_competitor_signal = 1 if ANY provider from Cat 1, Cat 2 (in corporate context), or Cat 3 is found.
  mYngle is NOT a competitor and must never appear in any signal field.

Binary field rules:
  is_public = 1 if the company appears publicly listed or has clear public company evidence.
  has_funding = 1 if funding rounds, venture backing, private equity, acquisition funding, or similar evidence is found.

Company: __COMPANY_NAME__
Domain: __DOMAIN__

Enrichment context:
__ENRICHMENT_CONTEXT__

Return a JSON object with EXACTLY these fields and no others:
{
  "sig_intl_footprint_score": <int 0-3>,
  "sig_intl_footprint_evidence": <str>,
  "sig_foreign_hq_score": <int 0-3>,
  "sig_foreign_hq_evidence": <str>,
  "sig_explicit_lnd_score": <int 0-3>,
  "sig_explicit_lnd_evidence": <str>,
  "sig_multicultural_score": <int 0-3>,
  "sig_multicultural_evidence": <str>,
  "sig_employer_branding_score": <int 0-3>,
  "sig_employer_branding_evidence": <str>,
  "sig_rapid_growth_score": <int 0-3>,
  "sig_rapid_growth_evidence": <str>,
  "sig_merger_acq_score": <int 0-3>,
  "sig_merger_acq_evidence": <str>,
  "sig_lnd_onboarding_score": <int 0-3>,
  "sig_lnd_onboarding_evidence": <str>,
  "ti_language_english_score": <int 0-3>,
  "ti_language_english_evidence": <str>,
  "ti_onboarding_score": <int 0-3>,
  "ti_onboarding_evidence": <str>,
  "ti_leadership_score": <int 0-3>,
  "ti_leadership_evidence": <str>,
  "ti_broader_professional_score": <int 0-3>,
  "ti_broader_professional_evidence": <str>,
  "ti_team_collab_score": <int 0-3>,
  "ti_team_collab_evidence": <str>,
  "ti_intercultural_score": <int 0-3>,
  "ti_intercultural_evidence": <str>,
  "ti_negotiation_sales_score": <int 0-3>,
  "ti_negotiation_sales_evidence": <str>,
  "has_competitor_signal": <0 or 1>,
  "has_competitor_signal_evidence": <str>,
  "has_language_competitor": <0 or 1>,
  "has_language_competitor_evidence": <str>,
  "has_online_learning_signal": <0 or 1>,
  "has_online_learning_signal_evidence": <str>,
  "has_lnd_platform_signal": <0 or 1>,
  "has_lnd_platform_signal_evidence": <str>,
  "is_public": <0 or 1>,
  "is_public_evidence": <str>,
  "has_funding": <0 or 1>,
  "has_funding_evidence": <str>,
  "competitor_signal_strength_score": <int 0-3>,
  "competitor_signal_strength_evidence": <str>,
  "language_competitor_strength_score": <int 0-3>,
  "language_competitor_strength_evidence": <str>,
  "online_learning_signal_strength_score": <int 0-3>,
  "online_learning_signal_strength_evidence": <str>,
  "lnd_platform_signal_strength_score": <int 0-3>,
  "lnd_platform_signal_strength_evidence": <str>,
  "model_signal_overall_confidence_score": <int 0-3>,
  "model_signal_needs_manual_review": <0 or 1>,
  "model_signal_manual_review_reason": <str>,
  "model_signal_sources_used": <str>,
  "model_signal_search_quality": <"good" or "partial" or "weak" or "failed">
}
"""


# ── Field lists ───────────────────────────────────────────────────────────────

# Step 1: full Lusha-equivalent firmographics
STEP1_FIELDS = [
    "lusha_company_name",
    "lusha_domain",
    "lusha_description",
    "lusha_founded_year",
    "lusha_employee_range",
    "lusha_revenue",
    "lusha_industry",
    "lusha_sub_industry",
    "lusha_company_type",
    "lusha_country",
    "lusha_city",
    "lusha_continent",
    "lusha_linkedin_url",
    "lusha_specialties",
    "lusha_technologies",
    "lusha_total_funding_amount",
    "lusha_total_funding_rounds",
    "lusha_last_round_type",
    "lusha_last_round_amount",
    "lusha_last_round_date",
    "lusha_ipo_status",
]

# Step 2: Mingo ICP buying signals
ICP_FIELDS = [
    "icp_lead_score",
    "icp_buying_signals",
    "icp_competitor_signal",
    "icp_direct_language_competitor_signal",
    "icp_online_language_learning_signal",
    "icp_broader_lnd_platform_signal",
    "icp_evidence",
    "icp_likely_training_interest",
    "icp_why_relevant",
    "icp_potential_buyer_function",
]

# Metadata added per row
META_FIELDS = [
    "enrichment_status",
    "step1_status",
    "step2_status",
    "step2_provider_used",
    "lucia_data_status",
    "lucia_api_called",
    "step1_run_status",
    "needs_manual_review",
    "match_notes",
    "error_message",
    "step1_tokens_in",
    "step1_tokens_out",
    "step1_cost_usd",
    "step2_tokens_in",
    "step2_tokens_out",
    "step2_cost_usd",
    "total_tokens_in",
    "total_tokens_out",
    "total_cost_usd",
    "anthropic_web_search_used",
    "anthropic_tools_used",
    "serper_search_used",
    # Serper evidence handoff (compact aggregated for Opportunity Radar / caller brief)
    "serper_query_summary",
    "serper_source_urls",
    "serper_result_titles",
    "serper_snippets",
    "raw_evidence_summary",
    "evidence_source_urls",
    # Raw Google evidence (full detail for webapp / Lovable handoff)
    "raw_google_evidence_count",
    "raw_google_evidence_urls",
    "raw_google_evidence_combined",
    "raw_google_evidence_json",
    "raw_google_evidence_json_01",
    "raw_google_evidence_json_02",
    "raw_google_evidence_json_03",
    "raw_google_evidence_json_parts",
    "raw_google_evidence_truncated",
    *[
        f"google_snippet_{i:02d}_{field}"
        for i in range(1, 11)
        for field in ("query_type", "query", "rank", "title", "url", "source_domain", "text")
    ],
]

# Real Lusha API enrichment fields (prefix "lusha_api_")
LUSHA_API_FIELDS = [
    "lusha_api_company_name",
    "lusha_api_domain",
    "lusha_api_description",
    "lusha_api_founded_year",
    "lusha_api_employee_range",
    "lusha_api_revenue_range",
    "lusha_api_industry",
    "lusha_api_sub_industry",
    "lusha_api_company_type",
    "lusha_api_country",
    "lusha_api_city",
    "lusha_api_continent",
    "lusha_api_linkedin_url",
    "lusha_api_specialties",
    "lusha_api_technologies",
    "lusha_api_total_funding_amount",
    "lusha_api_total_funding_rounds",
    "lusha_api_last_round_type",
    "lusha_api_last_round_amount",
    "lusha_api_last_round_date",
    "lusha_api_ipo_status",
]

LUSHA_API_META_FIELDS = [
    "lusha_api_status",
    "lusha_api_error",
    "lusha_api_match_confidence",
    "lusha_api_needs_review",
    "lusha_api_match_notes",
    "lusha_api_raw_keys",
]

# Model-signal fields — added by the structured signal-extraction layer
# Ordinal score fields (integer 0–3)
MODEL_SIGNAL_SCORE_FIELDS = [
    "sig_intl_footprint_score",
    "sig_foreign_hq_score",
    "sig_explicit_lnd_score",
    "sig_multicultural_score",
    "sig_employer_branding_score",
    "sig_rapid_growth_score",
    "sig_merger_acq_score",
    "sig_lnd_onboarding_score",
    "ti_language_english_score",
    "ti_onboarding_score",
    "ti_leadership_score",
    "ti_broader_professional_score",
    "ti_team_collab_score",
    "ti_intercultural_score",
    "ti_negotiation_sales_score",
    "competitor_signal_strength_score",
    "language_competitor_strength_score",
    "online_learning_signal_strength_score",
    "lnd_platform_signal_strength_score",
    "model_signal_overall_confidence_score",
]

# Binary fields (0 or 1)
MODEL_SIGNAL_BINARY_FIELDS = [
    "has_competitor_signal",
    "has_language_competitor",
    "has_online_learning_signal",
    "has_lnd_platform_signal",
    "is_public",
    "has_funding",
    "model_signal_needs_manual_review",
]

# Evidence columns (one per score/binary field, using same base name + _evidence)
_MODEL_SIGNAL_SCORED_BASES = [
    "sig_intl_footprint",
    "sig_foreign_hq",
    "sig_explicit_lnd",
    "sig_multicultural",
    "sig_employer_branding",
    "sig_rapid_growth",
    "sig_merger_acq",
    "sig_lnd_onboarding",
    "ti_language_english",
    "ti_onboarding",
    "ti_leadership",
    "ti_broader_professional",
    "ti_team_collab",
    "ti_intercultural",
    "ti_negotiation_sales",
    "competitor_signal_strength",
    "language_competitor_strength",
    "online_learning_signal_strength",
    "lnd_platform_signal_strength",
    "has_competitor_signal",
    "has_language_competitor",
    "has_online_learning_signal",
    "has_lnd_platform_signal",
    "is_public",
    "has_funding",
]
MODEL_SIGNAL_EVIDENCE_FIELDS = [f"{b}_evidence" for b in _MODEL_SIGNAL_SCORED_BASES]

# QA / metadata fields
MODEL_SIGNAL_QA_FIELDS = [
    "model_signal_manual_review_reason",
    "model_signal_sources_used",
    "model_signal_search_quality",
    # Foreign HQ hygiene audit fields (populated by sanitize_foreign_hq_signal)
    "foreign_hq_sanitized",
    "foreign_hq_sanitizer_reason",
    "foreign_hq_original_score",
    "foreign_hq_original_evidence",
    "inferred_input_country",
    "foreign_hq_uncertain",
]

MODEL_SIGNAL_FIELDS = (
    MODEL_SIGNAL_SCORE_FIELDS
    + MODEL_SIGNAL_BINARY_FIELDS
    + MODEL_SIGNAL_EVIDENCE_FIELDS
    + MODEL_SIGNAL_QA_FIELDS
)

DOMAIN_VALIDATION_FIELDS = [
    "input_domain",
    "validated_domain",
    "domain_used_for_enrichment",
    "domain_match_confidence",
    "possible_domain_mismatch",
    "suggested_domain",
    "domain_check_reason",
    "domain_source",
    "needs_domain_review",
]

# ICP override fields — populated per-row (search + classify) then finalised
# by apply_competitor_icp_override() after scoring.
ICP_OVERRIDE_FIELDS: list[str] = [
    "base_commercial_fit_score",
    "competitor_customer_signal",
    "competitor_provider_detected",
    "competitor_signal_strength",
    "competitor_signal_type",
    "competitor_evidence",
    "competitor_evidence_url",
    "icp_override_applied",
    "icp_override_reason",
    "competitive_switch_opportunity",
    "sales_action_hint",
]

# Low-threshold attention fields — populated alongside ICP_OVERRIDE_FIELDS.
# These fire on any plausible competitor mention, even when no hard override applies.
COMPETITOR_ATTENTION_FIELDS: list[str] = [
    "competitor_attention_signal",
    "competitor_attention_provider_detected",
    "competitor_attention_strength",
    "competitor_attention_type",
    "competitor_attention_evidence",
    "competitor_attention_url",
    "competitor_attention_needs_review",
]

ALL_ENRICHMENT_FIELDS = (
    LUSHA_API_FIELDS + LUSHA_API_META_FIELDS + STEP1_FIELDS + ICP_FIELDS
    + META_FIELDS + DOMAIN_VALIDATION_FIELDS + MODEL_SIGNAL_FIELDS
    + ICP_OVERRIDE_FIELDS + COMPETITOR_ATTENTION_FIELDS
)

# Employee range resolver output fields (populated before scoring)
EMPLOYEE_RANGE_RESOLVER_FIELDS = [
    "employee_range_resolved",
    "employee_range_source",
    "employee_range_confidence",
    "employee_range_notes",
    "employee_range_for_scoring",
    "employee_range_for_scoring_source",
]

# Used when no real employee range can be resolved — scoring-continuity only.
# Never written to lusha_employee_range.
DEFAULT_EMPLOYEE_RANGE_FOR_SCORING = "51 - 200"

# ── Canonical size bands (must match commercial_fit_scoring.SIZE_BAND_LOOKUP) ─
_SIZE_BANDS_ORDERED = [
    (10,       "1 - 10"),
    (50,       "11 - 50"),
    (200,      "51 - 200"),
    (500,      "201 - 500"),
    (1_000,    "501 - 1000"),
    (5_000,    "1001 - 5000"),
    (10_000,   "5001 - 10000"),
    (100_000,  "10001 - 100000"),
    (10_000_000, "100001 - 10000000"),
]

_EMPLOYEE_NUMBER_RE = re.compile(
    r"(?:(?:over|more\s+than|beyond|circa|approximately|about|~)?\s*)"
    r"([\d][,\d]*(?:\.\d+)?)\s*"
    r"(?:k\b)?"
    r"(?:[,\s]+(?:to|[-–—])\s*([\d][,\d]*(?:\.\d+)?))?(?:\s*k\b)?"
    r"\s*(?:employees?|people|staff|workforce|collaboratori|dipendenti|"
    r"team\s+members?|risorse|headcount)",
    re.IGNORECASE,
)
_TEAM_OF_RE = re.compile(
    r"team\s+of\s+([\d][,\d]*)",
    re.IGNORECASE,
)
_EMPLOYS_RE = re.compile(
    r"employs?\s+([\d][,\d]*)",
    re.IGNORECASE,
)


def _num_to_size_band(n: float) -> str:
    """Map a numeric employee count to the nearest canonical size band."""
    for upper, band in _SIZE_BANDS_ORDERED:
        if n <= upper:
            return band
    return "100001 - 10000000"


def _parse_employee_number(text: str) -> tuple[float | None, str]:
    """Extract employee count from free text. Returns (midpoint, matched_text)."""
    def _clean(s: str) -> float:
        return float(s.replace(",", "").replace(" ", ""))

    for pat in (_EMPLOYEE_NUMBER_RE, _TEAM_OF_RE, _EMPLOYS_RE):
        m = pat.search(text)
        if m:
            groups = [g for g in m.groups() if g]
            try:
                if len(groups) >= 2:
                    lo, hi = _clean(groups[0]), _clean(groups[1])
                    return (lo + hi) / 2.0, m.group(0).strip()
                else:
                    n = _clean(groups[0])
                    if "k" in m.group(0).lower():
                        n *= 1_000
                    return n, m.group(0).strip()
            except (ValueError, IndexError):
                continue
    return None, ""


def resolve_employee_range(row: dict, company_name: str = "") -> dict:
    """
    Determine the best available employee range for a row.
    Returns a dict with: employee_range_resolved, employee_range_source,
    employee_range_confidence, employee_range_notes.

    Priority:
    1. Existing Lucia/Lusha API or input data (lusha_api_employee_range,
       lusha_employee_range, employee_range, company_size, Company Number of Employees)
    2. Explicit number/range in text fields (lusha_description, icp_evidence, etc.)
    3. Conservative heuristic from profile signals (Low confidence)
    4. Unknown — leave blank
    """
    def _is_blank(v) -> bool:
        return v is None or str(v).strip() in ("", "nan", "None", "N/A", "-")

    result = {
        "employee_range_resolved":  "",
        "employee_range_source":    "missing",
        "employee_range_confidence": "None",
        "employee_range_notes":     "",
    }

    # ── Priority 1: existing structured data ─────────────────────────────────
    for field in (
        "lusha_api_employee_range", "lusha_employee_range",
        "employee_range", "company_size",
        "Company Number of Employees",
    ):
        raw = row.get(field)
        if _is_blank(raw):
            continue
        raw_s = str(raw).strip()
        # Normalise to canonical band via midpoint
        n, _ = _parse_employee_number(raw_s)
        band = _num_to_size_band(n) if n is not None else raw_s
        # Accept if it looks like a valid range string or number
        if band or raw_s:
            result["employee_range_resolved"]  = band or raw_s
            result["employee_range_source"]    = "existing_lucia_or_input_employee_range"
            result["employee_range_confidence"] = "High"
            result["employee_range_notes"]     = f"From field '{field}': {raw_s}"
            return result

    # ── Priority 2: explicit text evidence ───────────────────────────────────
    text_fields = [
        "lusha_description", "icp_evidence", "icp_why_relevant",
        "scoring_notes",
    ]
    # Also include any *_evidence column
    for k in row:
        if k.endswith("_evidence") and k not in text_fields:
            text_fields.append(k)
    for field in text_fields:
        val = row.get(field)
        if _is_blank(val):
            continue
        n, matched = _parse_employee_number(str(val))
        if n is not None and n > 0:
            band = _num_to_size_band(n)
            conf = "High" if n >= 100 else "Medium"
            result["employee_range_resolved"]  = band
            result["employee_range_source"]    = "explicit_text_employee_evidence"
            result["employee_range_confidence"] = conf
            result["employee_range_notes"]     = f"Parsed from '{field}': \"{matched}\""
            return result

    # ── Priority 3: conservative heuristic ───────────────────────────────────
    # Combine available text for signal detection
    _profile_text = " ".join(
        str(row.get(f) or "")
        for f in ("lusha_description", "icp_evidence", "icp_why_relevant",
                  "lusha_specialties", "lusha_company_type", "lusha_industry",
                  "lusha_continent", "lusha_country")
    ).lower()

    _LARGE_SIGNALS = (
        "multinational", "global enterprise", "worldwide", "major bank",
        "large consulting", "listed", "publicly listed", "fortune", "nyse", "nasdaq",
        "stock exchange", "group of companies", "international group",
        "340,000", "100,000", "50,000",
    )
    _MID_SIGNALS = (
        "multiple plants", "multiple offices", "several offices", "national",
        "manufacturing company", "industrial company", "multi-site",
    )
    _SMALL_SIGNALS = (
        "local", "small agency", "boutique", "studio", "freelance",
        "startup", "start-up",
    )

    if any(s in _profile_text for s in _LARGE_SIGNALS):
        band = "10001 - 100000"
        note = "Heuristic: large/global enterprise signals detected. Conservative estimate — Low confidence."
    elif any(s in _profile_text for s in _MID_SIGNALS):
        band = "201 - 500"
        note = "Heuristic: mid-sized company signals detected. Conservative estimate — Low confidence."
    elif any(s in _profile_text for s in _SMALL_SIGNALS):
        band = "11 - 50"
        note = "Heuristic: small/local company signals detected. Conservative estimate — Low confidence."
    else:
        # No useful signal — leave blank
        result["employee_range_notes"] = "No employee range data found in any source."
        return result

    result["employee_range_resolved"]  = band
    result["employee_range_source"]    = "heuristic_size_estimate"
    result["employee_range_confidence"] = "Low"
    result["employee_range_notes"]     = note
    return result


def resolve_employee_range_from_serper(
    company_name: str, domain: str, serper_key: str
) -> dict:
    """
    Fallback: run a few Serper queries to find employee count evidence.
    Uses _call_serper + _parse_employee_number / _num_to_size_band.
    Does NOT call Claude or any LLM.

    Returns the same shape as resolve_employee_range:
      employee_range_resolved, employee_range_source,
      employee_range_confidence, employee_range_notes.
    """
    _empty = {
        "employee_range_resolved":   "",
        "employee_range_source":     "missing",
        "employee_range_confidence": "None",
        "employee_range_notes":      "",
    }
    if not company_name or not serper_key:
        return _empty

    cn = company_name.strip()
    queries = [
        f'"{cn}" employees',
        f'"{cn}" "number of employees"',
        f'"{cn}" dipendenti',
        f'site:linkedin.com/company "{cn}" employees',
    ]
    if domain and domain.strip():
        queries.append(
            f'site:{domain.strip()} employees OR dipendenti OR workforce OR "team of"'
        )

    # Confidence heuristic: LinkedIn company page or own domain → High; others → Medium
    _STRONG_DOMAINS = ("linkedin.com", )

    best_n:    float | None = None
    best_text: str          = ""
    best_conf: str          = "None"

    for q in queries:
        try:
            hits, _status, _raw, _err = _call_serper(q, serper_key, timeout=10)
        except Exception:
            continue
        if not hits:
            continue
        for hit in hits:
            snippet = (hit.get("snippet") or "") + " " + (hit.get("title") or "")
            n, matched = _parse_employee_number(snippet)
            if n is not None and n > 0:
                hit_domain = hit.get("link", "")
                is_strong  = any(sd in hit_domain for sd in _STRONG_DOMAINS)
                if domain and domain.strip() in hit_domain:
                    is_strong = True
                conf = "High" if is_strong else "Medium"
                # Prefer High over Medium; among same level prefer larger sample
                if best_n is None or (conf == "High" and best_conf != "High"):
                    best_n    = n
                    best_text = matched
                    best_conf = conf
        if best_conf == "High":
            break  # good enough — stop querying

    if best_n is None:
        return _empty

    band = _num_to_size_band(best_n)
    return {
        "employee_range_resolved":   band,
        "employee_range_source":     "serper_employee_search",
        "employee_range_confidence": best_conf,
        "employee_range_notes":      f"Serper search found: \"{best_text}\"",
    }


# ─────────────────────────────────────────────────────────────────────────────
# Extreme Light Mode (ELM) — zero-token, no API key, keyword-only extraction
# ─────────────────────────────────────────────────────────────────────────────

_ELM_SLUGS = ["", "/about", "/about-us", "/careers", "/jobs", "/locations", "/contact"]

_ELM_UA = "Mozilla/5.0 (compatible; CompanyResearchBot/1.0)"

_ELM_KW_INTERNATIONAL = [
    "international", "global", "worldwide", "multinational", "cross-border",
    "offices in", "presence in", "emea", "apac", "latam", "global team",
    "international team", "countries", "regions",
]
_ELM_KW_LANGUAGES = [
    "english", "french", "german", "spanish", "portuguese", "dutch", "italian",
    "chinese", "mandarin", "japanese", "korean", "arabic", "russian", "polish",
    "turkish", "swedish", "norwegian", "danish", "finnish", "hebrew",
]
_ELM_KW_HIRING = [
    "careers", "jobs", "hiring", "join us", "join our team", "open positions",
    "vacancies", "we're growing", "recruitment", "apply now", "job openings",
    "we are hiring", "current openings",
]
_ELM_KW_GROWTH = [
    "funding", "raised", "series a", "series b", "series c", "ipo",
    "acquisition", "acquired", "merger", "expansion", "hypergrowth",
    "fast-growing", "scaling",
]
_ELM_KW_TRAINING = [
    "training", "learning", "development", "upskilling", "reskilling",
    "e-learning", "coaching", "mentoring", "academy", "bootcamp",
    "certification", "corporate training", "language training",
]
_ELM_KW_OFFICES = [
    "offices", "headquarters", "locations", "branches", "regional office",
    "hub", "campus", "sites", "hq",
]
_ELM_KW_TECHNOLOGY = [
    "saas", "cloud", "platform", "software", "machine learning", "artificial intelligence",
    "automation", "digital", "engineering", "devops", "api", "data-driven",
]
_ELM_COUNTRIES = [
    "united states", "usa", "united kingdom", "uk", "germany", "france",
    "netherlands", "spain", "italy", "portugal", "belgium", "switzerland",
    "sweden", "norway", "denmark", "finland", "poland", "czech republic",
    "australia", "canada", "india", "china", "japan", "singapore", "brazil",
    "mexico", "south africa", "uae", "israel", "ireland", "austria",
]

ELM_STATUS_FIELDS = [
    "elm_fetch_status",   # "ok" / "partial" / "failed"
    "elm_pages_fetched",  # comma-separated slugs that returned 200
    "elm_pages_failed",   # comma-separated slugs that failed/non-200
    "elm_total_chars",    # total chars fetched across all pages
    "elm_error",          # any fetch-level error message
]
ELM_KEYWORD_FIELDS = [
    "elm_kw_international",
    "elm_kw_languages",
    "elm_kw_hiring",
    "elm_kw_growth",
    "elm_kw_training",
    "elm_kw_offices",
    "elm_kw_technology",
    "elm_kw_countries_found",
]
ELM_SCORE_FIELDS = [
    "elm_score_international",
    "elm_score_hiring",
    "elm_score_growth",
    "elm_score_training",
    "elm_score_technology",
    "elm_score_overall_icp",
]
ELM_ALL_FIELDS = ELM_STATUS_FIELDS + ELM_KEYWORD_FIELDS + ELM_SCORE_FIELDS

# Fields checked to decide if Step 1 returned usable data
_STEP1_DATA_FIELDS = [
    "lusha_company_name", "lusha_domain", "lusha_industry",
    "lusha_country", "lusha_description",
]

_COMPANY_HINTS = ["company", "account", "organisation", "organization", "name", "naam", "bedrijf"]
_DOMAIN_HINTS  = ["domain", "website", "url", "web", "site", "domein",
                   "lusha_domain", "company_domain", "company_url"]

# ── Input Cleaner output detection ───────────────────────────────────────────

# Sheet names produced by input_cleaner_register_edition.py
_CLEANER_SHEET_NAMES: frozenset = frozenset({
    "Best Guess Input", "Cleaned Register Input", "Commercial Input", "Original Input",
})

# Column signatures unique to cleaner output (any one of these → cleaner output)
_CLEANER_COL_SIGNATURES: frozenset = frozenset({
    "final_selected_domain", "validated_domain", "recommended_domain",
    "python_validated_domain", "python_recommended_domain",
    "organization_type", "myngle_target_eligibility", "final_confidence",
    "domain_action", "business_output_reason", "pre_filter_decision",
})

# Ordered priority lists for column selection in cleaner output
_CLEANER_NAME_PRIORITY: list = [
    "company_name", "canonical_company_name", "cleaned_company_name",
    "Company Name", "company_name_clean", "company_name_raw",
    "name", "company", "organization_name", "organisation_name",
]
_CLEANER_DOMAIN_PRIORITY: list = [
    "website_url", "final_selected_domain", "canonical_company_domain",
    "validated_domain", "recommended_domain",
    "python_validated_domain", "python_recommended_domain",
    "Company Domain", "Company Website", "Website",
    "website", "domain", "url", "homepage",
]

# Columns that are NEVER a company-name column
_META_COLS_EXCLUDE: frozenset = frozenset({
    "organization_type", "organisation_type",
    "myngle_target_eligibility", "pre_filter_decision", "pre_filter_reason",
    "pre_score", "pre_label", "domain_action", "domain_confidence",
    "domain_reason", "final_confidence", "final_decision_source",
    "verifier_decision", "website_discovery_method", "professional_site_level",
    "source", "source_row_id", "current_status", "jurisdiction_code",
    "legal_form_detected", "registered_address", "city_or_registered_office",
    "federal_state", "registrar", "register_art", "register_nummer",
    "retrieved_at", "positive_reasons", "exclude_reasons", "low_priority_reasons",
    "manual_review_needed", "business_output_reason",
})
_META_SUFFIX_EXCLUDE: tuple = (
    "_status", "_source", "_reason", "_signal", "_signals",
    "_score", "_confidence", "_type", "_decision", "_action",
    "_label", "_notes", "_evidence",
)


def _is_meta_col(col_name: str) -> bool:
    if col_name in _META_COLS_EXCLUDE:
        return True
    cl = col_name.lower()
    return any(cl.endswith(sfx) for sfx in _META_SUFFIX_EXCLUDE)


# Priority list for resolving a human-readable company / register number from a row
_COMPANY_NUMBER_COLS: tuple = (
    "company_number", "Company Number", "native_company_number",
    "register_nummer", "source_row_id", "id",
)


def _get_company_number(row: dict) -> str:
    """Return the first non-blank company/register number found in *row*, or 'n/a'."""
    for col in _COMPANY_NUMBER_COLS:
        v = str(row.get(col, "") or "").strip()
        if v and v.lower() not in ("nan", "none", ""):
            return v
    return "n/a"


def _detect_company_number_col(df: "pd.DataFrame") -> str | None:
    """Return the first _COMPANY_NUMBER_COLS column that exists in df, or None."""
    for col in _COMPANY_NUMBER_COLS:
        if col in df.columns:
            return col
    return None


# Filename patterns that hint at cleaner output (case-insensitive)
_CLEANER_FNAME_PATTERNS: tuple = (
    "register_cleaned_",
    "_cleaned_",
    "cleaned_",
)


def _fname_looks_like_cleaner(fname: str) -> bool:
    fl = fname.lower()
    return any(p in fl for p in _CLEANER_FNAME_PATTERNS)


def is_input_cleaner_output(df: pd.DataFrame, fname: str = "") -> bool:
    """Return True if df/workbook looks like input_cleaner_register_edition.py output.

    Detection uses column signatures — filename hint is secondary.
    """
    cols = set(df.columns.tolist())
    # Any cleaner-specific column is sufficient
    if cols & _CLEANER_COL_SIGNATURES:
        return True
    # All three canonical cleaner columns present
    if {"company_name", "website_url", "final_selected_domain"}.issubset(cols):
        return True
    # Filename hint + has company_name
    if _fname_looks_like_cleaner(fname) and "company_name" in cols:
        return True
    return False


def get_cleaner_name_col(df: pd.DataFrame) -> str | None:
    """Return the best company-name column for a cleaner output df."""
    cols = set(df.columns)
    for cand in _CLEANER_NAME_PRIORITY:
        if cand in cols and not _is_meta_col(cand):
            return cand
    # Fuzzy fallback: first non-meta column with enough unique values
    for col in df.columns:
        if not _is_meta_col(col):
            return col
    return None


def get_cleaner_domain_col(df: pd.DataFrame) -> str | None:
    """Return the best domain column for a cleaner output df, or None."""
    cols = set(df.columns)
    for cand in _CLEANER_DOMAIN_PRIORITY:
        if cand in cols:
            return cand
    return None


def _col_has_data(df: pd.DataFrame, col: str | None) -> bool:
    if not col or col not in df.columns:
        return False
    return df[col].astype(str).str.strip().replace("nan", "").ne("").any()


def load_cleaner_workbook(file_obj, fname: str) -> tuple:
    """Load the best sheet from a cleaner workbook.

    Returns (df, sheet_name).  file_obj may be a path or a file-like object.
    """
    xl = pd.ExcelFile(file_obj)
    sheets = xl.sheet_names

    # Priority 1: Best Guess Input with company_name data
    if "Best Guess Input" in sheets:
        df = xl.parse("Best Guess Input")
        if _col_has_data(df, "company_name"):
            return df, "Best Guess Input"

    # Priority 2: Commercial Input
    if "Commercial Input" in sheets:
        df = xl.parse("Commercial Input")
        cols = set(df.columns)
        if _col_has_data(df, get_cleaner_name_col(df) if cols else None):
            # Guard against empty/placeholder sheets like "No rows in this category."
            if len(df) > 0 and len(df.columns) > 2:
                return df, "Commercial Input"

    # Priority 3: Cleaned Register Input
    if "Cleaned Register Input" in sheets:
        df = xl.parse("Cleaned Register Input")
        name_c = get_cleaner_name_col(df)
        if _col_has_data(df, name_c):
            return df, "Cleaned Register Input"

    # Priority 4: Original Input
    if "Original Input" in sheets:
        df = xl.parse("Original Input")
        name_c = get_cleaner_name_col(df)
        if _col_has_data(df, name_c):
            return df, "Original Input"

    # Fallback: first sheet
    df = xl.parse(sheets[0])
    return df, sheets[0]

# Prefixes and field names used to detect existing Lusha/Lucia enrichment columns
_LUSHA_COL_PREFIXES = ("lusha_", "lusha_api_", "lucia_")
_LUSHA_COL_NAMES_EXACT = frozenset([
    "company_size", "employee_range", "employee_size_score",
    "revenue", "revenue_range", "industry", "sub_industry",
    "founded_year", "country", "city", "linkedin_url",
])

# Lucia/Lusha contact export column signatures (company-level fields)
_LUCIA_EXPORT_COMPANY_COLS = frozenset([
    "Company Name", "Company Domain", "Company Description",
    "Company Year Founded", "Company Website",
    "Company Number of Employees", "Company Revenue",
    "Company linkedin URL", "Total Funding Amount",
    "Total Number of Rounds", "Last Round/Event Amount",
    "Last Round/Event Type", "Last Round/Event Date",
    "IPO Status", "Company Main Industry", "Company Sub Industry",
    "Company Technologies", "Company Specialties",
    "Company Continent", "Company Country", "Company State",
    "Company City", "Company Country ISO",
])

# Mapping: Lucia/Lusha CSV column name → internal lusha_api_* field name
_LUCIA_COL_MAP = {
    "Company Name":                "lusha_api_company_name",
    "Company Domain":              "lusha_api_domain",
    "Company Website":             "lusha_api_domain",   # fallback if no Company Domain
    "Company Description":         "lusha_api_description",
    "Company Year Founded":        "lusha_api_founded_year",
    "Company Number of Employees": "lusha_api_employee_range",
    "Company Revenue":             "lusha_api_revenue_range",
    "Company Main Industry":       "lusha_api_industry",
    "Company Sub Industry":        "lusha_api_sub_industry",
    "Company Country":             "lusha_api_country",
    "Company City":                "lusha_api_city",
    "Company Continent":           "lusha_api_continent",
    "Company linkedin URL":        "lusha_api_linkedin_url",
    "Company Specialties":         "lusha_api_specialties",
    "Company Technologies":        "lusha_api_technologies",
    "Total Funding Amount":        "lusha_api_total_funding_amount",
    "Total Number of Rounds":      "lusha_api_total_funding_rounds",
    "Last Round/Event Type":       "lusha_api_last_round_type",
    "Last Round/Event Amount":     "lusha_api_last_round_amount",
    "Last Round/Event Date":       "lusha_api_last_round_date",
    "IPO Status":                  "lusha_api_ipo_status",
}

_STATUS_LABELS = {
    "enriched_jina":                 "Enriched via Jina",
    "enriched_search":               "Enriched via Google",
    "enriched_jina_step1_only":      "Jina — Step 1 only",
    "enriched_search_step1_only":    "Google — Step 1 only",
    "enriched":                      "Enriched (both steps)",
    "step1_only":                    "Step 1 only",
    "no_data":                       "No data returned",
    "api_error":                     "API error",
    "jina_error":                    "Page fetch error",
    "web_search_fallback":           "Enriched via web search fallback",
    "cached_fallback":               "From cache (web search fallback)",
    "skipped_resume":                "Skipped (resumed)",
    "enriched_playwright":           "Enriched via browser scrape",
    "playwright_blocked":            "Browser blocked (bot detection)",
    "zero_cost_preview":             "Zero-cost preview",
}


# ─────────────────────────────────────────────────────────────────────────────
# Utility
# ─────────────────────────────────────────────────────────────────────────────

def clean_domain(raw: str) -> str:
    if not raw or not isinstance(raw, str):
        return ""
    d = raw.strip().lower()
    d = re.sub(r"^https?://", "", d)
    d = re.sub(r"^www\.", "", d)
    d = d.split("/")[0].strip()
    return "" if d in {"nan", "none", ""} or " " in d else d


def normalize_url(raw: str) -> str:
    if not raw or not isinstance(raw, str):
        return ""
    raw = raw.strip()
    if raw.lower() in {"nan", "none", ""} or " " in raw:
        return ""
    return raw if raw.startswith(("http://", "https://")) else f"https://{raw}"


def safe_filename(text: str) -> str:
    text = unicodedata.normalize("NFKD", str(text))
    text = re.sub(r"[^\w\s\-.]", "", text)
    text = re.sub(r"\s+", "_", text).strip("_")
    return text[:120] or "unknown"


# ─────────────────────────────────────────────────────────────────────────────
# Step 2 debug logging helpers
# ─────────────────────────────────────────────────────────────────────────────

def ensure_debug_log_dir() -> None:
    DEBUG_LOG_DIR.mkdir(exist_ok=True)


def append_to_debug_file(path: Path, content: str) -> None:
    """Append content to an existing debug file (creates it if missing)."""
    if not path or not _is_safe_debug_path(str(path)):
        return
    with path.open("a", encoding="utf-8") as fh:
        fh.write(content)


def write_debug_log(company_name: str, content: str, prefix: str = "step2_prompt") -> Path:
    ensure_debug_log_dir()
    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    fname = DEBUG_LOG_DIR / f"{prefix}_{safe_filename(company_name)}_{stamp}.txt"
    fname.write_text(content, encoding="utf-8")
    return fname


def append_debug_log(message: str) -> None:
    """Append a timestamped message to the in-session debug log."""
    if is_cli_mode():
        return
    _st, _ = get_streamlit()
    stamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    entry = f"[{stamp}] {message}\n"
    current = _st.session_state.get("_step2_debug_log", "")
    _st.session_state["_step2_debug_log"] = current + entry


def format_step2_debug_content(
    company_name: str,
    model: str,
    timestamp: str,
    provider: str,
    search_prompt: str,
    full_prompt: str,
    notes: list,
) -> str:
    sep  = "=" * 60
    thin = "-" * 40
    note_block = "\n".join(notes) if notes else "(none)"
    return (
        f"{sep}\n"
        f"COMPANY:             {company_name}\n"
        f"MODEL:               {model}\n"
        f"TIMESTAMP:           {timestamp}\n"
        f"WEB SEARCH PROVIDER: {provider}\n"
        f"\nGENERATED SEARCH PROMPT\n{thin}\n{search_prompt}\n"
        f"\nFULL CLAUDE PROMPT\n{thin}\n{full_prompt}\n"
        f"\nSTATUS / NOTES\n{thin}\n{note_block}\n"
        f"{sep}\n"
    )


def safe_json_dump(obj) -> str:
    """Serialize API response objects to indented JSON without exposing secrets."""
    if obj is None:
        return "null"
    try:
        return json.dumps(obj, ensure_ascii=False, indent=2, default=str)
    except (TypeError, ValueError):
        pass
    if hasattr(obj, "model_dump"):
        try:
            return json.dumps(obj.model_dump(), ensure_ascii=False, indent=2, default=str)
        except Exception:
            pass
    if hasattr(obj, "dict"):
        try:
            return json.dumps(obj.dict(), ensure_ascii=False, indent=2, default=str)
        except Exception:
            pass
    return str(obj)


def ensure_search_output_dir() -> None:
    SEARCH_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def write_search_debug_file(
    company_name: str,
    provider_label: str,
    content: str,
    ext: str = "txt",
    index: int = 1,
) -> Path:
    """Write one search debug file; return the path."""
    ensure_search_output_dir()
    stamp    = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    safe_co  = safe_filename(company_name)
    fname    = SEARCH_OUTPUT_DIR / (
        f"step2_search_{safe_co}_{provider_label}_{stamp}_search{index:02d}.{ext}"
    )
    fname.write_text(content, encoding="utf-8")
    return fname


_SAFE_DEBUG_DIRS = (
    str(DEBUG_LOG_DIR.resolve()),
    str(SEARCH_OUTPUT_DIR.resolve()),
)


def _is_safe_debug_path(p: str) -> bool:
    """Return True only when p resolves inside debug_logs/ or its subdirectories."""
    try:
        return str(Path(p).resolve()).startswith(_SAFE_DEBUG_DIRS)
    except Exception:
        return False


def build_debug_zip(file_records: list) -> bytes:
    """
    Build an in-memory ZIP of all debug files in file_records.
    Only includes files that resolve inside the debug_logs/ tree.
    Returns raw bytes suitable for st.download_button.
    """
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        seen_arcnames: set = set()
        for rec in file_records:
            fpath = rec.get("path", "")
            if not fpath or not _is_safe_debug_path(fpath):
                continue
            p = Path(fpath)
            if not p.exists():
                continue
            # Use the path relative to the parent of debug_logs/ as the archive name
            try:
                arcname = str(p.relative_to(Path(".")))
            except ValueError:
                arcname = p.name
            # Deduplicate archive names
            base, suffix = arcname, ""
            counter = 1
            while arcname + suffix in seen_arcnames:
                suffix = f"_{counter}"
                counter += 1
            arcname = arcname + suffix
            seen_arcnames.add(arcname)
            try:
                zf.write(p, arcname=arcname)
            except Exception:
                pass
    return buf.getvalue()


def _read_debug_file_safe(fpath: str, max_chars: int = 8000) -> tuple[str, bool]:
    """
    Read a debug file for preview.  Returns (content, truncated).
    Only reads files inside the safe debug dirs.
    """
    if not fpath or not _is_safe_debug_path(fpath):
        return "(file not accessible)", False
    try:
        text = Path(fpath).read_text(encoding="utf-8", errors="replace")
        if len(text) > max_chars:
            return text[:max_chars], True
        return text, False
    except Exception as exc:
        return f"(could not read file: {exc})", False


def _debug_file_download_button(rec: dict, key_suffix: str) -> None:
    """Render a download button for one debug file record."""
    fpath = rec.get("path", "")
    if not fpath or not _is_safe_debug_path(fpath):
        return
    p = Path(fpath)
    _st, _ = get_streamlit()
    if not p.exists():
        _st.caption(f"_(file no longer on disk: `{p.name}`)_")
        return
    company = rec.get("company", "")
    label = f"⬇ Download {company} debug file" if company else f"⬇ Download {p.name}"
    try:
        data = p.read_bytes()
        _st.download_button(
            label=label,
            data=data,
            file_name=p.name,
            mime="text/plain",
            key=f"dl_dbg_{key_suffix}",
        )
    except Exception:
        _st.caption(f"_(download unavailable for `{p.name}`)_")


def _debug_file_preview_expander(rec: dict, key_suffix: str) -> None:
    """Render a collapsed preview expander for one debug file record."""
    fpath = rec.get("path", "")
    if not fpath:
        return
    p = Path(fpath)
    _st, _ = get_streamlit()
    with _st.expander(f"Preview: {p.name}", expanded=False):
        text, truncated = _read_debug_file_safe(fpath)
        _st.code(text, language=None)
        if truncated:
            _st.caption("_(preview truncated to 8 000 chars — download the full file above)_")


def _format_serper_query_debug(
    company_name: str,
    model: str,
    timestamp: str,
    query: str,
    index: int,
    total_queries: int,
    results: list,
    http_status: int,
    raw_json,
    error_str,
    dry_run: bool = False,
) -> str:
    sep  = "=" * 60
    thin = "-" * 40
    payload_safe = safe_json_dump({
        "q": query, "gl": "us", "hl": "en", "num": 10,
        # API key is intentionally excluded
    })
    if dry_run:
        result_block = "[DRY RUN: no web search output because no API call was made]"
        http_block   = "N/A (dry run)"
        raw_block    = "[DRY RUN: no raw response]"
    else:
        http_block = str(http_status) if http_status else "0 (network error)"
        if error_str:
            result_block = f"ERROR: {error_str}"
            raw_block    = safe_json_dump(raw_json) if raw_json else "(no response body)"
        else:
            raw_block = safe_json_dump(raw_json) if raw_json else "(not captured)"
            lines = []
            for r in results:
                lines.append(f"[{r.get('position', '?')}] {r.get('title', '(no title)')}")
                if r.get("date"):
                    lines.append(f"    Date: {r['date']}")
                lines.append(f"    URL:  {r.get('link', '')}")
                lines.append(f"    {r.get('snippet', '')}")
                sl = r.get("sitelinks", [])
                if sl:
                    lines.append(f"    Sitelinks: {sl}")
                lines.append("")
            result_block = "\n".join(lines).strip() or "(no results)"

    mode_note = "DRY RUN — no API call was made." if dry_run else ""
    return (
        f"{sep}\n"
        f"STEP 2 SERPER SEARCH DEBUG{' — DRY RUN' if dry_run else ''}\n"
        f"{sep}\n"
        f"COMPANY:              {company_name}\n"
        f"MODEL:                {model}\n"
        f"PROVIDER:             Serper Google Search\n"
        f"TIMESTAMP:            {timestamp} UTC\n"
        f"QUERY INDEX:          {index} of {total_queries}\n"
        f"\nSERPER QUERY\n{thin}\n{query}\n"
        f"\nSERPER REQUEST PAYLOAD (API key excluded)\n{thin}\n{payload_safe}\n"
        f"\nHTTP STATUS CODE\n{thin}\n{http_block}\n"
        f"\nRAW SERPER RESPONSE\n{thin}\n{raw_block}\n"
        f"\nEXTRACTED ORGANIC RESULTS ({len(results)} result(s))\n{thin}\n{result_block}\n"
        + (f"\nSTATUS / NOTES\n{thin}\n{mode_note or '(none)'}\n" if mode_note or error_str else "")
        + f"{sep}\n"
    )


def _format_claude_pre_debug(
    company_name: str,
    model: str,
    timestamp: str,
    prompt: str,
    dry_run: bool = False,
) -> str:
    sep  = "=" * 60
    thin = "-" * 40
    tools_safe = safe_json_dump([WEB_SEARCH_TOOL])
    mode = "DRY RUN — no API call will be made." if dry_run else ""
    return (
        f"{sep}\n"
        f"STEP 2 CLAUDE WEB SEARCH — PRE-CALL DEBUG{' (DRY RUN)' if dry_run else ''}\n"
        f"{sep}\n"
        f"COMPANY:              {company_name}\n"
        f"MODEL:                {model}\n"
        f"PROVIDER:             Claude Web Search\n"
        f"TIMESTAMP:            {timestamp} UTC\n"
        f"\nNOTE ON INTERNAL SEARCH QUERY\n{thin}\n"
        "The web_search_20250305 tool is a server-side built-in Anthropic tool.\n"
        "Claude generates the actual search query internally; it is NOT exposed\n"
        "by the API response.\n"
        "\n\"Exact internal Claude Web Search query was not exposed by the API response.\"\n"
        f"\nFULL STEP 2 PROMPT SENT TO CLAUDE\n{thin}\n{prompt}\n"
        f"\nTOOLS CONFIGURATION (secrets excluded)\n{thin}\n{tools_safe}\n"
        + (f"\nSTATUS / NOTES\n{thin}\n{mode}\n" if mode else "")
        + f"{sep}\n"
    )


def _format_claude_post_debug(
    company_name: str,
    model: str,
    timestamp: str,
    resp,
    raw_text: str,
    error_str: str = "",
    parsed_json=None,
    dry_run_skip: bool = False,
) -> str:
    sep  = "=" * 60
    thin = "-" * 40

    if dry_run_skip:
        return (
            f"\n\n{sep}\n"
            f"STEP 2 CLAUDE WEB SEARCH — POST-CALL DEBUG (NOT EXECUTED)\n"
            f"{sep}\n"
            f"COMPANY:   {company_name}\n"
            f"MODEL:     {model}\n"
            f"TIMESTAMP: {timestamp} UTC\n"
            f"\nPOST-CALL DEBUG: skipped because dry run / zero-cost preview was active.\n"
            f"No Anthropic API call was made.\n"
            f"{sep}\n"
        )

    if error_str:
        return (
            f"\n\n{sep}\n"
            f"STEP 2 CLAUDE WEB SEARCH — POST-CALL DEBUG\n"
            f"{sep}\n"
            f"COMPANY:   {company_name}\n"
            f"MODEL:     {model}\n"
            f"TIMESTAMP: {timestamp} UTC\n"
            f"\nERROR\n{thin}\n{error_str}\n"
            f"{sep}\n"
        )

    stop_reason = getattr(resp, "stop_reason", "unknown") if resp else "unknown"
    usage       = getattr(resp, "usage", None)
    in_tok      = getattr(usage, "input_tokens",  "?") if usage else "?"
    out_tok     = getattr(usage, "output_tokens", "?") if usage else "?"

    tool_use_blocks    = []
    tool_result_blocks = []
    web_search_blocks  = []
    citation_blocks    = []
    other_blocks       = []

    if resp and hasattr(resp, "content"):
        for blk in resp.content:
            btype = getattr(blk, "type", "")
            if btype == "tool_use":
                tool_use_blocks.append(blk)
            elif btype == "tool_result":
                tool_result_blocks.append(blk)
            elif btype in ("web_search_result", "server_tool_use"):
                web_search_blocks.append(blk)
            elif btype == "text":
                pass  # already in raw_text
            else:
                other_blocks.append(blk)
        # Collect citations if present on text blocks
        for blk in resp.content:
            if getattr(blk, "type", "") == "text":
                cits = getattr(blk, "citations", []) or []
                citation_blocks.extend(cits)

    def _blk_section(label, blocks):
        if not blocks:
            return f"\n{label}\n{thin}\n(none)\n"
        lines = [f"\n{label}\n{thin}"]
        for b in blocks:
            lines.append(safe_json_dump(b))
        return "\n".join(lines) + "\n"

    citation_section = ""
    if citation_blocks:
        citation_section = f"\nCITATIONS / SOURCE REFERENCES\n{thin}\n"
        for c in citation_blocks:
            citation_section += safe_json_dump(c) + "\n"
    else:
        citation_section = f"\nCITATIONS / SOURCE REFERENCES\n{thin}\n(none)\n"

    raw_resp_dump = safe_json_dump(resp) if resp else "(not available)"

    parsed_section = ""
    if parsed_json is not None:
        parsed_section = (
            f"\nPARSED JSON OUTPUT\n{thin}\n"
            + safe_json_dump(parsed_json) + "\n"
        )
    else:
        parsed_section = f"\nPARSED JSON OUTPUT\n{thin}\n(not available or parse failed)\n"

    return (
        f"\n\n{sep}\n"
        f"STEP 2 CLAUDE WEB SEARCH — POST-CALL DEBUG\n"
        f"{sep}\n"
        f"COMPANY:              {company_name}\n"
        f"MODEL:                {model}\n"
        f"PROVIDER:             Claude Web Search\n"
        f"TIMESTAMP:            {timestamp} UTC\n"
        f"STOP REASON:          {stop_reason}\n"
        f"USAGE:                input_tokens={in_tok}, output_tokens={out_tok}\n"
        f"\nRESPONSE TEXT\n{thin}\n{raw_text or '(empty)'}\n"
        + _blk_section("TOOL USE BLOCKS", tool_use_blocks)
        + _blk_section("TOOL RESULT BLOCKS", tool_result_blocks)
        + _blk_section("WEB SEARCH / SOURCE BLOCKS", web_search_blocks)
        + citation_section
        + parsed_section
        + f"\nRAW RESPONSE OBJECT\n{thin}\n{raw_resp_dump}\n"
        + f"{sep}\n"
    )


def str_similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


_LEGAL_RE = re.compile(
    r"[\s,\-\.]*\b("
    r"b\.?v\.?|n\.?v\.?|s\.?a\.?s?\.?|s\.?p\.?a\.?|a\.?/\.?s\.?|a\.?s\.?"
    r"|g\.?m\.?b\.?h\.?|ag|ltd\.?|limited|inc\.?|corp\.?|llc|llp|plc"
    r"|oy|ab|s\.?r\.?l\.?|s\.?n\.?c\.?|kft|s\.?r\.?o\.?|o\.?[üu]\.?"
    r"|pte\.?|pty\.?|cv|vof|gg"
    r")\b\.?",
    re.IGNORECASE,
)

def _strip_legal(name: str) -> str:
    return _LEGAL_RE.sub(" ", name).strip(" .,/-")

def _legal_suffix(name: str) -> str:
    hits = _LEGAL_RE.findall(name)
    return re.sub(r"[^a-z0-9]", "", hits[-1].lower()) if hits else ""


# ─────────────────────────────────────────────────────────────────────────────
# Domain validation helpers
# ─────────────────────────────────────────────────────────────────────────────

_GENERIC_DOMAINS: frozenset = frozenset({
    "linkedin.com", "facebook.com", "twitter.com", "x.com", "instagram.com",
    "youtube.com", "wikipedia.org", "bloomberg.com", "crunchbase.com",
    "glassdoor.com", "indeed.com", "xing.com", "angel.co", "pitchbook.com",
    "google.com", "bing.com", "yahoo.com", "reuters.com", "ft.com",
    "github.com", "amazon.com", "zoominfo.com", "dnb.com",
    "opencorporates.com", "companieshouse.gov.uk", "app.lusha.com",
})

_DV_NOISE_TOKENS: frozenset = frozenset({
    "the", "and", "for", "group", "holding", "holdings", "global",
    "international", "services", "solutions", "consulting", "management",
    "technology", "technologies", "systems", "software", "digital",
    "company", "corp", "enterprise", "enterprises", "partners",
    "nederland", "netherlands", "deutschland", "germany", "france",
    "belgium", "europe", "european", "asia",
})

_DV_TLDS: frozenset = frozenset({
    "com","net","org","io","co","nl","de","fr","be","uk","us","eu",
    "info","biz","ch","at","se","no","dk","fi","pl","es","it","pt",
    "ru","cn","jp","au","ca","br","in","sg","ae","nz","mx","za",
    "ie","lu","hu","cz","ro","gr","hr","sk","si","ee","lv","lt",
})


def _dv_domain_tokens(domain: str) -> set:
    """Extract meaningful stem tokens from a domain (strips TLDs and www)."""
    d = clean_domain(domain)
    if not d:
        return set()
    parts = d.split(".")
    while len(parts) > 1 and parts[-1].lower() in _DV_TLDS:
        parts = parts[:-1]
    base = ".".join(parts)
    return {t for t in re.split(r"[-.]", base.lower()) if t and len(t) >= 2}


def _dv_company_tokens(name: str) -> set:
    """Extract meaningful tokens from a company name (strips legal forms and noise)."""
    clean = _strip_legal(name)
    clean = re.sub(r"[^\w\s\-]", " ", clean)
    tokens = {t.lower() for t in re.split(r"[\s\-_]+", clean) if len(t) >= 3}
    return tokens - _DV_NOISE_TOKENS


def _dv_token_overlap(company_name: str, domain: str) -> float:
    """Overlap ratio [0..1] between company name tokens and domain stem tokens.
    Includes substring containment so 'capgemini nederland' matches 'capgemini.com'."""
    ctok = _dv_company_tokens(company_name)
    dtok = _dv_domain_tokens(domain)
    if not ctok or not dtok:
        return 0.0
    overlap: set = ctok & dtok
    for c in ctok:
        for d in dtok:
            if c in d or d in c:
                overlap.add(c)
    return len(overlap) / min(len(ctok), len(dtok))


def is_lucia_contact_export(df: pd.DataFrame) -> bool:
    """Return True if df is a pre-enriched Lucia/Lusha contact export (Type 2).

    Requires: 'Company Name' AND ('Company Domain' OR 'Company Website')
    plus at least 2 additional Lucia-specific company-level columns.
    """
    cols = set(df.columns.tolist())
    if "Company Name" not in cols:
        return False
    if not (cols & {"Company Domain", "Company Website"}):
        return False
    extra = _LUCIA_EXPORT_COMPANY_COLS - {"Company Name", "Company Domain", "Company Website"}
    return len(cols & extra) >= 2


def get_lucia_name_col(df: pd.DataFrame) -> str | None:
    """Return 'Company Name' if present, else None."""
    return "Company Name" if "Company Name" in df.columns else None


def get_lucia_domain_col(df: pd.DataFrame) -> str | None:
    """Return 'Company Domain' or 'Company Website' for Lucia exports, else None."""
    if "Company Domain" in df.columns:
        return "Company Domain"
    if "Company Website" in df.columns:
        return "Company Website"
    return None


def map_lucia_export_row(row_dict: dict) -> dict:
    """Map a Lucia/Lusha CSV export row to internal lusha_api_* field names.

    'Company Domain' takes priority over 'Company Website' for lusha_api_domain.
    Always adds status fields marking the data as reused (no API call needed).
    """
    mapped: dict = {}
    domain_set = False
    for src_col, tgt_field in _LUCIA_COL_MAP.items():
        val = row_dict.get(src_col)
        if val is None or str(val).strip().lower() in ("", "nan", "none"):
            continue
        val_str = str(val).strip()
        if tgt_field == "lusha_api_domain":
            if src_col == "Company Domain":
                mapped[tgt_field] = val_str
                domain_set = True
            elif src_col == "Company Website" and not domain_set:
                mapped[tgt_field] = val_str
        else:
            mapped[tgt_field] = val_str
    mapped["lusha_api_status"]    = "reused_existing_lucia_data"
    mapped["lucia_data_status"]   = "existing_lucia_export_reused"
    mapped["lucia_api_called"]    = "False"
    mapped["lusha_api_match_notes"] = (
        "Existing Lucia/Lusha export data reused; API call skipped."
    )
    return mapped


def deduplicate_lucia_export(df: pd.DataFrame) -> pd.DataFrame:
    """Deduplicate a Lucia contact export to one row per unique company.

    Primary dedup key: cleaned Company Domain.
    Fallback: cleaned Company Name.
    Adds 'source_contact_count' column.
    """
    df = df.copy()
    domain_col = (
        "Company Domain" if "Company Domain" in df.columns else
        "Company Website" if "Company Website" in df.columns else None
    )
    name_col = "Company Name" if "Company Name" in df.columns else None

    if domain_col:
        df["_dedup_key"] = df[domain_col].apply(
            lambda x: clean_domain(str(x)) if pd.notna(x) else ""
        )
        if name_col:
            df["_dedup_key"] = df.apply(
                lambda r: r["_dedup_key"] if r["_dedup_key"]
                else str(r.get(name_col, "")).strip().lower(),
                axis=1,
            )
    elif name_col:
        df["_dedup_key"] = df[name_col].apply(lambda x: str(x).strip().lower())
    else:
        df["_dedup_key"] = df.index.astype(str)

    counts = df.groupby("_dedup_key").size().rename("source_contact_count")
    df_dedup = df.drop_duplicates(subset=["_dedup_key"], keep="first").copy()
    df_dedup = df_dedup.merge(counts, on="_dedup_key", how="left")
    df_dedup = df_dedup.drop(columns=["_dedup_key"])
    return df_dedup.reset_index(drop=True)


def normalize_input_to_company_df(
    raw_input_df: pd.DataFrame,
    detected_input_type: str,
    company_name_col: str | None = None,
    domain_col: str | None = None,
) -> dict:
    """Convert raw input into a canonical company-level dataframe.

    Adds canonical_company_name, canonical_company_domain, canonical_company_url
    to the returned company_df. For pre_enriched_lucia_export also deduplicates
    and pre-maps all Lucia 'Company X' columns to lusha_api_* fields so the
    processing loop and scoring engine see them immediately.

    Returns:
        {
          "company_df":           pd.DataFrame,
          "raw_input_df":         pd.DataFrame (original unchanged),
          "input_type":           str,
          "company_name_col":     str | None,
          "domain_col":           str | None,
          "contact_row_count":    int,
          "unique_company_count": int,
          "mapping_notes":        str,
        }
    """
    contact_row_count = len(raw_input_df)
    notes: list[str] = []

    if detected_input_type == "pre_enriched_lucia_export":
        # Step 1: deduplicate contact rows → one row per company
        company_df = deduplicate_lucia_export(raw_input_df)

        # Step 2: resolve company-level columns
        name_col_use   = company_name_col or get_lucia_name_col(raw_input_df)
        domain_col_use = domain_col       or get_lucia_domain_col(raw_input_df)

        # Step 3: pre-map all Lucia 'Company X' columns → lusha_api_* fields
        mapped_rows = [map_lucia_export_row(row.to_dict())
                       for _, row in company_df.iterrows()]
        all_lusha_keys: set[str] = set()
        for m in mapped_rows:
            all_lusha_keys.update(m.keys())
        for lusha_col in all_lusha_keys:
            company_df[lusha_col] = [m.get(lusha_col, "") for m in mapped_rows]

        # Step 4: add canonical identity columns
        if name_col_use and name_col_use in company_df.columns:
            company_df["canonical_company_name"] = (
                company_df[name_col_use].astype(str).str.strip()
            )
        else:
            company_df["canonical_company_name"] = ""

        if domain_col_use and domain_col_use in company_df.columns:
            company_df["canonical_company_domain"] = company_df[domain_col_use].apply(
                lambda x: clean_domain(str(x)) if pd.notna(x) else ""
            )
            company_df["canonical_company_url"] = company_df[domain_col_use].apply(
                lambda x: normalize_url(str(x)) if pd.notna(x) and str(x).strip() else ""
            )
        else:
            company_df["canonical_company_domain"] = ""
            company_df["canonical_company_url"]    = ""

        company_df["input_type"] = "pre_enriched_lucia_export"
        unique_company_count = len(company_df)
        notes.append(
            f"Deduplicated {contact_row_count} contact rows → {unique_company_count} companies."
        )

    elif detected_input_type == "input_cleaner_output":
        company_df     = raw_input_df.copy()
        name_col_use   = company_name_col or get_cleaner_name_col(raw_input_df)
        domain_col_use = domain_col       or get_cleaner_domain_col(raw_input_df)

        # Domain fallback columns for per-row resolution
        _dom_fallbacks = [
            c for c in _CLEANER_DOMAIN_PRIORITY if c in company_df.columns
        ]

        if name_col_use and name_col_use in company_df.columns:
            company_df["canonical_company_name"] = (
                company_df[name_col_use].astype(str).str.strip()
            )
        else:
            company_df["canonical_company_name"] = ""

        def _resolve_domain(row):
            for fc in _dom_fallbacks:
                v = str(row.get(fc, "") or "").strip()
                if v and v.lower() not in ("nan", "none", ""):
                    d = clean_domain(v)
                    if d:
                        return d
            return ""

        def _resolve_url(row):
            for fc in _dom_fallbacks:
                v = str(row.get(fc, "") or "").strip()
                if v and v.lower() not in ("nan", "none", ""):
                    return normalize_url(v)
            return ""

        company_df["canonical_company_domain"] = [
            _resolve_domain(r) for _, r in company_df.iterrows()
        ]
        company_df["canonical_company_url"] = [
            _resolve_url(r) for _, r in company_df.iterrows()
        ]

        if "source_contact_count" not in company_df.columns:
            company_df["source_contact_count"] = 1
        company_df["input_type"] = "input_cleaner_output"
        unique_company_count = len(company_df)

    else:  # simple_company_list (default)
        company_df     = raw_input_df.copy()
        name_col_use   = company_name_col
        domain_col_use = domain_col

        if name_col_use and name_col_use in company_df.columns:
            company_df["canonical_company_name"] = (
                company_df[name_col_use].astype(str).str.strip()
            )
        elif company_df.columns.tolist():
            name_col_use = company_df.columns[0]
            company_df["canonical_company_name"] = (
                company_df[name_col_use].astype(str).str.strip()
            )
        else:
            company_df["canonical_company_name"] = ""

        if domain_col_use and domain_col_use in company_df.columns:
            company_df["canonical_company_domain"] = company_df[domain_col_use].apply(
                lambda x: clean_domain(str(x)) if pd.notna(x) else ""
            )
            company_df["canonical_company_url"] = company_df[domain_col_use].apply(
                lambda x: normalize_url(str(x)) if pd.notna(x) and str(x).strip() else ""
            )
        else:
            company_df["canonical_company_domain"] = ""
            company_df["canonical_company_url"]    = ""

        if "source_contact_count" not in company_df.columns:
            company_df["source_contact_count"] = 1
        company_df["input_type"] = "simple_company_list"
        unique_company_count = len(company_df)

    return {
        "company_df":           company_df,
        "raw_input_df":         raw_input_df,
        "input_type":           detected_input_type,
        "company_name_col":     name_col_use,
        "domain_col":           domain_col_use,
        "contact_row_count":    contact_row_count,
        "unique_company_count": unique_company_count,
        "mapping_notes":        " ".join(notes),
    }


def detect_columns(df: pd.DataFrame, fname: str = "") -> tuple:
    """Detect company name and domain/URL columns in df.

    Priority order:
    1. Lucia/Lusha contact export — exact column match, bypasses fuzzy matching.
    2. Input Cleaner output — deterministic priority lists, bypasses fuzzy matching.
    3. Simple company list — fuzzy similarity scoring.
    """
    # Priority 1: Lucia/Lusha contact exports
    if is_lucia_contact_export(df):
        return get_lucia_name_col(df), get_lucia_domain_col(df)

    # Priority 2: Input Cleaner output
    if is_input_cleaner_output(df, fname):
        return get_cleaner_name_col(df), get_cleaner_domain_col(df)

    cols      = df.columns.tolist()
    col_lower = [str(c).lower() for c in cols]

    def best(hints):
        scores = [(max(str_similarity(cl, h) for h in hints), i)
                  for i, cl in enumerate(col_lower)]
        score, idx = max(scores)
        return cols[idx], score

    # Priority 3: exact priority list before fuzzy — prefers _clean over _raw variants
    _NAME_PRIORITY_GENERIC = [
        "company_name", "canonical_company_name", "cleaned_company_name",
        "company_name_clean", "Company Name", "name", "company",
        "company_name_raw", "organisation_name", "organization_name",
    ]
    name_col = None
    for _np in _NAME_PRIORITY_GENERIC:
        if _np in cols and not _is_meta_col(_np):
            name_col = _np
            break

    if name_col is None:
        # fuzzy matching — exclude metadata columns from name detection
        _eligible = [
            (i, cl) for i, cl in enumerate(col_lower)
            if not _is_meta_col(cols[i])
        ]
        if _eligible:
            name_scores = [(max(str_similarity(cl, h) for h in _COMPANY_HINTS), i)
                           for i, cl in _eligible]
            ns, n_idx = max(name_scores)
            name_col  = cols[n_idx] if ns >= 0.45 else None

    domain_col, ds = best(_DOMAIN_HINTS)
    # Reject domain_col if it looks like a name column (e.g. company_name_raw)
    _dom_rejected = (
        ds < 0.55
        or domain_col == name_col
        or "company_name" in (domain_col or "").lower()
        or _is_meta_col(domain_col or "")
    )
    return (
        name_col,
        domain_col if not _dom_rejected else None,
    )


def detect_lusha_columns(df: pd.DataFrame) -> list:
    """Return list of column names in df that look like existing Lusha/Lucia enrichment fields."""
    # Lucia/Lusha contact export: return all recognised company-level columns
    if is_lucia_contact_export(df):
        return [c for c in df.columns if c in _LUCIA_EXPORT_COMPANY_COLS]

    found = []
    for col in df.columns:
        col_l = col.lower().strip()
        if any(col_l.startswith(p) for p in _LUSHA_COL_PREFIXES):
            found.append(col)
        elif col_l in _LUSHA_COL_NAMES_EXACT:
            found.append(col)
    return found


def calc_cost(in_tok: int, out_tok: int) -> float:
    return (in_tok * _COST_INPUT_PER_M + out_tok * _COST_OUTPUT_PER_M) / 1_000_000


def _parse_json_response(text: str) -> dict:
    """
    Extract and parse JSON from Claude's response.
    1. Try stripping markdown fences and parsing directly.
    2. Fall back to pulling the first {...} block via regex.
    Raises ValueError/JSONDecodeError when no valid JSON is found.
    """
    cleaned = re.sub(r"^```(?:json)?\s*", "", text.strip())
    cleaned = re.sub(r"\s*```$", "", cleaned.strip())
    try:
        return json.loads(cleaned)
    except (json.JSONDecodeError, ValueError):
        pass
    # Extract first {...} block from raw text (handles prose wrappers)
    m = re.search(r"\{.*\}", text, re.DOTALL)
    if m:
        return json.loads(m.group())
    raise ValueError(f"No JSON object found in response (first 200 chars): {text[:200]!r}")


# ─────────────────────────────────────────────────────────────────────────────
# Cache
# ─────────────────────────────────────────────────────────────────────────────

def load_cache(cache_key: str):
    path = CACHE_DIR / f"{safe_filename(cache_key)}.json"
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return None
    return None


def save_cache(cache_key: str, data: dict) -> None:
    CACHE_DIR.mkdir(exist_ok=True)
    (CACHE_DIR / f"{safe_filename(cache_key)}.json").write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def _delete_cache(cache_key: str) -> None:
    p = CACHE_DIR / f"{safe_filename(cache_key)}.json"
    p.unlink(missing_ok=True)


def get_cache_count() -> int:
    return len(list(CACHE_DIR.glob("*.json"))) if CACHE_DIR.exists() else 0


def list_cache_files() -> list:
    return sorted(CACHE_DIR.glob("*.json")) if CACHE_DIR.exists() else []


# ─────────────────────────────────────────────────────────────────────────────
# Jina AI  ← page / search fetching
# ─────────────────────────────────────────────────────────────────────────────

_JINA_HEADERS = {
    "Accept": "text/plain",
    "X-Return-Format": "text",
    "x-respond-with": "text",
}
_JINA_CHAR_LIMIT = 6_000   # cap Jina content to save tokens
_JINA_ABOUT_SLUGS  = ("/about-us", "/about")
_JINA_MIN_CONTENT  = 500   # fewer chars than this → treat as failed Jina fetch
_JINA_RETRY_WAITS  = (30, 60, 120)  # backoff seconds on 429

_BOT_DETECTION_KEYWORDS = (
    "automatically identified", "security system", "bot", "captcha",
    "cloudflare", "datadome",
)


class JinaRateLimitRetry(Exception):
    """Raised to signal the UI that we're waiting on a 429 before retrying."""
    def __init__(self, wait: int, company: str):
        self.wait    = wait
        self.company = company
        super().__init__(f"Rate limited — waiting {wait}s for {company}")


def _jina_get(url: str, company_hint: str = "") -> str:
    """
    GET a single URL via Jina Reader.
    On 429: waits 30 s → 60 s → 120 s (3 retries).
    Raises JinaRateLimitRetry to let the UI display the wait message.
    Raises requests.HTTPError on non-429 failures.
    Returns up to _JINA_CHAR_LIMIT characters.
    """
    for attempt, wait in enumerate(_JINA_RETRY_WAITS):
        resp = requests.get(
            f"{JINA_READER_URL}{url}",
            headers=_JINA_HEADERS,
            timeout=30,
        )
        if resp.status_code != 429:
            break
        raise JinaRateLimitRetry(wait, company_hint)
    resp.raise_for_status()
    return resp.text[:_JINA_CHAR_LIMIT]


def _jina_get_with_retry(url: str, company_hint: str = "") -> str:
    """Wrap _jina_get to actually sleep and retry when JinaRateLimitRetry is raised."""
    for attempt, wait in enumerate(_JINA_RETRY_WAITS):
        try:
            return _jina_get(url, company_hint)
        except JinaRateLimitRetry as exc:
            if is_cli_mode():
                print(f"[enricher] Jina rate limit - waiting {exc.wait}s for {company_hint or url}", flush=True)
            else:
                _st, _ = get_streamlit()
                _st.session_state["_jina_retry_count"] = (
                    _st.session_state.get("_jina_retry_count", 0) + 1
                )
                _st.session_state["_last_retry_msg"] = (
                    f"⏳ Rate limit — waiting {exc.wait}s for {company_hint or url}…"
                )
            time.sleep(wait)
    # Final attempt — let HTTPError propagate
    return _jina_get(url, company_hint)


def fetch_via_jina_reader(url: str, company_hint: str = "") -> str:
    """
    Fetch the homepage AND about-us page via Jina Reader; return whichever has
    more content.  Raises requests.HTTPError when content is below
    _JINA_MIN_CONTENT so the caller can fall through to the next tier.
    """
    base = url.rstrip("/")
    best = ""

    # Try homepage
    try:
        text = _jina_get_with_retry(url, company_hint)
        if len(text) > len(best):
            best = text
    except requests.HTTPError as e:
        code = e.response.status_code if e.response is not None else 0
        if code not in (403, 503):
            raise

    # Try about-us slugs
    for slug in _JINA_ABOUT_SLUGS:
        try:
            text = _jina_get_with_retry(base + slug, company_hint)
            if len(text) > len(best):
                best = text
        except requests.HTTPError as e:
            code = e.response.status_code if e.response is not None else 0
            if code not in (403, 503, 404):
                raise

    if len(best) < _JINA_MIN_CONTENT:
        # Simulate a 404 so the caller falls through to the search tier
        raise requests.HTTPError(
            f"Jina returned only {len(best)} chars (< {_JINA_MIN_CONTENT})",
            response=None,
        )
    return best


def fetch_via_jina_search(query: str) -> str:
    for attempt, wait in enumerate(_JINA_RETRY_WAITS):
        resp = requests.get(
            f"{JINA_SEARCH_URL}{quote(query)}",
            headers=_JINA_HEADERS,
            timeout=30,
        )
        if resp.status_code != 429:
            break
        time.sleep(wait)
    resp.raise_for_status()
    return resp.text[:_JINA_CHAR_LIMIT]


# ─────────────────────────────────────────────────────────────────────────────
# Extreme Light Mode — page fetching and keyword extraction
# ─────────────────────────────────────────────────────────────────────────────

def _elm_fetch_pages(base_url: str, domain: str) -> tuple[dict, list, list]:
    """
    Fetch _ELM_SLUGS pages via requests + BeautifulSoup.
    Returns (pages_dict, fetched_slugs, failed_slugs).
    Caches by domain so repeated runs skip HTTP calls.
    """
    ck = f"elm_{domain}"
    cached = load_cache(ck)
    if cached and "pages" in cached:
        return (
            cached["pages"],
            cached.get("fetched", []),
            cached.get("failed", []),
        )

    base    = base_url.rstrip("/")
    headers = {"User-Agent": _ELM_UA}
    pages: dict   = {}
    fetched: list = []
    failed: list  = []

    for slug in _ELM_SLUGS:
        target = base if slug == "" else f"{base}{slug}"
        label  = slug if slug else "/"
        try:
            resp = requests.get(target, headers=headers, timeout=12, allow_redirects=True)
            if resp.status_code == 200:
                soup = BeautifulSoup(resp.text, "html.parser")
                for tag in soup(["script", "style", "nav", "footer", "header"]):
                    tag.decompose()
                text = soup.get_text(separator=" ", strip=True)
                pages[label] = text[:8_000]
                fetched.append(label)
            else:
                failed.append(f"{label}({resp.status_code})")
        except Exception as exc:
            failed.append(f"{label}(err:{type(exc).__name__})")

    save_cache(ck, {"pages": pages, "fetched": fetched, "failed": failed})
    return pages, fetched, failed


def _elm_count(text: str, keywords: list) -> int:
    tl = text.lower()
    return sum(1 for kw in keywords if kw in tl)


def _elm_find(text: str, keywords: list) -> list:
    tl = text.lower()
    return [kw for kw in keywords if kw in tl]


def _elm_score(count: int, per_point: float = 2.0) -> int:
    """Convert raw keyword count to 0–10 score."""
    return min(round(count / per_point * 10), 10)


def _elm_extract_signals(pages: dict, fetched: list, failed: list) -> dict:
    all_text = " ".join(pages.values())

    kw_intl     = _elm_count(all_text, _ELM_KW_INTERNATIONAL)
    langs       = _elm_find(all_text, _ELM_KW_LANGUAGES)
    kw_hiring   = _elm_count(all_text, _ELM_KW_HIRING)
    kw_growth   = _elm_count(all_text, _ELM_KW_GROWTH)
    kw_training = _elm_count(all_text, _ELM_KW_TRAINING)
    kw_offices  = _elm_count(all_text, _ELM_KW_OFFICES)
    kw_tech     = _elm_count(all_text, _ELM_KW_TECHNOLOGY)
    countries   = _elm_find(all_text, _ELM_COUNTRIES)

    s_intl     = min(_elm_score(kw_intl, 1.5) + min(len(countries), 4), 10)
    s_hiring   = _elm_score(kw_hiring,   1.5)
    s_growth   = _elm_score(kw_growth,   1.5)
    s_training = _elm_score(kw_training, 1.5)
    s_tech     = _elm_score(kw_tech,     2.0)
    s_overall  = round(
        s_intl     * 0.30
        + s_hiring   * 0.20
        + s_training * 0.30
        + s_growth   * 0.10
        + s_tech     * 0.10,
        1,
    )

    n_fetched = len(fetched)
    n_failed  = len(failed)
    if n_fetched == 0:
        fetch_status = "failed"
    elif n_failed > n_fetched:
        fetch_status = "partial"
    else:
        fetch_status = "ok"

    return {
        # Status
        "elm_fetch_status":  fetch_status,
        "elm_pages_fetched": ", ".join(fetched),
        "elm_pages_failed":  ", ".join(failed),
        "elm_total_chars":   sum(len(v) for v in pages.values()),
        "elm_error":         "",
        # Keywords
        "elm_kw_international":  kw_intl,
        "elm_kw_languages":      ", ".join(sorted(set(langs))),
        "elm_kw_hiring":         kw_hiring,
        "elm_kw_growth":         kw_growth,
        "elm_kw_training":       kw_training,
        "elm_kw_offices":        kw_offices,
        "elm_kw_technology":     kw_tech,
        "elm_kw_countries_found": ", ".join(sorted(set(countries))),
        # Scores
        "elm_score_international": s_intl,
        "elm_score_hiring":        s_hiring,
        "elm_score_growth":        s_growth,
        "elm_score_training":      s_training,
        "elm_score_technology":    s_tech,
        "elm_score_overall_icp":   s_overall,
    }


def enrich_one_row_light(company_name: str, raw_url: str) -> tuple:
    """
    Extreme Light Mode row enrichment — no Claude API, no tokens.
    Returns (elm_fields_dict, debug_record_dict).
    """
    empty = {f: "" for f in ELM_ALL_FIELDS}
    url   = normalize_url(raw_url) if raw_url else ""

    if not url:
        row = {**empty,
               "elm_fetch_status": "failed",
               "elm_error": "No URL provided"}
        return row, {"company": company_name, "url": raw_url, "status": "no_url"}

    domain = clean_domain(url)
    try:
        pages, fetched, failed = _elm_fetch_pages(url, domain)
    except Exception as exc:
        row = {**empty,
               "elm_fetch_status": "failed",
               "elm_error": str(exc)[:200]}
        return row, {"company": company_name, "url": url, "status": "fetch_error", "error": str(exc)}

    signals = _elm_extract_signals(pages, fetched, failed)
    dbg = {
        "company":       company_name,
        "url":           url,
        "domain":        domain,
        "status":        signals["elm_fetch_status"],
        "pages_fetched": signals["elm_pages_fetched"],
        "total_chars":   signals["elm_total_chars"],
    }
    return signals, dbg


# ─────────────────────────────────────────────────────────────────────────────
# Step 1 — Basic extraction via Claude (Jina page → Claude)
# ─────────────────────────────────────────────────────────────────────────────

def _claude_extract(webpage_text: str, api_key: str,
                    model_id: str = MODEL_STEP1) -> tuple:
    """
    Send page text to Claude for structured extraction.
    Returns (raw_fields_dict, input_tokens, output_tokens).
    """
    client    = anthropic.Anthropic(api_key=api_key)
    truncated = webpage_text[:_JINA_CHAR_LIMIT]
    msg = client.messages.create(
        model=model_id,
        max_tokens=1024,
        messages=[{"role": "user", "content": f"{_STEP1_PROMPT}\n\n{truncated}"}],
    )
    return (
        _parse_json_response(msg.content[0].text),
        msg.usage.input_tokens,
        msg.usage.output_tokens,
    )


def _map_step1_fields(raw: dict, source_url: str) -> dict:
    domain  = clean_domain(source_url) if source_url else ""
    # Prefer the domain Claude extracted over the source URL when available
    if not domain:
        domain = clean_domain(str(raw.get("domain") or ""))
    country = str(raw.get("country")   or "").strip()
    city    = str(raw.get("city")      or "").strip()

    def s(key):
        return str(raw.get(key) or "").strip()

    return {
        "lusha_company_name":         s("company_name"),
        "lusha_domain":               domain,
        "lusha_description":          s("description"),
        "lusha_founded_year":         s("founded_year"),
        "lusha_employee_range":       s("employee_range"),
        "lusha_revenue":              s("revenue_range"),
        "lusha_industry":             s("main_industry"),
        "lusha_sub_industry":         s("sub_industry"),
        "lusha_company_type":         s("company_type"),
        "lusha_country":              country,
        "lusha_city":                 city,
        "lusha_continent":            s("continent"),
        "lusha_linkedin_url":         s("linkedin_url"),
        "lusha_specialties":          s("specialties"),
        "lusha_technologies":         s("technologies"),
        "lusha_total_funding_amount": s("total_funding_amount"),
        "lusha_total_funding_rounds": s("total_funding_rounds"),
        "lusha_last_round_type":      s("last_round_type"),
        "lusha_last_round_amount":    s("last_round_amount"),
        "lusha_last_round_date":      s("last_round_date"),
        "lusha_ipo_status":           s("ipo_status"),
    }


def _step1_has_data(fields: dict) -> bool:
    return any(fields.get(f, "") for f in _STEP1_DATA_FIELDS)


_STEP1_FALLBACK_PROMPT_TMPL = (
    "Find company information about __COMPANY_NAME__ (__URL__). "
    "Extract: industry, employee count, founding year, headquarters location, "
    "description, international presence, specialties. "
    "Return ONLY JSON with fields: company_name, description, main_industry, "
    "sub_industry, employee_range, revenue_range, founded_year, company_type, "
    "country, city, linkedin_url, specialties, continent, technologies, "
    "total_funding_amount, total_funding_rounds, last_round_type, last_round_amount, "
    "last_round_date, ipo_status. Use empty string for any field not found."
)


_PW_DBG_SKIP = {"playwright_attempted": False, "playwright_result": "skipped"}


def run_step1(
    url: str,
    company_name: str,
    api_key: str,
    delay: float,
    use_playwright: bool = True,
    model_step1: str = MODEL_STEP1,
) -> tuple:
    """
    Three-tier Step 1 enrichment:
      Tier 1a — Jina direct scrape       → status 'enriched_jina'
      Tier 1b — human Playwright scrape  → status 'enriched_playwright'
      Tier 2  — Claude web_search        → status 'enriched_search'
    Returns (step1_fields, raw_json, in_tok, out_tok, status, error_msg, pw_debug).
    pw_debug: {"playwright_attempted": bool, "playwright_result": str}
    """
    source_url = normalize_url(url) if url else ""
    jina_err   = ""
    total_in   = total_out = 0

    # ── Tier 1a: Jina direct scrape ───────────────────────────────────────────
    if source_url:
        ck = f"step1_url_{source_url}"
        cached = load_cache(ck)
        if cached is not None:
            fields = _map_step1_fields(cached.get("claude_data", {}), source_url)
            if _step1_has_data(fields):
                return (fields, cached,
                        int(cached.get("tokens_in",  0) or 0),
                        int(cached.get("tokens_out", 0) or 0),
                        "enriched_jina", "", _PW_DBG_SKIP)
            _delete_cache(ck)

        try:
            time.sleep(delay)
            text = fetch_via_jina_reader(source_url, company_hint=company_name)
            raw_fields, in_t, out_t = _claude_extract(text, api_key, model_id=model_step1)
            total_in  += in_t
            total_out += out_t
            payload = {"claude_data": raw_fields, "tokens_in": in_t, "tokens_out": out_t}
            save_cache(ck, payload)
            fields = _map_step1_fields(raw_fields, source_url)
            if _step1_has_data(fields):
                return (fields, payload, total_in, total_out, "enriched_jina", "", _PW_DBG_SKIP)
            jina_err = "Jina page fetched but Claude found no usable data"
        except requests.HTTPError as e:
            code = e.response.status_code if e.response is not None else 0
            jina_err = f"Jina HTTP {code}: {str(e)[:120]}"
        except (json.JSONDecodeError, ValueError) as e:
            jina_err = f"Jina parse error: {e}"
        except anthropic.APIError as e:
            return ({}, {}, total_in, total_out, "api_error", f"Claude API: {e}", _PW_DBG_SKIP)
        except Exception as e:
            jina_err = str(e)

    # ── Tier 1b: Human Playwright scrape ──────────────────────────────────────
    _pw_attempted = False
    _pw_result    = "skipped"
    if source_url and jina_err and use_playwright and _PLAYWRIGHT_AVAILABLE:
        _pw_attempted = True
        ck_pw  = f"step1_playwright_{source_url}"
        cached_pw = load_cache(ck_pw)
        if cached_pw is not None:
            fields = _map_step1_fields(cached_pw.get("claude_data", {}), source_url)
            if _step1_has_data(fields):
                _pw_result = "success"
                return (fields, cached_pw,
                        int(cached_pw.get("tokens_in",  0) or 0),
                        int(cached_pw.get("tokens_out", 0) or 0),
                        "enriched_playwright", "",
                        {"playwright_attempted": True, "playwright_result": "success"})
            _delete_cache(ck_pw)

        try:
            pw_res = scrape_with_human_behaviour(source_url, max_chars=_JINA_CHAR_LIMIT)
            if pw_res.get("success") and len(pw_res.get("text", "")) >= _JINA_MIN_CONTENT:
                pw_text  = pw_res["text"]
                pw_lower = pw_text[:2000].lower()
                if any(kw in pw_lower for kw in _BOT_DETECTION_KEYWORDS):
                    _pw_result = "blocked"
                    jina_err  += " | playwright: bot-detected"
                else:
                    raw_fields, in_t, out_t = _claude_extract(pw_text, api_key, model_id=model_step1)
                    total_in  += in_t
                    total_out += out_t
                    payload = {"claude_data": raw_fields, "tokens_in": in_t, "tokens_out": out_t}
                    save_cache(ck_pw, payload)
                    fields = _map_step1_fields(raw_fields, source_url)
                    if _step1_has_data(fields):
                        _pw_result = "success"
                        return (fields, payload, total_in, total_out, "enriched_playwright", "",
                                {"playwright_attempted": True, "playwright_result": "success"})
                    _pw_result = "failed"
            else:
                _pw_result = "failed"
        except anthropic.APIError as e:
            return ({}, {}, total_in, total_out, "api_error", f"Claude API: {e}",
                    {"playwright_attempted": True, "playwright_result": "failed"})
        except Exception as e:
            _pw_result = "failed"
            jina_err  += f" | playwright error: {str(e)[:80]}"

    _pw_dbg = {"playwright_attempted": _pw_attempted, "playwright_result": _pw_result}

    # ── Tier 2: Claude web_search fallback ────────────────────────────────────
    target = source_url or company_name
    if not target:
        return ({}, {}, total_in, total_out, "no_data", "No URL or company name provided", _pw_dbg)

    ck = f"step1_fallback_{target}"
    cached = load_cache(ck)
    if cached is not None:
        fields = _map_step1_fields(cached.get("claude_data", {}), url)
        if _step1_has_data(fields):
            return (fields, cached,
                    int(cached.get("tokens_in",  0) or 0),
                    int(cached.get("tokens_out", 0) or 0),
                    "enriched_search", "", _pw_dbg)
        _delete_cache(ck)

    try:
        prompt = (
            _STEP1_FALLBACK_PROMPT_TMPL
            .replace("__COMPANY_NAME__", company_name or target)
            .replace("__URL__", target)
        )
        raw_text, in_t, out_t = _claude_web_search_loop(prompt, api_key, model_id=model_step1)
        total_in  += in_t
        total_out += out_t
        raw_fields = _parse_json_response(raw_text)
        payload    = {"claude_data": raw_fields, "tokens_in": in_t, "tokens_out": out_t}
        save_cache(ck, payload)
        fields = _map_step1_fields(raw_fields, url)
        if _step1_has_data(fields):
            return (fields, payload, total_in, total_out, "enriched_search", "", _pw_dbg)
        return ({}, {}, total_in, total_out, "no_data",
                f"Google search returned no usable data. Jina: {jina_err}", _pw_dbg)
    except (json.JSONDecodeError, ValueError) as e:
        return ({}, {}, total_in, total_out, "no_data",
                f"Search parse error: {e}. Jina: {jina_err}", _pw_dbg)
    except anthropic.APIError as e:
        return ({}, {}, total_in, total_out, "api_error", f"Claude API: {e}", _pw_dbg)
    except Exception as e:
        return ({}, {}, total_in, total_out, "no_data",
                f"Search error: {e}. Jina: {jina_err}", _pw_dbg)


# ─────────────────────────────────────────────────────────────────────────────
# Step 2 — ICP signals via Claude web_search
# ─────────────────────────────────────────────────────────────────────────────

_ICP_EMPTY = {f: "" for f in ICP_FIELDS}


def _assert_no_web_search_tool(tools=None) -> None:
    """Hard safety check — raise immediately if web_search_20250305 is present."""
    if not tools:
        return
    for t in tools:
        if isinstance(t, dict) and t.get("type") == "web_search_20250305":
            raise RuntimeError(
                "Blocked: Anthropic web_search is disabled. Use Serper only."
            )


def _claude_web_search_loop(prompt: str, api_key: str, model_id: str = None) -> tuple:
    """
    DISABLED — Anthropic web_search_20250305 is blocked in this app.
    All Step 2 web search is performed via Serper Google Search.
    """
    raise RuntimeError(
        "Blocked: _claude_web_search_loop is disabled. "
        "Anthropic web_search_20250305 must not be called. Use Serper only."
    )
    # Dead code below — kept for reference only, never executed.
    if model_id is None:
        raise ValueError("model_id must be provided to _claude_web_search_loop")

    client = anthropic.Anthropic(api_key=api_key)

    for attempt in range(3):
        try:
            resp = client.messages.create(
                model=model_id,
                max_tokens=2048,
                tools=[WEB_SEARCH_TOOL],
                messages=[{"role": "user", "content": prompt}],
            )
            text = "".join(
                getattr(b, "text", "")
                for b in resp.content
                if getattr(b, "type", "") == "text"
            ).strip()
            return text, resp.usage.input_tokens, resp.usage.output_tokens

        except anthropic.RateLimitError as e:
            wait = 30
            if hasattr(e, "response") and e.response is not None:
                wait = int(e.response.headers.get("retry-after", 30))
            time.sleep(wait)

        except anthropic.APIStatusError as e:
            if e.status_code == 529:
                time.sleep(60)
            else:
                raise

    return "", 0, 0


def _claude_web_search_full(prompt: str, api_key: str, model_id: str) -> tuple:
    """
    Like _claude_web_search_loop but also returns the raw response object.
    Returns (text, in_t, out_t, resp_or_None).
    """
    client = anthropic.Anthropic(api_key=api_key)
    for _ in range(3):
        try:
            resp = client.messages.create(
                model=model_id,
                max_tokens=2048,
                tools=[WEB_SEARCH_TOOL],
                messages=[{"role": "user", "content": prompt}],
            )
            text = "".join(
                getattr(b, "text", "")
                for b in resp.content
                if getattr(b, "type", "") == "text"
            ).strip()
            return text, resp.usage.input_tokens, resp.usage.output_tokens, resp
        except anthropic.RateLimitError as e:
            wait = 30
            if hasattr(e, "response") and e.response is not None:
                wait = int(e.response.headers.get("retry-after", 30))
            time.sleep(wait)
        except anthropic.APIStatusError as e:
            if e.status_code == 529:
                time.sleep(60)
            else:
                raise
    return "", 0, 0, None


# ─────────────────────────────────────────────────────────────────────────────
# Serper Google Search helpers (Step 2 alternative provider)
# ─────────────────────────────────────────────────────────────────────────────

def _build_serper_queries(company_name: str, target: str) -> list:
    """Return 5 feature-driven queries for ICP signal extraction.

    Each query targets a distinct buying-signal dimension so Claude receives
    evidence that is grouped by intent rather than by brand co-mention.
    """
    name   = company_name or clean_domain(target) or target
    domain = clean_domain(target) if target else ""
    site_q = f'site:{domain} OR ' if domain else ""
    return [
        # Q1 — General company context (anchor to official site when possible)
        f'{site_q}"{name}" about company overview headquarters',
        # Q2 — International footprint / HQ structure
        f'"{name}" headquarters OR offices OR countries OR "international operations" OR "global presence" OR "regional HQ"',
        # Q3 — L&D / employee training
        f'"{name}" "learning and development" OR training OR academy OR onboarding OR "talent development" OR L&D',
        # Q4 — Language / global-team signals
        f'"{name}" English OR "language training" OR "global teams" OR multilingual OR "language program" OR intercultural',
        # Q5 — Competitor / online learning tool co-mention
        (
            f'"{name}" Preply OR Learnlight OR Speexx OR goFLUENT OR Learnship OR Voxy'
            f' OR "online learning" OR LMS OR Berlitz OR Talaera OR Busuu OR Duolingo'
        ),
    ]

# Human-readable label for each query position (1-to-1 with _build_serper_queries)
_SERPER_QUERY_LABELS = [
    "General company context",
    "International footprint / HQ",
    "L&D / employee training",
    "Language / global teams",
    "Competitor / online learning signal",
]


def _call_serper(query: str, serper_key: str, timeout: int = 15):
    """
    POST one query to the Serper API.
    Returns (organic_results, http_status, raw_json_or_None, error_str_or_None).
    organic_results is a list of dicts; empty on error.
    Never logs or returns the serper_key.
    """
    raw_json    = None
    http_status = 0
    try:
        resp = requests.post(
            SERPER_SEARCH_URL,
            headers={"X-API-KEY": serper_key, "Content-Type": "application/json"},
            json={"q": query, "gl": "us", "hl": "en", "num": 5},
            timeout=timeout,
        )
        http_status = resp.status_code
        resp.raise_for_status()
        raw_json = resp.json()
    except requests.Timeout:
        return [], 0, None, "Serper API timed out"
    except requests.HTTPError as e:
        code = e.response.status_code if e.response is not None else 0
        http_status = code
        try:
            raw_json = e.response.json() if e.response is not None else None
        except Exception:
            pass
        if code == 403:
            return [], code, raw_json, "Serper API key rejected (403)"
        if code == 429:
            return [], code, raw_json, "Serper quota exceeded (429)"
        return [], code, raw_json, f"Serper HTTP {code}: {e}"
    except (json.JSONDecodeError, ValueError) as e:
        return [], http_status, None, f"Serper returned invalid JSON: {e}"
    except Exception as e:
        return [], 0, None, f"Serper error: {e}"

    results = [
        {
            "title":     item.get("title", ""),
            "link":      item.get("link", ""),
            "snippet":   item.get("snippet", ""),
            "position":  item.get("position", ""),
            "date":      item.get("date", ""),
            "sitelinks": item.get("sitelinks", []),
        }
        for item in (raw_json or {}).get("organic", [])
    ]
    return results, http_status, raw_json, None


_SERPER_CAREER_SIGNALS = frozenset([
    "career", "/jobs", "glassdoor", "indeed", ".jobs", "workable",
    "greenhouse.io", "lever.co", "smartrecruiters", "bamboohr",
    "jobvite", "icims", "taleo", "recruiting",
])
_SERPER_NEWS_SIGNALS = frozenset([
    "reuters", "bloomberg", "wsj.com", "ft.com", "forbes",
    "techcrunch", "businesswire", "prnewswire", "globenewswire",
    "marketwatch", "apnews.com", "cnbc", "wired", "theguardian",
    "economist", "fortune", "venturebeat",
])
_SERPER_THIRD_PARTY_SIGNALS = frozenset([
    "linkedin.com", "crunchbase", "zoominfo", "pitchbook",
    "dnb.com", "hoovers", "owler", "manta", "trustpilot",
    "clutch.co", "g2.com", "capterra", "wellfound", "angellist",
])


def _classify_serper_source(link: str, title: str) -> str:
    """Return a short source-type label for a Serper result URL."""
    u = link.lower()
    if any(s in u for s in _SERPER_CAREER_SIGNALS):
        return "careers"
    if any(s in u for s in _SERPER_NEWS_SIGNALS):
        return "news"
    if any(s in u for s in _SERPER_THIRD_PARTY_SIGNALS):
        return "third_party"
    return "company_site"


def build_raw_google_evidence(
    query_groups: list,
    query_strings: list | None = None,
) -> list[dict]:
    """Flatten (label, hits) query_groups into per-result evidence records.

    Each record contains: query_type, query, rank, title, url, display_url,
    source_domain, snippet, date, sitelinks.
    """
    records: list[dict] = []
    for i, (label, hits) in enumerate(query_groups):
        q_str = query_strings[i] if query_strings and i < len(query_strings) else ""
        for rank, hit in enumerate(hits, 1):
            link = str(hit.get("link", "") or "")
            domain = re.sub(r"^https?://(www\.)?", "", link).split("/")[0].split("?")[0]
            records.append({
                "query_type":    label,
                "query":         q_str,
                "rank":          rank,
                "title":         str(hit.get("title", "") or ""),
                "url":           link,
                "display_url":   str(hit.get("displayLink", "") or ""),
                "source_domain": domain,
                "snippet":       str(hit.get("snippet", "") or ""),
                "date":          str(hit.get("date", "") or ""),
                "sitelinks":     json.dumps(hit["sitelinks"]) if hit.get("sitelinks") else "",
            })
    return records


def _pack_raw_evidence_fields(
    raw_records: list[dict],
) -> dict:
    """Convert raw evidence records into the set of handoff fields for Excel/export.

    Splits oversized JSON across up to 3 part columns to avoid Excel cell limits.
    """
    _MAX_CELL = 32_000

    # Unique URLs in encounter order
    seen_urls: set[str] = set()
    unique_urls: list[str] = []
    for r in raw_records:
        u = r.get("url", "")
        if u and u not in seen_urls:
            seen_urls.add(u)
            unique_urls.append(u)

    # Human-readable combined text
    lines: list[str] = []
    for r in raw_records:
        lines.append(f"[{r['query_type']}] {r['title']}")
        if r["url"]:
            lines.append(r["url"])
        if r["snippet"]:
            lines.append(r["snippet"])
        lines.append("")
    combined = "\n".join(lines).strip()

    # JSON: try single cell first, then split over parts
    full_json = json.dumps(raw_records, ensure_ascii=False)
    if len(full_json) <= _MAX_CELL:
        json_parts = [full_json, "", ""]
        truncated = False
    else:
        # Split records into at most 3 chunks that fit within _MAX_CELL each
        chunks: list[list] = [[], [], []]
        ci = 0
        for rec in raw_records:
            rec_str = json.dumps([rec], ensure_ascii=False)
            # Try to add to current chunk; advance chunk index if it won't fit
            while ci < 3:
                candidate = json.dumps(chunks[ci] + [rec], ensure_ascii=False)
                if len(candidate) <= _MAX_CELL:
                    chunks[ci].append(rec)
                    break
                ci += 1
        truncated = any(rec not in chunks[0] + chunks[1] + chunks[2] for rec in raw_records)
        json_parts = [
            json.dumps(c, ensure_ascii=False) if c else ""
            for c in chunks
        ]

    # Per-snippet flat columns (first 10 results)
    snippet_fields: dict = {}
    for si, rec in enumerate(raw_records[:10], 1):
        pfx = f"google_snippet_{si:02d}"
        snippet_fields[f"{pfx}_query_type"]    = rec.get("query_type", "")
        snippet_fields[f"{pfx}_query"]         = rec.get("query", "")
        snippet_fields[f"{pfx}_rank"]          = rec.get("rank", "")
        snippet_fields[f"{pfx}_title"]         = rec.get("title", "")
        snippet_fields[f"{pfx}_url"]           = rec.get("url", "")
        snippet_fields[f"{pfx}_source_domain"] = rec.get("source_domain", "")
        snippet_fields[f"{pfx}_text"]          = rec.get("snippet", "")

    return {
        "raw_google_evidence_count":    len(raw_records),
        "raw_google_evidence_urls":     "\n".join(unique_urls),
        "raw_google_evidence_combined": combined[:_MAX_CELL],
        "raw_google_evidence_json":     json_parts[0],
        "raw_google_evidence_json_01":  json_parts[0],
        "raw_google_evidence_json_02":  json_parts[1],
        "raw_google_evidence_json_03":  json_parts[2],
        "raw_google_evidence_json_parts": sum(1 for p in json_parts if p),
        "raw_google_evidence_truncated": truncated,
        **snippet_fields,
    }


def _format_serper_results(query_groups: list) -> str:
    """Format per-query Serper results as compact annotated text for Claude.

    Args:
        query_groups: list of (query_label, hits) tuples where hits is a list
                      of result dicts returned by _call_serper.
    """
    if not query_groups or not any(hits for _, hits in query_groups):
        return "(No web search results were found.)"
    blocks: list[str] = []
    for label, hits in query_groups:
        if not hits:
            continue
        lines = [f"=== {label} ==="]
        for i, r in enumerate(hits, 1):
            title   = (r.get("title", "") or "(no title)").strip()
            link    = r.get("link", "")
            snippet = (r.get("snippet", "") or "").strip()[:180]
            stype   = _classify_serper_source(link, title)
            lines.append(f"[{i}] {title}")
            lines.append(f"    URL: {link}")
            lines.append(f"    source_type: {stype}")
            if snippet:
                lines.append(f"    snippet: {snippet}")
        blocks.append("\n".join(lines))
    return "\n\n".join(blocks) if blocks else "(No web search results were found.)"


def run_step2_serper(
    url: str,
    company_name: str,
    api_key: str,
    serper_key: str,
    delay: float,
    model_step2: str = MODEL_STEP2,
    _debug_callback=None,
    dry_run: bool = False,
) -> tuple:
    """
    Step 2 via Serper Google Search + Claude analysis (no web_search tool).
    Returns the same 8-tuple as run_step2.
    dry_run=True: generate queries and Claude prompt without calling any API.
    """
    def _dlog(msg: str) -> None:
        if _debug_callback:
            _debug_callback("status", msg=msg)

    target = normalize_url(url) if url else company_name
    if not target:
        return (_ICP_EMPTY.copy(), {}, 0, 0, "no_input", "No URL or company name", 0, 0)

    if not dry_run:
        ck = f"step2_serper_{target}"
        cached = load_cache(ck)
        if cached is not None:
            icp = cached.get("icp_data", {})
            if any(icp.get(f, "") for f in ICP_FIELDS[:3]):
                in_t  = int(cached.get("tokens_in", 0) or 0)
                out_t = int(cached.get("tokens_out", 0) or 0)
                _dlog(f"Using cached Serper Step 2 result for {company_name}")
                return (_extract_icp_fields(icp), cached, in_t, out_t, "cached", "", 0, 0)
            _delete_cache(ck)

    # ── Serper searches ───────────────────────────────────────────────────────
    queries = _build_serper_queries(company_name, target)
    _dlog(f"Generating Serper queries for {company_name}")

    # DRY RUN GUARD: do not call Serper or Anthropic in prompt preview mode
    if dry_run:
        _dry_placeholder = "\n".join(
            f"=== {lbl} ===\n[DRY RUN: results would appear here]"
            for lbl in _SERPER_QUERY_LABELS
        )
        _dry_instruction = (
            f"Now analyze this company based on the web search results provided below.\n\n"
            f"Company: {target}\n\n"
            f"Web search results (retrieved via Serper Google Search, grouped by signal type):\n"
            f"{_dry_placeholder}\n\n"
            "Evidence quality rules:\n"
            "- Only mark a buying signal as present when a result contains company-specific, "
            "contextual evidence — not just a keyword in a URL or a generic snippet.\n"
            "- A competitor signal requires the provider name to appear in a meaningful context.\n"
            "- Set lead_score to High only when two or more clearly distinct strong signals appear.\n"
            "- Base your analysis ONLY on the search results above. Do not invent evidence."
        )
        _dry_full_prompt = STEP2_STATIC_PREFIX + f"\n\n{_dry_instruction}"
        _dlog(f"DRY RUN: Serper queries that would be sent: {queries}")
        _dlog(f"DRY RUN: skipping Serper + Anthropic calls for {company_name}")
        if _debug_callback:
            _ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
            for _qi, _q in enumerate(queries, 1):
                try:
                    _dry_content = _format_serper_query_debug(
                        company_name, model_step2, _ts, _q, _qi, len(queries),
                        [], 0, None, None, dry_run=True,
                    )
                    _fpath = write_search_debug_file(
                        company_name, "serper_google_search", _dry_content, index=_qi,
                    )
                    _debug_callback(
                        "search_output",
                        company=company_name,
                        provider=STEP2_PROVIDER_SERPER,
                        query=_q,
                        result_count=0,
                        top_results=[],
                        debug_file=str(_fpath),
                        dry_run=True,
                    )
                except Exception:
                    pass
            _debug_callback(
                "prompt",
                company=company_name,
                model=model_step2,
                provider=STEP2_PROVIDER_SERPER,
                search_prompt=(
                    "DRY RUN — Serper queries that would be sent:\n"
                    + "\n".join(f"  • {q}" for q in queries)
                ),
                full_prompt=_dry_full_prompt,
                notes=[
                    "DRY RUN ACTIVE: no Anthropic or Serper API calls are being made.",
                    f"Serper queries that would be sent: {queries}",
                    f"Selected model: {model_step2}",
                ],
                dry_run=True,
                queries=queries,
            )
        return (_ICP_EMPTY.copy(), {}, 0, 0, "dry_run", "DRY RUN: no API call made", 0, 0)

    # ── Collect results per-query, preserving query-label grouping ───────────
    query_groups: list = []   # [(label, hits), ...]
    labels = _SERPER_QUERY_LABELS
    for qi, (q, label) in enumerate(zip(queries, labels), 1):
        _dlog(f"Serper query [{label}]: {q}")
        hits, http_status, raw_json, err_str = _call_serper(q, serper_key)
        if err_str:
            _dlog(f"Serper warning [{label}] — {err_str}")
            hits = []
        else:
            _dlog(f"Serper returned {len(hits)} results for [{label}]")
        query_groups.append((label, hits))
        if _debug_callback:
            _ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
            try:
                _content = _format_serper_query_debug(
                    company_name, model_step2, _ts, q, qi, len(queries),
                    hits, http_status, raw_json, err_str,
                )
                _fpath = write_search_debug_file(
                    company_name, "serper_google_search", _content, index=qi,
                )
                _debug_callback(
                    "search_output",
                    company=company_name,
                    provider=STEP2_PROVIDER_SERPER,
                    query=q,
                    result_count=len(hits),
                    top_results=hits[:3],
                    debug_file=str(_fpath),
                    dry_run=False,
                )
            except Exception:
                pass

    total_hits = sum(len(h) for _, h in query_groups)
    _dlog(f"Total Serper results for {company_name}: {total_hits} across {len(query_groups)} queries")

    # ── Build compact Serper evidence handoff for Opportunity Radar / caller brief ──
    _s_labels:   list[str] = []
    _s_urls:     list[str] = []
    _s_titles:   list[str] = []
    _s_snippets: list[str] = []
    for _lbl, _hits in query_groups:
        _s_labels.append(_lbl)
        for _h in _hits[:2]:  # top 2 results per query
            if _h.get("link"):
                _s_urls.append(str(_h["link"])[:200])
            if _h.get("title"):
                _s_titles.append(str(_h["title"])[:120])
            if _h.get("snippet"):
                _s_snippets.append(str(_h["snippet"])[:160])
    _serper_evidence_handoff: dict = {
        "serper_query_summary":   " | ".join(_s_labels),
        "serper_source_urls":     " | ".join(_s_urls[:8]),
        "serper_result_titles":   " | ".join(_s_titles[:8]),
        "serper_snippets":        " | ".join(_s_snippets[:6]),
        "raw_evidence_summary":   f"{total_hits} Serper results across {len(query_groups)} queries",
        "evidence_source_urls":   " | ".join(_s_urls[:8]),
    }

    # ── Raw Google evidence expansion (full detail for webapp handoff) ────────
    _raw_records = build_raw_google_evidence(query_groups, queries)
    _serper_evidence_handoff.update(_pack_raw_evidence_fields(_raw_records))

    # ── Build Claude prompt ────────────────────────────────────────────────────
    results_text = _format_serper_results(query_groups)
    search_instruction = (
        f"Now analyze this company based on the web search results provided below.\n\n"
        f"Company: {target}\n\n"
        f"Web search results (retrieved via Serper Google Search, grouped by signal type):\n"
        f"{results_text}\n\n"
        "Evidence quality rules:\n"
        "- Only mark a buying signal as present when a result contains company-specific, "
        "contextual evidence — not just a keyword in a URL or a generic snippet.\n"
        "- A competitor signal requires the provider name to appear in a meaningful context "
        "(HR case study, employee benefit page, vendor review) not just a search result title.\n"
        "- Set lead_score to High only when two or more clearly distinct strong signals appear.\n"
        "- Base your analysis ONLY on the search results above. Do not invent or infer evidence."
    )
    full_prompt = STEP2_STATIC_PREFIX + f"\n\n{search_instruction}"

    _dlog(f"Selected model: {model_step2}")

    if _debug_callback:
        _debug_callback(
            "prompt",
            company=company_name,
            model=model_step2,
            provider=STEP2_PROVIDER_SERPER,
            search_prompt=f"Queries: {queries}\n\nTotal results: {total_hits}",
            full_prompt=full_prompt,
            notes=[
                f"Serper queries used: {queries}",
                f"Total results retrieved: {total_hits} across {len(query_groups)} queries",
                f"Selected model: {model_step2}",
            ],
        )

    # ── Claude analysis (no web_search tool — we supply the context) ──────────
    _STRICT_SUFFIX = (
        "\n\nReply with ONLY a JSON object, no explanation, no markdown, no backticks."
    )
    client = anthropic.Anthropic(api_key=api_key)

    _dlog(f"Calling Claude for Serper analysis of {company_name}")
    try:
        time.sleep(delay)
        resp = client.messages.create(
            model=model_step2,
            max_tokens=2048,
            messages=[{"role": "user", "content": full_prompt}],
        )
        raw_text = "".join(
            getattr(b, "text", "") for b in resp.content
            if getattr(b, "type", "") == "text"
        ).strip()
        in_t  = resp.usage.input_tokens
        out_t = resp.usage.output_tokens
        _dlog(f"Received Claude response for {company_name}")

        # Parse (with one retry on failure)
        _icp_raw = None
        try:
            _dlog(f"Parsing Step 2 Serper response for {company_name}")
            _icp_raw = _parse_json_response(raw_text)
        except (json.JSONDecodeError, ValueError):
            _dlog(f"Parse failed — retrying with strict suffix for {company_name}")
            time.sleep(delay)
            resp2 = client.messages.create(
                model=model_step2,
                max_tokens=2048,
                messages=[{"role": "user", "content": full_prompt + _STRICT_SUFFIX}],
            )
            raw_text2 = "".join(
                getattr(b, "text", "") for b in resp2.content
                if getattr(b, "type", "") == "text"
            ).strip()
            in_t  += resp2.usage.input_tokens
            out_t += resp2.usage.output_tokens
            _dlog(f"Parsing Step 2 Serper retry response for {company_name}")
            _icp_raw = _parse_json_response(raw_text2)

        # Write Claude analysis post-call debug file (after parsing so it includes parsed JSON)
        if _debug_callback:
            _ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
            try:
                _post_content = _format_claude_post_debug(
                    company_name, model_step2, _ts, resp, raw_text,
                    parsed_json=_icp_raw,
                )
                _serper_resp_path = write_search_debug_file(
                    company_name, "serper_google_search_claude_response",
                    _post_content, index=len(queries) + 1,
                )
                _debug_callback(
                    "search_output",
                    company=company_name,
                    provider=STEP2_PROVIDER_SERPER,
                    query="[Claude analysis of Serper results]",
                    result_count=0,
                    top_results=[],
                    debug_file=str(_serper_resp_path),
                    dry_run=False,
                )
            except Exception:
                pass

        payload = {"icp_data": _icp_raw, "tokens_in": in_t, "tokens_out": out_t,
                   "serper_evidence": _serper_evidence_handoff}
        save_cache(ck, payload)
        _dlog(f"Finished Step 2 (Serper) for {company_name}")
        _icp_out = _extract_icp_fields(_icp_raw)
        _icp_out.update(_serper_evidence_handoff)
        return (_icp_out, payload, in_t, out_t, "ok", "", 0, 0)

    except (json.JSONDecodeError, ValueError) as e:
        _dlog(f"Step 2 Serper parse error for {company_name}: {e}")
        return (_ICP_EMPTY.copy(), {}, 0, 0, "parse_error", f"Serper parse error: {type(e).__name__}: {e}", 0, 0)
    except anthropic.APIError as e:
        _dlog(f"Step 2 Serper API error for {company_name}: {e}")
        return (_ICP_EMPTY.copy(), {}, 0, 0, "api_error", f"Claude API {type(e).__name__}: {e}", 0, 0)
    except Exception as e:
        _dlog(f"Step 2 Serper error for {company_name}: {e}")
        return (_ICP_EMPTY.copy(), {}, 0, 0, "api_error", f"{type(e).__name__}: {e}", 0, 0)


def run_step2(
    url: str,
    company_name: str,
    api_key: str,
    delay: float,
    model_step2: str = MODEL_STEP2,
    _debug_callback=None,
    search_provider: str = STEP2_PROVIDER_SERPER,
    serper_key: str = "",
    dry_run: bool = False,
) -> tuple:
    """
    Research ICP signals via Serper Google Search + Claude analysis.
    Serper is the ONLY permitted provider. Any attempt to use Claude web_search
    (STEP2_PROVIDER_CLAUDE) raises RuntimeError immediately.

    Returns (icp_fields_dict, raw_json, in_tok, out_tok, status, error_msg,
             cache_creation_tokens, cache_read_tokens).
    """
    if search_provider == STEP2_PROVIDER_CLAUDE:
        raise RuntimeError(
            f"Blocked: '{STEP2_PROVIDER_CLAUDE}' is disabled. "
            "Anthropic web_search must not be called. Use Serper only."
        )

    # Serper path — the only valid route
    if not dry_run and not serper_key:
        return (
            _ICP_EMPTY.copy(), {}, 0, 0, "api_error",
            "SERPER_API_KEY is missing from .streamlit/secrets.toml", 0, 0,
        )
    return run_step2_serper(
        url, company_name, api_key, serper_key, delay,
        model_step2=model_step2, _debug_callback=_debug_callback,
        dry_run=dry_run,
    )

    def _dlog(msg: str) -> None:
        if _debug_callback:
            _debug_callback("status", msg=msg)

    target = normalize_url(url) if url else company_name
    if not target:
        return (_ICP_EMPTY.copy(), {}, 0, 0, "no_input", "No URL or company name", 0, 0)

    ck = f"step2_claude_{target}"
    cached = load_cache(ck)
    if cached is not None:
        icp = cached.get("icp_data", {})
        if any(icp.get(f, "") for f in ICP_FIELDS[:3]):  # basic sanity check
            in_t  = int(cached.get("tokens_in", 0) or 0)
            out_t = int(cached.get("tokens_out", 0) or 0)
            _dlog(f"Using cached Step 2 result for {company_name}")
            return (_extract_icp_fields(icp), cached, in_t, out_t, "cached", "", 0, 0)
        _delete_cache(ck)

    _STRICT_SUFFIX = (
        "\n\nReply with ONLY a JSON object, no explanation, no markdown, no backticks."
    )
    search_prompt = f"Now research this company: {target}"
    full_prompt   = STEP2_STATIC_PREFIX + f"\n\n{search_prompt}"

    _dlog(f"Generating Step 2 prompt for {company_name}")
    _dlog(f"Selected model: {model_step2}")

    _ts_pre    = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    _pre_fpath = None   # path of the pre-call file; post-call section is appended here

    if _debug_callback:
        _debug_callback(
            "prompt",
            company=company_name,
            model=model_step2,
            provider=STEP2_PROVIDER_CLAUDE,
            search_prompt=search_prompt,
            full_prompt=full_prompt,
            notes=[
                "DRY RUN ACTIVE: no Anthropic API calls are being made."
                if dry_run else f"Generating Step 2 prompt for {company_name}",
                f"Selected model: {model_step2}",
            ],
            dry_run=dry_run,
            queries=[],
        )
        try:
            _pre_content = _format_claude_pre_debug(
                company_name, model_step2, _ts_pre, full_prompt, dry_run=dry_run,
            )
            _pre_fpath = write_search_debug_file(
                company_name, "claude_web_search", _pre_content, index=1,
            )
            _debug_callback(
                "search_output",
                company=company_name,
                provider=STEP2_PROVIDER_CLAUDE,
                query=search_prompt,
                result_count=0,
                top_results=[],
                debug_file=str(_pre_fpath),
                dry_run=dry_run,
            )
        except Exception:
            pass

    # DRY RUN GUARD: do not call Anthropic in prompt preview mode
    if dry_run:
        _dlog(f"DRY RUN: skipping Anthropic call for {company_name}")
        if _pre_fpath:
            try:
                append_to_debug_file(
                    _pre_fpath,
                    _format_claude_post_debug(
                        company_name, model_step2,
                        datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
                        None, "", dry_run_skip=True,
                    ),
                )
            except Exception:
                pass
        return (_ICP_EMPTY.copy(), {}, 0, 0, "dry_run", "DRY RUN: no API call made", 0, 0)

    try:
        _dlog(f"Calling Claude web search for {company_name}")
        time.sleep(delay)
        raw_text, in_t, out_t, _resp = _claude_web_search_full(
            full_prompt, api_key, model_id=model_step2,
        )
        _dlog(f"Received Claude response for {company_name}")

        # Parse (with one retry on failure)
        _icp_raw = None
        try:
            _dlog(f"Parsing Step 2 response for {company_name}")
            _icp_raw = _parse_json_response(raw_text)
        except (json.JSONDecodeError, ValueError):
            _dlog(f"Parse failed — retrying with strict suffix for {company_name}")
            time.sleep(delay)
            raw_text2, in_t2, out_t2, _ = _claude_web_search_full(
                full_prompt + _STRICT_SUFFIX, api_key, model_id=model_step2,
            )
            in_t  += in_t2
            out_t += out_t2
            _dlog(f"Parsing Step 2 retry response for {company_name}")
            _icp_raw = _parse_json_response(raw_text2)  # raises if still bad

        # Append post-call section (with parsed JSON) to the pre-call file
        if _pre_fpath:
            try:
                append_to_debug_file(
                    _pre_fpath,
                    _format_claude_post_debug(
                        company_name, model_step2,
                        datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
                        _resp, raw_text, parsed_json=_icp_raw,
                    ),
                )
            except Exception:
                pass

        payload = {"icp_data": _icp_raw, "tokens_in": in_t, "tokens_out": out_t}
        save_cache(ck, payload)
        _dlog(f"Finished Step 2 for {company_name}")
        return (_extract_icp_fields(_icp_raw), payload, in_t, out_t, "ok", "", 0, 0)

    except (json.JSONDecodeError, ValueError) as e:
        _dlog(f"Step 2 parse error for {company_name}: {e}")
        if _pre_fpath:
            try:
                append_to_debug_file(
                    _pre_fpath,
                    _format_claude_post_debug(
                        company_name, model_step2,
                        datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
                        None, "", error_str=f"Parse error: {type(e).__name__}: {e}",
                    ),
                )
            except Exception:
                pass
        return (_ICP_EMPTY.copy(), {}, 0, 0, "parse_error", f"Claude parse error: {type(e).__name__}: {e}", 0, 0)

    except anthropic.APIError as e:
        _dlog(f"Step 2 API error for {company_name}: {e}")
        if _pre_fpath:
            try:
                append_to_debug_file(
                    _pre_fpath,
                    _format_claude_post_debug(
                        company_name, model_step2,
                        datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
                        None, "", error_str=f"Claude API {type(e).__name__}: {e}",
                    ),
                )
            except Exception:
                pass
        return (_ICP_EMPTY.copy(), {}, 0, 0, "api_error", f"Claude API {type(e).__name__}: {e}", 0, 0)

    except Exception as e:
        _dlog(f"Step 2 error for {company_name}: {e}")
        return (_ICP_EMPTY.copy(), {}, 0, 0, "api_error", f"{type(e).__name__}: {e}", 0, 0)


# ── Provider category membership — used by the post-processing sanitizer ─────
# Each provider is listed in exactly one set; membership is checked case-insensitively.

_CAT1_PROVIDERS: frozenset = frozenset({
    "gofluent", "learnlight", "speexx", "voxy", "learnship", "berlitz",
    "ef corporate solutions", "babbel for business", "rosetta stone enterprise",
    "preply business", "talaera", "busuu for business", "lingoda for business",
    "fluentify", "twenix", "cambly",
})

_CAT2_PROVIDERS: frozenset = frozenset({
    "duolingo", "babbel", "busuu", "rosetta stone", "preply", "memrise",
    "mondly", "elsa speak", "fluentu", "italki", "lingoda", "open english",
    "mango languages", "pimsleur", "drops", "hellotalk", "tandem",
})

_CAT3_PROVIDERS: frozenset = frozenset({
    "opensesame", "coursera for business", "udemy business", "linkedin learning",
    "skillsoft", "docebo", "degreed", "cornerstone", "360learning",
    "moodle workplace", "absorb lms", "talentlms", "learnupon", "pluralsight",
})

# mYngle must never appear in any competitor/provider signal field.
_MYNGLE_VARIANTS: frozenset = frozenset({"myngle", "mYngle"})

# ── mYngle competitor list ─────────────────────────────────────────────────────
# Ordered longest-to-shortest so _detect_competitor_in_text returns the most
# specific match first (e.g. "Preply Business" before "Preply").
# EF variants require "EF Corporate" / "EF Education First" phrasing — plain "ef"
# substrings are never matched. Duolingo requires corporate context (see below).
MYNGLE_COMPETITOR_KEYWORDS: list[str] = [
    # Specific / business variants first (longest match wins)
    "Preply Business",
    "Babbel for Business",
    "Busuu for Business",
    "EF Corporate Solutions",
    "EF Education First",
    "EF Corporate",
    "Rosetta Stone Enterprise",
    "Lingoda for Business",
    # Main brand names
    "Speexx", "goFLUENT", "Berlitz",
    "Babbel", "Busuu",
    "Learnship", "Learnlight",
    "Preply",
    "Rosetta Stone",
    "Wall Street English",
    "Linguarama", "Cegos", "Altissia",
    "Gymglish", "Voxy", "Lingoda",
    "Talaera", "Twenix", "Cambly", "Fluentify",
    "Duolingo", "Open English",
]

# Duolingo only counts as a competitor signal in a corporate / L&D / employee context.
_DUOLINGO_CORPORATE_PHRASES: frozenset = frozenset({
    "corporate", "employee", "employees", "enterprise", "business",
    "for business", "workforce", "l&d", "hr", "company",
    "learning and development", "training", "onboarding", "b2b",
})

# Subset used in Serper competitor-customer search queries.
# Covers all main brands; kept to a reasonable length to avoid query-limit issues.
_COMPETITOR_SEARCH_SUBSET: list[str] = [
    "Preply Business", "Preply",
    "Speexx", "goFLUENT", "Berlitz",
    "Babbel for Business", "Babbel",
    "Busuu for Business", "Busuu",
    "EF Corporate", "EF Education First",
    "Learnship", "Learnlight",
    "Rosetta Stone", "Wall Street English",
    "Linguarama", "Cegos", "Altissia",
    "Gymglish", "Voxy", "Lingoda",
    "Talaera", "Twenix", "Cambly", "Fluentify",
    "Open English",
]

# URL fragments that strongly suggest a case study or customer page
_COMPETITOR_CUSTOMER_URL_SIGNALS: tuple[str, ...] = (
    "case-study", "case_study", "customer-story", "customer_story",
    "success-story", "success_story", "client-story", "testimonial",
    "case-studies", "customer-stories", "references", "klanten",
    "kunden", "kundenstimmen", "clientes",
)

# Snippet phrases that strongly indicate the company is a customer/user
_COMPETITOR_CUSTOMER_SNIPPET_HIGH: tuple[str, ...] = (
    "selected", "chose", "deployed", "implemented",
    "is using", "uses", "has chosen", "has selected",
    "language training partner", "official language provider",
    "language learning platform", "training platform", "e-learning partner",
    "powered by", "in partnership with", "as their", "as its",
)

# Phrases indicating a medium-confidence connection
_COMPETITOR_CUSTOMER_SNIPPET_MEDIUM: tuple[str, ...] = (
    "access to", "benefit", "offers", "employees can",
    "learning benefit", "corporate benefit", "provided by",
    "as part of", "included in", "onboarding", "academy",
)


def _build_competitor_customer_empty() -> dict:
    """Return a zeroed-out ICP override + competitor attention fields dict."""
    return {
        "competitor_customer_signal":  "",
        "competitor_provider_detected": "",
        "competitor_signal_strength":  "",
        "competitor_signal_type":      "",
        "competitor_evidence":         "",
        "competitor_evidence_url":     "",
        # Low-threshold attention layer
        "competitor_attention_signal":            "No",
        "competitor_attention_provider_detected": "",
        "competitor_attention_strength":          "",
        "competitor_attention_type":              "",
        "competitor_attention_evidence":          "",
        "competitor_attention_url":               "",
        "competitor_attention_needs_review":      "False",
    }


def _run_competitor_customer_search(
    company_name: str,
    domain: str,
    serper_key: str,
) -> dict:
    """Run up to two targeted Serper searches for competitor-customer evidence.

    Query 1 (always): "<company_name>" (competitor OR competitor OR ...)
    Query 2 (when domain available): site:<domain> (competitor OR ...)

    Results are deduplicated by URL.
    Returns dict with keys 'hits', 'query_used', 'error'.
    Never logs or returns API key values.
    """
    competitors_q = " OR ".join(f'"{c}"' for c in _COMPETITOR_SEARCH_SUBSET)
    all_hits: list  = []
    queries_used: list[str] = []
    errors: list[str]       = []
    seen_urls: set[str]     = set()

    def _add_hits(hits: list) -> None:
        for h in (hits or []):
            url = str(h.get("link", "") or "")
            if url and url not in seen_urls:
                seen_urls.add(url)
                all_hits.append(h)

    # Query 1: company name + competitors (always run)
    q1 = f'"{company_name}" ({competitors_q})'
    hits1, _s1, _r1, err1 = _call_serper(q1, serper_key, timeout=12)
    _add_hits(hits1)
    queries_used.append(q1)
    if err1:
        errors.append(err1)

    # Query 2: site-scoped when a clean domain is available
    if domain and "." in domain:
        clean_domain = (
            domain.replace("https://", "").replace("http://", "").rstrip("/").split("/")[0]
        )
        if clean_domain:
            q2 = f'site:{clean_domain} ({competitors_q})'
            hits2, _s2, _r2, err2 = _call_serper(q2, serper_key, timeout=12)
            _add_hits(hits2)
            queries_used.append(q2)
            if err2:
                errors.append(err2)

    return {
        "hits":       all_hits,
        "query_used": " | ".join(queries_used),
        "error":      "; ".join(errors),
    }


def _detect_competitor_in_text(text: str) -> str:
    """Return the first competitor name found in text (case-insensitive), or ''.

    MYNGLE_COMPETITOR_KEYWORDS is ordered longest-first so the most specific
    variant is returned (e.g. 'Preply Business' before 'Preply').
    Duolingo requires a corporate/L&D context phrase to avoid consumer-app matches.
    EF variants already require specific context ('EF Corporate', etc.).
    mYngle is never in MYNGLE_COMPETITOR_KEYWORDS and is never returned.
    """
    text_lc = text.lower()
    for comp in MYNGLE_COMPETITOR_KEYWORDS:
        comp_lc = comp.lower()
        if comp_lc not in text_lc:
            continue
        if comp_lc == "duolingo":
            if not any(ph in text_lc for ph in _DUOLINGO_CORPORATE_PHRASES):
                continue
        return comp
    return ""


def _classify_competitor_customer_evidence(
    hits: list,
    company_name: str,
    existing_icp_signal: str = "",
    existing_evidence: str = "",
) -> dict:
    """Classify competitor-customer evidence from search hits and existing signals.

    Populates two layers:
      1. competitor_customer_* — hard override candidates
         High  → signal=Yes, strength=High
         Medium → signal=Unclear, strength=Medium
         Low   → signal=No, strength=Low
      2. competitor_attention_* — low-threshold alert for any plausible mention
         Fires on Low, Medium, or High hits where company name is connected.
         mYngle is never detected as a competitor.
    """
    result     = _build_competitor_customer_empty()
    company_lc = company_name.lower()

    # ── Walk each search hit ──────────────────────────────────────────────────
    best_strength  = ""
    best_type      = ""
    best_evidence  = ""
    best_url       = ""
    best_provider  = ""

    attn_strength  = ""
    attn_type      = ""
    attn_evidence  = ""
    attn_url       = ""
    attn_provider  = ""

    strength_rank  = {"High": 3, "Medium": 2, "Low": 1, "": 0}

    for hit in hits:
        title   = str(hit.get("title", ""))
        snippet = str(hit.get("snippet", ""))
        url     = str(hit.get("link", ""))
        text    = (title + " " + snippet).lower()

        provider = _detect_competitor_in_text(text)
        if not provider:
            continue

        # ── Determine signal type and strength ───────────────────────────────
        url_lc   = url.lower()
        strength = "Low"
        sig_type = "search_snippet_only"

        if any(frag in url_lc for frag in _COMPETITOR_CUSTOMER_URL_SIGNALS):
            strength = "High"
            sig_type = "case_study" if "case" in url_lc else "customer_story"

        elif company_lc in text and any(ph in text for ph in _COMPETITOR_CUSTOMER_SNIPPET_HIGH):
            strength = "High"
            if "tender" in url_lc or "procurement" in url_lc or "contract" in url_lc:
                sig_type = "procurement_or_tender"
            elif "job" in url_lc or "career" in url_lc or "vacature" in url_lc:
                sig_type = "job_posting_or_benefits"
            elif "supplier" in url_lc or "vendor" in url_lc:
                sig_type = "supplier_or_vendor_page"
            elif "academy" in url_lc or "training" in url_lc or "learning" in url_lc:
                sig_type = "training_platform_reference"
            else:
                sig_type = "company_website_mention"

        elif company_lc in text and any(ph in text for ph in _COMPETITOR_CUSTOMER_SNIPPET_MEDIUM):
            strength = "Medium"
            sig_type = "search_snippet_only"

        else:
            if company_lc not in text:
                continue  # no company connection — skip
            strength = "Low"
            sig_type = "weak_context_mention"

        # Update best-customer (hard override) hit
        if strength_rank[strength] > strength_rank[best_strength]:
            best_strength = strength
            best_type     = sig_type
            best_evidence = f"{title}: {snippet}"[:300]
            best_url      = url
            best_provider = provider

        # Update attention hit (same ranking — captures Low too)
        if strength_rank[strength] > strength_rank[attn_strength]:
            attn_strength = strength
            attn_type     = sig_type
            attn_evidence = f"{title}: {snippet}"[:300]
            attn_url      = url
            attn_provider = provider

    # ── Fall back to existing ICP signal fields ───────────────────────────────
    if not best_provider and existing_icp_signal and existing_icp_signal.strip():
        first_provider = existing_icp_signal.split(",")[0].strip()
        if first_provider:
            best_provider = first_provider
            best_strength = "Medium"
            best_type     = "search_snippet_only"
            best_evidence = (existing_evidence or "")[:300]
            best_url      = ""
            if not attn_provider:
                attn_provider = first_provider
                attn_strength = "Medium"
                attn_type     = "search_snippet_only"
                attn_evidence = best_evidence
                attn_url      = ""

    # ── Build customer (hard override) output ─────────────────────────────────
    if not best_provider:
        result["competitor_customer_signal"] = "No"
    elif best_strength == "High":
        result["competitor_customer_signal"] = "Yes"
    elif best_strength == "Medium":
        result["competitor_customer_signal"] = "Unclear"
    else:
        result["competitor_customer_signal"] = "No"

    result["competitor_provider_detected"] = best_provider
    result["competitor_signal_strength"]   = best_strength
    result["competitor_signal_type"]       = best_type
    result["competitor_evidence"]          = best_evidence.strip()
    result["competitor_evidence_url"]      = best_url

    # ── Build attention output ────────────────────────────────────────────────
    if attn_provider:
        result["competitor_attention_signal"]            = "Yes"
        result["competitor_attention_provider_detected"] = attn_provider
        result["competitor_attention_strength"]          = attn_strength
        result["competitor_attention_type"]              = attn_type
        result["competitor_attention_evidence"]          = attn_evidence.strip()
        result["competitor_attention_url"]               = attn_url
        result["competitor_attention_needs_review"]      = "True"
    else:
        result["competitor_attention_signal"]       = "No"
        result["competitor_attention_needs_review"] = "False"

    return result


def apply_competitor_icp_override(df: "pd.DataFrame") -> "pd.DataFrame":
    """Post-scoring override layer for confirmed competitor-customer evidence.

    Must be called AFTER apply_results_compatible_scoring().
    Adds / updates all ICP_OVERRIDE_FIELDS and COMPETITOR_ATTENTION_FIELDS columns.
    Never destroys the original score — it is saved to base_commercial_fit_score.

    Hard override (score=10 / Hot): only when competitor_customer_signal=Yes,
    competitor_signal_strength=High, competitor_provider_detected populated.

    Attention-only (no score change): when competitor_attention_signal=Yes but
    no hard override — sets CHECK COMPETITOR MENTION hint for caller review.
    """
    import pandas as _pd

    df = df.copy()

    if "final_commercial_fit_score" in df.columns:
        df["base_commercial_fit_score"] = (
            _pd.to_numeric(df["final_commercial_fit_score"], errors="coerce")
        )
    else:
        df["base_commercial_fit_score"] = _pd.NA

    for col in ICP_OVERRIDE_FIELDS:
        if col == "base_commercial_fit_score":
            continue
        if col not in df.columns:
            df[col] = ""

    for col in COMPETITOR_ATTENTION_FIELDS:
        if col not in df.columns:
            df[col] = "False" if col == "competitor_attention_needs_review" else ""

    for idx, row in df.iterrows():

        signal   = str(row.get("competitor_customer_signal", "") or "").strip()
        strength = str(row.get("competitor_signal_strength", "") or "").strip()
        provider = str(row.get("competitor_provider_detected", "") or "").strip()

        attn_signal   = str(row.get("competitor_attention_signal", "") or "").strip()
        attn_provider = str(row.get("competitor_attention_provider_detected", "") or "").strip()

        if signal == "Yes" and strength == "High" and provider:
            # ── Hard override ─────────────────────────────────────────────────
            df.at[idx, "final_commercial_fit_score"] = 10.0
            if "commercial_tier" in df.columns:
                df.at[idx, "commercial_tier"] = "🥇 Hot"
            df.at[idx, "icp_override_applied"]           = "Yes"
            df.at[idx, "icp_override_reason"]            = (
                f"Explicit competitor customer evidence: company appears connected to "
                f"{provider} as language training provider"
            )
            df.at[idx, "competitive_switch_opportunity"]  = "Strong"
            df.at[idx, "sales_action_hint"]              = "ATTACK: already buys language training category"
            df.at[idx, "competitor_attention_signal"]     = "Yes"
            df.at[idx, "competitor_attention_needs_review"] = "True"
            if not str(df.at[idx, "competitor_attention_provider_detected"] or "").strip():
                df.at[idx, "competitor_attention_provider_detected"] = provider
            if "scoring_notes" in df.columns:
                existing_note = str(df.at[idx, "scoring_notes"] or "")
                override_note = (
                    f"Competitive switch signal: Strong. Explicit evidence suggests the company "
                    f"already uses or works with {provider}. ICP override applied because the "
                    f"company already buys the language training category."
                )
                df.at[idx, "scoring_notes"] = (existing_note + " | " + override_note).lstrip(" | ")

        elif signal == "Unclear" and strength == "Medium":
            # ── Medium confidence — no override, flag for investigation ───────
            df.at[idx, "icp_override_applied"]            = "No"
            df.at[idx, "icp_override_reason"]             = ""
            df.at[idx, "competitive_switch_opportunity"]   = "Possible"
            df.at[idx, "sales_action_hint"]               = "Investigate current language training provider"
            df.at[idx, "competitor_attention_needs_review"] = "True"

        elif attn_signal == "Yes" and attn_provider and signal != "Yes":
            # ── Attention only — no score override, flag for caller review ────
            df.at[idx, "icp_override_applied"]            = "No"
            df.at[idx, "icp_override_reason"]             = ""
            df.at[idx, "competitive_switch_opportunity"]   = "Possible"
            df.at[idx, "sales_action_hint"]               = (
                "CHECK COMPETITOR MENTION: verify source URL before outreach"
            )
            df.at[idx, "competitor_attention_needs_review"] = "True"

        else:
            # ── No override ───────────────────────────────────────────────────
            df.at[idx, "icp_override_applied"]            = "No"
            df.at[idx, "icp_override_reason"]             = ""
            df.at[idx, "competitive_switch_opportunity"]   = "No clear signal"
            df.at[idx, "sales_action_hint"]               = ""

    return df


def run_competitor_override_selftest() -> dict:
    """Zero-cost self-test for the competitor-customer ICP override + attention logic.

    Creates synthetic DataFrames and synthetic Serper-like hits.
    Makes NO calls to Serper, Claude, Jina, Firecrawl, or any external service.

    Returns a dict:
        {
            "results": [{"name": str, "passed": bool, "details": str}, ...],
            "passed":  int,
            "failed":  int,
            "total":   int,
        }
    """
    import pandas as _pd

    results: list[dict] = []

    def _check(name: str, assertions: list[tuple[bool, str]]) -> dict:
        failures = [msg for ok, msg in assertions if not ok]
        passed   = not failures
        detail   = "OK" if passed else "; ".join(failures)
        return {"name": name, "passed": passed, "details": detail}

    # ── TC1: High-confidence override from pre-filled fields ─────────────────
    tc1_row = {
        "canonical_company_name":    "ACME TEST S.P.A.",
        "final_commercial_fit_score": 4.2,
        "commercial_tier":           "\U0001f9ca Cool",
        "competitor_customer_signal": "Yes",
        "competitor_signal_strength": "High",
        "competitor_provider_detected": "Preply Business",
        "competitor_signal_type":    "case_study",
        "competitor_evidence":       "ACME selected Preply Business as language training partner.",
        "competitor_evidence_url":   "https://preply.com/business/case-study/acme",
    }
    tc1_df  = _pd.DataFrame([tc1_row])
    tc1_out = apply_competitor_icp_override(tc1_df).iloc[0]
    results.append(_check("TC1: High-confidence override from pre-filled fields", [
        (float(tc1_out.get("base_commercial_fit_score", 0)) == 4.2,
            f"base_commercial_fit_score should be 4.2, got {tc1_out.get('base_commercial_fit_score')}"),
        (float(tc1_out.get("final_commercial_fit_score", 0)) == 10.0,
            f"final_commercial_fit_score should be 10.0, got {tc1_out.get('final_commercial_fit_score')}"),
        ("Hot" in str(tc1_out.get("commercial_tier", "")),
            f"commercial_tier should contain Hot, got '{tc1_out.get('commercial_tier')}'"),
        (str(tc1_out.get("icp_override_applied", "")) == "Yes",
            f"icp_override_applied should be 'Yes', got '{tc1_out.get('icp_override_applied')}'"),
        (str(tc1_out.get("competitive_switch_opportunity", "")) == "Strong",
            f"competitive_switch_opportunity should be 'Strong', got '{tc1_out.get('competitive_switch_opportunity')}'"),
        ("ATTACK" in str(tc1_out.get("sales_action_hint", "")),
            f"sales_action_hint should contain ATTACK, got '{tc1_out.get('sales_action_hint')}'"),
        ("Preply Business" in str(tc1_out.get("icp_override_reason", "")),
            f"icp_override_reason should mention Preply Business, got '{tc1_out.get('icp_override_reason')}'"),
    ]))

    # ── TC2: Medium signal should not override ────────────────────────────────
    tc2_row = {
        "final_commercial_fit_score": 5.1,
        "commercial_tier":           "\U0001f525 Warm",
        "competitor_customer_signal": "Unclear",
        "competitor_signal_strength": "Medium",
        "competitor_provider_detected": "Speexx",
    }
    tc2_df  = _pd.DataFrame([tc2_row])
    tc2_out = apply_competitor_icp_override(tc2_df).iloc[0]
    results.append(_check("TC2: Medium signal should not override", [
        (float(tc2_out.get("final_commercial_fit_score", 0)) == 5.1,
            f"score should remain 5.1, got {tc2_out.get('final_commercial_fit_score')}"),
        ("Warm" in str(tc2_out.get("commercial_tier", "")),
            f"commercial_tier should stay Warm, got '{tc2_out.get('commercial_tier')}'"),
        (str(tc2_out.get("icp_override_applied", "")) == "No",
            f"icp_override_applied should be 'No', got '{tc2_out.get('icp_override_applied')}'"),
        (str(tc2_out.get("competitive_switch_opportunity", "")) == "Possible",
            f"competitive_switch_opportunity should be 'Possible', got '{tc2_out.get('competitive_switch_opportunity')}'"),
        (str(tc2_out.get("sales_action_hint", "")) == "Investigate current language training provider",
            f"sales_action_hint wrong: '{tc2_out.get('sales_action_hint')}'"),
    ]))

    # ── TC3: No competitor signal should not override ─────────────────────────
    tc3_row = {
        "final_commercial_fit_score": 6.2,
        "commercial_tier":           "\U0001f525 Warm",
        "competitor_customer_signal": "No",
        "competitor_signal_strength": "",
        "competitor_provider_detected": "",
    }
    tc3_df  = _pd.DataFrame([tc3_row])
    tc3_out = apply_competitor_icp_override(tc3_df).iloc[0]
    results.append(_check("TC3: No competitor signal should not override", [
        (float(tc3_out.get("final_commercial_fit_score", 0)) == 6.2,
            f"score should remain 6.2, got {tc3_out.get('final_commercial_fit_score')}"),
        (str(tc3_out.get("icp_override_applied", "")) == "No",
            f"icp_override_applied should be 'No', got '{tc3_out.get('icp_override_applied')}'"),
        (str(tc3_out.get("competitive_switch_opportunity", "")) == "No clear signal",
            f"competitive_switch_opportunity wrong: '{tc3_out.get('competitive_switch_opportunity')}'"),
    ]))

    # ── TC4: High-confidence case-study hit from classifier ───────────────────
    tc4_hits = [{
        "title":   "ACME TEST selected Preply Business for employee language training",
        "snippet": "ACME TEST selected Preply Business as its language training partner for international teams.",
        "link":    "https://preply.com/business/case-study/acme-test",
    }]
    tc4_cls = _classify_competitor_customer_evidence(tc4_hits, "ACME TEST")
    results.append(_check("TC4: High-confidence case-study hit", [
        (tc4_cls.get("competitor_customer_signal") == "Yes",
            f"competitor_customer_signal should be Yes, got '{tc4_cls.get('competitor_customer_signal')}'"),
        (tc4_cls.get("competitor_provider_detected") == "Preply Business",
            f"competitor_provider_detected should be Preply Business, got '{tc4_cls.get('competitor_provider_detected')}'"),
        (tc4_cls.get("competitor_signal_strength") == "High",
            f"competitor_signal_strength should be High, got '{tc4_cls.get('competitor_signal_strength')}'"),
        (tc4_cls.get("competitor_signal_type") in ("case_study", "customer_story", "company_website_mention"),
            f"competitor_signal_type unexpected: '{tc4_cls.get('competitor_signal_type')}'"),
        (bool(tc4_cls.get("competitor_evidence_url")),
            "competitor_evidence_url should be populated"),
        (tc4_cls.get("competitor_attention_signal") == "Yes",
            f"competitor_attention_signal should be Yes, got '{tc4_cls.get('competitor_attention_signal')}'"),
    ]))

    # ── TC5: False positive generic article should not trigger ────────────────
    tc5_hits = [{
        "title":   "Best language learning platforms: Preply, Duolingo and others",
        "snippet": "A general market article comparing online language learning tools.",
        "link":    "https://example.com/best-language-learning-platforms",
    }]
    tc5_cls = _classify_competitor_customer_evidence(tc5_hits, "ACME TEST")
    results.append(_check("TC5: False positive generic article should not trigger", [
        (tc5_cls.get("competitor_customer_signal") == "No",
            f"competitor_customer_signal should be No, got '{tc5_cls.get('competitor_customer_signal')}'"),
        (tc5_cls.get("competitor_provider_detected", "") == "",
            f"competitor_provider_detected should be empty, got '{tc5_cls.get('competitor_provider_detected')}'"),
        (tc5_cls.get("competitor_attention_signal") == "No",
            f"competitor_attention_signal should be No for generic article, got '{tc5_cls.get('competitor_attention_signal')}'"),
    ]))

    # ── TC6: Weak context mention should not trigger hard override ─────────────
    tc6_hits = [{
        "title":   "ACME TEST and Preply mentioned in industry overview",
        "snippet": "The article mentions ACME TEST and Preply in a broad overview of education technology.",
        "link":    "https://example.com/industry-overview",
    }]
    tc6_cls    = _classify_competitor_customer_evidence(tc6_hits, "ACME TEST")
    tc6_signal = tc6_cls.get("competitor_customer_signal", "No")
    tc6_str    = tc6_cls.get("competitor_signal_strength", "")
    tc6_attn   = tc6_cls.get("competitor_attention_signal", "No")
    tc6_row = {
        "final_commercial_fit_score":             5.5,
        "commercial_tier":                        "\U0001f525 Warm",
        "competitor_customer_signal":             tc6_signal,
        "competitor_signal_strength":             tc6_str,
        "competitor_provider_detected":           tc6_cls.get("competitor_provider_detected", ""),
        "competitor_attention_signal":            tc6_attn,
        "competitor_attention_provider_detected": tc6_cls.get("competitor_attention_provider_detected", ""),
        "competitor_attention_url":               tc6_cls.get("competitor_attention_url", ""),
    }
    tc6_df  = _pd.DataFrame([tc6_row])
    tc6_out = apply_competitor_icp_override(tc6_df).iloc[0]
    results.append(_check("TC6: Weak context mention should not trigger hard override", [
        (tc6_signal in ("No", "Unclear") and tc6_str != "High",
            f"customer signal should not be High, got signal='{tc6_signal}' strength='{tc6_str}'"),
        (float(tc6_out.get("final_commercial_fit_score", 0)) != 10.0,
            f"score should NOT be 10.0, got {tc6_out.get('final_commercial_fit_score')}"),
        (str(tc6_out.get("icp_override_applied", "")) == "No",
            f"icp_override_applied should be No, got '{tc6_out.get('icp_override_applied')}'"),
    ]))

    # ── TC-Guard: mYngle must never be detected as competitor ─────────────────
    tc_myngle_hits = [{
        "title":   "ACME TEST uses mYngle for corporate language training",
        "snippet": "ACME TEST selected mYngle as their official language training partner.",
        "link":    "https://myngle.com/case-study/acme-test",
    }]
    tc_myngle_cls = _classify_competitor_customer_evidence(tc_myngle_hits, "ACME TEST")
    results.append(_check("TC-Guard: mYngle must never be detected as competitor/provider", [
        ("myngle" not in str(tc_myngle_cls.get("competitor_provider_detected", "")).lower(),
            f"mYngle in competitor_provider_detected: '{tc_myngle_cls.get('competitor_provider_detected')}'"),
        ("myngle" not in str(tc_myngle_cls.get("competitor_attention_provider_detected", "")).lower(),
            f"mYngle in attention_provider_detected: '{tc_myngle_cls.get('competitor_attention_provider_detected')}'"),
    ]))

    # ── TC-Att1: Preply low-context mention triggers attention signal ──────────
    tc_att1_hits = [{
        "title":   "ACME TEST and Preply mentioned in language learning overview",
        "snippet": "The article mentions ACME TEST and Preply in relation to corporate language learning.",
        "link":    "https://example.com/overview",
    }]
    tc_att1_cls = _classify_competitor_customer_evidence(tc_att1_hits, "ACME TEST")
    results.append(_check("TC-Att1: Preply low-context mention triggers attention signal", [
        (tc_att1_cls.get("competitor_attention_signal") == "Yes",
            f"competitor_attention_signal should be Yes, got '{tc_att1_cls.get('competitor_attention_signal')}'"),
        ("preply" in str(tc_att1_cls.get("competitor_attention_provider_detected", "")).lower(),
            f"attention_provider should contain Preply, got '{tc_att1_cls.get('competitor_attention_provider_detected')}'"),
        (tc_att1_cls.get("competitor_attention_strength") in ("Low", "Medium", "High"),
            f"attention_strength unexpected: '{tc_att1_cls.get('competitor_attention_strength')}'"),
        (tc_att1_cls.get("competitor_customer_signal") != "Yes",
            f"customer_signal should NOT be Yes for weak mention, got '{tc_att1_cls.get('competitor_customer_signal')}'"),
    ]))

    # ── TC-Att2: Attention-only row triggers CHECK hint, no score override ─────
    tc_att2_row = {
        "final_commercial_fit_score":             5.5,
        "commercial_tier":                        "\U0001f525 Warm",
        "competitor_customer_signal":             "No",
        "competitor_signal_strength":             "Low",
        "competitor_provider_detected":           "",
        "competitor_attention_signal":            "Yes",
        "competitor_attention_provider_detected": "Preply",
        "competitor_attention_url":               "https://example.com/overview",
    }
    tc_att2_df  = _pd.DataFrame([tc_att2_row])
    tc_att2_out = apply_competitor_icp_override(tc_att2_df).iloc[0]
    results.append(_check("TC-Att2: Attention-only row triggers CHECK hint, no score override", [
        (float(tc_att2_out.get("final_commercial_fit_score", 0)) == 5.5,
            f"score should stay 5.5, got {tc_att2_out.get('final_commercial_fit_score')}"),
        (str(tc_att2_out.get("icp_override_applied", "")) == "No",
            f"icp_override_applied should be No, got '{tc_att2_out.get('icp_override_applied')}'"),
        ("CHECK COMPETITOR MENTION" in str(tc_att2_out.get("sales_action_hint", "")),
            f"sales_action_hint should contain CHECK COMPETITOR MENTION, got '{tc_att2_out.get('sales_action_hint')}'"),
        (str(tc_att2_out.get("competitive_switch_opportunity", "")) == "Possible",
            f"competitive_switch_opportunity should be Possible, got '{tc_att2_out.get('competitive_switch_opportunity')}'"),
    ]))

    # ── TC-Att3: Generic article without company name = no attention ───────────
    tc_att3_hits = [{
        "title":   "Preply Business releases new enterprise language training features",
        "snippet": "Preply Business announced new features for enterprise language training clients.",
        "link":    "https://preply.com/blog/enterprise-features",
    }]
    tc_att3_cls = _classify_competitor_customer_evidence(tc_att3_hits, "ACME TEST")
    results.append(_check("TC-Att3: Generic article without company name = no attention signal", [
        (tc_att3_cls.get("competitor_attention_signal") == "No",
            f"competitor_attention_signal should be No, got '{tc_att3_cls.get('competitor_attention_signal')}'"),
        (tc_att3_cls.get("competitor_customer_signal") == "No",
            f"competitor_customer_signal should be No, got '{tc_att3_cls.get('competitor_customer_signal')}'"),
    ]))

    # ── TC-Att4: EF false-positive guard ─────────────────────────────────────
    tc_att4_fp = _detect_competitor_in_text(
        "acme test has an effective and efficient workflow"
    )
    tc_att4_ok = _detect_competitor_in_text(
        "acme test selected ef corporate for international language training"
    )
    results.append(_check("TC-Att4: EF false-positive guard", [
        (tc_att4_fp == "",
            f"ef substring should not match, got '{tc_att4_fp}'"),
        (tc_att4_ok.lower().startswith("ef"),
            f"ef corporate context should match, got '{tc_att4_ok}'"),
    ]))

    # ── TC-Att5: Duolingo corporate guard ─────────────────────────────────────
    tc_att5_nocorp = _detect_competitor_in_text(
        "acme test managers enjoy duolingo on weekends"
    )
    tc_att5_corp = _detect_competitor_in_text(
        "acme test provides duolingo as a corporate language training benefit for employees"
    )
    results.append(_check("TC-Att5: Duolingo requires corporate/L&D context", [
        (tc_att5_nocorp == "",
            f"Duolingo without corporate context should not match, got '{tc_att5_nocorp}'"),
        ("duolingo" in tc_att5_corp.lower(),
            f"Duolingo with corporate context should match, got '{tc_att5_corp}'"),
    ]))

    # ── TC-Att6: Preply Business matched over plain Preply ────────────────────
    tc_att6_provider = _detect_competitor_in_text(
        "acme test has chosen preply business as its official language training partner"
    )
    results.append(_check("TC-Att6: Preply Business matched over plain Preply", [
        (tc_att6_provider == "Preply Business",
            f"Should match Preply Business, got '{tc_att6_provider}'"),
    ]))

    passed = sum(1 for r in results if r["passed"])
    failed = len(results) - passed
    return {
        "results": results,
        "passed":  passed,
        "failed":  failed,
        "total":   len(results),
    }


def _sanitize_provider_list(raw_value: str, allowed: frozenset) -> str:
    """Return only items from raw_value (comma-separated) whose lowercase name
    matches a member of allowed, stripping mYngle and cross-category stragglers."""
    if not raw_value or not raw_value.strip():
        return ""
    kept = []
    for item in raw_value.split(","):
        name = item.strip()
        if not name:
            continue
        name_lc = name.lower()
        if name_lc in {v.lower() for v in _MYNGLE_VARIANTS}:
            continue  # never a competitor/provider
        if name_lc in allowed:
            kept.append(name)
    return ", ".join(kept)


def _sanitize_icp_provider_fields(fields: dict) -> dict:
    """Deterministic post-processing: enforce category membership rules and
    remove mYngle from all competitor/provider signal fields.

    icp_competitor_signal and icp_direct_language_competitor_signal both
    represent Category 1 providers; after individual sanitization their union
    is written back to both fields so neither is accidentally left empty when
    the other has a valid value.
    """
    cs = _sanitize_provider_list(fields.get("icp_competitor_signal", ""), _CAT1_PROVIDERS)
    dl = _sanitize_provider_list(fields.get("icp_direct_language_competitor_signal", ""), _CAT1_PROVIDERS)

    # Build deduplicated union preserving first-seen order
    seen: dict = {}
    for name in [n.strip() for n in (cs + ("," if cs and dl else "") + dl).split(",") if n.strip()]:
        seen.setdefault(name.lower(), name)
    merged = ", ".join(seen.values())

    fields["icp_competitor_signal"] = merged
    fields["icp_direct_language_competitor_signal"] = merged
    fields["icp_online_language_learning_signal"] = _sanitize_provider_list(
        fields.get("icp_online_language_learning_signal", ""), _CAT2_PROVIDERS,
    )
    fields["icp_broader_lnd_platform_signal"] = _sanitize_provider_list(
        fields.get("icp_broader_lnd_platform_signal", ""), _CAT3_PROVIDERS,
    )
    return fields


# ── Foreign HQ hygiene ────────────────────────────────────────────────────────

# Italian city names used to detect domestic HQ evidence
_ITALY_DOMESTIC_CITY_NAMES: frozenset = frozenset({
    "italy", "italia", "milan", "milano", "rome", "roma", "turin", "torino",
    "bologna", "florence", "firenze", "naples", "napoli", "venice", "venezia",
    "genoa", "genova", "bari", "catania", "palermo", "verona", "padova",
    "padua", "brescia", "modena", "parma", "bergamo", "reggio", "perugia",
    "trieste", "vicenza", "trento", "bolzano", "ancona", "ferrara",
})

# Phrases in evidence that indicate the company IS the domestic entity
_DOMESTIC_HQ_INDICATORS: tuple = (
    "headquartered in", "headquarters in", "sede in", "sede a",
    "based in", "fondata a", "fondata in",
    "head office in", "head office at",
)

# Phrases that indicate foreign parent / foreign ownership.
# IMPORTANT: keep these SPECIFIC to foreign ownership — do NOT include generic words like
# "part of", "group", "multinational", "subsidiaries", "global", "international"
# which describe legitimate Italian companies with international operations.
_FOREIGN_PARENT_INDICATORS: tuple = (
    "subsidiary of", "owned by", "acquired by", "controlled by",
    "branch of", "affiliate of",
    "wholly owned", "majority owned", "joint venture with",
    "parent company outside", "parent group outside", "holding company outside",
    "parent company in germany", "parent company in france", "parent company in us",
    "parent company in the us", "parent company in united states",
    "parent company in uk", "parent company in china", "parent company in japan",
    "owned by a foreign", "owned by foreign", "controlled by foreign",
    "external operating hq", "reporting line outside italy",
    "foreign parent", "foreign holding", "foreign ownership",
)

# Phrases proving the parent/owner is explicitly outside Italy.
# REQUIRED in addition to _FOREIGN_PARENT_INDICATORS to keep the FHQ score positive.
_EXPLICIT_FOREIGN_COUNTRY_INDICATORS: tuple = (
    # Country adjectives
    "japanese parent", "japanese group", "japanese company", "japanese owner",
    "german parent", "german group", "german company", "german owner",
    "american parent", "us parent", "us group", "us company",
    "french parent", "french group", "french company",
    "british parent", "uk parent", "uk group", "uk company",
    "dutch parent", "dutch group", "dutch owner",
    "swiss parent", "swiss group",
    "chinese parent", "chinese group",
    "swedish parent", "swedish group",
    "korean parent", "korean group",
    # Country-in-sentence patterns
    "headquartered in germany", "headquartered in japan", "headquartered in france",
    "headquartered in the us", "headquartered in the united states",
    "headquartered in united states", "headquartered in uk",
    "headquartered in the uk", "headquartered in united kingdom",
    "headquartered in china", "headquartered in sweden",
    "headquartered in netherlands", "headquartered in switzerland",
    "headquartered outside italy",
    "parent outside italy", "parent company outside italy",
    "parent group outside italy",
    "reports to parent outside", "hr decisions centralized outside",
    "decision structure outside italy",
    "owned by a company in germany", "owned by a company in japan",
    "owned by a company in france", "owned by a company in the us",
    # Named country patterns for "acquired by X, a [country] company"
    ", a japanese ", ", a german ", ", a french ", ", an american ",
    ", a british ", ", a dutch ", ", a swedish ", ", a swiss ",
    ", a chinese ", ", a korean ",
    "based in germany", "based in japan", "based in france",
    "based in the us", "based in united states", "based in uk",
    "based in netherlands", "based in sweden", "based in switzerland",
    "sanden corporation",  # known Japanese parent (Sandenvendo)
)

# Phrases indicating a domestic Italian company with international footprint.
# These FORCE sanitization to 0 when found, even before the city check.
_DOMESTIC_INTL_FOOTPRINT_INDICATORS: tuple = (
    "italy with international subsidiaries",
    "italy with usa operations",
    "italy with operations abroad",
    "italia con filiali",
    "italia con sussidiarie",
    "italy and international subsidiaries",
    "headquartered in italy",
    "based in italy",
    "sede in italy",
    "sede legale in italy",
    "italian company with",
    "italian group with",
)


_EMPLOYEE_RANGE_PATTERNS = [
    re.compile(r'\b\d{1,6}\s*[-–]\s*\d{1,6}\s*employees?\b', re.IGNORECASE),
    re.compile(r'\bhas\s+(?:approximately\s+|around\s+|about\s+)?\d{1,6}\s*employees?\b', re.IGNORECASE),
    re.compile(r'\bwith\s+(?:approximately\s+|around\s+|about\s+)?\d{1,6}\s*employees?\b', re.IGNORECASE),
    re.compile(r'\bemploying\s+(?:approximately\s+|around\s+|about\s+)?\d{1,6}\s+(?:people|staff|employees)\b', re.IGNORECASE),
    re.compile(r'\b(?:approximately|around|about|over|more than|fewer than)\s+\d{1,6}\s*employees?\b', re.IGNORECASE),
    re.compile(r'\bheadcount\s+of\s+\d{1,6}\b', re.IGNORECASE),
    re.compile(r'\b\d{1,6}\+?\s*employees?\b', re.IGNORECASE),
    re.compile(r'\bstaff\s+of\s+\d{1,6}\b', re.IGNORECASE),
]

# Qualitative size descriptors stripped from Italy register profiles
_QUALITATIVE_SIZE_PHRASES = (
    "small company", "small specialist", "small-sized company", "small-sized specialist",
    "mid-sized company", "mid-size company", "medium-sized company", "medium-size company",
    "large company", "large-sized company",
    "small business", "small enterprise", "medium enterprise", "large enterprise",
    "small firm", "medium firm", "large firm",
)


def _strip_employee_range_from_text(text: str) -> str:
    """Remove employee count/range and qualitative size mentions from visible profile text."""
    if not text:
        return text
    result = text
    for pat in _EMPLOYEE_RANGE_PATTERNS:
        result = pat.sub("", result)
    for phrase in _QUALITATIVE_SIZE_PHRASES:
        result = re.sub(r'\b' + re.escape(phrase) + r'\b', '', result, flags=re.IGNORECASE)
    # Clean up double spaces / dangling commas left by removal
    result = re.sub(r',\s*,', ',', result)
    result = re.sub(r'\s{2,}', ' ', result)
    result = re.sub(r',\s*\.', '.', result)
    return result.strip()


def _check_profile_consistency(rd: dict, text: str) -> str:
    """Remove internally inconsistent phrases from a visible profile text field.

    Rules applied (non-scoring, display-only):
    1. sig_intl_footprint_score >= 2 → strip "no evidence of international" / "no international" phrases
    2. sig_foreign_hq_score = 0 → strip foreign-parent / foreign-HQ mentions
    3. sig_explicit_lnd_score = 0 → strip L&D maturity claims
    4. sig_explicit_lnd_score >= 2 → strip "no clear L&D evidence" / "no L&D" phrases
    """
    if not text:
        return text

    result = text

    intl_score = int(rd.get("sig_intl_footprint_score", 0) or 0)
    fhq_score  = int(rd.get("sig_foreign_hq_score", 0) or 0)
    lnd_score  = int(rd.get("sig_explicit_lnd_score", 0) or 0)

    # Rule 1: has intl footprint but text says "no international presence"
    if intl_score >= 2:
        for phrase in (
            "no evidence of international footprint",
            "no evidence of international",
            "no international operations",
            "no international presence",
            "no sign of international",
            "lacks international presence",
        ):
            result = result.replace(phrase, "")
            result = result.replace(phrase.capitalize(), "")

    # Rule 2: fhq_score = 0 → strip foreign-parent/foreign-HQ text from display
    if fhq_score == 0:
        # Strip full gap sentences first (before stripping the shorter sub-phrases)
        for full_phrase in (
            "No clear foreign parent or external HQ decision structure found",
            "No clear foreign parent / external HQ decision structure found",
            "No clear foreign parent/external HQ decision structure found",
            "No clear Italian-based / external hq decision structure found",
            "No clear Italian-based or external hq decision structure found",
            "No foreign parent or external HQ decision structure found",
            "no clear foreign parent or external hq decision structure found",
        ):
            result = result.replace(full_phrase, "")
            result = result.replace(full_phrase.capitalize(), "")
        # Strip the shorter phrases by replacing with empty string
        for phrase in (
            "foreign parent", "foreign headquarters", "foreign HQ",
            "foreign holding", "foreign ownership", "foreign owner",
            "non-Italian parent", "non-Italian HQ",
            "external HQ", "external headquarters",
        ):
            result = result.replace(phrase, "")
            result = result.replace(phrase.capitalize(), "")

    # Rule 3: lnd_score = 0 → strip strong L&D maturity claims
    if lnd_score == 0:
        for phrase in (
            "strong L&D culture", "mature L&D function", "established L&D program",
            "dedicated L&D", "strong learning culture", "structured learning program",
            "strong training culture", "formal training programs",
        ):
            result = result.replace(phrase, "")
            result = result.replace(phrase.capitalize(), "")

    # Rule 4: lnd_score >= 2 → strip "no L&D evidence" negations
    if lnd_score >= 2:
        for phrase in (
            "no clear L&D evidence", "no L&D evidence", "no evidence of L&D",
            "no learning and development", "no L&D investment",
            "limited L&D evidence", "no formal L&D",
        ):
            result = result.replace(phrase, "")
            result = result.replace(phrase.capitalize(), "")

    # Cleanup
    result = re.sub(r'\s{2,}', ' ', result)
    result = re.sub(r',\s*,', ',', result)
    result = re.sub(r',\s*\.', '.', result)
    return result.strip()


def _get_input_country(row: dict, scoring_profile: str = "default") -> str:
    """Infer the input country for a row from available fields.

    Returns an ISO2 code ("IT", "DE", etc.) or empty string if unknown.
    Priority: explicit canonical fields → scoring profile → unknown.
    """
    for field in ("input_country", "country_code", "canonical_country",
                  "lusha_api_country", "lusha_country", "country", "Company Country"):
        v = str(row.get(field, "") or "").strip().upper()
        if len(v) == 2 and v.isalpha():
            return v
        # Resolve "Italy" → "IT", "Germany" → "DE"
        vl = v.lower()
        if vl in ("italy", "italia"):
            return "IT"
        if vl in ("germany", "deutschland"):
            return "DE"
    # Fall back to scoring profile
    if scoring_profile == "italy_register_icp_only":
        return "IT"
    return ""


def sanitize_foreign_hq_signal(row: dict, input_country: str = "") -> dict:
    """Post-processing sanitizer: correct sig_foreign_hq_score when evidence
    shows a domestic multinational rather than a foreign-owned company.

    Modifies row in-place and returns it.
    Adds audit fields: foreign_hq_sanitized, foreign_hq_sanitizer_reason,
    foreign_hq_original_score, foreign_hq_original_evidence, inferred_input_country.
    """
    row.setdefault("foreign_hq_sanitized", False)
    row.setdefault("foreign_hq_sanitizer_reason", "")
    row.setdefault("foreign_hq_original_score", "")
    row.setdefault("foreign_hq_original_evidence", "")
    row.setdefault("foreign_hq_uncertain", False)
    # Always overwrite blank inferred_input_country with the resolved value so that
    # a prior pass that wrote "" (because scoring_profile wasn't known yet) is corrected.
    _existing_ctry = str(row.get("inferred_input_country", "") or "").strip()
    if input_country or not _existing_ctry:
        row["inferred_input_country"] = input_country

    orig_score = row.get("sig_foreign_hq_score", 0)
    try:
        score_int = int(orig_score)
    except (TypeError, ValueError):
        score_int = 0

    if score_int == 0:
        return row  # nothing to sanitize

    # Gather all evidence text
    evidence_text = " ".join(filter(None, [
        str(row.get("sig_foreign_hq_evidence", "") or ""),
        str(row.get("icp_evidence", "") or ""),
        str(row.get("icp_why_relevant", "") or ""),
        str(row.get("sig_intl_footprint_evidence", "") or ""),
    ])).lower()

    # Rule 0: explicit domestic-Italy-with-intl-footprint phrases → always sanitize for IT input
    if input_country == "IT":
        has_domestic_intl = any(ind in evidence_text for ind in _DOMESTIC_INTL_FOOTPRINT_INDICATORS)
        if has_domestic_intl:
            row["foreign_hq_original_score"]    = score_int
            row["foreign_hq_original_evidence"] = str(row.get("sig_foreign_hq_evidence", ""))
            row["sig_foreign_hq_score"]         = 0
            row["foreign_hq_sanitized"]         = True
            row["foreign_hq_sanitizer_reason"]  = (
                "Sanitized: evidence shows domestic HQ in Italy with international footprint "
                "(explicit domestic-intl phrase detected) — not a foreign HQ signal."
            )
            return row

    # Rule A/B: check for foreign parent / ownership (do NOT sanitize).
    # Only triggers on SPECIFIC foreign-ownership phrases — NOT on "group", "multinational",
    # "subsidiaries", "part of", "international operations", "global operations", etc.
    has_foreign_parent = any(ind in evidence_text for ind in _FOREIGN_PARENT_INDICATORS)
    if has_foreign_parent:
        # Also require explicit outside-Italy country proof before keeping the score
        has_explicit_country = any(ind in evidence_text for ind in _EXPLICIT_FOREIGN_COUNTRY_INDICATORS)
        if has_explicit_country:
            # Proven foreign parent with known country — keep the score
            return row
        # Acquisition/group language found but NO explicit country proof
        # → mark as uncertain, zero the score, flag for manual review
        row["foreign_hq_original_score"]    = score_int
        row["foreign_hq_original_evidence"] = str(row.get("sig_foreign_hq_evidence", ""))
        row["sig_foreign_hq_score"]         = 0
        row["foreign_hq_sanitized"]         = True
        row["foreign_hq_uncertain"]         = True
        row["foreign_hq_sanitizer_reason"]  = (
            "Acquisition/group evidence found, but no explicit evidence that parent "
            "or decision structure is outside Italy. Score zeroed pending manual review."
        )
        existing_review = int(row.get("model_signal_needs_manual_review", 0) or 0)
        if not existing_review:
            row["model_signal_needs_manual_review"] = 1
            existing_reason = str(row.get("model_signal_manual_review_reason", "") or "")
            row["model_signal_manual_review_reason"] = (
                (existing_reason + " | " if existing_reason else "")
                + "Foreign HQ uncertain: acquisition/group language without explicit country proof."
            )
        return row

    # Rule A: domestic HQ evidence for Italy input
    if input_country == "IT":
        has_domestic_hq = any(ind in evidence_text for ind in _DOMESTIC_HQ_INDICATORS)
        if has_domestic_hq:
            # Check if a domestic city or "italy" is mentioned within 100 chars of the HQ indicator
            for ind in _DOMESTIC_HQ_INDICATORS:
                idx = evidence_text.find(ind)
                if idx >= 0:
                    after = evidence_text[idx:idx + 100]
                    if any(city in after for city in _ITALY_DOMESTIC_CITY_NAMES):
                        row["foreign_hq_original_score"]    = score_int
                        row["foreign_hq_original_evidence"] = str(row.get("sig_foreign_hq_evidence", ""))
                        row["sig_foreign_hq_score"]         = 0
                        row["foreign_hq_sanitized"]         = True
                        row["foreign_hq_sanitizer_reason"]  = (
                            "Sanitized: evidence shows domestic HQ in Italy with international "
                            "footprint — not a foreign HQ signal."
                        )
                        return row

    # Rule B: no domestic city found but still no foreign parent — check for
    # pure export/international operations language (no HQ signal)
    _export_only_words = (
        "export", "subsidiaries abroad", "offices abroad", "operations abroad",
        "international clients", "global clients", "worldwide", "multinational",
        "internationally", "foreign markets", "overseas",
    )
    has_export_only = any(w in evidence_text for w in _export_only_words)
    has_any_hq_ref  = any(w in evidence_text for w in ("headquartered", "headquarters",
                                                         "sede", "parent", "owned by"))
    if has_export_only and not has_any_hq_ref:
        row["foreign_hq_original_score"]    = score_int
        row["foreign_hq_original_evidence"] = str(row.get("sig_foreign_hq_evidence", ""))
        row["sig_foreign_hq_score"]         = 0
        row["foreign_hq_sanitized"]         = True
        row["foreign_hq_sanitizer_reason"]  = (
            "Sanitized: evidence shows export/international operations only — "
            "no foreign parent or foreign HQ evidence found."
        )
        return row

    # Rule D: ambiguous — flag for manual review but keep the score
    if input_country and not has_foreign_parent and not has_export_only:
        existing_review = int(row.get("model_signal_needs_manual_review", 0) or 0)
        if not existing_review:
            row["model_signal_needs_manual_review"] = 1
            existing_reason = str(row.get("model_signal_manual_review_reason", "") or "")
            row["model_signal_manual_review_reason"] = (
                (existing_reason + " | " if existing_reason else "")
                + "Ambiguous foreign HQ evidence — verify whether this is a foreign parent/HQ "
                  "or domestic multinational footprint."
            )

    return row


def _extract_icp_fields(raw: dict) -> dict:
    fields = {
        "icp_lead_score":                          str(raw.get("lead_score")                          or "").strip(),
        "icp_buying_signals":                      str(raw.get("buying_signals")                      or "").strip(),
        "icp_competitor_signal":                   str(raw.get("competitor_signal")                   or "").strip(),
        "icp_direct_language_competitor_signal":   str(raw.get("direct_language_competitor_signal")   or "").strip(),
        "icp_online_language_learning_signal":     str(raw.get("online_language_learning_signal")     or "").strip(),
        "icp_broader_lnd_platform_signal":         str(raw.get("broader_lnd_platform_signal")         or "").strip(),
        "icp_evidence":                            str(raw.get("evidence")                            or "").strip(),
        "icp_likely_training_interest":            str(raw.get("likely_training_interest")            or "").strip(),
        "icp_why_relevant":                        str(raw.get("why_relevant")                        or "").strip(),
        "icp_potential_buyer_function":            str(raw.get("potential_buyer_function")            or "").strip(),
    }
    return _sanitize_icp_provider_fields(fields)


# ─────────────────────────────────────────────────────────────────────────────
# Model-signal extraction — Step 3
# ─────────────────────────────────────────────────────────────────────────────

_MODEL_SIGNAL_EMPTY: dict = {}  # populated after field list is known at import time

def _build_model_signal_empty() -> dict:
    empty: dict = {}
    for f in MODEL_SIGNAL_SCORE_FIELDS:
        empty[f] = 0
    for f in MODEL_SIGNAL_BINARY_FIELDS:
        empty[f] = 0
    for f in MODEL_SIGNAL_EVIDENCE_FIELDS:
        empty[f] = ""
    for f in MODEL_SIGNAL_QA_FIELDS:
        empty[f] = ""
    return empty


def _build_enrichment_context(row: dict) -> str:
    """Format the existing Step 1 + Step 2 enrichment data into a concise text context."""
    lines: list[str] = []

    # Step 1 firmographics
    s1_parts = []
    for key, label in [
        ("lusha_description",    "Description"),
        ("lusha_industry",       "Industry"),
        ("lusha_sub_industry",   "Sub-industry"),
        ("lusha_company_type",   "Company type"),
        ("lusha_country",        "Country"),
        ("lusha_city",           "City"),
        ("lusha_continent",      "Continent"),
        ("lusha_founded_year",   "Founded"),
        ("lusha_employee_range", "Employee range"),
        ("lusha_revenue",        "Revenue"),
        ("lusha_specialties",    "Specialties"),
        ("lusha_technologies",   "Technologies"),
        ("lusha_total_funding_amount", "Total funding"),
        ("lusha_total_funding_rounds", "Funding rounds"),
        ("lusha_last_round_type",  "Last round type"),
        ("lusha_last_round_amount","Last round amount"),
        ("lusha_ipo_status",       "IPO status"),
    ]:
        v = str(row.get(key, "") or "").strip()
        if v:
            s1_parts.append(f"{label}: {v}")
    if s1_parts:
        lines.append("=== Step 1 firmographic data ===")
        lines.extend(s1_parts)

    # Step 2 ICP signals
    s2_parts = []
    for key, label in [
        ("icp_lead_score",                        "Lead score"),
        ("icp_buying_signals",                    "Buying signals"),
        ("icp_competitor_signal",                 "Competitor signal (Cat 1)"),
        ("icp_direct_language_competitor_signal", "Direct language competitor"),
        ("icp_online_language_learning_signal",   "Online language learning signal (Cat 2)"),
        ("icp_broader_lnd_platform_signal",       "Broader L&D platform signal (Cat 3)"),
        ("icp_evidence",                          "ICP evidence"),
        ("icp_likely_training_interest",          "Likely training interest"),
        ("icp_why_relevant",                      "Why relevant"),
        ("icp_potential_buyer_function",          "Potential buyer function"),
    ]:
        v = str(row.get(key, "") or "").strip()
        if v:
            s2_parts.append(f"{label}: {v}")
    if s2_parts:
        lines.append("=== Step 2 ICP signal data ===")
        lines.extend(s2_parts)

    return "\n".join(lines) if lines else "(No enrichment context available)"


def _coerce_model_signals(raw: dict) -> dict:
    """Validate and coerce parsed model-signal JSON into expected types."""
    out: dict = _build_model_signal_empty()

    for f in MODEL_SIGNAL_SCORE_FIELDS:
        try:
            v = int(raw.get(f, 0) or 0)
            out[f] = max(0, min(3, v))
        except (TypeError, ValueError):
            out[f] = 0

    for f in MODEL_SIGNAL_BINARY_FIELDS:
        try:
            v = int(raw.get(f, 0) or 0)
            out[f] = 1 if v else 0
        except (TypeError, ValueError):
            out[f] = 0

    for f in MODEL_SIGNAL_EVIDENCE_FIELDS:
        out[f] = str(raw.get(f, "") or "").strip()

    # QA text fields
    out["model_signal_manual_review_reason"] = str(
        raw.get("model_signal_manual_review_reason", "") or ""
    ).strip()
    out["model_signal_sources_used"] = str(
        raw.get("model_signal_sources_used", "") or ""
    ).strip()
    sq = str(raw.get("model_signal_search_quality", "") or "").strip().lower()
    out["model_signal_search_quality"] = sq if sq in ("good", "partial", "weak", "failed") else "weak"

    return out


def run_model_signal_extraction(
    company_name: str,
    raw_url: str,
    enrichment_row: dict,
    api_key: str,
    model_id: str = MODEL_STEP2,
    include_evidence: bool = True,
    search_provider: str = STEP2_PROVIDER_SERPER,
) -> dict:
    """
    Extract structured model signals from already-fetched enrichment context.
    Returns a dict with all MODEL_SIGNAL_FIELDS populated.
    Never calls Jina, Serper, or the web_search tool — uses only provided context.
    Cache key is provider-specific so Serper and Claude results don't overwrite each other.
    """
    empty = _build_model_signal_empty()

    domain = clean_domain(raw_url) or company_name
    _prov_code = "sg" if search_provider == STEP2_PROVIDER_SERPER else "cws"
    cache_key = f"model_signals_{_prov_code}_{domain or safe_filename(company_name or 'unknown')}"

    cached = load_cache(cache_key)
    if cached is not None and cached.get("version") == 1:
        try:
            return _coerce_model_signals(cached.get("signals", {}))
        except Exception:
            _delete_cache(cache_key)

    context_text = _build_enrichment_context(enrichment_row)

    prompt = (
        MODEL_SIGNAL_PROMPT_TEMPLATE
        .replace("__COMPANY_NAME__", company_name or "(unknown)")
        .replace("__DOMAIN__", domain or "(unknown)")
        .replace("__ENRICHMENT_CONTEXT__", context_text)
    )

    _STRICT_SUFFIX = "\n\nReturn ONLY raw JSON. No markdown, no backticks, no explanation."

    try:
        client = anthropic.Anthropic(api_key=api_key)
        resp = client.messages.create(
            model=model_id,
            max_tokens=2500,
            messages=[{"role": "user", "content": prompt}],
        )
        raw_text = "".join(
            getattr(b, "text", "") for b in resp.content
            if getattr(b, "type", "") == "text"
        ).strip()

        try:
            raw_json = _parse_json_response(raw_text)
        except (json.JSONDecodeError, ValueError):
            resp2 = client.messages.create(
                model=model_id,
                max_tokens=2500,
                messages=[{"role": "user", "content": prompt + _STRICT_SUFFIX}],
            )
            raw_text2 = "".join(
                getattr(b, "text", "") for b in resp2.content
                if getattr(b, "type", "") == "text"
            ).strip()
            raw_json = _parse_json_response(raw_text2)

        signals = _coerce_model_signals(raw_json)
        save_cache(cache_key, {"version": 1, "signals": signals})

        if not include_evidence:
            for f in MODEL_SIGNAL_EVIDENCE_FIELDS:
                signals[f] = ""

        return signals

    except (json.JSONDecodeError, ValueError) as e:
        err_empty = _build_model_signal_empty()
        err_empty["model_signal_needs_manual_review"] = 1
        err_empty["model_signal_manual_review_reason"] = f"JSON parse error: {str(e)[:150]}"
        err_empty["model_signal_search_quality"] = "failed"
        return err_empty
    except anthropic.APIError as e:
        err_empty = _build_model_signal_empty()
        err_empty["model_signal_needs_manual_review"] = 1
        err_empty["model_signal_manual_review_reason"] = f"API error: {str(e)[:150]}"
        err_empty["model_signal_search_quality"] = "failed"
        return err_empty
    except Exception as e:
        err_empty = _build_model_signal_empty()
        err_empty["model_signal_needs_manual_review"] = 1
        err_empty["model_signal_manual_review_reason"] = f"Error: {type(e).__name__}: {str(e)[:150]}"
        err_empty["model_signal_search_quality"] = "failed"
        return err_empty


# ─────────────────────────────────────────────────────────────────────────────
# Review flagging
# ─────────────────────────────────────────────────────────────────────────────

def flag_review(row: dict, input_company_name: str) -> dict:
    reasons: list[str] = []
    status   = row.get("enrichment_status", "")
    returned = row.get("lusha_company_name", "")
    inp      = (input_company_name or "").strip()

    _bad_statuses = ("no_data", "api_error", "jina_error", "parse_error", "no_input")
    if status in _bad_statuses or any(status.endswith(s) for s in _bad_statuses):
        reasons.append(f"enrichment status: {status}")

    if returned and inp:
        inp_core = _strip_legal(inp).strip() or inp
        ret_core = _strip_legal(returned).strip() or returned
        sim = str_similarity(inp_core, ret_core)
        if sim < 0.70:
            reasons.append(
                f"Returned name '{returned}' differs from input '{inp}' ({sim:.0%})"
            )

    if inp and returned:
        inp_sfx = _legal_suffix(inp)
        ret_sfx = _legal_suffix(returned)
        if inp_sfx and ret_sfx and inp_sfx != ret_sfx:
            reasons.append(
                f"Legal entity mismatch: input '{inp_sfx.upper()}' vs returned '{ret_sfx.upper()}'"
            )

    row["needs_manual_review"] = "TRUE" if reasons else "FALSE"
    row["match_notes"]         = "; ".join(reasons) if reasons else ""
    return row


# ─────────────────────────────────────────────────────────────────────────────
# Real Lusha API enrichment (optional additional layer)
# ─────────────────────────────────────────────────────────────────────────────

_LUSHA_API_BASE = "https://api.lusha.com/company"
_LUSHA_TIMEOUT  = 15


def _lusha_raw_keys_summary(raw: dict) -> str:
    """
    Return a compact key-path summary of the top two levels of a dict.
    Used only for debugging — contains no secret values.
    Example: "top: data,meta; data: name,domain,industry"
    """
    if not isinstance(raw, dict):
        return f"(not a dict: {type(raw).__name__})"
    top_keys = list(raw.keys())
    parts = [f"top: {','.join(str(k) for k in top_keys)}"]
    for k in top_keys[:5]:
        v = raw.get(k)
        if isinstance(v, dict) and v:
            parts.append(f"{k}: {','.join(str(sk) for sk in list(v.keys())[:15])}")
        elif isinstance(v, list) and v and isinstance(v[0], dict):
            parts.append(f"{k}[0]: {','.join(str(sk) for sk in list(v[0].keys())[:15])}")
    return "; ".join(parts)


def _resolve_lusha_company_node(raw: dict) -> dict:
    """
    Try all known Lusha response nesting paths and return the dict that most
    likely contains the actual company record.

    Lusha API v2 wraps company data under raw["data"]; some versions use
    raw["company"], raw["data"]["company"], raw["companies"][0], or raw["results"][0].
    Fall back to raw itself if nothing better is found.
    """
    if not isinstance(raw, dict):
        return {}

    # Ordered list of extraction strategies
    candidates = []

    # raw["data"] — most common v2 envelope
    d = raw.get("data")
    if isinstance(d, dict) and d:
        # raw["data"]["company"] — double-wrapped
        dd = d.get("company")
        if isinstance(dd, dict) and dd:
            candidates.append(dd)
        else:
            candidates.append(d)

    # raw["company"]
    c = raw.get("company")
    if isinstance(c, dict) and c:
        candidates.append(c)

    # raw["companies"][0]
    clist = raw.get("companies")
    if isinstance(clist, list) and clist and isinstance(clist[0], dict):
        candidates.append(clist[0])

    # raw["results"][0]
    rlist = raw.get("results")
    if isinstance(rlist, list) and rlist and isinstance(rlist[0], dict):
        candidates.append(rlist[0])

    # raw itself as last resort
    candidates.append(raw)

    # Score each candidate by how many recognisable company fields it has
    _score_keys = {
        "name", "company_name", "industry", "description", "size",
        "employee_range", "employees", "country", "city", "domain",
        "linkedin", "founded", "type",
    }
    def _score(node):
        return sum(1 for k in node if k.lower() in _score_keys)

    best = max(candidates, key=_score, default=raw)
    return best if isinstance(best, dict) else raw


def _map_lusha_api_fields(raw: dict, source_url: str) -> dict:
    """
    Map a raw Lusha API response to LUSHA_API_FIELDS keys.
    Probes all known nesting structures defensively.
    lusha_api_raw_keys contains only key names for debugging — no values.
    """
    raw_keys_summary = _lusha_raw_keys_summary(raw)

    company = _resolve_lusha_company_node(raw)

    def _sv(node: dict, *keys) -> str:
        """Extract the first non-empty string value for any of the given keys."""
        for k in keys:
            v = node.get(k)
            if v is not None and str(v).strip() not in ("", "None", "null", "0"):
                return str(v).strip()
        return ""

    # Location — Lusha often nests under company["location"]
    loc = company.get("location") or {}
    if not isinstance(loc, dict):
        loc = {}

    def _loc(key: str) -> str:
        v = loc.get(key)
        if v is not None and str(v).strip() not in ("", "None", "null"):
            return str(v).strip()
        return ""

    country   = _sv(company, "country", "hq_country") or _loc("country") or _loc("countryCode")
    city      = _sv(company, "city", "hq_city")       or _loc("city")
    continent = _sv(company, "continent")              or _loc("continent")

    # Domain
    domain = (
        clean_domain(_sv(company, "domain", "website"))
        or clean_domain(_sv(raw,     "domain"))
        or clean_domain(source_url)
    )

    # Lists: specialties and technologies
    def _join_list(node: dict, *keys) -> str:
        for k in keys:
            v = node.get(k)
            if isinstance(v, list):
                return ", ".join(str(x) for x in v if x)
            if isinstance(v, str) and v.strip():
                return v.strip()
        return ""

    specialties  = _join_list(company, "specialties",  "specialties_list",  "expertises")
    technologies = _join_list(company, "technologies", "technology_stack",  "tech_stack")

    # Funding — Lusha often nests under company["funding"]
    funding = company.get("funding") or {}
    if not isinstance(funding, dict):
        funding = {}

    def _fund(key: str, *fallback_keys) -> str:
        v = _sv(funding, key, *fallback_keys)
        if v:
            return v
        return _sv(company, key, *fallback_keys)

    return {
        "lusha_api_company_name":         _sv(company, "name", "company_name", "companyName"),
        "lusha_api_domain":               domain,
        "lusha_api_description":          _sv(company, "description", "about", "summary"),
        "lusha_api_founded_year":         _sv(company, "founded", "founded_year", "year_founded", "foundedYear"),
        "lusha_api_employee_range":       _sv(company, "size", "employee_range", "employees", "company_size",
                                              "employeeRange", "employeeCount", "headcount"),
        "lusha_api_revenue_range":        _sv(company, "revenue_range", "revenue", "annual_revenue",
                                              "revenueRange", "annualRevenue"),
        "lusha_api_industry":             _sv(company, "industry", "main_industry", "primaryIndustry"),
        "lusha_api_sub_industry":         _sv(company, "sub_industry", "sub_category", "subIndustry"),
        "lusha_api_company_type":         _sv(company, "type", "company_type", "companyType"),
        "lusha_api_country":              country,
        "lusha_api_city":                 city,
        "lusha_api_continent":            continent,
        "lusha_api_linkedin_url":         _sv(company, "linkedin", "linkedin_url", "linkedinUrl"),
        "lusha_api_specialties":          specialties,
        "lusha_api_technologies":         technologies,
        "lusha_api_total_funding_amount": _fund("totalAmount",    "total_funding_amount", "total_funding"),
        "lusha_api_total_funding_rounds": _fund("totalRounds",    "total_funding_rounds", "funding_rounds"),
        "lusha_api_last_round_type":      _fund("lastRoundType",  "last_round_type",      "last_funding_type"),
        "lusha_api_last_round_amount":    _fund("lastRoundAmount","last_round_amount",     "last_funding_amount"),
        "lusha_api_last_round_date":      _fund("lastRoundDate",  "last_round_date",       "last_funding_date"),
        "lusha_api_ipo_status":           _sv(company, "ipo", "ipo_status", "ipoStatus"),
        # Debug-only key summary (key names only, no values)
        "lusha_api_raw_keys":             raw_keys_summary,
    }


def flag_lusha_api_review(
    lusha_fields: dict,
    input_company_name: str,
    input_url: str,
) -> dict:
    """
    Evaluate match quality for a real Lusha API result.
    Returns dict with keys: lusha_api_match_confidence, lusha_api_needs_review,
    lusha_api_match_notes.
    Kept separate from flag_review() which evaluates Step 1 fields.
    """
    reasons: list = []
    confidence = "high"

    api_name   = lusha_fields.get("lusha_api_company_name", "")
    api_domain = lusha_fields.get("lusha_api_domain",       "")
    api_status = lusha_fields.get("lusha_api_status",       "")

    inp_name   = (input_company_name or "").strip()
    inp_domain = clean_domain(input_url or "")

    # ── No useful data returned ───────────────────────────────────────────────
    if not _lusha_has_useful_data(lusha_fields):
        reasons.append("No useful Lusha API data returned")
        confidence = "low"

    # ── Company name similarity ───────────────────────────────────────────────
    if inp_name and api_name:
        inp_core = _strip_legal(inp_name).strip() or inp_name
        api_core = _strip_legal(api_name).strip()  or api_name
        sim = str_similarity(inp_core, api_core)
        if sim < 0.60:
            reasons.append(
                f"Name mismatch: input '{inp_name}' vs Lusha API '{api_name}' ({sim:.0%})"
            )
            confidence = "low"
        elif sim < 0.80:
            reasons.append(
                f"Possible name mismatch: '{inp_name}' vs '{api_name}' ({sim:.0%})"
            )
            if confidence == "high":
                confidence = "medium"

    # ── Domain comparison ─────────────────────────────────────────────────────
    if inp_domain and api_domain:
        if inp_domain != api_domain:
            reasons.append(
                f"Domain conflict: input '{inp_domain}' vs Lusha API '{api_domain}'"
            )
            if confidence != "low":
                confidence = "medium"

    # ── API-level error ───────────────────────────────────────────────────────
    if api_status and api_status not in ("ok", "cached"):
        reasons.append(f"Lusha API status: {api_status}")
        if confidence == "high":
            confidence = "medium"

    needs_review = "TRUE" if (confidence == "low" or reasons) else "FALSE"
    return {
        "lusha_api_match_confidence": confidence,
        "lusha_api_needs_review":     needs_review,
        "lusha_api_match_notes":      "; ".join(reasons) if reasons else "",
    }


# Fields that determine whether a Lusha API response contains useful company data
_LUSHA_USEFUL_FIELDS = [
    "lusha_api_company_name",
    "lusha_api_industry",
    "lusha_api_employee_range",
    "lusha_api_country",
    "lusha_api_description",
]


def _lusha_has_useful_data(fields: dict) -> bool:
    return any(fields.get(f, "").strip() for f in _LUSHA_USEFUL_FIELDS)


def run_lusha_api_enrichment(
    company_name: str,
    raw_url: str,
    api_key: str,
) -> tuple:
    """Live Lusha API enrichment is disabled in Layer 1. Moved to Buyer Contact Finder (Layer 2.5)."""
    disabled_meta = {
        "lusha_api_status":           "disabled_in_layer_1",
        "lusha_api_error":            "Live Lusha API disabled in Layer 1. Use Buyer Contact Finder for contact enrichment.",
        "lusha_api_match_confidence": "",
        "lusha_api_needs_review":     False,
        "lusha_api_match_notes":      "",
        "lusha_api_raw_keys":         "",
    }
    empty_fields = {f: "" for f in LUSHA_API_FIELDS}
    return {**empty_fields, **disabled_meta}, {}, "disabled_in_layer_1", ""


# ─────────────────────────────────────────────────────────────────────────────
# Company-domain validation  (runs before enrichment, uses local + Serper)
# ─────────────────────────────────────────────────────────────────────────────

def _dv_search_for_domain(company_name: str, serper_key: str) -> str:
    """Run two Serper queries and return the most likely official domain.

    Parses result URLs directly — no Claude call.  Returns "" when nothing
    convincing is found.
    """
    queries = [
        f'"{company_name}" official website',
        f'"{company_name}" company',
    ]
    domain_hits: dict[str, int] = {}
    for q in queries:
        hits, _, _, err = _call_serper(q, serper_key, timeout=10)
        if err or not hits:
            continue
        for h in hits[:5]:
            link = h.get("link", "") or ""
            d = clean_domain(link)
            if not d or d in _GENERIC_DOMAINS:
                continue
            if _dv_token_overlap(company_name, d) >= 0.15:
                domain_hits[d] = domain_hits.get(d, 0) + 1
    if not domain_hits:
        return ""
    return max(domain_hits, key=lambda d: (domain_hits[d], _dv_token_overlap(company_name, d)))


def validate_company_domain(
    company_name: str,
    raw_url: str,
    serper_key: str = "",
    dry_run: bool = False,
) -> dict:
    """Pre-enrichment check: does the input company name match the input domain?

    Fast local token check runs first; Serper is called only for uncertain or
    suspicious cases when a key is provided and dry_run is False.

    Returns a dict with all DOMAIN_VALIDATION_FIELDS keys.
    """
    result: dict = {
        "input_domain":               "",
        "validated_domain":           "",
        "domain_used_for_enrichment": "unknown",
        "domain_match_confidence":    "Unknown",
        "possible_domain_mismatch":   "False",
        "suggested_domain":           "",
        "domain_check_reason":        "",
        "domain_source":              "",
        "needs_domain_review":        "False",
    }

    input_domain = clean_domain(raw_url) if raw_url else ""
    result["input_domain"] = input_domain
    company_name = (company_name or "").strip()

    if not company_name:
        result.update({
            "domain_check_reason":        "No company name provided.",
            "domain_source":              "no_input",
            "domain_used_for_enrichment": "unknown",
            "needs_domain_review":        "True",
        })
        return result

    # ── No domain ─────────────────────────────────────────────────────────────
    if not input_domain:
        result.update({
            "domain_used_for_enrichment": "company_name_only",
            "domain_check_reason":        "No domain provided in input.",
            "domain_source":              "missing_domain",
            "needs_domain_review":        "True",
        })
        if serper_key and not dry_run:
            suggested = _dv_search_for_domain(company_name, serper_key)
            if suggested:
                result.update({
                    "suggested_domain":           suggested,
                    "validated_domain":           suggested,
                    "domain_used_for_enrichment": "suggested_domain",
                    "domain_match_confidence":    "Medium",
                    "domain_check_reason":        (
                        f"No domain in input; Serper search suggests {suggested}."
                    ),
                    "domain_source": "serper_search",
                })
        return result

    # ── Generic / directory domain ────────────────────────────────────────────
    if input_domain in _GENERIC_DOMAINS:
        result.update({
            "possible_domain_mismatch":   "True",
            "domain_match_confidence":    "Low",
            "domain_check_reason":        (
                f"{input_domain} is a generic directory/social/media site, "
                "not a company website."
            ),
            "domain_source":              "local_generic_check",
            "needs_domain_review":        "True",
        })
        if serper_key and not dry_run:
            suggested = _dv_search_for_domain(company_name, serper_key)
            if suggested:
                result.update({
                    "suggested_domain":           suggested,
                    "validated_domain":           suggested,
                    "domain_used_for_enrichment": "suggested_domain",
                    "domain_match_confidence":    "Medium",
                    "domain_source":              "serper_search",
                    "domain_check_reason":        (
                        f"{input_domain} is a generic site. "
                        f"Serper search suggests {suggested} as official domain."
                    ),
                })
        return result

    # ── Fast local token check ────────────────────────────────────────────────
    overlap = _dv_token_overlap(company_name, input_domain)

    if overlap >= 0.5:
        result.update({
            "validated_domain":           input_domain,
            "domain_used_for_enrichment": "original_domain",
            "domain_match_confidence":    "High",
            "possible_domain_mismatch":   "False",
            "domain_check_reason":        (
                "Company name tokens strongly match domain tokens."
            ),
            "domain_source":              "local_token_match",
            "needs_domain_review":        "False",
        })
        return result

    if overlap >= 0.15:
        # Partial match — acceptable: may be group/parent domain or abbreviation
        result.update({
            "validated_domain":           input_domain,
            "domain_used_for_enrichment": "original_domain",
            "domain_match_confidence":    "Medium",
            "possible_domain_mismatch":   "False",
            "domain_check_reason":        (
                "Partial name-domain overlap — likely a group/parent domain "
                "or brand abbreviation. Treated as acceptable."
            ),
            "domain_source":              "local_token_match",
            "needs_domain_review":        "False",
        })
        return result

    # ── Low / no token overlap — suspicious ───────────────────────────────────
    result.update({
        "validated_domain":           input_domain,
        "domain_used_for_enrichment": "original_domain",
        "domain_match_confidence":    "Low",
        "possible_domain_mismatch":   "True",
        "domain_check_reason":        (
            f"Company name '{company_name}' shares no clear tokens with "
            f"domain '{input_domain}'. Possible mismatch."
        ),
        "domain_source":              "local_token_check",
        "needs_domain_review":        "True",
    })

    if serper_key and not dry_run:
        suggested = _dv_search_for_domain(company_name, serper_key)
        if suggested and suggested != input_domain:
            ov2 = _dv_token_overlap(company_name, suggested)
            result["suggested_domain"] = suggested
            result["domain_source"] = "serper_search"
            if ov2 >= 0.3:
                # Strong match → use suggested for enrichment
                result.update({
                    "validated_domain":           suggested,
                    "domain_used_for_enrichment": "suggested_domain",
                    "domain_match_confidence":    "High",
                    "domain_check_reason":        (
                        f"Input domain '{input_domain}' did not match. "
                        f"Serper search strongly suggests '{suggested}'."
                    ),
                })
            else:
                # Weak match → flag but don't replace
                result.update({
                    "domain_match_confidence": "Medium",
                    "domain_check_reason":     (
                        f"Input domain '{input_domain}' did not match. "
                        f"Serper search found '{suggested}' as a possible "
                        "alternative — verify before use."
                    ),
                })

    return result


# ─────────────────────────────────────────────────────────────────────────────
# Per-row enrichment  ← orchestrates both steps
# ─────────────────────────────────────────────────────────────────────────────

def enrich_one_row(
    company_name: str,
    raw_url: str,
    api_key: str,
    delay: float,
    use_playwright: bool = True,
    model_step1: str = MODEL_STEP1,
    model_step2: str = MODEL_STEP2,
    _debug_callback=None,
    search_provider: str = STEP2_PROVIDER_SERPER,
    serper_key: str = "",
    dry_run: bool = False,
    enable_lusha_api: bool = False,
    lusha_api_key: str = "",
    extract_model_signals: bool = True,
    include_signal_evidence: bool = True,
    run_step1_enrichment: bool = True,
    run_step2_enrichment: bool = True,
    existing_lusha_data: dict | None = None,
    _cli_verbose: bool = False,
    scoring_profile: str = "default",
) -> tuple:
    """
    Run optional Lusha API enrichment, then Step 1 (Jina + Claude extraction),
    then Step 2 (Claude web_search ICP), then model-signal extraction (Step 3).
    Returns (combined_fields_dict, debug_record_dict).
    """
    url          = raw_url.strip() if raw_url else ""
    company_name = company_name.strip() if company_name else ""

    # ── Domain validation (fast local check + optional Serper) ────────────────
    if is_cli_mode():
        print("[enricher]   Step 0: domain validation", flush=True)
    _dv = validate_company_domain(
        company_name, url,
        serper_key=serper_key,
        dry_run=dry_run,
    )
    # Use the suggested domain for enrichment only when confidence is High
    if _dv.get("domain_used_for_enrichment") == "suggested_domain" and _dv.get("suggested_domain"):
        url = normalize_url(_dv["suggested_domain"])

    row = {f: "" for f in ALL_ENRICHMENT_FIELDS}
    row.update(_dv)

    # Populate canonical identity fields from input parameters
    row["canonical_company_name"]   = company_name
    row["canonical_company_domain"] = clean_domain(url) if url else ""
    row["canonical_company_url"]    = normalize_url(url) if url else ""
    row["scoring_profile"]          = scoring_profile

    # Populate row with any pre-existing Lusha/Lucia data from the input file
    # so downstream steps (Step 2, Step 3) can use it as context.
    _lusha_fields_set = set(LUSHA_API_FIELDS + LUSHA_API_META_FIELDS + STEP1_FIELDS)
    if existing_lusha_data:
        for _k, _v in existing_lusha_data.items():
            if _k in _lusha_fields_set:
                row[_k] = _v

    # 8-second pause between companies to stay under the token/min rate limit
    time.sleep(8)

    # ── Lusha live API disabled in Layer 1 — see Buyer Contact Finder (Layer 2.5) ──
    _lusha_raw_json = {}
    # Lusha live API is disabled in Layer 1 — moved to Buyer Contact Finder (Layer 2.5)
    if False and enable_lusha_api and lusha_api_key:   # kept for reference only
        la_fields, _lusha_raw_json, la_status, la_err = run_lusha_api_enrichment(
            company_name, url, lusha_api_key,
        )
        row.update(la_fields)
        row["lusha_api_status"] = la_status
        row["lusha_api_error"]  = la_err
        review_meta = flag_lusha_api_review(la_fields, company_name, url)
        row.update(review_meta)
    else:
        la_fields = {f: "" for f in LUSHA_API_FIELDS}
        la_fields.update({
            "lusha_api_status":           "disabled_in_layer_1",
            "lusha_api_error":            "",
            "lusha_api_match_confidence": "",
            "lusha_api_needs_review":     False,
            "lusha_api_match_notes":      "Lusha/Lucia live enrichment moved to Buyer Contact Finder (Layer 2.5).",
            "lusha_api_raw_keys":         "",
        })
        row.update(la_fields)

    # ── Step 1 (three-tier: Jina → Playwright → web_search → no_data) ──────────
    if is_cli_mode():
        print("[enricher]   Step 1: Jina/Claude firmographics", flush=True)
    if run_step1_enrichment:
        s1_fields, s1_raw, s1_in, s1_out, s1_status, s1_err, s1_pw_dbg = run_step1(
            url, company_name, api_key, delay,
            use_playwright=use_playwright, model_step1=model_step1,
        )
        # Only overwrite existing non-empty Lusha fields if the new value is non-empty
        if existing_lusha_data:
            for _sf, _sv in s1_fields.items():
                if _sv or not row.get(_sf):
                    row[_sf] = _sv
        else:
            row.update(s1_fields)
        row["step1_status"]     = s1_status
        row["step1_run_status"] = s1_status
        row["step1_tokens_in"]  = str(s1_in)
        row["step1_tokens_out"] = str(s1_out)
        row["step1_cost_usd"]   = f"{calc_cost(s1_in, s1_out):.6f}"
    else:
        # Step 1 skipped — use existing data already loaded into row above
        s1_fields = {}
        s1_raw    = {}
        s1_in = s1_out = 0
        s1_status = "skipped_existing_data"
        s1_err    = ""
        s1_pw_dbg = {"playwright_attempted": False, "playwright_result": "skipped"}
        row["step1_status"]     = "skipped_existing_data"
        row["step1_run_status"] = "skipped_existing_data"
        row["step1_cost_usd"]   = "0.000000"
        row["lucia_api_called"] = 0
        if existing_lusha_data:
            row["lucia_data_status"] = "existing_preserved"
        else:
            row["lucia_data_status"] = "missing_not_requested"

    # ── Step 2 ────────────────────────────────────────────────────────────────
    if is_cli_mode():
        print("[enricher]   Step 2: Serper ICP search", flush=True)
    if run_step2_enrichment:
        s2_fields, s2_raw, s2_in, s2_out, s2_status, s2_err, s2_cache_create, s2_cache_read = run_step2(
            url, company_name, api_key, delay, model_step2=model_step2,
            _debug_callback=_debug_callback,
            search_provider=search_provider, serper_key=serper_key,
            dry_run=dry_run,
        )
        row.update(s2_fields)
        row["step2_status"]               = s2_status
        row["step2_provider_used"]        = search_provider
        row["anthropic_web_search_used"]  = False
        row["anthropic_tools_used"]       = ""
        row["serper_search_used"]         = (search_provider == STEP2_PROVIDER_SERPER)
        row["step2_tokens_in"]            = str(s2_in)
        row["step2_tokens_out"]           = str(s2_out)
        if is_cli_mode() and s2_status == "cached":
            print("[enricher]   Step 2: cached Serper result used", flush=True)
        row["step2_cost_usd"]      = f"{calc_cost(s2_in, s2_out):.6f}"
    else:
        s2_fields = _ICP_EMPTY.copy()
        s2_raw    = {}
        s2_in = s2_out = 0
        s2_status = "skipped"
        s2_err    = ""
        s2_cache_create = s2_cache_read = 0
        row["step2_status"]        = "skipped"
        row["step2_provider_used"] = search_provider
        row["step2_cost_usd"]      = "0.000000"

    # ── Combined metadata ─────────────────────────────────────────────────────
    total_in  = s1_in  + s2_in
    total_out = s1_out + s2_out
    row["total_tokens_in"]  = str(total_in)
    row["total_tokens_out"] = str(total_out)
    row["total_cost_usd"]   = f"{calc_cost(total_in, total_out):.6f}"

    has_s1 = _step1_has_data(s1_fields) if run_step1_enrichment else bool(existing_lusha_data)
    has_s2 = any(s2_fields.get(f, "") for f in ICP_FIELDS[:3])

    if not run_step1_enrichment and existing_lusha_data:
        row["enrichment_status"] = "existing_lusha_preserved" if has_s2 else "existing_lusha_only"
    elif has_s1:
        row["enrichment_status"] = s1_status if has_s2 else f"{s1_status}_step1_only"
    elif has_s2:
        # Cleaner input with no Lusha step1: Step 2 (model signals) succeeded
        row["enrichment_status"] = "enriched_step2_only"
    else:
        row["enrichment_status"] = "no_data"

    err_parts = [p for p in [s1_err, s2_err] if p]
    row["error_message"] = " | ".join(err_parts)

    flag_review(row, company_name)

    # ── Step 3 — Model-signal extraction ─────────────────────────────────────
    if is_cli_mode():
        print("[enricher]   Step 3: model signal extraction", flush=True)
    if extract_model_signals and api_key and not dry_run:
        try:
            ms_fields = run_model_signal_extraction(
                company_name=company_name,
                raw_url=raw_url,
                enrichment_row=row,
                api_key=api_key,
                model_id=model_step2,
                include_evidence=include_signal_evidence,
                search_provider=search_provider,
            )
            row.update(ms_fields)
        except Exception as _ms_exc:
            # Never let Step 3 failures abort the run
            _ms_err = _build_model_signal_empty()
            _ms_err["model_signal_needs_manual_review"] = 1
            _ms_err["model_signal_manual_review_reason"] = (
                f"Step 3 extraction failed: {type(_ms_exc).__name__}: {str(_ms_exc)[:120]}"
            )
            _ms_err["model_signal_search_quality"] = "failed"
            row.update(_ms_err)
    else:
        # Fill defaults when signal extraction is disabled or dry-run
        row.update(_build_model_signal_empty())

    # ── Step 3b — Foreign HQ hygiene sanitizer ───────────────────────────────
    # Runs unconditionally (even on dry-run or when Step 3 is disabled) so that
    # any pre-existing or zero-value foreign HQ fields are properly initialised.
    if is_cli_mode():
        print("[enricher]   Step 3b: foreign HQ hygiene", flush=True)
    _fhq_country = _get_input_country(row, scoring_profile)
    sanitize_foreign_hq_signal(row, _fhq_country)

    # ── Step 4 — Competitor customer search ──────────────────────────────────
    # Runs a targeted Serper query to find evidence that the company is already
    # a customer/user of a direct mYngle competitor.  Only runs when a Serper
    # key is available and not in dry-run mode.
    if is_cli_mode():
        print("[enricher]   Step 4: competitor customer search", flush=True)
    _comp_fields = _build_competitor_customer_empty()
    if serper_key and not dry_run:
        try:
            _cc_result = _run_competitor_customer_search(company_name, url, serper_key)
            _comp_fields = _classify_competitor_customer_evidence(
                hits=_cc_result.get("hits", []),
                company_name=company_name,
                existing_icp_signal=row.get("icp_competitor_signal", ""),
                existing_evidence=row.get("competitor_signal_strength_evidence", ""),
            )
        except Exception as _cc_exc:
            _comp_fields["competitor_evidence"] = (
                f"Step 4 search error: {type(_cc_exc).__name__}: {str(_cc_exc)[:120]}"
            )
    else:
        # No serper key or dry-run — still check existing ICP signal fields
        _existing_sig = row.get("icp_competitor_signal", "")
        if _existing_sig and _existing_sig.strip():
            _comp_fields = _classify_competitor_customer_evidence(
                hits=[],
                company_name=company_name,
                existing_icp_signal=_existing_sig,
                existing_evidence=row.get("competitor_signal_strength_evidence", ""),
            )
    row.update(_comp_fields)

    # ── Post-step-3 review flag update ───────────────────────────────────────
    # flag_review() ran before Step 3, so model signal flags and FHQ uncertainty
    # were not yet set.  Apply them now without overriding a TRUE already set.
    _post_review_reasons: list[str] = []
    if int(row.get("model_signal_needs_manual_review", 0) or 0):
        _ms_reason = str(row.get("model_signal_manual_review_reason", "") or "")
        _post_review_reasons.append(
            f"model signal review: {_ms_reason}" if _ms_reason else "model signal review flagged"
        )
    if row.get("foreign_hq_uncertain"):
        _post_review_reasons.append("foreign HQ uncertain - acquisition without country proof")
    if str(row.get("needs_domain_review", "")).lower() in ("true", "1"):
        _post_review_reasons.append("domain review required")
    if _post_review_reasons:
        _existing_notes = row.get("match_notes", "")
        _all_notes = "; ".join([p for p in [_existing_notes] + _post_review_reasons if p])
        row["needs_manual_review"] = "TRUE"
        row["match_notes"] = _all_notes

    # Debug record
    dbg = {
        "input_company_name":       company_name,
        "input_url":                raw_url,
        "normalized_url":           normalize_url(url),
        "lusha_api_status":         row.get("lusha_api_status", ""),
        "lusha_api_raw_json":       _lusha_raw_json,
        "step1_status":             s1_status,
        "step1_raw_json":           s1_raw,
        "step1_tokens_in":          s1_in,
        "step1_tokens_out":         s1_out,
        "step1_playwright_attempted": s1_pw_dbg.get("playwright_attempted", False),
        "step1_playwright_result":    s1_pw_dbg.get("playwright_result", "skipped"),
        "step2_status":                  s2_status,
        "step2_raw_json":                s2_raw,
        "step2_tokens_in":               s2_in,
        "step2_tokens_out":              s2_out,
        "step2_cache_creation_tokens":   s2_cache_create,
        "step2_cache_read_tokens":       s2_cache_read,
        "total_cost":                    calc_cost(total_in, total_out),
        "enrichment_status":        row["enrichment_status"],
        "error_message":            row["error_message"],
        "needs_manual_review":      row["needs_manual_review"],
        "match_notes":              row["match_notes"],
    }
    if is_cli_mode():
        print("[enricher]   Row completed", flush=True)
    return row, dbg


# ─────────────────────────────────────────────────────────────────────────────
# Feature engineering placeholder (not implemented yet)
# ─────────────────────────────────────────────────────────────────────────────

def build_model_features(df_enriched: pd.DataFrame) -> pd.DataFrame:
    """
    TODO:
    Convert raw enrichment data into model-ready scalar/discrete features.
    This should later create fields such as:
    - multi_country_presence: 0/1
    - foreign_hq_signal: 0/1
    - competitor_signal: 0/1
    - employee_size_score: 1-5
    - international_presence_score: 1-5
    - language_need_score: 1-5
    - industry_fit_score: 1-5
    - data_quality_score: 1-5

    This should stay separate from raw enrichment.
    """
    return pd.DataFrame()


# ─────────────────────────────────────────────────────────────────────────────
# Download helpers
# ─────────────────────────────────────────────────────────────────────────────

def _build_model_features_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Build a trimmed model_features sheet for the Excel export.
    Contains: original input columns + all model signal score/binary columns.
    Evidence columns are excluded here (they live in the main Enriched sheet).
    employee_size_score is included only when it already exists in df.
    """
    keep: list[str] = []

    # Preserve all original input columns (anything not in the enrichment field list)
    enrichment_col_set = set(ALL_ENRICHMENT_FIELDS)
    input_cols = [c for c in df.columns if c not in enrichment_col_set]
    keep.extend(input_cols)

    # Add score and binary signal columns (no evidence columns)
    signal_cols = [
        c for c in MODEL_SIGNAL_SCORE_FIELDS + MODEL_SIGNAL_BINARY_FIELDS
        if c in df.columns
    ]
    keep.extend(signal_cols)

    # Include employee_size_score only when already present
    if "employee_size_score" in df.columns:
        keep.append("employee_size_score")

    keep = list(dict.fromkeys(keep))  # deduplicate, preserve order
    available = [c for c in keep if c in df.columns]
    return df[available].copy()


def make_unique_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Return a copy of df with all duplicate column names made unique.

    Later occurrences of a repeated name are renamed  col__2, col__3, …
    No data is dropped or reordered.
    """
    seen: dict[str, int] = {}
    new_cols = []
    for col in df.columns:
        if col not in seen:
            seen[col] = 1
            new_cols.append(col)
        else:
            seen[col] += 1
            new_cols.append(f"{col}__{seen[col]}")
    if new_cols == list(df.columns):
        return df
    df = df.copy()
    df.columns = new_cols
    return df


def df_to_excel_bytes(df: pd.DataFrame) -> bytes:
    df = make_unique_columns(df)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Enriched")
        try:
            mf_df = _build_model_features_df(df)
            if not mf_df.empty:
                mf_df.to_excel(writer, index=False, sheet_name="model_features")
        except Exception:
            pass
        try:
            ev_cols = [c for c in df.columns if c.endswith("_evidence")]
            if ev_cols:
                # Include company name/domain columns plus all evidence columns
                id_cols = [c for c in df.columns if c not in set(ALL_ENRICHMENT_FIELDS)][:3]
                qa_cols = list(dict.fromkeys(id_cols + ev_cols))
                qa_df = df[[c for c in qa_cols if c in df.columns]]
                qa_df.to_excel(writer, index=False, sheet_name="qa_evidence")
        except Exception:
            pass
    return buf.getvalue()


def df_to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8-sig")


def _js_auto_download(df: pd.DataFrame, filename: str) -> None:
    """
    Inject a tiny JS snippet that silently downloads the DataFrame as Excel.
    Executes inside a zero-height iframe — does NOT trigger a Streamlit rerun.
    Note: browsers may prompt once to allow multiple automatic downloads.
    """
    b64 = base64.b64encode(df_to_excel_bytes(df)).decode()
    components.html(
        f"""<script>
        (function(){{
            var a = document.createElement('a');
            a.href = 'data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                     + ';base64,{b64}';
            a.download = '{filename}';
            document.body.appendChild(a);
            a.click();
            document.body.removeChild(a);
        }})();
        </script>""",
        height=1,   # must be ≥ 1; script still runs at 1px
    )


def _html_dl_buttons(partial_df: pd.DataFrame, n_done: int, stamp: str) -> None:
    """
    Render Excel + CSV download anchors as plain HTML — clicking them triggers
    a browser download WITHOUT causing a Streamlit rerun, so the processing
    loop is never interrupted.
    """
    xl_bytes  = df_to_excel_bytes(partial_df)
    csv_bytes = df_to_csv_bytes(partial_df)
    xl_b64    = base64.b64encode(xl_bytes).decode()
    csv_b64   = base64.b64encode(csv_bytes).decode()
    xl_kb     = max(len(xl_bytes)  // 1024, 1)
    csv_kb    = max(len(csv_bytes) // 1024, 1)
    _style = (
        "display:inline-block;padding:7px 16px;border-radius:6px;"
        "font-size:13px;font-family:sans-serif;font-weight:600;"
        "text-decoration:none;color:#fff;background:#0068c9;"
    )
    components.html(
        f"""<div style="display:flex;gap:10px;margin:2px 0;">
            <a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{xl_b64}"
               download="enrichedResults_partial_{stamp}.xlsx" style="{_style}">
               ⬇ Excel ({n_done} rows, ~{xl_kb} KB)
            </a>
            <a href="data:text/csv;base64,{csv_b64}"
               download="enrichedResults_partial_{stamp}.csv" style="{_style}">
               ⬇ CSV ({n_done} rows, ~{csv_kb} KB)
            </a>
        </div>""",
        height=48,
    )


def make_log_df(debug_records: list, elm_mode: bool = False) -> pd.DataFrame:
    rows = []
    for d in debug_records:
        if elm_mode:
            rows.append({
                "company":        d.get("company", d.get("input_company_name", "")),
                "url":            d.get("url", d.get("input_url", "")),
                "domain":         d.get("domain", ""),
                "fetch_status":   d.get("status", ""),
                "pages_fetched":  d.get("pages_fetched", ""),
                "total_chars":    d.get("total_chars", ""),
            })
        else:
            rows.append({
                "input_company_name":  d.get("input_company_name", ""),
                "input_url":           d.get("input_url", ""),
                "lusha_api_status":    d.get("lusha_api_status", ""),
                "step1_status":        d.get("step1_status", ""),
                "step2_status":        d.get("step2_status", ""),
                "enrichment_status":   d.get("enrichment_status", ""),
                "step1_tokens_in":     d.get("step1_tokens_in", ""),
                "step1_tokens_out":    d.get("step1_tokens_out", ""),
                "step2_tokens_in":     d.get("step2_tokens_in", ""),
                "step2_tokens_out":    d.get("step2_tokens_out", ""),
                "total_cost_usd":      f"{d.get('total_cost', 0):.6f}",
                "needs_manual_review": d.get("needs_manual_review", ""),
                "match_notes":         d.get("match_notes", ""),
                "error_message":       d.get("error_message", ""),
            })
    return pd.DataFrame(rows)


def cache_to_zip_bytes() -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in list_cache_files():
            zf.write(f, f.name)
    return buf.getvalue()


def build_partial_df(results: list, df_work: pd.DataFrame,
                     field_list: list | None = None) -> pd.DataFrame:
    """Build an enriched DataFrame from however many rows have been processed so far."""
    flist       = field_list if field_list is not None else ALL_ENRICHMENT_FIELDS
    df_out      = df_work.head(len(results)).copy().reset_index(drop=True)
    enriched_df = pd.DataFrame(results)
    for col in flist:
        df_out[col] = enriched_df[col].values if col in enriched_df.columns else ""
    return df_out


def ts() -> str:
    """Compact UTC timestamp for filenames: 20250516_143022"""
    return datetime.utcnow().strftime("%Y%m%d_%H%M%S")


def get_provider_code(provider_name: str) -> str:
    """Return a short filename-safe code for the Step 2 search provider."""
    if provider_name == STEP2_PROVIDER_CLAUDE:
        return "cws"
    if provider_name == STEP2_PROVIDER_SERPER:
        return "sg"
    return "unk"


def get_model_code(model_name: str) -> str:
    """Return a short filename-safe code for the Step 2 model."""
    m = (model_name or "").lower()
    if "haiku" in m:
        return "hq"
    if "sonnet" in m:
        return "sn"
    if "opus" in m:
        return "op"
    return "unk"


def build_run_tag() -> str:
    if is_cli_mode():
        return "sg_hq"
    _st, _ = get_streamlit()
    prov  = _st.session_state.get("_step2_provider",   STEP2_PROVIDER_SERPER)
    model = _st.session_state.get("_model_step2",       MODEL_STEP2)
    lusha = _st.session_state.get("_enable_lusha_api",  False)
    tag   = f"{get_provider_code(prov)}_{get_model_code(model)}"
    if lusha:
        tag = f"{tag}_lusha"
    return tag


def save_to_local_folder(df: pd.DataFrame, folder: str, run_tag: str = "") -> tuple[str, str]:
    """
    Write Excel + CSV to *folder* with timestamped filenames.
    Returns (excel_path, csv_path). Raises OSError on permission / path errors.
    Only usable when the app runs locally — not on Streamlit Cloud.
    """
    folder_path = Path(folder.strip())
    folder_path.mkdir(parents=True, exist_ok=True)
    stamp      = ts()
    tag        = f"_{run_tag}" if run_tag else ""
    excel_path = folder_path / f"enrichedResults_{stamp}.xlsx"
    csv_path   = folder_path / f"enrichedResults_{stamp}.csv"
    df_to_excel_bytes_write(df, str(excel_path))
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    return str(excel_path), str(csv_path)


def df_to_excel_bytes_write(df: pd.DataFrame, path: str) -> None:
    """Write DataFrame to an Excel file at *path* on disk (two sheets)."""
    df = make_unique_columns(df)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Enriched")
        try:
            mf_df = _build_model_features_df(df)
            if not mf_df.empty:
                mf_df.to_excel(writer, index=False, sheet_name="model_features")
        except Exception:
            pass
        try:
            ev_cols = [c for c in df.columns if c.endswith("_evidence")]
            if ev_cols:
                # Include company name/domain columns plus all evidence columns
                id_cols = [c for c in df.columns if c not in set(ALL_ENRICHMENT_FIELDS)][:3]
                qa_cols = list(dict.fromkeys(id_cols + ev_cols))
                qa_df = df[[c for c in qa_cols if c in df.columns]]
                qa_df.to_excel(writer, index=False, sheet_name="qa_evidence")
        except Exception:
            pass


# ─────────────────────────────────────────────────────────────────────────────
# Rich Excel report builder
# Creates: Lead Scores | Company Profiles (visible); Input hidden
#          Advanced Evidence | Scoring Settings | Enriched |
#          model_features | qa_evidence (hidden)
# ─────────────────────────────────────────────────────────────────────────────

def _xl_get(row_dict: dict, *keys: str, default: str = "") -> str:
    """Return first non-blank string value from row_dict for any of *keys."""
    _null = {"", "nan", "none", "n/a", "unknown", "null", "nat"}
    for k in keys:
        v = row_dict.get(k, "")
        if v is not None and str(v).strip().lower() not in _null:
            return str(v).strip()
    return default


def _xl_write_df(ws, df: pd.DataFrame) -> None:
    """Write DataFrame to worksheet with a styled header row."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    hdr_fill = PatternFill(start_color="0B4A92", end_color="0B4A92", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=str(col))
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for ri, record in enumerate(df.to_dict("records"), 2):
        for ci, col in enumerate(df.columns, 1):
            val = record[col]
            if val is None or (isinstance(val, float) and pd.isna(val)):
                val = ""
            ws.cell(row=ri, column=ci, value=val)

    for col_cells in ws.columns:
        try:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
                max(max_len + 2, 10), 60
            )
        except Exception:
            pass


def _xl_write_scoring_settings(ws, scoring_profile: str = "default") -> None:
    """Write scoring constants to the Scoring Settings sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment
    try:
        from commercial_fit_scoring import (
            INTERCEPT as _INT, LEAN_COEFFICIENTS as _LC,
            SIZE_BAND_LOOKUP as _SB,
            SCORING_PROFILES as _SPROFS,
        )
    except ImportError:
        ws.cell(row=1, column=1, value="Scoring module not available")
        return

    _sprof = _SPROFS.get(scoring_profile, _SPROFS["default"])
    _is_italy = (scoring_profile == "italy_register_icp_only")
    _TT   = _sprof["tier_thresholds"]
    _ICW  = _sprof["model_weight"]
    _CSW  = _sprof["size_weight"]
    _SIGMOID_K = _sprof["sigmoid_k"]

    hdr_fill = PatternFill(start_color="0B4A92", end_color="0B4A92", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    bold = Font(bold=True, size=10)
    norm = Font(size=10)

    r = 1
    ws.cell(row=r, column=1, value="Scoring Settings").font = Font(bold=True, size=13)
    r += 2

    ws.cell(row=r, column=1, value="Scoring profile").font = bold
    ws.cell(row=r, column=2, value=_sprof["label"]).font = norm
    r += 1
    ws.cell(row=r, column=1, value="Intercept").font = bold
    ws.cell(row=r, column=2, value=_INT).font = norm
    r += 1
    ws.cell(row=r, column=1, value="Sigmoid steepness (k)").font = bold
    ws.cell(row=r, column=2, value=_SIGMOID_K).font = norm
    r += 1
    ws.cell(row=r, column=1, value="Model weight").font = bold
    ws.cell(row=r, column=2, value=f"{round(_ICW*100):.0f}%").font = norm
    r += 1
    ws.cell(row=r, column=1, value="Size weight").font = bold
    ws.cell(row=r, column=2, value=f"{round(_CSW*100):.0f}%").font = norm
    r += 1
    if _is_italy:
        ws.cell(row=r, column=1,
                value="Final score is based on ICP/model signals only. "
                      "Company size is excluded from scoring because the Italian register input "
                      "is already pre-filtered for 100+ employees. "
                      "Company size remains available as audit/context data only.").font = norm
    else:
        ws.cell(row=r, column=1,
                value="Legacy 75/25 comparison score: final_commercial_fit_score_75_25_legacy "
                      "(temporary audit column — will be removed once distribution is stable)").font = norm
    r += 2

    ws.cell(row=r, column=1, value="Lean Model Coefficients").font = bold
    r += 1
    for h, c in (("Field", 1), ("Coefficient", 2)):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
    r += 1
    for field, coef in sorted(_LC.items(), key=lambda x: -x[1]):
        ws.cell(row=r, column=1, value=field).font = norm
        ws.cell(row=r, column=2, value=coef).font = norm
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="Size Bands").font = bold
    r += 1
    for h, c in (("Employee Range", 1), ("Score", 2)):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
    r += 1
    for rng, score in _SB.items():
        ws.cell(row=r, column=1, value=rng).font = norm
        ws.cell(row=r, column=2, value=score).font = norm
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="Tier Thresholds").font = bold
    r += 1
    for h, c in (("Min Score", 1), ("Tier", 2)):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
    r += 1
    for threshold, label in _TT:
        ws.cell(row=r, column=1, value=threshold).font = norm
        ws.cell(row=r, column=2, value=label).font = norm
        r += 1

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 20

    r += 1
    ws.cell(row=r, column=1, value="Layer 1 Notes").font = bold
    r += 1
    ws.cell(row=r, column=1, value="lusha_layer1_note").font = norm
    ws.cell(row=r, column=2, value=(
        "Lusha/Lucia live enrichment is disabled in Layer 1. "
        "Company size is resolved from input data, uploaded metadata, website/Jina evidence, "
        "or public snippets only. Contact enrichment belongs to Buyer Contact Finder (Layer 2.5)."
    )).font = norm


# Human-readable labels for model signal field names.
_SIGNAL_READABLE: dict[str, str] = {
    # Global complexity
    "sig_foreign_hq_score":                  "Foreign parent / external HQ decision structure",
    "sig_intl_footprint_score":              "International footprint",
    "sig_multicultural_score":               "Multicultural workforce",
    # People development
    "sig_explicit_lnd_score":                "Explicit learning and development focus",
    "sig_lnd_onboarding_score":              "Learning, development, and onboarding programmes",
    "sig_employer_branding_score":           "Employer branding and employee experience",
    # Growth and context
    "sig_rapid_growth_score":                "Rapid growth trajectory",
    "sig_merger_acq_score":                  "Mergers and acquisitions activity",
    # Training intent
    "ti_language_english_score":             "English language training interest",
    "ti_onboarding_score":                   "Onboarding training interest",
    "ti_leadership_score":                   "Leadership development training interest",
    "ti_intercultural_score":                "Intercultural and cross-cultural training interest",
    "ti_negotiation_sales_score":            "Negotiation and sales training interest",
    "ti_team_collab_score":                  "Team collaboration training interest",
    "ti_broader_professional_score":         "Broader professional skills training interest",
    # Competitor signals
    "competitor_signal_strength_score":      "Direct language training competitor signal",
    "language_competitor_strength_score":    "Language-specific competitor signal",
    "online_learning_signal_strength_score": "Online learning platform signal",
    "lnd_platform_signal_strength_score":    "Broader L&D platform signal",
}

# Ordered list of all signal fields used for profile and evidence output.
_SIGNAL_FIELDS_ORDERED: list[str] = [
    "sig_foreign_hq_score",
    "sig_intl_footprint_score",
    "sig_multicultural_score",
    "sig_explicit_lnd_score",
    "sig_lnd_onboarding_score",
    "sig_employer_branding_score",
    "sig_rapid_growth_score",
    "sig_merger_acq_score",
    "ti_language_english_score",
    "ti_onboarding_score",
    "ti_leadership_score",
    "ti_intercultural_score",
    "ti_negotiation_sales_score",
    "ti_team_collab_score",
    "ti_broader_professional_score",
    "competitor_signal_strength_score",
    "language_competitor_strength_score",
    "online_learning_signal_strength_score",
    "lnd_platform_signal_strength_score",
]

# Signal category mapping for Advanced Evidence sheet.
_SIGNAL_CATEGORY: dict[str, str] = {
    "sig_foreign_hq_score":                  "Global Complexity",
    "sig_intl_footprint_score":              "Global Complexity",
    "sig_multicultural_score":               "Global Complexity",
    "sig_explicit_lnd_score":                "People Development",
    "sig_lnd_onboarding_score":              "People Development",
    "sig_employer_branding_score":           "People Development",
    "sig_rapid_growth_score":                "Growth and Context",
    "sig_merger_acq_score":                  "Growth and Context",
    "ti_language_english_score":             "Training Intent",
    "ti_onboarding_score":                   "Training Intent",
    "ti_leadership_score":                   "Training Intent",
    "ti_intercultural_score":                "Training Intent",
    "ti_negotiation_sales_score":            "Training Intent",
    "ti_team_collab_score":                  "Training Intent",
    "ti_broader_professional_score":         "Training Intent",
    "competitor_signal_strength_score":      "Competitor Signals",
    "language_competitor_strength_score":    "Competitor Signals",
    "online_learning_signal_strength_score": "Competitor Signals",
    "lnd_platform_signal_strength_score":    "Competitor Signals",
}

# Fields with model coefficients >= 0.10 — meaningful for gap detection.
_GAP_CANDIDATE_FIELDS: frozenset[str] = frozenset({
    # sig_foreign_hq_score intentionally excluded: score=0 is normal (domestic company), not a gap
    "sig_explicit_lnd_score",
    "sig_intl_footprint_score",
    "sig_employer_branding_score",
    "sig_lnd_onboarding_score",
    "ti_onboarding_score",
})


def _signal_strength_label(score: float) -> str:
    if score >= 3:
        return "Strong"
    if score >= 2:
        return "Medium"
    if score >= 1:
        return "Weak"
    return "Missing"


def _build_caller_angle(rd: dict) -> str:
    """Build a 2-3 sentence caller-facing angle for a scored company row.

    Replaces generic scoring disclaimers (size weight, K-factor, audit notes)
    with tier-specific, signal-anchored guidance for the sales or calling team.
    The scoring_profile is read from rd["scoring_profile"] so the function is
    self-contained and requires no extra arguments.
    """
    tier = str(rd.get("commercial_tier", "") or "").strip()
    _profile = str(rd.get("scoring_profile", "default") or "default")
    _is_italy = (_profile == "italy_register_icp_only")

    def _f(field: str) -> float:
        try:
            return float(rd.get(field, 0) or 0)
        except (TypeError, ValueError):
            return 0.0

    # Signal scores
    _intl    = _f("sig_intl_footprint_score")
    _lnd     = _f("sig_explicit_lnd_score")
    _onboard = _f("sig_lnd_onboarding_score")
    _multi   = _f("sig_multicultural_score")
    _growth  = _f("sig_rapid_growth_score")
    _merger  = _f("sig_merger_acq_score")
    _employer= _f("sig_employer_branding_score")
    _fhq     = _f("sig_foreign_hq_score")
    _fhq_san = bool(rd.get("foreign_hq_sanitized"))
    _ti_eng  = _f("ti_language_english_score")
    _ti_lead = _f("ti_leadership_score")
    _ti_onb  = _f("ti_onboarding_score")
    _ti_inter= _f("ti_intercultural_score")
    _ti_neg  = _f("ti_negotiation_sales_score")

    buyer_fn = str(rd.get("icp_potential_buyer_function", "") or "").strip()
    comp_opp = str(rd.get("competitive_switch_opportunity", "") or "").strip()

    # Infer most relevant buyer entry function
    def _buyer_hint() -> str:
        if buyer_fn:
            return buyer_fn
        if _lnd >= 2 or _onboard >= 2:
            return "HR, L&D or People leadership"
        if _ti_lead >= 2:
            return "L&D, HR or Senior leadership"
        if _multi >= 2 or _ti_inter >= 2:
            return "HR, Operations or People leadership"
        if _ti_onb >= 2:
            return "HR, Talent or Operations"
        if _ti_eng >= 2:
            return "HR or L&D leadership"
        if _intl >= 2:
            return "HR, Operations or People leadership"
        return "HR or People leadership"

    # Describe the single strongest structural signal
    def _anchor() -> str:
        if _lnd >= 2:
            return "The company shows explicit learning and development focus"
        if _ti_lead >= 2:
            return "Leadership development interest is evident"
        if _onboard >= 2:
            return "The company shows onboarding and employee development activity"
        if _multi >= 2 or _ti_inter >= 2:
            return "The company has a multicultural or cross-cultural workforce"
        if _ti_eng >= 2:
            return "English language training interest is evident"
        if _intl >= 2:
            return "The company has significant international operations"
        # Only use foreign HQ signal when it was NOT sanitized away
        if _fhq >= 2 and not _fhq_san:
            return "The company operates under a foreign parent or external group structure"
        if _ti_neg >= 2:
            return "Sales or negotiation training interest is evident"
        if _growth >= 2:
            return "The company is in a rapid growth phase"
        if _merger >= 2:
            return "The company has recent merger or acquisition activity"
        if _employer >= 2:
            return "The company has visible employer branding and employee-experience focus"
        return "Some mYngle-relevant signals exist but are not yet strong"

    # Second sentence that adds a verifiable hypothesis
    def _second_sentence(hot: bool) -> str:
        if comp_opp == "Strong":
            return (
                "Competitor evidence suggests a possible switch opportunity — "
                "investigate the current provider before leading with a full pitch."
            )
        # True foreign parent (not sanitized)
        if _fhq >= 2 and not _fhq_san:
            return (
                "The foreign parent or external HQ structure may mean HR/L&D or "
                "procurement decisions are made centrally — verify whether this is "
                "the right entry point or whether the Italian entity has autonomy."
            )
        if _intl >= 2:
            return (
                "In the first call, test whether cross-country communication, "
                "onboarding, English, leadership or client-facing communication "
                "is handled centrally or locally."
            )
        if _lnd >= 2 or _ti_lead >= 2:
            return (
                "In the first call, test whether language training or leadership "
                "development is managed centrally by HR/L&D or delegated to individual teams."
            )
        if hot:
            return (
                "In the first call, test whether training needs are coordinated "
                "centrally or handled team by team."
            )
        return (
            "Use Opportunity Radar to look for hiring, expansion, new leadership, "
            "or training initiatives before prioritising a call."
        )

    bh = _buyer_hint()
    anchor = _anchor()
    _needs_domain = str(rd.get("needs_domain_review", "")).lower() in ("true", "1")
    _domain_note = " Note: domain flagged for review - verify the website before calling." if _needs_domain else ""

    if "Hot" in tier:
        p1 = f"Strong Layer 1 fit. {anchor}, making {bh} a relevant entry point."
        p2 = _second_sentence(hot=True)
        return _append_competitor_note(f"{p1} {p2}{_domain_note}", rd)

    if "Warm" in tier:
        p1 = f"Promising but incomplete fit. {anchor}"
        p2 = (
            _second_sentence(hot=False)
            if (comp_opp == "Strong" or (_fhq >= 2 and not _fhq_san))
            else (
                "Use Opportunity Radar to look for hiring, expansion, new leadership, "
                "integration, or training initiatives before prioritising a call."
            )
        )
        return _append_competitor_note(
            f"{p1}, but the training need is not yet explicit. {p2}{_domain_note}", rd
        )

    if "Cool" in tier:
        return _append_competitor_note(
            f"Some mYngle-relevant context exists ({anchor.lower()}), "
            "but the current evidence is still thin. Do not prioritise a cold call yet "
            "unless Opportunity Radar finds a concrete trigger such as expansion, hiring, "
            f"acquisition, leadership change, or a new training initiative.{_domain_note}",
            rd,
        )

    # Pass (or unknown tier)
    return _append_competitor_note(
        "Low Layer 1 priority. No strong mYngle-relevant signal was found yet. "
        f"Only move forward if Opportunity Radar finds a clear current trigger.{_domain_note}",
        rd,
    )


def _append_competitor_note(base_text: str, rd: dict) -> str:
    """Append a competitor signal note to a caller angle string if signals are present."""
    cust_sig  = str(rd.get("competitor_customer_signal", "") or "").strip()
    cust_str  = str(rd.get("competitor_signal_strength", "") or "").strip()
    cust_prov = str(rd.get("competitor_provider_detected", "") or "").strip()
    attn_sig  = str(rd.get("competitor_attention_signal", "") or "").strip()
    attn_prov = str(rd.get("competitor_attention_provider_detected", "") or "").strip()
    attn_url  = str(rd.get("competitor_attention_url", "") or "").strip()

    if cust_sig == "Yes" and cust_str == "High" and cust_prov:
        return (
            base_text
            + f" | COMPETITOR SIGNAL: Strong evidence the company may already use {cust_prov}."
            " Treat as a competitive switch opportunity."
        )
    if attn_sig == "Yes" and attn_prov:
        url_note = f" Source: {attn_url}" if attn_url else ""
        return (
            base_text
            + f" | COMPETITOR MENTION: {attn_prov} found in search results."
            f"{url_note} Verify before outreach."
        )
    return base_text


def _compute_outreach_readiness(row: dict) -> str:
    """Return an outreach readiness label based on tier and domain review status."""
    tier = str(row.get("commercial_tier", "") or "")
    needs_domain = str(row.get("needs_domain_review", "")).lower() in ("true", "1")
    if "Hot" in tier or "Warm" in tier:
        if needs_domain:
            return "Domain review required"
        return "Ready for Opportunity Radar"
    if "Cool" in tier:
        return "Low priority - use Opportunity Radar"
    return "Low priority"


def _build_profile_signals_gaps(rd: dict) -> tuple[str, str]:
    """Build Top Positive Signals and Gaps from actual sig_*/ti_* score values.

    A signal appears in positives if its score >= 2 (Medium or Strong).
    A signal appears in gaps only if its score == 0 AND it is a gap candidate field
    AND it is not already in positives.  No signal ever appears in both lists.
    """
    scored: list[tuple[str, float]] = []
    for field in _SIGNAL_FIELDS_ORDERED:
        try:
            val = float(rd.get(field, 0) or 0)
        except (ValueError, TypeError):
            val = 0.0
        scored.append((field, val))

    # Sort strongest first
    scored.sort(key=lambda x: -x[1])

    positive_fields: set[str] = set()
    positives: list[str] = []
    for field, val in scored:
        if val >= 2:
            label = _SIGNAL_READABLE.get(field, field)
            strength = "Strong" if val >= 3 else "Medium"
            positives.append(f"{label} ({strength})")
            positive_fields.add(field)

    gaps: list[str] = []
    for field in _SIGNAL_FIELDS_ORDERED:
        if field not in _GAP_CANDIDATE_FIELDS:
            continue
        if field in positive_fields:
            continue
        try:
            val = float(rd.get(field, 0) or 0)
        except (ValueError, TypeError):
            val = 0.0
        if val == 0:
            label = _SIGNAL_READABLE.get(field, field)
            gaps.append(f"No clear {label.lower()} found")

    signals_str = "; ".join(positives) if positives else ""
    gaps_str    = "; ".join(gaps) if gaps else "No significant gaps identified"
    return signals_str, gaps_str


def _format_score_drivers(raw: str) -> str:
    """Convert top_score_drivers text to human-readable labels (legacy helper)."""
    if not raw or str(raw).lower() in ("none", "nan", ""):
        return ""
    parts = []
    for token in str(raw).split(";"):
        token = token.strip()
        if not token:
            continue
        field = token.split("=")[0].split("(")[0].strip()
        label = _SIGNAL_READABLE.get(field, field)
        parts.append(label)
    return "; ".join(parts) if parts else ""


def _xl_write_company_profiles(ws, df: pd.DataFrame,
                                name_col: str | None,
                                domain_col: str | None) -> dict:
    """
    Write one formatted block per company.
    Returns {df_row_index: profile_start_row} (0-based index → 1-based Excel row).
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    # Column widths: A=28, B=80, C=14, D=18, E=70, F=75
    for ci, w in enumerate([28, 80, 14, 18, 70, 75], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    tier_fills = {
        "🥇 Hot":  PatternFill(start_color="0B4A92", end_color="0B4A92", fill_type="solid"),
        "🥈 Warm": PatternFill(start_color="1F7AC4", end_color="1F7AC4", fill_type="solid"),
        "🥉 Cool": PatternFill(start_color="E47228", end_color="E47228", fill_type="solid"),
        "❄️ Pass": PatternFill(start_color="7F7F7F", end_color="7F7F7F", fill_type="solid"),
    }
    default_fill = PatternFill(start_color="2B4C7E", end_color="2B4C7E", fill_type="solid")
    label_font  = Font(bold=True,  size=10, color="1A1A1A")
    value_font  = Font(bold=False, size=10, color="1A1A1A")
    header_font = Font(bold=True,  size=12, color="FFFFFF")
    wrap_align  = Alignment(horizontal="left", vertical="top",  wrap_text=True)
    left_align  = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    alt_fill    = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")

    profile_start_rows: dict = {}
    cur = 1

    for df_idx, (_, row) in enumerate(df.iterrows()):
        rd = row.to_dict()
        profile_start_rows[df_idx] = cur

        company = _xl_get(rd, "canonical_company_name",
                          name_col or "", "lusha_company_name", "lusha_api_company_name")
        domain  = _xl_get(rd, "canonical_company_domain",
                          domain_col or "", "lusha_domain", "lusha_api_domain")
        tier    = _xl_get(rd, "commercial_tier")
        try:
            score_val = float(rd.get("final_commercial_fit_score", 0) or 0)
            score_str = f"{score_val:.1f}"
        except (ValueError, TypeError):
            score_str = ""
        industry  = _xl_get(rd, "lusha_industry",       "lusha_api_industry")
        country   = _xl_get(rd, "lusha_country",        "lusha_api_country")
        _is_italy_profile = str(rd.get("scoring_profile", "")) == "italy_register_icp_only"
        employees = (
            "Register filter: 100+ employees"
            if _is_italy_profile
            else _xl_get(rd, "lusha_employee_range", "lusha_api_employee_range")
        )
        why       = _xl_get(rd, "icp_why_relevant")
        # Build signals/gaps from actual sig_*/ti_* scores — single consistent source.
        signals, gaps = _build_profile_signals_gaps(rd)
        evidence  = _xl_get(rd, "icp_evidence")
        interp    = rd.get("caller_angle") or _build_caller_angle(rd)

        hdr_text = company
        if tier:
            hdr_text = f"{company}  [{tier}]"
        if domain:
            hdr_text += f"  —  {domain}"
        fill = tier_fills.get(tier, default_fill)

        # ── Header row (merged A:F) ──────────────────────────────────────────
        ws.cell(row=cur, column=1, value=hdr_text)
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=6)
        for ci in range(1, 7):
            ws.cell(row=cur, column=ci).fill = fill
        ws.cell(row=cur, column=1).font = header_font
        ws.cell(row=cur, column=1).alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )
        ws.row_dimensions[cur].height = 24
        cur += 1

        # ── Score / Tier / Employees row ─────────────────────────────────────
        _compact = [
            (1, "Score",     label_font, right_align),
            (2, score_str,   value_font, left_align),
            (3, "Tier",      label_font, right_align),
            (4, tier,        value_font, left_align),
            (5, "Employees", label_font, right_align),
            (6, employees,   value_font, left_align),
        ]
        for ci, val, fnt, aln in _compact:
            c = ws.cell(row=cur, column=ci, value=val)
            c.font = fnt
            c.alignment = aln
            c.fill = alt_fill
        ws.row_dimensions[cur].height = 20
        cur += 1

        # ── Industry / Country row ────────────────────────────────────────────
        _ic = [
            (1, "Industry", label_font, right_align),
            (2, industry,   value_font, left_align),
            (3, "Country",  label_font, right_align),
            (4, country,    value_font, left_align),
        ]
        for ci, val, fnt, aln in _ic:
            c = ws.cell(row=cur, column=ci, value=val)
            c.font = fnt
            c.alignment = aln
        ws.row_dimensions[cur].height = 20
        cur += 1

        # ── Outreach readiness row ────────────────────────────────────────────
        _ors = _xl_get(rd, "outreach_readiness_status") or _compute_outreach_readiness(rd)
        _needs_dr = str(rd.get("needs_domain_review", "")).lower() in ("true", "1")
        _ors_style = [
            (1, "Outreach Status",    label_font, right_align),
            (2, _ors,                 value_font, left_align),
            (3, "Domain Review",      label_font, right_align),
            (4, "Yes" if _needs_dr else "No", value_font, left_align),
        ]
        for ci, val, fnt, aln in _ors_style:
            c = ws.cell(row=cur, column=ci, value=val)
            c.font = fnt
            c.alignment = aln
        ws.row_dimensions[cur].height = 20
        cur += 1

        # ── Long-text rows ────────────────────────────────────────────────────
        if _is_italy_profile:
            why      = _strip_employee_range_from_text(_check_profile_consistency(rd, why))
            signals  = _strip_employee_range_from_text(_check_profile_consistency(rd, signals))
            gaps     = _strip_employee_range_from_text(_check_profile_consistency(rd, gaps))
            evidence = _strip_employee_range_from_text(_check_profile_consistency(rd, evidence))
            interp   = _strip_employee_range_from_text(_check_profile_consistency(rd, interp))

        long_rows = [
            ("Why Relevant",             why,      55),
            ("Top Positive Signals",     signals,  55),
            ("Gaps / Missing Signals",   gaps,     45),
            ("Evidence",                 evidence, 70),
            ("Caller Angle",              interp,  55),
        ]
        for label, content, height in long_rows:
            lc = ws.cell(row=cur, column=1, value=label)
            lc.font = label_font
            lc.alignment = Alignment(horizontal="left", vertical="top")

            vc = ws.cell(row=cur, column=2, value=content)
            vc.font = value_font
            vc.alignment = wrap_align
            try:
                ws.merge_cells(start_row=cur, start_column=2,
                               end_row=cur, end_column=6)
            except Exception:
                pass
            ws.row_dimensions[cur].height = height
            cur += 1

        # ── Separator ─────────────────────────────────────────────────────────
        ws.row_dimensions[cur].height = 10
        cur += 1

    return profile_start_rows


def _xl_write_summary(ws, df: pd.DataFrame,
                      name_col: str | None,
                      domain_col: str | None,
                      profile_start_rows: dict) -> None:
    """Write the Summary sheet with company scores and Open Profile hyperlinks."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.formatting.rule import DataBarRule
    from openpyxl.utils import get_column_letter

    headers = [
        "Company Name",
        "Company Domain / URL",
        "Domain Check",
        "Mismatch?",
        "Final Commercial Fit Score",
        "Commercial Tier",
        "Outreach Readiness",
        "Open Profile",
    ]
    widths = [30, 35, 14, 10, 24, 16, 22, 18]

    hdr_fill = PatternFill(start_color="0B4A92", end_color="0B4A92", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    link_font = Font(color="0B4A92", underline="single", size=10)
    tier_colors = {
        "🥇 Hot":  "D6E4F7",
        "🥈 Warm": "D9EAD3",
        "🥉 Cool": "FCE5CD",
        "❄️ Pass": "F4CCCC",
    }

    for ci, (hdr, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=ci, value=hdr)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    for df_idx, (_, row) in enumerate(df.iterrows()):
        rd   = row.to_dict()
        xrow = df_idx + 2   # Excel row (1-indexed header + offset)

        company    = _xl_get(rd, "canonical_company_name",
                             name_col or "", "lusha_company_name", "lusha_api_company_name")
        domain     = _xl_get(rd, "canonical_company_domain",
                             domain_col or "", "lusha_domain", "lusha_api_domain")
        dom_conf   = str(rd.get("domain_match_confidence", "") or "").strip()
        mismatch   = str(rd.get("possible_domain_mismatch", "") or "").strip()
        tier       = _xl_get(rd, "commercial_tier")
        try:
            score = float(rd.get("final_commercial_fit_score", "") or "")
        except (ValueError, TypeError):
            score = ""

        row_fill = PatternFill(
            start_color=tier_colors.get(tier, "FFFFFF"),
            end_color=tier_colors.get(tier, "FFFFFF"),
            fill_type="solid",
        ) if tier else None

        _ors = _xl_get(rd, "outreach_readiness_status") or _compute_outreach_readiness(rd)
        for ci, val in enumerate([company, domain, dom_conf, mismatch, score, tier, _ors], 1):
            c = ws.cell(row=xrow, column=ci, value=val)
            if row_fill:
                c.fill = row_fill
            c.alignment = Alignment(horizontal="left", vertical="center")

        # Highlight domain confidence issues (cols 3 and 4)
        if dom_conf in ("Low", "Unknown"):
            ws.cell(row=xrow, column=3).fill = PatternFill(
                start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
            )
        if mismatch == "True":
            ws.cell(row=xrow, column=4).fill = PatternFill(
                start_color="FFD700", end_color="FFD700", fill_type="solid"
            )
        # Highlight "Domain review required" outreach status
        if _ors == "Domain review required":
            ws.cell(row=xrow, column=7).fill = PatternFill(
                start_color="FFD700", end_color="FFD700", fill_type="solid"
            )

        # Open Profile hyperlink — column 8 (shifted by 1 new col)
        prof_start = profile_start_rows.get(df_idx, 1)
        link_target_row = prof_start + 4   # lands near "Why Relevant"
        lc = ws.cell(row=xrow, column=8, value="Open Profile")
        lc.hyperlink = f"#'Company Profiles'!A{link_target_row}"
        lc.font = link_font
        lc.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[xrow].height = 20

    # Blue data bar on score column (E — was C, shifted by 2)
    n = len(df)
    if n > 0:
        try:
            rule = DataBarRule(
                start_type="num", start_value=0,
                end_type="num", end_value=10,
                color="0070C0",
            )
            ws.conditional_formatting.add(f"E2:E{n + 1}", rule)
        except Exception:
            pass


def _xl_write_run_settings(ws, run_config: dict, run_mode: str = "streamlit") -> None:
    """Write run configuration to the Run Settings sheet (hidden).

    Allows CLI vs Streamlit output comparison without guessing run parameters.
    """
    from openpyxl.styles import Font, Alignment

    bold     = Font(bold=True,  size=10)
    norm     = Font(bold=False, size=10)
    hdr_font = Font(bold=True,  size=13)
    lft      = Alignment(horizontal="left", vertical="center")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 55

    rows = [
        ("Run Settings",             None),
        (None,                       None),
        ("run_mode",                 run_mode),
        ("input_filename",           run_config.get("input_filename", "")),
        ("selected_sheet",           run_config.get("selected_sheet", "")),
        ("input_type",               run_config.get("input_type", "")),
        ("scoring_profile",          run_config.get("scoring_profile", "")),
        ("scoring_profile_source",   run_config.get("scoring_profile_source", "")),
        ("run_step1_enrichment",     str(run_config.get("run_step1_enrichment", ""))),
        ("run_step2_enrichment",     str(run_config.get("run_step2_enrichment", ""))),
        ("extract_model_signals",    str(run_config.get("extract_model_signals", ""))),
        ("include_signal_evidence",  str(run_config.get("include_signal_evidence", ""))),
        ("search_provider",          run_config.get("search_provider", "")),
        ("use_playwright",           str(run_config.get("use_playwright", ""))),
        ("model_step1",              run_config.get("model_step1", "")),
        ("model_step2",              run_config.get("model_step2", "")),
    ]

    for r, (label, value) in enumerate(rows, 1):
        if label == "Run Settings":
            ws.cell(row=r, column=1, value=label).font = hdr_font
        elif label is None:
            pass
        else:
            lc = ws.cell(row=r, column=1, value=label)
            lc.font = bold
            lc.alignment = lft
            vc = ws.cell(row=r, column=2, value=value)
            vc.font = norm
            vc.alignment = lft


def _xl_write_advanced_evidence(ws, df: pd.DataFrame) -> None:
    """Write the Advanced Evidence sheet in long format: one row per company per signal.

    Columns: Company Name | Company Domain/URL | Final Score | Commercial Tier |
             Signal Category | Signal Name | Signal Score | Signal Strength |
             Evidence | Caller Angle
    Uses canonical company identity only — no contact-level fields.
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    headers = [
        "Company Name",
        "Company Domain / URL",
        "Final Score",
        "Commercial Tier",
        "Signal Category",
        "Signal Name",
        "Signal Score",
        "Signal Strength",
        "Evidence",
        "Caller Angle",
    ]
    col_widths = [30, 32, 12, 14, 20, 38, 12, 14, 60, 50]

    hdr_fill = PatternFill(start_color="0B4A92", end_color="0B4A92", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    wrap_aln = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ctr_aln  = Alignment(horizontal="center", vertical="top")

    for ci, (hdr, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=ci, value=hdr)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    xrow = 2
    for _, row in df.iterrows():
        rd = row.to_dict()
        company = str(rd.get("canonical_company_name", "") or "").strip()
        domain  = str(rd.get("canonical_company_url",  "") or
                      rd.get("canonical_company_domain", "") or "").strip()
        try:
            score_val = float(rd.get("final_commercial_fit_score", 0) or 0)
            score_str = f"{score_val:.2f}"
        except (ValueError, TypeError):
            score_str = ""
        tier    = str(rd.get("commercial_tier", "") or "").strip()
        interp  = str(rd.get("caller_angle", "") or "").strip() or _build_caller_angle(rd)

        for field in _SIGNAL_FIELDS_ORDERED:
            try:
                sig_score = float(rd.get(field, 0) or 0)
            except (ValueError, TypeError):
                sig_score = 0.0
            strength = _signal_strength_label(sig_score)
            ev_field = field.replace("_score", "_evidence")
            evidence = str(rd.get(ev_field, "") or "").strip()
            label    = _SIGNAL_READABLE.get(field, field)
            category = _SIGNAL_CATEGORY.get(field, "")

            vals = [
                company, domain, score_str, tier,
                category, label, int(sig_score) if sig_score else "",
                strength, evidence, interp,
            ]
            alt = (xrow % 2 == 0)
            row_fill = PatternFill(
                start_color="F7F9FC", end_color="F7F9FC", fill_type="solid"
            ) if alt else None

            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=xrow, column=ci, value=val)
                c.font = Font(size=9)
                c.alignment = wrap_aln if ci in (9, 10) else ctr_aln
                if row_fill:
                    c.fill = row_fill
            ws.row_dimensions[xrow].height = 40 if evidence else 15
            xrow += 1


def _xl_write_opportunity_input(
    ws,
    df: pd.DataFrame,
    name_guess: str | None,
    domain_guess: str | None,
) -> None:
    """Write the flat machine-readable Opportunity Input sheet.

    One row per company, simple column headers, no merged cells.
    Numeric score/probability columns are written as real Excel numbers
    (not strings) so Excel does not show green warning triangles.
    Conditional formatting, frozen header, autofilter, and bold header
    are applied to make the sheet easy to use while remaining machine-readable.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, DataBarRule

    # Priority-ordered candidate lists for each output column.
    # First matching df column wins; missing columns get an empty series.
    COLUMN_MAP: list[tuple[str, list]] = [
        # ── Identity ─────────────────────────────────────────────────────────
        ("company_name",    [name_guess, "canonical_company_name", "Company Name"]),
        ("domain",          [domain_guess, "canonical_company_domain",
                             "canonical_company_url", "Company Domain", "Company Website"]),
        # ── Domain validation ─────────────────────────────────────────────────
        ("input_domain",              ["input_domain"]),
        ("validated_domain",          ["validated_domain"]),
        ("domain_used_for_enrichment",["domain_used_for_enrichment"]),
        ("domain_match_confidence",   ["domain_match_confidence"]),
        ("possible_domain_mismatch",  ["possible_domain_mismatch"]),
        ("suggested_domain",          ["suggested_domain"]),
        ("domain_check_reason",       ["domain_check_reason"]),
        ("domain_source",             ["domain_source"]),
        ("needs_domain_review",       ["needs_domain_review"]),
        ("country",         ["lusha_country", "lusha_api_country", "Company Country", "company_hq_country"]),
        ("city",            ["lusha_city", "lusha_api_city", "Company City"]),
        ("industry",        ["lusha_industry", "lusha_api_industry", "Company Main Industry"]),
        ("employee_range",  [
            "employee_range_for_scoring",     # scoring default if nothing else
            "employee_range_resolved",        # High/Medium resolver output
            "lusha_employee_range",           # back-filled or original Lusha field
            "lusha_api_employee_range",       # live Lusha API (when available)
            "Company Number of Employees",    # raw input column
            "employee_range",                 # generic fallback
            "company_size",
        ]),
        ("employee_range_source",              ["employee_range_source"]),
        ("employee_range_confidence",          ["employee_range_confidence"]),
        ("employee_range_notes",               ["employee_range_notes"]),
        ("employee_range_for_scoring_source",  ["employee_range_for_scoring_source"]),
        ("score_employee_range_source",        ["score_employee_range_source"]),
        ("score_employee_range_confidence",    ["score_employee_range_confidence"]),
        # ── Commercial scoring ────────────────────────────────────────────────
        ("commercial_fit_score",         ["final_commercial_fit_score"]),
        ("commercial_fit_score_75_25_legacy", ["final_commercial_fit_score_75_25_legacy"]),
        ("commercial_tier",              ["commercial_tier"]),
        ("outreach_readiness_status",    ["outreach_readiness_status"]),
        ("model_probability",    ["model_probability", "lean_model_prob"]),
        ("lean_model_prob",      ["lean_model_prob"]),
        ("scoring_notes",        ["scoring_notes"]),
        ("caller_angle",         ["caller_angle"]),
        ("needs_manual_review",  ["needs_manual_review"]),
        ("match_notes",          ["match_notes"]),
        # ── ICP context ───────────────────────────────────────────────────────
        ("icp_lead_score",               ["icp_lead_score"]),
        ("icp_buying_signals",           ["icp_buying_signals"]),
        ("icp_evidence",                 ["icp_evidence"]),
        ("icp_why_relevant",             ["icp_why_relevant"]),
        ("icp_likely_training_interest", ["icp_likely_training_interest"]),
        ("icp_potential_buyer_function", ["icp_potential_buyer_function"]),
        # ── Signal summaries ──────────────────────────────────────────────────
        ("top_positive_signals", ["top_score_drivers"]),   # best available equivalent
        ("gaps_missing_signals", []),                      # no current equivalent — left blank
        # ── Raw signal scores (sig_* and ti_*) ───────────────────────────────
        ("sig_intl_footprint_score",            ["sig_intl_footprint_score"]),
        ("sig_foreign_hq_score",                ["sig_foreign_hq_score"]),
        ("sig_explicit_lnd_score",              ["sig_explicit_lnd_score"]),
        ("sig_multicultural_score",             ["sig_multicultural_score"]),
        ("sig_employer_branding_score",         ["sig_employer_branding_score"]),
        ("sig_rapid_growth_score",              ["sig_rapid_growth_score"]),
        ("sig_merger_acq_score",                ["sig_merger_acq_score"]),
        ("sig_lnd_onboarding_score",            ["sig_lnd_onboarding_score"]),
        ("ti_language_english_score",           ["ti_language_english_score"]),
        ("ti_onboarding_score",                 ["ti_onboarding_score"]),
        ("ti_leadership_score",                 ["ti_leadership_score"]),
        ("ti_broader_professional_score",       ["ti_broader_professional_score"]),
        ("ti_team_collab_score",                ["ti_team_collab_score"]),
        ("ti_intercultural_score",              ["ti_intercultural_score"]),
        ("ti_negotiation_sales_score",          ["ti_negotiation_sales_score"]),
        # ── Priority evidence fields ──────────────────────────────────────────
        ("sig_intl_footprint_evidence",    ["sig_intl_footprint_evidence"]),
        ("sig_foreign_hq_evidence",        ["sig_foreign_hq_evidence"]),
        ("sig_explicit_lnd_evidence",      ["sig_explicit_lnd_evidence"]),
        ("sig_multicultural_evidence",     ["sig_multicultural_evidence"]),
        ("sig_employer_branding_evidence", ["sig_employer_branding_evidence"]),
        ("sig_rapid_growth_evidence",      ["sig_rapid_growth_evidence"]),
        ("sig_merger_acq_evidence",        ["sig_merger_acq_evidence"]),
        ("sig_lnd_onboarding_evidence",    ["sig_lnd_onboarding_evidence"]),
        ("ti_language_english_evidence",   ["ti_language_english_evidence"]),
        ("ti_onboarding_evidence",         ["ti_onboarding_evidence"]),
        ("ti_leadership_evidence",         ["ti_leadership_evidence"]),
        ("ti_intercultural_evidence",      ["ti_intercultural_evidence"]),
        ("ti_negotiation_sales_evidence",  ["ti_negotiation_sales_evidence"]),
        # ── Competitor hard override (customer signal) ────────────────────────
        ("competitor_customer_signal",     ["competitor_customer_signal"]),
        ("competitor_provider_detected",   ["competitor_provider_detected"]),
        ("competitor_signal_strength",     ["competitor_signal_strength"]),
        ("competitor_evidence_url",        ["competitor_evidence_url"]),
        ("competitive_switch_opportunity", ["competitive_switch_opportunity"]),
        ("sales_action_hint",              ["sales_action_hint"]),
        # ── Competitor attention layer (low-threshold) ────────────────────────
        ("competitor_attention_signal",            ["competitor_attention_signal"]),
        ("competitor_attention_provider_detected", ["competitor_attention_provider_detected"]),
        ("competitor_attention_strength",          ["competitor_attention_strength"]),
        ("competitor_attention_type",              ["competitor_attention_type"]),
        ("competitor_attention_evidence",          ["competitor_attention_evidence"]),
        ("competitor_attention_url",               ["competitor_attention_url"]),
        ("competitor_attention_needs_review",      ["competitor_attention_needs_review"]),
        # ── Canonical identity + handoff fields ───────────────────────────────
        ("canonical_company_url",          ["canonical_company_url"]),
        ("company_number",                 ["company_number", "native_company_number",
                                            "register_nummer", "source_row_id"]),
        ("scoring_profile",                ["scoring_profile"]),
        ("inferred_input_country",         ["inferred_input_country"]),
        ("size_scoring_note",              ["size_scoring_note"]),
        # ── Foreign HQ audit ─────────────────────────────────────────────────
        ("foreign_hq_sanitized",           ["foreign_hq_sanitized"]),
        ("foreign_hq_sanitizer_reason",    ["foreign_hq_sanitizer_reason"]),
        ("foreign_hq_original_score",      ["foreign_hq_original_score"]),
        # ── Raw Serper evidence handoff (for Opportunity Radar / caller brief) ──
        ("serper_query_summary",           ["serper_query_summary"]),
        ("serper_source_urls",             ["serper_source_urls"]),
        ("serper_result_titles",           ["serper_result_titles"]),
        ("serper_snippets",                ["serper_snippets"]),
        ("raw_evidence_summary",           ["raw_evidence_summary"]),
        ("evidence_source_urls",           ["evidence_source_urls"]),
        # ── Canonical identity handoff ────────────────────────────────────────
        ("canonical_company_name",         ["canonical_company_name"]),
        ("canonical_company_domain",       ["canonical_company_domain"]),
        ("input_type",                     ["input_type"]),
        ("company_number_canonical",       ["company_number", "native_company_number",
                                            "register_nummer", "source_row_id"]),
        # ── Raw Google evidence (full detail for webapp / Lovable handoff) ────
        ("raw_google_evidence_count",      ["raw_google_evidence_count"]),
        ("raw_google_evidence_urls",       ["raw_google_evidence_urls"]),
        ("raw_google_evidence_combined",   ["raw_google_evidence_combined"]),
        ("raw_google_evidence_json",       ["raw_google_evidence_json"]),
        ("raw_google_evidence_json_01",    ["raw_google_evidence_json_01"]),
        ("raw_google_evidence_json_02",    ["raw_google_evidence_json_02"]),
        ("raw_google_evidence_json_03",    ["raw_google_evidence_json_03"]),
        ("raw_google_evidence_json_parts", ["raw_google_evidence_json_parts"]),
        ("raw_google_evidence_truncated",  ["raw_google_evidence_truncated"]),
        *[
            (f"google_snippet_{i:02d}_{field}", [f"google_snippet_{i:02d}_{field}"])
            for i in range(1, 11)
            for field in ("query_type", "query", "rank", "title", "url", "source_domain", "text")
        ],
    ]

    # Columns that should be written as real numeric values (float/int).
    # Anything whose name ends in _score or _prob, plus the named pairs below.
    _NUMERIC_COLS = {
        "commercial_fit_score", "model_probability", "lean_model_prob",
        "icp_lead_score",
    }
    # Text columns that must never be coerced to numbers even if they look numeric.
    _TEXT_COLS = {
        "company_name", "domain", "country", "city", "industry",
        "employee_range", "commercial_tier", "outreach_readiness_status",
        "scoring_notes", "caller_angle", "match_notes",
        "sales_action_hint", "canonical_company_url", "scoring_profile", "inferred_input_country",
        "size_scoring_note", "competitor_evidence_url", "competitor_provider_detected",
        "icp_buying_signals", "icp_evidence", "icp_why_relevant",
        "icp_likely_training_interest", "icp_potential_buyer_function",
        "top_positive_signals", "gaps_missing_signals",
        "needs_manual_review",
        # domain validation
        "input_domain", "validated_domain", "domain_used_for_enrichment",
        "domain_match_confidence", "possible_domain_mismatch",
        "suggested_domain", "domain_check_reason", "domain_source",
        "needs_domain_review",
        # competitor attention
        "competitor_attention_signal", "competitor_attention_provider_detected",
        "competitor_attention_strength", "competitor_attention_type",
        "competitor_attention_evidence", "competitor_attention_url",
        "competitor_attention_needs_review",
        "competitor_customer_signal", "competitive_switch_opportunity",
        "competitor_signal_strength",
        # Serper evidence handoff
        "serper_query_summary", "serper_source_urls", "serper_result_titles",
        "serper_snippets", "raw_evidence_summary", "evidence_source_urls",
        # Canonical identity
        "canonical_company_name", "canonical_company_domain", "input_type",
        "company_number_canonical",
        # Raw Google evidence
        "raw_google_evidence_urls", "raw_google_evidence_combined",
        "raw_google_evidence_json", "raw_google_evidence_json_01",
        "raw_google_evidence_json_02", "raw_google_evidence_json_03",
        "raw_google_evidence_truncated",
        *[
            f"google_snippet_{i:02d}_{field}"
            for i in range(1, 11)
            for field in ("query_type", "query", "title", "url", "source_domain", "text")
        ],
    }

    def _is_numeric_col(col_name: str) -> bool:
        if col_name in _TEXT_COLS:
            return False
        if col_name in _NUMERIC_COLS:
            return True
        nl = col_name.lower()
        # sig_*_score, ti_*_score, any *_score, *_prob
        if nl.endswith("_score") or nl.endswith("_prob") or nl.endswith("_probability"):
            return True
        return False

    def _coerce_numeric(series: pd.Series) -> pd.Series:
        """Convert a series to float; non-parseable values and blanks become None."""
        result = []
        for v in series:
            if v is None or v == "" or str(v).strip() in ("", "nan", "None", "NaN"):
                result.append(None)
                continue
            try:
                result.append(float(str(v).strip()))
            except (ValueError, TypeError):
                result.append(None)
        return pd.Series(result, dtype=object)

    # Excel number formats per column
    _NUM_FORMATS: dict[str, str] = {
        "commercial_fit_score": "0.00",
        "model_probability":    "0.000",
        "lean_model_prob":      "0.000",
        "icp_lead_score":       "0.00",
    }
    _SCORE_FORMAT = "0.00"
    _PROB_FORMAT  = "0.000"

    def _num_format(col_name: str) -> str:
        if col_name in _NUM_FORMATS:
            return _NUM_FORMATS[col_name]
        nl = col_name.lower()
        if nl.endswith("_prob") or nl.endswith("_probability"):
            return _PROB_FORMAT
        return _SCORE_FORMAT

    # Long text columns that should wrap
    _WRAP_COLS = {
        "scoring_notes", "caller_angle", "match_notes", "icp_buying_signals", "icp_evidence",
        "icp_why_relevant", "icp_likely_training_interest",
        "top_positive_signals", "gaps_missing_signals",
        "domain_check_reason",
        "sig_intl_footprint_evidence", "sig_foreign_hq_evidence",
        "sig_explicit_lnd_evidence", "sig_multicultural_evidence",
        "sig_employer_branding_evidence", "sig_rapid_growth_evidence",
        "sig_merger_acq_evidence", "sig_lnd_onboarding_evidence",
        "ti_language_english_evidence", "ti_onboarding_evidence",
        "ti_leadership_evidence", "ti_intercultural_evidence",
        "ti_negotiation_sales_evidence",
    }

    # Build output DataFrame: one column per entry in COLUMN_MAP
    out: dict = {}
    numeric_flags: dict[str, bool] = {}
    empty_str  = pd.Series([""] * len(df), dtype=str)
    empty_none = pd.Series([None] * len(df), dtype=object)

    for output_col, candidates in COLUMN_MAP:
        is_num = _is_numeric_col(output_col)
        numeric_flags[output_col] = is_num
        series = empty_none if is_num else empty_str
        for cand in candidates:
            if cand and cand in df.columns:
                raw = df[cand]
                if is_num:
                    series = _coerce_numeric(raw)
                else:
                    series = raw.fillna("").astype(str)
                break
        out[output_col] = series

    out_df = pd.DataFrame(out).reset_index(drop=True)

    # ── Write header row ──────────────────────────────────────────────────────
    hdr_fill = PatternFill(start_color="0B4A92", end_color="0B4A92", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    col_letters: dict[str, str] = {}

    for ci, col in enumerate(out_df.columns, 1):
        letter = get_column_letter(ci)
        col_letters[col] = letter
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    # Row background colors keyed by commercial_tier (handles emoji and plain variants)
    _ROW_TIER_FILLS = {
        "🥇 Hot":  PatternFill(start_color="D6E4F7", end_color="D6E4F7", fill_type="solid"),
        "🥈 Warm": PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
        "🥉 Cool": PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid"),
        "❄️ Pass": PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
        "Hot":     PatternFill(start_color="D6E4F7", end_color="D6E4F7", fill_type="solid"),
        "Warm":    PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
        "Cool":    PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid"),
        "Pass":    PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
        "Low":     PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    }

    # ── Write data rows ───────────────────────────────────────────────────────
    n_rows = len(out_df)
    records = out_df.to_dict("records")
    for ri, record in enumerate(records, 2):
        tier_val = str(record.get("commercial_tier", "") or "").strip()
        row_fill = _ROW_TIER_FILLS.get(tier_val)
        for ci, col in enumerate(out_df.columns, 1):
            val = record[col]
            cell = ws.cell(row=ri, column=ci, value=val)
            if row_fill:
                cell.fill = row_fill
            if numeric_flags.get(col) and val is not None:
                cell.number_format = _num_format(col)
            if col in _WRAP_COLS:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ── Column widths ─────────────────────────────────────────────────────────
    _WIDE_COLS = _WRAP_COLS | {
        "scoring_notes", "icp_evidence", "icp_why_relevant", "top_positive_signals",
    }
    for ci, col in enumerate(out_df.columns, 1):
        letter = col_letters[col]
        if col in _WIDE_COLS:
            ws.column_dimensions[letter].width = 45
        elif numeric_flags.get(col):
            ws.column_dimensions[letter].width = 12
        elif col in ("company_name", "domain"):
            ws.column_dimensions[letter].width = 28
        elif col in ("commercial_tier", "outreach_readiness_status",
                     "needs_manual_review", "needs_domain_review",
                     "domain_match_confidence", "possible_domain_mismatch",
                     "domain_used_for_enrichment", "domain_source"):
            ws.column_dimensions[letter].width = 22
        elif col in ("input_domain", "validated_domain", "suggested_domain"):
            ws.column_dimensions[letter].width = 26
        else:
            # Auto-fit based on header name length
            ws.column_dimensions[letter].width = min(max(len(col) + 2, 12), 30)

    # ── Freeze header + autofilter ────────────────────────────────────────────
    ws.freeze_panes = "A2"
    if n_rows > 0:
        last_col_letter = get_column_letter(len(out_df.columns))
        ws.auto_filter.ref = f"A1:{last_col_letter}1"

    # ── Conditional formatting ────────────────────────────────────────────────
    if n_rows == 0:
        return

    data_range_end = n_rows + 1  # last data row (1-indexed, row 1 = header)

    # commercial_fit_score: blue data bar 0–10, matching Lead Scores sheet
    fit_col = col_letters.get("commercial_fit_score")
    if fit_col:
        fit_range = f"{fit_col}2:{fit_col}{data_range_end}"
        try:
            ws.conditional_formatting.add(
                fit_range,
                DataBarRule(
                    start_type="num", start_value=0,
                    end_type="num",   end_value=10,
                    color="0070C0",
                ),
            )
        except Exception:
            pass

    # model_probability / lean_model_prob: color scale
    for prob_col_name in ("model_probability", "lean_model_prob"):
        pcol = col_letters.get(prob_col_name)
        if pcol:
            pr = f"{pcol}2:{pcol}{data_range_end}"
            ws.conditional_formatting.add(
                pr,
                ColorScaleRule(
                    start_type="min", start_color="FFCCCC",
                    mid_type="percentile", mid_value=50, mid_color="FFEB9C",
                    end_type="max", end_color="C6EFCE",
                ),
            )

    # needs_manual_review: TRUE → light orange
    nmr_col = col_letters.get("needs_manual_review")
    if nmr_col:
        nmr_range = f"{nmr_col}2:{nmr_col}{data_range_end}"
        ws.conditional_formatting.add(
            nmr_range,
            CellIsRule(
                operator="equal",
                formula=['"True"'],
                fill=PatternFill(bgColor="FFE0B2", fill_type="solid"),
            ),
        )
        ws.conditional_formatting.add(
            nmr_range,
            CellIsRule(
                operator="equal",
                formula=['"False"'],
                fill=PatternFill(bgColor="E8F5E9", fill_type="solid"),
            ),
        )

    # possible_domain_mismatch: True → amber warning
    pdm_col = col_letters.get("possible_domain_mismatch")
    if pdm_col:
        pdm_range = f"{pdm_col}2:{pdm_col}{data_range_end}"
        ws.conditional_formatting.add(
            pdm_range,
            CellIsRule(
                operator="equal",
                formula=['"True"'],
                fill=PatternFill(bgColor="FFD700", fill_type="solid"),
                font=Font(color="5C3A00", bold=True),
            ),
        )

    # needs_domain_review: True → amber; False → light green
    ndr_col = col_letters.get("needs_domain_review")
    if ndr_col:
        ndr_range = f"{ndr_col}2:{ndr_col}{data_range_end}"
        ws.conditional_formatting.add(
            ndr_range,
            CellIsRule(
                operator="equal",
                formula=['"True"'],
                fill=PatternFill(bgColor="FFD700", fill_type="solid"),
                font=Font(color="5C3A00", bold=True),
            ),
        )
        ws.conditional_formatting.add(
            ndr_range,
            CellIsRule(
                operator="equal",
                formula=['"False"'],
                fill=PatternFill(bgColor="E8F5E9", fill_type="solid"),
            ),
        )

    # outreach_readiness_status: "Domain review required" → amber
    ors_col = col_letters.get("outreach_readiness_status")
    if ors_col:
        ors_range = f"{ors_col}2:{ors_col}{data_range_end}"
        ws.conditional_formatting.add(
            ors_range,
            CellIsRule(
                operator="equal",
                formula=['"Domain review required"'],
                fill=PatternFill(bgColor="FFD700", fill_type="solid"),
                font=Font(color="5C3A00", bold=True),
            ),
        )

    # domain_match_confidence: Low/Unknown → light yellow caution
    dmc_col = col_letters.get("domain_match_confidence")
    if dmc_col:
        dmc_range = f"{dmc_col}2:{dmc_col}{data_range_end}"
        for val in ("Low", "Unknown"):
            ws.conditional_formatting.add(
                dmc_range,
                CellIsRule(
                    operator="equal",
                    formula=[f'"{val}"'],
                    fill=PatternFill(bgColor="FFF2CC", fill_type="solid"),
                ),
            )


def build_rich_excel_bytes(
    df: pd.DataFrame,
    name_col: str | None = None,
    domain_col: str | None = None,
    df_input_original: pd.DataFrame | None = None,
    scoring_profile: str = "default",
    run_config: dict | None = None,
    run_mode: str = "streamlit",
) -> bytes:
    """
    Build a fully formatted multi-sheet Excel workbook.

    Visible sheets  : Lead Scores | Company Profiles | Opportunity Input
    Hidden sheets   : Input | Advanced Evidence | Scoring Settings |
                      Enriched | model_features | qa_evidence

    name_col / domain_col: when provided (e.g. from detect_columns), these are
    used directly so person-level columns in Lucia exports are never chosen.
    df_input_original: original contact-level df to write to the Input sheet
    (for Lucia exports where df is already deduplicated to company level).
    """
    import openpyxl
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)   # discard the default empty sheet

    # Identify original input columns (not added by enrichment pipeline)
    _all_enrich = set(ALL_ENRICHMENT_FIELDS + list(_SCORE_OUTPUT_COLS or []))
    input_cols_list = [c for c in df.columns if c not in _all_enrich]

    # Determine name / domain columns for Summary and Company Profiles.
    # Priority: canonical columns (from normalize_input_to_company_df) > explicit
    # params > Lucia exact columns > heuristic fallback.
    _df_cols = set(df.columns)
    if "canonical_company_name" in _df_cols:
        _name_guess = "canonical_company_name"
    elif name_col is not None:
        _name_guess = name_col
    else:
        _col_set = set(input_cols_list)
        if "Company Name" in _col_set:
            _name_guess = "Company Name"
        elif "lusha_api_company_name" in _col_set:
            _name_guess = "lusha_api_company_name"
        else:
            _name_guess = None
        _person_prefixes = ("first ", "last ", "middle ", "contact ")
        if _name_guess is None:
            for c in input_cols_list:
                cl = c.lower()
                if any(h in cl for h in _COMPANY_HINTS) and not cl.startswith(_person_prefixes):
                    _name_guess = c
                    break
        if _name_guess is None and input_cols_list:
            _name_guess = input_cols_list[0]

    if "canonical_company_domain" in _df_cols:
        _domain_guess = "canonical_company_domain"
    elif domain_col is not None:
        _domain_guess = domain_col
    else:
        _col_set = set(input_cols_list)
        if "Company Domain" in _col_set:
            _domain_guess = "Company Domain"
        elif "Company Website" in _col_set:
            _domain_guess = "Company Website"
        elif "lusha_api_domain" in _col_set:
            _domain_guess = "lusha_api_domain"
        else:
            _domain_guess = None
        if _domain_guess is None:
            for c in input_cols_list:
                cl = c.lower()
                if any(h in cl for h in _DOMAIN_HINTS) and "linkedin" not in cl:
                    _domain_guess = c
                    break
        if _domain_guess is None and len(input_cols_list) > 1:
            _domain_guess = input_cols_list[1]

    # ── Export-time guardrails ────────────────────────────────────────────────
    _export_input_type = (
        str(df["input_type"].iloc[0])
        if "input_type" in df.columns and not df.empty
        else ""
    )
    _is_lucia_export_guard = (_export_input_type == "pre_enriched_lucia_export")
    if _is_lucia_export_guard:
        _guard_bad_domains = [
            str(d) for d in df.get("canonical_company_domain", pd.Series(dtype=str))
            if "linkedin.com" in str(d).lower()
        ]
        _guard_bad_urls = [
            str(u) for u in df.get("canonical_company_url", pd.Series(dtype=str))
            if "linkedin.com/in/" in str(u).lower()
        ]
        _guard_missing = [
            str(n) for n in df.get("canonical_company_name", pd.Series(dtype=str))
            if not str(n).strip() or str(n).strip().lower() in ("nan", "none", "")
        ]
        _guard_errors: list[str] = []
        if _guard_bad_domains:
            _guard_errors.append(f"LinkedIn URLs in canonical_company_domain: {_guard_bad_domains}")
        if _guard_bad_urls:
            _guard_errors.append(f"LinkedIn URLs in canonical_company_url: {_guard_bad_urls}")
        if _guard_missing:
            _guard_errors.append(f"Missing canonical_company_name in {len(_guard_missing)} row(s)")
        if "source_contact_count" not in df.columns:
            _guard_errors.append("source_contact_count column missing for Lucia export")
        if df_input_original is not None and len(df_input_original) == 0:
            _guard_errors.append("Input sheet has 0 rows — original CSV not preserved")
        if _guard_errors:
            raise ValueError(
                "Export guardrail failed — workbook not written:\n" + "\n".join(_guard_errors)
            )

    # ── Input (hidden) — kept for audit; not shown by default ───────────────────
    ws_input = wb.create_sheet("Input")
    if df_input_original is not None:
        # Lucia exports: write the original contact-level CSV unchanged
        _xl_write_df(ws_input, df_input_original)
    else:
        _xl_write_df(ws_input, df[input_cols_list] if input_cols_list else df)
    ws_input.sheet_state = "hidden"

    # ── Compute caller_angle once for all rows ────────────────────────────────
    # _build_caller_angle reads scoring_profile from rd["scoring_profile"].
    if "caller_angle" not in df.columns or df["caller_angle"].astype(str).str.strip().eq("").all():
        df = df.copy()
        df["caller_angle"] = [_build_caller_angle(r) for r in df.to_dict("records")]

    # ── Lead Scores (visible) ─────────────────────────────────────────────────
    ws_lead_scores = wb.create_sheet("Lead Scores")
    profile_rows = None   # filled below after Company Profiles is built

    # ── Company Profiles (visible) ────────────────────────────────────────────
    ws_profiles = wb.create_sheet("Company Profiles")
    profile_rows = _xl_write_company_profiles(
        ws_profiles, df, _name_guess, _domain_guess
    )
    _xl_write_summary(ws_lead_scores, df, _name_guess, _domain_guess, profile_rows)

    # ── Opportunity Input (visible) — flat machine-readable sheet ────────────
    ws_opp_input = wb.create_sheet("Opportunity Input")
    _xl_write_opportunity_input(ws_opp_input, df, _name_guess, _domain_guess)

    # Visible order: Lead Scores → Company Profiles → Opportunity Input
    wb._sheets = [ws_lead_scores, ws_profiles, ws_opp_input, ws_input]

    # ── Advanced Evidence (hidden) — long-format, one row per company per signal ─
    try:
        ws_ev = wb.create_sheet("Advanced Evidence")
        _xl_write_advanced_evidence(ws_ev, df)
        ws_ev.sheet_state = "hidden"
    except Exception:
        pass

    # ── Scoring Settings (hidden) ─────────────────────────────────────────────
    try:
        ws_sc = wb.create_sheet("Scoring Settings")
        _xl_write_scoring_settings(ws_sc, scoring_profile=scoring_profile)
        ws_sc.sheet_state = "hidden"
    except Exception:
        pass

    # ── Run Settings (hidden) ─────────────────────────────────────────────────
    if run_config:
        try:
            ws_rs = wb.create_sheet("Run Settings")
            _xl_write_run_settings(ws_rs, run_config=run_config, run_mode=run_mode)
            ws_rs.sheet_state = "hidden"
        except Exception:
            pass

    # ── Enriched (hidden) ─────────────────────────────────────────────────────
    try:
        ws_en = wb.create_sheet("Enriched")
        _xl_write_df(ws_en, df)
        ws_en.sheet_state = "hidden"
    except Exception:
        pass

    # ── model_features (hidden) ───────────────────────────────────────────────
    try:
        mf = _build_model_features_df(df)
        if not mf.empty:
            ws_mf = wb.create_sheet("model_features")
            _xl_write_df(ws_mf, mf)
            ws_mf.sheet_state = "hidden"
    except Exception:
        pass

    # ── qa_evidence (hidden) ──────────────────────────────────────────────────
    try:
        ev_cols = [c for c in df.columns if c.endswith("_evidence")]
        if ev_cols:
            id_cols = input_cols_list[:3]
            qa_cols = list(dict.fromkeys(id_cols + ev_cols))
            qa_df   = df[[c for c in qa_cols if c in df.columns]]
            ws_qa   = wb.create_sheet("qa_evidence")
            _xl_write_df(ws_qa, qa_df)
            ws_qa.sheet_state = "hidden"
    except Exception:
        pass

    # ── Validate and return bytes ─────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _validate_rich_excel(xl_bytes: bytes) -> dict:
    """
    Reload workbook and verify sheet visibility.
    Returns {"valid": bool, "visible": [...], "hidden": [...], "issues": [...]}.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(xl_bytes))
    except Exception as exc:
        return {"valid": False, "visible": [], "hidden": [], "issues": [str(exc)]}

    visible = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
    hidden  = [ws.title for ws in wb.worksheets if ws.sheet_state != "visible"]
    issues  = []

    expected_visible = {"Lead Scores", "Company Profiles"}
    for s in expected_visible - set(visible):
        issues.append(f"'{s}' should be visible but is missing or hidden")
    for s in set(visible) - expected_visible:
        issues.append(f"'{s}' is visible but should be hidden")

    return {"valid": not issues, "visible": visible, "hidden": hidden, "issues": issues}


# ─────────────────────────────────────────────────────────────────────────────
# Auto-save helpers
# ─────────────────────────────────────────────────────────────────────────────

def autosave_append(row_fields: dict, input_row: pd.Series) -> None:
    """Append one enriched row to the autosave CSV."""
    record = {**input_row.to_dict(), **row_fields}
    df_row = pd.DataFrame([record])
    write_header = not os.path.exists(AUTOSAVE_PATH)
    df_row.to_csv(AUTOSAVE_PATH, mode="a", header=write_header, index=False)


def autosave_load() -> pd.DataFrame | None:
    """Return the autosave DataFrame, or None if it doesn't exist / is unreadable."""
    if not os.path.exists(AUTOSAVE_PATH):
        return None
    try:
        df = pd.read_csv(AUTOSAVE_PATH)
        return df if len(df) > 0 else None
    except Exception:
        return None


def autosave_clear() -> None:
    try:
        os.remove(AUTOSAVE_PATH)
    except FileNotFoundError:
        pass


# ─────────────────────────────────────────────────────────────────────────────
# Simple local Excel autosave  (enrichment_outputs/Enriched results.xlsx)
# ─────────────────────────────────────────────────────────────────────────────

def _xl_autosave_output_dir(folder: str = _XL_AUTOSAVE_DIR) -> Path:
    """Return (and create if needed) the output folder next to the app file."""
    p = Path(__file__).parent / folder
    p.mkdir(parents=True, exist_ok=True)
    return p


def _xl_autosave_path(filename: str) -> Path:
    return _xl_autosave_output_dir() / filename


def _xl_autosave_write(
    df: pd.DataFrame,
    filename: str,
    name_col: str | None = None,
    domain_col: str | None = None,
    scoring_profile: str = "default",
) -> tuple[bool, str]:
    """
    Atomically write df to Excel using the full rich workbook format
    (Lead Scores, Company Profiles, Opportunity Input + hidden sheets).
    Returns (ok, message):
      ok=True  → message is the full path + HH:MM:SS timestamp
      ok=False → message is the human-readable error
    """
    try:
        dst = _xl_autosave_path(filename)
        tmp = dst.with_suffix(".tmp.xlsx")
        xl_bytes = build_rich_excel_bytes(df, name_col=name_col, domain_col=domain_col, scoring_profile=scoring_profile)
        tmp.write_bytes(xl_bytes)
        tmp.replace(dst)
        from datetime import datetime as _dt
        return True, f"{dst}  ({_dt.now().strftime('%H:%M:%S')})"
    except PermissionError:
        return (
            False,
            "Autosave failed because the Excel file may be open. "
            "Close it and the next autosave will try again.",
        )
    except Exception as exc:
        return False, str(exc)


def _xl_should_autosave(processed: int, every_n: int) -> bool:
    return every_n > 0 and processed > 0 and processed % every_n == 0


# ─────────────────────────────────────────────────────────────────────────────
# Per-company run-folder autosave
# ─────────────────────────────────────────────────────────────────────────────

def create_run_folder(base_folder: str, run_tag: str) -> str:
    """
    Create and return the path of a new timestamped run folder.
    Structure:
      {base_folder}/run_{YYYYMMDD_HHMMSS}_{run_tag}/
        rows/
        logs/
    Raises OSError on permission / path errors (caller must catch).
    """
    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    name  = f"run_{stamp}_{run_tag}" if run_tag else f"run_{stamp}"
    run_path = Path(base_folder.strip()) / name
    (run_path / "rows").mkdir(parents=True, exist_ok=True)
    (run_path / "logs").mkdir(parents=True, exist_ok=True)
    (run_path / "cache").mkdir(parents=True, exist_ok=True)
    return str(run_path)


def save_company_result_to_run_folder(
    row_index: int,
    company_name: str,
    input_row: pd.Series,
    enriched_fields: dict,
    debug_record: dict,
    run_dir: str,
) -> tuple:
    """
    Write one processed company as JSON to {run_dir}/rows/row_{NNNN}_{name}.json.
    Returns (success: bool, message: str).
    Does not include API keys or request headers.
    """
    try:
        safe_name = safe_filename(company_name or f"row_{row_index:04d}")
        fname = Path(run_dir) / "rows" / f"row_{row_index:04d}_{safe_name}.json"

        # Sanitise the debug record — strip any raw API response objects but keep
        # scalar metadata; the full raw JSON is already available in the cache files.
        _safe_debug: dict = {}
        for k, v in (debug_record or {}).items():
            if isinstance(v, (str, int, float, bool, type(None))):
                _safe_debug[k] = v
            elif isinstance(v, dict):
                # Truncate large nested dicts (e.g. raw API responses) to key list
                _safe_debug[k] = (
                    v if len(json.dumps(v, default=str)) < 4096
                    else {"_truncated": True, "keys": list(v.keys())}
                )
            else:
                _safe_debug[k] = str(v)[:500]

        record = {
            "row_index":       row_index,
            "company_name":    company_name,
            "saved_at_utc":    datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
            "input_row":       {str(k): str(v) for k, v in input_row.to_dict().items()},
            "enriched_fields": {k: str(v) for k, v in (enriched_fields or {}).items()},
            "debug_record":    _safe_debug,
        }
        fname.write_text(json.dumps(record, ensure_ascii=False, indent=2), encoding="utf-8")
        return True, str(fname)
    except Exception as exc:
        return False, f"Per-company save failed for '{company_name}': {exc}"


def save_partial_outputs_to_run_folder(
    results: list,
    debug_records: list,
    df_work: pd.DataFrame,
    active_fields: list,
    run_dir: str,
    elm_mode: bool = False,
    row_count: int = 0,
    name_col: str | None = None,
    domain_col: str | None = None,
    scoring_profile: str = "default",
) -> tuple:
    """
    Write cumulative files to run_dir:
      latest_results.csv                        — always overwritten (flat)
      latest_results.xlsx                       — always overwritten (full workbook)
      processing_log.csv                        — always overwritten
      enriched_results_partial_NNN_companies.xlsx — every CHECKPOINT_EVERY rows
    Returns (success: bool, message: str).
    """
    try:
        rdir = Path(run_dir)
        partial_df = build_partial_df(results, df_work, active_fields)
        log_df     = make_log_df(debug_records, elm_mode=elm_mode)

        partial_df.to_csv(rdir / "latest_results.csv", index=False, encoding="utf-8-sig")
        log_df.to_csv(rdir / "processing_log.csv",     index=False, encoding="utf-8-sig")

        xl_bytes = build_rich_excel_bytes(partial_df, name_col=name_col, domain_col=domain_col, scoring_profile=scoring_profile)
        (rdir / "latest_results.xlsx").write_bytes(xl_bytes)

        if row_count > 0 and row_count % CHECKPOINT_EVERY == 0:
            ckpt_name = f"enriched_results_partial_{row_count:03d}_companies.xlsx"
            (rdir / ckpt_name).write_bytes(xl_bytes)

        return True, f"{len(results)} rows written to {run_dir}"
    except Exception as exc:
        return False, f"Partial output save failed: {exc}"


def cleanup_old_streamlit_runs(
    base_dir: str = "enrichment_outputs/runs",
    keep_last_runs: int = 5,
    keep_days: int = 7,
    dry_run: bool = True,
    current_run_dir: str = "",
) -> tuple[list[str], list[str]]:
    """List (and optionally delete) old Streamlit run folders.

    Safety rules:
    - Never deletes current_run_dir
    - Never deletes outside base_dir
    - Keeps the most recent keep_last_runs folders
    - Only considers folders older than keep_days days for deletion
    - dry_run=True (default): returns what would be deleted without deleting

    Returns (would_delete: list[str], kept: list[str]).
    """
    import time as _time
    base = Path(__file__).parent / base_dir
    if not base.exists():
        return [], []

    # List only immediate subdirectories of base_dir
    all_runs = sorted(
        [d for d in base.iterdir() if d.is_dir()],
        key=lambda d: d.stat().st_mtime,
        reverse=True,  # newest first
    )

    cutoff_time = _time.time() - keep_days * 86400
    to_delete: list[str] = []
    to_keep: list[str] = []

    for i, run_dir in enumerate(all_runs):
        run_str = str(run_dir.resolve())
        cur_str = str(Path(current_run_dir).resolve()) if current_run_dir else ""

        # Never delete the current run
        if cur_str and run_str == cur_str:
            to_keep.append(run_str)
            continue

        # Keep the most recent keep_last_runs
        if i < keep_last_runs:
            to_keep.append(run_str)
            continue

        # Only delete if older than keep_days
        if run_dir.stat().st_mtime < cutoff_time:
            to_delete.append(run_str)
        else:
            to_keep.append(run_str)

    if not dry_run:
        import shutil as _shutil
        for run_str in to_delete:
            try:
                _shutil.rmtree(run_str)
            except Exception:
                pass

    return to_delete, to_keep


def autosave_already_done(df_saved: pd.DataFrame, name_col: str, domain_col: str | None,
                          company_name: str, raw_url: str) -> bool:
    """Return True if this company already appears in the autosave file."""
    if df_saved is None or df_saved.empty:
        return False
    # Match by domain first (more reliable), fall back to company name
    if domain_col and domain_col in df_saved.columns and raw_url:
        url_clean = clean_domain(raw_url)
        saved_domains = df_saved[domain_col].astype(str).apply(clean_domain)
        if url_clean and (saved_domains == url_clean).any():
            return True
    if name_col in df_saved.columns and company_name:
        saved_names = df_saved[name_col].astype(str).str.strip().str.lower()
        if company_name.lower() in saved_names.values:
            return True
    return False


# ─────────────────────────────────────────────────────────────────────────────
# Session-state helpers
# ─────────────────────────────────────────────────────────────────────────────

def ss(key, default=None):
    if is_cli_mode():
        return default
    _st, _ = get_streamlit()
    return _st.session_state.get(key, default)


def ss_set(**kwargs):
    if is_cli_mode():
        return
    _st, _ = get_streamlit()
    for k, v in kwargs.items():
        _st.session_state[k] = v


def reset_processing(clear_autosave: bool = False):
    if clear_autosave:
        autosave_clear()
    ss_set(
        processing=False, stop_requested=False,
        process_index=0, results=[], debug_records=[],
        enrichment_done=False, df_enriched=None,
        total_tokens_in=0, total_tokens_out=0, total_cost_usd=0.0,
        total_cache_read_tokens=0, total_cache_create_tokens=0,
        autosave_last_name="",
        _jina_retry_count=0, _last_retry_msg="",
        _final_auto_saved=False, _last_local_save="",
        _auto_dl_count=0, _auto_dl_last_msg="",
        _step2_debug_log="", _step2_prompt_records=[],
        _dry_run_records=[], _search_output_records=[], _step2_debug_files=[],
        _zero_cost_preview=False, _dry_run_preview_count=0,
        _per_company_autosave_run_dir="",
        _per_company_autosave_last_saved="",
        _per_company_autosave_last_error="",
        _final_save_path="", _final_save_error="",
        _xl_autosave_last_msg="", _xl_autosave_final_done=False,
    )


def _detect_italy_register_profile(fname: str, df: "pd.DataFrame | None" = None) -> bool:
    """Return True when strong indicators suggest this is an Italian register batch.

    Conservative: only matches clear Italy-register filename patterns or cleaner output
    with Italian country context.  Never triggers on Lucia/Lusha exports or generic lists.
    """
    fl = (fname or "").lower()
    # Filename patterns that indicate Italian register input
    _italy_patterns = (
        "italy100", "italy200", "italy200plus",
    )
    if any(p in fl for p in _italy_patterns):
        return True
    # Italy + cleaned suffix (e.g. Italy100_1_R0001_0500_cleaned_20260613.xlsx)
    if "italy" in fl and "cleaned" in fl:
        return True
    # Cleaner output with Italian country_code column populated with IT
    if df is not None and "country_code" in df.columns:
        _codes = df["country_code"].dropna().astype(str).str.strip().str.upper()
        if (_codes == "IT").any() and not (_codes == "DE").any():
            return True
    return False


def resolve_active_scoring_profile(
    input_filename: str = "",
    df: "pd.DataFrame | None" = None,
    selected_sheet: str = "",
    detected_input_type: str = "",
    user_override: str = "auto",
) -> str:
    """Single source of truth for scoring profile selection.

    Priority:
    1. user_override != "auto"  → return it directly
    2. Italy register detected   → "italy_register_icp_only"
    3. Otherwise                 → "default"
    """
    if user_override and user_override != "auto":
        return user_override
    if _detect_italy_register_profile(input_filename, df):
        return "italy_register_icp_only"
    return "default"


def resolve_active_run_config(
    input_filename: str = "",
    df: "pd.DataFrame | None" = None,
    selected_sheet: str = "",
    detected_input_type: str = "",
    has_existing_lusha_data: bool = False,
    user_overrides: dict | None = None,
) -> dict:
    """Single source of truth for all per-run enrichment parameters.

    Call once at run start; store the result in session_state._active_run_config.
    Use it unchanged for every enrich_one_row() call in that run.

    Step 1 (Jina/Firecrawl firmographic scrape) is skipped ONLY for true
    pre-enriched Lucia/Lusha exports where company-level data is already present.
    Input cleaner output must always run Step 1 regardless of what columns it has.
    """
    overrides = user_overrides or {}

    scoring_profile = resolve_active_scoring_profile(
        input_filename=input_filename,
        df=df,
        selected_sheet=selected_sheet,
        detected_input_type=detected_input_type,
        user_override=overrides.get("scoring_profile", "auto"),
    )
    sp_source = (
        "manual_override"
        if (overrides.get("scoring_profile") and overrides.get("scoring_profile") != "auto")
        else "auto_detected"
    )

    # Step 1 skipped ONLY for true pre-enriched Lucia/Lusha exports.
    # Cleaner output (detected_input_type == "input_cleaner_output") must always
    # run Step 1 — never treat it as pre-enriched even if it has Lusha-like cols.
    _is_lucia = (detected_input_type == "pre_enriched_lucia_export")
    _skip_s1_default = _is_lucia and has_existing_lusha_data
    run_step1 = bool(overrides.get("run_step1_enrichment", not _skip_s1_default))

    return {
        "scoring_profile":          scoring_profile,
        "scoring_profile_source":   sp_source,
        "run_step1_enrichment":     run_step1,
        "run_step2_enrichment":     bool(overrides.get("run_step2_enrichment", True)),
        "extract_model_signals":    bool(overrides.get("extract_model_signals", True)),
        "include_signal_evidence":  bool(overrides.get("include_signal_evidence", True)),
        "search_provider":          overrides.get("search_provider", STEP2_PROVIDER_SERPER),
        "use_playwright":           bool(overrides.get("use_playwright", _PLAYWRIGHT_AVAILABLE)),
        "model_step1":              overrides.get("model_step1", MODEL_STEP1),
        "model_step2":              overrides.get("model_step2", MODEL_STEP2),
        "input_type":               detected_input_type,
        "selected_sheet":           selected_sheet,
        "input_filename":           input_filename,
    }


def apply_results_compatible_scoring(
    df: pd.DataFrame,
    scoring_profile: str = "default",
) -> pd.DataFrame:
    """Single source of truth for scoring — always routes through Results(8).xlsx formula.

    scoring_profile: "default" | "italy_register_icp_only"

    Runs a dataframe-level foreign HQ hygiene pass (with the correct scoring profile)
    before scoring, so sanitized sig_foreign_hq_score values feed into the final score,
    top_score_drivers, and scoring_notes.
    """
    from commercial_fit_scoring import score_dataframe as _cfs_score_df

    # ── Foreign HQ hygiene — df-level pass with correct scoring profile ───────
    # enrich_one_row calls sanitize_foreign_hq_signal with _get_input_country(row)
    # (no scoring_profile arg), so Italy register rows get input_country="" and Rule A
    # never fires.  Re-run here where the active scoring_profile is known.
    _fhq_sanitize_cols = (
        "sig_foreign_hq_score", "foreign_hq_sanitized", "foreign_hq_sanitizer_reason",
        "foreign_hq_original_score", "foreign_hq_original_evidence", "inferred_input_country",
    )
    _records = df.to_dict("records")
    for _row in _records:
        _ctry = _get_input_country(_row, scoring_profile)
        sanitize_foreign_hq_signal(_row, _ctry)
    for _col in _fhq_sanitize_cols:
        df[_col] = [_r.get(_col, "") for _r in _records]

    df = _cfs_score_df(df, scoring_profile=scoring_profile)

    # Compute outreach_readiness_status after scoring (needs commercial_tier)
    _ors_records = df.to_dict("records")
    df["outreach_readiness_status"] = [_compute_outreach_readiness(r) for r in _ors_records]

    return df


def build_and_finish(results: list, debug_records: list, df_work: pd.DataFrame,
                     field_list: list | None = None) -> None:
    flist       = field_list if field_list is not None else ALL_ENRICHMENT_FIELDS
    df_out      = df_work.head(len(results)).copy().reset_index(drop=True)
    enriched_df = pd.DataFrame(results)
    for col in flist:
        df_out[col] = enriched_df[col].values if col in enriched_df.columns else ""

    # ── Employee range resolver — runs before scoring ─────────────────────────
    # Priority: existing fields → text parsing → Serper fallback → scoring default.
    # Only High/Medium confidence results are back-filled into lusha_employee_range.
    # The scoring default (51-200) is written only to employee_range_for_scoring,
    # never to lusha_employee_range.
    def _is_blank_val(v) -> bool:
        return v is None or str(v).strip() in ("", "nan", "None", "N/A", "-")

    _serper_key_for_er: str = ss("_serper_key", "") or ""
    _er_records  = df_out.to_dict("records")
    _er_results: list[dict] = []
    for _rec in _er_records:
        _cname  = str(_rec.get("lusha_company_name") or _rec.get("company_name") or "")
        _domain = str(_rec.get("canonical_company_domain") or _rec.get("domain") or "")
        _er = resolve_employee_range(_rec, company_name=_cname)

        # Serper fallback when local resolution produced nothing or only Low confidence
        if _er["employee_range_confidence"] in ("None", "Low") and _serper_key_for_er:
            _er_s = resolve_employee_range_from_serper(_cname, _domain, _serper_key_for_er)
            if _er_s.get("employee_range_resolved"):
                _er = _er_s

        # Populate employee_range_for_scoring and its source
        if _er.get("employee_range_resolved") and _er.get("employee_range_confidence") in ("High", "Medium"):
            _er["employee_range_for_scoring"]        = _er["employee_range_resolved"]
            _er["employee_range_for_scoring_source"] = _er["employee_range_source"]
        else:
            # Scoring-only default — not a found fact
            _er["employee_range_for_scoring"]        = DEFAULT_EMPLOYEE_RANGE_FOR_SCORING
            _er["employee_range_for_scoring_source"] = "default_commercial_minimum_assumption"
            if not _er.get("employee_range_notes"):
                _er["employee_range_notes"] = (
                    "No employee count found; default commercial minimum range used for scoring only."
                )
            if _er["employee_range_confidence"] in ("None",):
                _er["employee_range_confidence"] = "Low"

        _er_results.append(_er)

    for col in EMPLOYEE_RANGE_RESOLVER_FIELDS:
        df_out[col] = [r.get(col, "") for r in _er_results]

    # Back-fill lusha_employee_range with High/Medium resolved values only
    _resolved_vals  = [r.get("employee_range_resolved", "") for r in _er_results]
    _resolved_confs = [r.get("employee_range_confidence", "None") for r in _er_results]
    if "lusha_employee_range" not in df_out.columns:
        df_out["lusha_employee_range"] = ""
    _existing_lusha = df_out["lusha_employee_range"].tolist()
    df_out["lusha_employee_range"] = [
        _resolved_vals[i]
        if (_is_blank_val(_existing_lusha[i])
            and _resolved_vals[i]
            and _resolved_confs[i] in ("High", "Medium"))
        else _existing_lusha[i]
        for i in range(len(df_out))
    ]

    if not ss("_elm_mode", False):
        _baf_scoring_profile = ss("_scoring_profile", "default")
        try:
            df_out = apply_results_compatible_scoring(df_out, _baf_scoring_profile)
        except Exception:
            pass
        try:
            df_out = apply_competitor_icp_override(df_out)
        except Exception:
            pass
    ss_set(
        processing=False, stop_requested=False,
        enrichment_done=True, df_enriched=df_out,
        debug_records=debug_records,
    )
    if not is_cli_mode():
        _st, _ = get_streamlit()
        _st.rerun()


def _smoke_test_employee_range_resolver() -> None:
    """
    Offline smoke test for the employee range resolver pipeline.
    Run with:
        python -c "import enrich_clients_claude as e; e._smoke_test_employee_range_resolver()"
    """
    PASS = "\033[92m✓\033[0m"
    FAIL = "\033[91m✗\033[0m"
    failures: list[str] = []

    def chk(label: str, ok: bool, detail: str = "") -> None:
        if ok:
            print(f"  {PASS}  {label}")
        else:
            failures.append(label)
            print(f"  {FAIL}  {label}" + (f"  [{detail}]" if detail else ""))

    print("\n=== Employee range resolver smoke test ===\n")

    # ── Row 1: explicit employee count in icp_evidence ────────────────────────
    print("Case 1: explicit count in icp_evidence")
    row1 = {
        "company_name":           "Example Company",
        "lusha_api_employee_range": "",
        "lusha_employee_range":   "",
        "employee_range":         "",
        "company_size":           "",
        "Company Number of Employees": "",
        "icp_evidence": "The company has 500 employees and operates internationally.",
    }
    r1 = resolve_employee_range(row1, company_name="Example Company")
    chk("employee_range_resolved = '201 - 500'",
        r1["employee_range_resolved"] == "201 - 500",
        r1["employee_range_resolved"])
    chk("confidence is High or Medium",
        r1["employee_range_confidence"] in ("High", "Medium"),
        r1["employee_range_confidence"])

    # Simulate the for_scoring / backfill logic
    er_for_scoring1 = r1["employee_range_resolved"] if r1["employee_range_confidence"] in ("High", "Medium") else DEFAULT_EMPLOYEE_RANGE_FOR_SCORING
    chk("employee_range_for_scoring = resolved value",
        er_for_scoring1 == "201 - 500",
        er_for_scoring1)

    existing_lusha1 = ""
    def _is_blank_val(v):
        return v is None or str(v).strip() in ("", "nan", "None", "N/A", "-")
    backfilled1 = r1["employee_range_resolved"] if (
        _is_blank_val(existing_lusha1)
        and r1["employee_range_resolved"]
        and r1["employee_range_confidence"] in ("High", "Medium")
    ) else existing_lusha1
    chk("lusha_employee_range backfilled",
        backfilled1 == "201 - 500",
        backfilled1)

    # ── Row 2: no employee evidence at all ────────────────────────────────────
    print("\nCase 2: no employee evidence")
    row2 = {
        "company_name":           "Unknown Corp",
        "lusha_api_employee_range": "",
        "lusha_employee_range":   "",
        "employee_range":         "",
        "icp_evidence":           "They sell widgets and have an office in Milan.",
    }
    r2 = resolve_employee_range(row2, company_name="Unknown Corp")
    chk("employee_range_resolved is blank",
        r2["employee_range_resolved"] == "",
        r2["employee_range_resolved"])

    er_for_scoring2 = r2["employee_range_resolved"] if r2["employee_range_confidence"] in ("High", "Medium") else DEFAULT_EMPLOYEE_RANGE_FOR_SCORING
    chk("employee_range_for_scoring = default '51 - 200'",
        er_for_scoring2 == DEFAULT_EMPLOYEE_RANGE_FOR_SCORING,
        er_for_scoring2)

    existing_lusha2 = ""
    backfilled2 = r2["employee_range_resolved"] if (
        _is_blank_val(existing_lusha2)
        and r2["employee_range_resolved"]
        and r2["employee_range_confidence"] in ("High", "Medium")
    ) else existing_lusha2
    chk("lusha_employee_range NOT backfilled from default",
        backfilled2 == "",
        repr(backfilled2))

    print()
    if failures:
        print(f"FAILED: {len(failures)} check(s) - {failures}")
    else:
        print("All checks passed.")
    print()


def _validate_type1_type2_pipeline() -> None:
    """Offline smoke test — run with:
        python -c "import enrich_clients_claude as e; e._validate_type1_type2_pipeline()"

    Checks normalization, canonical columns, scoring, and that build_and_finish
    routes through apply_results_compatible_scoring (not the old _score_dataframe).
    """
    import sys
    from pathlib import Path
    import pandas as pd
    from commercial_fit_scoring import score_company

    PASS = "\033[92m✓\033[0m"
    FAIL = "\033[91m✗\033[0m"
    failures: list[str] = []

    def chk(label: str, ok: bool, detail: str = "") -> None:
        if ok:
            print(f"  {PASS}  {label}")
        else:
            failures.append(label)
            print(f"  {FAIL}  {label}" + (f"  [{detail}]" if detail else ""))

    print("\n=== Pipeline validation ===\n")

    # ── Type 1 ──
    print("Type 1: Capgemini synthetic row")
    t1_df = pd.DataFrame([{
        "Company": "Capgemini Nederland B.V.",
        "Website": "https://www.capgemini.com/",
    }])
    r1 = normalize_input_to_company_df(t1_df, "simple_company_list", "Company", "Website")
    chk("T1 input_type = simple_company_list", r1["input_type"] == "simple_company_list")
    chk("T1 company_df 1 row",               len(r1["company_df"]) == 1)
    chk("T1 canonical_company_name",
        r1["company_df"].iloc[0]["canonical_company_name"] == "Capgemini Nederland B.V.",
        repr(r1["company_df"].iloc[0]["canonical_company_name"]))
    chk("T1 canonical_company_domain = capgemini.com",
        r1["company_df"].iloc[0]["canonical_company_domain"] == "capgemini.com",
        repr(r1["company_df"].iloc[0]["canonical_company_domain"]))

    # Scoring
    cap_signals = {
        "sig_foreign_hq_score": 3, "sig_explicit_lnd_score": 3,
        "sig_intl_footprint_score": 3, "sig_employer_branding_score": 2,
        "sig_lnd_onboarding_score": 2, "ti_onboarding_score": 2,
        "sig_rapid_growth_score": 1, "lusha_api_employee_range": "100001 - 10000000",
    }
    rs = score_company(cap_signals)
    chk("T1 lean_model_prob ≈ 0.7285", abs(rs["lean_model_prob"] - 0.7285) < 0.001,
        str(round(rs["lean_model_prob"], 4)))
    chk("T1 final ≈ 9.54",             abs(rs["final_commercial_fit_score"] - 9.54) < 0.25,
        str(rs["final_commercial_fit_score"]))
    chk("T1 tier = 🥇 Hot",            rs["commercial_tier"] == "🥇 Hot",
        rs["commercial_tier"])
    chk("T1 NOT 9.99",                 abs(rs["final_commercial_fit_score"] - 9.99) > 0.1)
    chk("T1 NOT Tier 1",               rs["commercial_tier"] != "Tier 1")
    chk("T1 model_probability == lean_model_prob",
        abs(rs["model_probability"] - rs["lean_model_prob"]) < 0.001,
        f"{rs['model_probability']} vs {rs['lean_model_prob']}")

    # ── Type 2 ──
    print("\nType 2: Cold Caller CSV fixture")
    csv_path = Path(__file__).parent / "Example_Cold_Caller.csv"
    if not csv_path.exists():
        csv_path = Path(__file__).parent / "test.csv"
    assert csv_path.exists(), f"Fixture not found: {csv_path}"
    t2_df = pd.read_csv(csv_path)

    chk("T2 is_lucia_contact_export = True", is_lucia_contact_export(t2_df))
    nc, dc = detect_columns(t2_df)
    chk("T2 detect_columns → Company Name",   nc == "Company Name", repr(nc))
    chk("T2 detect_columns → Company Domain", dc == "Company Domain", repr(dc))

    r2 = normalize_input_to_company_df(t2_df, "pre_enriched_lucia_export",
                                       "Company Name", "Company Domain")
    chk("T2 input_type = pre_enriched_lucia_export",
        r2["input_type"] == "pre_enriched_lucia_export")
    chk("T2 contact_row_count = 3",     r2["contact_row_count"] == 3,
        str(r2["contact_row_count"]))
    chk("T2 unique_company_count = 3",  r2["unique_company_count"] == 3,
        str(r2["unique_company_count"]))
    chk("T2 company_df 3 rows",         len(r2["company_df"]) == 3,
        str(len(r2["company_df"])))

    cdf2 = r2["company_df"]
    names2   = cdf2["canonical_company_name"].tolist()
    domains2 = [clean_domain(str(d)) for d in cdf2["canonical_company_domain"].tolist()]

    for expected_name in ["Ali Lavoro", "Renovit", "S&you Italia"]:
        chk(f"T2 canonical_company_name contains '{expected_name}'",
            expected_name in names2, str(names2))
    for expected_dom in ["alilavoro.it", "renovit.it", "sandyou.it"]:
        chk(f"T2 canonical_company_domain contains '{expected_dom}'",
            expected_dom in domains2, str(domains2))

    row0 = cdf2.iloc[0].to_dict()
    chk("T2 lucia_api_called = False",
        str(row0.get("lucia_api_called", "")) == "False",
        repr(row0.get("lucia_api_called")))
    chk("T2 lusha_api_status = reused_existing_lucia_data",
        row0.get("lusha_api_status") == "reused_existing_lucia_data",
        repr(row0.get("lusha_api_status")))
    chk("T2 source_contact_count present",
        "source_contact_count" in row0)

    for pn in ("Anna", "Michela", "Annalisa"):
        chk(f"T2 '{pn}' NOT in canonical_company_name", pn not in names2)
    for d in cdf2["canonical_company_domain"].astype(str):
        chk(f"T2 no linkedin.com in canonical_company_domain: {d!r}",
            "linkedin.com" not in d.lower(), d)
    for u in cdf2["canonical_company_url"].astype(str):
        chk(f"T2 no linkedin.com/in/ in canonical_company_url: {u!r}",
            "linkedin.com/in/" not in u.lower(), u)

    # ── Scoring shared engine ──
    print("\nShared scoring engine")
    chk("apply_results_compatible_scoring is callable",
        callable(apply_results_compatible_scoring))
    # Confirm build_and_finish does NOT call _score_dataframe directly
    import inspect
    baf_src = inspect.getsource(build_and_finish)
    chk("build_and_finish does NOT call _score_dataframe",
        "_score_dataframe" not in baf_src, "old _score_dataframe still present")
    chk("build_and_finish calls apply_results_compatible_scoring",
        "apply_results_compatible_scoring" in baf_src)

    # ── Filename convention ──
    print("\nFilename convention")
    from datetime import datetime
    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    for prefix in ("enrichedResults_", "enrichedResults_partial_", "enrichedResults_snapshot_"):
        fname = f"{prefix}{stamp}.xlsx"
        chk(f"'{prefix}...' starts with enrichedResults_", fname.startswith("enrichedResults_"))
        for bad in ("enriched_results", "_sg_", "_hq_", "lusha", "lucia"):
            chk(f"  no '{bad}' in '{prefix}...'", bad not in fname.lower())

    # ── build_enriched_output_filename ──
    print("\nbuild_enriched_output_filename")
    _fixed_stamp = "20260617_133905"
    _fn_cases = [
        (
            "Italy200_01_R0001_0500_cleaned_20260613_1301.xlsx",
            "Italy200_01_R0001_0500_enriched_20260617_133905.xlsx",
        ),
        (
            "Italy200_15_R7001_7270_cleaned_20260613_2309.xlsx",
            "Italy200_15_R7001_7270_enriched_20260617_133905.xlsx",
        ),
        (
            "Germany_1_R0001_0500_cleaned_20260613_1613.xlsx",
            "Germany_1_R0001_0500_enriched_20260617_133905.xlsx",
        ),
        (
            "some_input.xlsx",
            "some_input_enriched_20260617_133905.xlsx",
        ),
    ]
    for _inp, _expected in _fn_cases:
        _got = build_enriched_output_filename(_inp, run_stamp=_fixed_stamp)
        chk(
            f"  {_inp} -> {_expected}",
            _got == _expected,
            f"got {_got!r}",
        )

    # ── Input Cleaner output detection regression tests (Part H) ─────────────
    print("\nInput Cleaner output detection")

    # H-T1: Generic register_cleaned output — Best Guess Input sheet
    _h_bgi_df = pd.DataFrame([
        {"company_name": f"Company {i}", "website_url": f"co{i}.de",
         "final_selected_domain": f"co{i}.de", "organization_type": "GmbH",
         "myngle_target_eligibility": "Yes"}
        for i in range(1, 5)
    ])
    chk("H-T1 is_input_cleaner_output (col sig)",
        is_input_cleaner_output(_h_bgi_df, "register_cleaned_20260613.xlsx"))
    _h_nc, _h_dc = detect_columns(_h_bgi_df, "register_cleaned_20260613.xlsx")
    chk("H-T1 name_col = company_name",    _h_nc == "company_name",   repr(_h_nc))
    chk("H-T1 domain_col = website_url",   _h_dc == "website_url",    repr(_h_dc))
    _h_r1 = normalize_input_to_company_df(_h_bgi_df, "input_cleaner_output",
                                          "company_name", "website_url")
    chk("H-T1 input_type = input_cleaner_output",
        _h_r1["input_type"] == "input_cleaner_output")
    chk("H-T1 rows_loaded = 4",   _h_r1["contact_row_count"] == 4, str(_h_r1["contact_row_count"]))
    chk("H-T1 unique_company_count = 4",
        _h_r1["unique_company_count"] == 4, str(_h_r1["unique_company_count"]))

    # H-T2: Domain fallback — website_url blank, final_selected_domain filled
    _h_fallback_df = pd.DataFrame([{
        "company_name": "Mustermann GmbH",
        "website_url": "",
        "final_selected_domain": "mustermann.de",
        "validated_domain": "",
    }])
    _h_r2 = normalize_input_to_company_df(_h_fallback_df, "input_cleaner_output",
                                          "company_name", "website_url")
    chk("H-T2 domain fallback to final_selected_domain",
        _h_r2["company_df"].iloc[0]["canonical_company_domain"] == "mustermann.de",
        repr(_h_r2["company_df"].iloc[0]["canonical_company_domain"]))

    # H-T3: Metadata trap — organization_type must never be company-name col
    _h_meta_df = pd.DataFrame([{
        "organization_type": "GmbH", "company_name": "Test GmbH",
        "website_url": "test.de",
    }])
    _h_nc3, _ = detect_columns(_h_meta_df, "")
    chk("H-T3 organization_type not selected as name col",
        _h_nc3 != "organization_type", repr(_h_nc3))

    # H-T4: German raw batch — company_name_clean, no domain
    _h_de_raw_df = pd.DataFrame([{
        "company_name_clean": "Muster GmbH", "company_name_raw": "MUSTER GMBH",
        "city_or_registered_office": "Berlin", "federal_state": "Berlin",
        "registered_address": "Musterstr. 1", "legal_form_detected": "GmbH",
    }])
    chk("H-T4 German raw NOT cleaner output",
        not is_input_cleaner_output(_h_de_raw_df, "Germany_1_R0001_0500.xlsx"))
    _h_nc4, _h_dc4 = detect_columns(_h_de_raw_df, "Germany_1_R0001_0500.xlsx")
    chk("H-T4 name_col = company_name_clean", _h_nc4 == "company_name_clean", repr(_h_nc4))
    chk("H-T4 domain_col = None",             _h_dc4 is None,                 repr(_h_dc4))

    # H-T5: German cleaned output
    _h_de_cleaned_df = pd.DataFrame([{
        "company_name": "Muster GmbH", "company_name_clean": "Muster GmbH",
        "website_url": "muster.de", "final_selected_domain": "muster.de",
        "validated_domain": "muster.de", "organization_type": "GmbH",
    }])
    chk("H-T5 German cleaned IS cleaner output",
        is_input_cleaner_output(_h_de_cleaned_df, "Germany_1_R0001_0500_cleaned.xlsx"))
    _h_nc5, _h_dc5 = detect_columns(_h_de_cleaned_df, "Germany_1_R0001_0500_cleaned.xlsx")
    chk("H-T5 name_col = company_name",   _h_nc5 == "company_name",  repr(_h_nc5))
    chk("H-T5 domain_col = website_url",  _h_dc5 == "website_url",   repr(_h_dc5))

    # H-T6: Lucia/Lusha export — existing behavior preserved
    csv_path2 = Path(__file__).parent / "Example_Cold_Caller.csv"
    if csv_path2.exists():
        _h_lucia_df = pd.read_csv(csv_path2)
        chk("H-T6 Lucia is_lucia_contact_export",     is_lucia_contact_export(_h_lucia_df))
        chk("H-T6 Lucia NOT is_input_cleaner_output", not is_input_cleaner_output(_h_lucia_df))
        _h_nc6, _h_dc6 = detect_columns(_h_lucia_df)
        chk("H-T6 Lucia name_col = Company Name",   _h_nc6 == "Company Name",   repr(_h_nc6))
        chk("H-T6 Lucia domain_col = Company Domain", _h_dc6 == "Company Domain", repr(_h_dc6))
    else:
        print("  [skip] H-T6: Example_Cold_Caller.csv not found, skipping Lucia test")

    # ── Company-number resolution (CLI progress) ──────────────────────────────
    print("\nCompany number resolution")
    _h_cn_row = {
        "company_number": "DE-HRB-12345", "company_name": "EUROPE Hotels GmbH",
        "website_url": "europehotels.de",
    }
    chk("H-CN1 resolves company_number", _get_company_number(_h_cn_row) == "DE-HRB-12345",
        repr(_get_company_number(_h_cn_row)))

    _h_rn_row = {
        "register_nummer": "HRB 99999", "company_name": "Test GmbH",
    }
    chk("H-CN2 resolves register_nummer", _get_company_number(_h_rn_row) == "HRB 99999",
        repr(_get_company_number(_h_rn_row)))

    _h_no_cn_row = {"company_name": "NoCN AG", "website": "nocn.de"}
    chk("H-CN3 missing company_number returns n/a",
        _get_company_number(_h_no_cn_row) == "n/a",
        repr(_get_company_number(_h_no_cn_row)))

    _h_cn_df = pd.DataFrame([_h_cn_row])
    chk("H-CN4 _detect_company_number_col finds company_number",
        _detect_company_number_col(_h_cn_df) == "company_number",
        repr(_detect_company_number_col(_h_cn_df)))

    _h_no_cn_df = pd.DataFrame([_h_no_cn_row])
    chk("H-CN5 _detect_company_number_col returns None when absent",
        _detect_company_number_col(_h_no_cn_df) is None,
        repr(_detect_company_number_col(_h_no_cn_df)))

    # ── Italy register scoring profile ────────────────────────────────────────
    print("\nItaly register scoring profile")
    from commercial_fit_scoring import score_company as _sc, SCORING_PROFILES as _SPROFS

    # Detection: Italy filename
    chk("ITP-1 Italy100 filename → italy_register_icp_only",
        _detect_italy_register_profile("Italy100_1_R0001_0500.xlsx"))
    chk("ITP-2 Italy200 filename → italy_register_icp_only",
        _detect_italy_register_profile("Italy200_1_R0001_0500.xlsx"))
    chk("ITP-3 Italy+cleaned filename → italy_register_icp_only",
        _detect_italy_register_profile("Italy100_1_R0001_0500_cleaned_20260614.xlsx"))
    chk("ITP-4 Germany filename → NOT italy_register",
        not _detect_italy_register_profile("Germany_1_R0001_0500.xlsx"))
    chk("ITP-5 generic cleaner → NOT italy_register",
        not _detect_italy_register_profile("register_cleaned_20260613.xlsx"))

    # Italy profile country_code column detection
    _h_it_df = pd.DataFrame([{"company_name": "Test SRL", "country_code": "IT"}])
    _h_de_df2 = pd.DataFrame([{"company_name": "Test GmbH", "country_code": "DE"}])
    chk("ITP-6 country_code=IT df → italy_register_icp_only",
        _detect_italy_register_profile("register_cleaned.xlsx", _h_it_df))
    chk("ITP-7 country_code=DE df → NOT italy_register",
        not _detect_italy_register_profile("register_cleaned.xlsx", _h_de_df2))

    # Scoring: Italy profile uses K=1, model_weight=1.0, size_weight=0.0
    _italy_prof = _SPROFS["italy_register_icp_only"]
    chk("ITP-8 italy profile model_weight = 1.0",  _italy_prof["model_weight"] == 1.0)
    chk("ITP-9 italy profile size_weight = 0.0",   _italy_prof["size_weight"]  == 0.0)
    chk("ITP-10 italy profile sigmoid_k = 1.0",    _italy_prof["sigmoid_k"]    == 1.0)

    # Score same row with default vs Italy — size must not affect Italy score
    _ref_row = {
        "sig_foreign_hq_score": 3, "sig_explicit_lnd_score": 3,
        "sig_intl_footprint_score": 3, "sig_employer_branding_score": 2,
        "sig_lnd_onboarding_score": 2, "ti_onboarding_score": 2,
        "sig_rapid_growth_score": 1,
        "lusha_api_employee_range": "1 - 10",   # tiny size → low size_score
    }
    _s_default = _sc(_ref_row, {"scoring_profile": "default"})
    _s_italy   = _sc(_ref_row, {"scoring_profile": "italy_register_icp_only"})
    chk("ITP-11 Italy final_score != default when size differs",
        abs(_s_italy["final_commercial_fit_score"] - _s_default["final_commercial_fit_score"]) > 0.01,
        f"italy={_s_italy['final_commercial_fit_score']} default={_s_default['final_commercial_fit_score']}")
    chk("ITP-12 Italy weighted_size_component = 0.0",
        _s_italy["weighted_size_component"] == 0.0,
        repr(_s_italy["weighted_size_component"]))
    chk("ITP-13 Italy scoring_profile field = italy_register_icp_only",
        _s_italy["scoring_profile"] == "italy_register_icp_only")
    chk("ITP-14 Italy sigmoid_k in output = 1.0",
        _s_italy["sigmoid_k"] == 1.0, repr(_s_italy["sigmoid_k"]))
    chk("ITP-15 Italy model_weight in output = 1.0",
        _s_italy["model_weight"] == 1.0, repr(_s_italy["model_weight"]))
    # Scores with large vs small size MUST be identical for Italy profile
    _ref_big  = dict(_ref_row, **{"lusha_api_employee_range": "100001 - 10000000"})
    _s_it_big = _sc(_ref_big,  {"scoring_profile": "italy_register_icp_only"})
    _s_it_sml = _sc(_ref_row,  {"scoring_profile": "italy_register_icp_only"})
    chk("ITP-16 Italy score invariant to company size",
        abs(_s_it_big["final_commercial_fit_score"] - _s_it_sml["final_commercial_fit_score"]) < 1e-6,
        f"big={_s_it_big['final_commercial_fit_score']} small={_s_it_sml['final_commercial_fit_score']}")

    # ── Foreign HQ hygiene sanitizer tests ───────────────────────────────────
    print("\nForeign HQ hygiene sanitizer")

    # FHQ-A: Italian domestic multinational — must be sanitized to 0
    _fhq_row_a = {
        "sig_foreign_hq_score": 2,
        "sig_foreign_hq_evidence": "The company is headquartered in Milan, Italy, and operates subsidiaries in the United States, Germany, and France.",
        "sig_intl_footprint_score": 3,
        "icp_evidence": "",
        "country_code": "IT",
    }
    _fhq_r_a = dict(_fhq_row_a)
    sanitize_foreign_hq_signal(_fhq_r_a, "IT")
    chk("FHQ-A sig_intl_footprint_score preserved", _fhq_r_a["sig_intl_footprint_score"] == 3)
    chk("FHQ-A sig_foreign_hq_score → 0",           int(_fhq_r_a["sig_foreign_hq_score"]) == 0,
        repr(_fhq_r_a["sig_foreign_hq_score"]))
    chk("FHQ-A foreign_hq_sanitized = True",         _fhq_r_a["foreign_hq_sanitized"] is True,
        repr(_fhq_r_a["foreign_hq_sanitized"]))

    # FHQ-B: Italian company owned by German parent — must NOT be sanitized
    _fhq_row_b = {
        "sig_foreign_hq_score": 3,
        "sig_foreign_hq_evidence": "The company is the Italian subsidiary of a German group headquartered in Munich.",
        "icp_evidence": "Part of a large German industrial group.",
        "country_code": "IT",
    }
    _fhq_r_b = dict(_fhq_row_b)
    sanitize_foreign_hq_signal(_fhq_r_b, "IT")
    chk("FHQ-B sig_foreign_hq_score stays 3",
        int(_fhq_r_b["sig_foreign_hq_score"]) == 3,
        repr(_fhq_r_b["sig_foreign_hq_score"]))
    chk("FHQ-B foreign_hq_sanitized = False",
        not _fhq_r_b["foreign_hq_sanitized"],
        repr(_fhq_r_b["foreign_hq_sanitized"]))

    # FHQ-C: Italian company with export activity only — no HQ reference
    _fhq_row_c = {
        "sig_foreign_hq_score": 1,
        "sig_foreign_hq_evidence": "The company exports to more than 40 countries worldwide.",
        "icp_evidence": "Strong export activity internationally.",
        "country_code": "IT",
    }
    _fhq_r_c = dict(_fhq_row_c)
    sanitize_foreign_hq_signal(_fhq_r_c, "IT")
    chk("FHQ-C sig_foreign_hq_score → 0 (export only)",
        int(_fhq_r_c["sig_foreign_hq_score"]) == 0,
        repr(_fhq_r_c["sig_foreign_hq_score"]))
    chk("FHQ-C foreign_hq_sanitized = True",
        _fhq_r_c["foreign_hq_sanitized"] is True)

    # FHQ-D: Unknown country — score 0 already, no-op
    _fhq_row_d = {
        "sig_foreign_hq_score": 0,
        "sig_foreign_hq_evidence": "",
        "icp_evidence": "The company is headquartered in Milan and has offices abroad.",
    }
    _fhq_r_d = dict(_fhq_row_d)
    sanitize_foreign_hq_signal(_fhq_r_d, "")
    chk("FHQ-D zero score remains 0 (no-op)", int(_fhq_r_d["sig_foreign_hq_score"]) == 0)
    chk("FHQ-D foreign_hq_sanitized = False", not _fhq_r_d["foreign_hq_sanitized"])

    # FHQ-E: _get_input_country resolves Italy scoring profile
    _row_e = {"company_name": "Test SRL"}
    chk("FHQ-E italy_register_icp_only profile → IT",
        _get_input_country(_row_e, "italy_register_icp_only") == "IT")

    # FHQ-F: country_code field takes priority
    _row_f = {"country_code": "DE", "company_name": "Test GmbH"}
    chk("FHQ-F country_code=DE → DE",
        _get_input_country(_row_f, "default") == "DE")

    # FHQ-G: foreign_hq_original_score preserved on sanitization
    _fhq_row_g = {
        "sig_foreign_hq_score": 2,
        "sig_foreign_hq_evidence": "Headquartered in Rome, Italy, and exports globally.",
        "icp_evidence": "",
    }
    _fhq_r_g = dict(_fhq_row_g)
    sanitize_foreign_hq_signal(_fhq_r_g, "IT")
    chk("FHQ-G foreign_hq_original_score = 2",
        str(_fhq_r_g["foreign_hq_original_score"]) == "2",
        repr(_fhq_r_g["foreign_hq_original_score"]))

    # FHQ-H: Ferrari-like — "headquartered in Maranello, Italy" → sanitize to 0
    # Simulates the df-level pass where scoring_profile drives input_country resolution.
    _fhq_row_h = {
        "sig_foreign_hq_score": 3,
        "sig_foreign_hq_evidence": (
            "Ferrari S.p.A. is headquartered in Maranello, Italy, "
            "confirmed in enrichment context."
        ),
        "icp_evidence": "",
    }
    _fhq_r_h = dict(_fhq_row_h)
    _ctry_h = _get_input_country(_fhq_r_h, "italy_register_icp_only")
    sanitize_foreign_hq_signal(_fhq_r_h, _ctry_h)
    chk("FHQ-H inferred_input_country = IT", _fhq_r_h.get("inferred_input_country") == "IT",
        repr(_fhq_r_h.get("inferred_input_country")))
    chk("FHQ-H sig_foreign_hq_score → 0", int(_fhq_r_h["sig_foreign_hq_score"]) == 0,
        repr(_fhq_r_h["sig_foreign_hq_score"]))
    chk("FHQ-H foreign_hq_sanitized = True", _fhq_r_h["foreign_hq_sanitized"] is True,
        repr(_fhq_r_h["foreign_hq_sanitized"]))

    # FHQ-I: Buzzi-like — domestic HQ + international subsidiaries → sanitize to 0
    _fhq_row_i = {
        "sig_foreign_hq_score": 3,
        "sig_foreign_hq_evidence": (
            "Headquartered in Casale Monferrato, Italy with international subsidiaries "
            "including USA operations."
        ),
        "sig_intl_footprint_score": 3,
        "icp_evidence": "",
    }
    _fhq_r_i = dict(_fhq_row_i)
    _ctry_i = _get_input_country(_fhq_r_i, "italy_register_icp_only")
    sanitize_foreign_hq_signal(_fhq_r_i, _ctry_i)
    chk("FHQ-I sig_foreign_hq_score → 0", int(_fhq_r_i["sig_foreign_hq_score"]) == 0,
        repr(_fhq_r_i["sig_foreign_hq_score"]))
    chk("FHQ-I sig_intl_footprint_score preserved", _fhq_r_i["sig_intl_footprint_score"] == 3)
    chk("FHQ-I foreign_hq_sanitized = True", _fhq_r_i["foreign_hq_sanitized"] is True,
        repr(_fhq_r_i["foreign_hq_sanitized"]))

    # FHQ-J: Foreign parent — "Italian subsidiary of a German group" → keep score 3
    _fhq_row_j = {
        "sig_foreign_hq_score": 3,
        "sig_foreign_hq_evidence": "Italian subsidiary of a German group headquartered in Munich.",
        "icp_evidence": "",
    }
    _fhq_r_j = dict(_fhq_row_j)
    _ctry_j = _get_input_country(_fhq_r_j, "italy_register_icp_only")
    sanitize_foreign_hq_signal(_fhq_r_j, _ctry_j)
    chk("FHQ-J sig_foreign_hq_score stays 3 (foreign parent)",
        int(_fhq_r_j["sig_foreign_hq_score"]) == 3,
        repr(_fhq_r_j["sig_foreign_hq_score"]))
    chk("FHQ-J foreign_hq_sanitized = False",
        not _fhq_r_j["foreign_hq_sanitized"],
        repr(_fhq_r_j["foreign_hq_sanitized"]))

    # FHQ-K: Sandenvendo — "acquired by Sanden Corporation" (Japanese parent) → keep positive
    _fhq_row_k = {
        "sig_foreign_hq_score": 3,
        "sig_foreign_hq_evidence": (
            "Sandenvendo S.p.A. was acquired by Sanden Corporation, a Japanese company "
            "headquartered in Isesaki, Japan."
        ),
        "icp_evidence": "",
        "country_code": "IT",
    }
    _fhq_r_k = dict(_fhq_row_k)
    sanitize_foreign_hq_signal(_fhq_r_k, "IT")
    chk("FHQ-K Sandenvendo: sig_foreign_hq_score stays 3 (explicit Japanese parent)",
        int(_fhq_r_k["sig_foreign_hq_score"]) == 3,
        repr(_fhq_r_k["sig_foreign_hq_score"]))
    chk("FHQ-K Sandenvendo: foreign_hq_sanitized = False",
        not _fhq_r_k["foreign_hq_sanitized"])
    chk("FHQ-K Sandenvendo: foreign_hq_uncertain = False",
        not _fhq_r_k.get("foreign_hq_uncertain", False))

    # FHQ-L: Gestione Ambiente — "acquired by Itelyum Group" (no country proof) → zero + uncertain
    _fhq_row_l = {
        "sig_foreign_hq_score": 2,
        "sig_foreign_hq_evidence": (
            "Gestione Ambiente S.r.l. was acquired by Itelyum Group in June 2025."
        ),
        "icp_evidence": "",
        "country_code": "IT",
    }
    _fhq_r_l = dict(_fhq_row_l)
    sanitize_foreign_hq_signal(_fhq_r_l, "IT")
    chk("FHQ-L Gestione: sig_foreign_hq_score → 0 (no country proof)",
        int(_fhq_r_l["sig_foreign_hq_score"]) == 0,
        repr(_fhq_r_l["sig_foreign_hq_score"]))
    chk("FHQ-L Gestione: foreign_hq_uncertain = True",
        _fhq_r_l.get("foreign_hq_uncertain") is True,
        repr(_fhq_r_l.get("foreign_hq_uncertain")))
    chk("FHQ-L Gestione: model_signal_needs_manual_review = 1",
        int(_fhq_r_l.get("model_signal_needs_manual_review", 0)) == 1,
        repr(_fhq_r_l.get("model_signal_needs_manual_review")))
    chk("FHQ-L Gestione: original evidence preserved",
        "Itelyum" in str(_fhq_r_l.get("foreign_hq_original_evidence", "")),
        repr(_fhq_r_l.get("foreign_hq_original_evidence", "")[:80]))

    # SIZE-1: _strip_employee_range_from_text removes employee counts
    _size_text = "The company has 1500 employees and exports globally."
    _size_stripped = _strip_employee_range_from_text(_size_text)
    chk("SIZE-1: '1500 employees' removed from text",
        "1500" not in _size_stripped, repr(_size_stripped))

    _size_text2 = "Employing approximately 800-1200 employees across Italy."
    _size_stripped2 = _strip_employee_range_from_text(_size_text2)
    chk("SIZE-2: '800-1200 employees' range removed",
        "800" not in _size_stripped2 and "1200" not in _size_stripped2, repr(_size_stripped2))

    # CONS-1: profile consistency — intl footprint score >= 2 strips "no international presence"
    _cons_rd1 = {"sig_intl_footprint_score": 3, "sig_foreign_hq_score": 0, "sig_explicit_lnd_score": 0}
    _cons_text1 = "Active exporter. No evidence of international footprint. Promising lead."
    _cons_out1 = _check_profile_consistency(_cons_rd1, _cons_text1)
    chk("CONS-1: 'no evidence of international footprint' stripped when score>=2",
        "no evidence of international footprint" not in _cons_out1.lower(), repr(_cons_out1))

    # CONS-2: lnd_score=0 strips "strong L&D culture"
    _cons_rd2 = {"sig_intl_footprint_score": 0, "sig_foreign_hq_score": 0, "sig_explicit_lnd_score": 0}
    _cons_text2 = "The company has a strong L&D culture and invests in training."
    _cons_out2 = _check_profile_consistency(_cons_rd2, _cons_text2)
    chk("CONS-2: 'strong L&D culture' stripped when lnd_score=0",
        "strong L&D culture" not in _cons_out2, repr(_cons_out2))

    # ── Caller Angle tests ────────────────────────────────────────────────────
    print("\nCaller Angle builder")

    # CA-1: Hot company with international footprint + explicit L&D
    _ca1 = {
        "commercial_tier": "🥇 Hot",
        "scoring_profile": "italy_register_icp_only",
        "sig_intl_footprint_score": 3,
        "sig_explicit_lnd_score": 3,
        "sig_foreign_hq_score": 0,
        "foreign_hq_sanitized": False,
        "icp_potential_buyer_function": "",
        "competitive_switch_opportunity": "",
    }
    _ca1_text = _build_caller_angle(_ca1)
    chk("CA-1 Hot: starts with 'Strong Layer 1'", _ca1_text.startswith("Strong Layer 1"),
        repr(_ca1_text[:60]))
    chk("CA-1 Hot: no size disclaimer", "size" not in _ca1_text.lower() and
        "100+ employees" not in _ca1_text, repr(_ca1_text[:80]))
    chk("CA-1 Hot: mentions L&D or HR",
        any(w in _ca1_text for w in ("L&D", "HR", "learning", "leadership")),
        repr(_ca1_text[:80]))

    # CA-2: Warm company with international footprint but no explicit L&D
    _ca2 = {
        "commercial_tier": "🥈 Warm",
        "scoring_profile": "italy_register_icp_only",
        "sig_intl_footprint_score": 2,
        "sig_explicit_lnd_score": 0,
        "sig_foreign_hq_score": 0,
        "foreign_hq_sanitized": False,
        "icp_potential_buyer_function": "",
        "competitive_switch_opportunity": "",
    }
    _ca2_text = _build_caller_angle(_ca2)
    chk("CA-2 Warm: mentions 'Promising' or incomplete",
        "romising" in _ca2_text or "incomplete" in _ca2_text, repr(_ca2_text[:80]))
    chk("CA-2 Warm: recommends Opportunity Radar or verification",
        "Opportunity Radar" in _ca2_text or "verif" in _ca2_text.lower(),
        repr(_ca2_text[:120]))

    # CA-3: Cool company with thin evidence
    _ca3 = {
        "commercial_tier": "🥉 Cool",
        "scoring_profile": "default",
        "sig_intl_footprint_score": 1,
        "sig_explicit_lnd_score": 0,
        "sig_foreign_hq_score": 0,
        "foreign_hq_sanitized": False,
        "icp_potential_buyer_function": "",
        "competitive_switch_opportunity": "",
    }
    _ca3_text = _build_caller_angle(_ca3)
    chk("CA-3 Cool: mentions thin evidence or Opportunity Radar",
        "thin" in _ca3_text or "Opportunity Radar" in _ca3_text, repr(_ca3_text[:80]))

    # CA-4: Pass company
    _ca4 = {
        "commercial_tier": "❄️ Pass",
        "scoring_profile": "default",
        "sig_intl_footprint_score": 0,
        "sig_explicit_lnd_score": 0,
        "sig_foreign_hq_score": 0,
        "foreign_hq_sanitized": False,
        "icp_potential_buyer_function": "",
        "competitive_switch_opportunity": "",
    }
    _ca4_text = _build_caller_angle(_ca4)
    chk("CA-4 Pass: says 'Low Layer 1 priority'",
        "Low Layer 1 priority" in _ca4_text, repr(_ca4_text[:80]))

    # CA-5: Ferrari-like domestic global — no foreign HQ language
    _ca5 = {
        "commercial_tier": "🥉 Cool",
        "scoring_profile": "italy_register_icp_only",
        "sig_intl_footprint_score": 3,
        "sig_foreign_hq_score": 0,      # sanitized
        "foreign_hq_sanitized": True,
        "sig_explicit_lnd_score": 0,
        "icp_potential_buyer_function": "",
        "competitive_switch_opportunity": "",
    }
    _ca5_text = _build_caller_angle(_ca5)
    _fhq_words = ("foreign HQ", "international headquarters", "foreign headquarters",
                  "Foreign HQ", "International headquarters")
    chk("CA-5 Ferrari: no foreign-HQ language in Caller Angle",
        not any(w in _ca5_text for w in _fhq_words), repr(_ca5_text[:120]))

    # CA-6: True Italian subsidiary of foreign parent — foreign parent language allowed
    _ca6 = {
        "commercial_tier": "🥇 Hot",
        "scoring_profile": "italy_register_icp_only",
        "sig_intl_footprint_score": 3,
        "sig_foreign_hq_score": 3,      # real foreign parent, NOT sanitized
        "foreign_hq_sanitized": False,
        "sig_explicit_lnd_score": 0,
        "icp_potential_buyer_function": "",
        "competitive_switch_opportunity": "",
    }
    _ca6_text = _build_caller_angle(_ca6)
    chk("CA-6 Foreign parent: mentions parent or external structure",
        any(w in _ca6_text for w in ("foreign parent", "external", "centrally")),
        repr(_ca6_text[:120]))

    print(f"\n{'='*60}")
    if failures:
        print(f"  FAILURES ({len(failures)}):")
        for f in failures:
            print(f"    - {f}")
        sys.exit(1)
    else:
        print("  OK: All pipeline validation checks passed.")
    print("=" * 60)




def format_duration(seconds: float) -> str:
    """Format a duration in seconds as HH:MM:SS or MM:SS.

    Examples:
        45   -> "00:45"
        125  -> "02:05"
        3725 -> "01:02:05"
    """
    s = max(0, int(seconds))
    h = s // 3600
    m = (s % 3600) // 60
    sec = s % 60
    if h:
        return f"{h:02d}:{m:02d}:{sec:02d}"
    return f"{m:02d}:{sec:02d}"


def estimate_cli_eta(
    completed: int,
    total: int,
    run_start_ts: float,
    row_durations: list,
    recent_window: int = 5,
) -> dict:
    """Return ETA stats dict for CLI progress reporting.

    Keys returned:
        elapsed_seconds, last_row_seconds, avg_row_seconds,
        recent_avg_row_seconds, remaining_seconds, estimated_finish_local
    """
    import time as _time_eta
    import datetime as _dt

    elapsed = _time_eta.monotonic() - run_start_ts
    last_row = row_durations[-1] if row_durations else 0.0

    if completed > 0 and row_durations:
        avg_row = sum(row_durations) / len(row_durations)
        window  = row_durations[-recent_window:] if len(row_durations) >= recent_window else row_durations
        recent_avg = sum(window) / len(window)
    else:
        avg_row    = 0.0
        recent_avg = 0.0

    remaining_rows = max(0, total - completed)
    remaining_secs = recent_avg * remaining_rows if recent_avg else avg_row * remaining_rows

    finish_local = (
        _dt.datetime.now() + _dt.timedelta(seconds=remaining_secs)
    ).strftime("%Y-%m-%d %H:%M")

    return {
        "elapsed_seconds":         elapsed,
        "last_row_seconds":        last_row,
        "avg_row_seconds":         avg_row,
        "recent_avg_row_seconds":  recent_avg,
        "remaining_seconds":       remaining_secs,
        "estimated_finish_local":  finish_local,
    }


def build_cli_output_filename(input_path: str, suffix: str = "lead_prioritized") -> str:
    """
    Derive CLI output filename from input filename.

    Italy200_02_R0501_1000_cleaned_20260612_1905.xlsx
    -> Italy200_02_R0501_1000_lead_prioritized_YYYYMMDD_HHMM.xlsx

    If input stem does not contain '_cleaned_', falls back to:
    {stem}_{suffix}_{YYYYMMDD_HHMM}.xlsx
    """
    import re as _re
    from datetime import datetime as _dt

    stem = Path(input_path).stem
    now  = _dt.now().strftime("%Y%m%d_%H%M")

    # Strip trailing _cleaned_YYYYMMDD_HHMM (or _cleaned_YYYYMMDDHHMMSS etc.)
    cleaned = _re.sub(r"_cleaned_\d{8}_?\d{4,6}$", "", stem, flags=_re.IGNORECASE)
    if cleaned != stem:
        return f"{cleaned}_{suffix}_{now}.xlsx"
    return f"{stem}_{suffix}_{now}.xlsx"


def build_enriched_output_filename(
    input_path: "str | Path | None",
    run_stamp: "str | None" = None,
) -> str:
    """
    Derive enricher output filename from the input (cleaned) filename.

    Italy200_01_R0001_0500_cleaned_20260613_1301.xlsx
    -> Italy200_01_R0001_0500_enriched_<run_stamp>.xlsx

    Rules:
    - Strip '_cleaned_YYYYMMDD_HHMMSS' (or _HHMM) suffix to get the batch prefix.
    - Append '_enriched_<run_stamp>.xlsx'.
    - If input is None/empty or contains no '_cleaned_' tag, fall back to:
      '<stem>_enriched_<run_stamp>.xlsx'.
    - run_stamp defaults to current UTC time formatted as YYYYMMDD_HHMMSS.
    """
    import re as _re
    from datetime import datetime as _dt

    if run_stamp is None:
        run_stamp = _dt.utcnow().strftime("%Y%m%d_%H%M%S")

    if not input_path:
        return f"enriched_{run_stamp}.xlsx"

    stem = Path(str(input_path)).stem
    batch_prefix = _re.sub(
        r"_cleaned_\d{8}_?\d{4,6}$", "", stem, flags=_re.IGNORECASE
    )
    return f"{batch_prefix}_enriched_{run_stamp}.xlsx"


def run_cli() -> None:
    """Non-Streamlit batch entry point."""
    import argparse

    parser = argparse.ArgumentParser(
        description="mYngle Lead Prioritizer — CLI batch mode",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--input",          required=False, default=None, help="Path to cleaned .xlsx or .csv input file")
    parser.add_argument("--project-root",   default=None,   help="Pipeline project root (for folder convention)")
    parser.add_argument("--output-dir",     default=None,   help="Output directory for enriched Excel")
    parser.add_argument("--anthropic-key",  default=None,   help="Anthropic API key")
    parser.add_argument("--serper-key",     default=None,   help="Serper API key")
    parser.add_argument("--max-rows",       type=int, default=0, help="Process first N rows (0 = all)")
    parser.add_argument("--debug",          action="store_true", help="Enable debug output")
    parser.add_argument("--dry-run-paths",  action="store_true", help="Print paths and exit, no processing")
    parser.add_argument("--output-name",    default=None,   help="Override output filename (without .xlsx)")
    parser.add_argument("--self-test-competitor-override", action="store_true",
                        help="Run zero-cost competitor override self-test and exit")
    parser.add_argument("--self-test-output", default=None,
                        help="Optional path for self-test results Excel (.xlsx)")
    parser.add_argument("--no-eta",           action="store_true",
                        help="Suppress per-row ETA progress lines")
    args = parser.parse_args()

    # ── Self-test mode: run and exit immediately, no API calls ────────────────
    if args.self_test_competitor_override:
        _selftest_results = run_competitor_override_selftest()
        print("\n" + "=" * 64)
        print("  mYngle Competitor Override Self-Test")
        print("=" * 64)
        for r in _selftest_results["results"]:
            status = "PASS" if r["passed"] else "FAIL"
            print(f"  [{status}] {r['name']}")
            if not r["passed"]:
                print(f"         → {r['details']}")
        print("=" * 64)
        print(
            f"  Summary: {_selftest_results['passed']} passed / "
            f"{_selftest_results['failed']} failed / "
            f"{_selftest_results['total']} total"
        )
        print("=" * 64)
        if args.self_test_output:
            try:
                import pandas as _pd_st
                _rows = []
                for r in _selftest_results["results"]:
                    _rows.append({
                        "test_name":   r["name"],
                        "result":      "PASS" if r["passed"] else "FAIL",
                        "details":     r["details"],
                    })
                _st_df = _pd_st.DataFrame(_rows)
                _out_path = Path(args.self_test_output)
                _st_df.to_excel(_out_path, index=False)
                print(f"  Self-test output written to: {_out_path}")
            except Exception as _ste:
                print(f"  WARNING: could not write self-test output: {_ste}", file=sys.stderr)
        if _selftest_results["failed"] > 0:
            sys.exit(1)
        sys.exit(0)

    # ── Normal mode requires --input ──────────────────────────────────────────
    if not args.input:
        print("ERROR: --input is required (or use --self-test-competitor-override)", file=sys.stderr)
        sys.exit(1)

    # ── Resolve input path ────────────────────────────────────────────────────
    input_path = Path(args.input).resolve()
    if not input_path.exists():
        print(f"ERROR: input file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    # ── Output dir ────────────────────────────────────────────────────────────
    out_dir = Path(args.output_dir).resolve() if args.output_dir else input_path.parent
    out_dir.mkdir(parents=True, exist_ok=True)

    # ── Load API keys: CLI arg → env var → secrets.toml ─────────────────────
    def _load_key_from_secrets(key_name: str) -> str:
        """Read a single key from .streamlit/secrets.toml without Streamlit."""
        try:
            import tomllib
        except ImportError:
            try:
                import tomli as tomllib  # type: ignore[no-redef]
            except ImportError:
                return ""
        secrets_path = Path(__file__).parent / ".streamlit" / "secrets.toml"
        if not secrets_path.exists():
            return ""
        try:
            with open(secrets_path, "rb") as f:
                data = tomllib.load(f)
            return str(data.get(key_name, "") or "")
        except Exception:
            return ""

    anthropic_key = (
        args.anthropic_key
        or os.environ.get("ANTHROPIC_API_KEY", "")
        or _load_key_from_secrets("ANTHROPIC_API_KEY")
        or _load_key_from_secrets("anthropic_api_key")
        or ""
    )
    serper_key = (
        args.serper_key
        or os.environ.get("SERPER_API_KEY", "")
        or _load_key_from_secrets("SERPER_API_KEY")
        or _load_key_from_secrets("serper_api_key")
        or ""
    )

    print(f"[enricher] Input:         {input_path}", flush=True)
    print(f"[enricher] Output dir:    {out_dir}", flush=True)
    print(f"[enricher] Anthropic key: {'found' if anthropic_key else 'missing'}", flush=True)
    print(f"[enricher] Serper key:    {'found' if serper_key else 'missing'}", flush=True)

    if args.dry_run_paths:
        print("[enricher] --dry-run-paths: exiting without processing.", flush=True)
        sys.exit(0)

    # ── Load input ────────────────────────────────────────────────────────────
    fname = input_path.name.lower()
    if fname.endswith(".csv"):
        df_in = pd.read_csv(input_path)
    else:
        df_in = pd.read_excel(input_path)

    if args.max_rows and args.max_rows > 0:
        df_in = df_in.head(args.max_rows)

    if len(df_in) == 0:
        print("[enricher] ERROR: input file has zero rows.", file=sys.stderr)
        sys.exit(2)

    print(f"[enricher] Rows to process: {len(df_in)}", flush=True)

    # ── Detect key columns ────────────────────────────────────────────────────
    # Columns that must never be selected as the company-name column
    _META_COLS_EXCLUDE: set[str] = {
        "organization_type", "myngle_target_eligibility", "pre_filter_decision",
        "pre_filter_reason", "pre_score", "pre_label", "domain_action",
        "domain_confidence", "domain_reason", "final_confidence",
        "final_decision_source", "verifier_decision", "website_discovery_method",
        "professional_site_level", "source", "source_row_id", "current_status",
        "jurisdiction_code", "legal_form_detected", "registered_address",
        "city_or_registered_office", "federal_state", "registrar",
        "register_art", "register_nummer", "retrieved_at",
        "positive_reasons", "exclude_reasons", "low_priority_reasons",
    }
    _META_SUFFIX_EXCLUDE = (
        "_status", "_source", "_reason", "_signal", "_score",
        "_confidence", "_type", "_decision", "_action", "_label",
        "_notes", "_evidence",
    )

    def _is_meta_col(col_name: str) -> bool:
        if col_name in _META_COLS_EXCLUDE:
            return True
        cl = col_name.lower()
        return any(cl.endswith(sfx) for sfx in _META_SUFFIX_EXCLUDE)

    _col_candidates = {
        # Priority order: cleaned output canonical → German raw → Italian raw → generic
        "company_name": [
            "company_name", "canonical_company_name", "cleaned_company_name",
            "Company Name", "company_name_clean", "company_name_raw",
            "name", "company", "organisation_name", "organization_name",
        ],
        "domain": [
            "website_url", "final_selected_domain", "canonical_company_domain",
            "validated_domain", "recommended_domain",
            "python_validated_domain", "python_recommended_domain",
            "Company Domain", "Company Website", "Website",
            "website", "domain", "url", "homepage",
        ],
        "city":           ["lusha_city", "city", "City", "lusha_api_city", "city_or_registered_office"],
        "country":        ["lusha_country", "country", "Country", "lusha_api_country"],
        "industry":       ["lusha_industry", "industry", "Industry", "lusha_api_industry"],
        "employee_range": [
            "employee_range_for_scoring", "employee_range_resolved",
            "lusha_employee_range", "lusha_api_employee_range", "employee_range",
        ],
    }

    def _find_col(role):
        for cand in _col_candidates.get(role, []):
            if cand in df_in.columns:
                # For company_name role, never select metadata columns
                if role == "company_name" and _is_meta_col(cand):
                    continue
                return cand
        # For company_name: fuzzy fallback — first non-metadata column that has unique values
        if role == "company_name":
            for col in df_in.columns:
                if not _is_meta_col(col):
                    n_unique = df_in[col].astype(str).str.strip().replace("", pd.NA).dropna().nunique()
                    if n_unique >= max(3, len(df_in) // 4):
                        return col
        return None

    company_col = _find_col("company_name") or df_in.columns[0]
    domain_col  = _find_col("domain")  # may be None for German raw files

    # Guard: never use a metadata column as company_name
    if _is_meta_col(company_col):
        # Explicit fallback to first non-meta column
        company_col = next((c for c in df_in.columns if not _is_meta_col(c)), df_in.columns[0])

    print(f"[enricher] Company col: {company_col}, Domain col: {domain_col or '(none)'}", flush=True)

    _cnum_col = _detect_company_number_col(df_in)
    print(f"[enricher] Company number column: {_cnum_col or 'none'}", flush=True)

    # ── Input type + unified run config ──────────────────────────────────────
    _cli_fname = str(input_path.name)
    _cli_det_itype = (
        "pre_enriched_lucia_export" if is_lucia_contact_export(df_in)
        else "input_cleaner_output"  if is_input_cleaner_output(df_in, _cli_fname)
        else "simple_company_list"
    )
    # For Lusha detection: cleaner output is never treated as pre-enriched Lusha
    _cli_has_lusha = (
        bool(detect_lusha_columns(df_in))
        and _cli_det_itype == "pre_enriched_lucia_export"
    )
    _run_config = resolve_active_run_config(
        input_filename=_cli_fname,
        df=df_in,
        detected_input_type=_cli_det_itype,
        has_existing_lusha_data=_cli_has_lusha,
    )
    _cli_scoring_profile = _run_config["scoring_profile"]

    from commercial_fit_scoring import SCORING_PROFILES as _SP
    _sp_info = _SP.get(_cli_scoring_profile, _SP["default"])
    print("", flush=True)
    print(f"[enricher] INPUT TYPE:          {_cli_det_itype}", flush=True)
    print(f"[enricher] SCORING PROFILE:     {_cli_scoring_profile} ({_run_config['scoring_profile_source']})", flush=True)
    print(f"[enricher] MODEL WEIGHT:        {_sp_info['model_weight']}", flush=True)
    print(f"[enricher] SIZE WEIGHT:         {_sp_info['size_weight']}", flush=True)
    print(f"[enricher] SIGMOID_K:           {_sp_info['sigmoid_k']}", flush=True)
    print(f"[enricher] run_step1:           {_run_config['run_step1_enrichment']}", flush=True)
    print(f"[enricher] run_step2:           {_run_config['run_step2_enrichment']}", flush=True)
    print(f"[enricher] extract_signals:     {_run_config['extract_model_signals']}", flush=True)
    print(f"[enricher] search_provider:     {_run_config['search_provider']}", flush=True)
    print(f"[enricher] use_playwright:      {_run_config['use_playwright']}", flush=True)
    print(f"[enricher] model_step1:         {_run_config['model_step1']}", flush=True)
    print(f"[enricher] model_step2:         {_run_config['model_step2']}", flush=True)
    if _cli_scoring_profile == "italy_register_icp_only":
        print("[enricher] COMPANY SIZE:        excluded from score, audit only", flush=True)
    else:
        print("[enricher] COMPANY SIZE:        included in score (10% weight)", flush=True)
    print("", flush=True)

    if not domain_col:
        print("[enricher] No domain column found - proceeding with company-name-only enrichment.", flush=True)

    # ── Progress helpers ──────────────────────────────────────────────────────
    import time as _time

    _show_eta = not getattr(args, "no_eta", False)

    # ── Process rows ──────────────────────────────────────────────────────────
    results       = []
    debug_records = []
    total         = len(df_in)
    _claude_calls = 0
    _serper_calls = 0
    _error_count  = 0
    _start_ts     = _time.monotonic()
    _row_durations: list = []

    for i, (_, row) in enumerate(df_in.iterrows(), 1):
        row_dict     = row.to_dict()
        company_name = str(row_dict.get(company_col, "") or "").strip()
        domain       = str(row_dict.get(domain_col, "") or "").strip() if domain_col else ""
        _cnum        = _get_company_number(row_dict)

        # Print START line for every row
        _eta_pre = estimate_cli_eta(i - 1, total, _start_ts, _row_durations)
        print(
            f"[enricher] START {i}/{total} | "
            f"company_number: {_cnum} | "
            f"company: {company_name[:60]} | "
            f"elapsed {format_duration(_eta_pre['elapsed_seconds'])} | "
            f"Claude: {_claude_calls} | Serper: {_serper_calls} | errors: {_error_count}",
            flush=True,
        )

        _row_start = _time.monotonic()
        _row_status = "ok"
        try:
            result, _debug_rec = enrich_one_row(
                company_name=company_name,
                raw_url=domain,
                api_key=anthropic_key or "",
                delay=0,
                serper_key=serper_key or "",
                _cli_verbose=True,
                scoring_profile=_run_config["scoring_profile"],
                run_step1_enrichment=_run_config["run_step1_enrichment"],
                run_step2_enrichment=_run_config["run_step2_enrichment"],
                extract_model_signals=_run_config["extract_model_signals"],
                include_signal_evidence=_run_config["include_signal_evidence"],
                search_provider=_run_config["search_provider"],
                use_playwright=_run_config["use_playwright"],
                model_step1=_run_config["model_step1"],
                model_step2=_run_config["model_step2"],
            )
            debug_records.append(_debug_rec)
            _claude_calls += int(result.get("claude_api_calls", 0) or 0)
            _serper_calls += int(result.get("serper_calls", 0) or 0)
        except Exception as exc:
            result = dict(row_dict)
            result["enrichment_error"] = f"{type(exc).__name__}: {exc}"
            _error_count += 1
            _row_status = f"error: {type(exc).__name__}"
        results.append(result)

        _row_elapsed = _time.monotonic() - _row_start
        _row_durations.append(_row_elapsed)

        # Print DONE line for every row
        print(
            f"[enricher] DONE  {i}/{total} | "
            f"company_number: {_cnum} | "
            f"name: {company_name[:60]} | "
            f"status: {_row_status} | "
            f"row_time: {format_duration(_row_elapsed)}",
            flush=True,
        )

        # Print ETA line (unless --no-eta)
        if _show_eta:
            _eta = estimate_cli_eta(i, total, _start_ts, _row_durations)
            print(
                f"[enricher] ETA   {i}/{total} | "
                f"elapsed {format_duration(_eta['elapsed_seconds'])} | "
                f"avg {format_duration(_eta['avg_row_seconds'])} | "
                f"recent avg {format_duration(_eta['recent_avg_row_seconds'])} | "
                f"remaining ~{format_duration(_eta['remaining_seconds'])} | "
                f"finish around {_eta['estimated_finish_local']}",
                flush=True,
            )

    _enrichment_elapsed = _time.monotonic() - _start_ts
    print(
        f"[enricher] Enrichment complete | companies={total} | errors={_error_count} | "
        f"enrichment time {format_duration(_enrichment_elapsed)}",
        flush=True,
    )

    # ── Buzzi runtime assertion (Italy profile) ──────────────────────────────
    if _cli_scoring_profile == "italy_register_icp_only":
        for _r in results:
            _rname = str(_r.get("company_name") or _r.get("canonical_company_name") or "").upper()
            _r_ev  = str(_r.get("sig_foreign_hq_evidence") or "").lower()
            if "BUZZI" in _rname and "casale monferrato" in _r_ev:
                _r_score = _r.get("sig_foreign_hq_score", "")
                try:
                    _r_score_int = int(float(_r_score))
                except (TypeError, ValueError):
                    _r_score_int = -1
                if _r_score_int != 0:
                    print(
                        f"[enricher] ASSERTION FAILED: {_rname} sig_foreign_hq_score={_r_score_int} "
                        f"(expected 0). Evidence: {str(_r.get('sig_foreign_hq_evidence', ''))[:200]}",
                        flush=True,
                    )
                    print("[enricher] ERROR: Buzzi foreign HQ sanitization did not work. Aborting.", file=__import__('sys').stderr)
                    __import__('sys').exit(3)
                else:
                    print(f"[enricher] ASSERTION OK: {_rname} sig_foreign_hq_score=0 (correctly sanitized)", flush=True)

    # ── Foreign HQ hygiene summary ────────────────────────────────────────────
    _fhq_sanitized   = sum(1 for r in results if r.get("foreign_hq_sanitized"))
    _fhq_review      = sum(1 for r in results
                           if "Ambiguous foreign HQ" in str(r.get("model_signal_manual_review_reason", "")))
    _inferred_ctry   = results[0].get("inferred_input_country", "") if results else ""
    print("", flush=True)
    print("[enricher] FOREIGN HQ HYGIENE:", flush=True)
    print(f"[enricher]   inferred input country:              {_inferred_ctry or 'unknown'}", flush=True)
    print(f"[enricher]   sanitized domestic intl footprints:  {_fhq_sanitized}", flush=True)
    print(f"[enricher]   ambiguous foreign HQ rows for review:{_fhq_review}", flush=True)
    print("", flush=True)

    # ── Build output dataframe ────────────────────────────────────────────────
    _active_fields = ALL_ENRICHMENT_FIELDS + EMPLOYEE_RANGE_RESOLVER_FIELDS
    df_out = df_in.head(len(results)).copy().reset_index(drop=True)
    enriched_df_raw = pd.DataFrame(results)
    for col in _active_fields:
        df_out[col] = enriched_df_raw[col].values if col in enriched_df_raw.columns else ""

    # ── Patch canonical identity + scoring_profile from enrichment results ────
    # These fields are NOT in ALL_ENRICHMENT_FIELDS (they come from input normalization)
    # but enrich_one_row now writes them. Backfill blanks from enrichment.
    _CANONICAL_PATCH_COLS = (
        "canonical_company_name", "canonical_company_domain",
        "canonical_company_url", "scoring_profile",
    )
    for _cp_col in _CANONICAL_PATCH_COLS:
        if _cp_col in enriched_df_raw.columns:
            _enr_vals = list(enriched_df_raw[_cp_col])
            if _cp_col not in df_out.columns:
                df_out[_cp_col] = _enr_vals
            else:
                _cur = df_out[_cp_col].astype(str).str.strip()
                _blank_mask = _cur.isin(["", "nan", "None"])
                df_out.loc[_blank_mask, _cp_col] = [
                    _enr_vals[j] for j in df_out.index[_blank_mask]
                ]

    # ── Employee range resolver ───────────────────────────────────────────────
    def _is_blank_val(v) -> bool:
        return v is None or str(v).strip() in ("", "nan", "None", "N/A", "-")

    _er_records = df_out.to_dict("records")
    _er_results: list[dict] = []
    for _rec in _er_records:
        _cname  = str(_rec.get("lusha_company_name") or _rec.get("company_name") or "")
        _domain = str(_rec.get("canonical_company_domain") or _rec.get("domain") or "")
        _er = resolve_employee_range(_rec, company_name=_cname)
        if _er["employee_range_confidence"] in ("None", "Low") and serper_key:
            _er_s = resolve_employee_range_from_serper(_cname, _domain, serper_key)
            if _er_s.get("employee_range_resolved"):
                _er = _er_s
        if _er.get("employee_range_resolved") and _er.get("employee_range_confidence") in ("High", "Medium"):
            _er["employee_range_for_scoring"]        = _er["employee_range_resolved"]
            _er["employee_range_for_scoring_source"] = _er["employee_range_source"]
        else:
            _er["employee_range_for_scoring"]        = DEFAULT_EMPLOYEE_RANGE_FOR_SCORING
            _er["employee_range_for_scoring_source"] = "default_commercial_minimum_assumption"
            if not _er.get("employee_range_notes"):
                _er["employee_range_notes"] = (
                    "No employee count found; default commercial minimum range used for scoring only."
                )
            if _er["employee_range_confidence"] in ("None",):
                _er["employee_range_confidence"] = "Low"
        _er_results.append(_er)

    for col in EMPLOYEE_RANGE_RESOLVER_FIELDS:
        df_out[col] = [r.get(col, "") for r in _er_results]

    _resolved_vals  = [r.get("employee_range_resolved", "") for r in _er_results]
    _resolved_confs = [r.get("employee_range_confidence", "None") for r in _er_results]
    if "lusha_employee_range" not in df_out.columns:
        df_out["lusha_employee_range"] = ""
    _existing_lusha = df_out["lusha_employee_range"].tolist()
    df_out["lusha_employee_range"] = [
        _resolved_vals[i]
        if (_is_blank_val(_existing_lusha[i])
            and _resolved_vals[i]
            and _resolved_confs[i] in ("High", "Medium"))
        else _existing_lusha[i]
        for i in range(len(df_out))
    ]

    # ── Commercial scoring ────────────────────────────────────────────────────
    print(f"[enricher] Applying scoring profile: {_cli_scoring_profile}", flush=True)
    try:
        df_out = apply_results_compatible_scoring(df_out, _cli_scoring_profile)
    except Exception as _score_exc:
        print(f"[enricher] Scoring skipped: {_score_exc}", flush=True)

    try:
        df_out = apply_competitor_icp_override(df_out)
    except Exception as _ov_exc:
        print(f"[enricher] ICP override skipped: {_ov_exc}", flush=True)

    # ── Write output ──────────────────────────────────────────────────────────
    if args.output_name:
        _out_fname = f"{args.output_name}.xlsx"
    else:
        _out_fname = build_enriched_output_filename(str(input_path))
    xl_path = out_dir / _out_fname
    print(f"[enricher] Output filename:  {_out_fname}", flush=True)
    print(f"[enricher] Output path:      {xl_path}", flush=True)

    _export_start = _time.monotonic()
    try:
        xl_bytes = build_rich_excel_bytes(
            df_out,
            name_col=company_col,
            domain_col=domain_col or None,
            df_input_original=df_in,
            scoring_profile=_cli_scoring_profile,
            run_config=_run_config,
            run_mode="cli",
        )
        xl_path.write_bytes(xl_bytes)
        print(f"[enricher] Saved: {xl_path}", flush=True)
    except Exception as exc:
        print(f"[enricher] Rich Excel failed ({exc}), falling back to flat Excel.", flush=True)
        try:
            df_to_excel_bytes_write(df_out, str(xl_path))
            print(f"[enricher] Saved (flat): {xl_path}", flush=True)
        except Exception as exc2:
            print(f"[enricher] ERROR: could not write output file: {exc2}", file=sys.stderr)
            sys.exit(2)
    _export_elapsed = _time.monotonic() - _export_start

    if not xl_path.exists():
        print(f"[enricher] ERROR: output file was not created: {xl_path}", file=sys.stderr)
        sys.exit(2)

    _total_runtime = _time.monotonic() - _start_ts

    # ── Evidence QA summary ───────────────────────────────────────────────────
    _ev_with    = sum(1 for r in results if r.get("raw_google_evidence_count", 0))
    _ev_without = len(results) - _ev_with
    _ev_total   = sum(int(r.get("raw_google_evidence_count", 0) or 0) for r in results)
    _url_total  = sum(
        len([u for u in str(r.get("raw_google_evidence_urls", "") or "").split("\n") if u])
        for r in results
    )
    _new_cols = [
        "raw_google_evidence_count", "raw_google_evidence_urls",
        "raw_google_evidence_combined", "raw_google_evidence_json",
        "raw_google_evidence_json_01", "raw_google_evidence_json_02",
        "raw_google_evidence_json_03", "raw_google_evidence_json_parts",
        "raw_google_evidence_truncated",
        "google_snippet_01_title", "google_snippet_01_url",
    ]
    print(f"[QA] Output Excel:           {xl_path}", flush=True)
    print(f"[QA] Sheet with evidence:    Opportunity Input", flush=True)
    print(f"[QA] Companies with evidence:{_ev_with}", flush=True)
    print(f"[QA] Companies without:      {_ev_without}", flush=True)
    print(f"[QA] Total snippet records:  {_ev_total}", flush=True)
    print(f"[QA] Total URLs written:     {_url_total}", flush=True)
    print(f"[QA] New evidence columns:   {_new_cols}", flush=True)

    print(
        f"[enricher] BATCH COMPLETE | "
        f"companies={total} | "
        f"enrichment time {format_duration(_enrichment_elapsed)} | "
        f"final export time {format_duration(_export_elapsed)} | "
        f"total runtime {format_duration(_total_runtime)} | "
        f"output={xl_path}",
        flush=True,
    )




def run_streamlit_app() -> None:
    _st, components = get_streamlit()
    # =============================================================================
    # UI
    # =============================================================================

    import pathlib as _pl
    import base64 as _b64
    _st.set_page_config(
        page_title="mYngle · lead prioritizer",
        page_icon="🏢",
        layout="wide",
        initial_sidebar_state="collapsed",
    )

    _logo = _pl.Path(__file__).parent / "mingle_local_final_fixed.png"
    _logo_src = (
        f"data:image/png;base64," + _b64.b64encode(_logo.read_bytes()).decode()
        if _logo.exists() else ""
    )
    _img_tag = (
        f'<img src="{_logo_src}" class="brand-logo" alt="mYngle" />'
        if _logo_src else ""
    )
    _st.markdown(
        f"""
        <style>
        .block-container {{
            max-width: 880px;
            padding-top: 2.2rem;
            padding-bottom: 3rem;
            padding-left: 2rem;
            padding-right: 2rem;
        }}
        div[data-testid="stMarkdownContainer"]:has(.brand-header) {{
            overflow: visible !important;
            margin-bottom: 1.0rem;
        }}
        .brand-header {{
            display: grid;
            grid-template-columns: 43% 57%;
            align-items: center;
            min-height: 140px;
            padding-top: 10px;
            padding-bottom: 6px;
            overflow: visible !important;
        }}
        .brand-title-block {{
            display: flex;
            align-items: center;
            justify-content: flex-start;
            overflow: visible !important;
        }}
        .brand-title {{
            font-size: 42px;
            font-weight: 700;
            color: #0B1F3A;
            line-height: 1.1;
            white-space: nowrap;
            margin: 0;
            padding: 0;
        }}
        .brand-logo-block {{
            display: flex;
            justify-content: flex-end;
            align-items: center;
            padding: 0;
            line-height: 0;
            overflow: visible !important;
        }}
        .brand-logo {{
            width: 430px;
            max-width: 100%;
            height: auto;
            display: block;
            object-fit: contain;
            object-position: center center;
            overflow: visible !important;
        }}
        </style>
        <div class="brand-header">
          <div class="brand-title-block">
            <span class="brand-title">lead prioritizer</span>
          </div>
          <div class="brand-logo-block">
            {_img_tag}
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # =============================================================================
    # API KEY — secrets only, no sidebar input
    # =============================================================================

    api_key = ""
    _api_key_error = ""
    try:
        api_key = (_st.secrets.get("ANTHROPIC_API_KEY", "") or "").strip()
    except Exception:
        pass

    if not api_key:
        _api_key_error = (
            "**ANTHROPIC_API_KEY not found in Streamlit secrets.** "
            "Add it to your app's secrets: Settings → Secrets → "
            "`ANTHROPIC_API_KEY = \"sk-ant-...\"`"
        )

    # Serper key — optional, only needed when Step 2 provider is Serper Google Search.
    # Required entry in .streamlit/secrets.toml: SERPER_API_KEY = "your-key"
    serper_key = ""
    try:
        serper_key = (_st.secrets.get("SERPER_API_KEY", "") or "").strip()
    except Exception:
        pass

    # Lusha API key — optional, only needed when Lusha API enrichment is enabled.
    # Required entry in .streamlit/secrets.toml: LUSHA_API_KEY = "your-key"
    lusha_api_key = ""
    try:
        lusha_api_key = (_st.secrets.get("LUSHA_API_KEY", "") or "").strip()
    except Exception:
        pass

    # =============================================================================
    # SIDEBAR — shown only when SHOW_ADVANCED_SETTINGS is True
    # =============================================================================

    # Re-check at runtime so the Streamlit secret can override the module constant.
    _show_adv: bool = SHOW_ADVANCED_SETTINGS
    try:
        _show_adv = bool(_st.secrets.get("SHOW_ADVANCED_SETTINGS", SHOW_ADVANCED_SETTINGS))
    except Exception:
        pass

    # _adv_main gates verbose sections on the MAIN PAGE (step previews, column
    # selectors, live metrics, intermediate downloads, cost estimates, etc.).
    # Kept False so the main page stays clean even when the sidebar is visible.
    # Column/row controls live in the sidebar "Advanced data options" expander.
    _adv_main: bool = False

    # ── Hardcoded defaults for all removed sidebar controls ───────────────────────
    _elm_mode  = False
    debug_mode = False
    delay_sec  = 1.0
    ss_set(
        _enable_lusha_api          = False,  # Lusha live API disabled in Layer 1
        _run_step1_enrichment      = not ss("_has_lusha_input", False),
        _run_step2_enrichment      = True,
        _extract_model_signals     = True,
        _include_signal_evidence   = True,
        _step2_dry_run             = False,
        _zero_cost_preview         = False,
        _show_step2_debug          = False,
        _save_step2_debug          = False,
        _use_playwright            = _PLAYWRIGHT_AVAILABLE,
        _local_save_enabled        = False,
        _local_save_path           = _DEFAULT_DOWNLOAD_DIR,
        _step2_provider            = STEP2_PROVIDER_SERPER,
        _elm_mode                  = False,
        _per_company_autosave_enabled = False,
    )

    # ── Simplified sidebar ─────────────────────────────────────────────────────────
    with _st.sidebar:
        _st.header("Settings")

        # API key statuses
        if api_key:
            _st.success("✓ Anthropic API key loaded")
        else:
            _st.error("⚠ Anthropic API key missing")
        if serper_key:
            _st.success("✓ Serper API key loaded")
        else:
            _st.error("⚠ Serper API key missing — required for Step 2 search")
        _st.sidebar.info(
            "**Lusha/Lucia live enrichment** is disabled in this layer. "
            "Contact enrichment runs in **Buyer Contact Finder (Layer 2.5)** "
            "after company selection, to preserve credits."
        )

        _st.divider()

        # Model selectors
        model_step1_label = _st.selectbox(
            "Model — Step 1 (firmographics)",
            options=list(AVAILABLE_MODELS.keys()),
            index=0,
            help="Used for extracting structured company data from scraped pages. Haiku is sufficient here.",
        )
        model_step2_label = _st.selectbox(
            "Model — Step 2 (ICP web search)",
            options=list(AVAILABLE_MODELS.keys()),
            index=0,
            help="Used for the agentic web search. Sonnet gives better signal detection but costs ~5x more.",
        )
        selected_model_step1 = AVAILABLE_MODELS[model_step1_label]
        selected_model_step2 = AVAILABLE_MODELS[model_step2_label]
        _st.session_state["_model_step1"] = selected_model_step1
        _st.session_state["_model_step2"] = selected_model_step2

        _st.divider()

        # Excel autosave
        _st.subheader("📄 Excel autosave")
        _xl_enabled = _st.checkbox(
            "Enable Excel autosave",
            value=ss("_xl_autosave_enabled", True),
            key="_xl_autosave_enabled",
            help=(
                "Writes intermediate results to a local Excel file while the app is running. "
                f"Saves to  {_XL_AUTOSAVE_DIR}/{_XL_AUTOSAVE_DEFAULT}  next to the app."
            ),
        )
        if _xl_enabled:
            _xl_every = _st.number_input(
                "Autosave every N companies",
                min_value=1,
                value=int(ss("_xl_autosave_every", _XL_AUTOSAVE_EVERY)),
                step=1,
                key="_xl_autosave_every",
                help="Writes intermediate results to a local Excel file while the app is running.",
            )
            _xl_fname = _st.text_input(
                "Autosave Excel filename",
                value=ss("_xl_autosave_filename", _XL_AUTOSAVE_DEFAULT),
                key="_xl_autosave_filename",
            )
            _st.caption(
                f"📁 Saves to: **{_XL_AUTOSAVE_DIR}/{_xl_fname or _XL_AUTOSAVE_DEFAULT}**"
            )
            _xl_last = ss("_xl_autosave_last_msg", "")
            if _xl_last:
                if _xl_last.startswith("⚠"):
                    _st.warning(_xl_last)
                else:
                    _st.caption(f"Last save: {_xl_last}")
        else:
            ss_set(_xl_autosave_every=_XL_AUTOSAVE_EVERY, _xl_autosave_filename=_XL_AUTOSAVE_DEFAULT)

    # =============================================================================
    # INPUT MODE
    # =============================================================================

    _sc_df: pd.DataFrame | None = None
    _sc_name_col   = "company_name"
    _sc_domain_col = "domain"

    # Input mode is always Batch Upload for the normal user flow.
    _app_mode = "Batch Upload"

    # =============================================================================
    # STEP 1 — Upload file  (Batch mode only)
    # =============================================================================

    uploaded = None
    if _app_mode == "Batch Upload":
        _st.divider()
        if _show_adv:
            _st.subheader("Step 1 · Upload your file")
        uploaded = _st.file_uploader(
            "Drag and drop here, or click to browse  (.xlsx · .xls · .csv)",
            type=["xlsx", "xls", "csv"],
            label_visibility="collapsed" if not _show_adv else "visible",
        )

    new_file_key = f"{uploaded.name}___{uploaded.size}" if uploaded else "__none__"
    if new_file_key != ss("_file_key"):
        ss_set(_file_key=new_file_key, df_raw=None, file_name=None, file_error=None,
               _selected_sheet=None, _detected_input_type_ui=None)
        reset_processing()
        if uploaded is not None:
            try:
                import io as _io
                fname = uploaded.name
                _raw_bytes = uploaded.read()
                if fname.lower().endswith(".csv"):
                    df_loaded    = pd.read_csv(_io.BytesIO(_raw_bytes))
                    _sel_sheet   = None
                    _det_itype   = (
                        "pre_enriched_lucia_export" if is_lucia_contact_export(df_loaded)
                        else "input_cleaner_output"  if is_input_cleaner_output(df_loaded, fname)
                        else "simple_company_list"
                    )
                else:
                    # Peek at sheet names to decide loading strategy
                    _xl_peek = pd.ExcelFile(_io.BytesIO(_raw_bytes))
                    _has_cleaner_sheets = bool(
                        set(_xl_peek.sheet_names) & _CLEANER_SHEET_NAMES
                    )
                    if _has_cleaner_sheets or _fname_looks_like_cleaner(fname):
                        df_loaded, _sel_sheet = load_cleaner_workbook(
                            _io.BytesIO(_raw_bytes), fname
                        )
                        _det_itype = (
                            "pre_enriched_lucia_export" if is_lucia_contact_export(df_loaded)
                            else "input_cleaner_output"  if is_input_cleaner_output(df_loaded, fname)
                            else "simple_company_list"
                        )
                    else:
                        df_loaded  = _xl_peek.parse(_xl_peek.sheet_names[0])
                        _sel_sheet = _xl_peek.sheet_names[0]
                        _det_itype = (
                            "pre_enriched_lucia_export" if is_lucia_contact_export(df_loaded)
                            else "input_cleaner_output"  if is_input_cleaner_output(df_loaded, fname)
                            else "simple_company_list"
                        )

                _det_scoring_profile = resolve_active_scoring_profile(
                    input_filename=fname,
                    df=df_loaded,
                    selected_sheet=_sel_sheet or "",
                    detected_input_type=_det_itype,
                )
                ss_set(df_raw=df_loaded, file_name=fname,
                       _selected_sheet=_sel_sheet, _detected_input_type_ui=_det_itype,
                       _scoring_profile=_det_scoring_profile)
                _is_lucia_loaded = (_det_itype == "pre_enriched_lucia_export")
                _detected_lusha  = detect_lusha_columns(df_loaded)
                # _has_lusha_input = True ONLY for true Lucia/Lusha exports, never
                # for cleaner output (even if it happens to contain Lusha-like cols).
                _has_lusha = bool(_detected_lusha) and _is_lucia_loaded
                ss_set(
                    _lusha_cols_in_input=_detected_lusha,
                    _has_lusha_input=_has_lusha,
                    _is_lucia_export=_is_lucia_loaded,
                    _is_cleaner_output=(_det_itype == "input_cleaner_output"),
                )
            except Exception as exc:
                ss_set(
                    file_error=str(exc),
                    _lusha_cols_in_input=[], _has_lusha_input=False,
                    _is_lucia_export=False, _is_cleaner_output=False,
                    _scoring_profile="default",
                )

    df_raw: pd.DataFrame | None = ss("df_raw")
    file_error: str | None      = ss("file_error")

    if file_error:
        _st.error(f"Could not read the file: {file_error}")
    elif uploaded and df_raw is not None:
        _ui_fname      = ss("file_name", "")
        _ui_sheet      = ss("_selected_sheet")
        _ui_itype      = ss("_detected_input_type_ui", "simple_company_list")
        _itype_labels  = {
            "pre_enriched_lucia_export": "Lucia/Lusha contact export",
            "input_cleaner_output":      "Input Cleaner output",
            "simple_company_list":       "Simple company list",
        }
        _itype_label = _itype_labels.get(_ui_itype, _ui_itype)

        if ss("_is_lucia_export", False):
            # Lucia/Lusha: count unique companies
            _l_domain_col = get_lucia_domain_col(df_raw)
            _l_name_col   = get_lucia_name_col(df_raw)
            if _l_domain_col:
                _dedup_keys = df_raw[_l_domain_col].apply(
                    lambda x: clean_domain(str(x)) if pd.notna(x) else ""
                )
                if _l_name_col:
                    _dedup_keys = _dedup_keys.where(
                        _dedup_keys != "",
                        df_raw[_l_name_col].apply(lambda x: str(x).strip().lower()),
                    )
            elif _l_name_col:
                _dedup_keys = df_raw[_l_name_col].apply(lambda x: str(x).strip().lower())
            else:
                _dedup_keys = pd.Series(range(len(df_raw))).astype(str)
            _n_contacts  = len(df_raw)
            _n_companies = _dedup_keys.nunique()
            _st.success(
                f"✓ **{_ui_fname}** loaded · "
                f"{_n_contacts:,} contact rows · {_n_companies:,} unique companies ready  \n"
                f"Detected input type: {_itype_label}  \n"
                f"Detected company column: `{_l_name_col or 'Company Name'}`  \n"
                f"Detected domain column: `{_l_domain_col or 'none'}`"
            )
        else:
            _det_name, _det_dom = detect_columns(df_raw, _ui_fname or "")
            if _det_name and _det_name in df_raw.columns:
                _t1_count = int(
                    df_raw[_det_name]
                    .dropna()
                    .astype(str)
                    .str.strip()
                    .replace("", pd.NA)
                    .dropna()
                    .nunique()
                )
            else:
                _t1_count = len(df_raw)
            _t1_row_count = len(df_raw)
            _sheet_line   = f"Selected sheet: `{_ui_sheet}`  \n" if _ui_sheet else ""
            _dom_label    = f"`{_det_dom}`" if _det_dom else "*none — will search*"
            _st.success(
                f"✓ **{_ui_fname}** loaded · "
                f"{_t1_row_count:,} rows · {_t1_count:,} "
                f"{'company' if _t1_count == 1 else 'companies'} ready  \n"
                f"Detected input type: {_itype_label}  \n"
                + _sheet_line +
                f"Detected company column: `{_det_name or 'none'}`  \n"
                f"Detected domain column: {_dom_label}"
            )
            if _t1_row_count >= 50 and _t1_count <= 10:
                _st.warning(
                    "Detected company column has very few unique values. "
                    "Please check column mapping."
                )
            # Show resolved run config so user sees it before starting
            _resolved_sp = ss("_scoring_profile", "default")
            from commercial_fit_scoring import SCORING_PROFILES as _SP
            _sp_display = _SP.get(_resolved_sp, _SP["default"])
            _sp_label = (
                "Italy register ICP only"
                if _resolved_sp == "italy_register_icp_only"
                else "Default (ICP 90% + size 10%)"
            )
            _ui_rc = resolve_active_run_config(
                input_filename=ss("file_name", ""),
                df=df_raw,
                selected_sheet=ss("_selected_sheet", ""),
                detected_input_type=ss("_detected_input_type_ui", ""),
                has_existing_lusha_data=bool(ss("_has_lusha_input", False)),
                user_overrides={
                    "model_step1": ss("_model_step1", MODEL_STEP1),
                    "model_step2": ss("_model_step2", MODEL_STEP2),
                },
            )
            _st.info(
                f"**Resolved scoring profile:** {_sp_label}  \n"
                f"K = {_sp_display.get('sigmoid_k', 10)} · "
                f"Model weight = {int(_sp_display.get('model_weight', 0.9)*100)}% · "
                f"Size weight = {int(_sp_display.get('size_weight', 0.1)*100)}%  \n"
                f"**Step 1 (firmographics):** "
                f"{'enabled' if _ui_rc['run_step1_enrichment'] else 'skipped - Lusha data present'}  \n"
                f"**Step 2 (ICP search):** "
                f"{'enabled' if _ui_rc['run_step2_enrichment'] else 'disabled'}  \n"
                f"**Model signals:** "
                f"{'enabled' if _ui_rc['extract_model_signals'] else 'disabled'}  \n"
                f"**Search provider:** {_ui_rc['search_provider']}  \n"
                f"**Input type:** {ss('_detected_input_type_ui', 'unknown')}"
            )
    # ── Column detection and processing scope ─────────────────────────────────────

    name_col     = _sc_name_col if _app_mode == "Single Company" else None
    domain_col   = _sc_domain_col if _app_mode == "Single Company" else None
    n_to_process = 1 if (_app_mode == "Single Company" and _sc_df is not None) else 0

    if df_raw is not None:
        # ── Sidebar: Advanced data options (column selector + row limiter) ─────────
        if _show_adv:
            with _st.sidebar:
                _st.divider()
                with _st.expander("⚙ Advanced data options", expanded=False):
                    _st.caption(f"{len(df_raw):,} rows · {len(df_raw.columns)} columns")
                    _adv_cols = df_raw.columns.tolist()
                    _adv_auto_n, _adv_auto_d = detect_columns(df_raw)
                    _adv_n_idx = _adv_cols.index(_adv_auto_n) if _adv_auto_n in _adv_cols else 0
                    _st.selectbox(
                        "Company name column",
                        options=_adv_cols,
                        index=_adv_n_idx,
                        key="adv_name_col",
                    )
                    _NODOM = "(none — auto)"
                    _adv_dom_opts = [_NODOM] + _adv_cols
                    _adv_dom_def = (
                        _adv_dom_opts.index(_adv_auto_d)
                        if _adv_auto_d and _adv_auto_d in _adv_dom_opts else 0
                    )
                    _st.selectbox(
                        "URL column (optional)",
                        options=_adv_dom_opts,
                        index=_adv_dom_def,
                        key="adv_domain_col_sel",
                    )
                    _st.checkbox("Limit rows for testing", value=False, key="adv_limit_rows")
                    if ss("adv_limit_rows", False):
                        _st.number_input(
                            "Number of rows",
                            min_value=1, max_value=len(df_raw),
                            value=min(5, len(df_raw)), step=1,
                            key="adv_row_limit_n",
                        )

        # ── Silently resolve columns and row count ─────────────────────────────────
        _det_name, _det_dom = detect_columns(df_raw)
        if _show_adv:
            _sid_n = ss("adv_name_col", None)
            name_col = _sid_n if (_sid_n and _sid_n in df_raw.columns) else _det_name
            _sid_d = ss("adv_domain_col_sel", None)
            domain_col = (
                _sid_d if (_sid_d and _sid_d not in ("(none — auto)", None)
                           and _sid_d in df_raw.columns)
                else _det_dom
            )
            if ss("adv_limit_rows", False):
                n_to_process = int(ss("adv_row_limit_n", len(df_raw)) or len(df_raw))
            else:
                n_to_process = len(df_raw)
        else:
            name_col     = _det_name
            domain_col   = _det_dom
            n_to_process = len(df_raw)

    # =============================================================================
    # STEP 5 — Start enrichment
    # =============================================================================

    _st.divider()
    currently_processing = ss("processing", False)
    enrichment_done      = ss("enrichment_done", False)

    blocking: list = []
    _active_provider     = ss("_step2_provider",    STEP2_PROVIDER_SERPER)
    _active_dry_run      = ss("_step2_dry_run",     False)
    _active_zero_cost    = ss("_zero_cost_preview", False)
    _is_preview_mode     = _active_dry_run or _active_zero_cost

    # (1) Claude real run → Anthropic key required
    # (2) Serper real run → both Anthropic + Serper keys required
    # (3) Dry run only   → no API keys required
    # (4) Zero-cost      → no API keys required
    if _api_key_error and not _elm_mode and not _is_preview_mode:
        blocking.append(_api_key_error)
    if _active_provider == STEP2_PROVIDER_CLAUDE:
        blocking.append(
            "Claude Web Search is disabled. Serper Google Search is the only permitted provider."
        )
    if (
        _active_provider == STEP2_PROVIDER_SERPER
        and not serper_key
        and not _elm_mode
        and not _is_preview_mode
    ):
        blocking.append("SERPER_API_KEY is missing from .streamlit/secrets.toml")
    _active_lusha_api = False  # Lusha live API disabled in Layer 1
    # (no blocking check needed — Lusha live API is disabled in Layer 1)
    if _app_mode == "Batch Upload" and uploaded is None:
        blocking.append("Upload a file to start.")
    if _app_mode == "Single Company" and (_sc_df is None or _sc_df.empty):
        blocking.append("Enter a company name to proceed.")
    if file_error:
        blocking.append(f"File could not be read: {file_error}")
    if df_raw is not None and name_col is None:
        blocking.append("No company name column selected.")
    if df_raw is not None and n_to_process == 0:
        blocking.append("Zero rows selected for processing.")

    if blocking and not currently_processing:
        for reason in blocking:
            _st.warning(f"⚠️ {reason}")
    elif not blocking and not currently_processing and not enrichment_done:
        if _adv_main:
            _s1 = ss("_model_step1", MODEL_STEP1)
            _s2 = ss("_model_step2", MODEL_STEP2)
            if _is_preview_mode or _elm_mode:
                _st.info(
                    f"Ready to preview **{n_to_process:,}** rows. "
                    "Estimated cost: **$0.00** — no API calls will be made."
                )
            else:
                _cost_per_row = _COST_EST.get((_s1, _s2), 0.05)
                est = n_to_process * _cost_per_row
                _st.info(
                    f"Ready to enrich **{n_to_process:,}** rows with two enrichment steps each. "
                    f"Rough estimated cost: ~${est:.2f} "
                    f"(~${_cost_per_row:.2f}/company with current model selection)."
                )

    _start_label = "▶ Enrich & Score" if _app_mode == "Single Company" else "▶ Start enrichment"
    start_btn = _st.button(
        _start_label,
        type="primary",
        use_container_width=True,
        disabled=(bool(blocking) or currently_processing),
        key="start_button",
    )

    if start_btn and not blocking and not currently_processing:
        if _app_mode == "Single Company" and _sc_df is not None:
            df_work    = _sc_df.copy()
            name_col   = _sc_name_col
            domain_col = _sc_domain_col
            _df_raw_for_input = None
        else:
            _is_lucia_run  = ss("_is_lucia_export", False)
            _is_cleaner_run = ss("_is_cleaner_output", False)
            _input_type    = (
                "pre_enriched_lucia_export" if _is_lucia_run
                else "input_cleaner_output" if _is_cleaner_run
                else "simple_company_list"
            )
            _norm_result   = normalize_input_to_company_df(
                df_raw, _input_type, name_col, domain_col
            )
            df_work            = _norm_result["company_df"].head(n_to_process).copy()
            # After deduplication Type 2 may have fewer rows than n_to_process;
            # update so the processing loop bound matches the actual work df.
            n_to_process       = len(df_work)

            # After normalization, ALL downstream processing uses canonical columns.
            # Preserve the original detected cols in metadata only.
            name_col   = "canonical_company_name"
            domain_col = "canonical_company_url"

            # Always preserve the original uploaded df for the Input sheet.
            _df_raw_for_input = df_raw.copy()

            # ── Hard validation: canonical columns must be present ────────────────
            _missing_canonical = [
                c for c in ("canonical_company_name", "canonical_company_domain",
                            "canonical_company_url", "input_type", "source_contact_count")
                if c not in df_work.columns
            ]
            if _missing_canonical:
                _st.error(
                    "❌ Normalization error — required canonical columns missing: "
                    + ", ".join(_missing_canonical)
                )
                _st.stop()

            _empty_names = df_work["canonical_company_name"].astype(str).str.strip().replace("", pd.NA).isna().sum()
            if _empty_names:
                _st.error(f"❌ {_empty_names} row(s) have an empty canonical_company_name after normalization.")
                _st.stop()

            # ── Canonical identity validation (Type 2) ────────────────────────────
            if _is_lucia_run:
                _person_names = set(
                    df_raw.get("First Name", pd.Series(dtype=str))
                    .dropna().astype(str).str.strip()
                    .replace("", pd.NA).dropna()
                )
                _canon_names  = set(df_work["canonical_company_name"].astype(str).str.strip())
                _leaked_names = _person_names & _canon_names
                _bad_domains  = [
                    d for d in df_work["canonical_company_domain"].astype(str)
                    if "linkedin.com" in d.lower()
                ]
                _bad_urls = [
                    u for u in df_work["canonical_company_url"].astype(str)
                    if "linkedin.com/in/" in u.lower()
                ]
                if _leaked_names or _bad_domains or _bad_urls:
                    _detail = []
                    if _leaked_names:
                        _detail.append(f"Person names in company identity: {sorted(_leaked_names)}")
                    if _bad_domains:
                        _detail.append(f"LinkedIn URLs in canonical_company_domain: {_bad_domains}")
                    if _bad_urls:
                        _detail.append(f"LinkedIn URLs in canonical_company_url: {_bad_urls}")
                    _st.error(
                        "❌ Canonical identity validation failed — contact fields leaked into "
                        "company identity. Fix the normalization layer before proceeding.\n\n"
                        + "\n".join(_detail)
                    )
                    _st.stop()

            # Store normalization metadata for downstream validation
            _norm_meta = {
                "input_type":           _norm_result["input_type"],
                "contact_row_count":    _norm_result["contact_row_count"],
                "unique_company_count": _norm_result["unique_company_count"],
                "mapping_notes":        _norm_result["mapping_notes"],
            }
            ss_set(_input_norm_meta=_norm_meta)
        resume_mode  = ss("_resume_mode", False)
        if not resume_mode:
            autosave_clear()   # wipe any previous autosave on a fresh start

        # ── Create per-company autosave run folder if feature is enabled ──────────
        _pca_run_dir_new = ""
        _pca_enabled_now = ss("_per_company_autosave_enabled", False)
        if _pca_enabled_now:
            _pca_base = (
                ss("_per_company_autosave_base_folder", "") or _PER_COMPANY_AUTOSAVE_DEFAULT_DIR
            )
            try:
                # Build a temporary run tag from current sidebar selections
                _tmp_prov  = ss("_step2_provider", STEP2_PROVIDER_SERPER)
                _tmp_model = ss("_model_step2",    MODEL_STEP2)
                _tmp_lusha = ss("_enable_lusha_api", False)
                _tmp_tag   = f"{get_provider_code(_tmp_prov)}_{get_model_code(_tmp_model)}"
                if _tmp_lusha:
                    _tmp_tag = f"{_tmp_tag}_lusha"
                _pca_run_dir_new = create_run_folder(_pca_base, _tmp_tag)
            except Exception as _pca_err:
                _st.warning(f"⚠ Could not create autosave run folder: {_pca_err}")
                _pca_run_dir_new = ""

        # ── Build and lock run config for this run ────────────────────────────
        _arc = resolve_active_run_config(
            input_filename=ss("file_name", ""),
            df=ss("df_raw"),
            selected_sheet=ss("_selected_sheet", ""),
            detected_input_type=ss("_detected_input_type_ui", ""),
            has_existing_lusha_data=bool(ss("_has_lusha_input", False)),
            user_overrides={
                "model_step1": ss("_model_step1", MODEL_STEP1),
                "model_step2": ss("_model_step2", MODEL_STEP2),
            },
        )

        ss_set(
            processing=True, stop_requested=False,
            process_index=0, results=[], debug_records=[],
            enrichment_done=False, df_enriched=None,
            _df_work=df_work, _name_col=name_col, _domain_col=domain_col,
            _n_to_process=n_to_process, _api_key=api_key, _delay=delay_sec,
            total_tokens_in=0, total_tokens_out=0, total_cost_usd=0.0,
            total_cache_read_tokens=0, total_cache_create_tokens=0,
            _resume_mode=resume_mode, autosave_last_name="",
            _run_start_time=__import__("time").time(),
            _elm_mode=_elm_mode,
            _active_fields=ELM_ALL_FIELDS if _elm_mode else ALL_ENRICHMENT_FIELDS,
            _local_save_enabled=ss("_local_save_enabled", True),
            _final_auto_saved=False, _last_local_save="",
            _serper_key=serper_key,
            _step2_dry_run=ss("_step2_dry_run", False),
            _zero_cost_preview=ss("_zero_cost_preview", False),
            _enable_lusha_api=False,  # Lusha live API disabled in Layer 1
            _lusha_api_key=lusha_api_key,
            _dry_run_records=[], _search_output_records=[], _step2_debug_files=[],
            _dry_run_preview_count=0,
            # Lucia export: original (contact-level) df for the Input sheet
            _df_raw_original=_df_raw_for_input,
            # Per-company autosave
            _per_company_autosave_run_dir=_pca_run_dir_new,
            _per_company_autosave_last_saved="",
            _per_company_autosave_last_error="",
            _final_save_path="", _final_save_error="",
            # Active run config — single source of truth, locked at run start
            _active_run_config=_arc,
            _active_scoring_profile=_arc["scoring_profile"],
            # Keep legacy individual keys in sync for any residual reads
            _use_playwright=_arc["use_playwright"],
            _model_step1=_arc["model_step1"],
            _model_step2=_arc["model_step2"],
            _step2_provider=_arc["search_provider"],
            _extract_model_signals=_arc["extract_model_signals"],
            _include_signal_evidence=_arc["include_signal_evidence"],
            _run_step1_enrichment=_arc["run_step1_enrichment"],
            _run_step2_enrichment=_arc["run_step2_enrichment"],
        )
        _st.rerun()

    # =============================================================================
    # PROCESSING LOOP — one row per Streamlit rerun
    # =============================================================================

    if ss("processing", False):
        idx           = ss("process_index", 0)
        results       = ss("results", [])
        debug_records = ss("debug_records", [])
        df_work       = ss("_df_work")
        _name_col     = ss("_name_col")
        _domain_col   = ss("_domain_col")
        _n            = ss("_n_to_process", 0)
        _api_key      = ss("_api_key", "")
        _delay        = ss("_delay", 1.0)
        _elm_mode_run  = ss("_elm_mode", False)
        _active_fields = ss("_active_fields", ALL_ENRICHMENT_FIELDS)
        # Read from locked run config — never recompute mid-run
        _arc_run = ss("_active_run_config") or {}
        _use_playwright_run          = _arc_run.get("use_playwright",        ss("_use_playwright", True))
        _model_step1_run             = _arc_run.get("model_step1",           ss("_model_step1", MODEL_STEP1))
        _model_step2_run             = _arc_run.get("model_step2",           ss("_model_step2", MODEL_STEP2))
        _step2_provider_run          = _arc_run.get("search_provider",       ss("_step2_provider", STEP2_PROVIDER_SERPER))
        _serper_key_run              = ss("_serper_key", "")
        _dry_run_run                 = ss("_step2_dry_run", False)
        _zero_cost_run               = ss("_zero_cost_preview", False)
        _enable_lusha_api_run        = ss("_enable_lusha_api", False)
        _lusha_api_key_run           = ss("_lusha_api_key", "")
        _extract_model_signals_run   = _arc_run.get("extract_model_signals",   ss("_extract_model_signals", True))
        _include_signal_evidence_run = _arc_run.get("include_signal_evidence", ss("_include_signal_evidence", True))
        _run_step1_enrichment_run    = _arc_run.get("run_step1_enrichment",    ss("_run_step1_enrichment", True))
        _run_step2_enrichment_run    = _arc_run.get("run_step2_enrichment",    ss("_run_step2_enrichment", True))
        _scoring_profile             = _arc_run.get("scoring_profile",         ss("_active_scoring_profile", "default"))
        _pca_enabled_run  = ss("_per_company_autosave_enabled", False)
        _pca_run_dir_run  = ss("_per_company_autosave_run_dir", "")
        total_in          = ss("total_tokens_in", 0)
        total_out         = ss("total_tokens_out", 0)
        total_cost        = ss("total_cost_usd", 0.0)
        total_cache_read  = ss("total_cache_read_tokens", 0)
        total_cache_create = ss("total_cache_create_tokens", 0)

        if _st.button("⏹ Stop", key="stop_button"):
            ss_set(stop_requested=True)
            _st.rerun()

        # Look ahead to get the current company name for the status line
        _cur_company = ""
        if idx < _n and df_work is not None:
            try:
                _cur_company = str(df_work.iloc[idx].get("canonical_company_name", "")).strip()
            except Exception:
                pass
        if idx >= _n:
            _progress_text = f"Completed {_n} of {_n}"
            _st.progress(1.0, text=_progress_text)
        else:
            _progress_text = (
                f"Processing {idx + 1} of {_n}"
                + (f" · Current company: {_cur_company}" if _cur_company else "")
            )
            _st.progress(idx / _n if _n else 1.0, text=_progress_text)

        # ── ETA display ───────────────────────────────────────────────────────────
        _completed  = len(results)
        _start_time = ss("_run_start_time", None)
        if _start_time is not None and _n > 0:
            import time as _time_mod
            _elapsed = _time_mod.time() - _start_time
            if _completed >= 2:
                _avg_sec      = _elapsed / _completed
                _remaining    = max(_n - _completed, 0)
                _eta_sec      = int(_avg_sec * _remaining)
                if _eta_sec < 60:
                    _eta_str = f"about {_eta_sec} sec"
                elif _eta_sec < 3600:
                    _eta_str = f"about {_eta_sec // 60} min {_eta_sec % 60} sec"
                else:
                    _eta_str = f"about {_eta_sec // 3600} hr {(_eta_sec % 3600) // 60} min"
                _st.caption(f"Estimated time remaining: {_eta_str}")
            else:
                _st.caption("Estimating time remaining…")

        # ── Autosave status (show only when a save has occurred or failed) ─────────
        _xl_in_progress_msg = ss("_xl_autosave_last_msg", "")
        if _xl_in_progress_msg:
            if _xl_in_progress_msg.startswith("⚠"):
                _st.warning(_xl_in_progress_msg)
            else:
                _st.caption(f"💾 {_xl_in_progress_msg}")

        if _adv_main:
            if _elm_mode_run:
                cnt_ok      = sum(1 for r in results if r.get("elm_fetch_status") == "ok")
                cnt_partial = sum(1 for r in results if r.get("elm_fetch_status") == "partial")
                cnt_failed  = sum(1 for r in results if r.get("elm_fetch_status") == "failed")
                avg_score   = (
                    sum(float(r.get("elm_score_overall_icp", 0) or 0) for r in results) / len(results)
                    if results else 0.0
                )
                mc1, mc2, mc3, mc4, mc5 = _st.columns(5)
                mc1.metric("Fetched OK",  cnt_ok)
                mc2.metric("Partial",     cnt_partial)
                mc3.metric("Failed",      cnt_failed)
                mc4.metric("Processed",   len(results))
                mc5.metric("Avg ICP score", f"{avg_score:.1f}/10")
            else:
                cnt_jina       = sum(1 for r in results if "enriched_jina"       in r.get("enrichment_status", ""))
                cnt_playwright = sum(1 for r in results if "enriched_playwright" in r.get("enrichment_status", ""))
                cnt_google     = sum(1 for r in results if "enriched_search"     in r.get("enrichment_status", ""))
                cnt_nodata     = sum(1 for r in results if r.get("enrichment_status") == "no_data")
                cnt_error      = sum(1 for r in results
                                     if r.get("enrichment_status") not in
                                     ("enriched_jina", "enriched_jina_step1_only",
                                      "enriched_playwright", "enriched_playwright_step1_only",
                                      "enriched_search", "enriched_search_step1_only",
                                      "no_data", "skipped_resume", "zero_cost_preview", ""))
                cnt_retries    = ss("_jina_retry_count", 0)
                cnt_previews   = ss("_dry_run_preview_count", 0)

                if _zero_cost_run and _dry_run_run:
                    mc1, mc2, mc3 = _st.columns(3)
                    mc1.metric("Dry-run previews generated", cnt_previews)
                    mc2.metric("Errors",                     cnt_error)
                    mc3.metric("Est. cost",                  "$0.00")
                else:
                    mc1, mc2, mc3, mc4, mc5, mc6, mc7 = _st.columns(7)
                    mc1.metric("Enriched (Jina)",    cnt_jina)
                    mc2.metric("Enriched (Browser)", cnt_playwright)
                    mc3.metric("Enriched (Google)",  cnt_google)
                    mc4.metric("429 Retries",        cnt_retries)
                    mc5.metric("No data",            cnt_nodata)
                    mc6.metric("Errors",             cnt_error)
                    mc7.metric("Est. cost",          f"${total_cost:.4f}")

                _retry_msg = ss("_last_retry_msg", "")
                if _retry_msg:
                    _st.info(_retry_msg)

        if _show_adv:
            if _zero_cost_run and _dry_run_run and not _elm_mode_run:
                _st.warning(
                    "⚠️ **ZERO-COST PREVIEW ACTIVE**: no Step 1 or Step 2 API calls are being made. "
                    "Using only uploaded row data and existing cache to generate Step 2 prompt previews."
                )
            elif _dry_run_run and not _elm_mode_run:
                _st.warning(
                    "⚠️ **DRY RUN ACTIVE**: no Anthropic or Serper API calls are being made "
                    "for Step 2. Prompts and search queries are generated and displayed only."
                )

        # ── Intermediate download buttons — advanced/debug only ───────────────────
        if _adv_main and results:
            _partial_df = build_partial_df(results, df_work, _active_fields)
            _n_done     = len(_partial_df)
            _stamp      = ts()
            with _st.expander(
                f"⬇ Download intermediate results ({_n_done} rows so far)", expanded=True
            ):
                _st.caption("These links download via the browser without interrupting processing.")
                _html_dl_buttons(_partial_df, _n_done, _stamp)

        # ── Step 2 dry run preview ────────────────────────────────────────────────
        if _adv_main and _dry_run_run and not _elm_mode_run:
            _dry_recs = ss("_dry_run_records", [])
            with _st.expander("Step 2 Dry Run Preview", expanded=True):
                if not _dry_recs:
                    _st.caption("Dry run preview will appear here as companies are processed.")
                for _dr in _dry_recs:
                    with _st.expander(f"Company: {_dr['company']}", expanded=False):
                        _st.markdown(
                            f"**Provider:** {_dr['provider']}  |  **Model:** `{_dr['model']}`"
                        )
                        if _dr.get("queries"):
                            _st.markdown("**Serper queries that would be sent:**")
                            for _q in _dr["queries"]:
                                _st.code(_q, language=None)
                        _st.markdown("**Generated search instruction / prompt suffix:**")
                        _st.code(_dr.get("search_prompt", ""), language=None)
                        _st.markdown("**Full Step 2 Claude prompt:**")
                        _st.code(_dr.get("full_prompt", ""), language=None)

        # ── Step 2 search output records (UI) ────────────────────────────────────
        if ss("_show_step2_debug", False) and not _elm_mode_run:
            _srecs = ss("_search_output_records", [])
            if _srecs:
                with _st.expander(
                    f"Search outputs ({len(_srecs)} action(s) so far)", expanded=False
                ):
                    _st.caption(
                        f"Detailed files saved in `{SEARCH_OUTPUT_DIR}/`. "
                        "Use the download buttons below to access them from the browser."
                    )
                for _sri, _sr in enumerate(_srecs):
                    _sr_label = (
                        f"Search output: {_sr['company']} "
                        f"({'DRY RUN' if _sr.get('dry_run') else _sr.get('provider', '')})"
                    )
                    with _st.expander(_sr_label, expanded=False):
                        _st.markdown(
                            f"**Provider:** {_sr.get('provider', '')}  |  "
                            f"**Dry run:** {_sr.get('dry_run', False)}"
                        )
                        _st.markdown("**Query / search instruction:**")
                        _st.code(_sr.get("query", ""), language=None)
                        _rc = _sr.get("result_count", 0)
                        _st.markdown(f"**Results returned:** {_rc}")
                        _tops = _sr.get("top_results", [])
                        if _tops:
                            _st.markdown("**Top results:**")
                            for _t in _tops:
                                _st.markdown(
                                    f"- [{_t.get('title','(no title)')}]({_t.get('link','')})"
                                    + (f"  — {_t.get('snippet','')[:120]}" if _t.get('snippet') else "")
                                )
                        _df = _sr.get("debug_file", "")
                        if _df:
                            _st.caption(f"`{_df}`")
                            _rec_wrap = {
                                "path":    _df,
                                "company": _sr.get("company", ""),
                                "dry_run": _sr.get("dry_run", False),
                            }
                            _debug_file_download_button(_rec_wrap, f"run_{idx}_{_sri}")
                            _debug_file_preview_expander(_rec_wrap, f"prev_{idx}_{_sri}")

        # ── Step 2 debug log window ───────────────────────────────────────────────
        if ss("_show_step2_debug", False) and not _elm_mode_run:
            _log_text     = ss("_step2_debug_log", "")
            _prompt_recs  = ss("_step2_prompt_records", [])
            with _st.expander("Step 2 Debug Log", expanded=True):
                if _log_text:
                    _st.code(_log_text, language=None)
                else:
                    _st.caption("No log entries yet — will appear as companies are processed.")
            for _pr in _prompt_recs:
                with _st.expander(f"Prompt sent for {_pr['company']}", expanded=False):
                    _st.code(_pr.get("prompt", ""), language=None)

        _resume_mode = ss("_resume_mode", False)
        _saved_df    = autosave_load() if _resume_mode else None

        if ss("stop_requested", False) or idx >= _n:
            build_and_finish(results, debug_records, df_work, _active_fields)
        else:
            input_row = df_work.iloc[idx]

            # Use only canonical columns set by normalize_input_to_company_df.
            # No fallback to original detected columns — canonical columns are
            # mandatory after the start handler validation passes.
            company_name = str(input_row.get("canonical_company_name", "")).strip()
            raw_url      = str(input_row.get("canonical_company_url",    "")).strip()
            # Normalize domain-only values to a URL (prepend https:// if needed)
            if raw_url and not raw_url.startswith(("http://", "https://")):
                raw_url = normalize_url(raw_url)

            # Extract existing Lusha/Lucia field values from the input row.
            # For Type 2 the lusha_api_* fields were already mapped upfront by
            # normalize_input_to_company_df — no second map_lucia_export_row call needed.
            _lusha_field_set = set(LUSHA_API_FIELDS + LUSHA_API_META_FIELDS + STEP1_FIELDS)
            _existing_lusha = {
                k: (str(v) if not isinstance(v, str) else v)
                for k, v in input_row.items()
                if k in _lusha_field_set and v is not None and str(v).strip() not in ("", "nan", "NaN", "None")
            }

            # For Type 2, skip the external Lusha/Lucia API regardless of the
            # sidebar setting — company data is already pre-mapped from the CSV.
            _row_input_type = str(input_row.get("input_type", "")).strip()
            _is_lucia_row   = (_row_input_type == "pre_enriched_lucia_export")
            _effective_lusha_api = False  # Lusha live API disabled in Layer 1

            # ── Resume: skip rows already in autosave ─────────────────────────────
            if _resume_mode and autosave_already_done(
                _saved_df, _name_col, _domain_col, company_name, raw_url
            ):
                _st.caption(
                    f"⏭ Skipping row {idx + 1} / {_n}: "
                    f"**{company_name or '(empty)'}** — already in autosave"
                )
                # Represent the skipped row with a minimal placeholder so build_and_finish
                # still has the right row count; the full data lives in the autosave CSV.
                skip_fields = {f: "" for f in _active_fields}
                if not _elm_mode_run:
                    skip_fields["enrichment_status"] = "skipped_resume"
                results.append(skip_fields)
                debug_records.append({"input_company_name": company_name, "skipped": True})
                ss_set(results=results, debug_records=debug_records, process_index=idx + 1)
                _st.rerun()

            # ── Build Step 2 debug callback if either debug option is enabled ────────
            _show_debug_ui = ss("_show_step2_debug", False)
            _save_debug_fs = ss("_save_step2_debug", False)

            def _make_step2_callback(cname: str, dry_run_mode: bool = False):
                def _cb(event: str, **kwargs) -> None:
                    if event == "status":
                        append_debug_log(kwargs.get("msg", ""))
                    elif event == "search_output":
                        _so_rec = {
                            "company":      kwargs.get("company", cname),
                            "provider":     kwargs.get("provider", ""),
                            "query":        kwargs.get("query", ""),
                            "result_count": kwargs.get("result_count", 0),
                            "top_results":  kwargs.get("top_results", []),
                            "debug_file":   kwargs.get("debug_file", ""),
                            "dry_run":      kwargs.get("dry_run", False),
                        }
                        if _show_debug_ui:
                            _srecs = _st.session_state.get("_search_output_records", [])
                            _srecs.append(_so_rec)
                            _st.session_state["_search_output_records"] = _srecs
                        # Always track search output files for download (when saving is on)
                        _so_path = kwargs.get("debug_file", "")
                        if _so_path and _save_debug_fs:
                            try:
                                _all_files = _st.session_state.get("_step2_debug_files", [])
                                _all_files.append({
                                    "path":     _so_path,
                                    "filename": Path(_so_path).name,
                                    "company":  kwargs.get("company", cname),
                                    "provider": kwargs.get("provider", ""),
                                    "kind":     "search",
                                    "dry_run":  kwargs.get("dry_run", False),
                                })
                                _st.session_state["_step2_debug_files"] = _all_files
                            except Exception:
                                pass
                    elif event == "prompt":
                        _is_dry   = kwargs.get("dry_run", False)
                        _file_pfx = "step2_dry_run" if _is_dry else "step2_prompt"
                        if _save_debug_fs:
                            _ts   = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
                            _body = format_step2_debug_content(
                                company_name=kwargs.get("company", cname),
                                model=kwargs.get("model", ""),
                                timestamp=_ts,
                                provider=kwargs.get("provider", STEP2_PROVIDER_CLAUDE),
                                search_prompt=kwargs.get("search_prompt", ""),
                                full_prompt=kwargs.get("full_prompt", ""),
                                notes=kwargs.get("notes", []),
                            )
                            try:
                                _prompt_path = write_debug_log(cname, _body, prefix=_file_pfx)
                                _all_files = _st.session_state.get("_step2_debug_files", [])
                                _all_files.append({
                                    "path":     str(_prompt_path),
                                    "filename": _prompt_path.name,
                                    "company":  kwargs.get("company", cname),
                                    "provider": kwargs.get("provider", STEP2_PROVIDER_CLAUDE),
                                    "kind":     "prompt_dry_run" if _is_dry else "prompt",
                                    "dry_run":  _is_dry,
                                })
                                _st.session_state["_step2_debug_files"] = _all_files
                            except Exception:
                                pass
                        if _show_debug_ui:
                            _recs = _st.session_state.get("_step2_prompt_records", [])
                            _recs.append({
                                "company":       kwargs.get("company", cname),
                                "prompt":        kwargs.get("full_prompt", ""),
                                "search_prompt": kwargs.get("search_prompt", ""),
                            })
                            _st.session_state["_step2_prompt_records"] = _recs
                        # Always accumulate dry-run records for the preview section
                        if dry_run_mode or _is_dry:
                            _drecs = _st.session_state.get("_dry_run_records", [])
                            _drecs.append({
                                "company":       kwargs.get("company", cname),
                                "provider":      kwargs.get("provider", ""),
                                "model":         kwargs.get("model", ""),
                                "queries":       kwargs.get("queries", []),
                                "search_prompt": kwargs.get("search_prompt", ""),
                                "full_prompt":   kwargs.get("full_prompt", ""),
                            })
                            _st.session_state["_dry_run_records"] = _drecs
                return _cb

            _debug_cb = (
                _make_step2_callback(company_name, dry_run_mode=_dry_run_run)
                if (_show_debug_ui or _save_debug_fs or _dry_run_run) and not _elm_mode_run
                else None
            )

            with _st.status(
                f"**{company_name or '(empty)'}**  ({idx + 1} of {_n})",
                expanded=False,
            ) as status_box:
                if _elm_mode_run:
                    status_box.write("⚡ Extreme Light Mode — fetching pages…")
                    fields, dbg = enrich_one_row_light(company_name, raw_url)
                    _fetch_status = fields.get("elm_fetch_status", "?")
                    status_box.write(
                        f"✅ Done — fetch: {_fetch_status} | "
                        f"{int(fields.get('elm_total_chars', 0) or 0):,} chars | "
                        f"score: {fields.get('elm_score_overall_icp', '?')}/10"
                    )
                    row_cost = 0.0
                else:
                    status_box.write(f"🤖 Step 2 model: `{_model_step2_run}`")
                    _prov_label = (
                        f"🔍 Step 2 search provider: {_step2_provider_run}"
                        + (" **(ZERO-COST PREVIEW)**" if (_zero_cost_run and _dry_run_run) else
                           " **(DRY RUN)**" if _dry_run_run else "")
                    )
                    status_box.write(_prov_label)

                    # ── ZERO-COST PREVIEW GUARD ───────────────────────────────────
                    # Do NOT call Jina, Claude, Serper, browser scraping, or any
                    # external API. Use only uploaded row data and existing cache.
                    if _zero_cost_run and _dry_run_run:
                        status_box.write(
                            "⚡ Zero-cost preview — skipping all Step 1 API calls, "
                            "running Step 2 dry run from row data only…"
                        )
                        # Build a minimal debug record (no real API calls)
                        dbg = {"input_company_name": company_name, "zero_cost_preview": True}
                        # Call run_step2 with dry_run=True; no Jina/browser scraping
                        icp_fields, _step2_cache, _s2_ti, _s2_to, _s2_status, _s2_msg, _s2_cr, _s2_cc = run_step2(
                            url=raw_url,
                            company_name=company_name,
                            api_key="",
                            delay=_delay,
                            model_step2=_model_step2_run,
                            _debug_callback=_debug_cb,
                            search_provider=_step2_provider_run,
                            serper_key="",
                            dry_run=True,
                        )
                        fields = {f: "" for f in ALL_ENRICHMENT_FIELDS}
                        fields["enrichment_status"]   = "zero_cost_preview"
                        fields["step2_status"]        = _s2_status
                        fields["step2_provider_used"] = _step2_provider_run
                        fields.update(icp_fields)
                        # Fill model-signal defaults (no API call in zero-cost mode)
                        fields.update(_build_model_signal_empty())
                        # Increment the preview counter
                        ss_set(_dry_run_preview_count=ss("_dry_run_preview_count", 0) + 1)
                        row_cost = 0.0
                        status_box.write("✅ Zero-cost preview done — no tokens used.")
                    # ── END ZERO-COST PREVIEW GUARD ──────────────────────────────
                    else:
                        status_box.write("⏳ Step 1 — Fetching page + extracting firmographics…")
                        fields, dbg = enrich_one_row(
                            company_name, raw_url, _api_key, _delay,
                            use_playwright=_use_playwright_run,
                            model_step1=_model_step1_run,
                            model_step2=_model_step2_run,
                            _debug_callback=_debug_cb,
                            search_provider=_step2_provider_run,
                            serper_key=_serper_key_run,
                            dry_run=_dry_run_run,
                            enable_lusha_api=_effective_lusha_api,
                            lusha_api_key=_lusha_api_key_run,
                            extract_model_signals=_extract_model_signals_run,
                            include_signal_evidence=_include_signal_evidence_run,
                            run_step1_enrichment=_run_step1_enrichment_run,
                            run_step2_enrichment=_run_step2_enrichment_run,
                            scoring_profile=_scoring_profile,
                            existing_lusha_data=_existing_lusha or None,
                        )
                    if not (_zero_cost_run and _dry_run_run):
                        s1_tok   = int(fields.get("step1_tokens_in",  0) or 0) + int(fields.get("step1_tokens_out", 0) or 0)
                        s2_tok   = int(fields.get("step2_tokens_in",  0) or 0) + int(fields.get("step2_tokens_out", 0) or 0)
                        row_cost = float(fields.get("total_cost_usd", 0) or 0)
                        _retry_note = ""
                        if "enriched_search" in fields.get("enrichment_status", ""):
                            _retry_note = " | ⚡ Google fallback used"
                        elif ss("_last_retry_msg", ""):
                            _retry_note = " | ⏳ Had Jina 429 retry"
                        if fields.get("step2_status") == "api_error":
                            status_box.write(
                                f"⚠️ Step 2 API error: {fields.get('error_message', '(no detail)')}"
                            )
                        status_box.write(
                            f"✅ Done — Step 1: {s1_tok} tokens | Step 2: {s2_tok} tokens | "
                            f"Row cost: ${row_cost:.5f}{_retry_note}"
                        )

            results.append(fields)
            debug_records.append(dbg)

            # ── Auto-save (crash recovery) — skipped in dry run / zero-cost preview ──
            if not _dry_run_run:
                try:
                    autosave_append(fields, input_row)
                    ss_set(autosave_last_name=company_name or raw_url or f"row {idx + 1}")
                except Exception:
                    pass  # never let autosave failure abort processing

            _new_idx = len(results)  # results already includes the row appended above

            # ── Per-company run-folder autosave ───────────────────────────────────
            if _pca_enabled_run and _pca_run_dir_run:
                # 1. Per-company JSON
                _pca_ok, _pca_msg = save_company_result_to_run_folder(
                    row_index=_new_idx,
                    company_name=company_name,
                    input_row=input_row,
                    enriched_fields=fields,
                    debug_record=dbg,
                    run_dir=_pca_run_dir_run,
                )
                if _pca_ok:
                    ss_set(
                        _per_company_autosave_last_saved=(
                            f"{company_name or f'row {_new_idx}'} → "
                            f"row_{_new_idx:04d}_{safe_filename(company_name or 'row')}.json"
                        ),
                        _per_company_autosave_last_error="",
                    )
                else:
                    ss_set(_per_company_autosave_last_error=_pca_msg[:200])

                # 2. Cumulative partial files (latest_results.* + checkpoint on multiples)
                _pca_ok2, _pca_msg2 = save_partial_outputs_to_run_folder(
                    results=results,
                    debug_records=debug_records,
                    df_work=df_work,
                    active_fields=_active_fields,
                    run_dir=_pca_run_dir_run,
                    elm_mode=_elm_mode_run,
                    row_count=_new_idx,
                    name_col=name_col,
                    scoring_profile=_scoring_profile,
                    domain_col=domain_col,
                )
                if _pca_ok2:
                    _save_label = f"{_new_idx} rows → latest_results.xlsx"
                    if _new_idx > 0 and _new_idx % CHECKPOINT_EVERY == 0:
                        _ckpt_label = f"enriched_results_partial_{_new_idx:03d}_companies.xlsx"
                        _save_label += f" + {_ckpt_label}"
                    ss_set(_last_local_save=_save_label)
                else:
                    ss_set(
                        _per_company_autosave_last_error=_pca_msg2[:200],
                        _last_local_save=f"⚠ Save failed: {_pca_msg2[:120]}",
                    )

            # ── Simple Excel autosave every N companies ───────────────────────────
            if ss("_xl_autosave_enabled", True) and _xl_should_autosave(
                _new_idx, int(ss("_xl_autosave_every", _XL_AUTOSAVE_EVERY))
            ):
                _xl_snap = build_partial_df(results, df_work, _active_fields)
                _xl_partial_fname = (
                    f"enriched_results_partial_{_new_idx:03d}_companies.xlsx"
                )
                _xl_ok, _xl_msg = _xl_autosave_write(
                    _xl_snap, _xl_partial_fname,
                    name_col=name_col, scoring_profile=_scoring_profile, domain_col=domain_col,
                )
                ss_set(_xl_autosave_last_msg=(
                    f"Autosaved {_new_idx} rows → {_xl_partial_fname}: {_xl_msg}"
                    if _xl_ok else f"⚠ {_xl_msg}"
                ))

            # ── Auto browser-download every _AUTO_DL_EVERY companies ─────────────
            _auto_dl_done = ss("_auto_dl_count", 0)
            if _new_idx % _AUTO_DL_EVERY == 0 and _new_idx // _AUTO_DL_EVERY > _auto_dl_done:
                _dl_snap = build_partial_df(results, df_work, _active_fields)
                _dl_name = f"enrichedResults_snapshot_{_new_idx}.xlsx"
                _js_auto_download(_dl_snap, _dl_name)
                ss_set(
                    _auto_dl_count=_new_idx // _AUTO_DL_EVERY,
                    _auto_dl_last_msg=f"Auto-downloaded at row {_new_idx} → {_dl_name}",
                )

            try:
                total_in   += int(fields.get("total_tokens_in",  0) or 0)
                total_out  += int(fields.get("total_tokens_out", 0) or 0)
                total_cost += float(fields.get("total_cost_usd", 0) or 0)
            except (ValueError, TypeError):
                pass

            # Accumulate cache token counts from the debug record of this row
            _dbg_last = debug_records[-1] if debug_records else {}
            try:
                total_cache_read   += int(_dbg_last.get("step2_cache_read_tokens",   0) or 0)
                total_cache_create += int(_dbg_last.get("step2_cache_creation_tokens", 0) or 0)
            except (ValueError, TypeError):
                pass

            ss_set(
                results=results, debug_records=debug_records, process_index=idx + 1,
                total_tokens_in=total_in, total_tokens_out=total_out, total_cost_usd=total_cost,
                total_cache_read_tokens=total_cache_read,
                total_cache_create_tokens=total_cache_create,
            )
            _st.rerun()

    # =============================================================================
    # RESULTS
    # =============================================================================



    def _render_advanced_results(
        df_enriched, debug_records_done, processed, _elm_done, debug_mode
    ):
        """Render advanced result sections — only shown when SHOW_ADVANCED_SETTINGS is True.

        Sections: status metrics, token cost, results table, commercial fit scoring,
        download buttons, debug details.
        """
        _done_fields = ELM_ALL_FIELDS if _elm_done else ALL_ENRICHMENT_FIELDS
        # ── Status summary ────────────────────────────────────────────────────────
        status_counts  = (
            df_enriched["elm_fetch_status"].value_counts().to_dict()
            if _elm_done and "elm_fetch_status" in df_enriched.columns
            else df_enriched.get("enrichment_status", pd.Series(dtype=str)).value_counts().to_dict()
        )
        needs_review_n = (
            int((df_enriched["needs_manual_review"] == "TRUE").sum())
            if "needs_manual_review" in df_enriched.columns else 0
        )
        all_sc = list(status_counts.items())
        web_search_n = (
            int((df_enriched["step2_status"] == "ok").sum())
            if "step2_status" in df_enriched.columns else 0
        )
        competitor_n = (
            int(df_enriched["icp_competitor_signal"].astype(str).str.strip().ne("").sum())
            if "icp_competitor_signal" in df_enriched.columns else 0
        )
        # In ELM mode there is no "Needs review" concept
        extra_metrics = [] if _elm_done else [("⚑ Needs review", needs_review_n)]
        extra_metrics += [("🔍 Web search used", web_search_n), ("🏁 Competitor found", competitor_n)]
        n_cols = min(len(all_sc) + len(extra_metrics), 6)
        if all_sc:
            rcols = _st.columns(max(n_cols, 1))
            for i, (s, c) in enumerate(all_sc):
                rcols[i % len(rcols)].metric(_STATUS_LABELS.get(s, s), c)
            for j, (label, val) in enumerate(extra_metrics):
                rcols[(len(all_sc) + j) % len(rcols)].metric(label, val)

        # ── Token usage ───────────────────────────────────────────────────────────
        t_in   = ss("total_tokens_in", 0)
        t_out  = ss("total_tokens_out", 0)
        t_cost = ss("total_cost_usd", 0.0)

        if not _elm_done:
            with _st.expander("💰 Token usage & cost", expanded=True):
                tc1, tc2, tc3 = _st.columns(3)
                tc1.metric("Total input tokens",   f"{t_in:,}")
                tc2.metric("Total output tokens",  f"{t_out:,}")
                tc3.metric("Estimated total cost", f"${t_cost:.4f}")

                _used_s1 = ss("_model_step1", MODEL_STEP1)
                _used_s2 = ss("_model_step2", MODEL_STEP2)
                _st.caption(
                    f"Step 1 model: `{_used_s1}` · Step 2 model: `{_used_s2}`. "
                    "Two API calls per row (Step 1 + Step 2). "
                    "Verify charges in your Anthropic dashboard."
                )

                _cache_read   = ss("total_cache_read_tokens",   0)
                _cache_create = ss("total_cache_create_tokens", 0)
                if _cache_read > 0 or _cache_create > 0:
                    _st.divider()
                    # Input price per M for the Step 2 model (approximate)
                    _s2_input_price_per_m = (
                        0.80 if "haiku"  in _used_s2 else
                        3.00 if "sonnet" in _used_s2 else 1.00
                    )
                    # Cache reads cost 10 % of normal input price; savings = 90 %
                    _savings_usd = _cache_read * _s2_input_price_per_m * 0.90 / 1_000_000
                    cc1, cc2, cc3 = _st.columns(3)
                    cc1.metric("Cache write tokens (Step 2)",  f"{_cache_create:,}",
                               help="Tokens written to Anthropic's prompt cache on the first company.")
                    cc2.metric("Cache read tokens (Step 2)",   f"{_cache_read:,}",
                               help="Tokens served from cache at 10 % of normal input cost.")
                    cc3.metric("Est. prompt-cache savings",    f"${_savings_usd:.4f}",
                               help="90 % discount on cache-read tokens vs full input price.")
                    _st.caption(
                        f"Prompt caching active on Step 2 ({_used_s2}). "
                        f"Static prefix cached once; each subsequent company reads it at "
                        f"~${_s2_input_price_per_m * 0.10:.3f}/M tokens instead of "
                        f"${_s2_input_price_per_m:.2f}/M."
                    )
        else:
            _st.info("⚡ Extreme Light Mode — no API calls, no tokens, no cost.")

        # ── Results table ─────────────────────────────────────────────────────────
        _st.subheader("Results")
        orig_cols = [c for c in df_enriched.columns if c not in _done_fields]

        if _elm_done:
            summary_cols = orig_cols + [c for c in ELM_ALL_FIELDS if c in df_enriched.columns]
            _st.dataframe(df_enriched[summary_cols], use_container_width=True, height=400)
            tab1, tab2, tab3 = _st.tabs(
                ["Status & fetch info", "Keyword counts", "Normalized scores"]
            )
            with tab1:
                _st.dataframe(
                    df_enriched[[c for c in ELM_STATUS_FIELDS if c in df_enriched.columns]],
                    use_container_width=True,
                )
            with tab2:
                _st.dataframe(
                    df_enriched[[c for c in ELM_KEYWORD_FIELDS if c in df_enriched.columns]],
                    use_container_width=True,
                )
            with tab3:
                _st.dataframe(
                    df_enriched[[c for c in ELM_SCORE_FIELDS if c in df_enriched.columns]],
                    use_container_width=True,
                )
        else:
            _lusha_done = ss("_enable_lusha_api", False)
            # Scoring columns first (if available)
            _sc_summary = [c for c in (_SCORE_OUTPUT_COLS or []) if c in df_enriched.columns]
            summary_cols = orig_cols + _sc_summary + [
                c for c in [
                    # Metadata
                    "enrichment_status", "step1_status", "step2_status",
                    "needs_manual_review", "match_notes",
                    # Lusha API (real) — key fields only in summary
                    "lusha_api_status", "lusha_api_match_confidence", "lusha_api_needs_review",
                    "lusha_api_company_name", "lusha_api_domain", "lusha_api_industry",
                    "lusha_api_employee_range", "lusha_api_country",
                    # Step 1 — firmographics
                    "lusha_company_name", "lusha_domain", "lusha_industry", "lusha_sub_industry",
                    "lusha_company_type", "lusha_employee_range", "lusha_revenue",
                    "lusha_country", "lusha_city", "lusha_continent",
                    "lusha_founded_year", "lusha_description",
                    "lusha_linkedin_url", "lusha_specialties", "lusha_technologies",
                    "lusha_total_funding_amount", "lusha_total_funding_rounds",
                    "lusha_last_round_type", "lusha_last_round_amount", "lusha_last_round_date",
                    "lusha_ipo_status",
                    # Step 2 — ICP buying signals
                    "icp_lead_score", "icp_buying_signals", "icp_competitor_signal",
                    "icp_direct_language_competitor_signal",
                    "icp_online_language_learning_signal",
                    "icp_broader_lnd_platform_signal",
                    "icp_evidence", "icp_likely_training_interest",
                    "icp_why_relevant", "icp_potential_buyer_function",
                    # Cost
                    "total_tokens_in", "total_tokens_out", "total_cost_usd",
                    "error_message",
                ]
                if c in df_enriched.columns
            ]
            _st.dataframe(df_enriched[summary_cols], use_container_width=True, height=400)
            _tabs = [
                "Step 1 — All firmographic columns",
                "Step 2 — All ICP columns",
                "Model signals — scores & binaries",
                "Model signals — QA evidence",
            ]
            if _lusha_done:
                _tabs.append("Lusha API fields")
            _tab_objs = _st.tabs(_tabs)
            with _tab_objs[0]:
                _st.dataframe(df_enriched[[c for c in STEP1_FIELDS if c in df_enriched.columns]],
                             use_container_width=True)
            with _tab_objs[1]:
                _st.dataframe(df_enriched[[c for c in ICP_FIELDS if c in df_enriched.columns]],
                             use_container_width=True)
            with _tab_objs[2]:
                _score_bin_cols = [
                    c for c in MODEL_SIGNAL_SCORE_FIELDS + MODEL_SIGNAL_BINARY_FIELDS
                    + MODEL_SIGNAL_QA_FIELDS
                    if c in df_enriched.columns
                ]
                if _score_bin_cols:
                    _st.dataframe(df_enriched[_score_bin_cols], use_container_width=True)
                else:
                    _st.info("Model signal extraction was not run or is disabled.")
            with _tab_objs[3]:
                _evid_cols = [c for c in MODEL_SIGNAL_EVIDENCE_FIELDS if c in df_enriched.columns]
                if _evid_cols:
                    _st.dataframe(df_enriched[_evid_cols], use_container_width=True)
                else:
                    _st.info("No evidence columns found (evidence may be disabled or extraction not run).")
            if _lusha_done and len(_tab_objs) > 4:
                with _tab_objs[4]:
                    _lusha_display_cols = [
                        c for c in LUSHA_API_FIELDS + LUSHA_API_META_FIELDS
                        if c in df_enriched.columns
                    ]
                    _st.dataframe(df_enriched[_lusha_display_cols], use_container_width=True)

        # ── Commercial Fit Scoring ────────────────────────────────────────────────
        if not _elm_done and _SCORING_AVAILABLE:
            _score_key_col = "final_commercial_fit_score"
            _score_cols    = _SCORE_OUTPUT_COLS
            if _score_key_col in df_enriched.columns:
                _st.divider()

                # ── Single company: score card ────────────────────────────────────
                if processed == 1:
                    _sc_row = df_enriched.iloc[0]
                    _fit   = _sc_row.get(_score_key_col, 0)
                    _tier  = _sc_row.get("commercial_tier", "—")
                    _icp   = _sc_row.get("icp_similarity_score", 0)
                    _sz    = _sc_row.get("company_size_score", 0)
                    _prob  = _sc_row.get("model_probability", 0)
                    _top_d = _sc_row.get("top_score_drivers", "")
                    _wk_d  = _sc_row.get("weak_score_drivers", "")
                    _notes = _sc_row.get("scoring_notes", "")
                    _dqf   = _sc_row.get("data_quality_flag", "")
                    _gc    = _sc_row.get("global_complexity_score", "")
                    _pd_s  = _sc_row.get("people_development_score", "")
                    _cc    = _sc_row.get("commercial_complexity_score", "")

                    _tier_emoji = {"🥇 Hot": "🟢", "🥈 Warm": "🟡", "🥉 Cool": "🟠", "❄️ Pass": "🔴"}.get(str(_tier), "⚪")
                    _dq_label   = {"high": "✅ High", "medium": "⚠️ Medium", "low": "🔴 Low"}.get(str(_dqf), str(_dqf))

                    _st.subheader("🎯 Commercial Fit Score")
                    _st.caption(
                        "The score combines ICP similarity and company size. "
                        "ICP similarity is based on enriched buying signals. "
                        "Company size is used as a commercial weighting factor."
                    )
                    _m1, _m2, _m3, _m4 = _st.columns(4)
                    _m1.metric("Commercial Fit Score", f"{float(_fit):.2f} / 10")
                    _m2.metric("Commercial Tier", f"{_tier_emoji} {_tier}")
                    _m3.metric("ICP Similarity", f"{float(_icp):.2f} / 10")
                    _m4.metric("Company Size", f"{_sz} / 10")

                    _d1, _d2, _d3 = _st.columns(3)
                    _d1.metric("Global Complexity", f"{_gc} / 10" if _gc != "" else "—")
                    _d2.metric("People Development", f"{_pd_s} / 10" if _pd_s != "" else "—")
                    _d3.metric("Commercial Complexity", f"{_cc} / 10" if _cc != "" else "—")

                    if _top_d and str(_top_d) != "none":
                        _st.markdown(f"**Top score drivers:** {_top_d}")
                    if _wk_d and str(_wk_d) != "none":
                        _st.markdown(f"**Weak / missing drivers:** {_wk_d}")

                    # Suggested cold-call opening angle from top driver
                    _top_driver_field = (str(_top_d) or "").split(";")[0].split("=")[0].strip()
                    _opening_map = {
                        "sig_explicit_lnd_score":    "I saw that {name} has a strong L&D programme — we help companies like yours scale language training across global teams.",
                        "sig_lnd_onboarding_score":  "Your structured onboarding process caught my attention — we specialise in language-ready onboarding for international hires.",
                        "sig_intl_footprint_score":  "{name}'s international presence is exactly the profile we work with — multilingual communication across offices is our core focus.",
                        "sig_foreign_hq_score":      "With {name}'s cross-border structure, language alignment between HQ and local teams is often a real friction point — that's where we come in.",
                        "sig_rapid_growth_score":    "{name}'s growth trajectory often brings language challenges — we help scaling companies maintain communication quality across markets.",
                        "sig_multicultural_score":   "The diverse workforce at {name} is a strong fit for our language learning programmes tailored for multicultural teams.",
                        "language_competitor_strength_score": "I noticed a competitor in your space is already investing in language training — this is a strong signal we can help {name} stay ahead.",
                    }
                    _company_display = str(df_enriched.iloc[0].get("company_name", "your company"))
                    _opening = _opening_map.get(
                        _top_driver_field,
                        "Based on {name}'s profile, your team could benefit from targeted language and communication training across international operations.",
                    ).replace("{name}", _company_display)
                    with _st.expander("💬 Suggested cold-call opening angle", expanded=True):
                        _st.markdown(f"_{_opening}_")
                        _st.caption("Generated from top score driver. Customise before use.")

                    if _notes:
                        with _st.expander("ℹ️ Scoring notes", expanded=False):
                            _st.caption(str(_notes))
                    if _dqf:
                        _st.caption(f"Data quality: {_dq_label} · Model probability: {float(_prob):.1%}")

                else:
                    # ── Batch: filtered + sorted score table ──────────────────────
                    _st.subheader("🎯 Commercial Fit Scoring")
                    _active_profile = ss("_scoring_profile", "default")
                    if _active_profile == "italy_register_icp_only":
                        _st.caption(
                            "Final score is based on ICP/model signals only. "
                            "Company size is excluded from Layer 1 scoring because the Italian register "
                            "input is already filtered for 100+ employees. "
                            "Employee estimates are audit-only and should be validated in Layer 2.5 "
                            "with Lucia/Lusha/contact enrichment."
                        )
                    else:
                        _st.caption(
                            "The score combines ICP similarity and company size. "
                            "ICP similarity is based on enriched buying signals. "
                            "Company size is used as a commercial weighting factor. "
                            "Final Commercial Fit Score = 0.75 × ICP Similarity + 0.25 × Company Size."
                        )

                    with _st.expander("ℹ️ About these scores", expanded=False):
                        if _active_profile == "italy_register_icp_only":
                            _st.markdown(
                                "1. **Model probability** — lean logistic regression on 7 key model-signal fields "
                                "(normalised 0–3 → 0–1).\n"
                                "2. **ICP Similarity Score [1–10]** — sigmoid-stretched model probability "
                                "(K=1 for Italy register profile).\n"
                                "3. **Company Size Score** — excluded from scoring for Italy register profile. "
                                "Input list is pre-filtered for 100+ employees; estimates are audit-only.\n"
                                "4. **Final Commercial Fit Score** = ICP Similarity only (100% model weight).\n"
                                "5. **Tier** — 🥇 Hot ≥ 6.50 · 🥈 Warm ≥ 5.00 · 🥉 Cool ≥ 3.00 · ❄️ Pass < 3.00.\n"
                                "6. **Composite scores** — global complexity, people development, commercial complexity "
                                "(each 0–10, from signal groupings).\n"
                            )
                        else:
                            _st.markdown(
                                "1. **Model probability** — lean logistic regression on 7 key model-signal fields "
                                "(normalised 0–3 → 0–1).\n"
                                "2. **ICP Similarity Score [1–10]** — sigmoid-stretched model probability.\n"
                                "3. **Company Size Score [1–10]** — 9-band employee-count mapping.\n"
                                "4. **Final Commercial Fit Score** = 0.75 × ICP Similarity + 0.25 × Company Size.\n"
                                "5. **Tier** — 🥇 Hot ≥ 8.66 · 🥈 Warm ≥ 7.19 · 🥉 Cool ≥ 4.23 · ❄️ Pass < 4.23.\n"
                                "6. **Composite scores** — global complexity, people development, commercial complexity "
                                "(each 0–10, from signal groupings).\n"
                            )

                    _tier_order  = ["🥇 Hot", "🥈 Warm", "🥉 Cool", "❄️ Pass"]
                    _tier_counts = df_enriched["commercial_tier"].value_counts()
                    _tier_colors = ["🟢", "🟡", "🟠", "🔴"]
                    _tc = _st.columns(4)
                    for _i, (_tier, _emoji) in enumerate(zip(_tier_order, _tier_colors)):
                        _cnt = int(_tier_counts.get(_tier, 0))
                        _pct = f"{_cnt / max(processed, 1):.0%}"
                        _tc[_i].metric(f"{_emoji} {_tier}", _cnt, delta=_pct, delta_color="off")

                    _fc1, _fc2, _fc3 = _st.columns([2, 2, 2])
                    with _fc1:
                        _tier_filter = _st.selectbox(
                            "Show tiers",
                            ["All", "🥇 Hot only", "🥇 Hot + 🥈 Warm"],
                            key="score_tier_filter",
                        )
                    with _fc2:
                        _min_score = _st.slider(
                            "Minimum fit score",
                            min_value=0.0, max_value=10.0,
                            value=0.0, step=0.5,
                            key="score_min_slider",
                        )
                    with _fc3:
                        _sort_by = _st.selectbox(
                            "Sort by",
                            ["final_commercial_fit_score", "model_probability", "company_size_score"],
                            key="score_sort_col",
                        )

                    _id_cols_sc = [
                        c for c in df_enriched.columns
                        if c not in set(ALL_ENRICHMENT_FIELDS + list(_score_cols))
                    ][:2]
                    _disp_core = [
                        "final_commercial_fit_score", "commercial_tier",
                        "model_probability", "icp_similarity_score", "company_size_score",
                        "top_score_drivers", "scoring_notes",
                        "global_complexity_score", "people_development_score",
                        "commercial_complexity_score", "data_quality_flag",
                    ]
                    _disp_cols = _id_cols_sc + [c for c in _disp_core if c in df_enriched.columns]
                    _score_disp_df = df_enriched[[c for c in _disp_cols if c in df_enriched.columns]].copy()

                    if _tier_filter == "🥇 Hot only":
                        _score_disp_df = _score_disp_df[_score_disp_df["commercial_tier"] == "🥇 Hot"]
                    elif _tier_filter == "🥇 Hot + 🥈 Warm":
                        _score_disp_df = _score_disp_df[_score_disp_df["commercial_tier"].isin(["🥇 Hot", "🥈 Warm"])]
                    if _min_score > 0 and "final_commercial_fit_score" in _score_disp_df.columns:
                        _score_disp_df = _score_disp_df[_score_disp_df["final_commercial_fit_score"] >= _min_score]
                    if _sort_by in _score_disp_df.columns:
                        _score_disp_df = _score_disp_df.sort_values(_sort_by, ascending=False)

                    _st.dataframe(_score_disp_df, use_container_width=True, height=420)
                    _st.caption(f"Showing {len(_score_disp_df)} of {processed} companies.")

            elif not ss("_extract_model_signals", True):
                _st.info(
                    "ℹ️ Commercial fit scoring requires Step 3 model signal extraction. "
                    "Enable **Extract model signals (Step 3)** in the sidebar and re-run."
                )

        # ── Downloads ─────────────────────────────────────────────────────────────
        _st.subheader("Download results")
        log_df = make_log_df(debug_records_done, elm_mode=_elm_done)

        if _elm_done:
            _fname_prefix = "elm_results"
            _log_fname    = "elm_fetch_log.csv"
        else:
            _run_tag      = build_run_tag()
            _fname_prefix = Path(build_enriched_output_filename(ss("file_name", ""))).stem
            _log_fname    = f"processing_log_{_run_tag}_{ts()}.csv"
        _xl_help      = (
            "All original columns + keyword counts + normalized scores."
            if _elm_done else
            "Sheet 1 (Enriched): all enrichment columns. "
            "Sheet 2 (model_features): input columns + model signal scores + binary columns only."
        )
        _log_help     = (
            "One row per company: fetch status, pages fetched, total chars."
            if _elm_done else
            "One row per company: step statuses, token counts, costs, review flags."
        )

        _rich_xl = (
            build_rich_excel_bytes(
                df_enriched,
                name_col=ss("_name_col"),
                domain_col=ss("_domain_col"),
                df_input_original=ss("_df_raw_original"),
                scoring_profile=ss("_scoring_profile", "default"),
                run_config=ss("_active_run_config"),
                run_mode="streamlit",
            )
            if not _elm_done else df_to_excel_bytes(df_enriched)
        )
        dl1, dl2, dl3 = _st.columns(3)
        with dl1:
            _st.download_button(
                "⬇ Results Excel (.xlsx)",
                data=_rich_xl,
                file_name=f"{_fname_prefix}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                help=_xl_help,
            )
        with dl2:
            _st.download_button(
                "⬇ Results CSV",
                data=df_to_csv_bytes(df_enriched),
                file_name=f"{_fname_prefix}.csv",
                mime="text/csv",
                use_container_width=True,
                help="Same data as Excel, in CSV format.",
            )
        with dl3:
            _st.download_button(
                "⬇ Processing log CSV",
                data=df_to_csv_bytes(log_df),
                file_name=_log_fname,
                mime="text/csv",
                use_container_width=True,
                help=_log_help,
            )

        if not _elm_done:
            _mf_df = _build_model_features_df(df_enriched)
            if not _mf_df.empty:
                _mf_fname = f"model_features_{_run_tag}_{ts()}.csv"
                _st.download_button(
                    "⬇ Model features CSV (scores + binaries only)",
                    data=df_to_csv_bytes(_mf_df),
                    file_name=_mf_fname,
                    mime="text/csv",
                    use_container_width=True,
                    help=(
                        "Input columns + all model signal score/binary columns. "
                        "No evidence columns. Ready for logistic regression."
                    ),
                )

        # ── Debug section ─────────────────────────────────────────────────────────
        if debug_mode and debug_records_done:
            _st.divider()
            _st.subheader("🐛 Per-row debug details")

            if _elm_done:
                debug_df = pd.DataFrame([
                    {
                        "row":          i + 1,
                        "company":      d.get("company", ""),
                        "url":          d.get("url", ""),
                        "domain":       d.get("domain", ""),
                        "fetch_status": d.get("status", ""),
                        "pages_fetched": d.get("pages_fetched", ""),
                        "total_chars":  d.get("total_chars", ""),
                    }
                    for i, d in enumerate(debug_records_done)
                ])
            else:
                debug_df = pd.DataFrame([
                    {
                        "row":               i + 1,
                        "company":           d.get("input_company_name", ""),
                        "url":               d.get("input_url", ""),
                        "lusha_api_status":  d.get("lusha_api_status", ""),
                        "step1_status":      d.get("step1_status", ""),
                        "step2_status":      d.get("step2_status", ""),
                        "enrichment_status": d.get("enrichment_status", ""),
                        "step1_tok_in":      d.get("step1_tokens_in", ""),
                        "step1_tok_out":     d.get("step1_tokens_out", ""),
                        "step2_tok_in":      d.get("step2_tokens_in", ""),
                        "step2_tok_out":     d.get("step2_tokens_out", ""),
                        "row_cost":          f"${d.get('total_cost', 0):.5f}",
                        "error":             d.get("error_message", ""),
                    }
                    for i, d in enumerate(debug_records_done)
                ])
            _st.dataframe(debug_df, use_container_width=True)

            if not _elm_done:
                _st.subheader("🐛 Raw JSON responses")
                company_labels = [
                    f"{i + 1}. {d.get('input_company_name') or '(empty)'} [{d.get('enrichment_status', '')}]"
                    for i, d in enumerate(debug_records_done)
                ]
                sel_idx = _st.selectbox(
                    "Select a company:",
                    options=range(len(company_labels)),
                    format_func=lambda i: company_labels[i],
                    key="raw_json_selector",
                )
                if sel_idx is not None:
                    d = debug_records_done[sel_idx]
                    col_lusha, col_a, col_b = _st.columns(3)
                    with col_lusha:
                        _st.markdown("**Lusha API — real company data**")
                        _st.caption(f"Status: `{d.get('lusha_api_status', '(not run)')}`")
                        if d.get("lusha_api_raw_json"):
                            _st.json(d["lusha_api_raw_json"])
                        else:
                            _st.info("No Lusha API response (disabled or no data).")
                    with col_a:
                        _st.markdown("**Step 1 — Jina + Claude extraction**")
                        if d.get("step1_raw_json"):
                            _st.json(d["step1_raw_json"])
                        else:
                            _st.info("No Step 1 response.")
                    with col_b:
                        _st.markdown("**Step 2 — Claude web_search ICP signals**")
                        if d.get("step2_raw_json"):
                            _st.json(d["step2_raw_json"])
                        else:
                            _st.info("No Step 2 response.")

            _st.subheader("🐛 Additional debug downloads")
            dbg_dl1, dbg_dl2 = _st.columns(2)
            with dbg_dl1:
                debug_enriched = df_enriched.copy()
                if not _elm_done:
                    debug_enriched["lusha_api_raw_json_preview"] = [
                        json.dumps(d.get("lusha_api_raw_json"), ensure_ascii=False)[:1500]
                        if d.get("lusha_api_raw_json") else ""
                        for d in debug_records_done
                    ] + [""] * max(0, len(debug_enriched) - len(debug_records_done))
                    debug_enriched["step1_json_preview"] = [
                        json.dumps(d.get("step1_raw_json"), ensure_ascii=False)[:1500]
                        if d.get("step1_raw_json") else ""
                        for d in debug_records_done
                    ] + [""] * max(0, len(debug_enriched) - len(debug_records_done))
                    debug_enriched["step2_json_preview"] = [
                        json.dumps(d.get("step2_raw_json"), ensure_ascii=False)[:1500]
                        if d.get("step2_raw_json") else ""
                        for d in debug_records_done
                    ] + [""] * max(0, len(debug_enriched) - len(debug_records_done))
                _dbg_fname = "elm_debug.xlsx" if _elm_done else "claude_enriched_debug.xlsx"
                _st.download_button(
                    "⬇ Debug Excel",
                    data=df_to_excel_bytes(debug_enriched),
                    file_name=_dbg_fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with dbg_dl2:
                cc = get_cache_count()
                if cc > 0:
                    _st.download_button(
                        f"⬇ Cache ZIP ({cc} files)",
                        data=cache_to_zip_bytes(),
                        file_name="claude_cache.zip",
                        mime="application/zip",
                        use_container_width=True,
                    )
                else:
                    _st.info("Cache is empty.")

            _st.subheader("🐛 Cache viewer")
            _st.caption(f"{CACHE_DIR.resolve()} — {get_cache_count()} file(s)")
            cache_files = list_cache_files()
            if cache_files:
                sel_cache = _st.selectbox(
                    "Select cached file:",
                    options=[f.stem for f in cache_files],
                    key="cache_file_selector",
                )
                if sel_cache:
                    try:
                        _st.json(json.loads(
                            (CACHE_DIR / f"{sel_cache}.json").read_text(encoding="utf-8")
                        ))
                    except Exception as exc:
                        _st.error(f"Could not read cache file: {exc}")
            else:
                _st.info("Cache is empty. Run an enrichment first.")



    if ss("enrichment_done", False):
        df_enriched: pd.DataFrame = ss("df_enriched")
        debug_records_done: list  = ss("debug_records", [])
        processed = len(df_enriched)
        _elm_done = ss("_elm_mode", False)
        _done_fields = ELM_ALL_FIELDS if _elm_done else ALL_ENRICHMENT_FIELDS

        _st.divider()
        _done_dry_run   = ss("_step2_dry_run",     False)
        _done_zero_cost = ss("_zero_cost_preview", False)
        _processed_word = "company" if processed == 1 else "companies"
        if ss("stop_requested", False):
            _st.warning(f"Enrichment stopped — **{processed:,}** {_processed_word} processed (partial).")
        else:
            _st.success(f"✅ Ready · **{processed:,}** {_processed_word} processed")

        if _show_adv:
            if _done_zero_cost and _done_dry_run and not _elm_done:
                _preview_count = ss("_dry_run_preview_count", 0)
                _st.info(
                    f"ℹ️ **Zero-cost preview completed.** No Step 1 or Step 2 API calls were made — "
                    f"**{_preview_count}** dry-run previews generated. "
                    "Disable zero-cost preview and dry run, then re-run to perform real enrichment."
                )
            elif _done_dry_run and not _elm_done:
                _st.info(
                    "ℹ️ **Dry run completed.** No Step 2 enrichment results were written — "
                    "Step 2 ICP columns are empty. Disable dry run and re-run to perform real enrichment."
                )

        # ── Simple Excel autosave — final write (runs exactly once per run) ─────────
        if not ss("_xl_autosave_final_done", False) and ss("_xl_autosave_enabled", True) and processed > 0:
            _xl_fin_fname_cfg = ss("_xl_autosave_filename", _XL_AUTOSAVE_DEFAULT)
            _xl_fin_ok, _xl_fin_msg = _xl_autosave_write(
                df_enriched, _xl_fin_fname_cfg,
                name_col=name_col, domain_col=domain_col,
                scoring_profile=ss("_scoring_profile", "default"),
            )
            if _xl_fin_ok:
                ss_set(
                    _xl_autosave_last_msg=f"Autosaved {processed} rows to: {_xl_fin_msg}",
                    _xl_autosave_final_done=True,
                )
                _st.caption(
                    f"📄 Autosaved to **{_xl_fin_msg}**"
                )
            else:
                ss_set(_xl_autosave_final_done=True)
                _st.warning(f"⚠ Excel autosave failed: {_xl_fin_msg}")

        # ── Auto-save final file into run folder (runs exactly once per completed run) ─
        _pca_done_enabled = ss("_per_company_autosave_enabled", False)
        _pca_done_dir     = ss("_per_company_autosave_run_dir", "")
        if not ss("_final_auto_saved", False):
            if _pca_done_enabled and _pca_done_dir:
                try:
                    _pca_rdir = Path(_pca_done_dir)
                    df_to_excel_bytes_write(df_enriched, str(_pca_rdir / "final_results.xlsx"))
                    df_enriched.to_csv(_pca_rdir / "final_results.csv", index=False, encoding="utf-8-sig")
                    # overwrite latest_results.* with the complete dataset too
                    df_to_excel_bytes_write(df_enriched, str(_pca_rdir / "latest_results.xlsx"))
                    df_enriched.to_csv(_pca_rdir / "latest_results.csv", index=False, encoding="utf-8-sig")
                    ss_set(_final_auto_saved=True, _final_save_path=str(_pca_rdir / "final_results.xlsx"))
                except Exception as _fin_err:
                    ss_set(_final_auto_saved=True, _final_save_path="",
                           _final_save_error=str(_fin_err))
            else:
                ss_set(_final_auto_saved=True)

        _final_xl_error = ss("_final_save_error", "")
        if _final_xl_error:
            _st.warning(f"⚠ Final auto-save failed: {_final_xl_error}")

        # ── Step 2 debug files — download + preview ───────────────────────────────
        if _show_adv and not _elm_done:
            _all_dbg_files = ss("_step2_debug_files", [])
            if _all_dbg_files:
                with _st.expander(
                    f"🔍 Step 2 debug files ({len(_all_dbg_files)} file(s))",
                    expanded=True,
                ):
                    # ── ZIP download (all files in one click) ─────────────────────
                    try:
                        _zip_bytes = build_debug_zip(_all_dbg_files)
                        _st.download_button(
                            label="⬇ Download all Step 2 debug files as ZIP",
                            data=_zip_bytes,
                            file_name=f"step2_debug_{build_run_tag()}_{ts()}.zip",
                            mime="application/zip",
                            use_container_width=True,
                            key="dl_all_debug_zip",
                        )
                    except Exception as _ze:
                        _st.warning(f"Could not build ZIP: {_ze}")

                    _st.divider()

                    # ── Per-file: download button + preview ───────────────────────
                    for _fi, _frec in enumerate(_all_dbg_files):
                        _tag  = " [DRY RUN]" if _frec.get("dry_run") else ""
                        _kind = _frec.get("kind", "")
                        _kind_label = {
                            "prompt":          "Prompt file",
                            "prompt_dry_run":  "Prompt file (dry run)",
                            "search":          "Search I/O file",
                        }.get(_kind, "Debug file")
                        _st.markdown(
                            f"**{_kind_label}{_tag}** — {_frec.get('company', '')} "
                            f"· {_frec.get('provider', '')}"
                        )
                        _st.caption(f"`{_frec.get('filename','')}`")
                        _debug_file_download_button(_frec, f"done_{_fi}")
                        _debug_file_preview_expander(_frec, f"done_{_fi}")
                        if _fi < len(_all_dbg_files) - 1:
                            _st.divider()

        # ── Primary browser download ──────────────────────────────────────────────
        if _elm_done:
            _fname_dl  = f"elm_results_{ts()}.xlsx"
            _dl_bytes  = df_to_excel_bytes(df_enriched)
        else:
            _fname_dl  = build_enriched_output_filename(ss("file_name", ""))
            _dl_bytes  = build_rich_excel_bytes(
                df_enriched,
                name_col=ss("_name_col"),
                domain_col=ss("_domain_col"),
                df_input_original=ss("_df_raw_original"),
                scoring_profile=ss("_scoring_profile", "default"),
                run_config=ss("_active_run_config"),
                run_mode="streamlit",
            )
        _st.download_button(
            label="⬇ Download lead scores",
            data=_dl_bytes,
            file_name=_fname_dl,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )

        if _adv_main:
            _render_advanced_results(
                df_enriched, debug_records_done, processed, _elm_done, debug_mode
            )

        # ── Restart ───────────────────────────────────────────────────────────────
        _st.divider()
        if _st.button("↺ Start a new enrichment", use_container_width=True, key="restart_btn"):
            reset_processing(clear_autosave=True)
            _st.rerun()

    # =============================================================================
    # STANDALONE COMMERCIAL FIT SCORER
    # Score a previously enriched file without running enrichment again.
    # =============================================================================

    if _adv_main and not ss("processing", False) and _SCORING_AVAILABLE:
        _st.divider()
        with _st.expander("🎯 Score an existing enrichment file", expanded=False):
            _st.caption(
                "Upload a previously enriched Excel or CSV file to apply commercial fit scoring "
                "without re-running the enrichment pipeline.  The file must contain the Step 3 "
                "model-signal columns (`sig_*`, `ti_*`, `has_*`, `is_public`, `has_funding`)."
            )
            _sa_upload = _st.file_uploader(
                "Upload enriched file (.xlsx · .xls · .csv)",
                type=["xlsx", "xls", "csv"],
                key="standalone_scorer_upload",
            )
            if _sa_upload is not None:
                try:
                    _sa_fname = _sa_upload.name
                    _sa_df = (
                        pd.read_csv(_sa_upload)
                        if _sa_fname.lower().endswith(".csv")
                        else pd.read_excel(_sa_upload)
                    )
                    _st.success(
                        f"**{_sa_fname}** loaded — "
                        f"{len(_sa_df):,} rows, {len(_sa_df.columns)} columns"
                    )

                    # Check for required signal columns
                    from commercial_fit_scoring import LEAN_COEFFICIENTS as _sa_lean_coeffs
                    _req_sig = [c for c in _sa_lean_coeffs if c in _sa_df.columns]
                    if not _req_sig:
                        _st.warning(
                            "⚠️ No model-signal columns found (`sig_*`, `ti_*`, `has_*`). "
                            "Run the enrichment pipeline with **Extract model signals (Step 3)** "
                            "enabled first, then upload that output here."
                        )
                    else:
                        _st.caption(f"Signal columns found: {len(_req_sig)} / {len(_sa_lean_coeffs)}")
                        _sa_df_scored = _score_dataframe(_sa_df.copy())

                        _sa_tier_order  = ["🥇 Hot", "🥈 Warm", "🥉 Cool", "❄️ Pass"]
                        _sa_tier_counts = _sa_df_scored["commercial_tier"].value_counts()
                        _sa_tc = _st.columns(4)
                        for _si, _st_tier in enumerate(_sa_tier_order):
                            _sa_tc[_si].metric(
                                _st_tier,
                                int(_sa_tier_counts.get(_st_tier, 0)),
                            )

                        _sa_score_cols = list(_SCORE_OUTPUT_COLS)
                        _sa_id_cols    = [c for c in _sa_df_scored.columns
                                          if c not in set(ALL_ENRICHMENT_FIELDS + _sa_score_cols)][:2]
                        _sa_disp_cols  = _sa_id_cols + _sa_score_cols
                        _sa_disp_df    = (
                            _sa_df_scored[[c for c in _sa_disp_cols if c in _sa_df_scored.columns]]
                            .copy()
                            .sort_values("final_commercial_fit_score", ascending=False)
                        )
                        _st.dataframe(_sa_disp_df, use_container_width=True, height=400)

                        _sa_stamp = ts()
                        _st.download_button(
                            "⬇ Download scored file (.xlsx)",
                            data=df_to_excel_bytes(_sa_df_scored),
                            file_name=f"scored_{_sa_fname.rsplit('.', 1)[0]}_{_sa_stamp}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            key="sa_dl_xlsx",
                        )
                        _st.download_button(
                            "⬇ Download scored file (.csv)",
                            data=df_to_csv_bytes(_sa_df_scored),
                            file_name=f"scored_{_sa_fname.rsplit('.', 1)[0]}_{_sa_stamp}.csv",
                            mime="text/csv",
                            use_container_width=True,
                            key="sa_dl_csv",
                        )
                except Exception as _sa_exc:
                    _st.error(f"Could not process file: {_sa_exc}")

    # ── Debug: competitor override self-test ──────────────────────────────────
    if _adv_main and not ss("processing", False):
        _st.divider()
        with _st.expander("🧪 Debug: Competitor override self-test", expanded=False):
            _st.caption(
                "Runs a zero-cost self-test of the competitor-customer ICP override logic. "
                "No API calls are made (no Serper, no Claude, no external web)."
            )
            if _st.button("Run competitor override self-test", key="btn_competitor_selftest"):
                with _st.spinner("Running self-test…"):
                    _st_res = run_competitor_override_selftest()
                _st.markdown(
                    f"**{_st_res['passed']} / {_st_res['total']} passed** "
                    + ("✅" if _st_res["failed"] == 0 else f"— ⚠️ {_st_res['failed']} failed")
                )
                _st_rows = []
                for _r in _st_res["results"]:
                    _st_rows.append({
                        "Test":    _r["name"],
                        "Result":  "✅ PASS" if _r["passed"] else "❌ FAIL",
                        "Details": _r["details"],
                    })
                import pandas as _pd_selftest
                _st.dataframe(
                    _pd_selftest.DataFrame(_st_rows),
                    use_container_width=True,
                    hide_index=True,
                )


if __name__ == "__main__":
    if "--self-test-competitor-override" in sys.argv or "--input" in sys.argv or cli_args_present():
        run_cli()
    elif running_under_streamlit():
        run_streamlit_app()
    else:
        print(
            "mYngle Lead Prioritizer\n"
            "  Batch mode:   python enrich_clients_claude.py --input FILE [--output-dir DIR] [--max-rows N]\n"
            "  Self-test:    python enrich_clients_claude.py --self-test-competitor-override\n"
            "  Streamlit UI: streamlit run enrich_clients_claude.py",
            file=sys.stderr,
        )
        sys.exit(0)
