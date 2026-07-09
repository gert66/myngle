"""
input_cleaner_lusha_edition.py — Layer 0: mYngle Input Cleaner · Lusha Edition
================================================================================
Prefilters raw Lusha company exports (e.g. Switzerland.xlsx, Spain.xlsx —
country-level "full export" files with thousands of rows x ~30 columns) on
commercially interesting companies BEFORE they go into the (expensive) Lead
Prioritizer. No Serper, no Firecrawl, no scraping — only the data Lusha
already provides, plus an optional cheap Anthropic Haiku prescreen on the
already-present Company Description for rows Lusha left unlabelled.

Column detection (``detect_lusha_columns``) only looks for the canonical
fields it needs (name, domain, description, employees, revenue, industries,
country, intent topics, LinkedIn URL); any extra firmographic columns a
given export happens to include (Company Year Founded, Total Funding
Amount, Company SIC/NAIC, Company Continent/State/City, Company Country
ISO, Company Intent Level, Topic Count Trend, ...) are simply ignored, so
new "full export" files are recognized automatically without code changes
as long as the core headers are present.

This is a standalone app (same pattern as input_cleaner_lite.py /
input_cleaner_register_edition.py): no coupling to the batch-core pipeline.

Entry point:  streamlit run input_cleaner_lusha_edition.py
"""

from __future__ import annotations

import io
import json
import re
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
import streamlit as st

try:
    import anthropic as _anthropic_lib
except ImportError:
    _anthropic_lib = None  # type: ignore[assignment]

from lead_hq_ai_interpreter import estimate_ai_cost_usd

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Input Cleaner · Lusha Edition",
    page_icon="🏢",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Same default Haiku model as lead_hq_ai_interpreter.py's AI-first HQ strategy
# (kept as a local constant rather than importing the private original, same
# convention as lead_prioritizer_core.py's own _DEFAULT_AI_MODEL).
_DEFAULT_AI_MODEL = "claude-haiku-4-5-20251001"


# ---------------------------------------------------------------------------
# Column detection — defensive, case-insensitive mapping of a raw Lusha
# company export onto the canonical fields this app works with.
# ---------------------------------------------------------------------------

_LUSHA_COLUMN_CANDIDATES: dict[str, tuple[str, ...]] = {
    "name": ("company name", "company_name", "name"),
    "domain": ("company domain", "company_domain", "domain", "website"),
    "description": ("company description", "company_description", "description"),
    "employees": (
        "company number of employees", "number of employees",
        "company employees", "employees",
    ),
    "revenue": ("company revenue", "revenue"),
    "main_industry": ("company main industry", "main industry", "industry"),
    "sub_industry": ("company sub industry", "sub industry", "sub-industry"),
    "country": ("company country", "country"),
    "intent_topics": (
        "company intent topics", "intent topics", "intent_topics",
    ),
    "linkedin_url": (
        "company linkedin url", "company linkedin", "linkedin url", "linkedin",
    ),
}

# The only two fields this app cannot function without.
REQUIRED_LUSHA_KEYS = ("name", "domain")


def _normalize_col_key(col: str) -> str:
    """Lowercase and collapse spaces/underscores/hyphens to a single space —
    tolerates the small header-spelling drift between Lusha exports."""
    return re.sub(r"[\s_\-]+", " ", str(col).strip().lower())


def detect_lusha_columns(df: pd.DataFrame) -> dict[str, "str | None"]:
    """Map each canonical field to the actual column name in ``df``, or
    ``None`` when that field isn't present in this particular export. Never
    raises — callers check ``missing_required_lusha_columns`` for the two
    fields (name, domain) this app truly cannot run without."""
    cols_norm = {_normalize_col_key(c): c for c in df.columns}
    mapping: dict[str, "str | None"] = {}
    for key, candidates in _LUSHA_COLUMN_CANDIDATES.items():
        found = None
        for cand in candidates:
            norm = _normalize_col_key(cand)
            if norm in cols_norm:
                found = cols_norm[norm]
                break
        mapping[key] = found
    return mapping


def missing_required_lusha_columns(mapping: dict) -> list[str]:
    return [k for k in REQUIRED_LUSHA_KEYS if not mapping.get(k)]


def _blank(value) -> bool:
    s = str(value if value is not None else "").strip()
    return not s or s.lower() in ("nan", "none")


# ---------------------------------------------------------------------------
# Step 1 — Deduplication on normalized domain.
# ---------------------------------------------------------------------------

def normalize_domain(raw) -> str:
    """Strip protocol, www, path, query. Return root domain lowercase, or
    "" when there is nothing usable (never raises)."""
    if not raw or not isinstance(raw, str):
        return ""
    d = raw.strip().lower()
    d = re.sub(r"^https?://", "", d)
    d = re.sub(r"^www\.", "", d)
    d = d.split("/")[0].split("?")[0].split("#")[0].strip()
    if not d or " " in d or d in ("nan", "none"):
        return ""
    return d


def _filled_field_count(row) -> int:
    """Number of non-blank fields in one row — used to pick a winner between
    duplicate-domain rows (most complete data wins)."""
    count = 0
    for v in row:
        if not _blank(v):
            count += 1
    return count


def dedupe_by_domain(df: pd.DataFrame, domain_col: str) -> "tuple[pd.DataFrame, int]":
    """Deduplicate on normalized domain. Among rows sharing a domain, the row
    with the most filled-in fields wins; ties keep the first-encountered row
    (stable, original file order). Rows with a blank/unusable domain are
    never deduplicated against each other or dropped here — a missing domain
    is not evidence two rows are the same company.

    Returns ``(deduped_df, removed_count)``.
    """
    if df.empty:
        return df.copy(), 0

    norm_domains = df[domain_col].apply(normalize_domain)
    filled_counts = df.apply(_filled_field_count, axis=1)

    work = df.copy()
    work["_norm_domain"] = norm_domains.values
    work["_filled_count"] = filled_counts.values
    work["_orig_order"] = range(len(work))

    blank_mask = work["_norm_domain"] == ""
    blank_rows = work[blank_mask]
    non_blank = work[~blank_mask]

    non_blank_sorted = non_blank.sort_values(
        ["_filled_count", "_orig_order"], ascending=[False, True], kind="stable")
    deduped_non_blank = non_blank_sorted.drop_duplicates(subset="_norm_domain", keep="first")

    removed = len(non_blank) - len(deduped_non_blank)
    result = pd.concat([deduped_non_blank, blank_rows]).sort_values("_orig_order")
    result = result.drop(columns=["_norm_domain", "_filled_count", "_orig_order"])
    return result.reset_index(drop=True), removed


# ---------------------------------------------------------------------------
# Step 2 — Industry exclusion rules (configurable, "scenario B" defaults:
# Higher Education / Training / E-Learning kept inside "Education").
# ---------------------------------------------------------------------------

_GOVERNMENT_INDUSTRY = "government"
_NONPROFIT_INDUSTRY = "community & nonprofit organizations"
_EDUCATION_INDUSTRY = "education"

# Education sub-industries that are NOT excluded even though the parent
# Main Industry ("Education") normally is — these are commercially relevant
# for corporate language training (institutions, corporate training vendors).
_EDUCATION_KEEP_SUB_INDUSTRIES = frozenset({
    "higher education", "training", "e-learning providers",
})

# Care-DELIVERY sub-industries only — hospitals and biotech/medtech are
# deliberately NOT in this set and stay IN.
_CARE_DELIVERY_SUB_INDUSTRIES = frozenset({
    "nursing homes & residential care facilities",
    "community & home healthcare services",
    "medical practices",
    "mental health care",
    "alternative medicine",
    "veterinary services",
})


@dataclass
class IndustryExclusionConfig:
    """Each rule can be toggled off per run from the sidebar. Defaults match
    the commercially agreed "scenario B" policy (Higher Education kept)."""
    exclude_government: bool = True
    exclude_nonprofit: bool = True
    exclude_education: bool = True
    exclude_care_delivery: bool = True


def classify_industry_exclusion(
    main_industry, sub_industry, config: IndustryExclusionConfig,
) -> str:
    """Returns a non-empty, human-readable exclusion reason when the row
    should be excluded on industry grounds, or ``""`` when it should not.

    A blank ``main_industry`` is NEVER excluded here — the absence of a
    label is not evidence of a non-ICP company (see the optional Haiku
    prescreen in Step 4 for these rows instead).
    """
    main = str(main_industry or "").strip()
    if not main:
        return ""
    main_lower = main.lower()
    sub = str(sub_industry or "").strip()
    sub_lower = sub.lower()

    if config.exclude_government and main_lower == _GOVERNMENT_INDUSTRY:
        return "Main Industry = Government"
    if config.exclude_nonprofit and main_lower == _NONPROFIT_INDUSTRY:
        return "Main Industry = Community & Nonprofit Organizations"
    if (config.exclude_education and main_lower == _EDUCATION_INDUSTRY
            and sub_lower not in _EDUCATION_KEEP_SUB_INDUSTRIES):
        return "Main Industry = Education (not Higher Education/Training/E-Learning)"
    if config.exclude_care_delivery and sub_lower in _CARE_DELIVERY_SUB_INDUSTRIES:
        return f"Sub Industry = {sub} (zorg-uitvoering)"
    return ""


# ---------------------------------------------------------------------------
# Step 3 — Hot list on Lusha intent topics. A filled Company Intent Topics
# field is a free buying signal and must never be silently dropped by an
# industry exclusion rule.
# ---------------------------------------------------------------------------

def has_intent_topics(value) -> bool:
    return not _blank(value)


def classify_rows(
    df: pd.DataFrame,
    mapping: dict,
    config: IndustryExclusionConfig,
) -> pd.DataFrame:
    """Adds ``hot_list``, ``industry_exclusion_reason``,
    ``intent_override_warning`` and ``excluded`` columns.

    An intent-flagged row that matches an exclusion rule is NEVER excluded —
    it is kept with ``intent_override_warning=True`` so a human decides,
    instead of the row silently disappearing.
    """
    out = df.copy()
    main_col = mapping.get("main_industry")
    sub_col = mapping.get("sub_industry")
    intent_col = mapping.get("intent_topics")

    main_series = out[main_col] if main_col else pd.Series([""] * len(out), index=out.index)
    sub_series = out[sub_col] if sub_col else pd.Series([""] * len(out), index=out.index)
    intent_series = out[intent_col] if intent_col else pd.Series([""] * len(out), index=out.index)

    out["hot_list"] = intent_series.apply(has_intent_topics)
    out["industry_exclusion_reason"] = [
        classify_industry_exclusion(m, s, config)
        for m, s in zip(main_series, sub_series)
    ]
    out["intent_override_warning"] = out["hot_list"] & (out["industry_exclusion_reason"] != "")
    out["excluded"] = (out["industry_exclusion_reason"] != "") & (~out["hot_list"])
    return out


# ---------------------------------------------------------------------------
# Step 4 — Optional Haiku prescreen for label-less rows (opt-in, off by
# default). ONLY uses the already-present Company Description — no Serper,
# no Firecrawl, no scraping. Same prompt-caching pattern as
# lead_hq_ai_interpreter.py's _call_anthropic_hq: a single, byte-for-byte
# static system block marked cache_control=ephemeral.
# ---------------------------------------------------------------------------

_PRESCREEN_SYSTEM_PROMPT = (
    "You are a B2B ICP (ideal customer profile) analyst for mYngle, a "
    "corporate language-training and coaching company. Given a company's own "
    "description, judge whether it plausibly needs corporate language "
    "training/coaching services for its employees. "
    "Reply ONLY with a valid JSON object — no prose, no markdown fences."
)

# IMPORTANT — prompt-cache boundary. This text, together with
# _PRESCREEN_SYSTEM_PROMPT, forms the single cache_control-marked system
# block sent on every prescreen call. For Anthropic's prompt cache to ever
# hit, this string MUST stay byte-for-byte identical across every call —
# never insert a per-row or per-run value here.
_PRESCREEN_STATIC_INSTRUCTIONS = """\
Classify and return JSON with these exact keys:
- "icp_prescreen": one of "likely_fit", "unclear", "unlikely_fit"
- "reason": one short sentence explaining your judgement (max 200 chars)

Rules:
- "likely_fit": the description describes a commercial company of meaningful
  size/complexity (multinational presence, professional services, industry,
  technology, manufacturing, finance, ...) where employees plausibly need
  business language training.
- "unlikely_fit": the description clearly describes a not-for-profit,
  government body, purely local micro-business, or an organisation with no
  plausible international/corporate training need.
- "unclear": the description is too thin, generic, or ambiguous to judge
  confidently either way.
- Never invent information not present in the description.
"""

_VALID_PRESCREEN_LABELS = frozenset({"likely_fit", "unclear", "unlikely_fit"})

# Rough token-estimate constants for the pre-run cost estimate shown in the
# UI — deliberately conservative (slightly over- rather than under-estimate)
# so a user is never surprised by the actual bill.
_PRESCREEN_STATIC_BLOCK_TOKENS_ESTIMATE = 220
_PRESCREEN_PER_LEAD_OVERHEAD_TOKENS = 40
_PRESCREEN_OUTPUT_TOKENS_ESTIMATE = 60
_PRESCREEN_DESCRIPTION_CHARS_PER_TOKEN = 4


def eligible_for_prescreen(df: pd.DataFrame, mapping: dict) -> "pd.Series":
    """Rows with a blank Main Industry AND a non-blank Company Description —
    exactly the rows Step 2 could never make an industry judgement about."""
    desc_col = mapping.get("description")
    if not desc_col:
        return pd.Series([False] * len(df), index=df.index)
    desc_present = df[desc_col].apply(lambda v: not _blank(v))
    main_col = mapping.get("main_industry")
    if not main_col:
        main_blank = pd.Series([True] * len(df), index=df.index)
    else:
        main_blank = df[main_col].apply(_blank)
    return main_blank & desc_present


def _extract_json_object(text: str) -> str:
    """Best-effort isolation of the JSON object in an AI response string —
    same pattern as lead_hq_ai_interpreter._extract_json_object: drops
    markdown fences and any prose before/after the object."""
    s = str(text or "").strip()
    if not s:
        return ""
    s = re.sub(r"^```(?:json|JSON)?\s*", "", s).strip()
    s = re.sub(r"\s*```$", "", s).strip()
    start, end = s.find("{"), s.rfind("}")
    if start != -1 and end != -1 and end > start:
        return s[start:end + 1].strip()
    return s


def parse_prescreen_response(raw: str) -> dict:
    """Defensive JSON parse of a prescreen response. Never raises — a broken,
    truncated, or unexpected-shape response degrades to a blank result
    (``icp_prescreen=""``) rather than crashing the run; the caller records
    that as a row-level miss, not a fatal error."""
    extracted = _extract_json_object(raw)
    try:
        data = json.loads(extracted)
    except Exception:
        return {"icp_prescreen": "", "reason": ""}
    if not isinstance(data, dict):
        return {"icp_prescreen": "", "reason": ""}
    label = str(data.get("icp_prescreen", "") or "").strip().lower()
    if label not in _VALID_PRESCREEN_LABELS:
        label = ""
    reason = str(data.get("reason", "") or "").strip()[:300]
    return {"icp_prescreen": label, "reason": reason}


def _build_prescreen_user_message(company_name: str, description: str, country: str) -> str:
    return (
        f"Company name: {company_name or '(unknown)'}\n"
        f"Country: {country or '(unknown)'}\n"
        f"Description:\n{(description or '').strip()[:1200]}\n"
    )


def _call_anthropic_prescreen(api_key: str, model: str, user_msg: str) -> "tuple[str, dict]":
    """One Anthropic prescreen call. Returns ``(raw_text, usage)``; raises on
    failure (caller isolates per-row, mirroring the rest of the pipeline)."""
    if _anthropic_lib is None:
        raise ImportError("anthropic package not installed")
    client = _anthropic_lib.Anthropic(api_key=api_key)
    response = client.messages.create(
        model=model,
        max_tokens=200,
        system=[
            {
                "type": "text",
                "text": _PRESCREEN_SYSTEM_PROMPT + "\n\n" + _PRESCREEN_STATIC_INSTRUCTIONS,
                "cache_control": {"type": "ephemeral"},
            }
        ],
        messages=[{"role": "user", "content": user_msg}],
    )
    raw_text = response.content[0].text if response.content else ""
    usage_obj = getattr(response, "usage", None)
    input_tokens = getattr(usage_obj, "input_tokens", None)
    output_tokens = getattr(usage_obj, "output_tokens", None)
    total = (input_tokens + output_tokens
             if input_tokens is not None and output_tokens is not None else None)
    return raw_text, {
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "total_tokens": total,
    }


def estimate_prescreen_cost(
    df: pd.DataFrame, mapping: dict, eligible_mask: "pd.Series",
    model: str = _DEFAULT_AI_MODEL,
) -> dict:
    """Rough pre-run cost estimate shown in the UI BEFORE the user starts the
    prescreen, so cost is never a surprise. Deliberately conservative."""
    n = int(eligible_mask.sum())
    if n == 0:
        return {"eligible_rows": 0, "estimated_input_tokens": 0,
                "estimated_output_tokens": 0, "estimated_cost_usd": 0.0}

    desc_col = mapping.get("description")
    if desc_col:
        avg_chars = df.loc[eligible_mask, desc_col].astype(str).str.len().mean()
    else:
        avg_chars = 0.0
    per_row_desc_tokens = int((avg_chars or 0) / _PRESCREEN_DESCRIPTION_CHARS_PER_TOKEN)

    total_input_tokens = (
        _PRESCREEN_STATIC_BLOCK_TOKENS_ESTIMATE
        + n * (per_row_desc_tokens + _PRESCREEN_PER_LEAD_OVERHEAD_TOKENS)
    )
    total_output_tokens = n * _PRESCREEN_OUTPUT_TOKENS_ESTIMATE
    cost = estimate_ai_cost_usd(model, total_input_tokens, total_output_tokens)
    return {
        "eligible_rows": n,
        "estimated_input_tokens": total_input_tokens,
        "estimated_output_tokens": total_output_tokens,
        "estimated_cost_usd": cost if cost is not None else 0.0,
    }


def prescreen_rows_with_ai(
    df: pd.DataFrame,
    mapping: dict,
    eligible_mask: "pd.Series",
    anthropic_api_key: str,
    model: str = _DEFAULT_AI_MODEL,
    progress_cb=None,
) -> pd.DataFrame:
    """Runs the Haiku prescreen for every eligible row. Adds
    ``icp_prescreen``/``icp_prescreen_reason``/``icp_prescreen_error``
    columns (all rows; blank for non-eligible or unprocessed rows).
    ``unlikely_fit`` rows are NEVER removed here — only marked, per the
    "human decides at download" rule.
    """
    out = df.copy()
    out["icp_prescreen"] = ""
    out["icp_prescreen_reason"] = ""
    out["icp_prescreen_error"] = ""

    name_col = mapping.get("name")
    desc_col = mapping.get("description")
    country_col = mapping.get("country")

    idxs = list(out.index[eligible_mask])
    n = len(idxs)
    for i, idx in enumerate(idxs):
        row = out.loc[idx]
        company_name = str(row.get(name_col, "") or "") if name_col else ""
        description = str(row.get(desc_col, "") or "") if desc_col else ""
        country = str(row.get(country_col, "") or "") if country_col else ""
        user_msg = _build_prescreen_user_message(company_name, description, country)
        try:
            raw_text, _usage = _call_anthropic_prescreen(anthropic_api_key, model, user_msg)
            parsed = parse_prescreen_response(raw_text)
            out.at[idx, "icp_prescreen"] = parsed["icp_prescreen"]
            out.at[idx, "icp_prescreen_reason"] = parsed["reason"]
        except Exception as exc:
            out.at[idx, "icp_prescreen_error"] = f"{type(exc).__name__}: {str(exc)[:200]}"
        if progress_cb:
            progress_cb(i + 1, n)
    return out


# ---------------------------------------------------------------------------
# Step 5 — Sorting, batch-app-compatible columns, and Excel export.
# ---------------------------------------------------------------------------

def parse_employee_count(value) -> float:
    """Numeric sort key for a Lusha employee-count field, which may be a
    plain number or a range ("51-200", "10,001+"). Uses the upper bound of a
    range (the more optimistic read of company size). Unparseable/blank
    values sort last (``-1.0``), never raise."""
    s = str(value if value is not None else "").strip()
    if not s or s.lower() in ("nan", "none"):
        return -1.0
    s = s.replace(",", "").replace(" ", "")
    nums = re.findall(r"\d+(?:\.\d+)?", s)
    if not nums:
        return -1.0
    return max(float(n) for n in nums)


def sort_selected_rows(df: pd.DataFrame, mapping: dict) -> pd.DataFrame:
    """Hot list first, then by company size (largest first)."""
    emp_col = mapping.get("employees")
    out = df.copy()
    if emp_col and emp_col in out.columns:
        out["_employee_sort_key"] = out[emp_col].apply(parse_employee_count)
    else:
        out["_employee_sort_key"] = -1.0

    if "hot_list" in out.columns:
        out = out.sort_values(
            ["hot_list", "_employee_sort_key"], ascending=[False, False], kind="stable")
    else:
        out = out.sort_values("_employee_sort_key", ascending=False, kind="stable")
    return out.drop(columns=["_employee_sort_key"]).reset_index(drop=True)


def add_batch_app_compatible_columns(df: pd.DataFrame, mapping: dict) -> pd.DataFrame:
    """Adds ``domain`` and ``country`` columns (non-destructive — the
    original Lusha columns are kept) so this sheet's columns are directly
    selectable in the Lead Prioritizer batch app's column-mapping dropdowns
    (``COMPANY_CANDIDATES`` / ``DOMAIN_CANDIDATES`` / ``COUNTRY_CANDIDATES``
    in lead_prioritizer_batch_app.py). "Company Name" already matches
    COMPANY_CANDIDATES verbatim, so no rename is needed for the name field.
    """
    out = df.copy()
    domain_col = mapping.get("domain")
    country_col = mapping.get("country")
    if domain_col and domain_col in out.columns and "domain" not in out.columns:
        out["domain"] = out[domain_col].apply(normalize_domain)
    if country_col and country_col in out.columns and "country" not in out.columns:
        out["country"] = out[country_col]
    return out


def _header_fill():
    from openpyxl.styles import Font, PatternFill
    fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=10)
    return fill, font


# Control characters (other than tab/newline/CR) that openpyxl's XML writer
# rejects with IllegalCharacterError — Lusha's scraped company descriptions
# occasionally contain them.
_ILLEGAL_XLSX_CHARS_RE = re.compile(
    "[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]"
)


def _sanitize_cell_value(val):
    if isinstance(val, str):
        return _ILLEGAL_XLSX_CHARS_RE.sub("", val)
    return val


def _write_sheet(
    ws, df: pd.DataFrame, highlight_rules: "list[tuple[str, str]] | None" = None,
) -> None:
    """Write df to an openpyxl worksheet with header styling, freeze panes,
    autofilter, and optional row highlighting. ``highlight_rules`` is an
    ordered list of ``(column_name, hex_color)``; the first rule whose column
    is truthy for a row wins."""
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    hdr_fill, hdr_font = _header_fill()
    cols = list(df.columns)

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = min(max(len(str(col)) + 2, 12), 50)

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        fill = None
        for col_name, color in (highlight_rules or []):
            if col_name in cols and bool(row.get(col_name)):
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                break
        for ci, col in enumerate(cols, 1):
            val = row[col]
            if isinstance(val, float) and val != val:
                val = ""
            val = _sanitize_cell_value(val)
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(wrap_text=False, vertical="top")
            if fill is not None:
                cell.fill = fill

    ws.freeze_panes = "A2"
    if len(df) > 0 and cols:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    ws.row_dimensions[1].height = 18


def build_summary_rows(funnel: dict) -> pd.DataFrame:
    return pd.DataFrame([{"Stap": k, "Aantal": v} for k, v in funnel.items()])


def build_excel(
    selected_df: pd.DataFrame,
    hot_df: pd.DataFrame,
    excluded_df: pd.DataFrame,
    funnel: dict,
) -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Geselecteerd"
    _write_sheet(ws1, selected_df, highlight_rules=[
        ("intent_override_warning", "FFEB9C"),  # warning yellow, wins over plain hot_list
        ("hot_list", "E2EFDA"),                  # light green
    ])

    ws2 = wb.create_sheet("Hot list (intent)")
    _write_sheet(ws2, hot_df)

    ws3 = wb.create_sheet("Uitgesloten")
    _write_sheet(ws3, excluded_df)

    ws4 = wb.create_sheet("Samenvatting")
    _write_sheet(ws4, build_summary_rows(funnel))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------

def _load_secrets_key(*names: str) -> "str | None":
    try:
        for name in names:
            val = st.secrets.get(name)
            if val:
                return val
    except Exception:
        pass
    return None


def _load_file(uploaded) -> "pd.DataFrame | None":
    raw = uploaded.read()
    name = uploaded.name.lower()
    try:
        if name.endswith(".csv"):
            return pd.read_csv(io.BytesIO(raw), dtype=str).fillna("")
        return pd.read_excel(io.BytesIO(raw), dtype=str).fillna("")
    except Exception as exc:
        st.error(f"Could not read file: {exc}")
        return None


# ---------------------------------------------------------------------------
# Path-mode input/output — reading a Lusha export directly from a local
# filesystem path (a browser upload never reveals the original file path),
# and writing the cleaned result back next to it. Kept as plain, Streamlit-
# free functions so they're independently testable.
# ---------------------------------------------------------------------------

_SUPPORTED_PATH_EXTENSIONS = (".xlsx", ".xls", ".csv")


def output_filename_for(source_name: str) -> str:
    """<source-stem>_cleaned.xlsx — e.g. Switzerland.xlsx -> Switzerland_cleaned.xlsx.
    Works for any source extension (.xlsx/.xls/.csv); the output is always .xlsx."""
    stem = Path(str(source_name or "")).stem
    return f"{stem}_cleaned.xlsx"


def output_path_for(source_path: "Path | str") -> Path:
    """Full output path: same directory as the source file, filename per
    ``output_filename_for()``."""
    return Path(source_path).with_name(output_filename_for(Path(source_path).name))


def write_output_next_to_source(source_path: "Path | str", excel_bytes: bytes) -> "tuple[Path, bool]":
    """Writes ``excel_bytes`` to ``output_path_for(source_path)``. Returns
    ``(output_path, overwritten)`` — ``overwritten`` is True when a file
    already existed at that path before writing. Can raise (a real
    filesystem error, e.g. permissions); callers wrap this in a try/except."""
    out_path = output_path_for(source_path)
    overwritten = out_path.exists()
    out_path.write_bytes(excel_bytes)
    return out_path, overwritten


def _clean_path_input(raw: str) -> str:
    """Strips surrounding whitespace and matching quote characters from a
    pasted file path. Windows Explorer's "Copy as path" wraps the result in
    double quotes (``"C:\\...\\Switzerland.xlsx"``); this also tolerates
    single quotes and repeated/nested quoting. Never raises."""
    s = str(raw or "").strip()
    while len(s) >= 2 and s[0] in ("'", '"') and s[-1] == s[0]:
        s = s[1:-1].strip()
    return s


def resolve_input_mode(manual_path_raw: str, uploaded_present: bool) -> str:
    """Decides which input source wins: ``"path"``, ``"upload"``, or
    ``"none"``. The path field always wins when filled, even if a browser
    upload is also present — pure decision logic kept separate from
    Streamlit I/O so the priority rule is independently testable."""
    if _clean_path_input(manual_path_raw):
        return "path"
    if uploaded_present:
        return "upload"
    return "none"


def load_dataframe_from_path(path: "Path | str") -> "tuple[pd.DataFrame | None, str]":
    """Reads a Lusha export directly from a local filesystem path. Returns
    ``(dataframe_or_None, error_message)`` — ``error_message`` is ``""`` on
    success. Never raises."""
    p = Path(path)
    if not p.exists():
        return None, f"Pad bestaat niet: {p}"
    if not p.is_file():
        return None, f"Pad is geen bestand: {p}"
    ext = p.suffix.lower()
    if ext not in _SUPPORTED_PATH_EXTENSIONS:
        return None, (
            f"Bestandstype niet ondersteund ({ext or '(geen extensie)'}) — "
            "verwacht .xlsx, .xls of .csv."
        )
    try:
        if ext == ".csv":
            df = pd.read_csv(p, dtype=str).fillna("")
        else:
            df = pd.read_excel(p, dtype=str).fillna("")
        return df, ""
    except Exception as exc:
        return None, f"Kon bestand niet lezen: {exc}"


def _sidebar_exclusion_config() -> IndustryExclusionConfig:
    st.sidebar.subheader("Industrie-uitsluitingen")
    exclude_government = st.sidebar.checkbox(
        "Government uitsluiten", value=True)
    exclude_nonprofit = st.sidebar.checkbox(
        "Community & Nonprofit Organizations uitsluiten", value=True)
    exclude_education = st.sidebar.checkbox(
        "Education uitsluiten (behalve Higher Education, Training, "
        "E-Learning Providers)", value=True)
    exclude_care_delivery = st.sidebar.checkbox(
        "Zorg-uitvoering uitsluiten (nursing homes, home healthcare, "
        "medical practices, mental health care, alternative medicine, "
        "veterinary services — ziekenhuizen/biotech blijven wél in)",
        value=True)
    return IndustryExclusionConfig(
        exclude_government=exclude_government,
        exclude_nonprofit=exclude_nonprofit,
        exclude_education=exclude_education,
        exclude_care_delivery=exclude_care_delivery,
    )


def _funnel_metrics(funnel: dict) -> None:
    cols = st.columns(len(funnel))
    for col, (label, val) in zip(cols, funnel.items()):
        col.metric(label, val)


def main():
    st.title("🏢 Input Cleaner · Lusha Edition")
    st.caption(
        "Layer 0 · mYngle Sales Intelligence · Prefilter raw Lusha company "
        "exports on commercially interesting companies before Lead Prioritizer")

    config = _sidebar_exclusion_config()

    st.sidebar.subheader("Haiku-prescreen (opt-in)")
    run_prescreen = st.sidebar.checkbox(
        "Prescreen rijen zonder Main Industry met Company Description "
        "(Anthropic Haiku, kost geld)", value=False,
        help="Alleen voor rijen zonder Main Industry maar mét Company "
             "Description. Gebruikt UITSLUITEND de al aanwezige description "
             "— geen Serper, geen Firecrawl, geen scraping. Rijen worden "
             "gemarkeerd (likely_fit / unclear / unlikely_fit), nooit "
             "automatisch verwijderd.")

    anthropic_key = _load_secrets_key("ANTHROPIC_API_KEY", "anthropic_api_key")
    if run_prescreen:
        if anthropic_key:
            st.sidebar.success("Anthropic API key loaded from secrets.")
        else:
            manual_key = st.sidebar.text_input(
                "Paste Anthropic API key", type="password", key="manual_anthropic")
            if manual_key.strip():
                anthropic_key = manual_key.strip()

    manual_path_raw = st.text_input(
        "Volledig pad naar het Lusha-bestand op deze computer",
        value="", key="lusha_path",
        placeholder=r"C:\Users\...\Switzerland.xlsx",
    )
    st.caption(
        "Tip: Shift + rechtermuisklik op het bestand in Verkenner → "
        "'Als pad kopiëren', en plak het hier. Het resultaat "
        "(`<naam>_cleaned.xlsx`) wordt dan automatisch naast het "
        "bronbestand weggeschreven.")

    with st.expander("Of: upload via de browser", expanded=False):
        uploaded = st.file_uploader(
            "Upload Lusha company export (CSV or Excel .xlsx)",
            type=["csv", "xlsx"], key="lusha_upload",
        )
        st.caption(
            "Bij een browser-upload is het oorspronkelijke bestandspad "
            "onbekend, dus automatisch wegschrijven naast de bron is niet "
            "mogelijk — alleen de downloadknop (naar de Downloads-map) is "
            "dan beschikbaar.")

    df = None
    source_name = ""
    source_path: "Path | None" = None
    manual_path = _clean_path_input(manual_path_raw)
    input_mode = resolve_input_mode(manual_path_raw, uploaded is not None)

    if input_mode == "path":
        if uploaded is not None:
            st.info(
                "Zowel een pad als een browser-upload zijn ingevuld — het "
                "pad heeft voorrang; de upload wordt genegeerd.")
        source_path = Path(manual_path)
        df, path_error = load_dataframe_from_path(source_path)
        if path_error:
            st.error(path_error)
            return
        source_name = source_path.name
    elif input_mode == "upload":
        df = _load_file(uploaded)
        if df is None:
            return
        source_name = uploaded.name
    else:
        st.info(
            "Vul een volledig bestandspad in, of upload een Lusha company "
            "export via de browser. Expected columns include **Company "
            "Name**, **Company Domain**, Company Description, Company "
            "Number of Employees, Company Revenue, Company Main Industry, "
            "Company Sub Industry, Company Country, Company Intent Topics, "
            "Company linkedin URL.")
        return

    st.success(f"Loaded {len(df)} rows, {len(df.columns)} columns from "
               f"**{source_name}**.")

    mapping = detect_lusha_columns(df)
    missing = missing_required_lusha_columns(mapping)
    if missing:
        st.error(
            "This file is missing required column(s): "
            f"**{', '.join(missing)}**. Expected a Company Name and a "
            "Company Domain column (case-insensitive). Detected columns: "
            f"{', '.join(df.columns)}"
        )
        return

    with st.expander("Column mapping (auto-detected)", expanded=False):
        st.json({k: v for k, v in mapping.items()})

    total_rows = len(df)

    # ── Step 1: dedupe ────────────────────────────────────────────────────
    deduped_df, removed_dupes = dedupe_by_domain(df, mapping["domain"])
    st.info(f"Removed **{removed_dupes}** duplicate row(s) on domain "
            f"(kept the most-complete row per domain).")

    # ── Step 2 + 3: industry exclusions + hot list ───────────────────────
    classified_df = classify_rows(deduped_df, mapping, config)

    excluded_df = classified_df[classified_df["excluded"]].copy()
    if "industry_exclusion_reason" in excluded_df.columns:
        excluded_df = excluded_df.rename(
            columns={"industry_exclusion_reason": "Uitsluitreden"})
    selected_df = classified_df[~classified_df["excluded"]].copy()

    override_count = int(classified_df["intent_override_warning"].sum())
    if override_count:
        st.warning(
            f"**{override_count}** rij(en) met een intent-signaal vielen "
            "onder een uitsluitregel maar zijn BEHOUDEN (waarschuwingsvlag "
            "`intent_override_warning`) — controleer deze handmatig.")

    # ── Step 4: optional Haiku prescreen ──────────────────────────────────
    prescreen_ran = False
    eligible_mask = eligible_for_prescreen(selected_df, mapping)
    if run_prescreen:
        cost = estimate_prescreen_cost(selected_df, mapping, eligible_mask)
        st.subheader("Haiku-prescreen kostenschatting")
        c1, c2, c3 = st.columns(3)
        c1.metric("Rijen zonder Main Industry (met description)", cost["eligible_rows"])
        c2.metric("Geschatte tokens (in/out)",
                   f"{cost['estimated_input_tokens']}/{cost['estimated_output_tokens']}")
        c3.metric("Geschatte kosten (USD)", f"${cost['estimated_cost_usd']:.4f}")

        if cost["eligible_rows"] == 0:
            st.caption("No eligible rows for prescreen — nothing to run.")
        elif not anthropic_key:
            st.error("No Anthropic API key available — cannot run the prescreen.")
        elif st.button("🤖 Start Haiku-prescreen", type="primary"):
            progress_bar = st.progress(0)
            status_text = st.empty()

            def _cb(i, n):
                if n:
                    progress_bar.progress(i / n)
                status_text.caption(f"Prescreening row {i} of {n}…")

            selected_df = prescreen_rows_with_ai(
                selected_df, mapping, eligible_mask, anthropic_key, progress_cb=_cb)
            progress_bar.progress(1.0)
            status_text.caption("Prescreen done.")
            st.session_state["lusha_selected_prescreened"] = selected_df
            prescreen_ran = True

    if "lusha_selected_prescreened" in st.session_state and run_prescreen:
        selected_df = st.session_state["lusha_selected_prescreened"]
        prescreen_ran = True

    # ── Step 5: sort + batch-app-compatible columns ───────────────────────
    selected_df = sort_selected_rows(selected_df, mapping)
    selected_df = add_batch_app_compatible_columns(selected_df, mapping)
    hot_df = selected_df[selected_df["hot_list"]].copy() if "hot_list" in selected_df.columns \
        else selected_df.iloc[0:0].copy()

    st.markdown("---")
    st.markdown("### Resultaat")

    funnel = {
        "Rijen in bronbestand": total_rows,
        "Duplicaten verwijderd": removed_dupes,
        "Na ontdubbeling": len(deduped_df),
        "Uitgesloten op industrie": len(excluded_df),
        "Behouden ondanks uitsluitregel (intent)": override_count,
        "Geselecteerd (totaal)": len(selected_df),
        "Waarvan hot list (intent topics)": len(hot_df),
    }
    if prescreen_ran and "icp_prescreen" in selected_df.columns:
        funnel["Prescreen: likely_fit"] = int((selected_df["icp_prescreen"] == "likely_fit").sum())
        funnel["Prescreen: unclear"] = int((selected_df["icp_prescreen"] == "unclear").sum())
        funnel["Prescreen: unlikely_fit"] = int((selected_df["icp_prescreen"] == "unlikely_fit").sum())

    _funnel_metrics({k: v for k, v in list(funnel.items())[:5]})

    show_cols = [c for c in [
        mapping.get("name"), "domain", "country", mapping.get("main_industry"),
        mapping.get("sub_industry"), "hot_list", "intent_override_warning",
        "icp_prescreen", "icp_prescreen_reason",
    ] if c and c in selected_df.columns]
    st.dataframe(selected_df[show_cols], use_container_width=True, height=350)

    with st.expander(f"Uitgesloten rijen ({len(excluded_df)})", expanded=False):
        st.dataframe(excluded_df, use_container_width=True)

    # ── Download / write to disk ──────────────────────────────────────────
    excel_bytes = build_excel(selected_df, hot_df, excluded_df, funnel)
    file_name = output_filename_for(source_name)

    if source_path is not None:
        try:
            written_path, overwritten = write_output_next_to_source(source_path, excel_bytes)
            if overwritten:
                st.success(
                    f"Resultaat weggeschreven naar `{written_path}` "
                    "(bestaand bestand overschreven).")
            else:
                st.success(f"Resultaat weggeschreven naar `{written_path}`.")
        except Exception as exc:
            st.error(f"Kon resultaat niet wegschrijven naar schijf: {exc}")

    st.download_button(
        "⬇ Download cleaned Excel",
        data=excel_bytes,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    if source_path is None:
        st.caption(
            "Bestand geüpload via de browser — het originele bestandspad is "
            "onbekend, dus automatisch wegschrijven naast de bron is niet "
            "mogelijk. Gebruik de downloadknop hierboven, of vul het "
            "volledige pad naar het bestand in om het resultaat automatisch "
            "naast het inputbestand weg te schrijven.")


if __name__ == "__main__":
    main()
