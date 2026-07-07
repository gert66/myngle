"""
Validate an HQ Recovery export workbook.

Usage:
    python validate_hq_recovery_export.py <path-to-export.xlsx>

The export workbook must contain the sheet produced by the HQ Recovery
workflow (default: 'Opportunity Input Full').  Run this script after
exporting the 100-row result from the Streamlit app.

Reports:
  1. Run summary
  2. Domestic HQ incorrectly scored 3
  3. Score-3 quality check
  4. Manual review quality check
  5. Query discipline check
  6. Row-limit check
"""

import sys
import re
from openpyxl import load_workbook

# ── Country normalizer (mirrors _normalize_country_for_hq in the app) ────────

def _normalize(value: object) -> str:
    text = re.sub(r"\s+", " ", re.sub(r"\.", "", str(value or "").strip().lower()))
    _MAP = {
        "it": "italy", "ita": "italy", "italia": "italy",
        "italy": "italy", "italian": "italy",
        "de": "germany", "deu": "germany",
        "germany": "germany", "deutschland": "germany", "german": "germany",
        "fr": "france", "fra": "france",
        "france": "france", "french": "france",
        "uk": "united kingdom", "gb": "united kingdom", "gbr": "united kingdom",
        "united kingdom": "united kingdom", "great britain": "united kingdom",
        "england": "united kingdom",
        "us": "united states", "usa": "united states",
        "united states": "united states",
        "united states of america": "united states",
        "ch": "switzerland", "che": "switzerland",
        "switzerland": "switzerland", "swiss": "switzerland",
        "lu": "luxembourg", "lux": "luxembourg", "luxembourg": "luxembourg",
        "nl": "netherlands", "nld": "netherlands",
        "netherlands": "netherlands", "holland": "netherlands",
    }
    return _MAP.get(text, text)


def _safe_float(v):
    try:
        return float(v)
    except Exception:
        return None


def load_sheet(path: str) -> tuple[list[str], list[dict]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    # Prefer 'Opportunity Input Full', else first sheet
    target = "Opportunity Input Full"
    if target not in wb.sheetnames:
        target = wb.sheetnames[0]
    ws = wb[target]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c or "").strip() for c in rows[0]]
    data = []
    for row in rows[1:]:
        data.append({headers[i]: row[i] for i in range(len(headers))})
    wb.close()
    return headers, data


def main():
    if len(sys.argv) < 2:
        print("Usage: python validate_hq_recovery_export.py <export.xlsx>")
        sys.exit(1)

    path = sys.argv[1]
    print(f"\n{'='*72}")
    print(f"Validating: {path}")
    print(f"{'='*72}\n")

    headers, rows = load_sheet(path)
    total_rows = len(rows)

    # ── Filter to rows touched by HQ Recovery ───────────────────────────────
    processed    = [r for r in rows if str(r.get("hq_recovery_processed") or "").strip().lower() == "yes"]
    skipped_rl   = [r for r in rows if str(r.get("hq_recovery_skip_reason") or "").strip().lower() == "row_limit"]
    selected_all = [r for r in rows if str(r.get("hq_recovery_selected") or "").strip().lower() == "yes"]
    unchanged    = [r for r in rows if str(r.get("hq_recovery_selected") or "").strip().lower() not in ("yes",)]

    n_processed  = len(processed)
    n_skipped_rl = len(skipped_rl)
    n_selected   = len(selected_all)
    n_unchanged  = total_rows - n_selected

    # Updated to score 3
    score3_rows  = [r for r in processed if _safe_float(r.get("sig_foreign_hq_score_reviewed")) == 3]
    n_score3     = len(score3_rows)

    # Needs manual review
    manual_rows  = [r for r in processed if str(r.get("needs_manual_review") or "").strip().lower() == "yes"]
    n_manual     = len(manual_rows)

    # Domestic incorrectly scored 3
    domestic_wrong = []
    for r in processed:
        det_norm = _normalize(r.get("hq_detected_country"))
        inp_norm = _normalize(r.get("input_country_used") or r.get("country") or "IT")
        score    = _safe_float(r.get("sig_foreign_hq_score_reviewed"))
        if det_norm and inp_norm and det_norm == inp_norm and score == 3:
            domestic_wrong.append(r)

    # ── 1. Run Summary ───────────────────────────────────────────────────────
    print("1. RUN SUMMARY")
    print(f"   Total rows in workbook       : {total_rows}")
    print(f"   Selected (full set)          : {n_selected}")
    print(f"   Rows processed               : {n_processed}")
    print(f"   Rows skipped by row limit    : {n_skipped_rl}")
    print(f"   Rows unchanged               : {n_unchanged}")
    print(f"   Updated to score 3           : {n_score3}")
    print(f"   Needs manual review          : {n_manual}")
    print(f"   Domestic HQ incorrectly → 3 : {len(domestic_wrong)}")

    # ── 2. Domestic HQ Safety Check ─────────────────────────────────────────
    print(f"\n2. DOMESTIC HQ SAFETY CHECK  (target: 0)")
    if not domestic_wrong:
        print("   ✓ Domestic HQ incorrectly scored 3 = 0 — PASS")
    else:
        print(f"   ✗ {len(domestic_wrong)} row(s) incorrectly scored 3:")
        for r in domestic_wrong:
            print(f"     company : {r.get('company_name') or r.get('name','?')}")
            print(f"     domain  : {r.get('domain','?')}")
            print(f"     input_c : {r.get('input_country_used') or r.get('country','?')}")
            print(f"     det_c   : {r.get('hq_detected_country','?')}")
            print(f"     det_city: {r.get('hq_detected_city','?')}")
            print(f"     score   : {r.get('sig_foreign_hq_score_reviewed','?')}")
            print(f"     quote   : {str(r.get('hq_evidence_quote') or r.get('domain_root_hq_evidence_quote',''))[:120]}")
            print()

    # ── 3. Score-3 Quality Check ─────────────────────────────────────────────
    print(f"\n3. SCORE-3 QUALITY CHECK  ({n_score3} rows)")
    _ITALY_RE = re.compile(
        r"\b(?:Italy|Italia|Italian|Modena|Ancona|Osimo|Senigallia|Bergamo|"
        r"Milano|Rome|Torino|Firenze|Bologna|Napoli|Venezia|Brescia)\b",
        re.IGNORECASE,
    )
    suspicious = []
    for r in score3_rows:
        company = r.get("company_name") or r.get("name", "?")
        domain  = r.get("domain", "?")
        inp_c   = r.get("input_country_used") or r.get("country", "?")
        det_c   = r.get("hq_detected_country", "?")
        det_city= r.get("hq_detected_city", "?")
        par_c   = r.get("parent_group_hq_country", "?")
        struct  = r.get("hq_structure_type", "?")
        conf    = r.get("hq_confidence", "?")
        manual  = r.get("needs_manual_review", "?")
        ev_url  = (r.get("domain_root_hq_evidence_url") or r.get("hq_evidence_url") or "")
        ev_quote= str(r.get("domain_root_hq_evidence_quote") or r.get("hq_evidence_quote") or "")

        # Suspicion flags
        flags = []
        if _ITALY_RE.search(ev_quote) and _normalize(inp_c) == "italy":
            flags.append("Italy token in evidence quote despite score=3")
        if _normalize(det_c) == "italy" and _normalize(inp_c) == "italy":
            flags.append("domestic Italy misclassified as score=3")
        if "regional" in ev_quote.lower() or "north america" in ev_quote.lower():
            flags.append("regional/North America HQ — not true global parent")

        if flags:
            suspicious.append((company, domain, det_c, ev_quote[:120], flags))

        print(f"   • {company[:50]}")
        print(f"     domain  : {domain}  /  input={inp_c}  det={det_c} ({det_city})")
        print(f"     parent  : {par_c}  struct={struct}  conf={conf}  manual={manual}")
        print(f"     ev_url  : {ev_url[:80]}")
        print(f"     quote   : {ev_quote[:130]}")
        if flags:
            print(f"     ⚠ FLAGS : {' | '.join(flags)}")
        print()

    print(f"   Suspicious score-3 rows: {len(suspicious)}")

    # ── 4. Manual Review Reasons ─────────────────────────────────────────────
    print(f"\n4. MANUAL REVIEW QUALITY CHECK  ({n_manual} rows)")
    reason_counts: dict[str, int] = {}
    for r in manual_rows:
        trigger = str(r.get("hq_review_trigger") or r.get("sig_foreign_hq_review_reason") or "unresolved")
        bucket = "other"
        tl = trigger.lower()
        if "unresolved" in tl or not trigger.strip():
            bucket = "unresolved"
        elif "domestic" in tl or "italy" in tl:
            bucket = "domestic evidence"
        elif "weak" in tl or "low" in tl:
            bucket = "weak evidence"
        elif "collision" in tl or "generic" in tl:
            bucket = "generic name"
        elif "regional" in tl:
            bucket = "regional HQ"
        elif "unrelated" in tl:
            bucket = "unrelated domain"
        reason_counts[bucket] = reason_counts.get(bucket, 0) + 1
    for k, v in sorted(reason_counts.items(), key=lambda x: -x[1]):
        print(f"   {k:30s}: {v}")

    # ── 5. Query Discipline Check ─────────────────────────────────────────────
    print(f"\n5. QUERY DISCIPLINE CHECK")
    bad_queries = []
    for r in processed:
        qs = str(r.get("serper_queries_used") or "")
        company = r.get("company_name") or r.get("name", "?")
        domain  = r.get("domain", "?")
        dr_query = str(r.get("domain_root_hq_query") or "")
        issues = []
        if '"' in qs:
            issues.append("quoted query")
        if re.search(r"\bsede\b|\blegale\b", qs, re.I):
            issues.append("sede legale query")
        if re.search(r"\bsite:", qs):
            issues.append("site: query")
        if re.search(r"\bhead\s+office\b", qs, re.I):
            issues.append("head office query")
        if str(r.get("anthropic_hq_review_used") or "").lower() == "yes":
            issues.append("Anthropic called")
        if issues:
            bad_queries.append((company, domain, qs[:100], issues))
    if bad_queries:
        print(f"   ⚠ {len(bad_queries)} row(s) with unexpected queries:")
        for company, domain, qs, issues in bad_queries[:10]:
            print(f"     {company[:40]} / {domain}  → {', '.join(issues)}")
            print(f"     queries: {qs}")
    else:
        print(f"   ✓ No quoted/sede/site:/head-office/Anthropic queries in processed rows")
    n_dr_used = sum(1 for r in processed if str(r.get("domain_root_hq_search_used") or "").lower() == "true")
    print(f"   domain_root_hq_search used : {n_dr_used} / {n_processed} rows")

    # ── 6. Row-Limit Check ───────────────────────────────────────────────────
    print(f"\n6. ROW-LIMIT CHECK")
    print(f"   Rows processed              : {n_processed}")
    print(f"   Rows skipped by limit       : {n_skipped_rl}")
    print(f"   Selected (full set)         : {n_selected}")
    calc_skipped = n_selected - n_processed
    ok_math = calc_skipped == n_skipped_rl
    print(f"   selected - processed = {calc_skipped}  (skip_reason count={n_skipped_rl}) {'✓' if ok_math else '✗'}")

    # ── 7. Final Recommendation ──────────────────────────────────────────────
    print(f"\n7. RECOMMENDATION")
    if len(domestic_wrong) == 0 and len(suspicious) == 0:
        print("   ✓ SAFE to proceed to row limit 500.")
        print("   No domestic-Italy misclassifications found.")
        print("   All score-3 rows appear to have genuine foreign-parent evidence.")
    elif len(domestic_wrong) == 0 and len(suspicious) > 0:
        print(f"   ⚠ Domestic guard OK, but {len(suspicious)} suspicious score-3 row(s) need review.")
        print("   Consider manual review before running on 500 rows.")
    else:
        print(f"   ✗ STOP: {len(domestic_wrong)} domestic row(s) incorrectly scored 3.")
        print("   Fix the domestic guard before proceeding.")

    print(f"\n{'='*72}\n")


if __name__ == "__main__":
    main()
