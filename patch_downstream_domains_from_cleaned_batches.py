"""
patch_downstream_domains_from_cleaned_batches.py

Standalone utility: patch corrected domains from cleaned-domain batch files
into downstream enriched/output Excel files for a queue such as Italy200.

Does NOT run enrichment, call APIs, use Lusha, or recalculate scores.
"""

import argparse
import csv
import os
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

EXCLUDE_FOLDERS = {"_domain_patched", "url_patched", "_archive", "_logs"}

SOURCE_SHEET_PREFERENCE = ["Best Guess Input", "Cleaned Register Input"]

SOURCE_COL_COMPANY = "company_name"
SOURCE_COL_URL = "website_url"
SOURCE_COL_CANONICAL_URL = "canonical_company_url"
SOURCE_COL_CANONICAL_DOMAIN = "canonical_company_domain"
SOURCE_COL_CITY = "city"
SOURCE_COL_PROVINCE = "province"

URL_LIKE_COLS = {"website_url", "canonical_company_url"}
DOMAIN_LIKE_COLS = {
    "domain",
    "input_domain",
    "validated_domain",
    "canonical_company_domain",
    "domain_used_for_enrichment",
    "suggested_domain",
    "final_selected_domain",
    "recommended_domain",
}

DO_NOT_PATCH_COLS = {
    "commercial_fit_score",
    "final_commercial_fit_score",
    "commercial_tier",
}

AUDIT_COLS = [
    "domain_patch_status",
    "domain_patch_old_domain",
    "domain_patch_new_domain",
    "domain_patch_old_url",
    "domain_patch_new_url",
    "domain_patch_source_file",
    "domain_patch_match_key",
    "needs_reenrichment",
]

STATUS_PATCHED_CHANGED = "PATCHED_DOMAIN_CHANGED"
STATUS_PATCHED_SAME = "PATCHED_DOMAIN_SAME"
STATUS_NO_MATCH = "NO_SOURCE_MATCH"
STATUS_DUPLICATE_SKIPPED = "DUPLICATE_SOURCE_MATCH_SKIPPED"
STATUS_SOURCE_EMPTY = "SOURCE_DOMAIN_EMPTY"
STATUS_TARGET_EMPTY = "TARGET_COMPANY_EMPTY"

REPORT_COLS = [
    "queue",
    "source_file",
    "target_file",
    "sheet",
    "row",
    "company_name",
    "old_domain",
    "new_domain",
    "old_url",
    "new_url",
    "status",
    "match_key",
    "needs_reenrichment",
    "notes",
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def normalize_name(name: str) -> str:
    if not name:
        return ""
    return re.sub(r"\s+", " ", str(name).strip().lower())


def extract_domain(raw: str) -> str:
    """Return bare domain (no scheme, no path, no trailing slash)."""
    if not raw:
        return ""
    s = str(raw).strip()
    s = re.sub(r"^https?://", "", s, flags=re.IGNORECASE)
    s = re.sub(r"^www\.", "", s, flags=re.IGNORECASE)
    s = s.split("/")[0].split("?")[0].split("#")[0].strip().lower()
    return s


def make_url(domain: str) -> str:
    if not domain:
        return ""
    return f"https://{domain}"


def col_index(headers: list, name: str) -> int | None:
    nl = name.lower()
    for i, h in enumerate(headers):
        if str(h or "").strip().lower() == nl:
            return i
    return None


def is_excluded(path: Path) -> bool:
    for part in path.parts:
        if part.lower() in {f.lower() for f in EXCLUDE_FOLDERS}:
            return True
    return False


def sheet_has_company_col(ws) -> bool:
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    return any(str(h or "").strip().lower() == SOURCE_COL_COMPANY for h in headers)


# ---------------------------------------------------------------------------
# Load source lookup
# ---------------------------------------------------------------------------


def load_source_lookup(source_dir: Path, queue: str):
    """
    Returns:
        lookup_primary   : {norm_name: (domain, source_file)}
        lookup_composite : {(norm_name, norm_city, norm_province): (domain, source_file)}
        duplicates_primary : set of norm_names with conflicting domains
        stats
    """
    pattern = re.compile(rf"^{re.escape(queue)}_.*_cleaned_.*\.xlsx$", re.IGNORECASE)
    source_files = sorted(
        f for f in source_dir.iterdir() if f.is_file() and pattern.match(f.name)
    )

    lookup_primary: dict[str, tuple[str, str]] = {}
    lookup_composite: dict[tuple, tuple[str, str]] = {}
    conflict_primary: dict[str, set] = {}
    duplicates_primary: set[str] = set()

    total_rows = 0
    rows_with_domain = 0

    for sf in source_files:
        try:
            wb = load_workbook(sf, read_only=True, data_only=True)
        except Exception as e:
            print(f"  [WARN] Cannot open source file {sf.name}: {e}")
            continue

        ws = None
        for pref in SOURCE_SHEET_PREFERENCE:
            if pref in wb.sheetnames:
                ws = wb[pref]
                break
        if ws is None:
            ws = wb.worksheets[0]

        rows_iter = ws.iter_rows(values_only=True)
        try:
            raw_headers = next(rows_iter)
        except StopIteration:
            wb.close()
            continue

        headers = [str(h or "").strip() for h in raw_headers]
        ci_name = col_index(headers, SOURCE_COL_COMPANY)
        ci_url = col_index(headers, SOURCE_COL_URL)
        ci_curl = col_index(headers, SOURCE_COL_CANONICAL_URL)
        ci_cdom = col_index(headers, SOURCE_COL_CANONICAL_DOMAIN)
        ci_city = col_index(headers, SOURCE_COL_CITY)
        ci_prov = col_index(headers, SOURCE_COL_PROVINCE)

        if ci_name is None:
            print(f"  [WARN] No company_name column in {sf.name}, skipping.")
            wb.close()
            continue

        for row in rows_iter:
            total_rows += 1
            raw_name = row[ci_name] if ci_name is not None else None
            norm = normalize_name(raw_name)
            if not norm:
                continue

            raw_url = row[ci_url] if ci_url is not None else None
            raw_curl = row[ci_curl] if ci_curl is not None else None
            raw_cdom = row[ci_cdom] if ci_cdom is not None else None

            domain = (
                extract_domain(raw_url)
                or extract_domain(raw_curl)
                or extract_domain(raw_cdom)
            )

            if not domain:
                continue

            rows_with_domain += 1

            # Primary key
            if norm in conflict_primary:
                if domain not in conflict_primary[norm]:
                    conflict_primary[norm].add(domain)
                    duplicates_primary.add(norm)
            elif norm in lookup_primary:
                existing_domain = lookup_primary[norm][0]
                if existing_domain != domain:
                    conflict_primary[norm] = {existing_domain, domain}
                    duplicates_primary.add(norm)
            else:
                lookup_primary[norm] = (domain, sf.name)

            # Composite key (name + city + province)
            raw_city = row[ci_city] if ci_city is not None else None
            raw_prov = row[ci_prov] if ci_prov is not None else None
            if raw_city or raw_prov:
                comp_key = (
                    norm,
                    normalize_name(raw_city),
                    normalize_name(raw_prov),
                )
                if comp_key not in lookup_composite:
                    lookup_composite[comp_key] = (domain, sf.name)
                # composite duplicates: keep first, safe for unique matches

        wb.close()

    stats = {
        "source_files": len(source_files),
        "source_rows": total_rows,
        "rows_with_domain": rows_with_domain,
        "duplicate_keys": len(duplicates_primary),
    }
    return lookup_primary, lookup_composite, duplicates_primary, stats


# ---------------------------------------------------------------------------
# Resolve match
# ---------------------------------------------------------------------------


def resolve_match(norm_name, norm_city, norm_province, lookup_primary,
                  lookup_composite, duplicates_primary):
    """
    Returns (domain, source_file, match_key, status_hint)
    status_hint: 'ok', 'duplicate', 'no_match', 'empty'
    """
    if not norm_name:
        return None, None, None, "empty"

    # Try composite first (stronger)
    if norm_city or norm_province:
        comp_key = (norm_name, norm_city or "", norm_province or "")
        if comp_key in lookup_composite:
            domain, sf = lookup_composite[comp_key]
            key_str = f"name+city+province:{norm_name}|{norm_city}|{norm_province}"
            return domain, sf, key_str, "ok"

    # Primary with duplicate check
    if norm_name in duplicates_primary:
        return None, None, f"name:{norm_name}", "duplicate"

    if norm_name in lookup_primary:
        domain, sf = lookup_primary[norm_name]
        return domain, sf, f"name:{norm_name}", "ok"

    return None, None, f"name:{norm_name}", "no_match"


# ---------------------------------------------------------------------------
# Patch a single sheet
# ---------------------------------------------------------------------------


def patch_sheet(ws, lookup_primary, lookup_composite, duplicates_primary,
                target_file_name, sheet_name, queue, source_dir_name,
                only_changed, dry_run):
    """Patch one worksheet in-place. Returns list of report row dicts."""
    report_rows = []

    raw_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = [str(h or "").strip() for h in raw_headers]
    headers_lower = [h.lower() for h in headers]

    ci_name = col_index(headers, SOURCE_COL_COMPANY)
    if ci_name is None:
        return []

    # Identify target columns to patch
    patch_url_cols: dict[int, str] = {}
    patch_dom_cols: dict[int, str] = {}
    for i, hl in enumerate(headers_lower):
        if hl in {c.lower() for c in DO_NOT_PATCH_COLS}:
            continue
        if hl.startswith("sig_"):
            continue
        if hl in {c.lower() for c in URL_LIKE_COLS}:
            patch_url_cols[i] = headers[i]
        elif hl in {c.lower() for c in DOMAIN_LIKE_COLS}:
            patch_dom_cols[i] = headers[i]

    # Locate or create audit columns
    audit_col_indices: dict[str, int] = {}
    next_col = len(headers)
    for ac in AUDIT_COLS:
        found = col_index(headers, ac)
        if found is not None:
            audit_col_indices[ac] = found
        else:
            audit_col_indices[ac] = next_col
            next_col += 1

    # Write audit headers if this is first pass (dry_run skips writes)
    if not dry_run:
        for ac, ci in audit_col_indices.items():
            if ci >= len(headers):
                ws.cell(row=1, column=ci + 1, value=ac)

    # Optional city/province cols for composite key
    ci_city = col_index(headers, SOURCE_COL_CITY)
    ci_prov = col_index(headers, SOURCE_COL_PROVINCE)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        cells = row

        raw_name = cells[ci_name].value if ci_name < len(cells) else None
        norm_name = normalize_name(raw_name)

        norm_city = normalize_name(
            cells[ci_city].value if ci_city is not None and ci_city < len(cells) else None
        )
        norm_prov = normalize_name(
            cells[ci_prov].value if ci_prov is not None and ci_prov < len(cells) else None
        )

        domain, src_file, match_key, hint = resolve_match(
            norm_name, norm_city, norm_prov,
            lookup_primary, lookup_composite, duplicates_primary
        )

        # Collect old values
        old_domain = ""
        old_url = ""
        for ci, col_name in patch_dom_cols.items():
            v = cells[ci].value if ci < len(cells) else None
            if v:
                old_domain = extract_domain(str(v))
                break
        for ci, col_name in patch_url_cols.items():
            v = cells[ci].value if ci < len(cells) else None
            if v:
                old_url = str(v).strip()
                break

        # Determine status
        if hint == "empty":
            status = STATUS_TARGET_EMPTY
            new_domain = ""
        elif hint == "duplicate":
            status = STATUS_DUPLICATE_SKIPPED
            new_domain = ""
        elif hint == "no_match":
            status = STATUS_NO_MATCH
            new_domain = ""
        elif not domain:
            status = STATUS_SOURCE_EMPTY
            new_domain = ""
        else:
            new_domain = domain
            if old_domain == new_domain:
                status = STATUS_PATCHED_SAME
            else:
                status = STATUS_PATCHED_CHANGED

        new_url = make_url(new_domain) if new_domain else ""
        needs_reenrichment = "YES" if status == STATUS_PATCHED_CHANGED else ""

        # Apply patches
        do_write = (
            not dry_run
            and status in (STATUS_PATCHED_CHANGED, STATUS_PATCHED_SAME)
            and new_domain
        )
        if only_changed and status == STATUS_PATCHED_SAME:
            do_write = False

        if do_write:
            for ci in patch_url_cols:
                ws.cell(row=row_idx, column=ci + 1, value=new_url)
            for ci in patch_dom_cols:
                ws.cell(row=row_idx, column=ci + 1, value=new_domain)

        if not dry_run:
            ws.cell(row=row_idx, column=audit_col_indices["domain_patch_status"] + 1,
                    value=status)
            ws.cell(row=row_idx, column=audit_col_indices["domain_patch_old_domain"] + 1,
                    value=old_domain)
            ws.cell(row=row_idx, column=audit_col_indices["domain_patch_new_domain"] + 1,
                    value=new_domain)
            ws.cell(row=row_idx, column=audit_col_indices["domain_patch_old_url"] + 1,
                    value=old_url)
            ws.cell(row=row_idx, column=audit_col_indices["domain_patch_new_url"] + 1,
                    value=new_url)
            ws.cell(row=row_idx, column=audit_col_indices["domain_patch_source_file"] + 1,
                    value=src_file or "")
            ws.cell(row=row_idx, column=audit_col_indices["domain_patch_match_key"] + 1,
                    value=match_key or "")
            ws.cell(row=row_idx, column=audit_col_indices["needs_reenrichment"] + 1,
                    value=needs_reenrichment)

        report_rows.append({
            "queue": queue,
            "source_file": src_file or "",
            "target_file": target_file_name,
            "sheet": sheet_name,
            "row": row_idx,
            "company_name": raw_name or "",
            "old_domain": old_domain,
            "new_domain": new_domain,
            "old_url": old_url,
            "new_url": new_url,
            "status": status,
            "match_key": match_key or "",
            "needs_reenrichment": needs_reenrichment,
            "notes": "",
        })

    return report_rows


# ---------------------------------------------------------------------------
# Collect target files
# ---------------------------------------------------------------------------


def collect_target_files(queue_root: Path, queue: str):
    targets = []

    def add_glob(folder: Path, label: str):
        if not folder.exists():
            return
        for f in sorted(folder.rglob("*.xlsx")):
            if is_excluded(f.relative_to(queue_root)):
                continue
            targets.append((f, label))

    add_glob(queue_root / "02_lead_prioritized", "02_lead_prioritized")
    add_glob(queue_root / "03_opportunity_radar", "03_opportunity_radar")

    for name_pat in [f"{queue}_ALL_lead_prioritized.xlsx",
                     f"{queue}_ALL_opportunity_input.xlsx"]:
        p = queue_root / name_pat
        if p.exists():
            targets.append((p, "root"))

    return targets


# ---------------------------------------------------------------------------
# Decide which sheets to patch
# ---------------------------------------------------------------------------


def sheets_to_patch(wb, patch_all_sheets: bool):
    preferred = {"opportunity input", "leads", "lead scores"}
    result = []
    for name in wb.sheetnames:
        ws = wb[name]
        if not sheet_has_company_col(ws):
            continue
        if patch_all_sheets:
            result.append(name)
        elif name.lower() in preferred:
            result.append(name)
        elif not result:
            # first sheet fallback
            result.append(name)
    return result


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def parse_args():
    p = argparse.ArgumentParser(
        description="Patch corrected domains into downstream Excel files."
    )
    p.add_argument("--target-root", required=True,
                   help="Root folder, e.g. C:\\Users\\<user>\\Nextcloud\\Myngle")
    p.add_argument("--queue", required=True,
                   help="Queue name, e.g. Italy200")
    p.add_argument("--dry-run", action="store_true",
                   help="Scan and report only, do not write files.")
    p.add_argument("--apply", action="store_true",
                   help="Write patched output files.")
    p.add_argument("--in-place", action="store_true", default=False,
                   help="Overwrite original files (requires --apply).")
    p.add_argument("--overwrite", action="store_true", default=False,
                   help="Overwrite existing output files in the output folder.")
    p.add_argument("--only-changed", action="store_true", default=False,
                   help="Only write cells where the domain actually changed.")
    p.add_argument("--patch-all-sheets", action="store_true", default=False,
                   help="Patch every sheet containing company_name.")
    return p.parse_args()


def main():
    args = parse_args()

    # Safety: default to dry-run
    if not args.dry_run and not args.apply:
        print("[INFO] Neither --dry-run nor --apply specified. Defaulting to --dry-run.")
        args.dry_run = True

    if args.dry_run and args.apply:
        print("[ERROR] Cannot use both --dry-run and --apply at the same time.")
        sys.exit(1)

    target_root = Path(args.target_root)
    queue = args.queue
    queue_root = target_root / queue

    if not target_root.exists():
        print(f"[WARN] target-root does not exist: {target_root}")
    if not queue_root.exists():
        print(f"[WARN] Queue folder does not exist: {queue_root}")

    # --- Load source ---
    source_dir = queue_root / "01_cleaned_domains"
    if not source_dir.exists():
        print(f"[ERROR] Source folder not found: {source_dir}")
        sys.exit(1)

    print(f"\nLoading source files from: {source_dir}")
    lookup_primary, lookup_composite, duplicates_primary, src_stats = load_source_lookup(
        source_dir, queue
    )

    print(f"  Source files found   : {src_stats['source_files']}")
    print(f"  Source rows loaded   : {src_stats['source_rows']}")
    print(f"  Rows with domain     : {src_stats['rows_with_domain']}")
    print(f"  Duplicate keys       : {src_stats['duplicate_keys']}")

    # --- Collect targets ---
    target_files = collect_target_files(queue_root, queue)
    print(f"\nTarget files found: {len(target_files)}")
    for tf, label in target_files:
        print(f"  [{label}] {tf.name}")

    # --- Output folder ---
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_root = queue_root / f"_domain_patched_{timestamp}" if not args.in_place else None

    if args.apply and out_root and not out_root.exists():
        out_root.mkdir(parents=True, exist_ok=True)

    report_path = None
    if args.apply and out_root:
        report_path = out_root / f"domain_patch_report_{timestamp}_{queue}.csv"

    # --- Process targets ---
    all_report_rows = []
    counters = {
        "files_processed": 0,
        "rows_scanned": 0,
        "rows_changed": 0,
        "rows_same": 0,
        "rows_no_match": 0,
        "rows_skipped": 0,
    }

    mode_label = "DRY-RUN" if args.dry_run else "APPLY"
    print(f"\n[{mode_label}] Processing target files...\n")

    for tf, label in target_files:
        try:
            wb = load_workbook(tf, data_only=True)
        except Exception as e:
            print(f"  [WARN] Cannot open {tf.name}: {e}")
            all_report_rows.append({
                "queue": queue,
                "source_file": "",
                "target_file": str(tf),
                "sheet": "",
                "row": "",
                "company_name": "",
                "old_domain": "",
                "new_domain": "",
                "old_url": "",
                "new_url": "",
                "status": "SKIPPED_INVALID_WORKBOOK",
                "match_key": "",
                "needs_reenrichment": "",
                "notes": str(e),
            })
            continue

        sheets = sheets_to_patch(wb, args.patch_all_sheets)
        if not sheets:
            wb.close()
            continue

        counters["files_processed"] += 1

        for sheet_name in sheets:
            ws = wb[sheet_name]
            rows = patch_sheet(
                ws, lookup_primary, lookup_composite, duplicates_primary,
                str(tf), sheet_name, queue, source_dir.name,
                args.only_changed, args.dry_run
            )
            all_report_rows.extend(rows)
            counters["rows_scanned"] += len(rows)
            for r in rows:
                s = r["status"]
                if s == STATUS_PATCHED_CHANGED:
                    counters["rows_changed"] += 1
                elif s == STATUS_PATCHED_SAME:
                    counters["rows_same"] += 1
                elif s == STATUS_NO_MATCH:
                    counters["rows_no_match"] += 1
                elif s in (STATUS_DUPLICATE_SKIPPED, STATUS_SOURCE_EMPTY, STATUS_TARGET_EMPTY):
                    counters["rows_skipped"] += 1

        if args.apply:
            if args.in_place:
                dest = tf
            else:
                rel = tf.relative_to(queue_root)
                dest = out_root / label / rel.name if label != "root" else out_root / "root" / rel.name
                dest.parent.mkdir(parents=True, exist_ok=True)

            if dest.exists() and not args.overwrite and not args.in_place:
                print(f"  [SKIP] Output exists (use --overwrite): {dest.name}")
            else:
                try:
                    wb.save(dest)
                    print(f"  [SAVED] {dest}")
                except Exception as e:
                    print(f"  [ERROR] Cannot save {dest}: {e}")

        wb.close()

    # --- Write CSV report ---
    if args.apply and report_path:
        with open(report_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=REPORT_COLS)
            writer.writeheader()
            writer.writerows(all_report_rows)
        print(f"\nReport written: {report_path}")

    # --- Summary ---
    print(f"""
=== Summary ===
Mode                : {mode_label}
Source files loaded : {src_stats['source_files']}
Source rows loaded  : {src_stats['source_rows']}
Rows with domain    : {src_stats['rows_with_domain']}
Duplicate keys      : {src_stats['duplicate_keys']}
Target files found  : {len(target_files)}
Files processed     : {counters['files_processed']}
Rows scanned        : {counters['rows_scanned']}
Rows changed        : {counters['rows_changed']}
Rows same           : {counters['rows_same']}
Rows no match       : {counters['rows_no_match']}
Rows skipped        : {counters['rows_skipped']}
Output folder       : {out_root if out_root else '(in-place)' if args.apply else 'n/a (dry-run)'}
Report              : {report_path if report_path else 'n/a'}
""")


if __name__ == "__main__":
    main()
