"""
Recalculate commercial fit scores after HQ corrections and/or competitor signal removal.

Scoring note
------------
``competitor_signal_strength_score`` and ``language_competitor_strength_score`` are
NOT part of LEAN_COEFFICIENTS and therefore do NOT directly affect
``final_commercial_fit_score``.  They appear only in the display-only
COMMERCIAL_COMPLEXITY_FIELDS grouping.  The competitor-removal mode sets them to 0 in
the scoring row copy as an audit measure and to eliminate any indirect display effects,
but the ``final_commercial_fit_score`` delta for pure competitor removal will typically
be 0 unless some future model update adds these fields to the coefficients.

Usage (CLI)
-----------
python recalculate_hq_changed_scores.py \\
    --enriched-workbook   enriched.xlsx \\
    --hq-recovery-workbook  hq_recovery.xlsx \\
    --output              recalculated.xlsx \\
    [--sheet "Opportunity Input Full"] \\
    [--recalculation-scope hq|competitor|both] \\
    [--max-recalculated-rows 10]

Recalculation scopes
--------------------
hq          – rows where HQ Recovery changed sig_foreign_hq_score  (default)
competitor  – rows with non-zero competitor signal
both        – union of hq and competitor rows
"""

from __future__ import annotations

import argparse
import io
import sys
from typing import Any

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from commercial_fit_scoring import SCORE_OUTPUT_COLS, score_company

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

SCORING_PROFILE = "italy_register_icp_only"
DEFAULT_SHEET   = "Opportunity Input Full"
SUMMARY_SHEET   = "HQ Score Recalc Summary"

SCOPE_HQ         = "hq"
SCOPE_COMPETITOR = "competitor"
SCOPE_BOTH       = "both"
VALID_SCOPES     = (SCOPE_HQ, SCOPE_COMPETITOR, SCOPE_BOTH)

# Numeric fields zeroed in the scoring row copy when competitor removal is active.
# NOTE: these are NOT in LEAN_COEFFICIENTS, so zeroing them does not
# change final_commercial_fit_score unless the model is updated.
_COMPETITOR_NUMERIC_FIELDS = (
    "competitor_signal_strength_score",
    "language_competitor_strength_score",
    "competitor_signal_strength",
    "competitor_attention_strength",
)

# Text/evidence fields used to *detect* competitor signal (not zeroed in scoring copy).
_COMPETITOR_TEXT_FIELDS = (
    # original columns
    "competitor_customer_match",
    "competitor_customer_evidence",
    "competitor_signal",
    "competitor_mentions",
    # real workbook columns
    "competitor_customer_signal",
    "competitor_provider_detected",
    "competitor_evidence_url",
    "competitive_switch_opportunity",
    "sales_action_hint",
    "competitor_attention_signal",
    "competitor_attention_provider_detected",
    "competitor_attention_type",
    "competitor_attention_evidence",
    "competitor_attention_url",
    "competitor_attention_needs_review",
    "competitor_signal_excluded_from_next_scoring",
)

HQ_AUDIT_COLS = [
    "hq_recalc_applied",
    "hq_recalc_reason",
    "hq_score_before_recalc",
    "hq_score_after_recalc",
    "commercial_fit_score_before_hq_recalc",
    "commercial_fit_score_after_hq_recalc",
    "commercial_fit_score_delta_hq_recalc",
    "final_commercial_fit_score_before_hq_recalc",
    "final_commercial_fit_score_after_hq_recalc",
    "final_commercial_fit_score_delta_hq_recalc",
    "commercial_fit_score_before_source_column",
    # Reviewed HQ evidence — only populated when hq_recalc_applied = Yes
    "hq_recalc_reviewed_country",
    "hq_recalc_reviewed_city",
    "hq_recalc_evidence_url",
    "hq_recalc_evidence_quote",
    "hq_recalc_ai_reason",
    "old_sig_foreign_hq_evidence",
    "hq_recalc_domain_mismatch_warning",
]

# Columns to clear on rows where HQ recalc was NOT applied, to prevent stale
# values from a prior recalc run passing through as if they apply to this run.
_HQ_STALE_CLEAR_COLS = (
    "hq_recalc_reason",
    "hq_score_before_recalc",
    "hq_score_after_recalc",
    "commercial_fit_score_before_hq_recalc",
    "commercial_fit_score_after_hq_recalc",
    "commercial_fit_score_delta_hq_recalc",
    "final_commercial_fit_score_before_hq_recalc",
    "final_commercial_fit_score_after_hq_recalc",
    "final_commercial_fit_score_delta_hq_recalc",
    "commercial_fit_score_before_source_column",
    "hq_recalc_reviewed_country",
    "hq_recalc_reviewed_city",
    "hq_recalc_evidence_url",
    "hq_recalc_evidence_quote",
    "hq_recalc_ai_reason",
    "old_sig_foreign_hq_evidence",
    "hq_recalc_domain_mismatch_warning",
)

COMPETITOR_AUDIT_COLS = [
    "competitor_recalc_applied",
    "competitor_recalc_reason",
    "competitor_signal_before_recalc",
    "competitor_signal_after_recalc",
    "language_competitor_signal_before_recalc",
    "language_competitor_signal_after_recalc",
    "competitor_signal_neutralized_for_scoring",
    "competitor_signal_used_for_scoring",
    "competitor_signal_suppressed",
]

GENERAL_AUDIT_COLS = [
    "recalc_scope_applied",
    "cfs_before_recalc",
    "cfs_after_recalc",
    "cfs_delta_recalc",
    "cfs_source_col_used",
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_float(v: Any) -> float | None:
    try:
        return float(v)
    except Exception:
        return None


def _first_numeric(row: dict, cols: tuple) -> float:
    """Return the first non-None numeric value from cols, or 0.0."""
    for col in cols:
        v = _safe_float(row.get(col))
        if v is not None:
            return v
    return 0.0


def _get_existing_commercial_fit_score(row: dict) -> tuple[float, str]:
    """Return (score, column_name) from the first populated CFS column found."""
    for col in (
        "final_commercial_fit_score",
        "commercial_fit_score",
        "commercial_fit_score_final",
        "Commercial Fit Score",
        "cfs",
        "score",
    ):
        if col in row:
            v = _safe_float(row.get(col))
            if v is not None:
                return v, col
    return 0.0, ""


def _norm_key(s: Any) -> str:
    return str(s or "").strip().lower()


def _has_competitor_signal(row: dict) -> bool:
    """True if the row carries any non-zero / non-empty competitor signal."""
    for col in _COMPETITOR_NUMERIC_FIELDS:
        v = _safe_float(row.get(col))
        if v is not None and v > 0:
            return True
    for col in _COMPETITOR_TEXT_FIELDS:
        v = row.get(col)
        if v is not None and str(v).strip():
            return True
    return False


# ---------------------------------------------------------------------------
# App-text refresh helpers
# ---------------------------------------------------------------------------

import re as _re

APP_TEXT_COLS = [
    "commercial_fit_score_app",
    "commercial_tier_app",
    "what_is_hot_app",
    "what_is_not_app",
    "why_relevant_app",
    "caller_angle_app",
    "call_starter_app",
    "caution_app",
    "evidence_summary_app",
    "key_source_links_app",
    "advanced_notes_app",
    "foreign_hq_signal_used_in_app",
    "foreign_hq_country_app",
    "foreign_hq_city_app",
    "competitor_signal_used_in_app",
    "app_text_refresh_applied",
    "app_text_refresh_reason",
    "app_text_hq_note_added",
    "app_text_competitor_note_added",
    "app_text_conflicting_text_removed",
]

_COMPETITOR_TERMS_RE = _re.compile(
    r"\b(?:"
    r"competitor|competing|rivalry|rival|alternative\s+brand|competitive\s+threat|"
    r"competitive\s+switch(?:\s+opportunity)?|"
    r"competitive\s+pressure|"
    r"competitor\s+signal|competitor\s+provider|"
    r"already\s+buys?\s+(?:the\s+)?language\s+training(?:\s+category)?|"
    r"language\s+training\s+category\s+buyer|"
    r"treat\s+as\s+a\s+competitive\s+switch\s+opportunity"
    r")\b",
    _re.IGNORECASE,
)

# Known language-training competitor / provider names — used when competitor is suppressed
_COMPETITOR_PROVIDERS_RE = _re.compile(
    r"\b(?:Berlitz|Speexx|Learnlight|Babbel|Busuu|Rosetta\s+Stone|"
    r"EF\s+(?:Education|English|Corporate|Language)|Wall\s+Street\s+English|"
    r"goFLUENT|Preply|Cambly|Duolingo|Pearson(?:\s+English)?|"
    r"Linguarama|Hult\s+EF|Global\s+LT)\b",
    _re.IGNORECASE,
)

_FOREIGN_HQ_CLAIM_RE = _re.compile(
    r"\b(?:"
    r"foreign\s+hq|"
    r"foreign\s+headquarters(?:\s+or\s+group\s+ownership)?|"
    r"foreign\s+headquarters[/\\]parent\s+company|"
    r"foreign\s+parent\s+company|"
    r"foreign\s+group\s+ownership|"
    r"international\s+hq|global\s+hq|"
    r"headquartered\s+abroad|"
    r"hq\s+(?:in|confirmed|detected)|"
    r"confirmed\s+(?:foreign|international)\s+hq|"
    r"headquarters?\s+(?:in|based\s+in|located\s+in)|"
    r"parent\s+company\s+(?:is\s+)?(?:located|based|headquartered)|"
    r"group\s+structure.*foreign|foreign.*group\s+structure"
    r")\b",
    _re.IGNORECASE,
)

# Sentences/items that claim the company has NO foreign HQ or a domestic HQ —
# must be removed when HQ Recovery confirms a foreign HQ after recalculation.
_CONTRADICTORY_DOMESTIC_HQ_RE = _re.compile(
    r"\b(?:"
    r"no\s+(?:evidence\s+of\s+)?foreign\s+(?:parent|hq|headquarters?)"
    r"|not\s+outside\s+(?:the\s+)?input\s+country"
    r"|italian\s+(?:company\s+)?(?:headquarters?|hq)"
    r"|italian[- ]headquartered"
    r"|domestic\s+headquarters?"
    r"|italy[- ]based\s+(?:parent\s+(?:company|group)|company|group)"
    r")\b"
    r"|headquartered?\s+in\s+\w[\w\s]*,?\s*Italy\b"
    r"|(?:parent|group)\s+(?:is\s+)?(?:based|located|headquartered)\s+in\s+\w[\w\s]*,?\s*Italy\b",
    _re.IGNORECASE,
)


def _v(row: dict, *keys: str) -> str:
    for k in keys:
        val = row.get(k)
        if val is not None and str(val).strip():
            return str(val).strip()
    return ""


def _strip_sentences(text: str, pattern: _re.Pattern) -> tuple[str, bool]:
    """Remove sentences matching pattern. Returns (cleaned_text, was_changed)."""
    if not text:
        return text, False
    sentences = [s.strip() for s in text.split(".") if s.strip()]
    clean = [s for s in sentences if not pattern.search(s)]
    changed = len(clean) < len(sentences)
    return ". ".join(clean) + ("." if clean else ""), changed


def _clean_pipe_or_sentence(text: str, *patterns: _re.Pattern) -> tuple[str, bool]:
    """Remove items from pipe-separated OR sentence-separated text matching any pattern.

    Handles both "tag1 | tag2 | tag3" and "Sentence one. Sentence two." formats.
    Returns (cleaned_text, was_changed).
    """
    if not text:
        return text, False

    def _matches(s: str) -> bool:
        return any(p.search(s) for p in patterns)

    # Pipe-separated (what_is_hot_app / what_is_not_app style)
    if "|" in text:
        parts = [p.strip() for p in text.split("|")]
        clean = [p for p in parts if p and not _matches(p)]
        changed = len(clean) < len([p for p in parts if p])
        return " | ".join(clean), changed

    # Sentence-separated
    sentences = [s.strip() for s in text.split(".") if s.strip()]
    clean = [s for s in sentences if not _matches(s)]
    changed = len(clean) < len(sentences)
    return ". ".join(clean) + ("." if clean else ""), changed


def _remove_suppressed_competitor_claims(text: str) -> tuple[str, bool]:
    """Remove items/sentences that mention competitor terms or known provider names.

    Applied to all app-facing commercial fields when competitor is suppressed.
    Returns (cleaned_text, was_changed).
    """
    return _clean_pipe_or_sentence(text, _COMPETITOR_TERMS_RE, _COMPETITOR_PROVIDERS_RE)


def _remove_unconfirmed_hq_claims(text: str) -> tuple[str, bool]:
    """Remove items/sentences that assert a foreign HQ when HQ is not confirmed.

    Preserves neutral group-structure or parent-company mentions that do not
    imply a confirmed foreign headquarters.
    Returns (cleaned_text, was_changed).
    """
    return _clean_pipe_or_sentence(text, _FOREIGN_HQ_CLAIM_RE)


def _remove_contradictory_domestic_hq_claims(text: str) -> tuple[str, bool]:
    """Remove items/sentences claiming domestic/no-foreign HQ when HQ IS confirmed.

    Applied to app-text and evidence fields when hq_upgraded=True, to neutralise
    stale evidence like "Italian headquarters" or "no foreign parent" that would
    contradict the newly confirmed foreign HQ signal.
    Returns (cleaned_text, was_changed).
    """
    return _clean_pipe_or_sentence(text, _CONTRADICTORY_DOMESTIC_HQ_RE)


def _collect_source_links(
    enr_row: dict,
    hqr_row: dict,
    state: dict,
    max_links: int = 8,
) -> str:
    """Collect deduplicated evidence URLs from all available columns."""
    links: list[str] = []

    def _add(url: str) -> None:
        url = url.strip()
        if url and url not in links and len(links) < max_links:
            links.append(url)

    # HQ evidence URL first when HQ is confirmed
    if state["foreign_hq_active"] and state["hq_evidence_url"]:
        _add(state["hq_evidence_url"])

    # Google snippets (up to 10) — skip competitor-tainted ones when suppressed
    for i in range(1, 11):
        url = _v(enr_row, f"google_snippet_{i}_url")
        if not url:
            continue
        if state["competitor_suppressed"]:
            snippet_text = _v(enr_row, f"google_snippet_{i}")
            if (_COMPETITOR_TERMS_RE.search(snippet_text)
                    or _COMPETITOR_PROVIDERS_RE.search(snippet_text)):
                continue
        _add(url)

    # Additional URL columns from enriched workbook
    for col in (
        "sig_foreign_hq_evidence_url",
        "source_url", "homepage_url",
        "competitor_evidence_url",
        "competitor_attention_url",
        "evidence_url",
    ):
        url = _v(enr_row, col)
        if not url:
            continue
        # Skip competitor evidence URLs when suppressed
        if state["competitor_suppressed"] and col in (
            "competitor_evidence_url", "competitor_attention_url"
        ):
            continue
        _add(url)

    # HQR evidence URL as fallback
    hqr_ev = _hqr_evidence_url(hqr_row)
    if hqr_ev:
        _add(hqr_ev)

    return " | ".join(links)


def _hqr_country(hqr_row: dict) -> str:
    return _v(hqr_row,
              "hq_detected_country", "sig_foreign_hq_country",
              "detected_country", "hq_country", "country_reviewed", "reviewed_country")


def _hqr_city(hqr_row: dict) -> str:
    return _v(hqr_row,
              "hq_detected_city", "sig_foreign_hq_city",
              "detected_city", "hq_city", "city_reviewed", "reviewed_city")


def _hqr_evidence_url(hqr_row: dict) -> str:
    return _v(hqr_row,
              "hq_evidence_url", "sig_foreign_hq_evidence_url",
              "evidence_url", "hq_url")


def _score_input(score_out: dict, enr_row: dict, out_key: str, enr_key: str) -> float:
    v = _safe_float(score_out.get(out_key))
    if v is not None:
        return v
    v = _safe_float(enr_row.get(enr_key))
    return v if v is not None else 0.0


def _build_final_signal_state(
    enr_row: dict,
    hqr_row: dict,
    score_out: dict,
    hq_eligible: bool,
    competitor_eligible: bool,
    hq_reviewed_val: float | None,
    hq_original_val: float | None,
) -> dict:
    """Return the definitive signal state dict used for all app-text construction."""
    hq_new = (hq_reviewed_val if hq_eligible and hq_reviewed_val is not None
              else _safe_float(enr_row.get("sig_foreign_hq_score")) or 0.0)
    hq_old = hq_original_val if hq_original_val is not None else 0.0

    # HQ location: prefer HQR columns if eligible (reviewed data), else fall back to enr
    if hq_eligible:
        country = _hqr_country(hqr_row)
        city    = _hqr_city(hqr_row)
    else:
        country = _v(enr_row, "sig_foreign_hq_country", "foreign_hq_country_app")
        city    = _v(enr_row, "sig_foreign_hq_city",    "foreign_hq_city_app")
    ev_url = _hqr_evidence_url(hqr_row) or _v(enr_row, "sig_foreign_hq_evidence_url")

    cfs  = _safe_float(score_out.get("final_commercial_fit_score")) or 0.0
    tier = score_out.get("commercial_tier") or _v(enr_row, "commercial_tier")

    return {
        "hq_score":           hq_new,
        "hq_score_old":       hq_old,
        "foreign_hq_active":  hq_new >= 3,
        "hq_upgraded":        hq_old < 3 and hq_new >= 3,
        "hq_downgraded":      hq_old >= 3 and hq_new < 3,
        "hq_country":         country,
        "hq_city":            city,
        "hq_evidence_url":    ev_url,
        "competitor_suppressed": competitor_eligible,
        "cfs":                cfs,
        "tier":               str(tier).strip() if tier else "",
        # Scoring signal strengths for what_is_hot / what_is_not construction
        "sig_explicit_lnd":      _score_input(score_out, enr_row, "score_input_explicit_lnd",      "sig_explicit_lnd_score"),
        "sig_intl_footprint":    _score_input(score_out, enr_row, "score_input_intl_footprint",    "sig_intl_footprint_score"),
        "sig_employer_branding": _score_input(score_out, enr_row, "score_input_employer_branding", "sig_employer_branding_score"),
        "sig_lnd_onboarding":    _score_input(score_out, enr_row, "score_input_lnd_onboarding",    "sig_lnd_onboarding_score"),
        "ti_onboarding":         _score_input(score_out, enr_row, "score_input_ti_onboarding",     "ti_onboarding_score"),
        "sig_rapid_growth":      _score_input(score_out, enr_row, "score_input_rapid_growth",      "sig_rapid_growth_score"),
        "domain_mismatch_warning": (
            str(enr_row.get("possible_domain_mismatch") or "").strip().lower() in ("yes", "true", "1")
            or str(hqr_row.get("possible_domain_mismatch") or "").strip().lower() in ("yes", "true", "1")
        ),
    }


def _build_lovable_app_fields(
    enr_row: dict,
    hqr_row: dict,
    score_out: dict,
    state: dict,
) -> tuple[dict, bool]:
    """Build all Lovable app-facing fields from the final signal state.

    Returns (app_fields_dict, conflict_removed_flag).
    """
    out: dict = {}
    conflict_removed = False

    # ── commercial score / tier ───────────────────────────────────────────────
    out["commercial_fit_score_app"] = f"{state['cfs']:.4f}"
    out["commercial_tier_app"] = state["tier"]

    # ── what_is_hot_app — built from active signals ───────────────────────────
    hot: list[str] = []
    if state["foreign_hq_active"]:
        tag = (f"Foreign HQ: {state['hq_country']}"
               if state["hq_country"] else "Foreign HQ confirmed")
        hot.append(tag)
    for sig_key, label in [
        ("sig_explicit_lnd",      "Explicit L&D signal"),
        ("sig_intl_footprint",    "International footprint"),
        ("sig_employer_branding", "Employer branding"),
        ("sig_lnd_onboarding",    "L&D / onboarding signal"),
        ("ti_onboarding",         "Onboarding signal"),
    ]:
        if state.get(sig_key, 0) > 0:
            hot.append(label)
    if state["tier"].upper() in ("A", "B"):
        hot.append(f"Commercial tier: {state['tier'].upper()}")
    out["what_is_hot_app"] = " | ".join(hot)

    # ── what_is_not_app — built from missing/weak/suppressed signals ──────────
    not_tags: list[str] = []
    if not state["foreign_hq_active"]:
        not_tags.append("No confirmed foreign HQ after review")
    lnd_present = (state.get("sig_explicit_lnd", 0) > 0
                   or state.get("sig_lnd_onboarding", 0) > 0
                   or state.get("ti_onboarding", 0) > 0)
    if not lnd_present:
        not_tags.append("No clear L&D/onboarding signal")
    if state.get("sig_intl_footprint", 0) == 0:
        not_tags.append("No international footprint detected")
    if state["competitor_suppressed"]:
        not_tags.append("Competitor signal suppressed (low reliability)")
    out["what_is_not_app"] = " | ".join(not_tags)

    # ── why_relevant_app ──────────────────────────────────────────────────────
    base = _v(enr_row, "icp_why_relevant", "why_relevant_app")
    if state["competitor_suppressed"]:
        base, chg = _remove_suppressed_competitor_claims(base)
        if chg:
            conflict_removed = True
    if state["hq_upgraded"]:
        # Strip old contradictory domestic/no-foreign HQ claims first
        base, chg = _remove_contradictory_domestic_hq_claims(base)
        if chg:
            conflict_removed = True
        hq_loc = state["hq_country"]
        if state["hq_city"] and hq_loc:
            hq_loc = f"{state['hq_city']}, {hq_loc}"
        elif state["hq_city"]:
            hq_loc = state["hq_city"]
        prefix = f"Confirmed foreign HQ ({hq_loc}). " if hq_loc else "Confirmed foreign HQ. "
        base = prefix + base if base else prefix.strip()
    elif not state["foreign_hq_active"]:
        base, chg = _remove_unconfirmed_hq_claims(base)
        if chg:
            conflict_removed = True
    out["why_relevant_app"] = base

    # ── caller_angle_app ──────────────────────────────────────────────────────
    base = _v(enr_row, "caller_angle", "caller_angle_app")
    if state["competitor_suppressed"]:
        base, chg = _remove_suppressed_competitor_claims(base)
        if chg:
            conflict_removed = True
    if not state["foreign_hq_active"]:
        base, chg = _remove_unconfirmed_hq_claims(base)
        if chg:
            conflict_removed = True
    if state["hq_upgraded"] and state["hq_country"]:
        if state["hq_country"].lower() not in base.lower():
            base = (base.rstrip(". ") + f". HQ confirmed in {state['hq_country']}.") if base else f"HQ confirmed in {state['hq_country']}."
    out["caller_angle_app"] = base

    # ── call_starter_app ──────────────────────────────────────────────────────
    base = _v(enr_row, "icp_buying_signals", "call_starter_app")
    if state["competitor_suppressed"]:
        base, chg = _remove_suppressed_competitor_claims(base)
        if chg:
            conflict_removed = True
    if not state["foreign_hq_active"]:
        base, chg = _remove_unconfirmed_hq_claims(base)
        if chg:
            conflict_removed = True
    out["call_starter_app"] = base

    # ── caution_app ───────────────────────────────────────────────────────────
    base = _v(enr_row, "scoring_notes", "caution_app")
    if state["competitor_suppressed"]:
        base, chg = _remove_suppressed_competitor_claims(base)
        if chg:
            conflict_removed = True
    if state["hq_upgraded"]:
        base, chg = _remove_contradictory_domestic_hq_claims(base)
        if chg:
            conflict_removed = True
    if state.get("sig_rapid_growth", 0) > 0:
        growth_note = "Growth signal present — may indicate cost sensitivity."
        if growth_note not in base:
            base = (base + " " + growth_note).strip() if base else growth_note
    if state.get("hq_upgraded") and state.get("domain_mismatch_warning"):
        dm_note = "HQ was reviewed, but domain match was flagged. Verify before outreach."
        if dm_note not in base:
            base = (base + " " + dm_note).strip() if base else dm_note
    out["caution_app"] = base

    # ── evidence_summary_app ──────────────────────────────────────────────────
    base = _v(enr_row, "icp_evidence", "raw_evidence_summary", "evidence_summary_app")
    if state["competitor_suppressed"]:
        base, chg = _remove_suppressed_competitor_claims(base)
        if chg:
            conflict_removed = True
    if state["hq_upgraded"]:
        # Remove old "Italian headquarters" / "no foreign parent" claims before
        # appending the confirmed foreign HQ note below.
        base, chg = _remove_contradictory_domestic_hq_claims(base)
        if chg:
            conflict_removed = True
    elif not state["foreign_hq_active"]:
        base, chg = _remove_unconfirmed_hq_claims(base)
        if chg:
            conflict_removed = True
    if state["foreign_hq_active"]:
        hq_note = "Foreign HQ signal confirmed"
        if state["hq_country"]:
            hq_note += f": {state['hq_country']}"
        if state["hq_city"]:
            hq_note += f" ({state['hq_city']})"
        hq_note += "."
        if hq_note.rstrip(".").lower() not in base.lower():
            base = (base + " " + hq_note).strip() if base else hq_note
    out["evidence_summary_app"] = base

    # ── key_source_links_app ──────────────────────────────────────────────────
    out["key_source_links_app"] = _collect_source_links(enr_row, hqr_row, state)

    # ── advanced_notes_app ────────────────────────────────────────────────────
    notes: list[str] = []
    if state["hq_score"] != state["hq_score_old"]:
        country_str = state["hq_country"] or "unknown"
        notes.append(
            f"HQ score updated: {state['hq_score_old']:.0f} → {state['hq_score']:.0f}"
            f" (country: {country_str})"
        )
    if state["competitor_suppressed"]:
        notes.append(
            "Competitor signal suppressed in final app export due to low reliability."
        )
    out["advanced_notes_app"] = " | ".join(notes)

    # ── meta flags ────────────────────────────────────────────────────────────
    out["foreign_hq_signal_used_in_app"] = "Yes" if state["foreign_hq_active"] else "No"
    out["foreign_hq_country_app"] = state["hq_country"] if state["foreign_hq_active"] else ""
    out["foreign_hq_city_app"]    = state["hq_city"]    if state["foreign_hq_active"] else ""
    out["competitor_signal_used_in_app"] = (
        "No" if state["competitor_suppressed"]
        else ("Yes" if _has_competitor_signal(enr_row) else "No")
    )

    return out, conflict_removed


def _refresh_app_text(
    row_out: dict,
    enr_row: dict,
    hqr_row: dict,
    suppress_competitor: bool,
    score_out: dict,
    hq_eligible: bool,
    hq_reviewed_val: float | None,
    hq_original_val: float | None,
) -> dict:
    """Build all Lovable app-facing fields and write them to row_out.

    Returns audit flag dict: {hq_note_added, comp_note_added, conflict_removed}.
    """
    state = _build_final_signal_state(
        enr_row, hqr_row, score_out,
        hq_eligible=hq_eligible,
        competitor_eligible=suppress_competitor,
        hq_reviewed_val=hq_reviewed_val,
        hq_original_val=hq_original_val,
    )
    app_fields, conflict_removed = _build_lovable_app_fields(enr_row, hqr_row, score_out, state)
    row_out.update(app_fields)

    return {
        "hq_note_added":    state["hq_upgraded"],
        "comp_note_added":  suppress_competitor,
        "conflict_removed": conflict_removed,
    }


# ---------------------------------------------------------------------------
# Excel I/O helpers
# ---------------------------------------------------------------------------

def _wb_to_rows(wb, sheet_name: str) -> tuple[list[str], list[dict]]:
    target = sheet_name if sheet_name in wb.sheetnames else wb.sheetnames[0]
    ws = wb[target]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(c or "").strip() for c in rows[0]]
    data = [
        {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        for row in rows[1:]
    ]
    return headers, data


def _build_match_index(
    rows: list[dict],
    has_domain: bool,
    has_company: bool,
    has_country: bool,
) -> tuple[dict, str]:
    if has_domain and has_company and has_country:
        strategy = "domain+company_name+input_country"
        idx: dict[str, int] = {}
        for i, r in enumerate(rows):
            k = (
                _norm_key(r.get("domain"))
                + "|" + _norm_key(r.get("company_name") or r.get("name"))
                + "|" + _norm_key(r.get("input_country_used") or r.get("country"))
            )
            idx.setdefault(k, i)
        if len(idx) >= len(rows) * 0.9:
            return idx, strategy

    if has_domain and has_company:
        strategy = "domain+company_name"
        idx = {}
        for i, r in enumerate(rows):
            k = (
                _norm_key(r.get("domain"))
                + "|" + _norm_key(r.get("company_name") or r.get("name"))
            )
            idx.setdefault(k, i)
        return idx, strategy

    return {}, "row_order_fallback"


def _match_key_for_row(r: dict, strategy: str) -> str:
    if "input_country" in strategy:
        return (
            _norm_key(r.get("domain"))
            + "|" + _norm_key(r.get("company_name") or r.get("name"))
            + "|" + _norm_key(r.get("input_country_used") or r.get("country"))
        )
    if "company_name" in strategy:
        return (
            _norm_key(r.get("domain"))
            + "|" + _norm_key(r.get("company_name") or r.get("name"))
        )
    return ""


def _build_output_wb(
    out_headers: list[str],
    out_rows: list[dict],
    sheet_name: str,
    summary: dict,
    deltas: list[tuple],
    fast_output: bool = True,
) -> Workbook:
    _hdr_fill = PatternFill("solid", fgColor="D9EAF7")
    _hdr_font = Font(bold=True)

    wb_out = Workbook()
    ws_data = wb_out.active
    ws_data.title = sheet_name
    ws_data.append(out_headers)
    for r in out_rows:
        ws_data.append([r.get(h) for h in out_headers])
    ws_data.freeze_panes = "A2"
    ws_data.auto_filter.ref = ws_data.dimensions
    ws_data.row_dimensions[1].height = 22
    for cell in ws_data[1]:
        cell.font = _hdr_font
        cell.fill = _hdr_fill

    if not fast_output and ws_data.max_column <= 250:
        _MAX_WIDTH, _SAMPLE = 50, 25
        for col_idx in range(1, ws_data.max_column + 1):
            header = ws_data.cell(row=1, column=col_idx).value
            max_len = len(str(header or ""))
            for row_idx in range(2, min(ws_data.max_row, _SAMPLE + 1) + 1):
                v = ws_data.cell(row=row_idx, column=col_idx).value
                if v is not None:
                    max_len = max(max_len, min(len(str(v)), _MAX_WIDTH))
            ws_data.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, _MAX_WIDTH)

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_sum = wb_out.create_sheet(SUMMARY_SHEET)
    bold = Font(bold=True)

    def _add(label: str, value: Any) -> None:
        ws_sum.append([label, value])

    def _section(title: str) -> None:
        ws_sum.append([])
        ws_sum.append([title])
        ws_sum.cell(ws_sum.max_row, 1).font = bold

    ws_sum.append(["HQ Score Recalculation Summary"])
    ws_sum["A1"].font = Font(bold=True, size=12)
    _section("Run parameters")
    _add("Sheet",                   sheet_name)
    _add("Recalculation scope",     summary.get("scope", "hq"))
    _add("Scoring profile",         SCORING_PROFILE)
    _add("Matching strategy",       summary.get("strategy", ""))
    _add("Test mode active",        "Yes" if summary.get("test_mode_active") else "No")
    _add("Max recalculated rows",   summary.get("max_recalculated_rows", 0) or "unlimited")

    _section("Row counts")
    _add("Total enriched rows",             summary.get("n_enr", 0))
    _add("Total HQ Recovery rows",          summary.get("n_hqr", 0))
    _add("Matched rows",                    summary.get("n_matched", 0))
    _add("Recalculated rows",               summary.get("n_recalculated", 0))
    _add("Skipped by row limit",            summary.get("skipped_by_recalc_limit", 0))
    _add("Unchanged rows",                  summary.get("n_enr", 0) - summary.get("n_recalculated", 0))

    if summary.get("scope") in (SCOPE_HQ, SCOPE_BOTH):
        _section("HQ changes")
        _add("Eligible HQ-changed rows",          summary.get("n_hq_eligible", 0))
        _add("HQ rows recalculated",              summary.get("n_hq_recalculated", 0))
        _add("HQ rows skipped by test limit",     summary.get("n_hq_skipped_limit", 0))
        _add("HQ upgrades  0/blank → 3",          summary.get("n_upgrades", 0))
        _add("HQ downgrades 3 → 0",               summary.get("n_downgrades", 0))
        _add("Other HQ numeric changes",          summary.get("n_other", 0))

    if summary.get("scope") in (SCOPE_COMPETITOR, SCOPE_BOTH):
        _section("Competitor signal")
        _add("Competitor-signal rows detected",      summary.get("n_competitor_detected", 0))
        _add("Competitor rows recalculated",         summary.get("n_competitor_recalculated", 0))
        _add("Competitor rows skipped by test limit", summary.get("n_competitor_skipped_limit", 0))
        _add("Avg competitor signal before (non-zero rows)",
             round(summary.get("avg_competitor_before", 0.0), 4))
        _add("Avg competitor signal after  (should be 0)",
             round(summary.get("avg_competitor_after", 0.0), 4))
        _add("Note: competitor fields are NOT in LEAN_COEFFICIENTS",
             "final_commercial_fit_score delta will be 0 unless model changes")

    if summary.get("app_text_refreshed"):
        _section("Lovable app text refresh")
        _add("App text rows refreshed",     summary.get("n_app_text_refreshed", 0))
        _add("HQ notes added",              summary.get("n_hq_notes", 0))
        _add("Competitor notes added",      summary.get("n_comp_notes", 0))
        _add("Conflicting text removed",    summary.get("n_conflict_removed", 0))

    _section("Score columns — which is final")
    _add("Authoritative final score column",  "final_commercial_fit_score")
    _add("Legacy/original score column",      "commercial_fit_score  (original enriched value, not updated by recalc)")
    _add("Note", "Downstream tools that read commercial_fit_score will see the pre-recalc score.")

    _section("Validation checks (should all be 0)")
    _add("Rows with hq_score_after_recalc=3 but hq_recalc_applied=No (inconsistency)",
         summary.get("n_val_inconsistent_hq", 0))
    _add("Rows with hq_recalc_applied=Yes but missing commercial score before/after",
         summary.get("n_val_recalc_missing_scores", 0))
    _add("Rows with hq_recalc_applied=Yes but final_commercial_fit_score blank",
         summary.get("n_val_blank_final_score", 0))
    _add("Rows with recalculated HQ=3 but sig_foreign_hq_evidence still contradicts (should be 0)",
         summary.get("n_val_contradictory_evidence", 0))
    _add("HQ recalculated rows with domain mismatch warning (verify before outreach)",
         summary.get("n_val_domain_mismatch_warning", 0))
    _add("HQ recalculated rows where app text still contains contradictory HQ wording (should be 0)",
         summary.get("n_val_app_text_still_contradictory", 0))
    _add("HQ recalculated rows with evidence URL populated",
         summary.get("n_val_has_evidence_url", 0))
    _add("HQ recalculated rows with evidence quote populated",
         summary.get("n_val_has_evidence_quote", 0))

    if deltas:
        _section("Score delta statistics")
        all_d = [x[4] for x in deltas]
        pos_d = [d for d in all_d if d > 0]
        neg_d = [d for d in all_d if d < 0]
        _add("Score increases",  len(pos_d))
        _add("Score decreases",  len(neg_d))
        _add("Max positive delta (biggest increase)", max(pos_d) if pos_d else "n/a")
        _add("Min positive delta (smallest increase)", min(pos_d) if pos_d else "n/a")
        _add("Max negative delta (biggest decrease)", min(neg_d) if neg_d else "n/a — no score decreases")
        _add("Mean delta",         round(sum(all_d) / len(all_d), 4))

        top_pos = sorted([d for d in deltas if d[4] > 0], key=lambda x: -x[4])[:20]
        top_neg = sorted([d for d in deltas if d[4] < 0], key=lambda x:  x[4])[:20]
        for title, subset in [
            ("Top 20 positive score deltas (biggest increase)", top_pos),
            ("Top 20 negative score deltas (biggest decrease)", top_neg),
        ]:
            if subset:
                ws_sum.append([])
                ws_sum.append([title])
                ws_sum.cell(ws_sum.max_row, 1).font = bold
                ws_sum.append(["company", "domain", "score_before", "score_after", "delta"])
                for company, domain, before, after, delta in subset:
                    ws_sum.append([company, domain,
                                   round(before, 4), round(after, 4), round(delta, 4)])

    ws_sum.column_dimensions["A"].width = 50
    ws_sum.column_dimensions["B"].width = 50

    # ── Lovable App Export sheet ───────────────────────────────────────────────
    lov_stats = _append_lovable_export_sheet(wb_out, out_rows, out_headers)

    # Legend in summary sheet
    _section("Lovable App Export — row color legend")
    ws_sum.append(["Light blue  (DDEBF7)", "HQ recalculated"])
    ws_sum.append(["Light orange (FCE4D6)", "Competitor suppressed"])
    ws_sum.append(["Light purple (EADCF8)", "Both HQ and competitor applied"])

    _section("Lovable App Export metrics")
    _add("Lovable App Export columns",            lov_stats["n_cols"])
    _add("Lovable App Export rows",               lov_stats["n_rows"])
    _add("HQ-colored rows",                       lov_stats["n_hq"])
    _add("Competitor-colored rows",               lov_stats["n_comp"])
    _add("Both-colored rows",                     lov_stats["n_both"])

    return wb_out, lov_stats


# ---------------------------------------------------------------------------
# Lovable App Export helpers
# ---------------------------------------------------------------------------

LOVABLE_EXPORT_SHEET = "Lovable App Export"

# Ordered column blueprint — only those present in out_headers are written
_LOVABLE_COL_ORDER = [
    # Identification
    "company_name", "name", "domain", "website", "homepage_url",
    "input_country_used", "country", "city", "employee_range", "industry", "sector",
    # Score
    "commercial_fit_score_app", "commercial_tier_app",
    "final_commercial_fit_score", "commercial_fit_score",
    "recalc_scope_applied", "app_text_refresh_applied", "app_text_refresh_reason",
    # Main app-facing text
    "what_is_hot_app", "what_is_not_app", "why_relevant_app",
    "caller_angle_app", "call_starter_app", "caution_app",
    "evidence_summary_app", "key_source_links_app",
    # Contact / call prep
    "how_to_contact_app", "contact_angle_app", "sales_action_hint_app", "sales_action_hint",
    # HQ status
    "foreign_hq_signal_used_in_app", "foreign_hq_country_app", "foreign_hq_city_app",
    "sig_foreign_hq_score",
    "hq_recalc_applied", "hq_recalc_reason",
    "hq_score_before_recalc", "hq_score_after_recalc",
    # Competitor status
    "competitor_signal_used_in_app", "competitor_signal_suppressed",
    "competitor_signal_used_for_scoring", "competitor_recalc_applied", "competitor_recalc_reason",
    "competitor_provider_detected", "competitor_attention_provider_detected",
    "competitor_signal_strength", "competitor_attention_strength",
    "competitor_signal_excluded_from_next_scoring",
    # Important signals (explicit, in coefficient order)
    "sig_foreign_hq_score",      # already above but deduplicated below
    "sig_explicit_lnd_score", "sig_intl_footprint_score",
    "sig_employer_branding_score", "sig_lnd_onboarding_score",
    "ti_onboarding_score", "sig_rapid_growth_score",
    # Snippets
    "google_snippet_1", "google_snippet_1_url",
    "google_snippet_2", "google_snippet_2_url",
    "google_snippet_3", "google_snippet_3_url",
    "google_snippet_4", "google_snippet_4_url",
    "google_snippet_5", "google_snippet_5_url",
    "google_snippet_6", "google_snippet_6_url",
    "google_snippet_7", "google_snippet_7_url",
    "google_snippet_8", "google_snippet_8_url",
    "google_snippet_9", "google_snippet_9_url",
    "google_snippet_10", "google_snippet_10_url",
    # Advanced / audit
    "advanced_notes_app",
    "app_text_hq_note_added", "app_text_competitor_note_added",
    "app_text_conflicting_text_removed",
]

_FILL_HQ   = PatternFill("solid", fgColor="DDEBF7")   # light blue
_FILL_COMP = PatternFill("solid", fgColor="FCE4D6")    # light orange
_FILL_BOTH = PatternFill("solid", fgColor="EADCF8")    # light purple
_FILL_CELL_CONFLICT = PatternFill("solid", fgColor="FFF2CC")  # yellow — conflict removed
_FILL_CELL_HQ_YES   = PatternFill("solid", fgColor="E2F0D9")  # green  — HQ used
_FILL_CELL_COMP_SUP = PatternFill("solid", fgColor="FCE4D6")  # orange — competitor suppressed


def _build_lovable_export_headers(all_headers: list[str]) -> list[str]:
    """Return ordered, deduplicated export columns that exist in all_headers."""
    present = set(all_headers)
    seen: set[str] = set()
    result: list[str] = []
    # 1. Explicit blueprint columns
    for col in _LOVABLE_COL_ORDER:
        if col in present and col not in seen:
            result.append(col)
            seen.add(col)
    # 2. Remaining sig_* / ti_* columns not already included
    for col in all_headers:
        if col not in seen and (col.startswith("sig_") or col.startswith("ti_")):
            result.append(col)
            seen.add(col)
    return result


def _append_lovable_export_sheet(
    wb_out: Workbook,
    out_rows: list[dict],
    out_headers: list[str],
) -> dict:
    """Create the Lovable App Export sheet and return coloring statistics."""
    lov_headers = _build_lovable_export_headers(out_headers)
    ws = wb_out.create_sheet(LOVABLE_EXPORT_SHEET)

    _hdr_fill = PatternFill("solid", fgColor="D9EAF7")
    _hdr_font = Font(bold=True)

    # Header row
    ws.append(lov_headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 22
    for cell in ws[1]:
        cell.font = _hdr_font
        cell.fill = _hdr_fill

    # Pre-compute column indexes for special cell highlights
    col_idx = {h: idx + 1 for idx, h in enumerate(lov_headers)}
    ci_conflict = col_idx.get("app_text_conflicting_text_removed")
    ci_hq_used  = col_idx.get("foreign_hq_signal_used_in_app")
    ci_comp_sup = col_idx.get("competitor_signal_suppressed")

    # Column widths — capped at 50
    _MAX_W = 50
    _SAMPLE = 30
    col_widths: dict[int, int] = {
        i + 1: min(len(str(h)), _MAX_W) for i, h in enumerate(lov_headers)
    }

    n_hq = n_comp = n_both = 0

    for row_idx, r in enumerate(out_rows, start=2):
        ws.append([r.get(h) for h in lov_headers])

        hq_yes   = str(r.get("hq_recalc_applied",   "")).strip() == "Yes"
        comp_yes = str(r.get("competitor_recalc_applied", "")).strip() == "Yes"

        if hq_yes and comp_yes:
            row_fill = _FILL_BOTH
            n_both += 1
        elif hq_yes:
            row_fill = _FILL_HQ
            n_hq += 1
        elif comp_yes:
            row_fill = _FILL_COMP
            n_comp += 1
        else:
            row_fill = None

        if row_fill is not None:
            for cell in ws[row_idx]:
                cell.fill = row_fill

        # Optional cell-level highlights (applied on top of row fill)
        if ci_conflict and str(r.get("app_text_conflicting_text_removed", "")) == "Yes":
            ws.cell(row_idx, ci_conflict).fill = _FILL_CELL_CONFLICT
        if ci_hq_used and str(r.get("foreign_hq_signal_used_in_app", "")) == "Yes":
            ws.cell(row_idx, ci_hq_used).fill = _FILL_CELL_HQ_YES
        if ci_comp_sup and str(r.get("competitor_signal_suppressed", "")) == "Yes":
            ws.cell(row_idx, ci_comp_sup).fill = _FILL_CELL_COMP_SUP

        # Sample-based column width tracking
        if row_idx <= _SAMPLE + 1:
            for ci, h in enumerate(lov_headers, start=1):
                v = r.get(h)
                if v is not None:
                    col_widths[ci] = min(max(col_widths[ci], len(str(v))), _MAX_W)

    for ci, width in col_widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = min(width + 2, _MAX_W)

    return {
        "n_cols": len(lov_headers),
        "n_rows": len(out_rows),
        "n_hq":   n_hq,
        "n_comp": n_comp,
        "n_both": n_both,
    }


# ---------------------------------------------------------------------------
# Core recalculation logic
# ---------------------------------------------------------------------------

def recalculate_hq_changed_scores_workbook(
    enriched_workbook_file,
    hq_recovery_workbook_file,
    sheet_name: str = DEFAULT_SHEET,
    fast_output: bool = True,
    max_eligible_rows: int = 0,   # kept for backwards compat; alias of max_recalculated_rows
    max_recalculated_rows: int = 0,
    scope: str = SCOPE_HQ,
    refresh_app_text: bool = True,
    progress_callback=None,
) -> tuple[bytes, dict]:
    """Process two workbook file-like objects (or paths) and return (excel_bytes, summary).

    scope: "hq" | "competitor" | "both"
    max_recalculated_rows: 0 = unlimited
    refresh_app_text: regenerate Lovable app-facing text fields for recalculated rows
    progress_callback: optional callable(dict) called periodically during the row loop
    """
    if scope not in VALID_SCOPES:
        scope = SCOPE_HQ
    # max_eligible_rows is the old name — honour it if the new param wasn't set
    _limit = max_recalculated_rows or max_eligible_rows

    # ── Load workbooks ────────────────────────────────────────────────────────
    wb_enr = load_workbook(enriched_workbook_file, read_only=True, data_only=True)
    enr_headers, enr_rows = _wb_to_rows(wb_enr, sheet_name)
    wb_enr.close()

    wb_hqr = load_workbook(hq_recovery_workbook_file, read_only=True, data_only=True)
    hqr_headers, hqr_rows = _wb_to_rows(wb_hqr, sheet_name)
    wb_hqr.close()

    # ── Build HQ Recovery match index ────────────────────────────────────────
    hqr_has_domain  = "domain"        in hqr_headers
    hqr_has_company = ("company_name" in hqr_headers or "name" in hqr_headers)
    hqr_has_country = ("input_country_used" in hqr_headers or "country" in hqr_headers)
    hqr_index, strategy = _build_match_index(
        hqr_rows, hqr_has_domain, hqr_has_company, hqr_has_country
    )
    use_row_order = (strategy == "row_order_fallback")

    if use_row_order and len(enr_rows) != len(hqr_rows):
        return b"", {
            "error": (
                f"Cannot use row-order fallback: enriched has {len(enr_rows)} rows, "
                f"HQ Recovery has {len(hqr_rows)} rows."
            ),
        }

    # ── Build output header list ──────────────────────────────────────────────
    out_headers = list(enr_headers)
    _extra = (GENERAL_AUDIT_COLS, HQ_AUDIT_COLS, COMPETITOR_AUDIT_COLS, SCORE_OUTPUT_COLS)
    if refresh_app_text:
        _extra = _extra + (APP_TEXT_COLS,)
    for cols in _extra:
        for c in cols:
            if c not in out_headers:
                out_headers.append(c)

    # ── Row loop ──────────────────────────────────────────────────────────────
    import time as _time
    out_rows: list[dict] = []
    n_matched = 0
    n_hq_eligible = n_upgrades = n_downgrades = n_other_hq = 0
    n_hq_recalculated = n_hq_skipped_limit = 0
    n_competitor_detected = 0
    n_recalculated = n_skipped_limit = 0
    n_competitor_recalculated = n_competitor_skipped_limit = 0
    n_app_text_refreshed = n_hq_notes = n_comp_notes = n_conflict_removed = 0
    competitor_before_vals: list[float] = []
    competitor_after_vals:  list[float] = []
    deltas: list[tuple[str, str, float, float, float]] = []
    _total_rows = len(enr_rows)
    _loop_start = _time.monotonic()

    def _emit_progress(done: bool = False) -> None:
        if progress_callback is None:
            return
        elapsed = _time.monotonic() - _loop_start
        progress_callback({
            "phase":              "complete" if done else "processing",
            "row_index":          i + 1,
            "total_rows":         _total_rows,
            "matched_rows":       n_matched,
            "eligible_rows_seen": n_hq_eligible + n_competitor_detected,
            "recalculated_rows":  n_recalculated,
            "skipped_by_limit":   n_skipped_limit,
            "scope":              scope,
            "limit":              _limit,
            "elapsed_seconds":    elapsed,
            "done":               done,
        })

    for i, enr_row in enumerate(enr_rows):
        row_out = dict(enr_row)

        # ── Match to HQ Recovery ──────────────────────────────────────────────
        if use_row_order:
            hqr_row, matched = hqr_rows[i], True
        else:
            hqr_idx = hqr_index.get(_match_key_for_row(enr_row, strategy))
            hqr_row, matched = (hqr_rows[hqr_idx], True) if hqr_idx is not None else ({}, False)

        if matched:
            n_matched += 1

        # ── Determine HQ eligibility ──────────────────────────────────────────
        hq_eligible      = False
        hq_reviewed_val  = None
        hq_original_val  = None
        hq_reason        = ""
        if matched and scope in (SCOPE_HQ, SCOPE_BOTH):
            # Reviewed score: sig_foreign_hq_score_for_next_scoring is authoritative
            # (written by the AI-first probe app); fall back to sig_foreign_hq_score_reviewed.
            reviewed_raw = None
            for _rev_col in ("sig_foreign_hq_score_for_next_scoring", "sig_foreign_hq_score_reviewed"):
                _v_raw = hqr_row.get(_rev_col)
                if _v_raw is not None and _v_raw != "":
                    reviewed_raw = _v_raw
                    break

            # Original score: prefer HQR-tracked pre-recovery value so a previously
            # recalculated enriched file (where sig_foreign_hq_score is already updated)
            # does not make the row look like it was never changed.
            original_raw = None
            for _orig_col in ("sig_foreign_hq_score_original", "sig_foreign_hq_score_original_before_recovery"):
                _o_raw = hqr_row.get(_orig_col)
                if _o_raw is not None and _o_raw != "":
                    original_raw = _o_raw
                    break
            if original_raw is None:
                original_raw = enr_row.get("sig_foreign_hq_score")

            hq_reviewed_val = _safe_float(reviewed_raw)
            hq_original_val = _safe_float(original_raw)
            if reviewed_raw is not None and reviewed_raw != "" and hq_reviewed_val != hq_original_val:
                hq_eligible = True
                hq_reason   = f"HQ Recovery changed score {hq_original_val!r} → {hq_reviewed_val!r}"
                n_hq_eligible += 1

        # ── Determine competitor eligibility ──────────────────────────────────
        competitor_eligible = False
        if scope in (SCOPE_COMPETITOR, SCOPE_BOTH):
            if _has_competitor_signal(enr_row):
                competitor_eligible = True
                n_competitor_detected += 1

        row_is_eligible = hq_eligible or competitor_eligible

        if not row_is_eligible:
            row_out["recalc_scope_applied"] = "No"
            row_out["hq_recalc_applied"]    = "No"
            row_out["competitor_recalc_applied"] = "No"
            # Wipe stale HQ audit columns that may have carried over from a prior
            # recalc run on the enriched workbook — prevents inconsistent state where
            # hq_score_after_recalc=3 but hq_recalc_applied=No.
            for _c in _HQ_STALE_CLEAR_COLS:
                row_out[_c] = None
            out_rows.append(row_out)
            continue

        # ── Test-mode limit check ─────────────────────────────────────────────
        if _limit > 0 and n_recalculated >= _limit:
            n_skipped_limit += 1
            if hq_eligible:
                row_out["hq_recalc_applied"] = "No - skipped by test row limit"
                n_hq_skipped_limit += 1
                for _c in _HQ_STALE_CLEAR_COLS:
                    row_out[_c] = None
            if competitor_eligible:
                row_out["competitor_recalc_applied"] = "No - skipped by test row limit"
                n_competitor_skipped_limit += 1
            row_out["recalc_scope_applied"] = "Skipped (test row limit)"
            out_rows.append(row_out)
            continue

        # ── Build scoring row copy ────────────────────────────────────────────
        old_cfs, _cfs_col = _get_existing_commercial_fit_score(row_out)
        row_copy = dict(row_out)

        if hq_eligible:
            new_hq = hq_reviewed_val if hq_reviewed_val is not None else 0.0
            row_copy["sig_foreign_hq_score"]                  = new_hq
            row_copy["sig_foreign_hq_score_for_next_scoring"] = new_hq

        comp_before = {f: _safe_float(row_out.get(f)) or 0.0 for f in _COMPETITOR_NUMERIC_FIELDS}
        if competitor_eligible:
            for f in _COMPETITOR_NUMERIC_FIELDS:
                row_copy[f] = 0.0

        # ── Single score_company call ─────────────────────────────────────────
        try:
            score_out = score_company(row_copy, {"scoring_profile": SCORING_PROFILE})
            for col in SCORE_OUTPUT_COLS:
                if col in score_out:
                    row_out[col] = score_out[col]

            new_cfs = _safe_float(score_out.get("final_commercial_fit_score")) or 0.0
            delta   = round(new_cfs - old_cfs, 4)

            row_out["recalc_scope_applied"] = scope
            row_out["cfs_before_recalc"]    = old_cfs
            row_out["cfs_after_recalc"]     = new_cfs
            row_out["cfs_delta_recalc"]     = delta
            row_out["cfs_source_col_used"]  = _cfs_col

            # HQ audit
            if hq_eligible:
                new_hq = hq_reviewed_val if hq_reviewed_val is not None else 0.0
                old_hq = hq_original_val or 0.0
                row_out.update({
                    "hq_recalc_applied":                           "Yes",
                    "hq_recalc_reason":                            hq_reason,
                    "hq_score_before_recalc":                      old_hq,
                    "hq_score_after_recalc":                       new_hq,
                    "commercial_fit_score_before_hq_recalc":       old_cfs,
                    "commercial_fit_score_after_hq_recalc":        new_cfs,
                    "commercial_fit_score_delta_hq_recalc":        delta,
                    "final_commercial_fit_score_before_hq_recalc": old_cfs,
                    "final_commercial_fit_score_after_hq_recalc":  new_cfs,
                    "final_commercial_fit_score_delta_hq_recalc":  delta,
                    "commercial_fit_score_before_source_column":   _cfs_col,
                    "sig_foreign_hq_score":                        new_hq,
                    "sig_foreign_hq_score_for_next_scoring":       new_hq,
                    # Reviewed HQ evidence — prevent "no foreign parent" old evidence
                    # from being the only visible proof when new HQ score = 3.
                    "hq_recalc_reviewed_country": _hqr_country(hqr_row),
                    "hq_recalc_reviewed_city":    _hqr_city(hqr_row),
                    "hq_recalc_evidence_url":     _hqr_evidence_url(hqr_row),
                    "hq_recalc_evidence_quote":   _v(hqr_row, "hq_evidence_quote", "ai_evidence_quote"),
                    "hq_recalc_ai_reason":        _v(hqr_row, "hq_reason", "ai_hq_reason"),
                    "old_sig_foreign_hq_evidence": _v(enr_row, "sig_foreign_hq_evidence",
                                                      "sig_foreign_hq_evidence_url"),
                })
                # Overwrite sig_foreign_hq_evidence with a clean reviewed statement so
                # old contradictory text ("Italian headquarters", "no foreign parent")
                # is no longer the visible HQ evidence for this row.
                _rev_country = _hqr_country(hqr_row)
                _rev_city    = _hqr_city(hqr_row)
                _rev_url     = _hqr_evidence_url(hqr_row)
                _rev_quote   = _v(hqr_row, "hq_evidence_quote", "ai_evidence_quote")
                _rev_reason  = _v(hqr_row, "hq_reason", "ai_hq_reason")
                _rev_loc     = (f"{_rev_city}, {_rev_country}" if _rev_city and _rev_country
                                else _rev_country or _rev_city or "")
                _ev_parts: list[str] = [
                    f"HQ Recovery reviewed: foreign parent/group HQ in {_rev_loc}."
                    if _rev_loc else
                    "HQ Recovery reviewed: foreign parent/group HQ confirmed."
                ]
                if _rev_quote:
                    _ev_parts.append(f'Evidence: "{_rev_quote[:200]}"')
                if _rev_url:
                    _ev_parts.append(f"Source: {_rev_url}")
                if _rev_reason:
                    _ev_parts.append(f"Reason: {_rev_reason[:200]}")
                _ev_parts.append("Previous HQ evidence superseded.")
                row_out["sig_foreign_hq_evidence"] = " ".join(_ev_parts)

                # Domain mismatch safety: flag rows where domain match was uncertain
                _domain_mismatch = (
                    str(enr_row.get("possible_domain_mismatch") or "").strip().lower() in ("yes", "true", "1")
                    or str(hqr_row.get("possible_domain_mismatch") or "").strip().lower() in ("yes", "true", "1")
                )
                if _domain_mismatch:
                    row_out["hq_recalc_domain_mismatch_warning"] = "Yes"
                    if "needs_manual_review" in row_out:
                        row_out["needs_manual_review"] = "Yes"
                else:
                    row_out["hq_recalc_domain_mismatch_warning"] = "No"

                n_hq_recalculated += 1
                if old_hq in (0.0, None) and new_hq == 3.0:
                    n_upgrades += 1
                elif old_hq == 3.0 and new_hq in (0.0, None):
                    n_downgrades += 1
                else:
                    n_other_hq += 1
            else:
                row_out["hq_recalc_applied"] = "No"

            # Competitor audit
            if competitor_eligible:
                row_out.update({
                    "competitor_recalc_applied":               "Yes",
                    "competitor_recalc_reason":                "Competitor signal neutralized for scoring",
                    "competitor_signal_before_recalc":         _first_numeric(comp_before, ("competitor_signal_strength", "competitor_attention_strength", "competitor_signal_strength_score")),
                    "competitor_signal_after_recalc":          0.0,
                    "language_competitor_signal_before_recalc": comp_before.get("language_competitor_strength_score", 0.0),
                    "language_competitor_signal_after_recalc": 0.0,
                    "competitor_signal_neutralized_for_scoring": "Yes",
                    "competitor_signal_used_for_scoring":      "No",
                    "competitor_signal_suppressed":            "Yes",
                })
                for f, bval in comp_before.items():
                    if bval > 0:
                        competitor_before_vals.append(bval)
                competitor_after_vals.append(0.0)
                n_competitor_recalculated += 1
            else:
                row_out["competitor_recalc_applied"] = "No"

            # ── App-text refresh ──────────────────────────────────────────
            if refresh_app_text:
                _flags = _refresh_app_text(
                    row_out, enr_row, hqr_row,
                    suppress_competitor=competitor_eligible,
                    score_out=score_out,
                    hq_eligible=hq_eligible,
                    hq_reviewed_val=hq_reviewed_val,
                    hq_original_val=hq_original_val,
                )
                row_out["app_text_refresh_applied"] = "Yes"
                row_out["app_text_refresh_reason"]  = (
                    ("HQ changed" if hq_eligible else "")
                    + (" + competitor suppressed" if competitor_eligible else "")
                ).strip(" +")
                row_out["app_text_hq_note_added"]            = "Yes" if _flags["hq_note_added"]    else "No"
                row_out["app_text_competitor_note_added"]    = "Yes" if _flags["comp_note_added"]  else "No"
                row_out["app_text_conflicting_text_removed"] = "Yes" if _flags["conflict_removed"] else "No"
                n_app_text_refreshed += 1
                if _flags["hq_note_added"]:    n_hq_notes      += 1
                if _flags["comp_note_added"]:  n_comp_notes    += 1
                if _flags["conflict_removed"]: n_conflict_removed += 1

            n_recalculated += 1
            deltas.append((
                str(enr_row.get("company_name") or enr_row.get("name") or "?"),
                str(enr_row.get("domain") or "?"),
                old_cfs, new_cfs, delta,
            ))

        except Exception as exc:
            row_out["hq_recalc_applied"]         = f"Error: {exc}"
            row_out["competitor_recalc_applied"]  = f"Error: {exc}"
            row_out["recalc_scope_applied"]       = f"Error: {exc}"

        out_rows.append(row_out)

        # Emit progress: every 100 rows, every recalculated row for the first 20,
        # and whenever a recalculation just happened (cheap gate).
        if progress_callback is not None:
            _just_recalculated = (
                row_out.get("recalc_scope_applied") not in (None, "No", "Skipped (test row limit)")
                and not str(row_out.get("recalc_scope_applied", "")).startswith("Error")
            )
            if (i % 100 == 0) or (_just_recalculated and n_recalculated <= 20):
                _emit_progress()

    _emit_progress(done=True)

    avg_comp_before = (sum(competitor_before_vals) / len(competitor_before_vals)
                       if competitor_before_vals else 0.0)

    # ── Validation checks ─────────────────────────────────────────────────────
    # These should all be 0 after the fixes; kept as safety net for future regressions.
    n_val_inconsistent_hq = sum(
        1 for r in out_rows
        if _safe_float(r.get("hq_score_after_recalc")) == 3.0
        and str(r.get("hq_recalc_applied") or "").startswith("No")
        and str(r.get("hq_recalc_reason") or "")
    )
    n_val_recalc_missing_scores = sum(
        1 for r in out_rows
        if str(r.get("hq_recalc_applied") or "") == "Yes"
        and (r.get("commercial_fit_score_before_hq_recalc") is None
             or r.get("commercial_fit_score_after_hq_recalc") is None)
    )
    n_val_blank_final_score = sum(
        1 for r in out_rows
        if str(r.get("hq_recalc_applied") or "") == "Yes"
        and r.get("final_commercial_fit_score") is None
    )
    # Rows where HQ was recalculated to 3 but sig_foreign_hq_evidence still contains
    # contradictory domestic/no-foreign claims — should be 0 after the evidence overwrite.
    n_val_contradictory_evidence = sum(
        1 for r in out_rows
        if str(r.get("hq_recalc_applied") or "") == "Yes"
        and _safe_float(r.get("hq_score_after_recalc")) == 3.0
        and _CONTRADICTORY_DOMESTIC_HQ_RE.search(str(r.get("sig_foreign_hq_evidence") or ""))
    )

    _hq_recalc_rows = [r for r in out_rows if str(r.get("hq_recalc_applied") or "") == "Yes"]
    # Domain mismatch warning: HQ recalculated but possible domain mismatch detected
    n_val_domain_mismatch_warning = sum(
        1 for r in _hq_recalc_rows
        if str(r.get("hq_recalc_domain_mismatch_warning") or "").strip() == "Yes"
    )
    # App text still contradictory after cleanup (any of the four main app text fields)
    _app_text_check_cols = ("why_relevant_app", "caution_app", "evidence_summary_app", "what_is_hot_app")
    n_val_app_text_still_contradictory = sum(
        1 for r in _hq_recalc_rows
        if any(_CONTRADICTORY_DOMESTIC_HQ_RE.search(str(r.get(c) or "")) for c in _app_text_check_cols)
    )
    # Evidence URL and quote coverage
    n_val_has_evidence_url = sum(
        1 for r in _hq_recalc_rows if str(r.get("hq_recalc_evidence_url") or "").strip()
    )
    n_val_has_evidence_quote = sum(
        1 for r in _hq_recalc_rows if str(r.get("hq_recalc_evidence_quote") or "").strip()
    )

    summary = {
        "error":                     "",
        "scope":                     scope,
        "strategy":                  strategy,
        "n_enr":                     len(enr_rows),
        "n_hqr":                     len(hqr_rows),
        "n_matched":                 n_matched,
        "n_hq_eligible":             n_hq_eligible,
        "n_hq_recalculated":         n_hq_recalculated,
        "n_hq_skipped_limit":        n_hq_skipped_limit,
        "n_competitor_detected":     n_competitor_detected,
        "n_recalculated":            n_recalculated,
        "skipped_by_recalc_limit":   n_skipped_limit,
        "n_upgrades":                n_upgrades,
        "n_downgrades":              n_downgrades,
        "n_other":                   n_other_hq,
        "n_competitor_recalculated": n_competitor_recalculated,
        "n_competitor_skipped_limit": n_competitor_skipped_limit,
        "avg_competitor_before":     round(avg_comp_before, 4),
        "avg_competitor_after":      0.0,
        "app_text_refreshed":        refresh_app_text and n_app_text_refreshed > 0,
        "n_app_text_refreshed":      n_app_text_refreshed,
        "n_hq_notes":                n_hq_notes,
        "n_comp_notes":              n_comp_notes,
        "n_conflict_removed":        n_conflict_removed,
        "deltas":                    deltas,
        "test_mode_active":          _limit > 0,
        "max_recalculated_rows":     _limit,
        # Validation
        "n_val_inconsistent_hq":              n_val_inconsistent_hq,
        "n_val_recalc_missing_scores":        n_val_recalc_missing_scores,
        "n_val_blank_final_score":            n_val_blank_final_score,
        "n_val_contradictory_evidence":       n_val_contradictory_evidence,
        "n_val_domain_mismatch_warning":      n_val_domain_mismatch_warning,
        "n_val_app_text_still_contradictory": n_val_app_text_still_contradictory,
        "n_val_has_evidence_url":             n_val_has_evidence_url,
        "n_val_has_evidence_quote":           n_val_has_evidence_quote,
    }

    wb_out, lov_stats = _build_output_wb(
        out_headers, out_rows, sheet_name, summary, deltas, fast_output=fast_output
    )
    summary["lovable_export"] = lov_stats
    buf = io.BytesIO()
    wb_out.save(buf)
    return buf.getvalue(), summary


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--enriched-workbook",    required=True)
    ap.add_argument("--hq-recovery-workbook", required=True)
    ap.add_argument("--output",               required=True)
    ap.add_argument("--sheet",                default=DEFAULT_SHEET)
    ap.add_argument("--recalculation-scope",  default=SCOPE_HQ,
                    choices=list(VALID_SCOPES),
                    help="hq | competitor | both  (default: hq)")
    ap.add_argument("--max-recalculated-rows", type=int, default=0,
                    help="Limit recalculated rows (0 = unlimited)")
    args = ap.parse_args()

    print(f"\n{'='*72}")
    print("Score Recalculation")
    print(f"  enriched   : {args.enriched_workbook}")
    print(f"  hq-recovery: {args.hq_recovery_workbook}")
    print(f"  output     : {args.output}")
    print(f"  sheet      : {args.sheet}")
    print(f"  scope      : {args.recalculation_scope}")
    print(f"  profile    : {SCORING_PROFILE}")
    print(f"  row limit  : {args.max_recalculated_rows or 'unlimited'}")
    print(f"{'='*72}\n")

    excel_bytes, summary = recalculate_hq_changed_scores_workbook(
        args.enriched_workbook,
        args.hq_recovery_workbook,
        sheet_name=args.sheet,
        scope=args.recalculation_scope,
        max_recalculated_rows=args.max_recalculated_rows,
    )

    if summary.get("error"):
        print(f"ERROR: {summary['error']}")
        sys.exit(1)

    with open(args.output, "wb") as fh:
        fh.write(excel_bytes)

    deltas = summary["deltas"]
    print(f"\n{'='*72}")
    print("RESULTS")
    print(f"  Scope                : {summary['scope']}")
    print(f"  Matching strategy    : {summary['strategy']}")
    print(f"  Rows matched         : {summary['n_matched']} / {summary['n_enr']}")
    if summary["scope"] in (SCOPE_HQ, SCOPE_BOTH):
        print(f"  HQ-eligible rows     : {summary['n_hq_eligible']}")
        print(f"  HQ upgrades  0→3     : {summary['n_upgrades']}")
        print(f"  HQ downgrades 3→0    : {summary['n_downgrades']}")
    if summary["scope"] in (SCOPE_COMPETITOR, SCOPE_BOTH):
        print(f"  Competitor detected  : {summary['n_competitor_detected']}")
        print(f"  Competitor recalc'd  : {summary['n_competitor_recalculated']}")
    print(f"  Recalculated total   : {summary['n_recalculated']}")
    print(f"  Skipped (row limit)  : {summary['skipped_by_recalc_limit']}")
    if deltas:
        all_d = [x[4] for x in deltas]
        print(f"  Biggest increase     : +{max(all_d):.4f}")
        print(f"  Biggest decrease     : {min(all_d):.4f}")
    print(f"  Output file          : {args.output}")
    print(f"{'='*72}\n")


def _running_under_streamlit() -> bool:
    try:
        from streamlit.runtime.scriptrunner import get_script_run_ctx
        return get_script_run_ctx() is not None
    except Exception:
        return False


if __name__ == "__main__":
    if _running_under_streamlit():
        import streamlit as st
        st.error(
            "This is the command-line backend script. "
            "Please run `streamlit run hq_score_recalc_app.py` for the browser UI."
        )
        st.stop()
    main()
