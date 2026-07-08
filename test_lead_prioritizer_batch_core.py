"""Tests for the Lead Prioritizer v2 shared batch core.

``prioritize_single_lead`` is mocked — no live APIs, no real keys.
"""

from __future__ import annotations

import time
from unittest.mock import patch

import pandas as pd
import pytest

from lead_output_schema import LeadPrioritizationResult, LeadEvidence, LeadSignal
import lead_prioritizer_batch_core as bc
from lead_prioritizer_batch_core import (
    BatchRunConfig,
    resolve_pipeline_flags,
    select_batch_rows,
    flatten_result_for_excel,
    flatten_evidence_for_excel,
    flatten_signals_for_excel,
    flatten_deep_dive_for_excel,
    build_run_summary_dataframe,
    build_excel_workbook_bytes,
    run_batch_dataframe,
    should_run_deep_dive,
)
from deep_dive_schema import DeepDiveClaim, DeepDiveResult


def _sample_result(**kw) -> LeadPrioritizationResult:
    base = dict(
        company_name="Acme", domain="acme.com", input_country="Italy",
        v2_pipeline_mode="full_v2_single_lead",
        hq_detected_country="Germany", hq_structure_type="foreign_parent",
        sig_foreign_hq_score_for_next_scoring=3.0,
        sig_international_profile_score=2.0,
        final_commercial_fit_score=7.2, commercial_tier="🥇 Hot",
        what_is_hot_app="Foreign HQ signal: Germany",
        ai_hq_raw_json='{"classification":"foreign_parent"}',
        evidence_items=[
            LeadEvidence(evidence_id="international_profile:organic:1",
                         signal_name="international_profile",
                         source_url="https://acme.com/intl",
                         source_title="Global", source_snippet="global offices",
                         query_used="q", parser_source="serper_organic_1",
                         source_type="organic"),
        ],
        signals=[
            LeadSignal(signal_name="international_profile", signal_value="positive_evidence",
                       signal_score=2.0, signal_confidence="High",
                       signal_reason="2 hits", evidence_url="https://acme.com/intl"),
        ],
    )
    base.update(kw)
    return LeadPrioritizationResult(**base)


# ---------------------------------------------------------------------------
# resolve_pipeline_flags
# ---------------------------------------------------------------------------

class TestResolveFlags:
    def test_full(self):
        f = resolve_pipeline_flags("full")
        assert f["run_full_v2_pipeline"] is True
        assert all(not f[k] for k in f if k != "run_full_v2_pipeline")

    def test_hq_only(self):
        assert all(v is False for v in resolve_pipeline_flags("hq_only").values())

    def test_evidence_only(self):
        f = resolve_pipeline_flags("evidence_only")
        assert f["collect_non_hq_evidence"] is True
        assert f["extract_non_hq_signals_flag"] is False
        assert f["run_full_v2_pipeline"] is False

    def test_signals_no_score(self):
        f = resolve_pipeline_flags("signals_no_score")
        assert f["collect_non_hq_evidence"] is True
        assert f["extract_non_hq_signals_flag"] is True
        assert f["build_app_summary_fields_flag"] is True
        assert f["calculate_commercial_score_flag"] is False
        assert f["build_caller_app_fields_flag"] is False

    def test_full_no_score(self):
        f = resolve_pipeline_flags("full_no_score")
        assert f["collect_non_hq_evidence"] is True
        assert f["extract_non_hq_signals_flag"] is True
        assert f["build_app_summary_fields_flag"] is True
        assert f["build_caller_app_fields_flag"] is True
        assert f["calculate_commercial_score_flag"] is False
        assert f["run_full_v2_pipeline"] is False

    def test_unknown_raises(self):
        with pytest.raises(ValueError):
            resolve_pipeline_flags("nope")


# ---------------------------------------------------------------------------
# select_batch_rows
# ---------------------------------------------------------------------------

class TestSelectRows:
    _df = pd.DataFrame({"c": list(range(10)), "d": [f"d{i}" for i in range(10)]})

    def _cfg(self, **kw):
        base = dict(company_name_column="c", domain_column="d")
        base.update(kw)
        return BatchRunConfig(**base)

    def test_start_and_limit(self):
        sub = select_batch_rows(self._df, self._cfg(start_row=2, row_limit=3))
        assert list(sub["c"]) == [2, 3, 4]
        assert list(sub.index) == [2, 3, 4]  # original index preserved

    def test_limit_zero_means_all_remaining(self):
        sub = select_batch_rows(self._df, self._cfg(start_row=7, row_limit=0))
        assert list(sub["c"]) == [7, 8, 9]

    def test_limit_larger_than_frame(self):
        sub = select_batch_rows(self._df, self._cfg(start_row=0, row_limit=100))
        assert len(sub) == 10


# ---------------------------------------------------------------------------
# Flatten helpers
# ---------------------------------------------------------------------------

class TestFlatten:
    def test_result_includes_key_fields_and_originals(self):
        row = flatten_result_for_excel(
            _sample_result(), {"c": "Acme", "d": "acme.com", "extra": "x"},
            source_index=5, run_success=True, run_error="")
        # original columns preserved
        assert row["c"] == "Acme" and row["extra"] == "x"
        # metadata
        assert row["source_index"] == 5 and row["run_success"] is True
        # HQ / score / app
        assert row["hq_detected_country"] == "Germany"
        assert row["sig_foreign_hq_score_for_next_scoring"] == 3.0
        assert row["final_commercial_fit_score"] == 7.2
        assert row["commercial_tier"] == "🥇 Hot"
        assert row["what_is_hot_app"] == "Foreign HQ signal: Germany"
        assert row["v2_pipeline_mode"] == "full_v2_single_lead"

    def test_flattens_lusha_audit_fields(self):
        # Regression: lusha_main_industry/lusha_sub_industry (Stap 2) and
        # company_size_complexity_source/lusha_employees/lusha_revenue
        # (Stap 3) must actually reach the Enriched Leads export -- these
        # exist on LeadPrioritizationResult but only appear in the Excel
        # output when also listed in _RESULT_FLAT_FIELDS.
        row = flatten_result_for_excel(
            _sample_result(
                lusha_main_industry="Manufacturing",
                lusha_sub_industry="Industrial Machinery & Equipment",
                sector_source="lusha_mapped",
                company_size_complexity_source="lusha",
                lusha_employees="201-500",
                lusha_revenue="$50M - $100M",
            ), {"c": "Acme", "d": "acme.com"}, source_index=1, run_success=True, run_error="")
        assert row["lusha_main_industry"] == "Manufacturing"
        assert row["lusha_sub_industry"] == "Industrial Machinery & Equipment"
        assert row["sector_source"] == "lusha_mapped"
        assert row["company_size_complexity_source"] == "lusha"
        assert row["lusha_employees"] == "201-500"
        assert row["lusha_revenue"] == "$50M - $100M"

    def test_flattens_rich_icp_context_fields(self):
        row = flatten_result_for_excel(
            _sample_result(
                icp_buying_signals="Active onboarding academy.",
                icp_likely_training_interest="Onboarding ramp-up.",
                icp_potential_buyer_function="HR / Talent",
                icp_context_by_ai=True,
                icp_context_content_note="AI-composed ICP context used.",
            ), {"c": "Acme", "d": "acme.com"}, source_index=1, run_success=True, run_error="")
        assert row["icp_buying_signals"] == "Active onboarding academy."
        assert row["icp_likely_training_interest"] == "Onboarding ramp-up."
        assert row["icp_potential_buyer_function"] == "HR / Talent"
        assert row["icp_context_by_ai"] is True
        assert row["icp_context_content_note"] == "AI-composed ICP context used."
        assert row["evidence_count"] == 1 and row["signal_count"] == 1

    def test_flattens_legacy_enrichment_fields(self):
        row = flatten_result_for_excel(
            _sample_result(
                legacy_score=9.0, legacy_tier="High",
                legacy_icp_lead_score="High",
                legacy_icp_buying_signals="International footprint",
                legacy_icp_likely_training_interest="Language training / Business English",
                legacy_icp_potential_buyer_function="HR",
                legacy_icp_why_relevant="Clear fit.",
                legacy_icp_evidence="Found international offices.",
            ), {"c": "Acme", "d": "acme.com"}, source_index=1, run_success=True, run_error="")
        assert row["legacy_score"] == 9.0
        assert row["legacy_tier"] == "High"
        assert row["legacy_icp_lead_score"] == "High"
        assert row["legacy_icp_buying_signals"] == "International footprint"
        assert row["legacy_icp_likely_training_interest"] == \
            "Language training / Business English"
        assert row["legacy_icp_potential_buyer_function"] == "HR"
        assert row["legacy_icp_why_relevant"] == "Clear fit."
        assert row["legacy_icp_evidence"] == "Found international offices."
        assert row["legacy_enrichment_error"] is None

    def test_legacy_enrichment_fields_blank_when_flag_off(self):
        row = flatten_result_for_excel(
            _sample_result(), {"c": "Acme", "d": "acme.com"},
            source_index=1, run_success=True, run_error="")
        assert row["legacy_score"] is None
        assert row["legacy_tier"] is None
        assert row["legacy_icp_lead_score"] is None

    def test_raw_ai_json_excluded_by_default(self):
        row = flatten_result_for_excel(_sample_result(), {}, 0, True, "")
        assert "ai_hq_raw_json" not in row

    def test_raw_ai_json_included_when_flagged(self):
        row = flatten_result_for_excel(_sample_result(), {}, 0, True, "",
                                       include_raw_ai_json=True)
        assert row["ai_hq_raw_json"].startswith("{")

    def test_no_api_key_fields(self):
        row = flatten_result_for_excel(_sample_result(), {"c": "Acme"}, 0, True, "")
        for key in row:
            k = key.lower()
            assert "api_key" not in k and "apikey" not in k
            assert "serper" not in k and "anthropic_api" not in k
        # also no competitor field surfaced
        assert not any("competitor" in k.lower() for k in row)

    def test_c4_safety_fields_in_enriched_leads(self):
        result = _sample_result(
            hq_positive_score_suppressed_for_review="Yes",
            hq_evidence_domain_mismatch_warning="Yes",
            hq_review_reason="risky domain root and evidence URL does not match lead domain",
            hq_query_risk_flag="Yes",
            hq_evidence_domain_match="No",
        )
        row = flatten_result_for_excel(result, {"c": "Acme"}, 0, True, "")
        assert row["hq_positive_score_suppressed_for_review"] == "Yes"
        assert row["hq_evidence_domain_mismatch_warning"] == "Yes"
        assert row["hq_review_reason"].startswith("risky domain root")
        assert row["hq_query_risk_flag"] == "Yes"
        assert row["hq_evidence_domain_match"] == "No"

    def test_employer_branding_fields_included(self):
        result = _sample_result(
            sig_employer_branding_score=2.0,
            employer_branding_reason="2 distinct keyword match(es) in evidence: "
                                      "employer branding, employee satisfaction",
            employer_branding_evidence_url="https://acme.com/careers",
            employer_branding_evidence_quote="Recognized as a great place to work.",
        )
        row = flatten_result_for_excel(result, {"c": "Acme"}, 0, True, "")
        assert row["sig_employer_branding_score"] == 2.0
        assert row["employer_branding_reason"].startswith("2 distinct")
        assert row["employer_branding_evidence_url"] == "https://acme.com/careers"
        assert row["employer_branding_evidence_quote"] == \
            "Recognized as a great place to work."

    def test_employer_branding_fields_default_to_none_when_absent(self):
        # A result without any employer branding data must still emit the
        # columns (as None) instead of crashing or dropping them.
        row = flatten_result_for_excel(_sample_result(), {"c": "Acme"}, 0, True, "")
        assert row["sig_employer_branding_score"] is None
        assert row["employer_branding_reason"] is None
        assert row["employer_branding_evidence_url"] is None
        assert row["employer_branding_evidence_quote"] is None

    def test_sector_fields_included(self):
        result = _sample_result(
            detected_industry="Consumer goods",
            detected_sub_industry="Consumer electronics",
            detected_company_type="Subsidiary",
            sector_confidence="High",
            sector_reason="Matched sector keyword(s): consumer electronics.",
            sector_evidence_url="https://acme.com/about",
            sector_evidence_quote="Acme is a consumer electronics company.",
            sector_source_title="About Acme",
        )
        row = flatten_result_for_excel(result, {"c": "Acme"}, 0, True, "")
        assert row["detected_industry"] == "Consumer goods"
        assert row["detected_sub_industry"] == "Consumer electronics"
        assert row["detected_company_type"] == "Subsidiary"
        assert row["sector_confidence"] == "High"
        assert row["sector_reason"].startswith("Matched sector")
        assert row["sector_evidence_url"] == "https://acme.com/about"
        assert row["sector_evidence_quote"] == "Acme is a consumer electronics company."
        assert row["sector_source_title"] == "About Acme"

    def test_sector_industry_evidence_rows_flattened(self):
        result = _sample_result(evidence_items=[
            LeadEvidence(evidence_id="sector_industry:organic:1",
                         signal_name="sector_industry",
                         source_url="https://acme.com/about",
                         source_title="About", source_snippet="a retail company",
                         query_used="acme company industry sector",
                         parser_source="serper_organic_1", source_type="organic"),
        ])
        rows = flatten_evidence_for_excel(result, source_index=2)
        assert len(rows) == 1
        assert rows[0]["signal_name"] == "sector_industry"
        assert rows[0]["source_snippet"] == "a retail company"

    def test_error_row_has_metadata_only(self):
        row = flatten_result_for_excel(None, {"c": "Acme"}, 3, False, "Boom: x")
        assert row["run_success"] is False
        assert row["run_error"] == "Boom: x"
        assert row["evidence_count"] == 0 and row["signal_count"] == 0
        assert "hq_detected_country" not in row

    def test_evidence_one_row_per_item(self):
        rows = flatten_evidence_for_excel(_sample_result(), 2)
        assert len(rows) == 1
        assert rows[0]["source_index"] == 2
        assert rows[0]["signal_name"] == "international_profile"
        assert rows[0]["source_url"] == "https://acme.com/intl"

    def test_signals_one_row_per_signal(self):
        rows = flatten_signals_for_excel(_sample_result(), 4)
        assert len(rows) == 1
        assert rows[0]["source_index"] == 4
        assert rows[0]["signal_score"] == 2.0


# ---------------------------------------------------------------------------
# Run summary
# ---------------------------------------------------------------------------

class TestRunSummary:
    def test_summary_fields_no_keys(self):
        cfg = BatchRunConfig(company_name_column="c", domain_column="d",
                             run_mode="full", start_row=1, row_limit=5)
        df = build_run_summary_dataframe(cfg, total_input_rows=20, selected_rows=5,
                                         processed_rows=5, success_count=4, error_count=1)
        rec = df.iloc[0].to_dict()
        assert rec["run_mode"] == "full"
        assert rec["total_input_rows"] == 20
        assert rec["success_count"] == 4 and rec["error_count"] == 1
        for k in rec:
            assert "api_key" not in k.lower() and "serper" not in k.lower()


# ---------------------------------------------------------------------------
# run_batch_dataframe
# ---------------------------------------------------------------------------

class TestRunBatch:
    _df = pd.DataFrame({
        "company": ["Acme", "Beta", "Gamma"],
        "domain": ["acme.com", "beta.com", "gamma.com"],
    })

    def test_calls_pipeline_with_full_flags(self):
        seen = {}

        def _fake(lead_input, **kwargs):
            seen.update(kwargs)
            return _sample_result(company_name=lead_input.company_name)

        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="full", row_limit=1)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake):
            out = run_batch_dataframe(self._df, cfg, "SERP", "ANTH")
        assert seen.get("run_full_v2_pipeline") is True
        assert seen.get("default_input_country") == "Italy"
        # keys passed through but never surface in output
        assert out["enriched_leads"].shape[0] == 1
        blob = out["enriched_leads"].to_csv(index=False)
        assert "SERP" not in blob and "ANTH" not in blob

    @pytest.mark.parametrize("compose_caller_content,rich_icp_context", [
        (False, False), (True, False), (False, True), (True, True),
    ])
    def test_compose_flags_independent_passthrough(self, compose_caller_content, rich_icp_context):
        # Onderdeel A (rich_icp_context) and Step 3 (compose_caller_content)
        # must reach prioritize_single_lead as independent kwargs — every
        # combination must pass through with no cross-dependency.
        seen = {}

        def _fake(lead_input, **kwargs):
            seen.update(kwargs)
            return _sample_result(company_name=lead_input.company_name)

        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="full", row_limit=1,
                             compose_caller_content=compose_caller_content,
                             rich_icp_context=rich_icp_context)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake):
            run_batch_dataframe(self._df, cfg, "SERP", "ANTH")
        assert seen.get("compose_caller_content_flag") is compose_caller_content
        assert seen.get("compose_icp_context") is rich_icp_context

    @pytest.mark.parametrize("ai_signal_scoring", [False, True])
    def test_ai_signal_scoring_passthrough_independent_of_other_flags(self, ai_signal_scoring):
        # Onderdeel 2 must reach prioritize_single_lead as an independent
        # kwarg, off by default, with no cross-dependency on the compose
        # flags above.
        seen = {}

        def _fake(lead_input, **kwargs):
            seen.update(kwargs)
            return _sample_result(company_name=lead_input.company_name)

        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="full", row_limit=1,
                             ai_signal_scoring=ai_signal_scoring)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake):
            run_batch_dataframe(self._df, cfg, "SERP", "ANTH")
        assert seen.get("ai_signal_scoring") is ai_signal_scoring
        assert seen.get("compose_caller_content_flag") is False
        assert seen.get("compose_icp_context") is False

    def test_ai_signal_scoring_default_is_false(self):
        cfg = BatchRunConfig(company_name_column="company", domain_column="domain")
        assert cfg.ai_signal_scoring is False

    @pytest.mark.parametrize("legacy_enrichment_mode", [False, True])
    def test_legacy_enrichment_mode_passthrough_independent_of_other_flags(
        self, legacy_enrichment_mode,
    ):
        seen = {}

        def _fake(lead_input, **kwargs):
            seen.update(kwargs)
            return _sample_result(company_name=lead_input.company_name)

        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="full", row_limit=1,
                             legacy_enrichment_mode=legacy_enrichment_mode)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake):
            run_batch_dataframe(self._df, cfg, "SERP", "ANTH")
        assert seen.get("legacy_enrichment_mode") is legacy_enrichment_mode
        assert seen.get("compose_caller_content_flag") is False
        assert seen.get("compose_icp_context") is False
        assert seen.get("ai_signal_scoring") is False

    def test_legacy_enrichment_mode_default_is_false(self):
        cfg = BatchRunConfig(company_name_column="company", domain_column="domain")
        assert cfg.legacy_enrichment_mode is False

    def test_public_source_signal_enrichment_defaults(self):
        cfg = BatchRunConfig(company_name_column="company", domain_column="domain")
        assert cfg.public_source_signal_enrichment is False
        assert cfg.public_source_signal_query == "vacancies"
        assert cfg.public_source_base_url == ""
        assert cfg.public_source_label == ""
        assert cfg.public_source_max_pages == 3

    @pytest.mark.parametrize("public_source_signal_enrichment", [False, True])
    def test_public_source_signal_enrichment_passthrough_independent_of_other_flags(
        self, public_source_signal_enrichment,
    ):
        seen = {}

        def _fake(lead_input, **kwargs):
            seen.update(kwargs)
            return _sample_result(company_name=lead_input.company_name)

        cfg = BatchRunConfig(
            company_name_column="company", domain_column="domain",
            run_mode="full", row_limit=1,
            public_source_signal_enrichment=public_source_signal_enrichment,
            public_source_signal_query="internships",
            public_source_base_url="https://example-registry.test/search",
            public_source_label="Example registry",
            public_source_max_pages=5,
        )
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake):
            run_batch_dataframe(self._df, cfg, "SERP", "ANTH", firecrawl_api_key="fc-key")

        assert seen.get("public_source_signal_enrichment") is public_source_signal_enrichment
        assert seen.get("public_source_signal_query") == "internships"
        assert seen.get("public_source_base_url") == "https://example-registry.test/search"
        assert seen.get("public_source_label") == "Example registry"
        assert seen.get("public_source_max_pages") == 5
        assert seen.get("firecrawl_api_key") == "fc-key"
        # Independent of every other opt-in flag.
        assert seen.get("legacy_enrichment_mode") is False
        assert seen.get("compose_icp_context") is False
        assert seen.get("ai_signal_scoring") is False

    def test_public_source_signal_enrichment_off_still_passes_defaults(self):
        # Off by default: missing Firecrawl key / base URL blocks only the
        # feature itself (inside the collector), never a normal run.
        seen = {}

        def _fake(lead_input, **kwargs):
            seen.update(kwargs)
            return _sample_result(company_name=lead_input.company_name)

        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="full", row_limit=1)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake):
            run_batch_dataframe(self._df, cfg, "SERP", "ANTH")
        assert seen.get("public_source_signal_enrichment") is False
        assert seen.get("public_source_base_url") == ""

    def test_continue_on_error(self):
        calls = {"n": 0}

        def _fake(lead_input, **kwargs):
            calls["n"] += 1
            if lead_input.company_name == "Beta":
                raise RuntimeError("boom")
            return _sample_result(company_name=lead_input.company_name)

        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="full", row_limit=0, continue_on_error=True)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake):
            out = run_batch_dataframe(self._df, cfg, "k1", "k2")
        assert calls["n"] == 3  # did not stop at Beta
        enriched = out["enriched_leads"]
        assert enriched.shape[0] == 3
        assert (~enriched["run_success"]).sum() == 1
        assert out["run_summary"].iloc[0]["error_count"] == 1

    def test_progress_callback_once_per_row(self):
        payloads = []

        def _fake(lead_input, **kwargs):
            return _sample_result(company_name=lead_input.company_name)

        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="full", row_limit=0)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake):
            run_batch_dataframe(self._df, cfg, "k1", "k2",
                                progress_callback=lambda p: payloads.append(p))
        assert len(payloads) == 3  # one per selected row
        last = payloads[-1]
        assert last["selected_rows"] == 3
        assert last["processed_rows"] == 3
        assert last["success_count"] == 3
        assert last["error_count"] == 0
        # secret-free payload
        for p in payloads:
            for k in p:
                assert "api_key" not in k.lower() and "serper" not in k.lower()
        assert "k1" not in str(payloads) and "k2" not in str(payloads)

    def test_progress_callback_called_for_error_rows(self):
        seen = []

        def _fake(lead_input, **kwargs):
            if lead_input.company_name == "Beta":
                raise RuntimeError("boom")
            return _sample_result(company_name=lead_input.company_name)

        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="full", row_limit=0, continue_on_error=True)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake):
            run_batch_dataframe(self._df, cfg, "k1", "k2",
                                progress_callback=lambda p: seen.append(p))
        assert len(seen) == 3
        beta = [p for p in seen if p["current_company_name"] == "Beta"][0]
        assert beta["run_success"] is False
        assert "boom" in beta["run_error"]

    def test_progress_callback_exception_does_not_break_batch(self):
        def _fake(lead_input, **kwargs):
            return _sample_result(company_name=lead_input.company_name)

        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="full", row_limit=0)

        def _boom(_payload):
            raise ValueError("callback broke")

        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake):
            out = run_batch_dataframe(self._df, cfg, "k1", "k2", progress_callback=_boom)
        # Batch still completes fully despite the broken callback.
        assert out["enriched_leads"].shape[0] == 3
        assert out["run_summary"].iloc[0]["success_count"] == 3

    def test_backward_compatible_without_callback(self):
        def _fake(lead_input, **kwargs):
            return _sample_result(company_name=lead_input.company_name)

        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="full", row_limit=1)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake):
            out = run_batch_dataframe(self._df, cfg, "k1", "k2")  # no callback
        assert out["enriched_leads"].shape[0] == 1

    def test_stop_on_error_when_configured(self):
        def _fake(lead_input, **kwargs):
            if lead_input.company_name == "Beta":
                raise RuntimeError("boom")
            return _sample_result(company_name=lead_input.company_name)

        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="full", row_limit=0, continue_on_error=False)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake):
            out = run_batch_dataframe(self._df, cfg, "k1", "k2")
        # Acme (ok) + Beta (error) then stop → 2 rows, Gamma never processed.
        assert out["enriched_leads"].shape[0] == 2
        assert out["run_summary"].iloc[0]["processed_rows"] == 2


# ---------------------------------------------------------------------------
# Lusha row-field auto-detection (Lusha enrichment plan, Stap 2 + Stap 3).
# No new BatchRunConfig column-name setting: detected directly from a row
# dict's own keys, case-insensitively. A row with none of these column
# names (any non-Lusha dataset) yields all-None -- exactly "no Lusha data
# available", the existing fallback for every LeadInput.lusha_* consumer.
# ---------------------------------------------------------------------------

class TestLushaFieldsFromRow:
    def test_detects_exact_lusha_column_names(self):
        row = {
            "Company Name": "Acme AG",
            "Company Main Industry": "Manufacturing",
            "Company Sub Industry": "Industrial Machinery & Equipment",
            "Company Description": "We make things.",
            "Company Specialties": "machinery, engineering",
            "Company Number of Employees": "201-500",
            "Company Revenue": "$50M - $100M",
        }
        out = bc._lusha_fields_from_row(row)
        assert out == {
            "lusha_main_industry": "Manufacturing",
            "lusha_sub_industry": "Industrial Machinery & Equipment",
            "lusha_description": "We make things.",
            "lusha_specialties": "machinery, engineering",
            "lusha_employees": "201-500",
            "lusha_revenue": "$50M - $100M",
        }

    def test_case_and_spacing_insensitive(self):
        row = {"company_main_industry": "Healthcare", "COMPANY SUB INDUSTRY": "Hospitals & Clinics"}
        out = bc._lusha_fields_from_row(row)
        assert out["lusha_main_industry"] == "Healthcare"
        assert out["lusha_sub_industry"] == "Hospitals & Clinics"

    def test_non_lusha_row_yields_all_none(self):
        row = {"company": "Acme", "domain": "acme.com"}
        out = bc._lusha_fields_from_row(row)
        assert out == {
            "lusha_main_industry": None, "lusha_sub_industry": None,
            "lusha_description": None, "lusha_specialties": None,
            "lusha_employees": None, "lusha_revenue": None,
        }

    def test_blank_values_treated_as_none(self):
        row = {"Company Main Industry": "", "Company Description": "   "}
        out = bc._lusha_fields_from_row(row)
        assert out["lusha_main_industry"] is None
        assert out["lusha_description"] is None

    def test_nan_string_treated_as_none(self):
        row = {"Company Main Industry": "nan"}
        out = bc._lusha_fields_from_row(row)
        assert out["lusha_main_industry"] is None


class TestLushaFieldsWiredIntoLeadInput:
    def test_run_batch_dataframe_passes_lusha_fields_to_prioritize_single_lead(self):
        df = pd.DataFrame({
            "company": ["Acme"],
            "domain": ["acme.com"],
            "Company Main Industry": ["Manufacturing"],
            "Company Sub Industry": ["Industrial Machinery & Equipment"],
        })
        seen = {}

        def _fake(lead_input, **kwargs):
            seen["lusha_main_industry"] = lead_input.lusha_main_industry
            seen["lusha_sub_industry"] = lead_input.lusha_sub_industry
            return _sample_result(company_name=lead_input.company_name)

        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="full", row_limit=1)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake):
            run_batch_dataframe(df, cfg, "SERP", "ANTH")
        assert seen["lusha_main_industry"] == "Manufacturing"
        assert seen["lusha_sub_industry"] == "Industrial Machinery & Equipment"

    def test_non_lusha_dataframe_passes_none_lusha_fields(self):
        seen = {}

        def _fake(lead_input, **kwargs):
            seen["lusha_main_industry"] = lead_input.lusha_main_industry
            return _sample_result(company_name=lead_input.company_name)

        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="full", row_limit=1)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake):
            run_batch_dataframe(
                pd.DataFrame({"company": ["Acme"], "domain": ["acme.com"]}), cfg, "SERP", "ANTH")
        assert seen["lusha_main_industry"] is None


# ---------------------------------------------------------------------------
# Shared GCS enrichment cache (opt-in) — BatchRunConfig.use_enrichment_cache.
# Regression: default (off) must never load/save a cache index or pass a
# cache_index other than None. When on: one index download per distinct
# country actually present, correct per-row index for a mixed-country
# dataset, and the run_summary cache hit/miss delta columns.
# ---------------------------------------------------------------------------

class TestEnrichmentCacheWiring:
    _df = pd.DataFrame({
        "company": ["Acme", "Beta"],
        "domain": ["acme.com", "beta.com"],
    })

    def test_default_off_never_loads_or_passes_cache_index(self):
        seen = {}

        def _fake(lead_input, **kwargs):
            seen.update(kwargs)
            return _sample_result(company_name=lead_input.company_name)

        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="full", row_limit=1)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake), \
             patch("enrichment_cache.load_cache_index") as mock_load, \
             patch("enrichment_cache.save_cache_index") as mock_save:
            run_batch_dataframe(self._df, cfg, "SERP", "ANTH")
        assert seen.get("cache_index") is None
        mock_load.assert_not_called()
        mock_save.assert_not_called()

    def test_default_off_run_summary_has_no_cache_columns(self):
        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="full", row_limit=1)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=lambda li, **k: _sample_result(company_name=li.company_name)):
            out = run_batch_dataframe(self._df, cfg, "SERP", "ANTH")
        for col in ("serper_cache_hits", "serper_cache_misses",
                    "firecrawl_cache_hits", "firecrawl_cache_misses"):
            assert col not in out["run_summary"].columns

    def test_enabled_downloads_once_per_distinct_country_and_saves_at_end(self):
        loads = []

        def _fake_load(bucket, country_slug):
            loads.append((bucket, country_slug))
            return {}

        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="full", row_limit=0,
                             use_enrichment_cache=True,
                             enrichment_cache_bucket="test-bucket")
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=lambda li, **k: _sample_result(company_name=li.company_name)), \
             patch("enrichment_cache.load_cache_index", side_effect=_fake_load), \
             patch("enrichment_cache.save_cache_index") as mock_save:
            run_batch_dataframe(self._df, cfg, "SERP", "ANTH")
        # Both rows default to the same country (config.default_input_country),
        # so exactly one distinct-country download, not one per row.
        assert loads == [("test-bucket", "italy")]
        assert mock_save.call_count >= 1

    def test_enabled_passes_loaded_index_object_to_prioritize_single_lead(self):
        sentinel_index = {"serper|acme.com|hq": {"fetched_at": "x", "response": {}}}
        seen = {}

        def _fake(lead_input, **kwargs):
            seen.update(kwargs)
            return _sample_result(company_name=lead_input.company_name)

        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="full", row_limit=1,
                             use_enrichment_cache=True,
                             enrichment_cache_bucket="test-bucket")
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake), \
             patch("enrichment_cache.load_cache_index", return_value=sentinel_index), \
             patch("enrichment_cache.save_cache_index"):
            run_batch_dataframe(self._df, cfg, "SERP", "ANTH")
        assert seen.get("cache_index") is sentinel_index

    def test_mixed_country_dataset_uses_correct_per_row_index(self):
        df = pd.DataFrame({
            "company": ["Acme IT", "Beta NL"],
            "domain": ["acme.it", "beta.nl"],
            "country": ["Italy", "Netherlands"],
        })
        indexes_by_slug = {
            "italy": {"marker": "italy-index"},
            "netherlands": {"marker": "netherlands-index"},
        }
        seen_per_company = {}

        def _fake(lead_input, **kwargs):
            seen_per_company[lead_input.company_name] = kwargs.get("cache_index")
            return _sample_result(company_name=lead_input.company_name)

        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             input_country_column="country",
                             run_mode="full", row_limit=0,
                             use_enrichment_cache=True,
                             enrichment_cache_bucket="test-bucket")
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake), \
             patch("enrichment_cache.load_cache_index",
                   side_effect=lambda bucket, slug: indexes_by_slug[slug]), \
             patch("enrichment_cache.save_cache_index"):
            run_batch_dataframe(df, cfg, "SERP", "ANTH")

        assert seen_per_company["Acme IT"] is indexes_by_slug["italy"]
        assert seen_per_company["Beta NL"] is indexes_by_slug["netherlands"]

    def test_enabled_run_summary_reports_cache_hit_miss_delta(self):
        import usage_tracker

        def _fake(lead_input, **kwargs):
            usage_tracker.record_cache_hit("serper")
            usage_tracker.record_cache_miss("serper")
            usage_tracker.record_cache_hit("firecrawl")
            return _sample_result(company_name=lead_input.company_name)

        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="full", row_limit=0,
                             use_enrichment_cache=True,
                             enrichment_cache_bucket="test-bucket")
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake), \
             patch("enrichment_cache.load_cache_index", return_value={}), \
             patch("enrichment_cache.save_cache_index"):
            out = run_batch_dataframe(self._df, cfg, "SERP", "ANTH")
        summary = out["run_summary"].iloc[0]
        assert summary["serper_cache_hits"] == 2   # one per row (Acme, Beta)
        assert summary["serper_cache_misses"] == 2
        assert summary["firecrawl_cache_hits"] == 2
        assert summary["firecrawl_cache_misses"] == 0

    def test_failing_download_falls_back_to_none_index_without_crash(self):
        seen = {}

        def _fake(lead_input, **kwargs):
            seen.update(kwargs)
            return _sample_result(company_name=lead_input.company_name)

        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="full", row_limit=1,
                             use_enrichment_cache=True,
                             enrichment_cache_bucket="test-bucket")
        # A failing download degrades to {} (enrichment_cache.load_cache_index's
        # own contract) -- confirm the batch run still completes successfully
        # with an (empty-but-not-None) cache_index rather than crashing.
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake), \
             patch("enrichment_cache.load_cache_index", return_value={}), \
             patch("enrichment_cache.save_cache_index"):
            out = run_batch_dataframe(self._df, cfg, "SERP", "ANTH")
        assert seen.get("cache_index") == {}
        assert out["enriched_leads"].shape[0] == 1


# ---------------------------------------------------------------------------
# Deep Dive (Step B, opt-in) — trigger gate, per-row wiring, Excel output,
# and score-invariance (never fed back into scoring).
# ---------------------------------------------------------------------------

class TestShouldRunDeepDive:
    def _cfg(self, **kw):
        base = dict(company_name_column="company", domain_column="domain",
                    deep_dive=True, deep_dive_min_score=8.0, deep_dive_on_foreign_hq=True)
        base.update(kw)
        return BatchRunConfig(**base)

    def test_off_by_default_regardless_of_score(self):
        cfg = self._cfg(deep_dive=False)
        result = _sample_result(final_commercial_fit_score=10.0,
                                sig_foreign_hq_score_for_next_scoring=3.0)
        assert should_run_deep_dive(result, cfg) == (False, "")

    def test_score_threshold_triggers(self):
        cfg = self._cfg()
        result = _sample_result(final_commercial_fit_score=8.5,
                                sig_foreign_hq_score_for_next_scoring=0.0)
        assert should_run_deep_dive(result, cfg) == (True, "score_threshold")

    def test_below_threshold_and_no_foreign_hq_does_not_trigger(self):
        cfg = self._cfg()
        result = _sample_result(final_commercial_fit_score=5.0,
                                sig_foreign_hq_score_for_next_scoring=0.0)
        assert should_run_deep_dive(result, cfg) == (False, "")

    def test_foreign_hq_triggers_when_enabled(self):
        cfg = self._cfg()
        result = _sample_result(final_commercial_fit_score=5.0,
                                sig_foreign_hq_score_for_next_scoring=3.0)
        assert should_run_deep_dive(result, cfg) == (True, "foreign_hq")

    def test_foreign_hq_ignored_when_disabled(self):
        cfg = self._cfg(deep_dive_on_foreign_hq=False)
        result = _sample_result(final_commercial_fit_score=5.0,
                                sig_foreign_hq_score_for_next_scoring=3.0)
        assert should_run_deep_dive(result, cfg) == (False, "")

    def test_both_conditions_true_prefers_score_threshold(self):
        cfg = self._cfg()
        result = _sample_result(final_commercial_fit_score=9.0,
                                sig_foreign_hq_score_for_next_scoring=3.0)
        assert should_run_deep_dive(result, cfg) == (True, "score_threshold")

    def test_score_none_does_not_crash(self):
        cfg = self._cfg()
        result = _sample_result(final_commercial_fit_score=None,
                                sig_foreign_hq_score_for_next_scoring=0.0)
        assert should_run_deep_dive(result, cfg) == (False, "")


class TestFlattenDeepDiveForExcel:
    def test_one_row_per_claim(self):
        dd = DeepDiveResult(
            company_name="Acme", trigger_reason="score_threshold",
            claims=[
                DeepDiveClaim(claim_id="hq_structure:1", category="hq_structure",
                              statement="s1", quote="q1", source_url="https://acme.com/about",
                              source_kind="own_domain", domain_verified=True,
                              retrieval_method="firecrawl"),
                DeepDiveClaim(claim_id="workforce:1", category="workforce",
                              statement="s2", quote="q2", source_url="https://acme.com/careers",
                              source_kind="own_domain", domain_verified=True,
                              retrieval_method="firecrawl"),
            ],
        )
        rows = flatten_deep_dive_for_excel(dd, source_index=3)
        assert len(rows) == 2
        assert rows[0]["source_index"] == 3
        assert rows[0]["company_name"] == "Acme"
        assert rows[0]["trigger_reason"] == "score_threshold"
        assert rows[0]["category"] == "hq_structure"
        assert rows[1]["category"] == "workforce"

    def test_no_claims_yields_no_rows(self):
        dd = DeepDiveResult(company_name="Acme", error="no_anthropic_api_key")
        assert flatten_deep_dive_for_excel(dd, source_index=0) == []


class TestRunBatchDeepDiveWiring:
    _df = pd.DataFrame({
        "company": ["Acme", "Beta"],
        "domain": ["acme.com", "beta.com"],
    })

    def _fake_result(self, company_name, **kw):
        return _sample_result(company_name=company_name, **kw)

    def test_deep_dive_runs_only_for_triggered_rows(self):
        def _fake_pipeline(lead_input, **kwargs):
            score = 9.0 if lead_input.company_name == "Acme" else 3.0
            return self._fake_result(lead_input.company_name,
                                     final_commercial_fit_score=score,
                                     sig_foreign_hq_score_for_next_scoring=0.0)

        dd_result = DeepDiveResult(company_name="Acme", trigger_reason="score_threshold",
                                   claims=[DeepDiveClaim(claim_id="hq_structure:1",
                                                         category="hq_structure", statement="s",
                                                         quote="q", source_url="https://acme.com")])
        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="full", deep_dive=True, deep_dive_min_score=8.0,
                             deep_dive_on_foreign_hq=False)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake_pipeline), \
             patch("lead_prioritizer_batch_core.run_deep_dive", return_value=dd_result) as m_dd:
            out = run_batch_dataframe(self._df, cfg, "S", "A")
        m_dd.assert_called_once()  # only Acme cleared the gate
        assert len(out["deep_dive"]) == 1
        assert out["deep_dive"].iloc[0]["company_name"] == "Acme"

    def test_deep_dive_disabled_by_default_never_called(self):
        def _fake_pipeline(lead_input, **kwargs):
            return self._fake_result(lead_input.company_name,
                                     final_commercial_fit_score=10.0,
                                     sig_foreign_hq_score_for_next_scoring=3.0)

        cfg = BatchRunConfig(company_name_column="company", domain_column="domain", run_mode="full")
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake_pipeline), \
             patch("lead_prioritizer_batch_core.run_deep_dive") as m_dd:
            out = run_batch_dataframe(self._df, cfg, "S", "A")
        m_dd.assert_not_called()
        assert out["deep_dive"].empty

    def test_deep_dive_receives_firecrawl_key_and_max_pages(self):
        def _fake_pipeline(lead_input, **kwargs):
            return self._fake_result(lead_input.company_name,
                                     final_commercial_fit_score=9.0,
                                     sig_foreign_hq_score_for_next_scoring=0.0)

        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="full", row_limit=1, deep_dive=True,
                             deep_dive_min_score=8.0, deep_dive_max_pages=3)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake_pipeline), \
             patch("lead_prioritizer_batch_core.run_deep_dive",
                   return_value=DeepDiveResult(company_name="Acme")) as m_dd:
            run_batch_dataframe(self._df, cfg, "S", "A", firecrawl_api_key="FC")
        assert m_dd.call_args.kwargs["firecrawl_api_key"] == "FC"
        assert m_dd.call_args.kwargs["max_pages"] == 3
        assert m_dd.call_args.kwargs["parent_domain"] is None

    @pytest.mark.parametrize("verify_quotes,auto_correct_quotes", [
        (True, True), (True, False), (False, True), (False, False),
    ])
    def test_deep_dive_receives_verify_and_auto_correct_flags(
        self, verify_quotes, auto_correct_quotes,
    ):
        def _fake_pipeline(lead_input, **kwargs):
            return self._fake_result(lead_input.company_name,
                                     final_commercial_fit_score=9.0,
                                     sig_foreign_hq_score_for_next_scoring=0.0)

        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="full", row_limit=1, deep_dive=True,
                             deep_dive_min_score=8.0, verify_quotes=verify_quotes,
                             auto_correct_quotes=auto_correct_quotes)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake_pipeline), \
             patch("lead_prioritizer_batch_core.run_deep_dive",
                   return_value=DeepDiveResult(company_name="Acme")) as m_dd:
            run_batch_dataframe(self._df, cfg, "S", "A")
        assert m_dd.call_args.kwargs["verify_quotes"] is verify_quotes
        assert m_dd.call_args.kwargs["auto_correct_quotes"] is auto_correct_quotes

    def test_deep_dive_error_never_breaks_the_batch(self):
        def _fake_pipeline(lead_input, **kwargs):
            return self._fake_result(lead_input.company_name,
                                     final_commercial_fit_score=9.0,
                                     sig_foreign_hq_score_for_next_scoring=0.0)

        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="full", deep_dive=True, deep_dive_min_score=8.0)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake_pipeline), \
             patch("lead_prioritizer_batch_core.run_deep_dive",
                   return_value=DeepDiveResult(company_name="Acme", error="deep_dive_failed: boom")):
            out = run_batch_dataframe(self._df, cfg, "S", "A")
        assert out["run_summary"].iloc[0]["error_count"] == 0
        assert len(out["enriched_leads"]) == 2


class TestDeepDiveExcelSheetConditional:
    def test_sheet_present_when_deep_dive_rows_exist(self):
        tables = {
            "enriched_leads": pd.DataFrame([{"company_name": "Acme"}]),
            "evidence": pd.DataFrame(),
            "signals": pd.DataFrame(),
            "deep_dive": pd.DataFrame([{"company_name": "Acme", "category": "hq_structure"}]),
            "run_summary": pd.DataFrame([{"processed_rows": 1}]),
        }
        data = build_excel_workbook_bytes(tables)
        import io
        xls = pd.ExcelFile(io.BytesIO(data))
        assert "Deep Dive" in xls.sheet_names

    def test_sheet_absent_when_no_deep_dive_rows(self):
        tables = {
            "enriched_leads": pd.DataFrame([{"company_name": "Acme"}]),
            "evidence": pd.DataFrame(),
            "signals": pd.DataFrame(),
            "deep_dive": pd.DataFrame(),
            "run_summary": pd.DataFrame([{"processed_rows": 1}]),
        }
        data = build_excel_workbook_bytes(tables)
        import io
        xls = pd.ExcelFile(io.BytesIO(data))
        assert "Deep Dive" not in xls.sheet_names

    def test_sheet_absent_when_deep_dive_key_missing_entirely(self):
        # Backward compatibility: callers (e.g. run_batch_foreign_hq_only)
        # that don't produce a "deep_dive" key at all must not break.
        tables = {
            "enriched_leads": pd.DataFrame([{"company_name": "Acme"}]),
            "evidence": pd.DataFrame(),
            "signals": pd.DataFrame(),
            "run_summary": pd.DataFrame([{"processed_rows": 1}]),
        }
        data = build_excel_workbook_bytes(tables)
        import io
        xls = pd.ExcelFile(io.BytesIO(data))
        assert "Deep Dive" not in xls.sheet_names


class TestDeepDiveScoreInvariance:
    """Hard Requirement #1, Part B: Deep Dive must never affect scoring."""

    _SCORE_FIELDS = (
        "final_commercial_fit_score", "commercial_tier",
        "sig_foreign_hq_score_for_next_scoring",
        "sig_international_profile_score",
    )

    def test_identical_enriched_leads_scoring_with_and_without_deep_dive(self):
        def _fake_pipeline(lead_input, **kwargs):
            return _sample_result(company_name=lead_input.company_name,
                                  final_commercial_fit_score=9.0,
                                  sig_foreign_hq_score_for_next_scoring=3.0)

        df = pd.DataFrame({"company": ["Acme"], "domain": ["acme.com"]})
        cfg_without = BatchRunConfig(company_name_column="company", domain_column="domain",
                                     run_mode="full", deep_dive=False)
        cfg_with = BatchRunConfig(company_name_column="company", domain_column="domain",
                                  run_mode="full", deep_dive=True, deep_dive_min_score=8.0)

        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake_pipeline):
            out_without = run_batch_dataframe(df, cfg_without, "S", "A")

        dd_result = DeepDiveResult(
            company_name="Acme", trigger_reason="score_threshold",
            claims=[DeepDiveClaim(claim_id="hq_structure:1", category="hq_structure",
                                  statement="s", quote="q", source_url="https://acme.com")])
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake_pipeline), \
             patch("lead_prioritizer_batch_core.run_deep_dive", return_value=dd_result):
            out_with = run_batch_dataframe(df, cfg_with, "S", "A")

        for field in self._SCORE_FIELDS:
            assert (out_without["enriched_leads"].iloc[0][field] ==
                   out_with["enriched_leads"].iloc[0][field])
        # Deep Dive itself did run and produced its own separate table.
        assert len(out_with["deep_dive"]) == 1
        assert out_without["deep_dive"].empty

    def test_deep_dive_failure_also_leaves_scoring_untouched(self):
        def _fake_pipeline(lead_input, **kwargs):
            return _sample_result(company_name=lead_input.company_name,
                                  final_commercial_fit_score=9.0,
                                  sig_foreign_hq_score_for_next_scoring=3.0)

        df = pd.DataFrame({"company": ["Acme"], "domain": ["acme.com"]})
        cfg_without = BatchRunConfig(company_name_column="company", domain_column="domain",
                                     run_mode="full", deep_dive=False)
        cfg_with = BatchRunConfig(company_name_column="company", domain_column="domain",
                                  run_mode="full", deep_dive=True, deep_dive_min_score=8.0)

        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake_pipeline):
            out_without = run_batch_dataframe(df, cfg_without, "S", "A")
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake_pipeline), \
             patch("lead_prioritizer_batch_core.run_deep_dive",
                   return_value=DeepDiveResult(company_name="Acme", error="deep_dive_failed: boom")):
            out_with = run_batch_dataframe(df, cfg_with, "S", "A")

        for field in self._SCORE_FIELDS:
            assert (out_without["enriched_leads"].iloc[0][field] ==
                   out_with["enriched_leads"].iloc[0][field])

    @pytest.mark.parametrize("verify_quotes,auto_correct_quotes", [
        (True, True), (True, False), (False, True), (False, False),
    ])
    def test_identical_scoring_regardless_of_quote_verification_settings(
        self, verify_quotes, auto_correct_quotes,
    ):
        # Runs the REAL run_deep_dive() (not mocked) so the quote-verifier
        # actually executes -- only its own network-boundary collection
        # functions are mocked -- and proves that no combination of
        # verify_quotes/auto_correct_quotes ever touches a score field.
        def _fake_pipeline(lead_input, **kwargs):
            return _sample_result(company_name=lead_input.company_name,
                                  final_commercial_fit_score=9.0,
                                  sig_foreign_hq_score_for_next_scoring=3.0)

        df = pd.DataFrame({"company": ["Acme"], "domain": ["acme.com"]})
        cfg_without = BatchRunConfig(company_name_column="company", domain_column="domain",
                                     run_mode="full", deep_dive=False)
        cfg_with = BatchRunConfig(company_name_column="company", domain_column="domain",
                                  run_mode="full", deep_dive=True, deep_dive_min_score=8.0,
                                  verify_quotes=verify_quotes,
                                  auto_correct_quotes=auto_correct_quotes)

        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake_pipeline):
            out_without = run_batch_dataframe(df, cfg_without, "S", "A")

        pages = [{"pages": [{"url": "https://acme.com/about", "title": None,
                            "text": "Acme was founded in 1990 in Munich.",
                            "source_kind": "own_domain", "retrieval_method": "firecrawl"}],
                 "pages_crawled": [], "used": True}][0]
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake_pipeline), \
             patch("deep_dive_runner._collect_pages_via_firecrawl", return_value=pages), \
             patch("deep_dive_runner._distill_claims", return_value=(
                 [DeepDiveClaim(claim_id="hq_structure:1", category="hq_structure",
                               statement="s", quote="Acme founded in 1990 in Munich",
                               source_url="https://acme.com/about")], "")):
            out_with = run_batch_dataframe(df, cfg_with, "S", "A", firecrawl_api_key="fc")

        for field in self._SCORE_FIELDS:
            assert (out_without["enriched_leads"].iloc[0][field] ==
                   out_with["enriched_leads"].iloc[0][field])
        # Deep Dive still ran and populated its own separate table.
        assert len(out_with["deep_dive"]) == 1


# ---------------------------------------------------------------------------
# Onafhankelijkheidstest A×B: rich_icp_context (Onderdeel A) and deep_dive
# (Onderdeel B) run through the real run_batch_dataframe() with all four
# on/off combinations. prioritize_single_lead is mocked but reflects
# whether compose_icp_context was actually passed through (proving Onderdeel
# A reaches the pipeline); run_deep_dive is mocked directly (it is called by
# batch-core itself, not by prioritize_single_lead). Neither feature must
# ever require or block the other.
# ---------------------------------------------------------------------------

class TestRichIcpContextAndDeepDiveIndependence:
    _df = pd.DataFrame({"company": ["Acme"], "domain": ["acme.com"]})

    def _fake_pipeline(self, lead_input, **kwargs):
        result = _sample_result(
            company_name=lead_input.company_name,
            final_commercial_fit_score=9.0,
            sig_foreign_hq_score_for_next_scoring=0.0,
        )
        if kwargs.get("compose_icp_context"):
            result.icp_buying_signals = "AI-composed buying signals."
            result.icp_context_by_ai = True
        return result

    @pytest.mark.parametrize("rich_icp_context,deep_dive", [
        (False, False), (True, False), (False, True), (True, True),
    ])
    def test_all_four_combinations_run_without_conflict(self, rich_icp_context, deep_dive):
        dd_result = DeepDiveResult(
            company_name="Acme", trigger_reason="score_threshold",
            claims=[DeepDiveClaim(claim_id="hq_structure:1", category="hq_structure",
                                  statement="s", quote="q", source_url="https://acme.com")])
        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="full", rich_icp_context=rich_icp_context,
                             deep_dive=deep_dive, deep_dive_min_score=8.0)

        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=self._fake_pipeline), \
             patch("lead_prioritizer_batch_core.run_deep_dive",
                   return_value=dd_result) as m_dd:
            out = run_batch_dataframe(self._df, cfg, "S", "A")

        assert out["run_summary"].iloc[0]["error_count"] == 0
        enriched = out["enriched_leads"].iloc[0]

        # Onderdeel A ran (or not) exactly per its own flag.
        if rich_icp_context:
            assert enriched["icp_buying_signals"] == "AI-composed buying signals."
        else:
            assert pd.isna(enriched.get("icp_buying_signals"))

        # Onderdeel B ran (or not) exactly per its own flag -- never gated
        # on rich_icp_context, and never itself gating rich_icp_context.
        if deep_dive:
            m_dd.assert_called_once()
            assert len(out["deep_dive"]) == 1
        else:
            m_dd.assert_not_called()
            assert out["deep_dive"].empty

    def test_deep_dive_never_requires_rich_icp_context(self):
        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="full", rich_icp_context=False,
                             deep_dive=True, deep_dive_min_score=8.0)
        dd_result = DeepDiveResult(company_name="Acme")
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=self._fake_pipeline), \
             patch("lead_prioritizer_batch_core.run_deep_dive", return_value=dd_result) as m_dd:
            run_batch_dataframe(self._df, cfg, "S", "A")
        m_dd.assert_called_once()

    def test_rich_icp_context_never_requires_deep_dive(self):
        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="full", rich_icp_context=True, deep_dive=False)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=self._fake_pipeline), \
             patch("lead_prioritizer_batch_core.run_deep_dive") as m_dd:
            out = run_batch_dataframe(self._df, cfg, "S", "A")
        m_dd.assert_not_called()
        assert out["enriched_leads"].iloc[0]["icp_buying_signals"] == "AI-composed buying signals."


# ---------------------------------------------------------------------------
# Excel workbook
# ---------------------------------------------------------------------------

class TestWorkbook:
    def test_returns_bytes_with_expected_sheets(self):
        tables = {
            "enriched_leads": pd.DataFrame([{"company_name": "Acme"}]),
            "evidence": pd.DataFrame([{"signal_name": "international_profile"}]),
            "signals": pd.DataFrame([{"signal_name": "international_profile"}]),
            "run_summary": pd.DataFrame([{"run_mode": "full"}]),
        }
        data = build_excel_workbook_bytes(tables)
        assert isinstance(data, (bytes, bytearray)) and len(data) > 0
        from openpyxl import load_workbook
        import io as _io
        wb = load_workbook(_io.BytesIO(data))
        assert wb.sheetnames == ["Enriched Leads", "Evidence", "Signals", "Run Summary"]

    def test_handles_empty_tables(self):
        data = build_excel_workbook_bytes({})
        from openpyxl import load_workbook
        import io as _io
        wb = load_workbook(_io.BytesIO(data))
        assert wb.sheetnames == ["Enriched Leads", "Evidence", "Signals", "Run Summary"]


# ---------------------------------------------------------------------------
# C5 batch integration (optional, country-agnostic) — adjudicator mocked
# ---------------------------------------------------------------------------

from lead_output_schema import HQDetectionResult  # noqa: E402
from lead_hq_sonnet_adjudicator import SonnetHQAdjudicationResult  # noqa: E402
from lead_prioritizer_batch_core import (  # noqa: E402
    apply_c5_adjudication,
    add_c5_summary_fields,
    foreign_hq_skip_reason,
    is_confirmed_foreign_hq_for_full_enrichment,
    row_selected_for_c5,
    run_batch_foreign_hq_only,
    run_batch_non_english_foreign_hq_only,
    run_batch_dataframe_parallel,
    run_single_batch_unit,
    split_dataframe_into_chunks,
    aggregate_parallel_chunk_progress,
    classify_parent_hq_language_market,
    resolve_parent_hq_country_for_export,
    FOREIGN_HQ_ONLY_MODE,
    NON_ENGLISH_FOREIGN_HQ_ONLY_MODE,
    MAX_PARALLEL_WORKERS,
    SUPPORTED_RUN_MODES,
)


def _enriched_rows():
    """Two rows: one HQ score-3 foreign, one score-0 domestic (manual review)."""
    return [
        {"company_name": "Nissan do Brasil", "domain": "nissan.com.br",
         "input_country": "Brazil", "sig_foreign_hq_score_for_next_scoring": 3.0,
         "needs_manual_review": False, "hq_positive_score_suppressed_for_review": "No",
         "hq_detected_country": "Japan", "ai_parent_hq_country": "Japan"},
        {"company_name": "Empresa Uruguaya", "domain": "empresa.com.uy",
         "input_country": "Uruguay", "sig_foreign_hq_score_for_next_scoring": 0.0,
         "needs_manual_review": True, "hq_positive_score_suppressed_for_review": "No",
         "hq_detected_country": "", "ai_parent_hq_country": ""},
    ]


def _mk_result(adjudication="unclear", confidence="Low", target="unclear",
               call_success=True, error="", model="claude-sonnet-5"):
    return SonnetHQAdjudicationResult(
        adjudication=adjudication, confidence=confidence, target_company_match=target,
        parent_company="", parent_hq_country="", parent_hq_city="",
        reason="r", model=model, call_attempted=True,
        call_success=call_success, error=error)


def _patch_adjudicator(result_or_fn):
    # apply_c5_adjudication reuses adjudicate_row from the probe, which calls
    # adjudicate_hq_with_sonnet in the probe's namespace.
    if callable(result_or_fn):
        return patch("run_hq_sonnet_adjudication_probe.adjudicate_hq_with_sonnet",
                     side_effect=result_or_fn)
    return patch("run_hq_sonnet_adjudication_probe.adjudicate_hq_with_sonnet",
                 return_value=result_or_fn)


class TestC5RowSelection:
    def test_scope_variants(self):
        rows = _enriched_rows()
        assert [row_selected_for_c5(r, "all_rows") for r in rows] == [True, True]
        assert [row_selected_for_c5(r, "score_3_only") for r in rows] == [True, False]
        assert [row_selected_for_c5(r, "score_3_or_manual_review") for r in rows] == [True, True]
        assert [row_selected_for_c5(r, "manual_review_or_suppressed") for r in rows] == [False, True]

    def test_suppressed_selected(self):
        r = {"sig_foreign_hq_score_for_next_scoring": 0.0, "needs_manual_review": False,
             "hq_positive_score_suppressed_for_review": "Yes"}
        assert row_selected_for_c5(r, "manual_review_or_suppressed") is True


class TestC5Disabled:
    def test_run_batch_alone_has_no_c5_columns(self):
        df = pd.DataFrame({"company": ["A"], "domain": ["a.com"]})
        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="hq_only", row_limit=1)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   return_value=_sample_result()):
            out = run_batch_dataframe(df, cfg, "k1", "k2")
        assert not any(c.startswith("c5_") for c in out["enriched_leads"].columns)


class TestC5AppendOnly:
    def test_scores_unchanged_fields_added(self):
        rows = _enriched_rows()
        with _patch_adjudicator(_mk_result(adjudication="domestic_confirmed",
                                           confidence="High", target="yes")):
            out, counts = apply_c5_adjudication(
                rows, anthropic_api_key="k", model_used="claude-sonnet-5",
                model_tier="sonnet", scoring_behavior="append_only", scope="all_rows")
        # original scores untouched
        assert out[0]["sig_foreign_hq_score_for_next_scoring"] == 3.0
        assert out[1]["sig_foreign_hq_score_for_next_scoring"] == 0.0
        assert out[0]["needs_manual_review"] is False
        # C5 fields present
        assert out[0]["c5_adjudication"] == "domestic_confirmed"
        assert out[0]["c5_call_attempted"] is True
        assert "c5_possible_foreign_parent_for_review" in out[0]
        assert counts["c5_rows_attempted"] == 2


class TestC5RecomputesHqLocationSummary:
    """apply_c5_adjudication recomputes hq_location_summary so C5's richer
    parent HQ (C5 > AI > detected) takes priority."""

    def _foreign_row(self):
        return {
            "company_name": "Fujifilm NL", "domain": "fujifilmtilburg.nl",
            "input_country": "Netherlands",
            "sig_foreign_hq_score_for_next_scoring": 3.0,
            "needs_manual_review": False,
            "hq_positive_score_suppressed_for_review": "No",
            "hq_structure_type": "foreign_parent",
            "foreign_hq_simple": True,
            # Base summary from AI fields (no city) before C5.
            "ai_parent_hq_country": "Japan", "ai_parent_hq_city": "",
            "hq_detected_country": "Japan", "hq_detected_city": "",
            "hq_location_summary": "Parent company headquarters: Japan",
        }

    def test_c5_parent_city_upgrades_summary(self):
        result = _mk_result(adjudication="foreign_parent_confirmed",
                            confidence="High", target="yes")
        result.parent_hq_country = "Japan"
        result.parent_hq_city = "Tokyo"
        with _patch_adjudicator(result):
            out, _ = apply_c5_adjudication(
                [self._foreign_row()], anthropic_api_key="k",
                model_used="claude-sonnet-5", model_tier="sonnet",
                scoring_behavior="append_only", scope="all_rows")
        assert out[0]["hq_location_summary"] == (
            "Parent company headquarters: Tokyo, Japan")

    def test_blank_row_keeps_base_summary(self):
        # A row not selected for C5 keeps whatever summary it already had.
        row = self._foreign_row()
        with _patch_adjudicator(_mk_result()):
            out, _ = apply_c5_adjudication(
                [row], anthropic_api_key="k", model_used="claude-sonnet-5",
                model_tier="sonnet", scoring_behavior="append_only",
                scope="manual_review_or_suppressed")  # this row is not selected
        assert out[0]["hq_location_summary"] == "Parent company headquarters: Japan"


class TestC5Conservative:
    def test_confirms_old_score_3(self):
        rows = [_enriched_rows()[0]]  # score 3
        with _patch_adjudicator(_mk_result("foreign_parent_confirmed", "High", "yes")):
            out, counts = apply_c5_adjudication(
                rows, anthropic_api_key="k", model_used="m", model_tier="sonnet",
                scoring_behavior="conservative_adjustment", scope="all_rows")
        assert out[0]["sig_foreign_hq_score_for_next_scoring"] == 3.0
        assert counts["c5_downgraded_score_3_count"] == 0

    def test_downgrades_old_score_3_when_domestic(self):
        rows = [_enriched_rows()[0]]
        with _patch_adjudicator(_mk_result("domestic_confirmed", "High", "yes")):
            out, counts = apply_c5_adjudication(
                rows, anthropic_api_key="k", model_used="m", model_tier="sonnet",
                scoring_behavior="conservative_adjustment", scope="all_rows")
        assert out[0]["sig_foreign_hq_score_for_next_scoring"] == 0.0
        assert out[0]["needs_manual_review"] is True
        assert counts["c5_downgraded_score_3_count"] == 1
        assert "downgraded" in out[0]["hq_reason"].lower()

    def test_does_not_auto_upgrade_old_score_0(self):
        rows = [_enriched_rows()[1]]  # score 0
        with _patch_adjudicator(_mk_result("foreign_parent_confirmed", "High", "yes")):
            out, counts = apply_c5_adjudication(
                rows, anthropic_api_key="k", model_used="m", model_tier="sonnet",
                scoring_behavior="conservative_adjustment", scope="all_rows")
        assert out[0]["sig_foreign_hq_score_for_next_scoring"] == 0.0
        assert out[0]["c5_possible_foreign_parent_for_review"] is True
        assert out[0]["needs_manual_review"] is True
        assert counts["c5_possible_foreign_parent_for_review_count"] == 1

    def test_failure_old_3_downgrades_old_0_stays_with_error(self):
        rows = _enriched_rows()  # [score3, score0]
        fail = _mk_result(adjudication="unclear", call_success=False,
                          error="sonnet_parse_failed")
        with _patch_adjudicator(fail):
            out, counts = apply_c5_adjudication(
                rows, anthropic_api_key="k", model_used="m", model_tier="sonnet",
                scoring_behavior="conservative_adjustment", scope="all_rows")
        # old score 3 → 0 + review
        assert out[0]["sig_foreign_hq_score_for_next_scoring"] == 0.0
        assert out[0]["needs_manual_review"] is True
        assert out[0]["c5_error"] == "sonnet_parse_failed"
        # old score 0 → stays 0 + review, error present
        assert out[1]["sig_foreign_hq_score_for_next_scoring"] == 0.0
        assert out[1]["needs_manual_review"] is True
        assert out[1]["c5_error"] == "sonnet_parse_failed"
        assert counts["c5_error_count"] == 2


class TestC5ScopeFiltering:
    def test_score_3_only_sends_one(self):
        rows = _enriched_rows()
        with _patch_adjudicator(_mk_result()):
            out, counts = apply_c5_adjudication(
                rows, anthropic_api_key="k", model_used="m", model_tier="sonnet",
                scope="score_3_only")
        assert out[0]["c5_call_attempted"] is True
        assert out[1]["c5_call_attempted"] is False   # not sent → blank
        assert out[1]["c5_possible_foreign_parent_for_review"] is False
        assert counts["c5_rows_attempted"] == 1

    def test_all_rows_sends_all(self):
        rows = _enriched_rows()
        with _patch_adjudicator(_mk_result()):
            out, counts = apply_c5_adjudication(
                rows, anthropic_api_key="k", model_used="m", model_tier="sonnet",
                scope="all_rows")
        assert counts["c5_rows_attempted"] == 2


class TestC5CountryAgnostic:
    def test_input_country_passed_through(self):
        rows = _enriched_rows()  # Brazil + Uruguay
        seen = []

        def _capture(**kwargs):
            seen.append(kwargs.get("input_country"))
            return _mk_result()

        with _patch_adjudicator(_capture):
            apply_c5_adjudication(
                rows, anthropic_api_key="k", model_used="m", model_tier="sonnet",
                scope="all_rows")
        assert "Brazil" in seen
        assert "Uruguay" in seen
        # no hardcoded country substituted
        assert set(seen) == {"Brazil", "Uruguay"}


class TestC5RunSummary:
    def test_summary_includes_settings_and_counts(self):
        base = build_run_summary_dataframe(
            BatchRunConfig(company_name_column="c", domain_column="d",
                           run_mode="hq_only"),
            total_input_rows=2, selected_rows=2, processed_rows=2,
            success_count=2, error_count=0)
        counts = {
            "c5_rows_attempted": 2, "c5_success_count": 2, "c5_error_count": 0,
            "c5_foreign_parent_confirmed_count": 1, "c5_domestic_confirmed_count": 1,
            "c5_unclear_count": 0, "c5_recommended_score_3_count": 1,
            "c5_possible_foreign_parent_for_review_count": 0,
            "c5_downgraded_score_3_count": 0,
        }
        summ = add_c5_summary_fields(
            base, c5_enabled=True, c5_scoring_behavior="conservative_adjustment",
            c5_scope="all_rows", c5_model_tier="sonnet", c5_model_used="claude-sonnet-5",
            counts=counts)
        rec = summ.iloc[0].to_dict()
        assert rec["c5_enabled"] is True
        assert rec["c5_scoring_behavior"] == "conservative_adjustment"
        assert rec["c5_scope"] == "all_rows"
        assert rec["c5_model_tier"] == "sonnet"
        assert rec["c5_model_used"] == "claude-sonnet-5"
        assert rec["c5_rows_attempted"] == 2
        assert rec["c5_foreign_parent_confirmed_count"] == 1
        assert rec["c5_downgraded_score_3_count"] == 0


# ---------------------------------------------------------------------------
# "Full enrichment, confirmed foreign-HQ only" batch mode
# ---------------------------------------------------------------------------

def _fake_prioritize_for_foreign_hq(scores: dict):
    """side_effect for prioritize_single_lead: hq_only calls report `scores`;
    full calls (run_full_v2_pipeline=True) report the same score plus a
    final_commercial_fit_score marker so tests can detect enrichment ran."""
    def _fake(lead_input, **kwargs):
        score = scores.get(lead_input.company_name, 0.0)
        if kwargs.get("run_full_v2_pipeline"):
            return _sample_result(
                company_name=lead_input.company_name, domain=lead_input.domain,
                input_country=lead_input.input_country or "Brazil",
                sig_foreign_hq_score_for_next_scoring=score,
                final_commercial_fit_score=99.0,
            )
        return _sample_result(
            company_name=lead_input.company_name, domain=lead_input.domain,
            input_country=lead_input.input_country or "Brazil",
            sig_foreign_hq_score_for_next_scoring=score,
            final_commercial_fit_score=None,
            evidence_items=[], signals=[],
        )
    return _fake


class TestConfirmedForeignHQGate:
    """Unit tests for is_confirmed_foreign_hq_for_full_enrichment /
    foreign_hq_skip_reason — the phase-3 eligibility gate."""

    def _c5_row(self, **kw):
        row = {
            "sig_foreign_hq_score_for_next_scoring": 0.0,
            "c5_call_attempted": True,
            "c5_call_success": True,
            "c5_adjudication": "foreign_parent_confirmed",
            "c5_target_company_match": "yes",
            "c5_recommended_hq_score": 3.0,
        }
        row.update(kw)
        return row

    def test_c5_confirmed_target_yes_score_3_is_eligible(self):
        assert is_confirmed_foreign_hq_for_full_enrichment(
            self._c5_row(), c5_enabled=True) is True

    def test_target_match_boolean_true_is_eligible(self):
        row = self._c5_row(c5_target_company_match=True)
        assert is_confirmed_foreign_hq_for_full_enrichment(row, c5_enabled=True) is True

    def test_uppercase_yes_and_string_score_are_normalized(self):
        row = self._c5_row(c5_target_company_match="YES",
                           c5_recommended_hq_score="3")
        assert is_confirmed_foreign_hq_for_full_enrichment(row, c5_enabled=True) is True

    def test_missing_recommended_score_column_does_not_block(self):
        row = self._c5_row()
        del row["c5_recommended_hq_score"]
        assert is_confirmed_foreign_hq_for_full_enrichment(row, c5_enabled=True) is True

    def test_recommended_score_other_than_3_blocks(self):
        row = self._c5_row(c5_recommended_hq_score=0.0)
        assert is_confirmed_foreign_hq_for_full_enrichment(row, c5_enabled=True) is False

    def test_target_mismatch_not_eligible(self):
        row = self._c5_row(c5_target_company_match="no")
        assert is_confirmed_foreign_hq_for_full_enrichment(row, c5_enabled=True) is False
        assert foreign_hq_skip_reason(row, c5_enabled=True) == "C5 target mismatch"

    def test_c5_unclear_not_eligible(self):
        row = self._c5_row(c5_adjudication="unclear", c5_target_company_match="unclear",
                           c5_recommended_hq_score=0.0)
        assert is_confirmed_foreign_hq_for_full_enrichment(row, c5_enabled=True) is False
        assert foreign_hq_skip_reason(row, c5_enabled=True) == "C5 unclear"

    def test_c5_error_not_eligible(self):
        row = self._c5_row(c5_call_success=False, c5_error="boom")
        assert is_confirmed_foreign_hq_for_full_enrichment(row, c5_enabled=True) is False
        assert foreign_hq_skip_reason(row, c5_enabled=True) == "C5 error"

    def test_c5_disabled_ignores_c5_columns(self):
        # Even a fully C5-confirmed row is not eligible via C5 when disabled;
        # only the final HQ score counts (previous behavior preserved).
        assert is_confirmed_foreign_hq_for_full_enrichment(
            self._c5_row(), c5_enabled=False) is False
        assert is_confirmed_foreign_hq_for_full_enrichment(
            {"sig_foreign_hq_score_for_next_scoring": 3.0}, c5_enabled=False) is True
        assert foreign_hq_skip_reason(self._c5_row(), c5_enabled=False) == \
            "Not confirmed foreign HQ"

    def test_final_score_3_is_always_eligible(self):
        row = self._c5_row(sig_foreign_hq_score_for_next_scoring="3")
        assert is_confirmed_foreign_hq_for_full_enrichment(row, c5_enabled=True) is True

    def test_c4_suppressed_then_c5_confirmed_is_eligible(self):
        row = self._c5_row(hq_positive_score_suppressed_for_review="Yes",
                           needs_manual_review=True)
        assert is_confirmed_foreign_hq_for_full_enrichment(row, c5_enabled=True) is True

    def test_c5_not_attempted_row_stays_conservative(self):
        row = {
            "sig_foreign_hq_score_for_next_scoring": 0.0,
            "c5_call_attempted": False, "c5_call_success": False,
            "c5_adjudication": "", "c5_target_company_match": "",
            "c5_recommended_hq_score": None,
        }
        assert is_confirmed_foreign_hq_for_full_enrichment(row, c5_enabled=True) is False
        assert foreign_hq_skip_reason(row, c5_enabled=True) == "Not confirmed foreign HQ"


class TestForeignHQOnlyMode:
    _df = pd.DataFrame({
        "company": ["Confirmed Foreign Co", "Local Domestic Co"],
        "domain": ["confirmed.com.br", "local.com.br"],
    })
    _cfg = BatchRunConfig(
        company_name_column="company", domain_column="domain",
        run_mode=FOREIGN_HQ_ONLY_MODE, row_limit=2,
        default_input_country="Brazil",
    )

    def test_skips_rows_with_final_hq_score_0(self):
        fake = _fake_prioritize_for_foreign_hq(
            {"Confirmed Foreign Co": 3.0, "Local Domestic Co": 0.0})
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=fake):
            out = run_batch_foreign_hq_only(self._df, self._cfg, "S", "A")
        rows = out["enriched_leads"].set_index("company_name")
        assert rows.loc["Local Domestic Co", "enrichment_skipped"] == True  # noqa: E712
        assert rows.loc["Local Domestic Co", "enrichment_skip_reason"] == "Not confirmed foreign HQ"
        assert pd.isna(rows.loc["Local Domestic Co", "final_commercial_fit_score"])

    def test_enriches_rows_with_final_hq_score_3(self):
        fake = _fake_prioritize_for_foreign_hq(
            {"Confirmed Foreign Co": 3.0, "Local Domestic Co": 0.0})
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=fake):
            out = run_batch_foreign_hq_only(self._df, self._cfg, "S", "A")
        rows = out["enriched_leads"].set_index("company_name")
        assert rows.loc["Confirmed Foreign Co", "enrichment_skipped"] == False  # noqa: E712
        assert rows.loc["Confirmed Foreign Co", "enrichment_skip_reason"] == ""
        assert rows.loc["Confirmed Foreign Co", "final_commercial_fit_score"] == 99.0

    def test_c5_conservative_downgrade_is_skipped(self):
        # Old HQ score 3, but C5 does not confirm → conservative_adjustment
        # downgrades to 0 → the row must be skipped, not enriched.
        fake = _fake_prioritize_for_foreign_hq(
            {"Confirmed Foreign Co": 3.0, "Local Domestic Co": 0.0})
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=fake), \
             _patch_adjudicator(_mk_result("domestic_confirmed", "High", "yes")):
            out = run_batch_foreign_hq_only(
                self._df, self._cfg, "S", "A",
                c5_enabled=True, c5_scoring_behavior="conservative_adjustment",
                c5_scope="all_rows", c5_model_used="claude-sonnet-5", c5_model_tier="sonnet",
            )
        rows = out["enriched_leads"].set_index("company_name")
        assert rows.loc["Confirmed Foreign Co", "sig_foreign_hq_score_for_next_scoring"] == 0.0
        assert rows.loc["Confirmed Foreign Co", "enrichment_skipped"] == True  # noqa: E712
        assert rows.loc["Confirmed Foreign Co", "enrichment_skip_reason"] == "Not confirmed foreign HQ"
        assert pd.isna(rows.loc["Confirmed Foreign Co", "final_commercial_fit_score"])
        summary = out["run_summary"].iloc[0].to_dict()
        assert summary["c5_downgraded_score_3_count"] == 1

    def test_c5_confirmed_score0_row_is_now_enriched(self):
        # Old HQ score 0; C5 confirms a foreign parent (target match yes,
        # High confidence → recommended score 3). The score is still never
        # auto-upgraded, but the row IS now eligible for full enrichment —
        # this was the Brazil bug where 22 C5-confirmed rows were skipped
        # with "Not confirmed foreign HQ".
        fake = _fake_prioritize_for_foreign_hq(
            {"Confirmed Foreign Co": 3.0, "Local Domestic Co": 0.0})
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=fake), \
             _patch_adjudicator(_mk_result("foreign_parent_confirmed", "High", "yes")):
            out = run_batch_foreign_hq_only(
                self._df, self._cfg, "S", "A",
                c5_enabled=True, c5_scoring_behavior="conservative_adjustment",
                c5_scope="all_rows", c5_model_used="claude-sonnet-5", c5_model_tier="sonnet",
            )
        rows = out["enriched_leads"].set_index("company_name")
        row = rows.loc["Local Domestic Co"]
        assert row["enrichment_skipped"] == False  # noqa: E712
        assert row["enrichment_skip_reason"] == ""
        assert row["full_enrichment_gate_reason"] == \
            "Confirmed by C5 foreign-parent adjudication"
        assert row["final_commercial_fit_score"] == 99.0
        # The screening score itself is never rewritten by the gate.
        assert row["sig_foreign_hq_score_for_next_scoring"] == 0.0
        summary = out["run_summary"].iloc[0].to_dict()
        assert summary["confirmed_foreign_hq_count"] == 2
        assert summary["full_enrichment_attempted_count"] == 2

    def test_run_summary_counts_reconcile_when_c5_confirms_all(self):
        # Brazil symptom was c5_foreign_parent_confirmed_count (110) >
        # confirmed_foreign_hq_count (88). With the fixed gate, all otherwise
        # valid C5-confirmed rows are eligible, so the counts reconcile.
        fake = _fake_prioritize_for_foreign_hq(
            {"Confirmed Foreign Co": 3.0, "Local Domestic Co": 0.0})
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=fake), \
             _patch_adjudicator(_mk_result("foreign_parent_confirmed", "High", "yes")):
            out = run_batch_foreign_hq_only(
                self._df, self._cfg, "S", "A",
                c5_enabled=True, c5_scoring_behavior="append_only",
                c5_scope="all_rows", c5_model_used="claude-sonnet-5", c5_model_tier="sonnet",
            )
        summary = out["run_summary"].iloc[0].to_dict()
        assert summary["c5_foreign_parent_confirmed_count"] == 2
        assert summary["confirmed_foreign_hq_count"] == 2
        assert summary["full_enrichment_attempted_count"] == 2
        assert summary["full_enrichment_skipped_count"] == 0
        # Scoring columns are untouched by the gate (append_only never
        # rewrites the screening score).
        rows = out["enriched_leads"].set_index("company_name")
        assert rows.loc["Confirmed Foreign Co", "sig_foreign_hq_score_for_next_scoring"] == 3.0
        assert rows.loc["Local Domestic Co", "sig_foreign_hq_score_for_next_scoring"] == 0.0
        assert rows.loc["Confirmed Foreign Co", "full_enrichment_gate_reason"] == \
            "Confirmed foreign HQ (final HQ score 3)"

    def test_c5_unclear_and_error_rows_get_specific_skip_reasons(self):
        def _adjudicate(**kwargs):
            if kwargs.get("company_name") == "Confirmed Foreign Co":
                return _mk_result("unclear", "Low", "unclear")
            return _mk_result("unclear", "Low", "unclear",
                              call_success=False, error="api down")

        fake = _fake_prioritize_for_foreign_hq(
            {"Confirmed Foreign Co": 0.0, "Local Domestic Co": 0.0})
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=fake), \
             _patch_adjudicator(_adjudicate):
            out = run_batch_foreign_hq_only(
                self._df, self._cfg, "S", "A",
                c5_enabled=True, c5_scoring_behavior="append_only",
                c5_scope="all_rows", c5_model_used="claude-sonnet-5", c5_model_tier="sonnet",
            )
        rows = out["enriched_leads"].set_index("company_name")
        assert rows.loc["Confirmed Foreign Co", "enrichment_skipped"] == True  # noqa: E712
        assert rows.loc["Confirmed Foreign Co", "enrichment_skip_reason"] == "C5 unclear"
        assert rows.loc["Local Domestic Co", "enrichment_skip_reason"] == "C5 error"
        summary = out["run_summary"].iloc[0].to_dict()
        assert summary["confirmed_foreign_hq_count"] == 0
        assert summary["full_enrichment_attempted_count"] == 0

    def test_output_columns_present_for_both_confirmed_and_skipped(self):
        fake = _fake_prioritize_for_foreign_hq(
            {"Confirmed Foreign Co": 3.0, "Local Domestic Co": 0.0})
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=fake):
            out = run_batch_foreign_hq_only(self._df, self._cfg, "S", "A")
        assert "enrichment_skipped" in out["enriched_leads"].columns
        assert "enrichment_skip_reason" in out["enriched_leads"].columns
        assert len(out["enriched_leads"]) == 2  # both confirmed and skipped rows present

    def test_run_summary_counts_and_mode(self):
        fake = _fake_prioritize_for_foreign_hq(
            {"Confirmed Foreign Co": 3.0, "Local Domestic Co": 0.0})
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=fake):
            out = run_batch_foreign_hq_only(self._df, self._cfg, "S", "A")
        summary = out["run_summary"].iloc[0].to_dict()
        assert summary["run_mode"] == FOREIGN_HQ_ONLY_MODE
        assert summary["total_processed_rows"] == 2
        assert summary["full_enrichment_attempted_count"] == 1
        assert summary["full_enrichment_skipped_count"] == 1
        assert summary["confirmed_foreign_hq_count"] == 1
        # C5 settings/counts columns are always present (enabled or not).
        assert summary["c5_enabled"] is False

    def test_country_agnostic_brazil_and_uruguay(self):
        df = pd.DataFrame({
            "company": ["Brazil Co", "Uruguay Co"],
            "domain": ["brco.com.br", "uyco.com.uy"],
            "country": ["Brazil", "Uruguay"],
        })
        cfg = BatchRunConfig(
            company_name_column="company", domain_column="domain",
            input_country_column="country", run_mode=FOREIGN_HQ_ONLY_MODE, row_limit=2,
        )
        seen_countries = []

        def _fake(lead_input, **kwargs):
            seen_countries.append(lead_input.input_country)
            score = 3.0 if lead_input.company_name == "Brazil Co" else 0.0
            if kwargs.get("run_full_v2_pipeline"):
                return _sample_result(
                    company_name=lead_input.company_name, domain=lead_input.domain,
                    input_country=lead_input.input_country,
                    sig_foreign_hq_score_for_next_scoring=score,
                    final_commercial_fit_score=99.0)
            return _sample_result(
                company_name=lead_input.company_name, domain=lead_input.domain,
                input_country=lead_input.input_country,
                sig_foreign_hq_score_for_next_scoring=score,
                final_commercial_fit_score=None, evidence_items=[], signals=[])

        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake):
            run_batch_foreign_hq_only(df, cfg, "S", "A")
        assert "Brazil" in seen_countries
        assert "Uruguay" in seen_countries
        # no hardcoded country substituted
        assert set(seen_countries) == {"Brazil", "Uruguay"}

    def test_existing_run_modes_unchanged(self):
        # The new mode is deliberately excluded from SUPPORTED_RUN_MODES / the
        # CLI's --mode choices; resolve_pipeline_flags for existing modes and
        # run_batch_dataframe's own output shape are both untouched.
        assert FOREIGN_HQ_ONLY_MODE not in SUPPORTED_RUN_MODES
        assert resolve_pipeline_flags("full")["run_full_v2_pipeline"] is True
        assert all(v is False for v in resolve_pipeline_flags("hq_only").values())

        df = pd.DataFrame({"company": ["Acme"], "domain": ["acme.com"]})
        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="hq_only", row_limit=1)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   return_value=_sample_result()):
            out = run_batch_dataframe(df, cfg, "S", "A")
        assert "enrichment_skipped" not in out["enriched_leads"].columns
        assert "full_enrichment_attempted_count" not in out["run_summary"].columns

    # ── Progress reporting (phase-aware payloads) ────────────────────────────

    def test_progress_phases_nonzero_totals_and_company(self):
        payloads = []
        fake = _fake_prioritize_for_foreign_hq(
            {"Confirmed Foreign Co": 3.0, "Local Domestic Co": 0.0})
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=fake):
            run_batch_foreign_hq_only(self._df, self._cfg, "S", "A",
                                      progress_callback=payloads.append)
        assert payloads, "expected progress payloads"
        assert all("phase" in p for p in payloads)
        # Phase 1: HQ screening over ALL selected rows — never 0/0.
        p1 = [p for p in payloads if p["phase"] == 1]
        assert p1
        assert all(p["phase_total"] == 2 for p in p1)
        assert all(p["phase_label"] == "HQ screening" for p in p1)
        assert p1[-1]["phase_processed"] == 2
        assert all(p.get("current_company_name") for p in p1)
        # Phase 2 absent when C5 is disabled.
        assert not [p for p in payloads if p["phase"] == 2]
        # Phase 3: total is the confirmed count (1), not a running counter.
        p3 = [p for p in payloads if p["phase"] == 3]
        assert p3
        assert all(p["phase_total"] == 1 for p in p3)
        assert p3[-1]["phase_processed"] == 1
        assert p3[-1]["current_company_name"] == "Confirmed Foreign Co"
        assert p3[-1]["success_count"] == 1
        assert p3[-1]["error_count"] == 0

    def test_progress_phase2_present_when_c5_enabled(self):
        payloads = []
        fake = _fake_prioritize_for_foreign_hq(
            {"Confirmed Foreign Co": 3.0, "Local Domestic Co": 0.0})
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=fake), \
             _patch_adjudicator(_mk_result("foreign_parent_confirmed", "High", "yes")):
            run_batch_foreign_hq_only(
                self._df, self._cfg, "S", "A",
                c5_enabled=True, c5_scoring_behavior="append_only",
                c5_scope="all_rows", c5_model_used="claude-sonnet-5", c5_model_tier="sonnet",
                progress_callback=payloads.append)
        p2 = [p for p in payloads if p["phase"] == 2]
        assert p2
        assert all(p["phase_label"] == "C5 adjudication" for p in p2)
        assert all(p["phase_total"] == 2 for p in p2)  # scope all_rows → both rows
        assert p2[-1]["phase_processed"] == 2
        assert all(p.get("current_company_name") for p in p2)

    def test_progress_callback_exception_does_not_break_run(self):
        def _boom(_payload):
            raise RuntimeError("ui broke")

        fake = _fake_prioritize_for_foreign_hq(
            {"Confirmed Foreign Co": 3.0, "Local Domestic Co": 0.0})
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=fake):
            out = run_batch_foreign_hq_only(self._df, self._cfg, "S", "A",
                                            progress_callback=_boom)
        assert len(out["enriched_leads"]) == 2


class TestForeignHqOnlyEnrichmentCacheWiring:
    """run_batch_foreign_hq_only's Phase 1 (hq_only run_batch_dataframe) and
    Phase 3 (_run_gated_full_enrichment) each run their own independent
    cache load/save cycle; the combined run_summary counts must cover BOTH
    phases, not just Phase 3."""

    _df = pd.DataFrame({
        "company": ["Confirmed Foreign Co", "Local Domestic Co"],
        "domain": ["confirmed.com.br", "local.com.br"],
    })
    _cfg = BatchRunConfig(
        company_name_column="company", domain_column="domain",
        run_mode=FOREIGN_HQ_ONLY_MODE, row_limit=2,
        default_input_country="Brazil",
        use_enrichment_cache=True, enrichment_cache_bucket="test-bucket",
    )

    def test_default_off_never_touches_cache(self):
        cfg = BatchRunConfig(
            company_name_column="company", domain_column="domain",
            run_mode=FOREIGN_HQ_ONLY_MODE, row_limit=2,
            default_input_country="Brazil",
        )
        fake = _fake_prioritize_for_foreign_hq(
            {"Confirmed Foreign Co": 3.0, "Local Domestic Co": 0.0})
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=fake), \
             patch("enrichment_cache.load_cache_index") as mock_load, \
             patch("enrichment_cache.save_cache_index") as mock_save:
            out = run_batch_foreign_hq_only(self._df, cfg, "S", "A")
        mock_load.assert_not_called()
        mock_save.assert_not_called()
        for col in ("serper_cache_hits", "serper_cache_misses",
                    "firecrawl_cache_hits", "firecrawl_cache_misses"):
            assert col not in out["run_summary"].columns

    def test_enabled_downloads_and_saves_for_both_phases(self):
        fake = _fake_prioritize_for_foreign_hq(
            {"Confirmed Foreign Co": 3.0, "Local Domestic Co": 0.0})
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=fake), \
             patch("enrichment_cache.load_cache_index", return_value={}) as mock_load, \
             patch("enrichment_cache.save_cache_index") as mock_save:
            out = run_batch_foreign_hq_only(self._df, self._cfg, "S", "A")
        # At least once for Phase 1 (hq_only run_batch_dataframe) and once
        # for Phase 3 (_run_gated_full_enrichment) — both run independently.
        assert mock_load.call_count >= 2
        assert mock_save.call_count >= 2
        assert len(out["enriched_leads"]) == 2

    def test_enabled_run_summary_reports_combined_hit_miss_delta(self):
        import usage_tracker

        def _fake(lead_input, **kwargs):
            usage_tracker.record_cache_miss("serper")
            score = 3.0 if lead_input.company_name == "Confirmed Foreign Co" else 0.0
            if kwargs.get("run_full_v2_pipeline"):
                return _sample_result(
                    company_name=lead_input.company_name, domain=lead_input.domain,
                    input_country=lead_input.input_country or "Brazil",
                    sig_foreign_hq_score_for_next_scoring=score,
                    final_commercial_fit_score=99.0,
                )
            return _sample_result(
                company_name=lead_input.company_name, domain=lead_input.domain,
                input_country=lead_input.input_country or "Brazil",
                sig_foreign_hq_score_for_next_scoring=score,
                final_commercial_fit_score=None, evidence_items=[], signals=[],
            )

        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake), \
             patch("enrichment_cache.load_cache_index", return_value={}), \
             patch("enrichment_cache.save_cache_index"):
            out = run_batch_foreign_hq_only(self._df, self._cfg, "S", "A")
        summary = out["run_summary"].iloc[0]
        # 2 rows in Phase 1 (hq_only) + 1 eligible row in Phase 3 (full) = 3.
        assert summary["serper_cache_misses"] == 3


# ---------------------------------------------------------------------------
# BatchRunConfig.gate_full_enrichment_on_foreign_hq: opt-in gating INSIDE the
# regular run_batch_dataframe path, sharing _run_gated_full_enrichment with
# run_batch_foreign_hq_only. The extraction itself is regression-tested by
# every existing TestForeignHQOnlyMode / TestConfirmedForeignHQGate /
# TestNonEnglishForeignHqOnlyMode* / TestExistingForeignHQOnlyModeUnaffected
# test above still passing unchanged (run_batch_foreign_hq_only now calls the
# shared helper internally instead of duplicating the loop).
# ---------------------------------------------------------------------------

class TestGatedFullEnrichmentInRunBatchDataframe:
    _df = pd.DataFrame({
        "company": ["Confirmed Foreign Co", "Local Domestic Co"],
        "domain": ["confirmed.com.br", "local.com.br"],
    })

    def _cfg(self, **overrides):
        base = dict(
            company_name_column="company", domain_column="domain",
            run_mode="full", row_limit=2, default_input_country="Brazil",
        )
        base.update(overrides)
        return BatchRunConfig(**base)

    def test_default_is_off(self):
        assert BatchRunConfig(
            company_name_column="company", domain_column="domain",
        ).gate_full_enrichment_on_foreign_hq is False

    def test_gate_off_is_byte_identical_to_pre_existing_behavior(self):
        # HARD RULE: gate off (the default) must be 100% the existing
        # per-row path -- no enrichment_skipped/gated_* columns anywhere,
        # every row enriched directly regardless of HQ score.
        fake = _fake_prioritize_for_foreign_hq(
            {"Confirmed Foreign Co": 3.0, "Local Domestic Co": 0.0})
        cfg = self._cfg(gate_full_enrichment_on_foreign_hq=False)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=fake):
            out = run_batch_dataframe(self._df, cfg, "S", "A")
        assert "enrichment_skipped" not in out["enriched_leads"].columns
        assert "gated_full_enrichment_attempted_count" not in out["run_summary"].columns
        assert "gated_full_enrichment_skipped_count" not in out["run_summary"].columns
        assert "gated_estimated_serper_calls_saved" not in out["run_summary"].columns
        rows = out["enriched_leads"].set_index("company_name")
        assert rows.loc["Local Domestic Co", "final_commercial_fit_score"] == 99.0
        assert rows.loc["Confirmed Foreign Co", "final_commercial_fit_score"] == 99.0

    def test_gate_on_skips_non_confirmed_rows(self):
        fake = _fake_prioritize_for_foreign_hq(
            {"Confirmed Foreign Co": 3.0, "Local Domestic Co": 0.0})
        cfg = self._cfg(gate_full_enrichment_on_foreign_hq=True)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=fake):
            out = run_batch_dataframe(self._df, cfg, "S", "A")
        rows = out["enriched_leads"].set_index("company_name")
        assert rows.loc["Local Domestic Co", "enrichment_skipped"] == True  # noqa: E712
        assert rows.loc["Local Domestic Co", "enrichment_skip_reason"] == "Not confirmed foreign HQ"
        assert pd.isna(rows.loc["Local Domestic Co", "final_commercial_fit_score"])

    def test_gate_on_enriches_confirmed_rows(self):
        fake = _fake_prioritize_for_foreign_hq(
            {"Confirmed Foreign Co": 3.0, "Local Domestic Co": 0.0})
        cfg = self._cfg(gate_full_enrichment_on_foreign_hq=True)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=fake):
            out = run_batch_dataframe(self._df, cfg, "S", "A")
        rows = out["enriched_leads"].set_index("company_name")
        assert rows.loc["Confirmed Foreign Co", "enrichment_skipped"] == False  # noqa: E712
        assert rows.loc["Confirmed Foreign Co", "enrichment_skip_reason"] == ""
        assert rows.loc["Confirmed Foreign Co", "final_commercial_fit_score"] == 99.0

    def test_gate_on_run_summary_counts(self):
        fake = _fake_prioritize_for_foreign_hq(
            {"Confirmed Foreign Co": 3.0, "Local Domestic Co": 0.0})
        cfg = self._cfg(gate_full_enrichment_on_foreign_hq=True)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=fake):
            out = run_batch_dataframe(self._df, cfg, "S", "A")
        summary = out["run_summary"].iloc[0].to_dict()
        assert summary["gated_full_enrichment_attempted_count"] == 1
        assert summary["gated_full_enrichment_skipped_count"] == 1
        # skipped_rows * 4 -- the 4 extra non-HQ Serper calls a full v2
        # enrichment would otherwise have made per row.
        assert summary["gated_estimated_serper_calls_saved"] == 4

    def test_gate_on_output_shape_matches_run_batch_dataframe_contract(self):
        # deep_dive key is always present, matching run_batch_dataframe's own
        # contract (unlike run_batch_foreign_hq_only, which omits it).
        fake = _fake_prioritize_for_foreign_hq(
            {"Confirmed Foreign Co": 3.0, "Local Domestic Co": 0.0})
        cfg = self._cfg(gate_full_enrichment_on_foreign_hq=True)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=fake):
            out = run_batch_dataframe(self._df, cfg, "S", "A")
        assert set(out.keys()) == {"enriched_leads", "evidence", "signals", "deep_dive", "run_summary"}
        assert len(out["deep_dive"]) == 0  # deep_dive off by default

    def test_no_infinite_recursion_when_gate_on(self):
        # Regression guard: Phase 1's internal hq_only sub-call must force
        # gate_full_enrichment_on_foreign_hq=False, or run_batch_dataframe
        # would re-enter its own gated branch forever.
        fake = _fake_prioritize_for_foreign_hq(
            {"Confirmed Foreign Co": 3.0, "Local Domestic Co": 0.0})
        cfg = self._cfg(gate_full_enrichment_on_foreign_hq=True)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=fake):
            out = run_batch_dataframe(self._df, cfg, "S", "A")
        assert len(out["enriched_leads"]) == 2

    def test_gate_on_reuses_full_ai_kwargs_not_a_stripped_down_call(self):
        # Unlike run_batch_foreign_hq_only's minimal run_full_v2_pipeline=True
        # call, the gated run_batch_dataframe path must pass the SAME
        # ai_kwargs/flags an ungated call would build for every row (e.g.
        # compose_caller_content_flag), so nothing already supported is
        # silently dropped just because gating is on.
        seen_kwargs = {}

        def _fake(lead_input, **kwargs):
            if kwargs.get("run_full_v2_pipeline"):
                seen_kwargs.update(kwargs)
                return _sample_result(
                    company_name=lead_input.company_name, domain=lead_input.domain,
                    input_country=lead_input.input_country or "Brazil",
                    sig_foreign_hq_score_for_next_scoring=3.0,
                    final_commercial_fit_score=99.0,
                )
            return _sample_result(
                company_name=lead_input.company_name, domain=lead_input.domain,
                input_country=lead_input.input_country or "Brazil",
                sig_foreign_hq_score_for_next_scoring=3.0,
                final_commercial_fit_score=None, evidence_items=[], signals=[],
            )

        cfg = self._cfg(gate_full_enrichment_on_foreign_hq=True,
                        compose_caller_content=True)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake):
            run_batch_dataframe(self._df, cfg, "S", "A")
        assert seen_kwargs.get("compose_caller_content_flag") is True


# ---------------------------------------------------------------------------
# Parallel chunk processing
# ---------------------------------------------------------------------------

class TestSplitDataframeIntoChunks:
    def test_splits_into_approximately_equal_chunks(self):
        df = pd.DataFrame({"c": list(range(10))})
        chunks = split_dataframe_into_chunks(df, 3)
        assert [len(c) for c in chunks] == [4, 3, 3]  # sizes differ by at most 1
        # order and original index preserved
        recombined = pd.concat(chunks)
        assert list(recombined["c"]) == list(range(10))
        assert list(recombined.index) == list(range(10))

    def test_never_produces_empty_chunks(self):
        df = pd.DataFrame({"c": [1, 2]})
        chunks = split_dataframe_into_chunks(df, 4)
        assert [len(c) for c in chunks] == [1, 1]

    def test_empty_frame_yields_no_chunks(self):
        assert split_dataframe_into_chunks(pd.DataFrame({"c": []}), 3) == []


def _fake_prioritize_named(lead_input, **kwargs):
    return _sample_result(company_name=lead_input.company_name,
                          domain=lead_input.domain)


class TestParallelBatch:
    _df9 = pd.DataFrame({
        "company": [f"Co{i}" for i in range(9)],
        "domain": [f"co{i}.com" for i in range(9)],
    })
    _cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                          run_mode="hq_only", row_limit=0)

    def test_parallel_disabled_path_unchanged(self):
        # The sequential entry point is untouched: same output shape as always,
        # no parallel fields on its Run Summary.
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=_fake_prioritize_named):
            out = run_batch_dataframe(self._df9, self._cfg, "S", "A")
        assert len(out["enriched_leads"]) == 9
        assert "parallel_processing_enabled" not in out["run_summary"].columns
        assert "chunk_reports" not in out

    def test_workers_1_single_chunk_matches_sequential(self):
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=_fake_prioritize_named):
            seq = run_batch_dataframe(self._df9, self._cfg, "S", "A")
            par = run_batch_dataframe_parallel(self._df9, self._cfg, "S", "A", workers=1)
        assert list(par["enriched_leads"]["company_name"]) == \
            list(seq["enriched_leads"]["company_name"])
        summary = par["run_summary"].iloc[0].to_dict()
        assert summary["parallel_chunk_count"] == 1
        assert summary["parallel_workers"] == 1

    def test_workers_3_processes_all_rows_preserves_order(self):
        payloads = []
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=_fake_prioritize_named):
            out = run_batch_dataframe_parallel(
                self._df9, self._cfg, "S", "A", workers=3,
                progress_callback=payloads.append)
        enriched = out["enriched_leads"]
        assert list(enriched["company_name"]) == [f"Co{i}" for i in range(9)]
        summary = out["run_summary"].iloc[0].to_dict()
        assert summary["processed_rows"] == 9
        assert summary["success_count"] == 9
        assert summary["error_count"] == 0
        assert summary["parallel_processing_enabled"] == True  # noqa: E712
        assert summary["parallel_workers"] == 3
        assert summary["parallel_chunk_count"] == 3
        assert summary["parallel_chunk_size_min"] == 3
        assert summary["parallel_chunk_size_max"] == 3
        assert summary["parallel_failed_chunk_count"] == 0
        assert summary["parallel_successful_chunk_count"] == 3
        # chunk progress is accurate: totals nonzero, completes at 3/3
        assert payloads
        assert all(p["parallel_chunks_total"] == 3 for p in payloads)
        assert payloads[-1]["parallel_chunks_completed"] == 3
        assert all(p["chunk_row_count"] == 3 for p in payloads)

    def test_combines_evidence_and_signals_from_chunks(self):
        # _sample_result carries 1 evidence item + 1 signal per row.
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=_fake_prioritize_named):
            out = run_batch_dataframe_parallel(self._df9, self._cfg, "S", "A", workers=3)
        assert len(out["evidence"]) == 9
        assert len(out["signals"]) == 9
        assert set(out["evidence"]["source_index"]) == set(range(9))
        assert set(out["signals"]["source_index"]) == set(range(9))

    def test_failed_chunk_reported_without_losing_successes(self):
        original_unit = bc.run_single_batch_unit

        def _unit(chunk_df, cfg, s, a, **kw):
            if "Co4" in list(chunk_df["company"]):
                raise RuntimeError("chunk exploded")
            return original_unit(chunk_df, cfg, s, a, **kw)

        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=_fake_prioritize_named), \
             patch("lead_prioritizer_batch_core.run_single_batch_unit",
                   side_effect=_unit):
            out = run_batch_dataframe_parallel(self._df9, self._cfg, "S", "A", workers=3)

        enriched = out["enriched_leads"]
        assert len(enriched) == 9  # nothing discarded
        # middle chunk (Co3-Co5) becomes placeholder error rows, order intact
        assert list(enriched["company"]) == [f"Co{i}" for i in range(9)]
        failed_rows = enriched[enriched["run_success"] == False]  # noqa: E712
        assert list(failed_rows["company"]) == ["Co3", "Co4", "Co5"]
        assert all(str(e).startswith("parallel_chunk_failed: RuntimeError")
                   for e in failed_rows["run_error"])
        summary = out["run_summary"].iloc[0].to_dict()
        assert summary["parallel_failed_chunk_count"] == 1
        assert summary["parallel_successful_chunk_count"] == 2
        assert summary["processed_rows"] == 6
        assert summary["success_count"] == 6
        assert summary["error_count"] == 3
        reports = out["chunk_reports"]
        failed = [r for r in reports if r["success"] is False]
        assert len(failed) == 1 and "chunk exploded" in failed[0]["error"]
        # successful chunks' data is intact
        assert len(out["evidence"]) == 6

    def test_fho_and_c5_config_passed_through_unchanged(self):
        seen: dict = {}

        def _unit(chunk_df, cfg, s, a, **kw):
            seen.update(kw)
            seen["run_mode"] = cfg.run_mode
            seen["start_row"] = cfg.start_row
            seen["row_limit"] = cfg.row_limit
            seen["default_input_country"] = cfg.default_input_country
            seen["input_country_column"] = cfg.input_country_column
            return {
                "enriched_leads": pd.DataFrame([{"company": "X"}] * len(chunk_df)),
                "evidence": pd.DataFrame(),
                "signals": pd.DataFrame(),
                "run_summary": pd.DataFrame([{
                    "processed_rows": len(chunk_df),
                    "success_count": len(chunk_df), "error_count": 0,
                }]),
            }

        cfg = BatchRunConfig(
            company_name_column="company", domain_column="domain",
            input_country_column="country", default_input_country="Uruguay",
            run_mode=FOREIGN_HQ_ONLY_MODE, start_row=1, row_limit=6)
        df = pd.DataFrame({
            "company": [f"Co{i}" for i in range(8)],
            "domain": [f"co{i}.com" for i in range(8)],
            "country": ["Uruguay"] * 8,
        })
        with patch("lead_prioritizer_batch_core.run_single_batch_unit",
                   side_effect=_unit):
            out = run_batch_dataframe_parallel(
                df, cfg, "S", "A", workers=2,
                c5_enabled=True, c5_scoring_behavior="conservative_adjustment",
                c5_scope="all_rows", c5_model_used="claude-sonnet-5",
                c5_model_tier="sonnet")
        # C5 config forwarded verbatim to every chunk unit
        assert seen["c5_enabled"] is True
        assert seen["c5_scoring_behavior"] == "conservative_adjustment"
        assert seen["c5_scope"] == "all_rows"
        assert seen["c5_model_used"] == "claude-sonnet-5"
        assert seen["c5_model_tier"] == "sonnet"
        # FHO mode + country config forwarded; selection already applied
        assert seen["run_mode"] == FOREIGN_HQ_ONLY_MODE
        assert seen["default_input_country"] == "Uruguay"
        assert seen["input_country_column"] == "country"
        assert seen["start_row"] == 0 and seen["row_limit"] == 0
        # start_row=1 + row_limit=6 → 6 selected rows across chunks
        summary = out["run_summary"].iloc[0].to_dict()
        assert summary["selected_rows"] == 6
        assert summary["processed_rows"] == 6

    def test_fho_mode_end_to_end_parallel(self):
        def _fake(lead_input, **kwargs):
            score = 3.0 if lead_input.company_name in ("Co1", "Co4") else 0.0
            return _sample_result(
                company_name=lead_input.company_name, domain=lead_input.domain,
                sig_foreign_hq_score_for_next_scoring=score,
                final_commercial_fit_score=(
                    99.0 if kwargs.get("run_full_v2_pipeline") else None),
            )

        df = pd.DataFrame({
            "company": [f"Co{i}" for i in range(6)],
            "domain": [f"co{i}.com" for i in range(6)],
        })
        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode=FOREIGN_HQ_ONLY_MODE, row_limit=0)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=_fake):
            out = run_batch_dataframe_parallel(df, cfg, "S", "A", workers=2)
        enriched = out["enriched_leads"].set_index("company_name")
        assert enriched.loc["Co1", "enrichment_skipped"] == False  # noqa: E712
        assert enriched.loc["Co4", "enrichment_skipped"] == False  # noqa: E712
        assert enriched.loc["Co0", "enrichment_skipped"] == True   # noqa: E712
        summary = out["run_summary"].iloc[0].to_dict()
        assert summary["confirmed_foreign_hq_count"] == 2
        assert summary["full_enrichment_attempted_count"] == 2
        assert summary["full_enrichment_skipped_count"] == 4

    def test_workers_capped_at_4(self):
        assert MAX_PARALLEL_WORKERS == 4
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=_fake_prioritize_named):
            out = run_batch_dataframe_parallel(self._df9, self._cfg, "S", "A",
                                               workers=10)
        summary = out["run_summary"].iloc[0].to_dict()
        assert summary["parallel_workers"] == 4
        assert summary["parallel_chunk_count"] == 4

    def test_chunk_result_callback_receives_successful_chunks(self):
        saved = []
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=_fake_prioritize_named):
            run_batch_dataframe_parallel(
                self._df9, self._cfg, "S", "A", workers=3,
                chunk_result_callback=lambda rep, tab: saved.append((rep, tab)))
        assert len(saved) == 3
        assert sorted(rep["chunk_index"] for rep, _ in saved) == [1, 2, 3]
        assert all("enriched_leads" in tab for _, tab in saved)


class TestGateFullEnrichmentOnForeignHqParallelConsistency:
    """Step 3: gate_full_enrichment_on_foreign_hq needs no extra plumbing in
    run_batch_dataframe_parallel / run_single_batch_unit for the regular run
    modes -- both just pass the same BatchRunConfig through to
    run_batch_dataframe, which branches on the flag internally."""

    _df = pd.DataFrame({
        "company": ["Confirmed Foreign Co", "Local Domestic Co"],
        "domain": ["confirmed.com.br", "local.com.br"],
    })
    _cfg = BatchRunConfig(
        company_name_column="company", domain_column="domain",
        run_mode="full", row_limit=2, default_input_country="Brazil",
        gate_full_enrichment_on_foreign_hq=True,
    )

    def test_parallel_path_applies_the_same_gate_as_sequential(self):
        fake = _fake_prioritize_for_foreign_hq(
            {"Confirmed Foreign Co": 3.0, "Local Domestic Co": 0.0})
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=fake):
            seq = run_batch_dataframe(self._df, self._cfg, "S", "A")
            par = run_batch_dataframe_parallel(self._df, self._cfg, "S", "A", workers=1)
        seq_rows = seq["enriched_leads"].set_index("company_name")
        par_rows = par["enriched_leads"].set_index("company_name")
        for name in ("Confirmed Foreign Co", "Local Domestic Co"):
            assert par_rows.loc[name, "enrichment_skipped"] == \
                seq_rows.loc[name, "enrichment_skipped"]
            assert par_rows.loc[name, "enrichment_skip_reason"] == \
                seq_rows.loc[name, "enrichment_skip_reason"]


# ---------------------------------------------------------------------------
# Australia non-English foreign-HQ mode
# ---------------------------------------------------------------------------

class TestClassifyParentHqLanguageMarket:
    @pytest.mark.parametrize("country,expected", [
        ("Germany", "non_english_speaking"),
        ("Japan", "non_english_speaking"),
        ("Brazil", "non_english_speaking"),
        ("United States", "english_speaking"),
        ("USA", "english_speaking"),
        ("UK", "english_speaking"),
        ("Canada", "english_speaking"),
        ("New Zealand", "english_speaking"),
        ("Ireland", "english_speaking"),
        ("Australia", "english_speaking"),
        ("Singapore", "review"),
        ("India", "review"),
        ("South Africa", "review"),
        ("United Arab Emirates", "review"),
        ("UAE", "review"),
        ("", "unclear"),
        (None, "unclear"),
        ("Atlantis", "unclear"),
    ])
    def test_classification(self, country, expected):
        assert classify_parent_hq_language_market(country) == expected

    def test_case_and_whitespace_insensitive(self):
        assert classify_parent_hq_language_market("  germany  ") == "non_english_speaking"
        assert classify_parent_hq_language_market("UNITED STATES") == "english_speaking"


class TestResolveParentHqCountryForExport:
    def test_c5_takes_priority(self):
        row = {"c5_parent_hq_country": "Germany", "ai_parent_hq_country": "France",
               "hq_detected_country": "Italy"}
        assert resolve_parent_hq_country_for_export(row) == "Germany"

    def test_falls_back_to_ai_then_detected(self):
        assert resolve_parent_hq_country_for_export(
            {"c5_parent_hq_country": "", "ai_parent_hq_country": "France",
             "hq_detected_country": "Italy"}) == "France"
        assert resolve_parent_hq_country_for_export(
            {"c5_parent_hq_country": "", "ai_parent_hq_country": "",
             "hq_detected_country": "Italy"}) == "Italy"

    def test_all_blank_returns_empty(self):
        assert resolve_parent_hq_country_for_export({}) == ""


def _au_result(company, score, parent_country, full=False):
    kw = dict(
        company_name=company, domain=company.lower() + ".com", input_country="Australia",
        hq_detected_country=parent_country,
        hq_structure_type="foreign_parent" if score == 3.0 else "domestic",
        sig_foreign_hq_score_for_next_scoring=score,
        ai_parent_hq_country=parent_country,
        domain_root=company.lower(), query_used=f"{company} headquarters", parser_source="ai",
        evidence_items=[], signals=[],
    )
    if full:
        kw["final_commercial_fit_score"] = 88.0
    return LeadPrioritizationResult(**kw)


def _fake_prioritize_for_au(scores: dict):
    def _fake(lead_input, **kwargs):
        score, parent = scores[lead_input.company_name]
        return _au_result(lead_input.company_name, score, parent,
                          full=bool(kwargs.get("run_full_v2_pipeline")))
    return _fake


class TestNonEnglishForeignHqOnlyMode:
    _rows = {
        "GermanCo": (3.0, "Germany"),        # confirmed + non-English -> enrich
        "USCo": (3.0, "United States"),       # confirmed but English -> skip
        "UKCo": (3.0, "United Kingdom"),      # confirmed but English -> skip
        "CanadaCo": (3.0, "Canada"),          # confirmed but English -> skip
        "NZCo": (3.0, "New Zealand"),         # confirmed but English -> skip
        "IrelandCo": (3.0, "Ireland"),        # confirmed but English -> skip
        "SGCo": (3.0, "Singapore"),           # confirmed but review -> skip
        "IndiaCo": (3.0, "India"),            # confirmed but review -> skip
        "SACo": (3.0, "South Africa"),        # confirmed but review -> skip
        "UAECo": (3.0, "UAE"),                # confirmed but review -> skip
        "DomesticCo": (0.0, "Australia"),     # not confirmed -> skip
    }
    _df = pd.DataFrame({"company": list(_rows), "domain": [c.lower() + ".com" for c in _rows]})
    _cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                          run_mode=NON_ENGLISH_FOREIGN_HQ_ONLY_MODE, row_limit=0,
                          default_input_country="Australia")

    def test_only_confirmed_non_english_parent_gets_full_enrichment(self):
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=_fake_prioritize_for_au(self._rows)):
            out = run_batch_non_english_foreign_hq_only(self._df, self._cfg, "S", "A")
        rows = out["enriched_leads"].set_index("company_name")
        assert rows.loc["GermanCo", "enrichment_skipped"] == False  # noqa: E712
        assert rows.loc["GermanCo", "recommended_for_non_english_foreign_hq_export"] == True  # noqa: E712
        # Backward-compat field: still True here since the row's country is Australia.
        assert rows.loc["GermanCo", "recommended_for_australia_export"] == True  # noqa: E712
        assert rows.loc["GermanCo", "final_commercial_fit_score"] == 88.0
        assert rows.loc["GermanCo", "parent_hq_language_market_type"] == "non_english_speaking"

    def test_english_speaking_parents_not_enriched(self):
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=_fake_prioritize_for_au(self._rows)):
            out = run_batch_non_english_foreign_hq_only(self._df, self._cfg, "S", "A")
        rows = out["enriched_leads"].set_index("company_name")
        for co in ("USCo", "UKCo", "CanadaCo", "NZCo", "IrelandCo"):
            assert rows.loc[co, "enrichment_skipped"] == True, co  # noqa: E712
            assert rows.loc[co, "parent_hq_language_market_type"] == "english_speaking", co
            assert rows.loc[co, "export_bucket"] == "skipped_not_relevant", co
            assert rows.loc[co, "enrichment_skip_reason"] == \
                "Parent HQ country is English-speaking/lower-priority for this non-English foreign-HQ run"
            assert pd.isna(rows.loc[co, "final_commercial_fit_score"])

    def test_review_markets_not_enriched(self):
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=_fake_prioritize_for_au(self._rows)):
            out = run_batch_non_english_foreign_hq_only(self._df, self._cfg, "S", "A")
        rows = out["enriched_leads"].set_index("company_name")
        for co in ("SGCo", "IndiaCo", "SACo", "UAECo"):
            assert rows.loc[co, "enrichment_skipped"] == True, co  # noqa: E712
            assert rows.loc[co, "parent_hq_language_market_type"] == "review", co
            assert rows.loc[co, "export_bucket"] == "manual_review", co
            assert rows.loc[co, "review_priority"] == "medium", co
            assert rows.loc[co, "enrichment_skip_reason"] == \
                "Parent HQ country is review/nuanced for language-market trigger"

    def test_not_confirmed_row_skipped(self):
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=_fake_prioritize_for_au(self._rows)):
            out = run_batch_non_english_foreign_hq_only(self._df, self._cfg, "S", "A")
        rows = out["enriched_leads"].set_index("company_name")
        # DomesticCo's parent country is "Australia" (English-speaking market),
        # so it now lands in the same skipped_not_relevant bucket as the other
        # English-speaking-parent rows, regardless of confirmation score.
        assert rows.loc["DomesticCo", "enrichment_skipped"] == True  # noqa: E712
        assert rows.loc["DomesticCo", "export_bucket"] == "skipped_not_relevant"

    def test_run_summary_counts(self):
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=_fake_prioritize_for_au(self._rows)):
            out = run_batch_non_english_foreign_hq_only(self._df, self._cfg, "S", "A")
        summary = out["run_summary"].iloc[0].to_dict()
        assert summary["confirmed_foreign_hq_count"] == 10  # all but DomesticCo
        assert summary["non_english_foreign_hq_count"] == 1  # GermanCo only
        # english_speaking_parent_hq_count is a pure market-type tally (not
        # gated by confirmation): 5 confirmed English parents + DomesticCo
        # (parent country "Australia", not confirmed, market still english).
        assert summary["english_speaking_parent_hq_count"] == 6
        assert summary["review_parent_hq_count"] == 4
        assert summary["unclear_parent_hq_count"] == 0
        assert summary["full_enrichment_attempted_count"] == 1
        assert summary["full_enrichment_skipped_count"] == 10

    def test_output_includes_all_export_columns(self):
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=_fake_prioritize_for_au(self._rows)):
            out = run_batch_non_english_foreign_hq_only(self._df, self._cfg, "S", "A")
        for col in (
            "foreign_hq_detected_for_export", "parent_hq_country_for_export",
            "parent_hq_language_market_type", "non_english_foreign_hq_detected",
            "non_english_foreign_hq_reason", "recommended_for_non_english_foreign_hq_export",
            "recommended_for_australia_export",
            "enrichment_skipped", "enrichment_skip_reason",
        ):
            assert col in out["enriched_leads"].columns

    def test_c5_conservative_downgrade_prevents_enrichment(self):
        rows = {"GermanCo": (3.0, "Germany")}
        df = pd.DataFrame({"company": ["GermanCo"], "domain": ["germanco.com"]})
        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode=NON_ENGLISH_FOREIGN_HQ_ONLY_MODE, row_limit=1,
                             default_input_country="Australia")
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=_fake_prioritize_for_au(rows)), \
             _patch_adjudicator(_mk_result("domestic_confirmed", "High", "yes")):
            out = run_batch_non_english_foreign_hq_only(
                df, cfg, "S", "A", c5_enabled=True,
                c5_scoring_behavior="conservative_adjustment", c5_scope="all_rows",
                c5_model_used="claude-sonnet-5", c5_model_tier="sonnet")
        row = out["enriched_leads"].iloc[0]
        assert row["sig_foreign_hq_score_for_next_scoring"] == 0.0
        assert row["enrichment_skipped"] == True  # noqa: E712
        assert row["recommended_for_non_english_foreign_hq_export"] == False  # noqa: E712
        assert row["recommended_for_australia_export"] == False  # noqa: E712


class TestNonEnglishForeignHqOnlyModeCountryAgnostic:
    """Regression guard: the mode must NOT be Australia-hardcoded.

    A non-Australia row (e.g. Italy, New Zealand) with a confirmed non-English
    foreign parent must now be ELIGIBLE for full enrichment — this is exactly
    the restriction that was removed.
    """

    def _row_result(self, company, score, parent_country, input_country, full=False):
        r = _au_result(company, score, parent_country, full=full)
        r.input_country = input_country
        return r

    def test_italy_row_with_german_parent_now_eligible(self):
        df = pd.DataFrame({"company": ["ItalyCo"], "domain": ["italyco.com"],
                           "country": ["Italy"]})
        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             input_country_column="country",
                             run_mode=NON_ENGLISH_FOREIGN_HQ_ONLY_MODE, row_limit=1)

        def _fake(lead_input, **kwargs):
            return self._row_result("ItalyCo", 3.0, "Germany", "Italy",
                                    full=bool(kwargs.get("run_full_v2_pipeline")))

        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake):
            out = run_batch_non_english_foreign_hq_only(df, cfg, "S", "A")
        row = out["enriched_leads"].iloc[0]
        assert row["enrichment_skipped"] == False  # noqa: E712
        assert row["enrichment_skip_reason"] == ""
        assert row["recommended_for_non_english_foreign_hq_export"] == True  # noqa: E712
        assert row["non_english_foreign_hq_detected"] == True  # noqa: E712
        # Old Australia-only flag correctly stays False (this isn't Australia).
        assert row["recommended_for_australia_export"] == False  # noqa: E712
        assert row["final_commercial_fit_score"] == 88.0

    # ── Acceptance test: New Zealand ────────────────────────────────────────

    _nz_rows = {
        "FranceCo": (3.0, "France"),         # confirmed non-English -> enrich
        "AustraliaCo": (3.0, "Australia"),   # confirmed but English -> skip
        "SingaporeCo": (3.0, "Singapore"),   # confirmed but review -> skip
        "DomesticCo": (3.0, "New Zealand"),  # same as input country -> skip
        "UnclearCo": (0.0, ""),              # not confirmed / unclear -> skip
    }

    def _nz_fake(self, lead_input, **kwargs):
        score, parent = self._nz_rows[lead_input.company_name]
        return self._row_result(lead_input.company_name, score, parent, "New Zealand",
                                full=bool(kwargs.get("run_full_v2_pipeline")))

    def _run_nz(self):
        df = pd.DataFrame({"company": list(self._nz_rows),
                           "domain": [f"{c.lower()}.com" for c in self._nz_rows]})
        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode=NON_ENGLISH_FOREIGN_HQ_ONLY_MODE, row_limit=0,
                             default_input_country="New Zealand")
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=self._nz_fake):
            return run_batch_non_english_foreign_hq_only(df, cfg, "S", "A")

    def test_nz_france_parent_eligible(self):
        rows = self._run_nz()["enriched_leads"].set_index("company_name")
        assert rows.loc["FranceCo", "enrichment_skipped"] == False  # noqa: E712
        assert rows.loc["FranceCo", "recommended_for_non_english_foreign_hq_export"] == True  # noqa: E712
        assert rows.loc["FranceCo", "parent_hq_language_market_type"] == "non_english_speaking"

    def test_nz_australia_parent_skipped_english_speaking(self):
        rows = self._run_nz()["enriched_leads"].set_index("company_name")
        assert rows.loc["AustraliaCo", "enrichment_skipped"] == True  # noqa: E712
        assert rows.loc["AustraliaCo", "parent_hq_language_market_type"] == "english_speaking"
        assert "English-speaking" in rows.loc["AustraliaCo", "enrichment_skip_reason"]

    def test_nz_singapore_parent_skipped_review(self):
        rows = self._run_nz()["enriched_leads"].set_index("company_name")
        assert rows.loc["SingaporeCo", "enrichment_skipped"] == True  # noqa: E712
        assert rows.loc["SingaporeCo", "parent_hq_language_market_type"] == "review"
        assert "review/nuanced" in rows.loc["SingaporeCo", "enrichment_skip_reason"]

    def test_nz_domestic_parent_skipped(self):
        rows = self._run_nz()["enriched_leads"].set_index("company_name")
        assert rows.loc["DomesticCo", "enrichment_skipped"] == True  # noqa: E712
        assert rows.loc["DomesticCo", "enrichment_skip_reason"] == \
            "Parent HQ country matches input country (not foreign)"

    def test_nz_unclear_result_skipped_manual_review(self):
        rows = self._run_nz()["enriched_leads"].set_index("company_name")
        assert rows.loc["UnclearCo", "enrichment_skipped"] == True  # noqa: E712
        assert rows.loc["UnclearCo", "export_bucket"] == "manual_review"
        assert rows.loc["UnclearCo", "review_priority"] == "medium"
        assert rows.loc["UnclearCo", "enrichment_skip_reason"] == \
            "Parent HQ country or language-market type unclear"

    def test_missing_input_country_skipped(self):
        df = pd.DataFrame({"company": ["NoCountryCo"], "domain": ["nocountryco.com"]})
        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode=NON_ENGLISH_FOREIGN_HQ_ONLY_MODE, row_limit=1)

        def _fake(lead_input, **kwargs):
            return self._row_result("NoCountryCo", 3.0, "Germany", "",
                                    full=bool(kwargs.get("run_full_v2_pipeline")))

        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake):
            out = run_batch_non_english_foreign_hq_only(df, cfg, "S", "A")
        row = out["enriched_leads"].iloc[0]
        assert row["enrichment_skipped"] == True  # noqa: E712
        assert row["enrichment_skip_reason"] == "Input country is missing"

    def test_missing_parent_country_skipped(self):
        df = pd.DataFrame({"company": ["NoParentCo"], "domain": ["noparentco.com"],
                           "country": ["New Zealand"]})
        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             input_country_column="country",
                             run_mode=NON_ENGLISH_FOREIGN_HQ_ONLY_MODE, row_limit=1)

        def _fake(lead_input, **kwargs):
            return self._row_result("NoParentCo", 3.0, "", "New Zealand",
                                    full=bool(kwargs.get("run_full_v2_pipeline")))

        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake):
            out = run_batch_non_english_foreign_hq_only(df, cfg, "S", "A")
        row = out["enriched_leads"].iloc[0]
        assert row["enrichment_skipped"] == True  # noqa: E712
        # A blank parent country classifies as an "unclear" language market,
        # so it now lands in the manual_review/medium bucket like any other
        # unrecognised parent country.
        assert row["export_bucket"] == "manual_review"
        assert row["review_priority"] == "medium"
        assert row["enrichment_skip_reason"] == "Parent HQ country or language-market type unclear"


class TestExistingForeignHqOnlyModeUnaffected:
    """The FOREIGN_HQ_ONLY_MODE selection/output behavior must be byte-for-byte
    unchanged after adding the non-English mode (shared-helper refactor)."""

    def test_selection_and_output_unchanged(self):
        rows = {"GermanCo": (3.0, "Germany"), "USCo": (3.0, "United States"),
                "DomesticCo": (0.0, "Australia")}
        df = pd.DataFrame({"company": list(rows), "domain": [f"{c.lower()}.com" for c in rows]})
        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode=FOREIGN_HQ_ONLY_MODE, row_limit=0)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=_fake_prioritize_for_au(rows)):
            out = run_batch_foreign_hq_only(df, cfg, "S", "A")
        enriched = out["enriched_leads"].set_index("company_name")
        # ALL confirmed rows enrich regardless of parent-HQ language (no filter)
        assert enriched.loc["GermanCo", "enrichment_skipped"] == False  # noqa: E712
        assert enriched.loc["USCo", "enrichment_skipped"] == False  # noqa: E712
        assert enriched.loc["DomesticCo", "enrichment_skipped"] == True  # noqa: E712
        # no non-English-mode-only columns leak into this mode's output
        for col in ("parent_hq_language_market_type", "recommended_for_australia_export"):
            assert col not in out["enriched_leads"].columns
        summary = out["run_summary"].iloc[0].to_dict()
        assert summary["confirmed_foreign_hq_count"] == 2
        assert summary["full_enrichment_attempted_count"] == 2
        assert summary["full_enrichment_skipped_count"] == 1


class TestParallelNonEnglishMode:
    def test_run_single_batch_unit_dispatches_to_non_english_mode(self):
        rows = {"GermanCo": (3.0, "Germany"), "USCo": (3.0, "United States")}
        df = pd.DataFrame({"company": list(rows), "domain": [f"{c.lower()}.com" for c in rows]})
        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode=NON_ENGLISH_FOREIGN_HQ_ONLY_MODE, row_limit=0,
                             default_input_country="Australia")
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=_fake_prioritize_for_au(rows)):
            out = run_single_batch_unit(df, cfg, "S", "A")
        assert "recommended_for_australia_export" in out["enriched_leads"].columns

    def test_parallel_run_aggregates_non_english_counts(self):
        rows = {f"Co{i}": (3.0, "Germany") for i in range(4)}
        rows["USCo"] = (3.0, "United States")
        rows["DomesticCo"] = (0.0, "Australia")
        df = pd.DataFrame({"company": list(rows), "domain": [f"{c.lower()}.com" for c in rows]})
        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode=NON_ENGLISH_FOREIGN_HQ_ONLY_MODE, row_limit=0,
                             default_input_country="Australia")
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=_fake_prioritize_for_au(rows)):
            out = run_batch_dataframe_parallel(df, cfg, "S", "A", workers=3)
        assert len(out["enriched_leads"]) == 6
        summary = out["run_summary"].iloc[0].to_dict()
        assert summary["non_english_foreign_hq_count"] == 4
        assert summary["full_enrichment_attempted_count"] == 4
        assert summary["full_enrichment_skipped_count"] == 2
        assert summary["parallel_chunk_count"] == 3


# ---------------------------------------------------------------------------
# aggregate_parallel_chunk_progress (pure helper) + live parallel progress
# ---------------------------------------------------------------------------

class TestAggregateParallelChunkProgress:
    def test_running_chunks_use_live_snapshot(self):
        reports = [
            {"chunk_index": 1, "row_count": 10, "success": None, "error": ""},
            {"chunk_index": 2, "row_count": 10, "success": None, "error": ""},
        ]
        snapshot = {
            0: {"processed": 4, "success": 4, "error": 0, "selected": 10,
                "current_company_name": "Co4", "last_update": 5.0,
                "phase_label": None, "phase": None, "phase_processed": 0, "phase_total": 0},
            1: {"processed": 2, "success": 1, "error": 1, "selected": 10,
                "current_company_name": "Co12", "last_update": 9.0,
                "phase_label": None, "phase": None, "phase_processed": 0, "phase_total": 0},
        }
        agg = aggregate_parallel_chunk_progress(snapshot, reports, total_selected_rows=20)
        assert agg["processed_rows"] == 6
        assert agg["success_count"] == 5
        assert agg["error_count"] == 1
        assert agg["selected_rows"] == 20
        # most recently updated running chunk wins for "current company"
        assert agg["current_company_name"] == "Co12"
        assert agg["chunks_active_count"] == 2
        assert {c["chunk_index"] for c in agg["active_chunks"]} == {1, 2}

    def test_finished_successful_chunk_trusts_row_count(self):
        reports = [{"chunk_index": 1, "row_count": 10, "success": True, "error": ""}]
        snapshot = {0: {"processed": 10, "success": 9, "error": 1, "selected": 10,
                       "current_company_name": "LastCo", "last_update": 1.0}}
        agg = aggregate_parallel_chunk_progress(snapshot, reports)
        assert agg["processed_rows"] == 10
        assert agg["success_count"] == 9
        assert agg["error_count"] == 1
        assert agg["chunks_active_count"] == 0  # finished chunks are not "active"

    def test_finished_failed_chunk_counts_all_rows_as_errors(self):
        reports = [{"chunk_index": 1, "row_count": 10, "success": False, "error": "boom"}]
        agg = aggregate_parallel_chunk_progress({}, reports)
        assert agg["processed_rows"] == 10
        assert agg["error_count"] == 10
        assert agg["success_count"] == 0

    def test_phase_info_surfaced_for_active_chunks(self):
        reports = [{"chunk_index": 1, "row_count": 5, "success": None, "error": ""}]
        snapshot = {0: {"processed": 2, "success": 2, "error": 0, "selected": 5,
                       "current_company_name": "Acme", "last_update": 1.0,
                       "phase": 1, "phase_label": "HQ screening",
                       "phase_processed": 2, "phase_total": 5}}
        agg = aggregate_parallel_chunk_progress(snapshot, reports)
        assert agg["active_chunks"][0]["phase_label"] == "HQ screening"
        assert agg["active_chunks"][0]["phase_total"] == 5

    def test_empty_reports_yields_zeroed_aggregate(self):
        agg = aggregate_parallel_chunk_progress({}, [])
        assert agg["processed_rows"] == 0
        assert agg["success_count"] == 0
        assert agg["error_count"] == 0
        assert agg["active_chunks"] == []
        assert agg["current_company_name"] == ""


class TestParallelLiveProgress:
    """Regression guard: parallel mode must emit progress before every chunk
    finishes (row-level heartbeat), not just once per chunk completion."""

    def _slow_prioritize(self, sleep_seconds):
        def _fake(lead_input, **kwargs):
            time.sleep(sleep_seconds)
            return _sample_result(company_name=lead_input.company_name,
                                  domain=lead_input.domain)
        return _fake

    def test_heartbeat_and_row_level_events_fire_before_completion(self):
        df = pd.DataFrame({"company": [f"Co{i}" for i in range(6)],
                           "domain": [f"co{i}.com" for i in range(6)]})
        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="hq_only", row_limit=0)
        events = []
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=self._slow_prioritize(0.05)):
            out = run_batch_dataframe_parallel(
                df, cfg, "S", "A", workers=2, progress_callback=events.append,
                heartbeat_interval_seconds=0.02,
            )
        assert len(events) > 2  # far more than the old "one event per chunk"
        heartbeats = [e for e in events if e.get("heartbeat")]
        completions = [e for e in events if not e.get("heartbeat")]
        assert heartbeats, "expected at least one heartbeat wake-up before chunks finished"
        assert len(completions) == 2  # one per chunk
        # heartbeat payloads carry live, incrementally-growing row counts —
        # not just a static "still waiting" message with no numbers.
        assert any(e["processed_rows"] > 0 for e in heartbeats)
        assert all("current_company_name" in e for e in events)
        assert all("parallel_workers" in e and e["parallel_workers"] == 2 for e in events)
        assert len(out["enriched_leads"]) == 6

    def test_progress_callback_never_called_from_worker_thread_unsafely(self):
        # The callback we pass to run_batch_dataframe_parallel must only ever
        # be invoked by the polling loop (main thread), never directly by a
        # chunk's internal row callback — verified by checking every payload
        # has the aggregate shape (not a bare per-row payload).
        df = pd.DataFrame({"company": ["A", "B", "C", "D"],
                           "domain": ["a.com", "b.com", "c.com", "d.com"]})
        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="hq_only", row_limit=0)
        events = []
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=_fake_prioritize_named):
            run_batch_dataframe_parallel(
                df, cfg, "S", "A", workers=2, progress_callback=events.append,
                heartbeat_interval_seconds=0.05,
            )
        assert events
        for e in events:
            assert "parallel_chunks_total" in e
            assert "active_chunks" in e

    def test_chunk_failure_reported_without_hiding_error(self):
        df = pd.DataFrame({"company": ["Good1", "Good2", "Bad1", "Bad2"],
                           "domain": ["g1.com", "g2.com", "b1.com", "b2.com"]})
        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode="hq_only", row_limit=0)

        def _fake(lead_input, **kwargs):
            return _sample_result(company_name=lead_input.company_name)

        original_unit = run_single_batch_unit

        def _unit(chunk_df, cfg2, s, a, **kw):
            if "Bad1" in list(chunk_df["company"]):
                raise RuntimeError("chunk exploded")
            return original_unit(chunk_df, cfg2, s, a, **kw)

        events = []
        with patch("lead_prioritizer_batch_core.prioritize_single_lead", side_effect=_fake), \
             patch("lead_prioritizer_batch_core.run_single_batch_unit", side_effect=_unit):
            out = run_batch_dataframe_parallel(
                df, cfg, "S", "A", workers=2, progress_callback=events.append,
                heartbeat_interval_seconds=0.02,
            )
        failure_events = [e for e in events if e.get("chunk_success") is False]
        assert failure_events
        assert "chunk exploded" in failure_events[0]["chunk_error"]
        assert failure_events[0]["error_count"] >= 2  # failed chunk's rows counted
        assert len(out["enriched_leads"]) == 4  # nothing silently dropped


# ---------------------------------------------------------------------------
# classify_non_english_foreign_hq_export_row — post-C5 export/review buckets
# ---------------------------------------------------------------------------
#
# Australia audit fix: C5 can confirm foreign_parent_confirmed + a target
# match against a non-English parent while the final HQ score stays 0 (the
# conservative C5 rule never auto-upgrades score-0 rows). This layer makes
# those NEC-style near-misses (e.g. NEC Australia -> NEC Corporation, Japan)
# visible as a high-priority manual_review bucket instead of silently
# dropping them in "Not confirmed foreign HQ".

class TestNonEnglishForeignHqExportBuckets:
    def _row(self, **kw):
        base = {
            "company_name": "Acme", "domain": "acme.com", "input_country": "Australia",
            "sig_foreign_hq_score_for_next_scoring": 0.0,
            "ai_parent_hq_country": "", "hq_detected_country": "", "c5_parent_hq_country": "",
            "c5_adjudication": "", "c5_target_company_match": "",
        }
        base.update(kw)
        return base

    def test_score_3_germany_parent_is_direct_target(self):
        row = self._row(sig_foreign_hq_score_for_next_scoring=3.0, ai_parent_hq_country="Germany")
        result = bc.classify_non_english_foreign_hq_export_row(row)
        assert result["export_bucket"] == "direct_target"
        assert result["recommended_for_non_english_foreign_hq_export"] is True
        assert result["enrichment_skipped"] is False

    def test_c5_confirmed_japan_parent_score_0_is_high_priority_manual_review(self):
        row = self._row(sig_foreign_hq_score_for_next_scoring=0.0,
                        ai_parent_hq_country="Japan",
                        c5_adjudication="foreign_parent_confirmed",
                        c5_target_company_match="yes")
        result = bc.classify_non_english_foreign_hq_export_row(row)
        assert result["export_bucket"] == "manual_review"
        assert result["review_priority"] == "high"
        assert result["non_english_foreign_hq_review"] is True
        assert result["recommended_for_non_english_foreign_hq_export"] is False
        assert result["enrichment_skipped"] is True  # not full enriched in this mode

    def test_c5_confirmed_united_states_parent_is_skipped_not_relevant(self):
        row = self._row(sig_foreign_hq_score_for_next_scoring=0.0,
                        ai_parent_hq_country="United States",
                        c5_adjudication="foreign_parent_confirmed",
                        c5_target_company_match="yes")
        result = bc.classify_non_english_foreign_hq_export_row(row)
        assert result["export_bucket"] == "skipped_not_relevant"
        assert result["recommended_for_non_english_foreign_hq_export"] is False

    def test_review_market_parents_are_manual_review_not_direct_target(self):
        for parent in ("Singapore", "India", "Hong Kong"):
            row = self._row(sig_foreign_hq_score_for_next_scoring=3.0, ai_parent_hq_country=parent)
            result = bc.classify_non_english_foreign_hq_export_row(row)
            assert result["export_bucket"] != "direct_target", parent
            assert result["export_bucket"] == "manual_review", parent
            assert result["review_priority"] == "medium", parent

    def test_bolivia_domain_with_australia_input_is_excluded(self):
        row = self._row(domain="entel.bo", input_country="Australia")
        result = bc.classify_non_english_foreign_hq_export_row(row)
        assert result["export_bucket"] == "excluded"
        assert result["exclude_from_export"] is True
        assert "Bolivia" in result["exclude_reason"]

    def test_company_name_entel_bolivia_is_excluded(self):
        row = self._row(company_name="Entel Bolivia", domain="entel.com", input_country="Australia")
        result = bc.classify_non_english_foreign_hq_export_row(row)
        assert result["export_bucket"] == "excluded"
        assert result["exclude_from_export"] is True

    def test_costa_rica_domain_and_name_also_excluded(self):
        row = self._row(domain="empresa.cr", input_country="New Zealand")
        result = bc.classify_non_english_foreign_hq_export_row(row)
        assert result["export_bucket"] == "excluded"

        row2 = self._row(company_name="Costa Rica Traders", domain="crt.com",
                         input_country="New Zealand")
        result2 = bc.classify_non_english_foreign_hq_export_row(row2)
        assert result2["export_bucket"] == "excluded"

    def test_exclusion_only_applies_for_australia_or_new_zealand_input(self):
        row = self._row(domain="entel.bo", input_country="Chile")
        result = bc.classify_non_english_foreign_hq_export_row(row)
        assert result["export_bucket"] != "excluded"

    def test_foreign_hq_only_mode_unaffected_by_new_bucket_layer(self):
        rows = {"GermanCo": (3.0, "Germany"), "USCo": (3.0, "United States"),
                "DomesticCo": (0.0, "Australia")}
        df = pd.DataFrame({"company": list(rows), "domain": [f"{c.lower()}.com" for c in rows]})
        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode=FOREIGN_HQ_ONLY_MODE, row_limit=0)
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=_fake_prioritize_for_au(rows)):
            out = run_batch_foreign_hq_only(df, cfg, "S", "A")
        enriched = out["enriched_leads"].set_index("company_name")
        assert enriched.loc["GermanCo", "enrichment_skipped"] == False  # noqa: E712
        assert enriched.loc["USCo", "enrichment_skipped"] == False  # noqa: E712
        assert enriched.loc["DomesticCo", "enrichment_skipped"] == True  # noqa: E712
        for col in ("export_bucket", "non_english_foreign_hq_review", "review_priority"):
            assert col not in out["enriched_leads"].columns

    def test_c5_conservative_adjustment_never_upgrades_score_0_to_3(self):
        rows = {"JapanCo": (0.0, "Japan")}
        df = pd.DataFrame({"company": ["JapanCo"], "domain": ["japanco.com"]})
        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode=NON_ENGLISH_FOREIGN_HQ_ONLY_MODE, row_limit=1,
                             default_input_country="Australia")
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=_fake_prioritize_for_au(rows)), \
             _patch_adjudicator(_mk_result("foreign_parent_confirmed", "High", "yes")):
            out = run_batch_non_english_foreign_hq_only(
                df, cfg, "S", "A", c5_enabled=True,
                c5_scoring_behavior="conservative_adjustment", c5_scope="all_rows",
                c5_model_used="claude-sonnet-5", c5_model_tier="sonnet")
        row = out["enriched_leads"].iloc[0]
        # Conservative C5 never auto-upgrades a score-0 row to 3 ...
        assert row["sig_foreign_hq_score_for_next_scoring"] == 0.0
        # ... but the row is still surfaced as a high-priority manual review
        # instead of silently disappearing (this is the whole point of the
        # export-bucket layer).
        assert row["enrichment_skipped"] == True  # noqa: E712
        assert row["export_bucket"] == "manual_review"
        assert row["review_priority"] == "high"
        assert row["non_english_foreign_hq_review"] == True  # noqa: E712

    def test_run_summary_includes_new_bucket_counts(self):
        rows = {"GermanCo": (3.0, "Germany"), "USCo": (3.0, "United States")}
        df = pd.DataFrame({"company": list(rows), "domain": [f"{c.lower()}.com" for c in rows]})
        cfg = BatchRunConfig(company_name_column="company", domain_column="domain",
                             run_mode=NON_ENGLISH_FOREIGN_HQ_ONLY_MODE, row_limit=0,
                             default_input_country="Australia")
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=_fake_prioritize_for_au(rows)):
            out = run_batch_non_english_foreign_hq_only(df, cfg, "S", "A")
        summary = out["run_summary"].iloc[0].to_dict()
        assert summary["direct_target_count"] == 1  # GermanCo
        assert summary["skipped_not_relevant_count"] == 1  # USCo
        assert summary["manual_review_count"] == 0
        assert summary["excluded_count"] == 0
        assert summary["high_priority_manual_review_count"] == 0
        assert summary["medium_priority_manual_review_count"] == 0
        assert summary["non_english_foreign_hq_review_count"] == 0


# ---------------------------------------------------------------------------
# Experimental AI-provider passthrough (defaults must stay Anthropic)
# ---------------------------------------------------------------------------

class TestAiProviderPassthrough:
    _df = pd.DataFrame({"c": ["Acme"], "d": ["acme.com"]})

    def _cfg(self, **kw):
        base = dict(company_name_column="c", domain_column="d",
                    run_mode="hq_only", row_limit=1)
        base.update(kw)
        return BatchRunConfig(**base)

    def test_config_defaults_stay_anthropic(self):
        cfg = self._cfg()
        assert cfg.ai_provider == "anthropic"
        assert cfg.ai_model == ""

    def test_default_run_keeps_anthropic_and_pipeline_model(self):
        captured = {}

        def _fake(lead, **kw):
            captured.update(kw)
            return _sample_result()

        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=_fake):
            run_batch_dataframe(self._df, self._cfg(), "S", "A")
        assert captured["ai_provider"] == "anthropic"
        assert captured["openai_api_key"] == ""
        # ai_model must NOT be overridden — the pipeline default stays in charge.
        assert "ai_model" not in captured

    def test_openai_provider_kwargs_passed_through(self):
        captured = {}

        def _fake(lead, **kw):
            captured.update(kw)
            return _sample_result()

        cfg = self._cfg(ai_provider="openai", ai_model="gpt-5.4-nano")
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=_fake):
            run_batch_dataframe(self._df, cfg, "S", "A", openai_api_key="OK")
        assert captured["ai_provider"] == "openai"
        assert captured["ai_model"] == "gpt-5.4-nano"
        assert captured["openai_api_key"] == "OK"

    def test_run_summary_records_provider_and_model_without_keys(self):
        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   return_value=_sample_result()):
            out = run_batch_dataframe(
                self._df, self._cfg(ai_provider="openai", ai_model="gpt-5.4-mini"),
                "S", "A", openai_api_key="SECRET-KEY")
        summary = out["run_summary"].iloc[0].to_dict()
        assert summary["ai_provider"] == "openai"
        assert summary["ai_model"] == "gpt-5.4-mini"
        # No API key may ever appear in any output table.
        for table in out.values():
            for col in table.columns:
                assert "SECRET-KEY" not in table[col].astype(str).str.cat(sep=" ")

    def test_single_batch_unit_passes_openai_key(self):
        captured = {}

        def _fake(lead, **kw):
            captured.update(kw)
            return _sample_result()

        with patch("lead_prioritizer_batch_core.prioritize_single_lead",
                   side_effect=_fake):
            run_single_batch_unit(
                self._df, self._cfg(ai_provider="openai"), "S", "A",
                openai_api_key="OK")
        assert captured["ai_provider"] == "openai"
        assert captured["openai_api_key"] == "OK"
