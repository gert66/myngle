"""
mYngle · Buyer Contact Finder  (Layer 2.5)
==========================================
Runs after Opportunity Radar and before Caller Prep Cockpit / Lovable.

Input  : Opportunity Radar export (.xlsx) — reads 'Caller Prep Input' sheet.
Output : Enriched Excel with the same structure, plus up to 3 buyer contacts
         per company, compatible with Lovable Caller Prep (Layer 3).

Entry point:  streamlit run buyer_contact_finder.py

Architecture note
-----------------
Layer 1 (Lead Prioritizer) handles company-level firmographics only.
Layer 2.5 (this tool) is the dedicated place for Lusha person/contact
enrichment — only for companies actually going to cold calling.
"""

import base64
import io
import json
import pathlib
import re
import time
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
import requests
import streamlit as st

# =============================================================================
# PAGE CONFIG
# =============================================================================

st.set_page_config(
    page_title="mYngle · Buyer Contact Finder",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# =============================================================================
# HEADER
# =============================================================================

_logo_path = pathlib.Path(__file__).parent / "mingle_local_final_fixed.png"
_logo_src = (
    f"data:image/png;base64,{base64.b64encode(_logo_path.read_bytes()).decode()}"
    if _logo_path.exists()
    else ""
)
_img_tag = (
    f'<img src="{_logo_src}" class="brand-logo" alt="mYngle" />'
    if _logo_src
    else ""
)

st.markdown(
    f"""
    <style>
    .block-container {{
        max-width: 880px;
        padding-top: 2.2rem;
        padding-bottom: 3rem;
        padding-left: 2rem;
        padding-right: 2rem;
    }}
    div[data-testid="stMarkdownContainer"]:has(.brand-header) {{
        overflow: visible !important;
        margin-bottom: 1.0rem;
    }}
    .brand-header {{
        display: grid;
        grid-template-columns: 43% 57%;
        align-items: center;
        min-height: 140px;
        padding-top: 10px;
        padding-bottom: 6px;
        overflow: visible !important;
    }}
    .brand-title-block {{
        display: flex;
        align-items: center;
        justify-content: flex-start;
        overflow: visible !important;
    }}
    .brand-title {{
        font-size: 42px;
        font-weight: 700;
        color: #0B1F3A;
        line-height: 1.1;
        white-space: nowrap;
        margin: 0;
        padding: 0;
    }}
    .brand-logo-block {{
        display: flex;
        justify-content: flex-end;
        align-items: center;
        padding: 0;
        line-height: 0;
        overflow: visible !important;
    }}
    .brand-logo {{
        width: 430px;
        max-width: 100%;
        height: auto;
        display: block;
        object-fit: contain;
        object-position: center center;
        overflow: visible !important;
    }}
    </style>
    <div class="brand-header">
      <div class="brand-title-block">
        <span class="brand-title">Buyer Contact Finder</span>
      </div>
      <div class="brand-logo-block">
        {_img_tag}
      </div>
    </div>
    """,
    unsafe_allow_html=True,
)

# =============================================================================
# CONSTANTS
# =============================================================================

SHEET_CALLER_PREP = "Caller Prep Input"

# Lusha person/contact endpoint candidates — tried in order during preflight.
# The /company endpoint (used by Layer 1) is confirmed working.
# Person endpoints vary by plan; we probe all known variants and stop at first 200.
_LUSHA_PERSON_CANDIDATES = [
    "https://api.lusha.com/contacts",          # most common v2 people endpoint
    "https://api.lusha.com/people",            # alternative name
    "https://api.lusha.com/person",            # v1 style (404 observed)
    "https://api.lusha.com/api/contacts",      # some account types
]
_LUSHA_COMPANY_BASE = "https://api.lusha.com/company"  # confirmed working
_LUSHA_TIMEOUT = 20

# HTTP status codes that mean "endpoint does not exist / not on your plan"
_LUSHA_UNSUPPORTED_CODES = {404, 405, 501}
# HTTP status codes that mean "API key problem"
_LUSHA_AUTH_CODES = {401, 403}

# Relevance tiers for buyer contact scoring
_HIGH_RELEVANCE_TITLES = [
    "head of hr", "hr director", "people director", "head of people",
    "head of learning", "l&d manager", "learning & development director",
    "learning and development director", "talent development manager",
    "training manager", "academy lead", "international hr manager",
    "international hr director", "global hr", "global head of people",
    "sales enablement lead", "sales enablement manager",
    "vp people", "chief people officer", "cpo",
    "head of talent", "director of people", "director of hr",
    "learning director", "l&d director", "people operations director",
    "head of people operations",
]

_MEDIUM_RELEVANCE_TITLES = [
    "hr manager", "talent acquisition", "people partner",
    "hr business partner", "hrbp", "operations manager",
    "customer success enablement", "commercial excellence",
    "people operations manager", "hr operations",
    "learning coordinator", "training coordinator",
    "employee development", "people development",
    "hr generalist", "talent manager", "learning specialist",
]

_LOW_RELEVANCE_TITLES = [
    "ceo", "chief executive", "cfo", "coo", "cto",
    "it manager", "legal", "finance", "general manager",
    "managing director", "president",
]

# Keywords used to score a returned contact title
_TITLE_SCORE_MAP: list[tuple[str, int]] = [
    # High relevance (score 3)
    ("head of hr", 3), ("hr director", 3), ("people director", 3),
    ("head of people", 3), ("head of learning", 3), ("l&d manager", 3),
    ("learning & development director", 3), ("learning and development director", 3),
    ("talent development manager", 3), ("training manager", 3),
    ("academy lead", 3), ("international hr", 3), ("global hr", 3),
    ("sales enablement lead", 3), ("sales enablement manager", 3),
    ("chief people officer", 3), ("cpo", 3), ("vp people", 3),
    # Medium relevance (score 2)
    ("hr manager", 2), ("talent acquisition", 2), ("people partner", 2),
    ("hr business partner", 2), ("hrbp", 2), ("operations manager", 2),
    ("people operations", 2), ("hr operations", 2),
    ("learning coordinator", 2), ("training coordinator", 2),
    ("employee development", 2), ("people development", 2),
    ("talent manager", 2), ("learning specialist", 2),
    ("customer success enablement", 2), ("commercial excellence", 2),
    # Broad catch-alls (score 1)
    ("hr", 1), ("people", 1), ("learning", 1), ("talent", 1),
    ("training", 1), ("l&d", 1), ("academy", 1), ("enablement", 1),
    # Lower relevance (score 0 — only used when nothing better found)
    ("ceo", 0), ("cfo", 0), ("coo", 0), ("cto", 0),
    ("managing director", 0), ("president", 0),
]

# Preferred buyer profiles by preferred_buyer_route keyword
_ROUTE_PRIORITY_TITLES: dict[str, list[str]] = {
    "international hr": [
        "international hr", "global hr", "head of people", "people director",
        "hr director", "vp people", "chief people officer",
    ],
    "l&d": [
        "head of learning", "l&d", "learning", "talent development",
        "training manager", "academy",
    ],
    "talent development": [
        "talent development", "l&d", "learning", "head of learning",
        "training manager", "academy",
    ],
    "sales enablement": [
        "sales enablement", "commercial enablement", "revenue enablement",
        "sales operations",
    ],
    "customer success": [
        "customer success enablement", "customer success director",
        "vp customer success", "head of customer success",
    ],
    "hr / people": [
        "hr director", "head of people", "people director",
        "chief people officer", "hr manager", "people partner",
    ],
    "people operations": [
        "people operations", "hr operations", "head of people operations",
        "onboarding", "hr business partner",
    ],
    "operations": [
        "operations director", "head of operations", "people operations",
    ],
}

# contact_data_status allowed values
STATUS_CONTACTS_FOUND = "contacts_found"
STATUS_NO_CONTACTS = "no_contacts_found"
STATUS_COMPANY_ONLY = "company_data_only"
STATUS_NOT_RUN = "contact_lookup_not_run"
STATUS_ERROR = "contact_lookup_error"
STATUS_NOT_SUPPORTED = "contact_lookup_not_supported"
STATUS_DEMO = "demo_contacts"
STATUS_SKIPPED_NOT_SELECTED = "skipped_not_selected"
STATUS_SKIPPED_HAS_CONTACTS = "skipped_already_has_contacts"

# Demo placeholder contact templates keyed by buyer route fragment
_DEMO_CONTACTS_BY_ROUTE: dict[str, list[dict]] = {
    "international hr": [
        {"title": "Head of International HR",    "department": "HR"},
        {"title": "HR Director",                 "department": "HR"},
        {"title": "People Operations Manager",   "department": "People Operations"},
    ],
    "l&d": [
        {"title": "Head of Learning & Development", "department": "L&D"},
        {"title": "Talent Development Manager",      "department": "L&D"},
        {"title": "Training Manager",                "department": "Training"},
    ],
    "talent": [
        {"title": "Head of Learning & Development", "department": "L&D"},
        {"title": "Talent Development Manager",      "department": "L&D"},
        {"title": "Training Manager",                "department": "Training"},
    ],
    "sales enablement": [
        {"title": "Sales Enablement Lead",          "department": "Sales Enablement"},
        {"title": "Commercial Enablement Manager",  "department": "Commercial"},
        {"title": "Revenue Enablement Manager",     "department": "Revenue Operations"},
    ],
    "customer success": [
        {"title": "Head of Customer Success",       "department": "Customer Success"},
        {"title": "Customer Success Enablement Manager", "department": "Customer Success"},
        {"title": "VP Customer Success",            "department": "Customer Success"},
    ],
    "hr / people": [
        {"title": "HR Director",       "department": "HR"},
        {"title": "Head of People",    "department": "People"},
        {"title": "People Manager",    "department": "People"},
    ],
    "people operations": [
        {"title": "Head of People Operations", "department": "People Operations"},
        {"title": "HR Operations Manager",     "department": "HR Operations"},
        {"title": "HR Business Partner",       "department": "HR"},
    ],
}

_DEMO_CONTACTS_DEFAULT = [
    {"title": "HR Director",       "department": "HR"},
    {"title": "People Manager",    "department": "People"},
    {"title": "Training Manager",  "department": "Training"},
]

_DEMO_FIT_NOTE = "Demo contact generated for interface preview only"
_DEMO_SOURCE = "Demo placeholder"
_DEMO_SEARCH_LABEL = "Demo placeholder mode — no API call made"

# Contact output field names
CONTACT_FIELDS_BASE = [
    "name", "title", "email", "phone", "linkedin_url",
    "department", "seniority", "source", "fit_notes",
]

CONTACT_OUTPUT_FIELDS: list[str] = []
for _i in range(1, 4):
    for _f in CONTACT_FIELDS_BASE:
        CONTACT_OUTPUT_FIELDS.append(f"contact_{_i}_{_f}")

CONTACT_META_FIELDS = [
    "contact_data_status",
    "contact_search_used",
    "contact_lookup_timestamp",
    "contact_lookup_error",
    "contact_credits_estimated",
    "contact_credits_used",
]

ALL_CONTACT_FIELDS = CONTACT_OUTPUT_FIELDS + CONTACT_META_FIELDS

# Excel sheet colour palette (mirrors Lead Prioritizer)
_TIER_FILLS: dict = {
    "🥇 Hot":  "D6E4F7",
    "🥈 Warm": "D9EAD3",
    "🥉 Cool": "FCE5CD",
    "❄️ Pass": "F4CCCC",
    "Hot":     "D6E4F7",
    "Warm":    "D9EAD3",
    "Cool":    "FCE5CD",
    "Pass":    "F4CCCC",
}

_REC_FILLS: dict = {
    "Call now":                 "D6E4F7",
    "Call this month":          "D9EAD3",
    "Call before budget cycle": "FCE5CD",
    "Manual research needed":   "FFF2CC",
    "Monitor":                  "F2F2F2",
    "Low priority":             "F2F2F2",
    "Internal / exclude":       "F4CCCC",
}

# Sheets to preserve from input (pass-through if present)
_PRESERVE_SHEETS = [
    "Opportunity Radar",
    "Trigger Evidence",
    "Buying Window",
    "Contact Route",
    "Caller Brief",
    "Raw Sources",
    "Original Lead Scores",
    "Original Company Profiles",
    "Original Opportunity Input",
]

# =============================================================================
# SESSION STATE HELPERS
# =============================================================================


def ss(key: str, default=None):
    return st.session_state.get(key, default)


def ss_set(**kwargs):
    for k, v in kwargs.items():
        st.session_state[k] = v


def reset_bcf():
    ss_set(
        _bcf_file_key="__none__",
        _bcf_df_raw=None,
        _bcf_df_original=None,
        _bcf_file_name=None,
        _bcf_file_error=None,
        _bcf_input_bytes=None,
        _bcf_selection_mode=None,
        _bcf_top_n=50,
        _bcf_selected_indices=None,
        _bcf_processing=False,
        _bcf_done=False,
        _bcf_results_df=None,
        _bcf_summary=None,
        _bcf_raw_evidence=None,
        _bcf_excel_bytes=None,
        _bcf_all_sheets=None,
    )


# =============================================================================
# COLUMN DETECTION
# =============================================================================

_NAME_CANDIDATES = [
    "company_name", "canonical_company_name", "company name",
    "company", "name",
]
_DOMAIN_CANDIDATES = [
    "validated_domain", "canonical_company_url", "canonical_company_domain",
    "company_domain", "company_url", "domain",
    "company website", "website", "url",
]
_REC_CANDIDATES = [
    "call_recommendation", "recommendation", "call recommendation",
]
_TIER_CANDIDATES = [
    "commercial_tier", "tier",
]
_SCORE_CANDIDATES = [
    "opportunity_score", "commercial_fit_score", "final_commercial_fit_score",
    "score",
]
_ROUTE_CANDIDATES = [
    "preferred_buyer_route", "buyer_route", "contact_route",
]


def _detect_col(df: pd.DataFrame, candidates: list) -> str | None:
    lower_map = {c.lower(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in lower_map:
            return lower_map[cand.lower()]
    return None


def _normalize_domain(raw: str) -> str:
    if not raw:
        return ""
    raw = str(raw).strip()
    raw = re.sub(r"^https?://", "", raw, flags=re.IGNORECASE)
    raw = raw.split("/")[0].split("?")[0]
    if raw.lower().startswith("www."):
        raw = raw[4:]
    return raw.strip().lower()


# =============================================================================
# EXISTING CONTACT DETECTION
# =============================================================================


def _has_existing_contacts(row: pd.Series) -> bool:
    """Return True if the row already has at least one contact name filled."""
    for i in range(1, 4):
        v = row.get(f"contact_{i}_name", "")
        if v and str(v).strip() not in ("", "nan", "None"):
            return True
    return False


def _is_val_blank(v) -> bool:
    return not v or str(v).strip() in ("", "nan", "None", "NaN")


# =============================================================================
# FILE LOADING
# =============================================================================


def _load_file(uploaded_file) -> tuple[pd.DataFrame | None, dict, str | None]:
    """
    Returns (caller_prep_df, all_sheet_data_dict, error_msg).
    caller_prep_df is None if the sheet is missing.
    all_sheet_data_dict maps sheet_name -> DataFrame for every sheet.
    """
    try:
        raw_bytes = uploaded_file.read()
        xf = pd.ExcelFile(io.BytesIO(raw_bytes))
        all_sheets: dict[str, pd.DataFrame] = {}
        for sn in xf.sheet_names:
            try:
                all_sheets[sn] = xf.parse(sn)
            except Exception:
                pass

        if SHEET_CALLER_PREP not in all_sheets:
            return None, all_sheets, (
                f"'{SHEET_CALLER_PREP}' sheet not found. "
                "Please upload an Opportunity Radar export."
            )

        df = all_sheets[SHEET_CALLER_PREP].copy()
        df = df.where(pd.notna(df), other="")
        return df, all_sheets, None
    except Exception as exc:
        return None, {}, f"Could not read file: {exc}"


# =============================================================================
# SUMMARY COMPUTATION
# =============================================================================


def _compute_summary(df: pd.DataFrame) -> dict:
    rec_col = _detect_col(df, _REC_CANDIDATES)
    tier_col = _detect_col(df, _TIER_CANDIDATES)

    total = len(df)

    def _count(col, value):
        if col and col in df.columns:
            return int((df[col].astype(str).str.strip() == value).sum())
        return 0

    already_have_contacts = int(
        df.apply(_has_existing_contacts, axis=1).sum()
    )
    missing_contacts = total - already_have_contacts

    summary = {
        "total_companies": total,
        "already_have_contacts": already_have_contacts,
        "missing_contacts": missing_contacts,
        "call_now": _count(rec_col, "Call now"),
        "call_this_month": _count(rec_col, "Call this month"),
        "manual_research_first": _count(rec_col, "Manual research needed"),
        "monitor": _count(rec_col, "Monitor"),
        "low_priority": _count(rec_col, "Low priority"),
        "tier_hot": _count(tier_col, "🥇 Hot") + _count(tier_col, "Hot"),
        "tier_warm": _count(tier_col, "🥈 Warm") + _count(tier_col, "Warm"),
        "tier_cool": _count(tier_col, "🥉 Cool") + _count(tier_col, "Cool"),
        "tier_pass": _count(tier_col, "❄️ Pass") + _count(tier_col, "Pass"),
    }
    return summary


# =============================================================================
# SELECTION LOGIC
# =============================================================================

SELECTION_MODES = [
    "Call now + Manual research first (top 50 by opportunity_score) — default",
    "All companies in Caller Prep Input",
    "Only Call now",
    "Only Manual research first",
    "Only Hot commercial tier",
    "Hot + Warm commercial tier",
    "Top N by opportunity_score",
    "Top N by commercial_fit_score",
    "Manual selection from table",
]


def _apply_selection(
    df: pd.DataFrame,
    mode: str,
    top_n: int = 50,
    manual_indices: list | None = None,
) -> tuple[list[int], int]:
    """
    Return (selected_row_indices, estimated_lookup_count).
    Indices refer to df.index positions (iloc-based).
    """
    rec_col = _detect_col(df, _REC_CANDIDATES)
    tier_col = _detect_col(df, _TIER_CANDIDATES)
    score_col = _detect_col(df, _SCORE_CANDIDATES)

    def _rec_match(values: list[str]) -> pd.Series:
        if rec_col:
            return df[rec_col].astype(str).str.strip().isin(values)
        return pd.Series([False] * len(df), index=df.index)

    def _tier_match(values: list[str]) -> pd.Series:
        if tier_col:
            tier_ser = df[tier_col].astype(str).str.strip()
            masks = [tier_ser == v for v in values]
            result = masks[0]
            for m in masks[1:]:
                result = result | m
            return result
        return pd.Series([False] * len(df), index=df.index)

    def _top_n_by(col: str | None, n: int) -> list[int]:
        if col and col in df.columns:
            try:
                scores = pd.to_numeric(df[col], errors="coerce").fillna(0)
                return scores.nlargest(n).index.tolist()
            except Exception:
                pass
        return list(range(min(n, len(df))))

    if "All companies" in mode:
        selected = list(range(len(df)))

    elif "Only Call now" in mode and "Manual" not in mode:
        mask = _rec_match(["Call now"])
        selected = df.index[mask].tolist()

    elif "Only Manual" in mode:
        mask = _rec_match(["Manual research needed"])
        selected = df.index[mask].tolist()

    elif "Call now + Manual" in mode:
        mask = _rec_match(["Call now", "Manual research needed", "Call this month"])
        sub = df[mask]
        if score_col and score_col in df.columns:
            try:
                scores = pd.to_numeric(sub[score_col], errors="coerce").fillna(0)
                selected = scores.nlargest(top_n).index.tolist()
            except Exception:
                selected = sub.index.tolist()[:top_n]
        else:
            selected = sub.index.tolist()[:top_n]

    elif "Only Hot commercial" in mode:
        mask = _tier_match(["🥇 Hot", "Hot"])
        selected = df.index[mask].tolist()

    elif "Hot + Warm" in mode:
        mask = _tier_match(["🥇 Hot", "Hot", "🥈 Warm", "Warm"])
        selected = df.index[mask].tolist()

    elif "Top N by opportunity_score" in mode:
        opp_col = _detect_col(df, ["opportunity_score"])
        selected = _top_n_by(opp_col, top_n)

    elif "Top N by commercial_fit_score" in mode:
        fit_col = _detect_col(df, ["commercial_fit_score", "final_commercial_fit_score"])
        selected = _top_n_by(fit_col, top_n)

    elif "Manual selection" in mode:
        selected = [i for i in (manual_indices or []) if 0 <= i < len(df)]

    else:
        selected = list(range(min(top_n, len(df))))

    # Exclude rows that already have contacts (unless explicitly refreshing)
    already = [i for i in selected if _has_existing_contacts(df.iloc[i])]
    net_lookups = len(selected) - len(already)
    return selected, net_lookups


# =============================================================================
# TITLE RELEVANCE SCORING
# =============================================================================


def _score_title(title: str, preferred_route: str = "") -> int:
    t = title.lower().strip()
    if not t:
        return -1

    # Route-specific priority boost
    route = (preferred_route or "").lower()
    route_key = None
    for k in _ROUTE_PRIORITY_TITLES:
        if k in route:
            route_key = k
            break

    if route_key:
        for priority_title in _ROUTE_PRIORITY_TITLES[route_key]:
            if priority_title in t:
                return 4  # top priority for route match

    for keyword, score in _TITLE_SCORE_MAP:
        if keyword in t:
            return score

    return -1  # unknown title


def _rank_contacts(contacts: list[dict], preferred_route: str = "") -> list[dict]:
    """Sort contacts by title relevance score (descending), return top 3."""
    scored = []
    for c in contacts:
        title = c.get("title", "")
        s = _score_title(title, preferred_route)
        scored.append((s, c))
    scored.sort(key=lambda x: x[0], reverse=True)
    return [c for _, c in scored[:3]]


def _contact_fit_notes(title: str, preferred_route: str = "") -> str:
    score = _score_title(title, preferred_route)
    if score >= 4:
        return f"Route match: aligns with preferred buyer route ({preferred_route})"
    if score >= 3:
        return "High relevance: direct buyer title"
    if score >= 2:
        return "Medium relevance: adjacent HR/L&D function"
    if score >= 1:
        return "Low relevance: broad HR/people keyword only"
    if score == 0:
        return "Low relevance: executive fallback (no buyer-function title found)"
    return "Unknown relevance"


# =============================================================================
# LUSHA PERSON API — PREFLIGHT + SEARCH
# =============================================================================

# Cached at session level so we only probe once per session
_LUSHA_ACTIVE_ENDPOINT: str | None = None   # set after successful preflight


def _lusha_headers(api_key: str) -> dict:
    return {"api_key": api_key, "Content-Type": "application/json"}


def _classify_http_error(status_code: int) -> str:
    """Return 'not_supported' | 'auth_error' | 'other'."""
    if status_code in _LUSHA_UNSUPPORTED_CODES:
        return "not_supported"
    if status_code in _LUSHA_AUTH_CODES:
        return "auth_error"
    return "other"


def _preflight_lusha_contact(api_key: str, test_domain: str) -> tuple[str, str, str]:
    """
    Probe all known Lusha person endpoint candidates with a single company.

    Returns (outcome, active_endpoint_or_empty, message).
    outcome: "ok" | "not_supported" | "auth_error" | "network_error"

    Side-effect: stores the working endpoint in the session state key
    "_bcf_lusha_active_endpoint" so subsequent calls skip the probe.
    """
    global _LUSHA_ACTIVE_ENDPOINT

    # Already resolved in this process (avoids re-probe on rerun within same session)
    cached = st.session_state.get("_bcf_lusha_active_endpoint")
    if cached:
        return "ok", cached, f"Using cached endpoint: {cached}"

    headers = _lusha_headers(api_key)
    norm = _normalize_domain(test_domain) or test_domain

    # Param variants — different endpoints use different param names
    param_variants = [
        {"companyDomain": norm, "limit": 1},
        {"company": norm, "limit": 1},
        {"domain": norm, "limit": 1},
    ]

    last_status = None
    last_msg = ""

    for endpoint in _LUSHA_PERSON_CANDIDATES:
        for params in param_variants:
            try:
                resp = requests.get(
                    endpoint, headers=headers, params=params, timeout=_LUSHA_TIMEOUT
                )
                if resp.status_code == 200:
                    st.session_state["_bcf_lusha_active_endpoint"] = endpoint
                    _LUSHA_ACTIVE_ENDPOINT = endpoint
                    return "ok", endpoint, f"Endpoint confirmed: {endpoint}"
                last_status = resp.status_code
                last_msg = f"HTTP {resp.status_code} from {endpoint}"
                classification = _classify_http_error(resp.status_code)
                if classification == "auth_error":
                    return "auth_error", "", (
                        f"Lusha API key rejected (HTTP {resp.status_code}). "
                        "Check that LUSHA_API_KEY is correct."
                    )
                # 404/405 — try next endpoint
            except requests.exceptions.Timeout:
                last_msg = f"Timeout on {endpoint}"
            except Exception as exc:
                last_msg = f"Network error on {endpoint}: {exc}"

    # All candidates failed
    if last_status in _LUSHA_UNSUPPORTED_CODES:
        return "not_supported", "", (
            "Lusha contact endpoint not available for this API setup "
            f"(last response: HTTP {last_status}). "
            "Company-level Lusha enrichment may still work, but contact "
            "enrichment requires a valid people/contact endpoint — "
            "check your Lusha plan or contact Lusha support."
        )
    return "network_error", "", f"Could not reach Lusha contact API: {last_msg}"


def _parse_lusha_people_response(raw, preferred_route: str) -> list[dict]:
    """Extract and rank up to 3 buyer contacts from a raw Lusha API response."""
    people_raw: list = []
    if isinstance(raw, list):
        people_raw = raw
    elif isinstance(raw, dict):
        for path in ("data.result", "data.contacts", "data.people", "data",
                     "results", "people", "contacts", "contact"):
            if "." in path:
                node = raw
                for p in path.split("."):
                    node = node.get(p, {}) if isinstance(node, dict) else {}
                if isinstance(node, list):
                    people_raw = node
                    break
            else:
                val = raw.get(path)
                if isinstance(val, list):
                    people_raw = val
                    break
                elif isinstance(val, dict) and val:
                    people_raw = [val]
                    break

    contacts = []
    for person in people_raw:
        if not isinstance(person, dict):
            continue

        def _sv(*keys):
            for k in keys:
                v = person.get(k)
                if v and str(v).strip() not in ("", "None", "null", "nan"):
                    return str(v).strip()
            return ""

        first = _sv("firstName", "first_name", "givenName")
        last = _sv("lastName", "last_name", "familyName")
        full_name = _sv("fullName", "full_name", "name")
        if not full_name and (first or last):
            full_name = f"{first} {last}".strip()

        email = _sv("email")
        if not email:
            emails = person.get("emails") or []
            if isinstance(emails, list) and emails:
                first_e = emails[0]
                email = str(first_e.get("email", "") if isinstance(first_e, dict) else first_e).strip()

        phone = _sv("phone", "phoneNumber", "phone_number")
        if not phone:
            phones = person.get("phones") or []
            if isinstance(phones, list) and phones:
                first_p = phones[0]
                phone = str(first_p.get("phoneNumber", "") if isinstance(first_p, dict) else first_p).strip()

        linkedin = _sv("linkedInUrl", "linkedin_url", "linkedinUrl", "linkedin")
        title = _sv("title", "jobTitle", "job_title", "position")
        dept = _sv("department", "dept")
        seniority = _sv("seniority", "seniorityLevel", "level")

        if not full_name and not title:
            continue

        contacts.append({
            "name": full_name,
            "title": title,
            "email": email,
            "phone": phone,
            "linkedin_url": linkedin,
            "department": dept,
            "seniority": seniority,
            "source": "Lusha",
            "fit_notes": _contact_fit_notes(title, preferred_route),
        })

    return _rank_contacts(contacts, preferred_route)


def _lusha_person_search(
    domain: str,
    company_name: str,
    preferred_route: str,
    api_key: str,
) -> tuple[list[dict], str, str | None, str]:
    """
    Query the active Lusha person endpoint for a single company.

    Returns (contacts, search_label, error_or_None, outcome).
    outcome: "ok" | "not_supported" | "auth_error" | "no_results" | "error"

    Caller must ensure preflight passed before calling this.
    Active endpoint is read from session state.
    """
    endpoint = st.session_state.get("_bcf_lusha_active_endpoint", "")
    if not endpoint:
        return [], "no active endpoint", "Preflight not run — no active endpoint.", "error"

    headers = _lusha_headers(api_key)
    norm = _normalize_domain(domain) if domain else ""

    if norm:
        params = {"companyDomain": norm, "limit": 10}
        search_label = f"companyDomain={norm}"
    elif company_name:
        params = {"companyName": company_name.strip(), "limit": 10}
        search_label = f"companyName={company_name.strip()}"
    else:
        return [], "no domain or name", "Missing domain and company name.", "error"

    try:
        resp = requests.get(endpoint, headers=headers, params=params, timeout=_LUSHA_TIMEOUT)
        if resp.status_code in _LUSHA_UNSUPPORTED_CODES:
            return [], search_label, (
                f"Contact endpoint not supported (HTTP {resp.status_code})"
            ), "not_supported"
        if resp.status_code in _LUSHA_AUTH_CODES:
            return [], search_label, (
                f"API key rejected (HTTP {resp.status_code})"
            ), "auth_error"
        resp.raise_for_status()
        raw = resp.json()
    except requests.exceptions.HTTPError as exc:
        code = exc.response.status_code if exc.response is not None else "?"
        return [], search_label, f"HTTP {code}: {exc}", "error"
    except requests.exceptions.Timeout:
        return [], search_label, "Request timed out", "error"
    except Exception as exc:
        return [], search_label, str(exc), "error"

    contacts = _parse_lusha_people_response(raw, preferred_route)
    if not contacts:
        return [], search_label, None, "no_results"
    return contacts, search_label, None, "ok"


# =============================================================================
# DEMO PLACEHOLDER GENERATOR
# =============================================================================


def _demo_contacts_for_route(preferred_route: str) -> list[dict]:
    """Return the 3 demo contact templates for the given buyer route."""
    route = (preferred_route or "").lower()
    for key, templates in _DEMO_CONTACTS_BY_ROUTE.items():
        if key in route:
            return templates
    return _DEMO_CONTACTS_DEFAULT


def _generate_demo_row(row: pd.Series, selected: bool, refresh_existing: bool = False) -> dict:
    """
    Return contact field updates populated with clearly artificial placeholder data.
    No API calls, no credits consumed.
    """
    result: dict = {}

    if not selected:
        existing_status = str(row.get("contact_data_status", "")).strip()
        if not existing_status or existing_status in (STATUS_NOT_RUN, ""):
            result["contact_data_status"] = STATUS_SKIPPED_NOT_SELECTED
        return result

    if _has_existing_contacts(row) and not refresh_existing:
        result["contact_data_status"] = STATUS_SKIPPED_HAS_CONTACTS
        return result

    route_val = ""
    for c in _ROUTE_CANDIDATES:
        v = row.get(c, "")
        if v and str(v).strip() not in ("", "nan", "Unknown"):
            route_val = str(v).strip()
            break

    templates = _demo_contacts_for_route(route_val)

    for i, tpl in enumerate(templates[:3], start=1):
        result[f"contact_{i}_name"] = f"Demo Contact {i}"
        result[f"contact_{i}_title"] = tpl["title"]
        result[f"contact_{i}_email"] = f"demo.contact{i}@example.com"
        result[f"contact_{i}_phone"] = ""
        result[f"contact_{i}_linkedin_url"] = ""
        result[f"contact_{i}_department"] = tpl.get("department", "")
        result[f"contact_{i}_seniority"] = ""
        result[f"contact_{i}_source"] = _DEMO_SOURCE
        result[f"contact_{i}_fit_notes"] = _DEMO_FIT_NOTE

    result["contact_data_status"] = STATUS_DEMO
    result["contact_search_used"] = _DEMO_SEARCH_LABEL
    result["contact_lookup_timestamp"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    result["contact_credits_estimated"] = 0
    result["contact_credits_used"] = 0
    result["contact_lookup_error"] = ""
    return result


# =============================================================================
# ENRICHMENT RUNNER
# =============================================================================


def _enrich_row(
    row: pd.Series,
    api_key: str,
    selected: bool,
    refresh_existing: bool = False,
) -> dict:
    """
    Returns a dict of contact fields to update for this row.
    Does NOT modify the original row directly.
    """
    result: dict = {}

    if not selected:
        existing_status = str(row.get("contact_data_status", "")).strip()
        if not existing_status or existing_status in (STATUS_NOT_RUN, ""):
            result["contact_data_status"] = STATUS_SKIPPED_NOT_SELECTED
        return result

    if _has_existing_contacts(row) and not refresh_existing:
        result["contact_data_status"] = STATUS_SKIPPED_HAS_CONTACTS
        return result

    name_col_val = ""
    for c in _NAME_CANDIDATES:
        v = row.get(c, "")
        if v and str(v).strip() not in ("", "nan"):
            name_col_val = str(v).strip()
            break

    domain_val = ""
    for c in _DOMAIN_CANDIDATES:
        v = row.get(c, "")
        if v and str(v).strip() not in ("", "nan"):
            domain_val = str(v).strip()
            break

    route_val = ""
    for c in _ROUTE_CANDIDATES:
        v = row.get(c, "")
        if v and str(v).strip() not in ("", "nan", "Unknown"):
            route_val = str(v).strip()
            break

    contacts, search_used, error, outcome = _lusha_person_search(
        domain_val, name_col_val, route_val, api_key
    )

    result["contact_search_used"] = search_used
    result["contact_lookup_timestamp"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    result["contact_credits_estimated"] = 1
    result["contact_credits_used"] = ""

    if outcome == "not_supported":
        result["contact_data_status"] = STATUS_NOT_SUPPORTED
        result["contact_lookup_error"] = (
            error or "Lusha contact endpoint not available for this API setup"
        )
        return result

    if error:
        result["contact_data_status"] = STATUS_ERROR
        result["contact_lookup_error"] = error
        return result

    result["contact_lookup_error"] = ""

    if not contacts:
        result["contact_data_status"] = STATUS_NO_CONTACTS
        return result

    result["contact_data_status"] = STATUS_CONTACTS_FOUND
    for i, contact in enumerate(contacts[:3], start=1):
        for field in CONTACT_FIELDS_BASE:
            result[f"contact_{i}_{field}"] = contact.get(field, "")

    return result


# =============================================================================
# EXCEL BUILDER
# =============================================================================


def _style_header_row(ws, header_fill: str = "0B1F3A", font_color: str = "FFFFFF"):
    hdr_fill = PatternFill("solid", fgColor=header_fill)
    hdr_font = Font(bold=True, color=font_color)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(wrap_text=False, vertical="center")


def _auto_col_widths(ws, max_width: int = 60):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                if len(val) > max_len:
                    max_len = len(val)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _df_to_sheet(ws, df: pd.DataFrame, freeze_row: bool = True):
    """Write a DataFrame into an existing worksheet."""
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))
    _style_header_row(ws)
    _auto_col_widths(ws)
    if freeze_row:
        ws.freeze_panes = "A2"


def _build_excel(
    enriched_df: pd.DataFrame,
    original_df: pd.DataFrame,
    summary: dict,
    raw_evidence: list[dict],
    all_sheets: dict[str, pd.DataFrame],
) -> bytes:
    wb = openpyxl.Workbook()

    # Sheet 1: Caller Prep Input (enriched) — must be first
    ws1 = wb.active
    ws1.title = SHEET_CALLER_PREP
    _df_to_sheet(ws1, enriched_df)

    # Apply tier/rec colouring on enriched sheet
    rec_col_idx = None
    tier_col_idx = None
    for idx, col in enumerate(enriched_df.columns, start=1):
        if col.lower() in ("call_recommendation", "recommendation"):
            rec_col_idx = idx
        if col.lower() in ("commercial_tier", "tier"):
            tier_col_idx = idx

    for row_idx, row in enumerate(enriched_df.itertuples(index=False, name=None), start=2):
        row_dict = dict(zip(enriched_df.columns, row))
        rec_val = str(row_dict.get("call_recommendation", row_dict.get("recommendation", ""))).strip()
        tier_val = str(row_dict.get("commercial_tier", row_dict.get("tier", ""))).strip()
        fill_hex = _TIER_FILLS.get(tier_val) or _REC_FILLS.get(rec_val)
        if fill_hex:
            fill = PatternFill("solid", fgColor=fill_hex)
            for col_idx in range(1, len(enriched_df.columns) + 1):
                ws1.cell(row=row_idx, column=col_idx).fill = fill

    # Sheet 2: Contact Enrichment Summary
    ws2 = wb.create_sheet("Contact Enrichment Summary")
    summary_rows = [
        ["Metric", "Value"],
        ["Total companies in Caller Prep Input", summary.get("total_companies", "")],
        ["Estimated lookups selected", summary.get("credits_estimated", "")],
        ["Actual lookups attempted", summary.get("attempted_lookups", "")],
        ["Selected for this run", summary.get("selected_count", "")],
        ["Skipped (not selected)", summary.get("skipped_not_selected", "")],
        ["Skipped (already had contacts)", summary.get("skipped_had_contacts", "")],
        ["Contacts found", summary.get("contacts_found", "")],
        ["No contacts found (searched, none returned)", summary.get("no_contacts_found", "")],
        ["Contact endpoint not supported", summary.get("endpoint_not_supported", "")],
        ["Demo placeholder contacts generated", summary.get("demo_contacts_generated", "")],
        ["Lookup errors", summary.get("lookup_errors", "")],
        ["Credits used (actual billed lookups)", summary.get("credits_used", "")],
        [],
        ["Preflight outcome", summary.get("preflight_outcome", "")],
        ["Preflight message", summary.get("preflight_message", "")],
        ["Run timestamp", summary.get("run_timestamp", "")],
        ["Input file", summary.get("input_file", "")],
    ]
    for r in summary_rows:
        ws2.append(r)
    _style_header_row(ws2)
    _auto_col_widths(ws2)

    # Sheet 3: Contact Review Needed
    review_mask = (
        enriched_df.get("contact_data_status", pd.Series(dtype=str))
        .astype(str)
        .isin([STATUS_NO_CONTACTS, STATUS_ERROR, STATUS_NOT_RUN])
    )
    # Also include rows where only low-relevance contacts found
    for i in range(1, 4):
        fit_col = f"contact_{i}_fit_notes"
        if fit_col in enriched_df.columns:
            low_mask = enriched_df[fit_col].astype(str).str.contains(
                "Low relevance|Unknown relevance", na=False
            )
            review_mask = review_mask | low_mask
    review_df = enriched_df[review_mask].copy() if review_mask.any() else enriched_df.iloc[:0].copy()
    ws3 = wb.create_sheet("Contact Review Needed")
    _df_to_sheet(ws3, review_df)

    # Sheet 4: Original Caller Prep Input (unchanged)
    ws4 = wb.create_sheet("Original Caller Prep Input")
    _df_to_sheet(ws4, original_df)

    # Sheet 5: Contact Raw Evidence
    if raw_evidence:
        ev_df = pd.DataFrame(raw_evidence)
    else:
        ev_df = pd.DataFrame(columns=[
            "company_name", "domain", "search_used",
            "contact_name", "contact_title", "contact_company",
            "contact_email", "contact_phone", "contact_linkedin_url",
            "relevance_score", "selected_or_rejected", "reason",
        ])
    ws5 = wb.create_sheet("Contact Raw Evidence")
    _df_to_sheet(ws5, ev_df)

    # Pass-through sheets from the input file (if present)
    for sheet_name in _PRESERVE_SHEETS:
        if sheet_name in all_sheets and sheet_name != SHEET_CALLER_PREP:
            ws_pt = wb.create_sheet(sheet_name)
            _df_to_sheet(ws_pt, all_sheets[sheet_name])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================================
# MAIN STREAMLIT UI
# =============================================================================

def _safe_str(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("nan", "none", "nat") else s


def main():
    # ── API key ───────────────────────────────────────────────────────────────
    lusha_api_key = ""
    try:
        lusha_api_key = (st.secrets.get("LUSHA_API_KEY", "") or "").strip()
    except Exception:
        pass

    # ── Upload ────────────────────────────────────────────────────────────────
    st.markdown("### Step 1 — Upload Opportunity Radar export")
    st.caption(
        "Upload the Excel file produced by Opportunity Radar. "
        "This tool reads the **Caller Prep Input** sheet."
    )

    uploaded = st.file_uploader(
        "Opportunity Radar export (.xlsx)",
        type=["xlsx"],
        key="bcf_uploader",
        help="Must contain a 'Caller Prep Input' sheet.",
    )

    if uploaded is None:
        st.info("Upload an Opportunity Radar export to begin.")
        return

    file_key = f"{uploaded.name}_{uploaded.size}"
    if ss("_bcf_file_key") != file_key:
        reset_bcf()
        ss_set(_bcf_file_key=file_key, _bcf_file_name=uploaded.name)

    if ss("_bcf_df_raw") is None:
        df_cpi, all_sheets, err = _load_file(uploaded)
        if err:
            st.error(f"❌ {err}")
            return
        ss_set(
            _bcf_df_raw=df_cpi.copy(),
            _bcf_df_original=df_cpi.copy(),
            _bcf_all_sheets=all_sheets,
        )

    df: pd.DataFrame = ss("_bcf_df_raw")
    all_sheets: dict = ss("_bcf_all_sheets") or {}

    st.success(
        f"✅ **{SHEET_CALLER_PREP}** sheet found — "
        f"{len(df)} companies loaded from `{ss('_bcf_file_name')}`"
    )

    # ── Summary ───────────────────────────────────────────────────────────────
    summary_data = _compute_summary(df)
    st.markdown("### Step 2 — Caller Prep Input summary")

    col_a, col_b, col_c = st.columns(3)
    with col_a:
        st.metric("Total companies", summary_data["total_companies"])
        st.metric("Already have contacts", summary_data["already_have_contacts"])
        st.metric("Missing contacts", summary_data["missing_contacts"])
        st.caption(
            "ℹ️ Not every company missing contacts will be enriched. "
            "Only the companies you select below will use Lusha credits in this run."
        )
    with col_b:
        st.metric("Call now", summary_data["call_now"])
        st.metric("Call this month", summary_data["call_this_month"])
        st.metric("Manual research first", summary_data["manual_research_first"])
        st.metric("Monitor / Low priority", summary_data["monitor"] + summary_data["low_priority"])
    with col_c:
        st.metric("🥇 Hot", summary_data["tier_hot"])
        st.metric("🥈 Warm", summary_data["tier_warm"])
        st.metric("🥉 Cool", summary_data["tier_cool"])
        st.metric("❄️ Pass", summary_data["tier_pass"])

    # ── Selection ─────────────────────────────────────────────────────────────
    st.markdown("### Step 3 — Select companies for contact enrichment")
    st.caption(
        "Lusha lookups consume credits. Choose carefully to avoid unnecessary spend."
    )

    sel_mode = st.selectbox(
        "Selection method",
        SELECTION_MODES,
        index=0,
    )

    top_n = 50
    manual_indices: list[int] = []

    if "Top N" in sel_mode:
        top_n = st.number_input(
            "Top N companies",
            min_value=1,
            max_value=len(df),
            value=min(50, len(df)),
            step=5,
        )

    elif "Manual selection" in sel_mode:
        name_col = _detect_col(df, _NAME_CANDIDATES)
        rec_col = _detect_col(df, _REC_CANDIDATES)
        display_cols = [c for c in [name_col, rec_col] if c]
        if not display_cols:
            display_cols = list(df.columns[:3])
        display_df = df[display_cols].copy()
        display_df.insert(0, "_select", False)
        edited = st.data_editor(
            display_df,
            use_container_width=True,
            num_rows="fixed",
            column_config={"_select": st.column_config.CheckboxColumn("Select")},
            key="bcf_manual_editor",
        )
        manual_indices = edited.index[edited["_select"] == True].tolist()

    refresh_existing = st.checkbox(
        "Refresh / overwrite existing contacts",
        value=False,
        help="By default, companies that already have contact data are skipped.",
    )

    selected_indices, net_lookups = _apply_selection(
        df, sel_mode, int(top_n), manual_indices
    )

    if refresh_existing:
        net_lookups = len(selected_indices)

    # Preview selected + run breakdown
    name_col = _detect_col(df, _NAME_CANDIDATES)
    rec_col = _detect_col(df, _REC_CANDIDATES)
    tier_col = _detect_col(df, _TIER_CANDIDATES)
    route_col = _detect_col(df, _ROUTE_CANDIDATES)

    already_selected = [i for i in selected_indices if _has_existing_contacts(df.iloc[i])]
    not_selected_count = len(df) - len(selected_indices)

    st.markdown("---")
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Selected for this run", len(selected_indices))
    m2.metric("Not selected", not_selected_count)
    m3.metric(
        "Already have contacts",
        len(already_selected),
        help="These are selected but will be skipped unless 'Refresh' is on.",
    )
    m4.metric("Estimated lookups selected", net_lookups)

    if net_lookups == 0 and len(selected_indices) > 0:
        st.info(
            "All selected companies already have contacts. "
            "Enable 'Refresh / overwrite existing contacts' to re-run them."
        )
    elif net_lookups == 0:
        st.info("No companies selected — adjust the selection method above.")

    preview_cols = [c for c in [name_col, rec_col, tier_col, route_col] if c]
    if selected_indices and preview_cols:
        preview_df = df.iloc[selected_indices][preview_cols].head(20).copy()
        st.markdown(f"**Preview — first 20 of {len(selected_indices)} selected companies:**")
        st.dataframe(preview_df, use_container_width=True, height=300)

    if not lusha_api_key:
        st.warning(
            "⚠ Lusha API key not found in Streamlit secrets. "
            "Add `LUSHA_API_KEY = \"your-key\"` to `.streamlit/secrets.toml`."
        )

    # ── Confirmation + Run ────────────────────────────────────────────────────
    st.markdown("### Step 4 — Run contact enrichment")

    demo_mode = st.checkbox(
        "🧪 Generate demo placeholder contacts instead of real Lusha lookup",
        value=False,
        key="bcf_demo_mode",
        help=(
            "For interface preview only. Populates artificial contacts based on "
            "preferred_buyer_route. No API calls, no credits used. "
            "MUST NOT be used for real outreach."
        ),
    )

    if demo_mode:
        st.warning(
            "⚠️ **Demo mode active.** Placeholder contacts will be generated — "
            "no real people, no real emails. "
            "These contacts are artificial and must not be used for real outreach."
        )
        confirm_label = (
            f"I understand these are demo placeholders and will enrich "
            f"{len(selected_indices)} selected companies with artificial contacts."
        )
        run_disabled_extra = len(selected_indices) == 0
    else:
        confirm_label = (
            f"I confirm I want to enrich the selected {net_lookups} companies with Lusha "
            f"(estimated {net_lookups} lookup credit{'' if net_lookups == 1 else 's'})."
        )
        run_disabled_extra = not lusha_api_key or net_lookups == 0

    confirmed = st.checkbox(
        confirm_label,
        value=False,
        key="bcf_confirm",
    )

    run_disabled = (
        not confirmed
        or run_disabled_extra
        or ss("_bcf_processing")
        or ss("_bcf_done")
    )

    if st.button(
        "▶ Run Demo Enrichment" if demo_mode else "▶ Run Contact Enrichment",
        disabled=run_disabled,
        type="primary",
        key="bcf_run",
    ):
        ss_set(_bcf_processing=True, _bcf_done=False, _bcf_demo_mode=demo_mode)
        st.rerun()

    # ── Processing ────────────────────────────────────────────────────────────
    if ss("_bcf_processing") and not ss("_bcf_done"):
        enriched_df = df.copy()

        # Ensure all contact output columns exist
        for f in ALL_CONTACT_FIELDS:
            if f not in enriched_df.columns:
                enriched_df[f] = ""

        selected_set = set(selected_indices)
        raw_evidence: list[dict] = []

        # ── Demo mode fast-path — no API calls ───────────────────────────────
        if ss("_bcf_demo_mode"):
            demo_counts = {
                "demo_generated": 0, "skipped_had_contacts": 0,
                "skipped_not_selected": 0,
            }
            for iloc_i in range(len(enriched_df)):
                row = enriched_df.iloc[iloc_i]
                is_selected = iloc_i in selected_set
                updates = _generate_demo_row(row, is_selected, refresh_existing)
                for field, value in updates.items():
                    enriched_df.at[enriched_df.index[iloc_i], field] = value
                status = updates.get("contact_data_status", "")
                if status == STATUS_DEMO:
                    demo_counts["demo_generated"] += 1
                elif status == STATUS_SKIPPED_HAS_CONTACTS:
                    demo_counts["skipped_had_contacts"] += 1
                elif status == STATUS_SKIPPED_NOT_SELECTED:
                    demo_counts["skipped_not_selected"] += 1

            run_summary = {
                **summary_data,
                "selected_count": len(selected_indices),
                "attempted_lookups": demo_counts["demo_generated"],
                "skipped_not_selected": demo_counts["skipped_not_selected"],
                "skipped_had_contacts": demo_counts["skipped_had_contacts"],
                "contacts_found": 0,
                "no_contacts_found": 0,
                "lookup_errors": 0,
                "endpoint_not_supported": 0,
                "demo_contacts_generated": demo_counts["demo_generated"],
                "credits_estimated": 0,
                "credits_used": 0,
                "preflight_outcome": "demo_mode",
                "preflight_message": "Demo placeholder mode — no Lusha API calls made",
                "run_timestamp": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
                "input_file": ss("_bcf_file_name", ""),
            }
            excel_bytes = _build_excel(
                enriched_df, ss("_bcf_df_original"), run_summary, [], all_sheets
            )
            ss_set(
                _bcf_processing=False,
                _bcf_done=True,
                _bcf_results_df=enriched_df,
                _bcf_summary=run_summary,
                _bcf_raw_evidence=[],
                _bcf_excel_bytes=excel_bytes,
                _bcf_preflight_error=None,
            )
            st.rerun()

        # ── Preflight: test one company before running the full batch ─────────
        preflight_placeholder = st.empty()
        preflight_placeholder.info("🔍 Running preflight check on Lusha contact endpoint…")

        # Find the first selected row with a domain to use as the test case
        test_domain = ""
        test_company = ""
        for idx in selected_indices:
            row_test = enriched_df.iloc[idx]
            for c in _DOMAIN_CANDIDATES:
                v = row_test.get(c, "")
                if v and str(v).strip() not in ("", "nan"):
                    test_domain = str(v).strip()
                    break
            for c in _NAME_CANDIDATES:
                v = row_test.get(c, "")
                if v and str(v).strip() not in ("", "nan"):
                    test_company = str(v).strip()
                    break
            if test_domain or test_company:
                break

        pf_outcome, pf_endpoint, pf_msg = _preflight_lusha_contact(
            lusha_api_key, test_domain or test_company
        )
        preflight_placeholder.empty()

        if pf_outcome != "ok":
            # Endpoint unavailable — mark all selected rows, build output, stop
            not_supported_msg = (
                "Lusha contact endpoint not available for this API setup"
                if pf_outcome == "not_supported"
                else pf_msg
            )
            for iloc_i in range(len(enriched_df)):
                if iloc_i in selected_set:
                    enriched_df.at[enriched_df.index[iloc_i], "contact_data_status"] = (
                        STATUS_NOT_SUPPORTED if pf_outcome == "not_supported" else STATUS_ERROR
                    )
                    enriched_df.at[enriched_df.index[iloc_i], "contact_lookup_error"] = not_supported_msg
                    enriched_df.at[enriched_df.index[iloc_i], "contact_lookup_timestamp"] = (
                        datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
                    )
                elif str(enriched_df.iloc[iloc_i].get("contact_data_status", "")).strip() == "":
                    enriched_df.at[enriched_df.index[iloc_i], "contact_data_status"] = STATUS_SKIPPED_NOT_SELECTED

            ui_label = (
                "⚠️ Contact lookup endpoint unavailable or unsupported. "
                "No contacts were retrieved. Company-level Lusha enrichment may "
                "still work, but contact enrichment requires a valid people/contact "
                "endpoint — check your Lusha plan or contact Lusha support."
                if pf_outcome == "not_supported"
                else f"⚠️ Lusha API error during preflight: {pf_msg}"
            )

            run_summary = {
                **summary_data,
                "selected_count": len(selected_indices),
                "attempted_lookups": 0,
                "skipped_not_selected": len(enriched_df) - len(selected_indices),
                "skipped_had_contacts": 0,
                "contacts_found": 0,
                "no_contacts_found": 0,
                "lookup_errors": 0,
                "endpoint_not_supported": len(selected_indices) if pf_outcome == "not_supported" else 0,
                "credits_estimated": net_lookups,
                "credits_used": 0,
                "preflight_outcome": pf_outcome,
                "preflight_message": pf_msg,
                "run_timestamp": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
                "input_file": ss("_bcf_file_name", ""),
            }
            excel_bytes = _build_excel(
                enriched_df, ss("_bcf_df_original"), run_summary, [], all_sheets
            )
            ss_set(
                _bcf_processing=False,
                _bcf_done=True,
                _bcf_results_df=enriched_df,
                _bcf_summary=run_summary,
                _bcf_raw_evidence=[],
                _bcf_excel_bytes=excel_bytes,
                _bcf_preflight_error=ui_label,
            )
            st.rerun()

        # ── Preflight passed — run full batch ─────────────────────────────────
        progress_bar = st.progress(0.0, text="Starting contact lookup…")
        status_text = st.empty()

        counts = {
            "contacts_found": 0,
            "no_contacts_found": 0,
            "lookup_errors": 0,
            "not_supported": 0,
            "skipped_had_contacts": 0,
            "skipped_not_selected": 0,
            "credits_used": 0,
            "attempted": 0,
        }

        total_rows = len(enriched_df)
        lookup_count = 0

        for iloc_i in range(total_rows):
            row = enriched_df.iloc[iloc_i]
            is_selected = iloc_i in selected_set

            company_name_val = ""
            for c in _NAME_CANDIDATES:
                v = row.get(c, "")
                if v and str(v).strip() not in ("", "nan"):
                    company_name_val = str(v).strip()
                    break

            if is_selected:
                lookup_count += 1
                status_text.text(
                    f"Looking up {lookup_count} / {net_lookups}: {company_name_val}"
                )

            updates = _enrich_row(row, lusha_api_key, is_selected, refresh_existing)
            for field, value in updates.items():
                enriched_df.at[enriched_df.index[iloc_i], field] = value

            # Update counts
            status = updates.get("contact_data_status", "")
            if status == STATUS_CONTACTS_FOUND:
                counts["contacts_found"] += 1
                counts["credits_used"] += 1
                counts["attempted"] += 1
            elif status == STATUS_NO_CONTACTS:
                counts["no_contacts_found"] += 1
                counts["credits_used"] += 1
                counts["attempted"] += 1
            elif status == STATUS_NOT_SUPPORTED:
                counts["not_supported"] += 1
                counts["attempted"] += 1
            elif status == STATUS_ERROR:
                counts["lookup_errors"] += 1
                counts["attempted"] += 1
            elif status == STATUS_SKIPPED_HAS_CONTACTS:
                counts["skipped_had_contacts"] += 1
            elif status == STATUS_SKIPPED_NOT_SELECTED:
                counts["skipped_not_selected"] += 1

            # Collect raw evidence for selected rows
            if is_selected:
                domain_val = ""
                for c in _DOMAIN_CANDIDATES:
                    v = row.get(c, "")
                    if v and str(v).strip() not in ("", "nan"):
                        domain_val = str(v).strip()
                        break

                search_used = updates.get("contact_search_used", "")
                error_val = updates.get("contact_lookup_error", "")

                for i in range(1, 4):
                    name_v = updates.get(f"contact_{i}_name", "")
                    title_v = updates.get(f"contact_{i}_title", "")
                    if not name_v and not title_v:
                        continue
                    score = _score_title(
                        title_v,
                        str(row.get(route_col or "", "") if route_col else ""),
                    )
                    raw_evidence.append({
                        "company_name": company_name_val,
                        "domain": domain_val,
                        "search_used": search_used,
                        "contact_name": name_v,
                        "contact_title": title_v,
                        "contact_company": company_name_val,
                        "contact_email": updates.get(f"contact_{i}_email", ""),
                        "contact_phone": updates.get(f"contact_{i}_phone", ""),
                        "contact_linkedin_url": updates.get(f"contact_{i}_linkedin_url", ""),
                        "relevance_score": score,
                        "selected_or_rejected": "selected",
                        "reason": updates.get(f"contact_{i}_fit_notes", ""),
                    })

                if error_val:
                    raw_evidence.append({
                        "company_name": company_name_val,
                        "domain": domain_val,
                        "search_used": search_used,
                        "contact_name": "",
                        "contact_title": "",
                        "contact_company": "",
                        "contact_email": "",
                        "contact_phone": "",
                        "contact_linkedin_url": "",
                        "relevance_score": -1,
                        "selected_or_rejected": "error",
                        "reason": error_val,
                    })

            if is_selected:
                progress_bar.progress(
                    min(lookup_count / max(net_lookups, 1), 1.0),
                    text=f"Completed {lookup_count} / {net_lookups} lookups",
                )
                time.sleep(0.3)  # rate-limit courtesy delay

        progress_bar.progress(1.0, text="All lookups complete.")
        status_text.empty()

        run_summary = {
            **summary_data,
            "selected_count": len(selected_indices),
            "attempted_lookups": counts["attempted"],
            "skipped_not_selected": counts["skipped_not_selected"],
            "skipped_had_contacts": counts["skipped_had_contacts"],
            "contacts_found": counts["contacts_found"],
            "no_contacts_found": counts["no_contacts_found"],
            "lookup_errors": counts["lookup_errors"],
            "endpoint_not_supported": counts["not_supported"],
            "credits_estimated": net_lookups,
            "credits_used": counts["credits_used"],
            "preflight_outcome": "ok",
            "preflight_message": pf_msg,
            "run_timestamp": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
            "input_file": ss("_bcf_file_name", ""),
        }

        excel_bytes = _build_excel(
            enriched_df,
            ss("_bcf_df_original"),
            run_summary,
            raw_evidence,
            all_sheets,
        )

        ss_set(
            _bcf_processing=False,
            _bcf_done=True,
            _bcf_results_df=enriched_df,
            _bcf_summary=run_summary,
            _bcf_raw_evidence=raw_evidence,
            _bcf_excel_bytes=excel_bytes,
            _bcf_preflight_error=None,
        )
        st.rerun()

    # ── Results ───────────────────────────────────────────────────────────────
    if ss("_bcf_done"):
        run_summary = ss("_bcf_summary") or {}
        preflight_error = ss("_bcf_preflight_error")

        is_demo_run = run_summary.get("preflight_outcome") == "demo_mode"

        if is_demo_run:
            st.warning(
                "⚠️ **Demo run completed.** "
                f"{run_summary.get('demo_contacts_generated', 0)} companies were populated "
                "with artificial placeholder contacts.  \n"
                "**These contacts are not real and must not be used for real outreach.**  \n"
                "Use this output to preview the Caller Prep interface layout only."
            )
            st.markdown("### ✅ Demo enrichment complete")
        elif preflight_error:
            st.error(preflight_error)
            st.markdown(
                "**What to do:**  \n"
                "- Check your Lusha plan — contact/person lookup may require a separate add-on.  \n"
                "- Verify your `LUSHA_API_KEY` in Streamlit secrets has people-search permissions.  \n"
                "- The output file below is still valid and Lovable-compatible — "
                "contact fields are blank but all other Opportunity Radar data is preserved."
            )
            st.markdown("### ⬇ Download output (no contacts — endpoint unavailable)")
        else:
            st.markdown("### ✅ Contact enrichment complete")

        r_a, r_b, r_c, r_d, r_e = st.columns(5)
        r_a.metric("Contacts found", run_summary.get("contacts_found", 0))
        r_b.metric("No contacts found", run_summary.get("no_contacts_found", 0))
        r_c.metric("Endpoint not supported", run_summary.get("endpoint_not_supported", 0))
        r_d.metric("Errors", run_summary.get("lookup_errors", 0))
        r_e.metric(
            "Credits used (actual)",
            run_summary.get("credits_used", 0),
            help="Counts only successful or billed lookups (contacts_found + no_contacts_found).",
        )

        st.markdown("---")
        ts = run_summary.get("run_timestamp", "")
        input_f = run_summary.get("input_file", "")
        safe_ts = ts.replace(":", "").replace(" ", "_").replace("UTC", "").strip()
        dl_name = f"BuyerContacts_{safe_ts}.xlsx"

        st.download_button(
            label="⬇ Download enriched Excel",
            data=ss("_bcf_excel_bytes"),
            file_name=dl_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )

        st.caption(
            f"Output file: **{dl_name}**  \n"
            "Sheet 1: Caller Prep Input (enriched, Lovable-compatible)  \n"
            "Sheet 2: Contact Enrichment Summary  \n"
            "Sheet 3: Contact Review Needed  \n"
            "Sheet 4: Original Caller Prep Input  \n"
            "Sheet 5: Contact Raw Evidence"
        )

        if st.button("🔄 Start over with a new file", key="bcf_reset"):
            reset_bcf()
            st.rerun()


# =============================================================================
# RUN
# =============================================================================

main()


# =============================================================================
# MANUAL TEST STEPS
# =============================================================================
#
# To verify correct behaviour without spending Lusha credits, run:
#   streamlit run buyer_contact_finder.py
# Then follow the steps below.
#
# Test 1 — Reads Caller Prep Input correctly
#   Upload an Opportunity Radar export. Confirm the sheet summary shows
#   the correct company count and recommendation breakdown.
#
# Test 2 — Missing sheet error
#   Upload any .xlsx without a 'Caller Prep Input' sheet.
#   Expect error: "Caller Prep Input sheet not found..."
#
# Test 3 — Selection logic: Call now + Manual
#   With default selection, confirm only "Call now", "Call this month",
#   and "Manual research needed" companies are shown in the preview.
#
# Test 4 — Top N selection
#   Select "Top N by opportunity_score" with N=10.
#   Confirm preview shows exactly 10 companies sorted by score.
#
# Test 5 — Skips companies already with contacts
#   Manually populate contact_1_name for a few rows.
#   Run enrichment without "Refresh existing".
#   Confirm those rows get contact_data_status = skipped_already_has_contacts.
#
# Test 6 — Does not overwrite contacts by default
#   Same as Test 5. Confirm contact_1_name is unchanged.
#
# Test 7 — Handles partial Lusha company-level fields gracefully
#   Upload a file with lusha_api_company_name, lusha_api_domain present.
#   Confirm the tool loads without error and those columns are preserved.
#
# Test 8 — Exports Caller Prep Input as first sheet
#   Open the downloaded Excel. Confirm sheet 1 is 'Caller Prep Input'.
#
# Test 9 — Lovable-compatible structure
#   Open the downloaded Excel. Load 'Caller Prep Input' in Caller Prep Cockpit.
#   Confirm it loads without error and existing fields are intact.
#
# Test 10 — Contact Enrichment Summary exists
#   Open the downloaded Excel. Confirm sheet 2 is 'Contact Enrichment Summary'
#   with correct counts.
#
# Test 11 — Contact Review Needed exists
#   Confirm sheet 3 is 'Contact Review Needed' and contains companies with
#   no_contacts_found or contact_lookup_error status.
#
# Test 12 — Original Caller Prep Input exists and is unchanged
#   Confirm sheet 4 matches the input sheet exactly (no contact columns added).
#
# Test 13 — Contact Raw Evidence exists
#   Confirm sheet 5 is 'Contact Raw Evidence' with one row per contact found.
#
# Test 14 — Demo batch (top 10 without spending credits)
#   Set LUSHA_API_KEY to a test/sandbox key.
#   Select "Top N by opportunity_score" with N=10.
#   Confirm the run completes and at least the structure is correct.
#
# Test 15 — No contacts found leaves fields blank
#   For a company where Lusha returns no people, confirm contact_1_name is blank
#   and contact_data_status = no_contacts_found.
