"""
mYngle · LinkedIn Signal Demo
==============================
Public growth and hiring signal extractor using Serper Google Search,
optional Firecrawl website crawls, and Claude extraction.

COMPLIANCE BOUNDARY:
This demo does NOT crawl LinkedIn directly, use cookies, browser automation,
or bypass any access controls. It only uses:
- Public Google search snippets (via Serper) that may reference LinkedIn URLs.
- Firecrawl on the company's own public website and likely public careers pages.
- Optional LinkedIn notes manually pasted by the user.

Entry point:
    streamlit run linkedin_signal_demo.py

Secrets (.streamlit/secrets.toml):
    SERPER_API_KEY      = "..."
    ANTHROPIC_API_KEY   = "..."
    FIRECRAWL_API_KEY   = "..."   # optional
"""

import hashlib
import io
import json
import pathlib
import re
import time
from dataclasses import dataclass, field, asdict
from datetime import datetime

import anthropic
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
import requests
import streamlit as st

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

APP_TITLE     = "mYngle · LinkedIn Signal Demo"
CACHE_DIR     = pathlib.Path("linkedin_signal_cache")
CACHE_VERSION = "v2"           # bump to invalidate old cache entries
SERPER_URL    = "https://google.serper.dev/search"
DEFAULT_MAX   = 4
HARD_CAP      = 10

CAREERS_PATHS = [
    "/careers", "/jobs", "/career",
    "/lavora-con-noi", "/posizioni-aperte",
]

# Search query templates — {name} and {domain} substituted per company
QUERY_TEMPLATES = [
    ("linkedin_company_domain",    'site:linkedin.com/company "{name}" "{domain}"'),
    ("linkedin_company_hiring",    'site:linkedin.com/company "{name}" hiring'),
    ("linkedin_company_jobs",      'site:linkedin.com/company "{name}" jobs'),
    ("linkedin_company_employees", 'site:linkedin.com/company "{name}" employees'),
    ("linkedin_posts_hiring",      'site:linkedin.com/posts "{name}" hiring'),
    ("linkedin_posts_growth",      'site:linkedin.com/posts "{name}" expansion OR growth OR hiring'),
    ("careers_general",            '"{name}" careers'),
    ("jobs_general",               '"{name}" jobs'),
    ("hiring_general",             '"{name}" hiring'),
    ("lavora_con_noi",             '"{name}" "lavora con noi"'),
    ("posizioni_aperte",           '"{name}" "posizioni aperte"'),
    ("expansion",                  '"{name}" expansion'),
    ("new_office",                 '"{name}" "new office"'),
    ("acquisition",                '"{name}" acquisition'),
    ("funding",                    '"{name}" funding'),
]
N_QUERIES = len(QUERY_TEMPLATES)

SIGNAL_SUMMARY_COLS = [
    "company_name", "company_url",
    "linkedin_url_given", "linkedin_url_found",
    "hiring_signal", "hiring_evidence", "hiring_source_url", "open_roles_hint",
    "growth_signal", "growth_evidence", "growth_source_url", "employee_growth_hint",
    "recent_activity_hint", "linkedin_signal_summary", "website_signal_summary",
    "confidence", "suggested_call_angle", "evidence_gaps",
    "entity_ambiguity", "entity_ambiguity_note",
]

COLD_CALLER_COLS = [
    "company_name", "company_url",
    "top_signal", "why_now", "suggested_call_angle",
    "evidence_1", "evidence_1_url",
    "evidence_2", "evidence_2_url",
    "confidence",
    "entity_ambiguity", "entity_ambiguity_note",
]

CLAUDE_MODEL = "claude-haiku-4-5-20251001"

# ---------------------------------------------------------------------------
# Login-wall detection
# ---------------------------------------------------------------------------

_LOGIN_WALL_PATTERNS = [
    r"please[,\s]+log\s*in",
    r"user\s*name\s*password",
    r"the username and/or password you entered is invalid",
    r"i forgot my password",
    r"keep me signed in",
    r"log in with an accredited partner account",
    r"username and/or password",
    r"forgot my password",
    r"sign\s+in\s+to\s+continue",
    r"you must be logged in",
]
_LOGIN_WALL_RE = re.compile("|".join(_LOGIN_WALL_PATTERNS), re.IGNORECASE)


def _is_login_wall(text: str) -> bool:
    return bool(_LOGIN_WALL_RE.search(text))


# ---------------------------------------------------------------------------
# Run usage tracking
# ---------------------------------------------------------------------------

@dataclass
class RunUsage:
    started_at:                   str  = ""
    finished_at:                  str  = ""
    elapsed_seconds:              float = 0.0
    companies_total:              int  = 0
    companies_processed:          int  = 0
    companies_failed:             int  = 0
    serper_searches_planned:      int  = 0
    serper_searches_attempted:    int  = 0
    serper_searches_failed:       int  = 0
    serper_cache_hits:            int  = 0
    firecrawl_crawls_planned:     int  = 0
    firecrawl_crawls_attempted:   int  = 0
    firecrawl_crawls_failed:      int  = 0
    firecrawl_login_wall_count:   int  = 0
    firecrawl_cache_hits:         int  = 0
    claude_calls_planned:         int  = 0
    claude_calls_attempted:       int  = 0
    claude_calls_failed:          int  = 0
    claude_cache_hits:            int  = 0
    anthropic_input_tokens:       int  = 0
    anthropic_output_tokens:      int  = 0
    anthropic_total_tokens:       int  = 0
    estimated_api_units:          float = 0.0
    actual_api_units:             float = 0.0


# ---------------------------------------------------------------------------
# Helpers — URL/domain normalisation
# ---------------------------------------------------------------------------

def ensure_scheme(val: str) -> str:
    """Prepend https:// if no scheme present."""
    val = str(val or "").strip()
    if val and not re.match(r"https?://", val, re.IGNORECASE):
        val = "https://" + val
    return val


def normalize_domain(raw: str) -> str:
    """https://www.example.com/path  →  example.com"""
    if not raw:
        return ""
    raw = raw.strip()
    raw = re.sub(r"^https?://", "", raw, flags=re.IGNORECASE)
    raw = raw.split("/")[0].split("?")[0]
    if raw.lower().startswith("www."):
        raw = raw[4:]
    return raw.lower().strip()


def base_url(raw: str) -> str:
    """Return scheme+host only: https://example.com"""
    raw = ensure_scheme(raw)
    m = re.match(r"(https?://[^/]+)", raw, re.IGNORECASE)
    return m.group(1).rstrip("/") if m else raw


# ---------------------------------------------------------------------------
# Cache
# ---------------------------------------------------------------------------

def _cache_key(name: str, domain: str) -> str:
    raw = f"{CACHE_VERSION}|{name.lower().strip()}|{domain.lower().strip()}"
    return hashlib.md5(raw.encode()).hexdigest()


def cache_load(name: str, domain: str) -> dict | None:
    CACHE_DIR.mkdir(exist_ok=True)
    p = CACHE_DIR / f"{_cache_key(name, domain)}.json"
    if p.exists():
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            return None
    return None


def cache_save(name: str, domain: str, data: dict) -> None:
    CACHE_DIR.mkdir(exist_ok=True)
    p = CACHE_DIR / f"{_cache_key(name, domain)}.json"
    try:
        p.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Serper search
# ---------------------------------------------------------------------------

def serper_search(query: str, api_key: str, n: int = 5) -> tuple[list[dict], bool]:
    """Returns (results, failed)."""
    try:
        resp = requests.post(
            SERPER_URL,
            headers={"X-API-KEY": api_key, "Content-Type": "application/json"},
            json={"q": query, "num": n},
            timeout=15,
        )
        resp.raise_for_status()
        data = resp.json()
        out = []
        for item in data.get("organic", [])[:n]:
            out.append({
                "title":   item.get("title", ""),
                "link":    item.get("link", ""),
                "snippet": item.get("snippet", ""),
                "date":    item.get("date", ""),
            })
        return out, False
    except Exception as exc:
        return [{"title": "", "link": "", "snippet": f"[search error: {exc}]", "date": ""}], True


def run_all_searches(
    name: str, domain: str, serper_key: str, usage: RunUsage
) -> dict[str, list[dict]]:
    """Run all query templates; update usage counters."""
    grouped: dict[str, list[dict]] = {}
    for label, template in QUERY_TEMPLATES:
        query = template.format(name=name, domain=domain)
        results, failed = serper_search(query, serper_key, n=5)
        usage.serper_searches_attempted += 1
        if failed:
            usage.serper_searches_failed += 1
        grouped[label] = results
        time.sleep(0.25)
    return grouped


def search_results_to_text(grouped: dict[str, list[dict]]) -> str:
    parts = []
    for label, results in grouped.items():
        parts.append(f"=== {label.upper()} ===")
        if not results:
            parts.append("(no results)\n")
            continue
        for i, r in enumerate(results, 1):
            line = f"[{i}] {r['title']}"
            if r.get("date"):
                line += f"  ({r['date']})"
            line += f"\n    {r['link']}"
            if r.get("snippet"):
                line += f"\n    {r['snippet']}"
            parts.append(line)
        parts.append("")
    return "\n".join(parts)


def find_linkedin_url_in_results(grouped: dict[str, list[dict]]) -> str:
    for results in grouped.values():
        for r in results:
            link = r.get("link", "")
            if "linkedin.com/company/" in link.lower():
                return link
    return ""


# ---------------------------------------------------------------------------
# Firecrawl website crawl
# ---------------------------------------------------------------------------

def firecrawl_scrape(url: str, api_key: str) -> tuple[str, str]:
    """
    Scrape one public URL with Firecrawl.
    Returns (extracted_text, error_message).
    Never crawls linkedin.com or any gated resource.
    """
    if not api_key:
        return "", "No Firecrawl API key"
    try:
        resp = requests.post(
            "https://api.firecrawl.dev/v1/scrape",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={"url": url, "formats": ["markdown"], "onlyMainContent": True},
            timeout=20,
        )
        if resp.status_code == 200:
            data = resp.json()
            md = data.get("data", {}).get("markdown", "") or ""
            return md[:4000], ""
        return "", f"HTTP {resp.status_code}"
    except Exception as exc:
        return "", str(exc)


def crawl_company_website(
    company_url: str,
    firecrawl_key: str,
    homepage_only: bool,
    usage: RunUsage,
) -> list[dict]:
    """
    Crawl company homepage + optional careers paths.
    status values: 'ok' | 'login_wall' | 'error'
    Login-wall pages are flagged and excluded from Claude evidence.
    """
    crawls: list[dict] = []
    if not company_url or not firecrawl_key:
        return crawls

    urls_to_try = [company_url.rstrip("/")]
    if not homepage_only:
        base = base_url(company_url)
        for path in CAREERS_PATHS:
            urls_to_try.append(base + path)

    for url in urls_to_try:
        usage.firecrawl_crawls_attempted += 1
        text, err = firecrawl_scrape(url, firecrawl_key)
        if text and _is_login_wall(text):
            status  = "login_wall"
            preview = text[:500]
            usage.firecrawl_login_wall_count += 1
        elif text:
            status  = "ok"
            preview = text[:500]
        else:
            status  = "error"
            preview = ""
            if err:
                usage.firecrawl_crawls_failed += 1
        crawls.append({
            "url_attempted":          url,
            "status":                 status,
            "extracted_text_preview": preview,
            "error":                  err,
        })
        time.sleep(0.2)
    return crawls


def crawls_to_text(crawls: list[dict]) -> str:
    """
    Serialise crawl results for the Claude prompt.
    Login-wall pages are noted as blocked — not passed as evidence.
    """
    parts = []
    for c in crawls:
        status = c.get("status", "")
        parts.append(f"URL: {c['url_attempted']}  [{status}]")
        if status == "login_wall":
            parts.append(
                "  [Page appears to be behind login and was excluded from evidence.]"
            )
        elif c.get("error"):
            parts.append(f"  Error: {c['error']}")
        elif c.get("extracted_text_preview"):
            parts.append(c["extracted_text_preview"])
        parts.append("")
    return "\n".join(parts)


# ---------------------------------------------------------------------------
# Claude extraction
# ---------------------------------------------------------------------------

EXTRACTION_PROMPT = """\
You are a B2B sales intelligence analyst for mYngle, which sells online language \
training and business communication support to international companies.

You have been given Google search snippets, optional website crawl text, and \
optionally some LinkedIn notes manually provided by the user. Your job is to \
extract factual hiring and growth signals for the company listed below.

STRICT RULES — follow every rule exactly:
- Only report signals explicitly supported by the provided text.
- Do not infer growth from generic company descriptions or marketing copy.
- Do not invent employee counts. Do not invent job counts.
- If evidence is missing, use "Unknown" — not "No".
- For Yes or Weak signals, always include the exact snippet or sentence as evidence.
- Prefer conservative scoring. When in doubt, use Weak or Unknown.
- Do NOT treat login-wall pages (marked [Page appears to be behind login]) as \
  evidence of hiring or growth.
- Do NOT infer current hiring from LinkedIn snippets that have no visible date. \
  If a LinkedIn snippet has no date, note "timing unclear" in the evidence field.
- Keep confidence LOW when all evidence comes only from search snippets with no \
  corroborating website text.
- hiring_signal / growth_signal must be exactly one of: Yes / Weak / No / Unknown
- confidence must be exactly one of: High / Medium / Low
- entity_ambiguity must be exactly one of: Yes / No
  Set entity_ambiguity = Yes when:
  * The company name appears to be a sub-brand, academy, department, business \
    unit, or local subsidiary.
  * Sources mostly refer to the parent company or global organisation rather \
    than the specific entity uploaded.
  * Evidence may not clearly belong to this exact legal entity.
  If Yes, explain briefly in entity_ambiguity_note.
- suggested_call_angle must be phrased as a soft, practical cold-caller opener — \
  preferably a question or tentative statement. Avoid assertive language such as \
  "you are expanding" unless the source directly states it.
  Good example: "I noticed signs of recent hiring linked to {name} — \
  are you currently expanding teams that might benefit from stronger business \
  English or cross-border communication support?"
  Bad example: "You are expanding into Germany and need language training."

TODAY: {today}
COMPANY NAME: {name}
COMPANY URL: {url}
LINKEDIN URL PROVIDED BY USER: {linkedin_url_given}
DOMAIN: {domain}

=== GOOGLE SEARCH SNIPPETS ===
{search_text}

=== WEBSITE CRAWL TEXT ===
{website_text}

=== USER-PROVIDED LINKEDIN NOTES ===
{linkedin_notes}

Respond with ONLY a valid JSON object matching this exact schema (no markdown fences):
{{
  "company_name": "",
  "company_url": "",
  "linkedin_url_given": "",
  "linkedin_url_found": "",
  "hiring_signal": "Yes|Weak|No|Unknown",
  "hiring_evidence": "",
  "hiring_source_url": "",
  "open_roles_hint": "",
  "growth_signal": "Yes|Weak|No|Unknown",
  "growth_evidence": "",
  "growth_source_url": "",
  "employee_growth_hint": "",
  "recent_activity_hint": "",
  "linkedin_signal_summary": "",
  "website_signal_summary": "",
  "suggested_call_angle": "",
  "confidence": "High|Medium|Low",
  "evidence_gaps": "",
  "entity_ambiguity": "Yes|No",
  "entity_ambiguity_note": ""
}}
"""


def extract_signals(
    name: str,
    url: str,
    domain: str,
    linkedin_url_given: str,
    linkedin_notes: str,
    search_text: str,
    website_text: str,
    anthropic_key: str,
    usage: RunUsage,
) -> dict:
    prompt = EXTRACTION_PROMPT.format(
        today=datetime.now().strftime("%Y-%m-%d"),
        name=name,
        url=url,
        domain=domain,
        linkedin_url_given=linkedin_url_given or "",
        search_text=search_text or "(none)",
        website_text=website_text or "(none)",
        linkedin_notes=linkedin_notes or "(none)",
    )
    usage.claude_calls_attempted += 1
    try:
        client = anthropic.Anthropic(api_key=anthropic_key)
        msg = client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=1200,
            messages=[{"role": "user", "content": prompt}],
        )
        # Capture token usage
        if hasattr(msg, "usage") and msg.usage:
            in_tok  = getattr(msg.usage, "input_tokens",  0) or 0
            out_tok = getattr(msg.usage, "output_tokens", 0) or 0
        else:
            in_tok  = len(prompt) // 4
            out_tok = 300
        usage.anthropic_input_tokens  += in_tok
        usage.anthropic_output_tokens += out_tok
        usage.anthropic_total_tokens  += in_tok + out_tok

        raw = msg.content[0].text.strip()
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        return json.loads(raw)
    except json.JSONDecodeError:
        usage.claude_calls_failed += 1
        return _empty_signal(name, url, linkedin_url_given, error="JSON parse error from Claude")
    except Exception as exc:
        usage.claude_calls_failed += 1
        return _empty_signal(name, url, linkedin_url_given, error=str(exc))


def _empty_signal(name: str, url: str, linkedin_url_given: str, error: str = "") -> dict:
    return {
        "company_name":         name,
        "company_url":          url,
        "linkedin_url_given":   linkedin_url_given,
        "linkedin_url_found":   "",
        "hiring_signal":        "Unknown",
        "hiring_evidence":      error,
        "hiring_source_url":    "",
        "open_roles_hint":      "",
        "growth_signal":        "Unknown",
        "growth_evidence":      "",
        "growth_source_url":    "",
        "employee_growth_hint": "",
        "recent_activity_hint": "",
        "linkedin_signal_summary": "",
        "website_signal_summary":  "",
        "suggested_call_angle": "",
        "confidence":           "Low",
        "evidence_gaps":        f"Extraction failed: {error}",
        "entity_ambiguity":     "No",
        "entity_ambiguity_note": "",
    }


# ---------------------------------------------------------------------------
# Per-company pipeline
# ---------------------------------------------------------------------------

def process_company(
    row: dict,
    serper_key: str,
    anthropic_key: str,
    firecrawl_key: str,
    usage: RunUsage,
    force_fresh: bool = False,
    homepage_only: bool = False,
    skip_firecrawl: bool = False,
    step_cb=None,           # optional callable(step_label: str)
) -> tuple[dict, list[dict], list[dict]]:
    """
    Process one company.
    Returns (signal_dict, raw_search_rows, raw_crawl_rows).
    Updates usage counters in-place.
    """
    name           = str(row.get("company_name", "")).strip()
    url            = ensure_scheme(str(row.get("company_url", "") or "").strip())
    linkedin_given = str(row.get("linkedin_url", "") or "").strip()
    linkedin_notes = str(row.get("linkedin_notes", "") or "").strip()
    domain         = normalize_domain(url)

    if not name:
        return _empty_signal("(empty)", url, linkedin_given, "Missing company_name"), [], []

    # ── Cache check ──────────────────────────────────────────────────────────
    if not force_fresh:
        cached = cache_load(name, domain)
        if cached:
            usage.serper_cache_hits  += 1
            usage.firecrawl_cache_hits += 1
            usage.claude_cache_hits  += 1
            return (
                cached.get("signal", {}),
                cached.get("search_rows", []),
                cached.get("crawl_rows", []),
            )

    # ── 1. Serper searches ───────────────────────────────────────────────────
    if step_cb:
        step_cb("🔎 Serper: zoeken naar signalen…")
    grouped       = run_all_searches(name, domain, serper_key, usage)
    search_text   = search_results_to_text(grouped)
    linkedin_found = find_linkedin_url_in_results(grouped)

    raw_search: list[dict] = []
    template_map = dict(QUERY_TEMPLATES)
    for label, results in grouped.items():
        tmpl  = template_map.get(label, "")
        query = tmpl.format(name=name, domain=domain)
        for r in results:
            raw_search.append({
                "company_name": name,
                "query":        query,
                "title":        r.get("title", ""),
                "snippet":      r.get("snippet", ""),
                "link":         r.get("link", ""),
                "source_type":  "linkedin" if "linkedin.com" in r.get("link", "") else "web",
            })

    # ── 2. Firecrawl ─────────────────────────────────────────────────────────
    raw_crawls: list[dict] = []
    website_text = ""
    if not skip_firecrawl and firecrawl_key and url:
        if step_cb:
            step_cb("🌐 Firecrawl: website crawlen…")
        crawls = crawl_company_website(url, firecrawl_key, homepage_only, usage)
        for c in crawls:
            raw_crawls.append({"company_name": name, **c})
        website_text = crawls_to_text(crawls)
    elif skip_firecrawl:
        website_text = "(Firecrawl skipped)"

    # ── 3. Claude extraction ─────────────────────────────────────────────────
    if step_cb:
        step_cb("🤖 Claude: signalen extraheren…")
    signal = extract_signals(
        name=name,
        url=url,
        domain=domain,
        linkedin_url_given=linkedin_given,
        linkedin_notes=linkedin_notes,
        search_text=search_text,
        website_text=website_text,
        anthropic_key=anthropic_key,
        usage=usage,
    )
    signal.setdefault("company_name", name)
    signal.setdefault("company_url", url)
    signal.setdefault("linkedin_url_given", linkedin_given)
    signal.setdefault("entity_ambiguity", "No")
    signal.setdefault("entity_ambiguity_note", "")
    if not signal.get("linkedin_url_found") and linkedin_found:
        signal["linkedin_url_found"] = linkedin_found

    # ── Cache save ────────────────────────────────────────────────────────────
    cache_save(name, domain, {
        "signal": signal,
        "search_rows": raw_search,
        "crawl_rows":  raw_crawls,
    })

    return signal, raw_search, raw_crawls


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

_HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT  = Font(bold=True, color="FFFFFF")


def _style_header_row(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _autofit(ws, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        length = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
            length + 4, max_width
        )


def _freeze(ws) -> None:
    ws.freeze_panes = ws["A2"]


def _trim(text: str, max_len: int) -> str:
    text = str(text or "").strip()
    return text[:max_len] + "…" if len(text) > max_len else text


def build_excel(
    signals: list[dict],
    raw_search: list[dict],
    raw_crawls: list[dict],
    usage: RunUsage,
) -> bytes:
    wb = openpyxl.Workbook()

    # ── Sheet 1: Signal Summary ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Signal Summary"
    if signals:
        df1 = pd.DataFrame(signals)
        for col in SIGNAL_SUMMARY_COLS:
            if col not in df1.columns:
                df1[col] = ""
        df1 = df1[SIGNAL_SUMMARY_COLS]
        ws1.append(list(df1.columns))
        _style_header_row(ws1)
        for _, r in df1.iterrows():
            ws1.append([str(v) if v is not None else "" for v in r])
    _autofit(ws1)
    _freeze(ws1)

    # ── Sheet 2: Raw Search Results ──────────────────────────────────────────
    ws2 = wb.create_sheet("Raw Search Results")
    search_cols = ["company_name", "query", "title", "snippet", "link", "source_type"]
    df2 = pd.DataFrame(raw_search) if raw_search else pd.DataFrame(columns=search_cols)
    for col in search_cols:
        if col not in df2.columns:
            df2[col] = ""
    ws2.append(search_cols)
    _style_header_row(ws2)
    for _, r in df2[search_cols].iterrows():
        ws2.append([str(v) if v is not None else "" for v in r])
    _autofit(ws2)
    _freeze(ws2)

    # ── Sheet 3: Raw Website Crawls ──────────────────────────────────────────
    ws3 = wb.create_sheet("Raw Website Crawls")
    crawl_cols = ["company_name", "url_attempted", "status", "extracted_text_preview", "error"]
    df3 = pd.DataFrame(raw_crawls) if raw_crawls else pd.DataFrame(columns=crawl_cols)
    for col in crawl_cols:
        if col not in df3.columns:
            df3[col] = ""
    ws3.append(crawl_cols)
    _style_header_row(ws3)
    for _, r in df3[crawl_cols].iterrows():
        ws3.append([str(v) if v is not None else "" for v in r])
    _autofit(ws3)
    _freeze(ws3)

    # ── Sheet 4: Cold Caller Input ───────────────────────────────────────────
    ws4 = wb.create_sheet("Cold Caller Input")
    cold_rows = []
    for s in signals:
        hiring_ev  = s.get("hiring_evidence", "")
        growth_ev  = s.get("growth_evidence", "")
        top_signal = ""
        if s.get("hiring_signal") in ("Yes", "Weak"):
            top_signal = f"Hiring: {s.get('open_roles_hint', '')}".strip()
        elif s.get("growth_signal") in ("Yes", "Weak"):
            top_signal = f"Growth: {s.get('employee_growth_hint', '')}".strip()
        why_now = s.get("recent_activity_hint", "") or s.get("linkedin_signal_summary", "")
        cold_rows.append({
            "company_name":          s.get("company_name", ""),
            "company_url":           s.get("company_url", ""),
            "top_signal":            _trim(top_signal, 180),
            "why_now":               _trim(why_now, 300),
            "suggested_call_angle":  _trim(s.get("suggested_call_angle", ""), 350),
            "evidence_1":            _trim(hiring_ev, 350),
            "evidence_1_url":        s.get("hiring_source_url", ""),
            "evidence_2":            _trim(growth_ev, 350),
            "evidence_2_url":        s.get("growth_source_url", ""),
            "confidence":            s.get("confidence", ""),
            "entity_ambiguity":      s.get("entity_ambiguity", ""),
            "entity_ambiguity_note": s.get("entity_ambiguity_note", ""),
        })
    df4 = pd.DataFrame(cold_rows) if cold_rows else pd.DataFrame(columns=COLD_CALLER_COLS)
    for col in COLD_CALLER_COLS:
        if col not in df4.columns:
            df4[col] = ""
    ws4.append(COLD_CALLER_COLS)
    _style_header_row(ws4)
    for _, r in df4[COLD_CALLER_COLS].iterrows():
        ws4.append([str(v) if v is not None else "" for v in r])
    _autofit(ws4)
    _freeze(ws4)

    # ── Sheet 5: Run Usage ───────────────────────────────────────────────────
    ws5 = wb.create_sheet("Run Usage")
    ws5.append(["Metric", "Value"])
    _style_header_row(ws5)
    for k, v in asdict(usage).items():
        ws5.append([k, str(v)])
    _autofit(ws5)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Estimation helpers
# ---------------------------------------------------------------------------

def compute_estimate(
    n_companies: int,
    skip_firecrawl: bool,
    homepage_only: bool,
    serper_units: float,
    fc_units: float,
    claude_units: float,
    avg_serper_s: float,
    avg_fc_s: float,
    avg_claude_s: float,
) -> dict:
    n_searches   = n_companies * N_QUERIES
    if skip_firecrawl:
        n_crawls = 0
    elif homepage_only:
        n_crawls = n_companies
    else:
        n_crawls = n_companies * (1 + len(CAREERS_PATHS))
    n_claude = n_companies

    est_seconds = (
        n_searches * avg_serper_s
        + n_crawls * avg_fc_s
        + n_claude * avg_claude_s
    )
    est_units = (
        n_searches * serper_units
        + n_crawls * fc_units
        + n_claude * claude_units
    )
    mins, secs = divmod(int(est_seconds), 60)
    return {
        "n_companies":  n_companies,
        "n_searches":   n_searches,
        "n_crawls":     n_crawls,
        "n_claude":     n_claude,
        "est_seconds":  est_seconds,
        "est_time_str": f"{mins}m {secs}s",
        "est_units":    round(est_units, 1),
    }


# ---------------------------------------------------------------------------
# Streamlit UI helpers
# ---------------------------------------------------------------------------

def _get_secret(key: str) -> str:
    try:
        return st.secrets[key]
    except (KeyError, FileNotFoundError):
        return ""


def _load_uploaded(uploaded) -> pd.DataFrame | None:
    fname = uploaded.name.lower()
    try:
        if fname.endswith(".csv"):
            return pd.read_csv(uploaded)
        xf = pd.ExcelFile(uploaded)
        for sheet in xf.sheet_names:
            df = xf.parse(sheet)
            lower = {c.lower() for c in df.columns}
            if "company_name" in lower or "company name" in lower:
                return df
        return xf.parse(xf.sheet_names[0])
    except Exception as exc:
        st.error(f"Kon bestand niet lezen: {exc}")
        return None


def _normalise_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Lowercase/strip column names, map common variants, ensure company_url exists.
    URL aliases (priority): company_domain > company_url > website_url >
      company_website > website > url > domain.
    Values without a scheme get https:// prepended.
    """
    df = df.copy()
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]

    for alias in ("company", "name"):
        if alias in df.columns and "company_name" not in df.columns:
            df.rename(columns={alias: "company_name"}, inplace=True)
            break

    url_aliases = [
        "company_domain", "company_url", "website_url",
        "company_website", "website", "url", "domain",
    ]
    if "company_url" not in df.columns:
        for alias in url_aliases:
            if alias in df.columns:
                df.rename(columns={alias: "company_url"}, inplace=True)
                break

    if "linkedin_url" not in df.columns and "linkedin" in df.columns:
        df.rename(columns={"linkedin": "linkedin_url"}, inplace=True)
    if "linkedin_notes" not in df.columns and "notes" in df.columns:
        df.rename(columns={"notes": "linkedin_notes"}, inplace=True)

    if "company_url" in df.columns:
        df["company_url"] = df["company_url"].apply(
            lambda v: ensure_scheme(str(v or "").strip())
        )

    return df


# ---------------------------------------------------------------------------
# Main Streamlit app
# ---------------------------------------------------------------------------

def main() -> None:
    st.set_page_config(page_title=APP_TITLE, page_icon="🔗", layout="wide")
    st.title("🔗 " + APP_TITLE)
    st.markdown(
        "> **This demo checks public growth and hiring signals from search snippets, "
        "company websites and optionally pasted LinkedIn notes. "
        "It does not crawl LinkedIn directly.**"
    )

    # ── Sidebar: API keys ─────────────────────────────────────────────────────
    st.sidebar.header("API-sleutels")
    serper_key    = _get_secret("SERPER_API_KEY")    or st.sidebar.text_input("Serper API key",    type="password")
    anthropic_key = _get_secret("ANTHROPIC_API_KEY") or st.sidebar.text_input("Anthropic API key", type="password")
    firecrawl_key = _get_secret("FIRECRAWL_API_KEY") or st.sidebar.text_input("Firecrawl API key (optioneel)", type="password")

    for label, val in [("Serper", serper_key), ("Anthropic", anthropic_key)]:
        if val:
            masked = val[:4] + "****" + val[-4:] if len(val) > 8 else "****"
            st.sidebar.success(f"{label}: `{masked}`")
        else:
            st.sidebar.error(f"{label}: niet ingesteld")
    if firecrawl_key:
        masked = firecrawl_key[:4] + "****" + firecrawl_key[-4:] if len(firecrawl_key) > 8 else "****"
        st.sidebar.info(f"Firecrawl: `{masked}`")
    else:
        st.sidebar.caption("Firecrawl: niet ingesteld — website crawl overgeslagen.")

    st.sidebar.markdown("---")
    st.sidebar.header("Run-instellingen")
    max_rows       = st.sidebar.number_input("Max. rijen", min_value=1, max_value=HARD_CAP, value=DEFAULT_MAX)
    force_fresh    = st.sidebar.checkbox("Cache negeren (force fresh)", value=False)
    homepage_only  = st.sidebar.checkbox("Alleen hoofdwebsite crawlen", value=False)
    skip_firecrawl = st.sidebar.checkbox("Firecrawl overslaan voor snelle test", value=False)
    if homepage_only and not skip_firecrawl:
        st.sidebar.caption("Alleen homepage — career-URL's overgeslagen.")
    if skip_firecrawl:
        st.sidebar.caption("Firecrawl volledig overgeslagen.")

    st.sidebar.markdown("---")
    st.sidebar.header("Schattingsparameters")
    serper_units  = st.sidebar.number_input("Serper units per search",           value=1.0, step=0.1)
    fc_units      = st.sidebar.number_input("Firecrawl units per crawl",          value=1.0, step=0.1)
    claude_units  = st.sidebar.number_input("Claude units per bedrijf",           value=1.0, step=0.1)
    avg_serper_s  = st.sidebar.number_input("Gem. seconden per Serper search",    value=1.5, step=0.1)
    avg_fc_s      = st.sidebar.number_input("Gem. seconden per Firecrawl crawl",  value=8.0, step=0.5)
    avg_claude_s  = st.sidebar.number_input("Gem. seconden per Claude extractie", value=25.0, step=1.0)

    # ── Input ─────────────────────────────────────────────────────────────────
    st.markdown("### Bedrijfsinput")
    input_mode = st.radio("Invoermethode", ["Handmatig invoeren", "CSV / XLSX uploaden"], horizontal=True)

    df_input: pd.DataFrame | None = None

    if input_mode == "Handmatig invoeren":
        st.caption(
            "Vul minimaal `company_name` en `company_url` in. "
            "`linkedin_url` en `linkedin_notes` zijn optioneel."
        )
        blank = {"company_name": "", "company_url": "", "linkedin_url": "", "linkedin_notes": ""}
        df_input = st.data_editor(
            pd.DataFrame([blank] * DEFAULT_MAX),
            num_rows="dynamic",
            use_container_width=True,
            key="manual_input",
        )
        df_input = df_input[df_input["company_name"].str.strip().astype(bool)].copy()
        if "company_url" in df_input.columns:
            df_input["company_url"] = df_input["company_url"].apply(
                lambda v: ensure_scheme(str(v or "").strip())
            )
    else:
        uploaded = st.file_uploader("Upload CSV of XLSX", type=["csv", "xlsx"])
        if uploaded:
            raw_df = _load_uploaded(uploaded)
            if raw_df is not None:
                df_input = _normalise_columns(raw_df)
                st.success(f"{len(df_input)} rijen geladen uit `{uploaded.name}`.")
                st.dataframe(df_input.head(10), use_container_width=True)

    # ── Pre-run estimate ──────────────────────────────────────────────────────
    n_for_estimate = 0
    if df_input is not None and not df_input.empty:
        n_for_estimate = min(len(df_input), int(max_rows), HARD_CAP)

    if n_for_estimate > 0:
        est = compute_estimate(
            n_for_estimate, skip_firecrawl, homepage_only,
            serper_units, fc_units, claude_units,
            avg_serper_s, avg_fc_s, avg_claude_s,
        )
        with st.expander("📊 Geschatte run", expanded=True):
            c1, c2, c3, c4, c5, c6 = st.columns(6)
            c1.metric("Bedrijven",         est["n_companies"])
            c2.metric("Serper searches",   est["n_searches"])
            c3.metric("Firecrawl crawls",  est["n_crawls"])
            c4.metric("Claude calls",      est["n_claude"])
            c5.metric("Geschatte tijd",    est["est_time_str"])
            c6.metric("Geschatte API-units", est["est_units"])
            st.caption(
                "Dit is een ruwe schatting. Werkelijke kosten/credits hangen af van "
                "provider-instellingen, caching, timeouts en response size."
            )

    # ── Run button ────────────────────────────────────────────────────────────
    st.markdown("---")
    run_btn = st.button("🚀 Start signaalanalyse", type="primary")

    if not run_btn:
        return

    if not serper_key or not anthropic_key:
        st.error("Serper API key en Anthropic API key zijn verplicht.")
        return

    if df_input is None or df_input.empty:
        st.error("Geen bedrijven opgegeven.")
        return

    missing = [c for c in ("company_name", "company_url") if c not in df_input.columns]
    if missing:
        st.error(
            f"Verplichte kolom(men) ontbreken: **{', '.join(missing)}**. "
            f"Aanwezige kolommen: {', '.join(df_input.columns.tolist())}. "
            f"Geaccepteerde aliassen voor company_url: "
            f"company_domain, website_url, company_website, website, url, domain."
        )
        return

    rows = df_input.to_dict("records")[: int(max_rows)]
    if len(rows) > HARD_CAP:
        rows = rows[:HARD_CAP]
        st.warning(f"Harde limiet: analyse beperkt tot {HARD_CAP} rijen.")

    # ── Run ───────────────────────────────────────────────────────────────────
    n_rows = len(rows)
    usage  = RunUsage(
        started_at=datetime.now().isoformat(timespec="seconds"),
        companies_total=n_rows,
        serper_searches_planned=n_rows * N_QUERIES,
        firecrawl_crawls_planned=(
            0 if skip_firecrawl else
            n_rows if homepage_only else
            n_rows * (1 + len(CAREERS_PATHS))
        ),
        claude_calls_planned=n_rows,
        estimated_api_units=est["est_units"] if n_for_estimate > 0 else 0.0,
    )

    all_signals: list[dict] = []
    all_search:  list[dict] = []
    all_crawls:  list[dict] = []

    progress_bar  = st.progress(0, text="Bezig…")
    status_box    = st.empty()
    step_box      = st.empty()
    metrics_row   = st.empty()
    run_start     = time.time()

    def _step(label: str) -> None:
        step_box.info(label)

    def _refresh_metrics() -> None:
        elapsed  = time.time() - run_start
        done     = usage.companies_processed + usage.companies_failed
        if done > 0:
            per_co   = elapsed / done
            remaining = per_co * (n_rows - done)
            remain_str = f"{int(remaining // 60)}m {int(remaining % 60)}s"
        else:
            remain_str = "—"
        with metrics_row.container():
            m1, m2, m3, m4, m5, m6, m7 = st.columns(7)
            m1.metric("Serper searches",       usage.serper_searches_attempted)
            m2.metric("Firecrawl crawls",      usage.firecrawl_crawls_attempted)
            m3.metric("Login walls",            usage.firecrawl_login_wall_count)
            m4.metric("FC fouten",             usage.firecrawl_crawls_failed)
            m5.metric("Claude calls",           usage.claude_calls_attempted)
            m6.metric("Verstreken",            f"{int(elapsed // 60)}m {int(elapsed % 60)}s")
            m7.metric("Resterend (schatting)", remain_str)

    for i, row in enumerate(rows):
        name = str(row.get("company_name", "")).strip() or f"rij {i+1}"
        status_box.info(f"🔍 **{name}** ({i+1}/{n_rows})")
        try:
            signal, search_rows, crawl_rows = process_company(
                row=row,
                serper_key=serper_key,
                anthropic_key=anthropic_key,
                firecrawl_key=firecrawl_key,
                usage=usage,
                force_fresh=force_fresh,
                homepage_only=homepage_only,
                skip_firecrawl=skip_firecrawl,
                step_cb=_step,
            )
            all_signals.append(signal)
            all_search.extend(search_rows)
            all_crawls.extend(crawl_rows)
            usage.companies_processed += 1
        except Exception as exc:
            st.warning(f"Fout bij {name}: {exc} — rij overgeslagen.")
            all_signals.append(
                _empty_signal(name, str(row.get("company_url", "")), "", str(exc))
            )
            usage.companies_failed += 1

        _refresh_metrics()
        progress_bar.progress((i + 1) / n_rows, text=f"{i+1}/{n_rows} verwerkt")

    # ── Finalise usage ────────────────────────────────────────────────────────
    usage.finished_at      = datetime.now().isoformat(timespec="seconds")
    usage.elapsed_seconds  = round(time.time() - run_start, 1)
    usage.actual_api_units = round(
        usage.serper_searches_attempted * serper_units
        + usage.firecrawl_crawls_attempted * fc_units
        + usage.claude_calls_attempted * claude_units,
        1,
    )

    step_box.empty()
    status_box.success(f"✅ Klaar — {usage.companies_processed} verwerkt, {usage.companies_failed} mislukt.")
    progress_bar.empty()

    # ── Actual usage summary ──────────────────────────────────────────────────
    st.markdown("### 📈 Werkelijke run")
    a1, a2, a3, a4, a5, a6, a7 = st.columns(7)
    a1.metric("Serper geprobeerd",    usage.serper_searches_attempted)
    a2.metric("Serper mislukt",       usage.serper_searches_failed)
    a3.metric("Firecrawl geprobeerd", usage.firecrawl_crawls_attempted)
    a4.metric("Login walls",          usage.firecrawl_login_wall_count)
    a5.metric("Claude calls",         usage.claude_calls_attempted)
    a6.metric("Tokens (in+out)",      usage.anthropic_total_tokens)
    a7.metric("Werkelijke API-units", usage.actual_api_units)

    mins_actual, secs_actual = divmod(int(usage.elapsed_seconds), 60)
    st.caption(
        f"Verstreken tijd: {mins_actual}m {secs_actual}s  |  "
        f"Cache hits: Serper {usage.serper_cache_hits}, "
        f"Firecrawl {usage.firecrawl_cache_hits}, "
        f"Claude {usage.claude_cache_hits}"
    )

    # ── Results table ─────────────────────────────────────────────────────────
    st.markdown("### Signaaloverzicht")
    df_summary   = pd.DataFrame(all_signals)
    display_cols = [c for c in SIGNAL_SUMMARY_COLS if c in df_summary.columns]
    st.dataframe(df_summary[display_cols], use_container_width=True)

    # ── Excel download ────────────────────────────────────────────────────────
    st.markdown("### Download")
    xlsx_bytes = build_excel(all_signals, all_search, all_crawls, usage)
    timestamp  = datetime.now().strftime("%Y%m%d_%H%M")
    st.download_button(
        "⬇ Download Excel (5 tabbladen)",
        data=xlsx_bytes,
        file_name=f"linkedin_signals_{timestamp}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    main()
