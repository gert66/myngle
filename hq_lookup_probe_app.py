"""
hq_lookup_probe_app.py

Streamlit UI for HQ headquarters detection.
Self-contained: all runtime logic is inlined here.

Run with:
    streamlit run hq_lookup_probe_app.py

This app is experimental and does not modify any production enrichment output.
"""

from __future__ import annotations

import csv
import io
import json
import os
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Any

import streamlit as st

# ---------------------------------------------------------------------------
# Dependency guard
# ---------------------------------------------------------------------------
try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    st.error("openpyxl is required:  pip install openpyxl")
    st.stop()

try:
    import requests as _requests
except ImportError:
    st.error("requests is required:  pip install requests")
    st.stop()

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_SERPER_URL = "https://google.serper.dev/search"
_REQUEST_TIMEOUT = 12
_INTER_REQUEST_SLEEP = 0.5

DEFAULT_COMPANY_COL   = "company_name"
DEFAULT_DOMAIN_COL    = "domain"
DEFAULT_COUNTRY_COL   = "input_country"
DEFAULT_INPUT_COUNTRY = "Italy"
DEFAULT_LIMIT         = 50

OLD_ENRICHMENT_COLS = [
    "sig_foreign_hq_score",
    "sig_foreign_hq_evidence",
    "foreign_hq_sanitized",
    "foreign_hq_sanitizer_reason",
    "foreign_hq_original_score",
    "final_commercial_fit_score",
    "commercial_fit_score",
]

PROBE_COLS = [
    "hq_detected_city",
    "hq_detected_region",
    "hq_detected_country",
    "hq_confidence",
    "foreign_hq_simple",
    "input_country_used",
    "needs_manual_review",
    "hq_reason",
    "hq_evidence_url",
    "hq_evidence_quote",
    "hq_query_used",
    "serper_knowledge_graph_location",
    "serper_answer_box",
    "top_organic_title_1",
    "top_organic_snippet_1",
    "top_organic_url_1",
    "top_organic_title_2",
    "top_organic_snippet_2",
    "top_organic_url_2",
    "top_organic_title_3",
    "top_organic_snippet_3",
    "top_organic_url_3",
    "probe_error",
    # Recovery review columns
    "has_multilingual_site",
    "website_language_count",
    "website_languages_detected",
    "has_global_structure_signal",
    "global_structure_evidence",
    "hq_review_trigger",
    "hq_structure_type",
    "local_entity_hq_country",
    "local_entity_hq_city",
    "parent_group_hq_country",
    "parent_group_hq_city",
    "review_foreign_parent_score",
    "review_global_network_score",
    "review_multilingual_website_score",
    "review_recommended_hq_signal",
    "sig_foreign_hq_score_original",
    "sig_foreign_hq_score_reviewed",
    "sig_foreign_hq_review_reason",
    "sig_foreign_hq_review_confidence",
    "sig_foreign_hq_review_source",
    "sig_foreign_hq_review_evidence_url",
    "sig_foreign_hq_review_evidence_quote",
    "needs_anthropic_hq_review",
    "anthropic_hq_review_used",
    "anthropic_web_search_used",
    "anthropic_review_evidence_mode",
    "anthropic_web_search_queries",
    "anthropic_web_search_result_count",
    "anthropic_error",
    # Serper usage audit
    "serper_calls_used",
    "serper_queries_used",
    "serper_cache_hit",
    # Website fetch
    "website_fetch_count",
    # Manual Google mimic HQ check
    "manual_google_mimic_used",
    "plain_company_search_query",
    "plain_company_top_result_url",
    "plain_company_top_result_title",
    "plain_company_top_result_snippet",
    "plain_company_official_domain",
    "plain_company_official_result_rank",
    "official_domain_from_plain_search",
    "official_domain_matches_input_domain",
    "official_page_fetch_used",
    "official_page_fetch_url",
    "official_page_fetch_count",
    "official_page_fetch_error",
    "official_page_hq_evidence_found",
    "official_page_hq_evidence_quote",
    "official_page_hq_country",
    "official_page_hq_city",
    "official_page_hq_evidence_strength",
    "brand_hq_search_used",
    "brand_hq_search_queries",
    "brand_hq_top_result_url",
    "brand_hq_top_result_snippet",
    "brand_hq_evidence_found",
    "brand_hq_evidence_quote",
    "brand_hq_top_result_domain",
    "brand_hq_top_result_is_official_domain",
    "brand_hq_result_selection_reason",
    # Performance / run metadata
    "run_mode",
    "max_serper_calls_per_row",
    "early_stop_used",
    "early_stop_reason",
    "early_stop_blocked_reason",
    "row_runtime_seconds",
    # Local-country page detection
    "official_top_result_is_local_country_page",
    "official_top_result_country_context",
    "official_top_result_local_page_reason",
    # Brand-root HQ check
    "brand_root_candidates",
    "brand_root_hq_search_used",
    "brand_root_hq_search_queries",
    "brand_root_hq_evidence_found",
    "brand_root_hq_evidence_quote",
    "brand_root_hq_evidence_url",
    "brand_root_hq_country",
    "brand_root_hq_city",
    "brand_root_hq_source_type",
    # Known-global-brand detection
    "known_global_brand_detected",
    "known_global_brand_name",
    "known_global_brand_type",
    "known_global_brand_reason",
    # Domain-root HQ search (very first query: "{domain_root} headquarters")
    "domain_root",
    "domain_root_hq_query",
    "domain_root_hq_search_used",
    "domain_root_hq_evidence_found",
    "domain_root_hq_evidence_quote",
    "domain_root_hq_evidence_url",
    "domain_root_hq_country",
    "domain_root_hq_city",
    "domain_root_hq_confidence",
    "domain_root_hq_evidence_rank",
    "domain_root_hq_rejected_evidence_reason",
    # Simple HQ mode
    "simple_hq_mode_used",
    "simple_hq_query",
    # HQ Recovery workflow columns
    "hq_recovery_selected",
    "hq_recovery_selection_reason",
    "sig_foreign_hq_score_original_before_recovery",
    # Final reviewed score carried forward to next scoring run
    "sig_foreign_hq_score_for_next_scoring",
    "competitor_signal_excluded_from_next_scoring",
]

# Token columns tracked internally but NOT exported to Excel/CSV visible columns.
# They are stored per-row only for the Run Summary sheet aggregation.
_INTERNAL_TOKEN_COLS = [
    "anthropic_model_used",
    "anthropic_input_tokens",
    "anthropic_output_tokens",
    "anthropic_total_tokens",
    "anthropic_stop_reason",
]

# ---------------------------------------------------------------------------
# Italian city / province maps
# ---------------------------------------------------------------------------

_ITALY_CITIES: dict[str, tuple[str, str]] = {
    "milan":          ("Milan",          "Italy"),
    "milano":         ("Milano",         "Italy"),
    "rome":           ("Rome",           "Italy"),
    "roma":           ("Roma",           "Italy"),
    "turin":          ("Turin",          "Italy"),
    "torino":         ("Torino",         "Italy"),
    "naples":         ("Naples",         "Italy"),
    "napoli":         ("Napoli",         "Italy"),
    "bologna":        ("Bologna",        "Italy"),
    "bergamo":        ("Bergamo",        "Italy"),
    "brescia":        ("Brescia",        "Italy"),
    "verona":         ("Verona",         "Italy"),
    "padova":         ("Padova",         "Italy"),
    "padua":          ("Padua",          "Italy"),
    "vicenza":        ("Vicenza",        "Italy"),
    "treviso":        ("Treviso",        "Italy"),
    "modena":         ("Modena",         "Italy"),
    "parma":          ("Parma",          "Italy"),
    "firenze":        ("Firenze",        "Italy"),
    "florence":       ("Florence",       "Italy"),
    "genova":         ("Genova",         "Italy"),
    "genoa":          ("Genoa",          "Italy"),
    "venice":         ("Venice",         "Italy"),
    "venezia":        ("Venezia",        "Italy"),
    "trieste":        ("Trieste",        "Italy"),
    "bari":           ("Bari",           "Italy"),
    "catania":        ("Catania",        "Italy"),
    "palermo":        ("Palermo",        "Italy"),
    "perugia":        ("Perugia",        "Italy"),
    "ancona":         ("Ancona",         "Italy"),
    "trento":         ("Trento",         "Italy"),
    "bolzano":        ("Bolzano",        "Italy"),
    "reggio":         ("Reggio",         "Italy"),
    "cagliari":       ("Cagliari",       "Italy"),
    "lecce":          ("Lecce",          "Italy"),
    "pescara":        ("Pescara",        "Italy"),
    "pisa":           ("Pisa",           "Italy"),
    "siena":          ("Siena",          "Italy"),
    "udine":          ("Udine",          "Italy"),
    "novara":         ("Novara",         "Italy"),
    "monza":          ("Monza",          "Italy"),
    "bentivoglio":    ("Bentivoglio",    "Italy"),
    "interporto":     ("Interporto Bologna", "Italy"),
    "castel maggiore":("Castel Maggiore","Italy"),
    "calderara":      ("Calderara di Reno", "Italy"),
    "imola":          ("Imola",          "Italy"),
    "faenza":         ("Faenza",         "Italy"),
    "rimini":         ("Rimini",         "Italy"),
    "ravenna":        ("Ravenna",        "Italy"),
    "ferrara":        ("Ferrara",        "Italy"),
    "forlì":          ("Forlì",          "Italy"),
    "forli":          ("Forlì",          "Italy"),
    "piacenza":       ("Piacenza",       "Italy"),
    "reggio emilia":  ("Reggio Emilia",  "Italy"),
    "mantova":        ("Mantova",        "Italy"),
    "mantua":         ("Mantua",         "Italy"),
    "cremona":        ("Cremona",        "Italy"),
    "como":           ("Como",           "Italy"),
    "varese":         ("Varese",         "Italy"),
    "lecco":          ("Lecco",          "Italy"),
    "pavia":          ("Pavia",          "Italy"),
    "lodi":           ("Lodi",           "Italy"),
    "sesto san giovanni": ("Sesto San Giovanni", "Italy"),
    "cinisello balsamo":  ("Cinisello Balsamo",  "Italy"),
    "busto arsizio":  ("Busto Arsizio",  "Italy"),
    "gallarate":      ("Gallarate",      "Italy"),
    "saronno":        ("Saronno",        "Italy"),
    "rho":            ("Rho",            "Italy"),
    "segrate":        ("Segrate",        "Italy"),
    "assago":         ("Assago",         "Italy"),
    "cernusco":       ("Cernusco sul Naviglio", "Italy"),
    "agrate brianza": ("Agrate Brianza", "Italy"),
    "vimercate":      ("Vimercate",      "Italy"),
    "cassina de' pecchi": ("Cassina de' Pecchi", "Italy"),
    "sesto fiorentino": ("Sesto Fiorentino", "Italy"),
    "prato":          ("Prato",          "Italy"),
    "livorno":        ("Livorno",        "Italy"),
    "lucca":          ("Lucca",          "Italy"),
    "pistoia":        ("Pistoia",        "Italy"),
    "arezzo":         ("Arezzo",         "Italy"),
    "grosseto":       ("Grosseto",       "Italy"),
    "la spezia":      ("La Spezia",      "Italy"),
    "savona":         ("Savona",         "Italy"),
    "imperia":        ("Imperia",        "Italy"),
    "salerno":        ("Salerno",        "Italy"),
    "caserta":        ("Caserta",        "Italy"),
    "foggia":         ("Foggia",         "Italy"),
    "taranto":        ("Taranto",        "Italy"),
    "brindisi":       ("Brindisi",       "Italy"),
    "cosenza":        ("Cosenza",        "Italy"),
    "catanzaro":      ("Catanzaro",      "Italy"),
    "reggio calabria": ("Reggio Calabria", "Italy"),
    "messina":        ("Messina",        "Italy"),
    "agrigento":      ("Agrigento",      "Italy"),
    "ragusa":         ("Ragusa",         "Italy"),
    "siracusa":       ("Siracusa",       "Italy"),
    "trapani":        ("Trapani",        "Italy"),
    "sassari":        ("Sassari",        "Italy"),
    "nuoro":          ("Nuoro",          "Italy"),
    "oristano":       ("Oristano",       "Italy"),
    "macerata":       ("Macerata",       "Italy"),
    "pesaro":         ("Pesaro",         "Italy"),
    "ascoli piceno":  ("Ascoli Piceno",  "Italy"),
    "teramo":         ("Teramo",         "Italy"),
    "chieti":         ("Chieti",         "Italy"),
    "campobasso":     ("Campobasso",     "Italy"),
    "potenza":        ("Potenza",        "Italy"),
    "matera":         ("Matera",         "Italy"),
    # Marche / Abruzzo / Campania small cities needed for recovery tests
    "osimo":          ("Osimo",          "Italy"),
    "senigallia":     ("Senigallia",     "Italy"),
    "san benedetto del tronto": ("San Benedetto del Tronto", "Italy"),
    "san benedetto":  ("San Benedetto del Tronto", "Italy"),
    "nusco":          ("Nusco",          "Italy"),
    "avellino":       ("Avellino",       "Italy"),
    "jesi":           ("Jesi",           "Italy"),
    "fabriano":       ("Fabriano",       "Italy"),
    "civitanova marche": ("Civitanova Marche", "Italy"),
}

_INTL_CITIES: dict[str, tuple[str, str]] = {
    "berlin": ("Berlin", "Germany"), "munich": ("Munich", "Germany"),
    "münchen": ("München", "Germany"), "hamburg": ("Hamburg", "Germany"),
    "frankfurt": ("Frankfurt", "Germany"), "cologne": ("Cologne", "Germany"),
    "köln": ("Köln", "Germany"), "düsseldorf": ("Düsseldorf", "Germany"),
    "stuttgart": ("Stuttgart", "Germany"), "dortmund": ("Dortmund", "Germany"),
    "mannheim": ("Mannheim", "Germany"),
    "paris": ("Paris", "France"), "lyon": ("Lyon", "France"),
    "marseille": ("Marseille", "France"), "toulouse": ("Toulouse", "France"),
    "bordeaux": ("Bordeaux", "France"), "strasbourg": ("Strasbourg", "France"),
    "amsterdam": ("Amsterdam", "Netherlands"), "rotterdam": ("Rotterdam", "Netherlands"),
    "the hague": ("The Hague", "Netherlands"), "den haag": ("Den Haag", "Netherlands"),
    "eindhoven": ("Eindhoven", "Netherlands"), "utrecht": ("Utrecht", "Netherlands"),
    "madrid": ("Madrid", "Spain"), "barcelona": ("Barcelona", "Spain"),
    "valencia": ("Valencia", "Spain"), "seville": ("Seville", "Spain"),
    "sevilla": ("Sevilla", "Spain"),
    "zurich": ("Zurich", "Switzerland"), "zürich": ("Zürich", "Switzerland"),
    "geneva": ("Geneva", "Switzerland"), "genève": ("Genève", "Switzerland"),
    "bern": ("Bern", "Switzerland"), "basel": ("Basel", "Switzerland"),
    "london": ("London", "United Kingdom"), "manchester": ("Manchester", "United Kingdom"),
    "birmingham": ("Birmingham", "United Kingdom"), "edinburgh": ("Edinburgh", "United Kingdom"),
    "macclesfield": ("Macclesfield", "United Kingdom"),
    "cheshire": ("Cheshire", "United Kingdom"),
    "leeds": ("Leeds", "United Kingdom"), "bristol": ("Bristol", "United Kingdom"),
    "glasgow": ("Glasgow", "United Kingdom"), "sheffield": ("Sheffield", "United Kingdom"),
    "vienna": ("Vienna", "Austria"), "wien": ("Wien", "Austria"),
    "graz": ("Graz", "Austria"), "salzburg": ("Salzburg", "Austria"),
    "brussels": ("Brussels", "Belgium"), "bruxelles": ("Bruxelles", "Belgium"),
    "antwerp": ("Antwerp", "Belgium"), "antwerpen": ("Antwerpen", "Belgium"),
    "new york": ("New York", "United States"), "san francisco": ("San Francisco", "United States"),
    "chicago": ("Chicago", "United States"), "boston": ("Boston", "United States"),
    "los angeles": ("Los Angeles", "United States"), "seattle": ("Seattle", "United States"),
    "armonk": ("Armonk", "United States"),
    "stockholm": ("Stockholm", "Sweden"), "oslo": ("Oslo", "Norway"),
    "copenhagen": ("Copenhagen", "Denmark"), "helsinki": ("Helsinki", "Finland"),
    "tokyo": ("Tokyo", "Japan"), "beijing": ("Beijing", "China"),
    "shanghai": ("Shanghai", "China"), "singapore": ("Singapore", "Singapore"),
    "dublin": ("Dublin", "Ireland"), "warsaw": ("Warsaw", "Poland"),
    "lisbon": ("Lisbon", "Portugal"), "lisboa": ("Lisboa", "Portugal"),
    "luxembourg": ("Luxembourg", "Luxembourg"),
    "bertrange": ("Bertrange", "Luxembourg"),
    "altdorf": ("Altdorf", "Switzerland"), "poschiavo": ("Poschiavo", "Switzerland"),
    "uri": ("Uri", "Switzerland"),
    "brande": ("Brande", "Denmark"), "aarhus": ("Aarhus", "Denmark"),
    "odense": ("Odense", "Denmark"), "aalborg": ("Aalborg", "Denmark"),
    "venice": ("Venice", "Italy"), "venezia": ("Venezia", "Italy"),
    "santa maria di sala": ("Santa Maria di Sala", "Italy"),
    "verona": ("Verona", "Italy"), "padova": ("Padova", "Italy"),
    "padua": ("Padua", "Italy"),
}

_COUNTRY_ALIASES: dict[str, str] = {
    "italy": "Italy", "italia": "Italy", "italian": "Italy",
    "it": "Italy", "ita": "Italy",
    "germany": "Germany", "deutschland": "Germany", "german": "Germany",
    "de": "Germany", "deu": "Germany",
    "france": "France", "french": "France",
    "fr": "France", "fra": "France",
    "spain": "Spain", "españa": "Spain", "spanish": "Spain",
    "netherlands": "Netherlands", "holland": "Netherlands", "dutch": "Netherlands",
    "belgium": "Belgium", "belgian": "Belgium",
    "switzerland": "Switzerland", "swiss": "Switzerland", "svizzera": "Switzerland",
    "austria": "Austria", "austrian": "Austria",
    "united kingdom": "United Kingdom", "uk": "United Kingdom",
    "great britain": "United Kingdom", "gb": "United Kingdom",
    "england": "United Kingdom", "scotland": "United Kingdom",
    "wales": "United Kingdom",
    "united states": "United States", "usa": "United States", "us": "United States",
    "america": "United States",
    "japan": "Japan", "japanese": "Japan",
    "china": "China", "chinese": "China",
    "sweden": "Sweden", "swedish": "Sweden",
    "denmark": "Denmark", "danish": "Denmark",
    "norway": "Norway", "norwegian": "Norway",
    "finland": "Finland", "finnish": "Finland",
    "portugal": "Portugal", "portuguese": "Portugal",
    "poland": "Poland", "polish": "Poland",
    "luxembourg": "Luxembourg",
    "ireland": "Ireland", "irish": "Ireland",
}

_ITALY_PROVINCE_CODES: dict[str, tuple[str, str]] = {
    "AG": ("Agrigento", "Italy"), "AL": ("Alessandria", "Italy"),
    "AN": ("Ancona",    "Italy"), "AO": ("Aosta",       "Italy"),
    "AR": ("Arezzo",    "Italy"), "AP": ("Ascoli Piceno","Italy"),
    "AT": ("Asti",      "Italy"), "AV": ("Avellino",    "Italy"),
    "BA": ("Bari",      "Italy"), "BT": ("Barletta",    "Italy"),
    "BL": ("Belluno",   "Italy"), "BN": ("Benevento",   "Italy"),
    "BG": ("Bergamo",   "Italy"), "BI": ("Biella",      "Italy"),
    "BO": ("Bologna",   "Italy"), "BZ": ("Bolzano",     "Italy"),
    "BS": ("Brescia",   "Italy"), "BR": ("Brindisi",    "Italy"),
    "CA": ("Cagliari",  "Italy"), "CL": ("Caltanissetta","Italy"),
    "CB": ("Campobasso","Italy"), "CE": ("Caserta",     "Italy"),
    "CT": ("Catania",   "Italy"), "CZ": ("Catanzaro",   "Italy"),
    "CH": ("Chieti",    "Italy"), "CO": ("Como",        "Italy"),
    "CS": ("Cosenza",   "Italy"), "CR": ("Cremona",     "Italy"),
    "KR": ("Crotone",   "Italy"), "CN": ("Cuneo",       "Italy"),
    "EN": ("Enna",      "Italy"), "FM": ("Fermo",       "Italy"),
    "FE": ("Ferrara",   "Italy"), "FI": ("Firenze",     "Italy"),
    "FG": ("Foggia",    "Italy"), "FC": ("Forlì",       "Italy"),
    "FR": ("Frosinone", "Italy"), "GE": ("Genova",      "Italy"),
    "GO": ("Gorizia",   "Italy"), "GR": ("Grosseto",    "Italy"),
    "IM": ("Imperia",   "Italy"), "IS": ("Isernia",     "Italy"),
    "SP": ("La Spezia", "Italy"), "AQ": ("L'Aquila",    "Italy"),
    "LT": ("Latina",    "Italy"), "LE": ("Lecce",       "Italy"),
    "LC": ("Lecco",     "Italy"), "LI": ("Livorno",     "Italy"),
    "LO": ("Lodi",      "Italy"), "LU": ("Lucca",       "Italy"),
    "MC": ("Macerata",  "Italy"), "MN": ("Mantova",     "Italy"),
    "MS": ("Massa",     "Italy"), "MT": ("Matera",      "Italy"),
    "ME": ("Messina",   "Italy"), "MI": ("Milano",      "Italy"),
    "MO": ("Modena",    "Italy"), "MB": ("Monza",       "Italy"),
    "NA": ("Napoli",    "Italy"), "NO": ("Novara",      "Italy"),
    "NU": ("Nuoro",     "Italy"), "OR": ("Oristano",    "Italy"),
    "PD": ("Padova",    "Italy"), "PA": ("Palermo",     "Italy"),
    "PR": ("Parma",     "Italy"), "PV": ("Pavia",       "Italy"),
    "PG": ("Perugia",   "Italy"), "PU": ("Pesaro",      "Italy"),
    "PE": ("Pescara",   "Italy"), "PC": ("Piacenza",    "Italy"),
    "PI": ("Pisa",      "Italy"), "PT": ("Pistoia",     "Italy"),
    "PN": ("Pordenone", "Italy"), "PZ": ("Potenza",     "Italy"),
    "PO": ("Prato",     "Italy"), "RG": ("Ragusa",      "Italy"),
    "RA": ("Ravenna",   "Italy"), "RC": ("Reggio Calabria","Italy"),
    "RE": ("Reggio Emilia","Italy"),"RI": ("Rieti",     "Italy"),
    "RN": ("Rimini",    "Italy"), "RM": ("Roma",        "Italy"),
    "RO": ("Rovigo",    "Italy"), "SA": ("Salerno",     "Italy"),
    "SS": ("Sassari",   "Italy"), "SV": ("Savona",      "Italy"),
    "SI": ("Siena",     "Italy"), "SR": ("Siracusa",    "Italy"),
    "SO": ("Sondrio",   "Italy"), "TA": ("Taranto",     "Italy"),
    "TE": ("Teramo",    "Italy"), "TR": ("Terni",       "Italy"),
    "TO": ("Torino",    "Italy"), "TP": ("Trapani",     "Italy"),
    "TN": ("Trento",    "Italy"), "TV": ("Treviso",     "Italy"),
    "TS": ("Trieste",   "Italy"), "UD": ("Udine",       "Italy"),
    "VA": ("Varese",    "Italy"), "VE": ("Venezia",     "Italy"),
    "VB": ("Verbania",  "Italy"), "VC": ("Vercelli",    "Italy"),
    "VR": ("Verona",    "Italy"), "VV": ("Vibo Valentia","Italy"),
    "VI": ("Vicenza",   "Italy"), "VT": ("Viterbo",     "Italy"),
}

_PROVINCE_CODE_RE = re.compile(
    r"(?:[\(,\s])([A-Z]{2})(?:\)|(?:\s*[-,]\s*(?:Ital(?:y|ia)))|\s*$)",
    re.MULTILINE,
)

_ITALIAN_CAP_RE = re.compile(r"\b([1-9]\d{4})\b")

_ITALIAN_FISCAL_RE = re.compile(
    r"\b(p\.?\s*iva|partita\s+iva|codice\s+fiscale|c\.f\.)\b", re.IGNORECASE
)

_ITALIAN_ADDR_PHRASES = (
    "sede legale", "sede principale", "sede amministrativa", "sede operativa",
    "ufficio", "uffici", "registered office in italy", "registered in italy",
    "incorporata in italia", "societa italiana", "società italiana",
)

_GROUP_CONTEXT_RE = re.compile(
    r"\b(?:"
    r"part\s+of|owned\s+by|subsidiary\s+of|member\s+of|division\s+of"
    r"|affiliated\s+(?:with|to)|belongs?\s+to"
    r"|parent\s+company|holding\s+company|holding\s+group"
    r"|gruppo\b|gruppo\s+\w+|appartenente\s+(?:a|al)"
    r"|fa\s+parte\s+(?:del|di)|filiale\s+di|controllata\s+da"
    r")\b",
    re.IGNORECASE,
)

_HQ_PATTERNS = [
    re.compile(
        r"headquarter(?:s|ed)\s+(?:in|at)\s+([A-Z][A-Za-zÀ-ÿ\s,]{2,40}?)(?:\.|,|\s*\b(?:with|and|since|in\s+\d))",
        re.IGNORECASE,
    ),
    re.compile(r"head\s+office\s+(?:in|at|is\s+in)\s+([A-Z][A-Za-zÀ-ÿ\s,]{2,40}?)(?:\.|,)", re.IGNORECASE),
    re.compile(r"based\s+in\s+([A-Z][A-Za-zÀ-ÿ\s,]{2,35}?)(?:\.|,|\s+(?:and|with|since))", re.IGNORECASE),
    re.compile(r"Headquarters?\s*:\s*([A-Za-zÀ-ÿ\s,]{2,50}?)(?:\n|$|\.)", re.IGNORECASE),
    re.compile(r"Head\s+[Oo]ffice\s*:\s*([A-Za-zÀ-ÿ\s,]{2,50}?)(?:\n|$|\.)", re.IGNORECASE),
    re.compile(r"registered\s+office\s*(?:in|at|:)\s*([A-Za-zÀ-ÿ\s,]{2,50}?)(?:\n|$|\.|,)", re.IGNORECASE),
    re.compile(r"corporate\s+office\s*(?:in|at|:)\s*([A-Za-zÀ-ÿ\s,]{2,50}?)(?:\n|$|\.|,)", re.IGNORECASE),
    re.compile(r"principal\s+office\s*(?:in|at|:)\s*([A-Za-zÀ-ÿ\s,]{2,50}?)(?:\n|$|\.|,)", re.IGNORECASE),
    re.compile(r"registered\s+in\s+([A-Za-zÀ-ÿ\s,]{2,40}?)(?:\n|$|\.|,)", re.IGNORECASE),
    re.compile(r"sede\s+legale\s*[:\s]+(?:in\s+)?([A-Za-zÀ-ÿ\s,]{2,50}?)(?:\n|$|\.|,)", re.IGNORECASE),
    re.compile(r"sede\s+principale\s*[:\s]+(?:in\s+)?([A-Za-zÀ-ÿ\s,]{2,50}?)(?:\n|$|\.|,)", re.IGNORECASE),
    re.compile(r"sede\s+amministrativa\s+(?:in\s+)?([A-Za-zÀ-ÿ\s,]{2,40}?)(?:\n|$|\.|,)", re.IGNORECASE),
    re.compile(r"sede\s+operativa\s+(?:in\s+)?([A-Za-zÀ-ÿ\s,]{2,40}?)(?:\n|$|\.|,)", re.IGNORECASE),
    re.compile(r"con\s+sede\s+(?:a|in)\s+([A-Za-zÀ-ÿ\s,]{2,40}?)(?:\n|$|\.|,)", re.IGNORECASE),
    re.compile(r"ha\s+sede\s+(?:a|in)\s+([A-Za-zÀ-ÿ\s,]{2,40}?)(?:\n|$|\.|,)", re.IGNORECASE),
    re.compile(r"(?:la\s+)?(?:sua\s+)?sede\s+(?:è\s+)?(?:a|in)\s+([A-Za-zÀ-ÿ\s,]{2,40}?)(?:\n|$|\.|,)", re.IGNORECASE),
    re.compile(r"uffici\s+(?:a|in|di)\s+([A-Za-zÀ-ÿ\s,]{2,40}?)(?:\n|$|\.|,)", re.IGNORECASE),
    re.compile(r"ufficio\s+(?:a|in|di)\s+([A-Za-zÀ-ÿ\s,]{2,40}?)(?:\n|$|\.|,)", re.IGNORECASE),
    re.compile(r"c/o\s+[A-Za-zÀ-ÿ\s]+?,\s*([A-Za-zÀ-ÿ\s,]{2,40}?)(?:\n|$|\.|,)", re.IGNORECASE),
    re.compile(r"located\s+(?:in|at)\s+([A-Za-zÀ-ÿ\s,]{2,40}?)(?:\n|$|\.|,)", re.IGNORECASE),
]

# ---------------------------------------------------------------------------
# Global structure / multilingual detection constants
# ---------------------------------------------------------------------------

_GLOBAL_STRUCTURE_TERMS = re.compile(
    r"\b(?:"
    # Strong parent/ownership signals
    r"parent\s+company|owned\s+by|controlled\s+by|part\s+of\s+\w"
    r"|member\s+firm|member\s+of\s+\w"
    r"|global\s+network|international\s+network"
    r"|subsidiar(?:y|ies)"
    r"|offices\s+worldwide|worldwide\s+presence|global\s+offices"
    r"|group\s+headquarters|group\s+hq\b"
    r"|belongs?\s+to\s+\w"
    # HQ in foreign location (must be followed by location, checked contextually)
    r"|head\s+office\s+in\s+[A-Z]|headquartered\s+in\s+[A-Z]"
    r"|corporate\s+headquarters\s+in\s+[A-Z]|hq\s+in\s+[A-Z]"
    r")\b",
    re.IGNORECASE,
)

_LANG_HREFLANG_RE = re.compile(r'hreflang=["\']([a-z]{2})(?:-[A-Za-z]{2,4})?["\']', re.IGNORECASE)
_LANG_URL_RE = re.compile(r'href=["\'][^"\']{0,200}/([a-z]{2})/', re.IGNORECASE)
_LANG_LABELS: dict[str, str] = {
    "english": "en", "italiano": "it", "deutsch": "de", "français": "fr",
    "español": "es", "português": "pt", "svenska": "sv", "русский": "ru",
    "中文": "zh", "한국인": "ko", "dutch": "nl", "polski": "pl",
    "română": "ro", "česky": "cs",
}
_LANG_KNOWN = {
    "en", "it", "de", "fr", "es", "pt", "sv", "ru", "zh", "ko",
    "nl", "pl", "ja", "da", "fi", "no", "cs", "ro", "tr", "ar", "hu", "el",
}
_FETCH_TIMEOUT = 8

_HIGH_INTL_COLS = [
    "sig_intl_footprint_score", "sig_multicultural_score",
    "ti_intercultural_score", "sig_lnd_onboarding_score",
    "ti_onboarding_score", "model_signal_overall_confidence_score",
]

# ---------------------------------------------------------------------------
# Core location / country resolution
# ---------------------------------------------------------------------------

def _resolve_city_country(text: str) -> tuple[str, str]:
    """Given a location string, identify city and country. Returns ("", "") if unrecognised."""
    text_lc = text.lower().strip(" ,.")
    for alias, (city, country) in _ITALY_CITIES.items():
        if alias in text_lc:
            return city, country
    for m in _PROVINCE_CODE_RE.finditer(text):
        code = m.group(1)
        if code in _ITALY_PROVINCE_CODES:
            return _ITALY_PROVINCE_CODES[code]
    if _ITALIAN_CAP_RE.search(text):
        return "", "Italy"
    if _ITALIAN_FISCAL_RE.search(text):
        return "", "Italy"
    for phrase in _ITALIAN_ADDR_PHRASES:
        if phrase in text_lc:
            return "", "Italy"
    for alias, (city, country) in _INTL_CITIES.items():
        if alias in text_lc:
            return city, country
    for alias, country in _COUNTRY_ALIASES.items():
        if alias in text_lc:
            return "", country
    return "", ""


def resolve_country_from_location(location_text: str) -> tuple[str, str]:
    """
    Public helper: (country, city) from free-form location text.

    Examples:
        resolve_country_from_location("Milano")           -> ("Italy", "Milano")
        resolve_country_from_location("Bentivoglio (BO)") -> ("Italy", "Bentivoglio")
        resolve_country_from_location("Prato, Tuscany")   -> ("Italy", "Prato")
        resolve_country_from_location("Berlin, Germany")  -> ("Germany", "Berlin")
    """
    city, country = _resolve_city_country(location_text)
    return country, city


def _has_group_context(text: str) -> bool:
    return bool(_GROUP_CONTEXT_RE.search(text))


def _italy_in_text(text: str) -> bool:
    text_lc = text.lower()
    if "italy" in text_lc or "italia" in text_lc:
        return True
    for alias in _ITALY_CITIES:
        if alias in text_lc:
            return True
    m = _PROVINCE_CODE_RE.search(text)
    if m and m.group(1) in _ITALY_PROVINCE_CODES:
        return True
    if _ITALIAN_CAP_RE.search(text):
        return True
    return False


def _std_country(raw: str) -> str:
    """Normalise a country string via aliases."""
    return _COUNTRY_ALIASES.get(raw.strip().lower(), raw.strip())


def _normalize_country_for_hq(value: object) -> str:
    """Robust country normalizer for domestic-guard comparisons.

    Handles ISO-2 (IT, DE, FR, UK, US, CH, …), ISO-3, full names, and
    common aliases. Returns a lowercase canonical key so == comparisons work.
    """
    text = re.sub(r"\s+", " ", re.sub(r"\.", "", str(value or "").strip().lower()))
    _MAP = {
        "it": "italy", "ita": "italy", "italia": "italy",
        "italy": "italy", "italian": "italy",
        "de": "germany", "deu": "germany",
        "germany": "germany", "deutschland": "germany", "german": "germany",
        "fr": "france", "fra": "france",
        "france": "france", "french": "france",
        "uk": "united kingdom", "gb": "united kingdom", "gbr": "united kingdom",
        "united kingdom": "united kingdom", "great britain": "united kingdom",
        "england": "united kingdom", "scotland": "united kingdom",
        "wales": "united kingdom",
        "us": "united states", "usa": "united states",
        "united states": "united states",
        "united states of america": "united states",
        "ch": "switzerland", "che": "switzerland",
        "switzerland": "switzerland", "swiss": "switzerland",
        "lu": "luxembourg", "lux": "luxembourg", "luxembourg": "luxembourg",
        "nl": "netherlands", "nld": "netherlands",
        "netherlands": "netherlands", "holland": "netherlands",
        "dk": "denmark", "dnk": "denmark",
        "denmark": "denmark", "danish": "denmark",
        "se": "sweden", "swe": "sweden",
        "sweden": "sweden", "swedish": "sweden",
        "no": "norway", "nor": "norway",
        "norway": "norway", "norwegian": "norway",
        "at": "austria", "aut": "austria",
        "austria": "austria", "austrian": "austria",
        "es": "spain", "esp": "spain",
        "spain": "spain", "españa": "spain", "spanish": "spain",
        "be": "belgium", "bel": "belgium",
        "belgium": "belgium", "belgian": "belgium",
        "pl": "poland", "pol": "poland",
        "poland": "poland", "polish": "poland",
        "pt": "portugal", "prt": "portugal",
        "portugal": "portugal", "portuguese": "portugal",
        "ie": "ireland", "irl": "ireland",
        "ireland": "ireland", "irish": "ireland",
        "fi": "finland", "fin": "finland",
        "finland": "finland", "finnish": "finland",
        "jp": "japan", "jpn": "japan",
        "japan": "japan", "japanese": "japan",
        "cn": "china", "chn": "china",
        "china": "china", "chinese": "china",
    }
    return _MAP.get(text, text)


def _set_manual_review(result: dict, value: str, reason: str = "") -> None:
    """Set needs_manual_review, but never downgrade 'Yes' to 'No'."""
    existing = str(result.get("needs_manual_review", "")).strip().lower()
    if existing == "yes":
        result["needs_manual_review"] = "Yes"
    else:
        result["needs_manual_review"] = value
    if reason:
        old = str(result.get("hq_review_trigger") or "").strip()
        if old and reason not in old:
            result["hq_review_trigger"] = old + "; " + reason
        elif not old:
            result["hq_review_trigger"] = reason



def _safe_float(val: Any, default: float = 0.0) -> float:
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def detect_multilingual_site(domain: str) -> dict[str, Any]:
    """Cheap homepage fetch to detect multilingual website. Cached by caller."""
    out: dict[str, Any] = {
        "has_multilingual_site": False, "website_language_count": 0,
        "website_languages_detected": "", "website_fetch_count": 0,
    }
    if not domain:
        return out
    domain = re.sub(r"^https?://", "", domain.strip()).rstrip("/")
    html = ""
    for scheme in ("https", "http"):
        out["website_fetch_count"] += 1
        try:
            resp = _requests.get(
                f"{scheme}://{domain}", timeout=_FETCH_TIMEOUT,
                headers={"User-Agent": "Mozilla/5.0 (compatible; HQProbe/1.0)"},
                allow_redirects=True,
            )
            html = resp.text[:80_000]
            break
        except Exception:
            continue
    if not html:
        return out

    langs: set[str] = set()
    for m in _LANG_HREFLANG_RE.finditer(html):
        lang = m.group(1).lower()
        if lang in _LANG_KNOWN:
            langs.add(lang)
    for m in _LANG_URL_RE.finditer(html):
        lang = m.group(1).lower()
        if lang in _LANG_KNOWN:
            langs.add(lang)
    html_lc = html.lower()
    for label, code in _LANG_LABELS.items():
        if label in html_lc:
            langs.add(code)

    count = len(langs)
    out["website_language_count"] = count
    out["website_languages_detected"] = ", ".join(sorted(langs))
    out["has_multilingual_site"] = count >= 2
    return out


def detect_global_structure(
    organic: list[dict],
    kg_location: str,
    answer_box: str,
    page_text: str = "",
) -> dict[str, Any]:
    """Scan Serper evidence for global/group/network structure signals."""
    out: dict[str, Any] = {"has_global_structure_signal": False, "global_structure_evidence": ""}
    texts: list[str] = []
    if kg_location:
        texts.append(kg_location)
    if answer_box:
        texts.append(answer_box)
    for item in organic[:5]:
        texts.append(f"{item.get('title', '')} {item.get('snippet', '')}")
    if page_text:
        texts.append(page_text[:3_000])

    for text in texts:
        m = _GLOBAL_STRUCTURE_TERMS.search(text)
        if m:
            start = max(0, m.start() - 60)
            end   = min(len(text), m.end() + 60)
            out["has_global_structure_signal"] = True
            out["global_structure_evidence"]   = text[start:end].strip()
            return out
    return out


# Strong HQ-related patterns that point to a specific foreign location
_FOREIGN_HQ_NEAR_RE = re.compile(
    r"\b(?:head\s+office\s+in|headquartered\s+in|headquarters\s+(?:is\s+)?in"
    r"|corporate\s+(?:hq|headquarters)\s+in|hq\s+in|main\s+office\s+in"
    r"|parent\s+company\s+(?:is\s+)?(?:in|based|located)"
    r"|owned\s+by|controlled\s+by|subsidiary\s+of|part\s+of\s+the\s+\w"
    r")\b",
    re.IGNORECASE,
)

# Short all-caps company name risk (e.g. IET, BEA)
_SHORT_ACRONYM_RE = re.compile(r"^[A-Z]{2,4}$")


def compute_review_trigger(
    probe: dict,
    input_row: dict,
    multilingual: dict,
    global_struct: dict,
    company_name: str = "",
) -> tuple[str, bool]:
    """Returns (hq_review_trigger, needs_anthropic_hq_review).

    Only returns True when at least one *strong* condition is met.
    Generic profile labels (Headquarters: Milan, Locations, Founded) do not qualify.
    """
    triggers: list[str] = []

    det_country  = _std_country(probe.get("hq_detected_country") or "")
    inp_country  = _std_country(probe.get("input_country_used") or "")
    confidence   = probe.get("hq_confidence", "")
    lang_count   = int(multilingual.get("website_language_count") or 0)

    # 1. Strong parent/group/network signal from detect_global_structure
    if global_struct.get("has_global_structure_signal"):
        triggers.append("strong_parent_group_signal")

    # 2. Foreign country explicitly mentioned near real HQ-intent phrases in organic snippets
    all_snip = " ".join(
        f"{probe.get(f'top_organic_title_{i}', '')} {probe.get(f'top_organic_snippet_{i}', '')}"
        for i in range(1, 4)
    ).lower()
    if _FOREIGN_HQ_NEAR_RE.search(all_snip):
        for alias, country in _COUNTRY_ALIASES.items():
            if alias in all_snip and _std_country(country).lower() != inp_country.lower():
                triggers.append("conflicting_hq_evidence")
                break

    # 3. HQ detected as foreign (non-input country) with any confidence
    if det_country and inp_country and det_country.lower() != inp_country.lower():
        triggers.append("conflicting_hq_evidence")

    # Deduplicate conflicting_hq_evidence if added twice
    triggers = list(dict.fromkeys(triggers))

    # 4. Multilingual + high international signals (language count >= 2 required,
    #    but alone is NOT enough — need additional international signal from input row)
    has_high_intl = any(
        _safe_float(input_row.get(col)) > 0 for col in _HIGH_INTL_COLS if col in input_row
    )
    if lang_count >= 4 and has_high_intl:
        triggers.append("multilingual_plus_high_international_signals")
    elif lang_count >= 2 and has_high_intl and global_struct.get("has_global_structure_signal"):
        triggers.append("multilingual_plus_high_international_signals")

    # 5. Low/medium confidence + global/parent/multilingual signals present
    if confidence in ("Medium", "Low", ""):
        has_any_signal = (
            global_struct.get("has_global_structure_signal")
            or lang_count >= 3
            or bool(triggers)
        )
        if has_any_signal and has_high_intl:
            triggers.append("medium_confidence_plus_global_signal")

    # 6. Old score = 0/blank but multiple high international signals
    try:
        old_score = float(input_row.get("sig_foreign_hq_score") or 0)
    except (TypeError, ValueError):
        old_score = 0.0
    high_intl_count = sum(
        1 for col in _HIGH_INTL_COLS
        if col in input_row and _safe_float(input_row.get(col)) > 0
    )
    if old_score == 0 and high_intl_count >= 2:
        triggers.append("zero_fhq_score_multiple_intl_signals")

    # 7. Short acronym with ambiguous/low evidence
    name = (company_name or "").strip()
    if _SHORT_ACRONYM_RE.match(name) and confidence in ("Low", ""):
        triggers.append("short_acronym_ambiguous_evidence")

    # Deduplicate
    seen_t: set[str] = set()
    triggers = [t for t in triggers if not (t in seen_t or seen_t.add(t))]  # type: ignore[func-returns-value]

    return "; ".join(triggers), bool(triggers)


# ---------------------------------------------------------------------------
# Deterministic extraction
# ---------------------------------------------------------------------------

def _extract_deterministic(
    organic: list[dict],
    kg_location: str,
    answer_box: str,
    input_country: str = "",
) -> dict[str, Any]:
    texts: list[tuple[str, str]] = []
    if kg_location:
        texts.append((kg_location, ""))
    if answer_box:
        texts.append((answer_box, ""))
    for item in organic:
        combined = f"{item.get('title', '')} {item.get('snippet', '')}"
        texts.append((combined, item.get("link", "")))

    input_country_std = _std_country(input_country)
    fallback_group_hit: dict[str, Any] = {}

    for text, url in texts:
        is_group = _has_group_context(text)
        for pat in _HQ_PATTERNS:
            m = pat.search(text)
            if not m:
                continue
            captured = m.group(1).strip(" ,.")
            city, country = _resolve_city_country(captured)
            if not city and not country:
                continue

            candidate = {
                "hq_detected_city":    city,
                "hq_detected_country": country,
                "hq_confidence":       "High",
                "hq_reason":           f"Pattern match: '{m.group(0)[:80]}'",
                "hq_evidence_url":     url,
                "hq_evidence_quote":   text[:250].strip(),
            }

            country_std = _std_country(country)
            if country_std.lower() == input_country_std.lower():
                return candidate

            if is_group:
                if _italy_in_text(text) and input_country_std == "Italy":
                    pass  # skip, look for cleaner Italian hit
                elif not fallback_group_hit:
                    candidate["hq_confidence"] = "Low"
                    candidate["hq_reason"] = "[group context detected] " + candidate["hq_reason"]
                    fallback_group_hit = candidate
                continue

            return candidate

    # Direct Italian city scan (handles addresses without explicit HQ pattern)
    for text, url in texts:
        if _italy_in_text(text):
            text_lc = text.lower()
            for alias, (city, country) in _ITALY_CITIES.items():
                if alias in text_lc:
                    return {
                        "hq_detected_city":    city,
                        "hq_detected_country": "Italy",
                        "hq_confidence":       "Medium",
                        "hq_reason":           f"Italian city '{city}' found in evidence (no explicit HQ pattern)",
                        "hq_evidence_url":     url,
                        "hq_evidence_quote":   text[:250].strip(),
                    }

    return fallback_group_hit

# ---------------------------------------------------------------------------
# Model extraction (optional Claude fallback)
# ---------------------------------------------------------------------------

_MODEL_EXTRACTION_PROMPT = """\
You are a headquarters-location extractor.

You will receive a JSON array of search result snippets for a specific company.
Your task is to identify the EXACT headquarters location of THIS specific company.

Rules:
- Extract headquarters of the EXACT company named, not of a parent group or subsidiary.
- Do NOT infer HQ from: international activity, branches, distributors, customers, \
training providers, competitor mentions, or group ownership.
- If the evidence says the company is PART OF a foreign group but also gives an Italian \
registered office or corporate address for the exact company itself, return Italy — \
group context is not the same as the company's own HQ.
- If the evidence is ONLY about a parent/holding group with no specific address for the \
exact company, set is_group_or_parent_only to true and confidence to "Low".
- entity_match: "Exact" if evidence explicitly names this company, "Likely" if very \
similar name, "Weak" if indirect, "No" if evidence is about a different entity entirely.
- If uncertain, set confidence to "Low" or "Unknown".

Return ONLY valid JSON, no other text:
{
  "hq_city": "",
  "hq_region": "",
  "hq_country": "",
  "confidence": "High|Medium|Low|Unknown",
  "reason": "",
  "evidence_url": "",
  "evidence_quote": "",
  "entity_match": "Exact|Likely|Weak|No",
  "is_group_or_parent_only": false
}
"""


def _model_extract(
    company_name: str,
    snippets: list[dict],
    anthropic_key: str,
    model: str = "claude-haiku-4-5-20251001",
) -> dict[str, Any]:
    try:
        import anthropic as _anthropic
    except ImportError:
        return {"probe_error": "anthropic package not installed (pip install anthropic)"}

    evidence = json.dumps(snippets[:5], ensure_ascii=False)
    client = _anthropic.Anthropic(api_key=anthropic_key)
    try:
        resp = client.messages.create(
            model=model,
            max_tokens=400,
            messages=[{"role": "user", "content": f"Company: {company_name}\n\nSearch results:\n{evidence}"}],
            system=_MODEL_EXTRACTION_PROMPT,
        )
        raw = resp.content[0].text.strip()
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        parsed = json.loads(raw)
        parsed.setdefault("entity_match", "")
        parsed.setdefault("is_group_or_parent_only", False)
        parsed["_usage"] = {
            "model": resp.model,
            "input_tokens":  getattr(resp.usage, "input_tokens",  0),
            "output_tokens": getattr(resp.usage, "output_tokens", 0),
            "stop_reason":   getattr(resp, "stop_reason", ""),
        }
        parsed["_usage"]["total_tokens"] = (
            parsed["_usage"]["input_tokens"] + parsed["_usage"]["output_tokens"]
        )
        return parsed
    except json.JSONDecodeError as exc:
        return {"probe_error": f"Model JSON parse error: {exc}"}
    except Exception as exc:
        return {"probe_error": f"Model error: {exc}"}

# ---------------------------------------------------------------------------
# Anthropic HQ structure adjudication (optional, for ambiguous cases)
# ---------------------------------------------------------------------------

_ANTHROPIC_HQ_REVIEW_PROMPT = """\
You are an expert HQ analyst reviewing company headquarters structure.

Based on the provided evidence classify the company's HQ structure:
- "domestic_italy": Company's own HQ is in Italy with no clear foreign parent or global network.
- "foreign_parent": Company has a non-Italian parent, or its own registered HQ is outside Italy.
- "global_network": Company belongs to a global professional network (e.g. KPMG, PwC networks) — \
local entity may be Italian but the brand/network is worldwide.
- "exporter_multilingual": Italian HQ, multilingual website or international activity, \
but no clear foreign parent or global network structure.
- "unclear": Evidence is ambiguous or insufficient.

Key rules:
- Multilingual website alone does NOT make "foreign_parent".
- A local entity in a professional services network should be "global_network", not "foreign_parent".
- If clear foreign registered office or foreign parent company → "foreign_parent".

Scoring 0–3:
- review_foreign_parent_score: 3 = clear foreign parent/HQ, 0 = no evidence
- review_global_network_score: 3 = clear global network, 0 = no evidence
- review_multilingual_website_score: 3 = clearly multilingual, 0 = Italian-only
- review_recommended_hq_signal: overall signal strength recommendation
- sig_foreign_hq_score_reviewed: 3 = clear foreign/global network; 1 = exporter/multilingual only; 0 = domestic Italy

Return ONLY valid JSON:
{
  "hq_structure_type": "domestic_italy|foreign_parent|global_network|exporter_multilingual|unclear",
  "local_entity_hq_country": "",
  "local_entity_hq_city": "",
  "parent_group_hq_country": "",
  "parent_group_hq_city": "",
  "review_foreign_parent_score": 0,
  "review_global_network_score": 0,
  "review_multilingual_website_score": 0,
  "review_recommended_hq_signal": 0,
  "sig_foreign_hq_score_reviewed": 0,
  "confidence": "High|Medium|Low|Unknown",
  "reason": "",
  "evidence_url": "",
  "evidence_quote": ""
}
"""


def _anthropic_hq_adjudicate(
    company_name: str,
    domain: str,
    probe: dict,
    multilingual: dict,
    global_struct: dict,
    anthropic_key: str,
    model: str = "claude-haiku-4-5-20251001",
) -> dict[str, Any]:
    try:
        import anthropic as _anthropic
    except ImportError:
        return {"probe_error": "anthropic package not installed (pip install anthropic)"}

    evidence = {
        "company": company_name,
        "domain": domain,
        "hq_detected_country": probe.get("hq_detected_country", ""),
        "hq_detected_city": probe.get("hq_detected_city", ""),
        "hq_confidence": probe.get("hq_confidence", ""),
        "hq_reason": probe.get("hq_reason", ""),
        "has_multilingual_site": multilingual.get("has_multilingual_site", False),
        "website_languages_detected": multilingual.get("website_languages_detected", ""),
        "has_global_structure_signal": global_struct.get("has_global_structure_signal", False),
        "global_structure_evidence": global_struct.get("global_structure_evidence", ""),
        "serper_snippets": [
            {
                "title":   probe.get(f"top_organic_title_{i}", ""),
                "snippet": probe.get(f"top_organic_snippet_{i}", ""),
                "url":     probe.get(f"top_organic_url_{i}", ""),
            }
            for i in range(1, 4)
            if probe.get(f"top_organic_snippet_{i}")
        ],
    }

    client = _anthropic.Anthropic(api_key=anthropic_key)
    try:
        resp = client.messages.create(
            model=model,
            max_tokens=600,
            messages=[{"role": "user", "content": json.dumps(evidence, ensure_ascii=False)}],
            system=_ANTHROPIC_HQ_REVIEW_PROMPT,
        )
        raw = resp.content[0].text.strip()
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        parsed = json.loads(raw)
        parsed["_usage"] = {
            "model":         resp.model,
            "input_tokens":  getattr(resp.usage, "input_tokens",  0),
            "output_tokens": getattr(resp.usage, "output_tokens", 0),
            "stop_reason":   getattr(resp, "stop_reason", ""),
        }
        parsed["_usage"]["total_tokens"] = (
            parsed["_usage"]["input_tokens"] + parsed["_usage"]["output_tokens"]
        )
        return parsed
    except json.JSONDecodeError as exc:
        return {"probe_error": f"Anthropic HQ review JSON parse error: {exc}"}
    except Exception as exc:
        return {"probe_error": f"Anthropic HQ review error: {exc}"}

# ---------------------------------------------------------------------------
# Classification
# ---------------------------------------------------------------------------

def _classify(
    hq_country: str,
    input_country: str,
    confidence: str,
    reason: str,
    organic: list[dict],
) -> tuple[str, bool]:
    """Returns (foreign_hq_simple, needs_manual_review)."""
    if not hq_country or confidence in ("Unknown", ""):
        return "", True

    hq_std    = _std_country(hq_country)
    input_std = _std_country(input_country)
    foreign   = hq_std.lower() != input_std.lower()

    needs_review = confidence in ("Low", "Unknown")
    if not needs_review and organic:
        all_text = " ".join(
            f"{r.get('title','')} {r.get('snippet','')}" for r in organic[:3]
        ).lower()
        if sum(1 for alias in _COUNTRY_ALIASES if alias in all_text) > 2:
            needs_review = True

    return ("True" if foreign else "False"), needs_review

# ---------------------------------------------------------------------------
# Serper search
# ---------------------------------------------------------------------------

def _serper_search(query: str, api_key: str) -> tuple[dict, str]:
    headers = {"X-API-KEY": api_key, "Content-Type": "application/json"}
    try:
        resp = _requests.post(
            _SERPER_URL, headers=headers,
            json={"q": query, "num": 5},
            timeout=_REQUEST_TIMEOUT,
        )
        resp.raise_for_status()
        return resp.json(), ""
    except _requests.HTTPError as exc:
        return {}, f"HTTP {exc.response.status_code}: {exc.response.text[:120]}"
    except Exception as exc:
        return {}, str(exc)[:200]


# ---------------------------------------------------------------------------
# Manual Google Mimic HQ Check
# ---------------------------------------------------------------------------

_DIRECTORY_DOMAINS = frozenset({
    "linkedin.com", "facebook.com", "instagram.com", "twitter.com", "x.com",
    "youtube.com", "wikipedia.org", "crunchbase.com", "bloomberg.com",
    "zoominfo.com", "dnb.com", "hoovers.com", "glassdoor.com", "indeed.com",
    "infobel.com", "paginegialle.it", "kompass.com", "europages.com",
    "europages.it", "registroimprese.it", "registro.imprese.it",
    "atoka.io", "cerved.com", "ilsole24ore.com", "corriere.it",
    "bizjournals.com", "wsj.com", "ft.com", "reuters.com",
    "yelp.com", "foursquare.com", "angellist.com", "pitchbook.com",
    "owler.com", "manta.com", "opencorporates.com", "sec.gov",
})

_LEGAL_SUFFIX_RE = re.compile(
    r"\s*\b(?:S\.?\s*P\.?\s*A\.?|S\.?\s*R\.?\s*L\.?|S\.?\s*A\.?\s*S\.?"
    r"|S\.?\s*N\.?\s*C\.?|S\.?\s*A\.?|GmbH|AG\b|Ltd\.?|LLC\.?|Corp\.?"
    r"|Corporation|Inc\.?|B\.?\s*V\.?\b|N\.?\s*V\.?\b"
    r"|ITALIA\b|ITALY\b|GROUP\b|GRUPPO\b|HOLDING\b|INTERNATIONAL\b)\s*$",
    re.IGNORECASE,
)

_OFFICIAL_PAGE_PATHS = [
    "/", "/about-us", "/about", "/company",
    "/en/about-us", "/en/about", "/ch/en/about-us",
    "/it/chi-siamo", "/chi-siamo", "/contatti", "/contact",
]

_OFFICIAL_HQ_PATTERNS: list[tuple[re.Pattern, str]] = [
    (re.compile(r"its?\s+head\s+office\s+is\s+in\s+([A-Za-zÀ-ÿ\s,\(\)]{3,60}?)(?:\.|;|\n|\Z)", re.I), "Strong"),
    (re.compile(r"head\s+office\s+(?:is\s+)?(?:located\s+)?in\s+([A-Za-zÀ-ÿ\s,\(\)]{3,60}?)(?:\.|;|,\s*(?:with|which)|\n|\Z)", re.I), "Strong"),
    (re.compile(r"headquartered\s+in\s+([A-Za-zÀ-ÿ\s,\(\)]{3,60}?)(?:\.|;|\n|\Z)", re.I), "Strong"),
    (re.compile(r"headquarters\s+(?:are\s+)?(?:is\s+)?(?:located\s+)?in\s+([A-Za-zÀ-ÿ\s,\(\)]{3,60}?)(?:\.|;|\n|\Z)", re.I), "Strong"),
    (re.compile(r"corporate\s+headquarters\s+in\s+([A-Za-zÀ-ÿ\s,\(\)]{3,60}?)(?:\.|;|\n|\Z)", re.I), "Strong"),
    (re.compile(r"group\s+headquarters\s+in\s+([A-Za-zÀ-ÿ\s,\(\)]{3,60}?)(?:\.|;|\n|\Z)", re.I), "Strong"),
    (re.compile(r"operational\s+headquarters\s+in\s+([A-Za-zÀ-ÿ\s,\(\)]{3,60}?)(?:\.|;|\n|\Z)", re.I), "Strong"),
    (re.compile(r"registered\s+office\s+(?:in|at|:)\s*([A-Za-zÀ-ÿ\s,\(\)]{3,60}?)(?:\.|;|\n|\Z)", re.I), "Strong"),
    (re.compile(r"sede\s+(?:principale|legale|centrale)\s*[:\s]+(?:in\s+)?([A-Za-zÀ-ÿ\s,]{3,50}?)(?:\.|;|,|\n)", re.I), "Strong"),
    (re.compile(r"parent\s+company\s+is\s+([A-Za-zÀ-ÿ\s,]{3,60}?)(?:\.|;|\n)", re.I), "Medium"),
    (re.compile(r"part\s+of\s+the\s+([A-Za-zÀ-ÿ\s]{3,40}?)\s+group\b", re.I), "Medium"),
    (re.compile(r"subsidiary\s+of\s+([A-Za-zÀ-ÿ\s,]{3,60}?)(?:\.|;|\n)", re.I), "Medium"),
    (re.compile(r"owned\s+by\s+([A-Za-zÀ-ÿ\s,]{3,60}?)(?:\.|;|,|\n)", re.I), "Medium"),
    (re.compile(r"based\s+in\s+([A-Za-zÀ-ÿ\s,\(\)]{3,50}?)(?:\.|;|,\s*(?:and|with|since)|\n)", re.I), "Medium"),
    # Structured data pattern: "Headquarters London" (no preposition), followed by non-location word
    (re.compile(r"\bheadquarters\s+([A-Z][A-Za-zÀ-ÿ\s]{2,40}?)(?:\s+(?:Type|Founded|Employees?|Industry|Revenue|CEO|Chairman|Website|Phone|Email)|[,;.]|\s*$)", re.I), "Medium"),
]


def _strip_html(html: str) -> str:
    html = re.sub(r"<(script|style|head)[^>]*>.*?</\1>", " ", html, flags=re.S | re.I)
    html = re.sub(r"<[^>]+>", " ", html)
    return re.sub(r"\s+", " ", html).strip()


def _is_directory_url(url: str) -> bool:
    try:
        from urllib.parse import urlparse
        host = urlparse(url).netloc.lower()
        host = re.sub(r"^www\.", "", host)
        return any(host == d or host.endswith("." + d) for d in _DIRECTORY_DOMAINS)
    except Exception:
        return False


def _get_domain_body(url_or_domain: str) -> str:
    """Return the main token of a domain (e.g. 'repower' from 'repower.com')."""
    s = url_or_domain.strip()
    try:
        from urllib.parse import urlparse
        parsed = urlparse(s if "://" in s else "https://" + s)
        host = re.sub(r"^www\.", "", parsed.netloc.lower()) or s
    except Exception:
        host = s
    parts = host.split(".")
    return parts[-2] if len(parts) >= 2 else parts[0]


def _extract_brand(company_name: str, domain: str) -> list[str]:
    """Strip legal suffixes; also extract domain body. Returns up to 3 brand candidates."""
    brands: list[str] = []
    cleaned = company_name.strip()
    for _ in range(6):
        new = _LEGAL_SUFFIX_RE.sub("", cleaned).strip().rstrip(",. ")
        if not new or new == cleaned:
            break
        cleaned = new
    if cleaned and cleaned.lower() != company_name.strip().lower():
        brands.append(cleaned)
    # Title-case variant
    if cleaned:
        tc = cleaned.title()
        if tc not in brands:
            brands.append(tc)
    # Domain body
    if domain:
        body = _get_domain_body(domain)
        if body and len(body) > 2:
            tc_body = body.title()
            if tc_body.lower() not in {b.lower() for b in brands}:
                brands.append(tc_body)
    seen: set[str] = set()
    out: list[str] = []
    for b in brands:
        if b.lower() not in seen and len(b) > 2:
            seen.add(b.lower())
            out.append(b)
    return out[:3]


# Terms to strip when building brand-root candidates
_LOCAL_ENTITY_RE = re.compile(
    r"\s*\b(?:"
    r"ITALIA(?:N[AE]?)?|ITALY|ITALIAN"
    r"|SEALING\s+SOLUTIONS|SOLUTIONS|INGREDIENTS"
    r"|ADVISORY|SERVICES|SERVICE|CONSULTING|MANAGEMENT"
    r"|S\.?\s*P\.?\s*A\.?|S\.?\s*R\.?\s*L\.?|S\.?\s*A\.?\s*S\.?"
    r"|S\.?\s*N\.?\s*C\.?|S\.?\s*A\.?|GmbH|AG\b|Ltd\.?|LLC\.?|Corp\.?"
    r"|Corporation|Inc\.?|B\.?\s*V\.?\b|N\.?\s*V\.?\b"
    r"|GROUP\b|GRUPPO\b|HOLDING\b|INTERNATIONAL\b|GLOBAL\b"
    r")\b\s*",
    re.IGNORECASE,
)

# Local-country page URL patterns
_LOCAL_COUNTRY_URL_RE = re.compile(
    r"(?:/|-|_|\.)"
    r"(?:it(?:al(?:ia?|ian)|y)|"
    r"in[-_]it(?:aly)?|"
    r"[-_]italy[-_/]?|"
    r"[-_]italia[-_/]?|"
    r"locations?/|"
    r"offices?/|"
    r"contact/|"
    r"filial[ei]/)",
    re.IGNORECASE,
)

# Local-country text signals in title/snippet
_LOCAL_COUNTRY_TEXT_RE = re.compile(
    r"\b(?:"
    r"Italy\s*[-–]|Italia\s*[-–]"
    r"|Italy S\.?[PR]\.?[LA]\.?|Italia S\.?[PR]\.?[LA]\.?"
    r"|Italy\s+(?:office|branch|location|operations|division|unit)"
    r"|Italian\s+(?:office|branch|subsidiary|division|operations)"
    r"|sede\s+italian[ae]"
    r"|ufficio\s+italiano"
    r")\b",
    re.IGNORECASE,
)


def _extract_brand_roots(company_name: str, domain: str) -> list[str]:
    """Generate brand-root candidates from most specific to broadest.

    For 'DATWYLER SEALING SOLUTIONS ITALY S.P.A.' + 'datwyler.com':
      → ['Datwyler Sealing Solutions', 'Datwyler']
    For 'NADARA ITALY S.P.A.' + 'nadara.com':
      → ['Nadara']
    For 'CSM INGREDIENTS ITALY S.P.A.' + 'csmingredients.com':
      → ['CSM Ingredients', 'CSM']
    """
    _VOWELS = set("aeiouAEIOU")

    def _smart_case(s: str) -> str:
        """Title-case, preserving consonant-only all-caps acronyms (CSM, KPMG, not CARE)."""
        parts = s.split()
        return " ".join(
            p if (p.isupper() and not any(c in _VOWELS for c in p)) else p.title()
            for p in parts
        )

    candidates: list[str] = []

    # 1. Strip legal + local suffixes iteratively; collect intermediate results
    cleaned = company_name.strip()
    prev_states: list[str] = []
    for _ in range(10):
        new = _LEGAL_SUFFIX_RE.sub("", cleaned).strip().rstrip(",. ")
        new = re.sub(r"\s+", " ", new)
        if not new or new == cleaned:
            break
        cleaned = new
        prev_states.append(cleaned)

    # After basic legal-suffix strip, also strip local-entity terms
    after_local = re.sub(r"\s+", " ", _LOCAL_ENTITY_RE.sub(" ", cleaned)).strip().rstrip(",. -")
    if after_local and after_local.lower() != cleaned.lower():
        # prev_states[-1] is most specific (e.g. "DATWYLER SEALING SOLUTIONS")
        if prev_states and prev_states[-1].strip().lower() != after_local.lower():
            candidates.append(_smart_case(prev_states[-1].strip()))  # e.g. "Datwyler Sealing Solutions"
        candidates.append(_smart_case(after_local))  # e.g. "Datwyler" after stripping "Sealing Solutions"
    elif prev_states:
        candidates.append(_smart_case(prev_states[-1].strip()))
    else:
        # name unchanged by legal strips — smart-case it
        after_loc2 = re.sub(r"\s+", " ", _LOCAL_ENTITY_RE.sub(" ", company_name)).strip().rstrip(",. -")
        if after_loc2:
            candidates.append(_smart_case(after_loc2))

    # 2. Also add domain body as a candidate
    if domain:
        body = _get_domain_body(domain)
        if body and len(body) > 2:
            body_tc = body.title()
            # Match meaningful name tokens (≥3 chars) against domain body
            name_tokens = re.findall(r"[A-Za-z]{3,}", company_name)
            matched = [t for t in name_tokens
                       if (t.lower() in body.lower() or body.lower() in t.lower())
                       and t.lower() not in {"italy", "italia", "italian", "advisory",
                                              "solutions", "services", "ingredients",
                                              "management", "consulting"}]
            domain_root = " ".join(_smart_case(t) for t in matched[:3]) if matched else body_tc
            if domain_root.lower() not in {c.lower() for c in candidates}:
                candidates.append(domain_root)

    # 3. Deduplicate; remove empty or too-short
    seen: set[str] = set()
    out: list[str] = []
    for c in candidates:
        c = _smart_case(c.strip())
        # Allow short tokens (≥2 chars) for known acronyms like "EY", "HP"
        if len(c) >= 2 and c.lower() not in seen:
            seen.add(c.lower())
            out.append(c)

    return out[:4]


def _is_local_country_page(
    url: str, title: str, snippet: str, country_keyword: str = "italy"
) -> tuple[bool, str, str]:
    """Return (is_local, country_context, reason).

    Checks whether the given URL / title / snippet indicate a local
    country office / location / branch page rather than a global brand page.
    """
    ck = country_keyword.lower()
    combined_url   = url.lower()
    combined_text  = f"{title} {snippet}".lower()

    # URL patterns
    url_local = bool(_LOCAL_COUNTRY_URL_RE.search(combined_url))
    # Text patterns (title/snippet)
    text_local = bool(_LOCAL_COUNTRY_TEXT_RE.search(f"{title} {snippet}"))

    # Generic contact/location path on a brand domain
    contact_path = bool(re.search(r"/(?:contact|location|office|branch|find-us|where)", combined_url))

    if url_local:
        return True, ck, f"url_contains_local_pattern:{combined_url}"
    if text_local:
        return True, ck, f"title_snippet_local_signal"
    if contact_path and ck in combined_url + combined_text:
        return True, ck, f"contact_path_with_country_keyword"

    return False, "", ""


def _find_official_result(
    organic: list[dict], input_domain: str
) -> tuple[int, dict, str]:
    """Return (1-based rank, result, clean_netloc) of best non-directory organic result."""
    input_body = _get_domain_body(input_domain) if input_domain else ""
    first_rank, first_result, first_netloc = 0, {}, ""
    for i, item in enumerate(organic[:10], start=1):
        url = item.get("link", "")
        if not url or _is_directory_url(url):
            continue
        try:
            from urllib.parse import urlparse
            netloc = re.sub(r"^www\.", "", urlparse(url).netloc.lower())
        except Exception:
            netloc = url
        body = _get_domain_body(netloc)
        if input_body and (input_body in body or body in input_body):
            return i, item, netloc
        if not first_rank:
            first_rank, first_result, first_netloc = i, item, netloc
    return first_rank, first_result, first_netloc


def _fetch_page_text(url: str, cache: dict) -> tuple[str, str]:
    """Fetch URL, strip HTML, return (text[:8000], error). Cached."""
    ck = ("fetch", url)
    if ck in cache:
        return cache[ck]
    try:
        resp = _requests.get(
            url, timeout=_FETCH_TIMEOUT,
            headers={"User-Agent": "Mozilla/5.0 (compatible; HQProbe/1.0)"},
            allow_redirects=True,
        )
        resp.raise_for_status()
        text = _strip_html(resp.text[:100_000])[:8_000]
        result: tuple[str, str] = (text, "")
    except Exception as exc:
        result = ("", str(exc)[:120])
    cache[ck] = result
    return result


def _scan_for_hq(text: str) -> tuple[str, str, str, str]:
    """Scan text with OFFICIAL_HQ_PATTERNS. Returns (quote, country, city, strength)."""
    for pat, strength in _OFFICIAL_HQ_PATTERNS:
        m = pat.search(text)
        if m:
            captured = m.group(1).strip(" ,.()")
            city, country = _resolve_city_country(captured)
            if country or city:
                start = max(0, m.start() - 30)
                end   = min(len(text), m.end() + 80)
                quote = text[start:end].strip()
                return quote[:300], country, city, strength
    return "", "", "", ""


_HQ_SIGNAL_RE = re.compile(
    r"\b(?:head\s+office|headquarters?|hq)\b",
    re.IGNORECASE,
)

# Regional/divisional/subsidiary HQ phrases — these should NOT auto-score as global parent HQ
_REGIONAL_HQ_RE = re.compile(
    r"\b(?:"
    r"(?:north\s*america[n]?|na|emea|apac|latam|asia[\s\-]pacific|european?|"
    r"regional?|division[al]?|segment|us|usa?)\s+(?:headquarters?|hq)|"
    r"headquarters?\s+for\s+(?:north\s*america[n]?|emea|apac|europe|us|usa?)|"
    r"usa?\s+corp(?:oration)?\s+headquarters?|"
    r"branch\s+(?:office|headquarters?)|local\s+(?:office|headquarters?)|"
    r"subsidiary\s+headquarters?|affiliate\s+headquarters?"
    r")\b",
    re.IGNORECASE,
)

# Subsidiary/third-party evidence phrases — downgrade confidence, require manual review
_SUBSIDIARY_SIGNAL_RE = re.compile(
    r"\b(?:usa?\s+corp(?:oration)?|subsidiary|affiliate|"
    r"branch\s+(?:office|headquarters?)|local\s+(?:office|headquarters?)|"
    r"regional\s+headquarters?|division\s+headquarters?)\b",
    re.IGNORECASE,
)

# ── Score-3 eligibility helper ────────────────────────────────────────────────
_S3_HQ_PHRASE_RE = re.compile(
    r"\b(?:headquarters?|headquartered|head\s+office|global\s+hq|"
    r"corporate\s+(?:headquarters?|hq)|group\s+(?:headquarters?|hq)|"
    r"international\s+(?:headquarters?|hq)|world\s+headquarters?|"
    r"principal\s+(?:office|place\s+of\s+business))\b",
    re.IGNORECASE,
)
_S3_BIO_REJECT_RE = re.compile(
    r"\b(?:brigadier\s+general|national\s+guard\s+biography|biography|military|"
    r"officer|commander|colonel|lieutenant|sergeant|captain|admiral|"
    r"general(?:\s+of)?|born\s+in|graduated|university|professor|physician|"
    r"surgeon|director\s+of\s+(?:the\s+)?(?:national|state)|"
    r"politician|senator|minister\s+of|secretary\s+of\s+state)\b",
    re.IGNORECASE,
)
_S3_BRANCH_REJECT_RE = re.compile(
    r"\b(?:branch\s+(?:in|opens?|office|at)|office\s+in|regional\s+office|"
    r"sales\s+office|service\s+(?:center|centre)|service\s+branch|"
    r"distribution\s+(?:center|centre)|north\s+america(?:n)?(?:\s+(?:hq|inc\.?|headquarters?))?|"
    r"us[a]?\s+subsidiary|subsidiary|affiliate|"
    r"get\s+directions?|locations\s+primary|primary\s+location)\b"
    r"|Holland,?\s*Michigan|Holland,?\s*MI\b",
    re.IGNORECASE,
)
_S3_DIRECTORY_REJECT_RE = re.compile(
    r"\b(?:company\s+profile|business\s+directory|yellow\s+pages?|"
    r"contact\s+details?|address\s+and\s+phone|jidipi|"
    r"dnb\.com|bloomberg\.com|zoominfo\.com|crunchbase\.com|"
    r"opencorporates\.com|companies\s+house|directory)\b",
    re.IGNORECASE,
)


def _is_score3_eligible_hq_evidence(
    quote: str,
    detected_country: str,
    input_country: str,
    evidence_url: str,
) -> tuple[bool, str]:
    """Return (eligible, reason). Score 3 requires a clear HQ phrase and no hard-reject signals."""
    q = quote or ""
    url = evidence_url or ""

    # Must have a detected country
    if not (detected_country or "").strip():
        return False, "no_detected_country"

    # Domestic: detected == input country
    _det_n = _normalize_country_for_hq(detected_country)
    _inp_n = _normalize_country_for_hq(input_country)
    if _det_n and _inp_n and _det_n == _inp_n:
        return False, "domestic_hq_not_foreign"

    if _S3_BIO_REJECT_RE.search(q):
        return False, "biography_or_personal_evidence"
    if _S3_BRANCH_REJECT_RE.search(q):
        return False, "branch_subsidiary_or_regional_hq"
    if _S3_DIRECTORY_REJECT_RE.search(q) or _S3_DIRECTORY_REJECT_RE.search(url):
        return False, "directory_source"
    if not _S3_HQ_PHRASE_RE.search(q):
        return False, "no_clear_hq_phrase_in_evidence"
    return True, ""


# ── Country correction map for simple-mode evidence quotes ────────────────────
_S3_CITY_COUNTRY_CORRECTIONS: list[tuple[re.Pattern, str]] = [
    # Michigan / US city that looks like Netherlands
    (re.compile(r"Holland,?\s*Michigan|Holland,?\s*MI\b", re.I), "United States"),
    # Sweden cities
    (re.compile(r"\b(?:Göteborg|Gothenburg|Goeteborg)\b", re.I), "Sweden"),
    (re.compile(r"\bStockholm\b", re.I), "Sweden"),
    (re.compile(r"\bM[aä]lm[oö]\b", re.I), "Sweden"),
    # Nordic capitals
    (re.compile(r"\bHelsinki\b", re.I), "Finland"),
    (re.compile(r"\bOslo\b", re.I), "Norway"),
    (re.compile(r"\bCopenhagen|K[øo]benhavn\b", re.I), "Denmark"),
    # Italian regions / small cities that the parser mislabels
    (re.compile(r"\bNogara\b|\bVeneto\b.*\bIT\b|\bNaz-?Sciaves\b|\bTrentino(?:-Alto\s+Adige)?\b", re.I), "Italy"),
    (re.compile(r"\bSant[''`]?Agata\s+de\s+Goti\b|\bSavigno\b|\bValsamoggia\b", re.I), "Italy"),
    (re.compile(r"\b(?:Modena|Ancona|Osimo|Senigallia|San\s+Benedetto\s+del\s+Tronto|"
                r"Bergamo|Milano|Milan|Brescia|Bologna|Torino|Turin|Firenze|Florence|"
                r"Napoli|Naples|Roma|Rome|Venezia|Venice|Genova|Genoa|Padova|Padua|"
                r"Verona|Vicenza|Perugia|Trieste|Palermo|Catania|Reggio\s+Emilia|"
                r"Parma|Piacenza|Ravenna|Ferrara|Rimini|Forlì|Cesena|"
                r"Jesi|Fabriano|Civitanova\s+Marche|Avellino|Nusco)\b", re.I), "Italy"),
    # Explicit country phrases in quote
    (re.compile(r"\bheadquarters?\s+(?:in\s+)?France\b|\bFRANCE\s+HEADQUARTERS?\b|\bFrance\s+headquarters?\b", re.I), "France"),
    (re.compile(r"\bheadquarters?\s+(?:in\s+)?Germany\b|\bGermany\s+headquarters?\b", re.I), "Germany"),
    (re.compile(r"\bheadquarters?\s+(?:in\s+)?Spain\b|\bSpain\s+headquarters?\b", re.I), "Spain"),
    (re.compile(r"\bheadquarters?\s+(?:in\s+)?Netherlands\b|\bNetherlands\s+headquarters?\b", re.I), "Netherlands"),
    (re.compile(r"\bheadquarters?\s+(?:in\s+)?Switzerland\b|\bSwitzerland\s+headquarters?\b", re.I), "Switzerland"),
    (re.compile(r"\bheadquarters?\s+(?:in\s+)?United\s+Kingdom\b|\bUK\s+headquarters?\b", re.I), "United Kingdom"),
]


def _correct_country_from_quote(quote: str, detected_country: str) -> str:
    """Apply city/phrase-based country corrections to catch common parser errors."""
    for pattern, correct_country in _S3_CITY_COUNTRY_CORRECTIONS:
        if pattern.search(quote):
            return correct_country
    return detected_country


# ── Haiku HQ review helper ────────────────────────────────────────────────────
_HAIKU_MODEL = "claude-3-5-haiku-20241022"

_HAIKU_SYSTEM = (
    "You are an expert B2B data analyst specialising in corporate ownership structures. "
    "You receive evidence from a web search about a company's headquarters and must decide "
    "whether the evidence clearly shows a GENUINE GLOBAL PARENT headquartered in a specific "
    "foreign country. Answer ONLY with a JSON object — no prose, no markdown fences."
)

_HAIKU_PROMPT_TMPL = """\
Company: {company_name}
Domain: {domain}
Input country (company registered in): {input_country}
Detected country from evidence: {detected_country}
Detected city: {detected_city}
Evidence URL: {evidence_url}
Evidence quote (from Serper): {evidence_quote}
Additional context (answer box / top organic snippets): {context}

Task:
Decide if the evidence CLEARLY shows that this company has a genuine global parent
headquartered in {detected_country} (which is different from {input_country}).

Return exactly this JSON (no other text):
{{
  "verdict": "score3" | "score0" | "uncertain",
  "confidence": "High" | "Medium" | "Low",
  "detected_country_corrected": "<country name or empty string if no correction>",
  "reason": "<one sentence>"
}}

Rules:
- verdict "score3": evidence unambiguously shows a real global/international HQ in a foreign country
- verdict "score0": evidence is domestic, a branch/subsidiary, biography, directory, or irrelevant
- verdict "uncertain": evidence is ambiguous; human review needed
- detected_country_corrected: if evidence actually points to a different country than {detected_country}, put the correct country; otherwise leave empty
- Do not use web search. Judge only the evidence provided.
"""


def _review_hq_with_haiku(
    *,
    company_name: str,
    domain: str,
    input_country: str,
    detected_country: str,
    detected_city: str,
    evidence_quote: str,
    evidence_url: str,
    answer_box: str = "",
    organic_titles_snippets: list[str] | None = None,
    anthropic_key: str,
) -> dict[str, Any]:
    """Call Haiku to review uncertain HQ evidence. Returns a result dict with keys:
    haiku_verdict, haiku_confidence, haiku_country_corrected, haiku_reason,
    haiku_input_tokens, haiku_output_tokens, haiku_error.
    Never raises — on any error returns haiku_error with details.
    """
    out: dict[str, Any] = {
        "haiku_verdict": "",
        "haiku_confidence": "",
        "haiku_country_corrected": "",
        "haiku_reason": "",
        "haiku_input_tokens": 0,
        "haiku_output_tokens": 0,
        "haiku_error": "",
    }
    if not anthropic_key:
        out["haiku_error"] = "no_api_key"
        return out

    context_parts: list[str] = []
    if answer_box:
        context_parts.append(f"[answerBox] {answer_box[:300]}")
    for snip in (organic_titles_snippets or [])[:5]:
        context_parts.append(snip[:200])
    context = " | ".join(context_parts) if context_parts else "(none)"

    prompt = _HAIKU_PROMPT_TMPL.format(
        company_name=company_name or "?",
        domain=domain or "?",
        input_country=input_country or "?",
        detected_country=detected_country or "?",
        detected_city=detected_city or "?",
        evidence_url=evidence_url or "",
        evidence_quote=(evidence_quote or "")[:400],
        context=context,
    )

    try:
        import anthropic as _anthropic_mod
        _client = _anthropic_mod.Anthropic(api_key=anthropic_key)
        _msg = _client.messages.create(
            model=_HAIKU_MODEL,
            max_tokens=256,
            system=_HAIKU_SYSTEM,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = (_msg.content[0].text or "").strip()
        out["haiku_input_tokens"]  = _msg.usage.input_tokens
        out["haiku_output_tokens"] = _msg.usage.output_tokens
        # Parse JSON response
        import json as _json_mod
        # Strip markdown fences if present
        _clean = re.sub(r"^```[a-z]*\n?|```$", "", raw, flags=re.M).strip()
        _parsed = _json_mod.loads(_clean)
        out["haiku_verdict"]            = str(_parsed.get("verdict", "")).strip()
        out["haiku_confidence"]         = str(_parsed.get("confidence", "")).strip()
        out["haiku_country_corrected"]  = str(_parsed.get("detected_country_corrected", "")).strip()
        out["haiku_reason"]             = str(_parsed.get("reason", "")).strip()
    except Exception as _he:
        out["haiku_error"] = str(_he)[:200]

    return out


# Text markers that indicate we've entered a "locations list" section
_LOCATION_SECTION_RE = re.compile(
    r"\b(?:locations?|get\s+directions?|other\s+offices?|all\s+offices?|"
    r"offices?\s+worldwide|global\s+locations?|primary\s+location)\b",
    re.IGNORECASE,
)


def _scan_hq_title_snippet(organic: list[dict]) -> tuple[str, str, str, str]:
    """Scan organic result titles/snippets for HQ evidence without requiring
    an explicit 'in/at' preposition.

    Handles cases like:
      title="Bodycote Macclesfield Head Office"
      snippet="Macclesfield, Cheshire, United Kingdom"

    Proximity fix: only look at text within ~80 chars AFTER the HQ signal word,
    stopping at location-section boundary phrases (e.g. "Locations", "Get directions").
    Regional HQ guard: skip results whose only HQ signal is regional (NA HQ, EMEA HQ).

    Returns (quote, country, city, strength).  Strength is always "Medium".
    """
    for item in organic[:5]:
        title   = item.get("title", "")
        snippet = item.get("snippet", "")
        combined = f"{title} {snippet}"

        sig_m = _HQ_SIGNAL_RE.search(combined)
        if not sig_m:
            continue

        # Skip purely regional HQ signals (e.g. "NA Headquarters Dallas")
        if _REGIONAL_HQ_RE.search(combined) and not _HQ_SIGNAL_RE.search(
            re.sub(_REGIONAL_HQ_RE.pattern, "", combined, flags=re.IGNORECASE)
        ):
            continue

        # Extract only the text within ~80 chars after the HQ signal, stopping
        # at a location-section boundary so we don't read a second city.
        sig_end = sig_m.end()
        window  = combined[sig_end: sig_end + 80]
        # Truncate at location boundary
        loc_boundary = _LOCATION_SECTION_RE.search(window)
        if loc_boundary:
            window = window[: loc_boundary.start()]

        # Also try text BEFORE the signal (for "Bodycote Macclesfield Head Office")
        sig_start  = sig_m.start()
        pre_window = combined[max(0, sig_start - 60): sig_start]

        # Try post-signal window first, then pre-signal
        for chunk in (window, pre_window):
            chunk = chunk.strip(" ,-–")
            if not chunk:
                continue
            city, country = _resolve_city_country(chunk)
            if country or city:
                quote = combined[:300].strip()
                return quote, country, city, "Medium"

    return "", "", "", ""


# Known global brands: brand_root_token (lowercase) → type
# "corporate"      = single parent company with a clear HQ country
# "global_network" = federated professional-services network of member firms
_KNOWN_GLOBAL_BRANDS: dict[str, str] = {
    # Corporate
    "ibm": "corporate", "microsoft": "corporate", "oracle": "corporate",
    "sap": "corporate", "siemens": "corporate", "bosch": "corporate",
    "fresenius": "corporate", "hp": "corporate", "hpe": "corporate",
    "cisco": "corporate", "ntt": "corporate", "fujitsu": "corporate",
    "accenture": "corporate", "capgemini": "corporate",
    # Global professional-services networks
    "kpmg": "global_network", "ey": "global_network", "pwc": "global_network",
    "deloitte": "global_network", "bdo": "global_network",
    "grant thornton": "global_network",
}

# Map brand root → canonical display name (for audit col)
_KNOWN_GLOBAL_BRAND_NAMES: dict[str, str] = {
    "ibm": "IBM", "microsoft": "Microsoft", "oracle": "Oracle",
    "sap": "SAP", "siemens": "Siemens", "bosch": "Bosch",
    "fresenius": "Fresenius", "hp": "HP", "hpe": "HPE",
    "cisco": "Cisco", "ntt": "NTT", "fujitsu": "Fujitsu",
    "accenture": "Accenture", "capgemini": "Capgemini",
    "kpmg": "KPMG", "ey": "EY", "pwc": "PwC",
    "deloitte": "Deloitte", "bdo": "BDO",
    "grant thornton": "Grant Thornton",
}

# Domain roots of global professional-services networks that get an extra
# "global network" follow-up query after the domain-root HQ search.
_GLOBAL_NETWORK_DOMAIN_ROOTS: frozenset[str] = frozenset({
    "kpmg", "ey", "pwc", "deloitte", "bdo", "grantthornton",
})


def _match_known_global_brand(brand_roots: list[str]) -> tuple[str, str, str]:
    """Return (brand_key, brand_type, canonical_name) or ('','','') if no match."""
    for root in brand_roots:
        key = root.lower().strip()
        if key in _KNOWN_GLOBAL_BRANDS:
            return key, _KNOWN_GLOBAL_BRANDS[key], _KNOWN_GLOBAL_BRAND_NAMES.get(key, root)
    return "", "", ""


def _domain_root_hq_query(domain: str) -> tuple[str, str]:
    """Return (domain_root, query) for a simple un-quoted domain-root HQ search.

    Examples:
        'datwyler.com'      → ('datwyler', 'datwyler headquarters')
        'www.ibm.com'       → ('ibm', 'ibm headquarters')
        'jobtech.it'        → ('jobtech', 'jobtech headquarters')
        'csmingredients.com'→ ('csmingredients', 'csmingredients headquarters')
    """
    root = _get_domain_body(domain) if domain else ""
    if not root:
        return "", ""
    return root, f"{root} headquarters"


def _mimic_google_hq_check(
    company_name: str,
    domain: str,
    input_country: str,
    serper_key: str,
    cache: dict,
    max_page_paths: int = 7,
    max_brand_queries: int = 4,
) -> dict[str, Any]:
    """Run manual Google mimic HQ check.

    Returns audit column dict plus internal keys:
        _mimic_serper_calls (int)
        _mimic_queries      (list[str])
        _mimic_serper_cache_hits (int)
        _mimic_hq_evidence  (dict)  — empty if nothing found
    """
    out: dict[str, Any] = {
        "manual_google_mimic_used":           True,
        "plain_company_search_query":         "",
        "plain_company_top_result_url":       "",
        "plain_company_top_result_title":     "",
        "plain_company_top_result_snippet":   "",
        "plain_company_official_domain":      "",
        "plain_company_official_result_rank": 0,
        "official_domain_from_plain_search":  "",
        "official_domain_matches_input_domain": False,
        "official_page_fetch_used":           False,
        "official_page_fetch_url":            "",
        "official_page_fetch_count":          0,
        "official_page_fetch_error":          "",
        "official_page_hq_evidence_found":    False,
        "official_page_hq_evidence_quote":    "",
        "official_page_hq_country":           "",
        "official_page_hq_city":              "",
        "official_page_hq_evidence_strength": "",
        "brand_hq_search_used":               False,
        "brand_hq_search_queries":            "",
        "brand_hq_top_result_url":            "",
        "brand_hq_top_result_snippet":        "",
        "brand_hq_evidence_found":            False,
        "brand_hq_evidence_quote":            "",
        "brand_hq_top_result_domain":         "",
        "brand_hq_top_result_is_official_domain": False,
        "brand_hq_result_selection_reason":   "",
        # Local-country page detection
        "official_top_result_is_local_country_page": False,
        "official_top_result_country_context":        "",
        "official_top_result_local_page_reason":      "",
        "early_stop_blocked_reason":                  "",
        # Brand-root HQ check
        "brand_root_candidates":        "",
        "brand_root_hq_search_used":    False,
        "brand_root_hq_search_queries": "",
        "brand_root_hq_evidence_found": False,
        "brand_root_hq_evidence_quote": "",
        "brand_root_hq_evidence_url":   "",
        "brand_root_hq_country":        "",
        "brand_root_hq_city":           "",
        "brand_root_hq_source_type":    "",
        # Known-global-brand detection
        "known_global_brand_detected":  False,
        "known_global_brand_name":      "",
        "known_global_brand_type":      "",
        "known_global_brand_reason":    "",
        # Domain-root HQ search
        "domain_root":                  "",
        "domain_root_hq_query":         "",
        "domain_root_hq_search_used":   False,
        "domain_root_hq_evidence_found": False,
        "domain_root_hq_evidence_quote": "",
        "domain_root_hq_evidence_url":   "",
        "domain_root_hq_country":        "",
        "domain_root_hq_city":           "",
        "domain_root_hq_confidence":     "",
        # Internal
        "_mimic_serper_calls":      0,
        "_mimic_queries":           [],
        "_mimic_serper_cache_hits": 0,
        "_mimic_hq_evidence":       {},
    }

    # ── 0. Domain-root HQ search (very first, simplest query) ────────────
    # Always runs first when a domain is provided. Query: "{domain_root} headquarters"
    # Not quoted, no company name, no legal suffix, no Italy.
    _dr_root, _dr_query = _domain_root_hq_query(domain)
    if _dr_root and _dr_query:
        out["domain_root"]            = _dr_root
        out["domain_root_hq_query"]   = _dr_query
        out["domain_root_hq_search_used"] = True
        out["_mimic_queries"].append(_dr_query)

        _dr_ck = ("serper", _dr_query)
        if _dr_ck in cache:
            _dr_data, _ = cache[_dr_ck]
            out["_mimic_serper_cache_hits"] += 1
        else:
            _dr_data, _ = _serper_search(_dr_query, serper_key)
            cache[_dr_ck] = (_dr_data, "")
            out["_mimic_serper_calls"] += 1
            time.sleep(_INTER_REQUEST_SLEEP)

        if _dr_data:
            _dr_org    = _dr_data.get("organic", [])
            _dr_kg     = _dr_data.get("knowledgeGraph", {})
            _dr_kg_loc = (_dr_kg.get("address", "") or _dr_kg.get("headquarters", "")
                          or _dr_kg.get("location", ""))
            _dr_ab     = _dr_data.get("answerBox", {})
            _dr_ab_t   = _dr_ab.get("answer", "") or _dr_ab.get("snippet", "")
            _dr_all    = " ".join([
                _dr_kg_loc, _dr_ab_t,
                *[f"{r.get('title','')} {r.get('snippet','')}" for r in _dr_org[:5]],
            ])
            _dr_q, _dr_country, _dr_city, _dr_strength = _scan_for_hq(_dr_all)
            if _dr_country or _dr_city:
                out["domain_root_hq_evidence_found"] = True
                out["domain_root_hq_evidence_quote"] = _dr_q
                out["domain_root_hq_country"]        = _dr_country
                out["domain_root_hq_city"]           = _dr_city
                out["domain_root_hq_confidence"]     = (
                    "High" if _dr_strength == "Strong" else
                    ("Medium" if _dr_strength == "Medium" else "Low")
                )
                # Pick best evidence URL
                from urllib.parse import urlparse as _up0
                _dr_input_nl = re.sub(r"^https?://", "", (domain or "").strip()).rstrip("/")
                _dr_input_nl = re.sub(r"^www\.", "", _dr_input_nl).lower()
                def _dr_url_rank(r: dict) -> int:
                    nl = re.sub(r"^www\.", "", _up0(r.get("link","")).netloc.lower())
                    if _is_directory_url(r.get("link","")):
                        return 10
                    if _dr_input_nl and (nl == _dr_input_nl or nl.endswith("." + _dr_input_nl)):
                        return 0
                    return 5
                _dr_top_r = sorted(_dr_org, key=_dr_url_rank)[0] if _dr_org else {}
                out["domain_root_hq_evidence_url"] = _dr_top_r.get("link", "") if _dr_top_r else ""

                # Populate _mimic_hq_evidence for strong/medium evidence
                _input_std_dr = _std_country(input_country or "Italy")
                _dr_country_std = _std_country(_dr_country)
                if _dr_strength in ("Strong", "Medium") and _dr_country_std.lower() != _input_std_dr.lower():
                    out["_mimic_hq_evidence"] = {
                        "hq_detected_city":    _dr_city,
                        "hq_detected_country": _dr_country,
                        "hq_confidence":       "High" if _dr_strength == "Strong" else "Medium",
                        "hq_reason":           f"[domain-root-hq:{_dr_query}] {_dr_q[:150]}",
                        "hq_evidence_url":     out["domain_root_hq_evidence_url"],
                        "hq_evidence_quote":   _dr_q,
                    }

        # Global-network follow-up: one extra query if domain root is a known network
        if _dr_root.lower() in _GLOBAL_NETWORK_DOMAIN_ROOTS and not out["_mimic_hq_evidence"]:
            _gn_query = f"{_dr_root} global network"
            out["_mimic_queries"].append(_gn_query)
            _gn_ck = ("serper", _gn_query)
            if _gn_ck in cache:
                _gn_data, _ = cache[_gn_ck]
                out["_mimic_serper_cache_hits"] += 1
            else:
                _gn_data, _ = _serper_search(_gn_query, serper_key)
                cache[_gn_ck] = (_gn_data, "")
                out["_mimic_serper_calls"] += 1
                time.sleep(_INTER_REQUEST_SLEEP)
            # We don't need to parse this result — the KGB detection in step 4 will
            # classify the brand as global_network. The query is logged for audit only.

    # ── 1. Plain company name search ──────────────────────────────────────
    # Skip if domain-root search already produced strong foreign-HQ evidence.
    _dr_found_foreign = bool(out.get("_mimic_hq_evidence"))
    plain_query = company_name.strip()
    out["plain_company_search_query"] = plain_query
    out["_mimic_queries"].append(plain_query)

    plain_data: dict = {}
    if not _dr_found_foreign:
        ck = ("serper", plain_query)
        if ck in cache:
            plain_data, _plain_err = cache[ck]
            out["_mimic_serper_cache_hits"] += 1
        else:
            plain_data, _plain_err = _serper_search(plain_query, serper_key)
            cache[ck] = (plain_data, "")
            out["_mimic_serper_calls"] += 1
            time.sleep(_INTER_REQUEST_SLEEP)

    plain_organic = plain_data.get("organic", []) if plain_data else []

    # ── 2. Find official result ────────────────────────────────────────────
    rank, off_result, off_netloc = _find_official_result(plain_organic, domain)
    if off_result:
        out["plain_company_top_result_url"]       = off_result.get("link", "")
        out["plain_company_top_result_title"]     = off_result.get("title", "")
        out["plain_company_top_result_snippet"]   = off_result.get("snippet", "")
        out["plain_company_official_domain"]      = off_netloc
        out["plain_company_official_result_rank"] = rank
        out["official_domain_from_plain_search"]  = off_netloc
        input_body  = _get_domain_body(domain) if domain else ""
        off_body    = _get_domain_body(off_netloc)
        out["official_domain_matches_input_domain"] = bool(
            input_body and (input_body in off_body or off_body in input_body)
        )
        # Detect if the top official result is a local-country page
        _lc_is, _lc_ctx, _lc_why = _is_local_country_page(
            off_result.get("link", ""),
            off_result.get("title", ""),
            off_result.get("snippet", ""),
            country_keyword=input_country or "italy",
        )
        out["official_top_result_is_local_country_page"] = _lc_is
        out["official_top_result_country_context"]       = _lc_ctx
        out["official_top_result_local_page_reason"]     = _lc_why

    # Pre-compute brand-root candidates (used in both fetch scan and brand-root search)
    _brand_roots = _extract_brand_roots(company_name, domain or off_netloc)
    out["brand_root_candidates"] = " | ".join(_brand_roots)

    # Known-global-brand detection
    _kgb_key, _kgb_type, _kgb_name = _match_known_global_brand(_brand_roots)
    if _kgb_key:
        out["known_global_brand_detected"] = True
        out["known_global_brand_name"]     = _kgb_name
        out["known_global_brand_type"]     = _kgb_type
        out["known_global_brand_reason"]   = (
            f"Brand root '{_kgb_name}' matched known global brand registry (type={_kgb_type})"
        )
        # Ensure queries use canonical brand name (e.g. 'IBM' not 'Ibm')
        _brand_roots = [
            _kgb_name if r.lower() == _kgb_key else r for r in _brand_roots
        ]
        out["brand_root_candidates"] = " | ".join(_brand_roots)

    # ── 3. Fetch official pages ────────────────────────────────────────────
    fetch_domain = off_netloc or (
        re.sub(r"^https?://", "", domain.strip()).rstrip("/") if domain else ""
    )
    combined_page_text = ""
    first_ok_url = ""
    fetch_count  = 0
    fetch_errors: list[str] = []

    if fetch_domain and not _dr_found_foreign:
        out["official_page_fetch_used"] = True
        for path in _OFFICIAL_PAGE_PATHS[:max_page_paths]:
            url_try = f"https://{fetch_domain}{path}"
            text, err = _fetch_page_text(url_try, cache)
            fetch_count += 1
            if text:
                if not first_ok_url:
                    first_ok_url = url_try
                # Early-exit scan for Strong evidence
                q, country, city, strength = _scan_for_hq(text)
                if strength == "Strong" and country:
                    out["official_page_fetch_url"]           = url_try
                    out["official_page_fetch_count"]         = fetch_count
                    out["official_page_hq_evidence_found"]   = True
                    out["official_page_hq_evidence_quote"]   = q
                    out["official_page_hq_country"]          = country
                    out["official_page_hq_city"]             = city
                    out["official_page_hq_evidence_strength"]= strength
                    out["_mimic_hq_evidence"] = {
                        "hq_detected_city":    city,
                        "hq_detected_country": country,
                        "hq_confidence":       "High",
                        "hq_reason":           f"[mimic-official-page:{path}] {q[:150]}",
                        "hq_evidence_url":     url_try,
                        "hq_evidence_quote":   q,
                    }
                    break
                combined_page_text += " " + text
            else:
                fetch_errors.append(f"{path}: {err}")

        out["official_page_fetch_url"]   = out["official_page_fetch_url"] or first_ok_url
        out["official_page_fetch_count"] = fetch_count
        out["official_page_fetch_error"] = "; ".join(fetch_errors[:3])

    # Fallback: scan combined text if single-page scan missed
    if combined_page_text and not out["official_page_hq_evidence_found"]:
        q, country, city, strength = _scan_for_hq(combined_page_text)
        if country or city:
            out["official_page_hq_evidence_found"]    = True
            out["official_page_hq_evidence_quote"]    = q
            out["official_page_hq_country"]           = country
            out["official_page_hq_city"]              = city
            out["official_page_hq_evidence_strength"] = strength
            if strength in ("Strong", "Medium") and not out["_mimic_hq_evidence"]:
                out["_mimic_hq_evidence"] = {
                    "hq_detected_city":    city,
                    "hq_detected_country": country,
                    "hq_confidence":       "High" if strength == "Strong" else "Medium",
                    "hq_reason":           f"[mimic-official-page-combined] {q[:150]}",
                    "hq_evidence_url":     out["official_page_fetch_url"],
                    "hq_evidence_quote":   q,
                }

    # If we found evidence from official page but the top result was a local-country page,
    # block early-stop and force brand-root check to validate parent HQ.
    _is_local_top = out["official_top_result_is_local_country_page"]
    if _is_local_top and out["_mimic_hq_evidence"]:
        # Only Italian local entity evidence — not enough to conclude parent/group HQ
        out["early_stop_blocked_reason"] = "local_country_page_requires_brand_root_hq_check"
        out["_mimic_hq_evidence"] = {}   # clear; brand-root check will decide

    # ── 4. Brand/domain HQ follow-up (only if no strong evidence yet) ──────
    # Also mandatory when local-country page was detected, even if some page evidence exists
    _need_brand_root = not out["_mimic_hq_evidence"] or _is_local_top
    if _need_brand_root:
        # Use brand roots (specific → broad) for brand-root HQ check
        brand_root_queries: list[str] = []
        if _brand_roots:
            primary = _brand_roots[0]
            brand_root_queries = [
                f'"{primary}" headquarters',
                f'"{primary}" group headquarters',
                f'"{primary}" head office',
                f'"{primary}" corporate headquarters',
                f'"{primary}" parent company',
            ]
        # Limit per mode; caller controls max_brand_queries
        brand_root_queries = brand_root_queries[:max_brand_queries]

        # ── 4a. Brand-root HQ search ──
        if brand_root_queries:
            out["brand_root_hq_search_used"]    = True
            out["brand_root_hq_search_queries"] = " | ".join(brand_root_queries)
            for bq in brand_root_queries:
                out["_mimic_queries"].append(bq)
                ck_b = ("serper", bq)
                if ck_b in cache:
                    bdata, _ = cache[ck_b]
                    out["_mimic_serper_cache_hits"] += 1
                else:
                    bdata, _ = _serper_search(bq, serper_key)
                    cache[ck_b] = (bdata, "")
                    out["_mimic_serper_calls"] += 1
                    time.sleep(_INTER_REQUEST_SLEEP)
                if not bdata:
                    continue
                b_org    = bdata.get("organic", [])
                b_kg     = bdata.get("knowledgeGraph", {})
                b_kg_loc = (b_kg.get("address", "") or b_kg.get("headquarters", "")
                            or b_kg.get("location", ""))
                b_ab     = bdata.get("answerBox", {})
                b_ab_t   = b_ab.get("answer", "") or b_ab.get("snippet", "")
                # Scan KG + answer box + top organic
                all_b = " ".join([
                    b_kg_loc, b_ab_t,
                    *[f"{r.get('title','')} {r.get('snippet','')}" for r in b_org[:5]],
                ])
                q, country, city, strength = _scan_for_hq(all_b)
                if country or city:
                    out["brand_root_hq_evidence_found"] = True
                    out["brand_root_hq_evidence_quote"] = q
                    out["brand_root_hq_country"]        = country
                    out["brand_root_hq_city"]           = city
                    # Determine source type
                    if b_kg_loc and (country.lower() in b_kg_loc.lower() or city.lower() in b_kg_loc.lower()):
                        src_type = "knowledge_graph"
                    elif b_ab_t and (country.lower() in b_ab_t.lower() or city.lower() in b_ab_t.lower()):
                        src_type = "answer_box"
                    else:
                        src_type = "organic_snippet"
                    out["brand_root_hq_source_type"] = src_type
                    # Pick best URL (prefer official/input domain)
                    from urllib.parse import urlparse as _up2
                    _official_nl = (fetch_domain or domain or "").lstrip("www.").lower()
                    def _br_rank(r: dict) -> int:
                        nl = _up2(r.get("link", "")).netloc.lstrip("www.").lower()
                        if _is_directory_url(r.get("link", "")):
                            return 10
                        if _official_nl and (nl == _official_nl or nl.endswith("." + _official_nl)):
                            return 0
                        return 5
                    top_r = sorted(b_org, key=_br_rank)[0] if b_org else {}
                    out["brand_root_hq_evidence_url"] = top_r.get("link", "") if top_r else ""
                    if strength in ("Strong", "Medium") and not out["_mimic_hq_evidence"]:
                        out["_mimic_hq_evidence"] = {
                            "hq_detected_city":    city,
                            "hq_detected_country": country,
                            "hq_confidence":       "High" if strength == "Strong" else "Medium",
                            "hq_reason":           f"[mimic-brand-root:{bq[:60]}] {q[:150]}",
                            "hq_evidence_url":     out["brand_root_hq_evidence_url"],
                            "hq_evidence_quote":   q,
                        }
                    break  # stop after first successful brand-root query

        # ── 4b. Original brand follow-up (fallback when brand-root produced nothing) ──
        brands = _extract_brand(company_name, fetch_domain or domain)
        brand_queries: list[str] = []
        if brands and not out["brand_root_hq_evidence_found"]:
            primary = brands[0]
            brand_queries = [
                f'"{primary}" headquarters',
                f'"{primary}" head office',
                f'"{primary}" group headquarters',
                f'"{primary}" parent company',
            ]
        brand_queries = brand_queries[:max_brand_queries]
        if brand_queries:
            out["brand_hq_search_used"]    = True
            out["brand_hq_search_queries"] = " | ".join(brand_queries)
            for bq in brand_queries:
                out["_mimic_queries"].append(bq)
                ck_b = ("serper", bq)
                if ck_b in cache:
                    bdata, _ = cache[ck_b]
                    out["_mimic_serper_cache_hits"] += 1
                else:
                    bdata, _ = _serper_search(bq, serper_key)
                    cache[ck_b] = (bdata, "")
                    out["_mimic_serper_calls"] += 1
                    time.sleep(_INTER_REQUEST_SLEEP)
                if not bdata:
                    continue
                b_org = bdata.get("organic", [])
                b_kg  = bdata.get("knowledgeGraph", {})
                b_kg_loc = b_kg.get("address","") or b_kg.get("headquarters","") or b_kg.get("location","")
                b_ab  = bdata.get("answerBox", {})
                b_ab_t = b_ab.get("answer","") or b_ab.get("snippet","")
                all_b = " ".join([
                    b_kg_loc, b_ab_t,
                    *[f"{r.get('title','')} {r.get('snippet','')}" for r in b_org[:3]],
                ])
                q, country, city, strength = _scan_for_hq(all_b)
                if country or city:
                    out["brand_hq_evidence_found"]   = True
                    out["brand_hq_evidence_quote"]   = q
                    # Prefer official/input domain result over directories or unrelated domains
                    from urllib.parse import urlparse as _up
                    _official_netloc = (fetch_domain or domain or "").lstrip("www.").lower()
                    _plain_official  = out.get("plain_company_official_domain", "").lstrip("www.").lower()
                    def _brand_rank(r: dict) -> int:
                        lnk = r.get("link", "")
                        nl  = _up(lnk).netloc.lstrip("www.").lower()
                        if _is_directory_url(lnk):
                            return 10
                        if _official_netloc and (nl == _official_netloc or nl.endswith("." + _official_netloc)):
                            return 0
                        if _plain_official and (nl == _plain_official or nl.endswith("." + _plain_official)):
                            return 1
                        return 5
                    sorted_org = sorted(b_org, key=_brand_rank) if b_org else []
                    top_r = sorted_org[0] if sorted_org else {}
                    top_nl = _up(top_r.get("link", "")).netloc.lstrip("www.").lower() if top_r else ""
                    _is_off = bool(_official_netloc and (top_nl == _official_netloc or top_nl.endswith("." + _official_netloc)))
                    _sel_reason = (
                        "input_domain_match" if _is_off
                        else ("plain_official_match" if _plain_official and top_nl == _plain_official else "best_non_directory")
                    )
                    out["brand_hq_top_result_url"]     = top_r.get("link", "")
                    out["brand_hq_top_result_snippet"] = top_r.get("snippet", "")
                    out["brand_hq_top_result_domain"]              = top_nl
                    out["brand_hq_top_result_is_official_domain"]  = _is_off
                    out["brand_hq_result_selection_reason"]        = _sel_reason
                    if strength in ("Strong", "Medium") and not out["_mimic_hq_evidence"]:
                        out["_mimic_hq_evidence"] = {
                            "hq_detected_city":    city,
                            "hq_detected_country": country,
                            "hq_confidence":       "High" if strength == "Strong" else "Medium",
                            "hq_reason":           f"[mimic-brand-search:{bq[:60]}] {q[:150]}",
                            "hq_evidence_url":     out["brand_hq_top_result_url"],
                            "hq_evidence_quote":   q,
                        }
                    break  # stop after first brand query with evidence

    return out


def _build_queries(company_name: str, domain: str) -> list[str]:
    n = company_name.strip()
    d = re.sub(r"^https?://", "", (domain or "").strip()).rstrip("/")
    queries = [
        f'"{n}" headquarters',
        f'"{n}" head office',
        f'"{n}" sede legale',
        f'"{n}" sede principale',
        f'"{n}" sede amministrativa',
    ]
    if d:
        queries += [f"site:{d} sede", f"site:{d} headquarters", f"site:{d} head office"]
    return queries

# ---------------------------------------------------------------------------
# Per-company probe
# ---------------------------------------------------------------------------

def probe_company(
    company_name: str,
    domain: str,
    input_country: str,
    serper_key: str,
    use_model: bool,
    anthropic_key: str,
    cache: dict,
    input_row: "dict | None" = None,
    use_multilingual_check: bool = True,
    use_anthropic_review: bool = False,
    use_mimic_check: bool = True,
    run_mode: str = "fast",
    use_simple_hq_mode: bool = True,
    use_haiku_uncertain: bool = False,
    haiku_calls_counter: "list[int] | None" = None,
    haiku_max_calls: int = 25,
) -> dict[str, Any]:
    """Run all queries for one company; return probe column dict."""
    import time as _time_mod
    _row_start = _time_mod.monotonic()

    # ── Mode-based limits ───────────────────────────────────────────────────
    # fast: ≤3 real Serper calls, 1 page path, 2 brand queries, 5s fetch timeout
    # deep/debug: ≤8 Serper calls, 7 page paths, 4 brand queries
    if run_mode == "fast":
        _max_serper        = 3
        _max_page_paths    = 1   # only homepage in fast mode
        _max_brand_queries = 2
        _fetch_timeout     = 5
    else:
        _max_serper        = 8
        _max_page_paths    = 7
        _max_brand_queries = 4
        _fetch_timeout     = 10

    result: dict[str, Any] = {col: "" for col in PROBE_COLS}
    result["run_mode"]               = run_mode
    result["max_serper_calls_per_row"] = _max_serper
    result["early_stop_used"]        = False
    result["early_stop_reason"]      = ""

    if not company_name.strip():
        result["probe_error"] = "blank company name"
        return result

    # ── Simple HQ mode (default) ─────────────────────────────────────────────
    # One query only: "{domain_root} headquarters" (or "{company_name} headquarters"
    # when domain is absent).  No mimic, no brand-root ladder, no Anthropic.
    if use_simple_hq_mode:
        result["simple_hq_mode_used"] = True
        result["max_serper_calls_per_row"] = 1

        if domain:
            _dr_root, _dr_query = _domain_root_hq_query(domain)
        else:
            _dr_root, _dr_query = "", f"{company_name.strip()} headquarters"

        result["domain_root"]             = _dr_root
        result["simple_hq_query"]         = _dr_query
        result["domain_root_hq_query"]    = _dr_query
        result["domain_root_hq_search_used"] = True

        _s_calls: int = 0
        _s_cache_hit: bool = False
        _s_queries: list[str] = [_dr_query] if _dr_query else []
        _s_data: dict = {}
        _s_err: str = ""

        if _dr_query and serper_key:
            _s_ck = ("serper", _dr_query)
            if _s_ck in cache:
                _s_data, _ = cache[_s_ck]
                _s_cache_hit = True
            else:
                _s_data, _s_err = _serper_search(_dr_query, serper_key)
                cache[_s_ck] = (_s_data, _s_err)
                _s_calls = 1
                time.sleep(_INTER_REQUEST_SLEEP)

        if _s_data:
            _s_org    = _s_data.get("organic", [])
            _s_kg     = _s_data.get("knowledgeGraph", {})
            _s_kg_loc = (_s_kg.get("address", "") or _s_kg.get("headquarters", "")
                         or _s_kg.get("location", ""))
            _s_ab     = _s_data.get("answerBox", {})
            _s_ab_t   = _s_ab.get("answer", "") or _s_ab.get("snippet", "")
            _s_places = _s_data.get("places", []) or _s_data.get("local", [])

            result["serper_knowledge_graph_location"] = _s_kg_loc
            result["serper_answer_box"]               = _s_ab_t
            for _si, _sitem in enumerate(_s_org[:3], start=1):
                result[f"top_organic_title_{_si}"]   = _sitem.get("title", "")
                result[f"top_organic_snippet_{_si}"] = _sitem.get("snippet", "")
                result[f"top_organic_url_{_si}"]     = _sitem.get("link", "")

            # Per-source scan in priority order — first hit wins.
            # Priority: KG > answerBox > organic[0] > organic[1] > … > places
            # This avoids concatenating all text (which causes Italian cities to
            # match before a later-occurring foreign city).
            _s_quote = _s_country = _s_city = _s_strength = ""
            _s_evidence_rank = ""
            _s_rejected_reasons: list[str] = []

            def _try_source(text: str, label: str) -> bool:
                nonlocal _s_quote, _s_country, _s_city, _s_strength, _s_evidence_rank
                q, co, ci, st = _scan_for_hq(text)
                if co or ci:
                    _s_quote, _s_country, _s_city, _s_strength = q, co, ci, st
                    _s_evidence_rank = label
                    return True
                return False

            if _s_kg_loc and not _try_source(_s_kg_loc, "knowledge_graph"):
                _s_rejected_reasons.append(f"kg_loc='{_s_kg_loc[:80]}' no match")
            if _s_ab_t and not (_s_country or _s_city) and not _try_source(_s_ab_t, "answer_box"):
                _s_rejected_reasons.append(f"answer_box='{_s_ab_t[:80]}' no match")

            for _s_rank_i, _s_r in enumerate(_s_org[:5], start=1):
                if _s_country or _s_city:
                    break
                _s_r_text = f"{_s_r.get('title','')} {_s_r.get('snippet','')}"
                if not _try_source(_s_r_text, f"organic_{_s_rank_i}"):
                    # Try proximity/signal scan as fallback for this single result
                    _fb_q, _fb_co, _fb_ci, _fb_st = _scan_hq_title_snippet([_s_r])
                    if _fb_co or _fb_ci:
                        _s_quote, _s_country, _s_city, _s_strength = _fb_q, _fb_co, _fb_ci, _fb_st
                        _s_evidence_rank = f"organic_{_s_rank_i}_signal"
                    else:
                        _s_rejected_reasons.append(f"organic_{_s_rank_i}='{_s_r_text[:60]}' no match")

            if not (_s_country or _s_city):
                for _s_rank_p, _s_p in enumerate(_s_places[:3], start=1):
                    _s_p_text = f"{_s_p.get('title','')} {_s_p.get('address','')}"
                    if _try_source(_s_p_text, f"places_{_s_rank_p}"):
                        break

            result["domain_root_hq_evidence_rank"]            = _s_evidence_rank
            result["domain_root_hq_rejected_evidence_reason"] = "; ".join(_s_rejected_reasons[:5])

            if _s_country or _s_city:
                _s_country_std = _std_country(_s_country)
                _s_input_std   = _std_country(input_country or "Italy")

                # Apply city/phrase-based country corrections before deciding foreign
                _s_country_corrected = _correct_country_from_quote(_s_quote, _s_country_std)
                if _s_country_corrected != _s_country_std:
                    _s_country_std = _s_country_corrected

                _s_is_foreign  = bool(_s_country_std and
                                      _s_country_std.lower() != _s_input_std.lower())
                _s_conf = ("High"   if _s_strength == "Strong"
                           else ("Medium" if _s_strength == "Medium" else "Low"))

                result["hq_detected_country"]          = _s_country_std
                result["hq_detected_city"]             = _s_city
                result["hq_confidence"]                = _s_conf
                result["hq_evidence_quote"]            = _s_quote
                result["hq_reason"]                    = f"[simple-hq:{_dr_query}] {_s_quote[:150]}"
                result["domain_root_hq_evidence_found"]  = True
                result["domain_root_hq_evidence_quote"]  = _s_quote
                result["domain_root_hq_country"]         = _s_country_std
                result["domain_root_hq_city"]            = _s_city
                result["domain_root_hq_confidence"]      = _s_conf

                # Best evidence URL: prefer official domain, avoid directories
                from urllib.parse import urlparse as _up_s
                _s_input_nl = re.sub(
                    r"^www\.", "",
                    re.sub(r"^https?://", "", (domain or "").strip()).rstrip("/").lower()
                )
                _s_best_url = ""
                for _sr in _s_org:
                    _sl = _sr.get("link", "")
                    if _is_directory_url(_sl):
                        continue
                    _snl = re.sub(r"^www\.", "", _up_s(_sl).netloc.lower())
                    if _s_input_nl and (_snl == _s_input_nl or _snl.endswith("." + _s_input_nl)):
                        _s_best_url = _sl
                        break
                if not _s_best_url:
                    _s_best_url = next(
                        (r.get("link", "") for r in _s_org if not _is_directory_url(r.get("link", ""))),
                        "",
                    )
                result["hq_evidence_url"]             = _s_best_url
                result["domain_root_hq_evidence_url"] = _s_best_url

                # ── Evidence quality checks ──────────────────────────────────────
                _s_evidence_domain = (
                    re.sub(r"^www\.", "", _up_s(_s_best_url).netloc.lower())
                    if _s_best_url else ""
                )
                _s_is_official_domain = bool(
                    _s_evidence_domain and _s_input_nl
                    and (_s_evidence_domain == _s_input_nl
                         or _s_evidence_domain.endswith("." + _s_input_nl))
                )
                _s_is_known_brand = bool(_dr_root and _dr_root in _KNOWN_GLOBAL_BRANDS)

                # Domain mismatch: best URL is from a different, non-directory domain
                _s_domain_mismatch = bool(
                    _s_evidence_domain and _s_input_nl
                    and not _s_is_official_domain
                    and not _is_directory_url(_s_best_url)
                    and not _s_is_known_brand
                )

                # Subsidiary/regional signal in evidence quote
                _s_is_subsidiary_regional = bool(_SUBSIDIARY_SIGNAL_RE.search(_s_quote))

                # Aggregate untrusted flag
                _s_untrusted = _s_domain_mismatch or _s_is_subsidiary_regional

                # Build rejection/trigger reason
                _s_trust_reason = ""
                if _s_domain_mismatch:
                    _s_trust_reason = f"unrelated_domain:{_s_evidence_domain}"
                    result["domain_root_hq_rejected_evidence_reason"] = (
                        f"evidence from {_s_evidence_domain}, not {_s_input_nl}"
                    )
                if _s_is_subsidiary_regional:
                    _s_sub_label = "subsidiary_or_regional_hq_evidence"
                    _s_trust_reason = (
                        (_s_trust_reason + "; " + _s_sub_label)
                        if _s_trust_reason else _s_sub_label
                    )

                if _s_untrusted:
                    _s_conf = "Low"
                    result["hq_confidence"]             = _s_conf
                    result["domain_root_hq_confidence"] = _s_conf
                    _set_manual_review(result, "Yes", _s_trust_reason)

                if _s_is_foreign:
                    result["foreign_hq_simple"]                    = "True"
                    result["hq_structure_type"]                    = "foreign_parent"
                    result["parent_group_hq_country"]              = _s_country_std
                    result["parent_group_hq_city"]                 = _s_city
                    result["local_entity_hq_country"]              = _s_input_std
                    result["sig_foreign_hq_review_source"]         = "simple_domain_root_hq_search"
                    result["sig_foreign_hq_review_evidence_url"]   = _s_best_url
                    result["sig_foreign_hq_review_evidence_quote"] = _s_quote
                    result["sig_foreign_hq_review_reason"] = (
                        f"Simple domain-root HQ search ('{_dr_query}') "
                        f"identifies {_s_country_std} as HQ"
                    )
                    result["sig_foreign_hq_review_confidence"] = _s_conf
                    _score3_ok, _score3_reason = _is_score3_eligible_hq_evidence(
                        quote=_s_quote,
                        detected_country=_s_country_std,
                        input_country=_s_input_std,
                        evidence_url=_s_best_url,
                    )
                    # ── Optional Haiku review for uncertain score-3 candidates ──
                    _haiku_used = False
                    _haiku_result: dict[str, Any] = {}
                    _haiku_eligible = (
                        use_haiku_uncertain
                        and anthropic_key
                        and _score3_ok           # only when det. logic would say score 3
                        and not _s_untrusted
                        and (
                            _s_conf in ("Medium", "Low")
                            or not _s_is_official_domain
                        )
                        and (haiku_calls_counter is None
                             or haiku_calls_counter[0] < haiku_max_calls)
                    )
                    if _haiku_eligible:
                        # Gather top organic snippets for context
                        _hk_snips = [
                            f"{r.get('title','')} — {r.get('snippet','')}"
                            for r in _s_org[:5]
                        ]
                        _hk_ab = str(_s_data.get("answerBox", {}).get("snippet", "")
                                     or _s_data.get("answerBox", {}).get("answer", ""))
                        _haiku_result = _review_hq_with_haiku(
                            company_name=company_name,
                            domain=domain,
                            input_country=_s_input_std,
                            detected_country=_s_country_std,
                            detected_city=_s_city,
                            evidence_quote=_s_quote,
                            evidence_url=_s_best_url,
                            answer_box=_hk_ab,
                            organic_titles_snippets=_hk_snips,
                            anthropic_key=anthropic_key,
                        )
                        if haiku_calls_counter is not None and not _haiku_result.get("haiku_error"):
                            haiku_calls_counter[0] += 1
                        _haiku_used = not bool(_haiku_result.get("haiku_error"))
                        # Apply country correction from Haiku if provided
                        _hk_cc = (_haiku_result.get("haiku_country_corrected") or "").strip()
                        if _hk_cc and _hk_cc.lower() not in ("", "none"):
                            _s_country_std = _std_country(_hk_cc)
                            result["hq_detected_country"]     = _s_country_std
                            result["domain_root_hq_country"]  = _s_country_std
                            result["parent_group_hq_country"] = _s_country_std
                            # Re-check foreign after Haiku correction
                            _s_is_foreign = bool(
                                _s_country_std
                                and _s_country_std.lower() != _s_input_std.lower()
                            )
                        # Override score decision if Haiku is confident
                        _hk_verdict = _haiku_result.get("haiku_verdict", "")
                        if _hk_verdict == "score0":
                            _score3_ok = False
                            _score3_reason = (
                                f"haiku_override_score0: {_haiku_result.get('haiku_reason','')}"
                            )
                        elif _hk_verdict == "score3":
                            _score3_ok = True
                            _score3_reason = ""

                    if _haiku_used:
                        result["anthropic_hq_review_used"]        = "Yes"
                        result["anthropic_review_evidence_mode"]  = "serper_evidence_only"
                        result["anthropic_web_search_used"]       = "No"
                        result["_anthr_model"]       = _HAIKU_MODEL
                        result["_anthr_input_tok"]   = _haiku_result.get("haiku_input_tokens", 0)
                        result["_anthr_output_tok"]  = _haiku_result.get("haiku_output_tokens", 0)
                        result["_anthr_total_tok"]   = (
                            _haiku_result.get("haiku_input_tokens", 0)
                            + _haiku_result.get("haiku_output_tokens", 0)
                        )
                        if _haiku_result.get("haiku_error"):
                            result["anthropic_error"] = _haiku_result["haiku_error"]
                        result["sig_foreign_hq_review_reason"] = (
                            result.get("sig_foreign_hq_review_reason", "")
                            + f" [haiku:{_haiku_result.get('haiku_verdict','')}:"
                            f"{_haiku_result.get('haiku_reason','')[:80]}]"
                        )

                    if _s_untrusted or not _score3_ok:
                        # Evidence not trusted or fails eligibility — score 0, manual review
                        result["sig_foreign_hq_score_reviewed"] = 0
                        result["review_foreign_parent_score"]   = 0
                        result["hq_confidence"] = "Low"
                        if _score3_reason and not result.get("domain_root_hq_rejected_evidence_reason"):
                            result["domain_root_hq_rejected_evidence_reason"] = _score3_reason
                        _reason_mr = _s_trust_reason or _score3_reason or "score3_ineligible"
                        _set_manual_review(result, "Yes", _reason_mr)
                    else:
                        result["sig_foreign_hq_score_reviewed"] = 3
                        result["review_foreign_parent_score"]   = 3
                        _set_manual_review(result, "No")
                else:
                    result["foreign_hq_simple"]             = "False"
                    result["hq_structure_type"]             = "domestic_italy"
                    result["sig_foreign_hq_score_reviewed"] = 0
                    if not _s_untrusted:
                        _set_manual_review(result, "No")
            else:
                _set_manual_review(result, "Yes")
                result["hq_reason"] = f"No HQ location found for query: '{_dr_query}'"
        else:
            result["needs_manual_review"] = "Yes"
            if _s_err:
                result["probe_error"] = _s_err

        result["input_country_used"]  = _std_country(input_country)
        result["hq_query_used"]       = _dr_query
        result["serper_calls_used"]   = _s_calls
        result["serper_queries_used"] = " | ".join(_s_queries)
        result["serper_cache_hit"]    = (
            "partial" if (_s_calls > 0 and _s_cache_hit)
            else ("True" if _s_cache_hit else "False")
        )
        result["row_runtime_seconds"] = round(time.monotonic() - _row_start, 2)
        result["_anthr_model"]        = ""
        result["_anthr_input_tok"]    = 0
        result["_anthr_output_tok"]   = 0
        result["_anthr_total_tok"]    = 0
        result["_anthr_stop"]         = ""

        _irow = input_row or {}
        result["sig_foreign_hq_score_original"] = _safe_float(_irow.get("sig_foreign_hq_score"))
        if result.get("sig_foreign_hq_score_reviewed") in ("", None):
            result["sig_foreign_hq_score_reviewed"] = _safe_float(_irow.get("sig_foreign_hq_score"))

        return result

    queries    = _build_queries(company_name, domain)
    best: dict = {}
    used_query = ""
    all_organic: list[dict] = []
    kg_location = ""
    answer_box_text = ""
    errors: list[str] = []

    # Usage tracking
    _serper_calls      = 0
    _serper_cache_hits = 0
    _queries_attempted: list[str] = []
    _anthr_input_tok  = 0
    _anthr_output_tok = 0
    _anthr_total_tok  = 0
    _anthr_model      = ""
    _anthr_stop       = ""
    _anthr_errors: list[str] = []

    # --- Manual Google Mimic HQ Check (runs before narrow queries) ---
    _mimic_audit: dict = {}
    _mimic_evidence: dict = {}
    if use_mimic_check and serper_key:
        _mimic_result = _mimic_google_hq_check(
            company_name, domain, input_country, serper_key, cache,
            max_page_paths=_max_page_paths,
            max_brand_queries=_max_brand_queries,
        )
        _mimic_serper_calls = _mimic_result.pop("_mimic_serper_calls", 0)
        _mimic_queries      = _mimic_result.pop("_mimic_queries", [])
        _mimic_cache_hits   = _mimic_result.pop("_mimic_serper_cache_hits", 0)
        _mimic_evidence     = _mimic_result.pop("_mimic_hq_evidence", {})
        # Accumulate into tracking
        _serper_calls      += _mimic_serper_calls
        _serper_cache_hits += _mimic_cache_hits
        _queries_attempted  = _mimic_queries + _queries_attempted
        # Store audit cols for later merge
        _mimic_audit = _mimic_result
        # If mimic found strong evidence, use it as best and skip narrow queries
        if _mimic_evidence.get("hq_detected_country"):
            best = {k: v for k, v in _mimic_evidence.items()}
            used_query = _mimic_queries[0] if _mimic_queries else ""
            result["early_stop_used"]   = True
            result["early_stop_reason"] = "clear_official_hq_evidence"
    else:
        _mimic_audit = {"manual_google_mimic_used": False}

    _narrow_serper_budget = max(0, _max_serper - _serper_calls)

    for query in queries:
        if best.get("hq_detected_country"):
            break  # mimic already found strong evidence — early stop
        if _narrow_serper_budget <= 0:
            break  # Serper budget exhausted for this row
        _queries_attempted.append(query)
        cache_key = ("serper", query)
        if cache_key in cache:
            data, err = cache[cache_key]
            _serper_cache_hits += 1
        else:
            if _serper_calls >= _max_serper:
                break  # double-guard
            data, err = _serper_search(query, serper_key)
            cache[cache_key] = (data, err)
            _serper_calls += 1
            _narrow_serper_budget -= 1
            time.sleep(_INTER_REQUEST_SLEEP)

        if err:
            errors.append(f"{query!r}: {err}")
            continue

        organic = data.get("organic", [])

        if not kg_location:
            kg = data.get("knowledgeGraph", {})
            kg_location = (
                kg.get("address", "") or kg.get("headquarters", "") or kg.get("location", "")
            )
        if not answer_box_text:
            ab = data.get("answerBox", {})
            answer_box_text = ab.get("answer", "") or ab.get("snippet", "")
        if not all_organic and organic:
            all_organic = organic

        extracted = _extract_deterministic(
            organic, kg_location, answer_box_text, input_country=input_country
        )
        if extracted and extracted.get("hq_detected_country"):
            best = extracted
            used_query = query
            all_organic = organic
            break

    # Model fallback
    if not best.get("hq_detected_country") and use_model and anthropic_key:
        snippets = [
            {"title": r.get("title", ""), "snippet": r.get("snippet", ""), "url": r.get("link", "")}
            for r in all_organic[:5]
        ]
        model_result = _model_extract(company_name, snippets, anthropic_key)
        _mu = model_result.pop("_usage", {})
        if _mu:
            _anthr_model      = _mu.get("model", "")
            _anthr_input_tok  += _mu.get("input_tokens", 0)
            _anthr_output_tok += _mu.get("output_tokens", 0)
            _anthr_total_tok  += _mu.get("total_tokens", 0)
            _anthr_stop        = _mu.get("stop_reason", "")
        is_group_only = model_result.get("is_group_or_parent_only", False)
        entity_match  = model_result.get("entity_match", "")
        if model_result.get("hq_country") and not is_group_only and entity_match != "No":
            prefix = f"[model entity={entity_match}] " if entity_match else "[model] "
            best = {
                "hq_detected_city":    model_result.get("hq_city", ""),
                "hq_detected_region":  model_result.get("hq_region", ""),
                "hq_detected_country": model_result.get("hq_country", ""),
                "hq_confidence":       model_result.get("confidence", ""),
                "hq_reason":           prefix + model_result.get("reason", ""),
                "hq_evidence_url":     model_result.get("evidence_url", ""),
                "hq_evidence_quote":   model_result.get("evidence_quote", ""),
            }
            if not used_query and queries:
                used_query = queries[0]
        elif model_result.get("hq_country") and is_group_only:
            best = {
                "hq_detected_city":    model_result.get("hq_city", ""),
                "hq_detected_region":  model_result.get("hq_region", ""),
                "hq_detected_country": model_result.get("hq_country", ""),
                "hq_confidence":       "Low",
                "hq_reason":           "[model – group/parent only] " + model_result.get("reason", ""),
                "hq_evidence_url":     model_result.get("evidence_url", ""),
                "hq_evidence_quote":   model_result.get("evidence_quote", ""),
            }
            if not used_query and queries:
                used_query = queries[0]
        if model_result.get("probe_error"):
            errors.append(model_result["probe_error"])
            _anthr_errors.append(model_result["probe_error"])

    for k, v in best.items():
        if k in result:
            result[k] = v

    result["hq_query_used"]                    = used_query
    result["serper_knowledge_graph_location"]  = kg_location
    result["serper_answer_box"]                = answer_box_text

    for i, item in enumerate(all_organic[:3], start=1):
        result[f"top_organic_title_{i}"]   = item.get("title", "")
        result[f"top_organic_snippet_{i}"] = item.get("snippet", "")
        result[f"top_organic_url_{i}"]     = item.get("link", "")

    foreign_str, needs_review = _classify(
        result["hq_detected_country"],
        input_country,
        result["hq_confidence"],
        result["hq_reason"],
        all_organic,
    )
    result["input_country_used"] = _std_country(input_country)
    result["foreign_hq_simple"]  = foreign_str
    result["needs_manual_review"] = "Yes" if needs_review else "No"

    if errors:
        result["probe_error"] = "; ".join(errors)

    # ----------------------------------------------------------------
    # Recovery review additions
    # ----------------------------------------------------------------
    _irow = input_row or {}

    # Preserve original sig_foreign_hq_score
    orig_score = _safe_float(_irow.get("sig_foreign_hq_score"))
    result["sig_foreign_hq_score_original"] = orig_score
    result["sig_foreign_hq_score_reviewed"] = orig_score  # default: keep original

    # Multilingual site detection
    ml_result: dict[str, Any] = {
        "has_multilingual_site": False, "website_language_count": 0,
        "website_languages_detected": "", "website_fetch_count": 0,
    }
    if use_multilingual_check and domain:
        _ml_key = ("multilingual", re.sub(r"^https?://", "", domain.strip()).rstrip("/"))
        if _ml_key in cache:
            ml_result = cache[_ml_key]
        else:
            ml_result = detect_multilingual_site(domain)
            cache[_ml_key] = ml_result
    result.update(ml_result)

    # Global structure detection
    gs_result = detect_global_structure(all_organic, kg_location, answer_box_text)
    result.update(gs_result)

    # HQ review trigger
    trigger_str, needs_adj = compute_review_trigger(
        result, _irow, ml_result, gs_result, company_name=company_name
    )
    result["hq_review_trigger"]              = trigger_str
    result["needs_anthropic_hq_review"]      = "Yes" if needs_adj else "No"
    result["anthropic_hq_review_used"]       = "No"
    result["anthropic_web_search_used"]      = "No"
    result["anthropic_review_evidence_mode"] = ""
    result["anthropic_web_search_queries"]   = ""
    result["anthropic_web_search_result_count"] = 0

    # ---- Merge mimic audit columns (before Anthropic so we can inspect evidence) ----
    for _k, _v in _mimic_audit.items():
        if _k in result:
            result[_k] = _v

    # ── Apply classification fields from all HQ evidence sources ─────────────
    # Priority: domain_root_hq → brand_root_hq → known_global_brand → Anthropic
    _input_std  = _std_country(input_country or "Italy")

    # Domain-root evidence (highest priority — clean, unambiguous query)
    _dr_country = result.get("domain_root_hq_country", "")
    _dr_city    = result.get("domain_root_hq_city", "")
    _dr_conf    = result.get("domain_root_hq_confidence", "")
    _dr_std     = _std_country(_dr_country)
    _dr_is_foreign = bool(_dr_country and _dr_std and _dr_std.lower() != _input_std.lower())

    if _dr_is_foreign and _dr_conf in ("High", "Medium"):
        result["hq_structure_type"]             = "foreign_parent"
        result["parent_group_hq_country"]       = _dr_country
        result["parent_group_hq_city"]          = _dr_city
        result["local_entity_hq_country"]       = _input_std
        result["sig_foreign_hq_score_reviewed"] = 3
        result["review_foreign_parent_score"]   = 3
        result["sig_foreign_hq_review_source"]  = "domain_root_hq_search"
        result["sig_foreign_hq_review_evidence_url"]   = result.get("domain_root_hq_evidence_url", "")
        result["sig_foreign_hq_review_evidence_quote"] = result.get("domain_root_hq_evidence_quote", "")
        result["sig_foreign_hq_review_reason"]  = (
            f"Domain-root HQ search ('{result.get('domain_root_hq_query','')}') "
            f"identifies {_dr_country} as parent/group HQ"
        )
        result["sig_foreign_hq_review_confidence"] = _dr_conf
        result["foreign_hq_simple"]             = "True"
        result["hq_detected_country"]           = _dr_std
        result["hq_detected_city"]              = _dr_city
        result["hq_confidence"]                 = _dr_conf
        result["needs_anthropic_hq_review"]     = "No"
        needs_adj = False

    _br_country = result.get("brand_root_hq_country", "")
    _br_city    = result.get("brand_root_hq_city", "")
    _br_std     = _std_country(_br_country)
    _kgb_type   = result.get("known_global_brand_type", "")
    _kgb_name   = result.get("known_global_brand_name", "")
    _kgb_detected = bool(result.get("known_global_brand_detected"))

    if not _dr_is_foreign and _br_country and _br_std and _br_std.lower() != _input_std.lower():
        # Brand-root found a foreign parent HQ → classify as foreign_parent
        result["hq_structure_type"]           = "foreign_parent"
        result["parent_group_hq_country"]     = _br_country
        result["parent_group_hq_city"]        = _br_city
        result["local_entity_hq_country"]     = _input_std
        result["sig_foreign_hq_score_reviewed"] = 3
        result["review_foreign_parent_score"] = 3
        result["sig_foreign_hq_review_source"] = "brand_root_hq_check"
        result["sig_foreign_hq_review_evidence_url"]   = result.get("brand_root_hq_evidence_url", "")
        result["sig_foreign_hq_review_evidence_quote"] = result.get("brand_root_hq_evidence_quote", "")
        result["sig_foreign_hq_review_reason"] = (
            f"Brand-root HQ check ({result.get('brand_root_hq_source_type','')}) "
            f"identifies {_br_country} as parent/group HQ"
        )
        result["sig_foreign_hq_review_confidence"] = (
            "High" if _mimic_evidence.get("hq_confidence") == "High" else "Medium"
        )
        result["foreign_hq_simple"]           = "True"
        result["hq_detected_country"]         = _br_std
        result["hq_detected_city"]            = _br_city
        result["hq_confidence"]               = "High"
        result["needs_anthropic_hq_review"]   = "No"
        needs_adj = False

    elif not _dr_is_foreign and _kgb_detected and _kgb_type == "global_network" and not _br_country:
        # Known global professional-services network — no single HQ country needed
        result["hq_structure_type"]              = "global_network"
        result["local_entity_hq_country"]        = _input_std
        result["sig_foreign_hq_score_reviewed"]  = 3
        result["review_global_network_score"]    = 3
        result["sig_foreign_hq_review_source"]   = "known_global_brand_check"
        result["sig_foreign_hq_review_reason"]   = (
            f"{_kgb_name} is a known global professional-services network; "
            "local entity classified as global_network"
        )
        result["sig_foreign_hq_review_confidence"] = "High"
        result["sig_foreign_hq_review_evidence_url"]   = result.get("brand_root_hq_evidence_url", "")
        result["sig_foreign_hq_review_evidence_quote"] = result.get("brand_root_hq_evidence_quote", "")
        result["foreign_hq_simple"]              = "True"
        result["needs_anthropic_hq_review"]      = "No"
        needs_adj = False

    elif not _dr_is_foreign and _kgb_detected and _kgb_type == "corporate" and not _br_country:
        # Known corporate brand but neither domain-root nor brand-root resolved a country yet.
        result["sig_foreign_hq_review_reason"] = (
            f"{_kgb_name} is a known global corporate brand; "
            "domain-root/brand-root HQ search did not resolve a foreign country"
        )

    # Suppress Anthropic when official page evidence is already strong and unambiguous
    _has_strong_official = (
        result.get("official_page_hq_evidence_strength") == "Strong"
        and result.get("official_page_hq_country")
        and not result.get("official_top_result_is_local_country_page")
    )
    _has_strong_brand = (
        result.get("brand_hq_evidence_found")
        and _mimic_evidence.get("hq_confidence") == "High"
        and _mimic_evidence.get("hq_detected_country")
    )
    _has_strong_brand_root  = bool(_br_country and _br_std.lower() != _input_std.lower())
    _has_global_network_kgb = bool(_kgb_detected and _kgb_type == "global_network")
    if _dr_is_foreign or _has_strong_official or _has_strong_brand or _has_strong_brand_root or _has_global_network_kgb:
        _ev_src = (
            "domain_root_hq_search" if _dr_is_foreign
            else ("brand_root_hq_check" if _has_strong_brand_root
                  else ("known_global_brand_check" if _has_global_network_kgb
                        else ("official_page_mimic_check" if _has_strong_official else "brand_hq_mimic_check")))
        )
        result["needs_anthropic_hq_review"] = "No"
        result["anthropic_hq_review_used"]  = "No"
        result["anthropic_web_search_used"] = "No"
        result["anthropic_review_evidence_mode"] = ""
        if not _dr_is_foreign and not _has_strong_brand_root:
            result["sig_foreign_hq_review_source"] = _ev_src
        if _has_strong_official and not _dr_is_foreign and not _has_strong_brand_root:
            result["sig_foreign_hq_review_evidence_url"]   = result.get("official_page_fetch_url", "")
            result["sig_foreign_hq_review_evidence_quote"] = result.get("official_page_hq_evidence_quote", "")
        needs_adj = False

    # Optional Anthropic HQ adjudication
    if use_anthropic_review and needs_adj and anthropic_key:
        adj = _anthropic_hq_adjudicate(
            company_name, domain, result, ml_result, gs_result, anthropic_key
        )
        _au = adj.pop("_usage", {})
        if _au:
            _anthr_model      = _anthr_model or _au.get("model", "")
            _anthr_input_tok  += _au.get("input_tokens", 0)
            _anthr_output_tok += _au.get("output_tokens", 0)
            _anthr_total_tok  += _au.get("total_tokens", 0)
            _anthr_stop        = _anthr_stop or _au.get("stop_reason", "")
        if not adj.get("probe_error"):
            for _field in (
                "hq_structure_type", "local_entity_hq_country", "local_entity_hq_city",
                "parent_group_hq_country", "parent_group_hq_city",
                "review_foreign_parent_score", "review_global_network_score",
                "review_multilingual_website_score", "review_recommended_hq_signal",
                "sig_foreign_hq_score_reviewed",
            ):
                if _field in adj:
                    result[_field] = adj[_field]
            result["sig_foreign_hq_review_reason"]         = adj.get("reason", "")
            result["sig_foreign_hq_review_confidence"]     = adj.get("confidence", "")
            result["sig_foreign_hq_review_source"]         = "anthropic"
            result["sig_foreign_hq_review_evidence_url"]   = adj.get("evidence_url", "")
            result["sig_foreign_hq_review_evidence_quote"] = adj.get("evidence_quote", "")
            result["anthropic_hq_review_used"]             = "Yes"
            result["anthropic_web_search_used"]            = "No"
            result["anthropic_review_evidence_mode"]       = "serper_evidence_only"
        else:
            _adj_err = adj["probe_error"]
            _anthr_errors.append(_adj_err)
            result["probe_error"] = (result["probe_error"] + "; " + _adj_err).lstrip("; ") if result.get("probe_error") else _adj_err

    # ── Final domestic-safety guard ──────────────────────────────────────────
    # Two triggers:
    # 1. hq_detected_country normalises to input_country (e.g. Italy)
    # 2. Evidence quote explicitly mentions an Italian location but
    #    hq_detected_country was misclassified as a foreign country
    #    (e.g. "emmegi headquarters" snippet says "Modena" but Serper KG
    #     shows a US address for a different entity)
    # Use display values for output, normalised values for equality comparisons.
    # _normalize_country_for_hq handles IT/ITA/Italia/Italy all as "italy".
    _final_det_display = _std_country(result.get("hq_detected_country") or "")
    _final_inp_display = _std_country(result.get("input_country_used") or input_country or "")
    _final_det_norm    = _normalize_country_for_hq(result.get("hq_detected_country"))
    _final_inp_norm    = _normalize_country_for_hq(result.get("input_country_used") or input_country)

    # Build a combined evidence string from all collected quotes
    _ev_quotes = " ".join(filter(None, [
        result.get("hq_evidence_quote", ""),
        result.get("domain_root_hq_evidence_quote", ""),
        result.get("brand_root_hq_evidence_quote", ""),
        result.get("official_page_hq_evidence_quote", ""),
        result.get("sig_foreign_hq_review_evidence_quote", ""),
    ]))

    # Italian location tokens that indicate domestic Italy HQ
    _ITALY_LOCATION_TOKENS = re.compile(
        r"\b(?:Italy|Italia|Italian|Modena|Ancona|Osimo|Senigallia|Venezia|Venice|"
        r"Milano|Milan|Roma|Rome|Torino|Turin|Firenze|Florence|Bologna|Napoli|Naples|"
        r"Bergamo|Brescia|Padova|Padua|Verona|Perugia|Bari|Catania|Palermo|Cagliari|"
        r"Reggio|Emilia|Toscana|Lombardia|Veneto|Lazio|Campania|Puglia|Sicilia|"
        r"Avellino|Nusco|San\s+Benedetto|Jesi|Fabriano|Civitanova)\b",
        re.IGNORECASE,
    )
    _quote_has_italy = bool(_ITALY_LOCATION_TOKENS.search(_ev_quotes))
    _quote_overrides_foreign = (
        _final_inp_norm == "italy"
        and _quote_has_italy
        and _final_det_norm
        and _final_det_norm != "italy"
    )

    _is_domestic = (
        (_final_det_norm and _final_inp_norm and _final_det_norm == _final_inp_norm)
        or _quote_overrides_foreign
    )

    if _is_domestic:
        _dom_country = _final_inp_display or "Italy"
        result["foreign_hq_simple"]             = "False"
        result["hq_structure_type"]             = "domestic"
        result["sig_foreign_hq_score_reviewed"] = 0
        result["sig_foreign_hq_score_for_next_scoring"] = 0
        result["parent_group_hq_country"]       = _dom_country
        result["hq_detected_country"]           = _dom_country
        result["review_foreign_parent_score"]   = 0
        result["review_global_network_score"]   = 0
        if result.get("hq_confidence") in ("High", "Medium"):
            result["needs_manual_review"] = "No"
        _guard_reason = (
            "quote_contains_italy_location_overrides_foreign"
            if _quote_overrides_foreign
            else f"hq_detected={_final_det_display}({_final_det_norm})==input={_final_inp_display}({_final_inp_norm})"
        )
        result["sig_foreign_hq_review_reason"] = (
            (result.get("sig_foreign_hq_review_reason") or "")
            + f" [domestic-guard: {_guard_reason}]"
        ).lstrip()
    else:
        # Keep reviewed score as-is; copy to next-scoring column
        result["sig_foreign_hq_score_for_next_scoring"] = result.get("sig_foreign_hq_score_reviewed", "")

    # ---- Populate usage columns ----
    result["serper_calls_used"]   = _serper_calls
    result["serper_queries_used"] = " | ".join(_queries_attempted)
    if _serper_calls > 0 and _serper_cache_hits > 0:
        result["serper_cache_hit"] = "partial"
    elif _serper_cache_hits > 0:
        result["serper_cache_hit"] = "True"
    else:
        result["serper_cache_hit"] = "False"
    result["website_fetch_count"] = ml_result.get("website_fetch_count", 0)
    result["anthropic_error"]     = "; ".join(_anthr_errors)
    result["row_runtime_seconds"] = round(_time_mod.monotonic() - _row_start, 2)
    # Internal token tracking (for run-summary aggregation only, not in Excel rows)
    result["_anthr_model"]        = _anthr_model
    result["_anthr_input_tok"]    = _anthr_input_tok
    result["_anthr_output_tok"]   = _anthr_output_tok
    result["_anthr_total_tok"]    = _anthr_total_tok
    result["_anthr_stop"]         = _anthr_stop

    return result

# ---------------------------------------------------------------------------
# Input helpers
# ---------------------------------------------------------------------------

def get_excel_sheet_names(fileobj: "io.IOBase | Path") -> list[str]:
    wb = load_workbook(fileobj, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def read_input_from_fileobj(
    fileobj: "io.IOBase",
    suffix: str,
    limit: int = DEFAULT_LIMIT,
    sheet_name: "str | None" = None,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    suffix = suffix.lower()
    if suffix in (".xlsx", ".xls"):
        wb = load_workbook(fileobj, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        iter_rows = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else "" for h in next(iter_rows)]
        for raw in iter_rows:
            if all(v is None or str(v).strip() == "" for v in raw):
                continue
            rows.append({headers[i]: raw[i] for i in range(len(headers))})
            if len(rows) >= limit:
                break
        wb.close()
    elif suffix == ".csv":
        text = fileobj.read()
        if isinstance(text, bytes):
            text = text.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append(dict(row))
            if len(rows) >= limit:
                break
    return rows

# ---------------------------------------------------------------------------
# Output builder
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_PROBE_FILL  = PatternFill("solid", fgColor="E2EFDA")
_PROBE_FONT  = Font(bold=True)
_OLD_FILL    = PatternFill("solid", fgColor="D6E4F0")
_WARN_FILL   = PatternFill("solid", fgColor="FCE4D6")

_WRAP_COLS = {
    "hq_reason", "hq_evidence_quote",
    "top_organic_snippet_1", "top_organic_snippet_2", "top_organic_snippet_3",
    "sig_foreign_hq_evidence",
}

_COL_WIDTHS: dict[str, int] = {
    "source_row": 10, "company_name": 38, "domain": 28, "input_country": 14,
    "hq_detected_city": 18, "hq_detected_region": 18, "hq_detected_country": 18,
    "input_country_used": 18,
    "hq_confidence": 12, "foreign_hq_simple": 16, "needs_manual_review": 18,
    "hq_reason": 50, "hq_evidence_url": 45, "hq_evidence_quote": 60,
    "hq_query_used": 45, "serper_knowledge_graph_location": 35, "serper_answer_box": 45,
    "top_organic_title_1": 45, "top_organic_snippet_1": 60, "top_organic_url_1": 45,
    "top_organic_title_2": 45, "top_organic_snippet_2": 60, "top_organic_url_2": 45,
    "top_organic_title_3": 45, "top_organic_snippet_3": 60, "top_organic_url_3": 45,
    "probe_error": 40,
    "sig_foreign_hq_score": 20, "sig_foreign_hq_evidence": 55,
    "foreign_hq_sanitized": 18, "foreign_hq_sanitizer_reason": 35,
    "foreign_hq_original_score": 22, "final_commercial_fit_score": 24,
    "commercial_fit_score": 22,
}


def _build_workbook(
    input_rows: list[dict],
    probe_results: list[dict],
    present_old_cols: list[str],
    company_col: str,
    domain_col: str,
    country_col: str,
    qa_meta: dict,
    output_label: str = "",
    run_usage: "dict | None" = None,
) -> "Workbook":
    wb = Workbook()
    ws = wb.active
    ws.title = "HQ Probe Results"

    input_identity_cols = ["source_row", company_col, domain_col, country_col]
    seen: set[str] = set()
    all_cols: list[str] = []
    for c in input_identity_cols + present_old_cols + PROBE_COLS:
        if c not in seen:
            all_cols.append(c)
            seen.add(c)

    ws.append(all_cols)
    ws.row_dimensions[1].height = 18
    probe_col_set = set(PROBE_COLS)
    old_col_set   = set(present_old_cols)
    for col_idx, h in enumerate(all_cols, start=1):
        cell = ws.cell(row=1, column=col_idx)
        if h in probe_col_set:
            cell.fill = _PROBE_FILL; cell.font = _PROBE_FONT
        elif h in old_col_set:
            cell.fill = _OLD_FILL; cell.font = Font(bold=True)
        else:
            cell.fill = _HEADER_FILL; cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=False)
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            _COL_WIDTHS.get(h, max(14, min(len(h) + 4, 35)))
        )

    for row_idx, (in_row, probe) in enumerate(zip(input_rows, probe_results), start=2):
        values: list[Any] = []
        for h in all_cols:
            if h == "source_row":
                values.append(row_idx - 1)
            elif h in probe:
                values.append(probe[h])
            else:
                v = in_row.get(h)
                values.append(v if v is not None else "")
        ws.append(values)
        ws.row_dimensions[row_idx].height = 15
        if probe.get("needs_manual_review") == "Yes":
            for col_idx in range(1, len(all_cols) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = _WARN_FILL
        for col_idx, h in enumerate(all_cols, start=1):
            if h in _WRAP_COLS:
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions

    # QA sheet
    ws_qa = wb.create_sheet("Run QA")
    detected_italy   = sum(1 for p in probe_results if _std_country(p.get("hq_detected_country") or "") == "Italy")
    detected_foreign = sum(1 for p in probe_results if p.get("foreign_hq_simple") == "True")
    detected_unknown = sum(1 for p in probe_results if not p.get("hq_detected_country"))
    needs_review_cnt = sum(1 for p in probe_results if p.get("needs_manual_review") == "Yes")
    has_errors       = sum(1 for p in probe_results if p.get("probe_error"))

    qa_rows_data = [
        ("Run QA – hq_lookup_probe_app.py", ""),
        ("", ""),
        ("timestamp",           qa_meta.get("timestamp", "")),
        ("input_file",          qa_meta.get("input_file", "")),
        ("output_file",         output_label or qa_meta.get("output_file", "")),
        ("company_col",         company_col),
        ("domain_col",          domain_col),
        ("country_col",         country_col or "(default)"),
        ("rows_processed",      len(probe_results)),
        ("", ""),
        ("── detection summary ──", ""),
        ("detected_italy_hq",   detected_italy),
        ("detected_foreign_hq", detected_foreign),
        ("detected_unknown",    detected_unknown),
        ("needs_manual_review", needs_review_cnt),
        ("rows_with_errors",    has_errors),
        ("", ""),
        ("── settings ──", ""),
        ("model_extraction_used", qa_meta.get("use_model", False)),
        ("model_used",          qa_meta.get("model", "")),
        ("serper_available",    qa_meta.get("serper_available", False)),
        ("limit",               qa_meta.get("limit", "")),
    ]
    ws_qa.column_dimensions["A"].width = 30
    ws_qa.column_dimensions["B"].width = 80
    for r_idx, (k, v) in enumerate(qa_rows_data, start=1):
        cell_a = ws_qa.cell(row=r_idx, column=1, value=k)
        ws_qa.cell(row=r_idx, column=2, value=v)
        if r_idx == 1:
            cell_a.font = Font(bold=True, size=13)
        elif k and not k.startswith(" "):
            cell_a.font = Font(bold=True)
        ws_qa.row_dimensions[r_idx].height = 15

    # Run Summary sheet
    _ru = run_usage or {}
    ws_sum = wb.create_sheet("Run Summary")
    sum_rows = [
        ("Run Summary – hq_lookup_probe_app.py", ""),
        ("", ""),
        ("── run info ──", ""),
        ("timestamp",           qa_meta.get("timestamp", "")),
        ("input_file",          qa_meta.get("input_file", "")),
        ("rows_processed",      len(probe_results)),
        ("", ""),
        ("── API usage ──", ""),
        ("serper_calls_used",          _ru.get("serper_calls_used", "")),
        ("rows_with_serper_cache_hit", _ru.get("rows_with_cache_hit", "")),
        ("website_fetches_attempted",  _ru.get("website_fetches", "")),
        ("anthropic_reviews_used",     _ru.get("anthropic_reviews_used", "")),
        ("anthropic_web_search_used",  _ru.get("anthropic_web_search_used", "")),
        ("anthropic_errors",           _ru.get("anthropic_errors", "")),
        ("", ""),
        ("── options used ──", ""),
        ("run_mode",                    qa_meta.get("run_mode", "")),
        ("multilingual_check_enabled",  qa_meta.get("use_multilingual_check", "")),
        ("anthropic_hq_review_enabled", qa_meta.get("use_anthropic_review", "")),
        ("model_fallback_enabled",      qa_meta.get("use_model", "")),
        ("model",                       qa_meta.get("model", "")),
        ("row_limit",                   qa_meta.get("limit", "")),
    ]
    ws_sum.column_dimensions["A"].width = 32
    ws_sum.column_dimensions["B"].width = 50
    for r_idx, (k, v) in enumerate(sum_rows, start=1):
        cell_a = ws_sum.cell(row=r_idx, column=1, value=k)
        ws_sum.cell(row=r_idx, column=2, value=v)
        if r_idx == 1:
            cell_a.font = Font(bold=True, size=13)
        elif k and not k.startswith("─"):
            cell_a.font = Font(bold=True)
        ws_sum.row_dimensions[r_idx].height = 15

    return wb


def build_excel_bytes(
    input_rows: list[dict],
    probe_results: list[dict],
    present_old_cols: list[str],
    company_col: str,
    domain_col: str,
    country_col: str,
    qa_meta: dict,
    run_usage: "dict | None" = None,
) -> bytes:
    wb = _build_workbook(
        input_rows=input_rows,
        probe_results=probe_results,
        present_old_cols=present_old_cols,
        company_col=company_col,
        domain_col=domain_col,
        country_col=country_col,
        qa_meta=qa_meta,
        output_label="(in-memory)",
        run_usage=run_usage,
    )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ---------------------------------------------------------------------------
# Streamlit app
# ---------------------------------------------------------------------------

st.set_page_config(
    page_title="mYngle HQ Lookup Probe",
    page_icon="🔍",
    layout="wide",
)

st.title("🔍 mYngle HQ Lookup Probe")
st.caption(
    "This tool tests a simple headquarters lookup approach. "
    "It is **experimental** and does not change production enrichment output."
)

# ---------------------------------------------------------------------------
# Column guessers
# ---------------------------------------------------------------------------

_COMPANY_GUESSES = ["company_name", "company", "naam", "azienda", "name"]
_DOMAIN_GUESSES  = ["domain", "website", "url", "website_url", "domein"]
_COUNTRY_GUESSES = ["inferred_input_country", "input_country", "country", "paese", "land"]

# Columns that must never be auto-selected as the country column
_NOT_COUNTRY_COLS = {
    "company_name", "company", "naam", "azienda", "name",
    "domain", "website", "url", "website_url", "domein",
}


def _guess_col(columns: list[str], guesses: list[str]) -> str:
    """Return best matching column from guesses, or first column as fallback."""
    cols_lower = {c.lower(): c for c in columns}
    for g in guesses:
        if g.lower() in cols_lower:
            return cols_lower[g.lower()]
    return columns[0] if columns else ""


def _guess_country_col(columns: list[str]) -> str:
    """Return best country column, or '' if none found (caller defaults to '(use default)')."""
    cols_lower = {c.lower(): c for c in columns}
    for g in _COUNTRY_GUESSES:
        match = cols_lower.get(g.lower())
        if match and match.lower() not in _NOT_COUNTRY_COLS:
            return match
    return ""  # → sidebar will show "(use default)"

# ---------------------------------------------------------------------------
# Session-state keys
# ---------------------------------------------------------------------------

_KEY_RESULTS    = "hq_probe_results"
_KEY_INPUT_ROWS = "hq_probe_input_rows"
_KEY_OLD_COLS   = "hq_probe_old_cols"
_KEY_META       = "hq_probe_meta"
_KEY_COLS_CFG   = "hq_probe_cols_cfg"

# ---------------------------------------------------------------------------
# Sidebar
# ---------------------------------------------------------------------------

with st.sidebar:
    st.header("Input")

    upload_tab, path_tab = st.tabs(["Upload file", "Local path"])

    uploaded_file  = None
    local_path_str = ""

    with upload_tab:
        uploaded_file = st.file_uploader(
            "Excel (.xlsx) or CSV",
            type=["xlsx", "csv"],
            help="Upload a file with company names and domains.",
        )

    with path_tab:
        local_path_str = st.text_input(
            "Path on disk",
            placeholder=r"C:\Users\...\input.xlsx",
            help="Use this for large files already on your machine.",
        )

    # Resolve file source
    file_source: str | None = None
    file_suffix  = ""
    file_label   = ""
    sheet_names: list[str] = []
    selected_sheet: str | None = None

    if uploaded_file is not None:
        file_suffix = Path(uploaded_file.name).suffix
        file_label  = uploaded_file.name
        if file_suffix.lower() in (".xlsx", ".xls"):
            try:
                sheet_names = get_excel_sheet_names(uploaded_file)
                uploaded_file.seek(0)
            except Exception:
                sheet_names = []
        file_source = "upload"

    elif local_path_str.strip():
        p = Path(local_path_str.strip())
        if p.exists():
            file_suffix = p.suffix
            file_label  = p.name
            if file_suffix.lower() in (".xlsx", ".xls"):
                try:
                    sheet_names = get_excel_sheet_names(p)
                except Exception:
                    sheet_names = []
            file_source = "path"
        else:
            st.warning(f"Path not found: {p}")

    if sheet_names and len(sheet_names) > 1:
        selected_sheet = st.selectbox("Sheet", sheet_names)
    elif sheet_names:
        selected_sheet = sheet_names[0]

    st.header("Columns")

    peeked_headers: list[str] = []
    peek_error = ""

    if file_source == "upload" and uploaded_file is not None:
        try:
            uploaded_file.seek(0)
            preview_rows = read_input_from_fileobj(
                uploaded_file, file_suffix, limit=1, sheet_name=selected_sheet
            )
            peeked_headers = list(preview_rows[0].keys()) if preview_rows else []
            uploaded_file.seek(0)
        except Exception as exc:
            peek_error = str(exc)

    elif file_source == "path":
        try:
            with open(local_path_str.strip(), "rb") as f:
                preview_rows = read_input_from_fileobj(
                    f, file_suffix, limit=1, sheet_name=selected_sheet
                )
            peeked_headers = list(preview_rows[0].keys()) if preview_rows else []
        except Exception as exc:
            peek_error = str(exc)

    if peek_error:
        st.warning(f"Could not read headers: {peek_error}")

    if peeked_headers:
        company_col = st.selectbox(
            "Company column",
            peeked_headers,
            index=peeked_headers.index(_guess_col(peeked_headers, _COMPANY_GUESSES)),
        )
        domain_col = st.selectbox(
            "Domain column",
            peeked_headers,
            index=peeked_headers.index(_guess_col(peeked_headers, _DOMAIN_GUESSES)),
        )
        # Country column: "(use default)" is always the safe first option
        country_opts = ["(use default)"] + peeked_headers
        country_guess = _guess_country_col(peeked_headers)
        country_guess_idx = (
            country_opts.index(country_guess)
            if country_guess and country_guess in country_opts
            else 0  # default to "(use default)" when no country column found
        )
        country_col_sel = st.selectbox(
            "Input country column (optional)",
            country_opts,
            index=country_guess_idx,
            help="Select the column with ISO country or country name. "
                 "If not in file, choose '(use default)' and set the fallback below.",
        )
        # Use "" as sentinel for "no column" so run logic uses default_country
        country_col = country_col_sel if country_col_sel != "(use default)" else ""
    else:
        company_col = st.text_input("Company column", DEFAULT_COMPANY_COL)
        domain_col  = st.text_input("Domain column",  DEFAULT_DOMAIN_COL)
        country_col = ""  # no file loaded yet, use default

    default_country = st.text_input(
        "Default country (fallback)",
        DEFAULT_INPUT_COUNTRY,
        help="Used for every row when no country column is selected or when the column is blank.",
    )

    st.header("Options")

    limit = st.number_input("Row limit", min_value=1, max_value=2000, value=50, step=50)

    only_fhq_signal = st.checkbox(
        "Only rows with old foreign HQ signal",
        value=False,
        help="Filter input to rows where sig_foreign_hq_score > 0 or foreign_hq_sanitized = True/Yes.",
    )

    only_recovery = st.checkbox(
        "Only recovery candidates (zero/blank old FHQ score)",
        value=False,
        help="Focus on companies where sig_foreign_hq_score is 0 or blank — likely under-scored.",
    )

    use_simple_hq_mode = st.checkbox(
        "Use simple HQ mode",
        value=True,
        help=(
            "When enabled, runs exactly one Serper query per row: "
            "{domain_root} headquarters.  No company-name queries, no page fetches, "
            "no brand-root ladder, no Anthropic."
        ),
    )

    run_mode = st.radio(
        "Run mode (legacy — only active when simple HQ mode is off)",
        options=["fast", "deep", "debug"],
        index=0,
        help=(
            "**Fast** (default): ≤3 Serper calls/row, 1 page fetch, 2 brand queries, 5 s timeout. "
            "Stop immediately when official HQ evidence is clear.\n\n"
            "**Deep**: ≤8 Serper calls/row, up to 7 page paths, 4 brand queries, 10 s timeout. "
            "Enables Anthropic review for ambiguous cases.\n\n"
            "**Debug**: same as Deep but keeps all audit detail visible in the results table."
        ),
        horizontal=True,
        disabled=use_simple_hq_mode,
    )

    use_mimic_check = True  # always on; mode controls depth

    use_multilingual_check = st.checkbox(
        "Detect multilingual website",
        value=(run_mode != "fast"),
        help="Fetches company homepage to detect language switchers and hreflang tags. Disabled by default in Fast mode.",
    )

    st.header("API keys")

    serper_key = st.text_input(
        "Serper API key",
        value=os.environ.get("SERPER_API_KEY", "") or os.environ.get("SERPER_KEY", ""),
        type="password",
        help="Required for live searches. Get one at serper.dev.",
    )

    use_model = st.checkbox(
        "Use model fallback for unresolved cases",
        value=False,
        help="Calls Claude Haiku for companies where pattern matching finds nothing.",
    )

    use_anthropic_review = st.checkbox(
        "Use Anthropic HQ review for ambiguous/global cases",
        value=(run_mode in ("deep", "debug")),
        help="Calls Claude to adjudicate hq_structure_type for rows with needs_anthropic_hq_review=Yes. Auto-enabled in Deep/Debug mode.",
    )

    use_haiku_uncertain = st.checkbox(
        "Use Haiku for uncertain HQ cases",
        value=False,
        help="Only reviews ambiguous score-3 candidates using existing Serper evidence. No extra web search.",
        key="use_haiku_uncertain_cb",
    )
    haiku_max_calls = 25
    if use_haiku_uncertain:
        haiku_max_calls = st.number_input(
            "Max Haiku calls per run",
            min_value=1,
            max_value=200,
            value=25,
            step=5,
            help="Limits total Haiku API calls per batch run.",
            key="haiku_max_calls_input",
        )

    anthropic_key = ""
    if use_model or use_anthropic_review or use_haiku_uncertain:
        anthropic_key = st.text_input(
            "Anthropic API key",
            value=os.environ.get("ANTHROPIC_API_KEY", ""),
            type="password",
        )

# ---------------------------------------------------------------------------
# Main area
# ---------------------------------------------------------------------------

if not file_source:
    st.info("Upload a file or enter a local path in the sidebar to get started.")
    st.stop()

_mode_max = 3 if run_mode == "fast" else 8
_mode_label = {"fast": "Fast", "deep": "Deep", "debug": "Debug"}.get(run_mode, run_mode)
st.markdown(
    f"**Mode: {_mode_label}** — up to **{_mode_max} Serper calls/row**"
    + (", + 1 website fetch for multilingual detection" if use_multilingual_check else "")
    + (", + Anthropic review for ambiguous rows" if use_anthropic_review else "")
    + f". With limit={int(limit)}, that is up to **{int(limit) * _mode_max:,} Serper calls**."
    + (" Early stopping is active: rows with clear official HQ evidence skip further queries." if run_mode == "fast" else "")
)

run_btn = st.button("▶ Run HQ Probe", type="primary", disabled=(not serper_key), key="hq_probe_run_button_main")
if not serper_key:
    st.warning("Enter a Serper API key in the sidebar to enable the run button.")

if run_btn:
    # Load input rows
    with st.spinner("Reading input file…"):
        try:
            if file_source == "upload":
                uploaded_file.seek(0)
                input_rows = read_input_from_fileobj(
                    uploaded_file, file_suffix, limit=int(limit), sheet_name=selected_sheet,
                )
            else:
                with open(local_path_str.strip(), "rb") as f:
                    input_rows = read_input_from_fileobj(
                        f, file_suffix, limit=int(limit), sheet_name=selected_sheet,
                    )
        except Exception as exc:
            st.error(f"Failed to read input: {exc}")
            st.stop()

    if not input_rows:
        st.warning("No rows found in the input file.")
        st.stop()

    # Old-FHQ signal filter
    if only_fhq_signal:
        def _has_fhq_signal(row: dict) -> bool:
            score = row.get("sig_foreign_hq_score")
            sanitized = str(row.get("foreign_hq_sanitized") or "").strip().lower()
            try:
                score_val = float(score)
            except (TypeError, ValueError):
                score_val = 0.0
            return score_val > 0 or sanitized in {"true", "yes", "1"}

        filtered = [r for r in input_rows if _has_fhq_signal(r)]
        if filtered:
            st.info(f"Old FHQ filter: {len(filtered)} / {len(input_rows)} rows have a signal.")
            input_rows = filtered
        else:
            st.warning("No rows matched the old FHQ signal filter. Running on all rows.")

    # Recovery candidates filter (zero/blank FHQ score)
    if only_recovery:
        def _is_recovery_candidate(row: dict) -> bool:
            try:
                return float(row.get("sig_foreign_hq_score") or 0) == 0
            except (TypeError, ValueError):
                return True

        filtered_r = [r for r in input_rows if _is_recovery_candidate(r)]
        if filtered_r:
            st.info(f"Recovery filter: {len(filtered_r)} / {len(input_rows)} rows have zero/blank FHQ score.")
            input_rows = filtered_r
        else:
            st.warning("No recovery candidates found (all rows have non-zero FHQ score). Running on all rows.")

    # Detect old enrichment cols
    sample_keys = set(input_rows[0].keys()) if input_rows else set()
    present_old_cols = [c for c in OLD_ENRICHMENT_COLS if c in sample_keys]

    # Run probe
    progress_bar = st.progress(0.0, text="Starting…")
    probe_results: list[dict] = []
    total = len(input_rows)
    cache: dict = {}
    error_rows: list[str] = []
    _haiku_calls_counter: list[int] = [0]  # mutable counter shared across rows

    for i, row in enumerate(input_rows):
        company = str(row.get(company_col) or "").strip()
        domain  = str(row.get(domain_col)  or "").strip()
        # Country: use column if selected and non-blank, else default
        if country_col:
            country = str(row.get(country_col) or "").strip() or default_country
        else:
            country = default_country

        probe = probe_company(
            company_name=company,
            domain=domain,
            input_country=country,
            serper_key=serper_key,
            use_model=use_model,
            anthropic_key=anthropic_key,
            cache=cache,
            input_row=row,
            use_multilingual_check=use_multilingual_check,
            use_anthropic_review=use_anthropic_review,
            use_mimic_check=use_mimic_check,
            run_mode=run_mode,
            use_haiku_uncertain=use_haiku_uncertain,
            haiku_calls_counter=_haiku_calls_counter,
            haiku_max_calls=haiku_max_calls,
        )

        # Belt-and-suspenders fallback (main guard is inside probe_company).
        _fb_det = _normalize_country_for_hq(probe.get("hq_detected_country"))
        _fb_inp = _normalize_country_for_hq(probe.get("input_country_used"))
        if _fb_det and _fb_inp and _fb_det == _fb_inp:
            probe["foreign_hq_simple"] = "False"
            probe["sig_foreign_hq_score_reviewed"] = 0
            probe["sig_foreign_hq_score_for_next_scoring"] = 0

        probe_results.append(probe)
        if probe.get("probe_error"):
            error_rows.append(f"Row {i+1} ({company}): {probe['probe_error']}")

        pct = (i + 1) / total
        country_hit = probe.get("hq_detected_country") or "…"
        progress_bar.progress(pct, text=f"[{i+1}/{total}] {company[:45]} → {country_hit}")

    progress_bar.empty()

    # Run-level usage aggregation
    run_usage = {
        "rows_processed":           len(probe_results),
        "serper_calls_used":        sum(int(p.get("serper_calls_used") or 0) for p in probe_results),
        "rows_with_cache_hit":      sum(1 for p in probe_results if p.get("serper_cache_hit") in ("True", "partial")),
        "website_fetches":          sum(int(p.get("website_fetch_count") or 0) for p in probe_results),
        "anthropic_reviews_used":   sum(1 for p in probe_results if p.get("anthropic_hq_review_used") == "Yes"),
        "anthropic_web_search_used":sum(1 for p in probe_results if p.get("anthropic_web_search_used") == "Yes"),
        "anthropic_input_tokens":   sum(int(p.get("_anthr_input_tok") or 0) for p in probe_results),
        "anthropic_output_tokens":  sum(int(p.get("_anthr_output_tok") or 0) for p in probe_results),
        "anthropic_total_tokens":   sum(int(p.get("_anthr_total_tok") or 0) for p in probe_results),
        "anthropic_errors":         sum(1 for p in probe_results if p.get("anthropic_error")),
    }

    # Store in session state
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    st.session_state[_KEY_RESULTS]    = probe_results
    st.session_state[_KEY_INPUT_ROWS] = input_rows
    st.session_state[_KEY_OLD_COLS]   = present_old_cols
    st.session_state[_KEY_COLS_CFG]   = (company_col, domain_col, country_col or "(default)")
    st.session_state[_KEY_META] = {
        "timestamp":              ts,
        "input_file":             file_label,
        "use_model":              use_model,
        "model":                  "claude-haiku-4-5-20251001" if (use_model or use_anthropic_review) else "",
        "serper_available":       bool(serper_key),
        "limit":                  int(limit),
        "use_multilingual_check": use_multilingual_check,
        "use_anthropic_review":   use_anthropic_review,
        "use_mimic_check":        use_mimic_check,
        "run_mode":               run_mode,
    }
    st.session_state["hq_probe_run_usage"] = run_usage

    if error_rows:
        with st.expander(f"⚠️ {len(error_rows)} row error(s)", expanded=False):
            for e in error_rows:
                st.text(e)

# ---------------------------------------------------------------------------
# Workflow mode selector
# ---------------------------------------------------------------------------

if "hq_app_mode" not in st.session_state:
    st.session_state["hq_app_mode"] = "probe"

_wf_c1, _wf_c2 = st.columns(2)
if _wf_c1.button(
    "🔍 HQ Probe",
    use_container_width=True,
    type="primary" if st.session_state["hq_app_mode"] == "probe" else "secondary",
):
    st.session_state["hq_app_mode"] = "probe"
    st.rerun()
if _wf_c2.button(
    "🔄 HQ Recovery",
    use_container_width=True,
    type="primary" if st.session_state["hq_app_mode"] == "recovery" else "secondary",
):
    st.session_state["hq_app_mode"] = "recovery"
    st.rerun()
st.markdown("---")

_app_mode = st.session_state["hq_app_mode"]

# ---------------------------------------------------------------------------
# HQ PROBE workflow (existing)
# ---------------------------------------------------------------------------
if _app_mode == "probe":
    _mode_max = 3 if run_mode == "fast" else 8
    _mode_label = {"fast": "Fast", "deep": "Deep", "debug": "Debug"}.get(run_mode, run_mode)
    if use_simple_hq_mode:
        st.warning("Simple mode uses 1 Serper call per row when a domain is present.")
        st.markdown(
            f"**Simple HQ mode** — 1 Serper call/row (domain-root query). "
            f"With limit={int(limit)}, that is up to **{int(limit):,} Serper calls**."
        )
    else:
        st.markdown(
            f"**Mode: {_mode_label}** — up to **{_mode_max} Serper calls/row**"
            + (", + 1 website fetch for multilingual detection" if use_multilingual_check else "")
            + (", + Anthropic review for ambiguous rows" if use_anthropic_review else "")
            + f". With limit={int(limit)}, that is up to **{int(limit) * _mode_max:,} Serper calls**."
            + (" Early stopping is active: rows with clear official HQ evidence skip further queries." if run_mode == "fast" else "")
        )

    run_btn = st.button("▶ Run HQ Probe", type="primary", disabled=(not serper_key), key="hq_probe_run_button_tab")
    if not serper_key:
        st.warning("Enter a Serper API key in the sidebar to enable the run button.")

    if run_btn:
        # Load input rows
        with st.spinner("Reading input file…"):
            try:
                if file_source == "upload":
                    uploaded_file.seek(0)
                    input_rows = read_input_from_fileobj(
                        uploaded_file, file_suffix, limit=int(limit), sheet_name=selected_sheet,
                    )
                else:
                    with open(local_path_str.strip(), "rb") as f:
                        input_rows = read_input_from_fileobj(
                            f, file_suffix, limit=int(limit), sheet_name=selected_sheet,
                        )
            except Exception as exc:
                st.error(f"Failed to read input: {exc}")
                st.stop()

        if not input_rows:
            st.warning("No rows found in the input file.")
            st.stop()

        # Old-FHQ signal filter
        if only_fhq_signal:
            def _has_fhq_signal(row: dict) -> bool:
                score = row.get("sig_foreign_hq_score")
                sanitized = str(row.get("foreign_hq_sanitized") or "").strip().lower()
                try:
                    score_val = float(score)
                except (TypeError, ValueError):
                    score_val = 0.0
                return score_val > 0 or sanitized in {"true", "yes", "1"}

            filtered = [r for r in input_rows if _has_fhq_signal(r)]
            if filtered:
                st.info(f"Old FHQ filter: {len(filtered)} / {len(input_rows)} rows have a signal.")
                input_rows = filtered
            else:
                st.warning("No rows matched the old FHQ signal filter. Running on all rows.")

        # Recovery candidates filter (zero/blank FHQ score)
        if only_recovery:
            def _is_recovery_candidate(row: dict) -> bool:
                try:
                    return float(row.get("sig_foreign_hq_score") or 0) == 0
                except (TypeError, ValueError):
                    return True

            filtered_r = [r for r in input_rows if _is_recovery_candidate(r)]
            if filtered_r:
                st.info(f"Recovery filter: {len(filtered_r)} / {len(input_rows)} rows have zero/blank FHQ score.")
                input_rows = filtered_r
            else:
                st.warning("No recovery candidates found (all rows have non-zero FHQ score). Running on all rows.")

        # Detect old enrichment cols
        sample_keys = set(input_rows[0].keys()) if input_rows else set()
        present_old_cols = [c for c in OLD_ENRICHMENT_COLS if c in sample_keys]

        # Run probe
        progress_bar = st.progress(0.0, text="Starting…")
        probe_results: list[dict] = []
        total = len(input_rows)
        cache: dict = {}
        error_rows: list[str] = []

        for i, row in enumerate(input_rows):
            company = str(row.get(company_col) or "").strip()
            domain  = str(row.get(domain_col)  or "").strip()
            # Country: use column if selected and non-blank, else default
            if country_col:
                country = str(row.get(country_col) or "").strip() or default_country
            else:
                country = default_country

            probe = probe_company(
                company_name=company,
                domain=domain,
                input_country=country,
                serper_key=serper_key,
                use_model=use_model,
                anthropic_key=anthropic_key,
                cache=cache,
                input_row=row,
                use_multilingual_check=use_multilingual_check,
                use_anthropic_review=use_anthropic_review,
                use_mimic_check=use_mimic_check,
                run_mode=run_mode,
                use_simple_hq_mode=use_simple_hq_mode,
                use_haiku_uncertain=use_haiku_uncertain,
                haiku_calls_counter=_haiku_calls_counter,
                haiku_max_calls=haiku_max_calls,
            )

            # Sanity guard: if detected country == input country, force foreign_hq_simple = False
            det = _std_country(probe.get("hq_detected_country") or "")
            inp = _std_country(probe.get("input_country_used") or "")
            if det and inp and det.lower() == inp.lower():
                probe["foreign_hq_simple"] = "False"

            probe_results.append(probe)
            if probe.get("probe_error"):
                error_rows.append(f"Row {i+1} ({company}): {probe['probe_error']}")

            pct = (i + 1) / total
            country_hit = probe.get("hq_detected_country") or "…"
            progress_bar.progress(pct, text=f"[{i+1}/{total}] {company[:45]} → {country_hit}")

        progress_bar.empty()

        # Run-level usage aggregation
        run_usage = {
            "rows_processed":           len(probe_results),
            "serper_calls_used":        sum(int(p.get("serper_calls_used") or 0) for p in probe_results),
            "rows_with_cache_hit":      sum(1 for p in probe_results if p.get("serper_cache_hit") in ("True", "partial")),
            "website_fetches":          sum(int(p.get("website_fetch_count") or 0) for p in probe_results),
            "anthropic_reviews_used":   sum(1 for p in probe_results if p.get("anthropic_hq_review_used") == "Yes"),
            "anthropic_web_search_used":sum(1 for p in probe_results if p.get("anthropic_web_search_used") == "Yes"),
            "anthropic_input_tokens":   sum(int(p.get("_anthr_input_tok") or 0) for p in probe_results),
            "anthropic_output_tokens":  sum(int(p.get("_anthr_output_tok") or 0) for p in probe_results),
            "anthropic_total_tokens":   sum(int(p.get("_anthr_total_tok") or 0) for p in probe_results),
            "anthropic_errors":         sum(1 for p in probe_results if p.get("anthropic_error")),
        }

        # Store in session state
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        st.session_state[_KEY_RESULTS]    = probe_results
        st.session_state[_KEY_INPUT_ROWS] = input_rows
        st.session_state[_KEY_OLD_COLS]   = present_old_cols
        st.session_state[_KEY_COLS_CFG]   = (company_col, domain_col, country_col or "(default)")
        st.session_state[_KEY_META] = {
            "timestamp":              ts,
            "input_file":             file_label,
            "use_model":              use_model,
            "model":                  "claude-haiku-4-5-20251001" if (use_model or use_anthropic_review) else "",
            "serper_available":       bool(serper_key),
            "limit":                  int(limit),
            "use_multilingual_check": use_multilingual_check,
            "use_anthropic_review":   use_anthropic_review,
            "use_mimic_check":        use_mimic_check,
            "run_mode":               run_mode,
            "use_simple_hq_mode":     use_simple_hq_mode,
        }
        st.session_state["hq_probe_run_usage"] = run_usage

        if error_rows:
            with st.expander(f"⚠️ {len(error_rows)} row error(s)", expanded=False):
                for e in error_rows:
                    st.text(e)


# ---------------------------------------------------------------------------
# HQ RECOVERY workflow  ("Opportunity Input Full HQ Recovery")
# ---------------------------------------------------------------------------
elif _app_mode == "recovery":
    _REC_TARGET_SHEET = "Opportunity Input Full"
    _KEY_REC_RESULTS  = "hq_recovery_results"
    _KEY_REC_ALL_ROWS = "hq_recovery_all_rows"
    _KEY_REC_META     = "hq_recovery_meta"

    st.subheader("Opportunity Input Full — HQ Recovery")
    st.caption(
        "Reads the full workbook, selects under-scored rows, runs simple HQ lookup, "
        "and writes back a revised workbook."
    )

    # ── Sheet selection ─────────────────────────────────────────────────────
    _rec_sheet = _REC_TARGET_SHEET
    if sheet_names:
        if _REC_TARGET_SHEET in sheet_names:
            st.success(f"Sheet '{_REC_TARGET_SHEET}' detected.")
        else:
            st.warning(
                f"Sheet '{_REC_TARGET_SHEET}' not found. "
                f"Available sheets: {sheet_names}. Using first sheet."
            )
            _rec_sheet = sheet_names[0]

    # ── Threshold setting ───────────────────────────────────────────────────
    _rec_threshold = st.number_input(
        "Commercial fit score threshold (high-score HQ-zero candidates)",
        min_value=0.0, max_value=100.0, value=5.0, step=0.5,
        help="Rows with commercial_fit_score ≥ threshold AND sig_foreign_hq_score = 0 are selected.",
    )

    # ── Load ALL rows from the target sheet ─────────────────────────────────
    _rec_all_rows: list[dict] = []
    _rec_load_err = ""
    try:
        if file_source == "upload":
            uploaded_file.seek(0)
            _rec_all_rows = read_input_from_fileobj(
                uploaded_file, file_suffix, limit=50000, sheet_name=_rec_sheet
            )
            uploaded_file.seek(0)
        else:
            with open(local_path_str.strip(), "rb") as _f:
                _rec_all_rows = read_input_from_fileobj(
                    _f, file_suffix, limit=50000, sheet_name=_rec_sheet
                )
    except Exception as _exc:
        _rec_load_err = str(_exc)

    if _rec_load_err:
        st.error(f"Failed to read '{_rec_sheet}': {_rec_load_err}")
        st.stop()

    if not _rec_all_rows:
        st.warning(f"No rows found in sheet '{_rec_sheet}'.")
        st.stop()

    # ── Selection logic ─────────────────────────────────────────────────────
    def _rec_fhq_score(row: dict) -> float:
        try:
            return float(row.get("sig_foreign_hq_score") or 0)
        except (TypeError, ValueError):
            return 0.0

    def _rec_is_sanitized(row: dict) -> bool:
        """sig_foreign_hq_score = 0 AND (sanitized=True OR original_score=3 OR reason non-empty)."""
        if _rec_fhq_score(row) != 0:
            return False
        sanitized = str(row.get("foreign_hq_sanitized") or "").strip().lower()
        try:
            orig = float(row.get("foreign_hq_original_score") or 0)
        except (TypeError, ValueError):
            orig = 0.0
        reason = str(row.get("foreign_hq_sanitizer_reason") or "").strip()
        return sanitized in ("true", "yes", "1") or orig == 3 or bool(reason)

    def _rec_is_highscore_hq_zero(row: dict, threshold: float) -> bool:
        """sig_foreign_hq_score = 0 AND commercial_fit_score >= threshold."""
        if _rec_fhq_score(row) != 0:
            return False
        try:
            comm = float(
                row.get("commercial_fit_score")
                or row.get("final_commercial_fit_score")
                or 0
            )
        except (TypeError, ValueError):
            comm = 0.0
        return comm >= threshold

    _rec_sanitized_idx   = [i for i, r in enumerate(_rec_all_rows) if _rec_is_sanitized(r)]
    _rec_highscore_idx   = [i for i, r in enumerate(_rec_all_rows)
                             if _rec_is_highscore_hq_zero(r, _rec_threshold)]
    _rec_selected_idx    = sorted(set(_rec_sanitized_idx) | set(_rec_highscore_idx))
    _rec_selected_full   = len(_rec_selected_idx)

    # Apply row limit (uses the same sidebar `limit` variable as the probe tab)
    _rec_row_limit = int(limit) if limit and int(limit) > 0 else 0
    if _rec_row_limit > 0:
        _rec_to_process_idx = _rec_selected_idx[:_rec_row_limit]
    else:
        _rec_to_process_idx = _rec_selected_idx
    _rec_to_process_set  = set(_rec_to_process_idx)
    _rec_skipped_idx     = [i for i in _rec_selected_idx if i not in _rec_to_process_set]
    _rec_unchanged_idx   = [i for i in range(len(_rec_all_rows)) if i not in set(_rec_selected_idx)]

    # ── Pre-run counts ──────────────────────────────────────────────────────
    _rc1, _rc2, _rc3, _rc4 = st.columns(4)
    _rc1.metric("Total rows",               len(_rec_all_rows))
    _rc2.metric("Sanitized candidates",     len(_rec_sanitized_idx))
    _rc3.metric("High-score HQ-zero",       len(_rec_highscore_idx))
    _rc4.metric("Selected (unique)",        _rec_selected_full)
    _rc5, _rc6, _rc7 = st.columns(3)
    _rc5.metric("Rows to process",          len(_rec_to_process_idx))
    _rc6.metric("Skipped by row limit",     len(_rec_skipped_idx))
    _rc7.metric("Rows left unchanged",      len(_rec_unchanged_idx))

    if _rec_row_limit > 0 and _rec_selected_full > _rec_row_limit:
        st.caption(
            f"Row limit {_rec_row_limit} applied — processing first {len(_rec_to_process_idx)} "
            f"of {_rec_selected_full} selected rows. Adjust 'Row limit' in the sidebar to change."
        )

    if not _rec_to_process_idx:
        st.info("No rows match the selection criteria. Adjust the threshold or check the input data.")
        st.stop()

    # ── Run button ──────────────────────────────────────────────────────────
    _rec_run_btn = st.button(
        f"▶ Run HQ Recovery ({len(_rec_to_process_idx)} rows)",
        type="primary",
        disabled=(not serper_key), key="hq_recovery_run_button",
    )
    if not serper_key:
        st.warning("Enter a Serper API key in the sidebar to enable the run button.")

    if _rec_run_btn:
        _rec_sample_keys = set(_rec_all_rows[0].keys()) if _rec_all_rows else set()
        _rec_present_old = [c for c in OLD_ENRICHMENT_COLS if c in _rec_sample_keys]

        _rec_progress = st.progress(0.0, text="Starting HQ Recovery…")
        _rec_probe_map: dict[int, dict] = {}  # row_index → probe result
        _rec_cache: dict = {}
        _rec_errors: list[str] = []
        _rec_total = len(_rec_to_process_idx)

        for _rec_i, _rec_row_idx in enumerate(_rec_to_process_idx):
            _rec_row     = _rec_all_rows[_rec_row_idx]
            _rec_company = str(_rec_row.get(company_col) or "").strip()
            _rec_domain  = str(_rec_row.get(domain_col)  or "").strip()
            _rec_country = (
                str(_rec_row.get(country_col) or "").strip()
                if country_col else default_country
            ) or default_country

            _rec_probe = probe_company(
                company_name=_rec_company,
                domain=_rec_domain,
                input_country=_rec_country,
                serper_key=serper_key,
                use_model=False,
                anthropic_key="",
                cache=_rec_cache,
                input_row=_rec_row,
                use_multilingual_check=False,
                use_anthropic_review=False,
                use_mimic_check=False,
                run_mode="fast",
                use_simple_hq_mode=True,
            )

            # ── Domestic-safety guard (must run before score columns are set) ──
            # Uses _normalize_country_for_hq so IT/ITA/Italia/Italy all match.
            # Two triggers:
            # 1. norm(hq_detected_country) == norm(input_country)
            # 2. Evidence quotes mention Italian location but detected country is foreign
            _rec_det_display = _std_country(_rec_probe.get("hq_detected_country") or "")
            _rec_inp_display = _std_country(_rec_probe.get("input_country_used") or _rec_country or "")
            _rec_det_norm = _normalize_country_for_hq(_rec_probe.get("hq_detected_country"))
            _rec_inp_norm = _normalize_country_for_hq(
                _rec_probe.get("input_country_used") or _rec_country
            )
            _rec_ev_quotes = " ".join(filter(None, [
                _rec_probe.get("hq_evidence_quote", ""),
                _rec_probe.get("domain_root_hq_evidence_quote", ""),
                _rec_probe.get("brand_root_hq_evidence_quote", ""),
                _rec_probe.get("official_page_hq_evidence_quote", ""),
                _rec_probe.get("sig_foreign_hq_review_evidence_quote", ""),
            ]))
            _REC_ITALY_TOKENS = re.compile(
                r"\b(?:Italy|Italia|Italian|Modena|Ancona|Osimo|Senigallia|"
                r"San\s+Benedetto|Nusco|Avellino|Venezia|Venice|Milano|Milan|"
                r"Roma|Rome|Torino|Turin|Firenze|Florence|Bologna|Napoli|Naples|"
                r"Bergamo|Brescia|Padova|Verona|Perugia|Bari|Palermo|Cagliari|"
                r"Macerata|Pesaro|Jesi|Fabriano|Civitanova)\b",
                re.IGNORECASE,
            )
            _rec_quote_has_italy = bool(_REC_ITALY_TOKENS.search(_rec_ev_quotes))
            _rec_quote_overrides = (
                _rec_inp_norm == "italy"
                and _rec_quote_has_italy
                and _rec_det_norm
                and _rec_det_norm != "italy"
            )
            _rec_is_domestic = (
                (_rec_det_norm and _rec_inp_norm and _rec_det_norm == _rec_inp_norm)
                or _rec_quote_overrides
            )
            if _rec_is_domestic:
                _dom = _rec_inp_display or "Italy"
                _rec_probe["foreign_hq_simple"]             = "False"
                _rec_probe["hq_structure_type"]             = "domestic"
                _rec_probe["sig_foreign_hq_score_reviewed"] = 0
                _rec_probe["parent_group_hq_country"]       = _dom
                _rec_probe["parent_group_hq_city"]          = _rec_probe.get("hq_detected_city", "")
                _rec_probe["review_foreign_parent_score"]   = 0
                _rec_probe["review_global_network_score"]   = 0
                _rec_probe["hq_detected_country"]           = _dom
                if _rec_probe.get("hq_confidence") in ("High", "Medium"):
                    _rec_probe["needs_manual_review"] = "No"
                _guard_reason = (
                    "quote_contains_italy_location_overrides_foreign"
                    if _rec_quote_overrides
                    else f"hq_detected={_rec_det_display}({_rec_det_norm})==input={_rec_inp_display}({_rec_inp_norm})"
                )
                _rec_probe["sig_foreign_hq_review_reason"] = (
                    "domestic_hq_matches_input_country"
                    f" [{_guard_reason}]"
                )

            # Recovery-specific columns
            _rec_probe["hq_recovery_selected"]           = "Yes"
            _rec_probe["hq_recovery_processed"]          = "Yes"
            _rec_probe["hq_recovery_skip_reason"]        = ""
            _rec_probe["hq_recovery_selection_reason"]   = (
                "sanitized_candidate" if _rec_row_idx in set(_rec_sanitized_idx)
                else "high_score_hq_zero"
            )
            _rec_probe["sig_foreign_hq_score_original_before_recovery"] = (
                _rec_row.get("sig_foreign_hq_score", "")
            )
            # Score for next scoring: use reviewed score (already 0 if domestic-guard fired)
            _rec_reviewed = _rec_probe.get("sig_foreign_hq_score_reviewed", 0)
            _rec_trusted  = (
                str(_rec_probe.get("needs_manual_review", "")).lower() != "yes"
            )
            _rec_probe["sig_foreign_hq_score_for_next_scoring"] = (
                int(_rec_reviewed) if _rec_trusted else
                _rec_row.get("sig_foreign_hq_score", 0)
            )
            # Competitor signal: exclude from next scoring
            _rec_probe["competitor_signal_excluded_from_next_scoring"] = "True"

            _rec_probe_map[_rec_row_idx] = _rec_probe

            if _rec_probe.get("probe_error"):
                _rec_errors.append(
                    f"Row {_rec_row_idx+1} ({_rec_company}): {_rec_probe['probe_error']}"
                )

            _rec_pct     = (_rec_i + 1) / _rec_total
            _rec_country_hit = _rec_probe.get("hq_detected_country") or "…"
            _rec_progress.progress(
                _rec_pct,
                text=f"[{_rec_i+1}/{_rec_total}] {_rec_company[:45]} → {_rec_country_hit}",
            )

        _rec_progress.empty()

        # ── Store in session state ──────────────────────────────────────────
        _rec_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        st.session_state[_KEY_REC_RESULTS]  = _rec_probe_map
        st.session_state[_KEY_REC_ALL_ROWS] = _rec_all_rows
        st.session_state[_KEY_REC_META]     = {
            "timestamp":        _rec_ts,
            "input_file":       file_label,
            "sheet":            _rec_sheet,
            "threshold":        _rec_threshold,
            "row_limit":        _rec_row_limit,
            "selected_full":    _rec_selected_full,
            "selected":         len(_rec_to_process_idx),  # actually processed
            "skipped":          len(_rec_skipped_idx),
            "total":            len(_rec_all_rows),
            "sanitized":        len(_rec_sanitized_idx),
            "highscore":        len(_rec_highscore_idx),
            "unchanged":        len(_rec_unchanged_idx),
        }
        # Store index sets so _build_recovery_rows can mark skipped rows
        st.session_state["hq_recovery_to_process_set"] = _rec_to_process_set
        st.session_state["hq_recovery_skipped_set"]    = set(_rec_skipped_idx)
        st.session_state["hq_recovery_selected_full"]  = set(_rec_selected_idx)

        if _rec_errors:
            with st.expander(f"⚠️ {len(_rec_errors)} row error(s)", expanded=False):
                for _e in _rec_errors:
                    st.text(_e)

    # ── Show recovery results (persist across reruns) ───────────────────────
    if _KEY_REC_RESULTS not in st.session_state:
        st.stop()

    _rec_probe_map_r: dict[int, dict] = st.session_state[_KEY_REC_RESULTS]
    _rec_all_rows_r: list[dict]       = st.session_state[_KEY_REC_ALL_ROWS]
    _rec_meta_r: dict                 = st.session_state[_KEY_REC_META]
    _rec_ts_r = _rec_meta_r.get("timestamp", "")

    # Summary metrics
    _rec_updated_to_3 = sum(
        1 for p in _rec_probe_map_r.values()
        if str(p.get("sig_foreign_hq_score_for_next_scoring", "")) == "3"
    )
    _rec_needs_review = sum(
        1 for p in _rec_probe_map_r.values()
        if p.get("needs_manual_review") == "Yes"
    )

    st.markdown("---")
    st.subheader("Recovery Results")
    _rr1, _rr2, _rr3, _rr4 = st.columns(4)
    _rr1.metric("Total rows",              _rec_meta_r.get("total", 0))
    _rr2.metric("Rows processed",          _rec_meta_r.get("selected", 0))
    if _rec_meta_r.get("skipped", 0):
        st.caption(
            f"Row limit {_rec_meta_r.get('row_limit', '—')} was applied: "
            f"{_rec_meta_r.get('skipped', 0)} selected rows were skipped "
            f"(hq_recovery_processed=No, hq_recovery_skip_reason=row_limit)."
        )
    _rr3.metric("Updated to score 3",      _rec_updated_to_3)
    _rr4.metric("Needs manual review",     _rec_needs_review)

    # ── Build revised rows (all rows, only processed rows get probe output) ──
    _rec_skipped_set_r  = st.session_state.get("hq_recovery_skipped_set", set())
    _rec_selected_all_r = st.session_state.get("hq_recovery_selected_full", set())

    def _build_recovery_rows() -> list[dict]:
        out = []
        _rec_probe_cols_set = set(PROBE_COLS + [
            "hq_recovery_selected",
            "hq_recovery_processed",
            "hq_recovery_skip_reason",
            "hq_recovery_selection_reason",
            "sig_foreign_hq_score_original_before_recovery",
            "sig_foreign_hq_score_for_next_scoring",
            "competitor_signal_excluded_from_next_scoring",
        ])
        for idx, row in enumerate(_rec_all_rows_r):
            merged = dict(row)
            if idx in _rec_probe_map_r:
                # Processed row — merge probe output
                probe = _rec_probe_map_r[idx]
                for k, v in probe.items():
                    if k in _rec_probe_cols_set and not k.startswith("_"):
                        merged[k] = v
                # Preserve original competitor columns
                for col_name in row:
                    if "competitor" in col_name.lower():
                        merged[col_name] = row[col_name]
            elif idx in _rec_skipped_set_r:
                # Selected but skipped due to row limit
                merged["hq_recovery_selected"]    = "Yes"
                merged["hq_recovery_processed"]   = "No"
                merged["hq_recovery_skip_reason"] = "row_limit"
            else:
                # Not selected at all
                merged["hq_recovery_selected"]    = "No"
                merged["hq_recovery_processed"]   = "No"
                merged["hq_recovery_skip_reason"] = ""
            out.append(merged)
        return out

    _rec_revised_rows = _build_recovery_rows()

    # ── Display table ───────────────────────────────────────────────────────
    _rec_display_cols = [
        company_col, domain_col,
        "hq_recovery_selected", "hq_recovery_processed", "hq_recovery_skip_reason",
        "hq_recovery_selection_reason",
        "sig_foreign_hq_score_original_before_recovery",
        "sig_foreign_hq_score_for_next_scoring",
        "sig_foreign_hq_score_reviewed",
        "hq_detected_country", "parent_group_hq_country",
        "hq_structure_type", "needs_manual_review",
        "sig_foreign_hq_review_evidence_quote",
        "serper_queries_used", "serper_calls_used",
        "competitor_signal_excluded_from_next_scoring",
    ]

    try:
        import pandas as _pd
        _rec_df_rows = [
            {c: r.get(c, "") for c in _rec_display_cols if c}
            for r in _rec_revised_rows
            if r.get("hq_recovery_selected") == "Yes"
        ]
        if _rec_df_rows:
            _rec_df = _pd.DataFrame(_rec_df_rows)
            _rec_present_cols = [c for c in _rec_display_cols if c and c in _rec_df.columns]
            st.dataframe(_rec_df[_rec_present_cols], use_container_width=True, height=400)
    except ImportError:
        pass

    # ── Export revised workbook ─────────────────────────────────────────────
    st.markdown("---")
    st.subheader("Download revised workbook")

    @st.cache_data(show_spinner=False)
    def _make_recovery_excel(_key: int) -> bytes:
        """Build a workbook with the revised 'Opportunity Input Full' sheet + run summary."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # ── Sheet 1: revised Opportunity Input Full ─────────────────────────
        ws = wb.active
        ws.title = _REC_TARGET_SHEET

        rows = _rec_revised_rows
        if not rows:
            buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

        # Column order: original cols first, then new probe cols appended
        _orig_cols   = list(_rec_all_rows_r[0].keys()) if _rec_all_rows_r else []
        _extra_cols  = [
            "hq_recovery_selected", "hq_recovery_selection_reason",
            "sig_foreign_hq_score_original_before_recovery",
            "sig_foreign_hq_score_for_next_scoring",
            "sig_foreign_hq_score_reviewed",
            "hq_detected_country", "hq_detected_city",
            "parent_group_hq_country", "parent_group_hq_city",
            "hq_structure_type", "needs_manual_review",
            "hq_confidence", "foreign_hq_simple",
            "sig_foreign_hq_review_reason", "sig_foreign_hq_review_confidence",
            "sig_foreign_hq_review_evidence_url", "sig_foreign_hq_review_evidence_quote",
            "serper_queries_used", "serper_calls_used",
            "domain_root_hq_evidence_rank", "domain_root_hq_rejected_evidence_reason",
            "competitor_signal_excluded_from_next_scoring",
            "hq_review_trigger",
        ]
        _all_cols = _orig_cols + [c for c in _extra_cols if c not in _orig_cols]

        # Header row
        _hdr_fill = PatternFill("solid", fgColor="1F4E79")
        _hdr_font = Font(bold=True, color="FFFFFF")
        for col_i, col_name in enumerate(_all_cols, start=1):
            cell = ws.cell(row=1, column=col_i, value=col_name)
            cell.fill = _hdr_fill
            cell.font = _hdr_font
            cell.alignment = Alignment(wrap_text=False)
            ws.column_dimensions[get_column_letter(col_i)].width = max(12, min(len(col_name) + 2, 40))

        # Data rows
        _new_col_set = set(_extra_cols)
        _new_fill    = PatternFill("solid", fgColor="E2EFDA")
        for row_i, row_data in enumerate(rows, start=2):
            for col_i, col_name in enumerate(_all_cols, start=1):
                val = row_data.get(col_name, "")
                if val is True:  val = "True"
                if val is False: val = "False"
                cell = ws.cell(row=row_i, column=col_i, value=val)
                if col_name in _new_col_set and val not in ("", None):
                    cell.fill = _new_fill

        ws.freeze_panes = "A2"

        # ── Sheet 2: Run Summary ────────────────────────────────────────────
        ws2 = wb.create_sheet("Recovery Run Summary")
        _sum_data = [
            ("Input file",          _rec_meta_r.get("input_file", "")),
            ("Sheet",               _rec_meta_r.get("sheet", "")),
            ("Timestamp",           _rec_meta_r.get("timestamp", "")),
            ("Total rows",          _rec_meta_r.get("total", 0)),
            ("Sanitized candidates",_rec_meta_r.get("sanitized", 0)),
            ("High-score HQ-zero",  _rec_meta_r.get("highscore", 0)),
            ("Selected (full)",     _rec_meta_r.get("selected_full", _rec_meta_r.get("selected", 0))),
            ("Rows processed",      _rec_meta_r.get("selected", 0)),
            ("Rows skipped (limit)",_rec_meta_r.get("skipped", 0)),
            ("Row limit applied",   _rec_meta_r.get("row_limit", 0) or "none"),
            ("Rows unchanged",      _rec_meta_r.get("unchanged", 0)),
            ("Updated to score 3",  _rec_updated_to_3),
            ("Needs manual review", _rec_needs_review),
            ("Threshold",           _rec_meta_r.get("threshold", 5.0)),
        ]
        for r_i, (k, v) in enumerate(_sum_data, start=1):
            ws2.cell(row=r_i, column=1, value=k).font = Font(bold=True)
            ws2.cell(row=r_i, column=2, value=v)
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 40

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    _rec_dl_fname = f"hq_recovery_{_rec_ts_r}.xlsx"
    st.download_button(
        label="⬇️ Download revised workbook",
        data=_make_recovery_excel(id(_rec_revised_rows)),
        file_name=_rec_dl_fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help=f"Revised '{_REC_TARGET_SHEET}' sheet + Recovery Run Summary. "
             f"Same number of rows as input ({len(_rec_all_rows_r)}).",
    key="hq_recovery_download_button",
    )
    st.caption(
        f"Output: **{_rec_dl_fname}** · "
        f"{len(_rec_all_rows_r)} rows (same as source) · "
        f"sheet: '{_REC_TARGET_SHEET}'"
    )


# Show results (persists across reruns via session state)
# ---------------------------------------------------------------------------

if _KEY_RESULTS not in st.session_state:
    st.stop()

probe_results: list[dict]   = st.session_state[_KEY_RESULTS]
input_rows: list[dict]      = st.session_state[_KEY_INPUT_ROWS]
present_old_cols: list[str] = st.session_state[_KEY_OLD_COLS]
company_col_r, domain_col_r, country_col_r = st.session_state[_KEY_COLS_CFG]
qa_meta: dict               = st.session_state[_KEY_META]

# Summary metrics
detected_italy   = sum(1 for p in probe_results if _std_country(p.get("hq_detected_country") or "") == "Italy")
detected_foreign = sum(1 for p in probe_results if p.get("foreign_hq_simple") == "True")
detected_unknown = sum(1 for p in probe_results if not p.get("hq_detected_country"))
needs_review_cnt = sum(1 for p in probe_results if p.get("needs_manual_review") == "Yes")

needs_adj_cnt    = sum(1 for p in probe_results if p.get("needs_anthropic_hq_review") == "Yes")
multilingual_cnt = sum(1 for p in probe_results if p.get("has_multilingual_site"))
global_net_cnt   = sum(1 for p in probe_results if p.get("has_global_structure_signal"))
run_usage        = st.session_state.get("hq_probe_run_usage", {})

st.markdown("---")
st.subheader("Results")
c1, c2, c3, c4, c5 = st.columns(5)
c1.metric("Rows processed",    len(probe_results))
c2.metric("Italy HQ detected", detected_italy)
c3.metric("Foreign HQ",        detected_foreign)
c4.metric("Unknown",           detected_unknown)
c5.metric("Needs review",      needs_review_cnt)

r1, r2, r3 = st.columns(3)
r1.metric("Multilingual site",       multilingual_cnt)
r2.metric("Global structure signal", global_net_cnt)
r3.metric("Needs Anthropic review",  needs_adj_cnt)

if run_usage:
    with st.expander("Usage / API call summary", expanded=False):
        u1, u2, u3, u4, u5 = st.columns(5)
        u1.metric("Serper calls",              run_usage.get("serper_calls_used", 0))
        u2.metric("Rows with cache hit",       run_usage.get("rows_with_cache_hit", 0))
        u3.metric("Website fetches",           run_usage.get("website_fetches", 0))
        u4.metric("Anthropic reviews used",    run_usage.get("anthropic_reviews_used", 0))
        u5.metric("Anthropic web search used", run_usage.get("anthropic_web_search_used", 0))
        t1, t2, t3, t4 = st.columns(4)
        t1.metric("Anthropic input tokens",    run_usage.get("anthropic_input_tokens", 0))
        t2.metric("Anthropic output tokens",   run_usage.get("anthropic_output_tokens", 0))
        t3.metric("Anthropic total tokens",    run_usage.get("anthropic_total_tokens", 0))
        t4.metric("Anthropic errors",          run_usage.get("anthropic_errors", 0))

# Display columns
_KEY_VIEW_COLS_RAW = [
    company_col_r, domain_col_r, country_col_r,
    "sig_foreign_hq_score_original",
    "sig_foreign_hq_score_reviewed",
    "sig_foreign_hq_score", "sig_foreign_hq_evidence",
    "foreign_hq_sanitized", "foreign_hq_sanitizer_reason",
    "hq_detected_city", "hq_detected_region", "hq_detected_country",
    "input_country_used",
    "hq_confidence", "foreign_hq_simple", "needs_manual_review",
    "hq_structure_type",
    "parent_group_hq_country", "parent_group_hq_city",
    "local_entity_hq_country", "local_entity_hq_city",
    "has_multilingual_site", "website_languages_detected",
    "has_global_structure_signal", "global_structure_evidence",
    "hq_review_trigger",
    "needs_anthropic_hq_review", "anthropic_hq_review_used",
    "anthropic_web_search_used", "anthropic_review_evidence_mode",
    "anthropic_error",
    "sig_foreign_hq_review_reason",
    "sig_foreign_hq_review_evidence_url",
    "run_mode", "early_stop_used", "early_stop_reason", "early_stop_blocked_reason",
    "serper_calls_used", "serper_cache_hit",
    "manual_google_mimic_used",
    "plain_company_official_domain", "official_domain_matches_input_domain",
    "official_top_result_is_local_country_page", "official_top_result_local_page_reason",
    "official_page_hq_evidence_found", "official_page_hq_evidence_strength",
    "official_page_hq_country", "official_page_hq_city",
    "brand_root_candidates",
    "brand_root_hq_search_used", "brand_root_hq_search_queries",
    "brand_root_hq_evidence_found", "brand_root_hq_evidence_quote",
    "brand_root_hq_evidence_url", "brand_root_hq_country", "brand_root_hq_city",
    "known_global_brand_detected", "known_global_brand_name", "known_global_brand_type",
    "brand_hq_evidence_found", "brand_hq_evidence_quote",
    "hq_reason", "hq_evidence_url", "hq_evidence_quote", "hq_query_used",
]
_seen_kvc: set[str] = set()
_KEY_VIEW_COLS: list[str] = []
for _c in _KEY_VIEW_COLS_RAW:
    if _c and _c not in _seen_kvc:
        _KEY_VIEW_COLS.append(_c)
        _seen_kvc.add(_c)


def _build_display_rows(
    input_rows: list[dict],
    probe_results: list[dict],
    cols: list[str],
) -> list[dict]:
    rows = []
    for i, (in_row, probe) in enumerate(zip(input_rows, probe_results), start=1):
        r: dict[str, Any] = {"#": i}
        for c in cols:
            r[c] = probe[c] if c in probe else in_row.get(c, "")
        rows.append(r)
    return rows


# Filters
st.markdown("**Filters**")
f1, f2, f3, f4 = st.columns(4)
f_review   = f1.checkbox("Needs manual review only")
f_foreign  = f2.checkbox("Foreign HQ only (new)")
f_unknown  = f3.checkbox("Unknown only")
f_disagree = f4.checkbox("Old/new disagreement")

all_display = _build_display_rows(input_rows, probe_results, _KEY_VIEW_COLS)


def _apply_filters(rows: list[dict]) -> list[dict]:
    out = rows
    if f_review:
        out = [r for r in out if r.get("needs_manual_review") == "Yes"]
    if f_foreign:
        out = [r for r in out if r.get("foreign_hq_simple") == "True"]
    if f_unknown:
        out = [r for r in out if not r.get("hq_detected_country")]
    if f_disagree:
        def _disagree(r: dict) -> bool:
            try:
                old_score = float(r.get("sig_foreign_hq_score") or 0)
            except (ValueError, TypeError):
                old_score = 0.0
            return old_score > 0 and r.get("foreign_hq_simple", "") in ("False", "")
        out = [r for r in out if _disagree(r)]
    return out


filtered_display = _apply_filters(all_display)
st.caption(f"Showing {len(filtered_display)} of {len(all_display)} rows")

# Only include columns that have data; deduplicate
_pvc_seen: set[str] = {"#"}
present_view_cols = ["#"]
for _c in _KEY_VIEW_COLS:
    if _c not in _pvc_seen and any(str(r.get(_c, "")).strip() for r in filtered_display):
        present_view_cols.append(_c)
        _pvc_seen.add(_c)

try:
    import pandas as pd
    df = pd.DataFrame(filtered_display)[present_view_cols]
    st.dataframe(df, use_container_width=True, height=420)
except ImportError:
    st.table(filtered_display)

# Downloads
st.markdown("---")
st.subheader("Download results")

dl1, dl2 = st.columns(2)

with dl1:
    @st.cache_data(show_spinner=False)
    def _make_excel(_input_key: int, _probe_key: int) -> bytes:
        return build_excel_bytes(
            input_rows=input_rows,
            probe_results=probe_results,
            present_old_cols=present_old_cols,
            company_col=company_col_r,
            domain_col=domain_col_r,
            country_col=country_col_r,
            qa_meta=qa_meta,
            run_usage=run_usage,
        )

    ts = qa_meta.get("timestamp", "")
    dl1.download_button(
        label="⬇️ Download Excel",
        data=_make_excel(id(input_rows), id(probe_results)),
        file_name=f"hq_probe_results_{ts}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    key="hq_probe_download_excel_button",
    )

with dl2:
    def _make_csv() -> bytes:
        if not all_display:
            return b""
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=list(all_display[0].keys()), extrasaction="ignore")
        writer.writeheader()
        writer.writerows(all_display)
        return buf.getvalue().encode("utf-8-sig")

    dl2.download_button(
        label="⬇️ Download CSV",
        data=_make_csv(),
        file_name=f"hq_probe_results_{ts}.csv",
        mime="text/csv",
    key="hq_probe_download_csv_button",
    )
