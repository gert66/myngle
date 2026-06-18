"""
patch_enriched_urls_from_cleaned.py

Copies corrected website_url values from a cleaned register file back into
existing enriched files for Italy100 and Italy200.

Default target sheet: "Opportunity Input" only.
Use --patch-all-sheets to also patch every sheet that contains company_name.
Use --target-sheets to specify a custom list of sheet names.

Usage:
  python patch_enriched_urls_from_cleaned.py \
      --source-cleaned "<path>" \
      --target-root "<path>" \
      --queues Italy100 Italy200 \
      --dry-run | --apply \
      [--target-sheets "Opportunity Input" "Best Guess Input"] \
      [--patch-all-sheets] \
      [--in-place] [--overwrite] [--report-dir "<path>"] \
      [--match-threshold strict] [--only-changed]
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
import unicodedata
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

# ---------------------------------------------------------------------------
# Optional dependency check
# ---------------------------------------------------------------------------
try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

PREFERRED_SOURCE_SHEETS = ["Best Guess Input", "Cleaned Register Input"]

DEFAULT_TARGET_SHEETS = ["Opportunity Input"]

ITALIAN_LEGAL_SUFFIXES = [
    r"s\.?p\.?a\.?",
    r"s\.?r\.?l\.?",
    r"societa[\s]*per[\s]*azioni",
    r"societa[\s]*a[\s]*responsabilita[\s]*limitata",
    r"in[\s]*forma[\s]*abbreviata",
    r"siglabile",
    r"s\.?n\.?c\.?",
    r"s\.?a\.?s\.?",
    r"s\.?c\.?a\.?r\.?l\.?",
    r"s\.?c\.?",
    r"onlus",
    r"ets",
    r"aps",
    r"asd",
    r"spa",
    r"srl",
    r"snc",
    r"sas",
    r"scarl",
]

URL_COLUMNS = [
    "website_url",
    "canonical_company_url",
]

DOMAIN_COLUMNS = [
    "domain",
    "validated_domain",
    "input_domain",
    "canonical_company_domain",
    "final_selected_domain",
    "recommended_domain",
]

AUDIT_COLUMNS = [
    "url_patch_status",
    "url_patch_old_website_url",
    "url_patch_new_website_url",
    "url_patch_old_domain",
    "url_patch_new_domain",
    "url_patch_match_key",
    "url_patch_source_sheet",
    "needs_reenrichment",
]

PROTECTED_COLUMNS = {
    "commercial_fit_score",
    "final_commercial_fit_score",
    "commercial_tier",
    "icp_evidence",
    "raw_google_evidence_json",
    "scoring_notes",
    "caller_angle",
}

STATUS_PATCHED_CHANGED = "PATCHED_URL_CHANGED"
STATUS_PATCHED_SAME = "PATCHED_URL_SAME"
STATUS_NO_MATCH = "NO_SOURCE_MATCH"
STATUS_SOURCE_EMPTY = "SOURCE_URL_EMPTY"
STATUS_DUPLICATE_SKIPPED = "DUPLICATE_SOURCE_MATCH_SKIPPED"
STATUS_TARGET_EMPTY = "TARGET_COMPANY_EMPTY"
STATUS_OPP_SHEET_MISSING = "OPPORTUNITY_INPUT_SHEET_MISSING"
STATUS_INVALID_WORKBOOK = "INVALID_WORKBOOK_SKIPPED"

# Folder name fragments that are never safe to read as input
UNSAFE_PATH_FRAGMENTS = [
    "_url_patched",
    "url_patched",
    "_url_patch_reports",
    "_logs",
    "_archive",
    "__pycache__",
]


# ---------------------------------------------------------------------------
# Normalisation helpers
# ---------------------------------------------------------------------------

def _remove_accents(text: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", text)
        if unicodedata.category(c) != "Mn"
    )


_LEGAL_RE = re.compile(
    r"\b(" + "|".join(ITALIAN_LEGAL_SUFFIXES) + r")\b",
    re.IGNORECASE,
)
_PUNCT_RE = re.compile(r"[^\w\s]", re.UNICODE)
_WS_RE = re.compile(r"\s+")


def normalize_company(name: Any) -> str:
    if not name or not str(name).strip():
        return ""
    s = str(name).strip()
    s = _remove_accents(s)
    s = s.lower()
    s = _LEGAL_RE.sub(" ", s)
    s = _PUNCT_RE.sub(" ", s)
    s = _WS_RE.sub(" ", s).strip()
    return s


def normalize_city(value: Any) -> str:
    if not value:
        return ""
    s = _remove_accents(str(value).strip().lower())
    s = _PUNCT_RE.sub(" ", s)
    return _WS_RE.sub(" ", s).strip()


# ---------------------------------------------------------------------------
# URL / domain helpers
# ---------------------------------------------------------------------------

def extract_domain(url_or_domain: Any) -> str:
    """Return bare domain from a full URL or plain domain string."""
    if not url_or_domain:
        return ""
    s = str(url_or_domain).strip()
    if not s:
        return ""
    if "://" not in s:
        s = "https://" + s
    parsed = urlparse(s)
    domain = parsed.netloc or parsed.path
    domain = re.sub(r"^www\.", "", domain.lower()).strip("/")
    return domain


def to_full_url(url_or_domain: Any) -> str:
    """Return https://<domain> from a full URL or plain domain string."""
    domain = extract_domain(url_or_domain)
    if not domain:
        return ""
    return f"https://{domain}"


# ---------------------------------------------------------------------------
# Source lookup builder
# ---------------------------------------------------------------------------

class SourceLookup:
    """
    Builds three-level lookup tables from the cleaned source file:
      primary   -> normalized company name
      secondary -> normalized company name + normalized city
      tertiary  -> normalized company name + normalized province
    """

    def __init__(self) -> None:
        self.primary: dict[str, str] = {}
        self.secondary: dict[str, str] = {}
        self.tertiary: dict[str, str] = {}
        self.duplicates: set[str] = set()
        self.source_sheet: str = ""
        self.total_rows = 0
        self.rows_with_url = 0

    def build(self, wb: openpyxl.Workbook) -> None:
        sheet = _pick_source_sheet(wb)
        self.source_sheet = sheet.title
        headers = _get_headers(sheet)

        col = _col_index(headers, "company_name")
        url_col = _col_index(headers, "website_url")
        city_col = _col_index(headers, "city")
        prov_col = _col_index(headers, "province")

        if col is None:
            sys.exit(f"Source sheet '{sheet.title}' has no 'company_name' column.")
        if url_col is None:
            sys.exit(f"Source sheet '{sheet.title}' has no 'website_url' column.")

        for row in sheet.iter_rows(min_row=2, values_only=True):
            self.total_rows += 1
            company_raw = row[col] if col < len(row) else None
            url_raw = row[url_col] if url_col < len(row) else None

            key = normalize_company(company_raw)
            if not key:
                continue

            url = str(url_raw).strip() if url_raw else ""
            if not url:
                continue

            self.rows_with_url += 1

            if city_col is not None and city_col < len(row):
                city = normalize_city(row[city_col])
                if city:
                    self.secondary[f"{key}||{city}"] = url

            if prov_col is not None and prov_col < len(row):
                prov = normalize_city(row[prov_col])
                if prov:
                    self.tertiary[f"{key}||prov||{prov}"] = url

            if key in self.primary and self.primary[key] != url:
                self.duplicates.add(key)
            else:
                self.primary[key] = url

    def lookup(
        self,
        company: Any,
        city: Any = None,
        province: Any = None,
    ) -> tuple[str, str, str]:
        """Returns (url, status, match_key). status: FOUND|DUPLICATE|NOT_FOUND|EMPTY"""
        key = normalize_company(company)
        if not key:
            return "", "EMPTY", ""

        if city:
            sec_key = f"{key}||{normalize_city(city)}"
            if sec_key in self.secondary:
                return self.secondary[sec_key], "FOUND", sec_key

        if province:
            ter_key = f"{key}||prov||{normalize_city(province)}"
            if ter_key in self.tertiary:
                return self.tertiary[ter_key], "FOUND", ter_key

        if key in self.duplicates:
            return "", "DUPLICATE", key

        if key in self.primary:
            return self.primary[key], "FOUND", key

        return "", "NOT_FOUND", key


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------

def _pick_source_sheet(wb: openpyxl.Workbook):
    for name in PREFERRED_SOURCE_SHEETS:
        if name in wb.sheetnames:
            return wb[name]
    return wb.active


def _get_headers(sheet) -> list[str]:
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
    return [str(h).strip() if h is not None else "" for h in first_row]


def _col_index(headers: list[str], name: str) -> int | None:
    try:
        return headers.index(name)
    except ValueError:
        return None


def _col_index_ci(headers: list[str], name: str) -> int | None:
    lo = name.lower()
    for i, h in enumerate(headers):
        if h.lower() == lo:
            return i
    return None


# ---------------------------------------------------------------------------
# Core patching logic (single sheet)
# ---------------------------------------------------------------------------

def patch_sheet(
    sheet,
    lookup: SourceLookup,
    dry_run: bool,
    only_changed: bool,
    source_sheet_name: str,
) -> list[dict]:
    """Patch URL/domain cells in a worksheet. Returns list of report rows."""
    headers = _get_headers(sheet)
    company_col = _col_index_ci(headers, "company_name")
    if company_col is None:
        return []

    city_col = _col_index_ci(headers, "city")
    prov_col = _col_index_ci(headers, "province")

    url_cols: dict[str, int] = {}
    for name in URL_COLUMNS:
        idx = _col_index_ci(headers, name)
        if idx is not None:
            url_cols[name] = idx

    domain_cols: dict[str, int] = {}
    for name in DOMAIN_COLUMNS:
        idx = _col_index_ci(headers, name)
        if idx is not None:
            domain_cols[name] = idx

    # Ensure audit columns exist (append at end if missing)
    audit_col_indices: dict[str, int] = {}
    for audit_col in AUDIT_COLUMNS:
        idx = _col_index_ci(headers, audit_col)
        if idx is None:
            new_col_idx = len(headers)
            headers.append(audit_col)
            if not dry_run:
                col_letter = get_column_letter(new_col_idx + 1)
                sheet[f"{col_letter}1"] = audit_col
            audit_col_indices[audit_col] = new_col_idx
        else:
            audit_col_indices[audit_col] = idx

    report_rows = []

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        company_raw = row[company_col].value if company_col < len(row) else None
        city_raw = row[city_col].value if city_col is not None and city_col < len(row) else None
        prov_raw = row[prov_col].value if prov_col is not None and prov_col < len(row) else None

        if not company_raw or not str(company_raw).strip():
            _set_audit(sheet, row_idx, audit_col_indices, {
                "url_patch_status": STATUS_TARGET_EMPTY,
                "needs_reenrichment": False,
                "url_patch_match_key": "",
                "url_patch_source_sheet": source_sheet_name,
                "url_patch_old_website_url": "",
                "url_patch_new_website_url": "",
                "url_patch_old_domain": "",
                "url_patch_new_domain": "",
            }, dry_run)
            report_rows.append({
                "company_name": "",
                "status": STATUS_TARGET_EMPTY,
                "old_website_url": "",
                "new_website_url": "",
                "old_domain": "",
                "new_domain": "",
                "match_key": "",
                "needs_reenrichment": False,
                "row": row_idx,
            })
            continue

        new_url, find_status, match_key = lookup.lookup(company_raw, city_raw, prov_raw)

        old_website_url = ""
        if "website_url" in url_cols:
            old_website_url = str(row[url_cols["website_url"]].value or "").strip()

        old_domain = ""
        for dc in DOMAIN_COLUMNS:
            if dc in domain_cols:
                v = str(row[domain_cols[dc]].value or "").strip()
                if v:
                    old_domain = v
                    break

        if find_status == "EMPTY":
            status = STATUS_TARGET_EMPTY
            new_url = ""
        elif find_status == "NOT_FOUND":
            status = STATUS_NO_MATCH
            new_url = ""
        elif find_status == "DUPLICATE":
            status = STATUS_DUPLICATE_SKIPPED
            new_url = ""
        else:
            if not new_url:
                status = STATUS_SOURCE_EMPTY
            else:
                new_domain = extract_domain(new_url)
                if new_domain == extract_domain(old_website_url) and new_domain:
                    status = STATUS_PATCHED_SAME
                else:
                    status = STATUS_PATCHED_CHANGED

        new_domain = extract_domain(new_url) if new_url else ""
        needs_reenrich = bool(new_domain and new_domain != extract_domain(old_website_url))

        if not dry_run and new_url and status in (STATUS_PATCHED_CHANGED, STATUS_PATCHED_SAME):
            if not only_changed or status == STATUS_PATCHED_CHANGED:
                full_url = to_full_url(new_url)
                domain_only = extract_domain(new_url)
                for col_idx in url_cols.values():
                    _set_cell(sheet, row_idx, col_idx, full_url)
                for col_idx in domain_cols.values():
                    _set_cell(sheet, row_idx, col_idx, domain_only)

        _set_audit(sheet, row_idx, audit_col_indices, {
            "url_patch_status": status,
            "url_patch_old_website_url": old_website_url,
            "url_patch_new_website_url": to_full_url(new_url) if new_url else "",
            "url_patch_old_domain": old_domain,
            "url_patch_new_domain": new_domain,
            "url_patch_match_key": match_key,
            "url_patch_source_sheet": source_sheet_name,
            "needs_reenrichment": needs_reenrich,
        }, dry_run)

        report_rows.append({
            "company_name": str(company_raw).strip(),
            "status": status,
            "old_website_url": old_website_url,
            "new_website_url": to_full_url(new_url) if new_url else "",
            "old_domain": old_domain,
            "new_domain": new_domain,
            "match_key": match_key,
            "needs_reenrichment": needs_reenrich,
            "row": row_idx,
        })

    return report_rows


def _set_cell(sheet, row_idx: int, col_idx: int, value: Any) -> None:
    col_letter = get_column_letter(col_idx + 1)
    sheet[f"{col_letter}{row_idx}"] = value


def _set_audit(sheet, row_idx: int, audit_indices: dict, values: dict, dry_run: bool) -> None:
    if dry_run:
        return
    for col_name, col_idx in audit_indices.items():
        v = values.get(col_name, "")
        _set_cell(sheet, row_idx, col_idx, v)


# ---------------------------------------------------------------------------
# File-level processing
# ---------------------------------------------------------------------------

def _resolve_output_path(output_dir: Path, filename: str, overwrite: bool) -> Path:
    stem = Path(filename).stem
    suffix = Path(filename).suffix
    candidate = output_dir / filename
    if not candidate.exists() or overwrite:
        return candidate
    n = 2
    while True:
        candidate = output_dir / f"{stem}_{n}{suffix}"
        if not candidate.exists():
            return candidate
        n += 1


def process_file(
    xlsx_path: Path,
    output_dir: Path,
    lookup: SourceLookup,
    dry_run: bool,
    in_place: bool,
    overwrite: bool,
    only_changed: bool,
    target_sheets: list[str],
    patch_all_sheets: bool,
) -> tuple[list[dict], bool]:
    """
    Process a single enriched file.
    Returns (report_rows, opportunity_sheet_found).
    Returns a sentinel INVALID_WORKBOOK_SKIPPED row on load failure.
    """
    try:
        wb = load_workbook(xlsx_path, data_only=False)
    except (zipfile.BadZipFile, KeyError, OSError, Exception) as exc:
        print(f"[patch] Skipping invalid workbook: {xlsx_path} | {exc}")
        return ([{
            "file": xlsx_path.name,
            "sheet": "",
            "row": "",
            "company_name": "",
            "status": STATUS_INVALID_WORKBOOK,
            "old_website_url": "",
            "new_website_url": "",
            "old_domain": "",
            "new_domain": "",
            "match_key": "",
            "needs_reenrichment": False,
            "notes": str(exc),
        }], False)
    all_report_rows = []
    opportunity_sheet_found = False

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        headers = _get_headers(sheet)
        company_col = _col_index_ci(headers, "company_name")

        if patch_all_sheets:
            # Patch any sheet that has company_name
            if company_col is None:
                continue
            should_patch = True
        else:
            # Only patch sheets in the target list
            should_patch = sheet_name in target_sheets and company_col is not None

        if sheet_name in target_sheets and company_col is not None:
            opportunity_sheet_found = True

        if not should_patch:
            continue

        rows = patch_sheet(sheet, lookup, dry_run, only_changed, lookup.source_sheet)
        for r in rows:
            r["file"] = xlsx_path.name
            r["sheet"] = sheet_name
        all_report_rows.extend(rows)

    if not opportunity_sheet_found and not patch_all_sheets:
        # Record one sentinel row per file so the report captures the gap
        all_report_rows.append({
            "file": xlsx_path.name,
            "sheet": "",
            "row": "",
            "company_name": "",
            "status": STATUS_OPP_SHEET_MISSING,
            "old_website_url": "",
            "new_website_url": "",
            "old_domain": "",
            "new_domain": "",
            "match_key": "",
            "needs_reenrichment": False,
            "notes": f"None of the target sheets {target_sheets} found in this workbook",
        })

    if not dry_run and all_report_rows:
        if in_place:
            wb.save(xlsx_path)
        else:
            output_dir.mkdir(parents=True, exist_ok=True)
            out_path = _resolve_output_path(output_dir, xlsx_path.name, overwrite)
            wb.save(out_path)

    return all_report_rows, opportunity_sheet_found


# ---------------------------------------------------------------------------
# Report writer
# ---------------------------------------------------------------------------

def write_report(report_rows: list[dict], report_dir: Path, queues: list[str]) -> Path:
    report_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    queues_str = "_".join(queues)
    report_path = report_dir / f"url_patch_report_{ts}_{queues_str}.csv"

    fieldnames = [
        "queue", "file", "sheet", "row", "company_name",
        "old_website_url", "new_website_url",
        "old_domain", "new_domain",
        "status", "needs_reenrichment", "match_key", "notes",
    ]

    with open(report_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in report_rows:
            writer.writerow({k: row.get(k, "") for k in fieldnames})

    return report_path


def print_summary(
    lookup: SourceLookup,
    report_rows: list[dict],
    target_sheets: list[str],
    files_with_opp_sheet: int,
    files_missing_opp_sheet: int,
) -> None:
    status_counts: dict[str, int] = {}
    for r in report_rows:
        s = r.get("status", "")
        status_counts[s] = status_counts.get(s, 0) + 1

    # Sheet-level breakdown for the primary target sheet
    opp_rows = [r for r in report_rows if r.get("sheet") in target_sheets]
    opp_status: dict[str, int] = {}
    for r in opp_rows:
        s = r.get("status", "")
        opp_status[s] = opp_status.get(s, 0) + 1

    needs_reenrich_opp = sum(1 for r in opp_rows if r.get("needs_reenrichment"))
    needs_reenrich_total = sum(1 for r in report_rows if r.get("needs_reenrichment"))

    print("\n" + "=" * 64)
    print("URL PATCH SUMMARY")
    print("=" * 64)
    print(f"  Source rows loaded               : {lookup.total_rows}")
    print(f"  Source rows with URL             : {lookup.rows_with_url}")
    print(f"  Duplicate source keys            : {len(lookup.duplicates)}")
    print()
    print(f"  Files with target sheet found    : {files_with_opp_sheet}")
    print(f"  Files missing target sheet       : {files_missing_opp_sheet}")
    print()
    ts_label = " / ".join(target_sheets)
    print(f"  [{ts_label}] sheet breakdown:")
    print(f"    Rows seen                      : {len(opp_rows)}")
    print(f"    Rows patched (URL changed)     : {opp_status.get(STATUS_PATCHED_CHANGED, 0)}")
    print(f"    Rows patched (URL same)        : {opp_status.get(STATUS_PATCHED_SAME, 0)}")
    print(f"    Rows no source match           : {opp_status.get(STATUS_NO_MATCH, 0)}")
    print(f"    Rows skipped (duplicate)       : {opp_status.get(STATUS_DUPLICATE_SKIPPED, 0)}")
    print(f"    Rows source URL empty          : {opp_status.get(STATUS_SOURCE_EMPTY, 0)}")
    print(f"    Rows target company empty      : {opp_status.get(STATUS_TARGET_EMPTY, 0)}")
    print(f"    Rows needing re-enrichment     : {needs_reenrich_opp}")
    print()
    print(f"  All sheets total rows seen       : {len(report_rows)}")
    print(f"  All sheets rows patched changed  : {status_counts.get(STATUS_PATCHED_CHANGED, 0)}")
    print(f"  All sheets needing re-enrichment : {needs_reenrich_total}")
    print(f"  Invalid workbooks skipped        : {status_counts.get(STATUS_INVALID_WORKBOOK, 0)}")
    print("=" * 64 + "\n")


# ---------------------------------------------------------------------------
# Queue scanning
# ---------------------------------------------------------------------------

def _is_safe_input_path(path: Path) -> bool:
    """Return False if any part of the path looks like a patched output folder."""
    parts = [p.lower() for p in path.parts]
    for fragment in UNSAFE_PATH_FRAGMENTS:
        for part in parts:
            if fragment in part:
                return False
    return True


def collect_xlsx_files(queue_dir: Path, subfolders: list[str]) -> list[tuple[Path, str]]:
    """
    Returns list of (xlsx_path, input_subfolder) for safe candidate files.
    Skips files located inside any patched output folder.
    """
    results = []
    for subfolder in subfolders:
        folder = queue_dir / subfolder
        if not folder.exists():
            continue
        if not _is_safe_input_path(folder):
            continue
        for f in sorted(folder.glob("*.xlsx")):
            if not f.name.startswith("~$") and _is_safe_input_path(f):
                results.append((f, subfolder))
    return results


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Patch website_url into enriched files from a cleaned register.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("--source-cleaned", required=True,
                   help="Path to the cleaned register .xlsx")
    p.add_argument("--target-root", required=True,
                   help="Root folder, e.g. C:\\Users\\...\\Myngle")
    p.add_argument("--queues", nargs="+", default=["Italy100", "Italy200"],
                   help="Queue folder names under target-root")
    p.add_argument("--dry-run", action="store_true",
                   help="Preview changes without writing any files (default when neither flag given)")
    p.add_argument("--apply", action="store_true",
                   help="Write patched files to output folders")
    p.add_argument("--target-sheets", nargs="+", default=DEFAULT_TARGET_SHEETS,
                   help=(
                       "Sheet names to patch (default: 'Opportunity Input'). "
                       "Example: --target-sheets \"Opportunity Input\" \"Best Guess Input\""
                   ))
    p.add_argument("--patch-all-sheets", action="store_true", default=False,
                   help="Patch every sheet that contains company_name, ignoring --target-sheets")
    p.add_argument("--in-place", action="store_true", default=False,
                   help="Overwrite original enriched files (dangerous; use with caution)")
    p.add_argument("--overwrite", action="store_true", default=False,
                   help="Overwrite existing patched output files instead of adding _2, _3 suffix")
    p.add_argument("--report-dir", default=None,
                   help="Directory for the CSV report (default: <target-root>\\_url_patch_reports)")
    p.add_argument("--match-threshold", default="strict", choices=["strict"],
                   help="Matching mode (currently: strict exact normalised match)")
    p.add_argument("--only-changed", action="store_true", default=False,
                   help="Only write cells where the URL actually changes")
    return p


def _output_subfolder_for(input_subfolder: str) -> str:
    """Map an input subfolder name to the safe output subfolder name."""
    mapping = {
        "02_lead_prioritized": "02_lead_prioritized_url_patched",
        "03_opportunity_radar": "03_opportunity_input_url_patched",
    }
    return mapping.get(input_subfolder, f"{input_subfolder}_url_patched")


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    if not args.dry_run and not args.apply:
        print("Neither --dry-run nor --apply specified. Defaulting to --dry-run.")
        args.dry_run = True

    if args.dry_run and args.apply:
        sys.exit("Pass either --dry-run or --apply, not both.")

    source_path = Path(args.source_cleaned)
    if not source_path.exists():
        sys.exit(f"Source file not found: {source_path}")

    target_root = Path(args.target_root)
    if not target_root.exists():
        sys.exit(f"Target root not found: {target_root}")

    report_dir = (
        Path(args.report_dir) if args.report_dir
        else target_root / "_url_patch_reports"
    )

    target_sheets: list[str] = args.target_sheets
    patch_all: bool = args.patch_all_sheets

    mode = "DRY-RUN" if args.dry_run else "APPLY"
    print(f"\n[{mode}] Loading source: {source_path.name}")

    source_wb = load_workbook(source_path, data_only=True, read_only=True)
    lookup = SourceLookup()
    lookup.build(source_wb)
    source_wb.close()

    print(f"  Source sheet   : '{lookup.source_sheet}'")
    print(f"  Rows loaded    : {lookup.total_rows}")
    print(f"  Rows with URL  : {lookup.rows_with_url}")
    print(f"  Duplicate keys : {len(lookup.duplicates)}")
    if lookup.duplicates:
        sample = sorted(lookup.duplicates)[:5]
        print(f"  Sample dupes   : {sample}")

    if patch_all:
        print("  Sheet mode     : ALL sheets with company_name (--patch-all-sheets)")
    else:
        print(f"  Target sheets  : {target_sheets}")

    # Input subfolders to scan per queue
    INPUT_SUBFOLDERS = [
        "02_lead_prioritized",
        "03_opportunity_radar",
    ]

    all_report_rows: list[dict] = []
    files_processed = 0
    files_with_opp_sheet = 0
    files_missing_opp_sheet = 0

    for queue in args.queues:
        queue_dir = target_root / queue
        if not queue_dir.exists():
            print(f"\n[WARN] Queue folder not found, skipping: {queue_dir}")
            continue

        # Warn if any output folders already exist (they will not be used as input)
        for out_sub in ["02_lead_prioritized_url_patched", "03_opportunity_input_url_patched"]:
            out_folder = queue_dir / out_sub
            if out_folder.exists():
                print(
                    f"  [INFO] Output folder already exists: {out_folder}\n"
                    f"         Existing files will not be used as input."
                )

        # Standalone opportunity input workbooks at queue root only (never from output folders)
        standalone_patterns = [
            f"{queue}_ALL_opportunity_input.xlsx",
            f"{queue}_opportunity_input.xlsx",
        ]
        standalone_files: list[tuple[Path, str]] = []
        for pat in standalone_patterns:
            candidate = queue_dir / pat
            if (
                candidate.exists()
                and not candidate.name.startswith("~$")
                and _is_safe_input_path(candidate)
            ):
                standalone_files.append((candidate, "03_opportunity_input_url_patched"))

        subfolder_files = collect_xlsx_files(queue_dir, INPUT_SUBFOLDERS)
        all_files = subfolder_files + standalone_files

        print(f"\nQueue: {queue}  |  {len(all_files)} file(s) found")

        for xlsx_path, input_subfolder in all_files:
            output_subfolder = _output_subfolder_for(input_subfolder)
            output_dir = queue_dir / output_subfolder

            print(f"  Processing [{input_subfolder}]: {xlsx_path.name}")
            rows, opp_found = process_file(
                xlsx_path=xlsx_path,
                output_dir=output_dir,
                lookup=lookup,
                dry_run=args.dry_run,
                in_place=args.in_place,
                overwrite=args.overwrite,
                only_changed=args.only_changed,
                target_sheets=target_sheets,
                patch_all_sheets=patch_all,
            )

            if opp_found:
                files_with_opp_sheet += 1
            else:
                files_missing_opp_sheet += 1
                print(f"    [WARN] Target sheet(s) {target_sheets} not found in this file")

            for r in rows:
                r.setdefault("queue", queue)
            all_report_rows.extend(rows)
            files_processed += 1

            opp_rows_this = [r for r in rows if r.get("sheet") in target_sheets]
            changed = sum(1 for r in opp_rows_this if r.get("status") == STATUS_PATCHED_CHANGED)
            no_match = sum(1 for r in opp_rows_this if r.get("status") == STATUS_NO_MATCH)
            print(
                f"    opp_rows={len(opp_rows_this)}"
                f"  changed={changed}"
                f"  no_match={no_match}"
                f"  opp_sheet={'YES' if opp_found else 'NO'}"
            )

    print(f"\nTotal files processed: {files_processed}")

    report_path = write_report(all_report_rows, report_dir, args.queues)
    print(f"Report written to   : {report_path}")

    print_summary(
        lookup,
        all_report_rows,
        target_sheets,
        files_with_opp_sheet,
        files_missing_opp_sheet,
    )

    if args.dry_run:
        print("Dry-run complete. No files were written.")
        print("Re-run with --apply to apply patches.\n")
    else:
        print("Apply complete.\n")


if __name__ == "__main__":
    main()
