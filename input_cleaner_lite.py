"""
input_cleaner_lite.py — Layer 0: mYngle Input Cleaner Lite
Validates and corrects company name / domain pairs before Lead Prioritizer enrichment.
No LLM calls. No scraping. Uses Serper for domain lookup when needed.
"""

import io
import re
import time
from pathlib import Path
from urllib.parse import urlparse

import pandas as pd
import requests
import streamlit as st

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Input Cleaner Lite",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── Constants ─────────────────────────────────────────────────────────────────
SERPER_URL = "https://google.serper.dev/search"

_GENERIC_DOMAINS: frozenset = frozenset({
    "linkedin.com", "facebook.com", "twitter.com", "x.com", "instagram.com",
    "youtube.com", "wikipedia.org", "bloomberg.com", "crunchbase.com",
    "glassdoor.com", "indeed.com", "xing.com", "angel.co", "pitchbook.com",
    "google.com", "bing.com", "yahoo.com", "reuters.com", "ft.com",
    "github.com", "amazon.com", "zoominfo.com", "dnb.com",
    "opencorporates.com", "companieshouse.gov.uk", "app.lusha.com",
    "rocketreach.co", "signalhire.com", "apollo.io", "hunter.io",
    "jobsite.co.uk", "totaljobs.com", "monster.com", "stepstone.de",
    "kununu.com", "trustpilot.com", "yelp.com",
})

_COMPANY_NAME_COLS = (
    "company_name", "company", "name", "account_name",
    "organization", "organisation", "business_name",
)
_DOMAIN_COLS = (
    "website_url", "website url", "website", "domain",
    "url", "company_url", "company url", "company_website", "company website",
    "homepage", "web_site", "web site", "site",
)


def _normalize_col_key(col: str) -> str:
    """Lowercase and collapse spaces/underscores/hyphens to a single space."""
    return re.sub(r"[\s_\-]+", " ", col.strip().lower())

# Legal suffixes to strip before token comparison
_LEGAL_TOKENS = re.compile(
    r"\b(ltd|limited|b\.?v\.?|n\.?v\.?|gmbh|ag|s\.?a\.?s?|spa|s\.?p\.?a\.?|"
    r"srl|s\.?r\.?l\.?|llc|inc|corp|corporation|plc|holding|holdings|"
    r"groep|group|co|company|pty|ges\.?mbh|kgaa|kg|og|ab|oy|as|asa|"
    r"se|pte|bhd|sarl|eurl|snc|scs|cv)\b\.?",
    re.IGNORECASE,
)

_NOISE_TOKENS: frozenset = frozenset({
    "the", "and", "for", "global", "international", "services", "solutions",
    "consulting", "management", "technology", "technologies", "systems",
    "software", "digital", "enterprise", "enterprises", "partners",
    "nederland", "netherlands", "deutschland", "germany", "france",
    "belgium", "europe", "european", "asia",
})

# TLDs to strip when extracting domain tokens
_TLDS: frozenset = frozenset({
    "com", "net", "org", "nl", "de", "fr", "be", "uk", "co", "io",
    "biz", "info", "eu", "at", "ch", "es", "it", "pl", "cz", "se",
    "no", "dk", "fi", "pt", "hu", "ro", "hr", "gr", "gov", "edu",
})

# Row colours for Excel (openpyxl ARGB hex)
_ACTION_COLORS = {
    "OK":                   "C6EFCE",  # green
    "LIKELY_OK":            "E2EFDA",  # light green
    "REVIEW":               "FFEB9C",  # yellow
    "SUGGEST_REPLACE":      "FCE4D6",  # orange
    "MISSING_DOMAIN_FIXED": "FCE4D6",  # orange
    "NO_CONFIDENT_MATCH":   "FFC7CE",  # red
    "MISSING_DOMAIN":       "FFC7CE",  # red
}


# ── Domain & name helpers ─────────────────────────────────────────────────────
def normalize_domain(raw: str) -> str:
    """Strip protocol, www, path, query. Return root domain lowercase."""
    if not raw or not isinstance(raw, str):
        return ""
    d = raw.strip().lower()
    d = re.sub(r"^https?://", "", d)
    d = re.sub(r"^www\.", "", d)
    d = d.split("/")[0].split("?")[0].split("#")[0].strip()
    if not d or " " in d or d in ("nan", "none"):
        return ""
    return d


def strip_legal(name: str) -> str:
    cleaned = _LEGAL_TOKENS.sub(" ", name)
    return re.sub(r"\s+", " ", cleaned).strip(" .,/-")


def company_tokens(name: str) -> set:
    clean = strip_legal(name)
    clean = re.sub(r"[^\w\s\-]", " ", clean)
    toks = {t.lower() for t in re.split(r"[\s\-_]+", clean) if len(t) >= 2}
    return toks - _NOISE_TOKENS


def domain_tokens(domain: str) -> set:
    if not domain:
        return set()
    parts = domain.split(".")
    while len(parts) > 1 and parts[-1].lower() in _TLDS:
        parts = parts[:-1]
    base = ".".join(parts)
    toks = {t for t in re.split(r"[-.]", base.lower()) if t and len(t) >= 2}
    return toks - _TLDS


def token_overlap(name: str, domain: str) -> float:
    """Return overlap ratio [0,1] with substring containment bonus."""
    ctok = company_tokens(name)
    dtok = domain_tokens(domain)
    if not ctok or not dtok:
        return 0.0
    overlap: set = ctok & dtok
    for c in ctok:
        for d in dtok:
            if c in d or d in c:
                overlap.add(c)
    return len(overlap) / min(len(ctok), len(dtok))


def is_generic(domain: str) -> bool:
    return domain.lower() in _GENERIC_DOMAINS


# ── Column detection ──────────────────────────────────────────────────────────
def detect_columns(df: pd.DataFrame) -> tuple[str | None, str | None]:
    # Map normalised key → original column name (both sides normalised)
    cols_norm = {_normalize_col_key(c): c for c in df.columns}
    name_col = next(
        (cols_norm[_normalize_col_key(k)] for k in _COMPANY_NAME_COLS
         if _normalize_col_key(k) in cols_norm), None
    )
    domain_col = next(
        (cols_norm[_normalize_col_key(k)] for k in _DOMAIN_COLS
         if _normalize_col_key(k) in cols_norm), None
    )
    return name_col, domain_col


# ── Serper search ─────────────────────────────────────────────────────────────
_SEARCH_QUERY_TEMPLATES = [
    '"{name}" official website',
    '"{name}" company website',
    '"{name}" official site',
]


def _call_serper(query: str, serper_key: str, timeout: int = 12) -> tuple:
    """POST one query to Serper. Returns (organic_results, error_str_or_None)."""
    try:
        resp = requests.post(
            SERPER_URL,
            headers={"X-API-KEY": serper_key, "Content-Type": "application/json"},
            json={"q": query, "gl": "us", "hl": "en", "num": 5},
            timeout=timeout,
        )
        resp.raise_for_status()
        data = resp.json()
        return data.get("organic", []), None
    except requests.Timeout:
        return [], "Serper timeout"
    except Exception as e:
        return [], str(e)


def _extract_domain(url: str) -> str:
    try:
        p = urlparse(url if url.startswith("http") else f"https://{url}")
        host = p.hostname or ""
        return re.sub(r"^www\.", "", host.lower())
    except Exception:
        return ""


def search_official_domain(
    company_name: str, serper_key: str
) -> tuple[str, float, str, list, str]:
    """
    Run up to 3 Serper queries for company_name.
    Returns (suggested_domain, confidence, reason, evidence_rows, query_used).
    confidence: 0.0–1.0
    """
    candidates: dict[str, int] = {}  # domain → vote count
    evidence: list[dict] = []
    query_used = ""

    for tmpl in _SEARCH_QUERY_TEMPLATES:
        query = tmpl.format(name=company_name)
        results, err = _call_serper(query, serper_key)
        if err:
            break
        if not query_used:
            query_used = query
        for item in results:
            url  = item.get("link", "")
            title = item.get("title", "")
            snippet = item.get("snippet", "")
            domain = _extract_domain(url)
            if not domain or is_generic(domain):
                evidence.append({
                    "query": query, "title": title,
                    "url": url, "domain": domain, "used": False,
                })
                continue
            overlap = token_overlap(company_name, domain)
            if overlap >= 0.15:
                candidates[domain] = candidates.get(domain, 0) + 1
            evidence.append({
                "query": query, "title": title,
                "url": url, "domain": domain,
                "overlap": round(overlap, 3), "used": overlap >= 0.15,
            })
        time.sleep(0.3)

    if not candidates:
        return "", 0.0, "No candidate domain found in search results.", evidence, query_used

    # Pick highest-voted; break ties by overlap
    best = max(candidates, key=lambda d: (candidates[d], token_overlap(company_name, d)))
    votes = candidates[best]
    overlap = token_overlap(company_name, best)

    if votes >= 2 and overlap >= 0.4:
        conf, reason = 0.85, "Domain appeared in multiple queries with strong name match."
    elif votes >= 2 or overlap >= 0.4:
        conf, reason = 0.65, "Domain appeared in search results with reasonable name match."
    elif overlap >= 0.2:
        conf, reason = 0.45, "Domain found but name match is weak. Review recommended."
    else:
        conf, reason = 0.25, "Domain found but name-domain overlap is very low."

    return best, conf, reason, evidence, query_used


# ── Core validation ───────────────────────────────────────────────────────────
def validate_row(
    company_name: str,
    raw_domain: str,
    serper_key: str | None,
) -> dict:
    """Validate one (company_name, domain) pair. Returns result fields dict."""
    name = str(company_name or "").strip()
    norm = normalize_domain(raw_domain)

    result = {
        "original_domain":           str(raw_domain or "").strip(),
        "normalized_input_domain":   norm,
        "validated_domain":          norm,
        "recommended_domain":        "",
        "domain_action":             "",
        "domain_confidence":         "",
        "domain_reason":             "",
        "manual_review_needed":      False,
        "serper_top_result_title":   "",
        "serper_top_result_url":     "",
        "serper_top_result_domain":  "",
        "search_query_used":         "",
    }

    if not name:
        result.update(
            domain_action="NO_CONFIDENT_MATCH",
            domain_confidence="None",
            domain_reason="Company name is blank.",
            manual_review_needed=True,
        )
        return result

    # ── Case 1: domain missing ────────────────────────────────────────────────
    if not norm:
        if serper_key:
            suggested, conf, reason, ev, query = search_official_domain(name, serper_key)
            _fill_serper_top(result, ev, query)
            if suggested and conf >= 0.45:
                result.update(
                    validated_domain=suggested,
                    recommended_domain=suggested,
                    domain_action="MISSING_DOMAIN_FIXED",
                    domain_confidence=_conf_label(conf),
                    domain_reason=reason,
                    manual_review_needed=True,
                )
            else:
                result.update(
                    validated_domain="",
                    domain_action="MISSING_DOMAIN",
                    domain_confidence="None",
                    domain_reason="Domain missing and no confident result found in search.",
                    manual_review_needed=True,
                )
        else:
            result.update(
                validated_domain="",
                domain_action="MISSING_DOMAIN",
                domain_confidence="None",
                domain_reason="Domain missing. No Serper key available to search.",
                manual_review_needed=True,
            )
        return result

    # ── Case 2: domain is a known generic/directory site ─────────────────────
    if is_generic(norm):
        if serper_key:
            suggested, conf, reason, ev, query = search_official_domain(name, serper_key)
            _fill_serper_top(result, ev, query)
            if suggested and conf >= 0.45:
                result.update(
                    validated_domain=suggested,
                    recommended_domain=suggested,
                    domain_action="SUGGEST_REPLACE",
                    domain_confidence=_conf_label(conf),
                    domain_reason=f"Input domain is a generic/directory site ({norm}). {reason}",
                    manual_review_needed=True,
                )
                return result
        result.update(
            domain_action="REVIEW",
            domain_confidence="Low",
            domain_reason=f"Input domain ({norm}) is a generic/directory site.",
            manual_review_needed=True,
        )
        return result

    # ── Case 3: compare tokens ───────────────────────────────────────────────
    overlap = token_overlap(name, norm)

    if overlap >= 0.5:
        result.update(
            domain_action="OK",
            domain_confidence="High",
            domain_reason="Company name tokens match domain with high confidence.",
            manual_review_needed=False,
        )
        return result

    if overlap >= 0.2:
        result.update(
            domain_action="LIKELY_OK",
            domain_confidence="Medium",
            domain_reason="Name and domain share some tokens. Likely a group or abbreviated domain.",
            manual_review_needed=False,
        )
        return result

    # Overlap too low — search for a better domain
    if serper_key:
        suggested, conf, reason, ev, query = search_official_domain(name, serper_key)
        _fill_serper_top(result, ev, query)
        if suggested and conf >= 0.45 and suggested != norm:
            result.update(
                validated_domain=suggested,
                recommended_domain=suggested,
                domain_action="SUGGEST_REPLACE",
                domain_confidence=_conf_label(conf),
                domain_reason=f"Low name-domain overlap ({overlap:.2f}). {reason}",
                manual_review_needed=True,
            )
        elif suggested and suggested == norm:
            result.update(
                domain_action="LIKELY_OK",
                domain_confidence="Medium",
                domain_reason=f"Search confirms input domain despite low token overlap ({overlap:.2f}).",
                manual_review_needed=False,
            )
        else:
            result.update(
                domain_action="NO_CONFIDENT_MATCH",
                domain_confidence="Low",
                domain_reason=f"Low name-domain overlap ({overlap:.2f}) and no strong replacement found.",
                manual_review_needed=True,
            )
    else:
        if overlap >= 0.1:
            result.update(
                domain_action="REVIEW",
                domain_confidence="Low",
                domain_reason=f"Name-domain overlap is low ({overlap:.2f}). Manual check recommended.",
                manual_review_needed=True,
            )
        else:
            result.update(
                domain_action="NO_CONFIDENT_MATCH",
                domain_confidence="None",
                domain_reason=f"Name-domain overlap is very low ({overlap:.2f}). Likely wrong domain.",
                manual_review_needed=True,
            )
    return result


def _conf_label(conf: float) -> str:
    if conf >= 0.75:
        return "High"
    if conf >= 0.45:
        return "Medium"
    return "Low"


def _fill_serper_top(result: dict, evidence: list, query: str) -> None:
    result["search_query_used"] = query
    used = [e for e in evidence if e.get("used")]
    top  = (used or evidence)
    if top:
        result["serper_top_result_title"]  = top[0].get("title", "")[:120]
        result["serper_top_result_url"]    = top[0].get("url", "")
        result["serper_top_result_domain"] = top[0].get("domain", "")


# ── Process dataframe ─────────────────────────────────────────────────────────
_OUTPUT_COLS = [
    "original_domain", "normalized_input_domain", "validated_domain",
    "recommended_domain", "domain_action", "domain_confidence",
    "domain_reason", "manual_review_needed",
    "serper_top_result_title", "serper_top_result_url",
    "serper_top_result_domain", "search_query_used",
]


def process_dataframe(
    df: pd.DataFrame,
    name_col: str,
    domain_col: str | None,
    serper_key: str | None,
    progress_cb=None,
) -> tuple[pd.DataFrame, list[dict]]:
    """
    Validate all rows. Returns (enriched_df, search_evidence_rows).
    progress_cb(i, n) called after each row.
    """
    results = []
    evidence_rows: list[dict] = []
    n = len(df)

    for i, (_, row) in enumerate(df.iterrows()):
        name       = str(row.get(name_col, "") or "").strip()
        raw_domain = str(row.get(domain_col, "") or "").strip() if domain_col else ""

        res = validate_row(name, raw_domain, serper_key)
        results.append(res)

        # Collect search evidence
        query = res.get("search_query_used", "")
        if query:
            evidence_rows.append({
                "company_name":      name,
                "search_query_used": query,
                "top_result_title":  res.get("serper_top_result_title", ""),
                "top_result_url":    res.get("serper_top_result_url", ""),
                "top_result_domain": res.get("serper_top_result_domain", ""),
                "recommended_domain": res.get("recommended_domain", ""),
                "domain_action":     res.get("domain_action", ""),
            })

        if progress_cb:
            progress_cb(i + 1, n)

    result_df = pd.DataFrame(results, index=df.index)
    enriched = pd.concat([df.copy(), result_df], axis=1)
    return enriched, evidence_rows


# ── Excel export ──────────────────────────────────────────────────────────────
def _action_fill(action: str):
    from openpyxl.styles import PatternFill
    hex_color = _ACTION_COLORS.get(action, "FFFFFF")
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _header_fill():
    from openpyxl.styles import PatternFill, Font
    fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=10)
    return fill, font


def _write_sheet(ws, df: pd.DataFrame, name_col: str | None = None) -> None:
    """Write df to an openpyxl worksheet with header + row colouring."""
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    hdr_fill, hdr_font = _header_fill()

    # Header
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        # Auto-width: cap at 50
        ws.column_dimensions[get_column_letter(ci)].width = min(
            max(len(str(col)) + 2, 12), 50
        )

    # Determine colour column index
    action_col_idx = None
    validated_col_idx = None
    original_col_idx = None
    cols = list(df.columns)
    if "domain_action" in cols:
        action_col_idx = cols.index("domain_action") + 1
    if "validated_domain" in cols:
        validated_col_idx = cols.index("validated_domain") + 1
    if "normalized_input_domain" in cols or "original_domain" in cols:
        check_col = "normalized_input_domain" if "normalized_input_domain" in cols else "original_domain"
        original_col_idx = cols.index(check_col) + 1

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        action = str(row.get("domain_action", "") or "")
        fill   = _action_fill(action)
        for ci, col in enumerate(cols, 1):
            val  = row[col]
            if isinstance(val, float) and val != val:
                val = ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=False, vertical="top")

        # Extra highlight on validated_domain if it differs from input
        if validated_col_idx and original_col_idx:
            v_val = str(row.get("validated_domain", "") or "")
            o_val = str(row.get("normalized_input_domain", row.get("original_domain", "")) or "")
            if v_val and v_val != o_val:
                from openpyxl.styles import Font
                ws.cell(row=ri, column=validated_col_idx).font = Font(bold=True, color="C00000")

    ws.freeze_panes = "A2"
    if len(df) > 0:
        from openpyxl.utils import get_column_letter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    ws.row_dimensions[1].height = 18


def _build_best_guess_df(
    enriched_df: pd.DataFrame, name_col: str
) -> pd.DataFrame:
    """Return a two-column dataframe (company_name, website_url) with best-guess domains."""
    rows = []
    for _, r in enriched_df.iterrows():
        action   = str(r.get("domain_action", "") or "")
        norm     = str(r.get("normalized_input_domain", "") or "").strip()
        recom    = str(r.get("recommended_domain", "") or "").strip()
        cname    = str(r.get(name_col, "") or "").strip()

        if action in ("OK", "LIKELY_OK"):
            url = norm
        elif action in ("SUGGEST_REPLACE", "MISSING_DOMAIN_FIXED"):
            url = recom
        elif action == "REVIEW":
            url = norm or recom
        else:  # NO_CONFIDENT_MATCH, MISSING_DOMAIN, unknown
            url = norm  # blank if norm is ""

        rows.append({"company_name": cname, "website_url": url})
    return pd.DataFrame(rows)


def _write_best_guess_sheet(ws, bg_df: pd.DataFrame, enriched_df: pd.DataFrame) -> None:
    """Write Best Guess Input sheet with simple highlighting."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    hdr_fill, hdr_font = _header_fill()

    # Header
    for ci, col in enumerate(bg_df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = 36

    # Fills
    changed_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    blank_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ok_fill      = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # We need the original normalised domain per row for comparison
    orig_norms = enriched_df.get("normalized_input_domain", pd.Series("", index=enriched_df.index))

    for ri, ((_, bg_row), orig_norm) in enumerate(
        zip(bg_df.iterrows(), orig_norms), 2
    ):
        cname = bg_row["company_name"]
        url   = bg_row["website_url"]
        orig  = str(orig_norm or "").strip()

        ws.cell(row=ri, column=1, value=cname)

        url_cell = ws.cell(row=ri, column=2, value=url)
        if not url:
            url_cell.fill = blank_fill
        elif url != orig:
            url_cell.fill = changed_fill
            url_cell.font = Font(bold=True, color="C00000")
        else:
            url_cell.fill = ok_fill

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 18


def build_excel(
    enriched_df: pd.DataFrame,
    original_df: pd.DataFrame,
    evidence_rows: list[dict],
    name_col: str = "company_name",
) -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()

    # Sheet 1: Best Guess Input
    ws0 = wb.active
    ws0.title = "Best Guess Input"
    bg_df = _build_best_guess_df(enriched_df, name_col)
    _write_best_guess_sheet(ws0, bg_df, enriched_df)

    # Sheet 2: Cleaned Input
    ws1 = wb.create_sheet("Cleaned Input")
    _write_sheet(ws1, enriched_df)

    # Sheet 3: Review Needed
    ws2 = wb.create_sheet("Review Needed")
    review_df = enriched_df[
        enriched_df.get("manual_review_needed", pd.Series(False, index=enriched_df.index))
        .astype(str)
        .str.lower()
        .isin(["true", "1", "yes"])
    ]
    if review_df.empty:
        review_df = pd.DataFrame(columns=enriched_df.columns)
    _write_sheet(ws2, review_df)

    # Sheet 3: Original Input
    ws3 = wb.create_sheet("Original Input")
    _write_sheet(ws3, original_df)

    # Sheet 4: Raw Search Evidence
    ws4 = wb.create_sheet("Raw Search Evidence")
    ev_df = pd.DataFrame(evidence_rows) if evidence_rows else pd.DataFrame(
        columns=["company_name", "search_query_used", "top_result_title",
                 "top_result_url", "top_result_domain", "recommended_domain", "domain_action"]
    )
    _write_sheet(ws4, ev_df)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Streamlit UI ──────────────────────────────────────────────────────────────
def _load_secrets_key() -> str | None:
    try:
        return st.secrets.get("SERPER_API_KEY") or st.secrets.get("serper_api_key")
    except Exception:
        return None


def _load_file(uploaded) -> pd.DataFrame | None:
    raw = uploaded.read()
    name = uploaded.name.lower()
    try:
        if name.endswith(".csv"):
            return pd.read_csv(io.BytesIO(raw), dtype=str).fillna("")
        else:
            return pd.read_excel(io.BytesIO(raw), dtype=str).fillna("")
    except Exception as e:
        st.error(f"Could not read file: {e}")
        return None


def _summary_metrics(df: pd.DataFrame) -> None:
    actions = df.get("domain_action", pd.Series(dtype=str))
    total  = len(df)
    ok     = int((actions == "OK").sum())
    lok    = int((actions == "LIKELY_OK").sum())
    rep    = int(actions.isin(["SUGGEST_REPLACE", "MISSING_DOMAIN_FIXED"]).sum())
    rev    = int(
        df.get("manual_review_needed", pd.Series(dtype=str))
        .astype(str).str.lower().isin(["true", "1", "yes"]).sum()
    )
    ncm    = int(actions.isin(["NO_CONFIDENT_MATCH", "MISSING_DOMAIN"]).sum())

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    def card(col, label, val, color):
        col.markdown(
            f"<div style='border-left:4px solid {color};padding:8px 12px;"
            f"background:#f7f9fc;border-radius:4px'>"
            f"<div style='font-size:1.5em;font-weight:700;color:{color}'>{val}</div>"
            f"<div style='font-size:0.78em;color:#555'>{label}</div></div>",
            unsafe_allow_html=True,
        )
    card(c1, "Total rows",       total, "#0B4A92")
    card(c2, "OK",               ok,    "#2E7D32")
    card(c3, "Likely OK",        lok,   "#558B2F")
    card(c4, "Suggested replace", rep,   "#E65100")
    card(c5, "Manual review",    rev,   "#F57F17")
    card(c6, "No confident match", ncm, "#B71C1C")


def main():
    st.title("🔍 Input Cleaner Lite")
    st.caption("Layer 0 · mYngle Sales Intelligence · Domain validation before Lead Prioritizer")

    serper_key = _load_secrets_key()
    if serper_key:
        st.sidebar.success("Serper API key loaded from secrets.")
    else:
        st.sidebar.warning(
            "No Serper API key found in `.streamlit/secrets.toml`.\n\n"
            "Domain search will be skipped. Only heuristic matching will run."
        )
        manual_key = st.sidebar.text_input(
            "Paste Serper API key (optional)", type="password", key="manual_serper"
        )
        if manual_key.strip():
            serper_key = manual_key.strip()

    uploaded = st.file_uploader(
        "Upload company list (CSV or Excel .xlsx)",
        type=["csv", "xlsx"],
        key="icl_upload",
    )

    if uploaded is None:
        st.info(
            "Upload a CSV or Excel file with at least a **company name** column "
            "and (optionally) a **domain / URL** column."
        )
        return

    df = _load_file(uploaded)
    if df is None:
        return

    st.success(f"Loaded {len(df)} rows, {len(df.columns)} columns.")

    # ── Column detection ──────────────────────────────────────────────────────
    name_col, domain_col = detect_columns(df)
    auto_detected = name_col is not None

    if not auto_detected:
        st.warning("Could not auto-detect company name column. Please select manually.")

    with st.expander("Column mapping", expanded=(not auto_detected)):
        col_options = ["(none)"] + list(df.columns)
        name_col = st.selectbox(
            "Company name column",
            list(df.columns),
            index=list(df.columns).index(name_col) if name_col else 0,
            key="col_name",
        )
        domain_col = st.selectbox(
            "Domain / URL column (optional)",
            col_options,
            index=col_options.index(domain_col) if domain_col and domain_col in col_options else 0,
            key="col_domain",
        )
        if domain_col == "(none)":
            domain_col = None
        if auto_detected and name_col and domain_col:
            st.caption(f"Auto-detected → company: **{name_col}** · domain: **{domain_col}**")

    # ── Preview ───────────────────────────────────────────────────────────────
    preview_cols = [c for c in [name_col, domain_col] if c]
    st.dataframe(df[preview_cols].head(5), use_container_width=True)

    # ── Run ───────────────────────────────────────────────────────────────────
    if st.button("🧹 Clean and validate", type="primary", use_container_width=True):
        progress_bar = st.progress(0)
        status_text  = st.empty()
        n = len(df)

        def progress_cb(i, total):
            progress_bar.progress(i / total)
            status_text.caption(f"Processing row {i} of {total}…")

        enriched_df, evidence_rows = process_dataframe(
            df, name_col, domain_col, serper_key, progress_cb
        )

        progress_bar.progress(1.0)
        status_text.caption("Done.")

        st.session_state["icl_enriched"]  = enriched_df
        st.session_state["icl_evidence"]  = evidence_rows
        st.session_state["icl_original"]  = df
        st.session_state["icl_name_col"]  = name_col

    # ── Results ───────────────────────────────────────────────────────────────
    enriched_df = st.session_state.get("icl_enriched")
    if enriched_df is None:
        return

    st.markdown("---")
    st.markdown("### Results")
    _summary_metrics(enriched_df)
    st.markdown("")

    # Show table — highlight key columns
    show_cols = [c for c in [
        name_col, domain_col, "normalized_input_domain", "validated_domain",
        "domain_action", "domain_confidence", "domain_reason", "manual_review_needed",
    ] if c and c in enriched_df.columns]
    st.dataframe(enriched_df[show_cols], use_container_width=True, height=350)

    # Review rows
    review_mask = (
        enriched_df.get("manual_review_needed", pd.Series(False))
        .astype(str).str.lower().isin(["true", "1", "yes"])
    )
    if review_mask.any():
        with st.expander(f"Rows needing manual review ({int(review_mask.sum())})", expanded=False):
            st.dataframe(enriched_df[review_mask][show_cols], use_container_width=True)

    # ── Download ──────────────────────────────────────────────────────────────
    evidence_rows = st.session_state.get("icl_evidence", [])
    original_df   = st.session_state.get("icl_original", df)
    excel_bytes   = build_excel(
        enriched_df, original_df, evidence_rows,
        name_col=st.session_state.get("icl_name_col", "company_name"),
    )

    st.download_button(
        "⬇ Download cleaned Excel",
        data=excel_bytes,
        file_name="input_cleaner_lite_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


if __name__ == "__main__":
    main()
