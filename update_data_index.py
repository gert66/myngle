"""
update_data_index.py

Scans configured Nextcloud/Myngle data root(s) and writes DATA_INDEX.md.
Reads config from local_paths.json in the repo root.
Only metadata is written — no cell values from Excel files.
"""

import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

REPO_ROOT = Path(__file__).parent
CONFIG_FILE = REPO_ROOT / "local_paths.json"


def load_config() -> dict:
    if not CONFIG_FILE.exists():
        print(f"Config not found: {CONFIG_FILE}")
        print("Copy local_paths.example.json to local_paths.json and adjust paths.")
        sys.exit(1)
    with CONFIG_FILE.open(encoding="utf-8") as f:
        return json.load(f)


def human_size(n_bytes: int) -> str:
    for unit in ("B", "KB", "MB", "GB"):
        if n_bytes < 1024:
            return f"{n_bytes:.0f} {unit}"
        n_bytes /= 1024
    return f"{n_bytes:.1f} TB"


def fmt_ts(ts: float) -> str:
    return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M")


def excel_metadata(path: Path, max_rows_sample: int = 5000) -> list[dict] | None:
    """Return sheet metadata without reading cell values."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            row_count = ws.max_row or 0
            # Read only first header row for column names
            cols = []
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                cols = [str(c) for c in row if c is not None][:10]
                break
            sheets.append({"name": name, "rows": row_count, "columns": cols})
        wb.close()
        return sheets
    except Exception as e:
        return [{"error": str(e)}]


def build_tree(
    root: Path,
    skip_dirs: set[str],
    max_depth: int,
    max_files: int,
    excel_meta: bool,
    excel_max: int,
    current_depth: int = 0,
) -> list[str]:
    lines: list[str] = []
    indent = "  " * current_depth

    try:
        entries = sorted(root.iterdir(), key=lambda p: (p.is_file(), p.name.lower()))
    except PermissionError:
        lines.append(f"{indent}- *(geen leestoegang)*")
        return lines

    dirs = [e for e in entries if e.is_dir() and e.name not in skip_dirs]
    files = [e for e in entries if e.is_file()]

    # Files in this folder
    shown_files = files[:max_files]
    excel_count = 0
    for f in shown_files:
        try:
            stat = f.stat()
            size = human_size(stat.st_size)
            mtime = fmt_ts(stat.st_mtime)
        except OSError:
            size, mtime = "?", "?"
        lines.append(f"{indent}- `{f.name}` — {size} — {mtime}")
        if excel_meta and f.suffix.lower() in (".xlsx", ".xls", ".xlsm") and excel_count < excel_max:
            excel_count += 1
            sheets = excel_metadata(f)
            if sheets:
                for s in sheets:
                    if "error" in s:
                        lines.append(f"{indent}  - *(fout bij lezen: {s['error']})*")
                    else:
                        col_str = ", ".join(s["columns"]) if s["columns"] else "—"
                        lines.append(f"{indent}  - Sheet **{s['name']}**: {s['rows']} rijen | kolommen: {col_str}")

    if len(files) > max_files:
        lines.append(f"{indent}- *… en nog {len(files) - max_files} bestand(en)*")

    # Recurse into subdirs
    if current_depth < max_depth:
        for d in dirs:
            lines.append(f"{indent}- **{d.name}/**")
            lines.extend(
                build_tree(d, skip_dirs, max_depth, max_files, excel_meta, excel_max, current_depth + 1)
            )
    else:
        for d in dirs:
            lines.append(f"{indent}- **{d.name}/** *(niet verder uitgeklapt)*")

    return lines


def generate_index(cfg: dict) -> str:
    data_roots: list[str] = cfg.get("data_roots", [])
    output_file: str = cfg.get("output_file", "DATA_INDEX.md")
    max_depth: int = int(cfg.get("max_depth", 4))
    max_files: int = int(cfg.get("max_files_per_folder", 20))
    excel_meta: bool = bool(cfg.get("excel_metadata", True))
    excel_max: int = int(cfg.get("excel_max_files_per_folder", 5))
    skip_dirs: set[str] = set(cfg.get("skip_dirs", []))

    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    try:
        import socket, getpass
        machine = f"{getpass.getuser()}@{socket.gethostname()}"
    except Exception:
        machine = "onbekend"

    parts = [
        "# DATA_INDEX",
        "",
        f"Gegenereerd: {now}  ",
        f"Machine: {machine}  ",
        "",
    ]

    for root_str in data_roots:
        root = Path(root_str)
        parts.append(f"## Data root: `{root_str}`")
        parts.append("")
        if not root.exists():
            parts.append(f"> Map niet gevonden: `{root_str}`")
            parts.append("")
            continue

        tree_lines = build_tree(root, skip_dirs, max_depth, max_files, excel_meta, excel_max)
        if tree_lines:
            parts.extend(tree_lines)
        else:
            parts.append("*(lege map)*")
        parts.append("")

    return "\n".join(parts)


def main():
    cfg = load_config()
    output_path = REPO_ROOT / cfg.get("output_file", "DATA_INDEX.md")
    content = generate_index(cfg)
    output_path.write_text(content, encoding="utf-8")
    print(f"Geschreven: {output_path}")


if __name__ == "__main__":
    main()
