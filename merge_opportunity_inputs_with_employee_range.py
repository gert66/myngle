"""
merge_opportunity_inputs_with_employee_range.py

Merges only the 'Opportunity Input' sheet from all .xlsx files in the
lead-prioritized output folders for Italy50, Italy100, Italy200.

Output workbook contains four sheets:
  1. Debug View              – outlier/manual-review sheet (foreign HQ + competitor focus)
  2. Opportunity Input Full  – all merged rows, source columns in original order,
                               trace columns (source_queue/file/row) appended at the end
  3. Scoring Logic           – explanatory notes on scoring methodology
  4. Merge QA                – run metadata and review counts

Adds to every row:
  - Chamber of Commerce Employee Range  (placed next to existing employee range column)
  - source_queue, source_file, source_row  (appended at the far right)
  - debug flags and review_priority  (computed, used in Debug View)

Usage:
  python merge_opportunity_inputs_with_employee_range.py \\
      --target-root "C:\\Users\\...\\Myngle" \\
      --queues Italy50 Italy100 Italy200

  # URL-patched variant:
  python merge_opportunity_inputs_with_employee_range.py \\
      --target-root "C:\\Users\\...\\Myngle" \\
      --queues Italy50 Italy100 Italy200 \\
      --input-subfolder "02_lead_prioritized_url_patched"
"""

from __future__ import annotations

import argparse
import csv
import sys
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Dependency check
# ---------------------------------------------------------------------------
try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule, DataBarRule
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

TARGET_SHEET = "Opportunity Input"
FULL_SHEET_NAME = "Opportunity Input Full"
DEBUG_SHEET_NAME = "Debug View"
EMP_RANGE_COL = "Chamber of Commerce Employee Range"
TRACE_COLS = ["source_queue", "source_file", "source_row"]

# Existing employee-range column names to detect (case-insensitive)
EMP_RANGE_ADJACENT_NAMES = {
    "employee_range",
    "employee_range_resolved",
    "employee_range_source",
    "employee_count_range",
    "employee_range_label",
    "employee range",
}

EMPLOYEE_RANGE_MAP: dict[str, str] = {
    "italy50":  "50-100 employees",
    "italy100": "100-200 employees",
    "italy200": "200+ employees",
}

FOREIGN_HQ_SANITIZER_FIELDS = [
    "sig_foreign_hq_score",
    "sig_foreign_hq_evidence",
    "foreign_hq_sanitized",
    "foreign_hq_sanitizer_reason",
    "foreign_hq_original_score",
    "foreign_hq_original_evidence",
    "inferred_input_country",
    "foreign_hq_uncertain",
]

REQUIRED_CORE_COLS = [
    "company_name", "domain",
    "final_commercial_fit_score", "commercial_fit_score", "commercial_tier",
    "icp_override_applied",
]

# Score columns that get blue data bars
DATABAR_COLS = {
    "display_score",
    "base_commercial_fit_score",
    "final_commercial_fit_score",
    "commercial_fit_score",
    "Final Commercial Fit Score",
    "commercial_fit_score_75_25_legacy",
}

# Priority order for resolving the canonical display/final score
_FINAL_SCORE_COLS = [
    "final_commercial_fit_score",
    "commercial_fit_score",
    "Final Commercial Fit Score",
    "commercial_fit_score_75_25_legacy",
]

# All debug flag column names
DEBUG_FLAG_COLS = [
    "debug_foreign_hq_flag",
    "debug_competitor_flag",
    "debug_score_origin_flag",
    "debug_domain_flag",
]
REVIEW_PRIORITY_COL = "review_priority"

# Domain-check warning keywords
_DOMAIN_WARN_WORDS = {"mismatch", "uncertain", "review", "fallback", "generic"}

# Header / fill styles
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TRACE_FILL  = PatternFill("solid", fgColor="D6E4F0")
EMP_FILL    = PatternFill("solid", fgColor="E2EFDA")
DEBUG_FILL  = PatternFill("solid", fgColor="FCE4D6")
FLAG_FILL   = PatternFill("solid", fgColor="FF0000")
FLAG_FONT   = Font(bold=True, color="FFFFFF")
HIGH_PRI_FILL   = PatternFill("solid", fgColor="FF0000")
HIGH_PRI_FONT   = Font(bold=True, color="FFFFFF")
MEDIUM_PRI_FILL = PatternFill("solid", fgColor="FFEB9C")
MEDIUM_PRI_FONT = Font(bold=False, color="000000")

# Tier-based row fills (fallback row coloring)
HOT_FILL  = PatternFill("solid", fgColor="C6EFCE")
WARM_FILL = PatternFill("solid", fgColor="DDEBF7")
COOL_FILL = PatternFill("solid", fgColor="FCE4D6")
PASS_FILL = PatternFill("solid", fgColor="FFC7CE")

DEFAULT_ROW_HEIGHT = 15

# Debug View ordered columns A–BB
DEBUG_VIEW_COLS = [
    # Identity
    "company_name",
    "domain",
    # Score block
    "display_score",
    "base_commercial_fit_score",
    "final_commercial_fit_score",
    "commercial_tier",
    "outreach_readiness_status",
    EMP_RANGE_COL,
    # Foreign HQ block
    "sig_foreign_hq_score",
    "foreign_hq_original_score",
    "foreign_hq_sanitized",
    "foreign_hq_sanitizer_reason",
    "sig_foreign_hq_evidence",
    "foreign_hq_original_evidence",
    "inferred_input_country",
    "foreign_hq_uncertain",
    # Competitor block
    "icp_override_applied",
    "icp_override_reason",
    "competitive_switch_opportunity",
    "sales_action_hint",
    "competitor_customer_signal",
    "competitor_signal_strength",
    "competitor_provider_detected",
    "competitor_evidence",
    "competitor_evidence_url",
    "competitor_confidence_url",
    "competitor_attention_signal",
    "competitor_attention_provider_detected",
    "competitor_attention_url",
    # Other signal block
    "sig_explicit_lnd_score",
    "sig_explicit_lnd_evidence",
    "sig_intl_footprint_score",
    "sig_intl_footprint_evidence",
    "sig_employer_branding_score",
    "sig_employer_branding_evidence",
    "sig_lnd_onboarding_score",
    "sig_lnd_onboarding_evidence",
    "ti_onboarding_score",
    "sig_rapid_growth_score",
    # Domain / review block
    "possible_domain_mismatch",
    "domain_check_reason",
    "needs_manual_review",
    "top_positive_signals",
    "gaps_missing_signals",
    "caller_angle",
    "scoring_notes",
    # Debug flags
    "debug_foreign_hq_flag",
    "debug_competitor_flag",
    "debug_score_origin_flag",
    "debug_domain_flag",
    REVIEW_PRIORITY_COL,
    # Trace
    "source_queue",
    "source_file",
    "source_row",
]

_DEBUG_COL_WIDTHS: dict[str, int] = {
    "company_name": 35,
    "domain": 28,
    "display_score": 16,
    "base_commercial_fit_score": 24,
    "final_commercial_fit_score": 24,
    "commercial_tier": 16,
    "outreach_readiness_status": 24,
    EMP_RANGE_COL: 26,
    "sig_foreign_hq_score": 20,
    "foreign_hq_original_score": 22,
    "foreign_hq_sanitized": 20,
    "foreign_hq_sanitizer_reason": 32,
    "sig_foreign_hq_evidence": 55,
    "foreign_hq_original_evidence": 55,
    "inferred_input_country": 22,
    "foreign_hq_uncertain": 20,
    "icp_override_applied": 22,
    "icp_override_reason": 40,
    "competitive_switch_opportunity": 28,
    "sales_action_hint": 35,
    "competitor_customer_signal": 28,
    "competitor_signal_strength": 24,
    "competitor_provider_detected": 28,
    "competitor_evidence": 50,
    "competitor_evidence_url": 45,
    "competitor_confidence_url": 45,
    "competitor_attention_signal": 28,
    "competitor_attention_provider_detected": 30,
    "competitor_attention_url": 45,
    "sig_explicit_lnd_score": 20,
    "sig_explicit_lnd_evidence": 45,
    "sig_intl_footprint_score": 22,
    "sig_intl_footprint_evidence": 45,
    "sig_employer_branding_score": 24,
    "sig_employer_branding_evidence": 45,
    "sig_lnd_onboarding_score": 22,
    "sig_lnd_onboarding_evidence": 45,
    "ti_onboarding_score": 18,
    "sig_rapid_growth_score": 22,
    "possible_domain_mismatch": 22,
    "domain_check_reason": 35,
    "needs_manual_review": 20,
    "top_positive_signals": 55,
    "gaps_missing_signals": 45,
    "caller_angle": 40,
    "scoring_notes": 60,
    "debug_foreign_hq_flag": 40,
    "debug_competitor_flag": 40,
    "debug_score_origin_flag": 40,
    "debug_domain_flag": 32,
    REVIEW_PRIORITY_COL: 16,
    "source_queue": 14,
    "source_file": 48,
    "source_row": 12,
}

_WRAP_COLS = {
    "sig_foreign_hq_evidence",
    "foreign_hq_original_evidence",
    "competitor_evidence",
    "top_positive_signals",
    "scoring_notes",
    "sig_intl_footprint_evidence",
    "sig_explicit_lnd_evidence",
    "sig_employer_branding_evidence",
    "sig_lnd_onboarding_evidence",
    "gaps_missing_signals",
    "caller_angle",
    "icp_override_reason",
}

# ---------------------------------------------------------------------------
# Safe workbook loader
# ---------------------------------------------------------------------------

def _safe_load(path: Path) -> tuple["openpyxl.Workbook | None", str]:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        return wb, ""
    except (zipfile.BadZipFile, KeyError, OSError, Exception) as exc:
        return None, str(exc)


# ---------------------------------------------------------------------------
# Safe output path
# ---------------------------------------------------------------------------

def _safe_output_path(path: Path, overwrite: bool) -> Path:
    if not path.exists() or overwrite:
        return path
    stem, suffix, parent = path.stem, path.suffix, path.parent
    n = 2
    while True:
        candidate = parent / f"{stem}_{n}{suffix}"
        if not candidate.exists():
            return candidate
        n += 1


# ---------------------------------------------------------------------------
# Value helpers
# ---------------------------------------------------------------------------

def _to_float(v: Any) -> "float | None":
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _is_truthy(v: Any) -> bool:
    if v is None:
        return False
    return str(v).strip().lower() in {"true", "yes", "1", "1.0", "x"}


def _is_blank_or_short(v: Any, threshold: int = 10) -> bool:
    return v is None or len(str(v).strip()) < threshold


def _str_lower(v: Any) -> str:
    return str(v or "").strip().lower()


# ---------------------------------------------------------------------------
# Column layout helpers
# ---------------------------------------------------------------------------

def _find_emp_range_insert_pos(source_headers: list[str]) -> int:
    """
    Return 0-based index of the position to insert EMP_RANGE_COL
    (immediately after the first existing employee-range column found).
    Returns len(source_headers) if none found (appended before trace cols).
    """
    for i, h in enumerate(source_headers):
        if h.strip().lower() in EMP_RANGE_ADJACENT_NAMES:
            return i + 1
    return len(source_headers)


def _build_output_headers(source_headers: list[str]) -> list[str]:
    """
    Build full output column list:
      original source cols (minus existing EMP_RANGE_COL)
      + EMP_RANGE_COL inserted next to existing employee-range column
      + TRACE_COLS appended at the end
    """
    clean_source = [h for h in source_headers if h != EMP_RANGE_COL]
    insert_pos = _find_emp_range_insert_pos(clean_source)
    result = clean_source[:insert_pos] + [EMP_RANGE_COL] + clean_source[insert_pos:]
    result += TRACE_COLS
    return result


def _row_to_output(
    source_row: tuple,
    source_headers: list[str],
    output_headers: list[str],
    queue: str,
    filename: str,
    row_num: int,
) -> list[Any]:
    emp_range_value = EMPLOYEE_RANGE_MAP.get(queue.lower(), "")
    clean_source = [h for h in source_headers if h != EMP_RANGE_COL]
    src_dict: dict[str, Any] = {
        h: (source_row[i] if i < len(source_row) else None)
        for i, h in enumerate(clean_source)
    }
    trace_dict = {
        "source_queue": queue,
        "source_file": filename,
        "source_row": row_num,
    }
    out_row: list[Any] = []
    for col_name in output_headers:
        if col_name in trace_dict:
            out_row.append(trace_dict[col_name])
        elif col_name == EMP_RANGE_COL:
            out_row.append(emp_range_value)
        elif col_name in src_dict:
            out_row.append(src_dict[col_name])
        else:
            out_row.append(None)
    return out_row


# ---------------------------------------------------------------------------
# Debug flag computation
# ---------------------------------------------------------------------------

def _final_score(row_dict: dict[str, Any]) -> "float | None":
    """Return the best available final/display score."""
    for col in ("display_score", "final_commercial_fit_score",
                "Final Commercial Fit Score", "commercial_fit_score",
                "commercial_fit_score_75_25_legacy"):
        v = _to_float(row_dict.get(col))
        if v is not None:
            return v
    return None


def _derive_score_fields(row_dict: dict[str, Any]) -> dict[str, Any]:
    """
    Derive canonical score fields for use in Debug View.
    Does not modify the original row_dict or Opportunity Input Full columns.

    Returns a dict with:
      display_score, base_commercial_fit_score, final_commercial_fit_score
    """
    # Resolve final/display score using priority list
    final = None
    for col in _FINAL_SCORE_COLS:
        v = _to_float(row_dict.get(col))
        if v is not None:
            final = v
            break

    # Resolve base score
    base = _to_float(row_dict.get("base_commercial_fit_score"))
    if base is None:
        # If we have both a final score and a commercial_fit_score that differs,
        # treat commercial_fit_score as the base (pre-override base)
        cs = _to_float(row_dict.get("commercial_fit_score"))
        fcs = _to_float(row_dict.get("final_commercial_fit_score"))
        if cs is not None and fcs is not None and cs != fcs:
            base = cs

    return {
        "display_score": final,
        "final_commercial_fit_score": final,
        "base_commercial_fit_score": base,
    }


def compute_debug_flags(
    row_dict: dict[str, Any],
) -> tuple[str, str, str, str, str]:
    """Returns (fhq_flag, competitor_flag, score_origin_flag, domain_flag, review_priority)."""

    score = _final_score(row_dict)
    fhq_score  = _to_float(row_dict.get("sig_foreign_hq_score"))
    fhq_orig   = _to_float(row_dict.get("foreign_hq_original_score"))
    fhq_sanit  = _is_truthy(row_dict.get("foreign_hq_sanitized"))
    fhq_evid   = row_dict.get("sig_foreign_hq_evidence")
    override   = _is_truthy(row_dict.get("icp_override_applied"))

    # ── foreign HQ flag ─────────────────────────────────────────────────────
    fhq_reasons: list[str] = []
    if fhq_sanit:
        fhq_reasons.append("Foreign HQ sanitized, review needed")
    if fhq_score == 0.0 and fhq_orig is not None and fhq_orig > 0:
        fhq_reasons.append("Foreign HQ score zero but original positive")
    if score is not None and score >= 8.5 and fhq_sanit:
        fhq_reasons.append("High score despite sanitized foreign HQ")
    if fhq_score is not None and fhq_score >= 7 and _is_blank_or_short(fhq_evid):
        fhq_reasons.append("Foreign HQ high but evidence weak")

    # ── competitor flag ─────────────────────────────────────────────────────
    comp_reasons: list[str] = []
    comp_url_cols = (
        "competitor_evidence_url", "competitor_confidence_url",
        "competitor_attention_url",
    )
    has_comp_url = any(
        str(row_dict.get(c) or "").strip() for c in comp_url_cols
    )
    comp_fields_populated = (
        str(row_dict.get("competitor_customer_signal") or "").strip()
        or str(row_dict.get("competitor_signal_strength") or "").strip()
        or str(row_dict.get("competitor_provider_detected") or "").strip()
    )

    if override:
        comp_reasons.append("Competitor override applied, verify evidence URL")
    if score is not None and score == 10 and override:
        comp_reasons.append("Score 10 from competitor override")
    if has_comp_url:
        comp_reasons.append("Competitor URL present, needs review")
    if _str_lower(row_dict.get("competitor_signal_strength")) == "high":
        comp_reasons.append("Competitor signal high, verify target-company proof")
    if (
        score is not None and score == 10
        and comp_fields_populated
        and not override
    ):
        comp_reasons.append("Competitor override suspected but weak fields")

    # ── score origin flag ────────────────────────────────────────────────────
    origin_reasons: list[str] = []
    if score is not None and score == 10 and override:
        origin_reasons.append("Score 10 from competitor override")
    elif score is not None and score == 10:
        origin_reasons.append("Score 10 from normal scoring - verify drivers")
    if score is not None and score >= 8.5 and fhq_sanit:
        origin_reasons.append("High score despite sanitized foreign HQ")

    # ── domain flag ─────────────────────────────────────────────────────────
    domain_reasons: list[str] = []
    if _is_truthy(row_dict.get("possible_domain_mismatch")):
        domain_reasons.append("Possible domain mismatch")
    if _is_truthy(row_dict.get("needs_manual_review")):
        domain_reasons.append("Manual review needed")
    dcr = _str_lower(row_dict.get("domain_check_reason"))
    if any(w in dcr for w in _DOMAIN_WARN_WORDS):
        domain_reasons.append("Domain check warning")

    # ── review priority ──────────────────────────────────────────────────────
    high = (
        (fhq_sanit and score is not None and score >= 8.5)
        or override
        or (score is not None and score == 10)
        or _is_truthy(row_dict.get("possible_domain_mismatch"))
    )
    any_flag = bool(fhq_reasons or comp_reasons or origin_reasons or domain_reasons)
    if high:
        priority = "High"
    elif any_flag:
        priority = "Medium"
    else:
        priority = "Low"

    return (
        "; ".join(fhq_reasons),
        "; ".join(comp_reasons),
        "; ".join(origin_reasons),
        "; ".join(domain_reasons),
        priority,
    )


# ---------------------------------------------------------------------------
# File discovery
# ---------------------------------------------------------------------------

def _discover_files(queue_dir: Path, subfolder: str) -> list[Path]:
    folder = queue_dir / subfolder
    if not folder.exists():
        return []
    return sorted(
        f for f in folder.glob("*.xlsx")
        if not f.name.startswith("~$")
    )


# ---------------------------------------------------------------------------
# Core merge
# ---------------------------------------------------------------------------

# Extra computed columns appended beyond output_headers in each merged row
_COMPUTED_COLS = DEBUG_FLAG_COLS + [REVIEW_PRIORITY_COL]


def merge_queues(
    target_root: Path,
    queues: list[str],
    input_subfolder: str,
) -> tuple[list[list[Any]], "list[str] | None", list[dict], dict[str, int], list[dict], list[str]]:
    all_data_rows: list[list[Any]] = []
    report_rows: list[dict] = []
    rows_by_queue: dict[str, int] = {}

    # ── Pass 1: read headers only ─────────────────────────────────────────────
    # valid_files: list of (queue, path, source_headers)
    valid_files: list[tuple[str, Path, list[str]]] = []

    for queue in queues:
        queue_dir = target_root / queue
        if not queue_dir.exists():
            print(f"  [WARN] Queue folder not found, skipping: {queue_dir}")
            continue

        files = _discover_files(queue_dir, input_subfolder)
        print(f"\n  Queue {queue}: {len(files)} file(s) in {queue_dir / input_subfolder}")
        rows_by_queue[queue] = 0

        for xlsx_path in files:
            wb, err = _safe_load(xlsx_path)
            if wb is None:
                print(f"    [SKIP-CORRUPT ] {xlsx_path.name}")
                report_rows.append({"queue": queue, "file": xlsx_path.name,
                                    "status": "SKIPPED_INVALID_WORKBOOK",
                                    "rows_read": 0, "notes": err})
                continue

            if TARGET_SHEET not in wb.sheetnames:
                wb.close()
                print(f"    [SKIP-NO-SHEET] {xlsx_path.name}")
                report_rows.append({"queue": queue, "file": xlsx_path.name,
                                    "status": "SKIPPED_MISSING_OPPORTUNITY_INPUT",
                                    "rows_read": 0,
                                    "notes": f"Sheet '{TARGET_SHEET}' not found. "
                                             f"Available: {wb.sheetnames}"})
                continue

            ws = wb[TARGET_SHEET]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                wb.close()
                report_rows.append({"queue": queue, "file": xlsx_path.name,
                                    "status": "SKIPPED_EMPTY_SHEET",
                                    "rows_read": 0, "notes": "Sheet has no rows"})
                continue

            source_headers = [str(h).strip() if h is not None else "" for h in header_row]
            wb.close()
            valid_files.append((queue, xlsx_path, source_headers))

    # ── Build union schema ────────────────────────────────────────────────────
    if not valid_files:
        return [], None, report_rows, rows_by_queue, [], []

    # Select the file with most non-empty headers as base
    base_queue, base_path, base_headers = max(
        valid_files,
        key=lambda t: sum(1 for h in t[2] if h),
    )
    # Build union: start with base, append any new headers from other files
    union_headers: list[str] = list(base_headers)
    union_set: set[str] = set(union_headers)
    for _q, _p, hdrs in valid_files:
        for h in hdrs:
            if h and h not in union_set:
                union_headers.append(h)
                union_set.add(h)

    # Track columns added beyond first-file's headers (for WARN-UNION-NEW)
    first_headers_set = set(valid_files[0][2])
    new_in_union: list[str] = [h for h in union_headers if h not in first_headers_set]

    # Remove EMP_RANGE_COL from union source headers before building output headers
    union_source = [h for h in union_headers if h != EMP_RANGE_COL]
    output_headers = _build_output_headers(union_source)

    # ── Console warnings ──────────────────────────────────────────────────────
    required_check_fields = REQUIRED_CORE_COLS + FOREIGN_HQ_SANITIZER_FIELDS
    for queue, xlsx_path, source_headers in valid_files:
        dups = [h for h in source_headers if h and source_headers.count(h) > 1]
        if dups:
            print(f"    [WARN-DUPS] {queue}/{xlsx_path.name}: duplicate headers: "
                  f"{', '.join(sorted(set(dups)))}")
        hdrs_set = set(source_headers)
        missing = [f for f in required_check_fields if f not in hdrs_set]
        if missing:
            print(f"    [WARN-MISSING-FIELDS] {queue}/{xlsx_path.name}: "
                  f"missing: {', '.join(missing)}")

    if new_in_union:
        print(f"    [WARN-UNION-NEW] union schema added {len(new_in_union)} column(s) "
              f"not in first-file headers: {', '.join(new_in_union[:20])}")

    # ── Build file_schema_info stubs (num_rows filled in pass 2) ─────────────
    file_schema_info: list[dict] = []
    base_headers_set = set(base_headers)
    for queue, xlsx_path, source_headers in valid_files:
        hdrs_set = set(source_headers)
        missing_core = [f for f in (REQUIRED_CORE_COLS + FOREIGN_HQ_SANITIZER_FIELDS)
                        if f not in hdrs_set]
        extra_vs_base = [h for h in source_headers if h and h not in base_headers_set]
        fhq_field_presence = {f: (f in hdrs_set) for f in FOREIGN_HQ_SANITIZER_FIELDS}
        fhq_field_positions: dict[str, "int | None"] = {}
        for f in FOREIGN_HQ_SANITIZER_FIELDS:
            try:
                fhq_field_positions[f] = source_headers.index(f) + 1  # 1-based
            except ValueError:
                fhq_field_positions[f] = None
        dups = [h for h in source_headers if h and source_headers.count(h) > 1]
        file_schema_info.append({
            "queue": queue,
            "file": xlsx_path.name,
            "num_headers": len(source_headers),
            "num_rows": 0,  # filled in pass 2
            "missing_core": missing_core,
            "extra_vs_base": extra_vs_base,
            "fhq_field_presence": fhq_field_presence,
            "fhq_field_positions": fhq_field_positions,
            "duplicate_headers": sorted(set(dups)),
        })

    # ── Pass 2: re-open each valid file and read data rows ────────────────────
    schema_info_by_key: dict[tuple[str, str], dict] = {
        (info["queue"], info["file"]): info for info in file_schema_info
    }

    for queue, xlsx_path, source_headers in valid_files:
        wb, err = _safe_load(xlsx_path)
        if wb is None:
            # Shouldn't normally happen since pass 1 succeeded, but handle gracefully
            report_rows.append({"queue": queue, "file": xlsx_path.name,
                                "status": "SKIPPED_INVALID_WORKBOOK",
                                "rows_read": 0, "notes": err})
            continue

        ws = wb[TARGET_SHEET]
        rows_iter = ws.iter_rows(values_only=True)
        next(rows_iter)  # skip header row

        rows_read = 0
        for src_row_idx, src_row in enumerate(rows_iter, start=2):
            if all(v is None or str(v).strip() == "" for v in src_row):
                continue
            out_row = _row_to_output(
                source_row=src_row,
                source_headers=source_headers,
                output_headers=output_headers,
                queue=queue,
                filename=xlsx_path.name,
                row_num=src_row_idx,
            )
            row_dict = dict(zip(output_headers, out_row))
            flags = compute_debug_flags(row_dict)   # 5 values
            out_row = out_row + list(flags)
            all_data_rows.append(out_row)
            rows_read += 1

        wb.close()
        rows_by_queue[queue] = rows_by_queue.get(queue, 0) + rows_read
        schema_info_by_key[(queue, xlsx_path.name)]["num_rows"] = rows_read
        print(f"    [OK  {rows_read:>6} rows] {xlsx_path.name}")
        report_rows.append({"queue": queue, "file": xlsx_path.name,
                             "status": "MERGED", "rows_read": rows_read, "notes": ""})

    return all_data_rows, output_headers, report_rows, rows_by_queue, file_schema_info, new_in_union


# ---------------------------------------------------------------------------
# Row dict helper
# ---------------------------------------------------------------------------

def _row_as_dict(row: list[Any], full_headers: list[str]) -> dict[str, Any]:
    extended = full_headers + _COMPUTED_COLS
    return {h: (row[i] if i < len(row) else None) for i, h in enumerate(extended)}


# ---------------------------------------------------------------------------
# Formatting helpers
# ---------------------------------------------------------------------------

def _get_row_fill(row_dict: dict[str, Any]) -> "PatternFill | None":
    tier = _str_lower(row_dict.get("commercial_tier"))
    score = _final_score(row_dict)
    # Use substring matching to handle emoji-prefixed tier labels like "🏅 Hot"
    if "hot" in tier or (score is not None and score >= 8.5):
        return HOT_FILL
    if "warm" in tier or (score is not None and score >= 5.0):
        return WARM_FILL
    if "cool" in tier:
        return COOL_FILL
    if "pass" in tier:
        return PASS_FILL
    return None


def _apply_databar_formatting(ws, headers: list[str], first_row: int, last_row: int) -> None:
    if first_row > last_row:
        return
    for col_idx, h in enumerate(headers, start=1):
        if h in DATABAR_COLS:
            col_letter = get_column_letter(col_idx)
            ws.conditional_formatting.add(
                f"{col_letter}{first_row}:{col_letter}{last_row}",
                DataBarRule(
                    start_type="num", start_value=0,
                    end_type="num", end_value=10,
                    color="638EC6",
                    showValue=True,
                ),
            )


# ---------------------------------------------------------------------------
# Debug View sort key
# ---------------------------------------------------------------------------

_PRIORITY_ORDER = {"high": 0, "medium": 1, "low": 2, "": 3}


def _debug_sort_key(rd: dict[str, Any]) -> tuple:
    priority = _str_lower(rd.get(REVIEW_PRIORITY_COL))
    override  = int(_is_truthy(rd.get("icp_override_applied")))
    fhq_sanit = int(_is_truthy(rd.get("foreign_hq_sanitized")))
    # Use derived score for stable sorting
    score = _to_float(_derive_score_fields(rd).get("display_score")) or _final_score(rd) or 0.0
    return (
        _PRIORITY_ORDER.get(priority, 3),
        -override,
        -fhq_sanit,
        -score,
    )


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _col_width_for_full(header: str) -> int:
    widths = {
        "source_queue": 14, "source_file": 50, "source_row": 10,
        EMP_RANGE_COL: 30, "company_name": 40,
        "website_url": 35, "domain": 28,
        "commercial_fit_score": 22, "final_commercial_fit_score": 22,
        "commercial_tier": 16,
    }
    return widths.get(header, max(12, min(len(header) + 4, 40)))


def _write_full_sheet(ws, data_rows: list[list[Any]], output_headers: list[str]) -> None:
    trace_set = set(TRACE_COLS)

    # Header row
    ws.append(output_headers)
    ws.row_dimensions[1].height = DEFAULT_ROW_HEIGHT
    for col_idx, header in enumerate(output_headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        if header in trace_set:
            cell.fill = TRACE_FILL
            cell.font = Font(bold=True)
        elif header == EMP_RANGE_COL:
            cell.fill = EMP_FILL
            cell.font = Font(bold=True)
        else:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=False)
        ws.column_dimensions[get_column_letter(col_idx)].width = _col_width_for_full(header)

    # Data rows
    for data_row_idx, row in enumerate(data_rows, start=2):
        row_values = row[:len(output_headers)]
        ws.append(row_values)
        ws.row_dimensions[data_row_idx].height = DEFAULT_ROW_HEIGHT
        row_dict = dict(zip(output_headers, row_values))
        fill = _get_row_fill(row_dict)
        if fill is not None:
            for col_idx in range(1, len(output_headers) + 1):
                ws.cell(row=data_row_idx, column=col_idx).fill = fill

    last_row = len(data_rows) + 1
    if last_row >= 2:
        _apply_databar_formatting(ws, output_headers, 2, last_row)

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions


def _write_debug_sheet(
    ws, data_rows: list[list[Any]], full_headers: list[str]
) -> tuple[int, int, int, int, int, dict[str, int]]:
    """Write Debug View. Returns (total, fhq_cnt, comp_cnt, origin_cnt, domain_cnt, priority_counts)."""
    row_dicts = [_row_as_dict(r, full_headers) for r in data_rows]
    row_dicts.sort(key=_debug_sort_key)

    # Header row
    ws.append(DEBUG_VIEW_COLS)
    ws.row_dimensions[1].height = DEFAULT_ROW_HEIGHT
    for col_idx, col_name in enumerate(DEBUG_VIEW_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        if col_name in set(DEBUG_FLAG_COLS) or col_name == REVIEW_PRIORITY_COL:
            cell.fill = DEBUG_FILL
            cell.font = Font(bold=True)
        elif col_name in set(TRACE_COLS):
            cell.fill = TRACE_FILL
            cell.font = Font(bold=True)
        elif col_name == EMP_RANGE_COL:
            cell.fill = EMP_FILL
            cell.font = Font(bold=True)
        else:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=False)
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            _DEBUG_COL_WIDTHS.get(col_name, 20)
        )

    # Pre-compute flag column indices (1-based)
    flag_col_map = {
        col_name: col_idx
        for col_idx, col_name in enumerate(DEBUG_VIEW_COLS, start=1)
        if col_name in set(DEBUG_FLAG_COLS)
    }
    priority_col_idx = next(
        (i for i, c in enumerate(DEBUG_VIEW_COLS, start=1) if c == REVIEW_PRIORITY_COL),
        None,
    )

    fhq_cnt = comp_cnt = origin_cnt = domain_cnt = 0
    priority_counts: dict[str, int] = {"High": 0, "Medium": 0, "Low": 0}

    for row_idx, rd in enumerate(row_dicts, start=2):
        # Derive canonical score fields and inject into the debug row dict
        derived = _derive_score_fields(rd)
        for k, v in derived.items():
            if v is not None:
                rd[k] = v

        row_values = [rd.get(col) for col in DEBUG_VIEW_COLS]
        ws.append(row_values)
        ws.row_dimensions[row_idx].height = DEFAULT_ROW_HEIGHT

        # Wrap text in long-text columns only
        for col_idx, col_name in enumerate(DEBUG_VIEW_COLS, start=1):
            if col_name in _WRAP_COLS:
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True)

        # Count flags
        if rd.get("debug_foreign_hq_flag"):
            fhq_cnt += 1
        if rd.get("debug_competitor_flag"):
            comp_cnt += 1
        if rd.get("debug_score_origin_flag"):
            origin_cnt += 1
        if rd.get("debug_domain_flag"):
            domain_cnt += 1

        pri = str(rd.get(REVIEW_PRIORITY_COL) or "").strip()
        if pri in priority_counts:
            priority_counts[pri] += 1

        # Highlight non-empty flag cells
        for flag_col, flag_col_idx in flag_col_map.items():
            if rd.get(flag_col):
                c = ws.cell(row=row_idx, column=flag_col_idx)
                c.fill = FLAG_FILL
                c.font = FLAG_FONT

        # Highlight review priority cell
        if priority_col_idx is not None:
            pri_cell = ws.cell(row=row_idx, column=priority_col_idx)
            if pri == "High":
                pri_cell.fill = HIGH_PRI_FILL
                pri_cell.font = HIGH_PRI_FONT
            elif pri == "Medium":
                pri_cell.fill = MEDIUM_PRI_FILL
                pri_cell.font = MEDIUM_PRI_FONT

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions

    last_row = len(row_dicts) + 1
    _apply_databar_formatting(ws, DEBUG_VIEW_COLS, 2, last_row)

    # CellIsRule highlights for score columns
    for score_col_name in ("final_commercial_fit_score", "display_score"):
        if score_col_name in DEBUG_VIEW_COLS:
            scl = get_column_letter(DEBUG_VIEW_COLS.index(score_col_name) + 1)
            ws.conditional_formatting.add(
                f"{scl}2:{scl}{last_row}",
                CellIsRule(operator="greaterThanOrEqual", formula=["8.5"],
                           fill=PatternFill("solid", fgColor="C6EFCE"),
                           font=Font(bold=True)),
            )
    if "sig_foreign_hq_score" in DEBUG_VIEW_COLS:
        fhq_cl = get_column_letter(DEBUG_VIEW_COLS.index("sig_foreign_hq_score") + 1)
        ws.conditional_formatting.add(
            f"{fhq_cl}2:{fhq_cl}{last_row}",
            CellIsRule(operator="greaterThanOrEqual", formula=["7"],
                       fill=PatternFill("solid", fgColor="FFEB9C")),
        )

    return len(row_dicts), fhq_cnt, comp_cnt, origin_cnt, domain_cnt, priority_counts


def _write_scoring_logic_sheet(ws) -> None:
    rows = [
        ("Scoring Logic – mYngle Opportunity Input Merge", ""),
        ("", ""),
        ("Normal ICP Scoring", "commercial_fit_scoring.py"),
        ("", ""),
        ("Signal fields used in regression/scoring model", ""),
        ("  sig_foreign_hq_score", "Weight: positive. Indicates foreign HQ structure."),
        ("  sig_explicit_lnd_score", "Weight: positive. Explicit L&D-related signals."),
        ("  sig_intl_footprint_score", "Weight: positive. International footprint."),
        ("  sig_employer_branding_score", "Weight: positive. Employer branding signals."),
        ("  sig_lnd_onboarding_score", "Weight: positive. L&D onboarding signals."),
        ("  ti_onboarding_score", "Weight: positive. Technology/implementation signals."),
        ("  sig_rapid_growth_score", "Weight: NEGATIVE coefficient. Rapid growth reduces score."),
        ("", ""),
        ("Competitor Signals", ""),
        ("  competitor_customer_signal", "Detected competitor customer signals."),
        ("  competitor_signal_strength", "Strength: Low / Medium / High."),
        ("  competitor_evidence_url", "URL used as proof of competitor customer signal."),
        ("  icp_override_applied", "YES = competitor override lifted final score to 10."),
        ("  icp_override_reason", "Reason for override (if applied)."),
        ("", ""),
        ("Important: competitor signals are NOT part of the normal LR score.", ""),
        ("Competitor override can lift final score to 10 ONLY when", ""),
        ("strong competitor-customer evidence is accepted.", ""),
        ("", ""),
        ("Current Review Concern", ""),
        ("  Competitor URLs may need manual validation before accepting override.", ""),
        ("  Rows with icp_override_applied=Yes should be reviewed in Debug View.", ""),
        ("", ""),
        ("Foreign HQ Sanitizer", ""),
        ("  If foreign_hq_sanitized=Yes, the original foreign HQ score was reduced.", ""),
        ("  foreign_hq_original_score shows the pre-sanitizer score.", ""),
        ("  Rows where original > final should be reviewed.", ""),
        ("", ""),
        ("Future Workflow", ""),
        ("  Reviewed and validated signals should feed an updated commercial fit score.", ""),
        ("  The Debug View is the primary tool for this review.", ""),
        ("", ""),
        ("Legacy Score Column", ""),
        ("  commercial_fit_score_75_25_legacy", "Fallback score only. Used when no newer final/commercial score is present."),
        ("  Score priority for display_score / final_commercial_fit_score:", ""),
        ("    1. final_commercial_fit_score", "Preferred."),
        ("    2. commercial_fit_score", "Standard."),
        ("    3. Final Commercial Fit Score", "Legacy header name variant."),
        ("    4. commercial_fit_score_75_25_legacy", "Last-resort fallback. Not the preferred score."),
    ]
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 70
    for r_idx, (k, v) in enumerate(rows, start=1):
        cell_a = ws.cell(row=r_idx, column=1, value=k)
        cell_b = ws.cell(row=r_idx, column=2, value=v)
        if r_idx == 1:
            cell_a.font = Font(bold=True, size=13)
        elif k and not k.startswith(" ") and v == "":
            cell_a.font = Font(bold=True)
        ws.row_dimensions[r_idx].height = DEFAULT_ROW_HEIGHT


def write_output_workbook(
    output_path: Path,
    data_rows: list[list[Any]],
    output_headers: list[str],
    qa_meta: dict,
    rows_by_queue: dict[str, int],
    file_schema_info: "list[dict] | None" = None,
    new_in_union: "list[str] | None" = None,
) -> tuple[int, int, int, int, int, dict[str, int], dict[str, int]]:
    """
    Write workbook with Debug View, Opportunity Input Full, Scoring Logic, Merge QA.
    Returns (debug_total, fhq_cnt, comp_cnt, origin_cnt, domain_cnt,
             priority_counts, override_and_score10_counts).
    """
    wb = Workbook()

    # Sheet 1: Debug View
    ws_debug = wb.active
    ws_debug.title = DEBUG_SHEET_NAME
    debug_total, fhq_cnt, comp_cnt, origin_cnt, domain_cnt, priority_counts = (
        _write_debug_sheet(ws_debug, data_rows, output_headers)
    )

    # Sheet 2: Opportunity Input Full
    ws_full = wb.create_sheet(FULL_SHEET_NAME)
    _write_full_sheet(ws_full, data_rows, output_headers)

    # Sheet 3: Scoring Logic
    ws_logic = wb.create_sheet("Scoring Logic")
    _write_scoring_logic_sheet(ws_logic)

    # Sheet 4: Merge QA — compute extra stats from data rows
    extended_headers = output_headers + _COMPUTED_COLS
    rows_fhq_sanitized = 0
    rows_fhq_zeroed = 0
    rows_override = 0
    rows_score_10 = 0
    rows_score_10_override = 0
    rows_comp_url = 0
    comp_url_cols = {"competitor_evidence_url", "competitor_confidence_url",
                     "competitor_attention_url"}

    for row in data_rows:
        rd = {h: (row[i] if i < len(row) else None)
              for i, h in enumerate(extended_headers)}
        score = _final_score(rd)
        if _is_truthy(rd.get("foreign_hq_sanitized")):
            rows_fhq_sanitized += 1
        fhq_s = _to_float(rd.get("sig_foreign_hq_score"))
        fhq_o = _to_float(rd.get("foreign_hq_original_score"))
        if fhq_s == 0.0 and fhq_o is not None and fhq_o > 0:
            rows_fhq_zeroed += 1
        override = _is_truthy(rd.get("icp_override_applied"))
        if override:
            rows_override += 1
        if score is not None and score == 10:
            rows_score_10 += 1
            if override:
                rows_score_10_override += 1
        if any(str(rd.get(c) or "").strip() for c in comp_url_cols):
            rows_comp_url += 1

    ws_qa = wb.create_sheet("Merge QA")
    qa_rows = [
        ("timestamp",                               qa_meta.get("timestamp", "")),
        ("target_root",                             qa_meta.get("target_root", "")),
        ("input_subfolder",                         qa_meta.get("input_subfolder", "")),
        ("queues_included",                         ", ".join(qa_meta.get("queues", []))),
        ("files_scanned",                           qa_meta.get("files_scanned", 0)),
        ("files_merged",                            qa_meta.get("files_merged", 0)),
        ("files_skipped_no_sheet",                  qa_meta.get("files_skipped_no_sheet", 0)),
        ("files_skipped_invalid",                   qa_meta.get("files_skipped_invalid", 0)),
        ("total_rows_merged",                       qa_meta.get("total_rows", 0)),
        ("", ""),
        ("── rows by queue ──", ""),
    ]
    for q, cnt in rows_by_queue.items():
        qa_rows.append((f"  {q}", cnt))
    qa_rows += [
        ("", ""),
        ("── foreign HQ review ──", ""),
        ("rows_with_foreign_hq_sanitized",          rows_fhq_sanitized),
        ("rows_fhq_original_positive_but_zeroed",   rows_fhq_zeroed),
        ("", ""),
        ("── competitor review ──", ""),
        ("rows_with_icp_override_applied",          rows_override),
        ("rows_with_final_score_10",                rows_score_10),
        ("rows_score_10_and_competitor_override",   rows_score_10_override),
        ("rows_with_competitor_url_populated",      rows_comp_url),
        ("", ""),
        ("── review priority ──", ""),
        ("review_priority_High",                    priority_counts.get("High", 0)),
        ("review_priority_Medium",                  priority_counts.get("Medium", 0)),
        ("review_priority_Low",                     priority_counts.get("Low", 0)),
        ("", ""),
        ("── debug view ──", ""),
        ("debug_view_rows",                         debug_total),
        ("rows_with_debug_foreign_hq_flag",         fhq_cnt),
        ("rows_with_debug_competitor_flag",         comp_cnt),
        ("rows_with_debug_score_origin_flag",       origin_cnt),
        ("rows_with_debug_domain_flag",             domain_cnt),
        ("", ""),
        ("output_path",                             qa_meta.get("output_path", "")),
    ]
    ws_qa.column_dimensions["A"].width = 42
    ws_qa.column_dimensions["B"].width = 80
    for r_idx, (k, v) in enumerate(qa_rows, start=1):
        ws_qa.cell(row=r_idx, column=1, value=k).font = Font(bold=True)
        ws_qa.cell(row=r_idx, column=2, value=v)
        ws_qa.row_dimensions[r_idx].height = DEFAULT_ROW_HEIGHT

    # ── Enhanced per-source-file schema section ───────────────────────────────
    if file_schema_info is not None:
        next_row = len(qa_rows) + 1

        # Blank separator + section heading
        next_row += 1  # blank row
        ws_qa.cell(row=next_row, column=1, value="── per-source-file schema ──").font = Font(bold=True)
        ws_qa.cell(row=next_row, column=2, value="")
        ws_qa.row_dimensions[next_row].height = DEFAULT_ROW_HEIGHT
        next_row += 1

        # Table header row across columns A–I
        table_headers = [
            "Queue", "File", "Headers", "Rows Merged",
            "Missing Core Cols", "Extra vs Base (count)",
            "FHQ Fields Present", "Missing FHQ Fields", "FHQ Field Positions",
        ]
        for col_idx, th in enumerate(table_headers, start=1):
            cell = ws_qa.cell(row=next_row, column=col_idx, value=th)
            cell.font = Font(bold=True)
        ws_qa.row_dimensions[next_row].height = DEFAULT_ROW_HEIGHT
        next_row += 1

        # Set column widths for C–I
        col_widths = {3: 20, 4: 20, 5: 50, 6: 25, 7: 22, 8: 50, 9: 70}
        for col_idx, width in col_widths.items():
            ws_qa.column_dimensions[get_column_letter(col_idx)].width = width

        # Data rows
        for info in file_schema_info:
            q = info["queue"]
            f = info["file"]
            num_headers = info["num_headers"]
            num_rows = info["num_rows"]
            missing_core = info["missing_core"]
            extra_vs_base = info["extra_vs_base"]
            fhq_field_presence = info["fhq_field_presence"]
            fhq_field_positions = info["fhq_field_positions"]

            missing_core_str = ", ".join(missing_core) if missing_core else "(none)"
            extra_count = len(extra_vs_base)
            fhq_present_str = (
                f"{sum(fhq_field_presence.values())}/{len(FOREIGN_HQ_SANITIZER_FIELDS)}"
            )
            missing_fhq_str = (
                ", ".join(
                    fld for fld in FOREIGN_HQ_SANITIZER_FIELDS
                    if not fhq_field_presence[fld]
                ) or "(none)"
            )
            fhq_positions_str = ", ".join(
                f"{fld}@col{fhq_field_positions[fld]}"
                for fld in FOREIGN_HQ_SANITIZER_FIELDS
                if fhq_field_positions[fld] is not None
            )

            row_values = [
                q, f, num_headers, num_rows,
                missing_core_str, extra_count,
                fhq_present_str, missing_fhq_str, fhq_positions_str,
            ]
            for col_idx, val in enumerate(row_values, start=1):
                ws_qa.cell(row=next_row, column=col_idx, value=val)
            ws_qa.row_dimensions[next_row].height = DEFAULT_ROW_HEIGHT
            next_row += 1

        # Schema warnings section
        next_row += 1  # blank row
        ws_qa.cell(row=next_row, column=1, value="── schema warnings ──").font = Font(bold=True)
        ws_qa.row_dimensions[next_row].height = DEFAULT_ROW_HEIGHT
        next_row += 1

        red_font = Font(color="FF0000")
        has_warnings = False

        for info in file_schema_info:
            q = info["queue"]
            f = info["file"]
            fhq_field_presence = info["fhq_field_presence"]
            missing_fhq = [
                fld for fld in FOREIGN_HQ_SANITIZER_FIELDS
                if not fhq_field_presence[fld]
            ]
            if missing_fhq:
                has_warnings = True
                ws_qa.cell(
                    row=next_row, column=1,
                    value=f"[WARN] {q}/{f}: missing FHQ fields",
                ).font = red_font
                ws_qa.cell(
                    row=next_row, column=2,
                    value=", ".join(missing_fhq),
                ).font = red_font
                ws_qa.row_dimensions[next_row].height = DEFAULT_ROW_HEIGHT
                next_row += 1

        if new_in_union:
            has_warnings = True
            ws_qa.cell(
                row=next_row, column=1,
                value=f"[WARN] union schema added {len(new_in_union)} column(s) not in first-file schema",
            ).font = red_font
            ws_qa.cell(
                row=next_row, column=2,
                value=", ".join(new_in_union[:20]),
            ).font = red_font
            ws_qa.row_dimensions[next_row].height = DEFAULT_ROW_HEIGHT
            next_row += 1

        if not has_warnings:
            ws_qa.cell(row=next_row, column=1, value="(no schema warnings)")
            ws_qa.row_dimensions[next_row].height = DEFAULT_ROW_HEIGHT

    wb.save(output_path)

    override_and_score10_counts = {
        "rows_override": rows_override,
        "rows_score_10": rows_score_10,
        "rows_score_10_override": rows_score_10_override,
        "rows_comp_url": rows_comp_url,
        "rows_fhq_sanitized": rows_fhq_sanitized,
    }
    return (debug_total, fhq_cnt, comp_cnt, origin_cnt, domain_cnt,
            priority_counts, override_and_score10_counts)


# ---------------------------------------------------------------------------
# CSV report
# ---------------------------------------------------------------------------

def write_csv_report(report_rows: list[dict], report_dir: Path, ts: str) -> Path:
    report_dir.mkdir(parents=True, exist_ok=True)
    csv_path = report_dir / f"merge_opportunity_inputs_report_{ts}.csv"
    fields = ["queue", "file", "status", "rows_read", "notes"]
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for r in report_rows:
            w.writerow({k: r.get(k, "") for k in fields})
    return csv_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description=(
            "Merge 'Opportunity Input' sheets from lead-prioritized folders "
            "and add Chamber of Commerce employee range + debug flag columns."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("--target-root", required=True,
                   help="Root folder, e.g. C:\\Users\\...\\Myngle")
    p.add_argument("--queues", nargs="+", default=["Italy50", "Italy100", "Italy200"],
                   help="Queue folder names (default: Italy50 Italy100 Italy200)")
    p.add_argument("--input-subfolder", default="02_lead_prioritized",
                   help="Subfolder inside each queue to read from "
                        "(default: 02_lead_prioritized)")
    p.add_argument("--output", default=None,
                   help="Full path for the output .xlsx (optional)")
    p.add_argument("--overwrite", action="store_true", default=False,
                   help="Overwrite existing output file")
    p.add_argument("--report-dir", default=None,
                   help="Directory for the CSV report "
                        "(default: <target-root>\\_merge_opportunity_inputs)")
    return p


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    target_root = Path(args.target_root)
    if not target_root.exists():
        sys.exit(f"Target root not found: {target_root}")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    queues: list[str] = args.queues
    input_subfolder: str = args.input_subfolder

    merge_dir = target_root / "_merge_opportunity_inputs"
    report_dir = Path(args.report_dir) if args.report_dir else merge_dir

    if args.output:
        raw_output = Path(args.output)
    else:
        queues_str = "_".join(queues)
        raw_output = merge_dir / (
            f"{queues_str}_ALL_opportunity_input_with_employee_range_{ts}.xlsx"
        )

    output_path = _safe_output_path(raw_output, args.overwrite)

    print(f"\n[MERGE] merge_opportunity_inputs_with_employee_range.py")
    print(f"  Target root      : {target_root}")
    print(f"  Queues           : {queues}")
    print(f"  Input subfolder  : {input_subfolder}")
    print(f"  Output           : {output_path}")

    data_rows, output_headers, report_rows, rows_by_queue, file_schema_info, new_in_union = merge_queues(
        target_root=target_root,
        queues=queues,
        input_subfolder=input_subfolder,
    )

    if output_headers is None:
        sys.exit(
            "\nNo data found. No files with an 'Opportunity Input' sheet were merged."
        )

    files_scanned  = len(report_rows)
    files_merged   = sum(1 for r in report_rows if r["status"] == "MERGED")
    files_no_sheet = sum(1 for r in report_rows
                         if r["status"] == "SKIPPED_MISSING_OPPORTUNITY_INPUT")
    files_invalid  = sum(1 for r in report_rows
                         if r["status"] == "SKIPPED_INVALID_WORKBOOK")
    total_rows     = len(data_rows)

    merge_dir.mkdir(parents=True, exist_ok=True)
    (debug_total, fhq_cnt, comp_cnt, origin_cnt, domain_cnt,
     priority_counts, extra_counts) = write_output_workbook(
        output_path=output_path,
        data_rows=data_rows,
        output_headers=output_headers,
        qa_meta={
            "timestamp": ts,
            "target_root": str(target_root),
            "input_subfolder": input_subfolder,
            "queues": queues,
            "files_scanned": files_scanned,
            "files_merged": files_merged,
            "files_skipped_no_sheet": files_no_sheet,
            "files_skipped_invalid": files_invalid,
            "total_rows": total_rows,
            "output_path": str(output_path),
        },
        rows_by_queue=rows_by_queue,
        file_schema_info=file_schema_info,
        new_in_union=new_in_union,
    )

    csv_path = write_csv_report(report_rows, report_dir, ts)

    print("\n" + "=" * 66)
    print("MERGE SUMMARY")
    print("=" * 66)
    print(f"  Files scanned                    : {files_scanned}")
    print(f"  Files merged                     : {files_merged}")
    print(f"  Total rows merged                : {total_rows}")
    for q, cnt in rows_by_queue.items():
        print(f"    {q:<20}           : {cnt}")
    print(f"  Skipped (no sheet)               : {files_no_sheet}")
    print(f"  Skipped (invalid wb)             : {files_invalid}")
    print()
    print(f"  Foreign HQ sanitized             : {extra_counts['rows_fhq_sanitized']}")
    print(f"  Competitor override applied      : {extra_counts['rows_override']}")
    print(f"  Final score = 10                 : {extra_counts['rows_score_10']}")
    print(f"  Score 10 + competitor override   : {extra_counts['rows_score_10_override']}")
    print(f"  Competitor URL populated         : {extra_counts['rows_comp_url']}")
    print()
    print(f"  Review priority High             : {priority_counts.get('High', 0)}")
    print(f"  Review priority Medium           : {priority_counts.get('Medium', 0)}")
    print(f"  Review priority Low              : {priority_counts.get('Low', 0)}")
    print()
    print(f"  Debug View rows                  : {debug_total}")
    print(f"    debug_foreign_hq_flag          : {fhq_cnt}")
    print(f"    debug_competitor_flag          : {comp_cnt}")
    print(f"    debug_score_origin_flag        : {origin_cnt}")
    print(f"    debug_domain_flag              : {domain_cnt}")
    print()
    print(f"  Output file                      : {output_path}")
    print(f"  CSV report                       : {csv_path}")
    print("=" * 66 + "\n")

    if files_no_sheet or files_invalid:
        print("Skipped files:")
        for r in report_rows:
            if r["status"] != "MERGED":
                print(f"  [{r['status']}] {r['queue']} / {r['file']}")
                if r["notes"]:
                    print(f"    {r['notes']}")
        print()


if __name__ == "__main__":
    main()
