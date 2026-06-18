"""
patch_lead_prioritized_urls_from_cleaned_domains.py

For each queue, reads corrected URLs from cleaned-domain files in:
  <target-root>\<queue>\01_cleaned_domains\*.xlsx

and patches them into enriched files in:
  <target-root>\<queue>\02_lead_prioritized\*.xlsx

Patched copies are written to:
  <target-root>\<queue>\02_lead_prioritized_url_patched\

Default target sheet: "Opportunity Input" only.

Usage:
  python patch_lead_prioritized_urls_from_cleaned_domains.py \
      --target-root "C:\\Users\\...\\Myngle" \
      --queues Italy100 Italy200 \
      --dry-run | --apply \
      [--target-sheets "Opportunity Input"] \
      [--patch-all-sheets] \
      [--in-place] [--overwrite] [--report-dir "<path>"]
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
import unicodedata
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

# ---------------------------------------------------------------------------
# Dependency check
# ---------------------------------------------------------------------------
try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

PREFERRED_SOURCE_SHEETS = ["Best Guess Input", "Cleaned Register Input"]
DEFAULT_TARGET_SHEETS = ["Opportunity Input"]

ITALIAN_LEGAL_SUFFIXES = [
    r"s\.?p\.?a\.?",
    r"s\.?r\.?l\.?",
    r"societa[\s]*per[\s]*azioni",
    r"societa[\s]*a[\s]*responsabilita[\s]*limitata",
    r"in[\s]*forma[\s]*abbreviata",
    r"siglabile",
    r"s\.?n\.?c\.?",
    r"s\.?a\.?s\.?",
    r"s\.?c\.?a\.?r\.?l\.?",
    r"s\.?c\.?",
    r"onlus",
    r"ets",
    r"aps",
    r"asd",
    r"spa",
    r"srl",
    r"snc",
    r"sas",
    r"scarl",
]

URL_COLUMNS = ["website_url", "canonical_company_url"]
DOMAIN_COLUMNS = [
    "domain",
    "validated_domain",
    "input_domain",
    "canonical_company_domain",
    "final_selected_domain",
    "recommended_domain",
]
AUDIT_COLUMNS = [
    "url_patch_status",
    "url_patch_old_website_url",
    "url_patch_new_website_url",
    "url_patch_old_domain",
    "url_patch_new_domain",
    "url_patch_match_key",
    "url_patch_source_file",
    "url_patch_source_sheet",
    "url_patch_match_scope",
    "needs_reenrichment",
]
UNSAFE_PATH_FRAGMENTS = [
    "_url_patched",
    "url_patched",
    "_url_patch_reports",
    "_logs",
    "_archive",
    "__pycache__",
]

# Statuses
ST_CHANGED = "PATCHED_URL_CHANGED"
ST_SAME = "PATCHED_URL_SAME"
ST_NO_MATCH = "NO_SOURCE_MATCH"
ST_SRC_EMPTY = "SOURCE_URL_EMPTY"
ST_DUPLICATE = "DUPLICATE_SOURCE_MATCH_SKIPPED"
ST_TGT_EMPTY = "TARGET_COMPANY_EMPTY"
ST_OPP_MISSING = "OPPORTUNITY_INPUT_SHEET_MISSING"
ST_INVALID_WB = "INVALID_WORKBOOK_SKIPPED"

# ---------------------------------------------------------------------------
# Normalisation helpers
# ---------------------------------------------------------------------------

def _remove_accents(text: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", text)
        if unicodedata.category(c) != "Mn"
    )


_LEGAL_RE = re.compile(
    r"\b(" + "|".join(ITALIAN_LEGAL_SUFFIXES) + r")\b",
    re.IGNORECASE,
)
_PUNCT_RE = re.compile(r"[^\w\s]", re.UNICODE)
_WS_RE = re.compile(r"\s+")


def normalize_company(name: Any) -> str:
    if not name or not str(name).strip():
        return ""
    s = _remove_accents(str(name).strip()).lower()
    s = _LEGAL_RE.sub(" ", s)
    s = _PUNCT_RE.sub(" ", s)
    return _WS_RE.sub(" ", s).strip()


def normalize_geo(value: Any) -> str:
    if not value:
        return ""
    s = _remove_accents(str(value).strip().lower())
    s = _PUNCT_RE.sub(" ", s)
    return _WS_RE.sub(" ", s).strip()


# ---------------------------------------------------------------------------
# URL / domain helpers
# ---------------------------------------------------------------------------

def extract_domain(url_or_domain: Any) -> str:
    if not url_or_domain:
        return ""
    s = str(url_or_domain).strip()
    if not s:
        return ""
    if "://" not in s:
        s = "https://" + s
    parsed = urlparse(s)
    domain = parsed.netloc or parsed.path
    return re.sub(r"^www\.", "", domain.lower()).strip("/")


def to_full_url(url_or_domain: Any) -> str:
    domain = extract_domain(url_or_domain)
    return f"https://{domain}" if domain else ""


# ---------------------------------------------------------------------------
# Batch range parsing
# ---------------------------------------------------------------------------

_RANGE_RE = re.compile(r"[Rr](\d+)[_\-](\d+)")


def parse_batch_range(filename: str) -> tuple[int, int] | None:
    """Return (start, end) row range from filename, or None."""
    m = _RANGE_RE.search(filename)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None


def ranges_overlap(a: tuple[int, int], b: tuple[int, int]) -> bool:
    return a[0] <= b[1] and b[0] <= a[1]


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------

def _safe_load(path: Path) -> openpyxl.Workbook | None:
    try:
        return load_workbook(path, data_only=False)
    except (zipfile.BadZipFile, KeyError, OSError, Exception) as exc:
        print(f"[patch] Skipping invalid workbook: {path} | {exc}")
        return None


def _pick_source_sheet(wb: openpyxl.Workbook):
    for name in PREFERRED_SOURCE_SHEETS:
        if name in wb.sheetnames:
            return wb[name]
    return wb.active


def _get_headers(sheet) -> list[str]:
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
    return [str(h).strip() if h is not None else "" for h in first_row]


def _col_index_ci(headers: list[str], name: str) -> int | None:
    lo = name.lower()
    for i, h in enumerate(headers):
        if h.lower() == lo:
            return i
    return None


def _is_safe_input_path(path: Path) -> bool:
    parts = [p.lower() for p in path.parts]
    for frag in UNSAFE_PATH_FRAGMENTS:
        if any(frag in part for part in parts):
            return False
    return True


# ---------------------------------------------------------------------------
# Lookup table for one set of source files
# ---------------------------------------------------------------------------

class BatchLookup:
    """
    Holds URL lookups built from one or more cleaned-domain source files.
    Supports primary (company), secondary (company+city), tertiary (company+province).
    """

    def __init__(self, scope: str = "batch") -> None:
        self.scope = scope          # "batch" | "queue-wide"
        self.primary: dict[str, str] = {}
        self.secondary: dict[str, str] = {}
        self.tertiary: dict[str, str] = {}
        self.duplicates: set[str] = set()
        self.source_file: str = ""
        self.source_sheet: str = ""
        self.rows_with_url = 0
        self.total_rows = 0

    def ingest(self, wb: openpyxl.Workbook, filename: str) -> None:
        sheet = _pick_source_sheet(wb)
        if not self.source_file:
            self.source_file = filename
        if not self.source_sheet:
            self.source_sheet = sheet.title
        headers = _get_headers(sheet)

        co_col = _col_index_ci(headers, "company_name")
        url_col = _col_index_ci(headers, "website_url")
        city_col = _col_index_ci(headers, "city")
        prov_col = _col_index_ci(headers, "province")

        if co_col is None or url_col is None:
            return

        for row in sheet.iter_rows(min_row=2, values_only=True):
            self.total_rows += 1
            company_raw = row[co_col] if co_col < len(row) else None
            url_raw = row[url_col] if url_col < len(row) else None

            key = normalize_company(company_raw)
            if not key:
                continue
            url = str(url_raw).strip() if url_raw else ""
            if not url:
                continue
            self.rows_with_url += 1

            if city_col is not None and city_col < len(row):
                city = normalize_geo(row[city_col])
                if city:
                    self.secondary[f"{key}||{city}"] = url
            if prov_col is not None and prov_col < len(row):
                prov = normalize_geo(row[prov_col])
                if prov:
                    self.tertiary[f"{key}||prov||{prov}"] = url

            if key in self.primary and self.primary[key] != url:
                self.duplicates.add(key)
            else:
                self.primary[key] = url

    def lookup(self, company: Any, city: Any = None, province: Any = None) -> tuple[str, str, str]:
        """Returns (url, find_status, match_key). find_status: FOUND|DUPLICATE|NOT_FOUND|EMPTY"""
        key = normalize_company(company)
        if not key:
            return "", "EMPTY", ""
        if city:
            sk = f"{key}||{normalize_geo(city)}"
            if sk in self.secondary:
                return self.secondary[sk], "FOUND", sk
        if province:
            tk = f"{key}||prov||{normalize_geo(province)}"
            if tk in self.tertiary:
                return self.tertiary[tk], "FOUND", tk
        if key in self.duplicates:
            return "", "DUPLICATE", key
        if key in self.primary:
            return self.primary[key], "FOUND", key
        return "", "NOT_FOUND", key


# ---------------------------------------------------------------------------
# Build lookup tables from source folder
# ---------------------------------------------------------------------------

def build_lookups_for_queue(
    cleaned_dir: Path,
) -> tuple[dict[tuple[int, int], BatchLookup], BatchLookup]:
    """
    Returns:
      batch_lookups: map of (start, end) -> BatchLookup for each ranged file
      queue_lookup:  single BatchLookup from ALL source files combined
    """
    batch_lookups: dict[tuple[int, int], BatchLookup] = {}
    queue_lookup = BatchLookup(scope="queue-wide")
    queue_lookup.source_file = "(all cleaned files)"
    queue_lookup.source_sheet = "(multiple)"

    xlsx_files = [
        f for f in sorted(cleaned_dir.glob("*.xlsx"))
        if not f.name.startswith("~$") and _is_safe_input_path(f)
    ]

    for src_path in xlsx_files:
        wb = _safe_load(src_path)
        if wb is None:
            continue

        batch_range = parse_batch_range(src_path.name)

        # Per-batch lookup
        if batch_range is not None:
            bl = batch_lookups.get(batch_range)
            if bl is None:
                bl = BatchLookup(scope="batch")
                batch_lookups[batch_range] = bl
            bl.ingest(wb, src_path.name)

        # Queue-wide lookup always gets all rows
        queue_lookup.ingest(wb, src_path.name)

    return batch_lookups, queue_lookup


def find_best_lookup(
    target_filename: str,
    batch_lookups: dict[tuple[int, int], BatchLookup],
    queue_lookup: BatchLookup,
) -> BatchLookup:
    """Return the most specific BatchLookup for a given target filename."""
    target_range = parse_batch_range(target_filename)
    if target_range is not None:
        # Prefer exact range match, then any overlapping batch
        for rng, bl in batch_lookups.items():
            if rng == target_range:
                return bl
        for rng, bl in batch_lookups.items():
            if ranges_overlap(rng, target_range):
                return bl
    # Fall back to queue-wide
    queue_lookup.scope = "queue-wide"
    return queue_lookup


# ---------------------------------------------------------------------------
# Core patching logic (single sheet)
# ---------------------------------------------------------------------------

def patch_sheet(
    sheet,
    lookup: BatchLookup,
    dry_run: bool,
    only_changed: bool,
) -> list[dict]:
    headers = _get_headers(sheet)
    company_col = _col_index_ci(headers, "company_name")
    if company_col is None:
        return []

    city_col = _col_index_ci(headers, "city")
    prov_col = _col_index_ci(headers, "province")

    url_cols: dict[str, int] = {
        n: i for n in URL_COLUMNS
        if (i := _col_index_ci(headers, n)) is not None
    }
    domain_cols: dict[str, int] = {
        n: i for n in DOMAIN_COLUMNS
        if (i := _col_index_ci(headers, n)) is not None
    }

    # Ensure audit columns exist (append at end if missing)
    audit_col_indices: dict[str, int] = {}
    for ac in AUDIT_COLUMNS:
        idx = _col_index_ci(headers, ac)
        if idx is None:
            new_idx = len(headers)
            headers.append(ac)
            if not dry_run:
                sheet[f"{get_column_letter(new_idx + 1)}1"] = ac
            audit_col_indices[ac] = new_idx
        else:
            audit_col_indices[ac] = idx

    report_rows: list[dict] = []

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        def cell_val(col_idx: int | None) -> Any:
            if col_idx is None or col_idx >= len(row):
                return None
            return row[col_idx].value

        company_raw = cell_val(company_col)
        city_raw = cell_val(city_col)
        prov_raw = cell_val(prov_col)

        if not company_raw or not str(company_raw).strip():
            _write_audit(sheet, row_idx, audit_col_indices, {
                "url_patch_status": ST_TGT_EMPTY,
                "url_patch_match_scope": lookup.scope,
                "url_patch_source_file": lookup.source_file,
                "url_patch_source_sheet": lookup.source_sheet,
                "needs_reenrichment": False,
            }, dry_run)
            report_rows.append(_make_row(
                company="", status=ST_TGT_EMPTY, match_scope=lookup.scope,
                source_file=lookup.source_file,
            ))
            report_rows[-1]["row_number"] = row_idx
            continue

        new_url, find_status, match_key = lookup.lookup(company_raw, city_raw, prov_raw)

        old_website_url = str(cell_val(url_cols.get("website_url")) or "").strip()
        old_domain = ""
        for dc in DOMAIN_COLUMNS:
            if dc in domain_cols:
                v = str(cell_val(domain_cols[dc]) or "").strip()
                if v:
                    old_domain = v
                    break

        if find_status == "EMPTY":
            status, new_url = ST_TGT_EMPTY, ""
        elif find_status == "NOT_FOUND":
            status, new_url = ST_NO_MATCH, ""
        elif find_status == "DUPLICATE":
            status, new_url = ST_DUPLICATE, ""
        else:
            if not new_url:
                status = ST_SRC_EMPTY
            else:
                new_domain = extract_domain(new_url)
                if new_domain and new_domain == extract_domain(old_website_url):
                    status = ST_SAME
                else:
                    status = ST_CHANGED

        new_domain = extract_domain(new_url) if new_url else ""
        needs_reenrich = bool(new_domain and new_domain != extract_domain(old_website_url))

        if not dry_run and new_url and status in (ST_CHANGED, ST_SAME):
            if not only_changed or status == ST_CHANGED:
                full_url = to_full_url(new_url)
                domain_only = extract_domain(new_url)
                for col_idx in url_cols.values():
                    sheet[f"{get_column_letter(col_idx + 1)}{row_idx}"] = full_url
                for col_idx in domain_cols.values():
                    sheet[f"{get_column_letter(col_idx + 1)}{row_idx}"] = domain_only

        _write_audit(sheet, row_idx, audit_col_indices, {
            "url_patch_status": status,
            "url_patch_old_website_url": old_website_url,
            "url_patch_new_website_url": to_full_url(new_url) if new_url else "",
            "url_patch_old_domain": old_domain,
            "url_patch_new_domain": new_domain,
            "url_patch_match_key": match_key,
            "url_patch_source_file": lookup.source_file,
            "url_patch_source_sheet": lookup.source_sheet,
            "url_patch_match_scope": lookup.scope,
            "needs_reenrichment": needs_reenrich,
        }, dry_run)

        r = _make_row(
            company=str(company_raw).strip(),
            status=status,
            old_website_url=old_website_url,
            new_website_url=to_full_url(new_url) if new_url else "",
            old_domain=old_domain,
            new_domain=new_domain,
            match_key=match_key,
            match_scope=lookup.scope,
            source_file=lookup.source_file,
            needs_reenrichment=needs_reenrich,
        )
        r["row_number"] = row_idx
        report_rows.append(r)

    return report_rows


def _make_row(**kwargs) -> dict:
    defaults = {
        "company_name": "",
        "status": "",
        "old_website_url": "",
        "new_website_url": "",
        "old_domain": "",
        "new_domain": "",
        "match_key": "",
        "match_scope": "",
        "source_file": "",
        "needs_reenrichment": False,
        "notes": "",
        "row_number": "",
    }
    defaults.update(kwargs)
    return defaults


def _write_audit(sheet, row_idx: int, audit_indices: dict, values: dict, dry_run: bool) -> None:
    if dry_run:
        return
    for col_name, col_idx in audit_indices.items():
        v = values.get(col_name, "")
        sheet[f"{get_column_letter(col_idx + 1)}{row_idx}"] = v


# ---------------------------------------------------------------------------
# File-level processing
# ---------------------------------------------------------------------------

def _resolve_output_path(output_dir: Path, filename: str, overwrite: bool) -> Path:
    candidate = output_dir / filename
    if not candidate.exists() or overwrite:
        return candidate
    stem, suffix = Path(filename).stem, Path(filename).suffix
    n = 2
    while True:
        candidate = output_dir / f"{stem}_{n}{suffix}"
        if not candidate.exists():
            return candidate
        n += 1


def process_target_file(
    xlsx_path: Path,
    output_dir: Path,
    batch_lookups: dict[tuple[int, int], BatchLookup],
    queue_lookup: BatchLookup,
    dry_run: bool,
    in_place: bool,
    overwrite: bool,
    target_sheets: list[str],
    patch_all_sheets: bool,
) -> tuple[list[dict], bool]:
    """Returns (report_rows, target_sheet_found)."""
    wb = _safe_load(xlsx_path)
    if wb is None:
        return ([{
            **_make_row(status=ST_INVALID_WB, notes="Could not open workbook"),
            "target_file": xlsx_path.name,
            "target_sheet": "",
            "source_file": "",
        }], False)

    lookup = find_best_lookup(xlsx_path.name, batch_lookups, queue_lookup)
    all_rows: list[dict] = []
    target_sheet_found = False

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        headers = _get_headers(sheet)
        has_company = _col_index_ci(headers, "company_name") is not None

        if patch_all_sheets:
            should_patch = has_company
        else:
            should_patch = sheet_name in target_sheets and has_company

        if sheet_name in target_sheets and has_company:
            target_sheet_found = True

        if not should_patch:
            continue

        rows = patch_sheet(sheet, lookup, dry_run, only_changed=False)
        for r in rows:
            r["target_file"] = xlsx_path.name
            r["target_sheet"] = sheet_name
            r.setdefault("source_file", lookup.source_file)
        all_rows.extend(rows)

    if not target_sheet_found and not patch_all_sheets:
        all_rows.append({
            **_make_row(status=ST_OPP_MISSING,
                        notes=f"None of {target_sheets} found"),
            "target_file": xlsx_path.name,
            "target_sheet": "",
            "source_file": lookup.source_file,
        })

    if not dry_run and all_rows:
        if in_place:
            wb.save(xlsx_path)
        else:
            output_dir.mkdir(parents=True, exist_ok=True)
            out_path = _resolve_output_path(output_dir, xlsx_path.name, overwrite)
            wb.save(out_path)

    return all_rows, target_sheet_found


# ---------------------------------------------------------------------------
# Report
# ---------------------------------------------------------------------------

REPORT_FIELDS = [
    "queue", "source_file", "target_file", "target_sheet", "row_number",
    "company_name", "old_website_url", "new_website_url",
    "old_domain", "new_domain", "status", "needs_reenrichment",
    "match_key", "match_scope", "notes",
]


def write_report(rows: list[dict], report_dir: Path, queues: list[str]) -> Path:
    report_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = report_dir / f"lead_prio_url_patch_{ts}_{'_'.join(queues)}.csv"
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=REPORT_FIELDS, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in REPORT_FIELDS})
    return path


def print_summary(
    all_rows: list[dict],
    queue_lookup: BatchLookup,
    target_sheets: list[str],
    files_processed: int,
    output_dir: Path,
) -> None:
    sc: dict[str, int] = {}
    for r in all_rows:
        sc[r.get("status", "")] = sc.get(r.get("status", ""), 0) + 1

    opp_rows = [r for r in all_rows if r.get("target_sheet") in target_sheets]
    osc: dict[str, int] = {}
    for r in opp_rows:
        osc[r.get("status", "")] = osc.get(r.get("status", ""), 0) + 1

    needs_re = sum(1 for r in opp_rows if r.get("needs_reenrichment"))

    print("\n" + "=" * 66)
    print("LEAD PRIORITIZED URL PATCH SUMMARY")
    print("=" * 66)
    print(f"  Source rows with URL (queue-wide) : {queue_lookup.rows_with_url}")
    print(f"  Duplicate source keys             : {len(queue_lookup.duplicates)}")
    print(f"  Target files processed            : {files_processed}")
    print()
    ts_label = " / ".join(target_sheets)
    print(f"  [{ts_label}]:")
    print(f"    Rows seen                       : {len(opp_rows)}")
    print(f"    Rows patched (URL changed)      : {osc.get(ST_CHANGED, 0)}")
    print(f"    Rows patched (URL same)         : {osc.get(ST_SAME, 0)}")
    print(f"    Rows no source match            : {osc.get(ST_NO_MATCH, 0)}")
    print(f"    Rows skipped (duplicate)        : {osc.get(ST_DUPLICATE, 0)}")
    print(f"    Rows source URL empty           : {osc.get(ST_SRC_EMPTY, 0)}")
    print(f"    Rows target company empty       : {osc.get(ST_TGT_EMPTY, 0)}")
    print(f"    Rows needing re-enrichment      : {needs_re}")
    print()
    print(f"  Files missing target sheet        : {sc.get(ST_OPP_MISSING, 0)}")
    print(f"  Invalid workbooks skipped         : {sc.get(ST_INVALID_WB, 0)}")
    print(f"  Output folder                     : {output_dir}")
    print("=" * 66 + "\n")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description=(
            "Patch corrected URLs from 01_cleaned_domains into 02_lead_prioritized files."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("--target-root", required=True,
                   help="Root folder, e.g. C:\\Users\\...\\Myngle")
    p.add_argument("--queues", nargs="+", default=["Italy100"],
                   help="Queue folder names under target-root")
    p.add_argument("--dry-run", action="store_true",
                   help="Preview changes without writing files")
    p.add_argument("--apply", action="store_true",
                   help="Write patched files to output folder")
    p.add_argument("--target-sheets", nargs="+", default=DEFAULT_TARGET_SHEETS,
                   help="Sheet names to patch (default: 'Opportunity Input')")
    p.add_argument("--patch-all-sheets", action="store_true", default=False,
                   help="Patch every sheet that has company_name")
    p.add_argument("--in-place", action="store_true", default=False,
                   help="Overwrite original enriched files (dangerous)")
    p.add_argument("--overwrite", action="store_true", default=False,
                   help="Overwrite existing output files instead of adding _2, _3 suffix")
    p.add_argument("--report-dir", default=None,
                   help="Directory for CSV report (default: <target-root>\\_url_patch_reports)")
    return p


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    if not args.dry_run and not args.apply:
        print("Neither --dry-run nor --apply specified. Defaulting to --dry-run.")
        args.dry_run = True
    if args.dry_run and args.apply:
        sys.exit("Pass either --dry-run or --apply, not both.")

    target_root = Path(args.target_root)
    if not target_root.exists():
        sys.exit(f"Target root not found: {target_root}")

    report_dir = (
        Path(args.report_dir) if args.report_dir
        else target_root / "_url_patch_reports"
    )
    target_sheets: list[str] = args.target_sheets
    patch_all: bool = args.patch_all_sheets

    mode = "DRY-RUN" if args.dry_run else "APPLY"
    print(f"\n[{mode}] patch_lead_prioritized_urls_from_cleaned_domains.py")

    all_report_rows: list[dict] = []
    last_output_dir = target_root

    for queue in args.queues:
        queue_dir = target_root / queue
        if not queue_dir.exists():
            print(f"\n[WARN] Queue folder not found, skipping: {queue_dir}")
            continue

        cleaned_dir = queue_dir / "01_cleaned_domains"
        lead_prio_dir = queue_dir / "02_lead_prioritized"
        output_dir = queue_dir / "02_lead_prioritized_url_patched"
        last_output_dir = output_dir

        if not cleaned_dir.exists():
            print(f"\n[WARN] Cleaned-domains folder not found, skipping: {cleaned_dir}")
            continue
        if not lead_prio_dir.exists():
            print(f"\n[WARN] Lead-prioritized folder not found, skipping: {lead_prio_dir}")
            continue

        if output_dir.exists():
            print(
                f"\n[INFO] Output folder already exists: {output_dir}\n"
                f"       Existing files will not be used as input."
            )

        print(f"\nQueue: {queue}")
        print(f"  Source : {cleaned_dir}")
        print(f"  Target : {lead_prio_dir}")
        print(f"  Output : {output_dir}")

        # Build lookup tables from cleaned-domain files
        batch_lookups, queue_lookup = build_lookups_for_queue(cleaned_dir)
        print(
            f"  Cleaned files read : {len(batch_lookups)} batch(es) + queue-wide"
        )
        print(f"  Source rows w/ URL : {queue_lookup.rows_with_url}")
        print(f"  Duplicate keys     : {len(queue_lookup.duplicates)}")
        if patch_all:
            print("  Sheet mode         : ALL sheets with company_name")
        else:
            print(f"  Target sheets      : {target_sheets}")

        # Collect target enriched files
        target_files = [
            f for f in sorted(lead_prio_dir.glob("*.xlsx"))
            if not f.name.startswith("~$") and _is_safe_input_path(f)
        ]
        print(f"  Target files       : {len(target_files)}")

        files_processed = 0
        for xlsx_path in target_files:
            print(f"  Processing: {xlsx_path.name}")
            rows, sheet_found = process_target_file(
                xlsx_path=xlsx_path,
                output_dir=output_dir,
                batch_lookups=batch_lookups,
                queue_lookup=queue_lookup,
                dry_run=args.dry_run,
                in_place=args.in_place,
                overwrite=args.overwrite,
                target_sheets=target_sheets,
                patch_all_sheets=patch_all,
            )
            opp_rows = [r for r in rows if r.get("target_sheet") in target_sheets]
            changed = sum(1 for r in opp_rows if r.get("status") == ST_CHANGED)
            no_match = sum(1 for r in opp_rows if r.get("status") == ST_NO_MATCH)
            print(
                f"    opp_rows={len(opp_rows)}"
                f"  changed={changed}"
                f"  no_match={no_match}"
                f"  sheet={'YES' if sheet_found else 'NO'}"
            )
            for r in rows:
                r["queue"] = queue
                r.setdefault("target_file", xlsx_path.name)
            all_report_rows.extend(rows)
            files_processed += 1

        print_summary(
            [r for r in all_report_rows if r.get("queue") == queue],
            queue_lookup,
            target_sheets,
            files_processed,
            output_dir,
        )

    report_path = write_report(all_report_rows, report_dir, args.queues)
    print(f"Report written to: {report_path}")

    if args.dry_run:
        print("Dry-run complete. No files were written.")
        print("Re-run with --apply to apply patches.\n")
    else:
        print("Apply complete.\n")


if __name__ == "__main__":
    main()
