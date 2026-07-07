"""
simple_serper_domain_benchmark.py
----------------------------------
Minimal benchmark cleaner: one Serper call per company, take the best
usable top-organic-result domain.

Purpose: compare a dead-simple "Google top result" baseline against the
full input_cleaner_register_edition.py pipeline.  Do NOT use this for
production enrichment.

Usage:
    # One file, 25 rows:
    python simple_serper_domain_benchmark.py \\
        --input Italy50/00_raw/Italy50_1_R0001_0500.xlsx \\
        --output-dir Italy50/02_benchmark \\
        --serper-key sk-xxx \\
        --limit 25

    # Entire folder:
    python simple_serper_domain_benchmark.py \\
        --input Italy50/00_raw \\
        --output-dir Italy50/02_benchmark \\
        --serper-key sk-xxx

    # Dry-run (no API calls):
    python simple_serper_domain_benchmark.py \\
        --input Italy50/00_raw/Italy50_1_R0001_0500.xlsx \\
        --output-dir /tmp \\
        --dry-run

    # Self-test (no API calls):
    python simple_serper_domain_benchmark.py --self-test
"""

import argparse
import json
import os
import re
import sys
import time
import urllib.parse
from datetime import datetime
from pathlib import Path

import requests

# ---------------------------------------------------------------------------
# Bad-domain blocklist
# These are directories, registers, social networks and aggregators that
# should never be accepted as a company's own official website.
# Keep this list simple and explicit — no regex, no substring logic.
# ---------------------------------------------------------------------------

_BAD_DOMAINS_EXACT: frozenset = frozenset({
    "linkedin.com", "facebook.com", "instagram.com", "twitter.com", "x.com",
    "wikipedia.org", "youtube.com",
    "kompass.com", "kompass.it",
    "fatturatoitalia.it", "fatturatoaziende.com",
    "visura.pro", "abbrevia.it", "altervista.org",
    "registroimprese.it", "infocamere.it",
    "paginegialle.it", "paginebianche.it",
    "informazione-aziende.it", "reportaziende.it", "companyreports.it",
    "crunchbase.com", "dnb.com", "zoominfo.com",
    "atoka.io", "europages.com", "europages.it",
    "cerved.com", "trustpilot.com",
})

# Wildcard patterns: any domain that ends with these base domains is bad.
_BAD_DOMAIN_BASES: tuple = (
    "linkedin.com", "facebook.com", "instagram.com",
    "wikipedia.org",
    "glassdoor.com",   # catches glassdoor.it, de.glassdoor.com, etc.
    "indeed.com",      # catches indeed.it, indeed.de, etc.
    "kompass.com", "kompass.it",
    "altervista.org",
    "wordpress.com", "blogspot.com", "wixsite.com", "weebly.com",
    "sites.google.com",
)


def _is_bad_domain(domain: str) -> bool:
    """Return True when domain is on the blocklist or matches a wildcard base."""
    if not domain:
        return True
    dl = domain.lower()
    if dl in _BAD_DOMAINS_EXACT:
        return True
    for base in _BAD_DOMAIN_BASES:
        if dl == base or dl.endswith("." + base):
            return True
    return False


# ---------------------------------------------------------------------------
# URL → root domain
# ---------------------------------------------------------------------------

def _extract_root_domain(url: str) -> str:
    """
    Normalise a URL to its root domain.

    Examples:
      https://www.example.it/about  →  example.it
      http://it.example.com/page    →  example.com  (country subdomain stripped)
      https://subsidiary.example.com →  example.com
    """
    if not url:
        return ""
    url = url.strip()
    if not url.startswith(("http://", "https://")):
        url = "https://" + url
    try:
        parsed = urllib.parse.urlparse(url)
        host   = parsed.netloc.lower()
    except Exception:
        return ""
    # Remove port
    host = host.split(":")[0]
    # Remove www.
    if host.startswith("www."):
        host = host[4:]
    # Remove single-letter or two-letter country-code subdomains that are not
    # part of the registered domain (e.g. "it.example.com" → "example.com",
    # but "q8.it" should stay "q8.it").
    # Heuristic: strip a leading segment that is 1-2 lower-case letters only.
    parts = host.split(".")
    if len(parts) >= 3 and re.fullmatch(r"[a-z]{1,2}", parts[0]):
        host = ".".join(parts[1:])
    return host


# ---------------------------------------------------------------------------
# Column detection
# ---------------------------------------------------------------------------

_COMPANY_COL_CANDIDATES = [
    "company name", "company", "ragione sociale", "denominazione",
    "business name", "name", "nome", "firma",
]
_CITY_COL_CANDIDATES = [
    "city", "città", "comune", "location", "sede", "town", "municipality",
]
_PROVINCE_COL_CANDIDATES = [
    "province", "provincia", "prov", "region", "regione",
    "national statistical institute province",
]
_WEBSITE_COL_CANDIDATES = [
    "website", "web", "url", "domain", "sito web", "sito", "homepage",
    "web site", "company website", "company url",
]


def _find_col(columns: list, candidates: list) -> str:
    """Return the first column name that matches any candidate (case-insensitive)."""
    lower_cols = {c.lower(): c for c in columns}
    for cand in candidates:
        if cand.lower() in lower_cols:
            return lower_cols[cand.lower()]
    return ""


# ---------------------------------------------------------------------------
# Serper API
# ---------------------------------------------------------------------------

_SERPER_URL     = "https://google.serper.dev/search"
_REQUEST_TIMEOUT = 10  # seconds


def _serper_search(query: str, api_key: str) -> tuple[list, str]:
    """
    Call Serper Google Search API.
    Returns (organic_results, error_message).
    organic_results is a list of dicts with keys: title, link, snippet.
    """
    headers = {
        "X-API-KEY":    api_key,
        "Content-Type": "application/json",
    }
    payload = {"q": query, "num": 10}
    try:
        resp = requests.post(
            _SERPER_URL, headers=headers, json=payload, timeout=_REQUEST_TIMEOUT,
        )
        resp.raise_for_status()
        data    = resp.json()
        organic = data.get("organic", [])
        return organic, ""
    except requests.HTTPError as exc:
        return [], f"HTTP {exc.response.status_code}: {exc.response.text[:120]}"
    except Exception as exc:
        return [], str(exc)[:200]


# ---------------------------------------------------------------------------
# Query builder
# ---------------------------------------------------------------------------

def _build_query(
    company_name: str,
    city: str = "",
    query_template: str = "",
    country_hint: str = "",
) -> str:
    """
    Build a single Serper search query for one company.

    Default template: "<company_name>" official website
    With city:        "<company_name>" "<city>" official website
    With country_hint: appends e.g. "Italy"
    """
    name = (company_name or "").strip()
    if not name:
        return ""

    if query_template:
        query = query_template.format(
            company_name=name,
            city=city or "",
            country_hint=country_hint or "",
        )
    else:
        if city and city.strip():
            query = f'"{name}" "{city.strip()}" official website'
        else:
            query = f'"{name}" official website'

    if country_hint and country_hint.strip():
        hint = country_hint.strip()
        if hint.lower() not in query.lower():
            query = f"{query} {hint}"

    return query


# ---------------------------------------------------------------------------
# Core: process one row
# ---------------------------------------------------------------------------

def _process_row(
    company_name: str,
    city: str,
    existing_website: str,
    api_key: str,
    query_template: str = "",
    country_hint: str = "",
    dry_run: bool = False,
) -> dict:
    """
    Process a single company row.

    Returns a dict with all simple_serper_* output columns.
    """
    out = {
        "simple_serper_query":           "",
        "simple_serper_result_rank":     "",
        "simple_serper_title":           "",
        "simple_serper_url":             "",
        "simple_serper_domain":          "",
        "simple_serper_status":          "",
        "simple_serper_reason":          "",
        "simple_serper_existing_domain": _extract_root_domain(existing_website or ""),
        "simple_serper_domain_changed":  "",
    }

    name = (company_name or "").strip()
    if not name:
        out["simple_serper_status"] = "NO_COMPANY_NAME"
        out["simple_serper_reason"] = "Company name column is blank."
        return out

    query = _build_query(name, city, query_template, country_hint)
    out["simple_serper_query"] = query

    if dry_run:
        out["simple_serper_status"] = "DRY_RUN"
        out["simple_serper_reason"] = f"Dry run — would search: {query}"
        return out

    organic, err = _serper_search(query, api_key)
    if err:
        out["simple_serper_status"] = "SERPER_ERROR"
        out["simple_serper_reason"] = err
        return out

    # Walk the top-5 organic results for the first usable domain
    skipped: list[str] = []
    for rank, item in enumerate(organic[:5], start=1):
        url    = item.get("link", "")
        title  = item.get("title", "")
        domain = _extract_root_domain(url)
        if _is_bad_domain(domain):
            skipped.append(f"#{rank}:{domain}")
            continue
        # First usable result
        out["simple_serper_result_rank"] = rank
        out["simple_serper_title"]       = title
        out["simple_serper_url"]         = url
        out["simple_serper_domain"]      = domain
        out["simple_serper_status"]      = (
            "FOUND_TOP_RESULT" if rank == 1 else "FOUND_AFTER_SKIPPING_BAD_RESULT"
        )
        out["simple_serper_reason"] = (
            f"Accepted rank {rank}."
            + (f" Skipped: {', '.join(skipped)}." if skipped else "")
        )
        # Compare with existing domain
        existing = out["simple_serper_existing_domain"]
        if existing:
            out["simple_serper_domain_changed"] = (
                "NO" if domain == existing else "YES"
            )
        return out

    out["simple_serper_status"] = "NO_USABLE_RESULT"
    out["simple_serper_reason"] = (
        f"All top-{min(5, len(organic))} results were bad domains."
        + (f" Skipped: {', '.join(skipped)}." if skipped else "")
        if organic else "Serper returned no organic results."
    )
    return out


# ---------------------------------------------------------------------------
# Excel I/O
# ---------------------------------------------------------------------------

def _read_excel(path: Path) -> "pd.DataFrame | None":
    try:
        import pandas as pd
        if path.suffix.lower() == ".csv":
            return pd.read_csv(path, dtype=str).fillna("")
        return pd.read_excel(path, dtype=str).fillna("")
    except Exception as exc:
        print(f"[ERROR] Could not read {path.name}: {exc}", file=sys.stderr)
        return None


_OUTPUT_COLS = [
    "simple_serper_query",
    "simple_serper_result_rank",
    "simple_serper_title",
    "simple_serper_url",
    "simple_serper_domain",
    "simple_serper_status",
    "simple_serper_reason",
    "simple_serper_existing_domain",
    "simple_serper_domain_changed",
]

_STATUS_COLORS = {
    "FOUND_TOP_RESULT":               "C6EFCE",  # green
    "FOUND_AFTER_SKIPPING_BAD_RESULT": "FFEB9C", # yellow
    "NO_USABLE_RESULT":               "FFC7CE",  # red
    "NO_COMPANY_NAME":                "D9D9D9",  # grey
    "SERPER_ERROR":                   "F4CCCC",  # light red
    "DRY_RUN":                        "EAD1DC",  # light purple
}


def _write_excel(df: "pd.DataFrame", out_path: Path) -> None:
    """Write df to a styled Excel file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    import pandas as pd

    out_path.parent.mkdir(parents=True, exist_ok=True)

    # Write with openpyxl for styling
    with pd.ExcelWriter(str(out_path), engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Benchmark")
        ws = writer.sheets["Benchmark"]

        # Bold header
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Colour status column rows
        status_col_idx = None
        for idx, col in enumerate(df.columns, start=1):
            if col == "simple_serper_status":
                status_col_idx = idx
                break

        if status_col_idx:
            for row_idx, status in enumerate(df["simple_serper_status"], start=2):
                color = _STATUS_COLORS.get(str(status), "FFFFFF")
                fill  = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
                ws.cell(row=row_idx, column=status_col_idx).fill = fill

        # Auto-width for output columns only (approximate)
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value or "")) for cell in col), default=0
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)


def _no_overwrite_path(p: Path) -> Path:
    """Return p if it doesn't exist; otherwise add _2, _3, … until free."""
    if not p.exists():
        return p
    stem, suffix, parent = p.stem, p.suffix, p.parent
    n = 2
    while True:
        c = parent / f"{stem}_{n}{suffix}"
        if not c.exists():
            return c
        n += 1


# ---------------------------------------------------------------------------
# Single-file processing
# ---------------------------------------------------------------------------

def _process_file(
    input_path: Path,
    output_dir: Path,
    api_key: str,
    limit: int = 0,
    max_serper_calls: int = 100,
    query_template: str = "",
    country_hint: str = "",
    dry_run: bool = False,
) -> dict:
    """
    Process one Excel file.  Returns a summary dict.
    """
    import pandas as pd

    ts      = datetime.now().strftime("%Y%m%d_%H%M")
    stem    = input_path.stem
    out_name = f"{stem}_simple_serper_{ts}.xlsx"
    out_path = _no_overwrite_path(output_dir / out_name)

    print(f"\n[benchmark] Input:      {input_path.name}", flush=True)
    print(f"[benchmark] Output:     {out_path}", flush=True)

    df = _read_excel(input_path)
    if df is None:
        return {"status": "read_error", "input": str(input_path)}

    cols = list(df.columns)
    company_col  = _find_col(cols, _COMPANY_COL_CANDIDATES)
    city_col     = _find_col(cols, _CITY_COL_CANDIDATES)
    province_col = _find_col(cols, _PROVINCE_COL_CANDIDATES)
    website_col  = _find_col(cols, _WEBSITE_COL_CANDIDATES)

    print(f"[benchmark] Columns:    company={company_col!r} city={city_col!r} "
          f"province={province_col!r} website={website_col!r}", flush=True)

    if not company_col:
        print(
            f"[benchmark] ERROR: no company name column found. "
            f"Available: {cols[:12]}",
            file=sys.stderr,
        )
        return {"status": "no_company_col", "input": str(input_path)}

    total = len(df)
    batch = df.head(limit).copy() if limit > 0 else df.copy()
    print(f"[benchmark] Rows:       {len(batch)} / {total}", flush=True)
    if not dry_run:
        effective = min(len(batch), max_serper_calls)
        print(f"[benchmark] Max calls:  {effective}", flush=True)

    results: list[dict] = []
    calls_made = 0

    for i, row in enumerate(batch.itertuples(index=False), start=1):
        name     = str(getattr(row, _safe_attr(company_col), "")).strip()
        city     = str(getattr(row, _safe_attr(city_col), "")).strip()     if city_col     else ""
        existing = str(getattr(row, _safe_attr(website_col), "")).strip()  if website_col  else ""

        # Enforce max Serper calls
        this_dry = dry_run or (calls_made >= max_serper_calls)

        res = _process_row(
            company_name=name,
            city=city,
            existing_website=existing,
            api_key=api_key,
            query_template=query_template,
            country_hint=country_hint,
            dry_run=this_dry,
        )
        results.append(res)

        if not dry_run and res["simple_serper_status"] not in ("NO_COMPANY_NAME", "DRY_RUN"):
            calls_made += 1

        # Progress
        if i % 25 == 0 or i == len(batch):
            found = sum(
                1 for r in results
                if r["simple_serper_status"] in (
                    "FOUND_TOP_RESULT", "FOUND_AFTER_SKIPPING_BAD_RESULT"
                )
            )
            print(f"[benchmark] Progress:   {i}/{len(batch)} rows | "
                  f"{found} found | {calls_made} API calls", flush=True)

        # Small sleep to avoid rate-limits
        if not this_dry and res["simple_serper_status"] not in ("NO_COMPANY_NAME",):
            time.sleep(0.25)

    # Build output dataframe
    results_df = pd.DataFrame(results, columns=_OUTPUT_COLS)
    out_df     = pd.concat(
        [batch.reset_index(drop=True), results_df.reset_index(drop=True)],
        axis=1,
    )

    _write_excel(out_df, out_path)
    print(f"[benchmark] Saved:      {out_path}", flush=True)

    # Summary
    status_counts = results_df["simple_serper_status"].value_counts().to_dict()
    return {
        "status":        "ok",
        "input":         str(input_path),
        "output":        str(out_path),
        "rows_total":    total,
        "rows_processed": len(batch),
        "serper_calls":  calls_made,
        "status_counts": status_counts,
    }


def _safe_attr(col: str) -> str:
    """Convert a column name to a valid Python attribute name for itertuples."""
    return re.sub(r"[^a-zA-Z0-9_]", "_", col)


# ---------------------------------------------------------------------------
# Self-test (no API calls required)
# ---------------------------------------------------------------------------

def _self_test() -> None:
    print("=" * 60, flush=True)
    print("simple_serper_domain_benchmark — Self-Test", flush=True)
    print("=" * 60, flush=True)

    failures: list[str] = []

    def check(label: str, actual, expected) -> None:
        if actual != expected:
            failures.append(f"  FAIL [{label}]: got {actual!r}, expected {expected!r}")
        else:
            print(f"  PASS [{label}]", flush=True)

    # --- Domain extraction ---
    print("\n── _extract_root_domain ──", flush=True)
    check("https with www",
          _extract_root_domain("https://www.example.it/about"), "example.it")
    check("http no www",
          _extract_root_domain("http://example.com/page?q=1"), "example.com")
    check("country subdomain stripped",
          _extract_root_domain("https://it.example.com/"), "example.com")
    check("short brand .it",
          _extract_root_domain("https://q8.it/"), "q8.it")
    check("no protocol",
          _extract_root_domain("example.com"), "example.com")
    check("empty string",
          _extract_root_domain(""), "")

    # --- Bad domain detection ---
    print("\n── _is_bad_domain ──", flush=True)
    check("linkedin.com rejected",
          _is_bad_domain("linkedin.com"), True)
    check("it.linkedin.com rejected",
          _is_bad_domain("it.linkedin.com"), True)
    check("wikipedia.org rejected",
          _is_bad_domain("wikipedia.org"), True)
    check("it.wikipedia.org rejected",
          _is_bad_domain("it.wikipedia.org"), True)
    check("fatturatoitalia.it rejected",
          _is_bad_domain("fatturatoitalia.it"), True)
    check("visura.pro rejected",
          _is_bad_domain("visura.pro"), True)
    check("indeed.com rejected",
          _is_bad_domain("indeed.com"), True)
    check("it.indeed.com rejected",
          _is_bad_domain("it.indeed.com"), True)
    check("de.glassdoor.com rejected",
          _is_bad_domain("de.glassdoor.com"), True)
    check("ibm.com accepted",
          _is_bad_domain("ibm.com"), False)
    check("zf.com accepted",
          _is_bad_domain("zf.com"), False)
    check("q8.it accepted",
          _is_bad_domain("q8.it"), False)
    check("solutions30.com accepted",
          _is_bad_domain("solutions30.com"), False)

    # --- Query builder ---
    print("\n── _build_query ──", flush=True)
    q1 = _build_query("Pompe Garbarino")
    check("default query",
          q1, '"Pompe Garbarino" official website')
    q2 = _build_query("Pompe Garbarino", city="Milano")
    check("with city",
          q2, '"Pompe Garbarino" "Milano" official website')
    q3 = _build_query("Pompe Garbarino", country_hint="Italy")
    check("with country_hint",
          q3, '"Pompe Garbarino" official website Italy')
    q4 = _build_query("")
    check("empty name returns empty string",
          q4, "")

    # --- _process_row dry-run ---
    print("\n── _process_row dry_run ──", flush=True)
    r = _process_row("IBM Italia SPA", "Milano", "", "", dry_run=True)
    check("dry_run status",
          r["simple_serper_status"], "DRY_RUN")
    check("dry_run query contains IBM",
          "IBM" in r["simple_serper_query"], True)

    r2 = _process_row("", "", "", "", dry_run=True)
    check("blank name → NO_COMPANY_NAME",
          r2["simple_serper_status"], "NO_COMPANY_NAME")

    print(flush=True)
    if failures:
        print("FAILED:", flush=True)
        for f in failures:
            print(f, flush=True)
        sys.exit(1)
    else:
        print(f"All tests passed ({14 + 4 + 2} checks).", flush=True)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Simple Serper Domain Benchmark — one Serper call per company, "
            "first usable organic result wins."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--input", default=None,
        help="Path to one .xlsx file or a folder of .xlsx files.",
    )
    parser.add_argument(
        "--output-dir", default=None,
        help="Directory where output files are written. "
             "Defaults to a 'benchmark_output' sub-folder next to the input.",
    )
    parser.add_argument(
        "--serper-key", default=None,
        help="Serper API key. Also read from SERPER_API_KEY or SERPER_KEY env vars.",
    )
    parser.add_argument(
        "--limit", type=int, default=0,
        help="Process only the first N rows (0 = all).",
    )
    parser.add_argument(
        "--max-serper-calls", type=int, default=100,
        help="Hard cap on Serper API calls per input file (default 100).",
    )
    parser.add_argument(
        "--query-template", default="",
        help=(
            'Custom query template. Use {company_name}, {city}, {country_hint}. '
            'Default: "\"<company_name>\" official website"'
        ),
    )
    parser.add_argument(
        "--country-hint", default="",
        help="Optional country string appended to every query, e.g. 'Italy'.",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Read rows and show planned queries without calling Serper.",
    )
    parser.add_argument(
        "--self-test", action="store_true",
        help="Run built-in self-tests and exit (no API calls required).",
    )
    return parser.parse_args()


def main() -> None:
    args = _parse_args()

    if args.self_test:
        _self_test()
        return

    if not args.input:
        print("ERROR: --input is required.", file=sys.stderr)
        sys.exit(1)

    # Resolve API key
    api_key = (
        args.serper_key
        or os.environ.get("SERPER_API_KEY", "")
        or os.environ.get("SERPER_KEY", "")
        or ""
    )
    if not api_key and not args.dry_run:
        print(
            "ERROR: Serper API key required. Pass --serper-key, "
            "or set SERPER_API_KEY / SERPER_KEY env var.",
            file=sys.stderr,
        )
        sys.exit(1)

    input_path = Path(args.input).resolve()
    if not input_path.exists():
        print(f"ERROR: input path not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    # Resolve input files
    if input_path.is_dir():
        files = sorted(input_path.glob("*.xlsx"))
        if not files:
            print(f"ERROR: no .xlsx files found in {input_path}", file=sys.stderr)
            sys.exit(1)
        print(
            f"[benchmark] Folder mode: {len(files)} file(s) in {input_path}",
            flush=True,
        )
    else:
        files = [input_path]

    # Resolve output dir
    if args.output_dir:
        output_dir = Path(args.output_dir).resolve()
    else:
        # Default: sibling folder "benchmark_output" or same folder for single file
        if input_path.is_dir():
            output_dir = input_path.parent / "benchmark_output"
        else:
            output_dir = input_path.parent / "benchmark_output"

    print(f"[benchmark] Output dir: {output_dir}", flush=True)
    output_dir.mkdir(parents=True, exist_ok=True)

    all_summaries: list[dict] = []
    for f in files:
        summary = _process_file(
            input_path       = f,
            output_dir       = output_dir,
            api_key          = api_key,
            limit            = args.limit,
            max_serper_calls = args.max_serper_calls,
            query_template   = args.query_template,
            country_hint     = args.country_hint,
            dry_run          = args.dry_run,
        )
        all_summaries.append(summary)

    # Final aggregate summary
    print("\n[benchmark] ═══ Run complete ═══", flush=True)
    total_rows   = sum(s.get("rows_processed", 0) for s in all_summaries)
    total_calls  = sum(s.get("serper_calls",   0) for s in all_summaries)
    total_found  = sum(
        s.get("status_counts", {}).get("FOUND_TOP_RESULT", 0)
        + s.get("status_counts", {}).get("FOUND_AFTER_SKIPPING_BAD_RESULT", 0)
        for s in all_summaries
    )
    print(f"  Files processed:  {len(all_summaries)}", flush=True)
    print(f"  Rows processed:   {total_rows}", flush=True)
    print(f"  Serper API calls: {total_calls}", flush=True)
    if total_rows > 0:
        print(
            f"  Domains found:    {total_found} / {total_rows} "
            f"({total_found / total_rows:.0%})",
            flush=True,
        )
    for s in all_summaries:
        if s.get("status_counts"):
            print(
                f"  {Path(s['input']).name}: "
                + " | ".join(f"{k}={v}" for k, v in s["status_counts"].items()),
                flush=True,
            )


if __name__ == "__main__":
    main()
