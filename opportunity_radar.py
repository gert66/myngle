"""
mYngle · Opportunity Radar
==========================
Researches buying-window signals and contact routes for company lists.
Accepts Lead Prioritizer exports (with Opportunity Input sheet) or simple
company lists.  Uses Serper Google Search + Claude Haiku for signal extraction.

Entry point:  streamlit run opportunity_radar.py
"""

import base64
import hashlib
import io
import json
import pathlib
import re
import time
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
import requests
import sys

st = None          # lazy — populated by _get_st() in UI mode only

def _get_st():
    global st
    if st is None:
        import streamlit as _st
        st = _st
    return st


_CLI_FLAGS = {
    "--input", "--dry-run-paths", "--project-root",
    "--serper-key", "--anthropic-key", "--max-rows", "--debug", "--force-fresh",
}


def cli_args_present() -> bool:
    for arg in sys.argv[1:]:
        if arg in _CLI_FLAGS:
            return True
        if any(arg.startswith(f + "=") for f in _CLI_FLAGS):
            return True
    return False


def running_under_streamlit() -> bool:
    if "streamlit" not in sys.modules:
        return False
    try:
        from streamlit.runtime.scriptrunner import get_script_run_ctx
        return get_script_run_ctx() is not None
    except Exception:
        return False


def is_cli_mode() -> bool:
    return cli_args_present() or not running_under_streamlit()


try:
    import anthropic as _anthropic_mod
    _ANTHROPIC_AVAILABLE = True
except ImportError:
    _ANTHROPIC_AVAILABLE = False


def _detect_col(df: pd.DataFrame, candidates: list) -> str | None:
    lower_map = {c.lower(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in lower_map:
            return lower_map[cand.lower()]
    return None


def _normalize_domain(raw: str) -> str:
    """Strip protocol, www, trailing path from a URL to get a bare domain."""
    if not raw:
        return raw
    raw = raw.strip()
    raw = re.sub(r"^https?://", "", raw, flags=re.IGNORECASE)
    raw = raw.split("/")[0].split("?")[0]
    if raw.lower().startswith("www."):
        raw = raw[4:]
    return raw.strip()


_INTERNAL_NAMES   = {"myngle"}
_INTERNAL_DOMAINS = {"myngle.com"}


def _is_internal(name: str, domain: str) -> bool:
    return (
        name.lower().strip() in _INTERNAL_NAMES
        or any(d in domain.lower() for d in _INTERNAL_DOMAINS)
    )


def _count_companies(df: pd.DataFrame, name_col: str | None) -> int:
    if name_col and name_col in df.columns:
        return int(
            df[name_col]
            .dropna()
            .astype(str)
            .str.strip()
            .replace("", pd.NA)
            .dropna()
            .nunique()
        )
    return len(df)


# =============================================================================
# FILE LOADING WITH SHEET PRIORITY
# =============================================================================

def _load_df_from_upload(uploaded_file) -> tuple:
    """
    Load the best DataFrame from an uploaded file.

    Sheet priority for Excel files:
      1. 'Opportunity Input' — preferred (Lead Prioritizer export)
      2. First sheet (usually 'Lead Scores')
      3. 'Enriched' — fallback if first sheet has no usable columns
      ('Company Profiles' is never read — it is formatted for humans)

    Returns (df, sheet_used) where sheet_used is a string label.
    CSV files always return (df, "csv").
    """
    fname = uploaded_file.name
    if fname.lower().endswith(".csv"):
        return pd.read_csv(uploaded_file), "csv"

    xf = pd.ExcelFile(uploaded_file)
    sheet_names = xf.sheet_names

    # Priority 1 — dedicated Opportunity Input sheet
    if "Opportunity Input" in sheet_names:
        return xf.parse("Opportunity Input"), "Opportunity Input"

    # Priority 2 — first sheet (skip Company Profiles if it is somehow first)
    first = next(
        (s for s in sheet_names if s != "Company Profiles"),
        sheet_names[0] if sheet_names else None,
    )
    if first:
        df_first = xf.parse(first)
        # Check if it has at least a name or domain column
        lower_cols = {c.lower() for c in df_first.columns}
        has_identity = any(
            c in lower_cols
            for c in ("company_name", "company name", "company", "name",
                      "domain", "company domain", "company website", "website")
        )
        if has_identity:
            return df_first, first

    # Priority 3 — 'Enriched' hidden sheet
    if "Enriched" in sheet_names:
        return xf.parse("Enriched"), "Enriched"

    # Last resort — first sheet regardless
    return xf.parse(sheet_names[0]), sheet_names[0]


def _detect_input_type(df: pd.DataFrame) -> str:
    """
    Classify the uploaded file as 'enriched_export' or 'simple_company_list'.

    An enriched export contains at least one of the enrichment signal columns.
    """
    col_set = {c.lower() for c in df.columns}
    for sig_col in _ENRICHED_SIGNAL_COLS:
        if sig_col.lower() in col_set:
            return "enriched_export"
    return "simple_company_list"


# =============================================================================
# CACHE
# =============================================================================

def _cache_key(name: str, domain: str, input_type: str = "") -> str:
    # CACHE_VERSION + input_type in key → old prompts never bleed into new runs
    raw = f"{CACHE_VERSION}|{name.lower().strip()}|{domain.lower().strip()}|{input_type}"
    return hashlib.md5(raw.encode()).hexdigest()


def _cache_load(name: str, domain: str, input_type: str = "") -> dict | None:
    RADAR_CACHE_DIR.mkdir(exist_ok=True)
    p = RADAR_CACHE_DIR / f"{_cache_key(name, domain, input_type)}.json"
    if p.exists():
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            return None
    return None


def _cache_save(name: str, domain: str, input_type: str, data: dict) -> None:
    RADAR_CACHE_DIR.mkdir(exist_ok=True)
    p = RADAR_CACHE_DIR / f"{_cache_key(name, domain, input_type)}.json"
    try:
        p.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    except Exception:
        pass


# =============================================================================
# SERPER
# =============================================================================

def _serper_search(query: str, api_key: str, n: int = 5) -> list:
    try:
        resp = requests.post(
            SERPER_URL,
            headers={"X-API-KEY": api_key, "Content-Type": "application/json"},
            json={"q": query, "num": n},
            timeout=15,
        )
        resp.raise_for_status()
        data = resp.json()
        results = []
        for item in data.get("organic", [])[:n]:
            results.append({
                "title":   item.get("title", ""),
                "url":     item.get("link", ""),
                "snippet": item.get("snippet", ""),
                "date":    item.get("date", ""),
            })
        return results
    except Exception:
        return []


def _run_searches(name: str, domain: str, api_key: str) -> dict:
    """Run all query groups; return {group_label: [results]}."""
    grouped: dict = {}
    for group_label, template in QUERY_GROUPS:
        query = template.format(name=name)
        results = _serper_search(query, api_key, n=5)
        grouped[group_label] = results
        time.sleep(0.3)  # stay within Serper rate limits
    return grouped


def _format_results_for_prompt(grouped: dict) -> str:
    parts = []
    for group_label, results in grouped.items():
        parts.append(f"=== {group_label.upper()} ===")
        if not results:
            parts.append("(no results)\n")
            continue
        for i, r in enumerate(results, 1):
            line = f"[{i}] {r['title']}"
            if r.get("date"):
                line += f"  ({r['date']})"
            line += f"\n    {r['url']}"
            if r.get("snippet"):
                line += f"\n    {r['snippet']}"
            parts.append(line)
        parts.append("")
    return "\n".join(parts)


def _collect_raw_sources(company_name: str, grouped: dict, input_type: str = "") -> list:
    rows = []
    for group_label, results in grouped.items():
        for r in results:
            rows.append({
                "company_name": company_name,
                "input_type":   input_type,
                "query_group":  group_label,
                "title":        r.get("title", ""),
                "url":          r.get("url", ""),
                "snippet":      r.get("snippet", ""),
                "date":         r.get("date", ""),
                "source_type":  "Organic search",
            })
    return rows


# =============================================================================
# CLAUDE EXTRACTION
# =============================================================================

_PROMPT_TEMPLATE = """\
You are a B2B sales intelligence analyst working for mYngle, a company that sells \
online language training and business communication support to international companies.

mYngle's target buyers are companies that have:
- International or multilingual teams
- Foreign HQ or group structures
- Client-facing international roles (sales, customer success, account management)
- Fast hiring or onboarding of international employees
- Post-merger or cross-border communication challenges
- Expanding into new countries or markets

mYngle does NOT sell: generic L&D platforms, e-learning tools, or broad HR software.
mYngle sells: language training, Business English, business communication coaching.

TODAY'S DATE: {today}

COMPANY PROFILE:
- Name: {name}
- Domain: {domain}
- Country: {country}
- Commercial Fit Score: {fit_score}
- Commercial Tier: {tier}
- ICP Evidence: {icp_evidence}

WEB SEARCH RESULTS:
{search_text}

ANALYSIS INSTRUCTIONS:

1. TRIGGER IDENTIFICATION
   Look for signals that a company is likely to need language or communication training:
   - International hiring, new foreign offices, global expansion
   - Growth of sales, customer success, or account management teams
   - M&A, integration activity, foreign group/HQ structure
   - Onboarding pressure from fast hiring
   - L&D or HR team growth suggesting new training budget
   - Annual report published (signals planning cycle timing)
   trigger_score: 0=no relevant signal, 1=weak/indirect, 2=clear signal, 3=strong + recent

   trigger_type MUST be exactly one of: {trigger_types}
   Choose the single strongest primary trigger. Do not combine or invent labels.
   If multiple triggers are present, pick the most commercially relevant one for mYngle
   and add context to trigger_evidence.
   Examples of correct mapping:
   - "International expansion + multinational workforce" → "Multilingual workforce growth"
   - "L&D / Onboarding Infrastructure + Annual Planning Cycle" → "HR / L&D hiring"
   - "Annual planning signal" → "Annual planning / budget window"
   - No relevant trigger found → "No clear trigger"

2. BUYING WINDOW
   - IMPORTANT: likely_buying_window must be FUTURE relative to today ({today})
   - If best evidence points to a past window, project forward one year and set
     buying_window_confidence to "Low"
   - Do NOT invent a specific quarter unless the fiscal year pattern and a planning
     logic chain clearly support it. Use hedged language:
       * "Possible Q3/Q4 planning conversation, assuming calendar-year budgeting"
       * "Possible annual planning window, confidence low"
       * "Possible fiscal-year planning conversation, timing not confirmed"
       * "Current trigger found, timing appears commercially relevant"
       * "No clear buying window found"
   - Only say "Q3 2026" if you can explain WHY Q3 — e.g. fiscal year ends Dec, so
     budget planning typically happens Aug/Sep. Otherwise use hedged wording.
   - If annual report found, it confirms a planning cycle exists but does NOT by itself
     confirm a specific quarter unless the fiscal year end is stated.
   buying_window_score: 0=no basis, 1=possible, 2=likely, 3=imminent

3. BUYER ROUTE — choose based on the dominant trigger signal:
   - "L&D / Talent Development": training, onboarding, people development, academy signals
   - "HR / People": general HR, people, workforce signals without strong L&D angle
   - "International HR": foreign HQ, global teams, multilingual workforce, cross-border structure
   - "Sales Enablement": sales expansion, account management, negotiation, international sales
   - "Customer Success": customer success growth, international client support
   - "People Operations": fast onboarding, operational employee growth, multi-site rollout
   - "Operations": only when operational coordination is the clearest angle
   - "Procurement": last resort only — never preferred first route
   - "Unknown": truly no signal to guide route choice
   preferred_buyer_route must be one of: {routes}

4. SUGGESTED TITLE SEARCHES (for LinkedIn Sales Navigator)
   Match to preferred and backup buyer routes. Output as a single string using OR syntax.
   Use ONLY the OR-syntax format. Do NOT use comma-separated lists.
   - L&D route: "Learning Development" OR "Talent Development" OR "L&D"
   - HR/People route: "HR Director" OR "People Director" OR "Head of People"
   - International HR route: "International HR" OR "Global HR" OR "People Operations"
   - Sales Enablement route: "Sales Enablement" OR "Revenue Enablement"
   - Customer Success route: "Customer Success Director" OR "VP Customer Success"
   - People Operations route: "Onboarding" OR "People Operations" OR "HR Operations"
   Combine preferred and backup routes: e.g.
   "Learning Development" OR "L&D" OR "HR Director" OR "Head of People"

5. WHY NOW — must connect evidence to mYngle's value proposition
   Focus on: language training, Business English, business communication, client-facing
   communication, international team communication, onboarding of international employees,
   multilingual workforce support, intercultural communication, foreign HQ communication,
   cross-border communication, business communication training.

   BANNED PHRASES — never write any of the following:
   "learning platform", "talent development tools", "upskilling solutions",
   "digital learning", "workforce solution", "talent tools", "e-learning",
   "learning tools", "broad L&D", "generic L&D", "training platform",
   "talent management", "HR platform", "learning management", "LMS",
   "skill development platform", "workforce training platform".

   Use instead: "language training", "Business English", "business communication training",
   "client-facing communication", "international team communication",
   "multilingual workforce support", "onboarding communication support",
   "cross-border communication", "intercultural communication".

   Example: "Capgemini has large international teams and active internal training
   infrastructure. Companies at that scale often review Business English and
   client-facing communication support during annual planning cycles."

6. CALLER OPENER — short, natural, specific
   - Mention one concrete signal from the search results
   - Connect it to language training or business communication
   - End with a soft discovery question
   - BANNED: "learning platform", "digital learning tools", "talent development tools",
     "upskilling", "workforce solutions", "HR software"
   - The question should feel like a natural cold-call opening, not a product pitch
   - Examples by trigger:
     * International hiring: "I noticed you're expanding internationally and hiring across
       new markets. Companies often use that moment to review language and communication
       support for new teams. Is business communication training already part of your
       L&D planning?"
     * Customer-facing growth: "I noticed growth in your international customer-facing
       teams. That often creates pressure around Business English and client communication.
       Is this something your team is already looking at?"
     * M&A/integration: "I noticed recent integration activity at {name}. Those transitions
       often bring communication and language alignment challenges across teams and countries.
       Is language training part of the integration plan?"
     * Annual planning: "I noticed your annual planning cycle may be coming up. Many
       companies review language and business communication training before finalising their
       L&D budget. Is this already on your agenda?"
     * Large international workforce: "I noticed {name} has extensive international
       operations. Companies at that scale often have ongoing needs around Business English
       and cross-border communication support. Is language training already part of your
       current L&D planning?"

7. CONFIDENCE AND EVIDENCE QUALITY
   - evidence_quality: Strong=multiple recent specific sources, Medium=1-2 relevant sources,
     Weak=snippets only or indirect signals, Insufficient=no relevant evidence
   - confidence_level: High only if Strong evidence + clear trigger + clear buyer route,
     Medium if some evidence present, Low/Unknown if mostly snippets or indirect
   - manual_review_needed: true if evidence is Weak/Insufficient OR buyer route unclear
     OR company seems relevant but signals are ambiguous

SCORING FIELD INSTRUCTIONS:
- trigger_score: 0-3 as above
- buying_window_score: 0-3 as above
- hiring_signal_score: 0-3 (overall hiring volume signal)
- international_hiring_signal: 0-3 (international/multilingual hiring specifically)
- lnd_hr_hiring_signal: 0-3 (L&D or HR hiring that signals training budget)
- sales_cs_hiring_signal: 0-3 (sales or customer success growth)
- onboarding_pressure_signal: 0-3 (fast hiring, headcount growth, new site openings)

Return ONLY the JSON object below — no markdown fences, no explanation text.

{schema}"""


def _extract_json_from_text(text: str) -> dict:
    text = re.sub(r"```(?:json)?", "", text).strip().rstrip("`").strip()
    start = text.find("{")
    end   = text.rfind("}")
    if start != -1 and end != -1 and end > start:
        try:
            return json.loads(text[start : end + 1])
        except json.JSONDecodeError:
            pass
    return {}


def _call_claude(
    name: str,
    domain: str,
    country: str,
    fit_score: str,
    tier: str,
    icp_evidence: str,
    grouped_results: dict,
    client,
) -> dict:
    search_text = _format_results_for_prompt(grouped_results)
    today_str = datetime.now().strftime("%Y-%m-%d")
    prompt = _PROMPT_TEMPLATE.format(
        today=today_str,
        name=name,
        domain=domain or "(unknown)",
        country=country or "(unknown)",
        fit_score=fit_score or "(not available)",
        tier=tier or "(not available)",
        icp_evidence=(icp_evidence or "(not available)")[:600],
        search_text=search_text,
        routes=", ".join(ALLOWED_ROUTES),
        trigger_types=", ".join(ALLOWED_TRIGGER_TYPES),
        schema=_CLAUDE_SCHEMA,
    )

    try:
        msg = client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=1500,
            messages=[{"role": "user", "content": prompt}],
        )
        raw_text = msg.content[0].text if msg.content else ""
        result   = _extract_json_from_text(raw_text)
        if not result:
            result = dict(_EMPTY_CLAUDE_RESULT)
            result["manual_review_reason"] = "Claude returned unparseable response"
        return result
    except Exception as exc:
        result = dict(_EMPTY_CLAUDE_RESULT)
        result["manual_review_reason"] = f"Claude error: {exc}"
        return result


# =============================================================================
# SCORING
# =============================================================================

def _fit_bucket(fit_score_raw, tier_raw) -> int:
    """Convert commercial fit score / tier to 0-3 bucket."""
    tier_str = str(tier_raw or "").lower()
    if "hot" in tier_str:
        return 3
    if "warm" in tier_str:
        return 2
    if "cool" in tier_str:
        return 1
    if "pass" in tier_str:
        return 0

    try:
        score = float(str(fit_score_raw).strip())
        if score >= 8.0:
            return 3
        if score >= 6.0:
            return 2
        if score >= 4.0:
            return 1
        return 0
    except (ValueError, TypeError):
        pass

    return 1  # default: something present but no score


def _contact_route_score(preferred_route: str) -> int:
    route = str(preferred_route or "").lower()
    if any(k in route for k in ("l&d", "talent development", "sales enablement", "international hr")):
        return 3
    if any(k in route for k in ("hr / people", "customer success", "people operations")):
        return 2
    if any(k in route for k in ("hr", "people", "operations")):
        return 1
    return 0  # Unknown or Procurement


def _opportunity_score(
    fit: int, trigger: int, window: int, route: int, input_type: str
) -> float:
    if input_type == "enriched_export":
        # 35% fit · 30% trigger · 25% window · 10% route
        return round(
            (fit / 3 * 10 * 0.35)
            + (trigger / 3 * 10 * 0.30)
            + (window / 3 * 10 * 0.25)
            + (route / 3 * 10 * 0.10),
            1,
        )
    else:
        # timing-only mode: 45% trigger · 35% window · 20% route
        return round(
            (trigger / 3 * 10 * 0.45)
            + (window / 3 * 10 * 0.35)
            + (route / 3 * 10 * 0.20),
            1,
        )


def _call_recommendation(
    fit: int,
    trigger: int,
    window: int,
    opp: float,
    manual: bool,
    input_type: str,
    evidence_quality: str = "",
) -> str:
    eq = evidence_quality.lower()
    eq_strong = eq in ("strong", "medium")

    if input_type == "enriched_export":
        # ICP fit is known — use full decision matrix
        if fit == 0:
            return "Low priority"
        # Cool fit (1) with no trigger → just monitor; don't create false urgency
        if fit == 1 and trigger == 0 and window == 0:
            return "Monitor"
        # Call now only when fit is strong, trigger is strong, AND evidence is credible
        if trigger >= 3 and fit >= 2 and eq_strong:
            return "Call now"
        if fit >= 2 and trigger >= 2 and eq_strong:
            return "Call this month"
        if fit >= 2 and window >= 2:
            return "Call this month"
        if fit >= 2 and trigger >= 1:
            return "Call before budget cycle" if window >= 1 else "Monitor"
        # Hot/Warm fit but no trigger and manual flag → human should research next
        if fit >= 2 and manual:
            return "Manual research needed"
        if trigger == 0 and window == 0:
            return "Monitor"
        if manual:
            return "Manual research needed"
        return "Monitor"
    else:
        # ICP fit unknown — conservative; weak evidence always routes to manual review
        if trigger == 0 and window == 0:
            return "Monitor"
        # Only the very strongest signal overrides a manual/weak-evidence flag
        if trigger >= 3 and window >= 2:
            return "Call this month"
        if manual:
            return "Manual research needed"
        if trigger >= 2 and window >= 1:
            return "Call this month"
        if trigger >= 2 or window >= 2:
            return "Call before budget cycle"
        if window >= 1 and trigger >= 1:
            return "Call before budget cycle"
        return "Manual research needed"


_QUARTER_RE = re.compile(
    r"Q([1-4])[- /](\d{4})|(\d{4})[- /]Q([1-4])", re.IGNORECASE
)
_YEAR_ONLY_RE = re.compile(r"\b(20\d{2})\b")

_QUARTER_STARTS = {1: (1, 1), 2: (4, 1), 3: (7, 1), 4: (10, 1)}


def _window_to_date(window_str: str):
    """Return the start date implied by a buying-window string, or None."""
    if not window_str:
        return None
    m = _QUARTER_RE.search(window_str)
    if m:
        q = int(m.group(1) or m.group(4))
        y = int(m.group(2) or m.group(3))
        month, day = _QUARTER_STARTS[q]
        try:
            return datetime(y, month, day)
        except ValueError:
            return None
    m = _YEAR_ONLY_RE.search(window_str)
    if m:
        try:
            return datetime(int(m.group(1)), 1, 1)
        except ValueError:
            return None
    return None


def _adjust_past_buying_window(claude_result: dict) -> dict:
    """
    If likely_buying_window refers to a date that has already passed, lower the
    buying_window_score to 0 and confidence to 'Low' so the scoring formulas
    do not recommend immediate action based on stale timing.
    Annual report evidence is preserved — only the recommended window is adjusted.
    """
    window = claude_result.get("likely_buying_window", "")
    if not window:
        return claude_result
    window_date = _window_to_date(window)
    if window_date is None:
        return claude_result
    today = datetime.now()
    if window_date >= today:
        return claude_result  # window is in the future — nothing to do

    result = dict(claude_result)
    result["buying_window_score"] = 0
    result["buying_window_confidence"] = "Low"
    old_reason = result.get("buying_window_reason", "")
    result["buying_window_reason"] = (
        f"[Window expired — {window} is in the past. "
        f"Next likely planning cycle estimated.] {old_reason}"
    ).strip()
    # Project forward by one year as a placeholder
    try:
        future_year = window_date.year + 1
        future_q = (window_date.month - 1) // 3 + 1
        result["likely_buying_window"] = f"Q{future_q} {future_year} (projected)"
    except Exception:
        result["likely_buying_window"] = "Next planning cycle (estimated)"
    return result


# Banned phrases that must never appear in caller-facing text fields
_BANNED_PHRASES = [
    "learning platform", "talent development tools", "upskilling solution",
    "digital learning tool", "workforce solution", "talent tools", "e-learning tool",
    "learning management", "talent management", "hr platform", "lms",
    "skill development platform", "workforce training platform",
    "broad l&d", "generic l&d", "learning tools", "training platform",
]

# Phrase-replacement map: banned generic phrase (lowercase) → mYngle-specific replacement
_PHRASE_REPLACEMENTS: list[tuple[str, str]] = [
    # Most specific multi-word first to prevent partial replacements
    ("l&d technology solutions",            "language and communication training"),
    ("learning and development technology", "language training"),
    ("talent development tools",            "language training"),
    ("talent development solutions",        "language training"),
    ("talent development platform",         "language training platform"),
    ("learning platforms",                  "language training"),
    ("learning platform",                   "language training"),
    ("upskilling solutions",                "language and communication training"),
    ("upskilling solution",                 "language and communication training"),
    ("digital learning tools",              "language training"),
    ("digital learning tool",               "language training"),
    ("digital learning",                    "language training"),
    ("workforce solutions",                 "language training"),
    ("workforce solution",                  "language training"),
    ("workforce training platform",         "language training"),
    ("talent tools",                        "language training"),
    ("e-learning tools",                    "language training"),
    ("e-learning tool",                     "language training"),
    ("learning management system",          "language training programme"),
    ("broad l&d",                           "language and communication training"),
    ("generic l&d",                         "language training"),
    ("l&d technology",                      "language and communication training"),
    ("training platform",                   "language training programme"),
    ("skill development platform",          "language training"),
    ("hr platform",                         "people development platform"),
    ("talent management",                   "people development"),
    ("evaluating new learning",             "reviewing language and communication training"),
    ("evaluating learning",                 "reviewing language training"),
    ("new learning tools",                  "language training"),
]


def _sanitize_text(text: str) -> str:
    """Replace banned generic phrases with mYngle-specific wording. Case-preserving."""
    if not text:
        return text
    result = text
    lower = result.lower()
    # Work on lowercase index to find replacement positions; rebuild preserving original case
    for banned, replacement in _PHRASE_REPLACEMENTS:
        start = 0
        while True:
            idx = lower.find(banned, start)
            if idx == -1:
                break
            result = result[:idx] + replacement + result[idx + len(banned):]
            lower  = result.lower()
            start  = idx + len(replacement)
    return result


def _sanitize_claude_result(cr: dict) -> dict:
    """Apply text sanitizer to all caller-facing text fields in a Claude result dict."""
    fields = ("why_now", "suggested_opener", "trigger_evidence",
              "buying_window_reason", "manual_review_reason")
    sanitized = dict(cr)
    for f in fields:
        if f in sanitized and isinstance(sanitized[f], str):
            sanitized[f] = _sanitize_text(sanitized[f])
    return sanitized

ALLOWED_TRIGGER_TYPES = [
    "International hiring",
    "Client-facing team expansion",
    "Sales / customer success growth",
    "New market or office expansion",
    "Multilingual workforce growth",
    "M&A / integration",
    "Funding / growth investment",
    "HR / L&D hiring",
    "Onboarding pressure",
    "Foreign HQ / group communication",
    "Employer branding / retention",
    "Annual planning / budget window",
    "No clear trigger",
    "Other",
]

ALLOWED_ROUTES = [
    "L&D / Talent Development",
    "HR / People",
    "International HR",
    "Sales Enablement",
    "Customer Success",
    "People Operations",
    "Operations",
    "Procurement",
    "Unknown",
]

ALLOWED_RECOMMENDATIONS = [
    "Call now",
    "Call this month",
    "Call before budget cycle",
    "Monitor",
    "Manual research needed",
    "Low priority",
    "Internal / exclude",
]

# JSON schema Claude must return
_CLAUDE_SCHEMA = """{
  "trigger_found": true/false,
  "trigger_type": "<one of the allowed trigger types>",
  "trigger_date": "<date or empty string>",
  "trigger_score": <0-3>,
  "trigger_evidence": "<1-3 sentence summary of evidence>",
  "annual_report_found": true/false,
  "annual_report_url": "<url or empty>",
  "annual_report_date": "<date or empty>",
  "fiscal_year_pattern": "<Calendar year | Non-calendar fiscal year | Unknown>",
  "likely_buying_window": "<e.g. Q1 2026 or empty>",
  "buying_window_score": <0-3>,
  "buying_window_confidence": "<High | Medium | Low | Unknown>",
  "buying_window_reason": "<brief explanation>",
  "hiring_signal_score": <0-3>,
  "international_hiring_signal": <0-3>,
  "lnd_hr_hiring_signal": <0-3>,
  "sales_cs_hiring_signal": <0-3>,
  "onboarding_pressure_signal": <0-3>,
  "preferred_buyer_route": "<one of the allowed routes>",
  "backup_buyer_route": "<one of the allowed routes or empty>",
  "suggested_title_searches": "<comma-separated titles to search for>",
  "suggested_opener": "<1-2 sentence caller opener referencing a specific signal>",
  "why_now": "<1-2 sentence reason this company is worth calling now>",
  "evidence_sources": "<comma-separated URLs that support the conclusions>",
  "evidence_quality": "<Strong | Medium | Weak | Insufficient>",
  "confidence_level": "<High | Medium | Low | Unknown>",
  "manual_review_needed": true/false,
  "manual_review_reason": "<reason or empty>"
}"""

_EMPTY_CLAUDE_RESULT: dict = {
    "trigger_found": False,
    "trigger_type": "None",
    "trigger_date": "",
    "trigger_score": 0,
    "trigger_evidence": "",
    "annual_report_found": False,
    "annual_report_url": "",
    "annual_report_date": "",
    "fiscal_year_pattern": "Unknown",
    "likely_buying_window": "",
    "buying_window_score": 0,
    "buying_window_confidence": "Unknown",
    "buying_window_reason": "",
    "hiring_signal_score": 0,
    "international_hiring_signal": 0,
    "lnd_hr_hiring_signal": 0,
    "sales_cs_hiring_signal": 0,
    "onboarding_pressure_signal": 0,
    "preferred_buyer_route": "Unknown",
    "backup_buyer_route": "",
    "suggested_title_searches": "",
    "suggested_opener": "",
    "why_now": "",
    "evidence_sources": "",
    "evidence_quality": "Insufficient",
    "confidence_level": "Unknown",
    "manual_review_needed": True,
    "manual_review_reason": "No search results available",
}

# Fuzzy mapping: substrings in Claude's free-text → canonical trigger type
_TRIGGER_MAP: list[tuple[str, str]] = [
    # Most specific first
    ("multilingual",              "Multilingual workforce growth"),
    ("foreign hq",                "Foreign HQ / group communication"),
    ("cross-border",              "Foreign HQ / group communication"),
    ("group communication",       "Foreign HQ / group communication"),
    ("m&a",                       "M&A / integration"),
    ("merger",                    "M&A / integration"),
    ("acquisition",               "M&A / integration"),
    ("integrat",                  "M&A / integration"),
    ("funding",                   "Funding / growth investment"),
    ("investment",                "Funding / growth investment"),
    ("private equity",            "Funding / growth investment"),
    ("client-facing",             "Client-facing team expansion"),
    ("customer success",          "Client-facing team expansion"),
    ("account manag",             "Client-facing team expansion"),
    ("sales",                     "Sales / customer success growth"),
    ("revenue",                   "Sales / customer success growth"),
    ("international hiring",      "International hiring"),
    ("international expansion",   "New market or office expansion"),
    ("new market",                "New market or office expansion"),
    ("new office",                "New market or office expansion"),
    ("global expansion",          "New market or office expansion"),
    ("onboard",                   "Onboarding pressure"),
    ("rapid growth",              "Onboarding pressure"),
    ("fast hiring",               "Onboarding pressure"),
    ("employer brand",            "Employer branding / retention"),
    ("retention",                 "Employer branding / retention"),
    ("l&d hiring",                "HR / L&D hiring"),
    ("hr hiring",                 "HR / L&D hiring"),
    ("learning and development",  "HR / L&D hiring"),
    ("annual plan",               "Annual planning / budget window"),
    ("budget",                    "Annual planning / budget window"),
    ("fiscal year",               "Annual planning / budget window"),
    ("annual report",             "Annual planning / budget window"),
    ("hiring",                    "International hiring"),
]

_ALLOWED_TRIGGER_SET = set(ALLOWED_TRIGGER_TYPES)


def _normalize_trigger_type(raw: str) -> str:
    """Map Claude's free-text trigger_type to the fixed taxonomy."""
    if raw in _ALLOWED_TRIGGER_SET:
        return raw
    raw_lower = raw.lower()
    for fragment, canonical in _TRIGGER_MAP:
        if fragment in raw_lower:
            return canonical
    if not raw or raw.lower() in ("none", "no trigger", "no clear trigger"):
        return "No clear trigger"
    return "Other"


_TITLE_SEARCH_ROUTE_MAP: dict[str, str] = {
    "l&d / talent development":  '"Learning Development" OR "Talent Development" OR "L&D"',
    "hr / people":               '"HR Director" OR "People Director" OR "Head of People"',
    "international hr":          '"International HR" OR "Global HR" OR "People Operations"',
    "sales enablement":          '"Sales Enablement" OR "Revenue Enablement"',
    "customer success":          '"Customer Success Director" OR "VP Customer Success"',
    "people operations":         '"Onboarding" OR "People Operations" OR "HR Operations"',
    "operations":                '"Operations Director" OR "Head of Operations"',
}


def _quote_or_titles(text: str) -> str:
    """Ensure every OR-separated title token is wrapped in double quotes."""
    if not text:
        return text
    tokens = [t.strip() for t in text.split(" OR ")]
    quoted = []
    for t in tokens:
        if not t:
            continue
        t = t.strip('"').strip("'").strip()
        if t:
            quoted.append(f'"{t}"')
    return " OR ".join(quoted)


def _normalize_title_searches(raw: str, preferred_route: str, backup_route: str) -> str:
    """
    Ensure title searches use quoted OR syntax for LinkedIn Sales Navigator.
    If Claude returned a comma-separated list or unquoted OR list, rebuild from routes.
    """
    # Rebuild from routes when no OR syntax present
    if not (raw and " OR " in raw):
        parts = []
        for route in (preferred_route, backup_route):
            key = str(route or "").lower()
            for route_key, titles in _TITLE_SEARCH_ROUTE_MAP.items():
                if route_key in key:
                    parts.append(titles)
                    break

        if parts:
            return " OR ".join(parts)

        # Fallback: comma-separated → quoted OR syntax
        if raw and "," in raw:
            pieces = [p.strip().strip('"') for p in raw.split(",") if p.strip()]
            return " OR ".join(f'"{p}"' for p in pieces[:6])

        if raw:
            return _quote_or_titles(raw)
        return ""

    # Has OR syntax — ensure every token is quoted
    return _quote_or_titles(raw)


def _cap_opportunity_score(opp: float, rec: str, input_type: str, eq: str, trigger_type: str) -> float:
    """Apply post-processing caps to opportunity_score for simple company list."""
    if input_type != "simple_company_list":
        return opp
    eq_lower = eq.lower()
    if rec == "Manual research needed":
        opp = min(opp, 6.0)
    if eq_lower in ("medium", "weak", "insufficient") and trigger_type in ("No clear trigger", "Annual planning / budget window", "Other"):
        opp = min(opp, 5.5)
    if trigger_type == "No clear trigger":
        opp = min(opp, 4.5)
    return round(opp, 1)


def _infer_buyer_route_from_context(icp_evidence: str) -> str:
    """Infer a non-Unknown buyer route from ICP evidence text."""
    ev = (icp_evidence or "").lower()
    if any(k in ev for k in ("international", "multilingual", "global", "cross-border", "foreign hq")):
        return "International HR"
    if any(k in ev for k in ("l&d", "learning and development", "talent development", "training", "academy")):
        return "L&D / Talent Development"
    if any(k in ev for k in ("sales", "account manag", "revenue", "commercial team")):
        return "Sales Enablement"
    if any(k in ev for k in ("customer success", "client-facing", "client facing")):
        return "Customer Success"
    if any(k in ev for k in ("onboard", "people operations", "hr operations")):
        return "People Operations"
    return "HR / People"


def _fallback_opener(name: str, route: str) -> str:
    """Generate a cautious, mYngle-specific cold-call opener when Claude left it blank."""
    r = route.lower()
    if "international" in r:
        return (
            f"Given {name}'s international teams and cross-border operations, "
            "I wanted to ask whether Business English or cross-border communication training "
            "is currently part of your L&D planning."
        )
    if "l&d" in r or "talent development" in r:
        return (
            f"I noticed {name} has an active L&D function. "
            "I wanted to check whether language training or business communication support "
            "is currently part of your training agenda."
        )
    if "sales enablement" in r:
        return (
            f"Given {name}'s international sales and account teams, "
            "I wanted to ask whether Business English or client communication training "
            "is currently being reviewed."
        )
    if "customer success" in r:
        return (
            f"Given {name}'s customer-facing teams, "
            "I wanted to ask whether business communication or language training "
            "is currently on your L&D agenda."
        )
    return (
        f"Given {name}'s international presence, "
        "I wanted to ask whether Business English or cross-border communication training "
        "is currently part of your L&D planning."
    )


def _apply_enriched_fallback(
    adj: dict,
    fit_score_raw,
    tier_raw,
    icp_evidence: str,
    company_name: str,
    input_type: str,
) -> dict:
    """
    For enriched exports where Claude returned blank guidance, fill in fallback
    why_now, buyer route, opener, and buying_window from commercial fit context.
    Never called for simple company lists.
    """
    if input_type != "enriched_export":
        return adj

    fit = _fit_bucket(fit_score_raw, tier_raw)

    # Low-fit / Pass companies: ensure why_now explains the low priority cautiously
    if fit == 0:
        adj = dict(adj)
        adj["why_now"] = (
            f"{company_name} has a low commercial fit and no current trigger was found. "
            "No immediate language or communication training need is evident. "
            "A stronger signal — such as international hiring, L&D activity, or market expansion "
            "— would be needed before calling."
        )
        return adj

    # Infer buyer route from ICP evidence when Claude returned Unknown or blank
    if not adj.get("preferred_buyer_route") or adj.get("preferred_buyer_route") == "Unknown":
        adj = dict(adj)
        inferred = _infer_buyer_route_from_context(icp_evidence)
        adj["preferred_buyer_route"] = inferred
        adj["suggested_title_searches"] = _normalize_title_searches("", inferred, "")

    # Fill blank buying window
    if not adj.get("likely_buying_window"):
        adj = dict(adj)
        adj["likely_buying_window"] = "No clear buying window found"
        adj["buying_window_confidence"] = adj.get("buying_window_confidence") or "Unknown"

    # Fill blank why_now for commercially relevant companies
    if not adj.get("why_now"):
        adj = dict(adj)
        tier_str = str(tier_raw or "").strip()
        score_str = str(fit_score_raw or "").strip()
        if fit >= 3:
            fit_desc = "a very strong mYngle-fit company (tier: Hot)"
        elif fit >= 2:
            fit_desc = "a good mYngle-fit company (tier: Warm)"
        else:
            fit_desc = "a potential mYngle-fit company (tier: Cool)"
        adj["why_now"] = (
            f"{company_name} is {fit_desc}. "
            "No concrete recent timing trigger was found in available sources. "
            "Best next step: manually research current hiring, L&D planning, "
            "international team growth, or customer-facing expansion before calling."
        )

    # Fill blank opener
    if not adj.get("suggested_opener"):
        adj = dict(adj)
        adj["suggested_opener"] = _fallback_opener(
            company_name, adj.get("preferred_buyer_route", "")
        )

    return adj


# Matches hard quarter strings like Q3 2026, Q3-Q4 2026, Q3 2026 or Q4 2026
_HARD_QUARTER_RE = re.compile(
    r"\bQ[1-4]\s*(?:[-/]\s*Q[1-4]\s+|\s+or\s+Q[1-4]\s+|\s+)20\d{2}\b",
    re.IGNORECASE,
)


def _soften_hard_buying_window(adj: dict) -> dict:
    """
    Replace a hard quarter string with hedged language unless confidence is High.
    Claude often speculates a quarter without source evidence; this normalises those.
    """
    window = str(adj.get("likely_buying_window", "") or "")
    if not window or not _HARD_QUARTER_RE.search(window):
        return adj
    # Keep hard quarter only when Claude explicitly rated confidence as High
    confidence = str(adj.get("buying_window_confidence", "") or "").lower()
    if confidence == "high":
        return adj
    adj = dict(adj)
    trigger_type = str(adj.get("trigger_type", "") or "")
    if trigger_type == "Annual planning / budget window":
        adj["likely_buying_window"] = "Possible annual planning window, confidence low"
    else:
        adj["likely_buying_window"] = "Possible Q3/Q4 planning conversation, timing not confirmed"
    adj["buying_window_confidence"] = "Low"
    return adj


_LOW_PRIORITY_OPENER = (
    "Manual research first. Only contact if a current HR, hiring, onboarding, "
    "integration, or international communication signal appears."
)


def _apply_low_priority_override(adj: dict, rec: str, company_name: str) -> dict:
    """Suppress the sales opener and ensure cautious why_now for Low priority companies."""
    if rec != "Low priority":
        return adj
    adj = dict(adj)
    adj["suggested_opener"] = _LOW_PRIORITY_OPENER
    existing_why = adj.get("why_now", "") or ""
    # Overwrite if why_now is blank or looks like a positive pitch
    if not existing_why or "low commercial fit" not in existing_why.lower():
        adj["why_now"] = (
            f"{company_name} is low priority: no strong commercial fit or timing trigger "
            "was found in available sources. "
            "A stronger signal — such as international hiring, L&D activity, or market "
            "expansion — would be needed before calling."
        )
    return adj


# =============================================================================
# TRIGGER RECENCY
# =============================================================================

_MONTH_MAP = {
    "january":1, "february":2, "march":3, "april":4, "may":5, "june":6,
    "july":7, "august":8, "september":9, "october":10, "november":11, "december":12,
    "jan":1, "feb":2, "mar":3, "apr":4, "jun":6, "jul":7, "aug":8,
    "sep":9, "oct":10, "nov":11, "dec":12,
}

_DATE_FORMATS_HIGH = ("%Y-%m-%d", "%d %B %Y", "%B %d, %Y", "%d/%m/%Y", "%m/%d/%Y")
_DATE_FORMATS_MED  = ("%B %Y", "%b %Y")


def _parse_trigger_date(raw: str):
    """Return (datetime | None, confidence: str)."""
    if not raw or not raw.strip():
        return None, "Unknown"
    d = raw.strip()

    for fmt in _DATE_FORMATS_HIGH:
        try:
            return datetime.strptime(d, fmt), "High"
        except ValueError:
            pass

    for fmt in _DATE_FORMATS_MED:
        try:
            return datetime.strptime(d, fmt), "Medium"
        except ValueError:
            pass

    # Q1-Q4 YYYY
    m = re.match(r"Q([1-4])\s*(20\d{2})", d, re.IGNORECASE)
    if m:
        month = (int(m.group(1)) - 1) * 3 + 1
        return datetime(int(m.group(2)), month, 1), "Medium"

    # Loose "Month YYYY" or "YYYY Month"
    parts = d.lower().split()
    if len(parts) == 2:
        for a, b in ((0, 1), (1, 0)):
            if parts[a] in _MONTH_MAP:
                try:
                    return datetime(int(parts[b]), _MONTH_MAP[parts[a]], 1), "Medium"
                except (ValueError, KeyError):
                    pass

    # Year only
    m = re.match(r"^(20\d{2})$", d.strip())
    if m:
        return datetime(int(m.group(1)), 1, 1), "Low"

    return None, "Unknown"


def _compute_trigger_recency(trigger_date: str, trigger_type: str) -> dict:
    today = datetime.now()
    parsed, date_confidence = _parse_trigger_date(trigger_date)

    if parsed is None:
        return {
            "trigger_age_days":  "",
            "recency_bucket":    "Unknown date",
            "is_current_trigger": False,
            "date_confidence":   "Unknown",
            "recency_note":      "No reliable trigger date found, treat conservatively.",
        }

    age_days = max(0, (today - parsed).days)
    if age_days <= 90:
        bucket = "Fresh"
    elif age_days <= 180:
        bucket = "Recent-ish"
    elif age_days <= 365:
        bucket = "Old context"
    else:
        bucket = "Stale"

    is_annual_only = (trigger_type == "Annual planning / budget window")
    is_current     = (bucket == "Fresh") and not is_annual_only

    if is_annual_only:
        note = "Annual report timing only, not a direct buying trigger."
    elif bucket == "Fresh":
        note = "Recent trigger, can support active outreach."
    elif bucket == "Recent-ish":
        note = "Older signal, use as context only. Verify if still active."
    elif bucket == "Old context":
        note = "Older signal, use as context only."
    else:
        note = "Stale signal. Recheck current status before outreach."

    return {
        "trigger_age_days":  age_days,
        "recency_bucket":    bucket,
        "is_current_trigger": is_current,
        "date_confidence":   date_confidence,
        "recency_note":      note,
    }


def _apply_recency_guardrail(rec: str, recency: dict) -> str:
    """Downgrade urgent recommendations when the trigger signal is old or unknown.
    Each step operates on the already-downgraded rec so multi-hop chains resolve."""
    bucket = recency.get("recency_bucket", "Unknown date")
    if bucket == "Fresh":
        return rec
    if rec == "Call now":
        rec = "Call this month" if bucket == "Recent-ish" else "Call before budget cycle"
    if rec == "Call this month" and bucket in ("Old context", "Stale", "Unknown date"):
        rec = "Call before budget cycle"
    if rec == "Call before budget cycle" and bucket == "Stale":
        rec = "Manual research needed"
    return rec


_STALE_URGENCY_WORDS = (
    " recent ", " recently ", " active ", " actively ", " current ", " currently ",
    " ongoing ", " live ", " now ", " right now", " this year's ",
)


def _apply_stale_recency_wording(adj: dict, rec: str, company_name: str) -> dict:
    """
    Override why_now and suggested_opener when trigger evidence is stale or unknown,
    so caller-facing text does not imply fresh activity.
    Skipped for Low priority and Internal companies (already handled).
    """
    if rec in ("Low priority", "Internal / exclude"):
        return adj

    bucket = adj.get("recency_bucket", "")
    if not bucket or bucket == "Fresh":
        return adj

    adj = dict(adj)

    if bucket == "Stale":
        adj["why_now"] = (
            f"{company_name} has historical signals that may be relevant, "
            "but the available trigger evidence is stale (older than 12 months). "
            "Manual research is needed to confirm whether there is current activity, "
            "budget ownership, or a live language-training need."
        )
        adj["suggested_opener"] = (
            f"We have some older background context on {company_name}, "
            "but our information may be outdated. "
            "Has language or business communication training come up recently in your planning?"
        )

    elif bucket == "Old context":
        why = adj.get("why_now", "")
        why_lower = why.lower()
        if any(w in why_lower for w in _STALE_URGENCY_WORDS):
            adj["why_now"] = (
                why.rstrip(".") + ". "
                "Note: the trigger signal is several months old — "
                "use as context only and verify current status before outreach."
            )

    elif bucket == "Unknown date":
        why = adj.get("why_now", "")
        if why and "date could not be confirmed" not in why.lower():
            adj["why_now"] = (
                why.rstrip(".") + ". "
                "Note: the trigger date could not be confirmed — "
                "treat timing conservatively and verify before outreach."
            )

    return adj


def _apply_simple_fallback(adj: dict) -> dict:
    """For simple company lists with no timing trigger, ensure why_now is never blank."""
    if adj.get("why_now"):
        return adj
    adj = dict(adj)
    adj["why_now"] = (
        "No clear current timing trigger was found from the available sources. "
        "Monitor for international hiring, L&D hiring, onboarding, "
        "sales/customer success expansion, or annual planning signals."
    )
    return adj


def _compute_caller_caution(adj: dict, rec: str) -> dict:
    """Derive caution_note, reason_not_to_call_now, missing_evidence from final rec + recency.

    These fields protect cold callers from overclaiming or calling at the wrong time.
    """
    bucket       = adj.get("recency_bucket", "")
    trigger_type = adj.get("trigger_type", "")

    caution_parts: list = []
    reason_parts:  list = []
    missing_parts: list = []

    if rec == "Low priority":
        caution_parts.append(
            "Low priority: do not actively contact unless a stronger current signal appears."
        )
        reason_parts.append(
            "No strong commercial fit or timing trigger found in available sources."
        )
        missing_parts.append(
            "Stronger signal needed: international hiring, L&D activity, "
            "onboarding pressure, M&A, or explicit training initiative."
        )

    if rec == "Call before budget cycle":
        caution_parts.append(
            "Soft timing angle only — no evidence of active buying intent. "
            "Use as context, not a primary hook."
        )

    if bucket in ("Stale", "Old context", "Unknown date"):
        _label = {
            "Stale":        "stale (older than 12 months)",
            "Old context":  "several months old",
            "Unknown date": "undated",
        }.get(bucket, bucket.lower())
        caution_parts.append(
            f"Trigger signal is {_label} — not confirmed as current. "
            "Do not imply the company is actively buying."
        )
        if not reason_parts:
            reason_parts.append(
                "Current evidence is stale or unconfirmed. "
                "Verify with current sources before outreach."
            )

    if trigger_type in ("No clear trigger", "No clear trigger found", "") and not reason_parts:
        reason_parts.append(
            "No current timing trigger was found in available sources."
        )

    if trigger_type in ("No clear trigger", "No clear trigger found", "") and not missing_parts:
        missing_parts.append(
            "Stronger signal needed: HR/L&D hiring, international expansion, "
            "onboarding pressure, M&A activity, or explicit training initiative."
        )

    return {
        "caution_note":           " | ".join(caution_parts),
        "reason_not_to_call_now": " | ".join(reason_parts),
        "missing_evidence":       " | ".join(missing_parts),
    }


def _compute_caller_prep_fields(r: dict) -> dict:
    """Derive final_opener, discovery_question_1/2, evidence_to_mention,
    what_not_to_overclaim from already-computed result fields.

    Returns a dict with the 5 keys.  For internal/excluded companies every
    value is left blank so no outreach copy is generated.
    """
    scores  = r.get("scores",  {})
    claude  = r.get("claude",  {})
    enriched = r.get("enriched_row", {})

    rec         = str(scores.get("call_recommendation", "") or "").strip()
    is_internal = rec in ("Internal / exclude",)

    if is_internal:
        return {
            "final_opener":        "",
            "discovery_question_1": "",
            "discovery_question_2": "",
            "evidence_to_mention": "",
            "what_not_to_overclaim": "Internal / exclude — do not generate outreach content.",
        }

    def _sv(*keys_dicts):
        """Return first non-blank value searching key/dict pairs."""
        for key, d in keys_dicts:
            v = str(d.get(key, "") or "").strip()
            if v and v.lower() not in ("nan", "none", "n/a"):
                return v
        return ""

    company         = str(r.get("company_name", "") or "").strip()
    suggested_opener = _sv(("suggested_opener", claude))
    icp_why         = _sv(("icp_why_relevant",       enriched),
                           ("icp_why_relevant",       claude))
    why_now         = _sv(("why_now",                 claude))
    buyer_route     = _sv(("preferred_buyer_route",   claude))
    training_interest = _sv(("icp_likely_training_interest", enriched),
                             ("icp_likely_training_interest", claude))
    trigger_type    = _sv(("trigger_type",            claude))
    is_current      = claude.get("is_current_trigger", False)
    recency_bucket  = _sv(("recency_bucket",          claude))
    evidence_quality = _sv(("evidence_quality",       claude))
    confidence_level = _sv(("confidence_level",       claude))
    caution_note    = _sv(("caution_note",             claude))
    missing_ev      = _sv(("missing_evidence",         claude))
    icp_evidence    = _sv(("icp_evidence",             enriched),
                           ("icp_evidence",             claude))
    icp_signals     = _sv(("icp_buying_signals",       enriched),
                           ("icp_buying_signals",       claude))
    ev_summary      = _sv(("trigger_evidence",         claude),
                           ("evidence_summary",         claude))
    latest_src_date = _sv(("latest_source_date",       scores))

    # ── 1. final_opener ───────────────────────────────────────────────────────
    if suggested_opener:
        final_opener = suggested_opener
    else:
        # Build a conservative fallback from available context
        parts = []
        if buyer_route:
            parts.append(f"I'm reaching out to {buyer_route}s at {company or 'your organisation'}.")
        elif company:
            parts.append(f"I'm reaching out regarding {company}.")
        if is_current and why_now:
            parts.append(why_now[:120].rstrip(".") + ".")
        elif icp_why:
            parts.append(icp_why[:120].rstrip(".") + ".")
        parts.append(
            "We work with international companies on Business English and "
            "cross-border communication training. I'd love to explore whether "
            "that's relevant for your team."
        )
        final_opener = " ".join(parts)

    # ── 2. discovery_question_1 ───────────────────────────────────────────────
    li = (training_interest or icp_why or "").lower()
    if "onboard" in li:
        dq1 = (
            "Is onboarding of new international employees currently creating "
            "new language or communication training needs?"
        )
    elif "l&d" in li or "learning" in li or "development" in li:
        dq1 = (
            "Is language training or business communication currently part "
            "of your L&D planning for this year?"
        )
    elif "client" in li or "client-facing" in li:
        dq1 = (
            "Are client-facing teams currently receiving support for "
            "business communication or professional English?"
        )
    elif buyer_route and "hr" in buyer_route.lower():
        dq1 = (
            "Are international teams currently facing communication or "
            "language challenges that HR is looking to address?"
        )
    else:
        dq1 = (
            "Is language training or business communication currently part "
            "of your team development plans?"
        )

    # ── 3. discovery_question_2 ───────────────────────────────────────────────
    if buyer_route and ("l&d" in buyer_route.lower() or "learning" in buyer_route.lower()):
        dq2 = (
            "Which teams would benefit most from stronger Business English "
            "or cross-border communication support?"
        )
    elif buyer_route and "hr" in buyer_route.lower():
        dq2 = (
            "Are these training needs handled centrally by HR and L&D, "
            "or by individual business units?"
        )
    else:
        dq2 = (
            "Do you currently work with external providers for language or "
            "communication training, or is that handled in-house?"
        )

    # ── 4. evidence_to_mention ────────────────────────────────────────────────
    ev_parts: list[str] = []
    for snippet in (icp_evidence, icp_signals, ev_summary):
        if snippet:
            ev_parts.append(snippet[:180].rstrip("."))
            break  # one clear snippet is enough
    date_note = f" (last signal: {latest_src_date})" if latest_src_date else ""
    if ev_parts:
        evidence_to_mention = ev_parts[0].strip() + date_note + "."
    else:
        evidence_to_mention = (
            "Public sources suggest this company has international operations "
            "or internal employee development activity."
            + date_note
        )

    # ── 5. what_not_to_overclaim ──────────────────────────────────────────────
    wno_parts: list[str] = []

    if not is_current:
        wno_parts.append(
            "Do not imply the company is actively buying or has confirmed training needs."
        )

    if recency_bucket in ("Stale", "Old context", "Unknown date"):
        wno_parts.append("Signal recency is uncertain — do not reference specific recent events.")

    if confidence_level.lower() in ("low", "unknown", ""):
        wno_parts.append(
            "Evidence confidence is low. Treat this as a fit-based lead, "
            "not a trigger-confirmed lead."
        )
    elif evidence_quality.lower() in ("weak", "insufficient"):
        wno_parts.append(
            "Evidence is thin. Do not cite specific sources or claim confirmed activity."
        )

    if rec in ("Monitor", "Low priority"):
        wno_parts.append(
            "This company is not yet ready to contact. Do not initiate outreach."
        )
    elif rec == "Call before budget cycle":
        wno_parts.append(
            "No clear buying window confirmed. Do not mention a specific budget cycle."
        )

    if trigger_type in ("No clear trigger", "No clear trigger found", ""):
        if not any("trigger" in p.lower() for p in wno_parts):
            wno_parts.append("No current timing trigger was confirmed.")

    if caution_note and caution_note not in " | ".join(wno_parts):
        wno_parts.append(caution_note)

    what_not_to_overclaim = " | ".join(wno_parts) if wno_parts else (
        "Standard caution: position mYngle as a relevant partner, "
        "not a response to a confirmed active need."
    )

    return {
        "final_opener":          final_opener,
        "discovery_question_1":  dq1,
        "discovery_question_2":  dq2,
        "evidence_to_mention":   evidence_to_mention,
        "what_not_to_overclaim": what_not_to_overclaim,
    }


def _compute_scores(
    claude_result: dict,
    fit_score_raw,
    tier_raw,
    input_type: str = "enriched_export",
    icp_evidence: str = "",
    company_name: str = "",
) -> tuple:
    """
    Returns (adjusted_claude_result, scores_dict).
    adjusted_claude_result has any expired buying window projected forward,
    banned phrases sanitized, and (for enriched exports) fallback guidance filled
    from commercial fit context when Claude returned no trigger.
    """
    adj = _adjust_past_buying_window(dict(claude_result))
    # Soften any hard-quarter buying window that lacks High confidence
    adj = _soften_hard_buying_window(adj)

    # Sanitize banned generic phrases — runs on cached AND fresh results
    adj = _sanitize_claude_result(adj)

    # Normalize trigger_type to fixed taxonomy
    adj["trigger_type"] = _normalize_trigger_type(adj.get("trigger_type", ""))

    # Normalize title searches to OR syntax
    adj["suggested_title_searches"] = _normalize_title_searches(
        adj.get("suggested_title_searches", ""),
        adj.get("preferred_buyer_route", ""),
        adj.get("backup_buyer_route", ""),
    )

    # Annual report alone is weak evidence — cap trigger_score at 1
    if adj.get("trigger_type") == "Annual planning / budget window":
        raw_trigger = int(adj.get("trigger_score", 0) or 0)
        if raw_trigger > 1:
            adj = dict(adj)
            adj["trigger_score"] = 1

    # For simple lists, never use commercial fit in scoring
    fit     = _fit_bucket(fit_score_raw, tier_raw) if input_type == "enriched_export" else 1
    trigger = int(adj.get("trigger_score", 0) or 0)
    window  = int(adj.get("buying_window_score", 0) or 0)
    route   = _contact_route_score(adj.get("preferred_buyer_route", ""))
    opp     = _opportunity_score(fit, trigger, window, route, input_type)
    eq      = str(adj.get("evidence_quality", ""))
    manual  = bool(adj.get("manual_review_needed", False))

    # For simple lists: cap confidence and force manual when evidence is weak
    if input_type == "simple_company_list":
        eq_low = eq.lower()
        if eq_low in ("weak", "insufficient"):
            manual = True
        cl = str(adj.get("confidence_level", "")).lower()
        if cl == "high":
            adj["confidence_level"] = "Medium"

    rec = _call_recommendation(fit, trigger, window, opp, manual, input_type, eq)

    # Compute trigger recency and apply guardrail before capping score
    recency = _compute_trigger_recency(adj.get("trigger_date", ""), adj.get("trigger_type", ""))
    rec = _apply_recency_guardrail(rec, recency)

    # Cap opportunity_score for simple inputs that receive conservative recommendations
    opp = _cap_opportunity_score(opp, rec, input_type, eq, adj.get("trigger_type", ""))

    # For enriched exports: fill blank guidance from commercial fit context
    adj = _apply_enriched_fallback(adj, fit_score_raw, tier_raw, icp_evidence, company_name, input_type)
    # For simple lists: ensure why_now is never blank when no trigger found
    if input_type == "simple_company_list":
        adj = _apply_simple_fallback(adj)
    # Suppress sales opener and enforce cautious copy for Low priority companies
    adj = _apply_low_priority_override(adj, rec, company_name)
    # Store recency fields in adj so they flow into Excel output
    adj = dict(adj)
    adj.update(recency)
    # Override stale/old caller-facing wording so text matches recency bucket
    adj = _apply_stale_recency_wording(adj, rec, company_name)
    # Add caution fields so cold callers don't overclaim
    adj.update(_compute_caller_caution(adj, rec))
    # Re-normalise route score after possible fallback route assignment
    route = _contact_route_score(adj.get("preferred_buyer_route", ""))

    scores = {
        "trigger_score":        trigger,
        "buying_window_score":  window,
        "contact_route_score":  route,
        "opportunity_score":    opp,
        "call_recommendation":  rec,
    }
    return adj, scores


# =============================================================================
# COMPANY LIST BUILDER
# =============================================================================

def _make_company_key(name: str, domain: str) -> str:
    """Return a normalized slug 'name__domain' suitable for dedup and traceability."""
    n = re.sub(r"[^\w\s]", "", name.lower())
    n = re.sub(r"\s+", "_", n.strip())
    d = re.sub(r"[.\-]", "_", domain.lower().strip())
    return f"{n}__{d}" if d else n


def _build_company_list(
    df: pd.DataFrame,
    name_col: str | None,
    domain_col: str | None,
    country_col: str | None,
    score_col: str | None,
    tier_col: str | None,
    icp_col: str | None,
    input_type: str = "enriched_export",
) -> list:
    """Deduplicate by company name and return list of dicts."""
    # If name_col wasn't detected, fall back to the first string-like column
    if name_col is None:
        for col in df.columns:
            if df[col].dtype == object:
                name_col = col
                break

    def _val(row, col):
        if col and col in row.index:
            v = row[col]
            return "" if pd.isna(v) else str(v).strip()
        return ""

    seen: set = set()
    companies = []
    # Find optional traceability columns added by Lead Prioritizer domain validation
    _validated_domain_col = _detect_col(df, ["validated_domain"])
    _input_domain_col     = _detect_col(df, ["input_domain"])

    for i, row in df.iterrows():
        name   = _val(row, name_col)
        # Prefer validated_domain when present and non-empty; fall back to domain_col
        _vd = _normalize_domain(_val(row, _validated_domain_col)) if _validated_domain_col else ""
        domain = _vd or _normalize_domain(_val(row, domain_col))

        # Fallback: use domain as display key when name is absent
        key = name or domain
        if not key:
            continue  # skip rows with no identity at all

        # Deduplicate by normalised name (prefer name over domain)
        dedup_key = name.lower() if name else domain.lower()
        if dedup_key in seen:
            continue
        seen.add(dedup_key)

        # Carry the full input row as a dict for Caller Prep Input passthrough
        def _sv(v):
            try:
                return "" if pd.isna(v) else v
            except Exception:
                return v
        enriched_row = {col: _sv(row[col]) for col in df.columns if col in row.index}

        # Stable fallback IDs — preserve existing values when present
        seq = len(companies) + 1
        if not str(enriched_row.get("lead_id", "") or "").strip():
            enriched_row["lead_id"] = f"OR_{seq:06d}"
        if not str(enriched_row.get("company_key", "") or "").strip():
            enriched_row["company_key"] = _make_company_key(name, domain)

        # Exclude internal / self entries
        if _is_internal(name, domain):
            companies.append({
                "company_name":             name,
                "domain":                   domain,
                "country":                  _val(row, country_col),
                "fit_score":                "",
                "tier":                     "",
                "icp_evidence":             "",
                "input_type":               input_type,
                "commercial_fit_available": False,
                "internal":                 True,
                "enriched_row":             enriched_row,
            })
            continue

        fit_score = _val(row, score_col)
        tier      = _val(row, tier_col)
        fit_avail = bool(fit_score or tier) and input_type == "enriched_export"
        companies.append({
            "company_name":             name,
            "domain":                   domain,
            "country":                  _val(row, country_col),
            "fit_score":                fit_score,
            "tier":                     tier,
            "icp_evidence":             _val(row, icp_col),
            "input_type":               input_type,
            "commercial_fit_available": fit_avail,
            "internal":                 False,
            "enriched_row":             enriched_row,
        })
    return companies


# =============================================================================
# CALLER PREP INPUT SHEET
# =============================================================================

# Column spec: (col_name, src, src_key, is_numeric, num_fmt, width, wrap)
# src: "result" | "enriched" | "claude" | "scores" | "placeholder"
_CPI_COLUMN_SPEC = [
    # Group 1 — Company identity
    ("lead_id",                   "enriched",     "lead_id",                   False, None,    16, False),
    ("company_key",               "enriched",     "company_key",               False, None,    16, False),
    ("company_name",              "result",       "company_name",              False, None,    30, False),
    ("input_company_name",        "enriched",     "input_company_name",        False, None,    30, False),
    ("normalized_company_name",   "enriched",     "normalized_company_name",   False, None,    26, False),
    ("domain",                    "result",       "domain",                    False, None,    26, False),
    ("input_domain",              "enriched",     "input_domain",              False, None,    26, False),
    ("validated_domain",          "enriched",     "validated_domain",          False, None,    26, False),
    ("domain_used_for_enrichment","enriched",     "domain_used_for_enrichment",False, None,    20, False),
    ("country",                   "result",       "country",                   False, None,    14, False),
    ("city",                      "enriched",     "city",                      False, None,    14, False),
    ("industry",                  "enriched",     "industry",                  False, None,    22, False),
    ("employee_range",            "enriched",     "employee_range",            False, None,    14, False),
    # Group 2 — Data quality
    ("domain_match_confidence",   "enriched",     "domain_match_confidence",   False, None,    16, False),
    ("possible_domain_mismatch",  "enriched",     "possible_domain_mismatch",  False, None,    14, False),
    ("suggested_domain",          "enriched",     "suggested_domain",          False, None,    26, False),
    ("domain_check_reason",       "enriched",     "domain_check_reason",       False, None,    40, True),
    ("domain_source",             "enriched",     "domain_source",             False, None,    16, False),
    ("needs_domain_review",       "enriched",     "needs_domain_review",       False, None,    14, False),
    ("needs_manual_review",       "enriched",     "needs_manual_review",       False, None,    14, False),
    ("match_notes",               "enriched",     "match_notes",               False, None,    40, True),
    ("scoring_notes",             "enriched",     "scoring_notes",             False, None,    40, True),
    # Group 3 — Commercial fit / ICP
    ("commercial_fit_score",      "result",       "fit_score",                 True,  "0.00",  12, False),
    ("commercial_tier",           "result",       "tier",                      False, None,    14, False),
    ("model_probability",         "enriched",     "model_probability",         True,  "0.000", 12, False),
    ("lean_model_prob",           "enriched",     "lean_model_prob",           True,  "0.000", 12, False),
    ("icp_lead_score",            "enriched",     "icp_lead_score",            True,  "0.00",  12, False),
    ("icp_buying_signals",        "enriched",     "icp_buying_signals",        False, None,    40, True),
    ("icp_evidence",              "enriched",     "icp_evidence",              False, None,    50, True),
    ("icp_why_relevant",          "enriched",     "icp_why_relevant",          False, None,    50, True),
    ("icp_likely_training_interest","enriched",   "icp_likely_training_interest",False,None,   40, True),
    ("icp_potential_buyer_function","enriched",   "icp_potential_buyer_function",False,None,   30, False),
    ("top_positive_signals",      "enriched",     "top_positive_signals",      False, None,    40, True),
    ("gaps_missing_signals",      "enriched",     "gaps_missing_signals",      False, None,    40, True),
    # Group 4 — Opportunity Radar / timing
    ("call_recommendation",       "scores",       "call_recommendation",       False, None,    22, False),
    ("opportunity_score",         "scores",       "opportunity_score",         True,  "0.00",  12, False),
    ("trigger_type",              "claude",       "trigger_type",              False, None,    28, False),
    ("trigger_score",             "scores",       "trigger_score",             True,  "0",     10, False),
    ("trigger_date",              "claude",       "trigger_date",              False, None,    14, False),
    ("trigger_age_days",          "claude",       "trigger_age_days",          True,  "0",     12, False),
    ("recency_bucket",            "claude",       "recency_bucket",            False, None,    16, False),
    ("is_current_trigger",        "claude",       "is_current_trigger",        False, None,    14, False),
    ("date_confidence",           "claude",       "date_confidence",           False, None,    14, False),
    ("recency_note",              "claude",       "recency_note",              False, None,    40, True),
    ("likely_buying_window",      "claude",       "likely_buying_window",      False, None,    36, False),
    ("buying_window_score",       "scores",       "buying_window_score",       True,  "0",     12, False),
    ("evidence_quality",          "claude",       "evidence_quality",          False, None,    14, False),
    ("confidence_level",          "claude",       "confidence_level",          False, None,    14, False),
    ("why_now",                   "claude",       "why_now",                   False, None,    55, True),
    ("evidence_summary",          "claude",       "trigger_evidence",          False, None,    55, True),
    # Group 5 — Buyer route / call prep
    ("preferred_buyer_route",     "claude",       "preferred_buyer_route",     False, None,    28, False),
    ("backup_buyer_route",        "claude",       "backup_buyer_route",        False, None,    28, False),
    ("suggested_title_searches",  "claude",       "suggested_title_searches",  False, None,    45, True),
    ("suggested_opener",          "claude",       "suggested_opener",          False, None,    55, True),
    ("caution_note",              "claude",       "caution_note",              False, None,    55, True),
    ("reason_not_to_call_now",    "claude",       "reason_not_to_call_now",    False, None,    55, True),
    ("missing_evidence",          "claude",       "missing_evidence",          False, None,    55, True),
    # Group 6 — Source summary
    ("top_source_urls",           "computed",     "top_source_urls",           False, None,    45, True),
    ("raw_source_summary",        "computed",     "raw_source_summary",        False, None,    30, False),
    ("source_count",              "computed",     "source_count",              True,  "0",     12, False),
    ("latest_source_date",        "computed",     "latest_source_date",        False, None,    16, False),
    # Group 7 — Layer 3 workflow placeholders (all blank)
    ("selected_for_calling",      "placeholder",  None,                        False, None,    16, False),
    ("assigned_to",               "placeholder",  None,                        False, None,    16, False),
    ("call_batch",                "placeholder",  None,                        False, None,    14, False),
    ("contact_1_name",            "placeholder",  None,                        False, None,    20, False),
    ("contact_1_title",           "placeholder",  None,                        False, None,    22, False),
    ("contact_1_linkedin_url",    "placeholder",  None,                        False, None,    30, False),
    ("contact_1_email",           "placeholder",  None,                        False, None,    24, False),
    ("contact_1_fit",             "placeholder",  None,                        False, None,    14, False),
    ("contact_1_notes",           "placeholder",  None,                        False, None,    30, True),
    ("contact_2_name",            "placeholder",  None,                        False, None,    20, False),
    ("contact_2_title",           "placeholder",  None,                        False, None,    22, False),
    ("contact_2_linkedin_url",    "placeholder",  None,                        False, None,    30, False),
    ("contact_2_email",           "placeholder",  None,                        False, None,    24, False),
    ("contact_2_fit",             "placeholder",  None,                        False, None,    14, False),
    ("contact_2_notes",           "placeholder",  None,                        False, None,    30, True),
    ("contact_3_name",            "placeholder",  None,                        False, None,    20, False),
    ("contact_3_title",           "placeholder",  None,                        False, None,    22, False),
    ("contact_3_linkedin_url",    "placeholder",  None,                        False, None,    30, False),
    ("contact_3_email",           "placeholder",  None,                        False, None,    24, False),
    ("contact_3_fit",             "placeholder",  None,                        False, None,    14, False),
    ("contact_3_notes",           "placeholder",  None,                        False, None,    30, True),
    ("final_opener",              "placeholder",  None,                        False, None,    40, True),
    ("discovery_question_1",      "placeholder",  None,                        False, None,    40, True),
    ("discovery_question_2",      "placeholder",  None,                        False, None,    40, True),
    ("evidence_to_mention",       "placeholder",  None,                        False, None,    40, True),
    ("what_not_to_overclaim",     "placeholder",  None,                        False, None,    40, True),
    ("call_notes",                "placeholder",  None,                        False, None,    40, True),
    ("call_outcome",              "placeholder",  None,                        False, None,    20, False),
    ("next_step",                 "placeholder",  None,                        False, None,    30, False),
    ("sales_feedback",            "placeholder",  None,                        False, None,    40, True),
]


def _cpi_get(r: dict, src: str, key: str | None, is_numeric: bool):
    """Extract a value from the correct sub-dict in a result record."""
    if src == "placeholder" or key is None:
        return None
    if src == "result":
        raw = r.get(key, "")
    elif src == "enriched":
        raw = r.get("enriched_row", {}).get(key, "")
    elif src == "claude":
        raw = r.get("claude", {}).get(key, "")
    elif src == "scores":
        raw = r.get("scores", {}).get(key, "")
    else:
        raw = ""

    # Normalize blank-ish values
    if raw is None or (isinstance(raw, float) and raw != raw):  # NaN check
        return None
    if isinstance(raw, str) and raw.strip().lower() in ("", "nan", "none"):
        return None

    if is_numeric:
        try:
            return float(str(raw).strip())
        except (ValueError, TypeError):
            return None
    return raw if not isinstance(raw, bool) else str(raw)


def _write_caller_prep_sheet(ws, results: list) -> None:
    """Write the Caller Prep Input sheet — one row per company, fully formatted."""
    hdr_fill = PatternFill(start_color="0B4A92", end_color="0B4A92", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)

    n_cols = len(_CPI_COLUMN_SPEC)

    # Header row
    for ci, (col_name, *_rest) in enumerate(_CPI_COLUMN_SPEC, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    # Column widths
    for ci, (_, _src, _key, _num, _fmt, width, _wrap) in enumerate(_CPI_COLUMN_SPEC, 1):
        ws.column_dimensions[get_column_letter(ci)].width = width

    # Data rows
    for ri, r in enumerate(results, 2):
        # Determine row background from tier > recommendation
        tier = str(r.get("tier", "") or "").strip()
        rec  = str(r.get("scores", {}).get("call_recommendation", "") or "").strip()
        row_color = _CPI_TIER_FILLS.get(tier) or _CPI_REC_FILLS.get(rec)
        row_fill  = (
            PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
            if row_color else None
        )

        # Compute source summary fields once per row
        raw_srcs = r.get("raw_sources", [])
        urls = list(dict.fromkeys(
            s.get("url", "") for s in raw_srcs if s.get("url")
        ))[:5]
        src_dates = [s.get("date", "") for s in raw_srcs if s.get("date")]
        src_dates_clean = sorted([d for d in src_dates if d], reverse=True)
        computed = {
            "top_source_urls":   " | ".join(urls),
            "raw_source_summary": (
                f"{len(raw_srcs)} sources"
                + (f" across {len({s.get('query_group','') for s in raw_srcs} - {''})} query groups"
                   if raw_srcs else "")
            ),
            "source_count":     float(len(raw_srcs)),
            "latest_source_date": src_dates_clean[0] if src_dates_clean else "",
        }

        domain = str(r.get("domain", "") or "")

        # Compute the 5 caller-prep fields once per row
        caller_prep = _compute_caller_prep_fields(r)

        for ci, (col_name, src, key, is_numeric, num_fmt, _width, wrap) in \
                enumerate(_CPI_COLUMN_SPEC, 1):
            if src == "computed":
                raw = computed.get(key)
                if raw is not None and is_numeric:
                    try:
                        val = float(raw)
                    except Exception:
                        val = raw
                else:
                    val = raw
            elif col_name in caller_prep:
                # Caller-prep fields: use generated value instead of blank placeholder
                val = caller_prep[col_name]
            else:
                val = _cpi_get(r, src, key, is_numeric)

            # Domain traceability fallbacks for older exports without Layer 0 fields
            if not val or val == "":
                if col_name == "input_domain":
                    val = domain
                elif col_name == "validated_domain":
                    val = domain
                elif col_name == "domain_used_for_enrichment":
                    val = "domain"
                elif col_name == "domain_match_confidence":
                    val = "Unknown"

            cell = ws.cell(row=ri, column=ci, value=val)
            if row_fill:
                cell.fill = row_fill
            if is_numeric and val is not None and num_fmt:
                cell.number_format = num_fmt
            if wrap:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Compact row height — prevent openpyxl from auto-expanding to text size
        ws.row_dimensions[ri].height = 20

    # Freeze top row + autofilter
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22
    if results:
        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"


# =============================================================================
# EXCEL BUILDER
# =============================================================================

def _copy_sheet(src_ws, dst_wb, dest_name: str) -> None:
    """Copy a worksheet cell-by-cell into dst_wb with basic style preservation."""
    from copy import copy as _copy
    dst_ws = dst_wb.create_sheet(dest_name)
    for row in src_ws.iter_rows():
        for cell in row:
            dst = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                try:
                    dst.font        = _copy(cell.font)
                    dst.fill        = _copy(cell.fill)
                    dst.border      = _copy(cell.border)
                    dst.alignment   = _copy(cell.alignment)
                    dst.number_format = cell.number_format
                except Exception:
                    pass
    for col_letter, col_dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = col_dim.width
    for row_num, row_dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_num].height = row_dim.height
    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes
    if src_ws.auto_filter.ref:
        dst_ws.auto_filter.ref = src_ws.auto_filter.ref


def _build_excel_bytes(results: list, raw_sources: list,
                       input_bytes: bytes | None = None) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:

        # ── Sheet 1: Opportunity Radar (main summary) ─────────────────────────
        radar_cols = [
            "company_name", "domain", "country",
            "input_type", "commercial_fit_available",
            "commercial_fit_score", "commercial_tier",
            "trigger_score", "buying_window_score",
            "contact_route_score", "opportunity_score",
            "call_recommendation", "is_current_trigger", "recency_bucket",
            "why_now", "likely_buying_window",
            "preferred_buyer_route", "backup_buyer_route",
            "suggested_title_searches", "suggested_opener",
            "confidence_level", "evidence_quality", "manual_review_needed",
        ]
        radar_rows = []
        for r in results:
            c = r.get("claude", {})
            s = r.get("scores", {})
            radar_rows.append({
                "company_name":             r.get("company_name", ""),
                "domain":                   r.get("domain", ""),
                "country":                  r.get("country", ""),
                "input_type":               r.get("input_type", ""),
                "commercial_fit_available": r.get("commercial_fit_available", False),
                "commercial_fit_score":     r.get("fit_score", ""),
                "commercial_tier":          r.get("tier", ""),
                "trigger_score":            s.get("trigger_score", 0),
                "buying_window_score":      s.get("buying_window_score", 0),
                "contact_route_score":      s.get("contact_route_score", 0),
                "opportunity_score":        s.get("opportunity_score", 0),
                "call_recommendation":      s.get("call_recommendation", ""),
                "is_current_trigger":       c.get("is_current_trigger", False),
                "recency_bucket":           c.get("recency_bucket", ""),
                "why_now":                  c.get("why_now", ""),
                "likely_buying_window":     c.get("likely_buying_window", ""),
                "preferred_buyer_route":    c.get("preferred_buyer_route", ""),
                "backup_buyer_route":       c.get("backup_buyer_route", ""),
                "suggested_title_searches": c.get("suggested_title_searches", ""),
                "suggested_opener":         c.get("suggested_opener", ""),
                "confidence_level":         c.get("confidence_level", ""),
                "evidence_quality":         c.get("evidence_quality", ""),
                "manual_review_needed":     c.get("manual_review_needed", False),
            })
        pd.DataFrame(radar_rows, columns=radar_cols).to_excel(
            writer, index=False, sheet_name="Opportunity Radar"
        )

        # ── Sheet 2: Trigger Evidence ─────────────────────────────────────────
        trig_rows = []
        for r in results:
            c = r.get("claude", {})
            trig_rows.append({
                "company_name":    r.get("company_name", ""),
                "trigger_found":   c.get("trigger_found", False),
                "trigger_type":    c.get("trigger_type", ""),
                "trigger_date":    c.get("trigger_date", ""),
                "trigger_age_days": c.get("trigger_age_days", ""),
                "recency_bucket":  c.get("recency_bucket", ""),
                "is_current_trigger": c.get("is_current_trigger", False),
                "date_confidence": c.get("date_confidence", ""),
                "recency_note":    c.get("recency_note", ""),
                "trigger_score":   c.get("trigger_score", 0),
                "trigger_evidence": c.get("trigger_evidence", ""),
                "source_urls":     c.get("evidence_sources", ""),
            })
        pd.DataFrame(trig_rows).to_excel(
            writer, index=False, sheet_name="Trigger Evidence"
        )

        # ── Sheet 3: Buying Window ────────────────────────────────────────────
        bw_rows = []
        for r in results:
            c = r.get("claude", {})
            bw_rows.append({
                "company_name":          r.get("company_name", ""),
                "annual_report_found":   c.get("annual_report_found", False),
                "annual_report_url":     c.get("annual_report_url", ""),
                "annual_report_date":    c.get("annual_report_date", ""),
                "fiscal_year_pattern":   c.get("fiscal_year_pattern", ""),
                "likely_buying_window":  c.get("likely_buying_window", ""),
                "buying_window_score":   c.get("buying_window_score", 0),
                "buying_window_confidence": c.get("buying_window_confidence", ""),
                "buying_window_reason":  c.get("buying_window_reason", ""),
            })
        pd.DataFrame(bw_rows).to_excel(
            writer, index=False, sheet_name="Buying Window"
        )

        # ── Sheet 4: Contact Route ────────────────────────────────────────────
        cr_rows = []
        for r in results:
            c = r.get("claude", {})
            cr_rows.append({
                "company_name":            r.get("company_name", ""),
                "preferred_buyer_route":   c.get("preferred_buyer_route", ""),
                "backup_buyer_route":      c.get("backup_buyer_route", ""),
                "suggested_title_searches": c.get("suggested_title_searches", ""),
                "suggested_opener":        c.get("suggested_opener", ""),
            })
        pd.DataFrame(cr_rows).to_excel(
            writer, index=False, sheet_name="Contact Route"
        )

        # ── Sheet 5: Caller Brief ─────────────────────────────────────────────
        brief_rows = []
        for r in results:
            c = r.get("claude", {})
            s = r.get("scores", {})
            brief_rows.append({
                "company_name":         r.get("company_name", ""),
                "domain":               r.get("domain", ""),
                "call_recommendation":  s.get("call_recommendation", ""),
                "why_now":              c.get("why_now", ""),
                "trigger_type":         c.get("trigger_type", ""),
                "recency_bucket":       c.get("recency_bucket", ""),
                "is_current_trigger":   c.get("is_current_trigger", False),
                "recency_note":         c.get("recency_note", ""),
                "buying_window":        c.get("likely_buying_window", ""),
                "preferred_buyer_route": c.get("preferred_buyer_route", ""),
                "backup_buyer_route":   c.get("backup_buyer_route", ""),
                "title_searches":       c.get("suggested_title_searches", ""),
                "opener":               c.get("suggested_opener", ""),
                "evidence_summary":     c.get("trigger_evidence", ""),
                "evidence_quality":     c.get("evidence_quality", ""),
                "confidence_level":     c.get("confidence_level", ""),
            })
        pd.DataFrame(brief_rows).to_excel(
            writer, index=False, sheet_name="Caller Brief"
        )

        # ── Sheet 6: Raw Sources ──────────────────────────────────────────────
        raw_cols = [
            "company_name", "input_type", "query_group", "title", "url",
            "snippet", "date", "source_type",
        ]
        pd.DataFrame(raw_sources or [], columns=raw_cols).to_excel(
            writer, index=False, sheet_name="Raw Sources"
        )

    # ── Sheet 7: Caller Prep Input + optional original sheets ─────────────────
    buf.seek(0)
    wb = openpyxl.load_workbook(buf)

    ws_cp = wb.create_sheet("Caller Prep Input")
    _write_caller_prep_sheet(ws_cp, results)

    # Copy original Lead Scores / Company Profiles / Opportunity Input if present
    if input_bytes:
        try:
            src_wb = openpyxl.load_workbook(io.BytesIO(input_bytes))
            for src_name, dst_name in [
                ("Lead Scores",      "Original Lead Scores"),
                ("Company Profiles", "Original Company Profiles"),
                ("Opportunity Input","Original Opportunity Input"),
            ]:
                if src_name in src_wb.sheetnames:
                    try:
                        _copy_sheet(src_wb[src_name], wb, dst_name)
                    except Exception:
                        pass
        except Exception:
            pass

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# =============================================================================
# AUTOSAVE HELPERS
# =============================================================================

def _output_folder() -> pathlib.Path:
    """Return the autosave folder: ~/Downloads if it exists, else ./output/."""
    downloads = pathlib.Path.home() / "Downloads"
    if downloads.exists():
        return downloads
    fallback = pathlib.Path(__file__).parent / "output"
    fallback.mkdir(exist_ok=True)
    return fallback


def _unique_path(folder: pathlib.Path, name: str) -> pathlib.Path:
    """Return a path in folder that does not already exist."""
    p = folder / name
    if not p.exists():
        return p
    stem, suffix = p.stem, p.suffix
    for i in range(1, 1000):
        p2 = folder / f"{stem}_{i}{suffix}"
        if not p2.exists():
            return p2
    return folder / f"{stem}_dup{suffix}"


def _autosave_excel(excel_bytes: bytes, filename: str) -> pathlib.Path:
    """Write excel_bytes to the output folder and return the path written."""
    folder = _output_folder()
    p = _unique_path(folder, filename)
    p.write_bytes(excel_bytes)
    return p


# =============================================================================
# PIPELINE FOLDER HELPERS
# =============================================================================

def _parse_pipeline_filename(input_path: str) -> dict:
    """
    Parse cohort/batch_number/row_range from a prioritized xlsx filename.
    Scans left-to-right for first purely-numeric segment = batch_number.
    """
    stem = pathlib.Path(input_path).stem
    stem = re.sub(
        r'_(prioritized|cleaned|enriched|opportunity|lead_prioritized)_\d{8}_\d{4}.*$',
        '', stem, flags=re.IGNORECASE,
    )
    parts = stem.split("_")
    batch_number = ""
    batch_idx    = -1
    for i, p in enumerate(parts):
        if re.fullmatch(r'\d+', p):
            batch_number = p
            batch_idx    = i
            break
    if batch_idx < 0:
        return {"cohort": stem, "batch_number": "", "row_range": "", "batch_stem": stem, "valid": False}
    cohort    = "_".join(parts[:batch_idx])
    row_range = ""
    for j in range(batch_idx + 1, len(parts) - 1):
        candidate = parts[j] + "_" + parts[j + 1]
        if re.fullmatch(r'R\d+_\d+', candidate):
            row_range = candidate
            break
    batch_stem = cohort + "_" + batch_number
    if row_range:
        batch_stem = batch_stem + "_" + row_range
    return {
        "cohort":       cohort,
        "batch_number": batch_number,
        "row_range":    row_range,
        "batch_stem":   batch_stem,
        "valid":        bool(cohort and batch_number),
    }


def _resolve_opportunity_paths(
    input_path: str,
    project_root: str | None = None,
    ts: str | None = None,
) -> dict:
    """Resolve pipeline output paths for the Opportunity Radar."""
    if ts is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
    parsed = _parse_pipeline_filename(input_path)
    cohort = parsed.get("cohort", "unknown")
    stem   = parsed.get("batch_stem", pathlib.Path(input_path).stem)
    # Auto-detect project_root by walking up from 02_lead_prioritized/
    if not project_root:
        p = pathlib.Path(input_path).resolve()
        for parent in p.parents:
            if parent.name.startswith("02_"):
                project_root = str(parent.parent.parent)
                break
        if not project_root:
            project_root = str(pathlib.Path(input_path).resolve().parent.parent.parent)
    root       = pathlib.Path(project_root)
    cohort_dir = root / cohort
    output_dir = cohort_dir / "03_opportunity_radar"
    logs_dir   = cohort_dir / "_logs"
    return {
        "cohort":       cohort,
        "batch_stem":   stem,
        "batch_number": parsed.get("batch_number", ""),
        "row_range":    parsed.get("row_range", ""),
        "valid_name":   parsed.get("valid", False),
        "project_root": str(root),
        "cohort_dir":   str(cohort_dir),
        "output_dir":   str(output_dir),
        "logs_dir":     str(logs_dir),
        "output_xlsx":  str(output_dir / f"{stem}_opportunity_{ts}.xlsx"),
        "run_log_csv":  str(logs_dir   / f"{cohort}_opportunity_runlog.csv"),
        "ts":           ts,
    }


def _load_cli_secrets(
    cli_anthropic: str | None,
    cli_serper: str | None,
) -> tuple[str, str]:
    """Load API keys: CLI arg → env var → secrets.toml. Never logs key values."""
    import os as _os

    def _from_toml(name: str) -> str:
        try:
            import tomllib
        except ImportError:
            try:
                import tomli as tomllib  # type: ignore[no-redef]
            except ImportError:
                return ""
        p = pathlib.Path(__file__).parent / ".streamlit" / "secrets.toml"
        if not p.exists():
            return ""
        try:
            with open(p, "rb") as f:
                data = tomllib.load(f)
            return str(data.get(name, "") or "")
        except Exception:
            return ""

    anthropic_key = (
        cli_anthropic
        or _os.environ.get("ANTHROPIC_API_KEY", "")
        or _from_toml("ANTHROPIC_API_KEY")
        or _from_toml("anthropic_api_key")
        or ""
    )
    serper_key = (
        cli_serper
        or _os.environ.get("SERPER_API_KEY", "")
        or _from_toml("SERPER_API_KEY")
        or _from_toml("serper_api_key")
        or ""
    )
    return anthropic_key.strip(), serper_key.strip()


def _append_opportunity_runlog(csv_path: pathlib.Path, row: dict) -> None:
    """Append one row to the opportunity run log CSV."""
    import csv as _csv
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "timestamp", "cohort", "batch_stem", "input_file", "output_file",
        "mode", "max_rows", "rows_in_input", "rows_processed",
        "api_calls_attempted", "api_calls_successful", "api_calls_failed",
        "serper_key_loaded", "anthropic_key_loaded", "status", "error_message",
    ]
    write_header = not csv_path.exists()
    with open(csv_path, "a", newline="", encoding="utf-8") as f:
        writer = _csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        if write_header:
            writer.writeheader()
        writer.writerow(row)


# =============================================================================
# CLI ENTRY POINT
# =============================================================================

def run_cli() -> None:
    """Non-Streamlit batch entry point for the Opportunity Radar."""
    import argparse

    parser = argparse.ArgumentParser(
        description="mYngle Opportunity Radar — CLI batch mode",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--input",          required=True,  help="Path to prioritized .xlsx input")
    parser.add_argument("--project-root",   default=None,   help="Pipeline project root folder")
    parser.add_argument("--serper-key",     default=None,   help="Serper API key")
    parser.add_argument("--anthropic-key",  default=None,   help="Anthropic API key")
    parser.add_argument("--max-rows",       type=int, default=0, help="Process first N rows (0=all)")
    parser.add_argument("--debug",          action="store_true", help="Enable debug output")
    parser.add_argument("--dry-run-paths",  action="store_true", help="Print resolved paths and exit")
    parser.add_argument("--force-fresh",    action="store_true", help="Bypass result cache")
    args = parser.parse_args()

    input_path = pathlib.Path(args.input).resolve()
    if not input_path.exists():
        print(f"ERROR: input file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    ts           = datetime.now().strftime("%Y%m%d_%H%M")
    pl           = _resolve_opportunity_paths(str(input_path), args.project_root, ts=ts)
    anthropic_key, serper_key = _load_cli_secrets(args.anthropic_key, args.serper_key)

    print("[radar] CLI mode detected", flush=True)
    print(f"[radar] Input:            {input_path}", flush=True)
    print(f"[radar] Project root:     {pl['project_root']}", flush=True)
    print(f"[radar] Cohort:           {pl['cohort']}", flush=True)
    print(f"[radar] Batch stem:       {pl['batch_stem']}", flush=True)
    print(f"[radar] Output file:      {pl['output_xlsx']}", flush=True)
    print(f"[radar] Max rows:         {args.max_rows if args.max_rows > 0 else 'all'}", flush=True)
    print(f"[radar] Debug:            {args.debug}", flush=True)
    print(f"[radar] Force fresh:      {args.force_fresh}", flush=True)
    print(f"[radar] Serper key:       {'loaded' if serper_key else 'missing'}", flush=True)
    print(f"[radar] Anthropic key:    {'loaded' if anthropic_key else 'missing'}", flush=True)

    if args.dry_run_paths:
        out_dir  = pathlib.Path(pl["output_dir"])
        logs_dir = pathlib.Path(pl["logs_dir"])
        out_dir.mkdir(parents=True, exist_ok=True)
        logs_dir.mkdir(parents=True, exist_ok=True)
        print(f"\n[radar] --dry-run-paths results:")
        print(f"  input exists:         {input_path.exists()}")
        print(f"  project root exists:  {pathlib.Path(pl['project_root']).exists()}")
        print(f"  cohort dir exists:    {pathlib.Path(pl['cohort_dir']).exists()}")
        print(f"  output dir:           {pl['output_dir']}")
        print(f"  run log:              {pl['run_log_csv']}")
        print(f"  output dir ready:     True")
        sys.exit(0)

    # ── Load input ────────────────────────────────────────────────────────────
    fname = input_path.name.lower()
    df_in = pd.read_csv(input_path) if fname.endswith(".csv") else pd.read_excel(input_path)
    if args.max_rows and args.max_rows > 0:
        df_in = df_in.head(args.max_rows)

    input_type  = _detect_input_type(df_in)
    name_col    = _detect_col(df_in, _NAME_CANDIDATES)
    domain_col  = _detect_col(df_in, _DOMAIN_CANDIDATES)
    score_col   = _detect_col(df_in, _SCORE_CANDIDATES)
    tier_col    = _detect_col(df_in, _TIER_CANDIDATES)
    icp_col     = _detect_col(df_in, _ICP_CANDIDATES)
    country_col = _detect_col(df_in, _COUNTRY_CANDIDATES)

    print(f"[radar] Rows to process:  {len(df_in)}", flush=True)
    print(f"[radar] Input type:       {input_type}", flush=True)
    print(f"[radar] Name col:         {name_col}", flush=True)
    print(f"[radar] Domain col:       {domain_col}", flush=True)

    if not name_col:
        print("ERROR: could not detect company name column.", file=sys.stderr)
        sys.exit(1)

    company_list = _build_company_list(
        df_in, name_col, domain_col, country_col, score_col, tier_col, icp_col,
        input_type=input_type,
    )
    n_total = len(company_list)

    # ── Build Anthropic client ────────────────────────────────────────────────
    client = None
    if anthropic_key and _ANTHROPIC_AVAILABLE:
        try:
            client = _anthropic_mod.Anthropic(api_key=anthropic_key)
        except Exception as _ce:
            print(f"[radar] Anthropic client error: {_ce}", flush=True)

    # ── Process companies ─────────────────────────────────────────────────────
    results:      list = []
    raw_sources:  list = []
    n_api_att = n_api_ok = n_api_fail = 0
    out_dir   = pathlib.Path(pl["output_dir"])
    out_dir.mkdir(parents=True, exist_ok=True)
    partial_path = out_dir / f"{pl['batch_stem']}_opportunity_PARTIAL_{ts}.xlsx"
    error_msg = ""

    try:
        for i, company in enumerate(company_list, 1):
            print(f"\r[radar] {i}/{n_total} — {str(company.get('name', ''))[:40]:<40}",
                  end="", flush=True)
            name   = company.get("name", "")
            domain = company.get("domain", "")

            # Cache check
            cached = None if args.force_fresh else _cache_load(name, domain, input_type)
            if cached:
                results.append(cached)
                raw_sources.extend(cached.get("_raw_sources", []))
                continue

            # Serper searches
            grouped: dict = {}
            if serper_key:
                try:
                    grouped = _run_searches(name, domain, serper_key)
                    n_api_att += 1
                    n_api_ok  += 1
                except Exception as _se:
                    n_api_att  += 1
                    n_api_fail += 1
                    if args.debug:
                        print(f"\n[radar] Serper error {name}: {_se}", flush=True)

            company_raw_sources = _collect_raw_sources(name, grouped, input_type)

            # Claude analysis
            claude_result: dict = {}
            if client:
                try:
                    claude_result = _call_claude(
                        name=name,
                        domain=domain,
                        country=company.get("country", ""),
                        fit_score=str(company.get("fit_score", "") or ""),
                        tier=str(company.get("tier", "") or ""),
                        icp_evidence=str(company.get("icp_evidence", "") or ""),
                        grouped_results=grouped,
                        client=client,
                    )
                    n_api_att += 1
                    n_api_ok  += 1
                except Exception as _ce:
                    n_api_att  += 1
                    n_api_fail += 1
                    if args.debug:
                        print(f"\n[radar] Claude error {name}: {_ce}", flush=True)

            # Scores + fallbacks
            adj, scores = _compute_scores(
                claude_result,
                fit_score_raw  = company.get("fit_score"),
                tier_raw       = company.get("tier", ""),
                input_type     = input_type,
                icp_evidence   = str(company.get("icp_evidence", "") or ""),
                company_name   = name,
            )
            row_out = {
                **adj,
                **scores,
                "company_name": name,
                "domain":       domain,
                "country":      company.get("country", ""),
                "fit_score":    company.get("fit_score", ""),
                "tier":         company.get("tier", ""),
                "icp_evidence": company.get("icp_evidence", ""),
                "input_type":   input_type,
                "commercial_fit_available": input_type == "enriched_export",
                "_raw_sources": company_raw_sources,
            }
            _cache_save(name, domain, input_type, row_out)
            results.append(row_out)
            raw_sources.extend(company_raw_sources)

        print(f"\n[radar] Processing complete — {len(results)}/{n_total}.", flush=True)

    except KeyboardInterrupt:
        print(f"\n[radar] Interrupted — saving partial output ({len(results)} rows).", flush=True)
    except Exception as _exc:
        error_msg = f"{type(_exc).__name__}: {_exc}"
        print(f"\n[radar] Error: {error_msg}", flush=True)

    # ── Write Excel ───────────────────────────────────────────────────────────
    out_path = pathlib.Path("")
    if results:
        try:
            xl_bytes = _build_excel_bytes(results, raw_sources)
            is_partial = len(results) < n_total
            out_path = partial_path if is_partial else out_dir / f"{pl['batch_stem']}_opportunity_{ts}.xlsx"
            out_path.write_bytes(xl_bytes)
            print(f"[radar] Saved: {out_path}", flush=True)
        except Exception as _we:
            print(f"[radar] Excel write error: {_we}", flush=True)
            error_msg = error_msg or str(_we)
    else:
        print("[radar] No results — no output written.", flush=True)

    # ── Run log ───────────────────────────────────────────────────────────────
    _append_opportunity_runlog(
        pathlib.Path(pl["run_log_csv"]),
        {
            "timestamp":           ts,
            "cohort":              pl["cohort"],
            "batch_stem":          pl["batch_stem"],
            "input_file":          str(input_path),
            "output_file":         str(out_path),
            "mode":                "cli",
            "max_rows":            args.max_rows if args.max_rows > 0 else "all",
            "rows_in_input":       len(df_in),
            "rows_processed":      len(results),
            "api_calls_attempted": n_api_att,
            "api_calls_successful":n_api_ok,
            "api_calls_failed":    n_api_fail,
            "serper_key_loaded":   bool(serper_key),
            "anthropic_key_loaded":bool(anthropic_key),
            "status":              "complete" if (len(results) == n_total and not error_msg) else "partial",
            "error_message":       error_msg,
        },
    )
    print(f"[radar] Run log: {pl['run_log_csv']}", flush=True)
    print("[radar] Done.", flush=True)



# =============================================================================
# CONSTANTS
# =============================================================================

CLAUDE_MODEL    = "claude-haiku-4-5-20251001"
SERPER_URL      = "https://google.serper.dev/search"
RADAR_CACHE_DIR = pathlib.Path("radar_cache")

# Bump this string whenever the prompt or interpretation logic changes.
# It is included in cache keys so old Claude outputs are never reused after a prompt update.
CACHE_VERSION = "v3_myngle_20260608"

# 5 query groups — one Serper call each
QUERY_GROUPS = [
    (
        "Annual Report / Financial",
        '"{name}" annual report fiscal year results revenue 2024 2025',
    ),
    (
        "International Hiring / Growth",
        '"{name}" hiring international careers jobs "new office" global expansion 2024 2025',
    ),
    (
        "Language / Communication / L&D",
        '"{name}" "language training" OR "business English" OR "communication training" '
        'OR "learning and development" OR "talent development" OR training academy HR',
    ),
    (
        "Sales / Customer Success Expansion",
        '"{name}" "sales team" OR "customer success" OR "account management" '
        'OR "sales enablement" OR "client-facing" international expansion',
    ),
    (
        "M&A / Funding / Integration",
        '"{name}" acquisition OR merger OR integration OR funding OR investment OR "private equity"',
    ),
]


def run_streamlit_app() -> None:
    _st = _get_st()

    # =============================================================================
    # PAGE CONFIG
    # =============================================================================

    _st.set_page_config(
        page_title="mYngle · Opportunity Radar",
        page_icon="📡",
        layout="wide",
        initial_sidebar_state="collapsed",
    )

    # =============================================================================
    # HEADER  (logo + title — mirrors lead prioritizer layout exactly)
    # =============================================================================

    _logo_path = pathlib.Path(__file__).parent / "mingle_local_final_fixed.png"
    _logo_src  = (
        f"data:image/png;base64,{base64.b64encode(_logo_path.read_bytes()).decode()}"
        if _logo_path.exists() else ""
    )
    _img_tag = (
        f'<img src="{_logo_src}" class="brand-logo" alt="mYngle" />'
        if _logo_src else ""
    )

    _st.markdown(
        f"""
        <style>
        .block-container {{
            max-width: 880px;
            padding-top: 2.2rem;
            padding-bottom: 3rem;
            padding-left: 2rem;
            padding-right: 2rem;
        }}

        div[data-testid="stMarkdownContainer"]:has(.brand-header) {{
            overflow: visible !important;
            margin-bottom: 1.0rem;
        }}

        .brand-header {{
            display: grid;
            grid-template-columns: 43% 57%;
            align-items: center;
            min-height: 140px;
            padding-top: 10px;
            padding-bottom: 6px;
            overflow: visible !important;
        }}

        .brand-title-block {{
            display: flex;
            align-items: center;
            justify-content: flex-start;
            overflow: visible !important;
        }}

        .brand-title {{
            font-size: 42px;
            font-weight: 700;
            color: #0B1F3A;
            line-height: 1.1;
            white-space: nowrap;
            margin: 0;
            padding: 0;
        }}

        .brand-logo-block {{
            display: flex;
            justify-content: flex-end;
            align-items: center;
            padding: 0;
            line-height: 0;
            overflow: visible !important;
        }}

        .brand-logo {{
            width: 430px;
            max-width: 100%;
            height: auto;
            display: block;
            object-fit: contain;
            object-position: center center;
            overflow: visible !important;
        }}
        </style>

        <div class="brand-header">
          <div class="brand-title-block">
            <span class="brand-title">Opportunity Radar</span>
          </div>
          <div class="brand-logo-block">
            {_img_tag}
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # STEP 1 — UPLOAD
    # =============================================================================

    _st.divider()
    _st.subheader("Step 1 · Upload your file")
    _st.caption(
        "Upload a Lead Prioritizer export, an Opportunity Input sheet, "
        "or a simple company list with company name and website."
    )

    uploaded = _st.file_uploader(
        "Drag and drop here, or click to browse  (.xlsx · .xls · .csv)",
        type=["xlsx", "xls", "csv"],
        label_visibility="collapsed",
    )

    new_key = f"{uploaded.name}___{uploaded.size}" if uploaded else "__none__"
    if new_key != ss("_or_file_key", "__none__"):
        ss_set(
            _or_file_key      = new_key,
            _or_df_raw        = None,
            _or_file_name     = None,
            _or_file_error    = None,
            _or_name_col      = None,
            _or_domain_col    = None,
            _or_country_col   = None,
            _or_score_col     = None,
            _or_tier_col      = None,
            _or_icp_col       = None,
            _or_n_companies   = 0,
            _or_input_type    = None,
            _or_processing    = False,
            _or_done          = False,
            _or_process_index = 0,
            _or_company_list  = None,
            _or_results       = None,
            _or_raw_sources   = None,
            _or_excel_bytes   = None,
            _or_stop          = False,
        )
        if uploaded is not None:
            try:
                fname             = uploaded.name
                # Capture raw bytes before _load_df_from_upload consumes the file
                try:
                    uploaded.seek(0)
                    _raw_bytes = uploaded.read()
                    uploaded.seek(0)
                except Exception:
                    _raw_bytes = None
                df_loaded, sheet  = _load_df_from_upload(uploaded)
                input_type        = _detect_input_type(df_loaded)
                name_col          = _detect_col(df_loaded, _NAME_CANDIDATES)
                domain_col        = _detect_col(df_loaded, _DOMAIN_CANDIDATES)
                country_col       = _detect_col(df_loaded, _COUNTRY_CANDIDATES)
                score_col         = _detect_col(df_loaded, _SCORE_CANDIDATES)
                tier_col          = _detect_col(df_loaded, _TIER_CANDIDATES)
                icp_col           = _detect_col(df_loaded, _ICP_CANDIDATES)
                n                 = _count_companies(df_loaded, name_col)
                ss_set(
                    _or_df_raw      = df_loaded,
                    _or_file_name   = fname,
                    _or_name_col    = name_col,
                    _or_domain_col  = domain_col,
                    _or_country_col = country_col,
                    _or_score_col   = score_col,
                    _or_tier_col    = tier_col,
                    _or_icp_col     = icp_col,
                    _or_n_companies = n,
                    _or_input_type  = input_type,
                    _or_input_bytes = _raw_bytes,
                )
            except Exception as exc:
                ss_set(_or_file_error=str(exc))

    if ss("_or_file_error"):
        _st.error(f"Could not read the file: {ss('_or_file_error')}")
    elif ss("_or_df_raw") is not None and not ss("_or_processing", False) and not ss("_or_done", False):
        n          = ss("_or_n_companies", 0)
        itype      = ss("_or_input_type", "")
        itype_label = (
            "enriched export detected" if itype == "enriched_export"
            else "simple company list detected"
        )
        _st.success(
            f"✓ **{ss('_or_file_name')}** loaded · "
            f"{n:,} {'company' if n == 1 else 'companies'} ready · "
            f"{itype_label}"
        )

    # =============================================================================
    # API KEY STATUS
    # =============================================================================

    if not ss("_or_done", False):
        if _keys_ok:
            _st.success("🔑 API keys detected: Serper and Claude ready")
        else:
            if not _serper_key:
                _st.error(
                    "Missing SERPER_API_KEY. "
                    "Opportunity Radar cannot search for company triggers without it. "
                    "Add it to .streamlit/secrets.toml."
                )
            if not _anthropic_key or not _ANTHROPIC_AVAILABLE:
                _st.error(
                    "Missing ANTHROPIC_API_KEY. "
                    "Opportunity Radar cannot interpret evidence without it. "
                    "Add it to .streamlit/secrets.toml."
                )

    # =============================================================================
    # STEP 2 — START / PROCESSING LOOP
    # =============================================================================

    _ready      = ss("_or_df_raw") is not None and _keys_ok
    _processing = ss("_or_processing", False)
    _done       = ss("_or_done", False)

    if not _done and not _processing:
        use_cache = _st.checkbox(
            "Use cached results when available",
            value=ss("_or_force_refresh", False),
            key="or_force_refresh_cb",
            help=(
                f"When checked, previously cached analysis may be reused. "
                f"When unchecked (default), every company is re-fetched and re-analysed from scratch. "
                f"Cache version: {CACHE_VERSION}"
            ),
        )
        ss_set(_or_force_refresh=use_cache)
        mode_text = (
            "Scan mode: cached results may be reused"
            if ss("_or_force_refresh", False)
            else "Scan mode: fresh search and fresh analysis"
        )
        _st.caption(mode_text)

        _col_a, _col_b = _st.columns([3, 2])
        with _col_a:
            autosave_cb = _st.checkbox(
                "Autosave results to Downloads",
                value=ss("_or_autosave", True),
                key="or_autosave_cb",
            )
            ss_set(_or_autosave=autosave_cb)
        with _col_b:
            chk_n_val = _st.number_input(
                "Checkpoint every N companies",
                min_value=1, max_value=50,
                value=int(ss("_or_checkpoint_n", 10)),
                step=1, key="or_checkpoint_n_input",
            )
            ss_set(_or_checkpoint_n=int(chk_n_val))

        if autosave_cb:
            _st.caption(f"Autosave destination: {_output_folder()}")

        start_btn = _st.button(
            "▶ Start radar scan",
            type="primary",
            use_container_width=True,
            disabled=not _ready,
            key="or_start_btn",
        )

        if start_btn and _ready:
            df_raw = ss("_or_df_raw")
            company_list = _build_company_list(
                df_raw,
                ss("_or_name_col"),
                ss("_or_domain_col"),
                ss("_or_country_col"),
                ss("_or_score_col"),
                ss("_or_tier_col"),
                ss("_or_icp_col"),
                ss("_or_input_type", "simple_company_list"),
            )
            ss_set(
                _or_processing     = True,
                _or_done           = False,
                _or_process_index  = 0,
                _or_company_list   = company_list,
                _or_results        = [],
                _or_raw_sources    = [],
                _or_excel_bytes    = None,
                _or_stop           = False,
                _or_scan_start_time = time.time(),
            )
            _st.rerun()

    if _processing and not _done:
        company_list = ss("_or_company_list", [])
        results      = ss("_or_results", [])
        raw_sources  = ss("_or_raw_sources", [])
        idx          = ss("_or_process_index", 0)
        n_total      = len(company_list)

        # ── Stop button ───────────────────────────────────────────────────────────
        if _st.button("⏹ Stop", key="or_stop_btn"):
            ss_set(_or_stop=True)

        if ss("_or_stop", False):
            ss_set(_or_processing=False, _or_done=True)
            _st.rerun()

        # ── Progress display ──────────────────────────────────────────────────────
        if n_total:
            frac = idx / n_total if n_total else 0
            _st.progress(frac)
            _st.markdown("**Scanning opportunities…**")

            if idx < n_total:
                current_name = (
                    company_list[idx].get("company_name")
                    or company_list[idx].get("domain", "")
                )
                _st.caption(f"Company {idx + 1} of {n_total}: {current_name}")

            scan_start = ss("_or_scan_start_time")
            if scan_start:
                elapsed_sec = time.time() - scan_start
                elapsed_str = time.strftime("%M:%S", time.gmtime(int(elapsed_sec)))
                _st.caption(f"Elapsed: {elapsed_str}")
                completed = idx  # companies fully processed so far
                if completed >= 2 and n_total > idx:
                    avg_sec        = elapsed_sec / completed
                    remaining_sec  = (n_total - idx) * avg_sec
                    remaining_str  = time.strftime("%M:%S", time.gmtime(int(remaining_sec)))
                    ready_time     = datetime.now() + timedelta(seconds=remaining_sec)
                    ready_str      = ready_time.strftime("%H:%M")
                    _st.caption(f"Estimated remaining: {remaining_str}")
                    _st.caption(f"Estimated ready around: {ready_str}")
                elif idx < n_total:
                    _st.caption("Estimating time remaining…")

            # Show latest checkpoint status
            chk_path = ss("_or_last_checkpoint_path")
            chk_count = ss("_or_last_checkpoint_count")
            if chk_path and chk_count:
                _st.caption(
                    f"Latest checkpoint saved after {chk_count} companies: {chk_path}"
                )

        # ── Process one company ───────────────────────────────────────────────────
        if idx < n_total:
            company     = company_list[idx]
            name        = company.get("company_name", "")
            domain      = company.get("domain", "")
            country     = company.get("country", "")
            fit_score   = company.get("fit_score", "")
            tier        = company.get("tier", "")
            icp_ev      = company.get("icp_evidence", "")
            c_itype     = company.get("input_type", "simple_company_list")
            c_fit_avail = company.get("commercial_fit_available", False)
            is_internal = company.get("internal", False)

            enriched_row = company.get("enriched_row", {})

            if is_internal:
                # Mark without any research
                record = {
                    "company_name":             name,
                    "domain":                   domain,
                    "country":                  country,
                    "fit_score":                "",
                    "tier":                     "",
                    "input_type":               c_itype,
                    "commercial_fit_available": False,
                    "claude": {
                        **_EMPTY_CLAUDE_RESULT,
                        "why_now": "Internal / exclude",
                        "trigger_evidence": "Internal company — excluded from radar.",
                    },
                    "scores": {
                        "trigger_score":       0,
                        "buying_window_score": 0,
                        "contact_route_score": 0,
                        "opportunity_score":   0,
                        "call_recommendation": "Internal / exclude",
                    },
                    "enriched_row": enriched_row,
                }
                results.append(record)
            else:
                # Check cache — only use if "Use cached results" is checked
                use_cache = ss("_or_force_refresh", False)
                cached = _cache_load(name, domain, c_itype) if use_cache else None
                if cached is not None:
                    # Enforce correct fit data for this input type
                    cached["input_type"]              = c_itype
                    cached["commercial_fit_available"] = c_fit_avail
                    if c_itype == "simple_company_list":
                        # Strip any enriched fit values that crept into the cache
                        cached["fit_score"] = ""
                        cached["tier"]      = ""
                    # Reattach enriched_row (not stored in cache)
                    cached["enriched_row"] = enriched_row
                    # Re-apply window adjustment and recompute scores (in case window aged)
                    adj_claude, fresh_scores = _compute_scores(
                        cached.get("claude", {}),
                        cached.get("fit_score", ""),
                        cached.get("tier", ""),
                        c_itype,
                        icp_evidence=icp_ev,
                        company_name=name,
                    )
                    cached["claude"] = adj_claude
                    cached["scores"] = fresh_scores
                    results.append(cached)
                    # Restore raw sources stored in cache (if any)
                    raw_sources.extend(cached.get("raw_sources", []))
                else:
                    # Run Serper searches
                    grouped_results = _run_searches(name or domain, domain, _serper_key)
                    sources         = _collect_raw_sources(name, grouped_results, c_itype)

                    # Call Claude
                    client = _anthropic_mod.Anthropic(api_key=_anthropic_key)
                    raw_claude = _call_claude(
                        name, domain, country, fit_score, tier, icp_ev,
                        grouped_results, client,
                    )

                    # Adjust window + compute scores; formula depends on input type
                    adj_claude, scores = _compute_scores(
                        raw_claude, fit_score, tier, c_itype,
                        icp_evidence=icp_ev, company_name=name,
                    )

                    # For simple lists: never carry commercial fit values
                    out_fit_score = fit_score if c_itype == "enriched_export" else ""
                    out_tier      = tier      if c_itype == "enriched_export" else ""

                    record = {
                        "company_name":             name,
                        "domain":                   domain,
                        "country":                  country,
                        "fit_score":                out_fit_score,
                        "tier":                     out_tier,
                        "input_type":               c_itype,
                        "commercial_fit_available": c_fit_avail,
                        "claude":                   adj_claude,
                        "scores":                   scores,
                        "raw_sources":              sources,
                        "enriched_row":             enriched_row,
                    }
                    _cache_save(name, domain, c_itype, record)
                    results.append(record)
                    raw_sources.extend(sources)

            new_idx = idx + 1
            ss_set(
                _or_results       = results,
                _or_raw_sources   = raw_sources,
                _or_process_index = new_idx,
            )

            # ── Checkpoint autosave ───────────────────────────────────────────────
            if ss("_or_autosave", True):
                chk_n = ss("_or_checkpoint_n", 10)
                if new_idx > 0 and new_idx % chk_n == 0:
                    try:
                        chk_bytes = _build_excel_bytes(results, raw_sources,
                                                       input_bytes=ss("_or_input_bytes"))
                        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                        chk_name = (
                            f"opportunity_radar_checkpoint_{ts}_after_{new_idx:03d}.xlsx"
                        )
                        chk_path = _autosave_excel(chk_bytes, chk_name)
                        ss_set(
                            _or_last_checkpoint_path=str(chk_path),
                            _or_last_checkpoint_count=new_idx,
                        )
                    except Exception:
                        pass  # never block the scan on a save failure

            _st.rerun()

        else:
            # All companies processed
            ss_set(_or_processing=False, _or_done=True)
            _st.rerun()

    # =============================================================================
    # STEP 3 — RESULTS + DOWNLOAD
    # =============================================================================

    if _done:
        results     = ss("_or_results", [])
        raw_sources = ss("_or_raw_sources", [])
        processed   = len(results)
        n           = ss("_or_n_companies", 0)

        scan_start = ss("_or_scan_start_time")
        avg_note = ""
        if scan_start and processed > 0:
            total_sec = time.time() - scan_start
            avg_sec   = total_sec / processed
            avg_note  = f" · avg {avg_sec:.0f} s/company"

        _st.success(
            f"✅ Ready · **{processed:,}** "
            f"{'company' if processed == 1 else 'companies'} scanned{avg_note}"
        )

        # Build Excel once, cache bytes in session state
        if ss("_or_excel_bytes") is None:
            ss_set(_or_excel_bytes=_build_excel_bytes(results, raw_sources,
                                                       input_bytes=ss("_or_input_bytes")))

        # Final autosave — runs once per completed scan
        if ss("_or_autosave", True) and ss("_or_final_autosave_path") is None:
            try:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                final_name = f"opportunity_radar_autosave_{ts}.xlsx"
                final_path = _autosave_excel(ss("_or_excel_bytes"), final_name)
                ss_set(_or_final_autosave_path=str(final_path))
            except Exception:
                pass  # autosave failure never blocks download

        if ss("_or_final_autosave_path"):
            _st.caption(f"Final autosave saved: {ss('_or_final_autosave_path')}")

        _st.download_button(
            label="⬇ Download opportunity radar",
            data=ss("_or_excel_bytes"),
            file_name=f"opportunity_radar_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )

        _st.divider()
        if _st.button("↺ Start a new radar scan", use_container_width=True, key="or_restart_btn"):
            reset()
            _st.rerun()



if __name__ == "__main__":
    if cli_args_present():
        run_cli()
    elif running_under_streamlit():
        run_streamlit_app()
    else:
        run_cli()